"""Build Cafe & Coffee Shop Command Center™ — The Complete Coffee-Shop System.

14 tabs · a premium cafe operating system in Google Sheets & Excel. Dashboard, a
cup-cost engine, a menu board with beverage-cost %, daypart sales, weekly sales,
labor & prime cost, bean & milk usage, inventory & par, a waste log, ordering,
cash & tips and a regulars/loyalty tracker — one dashboard. Cost every cup, watch
your prime cost & pour more profit.

Run: python3 build_xlsx.py   ->  ../Cafe_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["g", "oz", "each", "ml", "shot", "pump", "lb", "gal"]
CATS = ["Espresso", "Brewed", "Cold", "Tea", "Food", "Pastry", "Retail"]
ROLES = ["Lead Barista", "Barista", "Shift Sup", "Baker", "Cashier"]

TARGET_BEV = 0.28
TARGET_LABOR = 0.30

# Cup cost engine — flagship Latte: (component, qty, unit, cost/unit)
LATTE = [
    ("Espresso beans", 18, "g", 0.024),
    ("Whole milk", 12, "oz", 0.048),
    ("12 oz cup", 1, "each", 0.12),
    ("Lid", 1, "each", 0.03),
    ("Sleeve", 1, "each", 0.02),
]

# Menu board: (item, category, cup cost, price, units/day). Latte cost = LatteCost.
MENU = [
    ("Latte", "Espresso", None, 5.25, 90),
    ("Cappuccino", "Espresso", 0.98, 4.75, 55),
    ("Drip Coffee", "Brewed", 0.52, 3.25, 120),
    ("Cold Brew", "Cold", 1.10, 5.00, 70),
    ("Mocha", "Espresso", 1.45, 5.75, 45),
    ("Americano", "Espresso", 0.58, 4.00, 40),
    ("Espresso", "Espresso", 0.48, 3.50, 30),
    ("Chai Latte", "Tea", 1.22, 5.25, 35),
    ("Croissant", "Pastry", 1.20, 4.25, 60),
    ("Avocado Toast", "Food", 2.60, 9.50, 35),
]

# Daypart sales (representative day): (daypart, transactions, sales)
DAYPART = [
    ("Morning rush (6-10)", 190, 1329),
    ("Midday (10-2)", 100, 831),
    ("Afternoon (2-6)", 62, 609),
]

# Weekly sales: (day, sales, transactions)
WEEKLY = [
    ("Monday", 2450, 312), ("Tuesday", 2680, 338), ("Wednesday", 2769, 352),
    ("Thursday", 2900, 366), ("Friday", 3350, 410), ("Saturday", 3600, 430),
    ("Sunday", 2100, 270),
]

# Labor (7 days): (day, labor cost)
LABOR = [
    ("Monday", 686), ("Tuesday", 750), ("Wednesday", 775), ("Thursday", 812),
    ("Friday", 938), ("Saturday", 1008), ("Sunday", 588),
]

# Bean & milk usage (weekly): (item, qty, unit, cost)
USAGE = [
    ("Espresso beans", 22, "lb", 286.00),
    ("Whole milk", 48, "gal", 216.00),
    ("Oat milk", 18, "gal", 144.00),
    ("Cold brew concentrate", 6, "gal", 90.00),
    ("Syrups (asst.)", 12, "each", 96.00),
]

# Inventory & par: (item, par, on hand, unit)
INVENTORY = [
    ("Espresso beans", 30, 14, "lb"), ("Whole milk", 60, 28, "gal"), ("Oat milk", 24, 10, "gal"),
    ("12 oz cups", 1000, 420, "each"), ("Lids", 1000, 380, "each"), ("Pastries", 120, 45, "each"),
    ("Syrups", 18, 8, "each"), ("Cold cups", 800, 300, "each"),
]

# Waste log: (item, reason, cost)
WASTE = [
    ("Spoiled milk", "Over-ordered", 180.00),
    ("Stale pastries", "End of day", 145.00),
    ("Wasted shots", "Training", 90.00),
    ("Dumped brew", "Slow afternoon", 120.00),
    ("Dropped inventory", "Breakage", 80.00),
]

# Ordering / suppliers: (item, supplier, par order, cost)
ORDERING = [
    ("Espresso beans (bulk)", "Roaster Co.", 30, 390.00),
    ("Milk (case)", "Dairy Direct", 12, 54.00),
    ("Cups & lids", "PackSupply", 2000, 168.00),
    ("Pastries (par)", "Local Bakery", 120, 144.00),
    ("Syrups (case)", "FlavorHouse", 12, 96.00),
]

# Cash & tips (7 days): (day, cash, card, tips)
CASHTIPS = [
    ("Monday", 380, 2070, 210), ("Tuesday", 410, 2270, 230), ("Wednesday", 420, 2349, 245),
    ("Thursday", 440, 2460, 255), ("Friday", 520, 2830, 300), ("Saturday", 560, 3040, 330),
    ("Sunday", 320, 1780, 190),
]

# Regulars / loyalty: (name, favorite, visits/wk, spend/wk)
REGULARS = [
    ("Morning Marcus", "Oat Latte", 5, 26.25),
    ("Study Sara", "Cold Brew + Toast", 4, 58.00),
    ("Quick Quinn", "Drip Coffee", 6, 19.50),
    ("Weekend Wes", "Mocha", 2, 11.50),
    ("Chai Chloe", "Chai Latte", 5, 26.25),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your cafe, targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR CAFE", "section")
    controls = [
        ("Cafe name", "Wildroot Coffee", None, "Cafe"),
        ("Owner", "Priya", None, "Owner"),
        ("Target beverage-cost %", TARGET_BEV, "0%", "TargetBev"),
        ("Target labor %", TARGET_LABOR, "0%", "TargetLabor"),
        ("Avg-ticket goal", 8, '"$"#,##0', "TicketGoal"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Category", CATS, "CatList"),
             ("G", "Role", ROLES, "RoleList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ☕  CAFE & COFFEE SHOP COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every cup, watch your prime cost & pour more profit.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE CAFE, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Cost every drink down to the cup — beans, milk, cup, lid & sleeve — and see each menu item's "
                      "beverage-cost % and margin. Track sales by daypart so you know where the money is, keep labor "
                      "and prime cost in check, and watch bean & milk usage, inventory, waste, ordering, cash & tips "
                      "and your regulars — all in ONE premium Google Sheets & Excel system built for coffee shops.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your cafe name & target beverage-cost / labor %.",
             "2.  Build a Cup Cost — beans, milk, cup & lid give the true cost per drink.",
             "3.  Fill the Menu Board — beverage-cost % and margin calculate per item.",
             "4.  Log Daypart & Weekly Sales; track Labor to see your prime cost.",
             "5.  Keep Bean & Milk, Inventory, Waste, Ordering & Cash-&-Tips current.",
             "6.  Check the Dashboard: sales, avg ticket, prime cost & a Cafe Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional cafe (Wildroot Coffee) is included so you can see how it all connects — "
               "just type over it with your own drinks and numbers. Prime cost (beverage + labor) is the number that "
               "makes or breaks a cafe, and it rolls into a live Cafe Score. Twelve matching printable pages (cup-cost "
               "card, daypart log, prep & par list, waste log & more) are included for the bar. This is a business "
               "tool, not financial or accounting advice — confirm figures with your own books.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "A café lives and dies on prime cost — keep it under 60% and pour profit.", "section_gold")


# ===========================================================================
def build_cupcost(wb):
    ws = wb.create_sheet("Cup Cost"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 10, 10, 14, 14, 2])
    luxe_header(ws, "F", "☕  CUP COST",
                "Cost a drink to the cup — beans, milk, cup, lid & sleeve. The engine behind your menu.")
    merge_set(ws, "B5:C5", "DRINK", "section_gold")
    ws.cell(row=5, column=4, value="Latte (12 oz)").font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Component", "Qty", "Unit", "Cost/Unit", "Ext. Cost"], start_col=2)
    start = 7
    for i, (comp, qty, unit, cu) in enumerate(LATTE):
        r = start + i
        ws.cell(row=r, column=2, value=comp).style = "td_left"
        cq = ws.cell(row=r, column=3, value=qty); cq.style = "input"; cq.number_format = "0.##"
        ws.cell(row=r, column=4, value=unit).style = "td"
        cc = ws.cell(row=r, column=5, value=cu); cc.style = "input"; cc.number_format = '"$"#,##0.000'
        ce = ws.cell(row=r, column=6, value=f"=C{r}*E{r}"); ce.style = "td"; ce.number_format = '"$"#,##0.000'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LATTE) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="CUP COST").style = "th"
    for c in range(3, 6):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0.00'
    cell_name(wb, "LatteCost", "Cup Cost", f"$F${tot}")
    ws.cell(row=tot + 2, column=2, value="Copy this build for every drink — swap milk, add syrups, change the cup.").style = "section"


def build_menu(wb):
    ws = wb.create_sheet("Menu Board"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 12, 12, 12, 10, 12, 12, 2])
    luxe_header(ws, "H", "📋  MENU BOARD",
                "Every drink's cup cost, price, beverage-cost % & margin — price for a healthy pour.")
    table_headers(ws, 4, ["Item", "Category", "Cup Cost", "Price", "Units/day", "Bev %", "Margin $"], start_col=2)
    start = L0
    for i, (item, cat, cost, price, units) in enumerate(MENU):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        cc = ws.cell(row=r, column=4, value="=LatteCost" if cost is None else cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cu = ws.cell(row=r, column=6, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cb = ws.cell(row=r, column=7, value=f"=IFERROR(D{r}/E{r},0)"); cb.style = "td"; cb.number_format = "0%"
        cm = ws.cell(row=r, column=8, value=f"=E{r}-D{r}"); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MENU) - 1
    nrange(wb, "MenuItem", "Menu Board", "B", start, end)
    nrange(wb, "MenuCost", "Menu Board", "D", start, end)
    nrange(wb, "MenuPrice", "Menu Board", "E", start, end)
    nrange(wb, "MenuUnits", "Menu Board", "F", start, end)
    nrange(wb, "MenuMargin", "Menu Board", "H", start, end)
    # overall beverage-cost % (weighted by units sold) in a named cell
    br = end + 2
    ws.cell(row=br, column=2, value="Overall beverage-cost %").style = "field_label"
    c = ws.cell(row=br, column=4, value="=IFERROR(SUMPRODUCT(MenuCost,MenuUnits)/SUMPRODUCT(MenuPrice,MenuUnits),0)")
    c.style = "field_value"; c.number_format = "0%"; c.fill = fill(MINT_BG)
    cell_name(wb, "BevCostPct", "Menu Board", f"$D${br}")
    ws.conditional_formatting.add(f"G{start}:G{end}",
        ColorScaleRule(start_type="num", start_value=0.12, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.28, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.40, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


def build_daypart(wb):
    ws = wb.create_sheet("Daypart Sales"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 14, 14, 14, 2])
    luxe_header(ws, "E", "🕗  DAYPART SALES",
                "Where the money is by time of day — feed the rush, fix the lull.")
    table_headers(ws, 4, ["Daypart", "Transactions", "Sales", "Avg Ticket"], start_col=2)
    start = L0
    for i, (dp, txn, sales) in enumerate(DAYPART):
        r = start + i
        ws.cell(row=r, column=2, value=dp).style = "td_left"
        ct = ws.cell(row=r, column=3, value=txn); ct.style = "input"; ct.number_format = "#,##0"
        cs = ws.cell(row=r, column=4, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); ca.style = "td"; ca.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DAYPART) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="DAY TOTAL").style = "th"
    ct = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = "#,##0"
    cs = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    ca = ws.cell(row=tot, column=5, value=f"=IFERROR(D{tot}/C{tot},0)"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = '"$"#,##0.00'
    nrange(wb, "DaypartName", "Daypart Sales", "B", start, end)
    nrange(wb, "DaypartSales", "Daypart Sales", "D", start, end)
    cell_name(wb, "DayTotal", "Daypart Sales", f"$D${tot}")
    cell_name(wb, "TxnTotal", "Daypart Sales", f"$C${tot}")
    d = DoughnutChart(); d.title = "Sales by Daypart"; d.height = 7.4; d.width = 9
    d.add_data(Reference(ws, min_col=4, min_row=start, max_row=end), titles_from_data=False)
    d.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end)); d.dataLabels = no_labels()
    ws.add_chart(d, "G5")
    ws.freeze_panes = "A5"


def build_weekly(wb):
    ws = wb.create_sheet("Weekly Sales"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 2])
    luxe_header(ws, "E", "📈  WEEKLY SALES",
                "The week at a glance — sales, transactions & average ticket by day.")
    table_headers(ws, 4, ["Day", "Sales", "Transactions", "Avg Ticket"], start_col=2)
    start = L0
    for i, (day, sales, txn) in enumerate(WEEKLY):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        ct = ws.cell(row=r, column=4, value=txn); ct.style = "input"; ct.number_format = "#,##0"
        ca = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); ca.style = "td"; ca.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(WEEKLY) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    cs = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    ct = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = "#,##0"
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "WeekSalesTotal", "Weekly Sales", f"$C${tot}")
    ws.add_chart(_barchart(ws, "Sales by Day", start, end, 3, 2), "G4")
    ws.freeze_panes = "A5"


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


def build_labor(wb):
    ws = wb.create_sheet("Labor & Prime"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 2])
    luxe_header(ws, "E", "👥  LABOR & PRIME COST",
                "Daily labor vs sales — plus the prime cost (beverage + labor) that runs your cafe.")
    table_headers(ws, 4, ["Day", "Labor $", "Sales", "Labor %"], start_col=2)
    start = L0
    for i, (day, labor) in enumerate(LABOR):
        r = start + i
        sales = WEEKLY[i][1]
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cl = ws.cell(row=r, column=3, value=labor); cl.style = "input"; cl.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LABOR) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    cl = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cl.style = "td"; cl.font = Font(bold=True, color=PRIMARY); cl.fill = fill(SURFACE); cl.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cp = ws.cell(row=tot, column=5, value=f"=IFERROR(C{tot}/D{tot},0)"); cp.style = "td"; cp.font = Font(bold=True, color=DANGER); cp.fill = fill(SURFACE); cp.number_format = "0%"
    cell_name(wb, "LaborTotal", "Labor & Prime", f"$C${tot}")
    cell_name(wb, "LaborPct", "Labor & Prime", f"$E${tot}")
    pr = tot + 2
    merge_set(ws, f"B{pr}:D{pr}", "PRIME COST  (beverage % + labor %)", "section")
    c = ws.cell(row=pr, column=5, value="=BevCostPct+LaborPct"); c.style = "field_value"; c.number_format = "0%"; c.font = Font(bold=True, size=13, color=PRIMARY); c.fill = fill(MINT_BG)
    cell_name(wb, "PrimeCost", "Labor & Prime", f"$E${pr}")
    ws.freeze_panes = "A5"


def build_usage(wb):
    ws, start, end = build_log(
        wb, "Bean & Milk", "🥛", "BEAN & MILK USAGE",
        "Weekly usage & cost of your biggest inputs — where beverage cost really comes from.",
        ["Item", "Qty", "Unit", "Weekly Cost"],
        USAGE, [2, 26, 10, 12, 14, 2], text_left={2}, ints={3}, money2={5}, reserved=20,
        validations=[("D", "UnitList")], start_col=2)
    nrange(wb, "UsageCost", "Bean & Milk", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="WEEKLY INPUT COST").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(UsageCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_inventory(wb):
    ws = wb.create_sheet("Inventory & Par"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 12, 10, 14, 2])
    luxe_header(ws, "F", "📦  INVENTORY & PAR",
                "Par vs on hand — see what to order before you 86 the oat milk mid-rush.")
    table_headers(ws, 4, ["Item", "Par", "On Hand", "Unit", "To Order"], start_col=2)
    start = L0
    for i, (item, par, oh, unit) in enumerate(INVENTORY):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=par); cp.style = "input"; cp.number_format = "#,##0"
        co = ws.cell(row=r, column=4, value=oh); co.style = "input"; co.number_format = "#,##0"
        ws.cell(row=r, column=5, value=unit).style = "td"
        cb = ws.cell(row=r, column=6, value=f"=MAX(C{r}-D{r},0)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.number_format = "#,##0"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVENTORY) - 1
    for c in range(2, 7):
        add_dv(ws, f"E{start}:E{end}", "UnitList")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=[f"C{start}*0.5"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Dumped shots, spoiled milk & stale pastries — every dollar of waste, logged.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="WEEKLY WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_ordering(wb):
    ws, start, end = build_log(
        wb, "Ordering", "🚚", "ORDERING & SUPPLIERS",
        "Your standing order — par quantities & cost, by supplier.",
        ["Item", "Supplier", "Par Order", "Cost"],
        ORDERING, [2, 26, 20, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)
    nrange(wb, "OrderCost", "Ordering", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ORDER TOTAL").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(OrderCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_cash(wb):
    ws = wb.create_sheet("Cash & Tips"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 14, 2])
    luxe_header(ws, "F", "💵  CASH & TIPS",
                "Reconcile each day — cash, card & tips, so the drawer always balances.")
    table_headers(ws, 4, ["Day", "Cash", "Card", "Tips", "Total"], start_col=2)
    start = L0
    for i, (day, cash, card, tips) in enumerate(CASHTIPS):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cash); cc.style = "input"; cc.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=4, value=card); cd.style = "input"; cd.number_format = '"$"#,##0'
        ct = ws.cell(row=r, column=5, value=tips); ct.style = "input"; ct.number_format = '"$"#,##0'
        cto = ws.cell(row=r, column=6, value=f"=C{r}+D{r}+E{r}"); cto.style = "td"; cto.font = Font(bold=True, color=PRIMARY); cto.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CASHTIPS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    for col in (3, 4, 5, 6):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "TipsTotal", "Cash & Tips", f"$E${tot}")
    ws.freeze_panes = "A5"


def build_regulars(wb):
    ws = wb.create_sheet("Regulars"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 24, 12, 14, 2])
    luxe_header(ws, "E", "⭐  REGULARS",
                "Your best customers & their usual — regulars are the backbone of a cafe.")
    table_headers(ws, 4, ["Regular", "Usual Order", "Visits/wk", "Spend/wk"], start_col=2)
    start = L0
    for i, (name, fav, visits, spend) in enumerate(REGULARS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=fav).style = "td_left"
        cv = ws.cell(row=r, column=4, value=visits); cv.style = "input"; cv.number_format = "#,##0"
        cs = ws.cell(row=r, column=5, value=spend); cs.style = "input"; cs.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(REGULARS) - 1
    nrange(wb, "RegSpend", "Regulars", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="REGULARS / WK").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(RegSpend)*4"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).value = "=SUM(RegSpend)"
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ☕  CAFE & COFFEE SHOP COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Cup cost, beverage & labor %, prime cost & a Cafe Score — your whole shop, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MENU ITEMS", "=COUNTA(MenuItem)", "num"),
        ("AVG CUP COST", "=AVERAGE(MenuCost)", "money2"),
        ("AVG PRICE", "=AVERAGE(MenuPrice)", "money2"),
        ("AVG MARGIN", "=AVERAGE(MenuMargin)", "money2"),
        ("BEV COST", "=BevCostPct", "pct"),
        ("LABOR COST", "=LaborPct", "pct"),
    ]
    row2 = [
        ("DAILY SALES", "=DayTotal", "money"),
        ("TRANSACTIONS", "=TxnTotal", "num"),
        ("AVG TICKET", "=IFERROR(DayTotal/TxnTotal,0)", "money2"),
        ("TOP DAYPART", "=INDEX(DaypartName,MATCH(MAX(DaypartSales),DaypartSales,0))", "text"),
        ("PRIME COST", "=PrimeCost", "pct"),
        ("CAFE SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "CAFE HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SALES BY DAYPART", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Beverage cost on target", "=IFERROR(MIN(TargetBev/BevCostPct,1),0)"),
        ("Labor on target", "=IFERROR(MIN(TargetLabor/LaborPct,1),0)"),
        ("Prime cost healthy", "=IFERROR(MIN(0.62/PrimeCost,1),0)"),
        ("Avg ticket vs goal", "=IFERROR(MIN((DayTotal/TxnTotal)/TicketGoal,1),0)"),
        ("Margin per cup", "=IFERROR(MIN(AVERAGE(MenuMargin)/4,1),0)"),
        ("Low waste", "=IFERROR(1-MIN((WasteTotal/WeekSalesTotal)/0.06,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    dp = wb["Daypart Sales"]
    d = DoughnutChart(); d.title = "Sales by Daypart"; d.height = 7.4; d.width = 8.6
    d.add_data(Reference(dp, min_col=4, min_row=5, max_row=7), titles_from_data=False)
    d.set_categories(Reference(dp, min_col=2, min_row=5, max_row=7)); d.dataLabels = no_labels()
    ws.add_chart(d, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Cafe & Coffee Shop Command Center™ — cost every cup & watch your prime cost.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_cupcost(wb); build_menu(wb)
    build_daypart(wb); build_weekly(wb); build_labor(wb); build_usage(wb)
    build_inventory(wb); build_waste(wb); build_ordering(wb); build_cash(wb)
    build_regulars(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Cup Cost", "Menu Board", "Daypart Sales", "Weekly Sales",
             "Labor & Prime", "Bean & Milk", "Inventory & Par", "Waste Log", "Ordering",
             "Cash & Tips", "Regulars", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Cafe_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
