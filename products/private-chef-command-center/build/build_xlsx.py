"""Build Personal & Private Chef Command Center™ — The Private-Chef Operating System.

14 tabs · a premium private / personal-chef operating system in Google Sheets & Excel.
Dashboard, a per-event pricing engine (cost a dinner per guest → your take-home & real
hourly rate), a service menu, dish costing, a client roster & monthly revenue, a
booking calendar, grocery & kitchen-kit lists, a mileage log, waste, income & expenses
and a monthly summary — one dashboard. Cost every dinner, and pay yourself properly.

Run: python3 build_xlsx.py   ->  ../Private_Chef_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
BASIS = ["per guest", "flat rate", "per hour"]
SERVICE = ["Private dinner", "Dinner party", "Meal prep", "Date night", "Event", "Consult"]
STATUS = ["Booked", "Enquiry", "Confirmed", "Done", "Cancelled"]

TARGET_FC = 0.30
MARGIN_GOAL = 0.55
CLIENT_GOAL = 5
PROFIT_GOAL = 3000
WASTE_LIMIT = 0.04

# Event pricing engine — flagship Private Dinner (4 courses): (component, cost/guest)
EVENT = [
    ("Appetizer", 4.00), ("Main protein", 10.00), ("Sides & vegetables", 3.50),
    ("Dessert", 2.50), ("Pantry, oil & garnish", 2.00),
]
GUESTS = 6
PRICE_PER_GUEST = 90
TRAVEL = 30
HOURS = 7

# Service menu — the chef's packages: (package, guests, basis, price)
MENU = [
    ("Intimate Dinner (3-course)", 2, "per guest", 110),
    ("Private Dinner (4-course)", 6, "per guest", 90),
    ("Dinner Party (4-course)", 10, "per guest", 85),
    ("Weekly Meal Prep", 1, "flat rate", 280),
    ("Date Night for Two", 2, "flat rate", 300),
]

# Dish costing — signature dishes: (dish, course, cost/guest, menu price/guest)
DISHES = [
    ("Seared Scallops", "Appetizer", 5.20, 18), ("Burrata & Heirloom", "Appetizer", 3.80, 14),
    ("Filet Mignon", "Main", 12.50, 42), ("Herb-Roast Chicken", "Main", 7.20, 30),
    ("Pan-Seared Salmon", "Main", 9.40, 34), ("Truffle Risotto", "Main", 5.60, 26),
    ("Chocolate Torte", "Dessert", 2.90, 12), ("Lemon Tart", "Dessert", 2.60, 11),
]

# Clients roster (monthly): (client, service, events/mo, price/event)
CLIENTS = [
    ("Anderson Family", "Private dinner", 4, 540), ("Bello", "Private dinner", 2, 540),
    ("Chen", "Meal prep", 4, 280), ("Ruiz (party)", "Dinner party", 1, 990),
    ("Okoro", "Date night", 2, 300),
]

# Booking calendar (upcoming): (date, client, service, guests, status)
BOOKINGS = [
    ("Fri", "Anderson Family", "Private dinner", 6, "Confirmed"),
    ("Sat", "Ruiz (party)", "Dinner party", 10, "Confirmed"),
    ("Sun", "Chen", "Meal prep", 1, "Booked"),
    ("Tue", "Okoro", "Date night", 2, "Booked"),
    ("Thu", "Bello", "Private dinner", 6, "Enquiry"),
]

# Grocery / shopping list: (item, vendor, cost)
GROCERIES = [
    ("Proteins (butcher)", "Local butcher", 620.00), ("Produce & herbs", "Farmers market", 340.00),
    ("Dairy & eggs", "Wholesaler", 190.00), ("Pantry & dry goods", "Restaurant depot", 260.00),
    ("Specialty & garnish", "Specialty grocer", 180.00), ("Beverages & extras", "Wholesaler", 110.00),
]

# Kitchen kit / equipment checklist: (item, have, notes)
KIT = [
    ("Chef knives & steel", "Yes", "Sharpen weekly"), ("Portable induction burner", "Yes", "Backup heat"),
    ("Cambro / transport bins", "Yes", "Meal-prep drops"), ("Immersion circulator", "Yes", "Sous-vide mains"),
    ("Sheet pans & liners", "Yes", "Restock monthly"), ("Serviceware kit", "No", "Rent for parties"),
]

# Mileage & travel log: (date, client, miles, cost)
MILEAGE = [
    ("Fri", "Anderson Family", 24, 16.08), ("Sat", "Ruiz (party)", 38, 25.46),
    ("Sun", "Chen", 12, 8.04), ("Tue", "Okoro", 18, 12.06), ("Thu", "Bello", 24, 16.08),
]

# Waste log: (item, reason, cost) — monthly total $143
WASTE = [
    ("Over-shopping", "Guest count dropped", 60.00), ("Spoiled produce", "Menu change", 45.00),
    ("Recipe testing", "New dish trial", 38.00),
]

# Income & expenses (monthly): (item, amount)
LEDGER = [
    ("Groceries & food", 1700, "E"), ("Mileage & travel", 360, "E"), ("Insurance & licenses", 180, "E"),
    ("Equipment & supplies", 220, "E"), ("Marketing", 150, "E"),
]

# Monthly summary: (month, revenue, expenses)
MONTHS = [("Jul", 4800, 2300), ("Aug", 5200, 2450), ("Sep", 5950, 2610),
          ("Oct", 6400, 2760), ("Nov", 7100, 2980), ("Dec", 8200, 3300)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Business name", "Chef's Table Co.", None, "Business"),
        ("Chef", "Rowan", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Client goal", CLIENT_GOAL, "#,##0", "ClientGoal"),
        ("Monthly profit goal", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Waste limit %", WASTE_LIMIT, "0%", "WasteLimit"),
        ("Mileage rate ($/mi)", 0.67, '"$"#,##0.00', "MileRate"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Basis", BASIS, "BasisList"), ("F", "Service", SERVICE, "ServiceList"),
             ("G", "Status", STATUS, "StatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🍽  PERSONAL & PRIVATE CHEF COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every dinner, and pay yourself properly.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE PRIVATE-CHEF BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A private chef lives or dies on two numbers: what a dinner costs to put on the table, and what you "
                      "actually take home per hour after food and travel. This makes both visible: a per-event pricing "
                      "engine (cost a dinner per guest, set your per-guest price, and see your take-home and real hourly "
                      "rate), a service menu and dish costing, and a client roster that turns bookings into monthly "
                      "revenue. Manage your booking calendar, grocery and kitchen-kit lists, mileage, waste and income & "
                      "expenses — all in ONE premium Google Sheets & Excel system built for personal and private chefs.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your target food-cost % & client goal.",
             "2.  Cost a dinner in Event Pricing — per guest, plus travel & hours.",
             "3.  Set your per-guest price; your take-home & hourly rate calculate live.",
             "4.  Build your Service Menu & cost your signature dishes.",
             "5.  Add Clients & bookings — monthly revenue rolls up.",
             "6.  Check the Dashboard: revenue, profit & a Chef Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional private chef (Chef's Table Co.) is included so you can see how it all "
               "connects — just type over it with your own dinners and clients. Cost per dinner and your real hourly "
               "take-home are the two numbers that decide whether private cheffing pays, and they roll into a live Chef "
               "Score. Twelve matching printable pages (event quote, dish cost card, prep list, shopping list, run sheet "
               "& more) are included. This is a business tool, not financial, legal or food-safety advice — follow your "
               "local food-handling rules and confirm figures with your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Cost the dinner, price per guest, and pay yourself for your craft.", "section_gold")


# ===========================================================================
def build_eventpricing(wb):
    ws = wb.create_sheet("Event Pricing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "🍲  EVENT PRICING",
                "Cost a dinner per guest, set your per-guest price, and see your take-home and real hourly rate.")
    ws.cell(row=5, column=2, value="EVENT").style = "section_gold"
    ws.cell(row=5, column=3, value="Private Dinner (4-course)").font = Font(bold=True, color=PRIMARY)
    ws.cell(row=6, column=2, value="Guests").style = "field_label"
    cg = ws.cell(row=6, column=3, value=GUESTS); cg.style = "input"; cg.number_format = "#,##0"
    cell_name(wb, "Guests", "Event Pricing", "$C$6")
    table_headers(ws, 8, ["Course / component", "Cost / guest"], start_col=2)
    start = 9
    for i, (comp, cost) in enumerate(EVENT):
        r = start + i
        ws.cell(row=r, column=2, value=comp).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EVENT) - 1; fpg = end + 1
    ws.cell(row=fpg, column=2, value="FOOD / GUEST").style = "th"
    cf = ws.cell(row=fpg, column=3, value=f"=SUM(C{start}:C{end})"); cf.style = "td"; cf.font = Font(bold=True, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = '"$"#,##0.00'
    cell_name(wb, "FoodPerGuest", "Event Pricing", f"$C${fpg}")
    # pricing block
    pr = fpg + 2
    ws.cell(row=pr, column=2, value="YOUR PRICE / GUEST").style = "section_gold"
    cp = ws.cell(row=pr, column=3, value=PRICE_PER_GUEST); cp.style = "input"; cp.number_format = '"$"#,##0'
    cell_name(wb, "PricePerGuest", "Event Pricing", f"$C${pr}")
    rows = [
        ("Event food cost", "=FoodPerGuest*Guests", "EventFood", '"$"#,##0.00', None),
        ("Event price (per guest × guests)", "=PricePerGuest*Guests", "EventPrice", '"$"#,##0', MINT_BG),
        ("− Travel / mileage", TRAVEL, "Travel", '"$"#,##0', None),
        ("Hours worked (shop, prep, cook, serve)", HOURS, "Hours", "0.0", None),
    ]
    rr = pr + 1
    for i, (lab, val, nm, fmt, bg) in enumerate(rows):
        r = rr + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input" if not str(val).startswith("=") else "field_value"
        c.number_format = fmt
        if bg:
            c.fill = fill(bg)
        cell_name(wb, nm, "Event Pricing", f"$C${r}")
    take_r = rr + len(rows)
    ws.cell(row=take_r, column=2, value="= YOUR TAKE-HOME (price − food − travel)").style = "th"
    ct = ws.cell(row=take_r, column=3, value="=EventPrice-EventFood-Travel"); ct.style = "td"; ct.font = Font(bold=True, size=12, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0'
    cell_name(wb, "ChefTake", "Event Pricing", f"$C${take_r}")
    hr = take_r + 1
    ws.cell(row=hr, column=2, value="= YOUR REAL HOURLY RATE").style = "th"
    ch = ws.cell(row=hr, column=3, value="=IFERROR(ChefTake/Hours,0)"); ch.style = "td"; ch.font = Font(bold=True, size=12, color=PRIMARY); ch.fill = fill(MINT_BG); ch.number_format = '"$"#,##0.00'
    cell_name(wb, "ChefHourly", "Event Pricing", f"$C${hr}")
    fcr = hr + 2
    ws.cell(row=fcr, column=2, value="Food cost % (of event price)").style = "field_label"
    cfc = ws.cell(row=fcr, column=3, value="=IFERROR(EventFood/EventPrice,0)"); cfc.style = "field_value"; cfc.number_format = "0%"; cfc.fill = fill(MINT_BG)
    cell_name(wb, "FoodCostPct", "Event Pricing", f"$C${fcr}")
    ws.cell(row=fcr + 1, column=2, value="Margin % (take-home ÷ price)").style = "field_label"
    cm = ws.cell(row=fcr + 1, column=3, value="=IFERROR(ChefTake/EventPrice,0)"); cm.style = "field_value"; cm.number_format = "0%"; cm.fill = fill(MINT_BG)
    cell_name(wb, "MarginPct", "Event Pricing", f"$C${fcr+1}")
    ws.cell(row=fcr + 3, column=2, value="Copy this build for every event — swap the courses & guest count.").style = "section"


def build_menu(wb):
    ws = wb.create_sheet("Service Menu"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 12, 16, 14, 2])
    luxe_header(ws, "E", "📖  SERVICE MENU",
                "Your packages — guests, how they're priced, and the rate. Keep it simple to quote.")
    table_headers(ws, 4, ["Package", "Guests", "Basis", "Price"], start_col=2)
    start = L0
    for i, (pkg, guests, basis, price) in enumerate(MENU):
        r = start + i
        ws.cell(row=r, column=2, value=pkg).style = "td_left"
        cg = ws.cell(row=r, column=3, value=guests); cg.style = "input"; cg.number_format = "#,##0"
        ws.cell(row=r, column=4, value=basis).style = "td"
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"D{r}", "BasisList")
    end = start + len(MENU) - 1
    nrange(wb, "MenuName", "Service Menu", "B", start, end)
    nrange(wb, "MenuPrice", "Service Menu", "E", start, end)
    ws.freeze_panes = "A5"


def build_dishes(wb):
    ws = wb.create_sheet("Dish Costing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 14, 12, 2])
    luxe_header(ws, "F", "🍳  DISH COSTING",
                "Cost each signature dish per guest — cost, menu price, margin & food-cost %.")
    table_headers(ws, 4, ["Dish", "Course", "Cost/guest", "Price/guest", "Food %"], start_col=2)
    start = L0
    for i, (dish, course, cost, price) in enumerate(DISHES):
        r = start + i
        ws.cell(row=r, column=2, value=dish).style = "td_left"
        ws.cell(row=r, column=3, value=course).style = "td"
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cf = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/E{r},0)"); cf.style = "td"; cf.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DISHES) - 1
    ws.conditional_formatting.add(f"F{start}:F{end}",
        ColorScaleRule(start_type="num", start_value=0.20, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.35, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.50, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


def build_clients(wb):
    ws = wb.create_sheet("Clients"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 18, 14, 14, 16, 2])
    luxe_header(ws, "F", "👥  CLIENTS",
                "Your roster — service, events a month and the rate, so monthly revenue rolls up.")
    table_headers(ws, 4, ["Client", "Service", "Events/mo", "Price/event", "Monthly Rev"], start_col=2)
    start = L0
    for i, (client, service, events, price) in enumerate(CLIENTS):
        r = start + i
        ws.cell(row=r, column=2, value=client).style = "td_left"
        ws.cell(row=r, column=3, value=service).style = "td_left"
        ce = ws.cell(row=r, column=4, value=events); ce.style = "input"; ce.number_format = "#,##0"
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        crv = ws.cell(row=r, column=6, value=f"=D{r}*E{r}"); crv.style = "td"; crv.font = Font(bold=True, color=PRIMARY); crv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "ServiceList")
    end = start + len(CLIENTS) - 1
    nrange(wb, "ClientName", "Clients", "B", start, end)
    nrange(wb, "ClientEvents", "Clients", "D", start, end)
    nrange(wb, "ClientRev", "Clients", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTALS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ce = ws.cell(row=tot, column=4, value="=SUM(ClientEvents)"); ce.style = "td"; ce.font = Font(bold=True, color=PRIMARY); ce.fill = fill(SURFACE); ce.number_format = "#,##0"
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    crv = ws.cell(row=tot, column=6, value="=SUM(ClientRev)"); crv.style = "td"; crv.font = Font(bold=True, color=PRIMARY); crv.fill = fill(MINT_BG); crv.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyRev", "Clients", f"$F${tot}")
    cell_name(wb, "EventsMonth", "Clients", f"$D${tot}")
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Active clients").style = "field_label"
    c = ws.cell(row=sr, column=6, value="=COUNTA(ClientName)"); c.style = "field_value"; c.number_format = "#,##0"; c.fill = fill(MINT_BG)
    cell_name(wb, "TotalClients", "Clients", f"$F${sr}")
    ws.cell(row=sr + 1, column=2, value="Top client (by revenue)").style = "field_label"
    c2 = ws.cell(row=sr + 1, column=6, value="=IFERROR(INDEX(ClientName,MATCH(MAX(ClientRev),ClientRev,0)),\"\")"); c2.style = "field_value"; c2.fill = fill(MINT_BG)
    cell_name(wb, "TopClient", "Clients", f"$F${sr+1}")
    ws.freeze_panes = "A5"


def build_calendar(wb):
    ws = wb.create_sheet("Booking Calendar"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 24, 18, 12, 16, 2])
    luxe_header(ws, "F", "📅  BOOKING CALENDAR",
                "This week's bookings — who, what service, how many guests & the status.")
    table_headers(ws, 4, ["Day", "Client", "Service", "Guests", "Status"], start_col=2)
    start = L0
    for i, (day, client, service, guests, status) in enumerate(BOOKINGS):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        ws.cell(row=r, column=3, value=client).style = "td_left"
        ws.cell(row=r, column=4, value=service).style = "td_left"
        cgt = ws.cell(row=r, column=5, value=guests); cgt.style = "input"; cgt.number_format = "#,##0"
        ws.cell(row=r, column=6, value=status).style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"D{r}", "ServiceList")
        add_dv(ws, f"F{r}", "StatusList")
    end = start + len(BOOKINGS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="THIS WEEK").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cgt = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cgt.style = "td"; cgt.font = Font(bold=True, color=PRIMARY); cgt.fill = fill(SURFACE); cgt.number_format = "#,##0"
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


def build_groceries(wb):
    ws, start, end = build_log(
        wb, "Grocery List", "🛒", "GROCERY LIST",
        "Your shopping by vendor — what the food actually costs, so pricing stays honest.",
        ["Item / category", "Vendor", "Cost"],
        GROCERIES, [2, 26, 22, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "GroceryCost", "Grocery List", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL GROCERIES").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(GroceryCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_kit(wb):
    ws, start, end = build_log(
        wb, "Kitchen Kit", "🔪", "KITCHEN KIT",
        "Your travelling kitchen — what you own, what you rent, and what to restock.",
        ["Equipment", "Have?", "Notes"],
        KIT, [2, 28, 12, 26, 2], text_left={2, 4}, reserved=24, start_col=2,
        validations=[("C", "YesNoList")])


def build_mileage(wb):
    ws, start, end = build_log(
        wb, "Mileage & Travel", "🚗", "MILEAGE & TRAVEL",
        "Every drive to a client — miles and cost, for pricing and for taxes.",
        ["Day", "Client", "Miles", "Cost"],
        MILEAGE, [2, 12, 24, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)
    nrange(wb, "MileCost", "Mileage & Travel", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(MileCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Over-shopping, spoilage & recipe testing — the leaks that eat your margin.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_ledger(wb):
    ws = wb.create_sheet("Income & Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  INCOME & EXPENSES",
                "Monthly books — client revenue in, costs out, and the profit that's really yours.")
    ws.cell(row=5, column=2, value="INCOME").style = "section_gold"
    ws.cell(row=6, column=2, value="Monthly client revenue").style = "td_left"
    ci = ws.cell(row=6, column=3, value="=MonthlyRev"); ci.style = "td"; ci.font = Font(bold=True, color=PRIMARY); ci.number_format = '"$"#,##0'; ci.fill = fill(MINT_BG)
    cell_name(wb, "TotalIncome", "Income & Expenses", "$C$6")
    ws.cell(row=8, column=2, value="EXPENSES").style = "section_gold"
    table_headers(ws, 9, ["Expense", "Amount"], start_col=2)
    start = 10
    for i, (item, amt, _t) in enumerate(LEDGER):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LEDGER) - 1
    nrange(wb, "ExpAmt", "Income & Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Income & Expenses", f"$C${tot}")
    nr = tot + 2
    ws.cell(row=nr, column=2, value="= MONTHLY PROFIT").style = "th"
    cn = ws.cell(row=nr, column=3, value="=TotalIncome-ExpTotal"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetProfit", "Income & Expenses", f"$C${nr}")


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 14, 2])
    luxe_header(ws, "E", "📈  MONTHLY SUMMARY",
                "Revenue, expenses & profit by month — watch the season build.")
    table_headers(ws, 4, ["Month", "Revenue", "Expenses", "Profit"], start_col=2)
    start = L0
    for i, (m, rev, exp) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ci = ws.cell(row=r, column=3, value=rev); ci.style = "input"; ci.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=4, value=exp); ce.style = "input"; ce.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthRev", "Monthly Summary", "C", start, end)
    ws.add_chart(_barchart(ws, "Revenue by Month", start, end, 3, 2), "G4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🍽  PERSONAL & PRIVATE CHEF COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Cost per dinner, your real hourly rate, monthly revenue & a Chef Score — at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("FOOD / GUEST", "=FoodPerGuest", "money2"),
        ("PRICE / GUEST", "=PricePerGuest", "money"),
        ("FOOD COST", "=FoodCostPct", "pct"),
        ("MARGIN / EVENT", "=MarginPct", "pct"),
        ("YOUR HOURLY", "=ChefHourly", "money2"),
        ("TOP CLIENT", "=TopClient", "text"),
    ]
    row2 = [
        ("MONTHLY REVENUE", "=MonthlyRev", "money"),
        ("ACTIVE CLIENTS", "=TotalClients", "num"),
        ("EVENTS / MONTH", "=EventsMonth", "num"),
        ("MONTHLY PROFIT", "=NetProfit", "money"),
        ("WASTE %", "=IFERROR(WasteTotal/MonthlyRev,0)", "pct"),
        ("CHEF SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "CHEF HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Food cost on target", "=IFERROR(MIN(TargetFC/FoodCostPct,1),0)"),
        ("Margin healthy", "=IFERROR(MIN(MarginPct/MarginGoal,1),0)"),
        ("Menu priced", "=IFERROR(COUNTIF(MenuPrice,\">0\")/COUNTA(MenuName),0)"),
        ("Clients vs goal", "=IFERROR(MIN(TotalClients/ClientGoal,1),0)"),
        ("Profitable", "=IFERROR(MIN(NetProfit/ProfitGoal,1),0)"),
        ("Waste low", "=IFERROR(1-MIN((WasteTotal/MonthlyRev)/WasteLimit,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Personal & Private Chef Command Center™ — cost every dinner, and pay yourself properly.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_eventpricing(wb); build_menu(wb)
    build_dishes(wb); build_clients(wb); build_calendar(wb); build_groceries(wb)
    build_kit(wb); build_mileage(wb); build_waste(wb); build_ledger(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Event Pricing", "Service Menu", "Dish Costing", "Clients",
             "Booking Calendar", "Grocery List", "Kitchen Kit", "Mileage & Travel", "Waste Log",
             "Income & Expenses", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Private_Chef_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
