"""Build Lyft Driver Command Center™ — The Ultimate Rideshare Earnings & Budget System.

19 sheets + Welcome · a premium rideshare-driver operating system in Excel & Sheets.
Shift log, earnings, mileage, fuel, vehicle & maintenance, business expenses, a
household budget, a tax center (mileage deduction), savings & goals, bonuses,
hot zones, ratings and an Analytics driver-health score — one dashboard.
Works for Lyft, Uber or any rideshare / delivery gig.

Run: python3 build_xlsx.py   ->  ../Lyft_Driver_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

PLATFORMS = ["Lyft", "Uber", "Both", "Delivery", "Other"]
EXP_CATS = ["Fuel", "Car Payment", "Insurance", "Maintenance & Repairs", "Phone & Data",
            "Tolls", "Car Wash & Cleaning", "Supplies", "Miscellaneous"]
BUDGET_CATS = ["Housing", "Utilities", "Groceries", "Health", "Debt", "Savings",
               "Personal", "Insurance", "Other"]
MAINT_TYPES = ["Oil Change", "Tires", "Brakes", "Inspection", "Rotation", "Battery", "Repair", "Other"]
GOAL_CATS = ["Earnings", "Hourly", "Trips", "Savings", "Miles", "Tax", "Debt"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your goals & vehicle details once — every dashboard follows.")
    merge_set(ws, "B5:C5", "DRIVER INPUTS", "section")
    controls = [
        ("Driver Name", "Jordan Rivera", None, "DriverName"),
        ("Vehicle", "2021 Toyota Camry", None, "VehicleName"),
        ("Home City", "Austin, TX", None, "HomeCity"),
        ("Monthly Net Goal", 3000, '"$"#,##0', "NetGoal"),
        ("Target Net $ / Hour", 20, '"$"#,##0.00', "HourlyTarget"),
        ("Monthly Trips Goal", 300, "#,##0", "TripsGoal"),
        ("Monthly Shifts Goal", 16, "0", "ShiftGoal"),
        ("Emergency Fund Goal", 6000, '"$"#,##0', "SavingsGoal"),
        ("Tax Reserve Goal (mo)", 1800, '"$"#,##0', "TaxReserveGoal"),
        ("IRS Mileage Rate", 0.70, '"$"#,##0.00', "MileageRate"),
        ("Tax Set-Aside %", 0.25, "0%", "TaxRate"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Platform", PLATFORMS, "PlatformList"), ("F", "Expense Category", EXP_CATS, "ExpCatList"),
             ("G", "Budget Category", BUDGET_CATS, "BudgetCatList"), ("H", "Maintenance", MAINT_TYPES, "MaintList"),
             ("I", "Goal Category", GOAL_CATS, "GoalCatList"), ("J", "Priority", PRIORITIES, "PriorityList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    ws.cell(row=17, column=5, value="Yes / No").style = "th"
    for ri, v in enumerate(YESNO):
        ws.cell(row=18 + ri, column=5, value=v).style = "td_left"
    wb.defined_names["YesNoList"] = DefinedName("YesNoList", attr_text="Settings!$E$18:$E$19")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🚗  LYFT DRIVER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate rideshare earnings & budget system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "RUN YOUR WHOLE DRIVING BUSINESS FROM ONE FILE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Every shift, mile and dollar in ONE premium Excel & Google Sheets system. Log your shifts and "
                      "the Command Center instantly shows your real take-home — net earnings, true $/hour and $/mile "
                      "after fuel and expenses, your mileage tax deduction, your budget and your savings. Stop guessing "
                      "whether a day was worth it: this is the difference between driving for cash and running a "
                      "profitable business. Built for Lyft — works for Uber, delivery or any gig.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your vehicle, goals & the IRS mileage rate.",
             "2.  Log each shift in the Shift Log — hours, trips, fares, tips, miles & fuel.",
             "3.  Track Mileage, Fuel & Vehicle Maintenance as you go.",
             "4.  Set your Business Expenses & the Monthly Budget — net income flows in live.",
             "5.  Use the Tax Center to bank your mileage deduction & set-aside.",
             "6.  Watch the Dashboard track net $/hour, savings & a Driver Health Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (Jordan, a full-time Austin driver doing ~$4,200/mo across 16 shifts, ~2,870 "
               "miles and 316 trips) is included so you can see how everything connects — just type over it with your "
               "own numbers. Net earnings, true $/hour and $/mile, the mileage tax deduction, budget leftover and the "
               "Driver Health Score all update automatically. Every sheet is print-friendly and works in Excel and "
               "Google Sheets, on desktop and phone — perfect for logging from the driver's seat.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, every mile counted — let's grow your take-home.", "section_gold")


# ===========================================================================
# Driver Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Driver Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "🪪  DRIVER PROFILE", "Your driving business, defined — the details every number flows from.")
    blocks = [
        ("THE DRIVER", [("Driver Name", "=DriverName"), ("Home City", "=HomeCity"),
                        ("Platforms", "Lyft + Uber"), ("Driving Since", "2021"),
                        ("Status", "Full-time"), ("Weekly Target", "5 shifts")]),
        ("THE VEHICLE", [("Vehicle", "=VehicleName"), ("MPG (avg)", "32 city/hwy"),
                         ("Odometer", "68,400 mi"), ("Plate / Unit", "ABC-1234"),
                         ("Insurance", "Rideshare rider"), ("Est. Value", "$21,500")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "RATINGS & STANDING", "section_gold"); ws.row_dimensions[15].height = 22
    stand = [("Lyft Rating", "4.98 ★"), ("Uber Rating", "4.96 ★"), ("Acceptance", "84%"),
             ("Cancellation", "2%"), ("Lifetime Trips", "9,240"), ("Rewards Tier", "Platinum")]
    for i, (p, h) in enumerate(stand):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# Shift Log — core earnings engine
# ===========================================================================
SHIFTS = [
    (23, "Lyft", 8.5, 21, 214, 44, 20, 188, 34),
    (22, "Lyft", 9.0, 23, 232, 51, 35, 205, 37),
    (20, "Uber", 7.5, 18, 176, 33, 15, 158, 29),
    (19, "Both", 8.0, 20, 205, 42, 25, 182, 33),
    (17, "Lyft", 6.5, 15, 152, 28, 10, 140, 26),
    (16, "Lyft", 9.5, 24, 245, 55, 40, 214, 39),
    (15, "Uber", 8.0, 19, 192, 36, 18, 176, 32),
    (13, "Both", 8.5, 22, 220, 47, 30, 198, 35),
    (12, "Lyft", 7.0, 17, 168, 31, 12, 152, 28),
    (10, "Lyft", 9.0, 23, 236, 52, 38, 208, 38),
    (9, "Uber", 7.5, 18, 180, 34, 16, 164, 30),
    (8, "Both", 8.0, 20, 208, 43, 22, 186, 34),
    (6, "Lyft", 6.0, 14, 142, 26, 8, 132, 24),
    (5, "Lyft", 9.0, 22, 228, 49, 33, 200, 36),
    (3, "Uber", 8.5, 21, 214, 40, 20, 190, 34),
    (2, "Both", 8.0, 19, 196, 41, 24, 178, 33),
]


def build_shiftlog(wb):
    ws = wb.create_sheet("Shift Log"); ws.sheet_view.showGridLines = False
    headers = ["Date", "Platform", "Hours", "Trips", "Fares", "Tips", "Bonus/Surge",
               "Earnings", "Miles", "Fuel $"]
    set_widths(ws, [13, 12, 9, 8, 11, 10, 13, 12, 10, 10])
    luxe_header(ws, "J", "🚗  SHIFT LOG",
                "Log every shift — earnings, miles & fuel roll straight into your dashboard.")
    table_headers(ws, 4, headers)
    start = L0
    reserved = 60
    end = start + reserved - 1
    for i, (off, plat, hrs, trips, fares, tips, bonus, miles, fuel) in enumerate(SHIFTS):
        r = start + i
        ws.cell(row=r, column=1, value=dminus(off))
        ws.cell(row=r, column=2, value=plat)
        ws.cell(row=r, column=3, value=hrs)
        ws.cell(row=r, column=4, value=trips)
        ws.cell(row=r, column=5, value=fares)
        ws.cell(row=r, column=6, value=tips)
        ws.cell(row=r, column=7, value=bonus)
        ws.cell(row=r, column=8, value=f"=E{r}+F{r}+G{r}")
        ws.cell(row=r, column=9, value=miles)
        ws.cell(row=r, column=10, value=fuel)
    style_rows(ws, start, end, len(headers), text_left=set(), dates={1},
               money={5, 6, 7, 8, 10}, ints={4, 9}, dec={3})
    add_dv(ws, f"B{start}:B{end}", "PlatformList")
    ws.freeze_panes = "A5"
    # totals row
    trow = end + 1
    ws.cell(row=trow, column=1, value="TOTAL").style = "th"
    ws.cell(row=trow, column=2).style = "th"
    fmts = {3: "0.0", 4: "#,##0", 5: '"$"#,##0', 6: '"$"#,##0', 7: '"$"#,##0',
            8: '"$"#,##0', 9: "#,##0", 10: '"$"#,##0'}
    for col in range(3, 11):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmts[col]
    # named ranges
    nrange(wb, "ShiftDate", "Shift Log", "A", start, end)
    nrange(wb, "ShiftPlatform", "Shift Log", "B", start, end)
    nrange(wb, "ShiftHours", "Shift Log", "C", start, end)
    nrange(wb, "ShiftTrips", "Shift Log", "D", start, end)
    nrange(wb, "ShiftFares", "Shift Log", "E", start, end)
    nrange(wb, "ShiftTips", "Shift Log", "F", start, end)
    nrange(wb, "ShiftBonus", "Shift Log", "G", start, end)
    nrange(wb, "ShiftEarn", "Shift Log", "H", start, end)
    nrange(wb, "ShiftMiles", "Shift Log", "I", start, end)
    nrange(wb, "ShiftFuel", "Shift Log", "J", start, end)
    cell_name(wb, "TotalHours", "Shift Log", f"$C${trow}")
    cell_name(wb, "TotalTrips", "Shift Log", f"$D${trow}")
    cell_name(wb, "TotalFares", "Shift Log", f"$E${trow}")
    cell_name(wb, "TotalTips", "Shift Log", f"$F${trow}")
    cell_name(wb, "TotalBonus", "Shift Log", f"$G${trow}")
    cell_name(wb, "GrossEarn", "Shift Log", f"$H${trow}")
    cell_name(wb, "TotalMiles", "Shift Log", f"$I${trow}")
    cell_name(wb, "FuelTotal", "Shift Log", f"$J${trow}")
    ws.conditional_formatting.add(f"H{start}:H{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=340, color=PRIMARY, showValue=True))


# ===========================================================================
# Earnings Breakdown
# ===========================================================================
def build_earnings(wb):
    ws = wb.create_sheet("Earnings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 14, 12, 3, 18, 14, 12, 2])
    luxe_header(ws, "H", "💵  EARNINGS BREAKDOWN",
                "Where the money comes from — earnings mix and platform split, live from your log.")
    # earnings mix
    merge_set(ws, "B5:D5", "EARNINGS MIX (THIS MONTH)", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Source", "Amount", "Share"], start_col=2)
    mix = [("Fares", "=TotalFares"), ("Tips", "=TotalTips"), ("Bonus / Surge", "=TotalBonus")]
    ms = 7
    for i, (lab, fml) in enumerate(mix):
        r = ms + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cv = ws.cell(row=r, column=3, value=fml); cv.style = "td"; cv.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/GrossEarn,0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    me = ms + len(mix) - 1
    mt = me + 1
    ws.cell(row=mt, column=2, value="TOTAL EARNINGS").style = "th"
    ct = ws.cell(row=mt, column=3, value="=GrossEarn"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    ws.cell(row=mt, column=4).style = "td"; ws.cell(row=mt, column=4).fill = fill(SURFACE)
    nrange(wb, "MixSource", "Earnings", "B", ms, me)
    nrange(wb, "MixVal", "Earnings", "C", ms, me)
    ws.conditional_formatting.add(f"C{ms}:C{me}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=3300, color=GOLD_LT, showValue=True))
    # by platform
    merge_set(ws, "F5:H5", "BY PLATFORM", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Platform", "Earnings", "Share"], start_col=6)
    plats = ["Lyft", "Uber", "Both"]
    ps = 7
    for i, p in enumerate(plats):
        r = ps + i
        ws.cell(row=r, column=6, value=p).style = "td_left"
        cv = ws.cell(row=r, column=7, value=f'=SUMIF(ShiftPlatform,"{p}",ShiftEarn)'); cv.style = "td"; cv.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/GrossEarn,0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(6, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    pe = ps + len(plats) - 1
    nrange(wb, "PlatName", "Earnings", "F", ps, pe)
    nrange(wb, "PlatEarn", "Earnings", "G", ps, pe)
    # per-shift efficiency
    merge_set(ws, "B12:D12", "PER-SHIFT AVERAGES", "section_gold"); ws.row_dimensions[12].height = 22
    avgs = [("Avg / shift", "=IFERROR(GrossEarn/COUNT(ShiftEarn),0)", '"$"#,##0'),
            ("Avg / trip", "=IFERROR(GrossEarn/TotalTrips,0)", '"$"#,##0.00'),
            ("Gross $ / hour", "=IFERROR(GrossEarn/TotalHours,0)", '"$"#,##0.00'),
            ("Tips % of fares", "=IFERROR(TotalTips/TotalFares,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(avgs):
        r = 13 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        ws.cell(row=r, column=4).style = "td"; ws.cell(row=r, column=4).fill = fill(WHITE)
    merge_set(ws, "F12:H12", "TAKE-HOME (AFTER EXPENSES)", "section_gold"); ws.row_dimensions[12].height = 22
    take = [("Net earnings", "=NetEarn", '"$"#,##0'), ("Net $ / hour", "=PerHour", '"$"#,##0.00'),
            ("Net $ / mile", "=PerMile", '"$"#,##0.00'), ("Expense ratio", "=IFERROR(ExpenseTotal/GrossEarn,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(take):
        r = 13 + i
        ws.cell(row=r, column=6, value=lab).style = "field_label"
        c = ws.cell(row=r, column=7, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net earnings", "Net $ / hour"):
            c.fill = fill(MINT_BG)
        ws.cell(row=r, column=8).style = "td"; ws.cell(row=r, column=8).fill = fill(WHITE)


# ===========================================================================
# Mileage Tracker
# ===========================================================================
def build_mileage(wb):
    rows = [
        (dminus(23), "Home → downtown zone", "Business", 188, "Morning rush shift"),
        (dminus(22), "Airport runs", "Business", 205, "Airport queue day"),
        (dminus(20), "Suburbs + events", "Business", 158, "Slow afternoon"),
        (dminus(19), "Mixed platform day", "Business", 182, "Lyft + Uber"),
        (dminus(16), "Concert surge", "Business", 214, "Big surge night"),
        (dminus(13), "Downtown + campus", "Business", 198, "Steady"),
        (dminus(10), "Airport + late bars", "Business", 208, "Long day"),
        (dminus(8), "Mixed + delivery", "Business", 186, "Tried delivery"),
        (dminus(5), "Weekend rush", "Business", 200, "Best day this week"),
        (dminus(2), "Downtown day", "Business", 178, "Solid"),
        (dminus(11), "Personal errands", "Personal", 24, "Not deductible"),
    ]
    ws, start, end = build_log(
        wb, "Mileage", "🛣", "MILEAGE TRACKER",
        "Log business miles for the IRS deduction — the biggest tax break drivers miss.",
        ["Date", "Route / Purpose", "Type", "Miles", "Notes"],
        rows, [13, 30, 14, 12, 26],
        text_left={2, 5}, dates={1}, ints={4}, reserved=40)
    nrange(wb, "MileType", "Mileage", "C", start, end)
    nrange(wb, "MileMiles", "Mileage", "D", start, end)
    for t, cc in {"Business": MINT_BG, "Personal": WARN_BG}.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{t}"'], fill=fill(cc)))


# ===========================================================================
# Fuel Log
# ===========================================================================
def build_fuel(wb):
    rows = [
        (dminus(23), "Shell", 11.2, 3.29, "Full tank"),
        (dminus(20), "Costco", 10.8, 3.05, "Cheapest in town"),
        (dminus(17), "QuikTrip", 11.5, 3.19, "Topped off"),
        (dminus(14), "Shell", 10.9, 3.29, "Before airport day"),
        (dminus(11), "Costco", 11.3, 3.09, "Membership pays off"),
        (dminus(8), "Exxon", 10.6, 3.35, "Highway trip"),
        (dminus(5), "QuikTrip", 11.4, 3.15, "Weekend fill"),
        (dminus(2), "Costco", 11.0, 3.05, "End of month"),
    ]
    sample = [(d, st, g, ppg, f"=ROUND(C{i+L0}*D{i+L0},2)", note)
              for i, (d, st, g, ppg, note) in enumerate(rows)]
    ws, start, end = build_log(
        wb, "Fuel Log", "⛽", "FUEL LOG",
        "Track every fill-up — gallons, price & cost so your real fuel spend is never a guess.",
        ["Date", "Station", "Gallons", "$ / Gal", "Cost", "Notes"],
        sample, [13, 16, 12, 12, 12, 24],
        text_left={2, 6}, dates={1}, dec={3}, money2={4, 5}, reserved=40)
    trow = end + 1
    ws.cell(row=trow, column=2, value="TOTAL").style = "th"
    for col, fmt in ((3, "0.0"), (5, '"$"#,##0.00')):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt
    for col in (1, 4, 6):
        ws.cell(row=trow, column=col).style = "td"; ws.cell(row=trow, column=col).fill = fill(SURFACE)


# ===========================================================================
# Vehicle & Maintenance
# ===========================================================================
def build_vehicle(wb):
    rows = [
        (dminus(18), "Oil Change", 62, 66200, "Synthetic — every 5k mi", dplus(60)),
        (dminus(40), "Rotation", 25, 63800, "With oil change", dplus(38)),
        (dminus(120), "Tires", 640, 58900, "4 new — Michelin", dplus(600)),
        (dminus(70), "Brakes", 285, 61200, "Front pads + rotors", dplus(700)),
        (dminus(9), "Inspection", 40, 68100, "State + emissions", dplus(355)),
        (dminus(200), "Battery", 145, 55400, "3-yr warranty", dplus(900)),
        (dminus(30), "Repair", 180, 64500, "AC recharge + cabin filter", "—"),
    ]
    ws, start, end = build_log(
        wb, "Vehicle", "🔧", "VEHICLE & MAINTENANCE",
        "Keep your money-maker running — service history, cost & what's due next.",
        ["Date", "Service", "Cost", "Odometer", "Notes", "Next Due"],
        rows, [13, 16, 12, 13, 26, 13],
        text_left={2, 5}, dates={1, 6}, money={3}, ints={4}, reserved=30,
        validations=[("B", "MaintList")])
    trow = end + 1
    ws.cell(row=trow, column=2, value="TOTAL SPENT").style = "th"
    c = ws.cell(row=trow, column=3, value=f"=SUM(C{start}:C{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    for col in (1, 4, 5, 6):
        ws.cell(row=trow, column=col).style = "td"; ws.cell(row=trow, column=col).fill = fill(SURFACE)


# ===========================================================================
# Business Expenses  — defines ExpenseTotal
# ===========================================================================
def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 16, 3, 30, 2])
    luxe_header(ws, "F", "🧾  BUSINESS EXPENSES",
                "Every cost of driving — because deductions are money back in your pocket.")
    table_headers(ws, 4, ["Category", "This Month", "Annual (est.)"])
    exp = {"Fuel": "=FuelTotal", "Car Payment": 420, "Insurance": 185, "Maintenance & Repairs": 145,
           "Phone & Data": 45, "Tolls": 58, "Car Wash & Cleaning": 40, "Supplies": 25, "Miscellaneous": 0}
    start = L0; eend = start + len(EXP_CATS) - 1
    for i, cat in enumerate(EXP_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        val = exp[cat]
        cm = ws.cell(row=r, column=2, value=val)
        cm.style = "field_value" if cat == "Fuel" else "input"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    etot = eend + 1
    ws.cell(row=etot, column=1, value="TOTAL EXPENSES").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=etot, column=col, value=f"=SUM({L}{start}:{L}{eend})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "ExpCat", "Expenses", "A", start, eend)
    nrange(wb, "ExpMonthly", "Expenses", "B", start, eend)
    cell_name(wb, "ExpenseTotal", "Expenses", f"$B${etot}")
    ws.conditional_formatting.add(f"B{start}:B{eend}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=520, color=ACCENT, showValue=True))
    # bottom line
    merge_set(ws, "E4:F4", "COST PER MILE", "section_gold")
    rows2 = [("Total business miles", "=TotalMiles", "#,##0"),
             ("Expenses this month", "=ExpenseTotal", '"$"#,##0'),
             ("Cost per mile", "=IFERROR(ExpenseTotal/TotalMiles,0)", '"$"#,##0.00'),
             ("Fuel per mile", "=IFERROR(FuelTotal/TotalMiles,0)", '"$"#,##0.00'),
             ("Expense ratio", "=IFERROR(ExpenseTotal/GrossEarn,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 5 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Cost per mile":
            c.fill = fill(WARN_BG)


# ===========================================================================
# Monthly Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "📊  MONTHLY BUDGET",
                "Your take-home, put to work — driving net income flows in; plan every dollar.")
    merge_set(ws, "B5:E5", "INCOME", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Source", "Planned", "Actual", ""], start_col=2)
    inc = [("Driving (net)", "=NetGoal", "=NetEarn"), ("Side income", 300, 250), ("Other", 0, 0)]
    is_ = 7
    for i, (lab, pl, ac) in enumerate(inc):
        r = is_ + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cp = ws.cell(row=r, column=3, value=pl); cp.style = "td" if lab == "Driving (net)" else "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=ac); ca.style = "td" if lab == "Driving (net)" else "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=5).style = "td"; ws.cell(row=r, column=5).fill = fill(WHITE)
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ie = is_ + len(inc) - 1; itot = ie + 1
    ws.cell(row=itot, column=2, value="TOTAL INCOME").style = "th"
    for col in (3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=itot, column=col, value=f"=SUM({L}{is_}:{L}{ie})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=itot, column=5).style = "td"; ws.cell(row=itot, column=5).fill = fill(SURFACE)
    cell_name(wb, "BudgetIncome", "Budget", f"$D${itot}")
    # expenses
    merge_set(ws, "B12:E12", "HOUSEHOLD EXPENSES", "section_gold"); ws.row_dimensions[12].height = 22
    table_headers(ws, 13, ["Category", "Planned", "Actual", "Δ"], start_col=2)
    hh = [("Housing", 1200, 1200), ("Utilities", 190, 205), ("Groceries", 450, 470),
          ("Health", 250, 250), ("Debt", 300, 300), ("Savings", 400, 400),
          ("Personal", 250, 220), ("Insurance", 140, 140), ("Other", 120, 95)]
    hs = 14
    for i, (cat, pl, ac) in enumerate(hh):
        r = hs + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=3, value=pl); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=ac); ca.style = "input"; ca.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cd.style = "td"; cd.number_format = '"$"#,##0;[Red]-"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    he = hs + len(hh) - 1; htot = he + 1
    ws.cell(row=htot, column=2, value="TOTAL EXPENSES").style = "th"
    for col in (3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=htot, column=col, value=f"=SUM({L}{hs}:{L}{he})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=htot, column=5).style = "td"; ws.cell(row=htot, column=5).fill = fill(SURFACE)
    nrange(wb, "BudgetCat", "Budget", "B", hs, he)
    nrange(wb, "BudgetActual", "Budget", "D", hs, he)
    cell_name(wb, "BudgetExpense", "Budget", f"$D${htot}")
    # leftover
    lo = htot + 2
    merge_set(ws, f"B{lo}:C{lo}", "LEFTOVER (INCOME − EXPENSES)", "section_gold")
    c = ws.cell(row=lo, column=4, value="=BudgetIncome-BudgetExpense"); c.style = "field_value"; c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    ws.conditional_formatting.add(f"D{lo}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))


# ===========================================================================
# Tax Center
# ===========================================================================
def build_tax(wb):
    ws = wb.create_sheet("Tax Center"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 4, 26, 16, 2])
    luxe_header(ws, "G", "🧮  TAX CENTER",
                "Keep more of what you earn — the mileage deduction, set-aside & quarterly estimates.")
    merge_set(ws, "B5:C5", "MILEAGE VS ACTUAL METHOD", "section_gold"); ws.row_dimensions[5].height = 22
    left = [("Business miles", "=TotalMiles", "#,##0"),
            ("IRS mileage rate", "=MileageRate", '"$"#,##0.00'),
            ("Mileage deduction", "=TotalMiles*MileageRate", '"$"#,##0'),
            ("Actual costs (expenses)", "=ExpenseTotal", '"$"#,##0'),
            ("Better method", '=IF(TotalMiles*MileageRate>=ExpenseTotal,"Mileage","Actual")', "text"),
            ("Deduction (best)", "=MAX(TotalMiles*MileageRate,ExpenseTotal)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(left):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"
        if fmt != "text":
            c.number_format = fmt
        if lab in ("Mileage deduction", "Deduction (best)"):
            c.fill = fill(MINT_BG)
    cell_name(wb, "TaxDeduction", "Tax Center", "$C$8")
    merge_set(ws, "E5:F5", "SET-ASIDE & ESTIMATE", "section_gold"); ws.row_dimensions[5].height = 22
    right = [("Gross earnings", "=GrossEarn", '"$"#,##0'),
             ("− Mileage deduction", "=-TotalMiles*MileageRate", '"$"#,##0'),
             ("Taxable (est.)", "=MAX(GrossEarn-TotalMiles*MileageRate,0)", '"$"#,##0'),
             ("Set-aside rate", "=TaxRate", "0%"),
             ("Set aside this month", "=MAX(GrossEarn-TotalMiles*MileageRate,0)*TaxRate", '"$"#,##0'),
             ("Quarterly estimate", "=MAX(GrossEarn-TotalMiles*MileageRate,0)*TaxRate*3", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(right):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Set aside this month", "Quarterly estimate"):
            c.fill = fill(WARN_BG)
    merge_set(ws, "B13:F13", "DEDUCTION CHECKLIST — DON'T LEAVE MONEY ON THE TABLE", "section_gold"); ws.row_dimensions[13].height = 22
    checks = [("Business mileage (biggest one)", "Yes"), ("Phone & data (business %)", "Yes"),
              ("Car wash & cleaning", "Yes"), ("Tolls & parking", "Yes"),
              ("Water/snacks for riders", "Yes"), ("Phone mount & chargers", "Yes"),
              ("Dashcam & accessories", "Yes"), ("Platform & bank fees", "Yes")]
    for i, (item, yn) in enumerate(checks):
        r = 14 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=item).style = "td_left"
        cc = ws.cell(row=r, column=col + 1, value=yn); cc.style = "td"
        add_dv(ws, f"{get_column_letter(col+1)}{r}", "YesNoList")
        ws.conditional_formatting.add(f"{get_column_letter(col+1)}{r}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# Savings & Goals  — defines SavingsSaved, TaxSetAside
# ===========================================================================
def build_savings(wb):
    ws = wb.create_sheet("Savings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 12, 2])
    luxe_header(ws, "E", "🐷  SAVINGS & GOALS",
                "Pay yourself first — build an emergency fund, tax reserve & drive toward big goals.")
    merge_set(ws, "B5:E5", "SAVINGS BUCKETS", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Bucket", "Goal", "Saved", "Progress"], start_col=2)
    buckets = [("Emergency fund", "=SavingsGoal", 4200), ("Tax reserve", "=TaxReserveGoal", 1400),
               ("New(er) car fund", 8000, 3100), ("Vacation", 2500, 900)]
    bs = 7
    for i, (lab, goal, saved) in enumerate(buckets):
        r = bs + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cg = ws.cell(row=r, column=3, value=goal); cg.style = "field_value" if isinstance(goal, str) else "input"; cg.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(MIN(D{r}/C{r},1),0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    be = bs + len(buckets) - 1
    cell_name(wb, "SavingsSaved", "Savings", f"$D${bs}")
    cell_name(wb, "TaxSetAside", "Savings", f"$D${bs+1}")
    ws.conditional_formatting.add(f"E{bs}:E{be}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    # goals
    merge_set(ws, "B13:E13", "DRIVER GOALS & OKRs", "section_gold"); ws.row_dimensions[13].height = 22
    table_headers(ws, 14, ["Goal", "Target", "Current", "Progress"], start_col=2)
    goals = [("$3,000 net / month", "$3,000", "$2,786", 0.93),
             ("$20 net / hour", "$20.00", "$21.68", 1.00),
             ("300 trips / month", "300", "316", 1.00),
             ("$6k emergency fund", "$6,000", "$4,200", 0.70),
             ("Pay off $4k card", "$4,000", "$2,600", 0.65),
             ("2,900 miles / month", "2,900", "2,871", 0.99)]
    gs = 15
    for i, (g, tgt, cur, prog) in enumerate(goals):
        r = gs + i
        ws.cell(row=r, column=2, value=g).style = "td_left"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ge = gs + len(goals) - 1
    ws.conditional_formatting.add(f"E{gs}:E{ge}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GOLD_LT, showValue=True))


# ===========================================================================
# Bonuses & Quests
# ===========================================================================
def build_bonuses(wb):
    rows = [
        ("Lyft", "Ride Challenge", "50 rides / week", "Completed", 120, dminus(3), "Hit with room to spare"),
        ("Uber", "Quest", "40 trips (Fri-Sun)", "In Progress", 90, dplus(2), "31 / 40 done"),
        ("Lyft", "Streak Bonus", "3 rides in a row", "Completed", 18, dminus(5), "Stacked 4 streaks"),
        ("Uber", "Consecutive Trips", "6 in a row", "Available", 45, dplus(4), "Airport is best for this"),
        ("Lyft", "Peak Bonus", "Downtown 6-9pm", "Completed", 64, dminus(6), "Concert night"),
        ("Uber", "Boost Zone", "1.5x East side", "Available", 0, dplus(1), "Watch the map"),
        ("Lyft", "Weekly Guarantee", "$1,100 / 55 rides", "In Progress", 75, dplus(3), "On pace"),
    ]
    ws, start, end = build_log(
        wb, "Bonuses", "🎯", "BONUSES & QUESTS",
        "Never miss a promo — track every quest, streak & guarantee and the extra it pays.",
        ["Platform", "Type", "Requirement", "Status", "Reward", "Deadline", "Notes"],
        rows, [12, 16, 20, 14, 11, 13, 24],
        text_left={3, 7}, money={5}, dates={6}, reserved=30,
        validations=[("A", "PlatformList")])
    nrange(wb, "BonusReward", "Bonuses", "E", start, end)
    for st, cc in {"Completed": MINT_BG, "In Progress": WARN_BG, "Available": SOFT_BG}.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# Hot Zones & Hours
# ===========================================================================
def build_zones(wb):
    ws = wb.create_sheet("Hot Zones"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 16, 14, 14, 26, 2])
    luxe_header(ws, "G", "📍  HOT ZONES & HOURS",
                "Drive where it pays — the zones, days & hours that earn the most per hour.")
    merge_set(ws, "B5:F5", "BEST ZONES", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Zone", "Best Time", "Avg $/hr", "Surge", "Notes"], start_col=2)
    zones = [
        ("Airport (AUS)", "6-9am, 4-7pm", 34, "Medium", "Long trips, worth the queue"),
        ("Downtown / 6th St", "Fri-Sat 9pm-2am", 41, "High", "Bar close = surge gold"),
        ("Domain / North", "Weekday PM", 28, "Low", "Steady office runs"),
        ("Campus (UT)", "Game days", 38, "High", "Events spike demand"),
        ("Convention Center", "Event nights", 45, "High", "Check the event calendar"),
        ("Suburbs", "Weekday AM", 22, "Low", "Slow — reposition"),
    ]
    zs = 7
    for i, (z, t, hr, surge, note) in enumerate(zones):
        r = zs + i
        ws.cell(row=r, column=2, value=z).style = "td_left"
        ws.cell(row=r, column=3, value=t).style = "td"
        ch = ws.cell(row=r, column=4, value=hr); ch.style = "td"; ch.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=surge).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ze = zs + len(zones) - 1
    nrange(wb, "ZoneName", "Hot Zones", "B", zs, ze)
    nrange(wb, "ZoneRate", "Hot Zones", "D", zs, ze)
    ws.conditional_formatting.add(f"D{zs}:D{ze}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=45, color=PRIMARY, showValue=True))
    for st, cc in {"High": MINT_BG, "Medium": WARN_BG, "Low": WHITE}.items():
        ws.conditional_formatting.add(f"E{zs}:E{ze}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    merge_set(ws, "B15:F15", "BEST DAYS OF WEEK", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Day", "Avg $/hr", "Best Window", "Rating"], start_col=2)
    days = [("Friday", 36, "5pm-2am", "★★★★★"), ("Saturday", 39, "6pm-3am", "★★★★★"),
            ("Sunday", 27, "9am-2pm brunch", "★★★☆☆"), ("Thursday", 30, "5-9pm", "★★★★☆"),
            ("Weekday AM", 26, "6-9am", "★★★☆☆")]
    ds = 17
    for i, (d, hr, win, rate) in enumerate(days):
        r = ds + i
        ws.cell(row=r, column=2, value=d).style = "td_left"
        ch = ws.cell(row=r, column=4, value=win); ch.style = "td"
        cr = ws.cell(row=r, column=3, value=hr); cr.style = "td"; cr.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=rate).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)


# ===========================================================================
# Ratings & Feedback
# ===========================================================================
def build_ratings(wb):
    ws = wb.create_sheet("Ratings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 12, 12, 14, 30, 2])
    luxe_header(ws, "G", "⭐  RATINGS & FEEDBACK",
                "Protect your rating — it drives your access to the best rides & bonuses.")
    merge_set(ws, "B5:F5", "PLATFORM STANDING", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Platform", "Rating", "Trips", "Trend", "Riders Say"], start_col=2)
    rows = [
        ("Lyft", 4.98, 5120, "▲ +0.01", "Clean car, smooth ride"),
        ("Uber", 4.96, 4120, "► flat", "Great conversation"),
        ("Lyft Acceptance", 0.84, 0, "▲ +3%", "Above Platinum line"),
        ("Cancellation", 0.02, 0, "► flat", "Well under limit"),
    ]
    start = 7
    for i, (plat, rt, cnt, trend, note) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=plat).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rt); cr.style = "td"
        cr.number_format = "0.00" if rt > 3 else "0%"
        cc = ws.cell(row=r, column=4, value=cnt if cnt else "—"); cc.style = "td"
        if cnt:
            cc.number_format = "#,##0"
        ws.cell(row=r, column=5, value=trend).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    ws.conditional_formatting.add(f"C{start}:C{start+1}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=5, color=GOLD_LT, showValue=True))
    merge_set(ws, "B12:F12", "5-STAR CHECKLIST", "section_gold"); ws.row_dimensions[12].height = 22
    tips = ["Offer water, gum & a phone charger", "Confirm the rider's name at pickup",
            "Keep the car spotless & smelling fresh", "Let the rider pick music & temperature",
            "Smooth acceleration & braking", "Know the fastest route — don't rely only on GPS"]
    for i, t in enumerate(tips):
        r = 13 + (i // 2)
        col = 2 if i % 2 == 0 else 4
        merge_set(ws, f"{get_column_letter(col)}{r}:{get_column_letter(col+1)}{r}", "✓  " + t, "td_left")


# ===========================================================================
# Analytics — Driver Health Score + computed cells
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 3, 20, 12, 12, 2])
    luxe_header(ws, "G", "📈  ANALYTICS COMMAND CENTER",
                "Your driving business by the numbers — take-home, health dimensions & the Driver Health Score.")
    merge_set(ws, "B5:C5", "TAKE-HOME SNAPSHOT", "section")
    snap = [("Gross earnings", "=GrossEarn", '"$"#,##0', None),
            ("Business expenses", "=ExpenseTotal", '"$"#,##0', None),
            ("Net earnings", "=GrossEarn-ExpenseTotal", '"$"#,##0', "NetEarn"),
            ("Online hours", "=TotalHours", "0.0", None),
            ("Net $ / hour", "=IFERROR((GrossEarn-ExpenseTotal)/TotalHours,0)", '"$"#,##0.00', "PerHour"),
            ("Net $ / mile", "=IFERROR((GrossEarn-ExpenseTotal)/TotalMiles,0)", '"$"#,##0.00', "PerMile"),
            ("Trips", "=TotalTrips", "#,##0", None),
            ("Mileage deduction", "=TotalMiles*MileageRate", '"$"#,##0', None)]
    for i, (lab, fml, fmt, nm) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if nm:
            cell_name(wb, nm, "Analytics", f"$C${r}")
        if lab in ("Net earnings", "Net $ / hour"):
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    merge_set(ws, "E5:G5", "DRIVER HEALTH", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=5)
    metrics = [
        ("Net earnings vs goal", "=IFERROR(MIN(NetEarn/NetGoal,1),0)"),
        ("Net $/hour vs target", "=IFERROR(MIN(PerHour/HourlyTarget,1),0)"),
        ("Trips vs goal", "=IFERROR(MIN(TotalTrips/TripsGoal,1),0)"),
        ("Emergency fund", "=IFERROR(MIN(SavingsSaved/SavingsGoal,1),0)"),
        ("Driving consistency", "=IFERROR(MIN(COUNT(ShiftMiles)/ShiftGoal,1),0)"),
        ("Tax reserve", "=IFERROR(MIN(TaxSetAside/TaxReserveGoal,1),0)"),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(metrics):
        r = hs + i
        ws.cell(row=r, column=5, value=dim).style = "td_left"
        c = ws.cell(row=r, column=6, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=7, value=f'=IF(F{r}>=0.75,"Strong",IF(F{r}>=0.5,"Growing","Focus"))').style = "td"
        if i % 2:
            for c2 in range(5, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(metrics) - 1
    cell_name(wb, "HealthRange", "Analytics", f"$F${hs}:$F${he}")
    ws.conditional_formatting.add(f"F{hs}:F{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # top days by earnings
    merge_set(ws, "B15:G15", "BEST SHIFTS (BY EARNINGS)", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Shift", "Earnings", "Hours", "$/hr", "Miles"], start_col=2)
    best = [("Concert surge night", 340, 9.5, 214), ("Weekend rush", 310, 9.0, 200),
            ("Airport queue day", 318, 9.0, 205), ("Long day (airport+bars)", 326, 9.0, 208),
            ("Mixed platform day", 297, 8.5, 198), ("Downtown day", 278, 8.5, 188)]
    vs = 17
    for i, (title, earn, hrs, miles) in enumerate(best):
        r = vs + i
        ws.cell(row=r, column=2, value=title).style = "td_left"
        ce = ws.cell(row=r, column=3, value=earn); ce.style = "td"; ce.number_format = '"$"#,##0'
        ch = ws.cell(row=r, column=4, value=hrs); ch.style = "td"; ch.number_format = "0.0"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cp.style = "td"; cp.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=6, value=miles); cm.style = "td"; cm.number_format = "#,##0"
        if i % 2:
            for c2 in range(2, 7):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    ve = vs + len(best) - 1
    nrange(wb, "BestShift", "Analytics", "B", vs, ve)
    nrange(wb, "BestEarn", "Analytics", "C", vs, ve)
    ws.conditional_formatting.add(f"C{vs}:C{ve}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=340, color=PRIMARY, showValue=True))
    # monthly earnings trend
    merge_set(ws, "B25:C25", "NET EARNINGS — 6 MONTHS", "section")
    ws.cell(row=26, column=2, value="Month").style = "th"; ws.cell(row=26, column=3, value="Net ($)").style = "th"
    months = _recent_months(6); vals = [2380, 2510, 2690, 2620, 2740, 2786]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 27 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "TrendMonth", "Analytics", "$B$27:$B$32")
    cell_name(wb, "TrendVal", "Analytics", "$C$27:$C$32")


# ===========================================================================
# Weekly Planner
# ===========================================================================
def build_planner(wb):
    ws = wb.create_sheet("Planner"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 16, 14, 14, 26, 2])
    luxe_header(ws, "G", "🗓  WEEKLY PLANNER",
                "Plan the week that hits your goal — target hours, zones & the earnings you need.")
    table_headers(ws, 4, ["Day", "Plan", "Target Hrs", "$ Goal", "Zone / Focus"])
    days = [
        ("Monday", "Off", 0, 0, "Rest / car maintenance"),
        ("Tuesday", "PM shift", 6, 160, "Downtown offices"),
        ("Wednesday", "PM shift", 6, 160, "Airport evening"),
        ("Thursday", "Evening", 5, 150, "Happy hour → dinner"),
        ("Friday", "Night", 8, 300, "6th St bar close"),
        ("Saturday", "Night", 9, 330, "Events + downtown"),
        ("Sunday", "Brunch", 6, 180, "Brunch + airport"),
    ]
    start = L0
    for i, (d, plan, hrs, goal, zone) in enumerate(days):
        r = start + i
        ws.cell(row=r, column=1, value=d).style = "td_left"
        ws.cell(row=r, column=2, value=plan).style = "input"
        ch = ws.cell(row=r, column=3, value=hrs); ch.style = "input"; ch.number_format = "0.0"
        cg = ws.cell(row=r, column=4, value=goal); cg.style = "input"; cg.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=zone).style = "td_left"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(days) - 1
    trow = end + 1
    ws.cell(row=trow, column=1, value="WEEK TOTAL").style = "th"
    ws.cell(row=trow, column=2).style = "th"
    for col, fmt in ((3, "0.0"), (4, '"$"#,##0')):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt
    ws.cell(row=trow, column=5).style = "td"; ws.cell(row=trow, column=5).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


# ===========================================================================
# Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  RECEIPTS & VEHICLE GALLERY",
                "Your paper trail — snap receipts, odometer & maintenance photos for tax time.")
    merge_set(ws, "B5:D5", "HOW TO ADD PHOTOS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Excel: Insert ▸ Pictures ▸ Place in Cell (or drag an image) into any framed box below. "
                      "Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"paste-link-here\"). "
                      "Caption each one — date, amount & category — so tax season is a five-minute job.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Gas Receipt", "Maintenance", "Odometer", "Car Wash", "Repair Invoice", "Supplies"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add photo)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Date · amount · category…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🚗  LYFT DRIVER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Earnings, miles, taxes & savings — your whole driving business, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("NET EARNINGS", "=NetEarn", "money"),
        ("GROSS EARNINGS", "=GrossEarn", "money"),
        ("TIPS", "=TotalTips", "money"),
        ("NET $ / HOUR", "=PerHour", "money2"),
        ("NET $ / MILE", "=PerMile", "money2"),
        ("MILES DRIVEN", "=TotalMiles", "num"),
    ]
    row2 = [
        ("ONLINE HOURS", "=TotalHours", "dec"),
        ("TRIPS", "=TotalTrips", "num"),
        ("EXPENSES", "=ExpenseTotal", "money"),
        ("TAX DEDUCTION", "=TotalMiles*MileageRate", "money"),
        ("SAVINGS", "=IFERROR(MIN(SavingsSaved/SavingsGoal,1),0)", "pct"),
        ("DRIVER HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "EARNINGS & TAKE-HOME", "section_gold")
    ln = LineChart(); ln.title = "Net Earnings ($) — 6 mo"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=26, max_row=32), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=27, max_row=32)); ln.legend = None
    ws.add_chart(ln, "B12")
    d1 = DoughnutChart(); d1.title = "Earnings Mix"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Earnings"], min_col=3, min_row=6, max_row=9), titles_from_data=True)
    d1.set_categories(Reference(wb["Earnings"], min_col=2, min_row=7, max_row=9)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "BEST SHIFTS & EXPENSES", "section_gold")
    cb = BarChart(); cb.type = "bar"; cb.title = "Best Shifts by Earnings"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); cb.legend = None
    ws.add_chart(cb, "B30")
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Expenses"], min_col=2, min_row=4, max_row=13), titles_from_data=True)
    eb.set_categories(Reference(wb["Expenses"], min_col=1, min_row=5, max_row=13)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Lyft Driver Command Center™ — every shift, mile & dollar in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_shiftlog(wb)
    build_earnings(wb); build_mileage(wb); build_fuel(wb); build_vehicle(wb)
    build_expenses(wb); build_budget(wb); build_tax(wb); build_savings(wb)
    build_bonuses(wb); build_zones(wb); build_ratings(wb); build_analytics(wb)
    build_planner(wb); build_gallery(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Driver Profile", "Shift Log", "Earnings", "Mileage", "Fuel Log",
             "Vehicle", "Expenses", "Budget", "Tax Center", "Savings", "Bonuses", "Hot Zones", "Ratings",
             "Analytics", "Planner", "Gallery", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Lyft_Driver_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
