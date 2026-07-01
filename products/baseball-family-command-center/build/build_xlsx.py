"""Build Baseball Family Command Center™ — The Ultimate Baseball Season OS.

24 sheets + Welcome · a premium operating system for baseball families:
schedule, budget, equipment, stats, pitching, travel, development & team.

Run: python3 build_xlsx.py   ->  ../Baseball_Family_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"
ACCENT = "937356"
GOLD_LT = "C9A86A"
SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"
MINT_BG = "E3F8EF"
WHITE = "FFFFFF"
TEXT = "333333"
DANGER = "C94C4C"
RED_BG = "FBE6E6"
WARN_BG = "FBF0E2"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"
IVORY = "FBF8F2"

# ---- dropdown vocabularies ----
EVENT_TYPES = ["Game", "Practice", "Tournament", "Batting Cage", "Strength",
               "Private Lesson", "Team Meeting", "School", "Family", "Volunteer"]
LEAGUES = ["Little League", "Travel", "Select", "Club", "Pony", "Cal Ripken", "AAU", "Rec"]
DIVISIONS = ["6U", "8U", "10U", "12U", "14U", "16U", "18U"]
POSITIONS = ["Pitcher", "Catcher", "1B", "2B", "3B", "SS", "LF", "CF", "RF", "DH", "Utility"]
EXPENSE_CATS = ["Registration", "League Fees", "Uniforms", "Cleats", "Gloves",
                "Bats", "Helmets", "Batting Gloves", "Bags", "Protective Gear",
                "Tournaments", "Hotels", "Fuel", "Meals", "Private Lessons",
                "Batting Cage", "Team Apparel", "Photos", "Fundraising", "Miscellaneous"]
EXPENSE_GROUPS = ["Fees & Registration", "Equipment", "Travel", "Training", "Extras"]
EQUIP_TYPES = ["Glove", "Bat", "Cleats", "Helmet", "Catcher's Gear",
               "Batting Gloves", "Bat Bag", "Baseballs", "Training Aid", "Protective"]
EQUIP_STATUS = ["New", "Good", "Fair", "Replace Soon", "Replace Now"]
VOL_ROLES = ["Snack", "Dugout Helper", "Scorekeeper", "Pitch Count",
             "Team Parent", "Fundraising", "Field Prep", "Concessions"]
PRACTICE_TYPES = ["Team Practice", "Batting Cage", "Fielding", "Pitching",
                  "Strength", "Private Lesson", "Scrimmage"]
TOURNEY_TYPES = ["Pool Play", "Bracket", "Showcase", "Championship", "Friendly"]
SKILL_LEVELS = ["Developing", "Average", "Strong", "Elite"]
BATS = ["Left", "Right", "Switch"]
THROWS = ["Left", "Right"]
ATTEND = ["Present", "Absent", "Excused"]
YESNO = ["Yes", "No"]

CAT_GROUP = {
    "Registration": "Fees & Registration", "League Fees": "Fees & Registration",
    "Tournaments": "Fees & Registration",
    "Uniforms": "Equipment", "Cleats": "Equipment", "Gloves": "Equipment",
    "Bats": "Equipment", "Helmets": "Equipment", "Batting Gloves": "Equipment",
    "Bags": "Equipment", "Protective Gear": "Equipment",
    "Hotels": "Travel", "Fuel": "Travel", "Meals": "Travel",
    "Private Lessons": "Training", "Batting Cage": "Training",
    "Team Apparel": "Extras", "Photos": "Extras", "Fundraising": "Extras",
    "Miscellaneous": "Extras",
}

LOG_ROWS = 40
L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


# ===========================================================================
# Styles & shared helpers (premium pattern)
# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"),
                            fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True),
                               fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY),
                              alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT),
                                   alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"),
                         fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT),
                              alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
                              border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY),
                            fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True),
                             fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "prompt": NamedStyle(name="prompt", font=f(11, False, TEXT, italic=True),
                             alignment=Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True),
                             border=BOX, fill=PatternFill("solid", fgColor=IVORY)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT),
                           alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None,
               dates=None, pcts=None, start_col=1):
    text_left = text_left or set()
    money = money or set()
    ints = ints or set()
    dates = dates or set()
    pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}")
    ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label
    lc.font = Font(size=9, bold=True, color=ACCENT)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vc.font = Font(size=18, bold=True, color=PRIMARY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pct": "0%",
                        "days": "0", "dec": "0.0"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc)
            c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN,
                              top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers),
               text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def no_labels():
    dl = DataLabelList()
    dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 3] + [17] * 7)
    luxe_header(ws, "K", "⚙  SETTINGS", "Set your season inputs once — every dashboard follows. Edit the lists to fit your program.")

    merge_set(ws, "B5:C5", "SEASON INPUTS", "section")
    controls = [
        ("Player Name", "Cody Reyes", None, "PlayerName"),
        ("Team", "Bandits 12U Select", None, "TeamName"),
        ("Monthly Baseball Budget", 900, '"$"#,##0', "MonthlyBudget"),
        ("Season Start", dminus(45), "mm/dd/yyyy", "SeasonStart"),
        ("Season End", dplus(120), "mm/dd/yyyy", "SeasonEnd"),
        ("Pitch Count Limit", 75, "0", "PitchLimit"),
        ("Innings / Game", 7, "0", "GameInnings"),
        ("Home Ballpark", "Miller Field", None, "HomePark"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")

    bank1 = [("E", "Event Type", EVENT_TYPES, "EventTypeList"),
             ("F", "League", LEAGUES, "LeagueList"),
             ("G", "Division", DIVISIONS, "DivisionList"),
             ("H", "Position", POSITIONS, "PositionList"),
             ("I", "Expense Category", EXPENSE_CATS, "ExpenseCatList"),
             ("J", "Equipment Type", EQUIP_TYPES, "EquipTypeList"),
             ("K", "Equipment Status", EQUIP_STATUS, "EquipStatusList")]
    _emit_lists(wb, ws, bank1, 5, 6)

    bank2 = [("E", "Volunteer Role", VOL_ROLES, "VolRoleList"),
             ("F", "Practice Type", PRACTICE_TYPES, "PracticeTypeList"),
             ("G", "Tournament Type", TOURNEY_TYPES, "TourneyTypeList"),
             ("H", "Skill Level", SKILL_LEVELS, "SkillLevelList"),
             ("I", "Attendance", ATTEND, "AttendList"),
             ("J", "Yes / No", YESNO, "YesNoList"),
             ("K", "Expense Group", EXPENSE_GROUPS, "ExpenseGroupList")]
    _emit_lists(wb, ws, bank2, 30, 31)

    # extra single-column lists (Bats / Throws) tucked below bank2
    ws.cell(row=40, column=5, value="Bats").style = "th"
    for i, v in enumerate(BATS):
        ws.cell(row=41 + i, column=5, value=v).style = "td_left"
    wb.defined_names["BatsList"] = DefinedName("BatsList", attr_text="Settings!$E$41:$E$43")
    ws.cell(row=40, column=6, value="Throws").style = "th"
    for i, v in enumerate(THROWS):
        ws.cell(row=41 + i, column=6, value=v).style = "td_left"
    wb.defined_names["ThrowsList"] = DefinedName("ThrowsList", attr_text="Settings!$F$41:$F$42")


def _emit_lists(wb, ws, lists, header_row, data_row):
    merge_set(ws, f"E{header_row}:K{header_row}", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in lists:
        ci = column_index_from_string(col)
        ws.cell(row=header_row + 1, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=data_row + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(
            nm, attr_text=f"Settings!${col}${data_row + 1}:${col}${data_row + len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 74, 3])
    ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ⚾  BASEBALL FAMILY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate baseball season operating system — your whole season in one place.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    merge_set(ws, "B5:B5", "WELCOME TO THE SEASON", "section_gold")
    ws.merge_cells("B6:B8")
    ws["B6"].value = (
        "Baseball season means dozens of games and practices, tournaments, travel, "
        "equipment, fundraising and player development. Baseball Family Command Center™ "
        "brings all of it into one elegant dashboard so nothing gets missed, costs stay "
        "under control, and the whole family runs on a system instead of from memory.")
    ws["B6"].style = "body"
    for r in (6, 7, 8):
        ws.row_dimensions[r].height = 22

    merge_set(ws, "B10:B10", "HOW TO USE IT", "section")
    steps = [
        "1.  Open Settings and add your player, team, budget & season dates.",
        "2.  Fill the Player Profile, Season Calendar, and Baseball Budget.",
        "3.  Log games, practices, batting stats & pitching outings.",
        "4.  Plan tournaments, travel & the snack / volunteer schedule.",
        "5.  Watch the Executive Baseball Dashboard update itself automatically.",
    ]
    for i, s in enumerate(steps):
        r = 11 + i
        ws.merge_cells(f"B{r}:B{r}")
        ws[f"B{r}"].value = s
        ws[f"B{r}"].style = "body"
        ws.row_dimensions[r].height = 22

    dr = 18
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th")
    ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = (
        "Sample data is included to show how everything connects — just type over it with "
        "your own. Batting average, OBP, SLG, OPS and ERA calculate automatically. Every "
        "sheet is print-friendly and works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT)
    c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22
        ws.cell(row=rr, column=2).fill = fill(WARN_BG)

    merge_set(ws, f"B{dr+5}:B{dr+5}",
              "One organized season, less stress, more baseball — let's play ball.", "section_gold")


# ===========================================================================
# 2 — Player Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Player Profile")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 26, 6, 24, 26, 2])
    luxe_header(ws, "G", "⚾  PLAYER PROFILE",
                "Everything about your player in one place — stats card, contacts & medical.")
    blocks = [
        ("PLAYER", [("Player Name", "=PlayerName"), ("Date of Birth", dt.date(2013, 5, 8)),
                    ("Division", "12U"), ("Team", "=TeamName"), ("League", "Select"),
                    ("Jersey Number", 7), ("Primary Position", "Shortstop"),
                    ("Secondary Position", "Pitcher"), ("Bats", "Right"), ("Throws", "Right"),
                    ("Height", "5'0\""), ("Weight", "92 lb")]),
        ("TEAM & STAFF", [("Head Coach", "Coach Ramirez"), ("Assistant Coach", "Coach Webb"),
                          ("Team Manager", "Dana R."), ("Home Ballpark", "=HomePark"),
                          ("Practice Night", "Mon / Wed"), ("Team Colors", "Navy / Gold")]),
        ("MEDICAL & EMERGENCY", [("Emergency Contact", "Mom — (555) 210-7788"),
                                 ("Secondary Contact", "Dad — (555) 210-4521"),
                                 ("Allergies", "Bee stings"), ("Medical Notes", "Carries EpiPen"),
                                 ("Insurance Provider", "____________"),
                                 ("Policy #", "____________")]),
        ("SEASON GOALS", [("Primary Goal", "Bat .350+"), ("Hitting", "Cut down on strikeouts"),
                          ("Fielding", "Clean up throws from SS"), ("Off-Field", "Make honor roll")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold")
        ws.row_dimensions[row].height = 22
        row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            cv = ws.cell(row=row, column=3, value=fields[i][1]); cv.style = "field_value"
            if isinstance(fields[i][1], dt.date):
                cv.number_format = "mm/dd/yyyy"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24
            i += 2
            row += 1
        row += 1


# ===========================================================================
# 3 — Master Season Calendar
# ===========================================================================
def build_calendar(wb):
    sample = [
        ("Practice", "Team practice", dplus(1), "5:30 PM", "Miller Field", "Bring full gear"),
        ("Game", "vs Bandits", dplus(2), "6:00 PM", "Miller Field", "Home — navy"),
        ("Practice", "Fielding focus", dplus(3), "5:30 PM", "Miller Field", ""),
        ("Game", "vs Rockets", dplus(5), "11:00 AM", "Eastside Park", "Away — white"),
        ("Batting Cage", "Cage session", dplus(4), "7:00 PM", "Grand Slam Cages", ""),
        ("Practice", "Pitching + bullpen", dplus(7), "5:30 PM", "Miller Field", ""),
        ("Game", "vs Cobras", dplus(9), "1:00 PM", "North Diamond", "Away"),
        ("Private Lesson", "Hitting lesson", dplus(6), "4:00 PM", "Pro Swing Academy", "Coach Webb"),
        ("School", "Science fair", dplus(5), "6:00 PM", "Lincoln Middle", "Conflict w/ game!"),
        ("Game", "vs Storm", dplus(12), "10:00 AM", "Miller Field", "Home"),
        ("Tournament", "Summer Slam", dplus(14), "All weekend", "Capital City", "3 nights"),
        ("Family", "Cousin's graduation", dplus(14), "2:00 PM", "Home", "Conflict w/ tournament"),
        ("Strength", "Speed & agility", dplus(8), "4:30 PM", "Field House", ""),
        ("Volunteer", "Snack duty", dplus(2), "6:00 PM", "Miller Field", "Reyes family"),
        ("Game", "vs Titans (W 8-5)", dminus(3), "6:00 PM", "Miller Field", "Home — won"),
        ("Game", "vs Rockets (L 3-6)", dminus(6), "11:00 AM", "Eastside Park", "Away"),
        ("Game", "vs Cobras (W 7-2)", dminus(10), "1:00 PM", "North Diamond", "Away — won"),
        ("Game", "vs Storm (W 5-4)", dminus(13), "10:00 AM", "Miller Field", "Home — won"),
    ]
    ws, start, end = build_log(
        wb, "Calendar", "📅", "MASTER SEASON CALENDAR",
        "Every game, practice & tournament — with countdowns and conflict detection.",
        ["Type", "Event / Opponent", "Date", "Time", "Location", "Notes"],
        sample, [15, 26, 13, 13, 20, 24],
        text_left={2, 5, 6}, dates={3},
        validations=[("A", "EventTypeList")], reserved=70)
    nrange(wb, "CalType", "Calendar", "A", start, end)
    nrange(wb, "CalEvent", "Calendar", "B", start, end)
    nrange(wb, "CalDate", "Calendar", "C", start, end)
    ws.cell(row=4, column=7, value="In").style = "th"
    ws.cell(row=4, column=8, value="Clash").style = "th"
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 8
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=7, value=f'=IF(C{r}="","",C{r}-TODAY())')
        c.style = "td"; c.number_format = "0;[Red]-0"
        c.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
        cf = ws.cell(row=r, column=8,
                     value=f'=IF(C{r}="","",IF(AND(COUNTIF($C${start}:C{r},C{r})=1,COUNTIF(CalDate,C{r})>1),1,0))')
        cf.style = "td"; cf.number_format = "0;;"
        cf.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
    nrange(wb, "CalConflict", "Calendar", "H", start, end)
    ws.conditional_formatting.add(
        f"A{start}:F{end}",
        FormulaRule(formula=[f'AND($C{start}<>"",COUNTIF(CalDate,$C{start})>1)'], fill=fill(WARN_BG)))
    ws.conditional_formatting.add(
        f"A{start}:F{end}",
        FormulaRule(formula=[f'AND($A{start}="Game",$C{start}>=TODAY(),$C{start}-TODAY()<=3)'], fill=fill(MINT_BG)))
    ws.cell(row=4, column=10, value="SEASON COUNTS").style = "section_gold"
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 10
    counts = [("Games", '=COUNTIF(CalType,"Game")'), ("Practices", '=COUNTIF(CalType,"Practice")'),
              ("Tournaments", '=COUNTIF(CalType,"Tournament")'), ("Other events",
               '=SUMPRODUCT((CalType<>"")*(CalType<>"Game")*(CalType<>"Practice")*(CalType<>"Tournament"))')]
    for i, (lab, fml) in enumerate(counts):
        r = 5 + i
        ws.cell(row=r, column=10, value=lab).style = "td_left"
        ws.cell(row=r, column=11, value=fml).style = "td"
    wb.defined_names["CalCountLabels"] = DefinedName("CalCountLabels", attr_text="Calendar!$J$5:$J$8")
    wb.defined_names["CalCountVals"] = DefinedName("CalCountVals", attr_text="Calendar!$K$5:$K$8")


# ===========================================================================
# 4 — Game Day Command Center
# ===========================================================================
def build_gameday(wb):
    build_log(
        wb, "Game Day", "🏟", "GAME DAY COMMAND CENTER",
        "Everything for game day — ballpark, times, uniform, lineup & the final score.",
        ["Opponent", "Date", "Ballpark / Address", "Field", "Game Time", "Arrive By",
         "Uniform", "Result", "Score", "Player Notes"],
        [
            ("Bandits", dplus(2), "Miller Field (home)", "1", "6:00 PM", "5:00 PM", "Navy", "—", "", ""),
            ("Rockets", dplus(5), "Eastside Park, 40 Park Rd", "3", "11:00 AM", "10:00 AM", "White", "—", "", ""),
            ("Cobras", dplus(9), "North Diamond, 8 Field St", "2", "1:00 PM", "12:00 PM", "White", "—", "", ""),
            ("Storm", dplus(12), "Miller Field (home)", "1", "10:00 AM", "9:00 AM", "Navy", "—", "", ""),
            ("Titans", dminus(3), "Miller Field (home)", "1", "6:00 PM", "5:00 PM", "Navy", "W", "8-5", "2-4, 2 RBI"),
            ("Storm", dminus(13), "Miller Field (home)", "1", "10:00 AM", "9:00 AM", "Navy", "W", "5-4", "1-3, SB"),
        ],
        [16, 13, 24, 8, 12, 11, 10, 9, 9, 18],
        text_left={1, 3, 10}, dates={2}, reserved=40)


# ===========================================================================
# 5 — Practice Planner
# ===========================================================================
def build_practice(wb):
    sample = [
        (dminus(2), "Team Practice", "Coach Ramirez", "Present", "Infield + BP", 2.0, "Great energy"),
        (dminus(4), "Batting Cage", "Coach Webb", "Present", "Timing & contact", 1.0, ""),
        (dminus(6), "Fielding", "Coach Ramirez", "Present", "Double plays", 1.5, ""),
        (dminus(8), "Pitching", "Coach Ramirez", "Present", "Bullpen", 1.0, "Worked changeup"),
        (dminus(10), "Team Practice", "Coach Ramirez", "Excused", "Situational", 0, "Family trip"),
        (dminus(12), "Strength", "Coach Webb", "Present", "Speed & agility", 1.0, ""),
        (dminus(14), "Scrimmage", "Coach Ramirez", "Present", "Live at-bats", 2.0, "2 hits"),
        (dminus(16), "Batting Cage", "Coach Webb", "Present", "Opposite field", 1.0, ""),
        (dminus(18), "Team Practice", "Coach Ramirez", "Present", "Base running", 1.5, ""),
        (dminus(20), "Private Lesson", "Pro Swing", "Present", "Swing path", 1.0, "Great progress"),
    ]
    ws, start, end = build_log(
        wb, "Practice", "🧢", "PRACTICE PLANNER",
        "Log every session — attendance and hours roll up to your dashboard.",
        ["Date", "Type", "Coach", "Attendance", "Skills Covered", "Hours", "Notes"],
        sample, [13, 16, 16, 13, 22, 10, 22],
        text_left={3, 5, 7}, dates={1},
        validations=[("B", "PracticeTypeList"), ("D", "AttendList")], reserved=60)
    nrange(wb, "PracDate", "Practice", "A", start, end)
    nrange(wb, "PracType", "Practice", "B", start, end)
    nrange(wb, "PracAttend", "Practice", "D", start, end)
    nrange(wb, "PracHours", "Practice", "F", start, end)
    for r in range(start, end + 1):
        ws.cell(row=r, column=6).number_format = "0.0"
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Present"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Absent"'], fill=fill(RED_BG)))


# ===========================================================================
# 6 — Baseball Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 18, 14, 14, 14, 12, 3, 20, 14])
    luxe_header(ws, "I", "💰  BASEBALL BUDGET",
                "Every season cost in one place — spending, remaining & cost-per-game, automatic.")
    table_headers(ws, 4, ["Category", "Group", "Season Budget", "Spent", "Remaining", "% Used"])
    planned = {
        "Registration": 250, "League Fees": 600, "Uniforms": 180, "Cleats": 90,
        "Gloves": 200, "Bats": 350, "Helmets": 60, "Batting Gloves": 60, "Bags": 80,
        "Protective Gear": 120, "Tournaments": 1400, "Hotels": 1200, "Fuel": 600,
        "Meals": 700, "Private Lessons": 800, "Batting Cage": 300, "Team Apparel": 150,
        "Photos": 120, "Fundraising": 0, "Miscellaneous": 150,
    }
    spent = {
        "Registration": 250, "League Fees": 600, "Uniforms": 180, "Cleats": 90,
        "Gloves": 200, "Bats": 200, "Helmets": 60, "Batting Gloves": 35, "Bags": 80,
        "Protective Gear": 0, "Tournaments": 900, "Hotels": 720, "Fuel": 380,
        "Meals": 360, "Private Lessons": 480, "Batting Cage": 180, "Team Apparel": 150,
        "Photos": 0, "Fundraising": 0, "Miscellaneous": 90,
    }
    start = L0
    end = start + len(EXPENSE_CATS) - 1
    for i, cat in enumerate(EXPENSE_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        ws.cell(row=r, column=2, value=CAT_GROUP[cat]).style = "td_left"
        cp = ws.cell(row=r, column=3, value=planned[cat]); cp.style = "input"; cp.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=spent[cat]); cs.style = "input"; cs.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/C{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    ws.cell(row=total, column=2, value="").style = "th"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=total, column=6, value=f"=IFERROR(D{total}/C{total},0)")
    cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.2, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))
    nrange(wb, "BudgetCat", "Budget", "A", start, end)
    nrange(wb, "BudgetGroup", "Budget", "B", start, end)
    nrange(wb, "BudgetPlanned", "Budget", "C", start, end)
    nrange(wb, "BudgetActual", "Budget", "D", start, end)
    wb.defined_names["BudgetTotalPlanned"] = DefinedName("BudgetTotalPlanned", attr_text=f"Budget!$C${total}")
    wb.defined_names["BudgetTotalActual"] = DefinedName("BudgetTotalActual", attr_text=f"Budget!$D${total}")

    ws.cell(row=4, column=8, value="SEASON TO DATE").style = "section_gold"
    kpis = [("Total Budget", "=BudgetTotalPlanned", '"$"#,##0'),
            ("Total Spent", "=BudgetTotalActual", '"$"#,##0'),
            ("Remaining", "=BudgetTotalPlanned-BudgetTotalActual", '"$"#,##0;[Red]-"$"#,##0'),
            ("Cost / Game", '=IFERROR(BudgetTotalActual/MAX(COUNTIF(CalType,"Game"),1),0)', '"$"#,##0'),
            ("Cost / Tournament", '=IFERROR(BudgetTotalActual/MAX(COUNTIF(CalType,"Tournament"),1),0)', '"$"#,##0'),
            ("Cost / Practice", '=IFERROR(BudgetTotalActual/MAX(COUNTIF(CalType,"Practice"),1),0)', '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=8, value=lab).style = "field_label"
        c = ws.cell(row=r, column=9, value=fml); c.style = "field_value"; c.number_format = fmt

    gstart = total + 3
    merge_set(ws, f"A{gstart-1}:F{gstart-1}", "SPENDING BY GROUP", "section_gold")
    ws.cell(row=gstart, column=1, value="Group").style = "th"
    ws.cell(row=gstart, column=2, value="Spent").style = "th"
    for i, g in enumerate(EXPENSE_GROUPS):
        r = gstart + 1 + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        c = ws.cell(row=r, column=2, value=f'=SUMIF(BudgetGroup,A{r},BudgetActual)')
        c.style = "td"; c.number_format = '"$"#,##0'
    gend = gstart + len(EXPENSE_GROUPS)
    wb.defined_names["GroupLabels"] = DefinedName("GroupLabels", attr_text=f"Budget!$A${gstart+1}:$A${gend}")
    wb.defined_names["GroupVals"] = DefinedName("GroupVals", attr_text=f"Budget!$B${gstart+1}:$B${gend}")
    donut = DoughnutChart(); donut.title = "Spending by Group"; donut.height = 8; donut.width = 12
    donut.add_data(Reference(ws, min_col=2, min_row=gstart, max_row=gend), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=1, min_row=gstart + 1, max_row=gend))
    donut.dataLabels = no_labels()
    ws.add_chart(donut, "H13")
    ws.freeze_panes = "A5"


# ===========================================================================
# 7 — Equipment Command Center
# ===========================================================================
def build_equipment(wb):
    sample = [
        ("Infield Glove", "Rawlings", "11.5\"", dminus(200), 200, "Good", dplus(400), ""),
        ("Bat (BBCOR)", "Marucci", "31/28", dminus(120), 350, "Good", dplus(180), "USSSA stamp"),
        ("Backup Bat", "Easton", "30/27", dminus(300), 150, "Fair", dplus(40), "Outgrowing"),
        ("Cleats", "New Balance", "5.5", dminus(150), 90, "Replace Soon", dplus(30), "Feet growing"),
        ("Batting Helmet", "Rawlings", "M", dminus(220), 60, "Good", dplus(500), ""),
        ("Catcher's Gear", "All-Star", "Youth", dminus(400), 260, "Good", dplus(300), "Backup catcher"),
        ("Batting Gloves", "Franklin", "M", dminus(60), 35, "Replace Soon", dplus(25), "Worn palms"),
        ("Bat Bag", "Boombah", "—", dminus(300), 80, "Good", dplus(400), ""),
        ("Sliding Shorts", "McDavid", "M", dminus(150), 30, "Good", dplus(300), ""),
        ("Turf Shoes", "Mizuno", "5.5", dminus(150), 65, "Replace Soon", dplus(30), "For cages"),
        ("Baseballs (dozen)", "Rawlings", "—", dminus(30), 45, "New", dplus(120), "Practice"),
        ("Sunglasses", "Oakley", "—", dminus(90), 90, "Good", dplus(400), ""),
    ]
    ws, start, end = build_log(
        wb, "Equipment", "🧤", "EQUIPMENT COMMAND CENTER",
        "Track every piece — gear due for replacement in 45 days flags automatically.",
        ["Equipment", "Brand", "Size", "Purchased", "Price", "Condition", "Replace By", "Notes"],
        sample, [16, 14, 9, 13, 11, 14, 13, 22],
        text_left={1, 8}, dates={4, 7}, money={5},
        validations=[("F", "EquipStatusList")], reserved=40)
    nrange(wb, "EquipName", "Equipment", "A", start, end)
    nrange(wb, "EquipPrice", "Equipment", "E", start, end)
    nrange(wb, "EquipCond", "Equipment", "F", start, end)
    nrange(wb, "EquipReplace", "Equipment", "G", start, end)
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'AND($G{start}<>"",$G{start}<=TODAY()+45)'], fill=fill(WARN_BG)))
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Replace Now"'], fill=fill(RED_BG)))
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"New"'], fill=fill(MINT_BG)))


# ===========================================================================
# 8 — Bat Inventory
# ===========================================================================
def build_bats(wb):
    build_log(
        wb, "Bats", "🏏", "BAT INVENTORY",
        "Track every bat — length, drop, certification, usage hours & condition.",
        ["Brand / Model", "Length", "Weight", "Drop", "Material", "Purchased", "Cost", "Cert", "Usage Hrs", "Condition"],
        [
            ("Marucci CAT9", "31\"", "28 oz", "-3", "Alloy", dminus(120), 350, "BBCOR", 42, "Good"),
            ("Easton Ghost", "30\"", "27 oz", "-3", "Composite", dminus(300), 300, "USSSA", 60, "Fair"),
            ("DeMarini Voodoo", "31\"", "28 oz", "-3", "Hybrid", dminus(60), 320, "BBCOR", 12, "Good"),
            ("Louisville Solo", "30\"", "20 oz", "-10", "Alloy", dminus(500), 120, "USA", 90, "Retired"),
            ("Axe Elite (backup)", "31\"", "28 oz", "-3", "Alloy", dminus(200), 200, "BBCOR", 30, "Good"),
        ],
        [22, 10, 10, 8, 12, 13, 11, 10, 11, 12],
        text_left={1}, dates={6}, money={7}, ints={9}, reserved=30)


# ===========================================================================
# 9 — Glove Care Log
# ===========================================================================
def build_gloves(wb):
    build_log(
        wb, "Gloves", "🧴", "GLOVE CARE LOG",
        "Keep every glove game-ready — conditioning, re-lacing and repairs.",
        ["Brand / Model", "Position", "Purchased", "Last Conditioned", "Re-Laced", "Repairs", "Notes"],
        [
            ("Rawlings HOH", "Infield / SS", dminus(200), dminus(20), "No", "—", "Break in pocket more"),
            ("Wilson A2000", "Outfield", dminus(400), dminus(35), "Yes", "Web re-laced", ""),
            ("Mizuno Pro", "Pitcher", dminus(300), dminus(15), "No", "—", "Closed web"),
            ("Easton (backup)", "1B mitt", dminus(500), dminus(60), "No", "Needs oil", "Stiff"),
        ],
        [22, 14, 13, 15, 11, 16, 22],
        text_left={1, 6, 7}, dates={3, 4},
        validations=[("E", "YesNoList")], reserved=25)


# ===========================================================================
# 10 — Player Statistics (batting)
# ===========================================================================
def build_stats(wb):
    ws = wb.create_sheet("Stats")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 16, 7, 7, 7, 7, 7, 7, 7, 7, 7, 7, 4, 16, 12])
    luxe_header(ws, "O", "📊  PLAYER STATISTICS",
                "Log every game — AVG, OBP, SLG & OPS calculate themselves all season.")
    hdr = ["Date", "Opponent", "AB", "H", "R", "RBI", "2B", "3B", "HR", "BB", "K", "SB"]
    table_headers(ws, 4, hdr)
    games = [
        (dminus(13), "Storm", 4, 2, 1, 1, 1, 0, 0, 0, 1, 1),
        (dminus(10), "Cobras", 3, 1, 1, 0, 0, 0, 0, 1, 0, 0),
        (dminus(6), "Rockets", 4, 2, 2, 3, 0, 0, 1, 0, 1, 0),
        (dminus(3), "Titans", 3, 1, 0, 1, 1, 0, 0, 1, 1, 1),
    ]
    start = L0
    end = start + 30 - 1
    for i, row in enumerate(games):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, 12, text_left={2}, dates={1}, ints=set(range(3, 13)))
    total = end + 1
    ws.cell(row=total, column=1, value="SEASON").style = "th"
    ws.cell(row=total, column=2, value="Totals").style = "th"
    for col in range(3, 13):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE)
    nrange(wb, "StatAB", "Stats", "C", start, end)
    nrange(wb, "StatH", "Stats", "D", start, end)
    nrange(wb, "StatBB", "Stats", "J", start, end)
    wb.defined_names["StatGamesRange"] = DefinedName("StatGamesRange", attr_text=f"Stats!$A${start}:$A${end}")

    # rate-stat summary block
    ws.cell(row=4, column=14, value="SLASH LINE").style = "section_gold"
    tb = f"(SUM({get_column_letter(4)}{start}:{get_column_letter(4)}{end})" \
         f"-SUM(G{start}:G{end})-SUM(H{start}:H{end})-SUM(I{start}:I{end}))" \
         f"+2*SUM(G{start}:G{end})+3*SUM(H{start}:H{end})+4*SUM(I{start}:I{end})"
    rate = [
        ("AVG", f"=IFERROR(SUM(D{start}:D{end})/SUM(C{start}:C{end}),0)"),
        ("OBP", f"=IFERROR((SUM(D{start}:D{end})+SUM(J{start}:J{end}))/(SUM(C{start}:C{end})+SUM(J{start}:J{end})),0)"),
        ("SLG", f"=IFERROR(({tb})/SUM(C{start}:C{end}),0)"),
        ("OPS", "=N15+N16"),
    ]
    for i, (lab, fml) in enumerate(rate):
        r = 5 + i
        ws.cell(row=r, column=14, value=lab).style = "field_label"
        c = ws.cell(row=r, column=15, value=fml); c.style = "field_value"; c.number_format = ".000"
    # fix OPS reference (OBP row + SLG row = rows 6 & 7 -> N6+N7)
    ws["O8"].value = "=O6+O7"

    # counting-stat block for a chart
    ws.cell(row=10, column=14, value="KEY TOTALS").style = "section_gold"
    tots = [("Hits", f"=SUM(D{start}:D{end})"), ("RBIs", f"=SUM(F{start}:F{end})"),
            ("Runs", f"=SUM(E{start}:E{end})"), ("SB", f"=SUM(L{start}:L{end})")]
    for i, (lab, fml) in enumerate(tots):
        r = 11 + i
        ws.cell(row=r, column=14, value=lab).style = "td_left"
        ws.cell(row=r, column=15, value=fml).style = "td"
    wb.defined_names["StatTotLabels"] = DefinedName("StatTotLabels", attr_text="Stats!$N$11:$N$14")
    wb.defined_names["StatTotVals"] = DefinedName("StatTotVals", attr_text="Stats!$O$11:$O$14")
    ws.freeze_panes = "A5"


# ===========================================================================
# 11 — Pitching Tracker
# ===========================================================================
def build_pitching(wb):
    ws = wb.create_sheet("Pitching")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 16, 10, 11, 11, 9, 9, 9, 11, 4, 16, 12])
    luxe_header(ws, "L", "🥎  PITCHING TRACKER",
                "Track outings, pitch counts & ERA — with your own pitch-count limit alerts.")
    hdr = ["Date", "Opponent", "IP", "Pitches", "Strikes", "BB", "K", "ER", "Strike %"]
    table_headers(ws, 4, hdr)
    outings = [
        (dminus(13), "Storm", 3, 52, 34, 1, 4, 1),
        (dminus(6), "Rockets", 2, 41, 25, 2, 2, 2),
        (dminus(3), "Titans", 4, 68, 45, 1, 6, 1),
    ]
    start = L0
    end = start + 25 - 1
    for i, row in enumerate(outings):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, 8, text_left={2}, dates={1}, ints={4, 5, 6, 7, 8})
    for r in range(start, end + 1):
        ws.cell(row=r, column=3).number_format = "0.0"
        c = ws.cell(row=r, column=9, value=f'=IF(D{r}="","",E{r}/D{r})'); c.style = "td"; c.number_format = "0%"
        c.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
    # pitch-count warning
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        FormulaRule(formula=[f'AND($D{start}<>"",$D{start}>PitchLimit)'], fill=fill(RED_BG)))
    nrange(wb, "PitchIP", "Pitching", "C", start, end)
    nrange(wb, "PitchCount", "Pitching", "D", start, end)
    nrange(wb, "PitchER", "Pitching", "H", start, end)
    # summary
    ws.cell(row=4, column=11, value="SEASON").style = "section_gold"
    summ = [("Innings", f"=SUM(C{start}:C{end})", "0.0"),
            ("Strikeouts", f"=SUM(G{start}:G{end})", "0"),
            ("Walks", f"=SUM(F{start}:F{end})", "0"),
            ("ERA", f"=IFERROR(SUM(H{start}:H{end})/SUM(C{start}:C{end})*GameInnings,0)", "0.00"),
            ("Last pitch count", f"=IFERROR(LOOKUP(2,1/(D{start}:D{end}<>\"\"),D{start}:D{end}),0)", "0")]
    for i, (lab, fml, fmt) in enumerate(summ):
        r = 5 + i
        ws.cell(row=r, column=11, value=lab).style = "field_label"
        c = ws.cell(row=r, column=12, value=fml); c.style = "field_value"; c.number_format = fmt
    line = LineChart(); line.title = "Pitch Count by Outing"; line.height = 7.5; line.width = 12
    line.add_data(Reference(ws, min_col=4, min_row=4, max_row=start + len(outings) - 1), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=2, min_row=start, max_row=start + len(outings) - 1))
    line.legend = None
    ws.add_chart(line, "K11")
    ws.freeze_panes = "A5"


# ===========================================================================
# 12 — Tournament Command Center
# ===========================================================================
def build_tournaments(wb):
    build_log(
        wb, "Tournaments", "🏆", "TOURNAMENT COMMAND CENTER",
        "Plan every tournament — venue, hotel, fees, schedule & checklists.",
        ["Tournament", "Type", "Location", "Start", "End", "Hotel", "Entry Fee", "Status", "Notes"],
        [
            ("Summer Slam", "Bracket", "Capital City", dplus(14), dplus(16), "Marriott Downtown", 450, "Booked", "3 nights, 5 games"),
            ("Independence Classic", "Pool Play", "Lakeside", dplus(40), dplus(42), "Lakeside Inn", 400, "Planned", "Need hotel"),
            ("Fall Championship", "Championship", "Metro City", dplus(75), dplus(78), "TBD", 500, "Planned", ""),
            ("Labor Day Showcase", "Showcase", "Home", dplus(95), dplus(97), "— (home)", 350, "Planned", "No travel"),
        ],
        [20, 14, 16, 13, 13, 18, 11, 12, 20],
        text_left={1, 3, 6, 9}, dates={4, 5}, money={7},
        validations=[("B", "TourneyTypeList")], reserved=30)


# ===========================================================================
# 13 — Travel Planner
# ===========================================================================
def build_travel(wb):
    sample = [
        ("Summer Slam", "Drive", dplus(14), "Capital City", 240, 105, 45, 20, 720, "Booked"),
        ("Cobras @ North", "Drive", dplus(9), "North Diamond", 55, 24, 8, 0, 0, "Planned"),
        ("Independence Classic", "Drive", dplus(40), "Lakeside", 190, 82, 30, 15, 360, "Planned"),
        ("Fall Championship", "Fly", dplus(75), "Metro City", 0, 0, 55, 0, 640, "Planned"),
    ]
    ws, start, end = build_log(
        wb, "Travel", "✈", "TRAVEL PLANNER",
        "Track miles, fuel, tolls & lodging for every road trip — costs total themselves.",
        ["Trip", "Mode", "Date", "Destination", "Miles", "Fuel", "Tolls", "Parking", "Lodging", "Status"],
        sample, [20, 10, 13, 18, 11, 11, 11, 11, 12, 12],
        text_left={1, 4}, dates={3}, money={6, 7, 8, 9}, ints={5}, reserved=40)
    nrange(wb, "TravelMiles", "Travel", "E", start, end)
    nrange(wb, "TravelFuel", "Travel", "F", start, end)
    nrange(wb, "TravelLodging", "Travel", "I", start, end)
    ws.cell(row=4, column=12, value="TOTALS").style = "section_gold"
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 12
    tot = [("Total miles", "=SUM(TravelMiles)", "0"),
           ("Fuel + tolls", "=SUM(TravelFuel)+SUM(Travel!G5:G44)", '"$"#,##0'),
           ("Lodging", "=SUM(TravelLodging)", '"$"#,##0'),
           ("Trip cost", "=SUM(Travel!F5:I44)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(tot):
        r = 5 + i
        ws.cell(row=r, column=12, value=lab).style = "field_label"
        c = ws.cell(row=r, column=13, value=fml); c.style = "field_value"; c.number_format = fmt


# ===========================================================================
# 14 — Team Roster
# ===========================================================================
def build_roster(wb):
    build_log(
        wb, "Roster", "📋", "TEAM ROSTER",
        "The whole team in one directory — players, parents, contacts & volunteer roles.",
        ["#", "Player", "Position", "Parent", "Phone", "Email", "Birthday", "Volunteer Role"],
        [
            (7, "Cody Reyes", "SS", "Dana Reyes", "(555) 210-7788", "dreyes@email.com", dt.date(2013, 5, 8), "Team Parent"),
            (12, "Mason Webb", "P / 1B", "Luke Webb", "(555) 332-1190", "lwebb@email.com", dt.date(2013, 2, 19), "Pitch Count"),
            (3, "Diego Ruiz", "2B", "Ana Ruiz", "(555) 778-3300", "aruiz@email.com", dt.date(2013, 9, 1), "Scorekeeper"),
            (22, "Eli Carter", "C", "Sam Carter", "(555) 909-2020", "scarter@email.com", dt.date(2012, 11, 30), "Dugout Helper"),
            (9, "Noah Kim", "CF", "Grace Kim", "(555) 661-2048", "gkim@email.com", dt.date(2013, 6, 14), "Snack"),
            (15, "Owen Bell", "3B", "Pat Bell", "(555) 443-7766", "pbell@email.com", dt.date(2013, 4, 3), "Field Prep"),
        ],
        [6, 18, 12, 18, 16, 22, 13, 16],
        text_left={2, 4, 6, 8}, dates={7}, ints={1},
        validations=[("C", "PositionList"), ("H", "VolRoleList")], reserved=40)


# ===========================================================================
# 15 — Snack & Volunteer Schedule
# ===========================================================================
def build_volunteer(wb):
    sample = [
        (dplus(2), "vs Bandits", "Reyes", "Snack", 2, "No", "Orange slices + water"),
        (dplus(5), "vs Rockets", "Webb", "Scorekeeper", 2, "No", ""),
        (dplus(9), "vs Cobras", "Ruiz", "Pitch Count", 2, "No", ""),
        (dplus(12), "vs Storm", "Kim", "Snack", 2, "No", "Open — sign up"),
        (dplus(14), "Summer Slam", "Bell", "Dugout Helper", 6, "No", "Tournament"),
        (dminus(3), "vs Titans", "Carter", "Snack", 2, "Yes", "Done"),
        (dminus(6), "vs Rockets", "Reyes", "Field Prep", 2, "Yes", "Done"),
    ]
    ws, start, end = build_log(
        wb, "Volunteers", "🙋", "SNACK & VOLUNTEER SCHEDULE",
        "Coordinate snack, scorekeeping, pitch count & team-parent duties — hours roll up.",
        ["Date", "Game / Event", "Family", "Role", "Hours", "Done?", "Notes"],
        sample, [13, 20, 14, 16, 10, 10, 24],
        text_left={2, 3, 7}, dates={1}, ints={5},
        validations=[("D", "VolRoleList"), ("F", "YesNoList")], reserved=40)
    nrange(wb, "VolHours", "Volunteers", "E", start, end)
    nrange(wb, "VolDone", "Volunteers", "F", start, end)
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))
    ws.cell(row=4, column=9, value="TOTALS").style = "section_gold"
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 10
    ws.cell(row=5, column=9, value="Volunteer hours").style = "field_label"
    ws.cell(row=5, column=10, value="=SUM(VolHours)").style = "field_value"
    ws.cell(row=6, column=9, value="Open slots").style = "field_label"
    ws.cell(row=6, column=10, value='=COUNTIF(VolDone,"No")').style = "field_value"


# ===========================================================================
# 16 — Player Development
# ===========================================================================
def build_development(wb):
    sample = [
        ("Hitting", "Strong", 0.72, "Great contact", "Drive the ball more"),
        ("Fielding", "Strong", 0.7, "Soft hands at SS", "Quicker transfer"),
        ("Throwing", "Average", 0.58, "Accurate", "Add arm strength"),
        ("Base Running", "Strong", 0.75, "Reads pitchers well", "First-step quickness"),
        ("Speed", "Average", 0.55, "Solid home-to-first", "Sprint work"),
        ("Agility", "Average", 0.6, "Good lateral range", "Ladder drills"),
        ("Strength", "Developing", 0.45, "Building core", "Off-season program"),
        ("Baseball IQ", "Strong", 0.72, "Knows situations", "Study pitch counts"),
    ]
    ws, start, end = build_log(
        wb, "Development", "📈", "PLAYER DEVELOPMENT",
        "Rate each skill, track progress and capture coach feedback all season.",
        ["Skill", "Level", "Progress", "Coach Feedback", "Next Focus"],
        sample, [18, 14, 12, 30, 28],
        text_left={4, 5}, pcts={3},
        validations=[("B", "SkillLevelList")], reserved=20)
    nrange(wb, "DevSkill", "Development", "A", start, end)
    nrange(wb, "DevProgress", "Development", "C", start, end)
    ws.conditional_formatting.add(
        f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))
    bar = BarChart(); bar.type = "bar"; bar.title = "Skill Progress"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    bar.legend = None
    ws.add_chart(bar, "G5")


# ===========================================================================
# 17 — Nutrition & Hydration
# ===========================================================================
def build_nutrition(wb):
    ws = wb.create_sheet("Nutrition")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 34, 4, 22, 30, 2])
    luxe_header(ws, "G", "🥗  NUTRITION & HYDRATION",
                "Fuel for performance — game-day, tournament & recovery meals + hydration.")
    blocks = [
        ("GAME-DAY FUEL", [
            ("3–4 hrs before", "Chicken + rice + veg"),
            ("1 hr before", "Banana, granola bar, water"),
            ("During game", "Water + electrolytes"),
            ("After game", "Protein + carbs within 45 min")]),
        ("TOURNAMENT WEEKEND", [
            ("Cooler packing", "Sandwiches, fruit, water, jerky"),
            ("Between games", "Light snacks, not heavy"),
            ("Hotel breakfast", "Eggs, oatmeal, fruit"),
            ("Hydration", "Water bottle refilled all day")]),
        ("RECOVERY SNACKS", [
            ("Post-practice", "Chocolate milk / smoothie"),
            ("High protein", "Greek yogurt, eggs, nut butter"),
            ("Quick carbs", "Fruit, pretzels, rice cakes"),
            ("Avoid", "Sugary sodas & fried food")]),
        ("HYDRATION & EXTRAS", [
            ("Daily water", "Half body-weight in oz"),
            ("Hot days", "Add electrolytes"),
            ("Supplements", "Per pediatrician only"),
            ("Grocery list", "See Budget / shopping")]),
    ]
    for bi, (title, items) in enumerate(blocks):
        cl, cr = (2, 3) if bi % 2 == 0 else (5, 6)
        base = 5 if bi < 2 else 13
        merge_set(ws, f"{get_column_letter(cl)}{base}:{get_column_letter(cr)}{base}", title, "section_gold")
        for i, (a, b) in enumerate(items):
            r = base + 1 + i
            ws.cell(row=r, column=cl, value=a).style = "field_label"
            ws.cell(row=r, column=cr, value=b).style = "field_value"
            ws.row_dimensions[r].height = 22


# ===========================================================================
# 18 — Medical Center
# ===========================================================================
def build_medical(wb):
    build_log(
        wb, "Medical", "🏥", "MEDICAL CENTER",
        "Injuries, appointments, recovery & insurance claims — organized and private.",
        ["Date", "Type", "Issue / Visit", "Provider", "Status", "Follow-Up", "Cost", "Claim #", "Notes"],
        [
            (dminus(50), "Injury", "Jammed thumb (sliding)", "Dr. Lee", "Recovered", "—", 0, "—", "Iced, taped"),
            (dminus(35), "Checkup", "Sports physical", "Dr. Lee", "Cleared", "Next yr", 0, "—", "Cleared to play"),
            (dminus(20), "PT", "Shoulder maintenance", "PeakPhysio", "Ongoing", "Weekly", 180, "CLM-3310", "Pitching prehab"),
            (dminus(5), "Dental", "Mouthguard", "Smile Dental", "Complete", "—", 45, "—", ""),
        ],
        [13, 12, 22, 16, 13, 12, 11, 12, 20],
        text_left={3, 9}, dates={1}, money={7}, reserved=40)


# ===========================================================================
# 19 — Packing Checklist
# ===========================================================================
def build_packing(wb):
    ws = wb.create_sheet("Packing")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 4, 30, 4, 4, 30, 4, 4, 30, 2])
    luxe_header(ws, "J", "🎒  PACKING CHECKLIST",
                "Never forget the gear again — ready-made lists for every kind of day.")
    lists = [
        ("GAME-DAY BAG", ["Glove", "Bat(s)", "Batting helmet", "Batting gloves", "Cleats",
                          "Turf shoes", "Uniform + belt", "Hat", "Water + snacks",
                          "Sunscreen", "Eye black", "Sunglasses"]),
        ("DOUBLE HEADER / CAGE", ["Everything in game bag", "2nd uniform / dry shirt", "Extra socks",
                                  "More water", "Cooler + lunch", "Chairs", "Bug spray",
                                  "Phone charger", "Cash for concessions"]),
        ("TOURNAMENT WEEKEND", ["Full game bag", "3+ uniform sets", "Extra cleats & socks",
                                "Cooler + meals", "Toiletries", "Hotel confirmation",
                                "First-aid kit", "Team apparel", "Chargers", "Schedule printout",
                                "Sunscreen + hats", "Camping chairs"]),
    ]
    cols = [(2, 3), (5, 6), (8, 9)]
    for li, (title, items) in enumerate(lists):
        cbox, ctext = cols[li]
        merge_set(ws, f"{get_column_letter(cbox)}5:{get_column_letter(ctext)}5", title, "section_gold")
        ws.row_dimensions[5].height = 22
        for i, it in enumerate(items):
            r = 6 + i
            cb = ws.cell(row=r, column=cbox, value="☐"); cb.alignment = Alignment(horizontal="center")
            cb.font = Font(size=12, color=ACCENT); cb.border = BOX
            ct = ws.cell(row=r, column=ctext, value=it); ct.style = "td_left"
            if i % 2:
                cb.fill = fill(MUTED_ROW); ct.fill = fill(MUTED_ROW)


# ===========================================================================
# 20 — Fundraising Tracker
# ===========================================================================
def build_fundraising(wb):
    sample = [
        ("Car wash", "Event", dminus(20), 0, 580, "Team", "Great turnout"),
        ("Local sponsor — Joe's Auto", "Sponsorship", dminus(35), 0, 500, "Manager", "Banner + patch"),
        ("Discount card sales", "Product Sales", dminus(10), 150, 900, "All families", "Sold 300 cards"),
        ("Raffle night", "Event", dplus(20), 0, 0, "Committee", "Tickets printing"),
        ("Restaurant night", "Event", dminus(5), 0, 320, "Team Parent", "15% of sales"),
        ("Volunteer hours (concessions)", "Volunteer", dminus(2), 0, 0, "Reyes", "8 / 20 hrs done"),
    ]
    ws, start, end = build_log(
        wb, "Fundraising", "💵", "FUNDRAISING TRACKER",
        "Offset the season's cost — track sponsors, sales, donations & volunteer hours.",
        ["Activity", "Type", "Date", "Cost", "Raised", "Lead", "Notes"],
        sample, [26, 14, 13, 11, 11, 16, 22],
        text_left={1, 6, 7}, dates={3}, money={4, 5}, reserved=40)
    nrange(wb, "FundRaised", "Fundraising", "E", start, end)
    ws.cell(row=4, column=9, value="TOTALS").style = "section_gold"
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 12
    ws.cell(row=5, column=9, value="Total raised").style = "field_label"
    c = ws.cell(row=5, column=10, value="=SUM(FundRaised)"); c.style = "field_value"; c.number_format = '"$"#,##0'
    ws.cell(row=6, column=9, value="Offsets budget").style = "field_label"
    c = ws.cell(row=6, column=10, value="=IFERROR(SUM(FundRaised)/BudgetTotalPlanned,0)"); c.style = "field_value"; c.number_format = "0%"


# ===========================================================================
# 21 — Photo & Memory Gallery (image placeholders)
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "📸  PHOTO & MEMORY GALLERY",
                "Keep the season's best moments — paste photos with captions & dates.")
    sections = ["Team Photo", "Action Shot", "Tournament Win", "Player Award",
                "Championship", "Family Memory"]
    top0 = 5
    card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th")
        ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+5}", "📷\nPaste photo here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 6):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 6, column=col, value="Caption").style = "field_label"
        merge_set(ws, f"{M}{row+6}:{R}{row+6}", "", "field_value")
        ws.cell(row=row + 7, column=col, value="Date / Event").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")


# ===========================================================================
# 22 — Baseball Goals
# ===========================================================================
def build_goals(wb):
    sample = [
        ("Bat .350+ this season", "Season", dplus(120), 0.55, "On Track", "Currently .429"),
        ("Cut strikeouts in half", "Skill", dplus(60), 0.5, "On Track", ""),
        ("Clean up SS throws", "Skill", dplus(45), 0.6, "On Track", ""),
        ("Sub-7.0 sixty time", "Fitness", dplus(90), 0.4, "On Track", "Sprint work"),
        ("Make the all-star team", "Team", dplus(75), 0.5, "On Track", ""),
        ("Honor roll all year", "Academic", dplus(120), 0.8, "On Track", "Balance school"),
        ("Be a great teammate", "Personal", dplus(120), 0.85, "On Track", ""),
    ]
    ws, start, end = build_log(
        wb, "Goals", "🎯", "BASEBALL GOALS",
        "Season, skill, fitness & academic goals — progress bars keep them in view.",
        ["Goal", "Category", "Target Date", "Progress", "Status", "Notes"],
        sample, [30, 14, 14, 12, 14, 24],
        text_left={1, 6}, dates={3}, pcts={4}, reserved=30)
    nrange(wb, "GoalName", "Goals", "A", start, end)
    nrange(wb, "GoalProgress", "Goals", "D", start, end)
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))


# ===========================================================================
# 23 — Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 18, 3, 16, 12, 12, 12, 12, 12, 2])
    luxe_header(ws, "L", "📈  ANALYTICS DASHBOARD",
                "The season by the numbers — readiness scores, costs & a Baseball Readiness Score.")

    merge_set(ws, "B5:D5", "FAMILY READINESS SCORES", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Budget Health", '=IFERROR(1-BudgetTotalActual/BudgetTotalPlanned,0)'),
        ("Attendance", '=IFERROR(COUNTIF(PracAttend,"Present")/MAX(COUNTA(PracDate),1),0)'),
        ("Equipment Ready", '=IFERROR(1-SUMPRODUCT((EquipReplace<>"")*(EquipReplace<=TODAY()+45))/MAX(COUNTA(EquipName),1),0)'),
        ("Skill Progress", '=IFERROR(AVERAGE(DevProgress),0)'),
        ("Goal Progress", '=IFERROR(AVERAGE(GoalProgress),0)'),
        ("Season Complete", '=IFERROR(COUNTIFS(CalType,"Game",CalDate,"<"&TODAY())/MAX(COUNTIF(CalType,"Game"),1),0)'),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4,
                value=f'=IF(C{r}>=0.75,"Great",IF(C{r}>=0.4,"On Track","Needs Work"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(
        f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF" + WARN_BG,
                       mid_type="num", mid_value=0.5, mid_color="FFFFF3CD",
                       end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))

    merge_set(ws, "F5:H5", "BASEBALL READINESS SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(
                top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "A blend of budget, attendance, gear, skills & schedule.", "subtitle")
    ws["F10"].fill = fill(IVORY)

    merge_set(ws, "F12:H12", "MONTHLY EXPENSES ($)", "section")
    ws.cell(row=13, column=6, value="Month").style = "th"
    ws.cell(row=13, column=7, value="Spent").style = "th"
    trend = [("Mar", 820), ("Apr", 1140), ("May", 980), ("Jun", 760), ("Jul", 1020), ("Aug", 235)]
    for i, (m, h) in enumerate(trend):
        r = 14 + i
        ws.cell(row=r, column=6, value=m).style = "td_left"
        c = ws.cell(row=r, column=7, value=h); c.style = "td"; c.number_format = '"$"#,##0'
    line = LineChart(); line.title = "Monthly Baseball Expenses"; line.height = 7.5; line.width = 13
    line.add_data(Reference(ws, min_col=7, min_row=13, max_row=13 + len(trend)), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=6, min_row=14, max_row=13 + len(trend)))
    line.legend = None
    ws.add_chart(line, "F21")

    bar = BarChart(); bar.type = "bar"; bar.title = "Readiness by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None
    ws.add_chart(bar, "B17")


# ===========================================================================
# 1 — Executive Baseball Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2])
    ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ⚾  BASEBALL FAMILY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2",
              "  Schedule, budget, equipment, stats, travel & development — your whole season, automatically organized.",
              "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    row1 = [
        ("DAYS TO NEXT GAME", '=IF(COUNTIFS(CalType,"Game",CalDate,">="&TODAY())=0,0,MINIFS(CalDate,CalType,"Game",CalDate,">="&TODAY())-TODAY())', "days"),
        ("GAMES / WEEK", '=COUNTIFS(CalType,"Game",CalDate,">="&TODAY(),CalDate,"<="&TODAY()+7)', "num"),
        ("PRACTICES / WEEK", '=COUNTIFS(CalType,"Practice",CalDate,">="&TODAY(),CalDate,"<="&TODAY()+7)', "num"),
        ("NEXT TOURNAMENT", '=IF(COUNTIFS(CalType,"Tournament",CalDate,">="&TODAY())=0,0,MINIFS(CalDate,CalType,"Tournament",CalDate,">="&TODAY())-TODAY())', "days"),
        ("MONTHLY BUDGET", "=MonthlyBudget", "money"),
        ("BUDGET LEFT", "=BudgetTotalPlanned-BudgetTotalActual", "money"),
    ]
    row2 = [
        ("EQUIP. ALERTS", '=SUMPRODUCT((EquipReplace<>"")*(EquipReplace<=TODAY()+45))', "num"),
        ("GAMES PLAYED", '=COUNTIFS(CalType,"Game",CalDate,"<"&TODAY())', "num"),
        ("TRAVEL MILES", "=SUM(TravelMiles)", "num"),
        ("VOLUNTEER HRS", "=SUM(VolHours)", "num"),
        ("DEV PROGRESS", "=IFERROR(AVERAGE(DevProgress),0)", "pct"),
        ("SEASON COMPLETE", '=IFERROR(COUNTIFS(CalType,"Game",CalDate,"<"&TODAY())/MAX(COUNTIF(CalType,"Game"),1),0)', "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)

    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "MONEY & SEASON AT A GLANCE", "section_gold")

    d1 = DoughnutChart(); d1.title = "Baseball Spending"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Budget"], min_col=2, min_row=29, max_row=33), titles_from_data=False)
    d1.set_categories(Reference(wb["Budget"], min_col=1, min_row=29, max_row=33))
    d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")

    d2 = DoughnutChart(); d2.title = "Games vs Practices"; d2.height = 8.2; d2.width = 11.5
    d2.add_data(Reference(wb["Calendar"], min_col=11, min_row=5, max_row=8), titles_from_data=False)
    d2.set_categories(Reference(wb["Calendar"], min_col=10, min_row=5, max_row=8))
    d2.dataLabels = no_labels()
    ws.add_chart(d2, "H12")

    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "DEVELOPMENT & RESULTS", "section_gold")

    s1 = BarChart(); s1.type = "bar"; s1.title = "Skill Progress"; s1.height = 8.2; s1.width = 11.5
    s1.add_data(Reference(wb["Development"], min_col=3, min_row=4, max_row=12), titles_from_data=True)
    s1.set_categories(Reference(wb["Development"], min_col=1, min_row=5, max_row=12))
    s1.legend = None
    ws.add_chart(s1, "B30")

    g1 = BarChart(); g1.type = "col"; g1.title = "Batting Totals"; g1.height = 8.2; g1.width = 11.5
    g1.add_data(Reference(wb["Stats"], min_col=15, min_row=11, max_row=14), titles_from_data=False)
    g1.set_categories(Reference(wb["Stats"], min_col=14, min_row=11, max_row=14))
    g1.legend = None
    ws.add_chart(g1, "H30")

    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47",
              "Baseball Family Command Center™ — your whole season, organized in one place. Edit anything in Settings.",
              "subtitle")


# ===========================================================================
# Build
# ===========================================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    build_settings(wb)
    build_welcome(wb)
    build_profile(wb)
    build_calendar(wb)
    build_gameday(wb)
    build_practice(wb)
    build_budget(wb)
    build_equipment(wb)
    build_bats(wb)
    build_gloves(wb)
    build_stats(wb)
    build_pitching(wb)
    build_tournaments(wb)
    build_travel(wb)
    build_roster(wb)
    build_volunteer(wb)
    build_development(wb)
    build_nutrition(wb)
    build_medical(wb)
    build_packing(wb)
    build_fundraising(wb)
    build_gallery(wb)
    build_goals(wb)
    build_analytics(wb)
    build_dashboard(wb)   # index 0

    order = ["Welcome", "Dashboard", "Player Profile", "Calendar", "Game Day",
             "Practice", "Budget", "Equipment", "Bats", "Gloves", "Stats",
             "Pitching", "Tournaments", "Travel", "Roster", "Volunteers",
             "Development", "Nutrition", "Medical", "Packing", "Fundraising",
             "Gallery", "Goals", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "Baseball_Family_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
