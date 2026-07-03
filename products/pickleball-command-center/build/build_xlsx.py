"""Build Pickleball Command Center™ — The Ultimate Pickleball Performance,
League & Tournament Management System.

17 sheets + Welcome · a premium pickleball operating system in Excel & Sheets.
Matches, tournaments, practice, skills, equipment, budget, fitness, partners,
travel, nutrition, goals, gallery & analytics — one elegant dashboard.

Run: python3 build_xlsx.py   ->  ../Pickleball_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

SKILLS = ["Serve", "Return", "Dinking", "Third Shot Drop", "Volleys", "Speed-Ups",
          "Lobs", "Court Positioning", "Footwork", "Strategy", "Communication", "Mental Game"]
MATCH_TYPES = ["Rec", "League", "Tournament", "Ladder", "Open Play"]
TOURN_LEVELS = ["2.5", "3.0", "3.5", "4.0", "4.5", "Open", "Senior"]
EXPENSE_CATS = ["Club Membership", "Court Fees", "Tournament Fees", "Lessons", "Coaching",
                "Equipment", "Apparel", "Travel", "Hotels", "Meals", "Miscellaneous"]
EQUIP_TYPES = ["Paddle", "Balls", "Shoes", "Bag", "Apparel", "Grips", "Eyewear"]
GOAL_CATS = ["Ranking", "Tournament", "Fitness", "Skill", "Practice", "Milestone"]
CLUBS = ["Riverside Pickleball Club", "Downtown Rec Center", "Sunset Courts", "Metro League"]
RESULTS = ["W", "L"]
CONDITIONS = ["New", "Excellent", "Good", "Worn", "Replace Soon"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 16 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pct": "0%", "dec": "0.0",
                        "rate": "0.00", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1))
        m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [18] * 7)
    luxe_header(ws, "K", "⚙  SETTINGS", "Set your player details once — every dashboard follows. Edit the lists to fit any level.")
    merge_set(ws, "B5:C5", "PLAYER INPUTS", "section")
    controls = [
        ("Player Name", "Jordan Reeves", None, "PlayerName"),
        ("Playing Level", "3.5", None, "PlayLevel"),
        ("Current Rating (DUPR)", 3.75, "0.00", "PlayerRating"),
        ("Age Division", "50+", None, "AgeDiv"),
        ("Home Club", "Riverside Pickleball Club", None, "HomeClub"),
        ("Monthly Budget", 180, '"$"#,##0', "MonthlyBudget"),
        ("Season Match Target", 40, "0", "SeasonTarget"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Skill", SKILLS, "SkillList"), ("F", "Match Type", MATCH_TYPES, "MatchTypeList"),
             ("G", "Tournament Level", TOURN_LEVELS, "TournLevelList"), ("H", "Expense Category", EXPENSE_CATS, "ExpenseCatList"),
             ("I", "Equipment Type", EQUIP_TYPES, "EquipTypeList"), ("J", "Goal Category", GOAL_CATS, "GoalCatList"),
             ("K", "Club", CLUBS, "ClubList")]
    merge_set(ws, "E5:K5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 21, "Result", RESULTS, "ResultList"),
             ("F", 21, "Condition", CONDITIONS, "CondList"),
             ("G", 21, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 82, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🏓  PICKLEBALL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate pickleball performance, league & tournament management system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE PICKLEBALL JOURNEY — IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Most pickleball spreadsheets only track scores. Pickleball Command Center™ brings match "
                      "tracking, player performance, tournaments, practice, skills, equipment, fitness, finances, "
                      "partners, travel, nutrition and goals into ONE premium Excel & Google Sheets dashboard — so "
                      "whether you play once a week or compete every weekend, you improve your game and stay "
                      "organized. It's not a score tracker — it's your pickleball operating system.")
    ws["B6"].style = "body"
    for r in (6, 7, 8, 9):
        ws.row_dimensions[r].height = 22
    merge_set(ws, "B11:B11", "START HERE", "section")
    steps = ["1.  Open Settings and add your name, level, rating & season target.",
             "2.  Fill in your Player Profile (division, club, coach, goals).",
             "3.  Log matches in the Match Tracker — win % calculates itself.",
             "4.  Plan practices, rate your skills & track equipment replacements.",
             "5.  Set budgets, log fitness & partners — everything rolls up automatically.",
             "6.  Watch the Executive Dashboard track wins, court time & your performance score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (a 3.5 league & tournament season) is included so you can see how everything "
               "connects — just type over it with your own. Win %, court time, skill progress, budget, partner "
               "records and the Pickleball Performance Score all update automatically. Every sheet is print-friendly "
               "and works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more time on the court — game on!", "section_gold")


# ===========================================================================
# 2 — Player Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Player Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 28, 6, 24, 24, 2])
    luxe_header(ws, "G", "🏓  PLAYER PROFILE", "Who you are on the court — the essentials, all in one place.")
    blocks = [
        ("THE PLAYER", [("Player Name", "=PlayerName"), ("Age Division", "=AgeDiv"),
                        ("Playing Level", "=PlayLevel"), ("Current Rating (DUPR)", "=PlayerRating"),
                        ("Dominant Hand", "Right"), ("Preferred Position", "Left / Even")]),
        ("TEAM & SUPPORT", [("Home Club", "=HomeClub"), ("Coach", "Coach Diaz"),
                            ("Primary Partner", "Chris P."), ("Playing Since", "2021"),
                            ("Emergency Contact", "Pat R. — (555) 210-4521"), ("Medical Notes", "None — carry inhaler")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            cv = ws.cell(row=row, column=3, value=fields[i][1]); cv.style = "field_value"
            if fields[i][0] == "Current Rating (DUPR)":
                cv.number_format = "0.00"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "SEASON GOALS", "section_gold"); ws.row_dimensions[15].height = 22
    goals = ["Reach a 4.0 DUPR rating", "Medal at a 3.5 sanctioned tournament",
             "Play 40 matches this season", "Master the third-shot drop"]
    for i, g in enumerate(goals):
        r = 16 + i
        cb = ws.cell(row=r, column=2, value="◆"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=11, color=ACCENT); cb.border = BOX
        merge_set(ws, f"C{r}:F{r}", g, "td_left")


# ===========================================================================
# 3 — Match Tracker
# ===========================================================================
def build_matches(wb):
    # (dago, opponents, partner, score, type, venue, result)
    P1, P2, P3 = "Chris P.", "Sam T.", "Alex D."
    matches = [
        (56, "Rodriguez / Lin", P1, "11-7, 11-9", "Rec", "Riverside", "W"),
        (54, "Boyd / Tran", P2, "9-11, 8-11", "League", "Downtown Rec", "L"),
        (51, "Okafor / Diaz", P1, "11-6, 11-8", "Rec", "Riverside", "W"),
        (49, "Nguyen / Park", P3, "11-9, 9-11, 11-7", "Ladder", "Sunset Courts", "W"),
        (46, "Kim / Foster", P1, "8-11, 11-9, 9-11", "League", "Downtown Rec", "L"),
        (44, "Ramos / Webb", P2, "11-5, 11-7", "Rec", "Riverside", "W"),
        (41, "Silva / Cho", P1, "11-8, 11-6", "Tournament", "Metro Complex", "W"),
        (41, "Hill / Grant", P1, "11-9, 12-10", "Tournament", "Metro Complex", "W"),
        (40, "Adler / Meyer", P1, "7-11, 9-11", "Tournament", "Metro Complex", "L"),
        (37, "Brooks / Vale", P3, "11-6, 11-9", "Rec", "Riverside", "L"),
        (35, "Ito / Sato", P1, "11-7, 11-5", "League", "Downtown Rec", "W"),
        (33, "Pope / Ruiz", P2, "11-9, 8-11, 11-8", "Ladder", "Sunset Courts", "W"),
        (30, "Cross / Blake", P1, "9-11, 11-8, 11-9", "Rec", "Riverside", "W"),
        (28, "Nash / Owens", P2, "10-12, 9-11", "League", "Downtown Rec", "L"),
        (25, "Frost / Lane", P1, "11-6, 11-8", "Rec", "Riverside", "W"),
        (23, "Bauer / Quinn", P3, "11-9, 11-7", "Ladder", "Sunset Courts", "W"),
        (20, "Shaw / Reed", P1, "8-11, 11-9, 8-11", "League", "Downtown Rec", "L"),
        (18, "Todd / Marsh", P2, "11-7, 9-11, 11-6", "Rec", "Riverside", "W"),
        (15, "Vega / Cole", P1, "11-8, 11-10", "Tournament", "Lakeside", "W"),
        (13, "Dunn / Pratt", P1, "11-9, 9-11, 11-7", "Tournament", "Lakeside", "L"),
        (10, "Ford / Beck", P1, "11-6, 11-7", "Rec", "Riverside", "W"),
        (7, "Lowe / Kerr", P3, "9-11, 8-11", "League", "Downtown Rec", "L"),
        (4, "Hart / Dean", P1, "11-8, 11-9", "Rec", "Riverside", "W"),
        (2, "Page / Munoz", P2, "11-9, 11-6", "Ladder", "Sunset Courts", "W"),
    ]
    sample = [(dminus(d), opp, pt, sc, ty, ve, res) for (d, opp, pt, sc, ty, ve, res) in matches]
    ws, start, end = build_log(
        wb, "Matches", "🎾", "MATCH TRACKER",
        "Every game logged — win %, partner records & form all calculate automatically.",
        ["Date", "Opponents", "Partner", "Score", "Match Type", "Venue", "Result", "Notes"],
        sample, [13, 20, 14, 20, 14, 15, 10, 20],
        text_left={2, 4, 8}, dates={1},
        validations=[("C", "ClubList"), ("E", "MatchTypeList"), ("G", "ResultList")], reserved=60)
    # widen headers set to 8 (build_log used 8 cols); result column highlight
    nrange(wb, "MatchDate", "Matches", "A", start, end)
    nrange(wb, "MatchPartner", "Matches", "C", start, end)
    nrange(wb, "MatchType", "Matches", "E", start, end)
    nrange(wb, "MatchResult", "Matches", "G", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"W"'], fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"L"'], fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    # results summary (feeds dashboard donut)
    ws.cell(row=4, column=10, value="Result").style = "th"
    ws.cell(row=4, column=11, value="Count").style = "th"
    ws.cell(row=5, column=10, value="Wins").style = "td_left"
    ws.cell(row=5, column=11, value='=COUNTIF(MatchResult,"W")').style = "td"
    ws.cell(row=6, column=10, value="Losses").style = "td_left"
    ws.cell(row=6, column=11, value='=COUNTIF(MatchResult,"L")').style = "td"
    set_widths(ws, [13, 20, 14, 20, 14, 15, 10, 20, 3, 12, 10])
    cell_name(wb, "ResultLabel", "Matches", "$J$5:$J$6")
    cell_name(wb, "ResultCount", "Matches", "$K$5:$K$6")


# ===========================================================================
# 4 — Tournament Manager
# ===========================================================================
def build_tournaments(wb):
    ws = wb.create_sheet("Tournaments"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 18, 13, 10, 11, 13, 11, 20])
    luxe_header(ws, "H", "🏆  TOURNAMENT MANAGER",
                "Every event tracked — registration, entry fees, placement & prize money.")
    table_headers(ws, 4, ["Tournament", "Location", "Date", "Level", "Entry Fee", "Placement", "Prize", "Notes"])
    # (name, loc, date_offset(+future/-past), level, fee, placement, prize, note)
    rows = [
        ("Spring Slam 3.5", "Metro Complex", -41, "3.5", 55, "1st", 150, "Gold — great run!"),
        ("Lakeside Open", "Lakeside", -14, "3.5", 60, "3rd", 0, "Bronze medal match"),
        ("Summer Shootout", "Riverside", -75, "3.5", 50, "1st", 200, "Undefeated"),
        ("Metro League Finals", "Downtown Rec", -28, "3.5", 45, "5th", 0, "Tough bracket"),
        ("Fall Classic", "Sunset Courts", 20, "3.5", 60, "Registered", 0, "Bring extra paddle"),
        ("Harvest Cup", "Metro Complex", 40, "4.0", 65, "Registered", 0, "Playing up a level"),
        ("Regional Championships", "State Center", 65, "3.5", 75, "Registered", 0, "Season goal event"),
    ]
    start = L0
    for i, (nm, loc, doff, lvl, fee, place, prize, note) in enumerate(rows):
        r = start + i
        d = dplus(doff) if doff >= 0 else dminus(-doff)
        ws.cell(row=r, column=1, value=nm)
        ws.cell(row=r, column=2, value=loc)
        ws.cell(row=r, column=3, value=d)
        ws.cell(row=r, column=4, value=lvl)
        ws.cell(row=r, column=5, value=fee)
        ws.cell(row=r, column=6, value=place)
        ws.cell(row=r, column=7, value=prize)
        ws.cell(row=r, column=8, value=note)
    end = start + 30 - 1
    style_rows(ws, start, end, 8, text_left={1, 2, 8}, dates={3}, money={5, 7})
    for col_letter, lst in [("D", "TournLevelList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTALS").style = "th"
    for col in (5, 7):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "TournName", "Tournaments", "A", start, end)
    nrange(wb, "TournDate", "Tournaments", "C", start, end)
    nrange(wb, "TournPlace", "Tournaments", "F", start, end)
    nrange(wb, "TournPrize", "Tournaments", "G", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"1st"'], fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Registered"'], fill=fill(WARN_BG)))


# ===========================================================================
# 5 — Practice Planner
# ===========================================================================
def build_practice(wb):
    rows = [
        (2, "Third shot drop + dinking", 2.0, "Drop targets, dink cross-court", "Softer hands, good"),
        (5, "Serve & return depth", 1.5, "Deep serve ladder, return placement", "Consistent depth"),
        (9, "Speed-ups & hands battles", 1.5, "Hand-speed drills, reset practice", "Quicker reactions"),
        (13, "Footwork & positioning", 2.0, "Split-step, kitchen line movement", "Stay balanced"),
        (18, "Third shot drop", 1.0, "100 drops, transition zone", "70% in kitchen"),
        (25, "Dinking patterns", 1.5, "Cross-court, figure-8, attack the middle", "Patience improving"),
        (34, "Serve consistency", 2.0, "Serve targets, spin serves", "Add topspin"),
        (41, "Reset & defense", 1.0, "Block volleys, reset from baseline", "Calmer under fire"),
        (48, "Match play sim", 1.5, "Games to 11, scenario drills", "Close out leads"),
        (55, "Strategy & stacking", 1.0, "Stacking, poaching signals", "Communicate more"),
    ]
    sample = [(dminus(d), foc, dur, dr, fb) for (d, foc, dur, dr, fb) in rows]
    ws, start, end = build_log(
        wb, "Practice", "🎯", "PRACTICE PLANNER",
        "Train with intent — focus areas, drills & coach feedback, session by session.",
        ["Date", "Focus Area", "Duration (hrs)", "Drills", "Coach Feedback"],
        sample, [13, 26, 15, 30, 26],
        text_left={2, 4, 5}, dates={1}, dec={3}, reserved=40)
    nrange(wb, "PracticeDate", "Practice", "A", start, end)
    nrange(wb, "PracticeHrs", "Practice", "C", start, end)


# ===========================================================================
# 6 — Skill Development
# ===========================================================================
def build_skills(wb):
    ws = wb.create_sheet("Skills"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 12, 12, 12, 12, 24])
    luxe_header(ws, "F", "📈  SKILL DEVELOPMENT",
                "Rate every part of your game — watch each skill climb from start to now.")
    table_headers(ws, 4, ["Skill", "Start", "Current", "Target", "Gain", "Focus Notes"])
    # (skill, start, current, target, note)
    data = [
        ("Serve", 6, 8, 9, "Depth + spin dialed in"),
        ("Return", 6, 7, 9, "Work on placement"),
        ("Dinking", 5, 7, 9, "Patience much better"),
        ("Third Shot Drop", 4, 6, 8, "Priority — drilling weekly"),
        ("Volleys", 6, 7, 8, "Punch volleys solid"),
        ("Speed-Ups", 5, 6, 8, "Pick better moments"),
        ("Lobs", 5, 6, 7, "Offensive lob developing"),
        ("Court Positioning", 6, 7, 9, "Get to the line faster"),
        ("Footwork", 5, 6, 8, "Split-step consistency"),
        ("Strategy", 6, 7, 9, "Attack the middle"),
        ("Communication", 7, 8, 9, "Great with Chris"),
        ("Mental Game", 5, 6, 8, "Reset after errors"),
    ]
    start = L0
    for i, (sk, st, cur, tgt, note) in enumerate(data):
        r = start + i
        ws.cell(row=r, column=1, value=sk).style = "td_left"
        for ci, v in [(2, st), (3, cur), (4, tgt)]:
            c = ws.cell(row=r, column=ci, value=v); c.style = "input" if ci == 3 else "td"; c.number_format = "0"
        ws.cell(row=r, column=5, value=f"=C{r}-B{r}").style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for cc in range(1, 7):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    end = start + len(data) - 1
    avg = end + 1
    ws.cell(row=avg, column=1, value="AVERAGE").style = "th"
    for col in (2, 3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=avg, column=col, value=f"=AVERAGE({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0.0"
    ws.freeze_panes = "A5"
    nrange(wb, "SkillName", "Skills", "A", start, end)
    nrange(wb, "SkillStart", "Skills", "B", start, end)
    nrange(wb, "SkillCurrent", "Skills", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=3, start_color="FF" + WARN_BG,
                       end_type="num", end_value=10, end_color="FF" + HIGHLIGHT))
    # chart: start vs current
    bar = BarChart(); bar.type = "bar"; bar.title = "Skill Progress — Start vs Now"; bar.height = 11; bar.width = 15
    bar.add_data(Reference(ws, min_col=2, min_row=4, max_row=end), titles_from_data=True)
    bar.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    ws.add_chart(bar, "H5")


# ===========================================================================
# 7 — Equipment Command Center
# ===========================================================================
def build_equipment(wb):
    ws = wb.create_sheet("Equipment"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 13, 16, 12, 11, 14, 13, 13, 18])
    luxe_header(ws, "I", "🎽  EQUIPMENT COMMAND CENTER",
                "Gear you can trust — track condition & get automatic replacement reminders.")
    table_headers(ws, 4, ["Item", "Type", "Brand", "Purchased", "Price", "Condition", "Replace By", "Status", "Notes"])
    # (item, type, brand, purch_dago, price, condition, replace_doff, note)
    rows = [
        ("Primary Paddle", "Paddle", "Selkirk Vanguard", 210, 220, "Worn", 25, "Grip wearing, dead spot"),
        ("Backup Paddle", "Paddle", "JOOLA Perseus", 90, 200, "Excellent", 400, "Match backup"),
        ("Court Shoes", "Shoes", "K-Swiss Express", 150, 130, "Replace Soon", 40, "Tread low"),
        ("Indoor Balls (x12)", "Balls", "Onix Fuse", 40, 30, "Good", 120, "Rotate stock"),
        ("Outdoor Balls (x12)", "Balls", "Franklin X-40", 30, 32, "Good", 90, "Cracks appear fast"),
        ("Paddle Bag", "Bag", "Selkirk Tour", 300, 90, "Excellent", 900, "Holds 4 paddles"),
        ("Overgrips (x6)", "Grips", "Gamma", 20, 15, "Good", 80, "Fresh stock on hand"),
        ("Court Eyewear", "Eyewear", "Tifosi", 260, 45, "Good", 300, "Anti-fog"),
        ("Performance Tee (x3)", "Apparel", "Nike Dri-FIT", 120, 75, "Good", 300, "Match-day kit"),
    ]
    start = L0
    for i, (item, ty, brand, pdago, price, cond, rdoff, note) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=ty)
        ws.cell(row=r, column=3, value=brand)
        ws.cell(row=r, column=4, value=dminus(pdago))
        ws.cell(row=r, column=5, value=price)
        ws.cell(row=r, column=6, value=cond)
        ws.cell(row=r, column=7, value=dplus(rdoff))
        ws.cell(row=r, column=8, value=f'=IF(G{r}="","",IF(G{r}<=TODAY()+60,"DUE SOON","OK"))')
        ws.cell(row=r, column=9, value=note)
    end = start + 30 - 1
    style_rows(ws, start, end, 9, text_left={1, 3, 9}, dates={4, 7}, money={5})
    for col_letter, lst in [("B", "EquipTypeList"), ("F", "CondList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL INVESTED").style = "th"
    c = ws.cell(row=total, column=5, value=f"=SUM(E{start}:E{end})")
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "EquipName", "Equipment", "A", start, end)
    nrange(wb, "EquipReplace", "Equipment", "G", start, end)
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"DUE SOON"'], fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Replace Soon"'], fill=fill(RED_BG)))


# ===========================================================================
# 8 — Pickleball Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 14, 14, 12, 3, 20, 14])
    luxe_header(ws, "H", "💰  PICKLEBALL BUDGET",
                "Know what the game costs — plan vs actual across every category, all year.")
    table_headers(ws, 4, ["Category", "Annual Budget", "Spent (YTD)", "Remaining", "% Used"])
    # (cat, planned, actual)
    data = [
        ("Club Membership", 480, 480), ("Court Fees", 240, 140), ("Tournament Fees", 360, 210),
        ("Lessons", 300, 180), ("Coaching", 200, 120), ("Equipment", 300, 220),
        ("Apparel", 120, 60), ("Travel", 120, 55), ("Hotels", 180, 90),
        ("Meals", 120, 70), ("Miscellaneous", 80, 40),
    ]
    start = L0; end = start + len(data) - 1
    for i, (cat, pl, ac) in enumerate(data):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=2, value=pl); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=ac); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in range(2, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})"); c.style = "td"
        c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=total, column=5, value=f"=IFERROR(C{total}/B{total},0)"); cu.style = "td"
    cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.1, color=PRIMARY, showValue=True))
    nrange(wb, "BudgetCat", "Budget", "A", start, end)
    nrange(wb, "BudgetActual", "Budget", "C", start, end)
    cell_name(wb, "BudgetTotalPlanned", "Budget", f"$B${total}")
    cell_name(wb, "BudgetTotalActual", "Budget", f"$C${total}")
    # side summary
    merge_set(ws, "G5:H5", "AT A GLANCE", "section_gold")
    rows2 = [("Annual budget", f"=B{total}", '"$"#,##0'), ("Spent YTD", f"=C{total}", '"$"#,##0'),
             ("Remaining", f"=B{total}-C{total}", '"$"#,##0'), ("Monthly target", "=MonthlyBudget", '"$"#,##0'),
             ("Prize money won", "=SUM(TournPrize)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 6 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Prize money won":
            ws.cell(row=r, column=8).fill = fill(MINT_BG)


# ===========================================================================
# 9 — Fitness & Recovery
# ===========================================================================
def build_fitness(wb):
    rows = [
        (2, "Strength — legs & core", "Strength", 45, 7.5, "Squats, lunges, planks"),
        (3, "Court cardio intervals", "Cardio", 30, 7.0, "Suicides, ladder drills"),
        (4, "Mobility & stretch", "Mobility", 20, 8.0, "Hips, shoulders, ankles"),
        (6, "Recovery day", "Recovery", 0, 8.5, "Foam roll, light walk"),
        (8, "Strength — upper body", "Strength", 40, 7.0, "Rows, presses, bands"),
        (10, "Zone 2 cardio", "Cardio", 35, 7.5, "Easy bike"),
        (12, "Yoga / balance", "Mobility", 30, 8.0, "Single-leg balance"),
        (15, "Strength — full body", "Strength", 45, 7.0, "Compound lifts"),
        (18, "Sprint intervals", "Cardio", 25, 6.5, "Left knee a little sore"),
        (22, "Recovery + massage", "Recovery", 0, 8.5, "Deep tissue"),
    ]
    sample = [(dminus(d), act, ty, mins, sleep, note) for (d, act, ty, mins, sleep, note) in rows]
    ws, start, end = build_log(
        wb, "Fitness", "💪", "FITNESS & RECOVERY",
        "Fit players win more — strength, cardio, mobility & recovery, all logged.",
        ["Date", "Activity", "Type", "Minutes", "Sleep (hrs)", "Notes"],
        sample, [13, 26, 14, 11, 12, 28],
        text_left={2, 6}, dates={1}, ints={4}, dec={5}, reserved=40)
    nrange(wb, "FitDate", "Fitness", "A", start, end)
    nrange(wb, "FitType", "Fitness", "C", start, end)


# ===========================================================================
# 10 — Doubles Partner Tracker
# ===========================================================================
def build_partners(wb):
    ws = wb.create_sheet("Partners"); ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 12, 9, 9, 10, 10, 26, 24])
    luxe_header(ws, "H", "🤝  DOUBLES PARTNER TRACKER",
                "Find your best match — records, chemistry notes & win % for every partner.")
    table_headers(ws, 4, ["Partner", "Matches", "Wins", "Losses", "Win %", "Record", "Strengths", "Areas to Improve"])
    # (partner, strengths, improve)
    data = [
        ("Chris P.", "Great hands, calm under pressure", "Poach more on returns"),
        ("Sam T.", "Aggressive net play, big serve", "Shot selection late in games"),
        ("Alex D.", "Fast, covers the court", "Third-shot consistency"),
    ]
    start = L0
    for i, (p, strong, improve) in enumerate(data):
        r = start + i
        ws.cell(row=r, column=1, value=p).style = "td_left"
        ws.cell(row=r, column=2, value=f'=C{r}+D{r}').style = "td"
        cw = ws.cell(row=r, column=3, value=f'=COUNTIFS(MatchPartner,A{r},MatchResult,"W")'); cw.style = "td"; cw.number_format = "0"
        cl = ws.cell(row=r, column=4, value=f'=COUNTIFS(MatchPartner,A{r},MatchResult,"L")'); cl.style = "td"; cl.number_format = "0"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)"); cp.style = "td"; cp.number_format = "0%"
        cr = ws.cell(row=r, column=6, value=f'=C{r}&"-"&D{r}'); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY)
        ws.cell(row=r, column=7, value=strong).style = "td_left"
        ws.cell(row=r, column=8, value=improve).style = "td_left"
        for cc in range(1, 9):
            ws.cell(row=r, column=cc).border = BOX
        if i % 2:
            for cc in range(1, 9):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    end = start + len(data) - 1
    ws.freeze_panes = "A5"
    nrange(wb, "PartnerName", "Partners", "A", start, end)
    nrange(wb, "PartnerWinPct", "Partners", "E", start, end)
    nrange(wb, "PartnerRecord", "Partners", "F", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))


# ===========================================================================
# 11 — Club & League Manager
# ===========================================================================
def build_league(wb):
    ws = wb.create_sheet("Club & League"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 22, 18, 14, 3, 22, 14, 2])
    luxe_header(ws, "I", "🏟  CLUB & LEAGUE MANAGER",
                "Stay on top of league night — schedule, standings, duties & contacts.")
    # league schedule
    merge_set(ws, "B5:E5", "LEAGUE SCHEDULE", "section"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Week", "Opponent Team", "Court / Time", "Result"], start_col=2)
    sched = [
        ("Wk 5", "Net Ninjas", "Court 3 · 6:30p", "W 2-1"),
        ("Wk 6", "Dink Dynasty", "Court 1 · 6:30p", "L 1-2"),
        ("Wk 7", "Kitchen Kings", "Court 2 · 7:00p", "W 3-0"),
        ("Wk 8", "Paddle Pushers", "Court 3 · 6:30p", "This week"),
        ("Wk 9", "Third Shot Heroes", "Court 1 · 7:00p", "—"),
        ("Wk 10", "Bangers United", "Court 2 · 6:30p", "—"),
    ]
    for i, (wk, opp, ct, res) in enumerate(sched):
        r = 7 + i
        ws.cell(row=r, column=2, value=wk).style = "td_left"
        ws.cell(row=r, column=3, value=opp).style = "td_left"
        ws.cell(row=r, column=4, value=ct).style = "td_left"
        ws.cell(row=r, column=5, value=res).style = "td"
        if i % 2:
            for cc in range(2, 6):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    # standings
    merge_set(ws, "G5:H5", "TEAM STANDINGS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.cell(row=6, column=7, value="Team").style = "th"
    ws.cell(row=6, column=8, value="W-L").style = "th"
    standings = [("Kitchen Kings", "7-1"), ("Riverside Risers ★", "6-2"),
                 ("Dink Dynasty", "5-3"), ("Net Ninjas", "4-4"), ("Paddle Pushers", "3-5")]
    for i, (tm, wl) in enumerate(standings):
        r = 7 + i
        ws.cell(row=r, column=7, value=tm).style = "td_left"
        ws.cell(row=r, column=8, value=wl).style = "td"
        if tm.endswith("★"):
            ws.cell(row=r, column=7).fill = fill(MINT_BG); ws.cell(row=r, column=8).fill = fill(MINT_BG)
        elif i % 2:
            ws.cell(row=r, column=7).fill = fill(MUTED_ROW); ws.cell(row=r, column=8).fill = fill(MUTED_ROW)
    # duties & contacts
    merge_set(ws, "B14:E14", "MY DUTIES & CLUB CONTACTS", "section_gold"); ws.row_dimensions[14].height = 22
    duties = [("Volunteer duty", "Set up nets — Wk 9"), ("League fee", "Paid ✓"),
              ("Club president", "Dana K. — (555) 330-1180"), ("League coordinator", "Marcus V. — (555) 774-2201")]
    for i, (lab, val) in enumerate(duties):
        r = 15 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        merge_set(ws, f"C{r}:E{r}", val, "field_value")


# ===========================================================================
# 12 — Travel Planner
# ===========================================================================
def build_travel(wb):
    rows = [
        ("Regional Championships", 65, "State Center", "Grand Hotel (2 nts)", "Drive · 180 mi", 320, "Book early rate"),
        ("Harvest Cup", 40, "Metro Complex", "Home", "Drive · 25 mi", 20, "Day trip"),
        ("Fall Classic", 20, "Sunset Courts", "Home", "Drive · 15 mi", 12, "Day trip"),
        ("Winter Slam (maybe)", 95, "Coastal Arena", "Seaside Inn (2 nts)", "Flight + rental", 540, "Pending registration"),
    ]
    sample = [(nm, dplus(doff), loc, hotel, transp, cost, note) for (nm, doff, loc, hotel, transp, cost, note) in rows]
    ws, start, end = build_log(
        wb, "Travel", "✈", "TRAVEL PLANNER",
        "Away tournaments made easy — hotels, transport & costs in one place.",
        ["Event", "Date", "Venue", "Lodging", "Transport", "Est. Cost", "Notes"],
        sample, [24, 13, 18, 20, 18, 12, 22],
        text_left={1, 3, 4, 5, 7}, dates={2}, money={6}, reserved=20)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    c = ws.cell(row=total, column=6, value=f"=SUM(F{start}:F{end})")
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'


# ===========================================================================
# 13 — Nutrition & Hydration
# ===========================================================================
def build_nutrition(wb):
    ws = wb.create_sheet("Nutrition"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 30, 4, 22, 26, 2])
    luxe_header(ws, "G", "🥗  NUTRITION & HYDRATION",
                "Fuel the win — match-day meals, hydration, recovery & supplements.")
    blocks = [
        ("MATCH-DAY FUEL", 5, [
            ("2–3 hrs before", "Oatmeal, banana, honey, coffee"),
            ("30 min before", "Half banana + electrolyte drink"),
            ("Between matches", "Trail mix, dates, water"),
            ("Post-match", "Protein shake + fruit")]),
        ("HYDRATION PLAN", 11, [
            ("Daily target", "3–4 L water"),
            ("On court", "500 ml + electrolytes / hour"),
            ("Hot days", "Add sodium tabs"),
            ("Recovery", "16 oz per lb lost")]),
    ]
    for title, top, items in blocks:
        merge_set(ws, f"B{top}:C{top}", title, "section_gold"); ws.row_dimensions[top].height = 22
        for i, (lab, val) in enumerate(items):
            r = top + 1 + i
            ws.cell(row=r, column=2, value=lab).style = "field_label"
            ws.cell(row=r, column=3, value=val).style = "field_value"
            ws.row_dimensions[r].height = 22
    blocks2 = [
        ("SUPPLEMENTS", 5, [("Morning", "Multivitamin, vitamin D"), ("Pre-match", "Electrolytes, caffeine"),
                            ("Post-match", "Whey protein, magnesium"), ("Joint care", "Fish oil, collagen")]),
        ("GROCERY STAPLES", 11, [("Proteins", "Chicken, eggs, Greek yogurt, salmon"),
                                 ("Carbs", "Oats, rice, sweet potato, bananas"),
                                 ("Recovery", "Berries, spinach, nuts, tart cherry"),
                                 ("Hydration", "Electrolyte packs, coconut water")]),
    ]
    for title, top, items in blocks2:
        merge_set(ws, f"E{top}:F{top}", title, "section_gold"); ws.row_dimensions[top].height = 22
        for i, (lab, val) in enumerate(items):
            r = top + 1 + i
            ws.cell(row=r, column=5, value=lab).style = "field_label"
            ws.cell(row=r, column=6, value=val).style = "field_value"


# ===========================================================================
# 14 — Goal Tracker
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 16, 12, 14, 14, 2])
    luxe_header(ws, "H", "🎯  GOAL TRACKER",
                "Turn ambition into a plan — targets, progress bars & season milestones.")
    table_headers(ws, 4, ["Goal", "Category", "Target", "Current", "Progress", "Deadline"])
    # (goal, category, target_txt, current_txt, progress, deadline_offset)
    goals = [
        ("Reach 4.0 DUPR rating", "Ranking", "4.00", "3.75", 0.75, 150),
        ("Medal at a 3.5 tournament", "Tournament", "Podium", "3rd @ Lakeside", 0.60, 65),
        ("3× strength sessions / week", "Fitness", "12 / mo", "8 / mo", 0.70, 30),
        ("Master the third-shot drop", "Skill", "8 / 10", "6 / 10", 0.65, 90),
        ("12 practice sessions / month", "Practice", "12", "6", 0.50, 30),
        ("Play 40 matches this season", "Milestone", "40", "24", 0.60, 120),
    ]
    start = L0
    for i, (g, cat, tgt, cur, prog, doff) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        cd = ws.cell(row=r, column=6, value=dplus(doff)); cd.style = "td"; cd.number_format = "mm/dd/yyyy"
        if i % 2:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalCatList")
    nrange(wb, "GoalCategory", "Goals", "B", start, end)
    nrange(wb, "GoalProgress", "Goals", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    merge_set(ws, "B13:F13", "SEASON MILESTONES", "section_gold"); ws.row_dimensions[13].height = 22
    ms = ["First tournament gold ✓ (Spring Slam)", "Broke 60% win rate ✓",
          "Reach 4.0 DUPR", "Win a match playing up at 4.0", "Complete a full injury-free season"]
    for i, m in enumerate(ms):
        r = 14 + i
        done = "✓" in m
        cb = ws.cell(row=r, column=2, value="☑" if done else "☐")
        cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=12, color=PRIMARY if done else ACCENT); cb.border = BOX
        merge_set(ws, f"C{r}:F{r}", m.replace(" ✓", ""), "td_left")
        if done:
            ws.cell(row=r, column=2).fill = fill(MINT_BG)


# ===========================================================================
# 15 — Photo & Memory Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "📸  PHOTO & MEMORY GALLERY",
                "Relive the season — paste photos of medals, teams & big wins.")
    sections = ["Tournament Gold", "Team Photo", "Medal Ceremony", "Action Shot", "New Paddle", "Club Event"]
    top0 = 5; card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th"); ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+5}", "📷\nPaste photo here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 6):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 6, column=col, value="Event / Date").style = "field_label"
        merge_set(ws, f"{M}{row+6}:{R}{row+6}", "", "field_value")
        ws.cell(row=row + 7, column=col, value="Caption").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")


# ===========================================================================
# 16 — Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 18, 3, 16, 12, 12, 2])
    luxe_header(ws, "H", "📊  ANALYTICS DASHBOARD",
                "Your season by the numbers — every area scored into one Performance Score.")
    merge_set(ws, "B5:D5", "PERFORMANCE DIMENSIONS", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Win rate", '=IFERROR(COUNTIF(MatchResult,"W")/COUNTA(MatchResult),0)'),
        ("Skill level", "=IFERROR(AVERAGE(SkillCurrent)/10,0)"),
        ("Fitness consistency", '=IFERROR(AVERAGEIF(GoalCategory,"Fitness",GoalProgress),0)'),
        ("Tournament success", '=IFERROR(COUNTIF(TournPlace,"1st")/MAX(COUNTA(TournPlace),1),0)'),
        ("Goal progress", "=IFERROR(AVERAGE(GoalProgress),0)"),
        ("Practice consistency", "=IFERROR(MIN(COUNTIFS(PracticeDate,\">=\"&TODAY()-30)/8,1),0)"),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.75,"Elite",IF(C{r}>=0.5,"Solid","Develop"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(f"C{start}:C{end}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "F5:H5", "PICKLEBALL PERFORMANCE SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY); cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "Wins · skills · fitness · tournaments · goals · practice.", "subtitle")
    ws["F10"].fill = fill(IVORY)
    cell_name(wb, "HealthRange", "Analytics", f"$C${start}:$C${end}")
    # court time trend table (feeds dashboard line chart)
    merge_set(ws, "B15:D15", "COURT TIME — LAST 6 MONTHS", "section")
    ws.cell(row=16, column=2, value="Month").style = "th"
    ws.cell(row=16, column=3, value="Hours").style = "th"
    months = _recent_months(6)
    hrs = [5, 6, 7, 6, 8, 8]
    for i, (m, h) in enumerate(zip(months, hrs)):
        r = 17 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=h); c.style = "td"; c.number_format = "0.0"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "CourtMonth", "Analytics", "$B$17:$B$22")
    cell_name(wb, "CourtHrs", "Analytics", "$C$17:$C$22")
    bar = BarChart(); bar.type = "bar"; bar.title = "Performance by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None; ws.add_chart(bar, "F15")


# ===========================================================================
# 1 — Executive Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🏓  PICKLEBALL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Matches, skills, tournaments, fitness & finances — your whole game, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MATCHES PLAYED", "=COUNTA(MatchResult)", "num"),
        ("WIN PERCENTAGE", '=IFERROR(COUNTIF(MatchResult,"W")/COUNTA(MatchResult),0)', "pct"),
        ("CURRENT RATING", "=PlayerRating", "rate"),
        ("TOURNAMENT WINS", '=COUNTIF(TournPlace,"1st")', "num"),
        ("PRACTICE (30 DAYS)", '=COUNTIFS(PracticeDate,">="&TODAY()-30)', "num"),
        ("COURT TIME (HRS)", "=SUM(PracticeHrs)+COUNTA(MatchResult)", "dec"),
    ]
    row2 = [
        ("MONTHLY BUDGET", "=MonthlyBudget", "money"),
        ("GEAR TO REPLACE", '=SUMPRODUCT((EquipReplace<=TODAY()+60)*(EquipReplace<>"")*(EquipName<>""))', "num"),
        ("FITNESS PROGRESS", '=IFERROR(AVERAGEIF(GoalCategory,"Fitness",GoalProgress),0)', "pct"),
        ("UPCOMING EVENTS", '=COUNTIF(TournDate,">="&TODAY())', "num"),
        ("TOP PARTNER", "=INDEX(PartnerRecord,MATCH(MAX(PartnerWinPct),PartnerWinPct,0))", "text"),
        ("SEASON PROGRESS", "=IFERROR(COUNTA(MatchResult)/SeasonTarget,0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "MATCHES & COURT TIME", "section_gold")
    # match results donut
    d1 = DoughnutChart(); d1.title = "Match Results (W/L)"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Matches"], min_col=11, min_row=5, max_row=6), titles_from_data=False)
    d1.set_categories(Reference(wb["Matches"], min_col=10, min_row=5, max_row=6)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")
    # court time line
    ln = LineChart(); ln.title = "Court Time by Month"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); ln.legend = None
    ws.add_chart(ln, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "SKILLS & SPENDING", "section_gold")
    # skill progress bar (start vs current)
    sk = BarChart(); sk.type = "bar"; sk.title = "Skill Progress — Start vs Now"; sk.height = 8.2; sk.width = 11.5
    sk.add_data(Reference(wb["Skills"], min_col=2, min_row=4, max_row=16), titles_from_data=True)
    sk.add_data(Reference(wb["Skills"], min_col=3, min_row=4, max_row=16), titles_from_data=True)
    sk.set_categories(Reference(wb["Skills"], min_col=1, min_row=5, max_row=16))
    ws.add_chart(sk, "B30")
    # spending donut
    sp = DoughnutChart(); sp.title = "Spending by Category"; sp.height = 8.2; sp.width = 11.5
    sp.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=15), titles_from_data=True)
    sp.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=15)); sp.dataLabels = no_labels()
    ws.add_chart(sp, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Pickleball Command Center™ — improve your game, one organized season at a time. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_matches(wb)
    build_tournaments(wb); build_practice(wb); build_skills(wb); build_equipment(wb)
    build_budget(wb); build_fitness(wb); build_partners(wb); build_league(wb)
    build_travel(wb); build_nutrition(wb); build_goals(wb); build_gallery(wb)
    build_analytics(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Player Profile", "Matches", "Tournaments", "Practice",
             "Skills", "Equipment", "Budget", "Fitness", "Partners", "Club & League",
             "Travel", "Nutrition", "Goals", "Gallery", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Pickleball_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
