"""Build the Soccer Mom Command Center (SMCC) Excel workbook.

14 sheets · family dashboard + 12 functional planners + settings.

Run: python3 build_xlsx.py
Outputs: ../Soccer_Mom_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
)
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"
ACCENT = "937356"
SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"
BG = "FFFFFF"
TEXT = "333333"
SUCCESS = "75E6C1"
WARNING = "937356"
DANGER = "C94C4C"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"

EXPENSE_CATEGORIES = [
    "Registration", "Uniforms", "Cleats", "Shin Guards", "Balls",
    "Training Equipment", "Tournaments", "Hotels", "Fuel", "Food",
    "Team Fees", "Private Coaching",
]
POSITIONS = ["Goalkeeper", "Defender", "Midfielder", "Forward", "Striker", "Sub"]
PRACTICE_TYPES = ["Field Practice", "Conditioning", "Scrimmage",
                  "Skills Clinic", "Tactical Session", "Indoor"]
CONDITIONS = ["New", "Good", "Fair", "Replace"]
HOME_AWAY = ["Home", "Away", "Neutral"]
RESULTS = ["Win", "Loss", "Draw", "—"]
LEAGUES = ["U6", "U8", "U10", "U12", "U14", "U16", "U18", "Adult", "Club", "Travel"]
DIVISIONS = ["Recreational", "Select", "Premier", "Elite", "Academy"]
FIELDS = ["Riverside Park", "Memorial Field", "Lincoln Complex",
          "Eastside Sports Center", "Indoor Arena", "Other"]
YESNO = ["Yes", "No"]
DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
SEASONS = ["Spring", "Summer", "Fall", "Winter"]
UNIFORM_COLORS = ["Home Green", "Away White", "Alternate Black", "Goalkeeper"]
ATTENDANCE_LEVELS = ["Present", "Late", "Absent", "Excused"]

THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Style registration
# ---------------------------------------------------------------------------

def clean_labels(pct=False, val=False):
    """Concise data labels: never show series/category text (it collides on
    multi-slice pies). Percentages or values only."""
    dl = DataLabelList()
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    dl.showVal = val
    dl.showPercent = pct
    return dl


def no_labels():
    """Suppress all on-chart labels (legend carries meaning)."""
    return clean_labels(pct=False, val=False)


def register_styles(wb: Workbook) -> None:
    styles = {
        "title": NamedStyle(
            name="title",
            font=Font(size=22, bold=True, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "subtitle": NamedStyle(
            name="subtitle",
            font=Font(size=11, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "section": NamedStyle(
            name="section",
            font=Font(size=12, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "th": NamedStyle(
            name="th",
            font=Font(size=11, bold=True, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=BOX,
        ),
        "td": NamedStyle(
            name="td",
            font=Font(size=11, color=TEXT),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=BOX,
        ),
        "td_left": NamedStyle(
            name="td_left",
            font=Font(size=11, color=TEXT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
            border=BOX,
        ),
        "kpi_label": NamedStyle(
            name="kpi_label",
            font=Font(size=10, bold=True, color=ACCENT),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
        ),
        "kpi_value": NamedStyle(
            name="kpi_value",
            font=Font(size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
        ),
        "kpi_money": NamedStyle(
            name="kpi_money",
            font=Font(size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
            number_format='"$"#,##0.00',
        ),
        "kpi_pct": NamedStyle(
            name="kpi_pct",
            font=Font(size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
            number_format="0%",
        ),
        "input": NamedStyle(
            name="input",
            font=Font(size=11, bold=True, color=PRIMARY),
            fill=PatternFill("solid", fgColor=SURFACE),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=BOX,
        ),
        "field_label": NamedStyle(
            name="field_label",
            font=Font(size=10, bold=True, color=ACCENT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=BOX,
            fill=PatternFill("solid", fgColor=SOFT_BG),
        ),
        "field_value": NamedStyle(
            name="field_value",
            font=Font(size=11, color=TEXT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=BOX,
        ),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def merge_set(ws, rng, value, style_name):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.style = style_name
    return cell


def page_header(ws, last_col: str, title: str, subtitle: str) -> None:
    ws.row_dimensions[1].height = 44
    ws.row_dimensions[2].height = 22
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")


def set_column_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng: str, list_name: str) -> None:
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def style_data_rows(ws, start_row: int, end_row: int, col_count: int,
                    text_left_cols: set[int] | None = None,
                    money_cols: set[int] | None = None,
                    int_cols: set[int] | None = None,
                    date_cols: set[int] | None = None) -> None:
    text_left_cols = text_left_cols or set()
    money_cols = money_cols or set()
    int_cols = int_cols or set()
    date_cols = date_cols or set()
    for r in range(start_row, end_row + 1):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left_cols else "td"
            cell.fill = fill(MUTED_ROW if (r - start_row) % 2 == 1 else BG)
            if c in money_cols:
                cell.number_format = '"$"#,##0.00'
            elif c in int_cols:
                cell.number_format = "0"
            elif c in date_cols:
                cell.number_format = "mm/dd/yyyy"


# ===========================================================================
# Sheet 14: Settings
# ===========================================================================
def build_settings(wb: Workbook) -> None:
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [2, 32, 22, 4, 22, 22, 22, 22, 22])
    page_header(ws, "I", "⚙  SETTINGS", "Season, budget, dropdowns. Edit once — drives every sheet.")

    merge_set(ws, "B4:C4", "USER CONTROLS", "section")
    controls = [
        ("Season",                 "Fall 2026",       None,                "Season"),
        ("Season Start Date",      dt.date(2026, 8, 24), "mm/dd/yyyy",     "SeasonStart"),
        ("Season End Date",        dt.date(2026, 11, 22),"mm/dd/yyyy",     "SeasonEnd"),
        ("League",                 "U10 Travel",      None,                "League"),
        ("Division",               "Premier",         None,                "Division"),
        ("Monthly Soccer Budget",  650,               '"$"#,##0.00',       "MonthlyBudget"),
        ("# of Players in Family", 2,                 "0",                 "PlayerCount"),
        ("Fuel $ per Mile",        0.18,              '"$"#,##0.000',      "FuelPerMile"),
        ("Today (auto)",           "=TODAY()",        "mm/dd/yyyy",        "TodayDate"),
    ]
    for i, (label, default, fmt, name) in enumerate(controls):
        r = 5 + i
        ws.cell(row=r, column=2, value=label).style = "field_label"
        cell = ws.cell(row=r, column=3, value=default)
        cell.style = "input"
        if fmt:
            cell.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Settings!$C${r}")

    merge_set(ws, "E4:I4", "DROPDOWN LISTS", "section")
    list_data = [
        ("E", "Expense Categories", EXPENSE_CATEGORIES, "ExpenseList"),
        ("F", "Positions",          POSITIONS,          "PositionList"),
        ("G", "Practice Types",     PRACTICE_TYPES,     "PracticeTypeList"),
        ("H", "Conditions",         CONDITIONS,         "ConditionList"),
        ("I", "Home / Away",        HOME_AWAY,          "VenueList"),
    ]
    for col, h, data, name in list_data:
        c_idx = ord(col) - 64
        ws.cell(row=5, column=c_idx, value=h).style = "th"
        for ri, val in enumerate(data):
            ws.cell(row=6 + ri, column=c_idx, value=val).style = "td_left"
        end = 5 + len(data)
        wb.defined_names[name] = DefinedName(
            name, attr_text=f"Settings!${col}$6:${col}${end}"
        )

    # Extra lists farther down
    extras = [
        ("E", "Results",         RESULTS),
        ("F", "Leagues",         LEAGUES),
        ("G", "Divisions",       DIVISIONS),
        ("H", "Fields",          FIELDS),
        ("I", "Yes / No",        YESNO),
    ]
    start_row = 20
    for col, h, data in extras:
        c_idx = ord(col) - 64
        ws.cell(row=start_row, column=c_idx, value=h).style = "th"
        for ri, val in enumerate(data):
            ws.cell(row=start_row + 1 + ri, column=c_idx, value=val).style = "td_left"
        end = start_row + len(data)
        if h == "Results":
            wb.defined_names["ResultList"] = DefinedName(
                "ResultList", attr_text=f"Settings!${col}$21:${col}${end}"
            )
        elif h == "Leagues":
            wb.defined_names["LeagueList"] = DefinedName(
                "LeagueList", attr_text=f"Settings!${col}$21:${col}${end}"
            )
        elif h == "Divisions":
            wb.defined_names["DivisionList"] = DefinedName(
                "DivisionList", attr_text=f"Settings!${col}$21:${col}${end}"
            )
        elif h == "Fields":
            wb.defined_names["FieldList"] = DefinedName(
                "FieldList", attr_text=f"Settings!${col}$21:${col}${end}"
            )
        elif h == "Yes / No":
            wb.defined_names["YesNoList"] = DefinedName(
                "YesNoList", attr_text=f"Settings!${col}$21:${col}${end}"
            )

    # Attendance list separate column
    ws.cell(row=20, column=2, value="Attendance").style = "th"
    for ri, val in enumerate(ATTENDANCE_LEVELS):
        ws.cell(row=21 + ri, column=2, value=val).style = "td_left"
    wb.defined_names["AttendanceList"] = DefinedName(
        "AttendanceList", attr_text=f"Settings!$B$21:$B${20 + len(ATTENDANCE_LEVELS)}"
    )

    # Uniform colors
    ws.cell(row=27, column=2, value="Uniforms").style = "th"
    for ri, val in enumerate(UNIFORM_COLORS):
        ws.cell(row=28 + ri, column=2, value=val).style = "td_left"
    wb.defined_names["UniformList"] = DefinedName(
        "UniformList", attr_text=f"Settings!$B$28:$B${27 + len(UNIFORM_COLORS)}"
    )


# ===========================================================================
# Sheet 2: Player Profiles
# ===========================================================================
def build_players(wb: Workbook) -> None:
    ws = wb.create_sheet("Players")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [2, 22, 28, 4, 22, 28])
    page_header(ws, "F", "⚽  PLAYER PROFILES",
                "One section per child. Duplicate the block as needed.")

    samples = [
        {
            "Name": "Maya Rivera", "Team": "FC United U10 Girls",
            "Age Group": "U10", "Jersey #": 7, "Position": "Midfielder",
            "Coach": "Coach Daniels", "League": "U10 Travel",
            "Practice Days": "Mon / Wed", "Medical Notes": "Mild asthma — inhaler in bag",
            "Emergency Contact": "Mom: (555) 412-7833",
            "Allergies": "Peanuts", "Preferred Foot": "Right",
            "Goal for Season": "10 assists + start every game",
        },
        {
            "Name": "Diego Rivera", "Team": "FC United U13 Boys",
            "Age Group": "U13", "Jersey #": 11, "Position": "Forward",
            "Coach": "Coach Marin", "League": "U13 Club",
            "Practice Days": "Tue / Thu / Sat", "Medical Notes": "—",
            "Emergency Contact": "Dad: (555) 412-7901",
            "Allergies": "None", "Preferred Foot": "Left",
            "Goal for Season": "15 goals + tournament MVP",
        },
    ]

    row = 4
    for player in samples:
        merge_set(ws, f"B{row}:F{row}", "  PLAYER", "section")
        ws.row_dimensions[row].height = 22
        row += 1
        fields = list(player.items())
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24
            i += 2
            row += 1
        row += 2

    add_dv(ws, "C5:C20", "PositionList")
    add_dv(ws, "F5:F20", "LeagueList")


# ===========================================================================
# Sheet 3: Season Schedule
# ===========================================================================
def build_schedule(wb: Workbook) -> None:
    ws = wb.create_sheet("Schedule")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [12, 24, 12, 12, 12, 24, 22, 18, 18, 14, 26, 14])
    page_header(ws, "L", "📅  SEASON SCHEDULE",
                "Every match — countdown, venue, weather, result, notes.")
    headers = ["Date", "Opponent", "Home/Away", "Kickoff", "Arrival",
               "Field", "Address", "Weather", "Uniform", "Result",
               "Notes", "Days Out"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)

    games = [
        (d(-14), "Riverside Rovers",   "Home",     "10:00",  "9:15",  "Riverside Park",      "1500 River Rd",        "Sunny 72°",    "Home Green",  "Win"),
        (d(-7),  "Lincoln Lions",      "Away",     "11:30",  "10:00", "Lincoln Complex",     "240 Lincoln Way",      "Cloudy 65°",   "Away White",  "Draw"),
        (d(-3),  "Eastside Eagles",    "Away",     "9:00",   "7:45",  "Eastside Sports Ctr", "88 Eastside Blvd",     "Rain 58°",     "Away White",  "Loss"),
        (d(2),   "Memorial Hawks",     "Home",     "10:00",  "9:15",  "Memorial Field",      "33 Memorial Pkwy",     "Partly Cloudy","Home Green",  "—"),
        (d(5),   "Northside FC",       "Away",     "1:30",   "12:00", "Northside Park",      "210 North Ave",        "Sunny 70°",    "Away White",  "—"),
        (d(9),   "Westside Wolves",    "Home",     "11:00",  "10:15", "Riverside Park",      "1500 River Rd",        "TBD",          "Home Green",  "—"),
        (d(12),  "Tournament: City Cup","Neutral", "9:00",   "7:30",  "Lincoln Complex",     "240 Lincoln Way",      "TBD",          "Home Green",  "—"),
        (d(16),  "Tournament: City Cup","Neutral", "1:00",   "11:30", "Lincoln Complex",     "240 Lincoln Way",      "TBD",          "Away White",  "—"),
        (d(20),  "Riverside Rovers",   "Away",     "10:00",  "8:45",  "Riverside Park",      "1500 River Rd",        "TBD",          "Away White",  "—"),
        (d(27),  "Memorial Hawks",     "Away",     "11:30",  "10:00", "Memorial Field",      "33 Memorial Pkwy",     "TBD",          "Away White",  "—"),
        (d(34),  "Eastside Eagles",    "Home",     "10:00",  "9:15",  "Riverside Park",      "1500 River Rd",        "TBD",          "Home Green",  "—"),
        (d(41),  "Lincoln Lions",      "Home",     "12:00",  "11:00", "Riverside Park",      "1500 River Rd",        "TBD",          "Home Green",  "—"),
    ]

    start = 5
    end = start + 40 - 1
    for i, g in enumerate(games):
        r = start + i
        for ci, val in enumerate(g, 1):
            ws.cell(row=r, column=ci, value=val)
        ws.cell(row=r, column=11, value="—")

    # Days Out + formatting on all reserved rows
    for r in range(start, end + 1):
        ws.cell(row=r, column=12,
                value=f'=IF(A{r}="","",A{r}-TODAY())')

    style_data_rows(ws, start, end, 12,
                    text_left_cols={2, 6, 7, 11},
                    date_cols={1})
    for r in range(start, end + 1):
        ws.cell(row=r, column=12).number_format = "0;[Red]-0"

    add_dv(ws, f"C{start}:C{end}", "VenueList")
    add_dv(ws, f"F{start}:F{end}", "FieldList")
    add_dv(ws, f"I{start}:I{end}", "UniformList")
    add_dv(ws, f"J{start}:J{end}", "ResultList")

    # Conditional formatting
    # Upcoming next 7 days
    ws.conditional_formatting.add(
        f"A{start}:L{end}",
        FormulaRule(formula=[f'AND($A{start}<>"",$A{start}-TODAY()>=0,$A{start}-TODAY()<=7)'],
                    fill=PatternFill("solid", fgColor="DCF5EC")),
    )
    # Past games muted
    ws.conditional_formatting.add(
        f"A{start}:L{end}",
        FormulaRule(formula=[f'AND($A{start}<>"",$A{start}<TODAY())'],
                    fill=PatternFill("solid", fgColor="F1F1F1")),
    )
    # Win/Loss/Draw color codes on Result column
    ws.conditional_formatting.add(
        f"J{start}:J{end}",
        CellIsRule(operator="equal", formula=['"Win"'],
                   fill=PatternFill("solid", fgColor="E3F8EF")),
    )
    ws.conditional_formatting.add(
        f"J{start}:J{end}",
        CellIsRule(operator="equal", formula=['"Loss"'],
                   fill=PatternFill("solid", fgColor="FBE6E6")),
    )

    wb.defined_names["GameDates"] = DefinedName(
        "GameDates", attr_text=f"Schedule!$A${start}:$A${end}"
    )
    wb.defined_names["GameOpponents"] = DefinedName(
        "GameOpponents", attr_text=f"Schedule!$B${start}:$B${end}"
    )
    wb.defined_names["GameResults"] = DefinedName(
        "GameResults", attr_text=f"Schedule!$J${start}:$J${end}"
    )

    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 4: Practice Planner
# ===========================================================================
def build_practices(wb: Workbook) -> None:
    ws = wb.create_sheet("Practices")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [12, 12, 22, 22, 14, 28, 26])
    page_header(ws, "G", "🏃  PRACTICE PLANNER",
                "Sessions, focus areas, attendance, coach notes & drills.")
    headers = ["Date", "Time", "Location", "Focus Area",
               "Attendance", "Coach Notes", "Homework Drills"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)

    practices = [
        (d(-10), "5:30 PM", "Riverside Park",  "Passing accuracy",       "Present", "Strong tempo. Focus next on receiving.", "20 wall passes x 3 days"),
        (d(-8),  "5:30 PM", "Riverside Park",  "1v1 defending",          "Present", "Body positioning improved.",            "Mirror drill 10 min"),
        (d(-5),  "5:30 PM", "Indoor Arena",    "Conditioning",           "Late",    "Showed up 15 min late.",                "Sprint ladder 3x"),
        (d(-3),  "5:30 PM", "Riverside Park",  "Shooting from distance", "Present", "Power was excellent today.",            "Ball-strike technique"),
        (d(-1),  "5:30 PM", "Riverside Park",  "Tactical: 3-4-3",        "Excused", "Doctor appt — light review at home.",   "Watch position video"),
        (d(2),   "5:30 PM", "Riverside Park",  "Scrimmage",              "—",       "—",                                     "—"),
        (d(4),   "5:30 PM", "Riverside Park",  "Crossing & finishing",   "—",       "—",                                     "—"),
        (d(7),   "5:30 PM", "Indoor Arena",    "Goalkeeper rotation",    "—",       "—",                                     "—"),
        (d(9),   "5:30 PM", "Riverside Park",  "Set pieces",             "—",       "—",                                     "—"),
        (d(11),  "5:30 PM", "Riverside Park",  "Match prep",             "—",       "—",                                     "—"),
    ]

    start = 5
    end = start + 50 - 1
    for i, p in enumerate(practices):
        r = start + i
        for ci, val in enumerate(p, 1):
            ws.cell(row=r, column=ci, value=val)

    style_data_rows(ws, start, end, 7,
                    text_left_cols={3, 4, 6, 7},
                    date_cols={1})

    add_dv(ws, f"C{start}:C{end}", "FieldList")
    add_dv(ws, f"D{start}:D{end}", "PracticeTypeList")
    add_dv(ws, f"E{start}:E{end}", "AttendanceList")

    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Present"'],
                   fill=PatternFill("solid", fgColor="E3F8EF")),
    )
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Absent"'],
                   fill=PatternFill("solid", fgColor="FBE6E6")),
    )
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Late"'],
                   fill=PatternFill("solid", fgColor="FBF0E2")),
    )

    wb.defined_names["PracDates"] = DefinedName(
        "PracDates", attr_text=f"Practices!$A${start}:$A${end}"
    )
    wb.defined_names["PracAttendance"] = DefinedName(
        "PracAttendance", attr_text=f"Practices!$E${start}:$E${end}"
    )

    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 5: Soccer Budget
# ===========================================================================
def build_budget(wb: Workbook) -> None:
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [24, 18, 18, 18, 16, 4, 24, 20])
    page_header(ws, "H", "💰  SOCCER BUDGET",
                "Plan, track, and check variance across every season cost.")

    headers = ["Category", "Planned", "Actual", "Variance", "% Spent"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    planned = {
        "Registration":       250, "Uniforms":           180, "Cleats":             95,
        "Shin Guards":         25, "Balls":               45, "Training Equipment": 80,
        "Tournaments":        450, "Hotels":             380, "Fuel":              240,
        "Food":               180, "Team Fees":          200, "Private Coaching":  300,
    }
    actuals = {
        "Registration":       250, "Uniforms":           165, "Cleats":             89.99,
        "Shin Guards":         22, "Balls":               42, "Training Equipment": 60,
        "Tournaments":        300, "Hotels":             280, "Fuel":              188,
        "Food":               130, "Team Fees":          200, "Private Coaching":  150,
    }

    start = 5
    end = start + len(EXPENSE_CATEGORIES) - 1
    for i, cat in enumerate(EXPENSE_CATEGORIES):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        c_plan = ws.cell(row=r, column=2, value=planned.get(cat, 0))
        c_plan.style = "input"; c_plan.number_format = '"$"#,##0.00'
        c_act = ws.cell(row=r, column=3, value=actuals.get(cat, 0))
        c_act.style = "td"; c_act.number_format = '"$"#,##0.00'
        c_var = ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        c_var.style = "td"; c_var.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
        c_pct = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)")
        c_pct.style = "td"; c_pct.number_format = "0%"
        ws.row_dimensions[r].height = 22
        if i % 2 == 1:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)

    total_row = end + 1
    ws.cell(row=total_row, column=1, value="TOTAL").style = "th"
    for col, fml in (
        (2, f"=SUM(B{start}:B{end})"),
        (3, f"=SUM(C{start}:C{end})"),
        (4, f"=SUM(D{start}:D{end})"),
        (5, f"=IFERROR(C{total_row}/B{total_row},0)"),
    ):
        c = ws.cell(row=total_row, column=col, value=fml)
        c.style = "td"
        c.font = Font(size=11, bold=True, color=PRIMARY)
        c.fill = fill(SURFACE)
        c.number_format = '"$"#,##0.00' if col in (2, 3, 4) else "0%"

    wb.defined_names["BudgetCategories"] = DefinedName(
        "BudgetCategories", attr_text=f"Budget!$A${start}:$A${end}"
    )
    wb.defined_names["BudgetPlanned"] = DefinedName(
        "BudgetPlanned", attr_text=f"Budget!$B${start}:$B${end}"
    )
    wb.defined_names["BudgetActual"] = DefinedName(
        "BudgetActual", attr_text=f"Budget!$C${start}:$C${end}"
    )
    wb.defined_names["BudgetTotalPlanned"] = DefinedName(
        "BudgetTotalPlanned", attr_text=f"Budget!$B${total_row}"
    )
    wb.defined_names["BudgetTotalActual"] = DefinedName(
        "BudgetTotalActual", attr_text=f"Budget!$C${total_row}"
    )

    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color=PRIMARY, showValue=True),
    )

    # KPI sidebar
    ws.cell(row=4, column=7, value="BUDGET KPIs").style = "section"
    kpis = [
        ("Total Planned",        "=BudgetTotalPlanned",   '"$"#,##0.00'),
        ("Total Actual",         "=BudgetTotalActual",    '"$"#,##0.00'),
        ("Variance",             "=BudgetTotalPlanned-BudgetTotalActual",
                                                          '"$"#,##0.00;[Red]-"$"#,##0.00'),
        ("Cost Per Month",       "=BudgetTotalActual/MAX(1,(SeasonEnd-SeasonStart)/30)",
                                                          '"$"#,##0.00'),
        ("Cost Per Player",      "=BudgetTotalActual/MAX(PlayerCount,1)",
                                                          '"$"#,##0.00'),
    ]
    for i, (lab, fml, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml)
        c.style = "field_value"
        c.number_format = fmt
        c.font = Font(size=12, bold=True, color=PRIMARY)

    # Charts
    pie = DoughnutChart()
    pie.title = "Spending Breakdown"
    pie.height = 9; pie.width = 14
    pie.add_data(Reference(ws, min_col=3, min_row=4, max_row=end),
                 titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    pie.dataLabels = no_labels()
    ws.add_chart(pie, "A20")

    bar = BarChart()
    bar.type = "col"; bar.title = "Budget vs Actual"
    bar.height = 9; bar.width = 14
    bar.add_data(Reference(ws, min_col=2, min_row=4, max_row=end), titles_from_data=True)
    bar.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    ws.add_chart(bar, "G20")


# ===========================================================================
# Sheet 6: Equipment Tracker
# ===========================================================================
def build_equipment(wb: Workbook) -> None:
    ws = wb.create_sheet("Equipment")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [18, 24, 12, 12, 14, 14, 16, 14, 24])
    page_header(ws, "I", "🎽  EQUIPMENT TRACKER",
                "Size, condition, replacement dates — never get caught short.")
    headers = ["Category", "Item", "Size", "Qty", "Purchased",
               "Condition", "Replace By", "Cost", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)

    items = [
        ("Cleats",          "Adidas Predator",       "5Y",  1, "Yes", "Good",   d(180), 89.99, "Worn for one season"),
        ("Cleats",          "Indoor turf shoes",     "5Y",  1, "Yes", "New",    d(365), 49.99, ""),
        ("Socks",           "Match socks (Green)",   "M",   4, "Yes", "Good",   d(180), 32.00, ""),
        ("Socks",           "Training socks",        "M",   6, "Yes", "Good",   d(120), 18.00, ""),
        ("Jerseys",         "Home jersey #7",        "YM",  1, "Yes", "New",    d(365), 65.00, ""),
        ("Jerseys",         "Away jersey #7",        "YM",  1, "Yes", "New",    d(365), 65.00, ""),
        ("Shorts",          "Match shorts (Green)",  "YM",  2, "Yes", "Good",   d(180), 36.00, ""),
        ("Water Bottles",   "Insulated bottle 24oz", "—",   2, "Yes", "Good",   d(180), 18.00, ""),
        ("Ball",            "Match ball Size 4",     "4",   1, "Yes", "Good",   d(120), 32.00, ""),
        ("Ball",            "Training ball Size 4",  "4",   2, "Yes", "Fair",   d(60),  18.00, "One losing air"),
        ("Backpack",        "Soccer pack",           "—",   1, "Yes", "Good",   d(360), 45.00, ""),
        ("Goalkeeper Gear", "GK gloves",             "5",   1, "No",  "Replace", d(30), 35.00, "Last year — too small"),
        ("Rain Jacket",     "Lightweight rain shell","YM",  1, "Yes", "Good",   d(300), 38.00, ""),
        ("First Aid Kit",   "Compact bag kit",       "—",   1, "Yes", "Good",   d(365), 22.00, "Restock ice packs"),
        ("Shin Guards",     "Slip-in shin guards",   "M",   1, "Yes", "Good",   d(180), 22.00, ""),
    ]
    start = 5
    end = start + 30 - 1
    for i, it in enumerate(items):
        r = start + i
        for ci, val in enumerate(it, 1):
            ws.cell(row=r, column=ci, value=val)

    style_data_rows(ws, start, end, 9,
                    text_left_cols={2, 9},
                    money_cols={8},
                    int_cols={4},
                    date_cols={7})

    add_dv(ws, f"E{start}:E{end}", "YesNoList")
    add_dv(ws, f"F{start}:F{end}", "ConditionList")

    # Replacement alert: Replace By within 60 days
    ws.conditional_formatting.add(
        f"A{start}:I{end}",
        FormulaRule(formula=[f'AND($G{start}<>"",$G{start}-TODAY()>=0,$G{start}-TODAY()<=60)'],
                    fill=PatternFill("solid", fgColor="FBF0E2")),
    )
    # Replace condition
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Replace"'],
                   fill=PatternFill("solid", fgColor="FBE6E6")),
    )
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"No"'],
                   fill=PatternFill("solid", fgColor="FBE6E6")),
    )
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill("solid", fgColor="E3F8EF")),
    )

    wb.defined_names["EquipItem"] = DefinedName(
        "EquipItem", attr_text=f"Equipment!$B${start}:$B${end}"
    )
    wb.defined_names["EquipPurchased"] = DefinedName(
        "EquipPurchased", attr_text=f"Equipment!$E${start}:$E${end}"
    )
    wb.defined_names["EquipReplaceBy"] = DefinedName(
        "EquipReplaceBy", attr_text=f"Equipment!$G${start}:$G${end}"
    )


# ===========================================================================
# Sheet 7: Tournament Planner
# ===========================================================================
def build_tournaments(wb: Workbook) -> None:
    ws = wb.create_sheet("Tournaments")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [24, 14, 14, 24, 24, 14, 14, 16, 20, 24])
    page_header(ws, "J", "🏆  TOURNAMENT PLANNER",
                "Dates, venues, hotels, entry fees, packing status.")
    headers = ["Tournament", "Start", "End", "Venue", "Hotel",
               "Entry Fee", "# Matches", "Travel Time", "Hotel Conf.", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)

    samples = [
        ("City Cup",            d(12), d(16), "Lincoln Complex",  "Hampton Inn Lincoln",   125, 4, "1h 10m", "HX-882211", "Pool play 9am/1pm"),
        ("Riverbend Classic",   d(34), d(36), "Riverbend Sports", "Marriott Riverbend",    150, 3, "2h 45m", "MR-441020", "Hotel block expires 11/1"),
        ("Fall Showcase",       d(60), d(63), "Capitol Fields",   "Hyatt Place Capitol",   195, 5, "3h 30m", "HP-771188", "Scout opportunities"),
        ("State Cup Qualifier", d(85), d(88), "State Complex",    "Holiday Inn State",     220, 3, "4h 00m", "—",         "Pending registration"),
    ]
    start = 5
    end = start + 12 - 1
    for i, t in enumerate(samples):
        r = start + i
        for ci, val in enumerate(t, 1):
            ws.cell(row=r, column=ci, value=val)

    style_data_rows(ws, start, end, 10,
                    text_left_cols={1, 4, 5, 9, 10},
                    money_cols={6},
                    int_cols={7},
                    date_cols={2, 3})

    # Highlight upcoming tournaments within 30 days
    ws.conditional_formatting.add(
        f"A{start}:J{end}",
        FormulaRule(formula=[f'AND($B{start}<>"",$B{start}-TODAY()>=0,$B{start}-TODAY()<=30)'],
                    fill=PatternFill("solid", fgColor="DCF5EC")),
    )

    wb.defined_names["TourStart"] = DefinedName(
        "TourStart", attr_text=f"Tournaments!$B${start}:$B${end}"
    )
    wb.defined_names["TourName"] = DefinedName(
        "TourName", attr_text=f"Tournaments!$A${start}:$A${end}"
    )


# ===========================================================================
# Sheet 8: Carpool Manager
# ===========================================================================
def build_carpool(wb: Workbook) -> None:
    ws = wb.create_sheet("Carpool")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [12, 22, 22, 24, 14, 14, 18, 18])
    page_header(ws, "H", "🚗  CARPOOL MANAGER",
                "Driver rotation, capacity, fuel contribution.")
    headers = ["Date", "Event", "Driver", "Players", "Pickup",
               "Return", "Contact", "Fuel Share"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)

    rows = [
        (d(2),  "Home vs Hawks",   "Sarah R.",   "Maya, Lily, Sophie",     "8:30 AM",  "12:30 PM", "(555) 412-7833", 10),
        (d(5),  "Away @ Northside","James C.",   "Maya, Aiden",            "11:30 AM", "4:30 PM",  "(555) 412-7901", 15),
        (d(9),  "Home vs Wolves",  "Priya M.",   "Maya, Diego, Liam",      "9:30 AM",  "1:30 PM",  "(555) 388-2210", 10),
        (d(12), "City Cup Day 1",  "Sarah R.",   "Maya, Lily, Sophie, Mia","6:30 AM",  "8:00 PM",  "(555) 412-7833", 25),
        (d(16), "City Cup Day 2",  "Marcus K.",  "Maya, Aiden",            "10:00 AM", "6:00 PM",  "(555) 901-2271", 25),
        (d(20), "Away @ Rovers",   "Aisha P.",   "Maya, Sophie, Lily",     "8:00 AM",  "1:00 PM",  "(555) 388-3344", 12),
        (d(27), "Away @ Memorial", "Sarah R.",   "Maya, Diego",            "9:30 AM",  "2:30 PM",  "(555) 412-7833", 8),
    ]
    start = 5
    end = start + 30 - 1
    for i, r_data in enumerate(rows):
        r = start + i
        for ci, val in enumerate(r_data, 1):
            ws.cell(row=r, column=ci, value=val)

    style_data_rows(ws, start, end, 8,
                    text_left_cols={2, 3, 4},
                    money_cols={8},
                    date_cols={1})


# ===========================================================================
# Sheet 9: Team Roster
# ===========================================================================
def build_roster(wb: Workbook) -> None:
    ws = wb.create_sheet("Roster")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [22, 22, 18, 28, 16, 14, 22])
    page_header(ws, "G", "👥  TEAM ROSTER",
                "Players, parents, contact info, positions, birthdays.")
    headers = ["Player", "Parent", "Phone", "Email",
               "Position", "Birthday", "Emergency Contact"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    rows = [
        ("Maya Rivera",    "Sarah Rivera",    "(555) 412-7833", "sarah@example.com",  "Midfielder", dt.date(2016,  3, 12), "James (Dad)"),
        ("Lily Chen",      "Priya Chen",      "(555) 388-2210", "priya@example.com",  "Defender",   dt.date(2016,  7,  4), "Mark (Dad)"),
        ("Sophie Park",    "Aisha Park",      "(555) 388-3344", "aisha@example.com",  "Forward",    dt.date(2016,  5, 22), "Alex (Stepdad)"),
        ("Aiden Brooks",   "Marcus Brooks",   "(555) 901-2271", "marcus@example.com", "Goalkeeper", dt.date(2016,  9, 18), "Emily (Mom)"),
        ("Mia Lopez",      "Carla Lopez",     "(555) 200-1188", "carla@example.com",  "Defender",   dt.date(2015, 12,  1), "—"),
        ("Olivia Reed",    "Hannah Reed",     "(555) 200-2244", "hannah@example.com", "Midfielder", dt.date(2016,  1, 30), "Tom (Dad)"),
        ("Sophia Diaz",    "Elena Diaz",      "(555) 200-7711", "elena@example.com",  "Forward",    dt.date(2016,  6, 11), "Luis (Dad)"),
        ("Ava Patel",      "Anika Patel",     "(555) 412-9032", "anika@example.com",  "Defender",   dt.date(2016,  4,  7), "Raj (Dad)"),
        ("Isla Wright",    "Beth Wright",     "(555) 412-7799", "beth@example.com",   "Midfielder", dt.date(2016,  8, 25), "Sam (Dad)"),
        ("Zara Khan",      "Ayesha Khan",     "(555) 388-9911", "ayesha@example.com", "Forward",    dt.date(2016, 10, 14), "Imran (Dad)"),
        ("Emma Carter",    "Linda Carter",    "(555) 412-7820", "linda@example.com",  "Sub",        dt.date(2016,  2, 16), "—"),
        ("Hannah Ross",    "Megan Ross",      "(555) 412-7822", "megan@example.com",  "Midfielder", dt.date(2016, 11,  3), "Dan (Dad)"),
    ]
    start = 5
    end = start + 20 - 1
    for i, r_data in enumerate(rows):
        r = start + i
        for ci, val in enumerate(r_data, 1):
            ws.cell(row=r, column=ci, value=val)

    style_data_rows(ws, start, end, 7,
                    text_left_cols={1, 2, 4, 7},
                    date_cols={6})
    add_dv(ws, f"E{start}:E{end}", "PositionList")


# ===========================================================================
# Sheet 10: Meal & Snack Planner
# ===========================================================================
def build_meals(wb: Workbook) -> None:
    ws = wb.create_sheet("Meals")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [16, 32, 32, 6, 22, 14])
    page_header(ws, "F", "🍎  MEAL · SNACK · HYDRATION",
                "Game-day fuel, post-game recovery, snack roster, shopping list.")

    sections = [
        ("PRE-GAME MEAL (2–3 hrs before)", [
            ("Saturday game day",  "Whole-grain pasta + grilled chicken + steamed veggies",
             "Slow-release carbs; avoid heavy sauces"),
            ("Sunday game day",    "Oatmeal + banana + scrambled eggs + toast",
             "Carbs + protein"),
            ("Tournament morning", "Bagel + peanut-free PB + Greek yogurt + berries",
             "Easy on stomach, sustained energy"),
        ]),
        ("POST-GAME RECOVERY (within 60 min)", [
            ("Light recovery",  "Chocolate milk + turkey wrap",
             "Carb + protein in 3:1 ratio"),
            ("Big recovery",    "Rice bowl + chicken + veggies + smoothie",
             "After 90+ min match"),
            ("Heat day",        "Electrolyte drink + watermelon + pretzels",
             "Replenish sodium"),
        ]),
        ("TOURNAMENT MEALS (multi-day)", [
            ("Friday dinner",       "Pasta night at hotel",                     "Carb load"),
            ("Saturday breakfast",  "Continental breakfast + protein bar packed","Eat by 7am if 9am kickoff"),
            ("Saturday lunch",      "Sub sandwich + pretzels + fruit",          "Between matches"),
            ("Sunday breakfast",    "Eggs + toast + smoothie",                  "Light if early game"),
        ]),
        ("SNACK ROSTER (per game)", [
            ("Game 1 vs Hawks",       "Rivera family", "Orange slices + pretzels"),
            ("Game 2 @ Northside",    "Chen family",   "Granola bars + apples"),
            ("Game 3 vs Wolves",      "Park family",   "Watermelon + crackers"),
            ("Tournament Day 1",      "Brooks family", "Bagels + bananas + Gatorade"),
            ("Tournament Day 2",      "Lopez family",  "Trail mix + grapes + water"),
        ]),
    ]

    row = 4
    for title, items in sections:
        merge_set(ws, f"A{row}:C{row}", title, "section")
        ws.row_dimensions[row].height = 22
        row += 1
        # Header for the section table
        if "ROSTER" in title:
            cols = ["Game", "Family Assigned", "Snack"]
        else:
            cols = ["Occasion", "Meal", "Notes"]
        for i, h in enumerate(cols, 1):
            ws.cell(row=row, column=i, value=h).style = "th"
        ws.row_dimensions[row].height = 26
        row += 1
        for i, item in enumerate(items):
            for ci, val in enumerate(item, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.style = "td_left"
                cell.fill = fill(MUTED_ROW if i % 2 == 1 else BG)
            ws.row_dimensions[row].height = 24
            row += 1
        row += 1

    # Hydration tracker (right side)
    merge_set(ws, "E4:F4", "HYDRATION TRACKER", "th")
    ws.row_dimensions[4].height = 30
    ws.cell(row=5, column=5, value="Day").style = "th"
    ws.cell(row=5, column=6, value="oz").style = "th"
    days_data = [("Mon", 64), ("Tue", 72), ("Wed", 68),
                 ("Thu", 80), ("Fri", 72), ("Sat (Game)", 96), ("Sun (Game)", 88)]
    for i, (day, oz) in enumerate(days_data):
        r = 6 + i
        ws.cell(row=r, column=5, value=day).style = "td_left"
        c = ws.cell(row=r, column=6, value=oz)
        c.style = "td"; c.number_format = "0"
        if i % 2 == 1:
            for col in (5, 6):
                ws.cell(row=r, column=col).fill = fill(MUTED_ROW)

    # Shopping list block at right column
    shop_row = 14
    merge_set(ws, f"E{shop_row}:F{shop_row}", "SHOPPING LIST", "th")
    ws.row_dimensions[shop_row].height = 28
    items_shop = [
        ("Chocolate milk (qt)",    2),
        ("Bananas",               10),
        ("Whole-grain pasta",      1),
        ("Chicken breast (lb)",    2),
        ("Greek yogurt (tubs)",    3),
        ("Bagels (pkg)",           1),
        ("Sub bread",              2),
        ("Watermelon",             1),
        ("Pretzels (bag)",         2),
        ("Granola bars (box)",     2),
        ("Apples",                 8),
        ("Trail mix (bag)",        1),
        ("Electrolyte mix",        1),
        ("Eggs (dozen)",           1),
        ("Smoothie kit",           2),
    ]
    for i, (item, qty) in enumerate(items_shop):
        r = shop_row + 1 + i
        ws.cell(row=r, column=5, value=item).style = "td_left"
        c = ws.cell(row=r, column=6, value=qty); c.style = "td"; c.number_format = "0"
        if i % 2 == 1:
            for col in (5, 6):
                ws.cell(row=r, column=col).fill = fill(MUTED_ROW)


# ===========================================================================
# Sheet 11: Packing Checklist
# ===========================================================================
def build_packing(wb: Workbook) -> None:
    ws = wb.create_sheet("Packing")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [22, 6, 22, 6, 22, 6, 22, 6, 22, 6, 22, 6])
    page_header(ws, "L", "🎒  PACKING CHECKLISTS",
                "Six pre-built lists. Check items off as you pack.")

    lists = [
        ("PRACTICE", ["Cleats", "Shin guards", "Socks", "Practice jersey",
                      "Shorts", "Water bottle", "Ball", "Pinnie", "Snack"]),
        ("GAME DAY", ["Match cleats", "Shin guards", "Match socks (home/away)",
                      "Match jersey + shorts", "Water bottle (filled)",
                      "Backup water", "Goalkeeper gloves (if GK)",
                      "Snack + electrolytes", "Sunscreen", "Hat / sunglasses",
                      "Camp chair", "Cash for entry"]),
        ("TOURNAMENT WEEKEND", ["All Game Day items × 3 matches",
                                "Spare cleats", "Spare socks (x3)",
                                "Both uniform sets", "Hotel essentials",
                                "Phone charger", "Pop-up tent / shade",
                                "Cooler with ice", "Roll of toilet paper",
                                "Spare ball", "First-aid kit",
                                "Game schedule printout", "Hotel confirmation",
                                "Driver / contact list"]),
        ("OVERNIGHT TOURNAMENT", ["Suitcase + toiletries", "Pajamas",
                                  "Casual outfit", "Recovery snacks",
                                  "Hydration pack", "Compression sleeves",
                                  "Foam roller", "Tablet + headphones",
                                  "Homework folder", "Spare phone charger"]),
        ("SUMMER CAMP", ["Cleats + indoor shoes", "Multiple training jerseys",
                         "Shorts (x5)", "Sun hat", "Sunscreen", "Insect repellent",
                         "Water bottle (large)", "Notebook + pen",
                         "Refillable snacks", "Inhaler / meds"]),
        ("WINTER TRAINING", ["Indoor turf shoes", "Long-sleeve base layer",
                             "Track pants", "Beanie", "Gloves",
                             "Water bottle (insulated)", "Layered jacket",
                             "Indoor ball", "Hand warmers"]),
    ]

    # Lay 6 lists across in pairs of (item, check) columns
    col_pairs = [(1, 2), (3, 4), (5, 6), (7, 8), (9, 10), (11, 12)]
    for (title, items), (cn, cc) in zip(lists, col_pairs):
        merge_set(ws,
                  f"{get_column_letter(cn)}4:{get_column_letter(cc)}4",
                  title, "th")
        ws.row_dimensions[4].height = 30
        for i, item in enumerate(items):
            r = 5 + i
            ci = ws.cell(row=r, column=cn, value=item)
            ci.style = "td_left"
            chk = ws.cell(row=r, column=cc, value="☐")
            chk.style = "td"
            chk.font = Font(size=14, bold=True, color=PRIMARY)
            if i % 2 == 1:
                for c in (cn, cc):
                    ws.cell(row=r, column=c).fill = fill(MUTED_ROW)


# ===========================================================================
# Sheet 12: Mileage & Travel
# ===========================================================================
def build_mileage(wb: Workbook) -> None:
    ws = wb.create_sheet("Mileage")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [12, 26, 12, 14, 14, 14, 14, 14])
    page_header(ws, "H", "🗺  MILEAGE & TRAVEL",
                "Miles, fuel, hotel, parking, tolls — auto-summed monthly.")
    headers = ["Date", "Destination", "Miles", "Fuel $",
               "Hotel $", "Parking", "Tolls", "Total"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)

    rows = [
        (d(-21), "Riverside Park (home)",   12,  None,   0,    0,    0),
        (d(-14), "Lincoln Complex",         45,  None,   0,    8,    3),
        (d(-7),  "Eastside Sports Ctr",     38,  None,   0,    5,    2),
        (d(-3),  "Memorial Field",          22,  None,   0,    0,    0),
        (d(2),   "Northside Park",          52,  None,   0,    10,   0),
        (d(9),   "Riverside Park (home)",   12,  None,   0,    0,    0),
        (d(12),  "City Cup — Lincoln",      45,  None, 220,    18,   8),
        (d(16),  "City Cup — Lincoln",      45,  None, 220,    18,   8),
        (d(34),  "Riverbend Sports",       145,  None, 280,    25,  14),
        (d(60),  "Capitol Fields",         210,  None, 340,    30,  18),
    ]
    start = 5
    end = start + 40 - 1
    for i, row in enumerate(rows):
        r = start + i
        date, dest, miles, _, hotel, parking, tolls = row
        ws.cell(row=r, column=1, value=date)
        ws.cell(row=r, column=2, value=dest)
        ws.cell(row=r, column=3, value=miles)
        ws.cell(row=r, column=5, value=hotel)
        ws.cell(row=r, column=6, value=parking)
        ws.cell(row=r, column=7, value=tolls)

    for r in range(start, end + 1):
        ws.cell(row=r, column=4, value=f'=IF(C{r}="","",C{r}*FuelPerMile)')
        ws.cell(row=r, column=8,
                value=f'=IF(C{r}="","",D{r}+IFERROR(E{r},0)+IFERROR(F{r},0)+IFERROR(G{r},0))')

    style_data_rows(ws, start, end, 8,
                    text_left_cols={2},
                    money_cols={4, 5, 6, 7, 8},
                    int_cols={3},
                    date_cols={1})

    # Monthly totals KPI row at top of side panel could go to dashboard
    wb.defined_names["MileageMiles"] = DefinedName(
        "MileageMiles", attr_text=f"Mileage!$C${start}:$C${end}"
    )
    wb.defined_names["MileageTotal"] = DefinedName(
        "MileageTotal", attr_text=f"Mileage!$H${start}:$H${end}"
    )
    wb.defined_names["MileageDate"] = DefinedName(
        "MileageDate", attr_text=f"Mileage!$A${start}:$A${end}"
    )


# ===========================================================================
# Sheet 13: Team Communication
# ===========================================================================
def build_communication(wb: Workbook) -> None:
    ws = wb.create_sheet("Communication")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [12, 22, 32, 18, 16])
    page_header(ws, "E", "💬  TEAM COMMUNICATION",
                "Coach messages · events · volunteers · fundraisers · tasks.")

    sections = [
        ("COACH MESSAGES", [
            (dt.date(2026, 8, 14), "Coach Daniels", "Welcome message + season expectations. Reply with confirm.",         "Email",        "Acknowledged"),
            (dt.date(2026, 8, 20), "Coach Daniels", "Bring shin guards + water to every practice. Practice 5:30pm sharp.","WhatsApp",     "Read"),
            (dt.date(2026, 9,  1), "Coach Daniels", "Game schedule attached. Print and put on fridge.",                   "Email + PDF",  "Saved"),
        ]),
        ("TEAM EVENTS", [
            (dt.date(2026, 8, 27), "Team Picture Day",           "Wear home jersey. Arrive 9:00am Riverside Park.",      "Required",     "Confirmed"),
            (dt.date(2026, 9, 14), "Team BBQ at the field",      "$10/family. Bring side dish.",                          "Optional",     "RSVP'd"),
            (dt.date(2026,10, 18), "Pumpkin Patch fundraiser",   "Volunteer parking lot helpers needed.",                 "Volunteer",    "Pending"),
        ]),
        ("PARENT VOLUNTEER ROSTER", [
            ("Concession stand",    "Sat 9/12",  "Sarah R. + James C.",  "Lead",     "Confirmed"),
            ("Field setup",         "Sun 9/13",  "Marcus B.",            "Setup",    "Confirmed"),
            ("Snack lead",          "Sat 9/19",  "Priya M.",             "Lead",     "Confirmed"),
            ("Photo coordinator",   "Aug 27",    "Aisha P.",             "Lead",     "Confirmed"),
        ]),
        ("FUNDRAISERS", [
            ("Pumpkin Patch",       "Oct 18",    "Goal $3,500",  "Per-family $200", "In progress"),
            ("Spirit wear sale",    "Sep 1–15",  "Goal $1,800",  "Margin 40%",      "Open"),
            ("Restaurant night",    "Sep 22",    "Goal $600",    "Chipotle 25%",    "Scheduled"),
        ]),
        ("TEAM TASKS", [
            ("Order new GK gloves",        "Sarah R.",      "Aug 22",   "—", "Done"),
            ("Print updated schedule",     "Priya M.",      "Sep 1",    "—", "Open"),
            ("Confirm tournament hotel",   "Marcus B.",     "Sep 5",    "—", "Open"),
            ("Update emergency contacts",  "All families",  "Sep 10",   "—", "In progress"),
        ]),
    ]
    row = 4
    for title, items in sections:
        merge_set(ws, f"A{row}:E{row}", title, "section")
        ws.row_dimensions[row].height = 22
        row += 1
        for h in ["Date / Owner", "From / Topic", "Detail", "Type / Role", "Status"]:
            pass  # placeholder so the headers below are consistent

        # Bespoke headers per section type
        if "COACH" in title:
            cols = ["Date", "From", "Message", "Channel", "Status"]
        elif "EVENTS" in title:
            cols = ["Date", "Event", "Detail", "Type", "Status"]
        elif "VOLUNTEER" in title:
            cols = ["Task", "When", "Owner(s)", "Role", "Status"]
        elif "FUNDRAISER" in title:
            cols = ["Activity", "When", "Goal", "Detail", "Status"]
        else:  # TASKS
            cols = ["Task", "Owner", "Due", "Detail", "Status"]
        for i, h in enumerate(cols, 1):
            ws.cell(row=row, column=i, value=h).style = "th"
        ws.row_dimensions[row].height = 26
        row += 1
        for i, item in enumerate(items):
            for ci, val in enumerate(item, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.style = "td_left" if ci in (2, 3) else "td"
                cell.fill = fill(MUTED_ROW if i % 2 == 1 else BG)
                if isinstance(val, dt.date):
                    cell.number_format = "mm/dd/yyyy"
            ws.row_dimensions[row].height = 22
            row += 1
        row += 1


# ===========================================================================
# Sheet 1: Family Command Dashboard
# ===========================================================================
def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [2, 18, 18, 18, 18, 18, 18, 18, 18, 2])
    ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:J1", "  ⚽  SOCCER MOM COMMAND CENTER", "title")
    ws.row_dimensions[2].height = 22
    merge_set(ws, "A2:J2",
              "  Plan · Track · Travel · Fuel · Win — one premium dashboard for the whole soccer season.",
              "subtitle")

    # KPI row 1 (4 cards)
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 56
    kpi_row1 = [
        ("⚽ Games This Month",
         '=SUMPRODUCT((GameDates<>"")*(MONTH(IFERROR(GameDates,0))=MONTH(TODAY()))*(YEAR(IFERROR(GameDates,0))=YEAR(TODAY())))',
         "kpi_value"),
        ("🏃 Practices This Week",
         '=SUMPRODUCT((PracDates<>"")*(PracDates-TODAY()>=0)*(PracDates-TODAY()<=7))',
         "kpi_value"),
        ("💰 Monthly Budget",
         '=MonthlyBudget',
         "kpi_money"),
        ("💵 Budget Remaining",
         '=MonthlyBudget-BudgetTotalActual',
         "kpi_money"),
    ]
    col = 2
    for label, formula, val_style in kpi_row1:
        merge_set(ws, f"{get_column_letter(col)}4:{get_column_letter(col+1)}4",
                  label, "kpi_label")
        merge_set(ws, f"{get_column_letter(col)}5:{get_column_letter(col+1)}5",
                  formula, val_style)
        for r in (4, 5):
            for cc in range(col, col + 2):
                ws.cell(row=r, column=cc).border = BOX
        col += 2

    # KPI row 2
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 56
    kpi_row2 = [
        ("🏆 Upcoming Tournaments",
         '=SUMPRODUCT((TourStart<>"")*(TourStart-TODAY()>=0))',
         "kpi_value"),
        ("🎽 Equipment Needed",
         '=SUMPRODUCT((EquipItem<>"")*(EquipPurchased<>"Yes"))',
         "kpi_value"),
        ("⚠ Family Conflicts",
         '=SUMPRODUCT((GameDates<>"")*(COUNTIF(PracDates,GameDates)>0))',
         "kpi_value"),
        ("✅ Practice Attendance",
         '=IFERROR(SUMPRODUCT((PracAttendance="Present")*1)/SUMPRODUCT((PracAttendance<>"")*(PracAttendance<>"—")*1),0)',
         "kpi_pct"),
    ]
    col = 2
    for label, formula, val_style in kpi_row2:
        merge_set(ws, f"{get_column_letter(col)}7:{get_column_letter(col+1)}7",
                  label, "kpi_label")
        merge_set(ws, f"{get_column_letter(col)}8:{get_column_letter(col+1)}8",
                  formula, val_style)
        for r in (7, 8):
            for cc in range(col, col + 2):
                ws.cell(row=r, column=cc).border = BOX
        col += 2

    # Quick navigation
    ws.row_dimensions[10].height = 26
    merge_set(ws, "B10:I10", "QUICK NAVIGATION", "section")
    ws.row_dimensions[11].height = 30
    nav = ["Schedule", "Budget", "Equipment", "Roster",
           "Travel", "Meals", "Carpool", "Comms"]
    for i, name in enumerate(nav):
        cell = ws.cell(row=11, column=2 + i, value=name)
        cell.fill = fill(PRIMARY)
        cell.font = Font(size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX

    # Analytics
    ws.row_dimensions[13].height = 26
    merge_set(ws, "B13:I13", "ANALYTICS", "section")

    # Donut: Budget by Category
    donut = DoughnutChart()
    donut.title = "Spending by Category"
    donut.height = 9; donut.width = 12
    end = 4 + len(EXPENSE_CATEGORIES)
    donut.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=end),
                   titles_from_data=True)
    donut.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=end))
    donut.dataLabels = no_labels()
    ws.add_chart(donut, "B14")

    # Bar: Budget vs Actual
    bar = BarChart()
    bar.type = "col"
    bar.title = "Budget vs Actual"
    bar.height = 9; bar.width = 12
    bar.add_data(Reference(wb["Budget"], min_col=2, min_row=4, max_row=end),
                 titles_from_data=True)
    bar.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=end),
                 titles_from_data=True)
    bar.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=end))
    ws.add_chart(bar, "H14")

    # Footer
    ws.row_dimensions[34].height = 26
    merge_set(ws, "B34:I34",
              "Soccer Mom Command Center v1.0  ·  Edit Settings → Season · Budget · Player Count  ·  Add games on Schedule",
              "subtitle")


# ===========================================================================
# Build
# ===========================================================================
def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    # Build in dependency order: settings + data sheets before dashboard
    build_settings(wb)
    build_players(wb)
    build_schedule(wb)
    build_practices(wb)
    build_budget(wb)
    build_equipment(wb)
    build_tournaments(wb)
    build_carpool(wb)
    build_roster(wb)
    build_meals(wb)
    build_packing(wb)
    build_mileage(wb)
    build_communication(wb)
    build_dashboard(wb)

    order = ["Dashboard", "Players", "Schedule", "Practices", "Budget",
             "Equipment", "Tournaments", "Carpool", "Roster", "Meals",
             "Packing", "Mileage", "Communication", "Settings"]
    wb._sheets = [wb[name] for name in order]

    tab_colors = {
        "Dashboard": PRIMARY,
        "Players":   ACCENT,
        "Schedule":  HIGHLIGHT,
        "Practices": HIGHLIGHT,
        "Budget":    PRIMARY,
        "Equipment": ACCENT,
        "Tournaments": PRIMARY,
        "Carpool":   ACCENT,
        "Roster":    HIGHLIGHT,
        "Meals":     ACCENT,
        "Packing":   ACCENT,
        "Mileage":   HIGHLIGHT,
        "Communication": PRIMARY,
        "Settings":  SURFACE,
    }
    for name, color in tab_colors.items():
        wb[name].sheet_properties.tabColor = color

    out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(out_dir, "Soccer_Mom_Command_Center.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
