"""Build Cleaning & Service Business Command Center™ — The Service-Business Operating System.

14 tabs · a premium cleaning / service-business operating system in Google Sheets &
Excel. Dashboard, a job-P&L engine (did this job actually pay, and what per hour?),
services & pricing, a client roster that becomes recurring revenue, a schedule, leads
& quotes, supplies, team & labor, mileage, expenses, reviews and a monthly summary —
one dashboard. Price every job, keep the calendar full, and grow recurring revenue.

Run: python3 build_xlsx.py   ->  ../Cleaning_Business_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
FREQ = ["Weekly", "Biweekly", "Monthly", "One-time"]
JOBSTATUS = ["Scheduled", "Done", "Invoiced", "Paid", "Cancelled"]
LEADSTAGE = ["New", "Quoted", "Booked", "Lost"]

MARGIN_GOAL = 0.45
HOUR_GOAL = 32
PROFIT_GOAL = 2500
MRR_GOAL = 4000
CLIENT_GOAL = 30

# Job P&L — flagship standard clean
JOB_PRICE = 180
JOB_SUPPLIES = 18
JOB_LABOR = 50
JOB_TRAVEL = 12
JOB_HOURS = 2.5

# Services & pricing: (service, price, hours)
SERVICES = [
    ("Standard Home Clean (3BR)", 180, 2.5), ("Deep Clean", 320, 5.0), ("Move-out Clean", 380, 6.0),
    ("Office Clean (weekly)", 220, 3.0), ("Apartment / Studio", 120, 1.5), ("Add-on: Inside Oven", 45, 0.75),
]

# Clients — recurring roster: (client, frequency, price, jobs/mo)
CLIENTS = [
    ("The Hardings", "Weekly", 180, 4), ("Cedar Offices", "Weekly", 180, 4), ("Marlow Home", "Weekly", 180, 4),
    ("Bexley Flat", "Biweekly", 180, 2), ("The Ortons", "Biweekly", 180, 2), ("Quinn Studio", "Biweekly", 180, 2),
    ("Vale House", "Biweekly", 180, 2), ("The Abbots", "Monthly", 180, 1), ("Rowan Clinic", "Monthly", 180, 1),
    ("Nash Loft", "Monthly", 180, 1), ("Piper Home", "Monthly", 180, 1), ("The Fenns", "Monthly", 180, 1),
]
ONE_TIME_REV = 1500
ONE_TIME_JOBS = 6

# Schedule — this week: (day, client, service, cleaner, status)
SCHEDULE = [
    ("Mon", "The Hardings", "Standard", "Ava", "Paid"), ("Mon", "Cedar Offices", "Office", "Ava", "Done"),
    ("Tue", "Bexley Flat", "Standard", "Leo", "Scheduled"), ("Wed", "Deep Clean — Voss", "Deep Clean", "Ava+Leo", "Scheduled"),
    ("Thu", "Marlow Home", "Standard", "Ava", "Scheduled"), ("Fri", "Move-out — Reyes", "Move-out", "Ava+Leo", "Scheduled"),
]

# Leads & quotes: (lead, source, quote, stage)
LEADS = [
    ("Tanaka move-out", "Google", 380, "Quoted"), ("Bright Dental", "Referral", 260, "Booked"),
    ("Okafor deep clean", "Instagram", 320, "New"), ("Hensley weekly", "Yard sign", 180, "Quoted"),
]

# Supplies & inventory: (item, par, unit cost)
SUPPLIES = [
    ("All-purpose cleaner", 6, 3.20), ("Microfiber cloths (pk)", 4, 8.50), ("Glass cleaner", 4, 2.80),
    ("Disinfectant", 5, 4.10), ("Trash bags (box)", 3, 9.00), ("Vacuum bags", 8, 1.40),
    ("Gloves (box)", 4, 6.50), ("Floor cleaner", 3, 5.20),
]

# Team & labor: (cleaner, role, rate, hours)
TEAM = [
    ("Ava (owner)", "Lead", 0, 68), ("Leo", "Cleaner", 20, 52), ("Mara", "Cleaner", 19, 30),
]

# Mileage & vehicle: (date, route, miles, cost)
MILEAGE = [
    ("Mon", "Harding + Cedar", 22, 14.74), ("Tue", "Bexley", 12, 8.04),
    ("Wed", "Voss deep clean", 18, 12.06), ("Thu", "Marlow", 14, 9.38), ("Fri", "Reyes move-out", 26, 17.42),
]

# Expenses (monthly): (item, amount)
EXPENSES = [
    ("Supplies", 560), ("Helper labor", 1800), ("Fuel & vehicle", 380),
    ("Insurance & bonding", 120), ("Marketing & leads", 150), ("Software & phone", 90),
]

# Reviews & referrals: (client, rating, source)
REVIEWS = [
    ("The Hardings", 5, "Google"), ("Cedar Offices", 5, "Referral"), ("Bexley Flat", 4, "Google"),
    ("The Ortons", 5, "Instagram"), ("Vale House", 5, "Referral"),
]

# Monthly summary: (month, revenue, expenses)
MONTHS = [("Jul", 6000, 3100), ("Aug", 6300, 3200), ("Sep", 6600, 3300),
          ("Oct", 6100, 3150), ("Nov", 5700, 3000), ("Dec", 6900, 3400)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Business name", "Bright & Tidy Co.", None, "Business"),
        ("Owner", "Ava", None, "Owner"),
        ("Job-margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Profit-per-hour goal ($)", HOUR_GOAL, '"$"#,##0', "HourGoal"),
        ("Monthly profit goal", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Recurring revenue goal", MRR_GOAL, '"$"#,##0', "MRRGoal"),
        ("Recurring-client goal", CLIENT_GOAL, "#,##0", "ClientGoal"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Frequency", FREQ, "FreqList"), ("F", "Job status", JOBSTATUS, "JobStatusList"),
             ("G", "Lead stage", LEADSTAGE, "LeadStageList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🧽  CLEANING & SERVICE BUSINESS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Price every job, keep the calendar full, and grow recurring revenue.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE SERVICE BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A service business lives or dies on two numbers: whether each job actually pays after supplies, "
                      "labor and travel, and how much recurring revenue keeps the lights on. This makes both visible: a "
                      "job-P&L engine that shows your profit and your profit per hour on any job, and a client roster "
                      "that turns weekly, biweekly and monthly clients into recurring revenue. Build your services & "
                      "pricing, run your schedule, track leads, supplies, team and mileage, and keep your books — all in "
                      "ONE premium Google Sheets & Excel system built for cleaners and home-service pros.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your margin, profit-per-hour & recurring goals.",
             "2.  Price a job in Job P&L — supplies, labor & travel → profit per hour.",
             "3.  Build Services & Pricing and add your Clients by frequency.",
             "4.  Run the Schedule and track Leads & Quotes.",
             "5.  Track supplies, team, mileage and expenses.",
             "6.  Check the Dashboard: profit, recurring revenue & a Service Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional cleaning business (Bright & Tidy Co.) is included so you can see how it all "
               "connects — just type over it with your own jobs and clients. Profit per hour and recurring revenue are "
               "the two numbers that decide whether a service business is worth running, and they roll into a live "
               "Service Score. Twelve matching printable pages (job quote, service price list, schedule, supply list & "
               "more) are included. This is a business tool, not financial, legal or tax advice — confirm figures with "
               "your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Price by the hour of profit, and build the recurring book — that's a service business that lasts.", "section_gold")


# ===========================================================================
def build_jobpnl(wb):
    ws = wb.create_sheet("Job P&L"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "🧮  JOB P&L",
                "Price a job to the profit — supplies, labor and travel out, and what you make per hour.")
    ws.cell(row=5, column=2, value="JOB").style = "section_gold"
    ws.cell(row=5, column=3, value="Standard Home Clean (3BR)").font = Font(bold=True, color=PRIMARY)
    ws.cell(row=6, column=2, value="Job price").style = "field_label"
    cp = ws.cell(row=6, column=3, value=JOB_PRICE); cp.style = "input"; cp.number_format = '"$"#,##0'
    cell_name(wb, "JobPrice", "Job P&L", "$C$6")
    rows = [("− Supplies", JOB_SUPPLIES, "JobSupplies"), ("− Labor (helper)", JOB_LABOR, "JobLabor"),
            ("− Travel & fuel", JOB_TRAVEL, "JobTravel")]
    for i, (lab, val, nm) in enumerate(rows):
        r = 7 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"; c.number_format = '"$"#,##0'
        cell_name(wb, nm, "Job P&L", f"$C${r}")
    ws.cell(row=10, column=2, value="= JOB COSTS").style = "th"
    cc = ws.cell(row=10, column=3, value="=JobSupplies+JobLabor+JobTravel"); cc.style = "td"; cc.font = Font(bold=True, color=DANGER); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0'
    cell_name(wb, "JobCosts", "Job P&L", "$C$10")
    ws.cell(row=11, column=2, value="= JOB PROFIT").style = "th"
    cpr = ws.cell(row=11, column=3, value="=JobPrice-JobCosts"); cpr.style = "td"; cpr.font = Font(bold=True, size=13, color=PRIMARY); cpr.fill = fill(MINT_BG); cpr.number_format = '"$"#,##0'
    cell_name(wb, "JobProfit", "Job P&L", "$C$11")
    ws.cell(row=13, column=2, value="Hours on the job (all cleaners)").style = "field_label"
    ch = ws.cell(row=13, column=3, value=JOB_HOURS); ch.style = "input"; ch.number_format = "0.0"
    cell_name(wb, "JobHours", "Job P&L", "$C$13")
    ws.cell(row=14, column=2, value="= PROFIT PER HOUR").style = "th"
    cph = ws.cell(row=14, column=3, value="=IFERROR(JobProfit/JobHours,0)"); cph.style = "td"; cph.font = Font(bold=True, size=12, color=PRIMARY); cph.fill = fill(MINT_BG); cph.number_format = '"$"#,##0.00'
    cell_name(wb, "ProfitHour", "Job P&L", "$C$14")
    ws.cell(row=15, column=2, value="Job margin").style = "field_label"
    cm = ws.cell(row=15, column=3, value="=IFERROR(JobProfit/JobPrice,0)"); cm.style = "field_value"; cm.number_format = "0%"; cm.fill = fill(MINT_BG)
    cell_name(wb, "JobMargin", "Job P&L", "$C$15")
    ws.cell(row=17, column=2, value="Copy this for every service — price by the hour of profit, not the sticker.").style = "section"


def build_services(wb):
    ws = wb.create_sheet("Services & Pricing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 12, 14, 2])
    luxe_header(ws, "E", "📋  SERVICES & PRICING",
                "Your menu — price, hours and the rate per hour, so every quote is priced for profit.")
    table_headers(ws, 4, ["Service", "Price", "Hours", "$ / Hour"], start_col=2)
    start = L0
    for i, (svc, price, hrs) in enumerate(SERVICES):
        r = start + i
        ws.cell(row=r, column=2, value=svc).style = "td_left"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        ch = ws.cell(row=r, column=4, value=hrs); ch.style = "input"; ch.number_format = "0.0"
        cr = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cr.style = "td"; cr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SERVICES) - 1
    nrange(wb, "SvcName", "Services & Pricing", "B", start, end)
    nrange(wb, "SvcPrice", "Services & Pricing", "C", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=40, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=60, mid_color="FFFFF3CD",
                       end_type="num", end_value=80, end_color="FF" + HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_clients(wb):
    ws = wb.create_sheet("Clients"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 16, 12, 12, 16, 2])
    luxe_header(ws, "F", "🔁  CLIENTS",
                "Your recurring book — frequency, price and jobs a month turn into recurring revenue.")
    table_headers(ws, 4, ["Client", "Frequency", "Price", "Jobs/mo", "Monthly Rev"], start_col=2)
    start = L0
    for i, (client, freq, price, jobs) in enumerate(CLIENTS):
        r = start + i
        ws.cell(row=r, column=2, value=client).style = "td_left"
        ws.cell(row=r, column=3, value=freq).style = "td"
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cj = ws.cell(row=r, column=5, value=jobs); cj.style = "input"; cj.number_format = "#,##0"
        cr = ws.cell(row=r, column=6, value=f"=D{r}*E{r}"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "FreqList")
    end = start + len(CLIENTS) - 1
    nrange(wb, "ClientName", "Clients", "B", start, end)
    nrange(wb, "ClientJobs", "Clients", "E", start, end)
    nrange(wb, "ClientRev", "Clients", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="RECURRING (MRR)").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cj = ws.cell(row=tot, column=5, value="=SUM(ClientJobs)"); cj.style = "td"; cj.font = Font(bold=True, color=PRIMARY); cj.fill = fill(SURFACE); cj.number_format = "#,##0"
    cr = ws.cell(row=tot, column=6, value="=SUM(ClientRev)"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0'
    cell_name(wb, "MRR", "Clients", f"$F${tot}")
    cell_name(wb, "RecurringJobs", "Clients", f"$E${tot}")
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Recurring clients").style = "field_label"
    c = ws.cell(row=sr, column=6, value="=COUNTA(ClientName)"); c.style = "field_value"; c.number_format = "#,##0"; c.fill = fill(MINT_BG)
    cell_name(wb, "RecurringClients", "Clients", f"$F${sr}")
    ws.cell(row=sr + 1, column=2, value="One-time / deep-clean revenue").style = "field_label"
    co = ws.cell(row=sr + 1, column=6, value=ONE_TIME_REV); co.style = "input"; co.number_format = '"$"#,##0'
    cell_name(wb, "OneTimeRev", "Clients", f"$F${sr+1}")
    ws.cell(row=sr + 2, column=2, value="One-time jobs this month").style = "field_label"
    coj = ws.cell(row=sr + 2, column=6, value=ONE_TIME_JOBS); coj.style = "input"; coj.number_format = "#,##0"
    cell_name(wb, "OneTimeJobs", "Clients", f"$F${sr+2}")
    ws.cell(row=sr + 3, column=2, value="= MONTHLY REVENUE (recurring + one-time)").style = "th"
    cmr = ws.cell(row=sr + 3, column=6, value="=MRR+OneTimeRev"); cmr.style = "td"; cmr.font = Font(bold=True, size=12, color=PRIMARY); cmr.fill = fill(MINT_BG); cmr.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyRevenue", "Clients", f"$F${sr+3}")
    ws.cell(row=sr + 4, column=2, value="Jobs this month").style = "field_label"
    cjm = ws.cell(row=sr + 4, column=6, value="=RecurringJobs+OneTimeJobs"); cjm.style = "field_value"; cjm.number_format = "#,##0"; cjm.fill = fill(MINT_BG)
    cell_name(wb, "JobsMonth", "Clients", f"$F${sr+4}")
    ws.cell(row=sr + 5, column=2, value="Average job value").style = "field_label"
    cav = ws.cell(row=sr + 5, column=6, value="=IFERROR(MonthlyRevenue/JobsMonth,0)"); cav.style = "field_value"; cav.number_format = '"$"#,##0'; cav.fill = fill(MINT_BG)
    cell_name(wb, "AvgJob", "Clients", f"$F${sr+5}")
    ws.freeze_panes = "A5"


def build_schedule(wb):
    ws, start, end = build_log(
        wb, "Schedule", "🗓", "SCHEDULE",
        "This week's jobs — who, what service, which cleaner and the status.",
        ["Day", "Client / Job", "Service", "Cleaner", "Status"],
        SCHEDULE, [2, 12, 24, 16, 14, 14, 2], text_left={2, 3, 4, 5}, reserved=28, start_col=2,
        validations=[("F", "JobStatusList")])


def build_leads(wb):
    ws, start, end = build_log(
        wb, "Leads & Quotes", "🔍", "LEADS & QUOTES",
        "Every enquiry — source, quote and stage, so nothing falls through the cracks.",
        ["Lead", "Source", "Quote", "Stage"],
        LEADS, [2, 22, 16, 14, 14, 2], text_left={2, 3}, money={4}, reserved=24, start_col=2,
        validations=[("E", "LeadStageList")])
    nrange(wb, "LeadQuote", "Leads & Quotes", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="PIPELINE VALUE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(LeadQuote)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)


def build_supplies(wb):
    ws = wb.create_sheet("Supplies"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 14, 14, 2])
    luxe_header(ws, "E", "🧴  SUPPLIES",
                "Your par levels — what to keep on the van and what a full restock costs.")
    table_headers(ws, 4, ["Item", "Par", "Unit Cost", "Restock $"], start_col=2)
    start = L0
    for i, (item, par, cost) in enumerate(SUPPLIES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=par); cp.style = "input"; cp.number_format = "#,##0"
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cr = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SUPPLIES) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="FULL RESTOCK").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cr = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = '"$"#,##0.00'
    ws.freeze_panes = "A5"


def build_team(wb):
    ws = wb.create_sheet("Team & Labor"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 16, 12, 12, 14, 2])
    luxe_header(ws, "F", "👥  TEAM & LABOR",
                "Your crew — role, rate and hours, so labor cost is never a mystery.")
    table_headers(ws, 4, ["Cleaner", "Role", "Rate/hr", "Hours", "Labor $"], start_col=2)
    start = L0
    for i, (name, role, rate, hrs) in enumerate(TEAM):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=role).style = "td"
        crt = ws.cell(row=r, column=4, value=rate); crt.style = "input"; crt.number_format = '"$"#,##0'
        chh = ws.cell(row=r, column=5, value=hrs); chh.style = "input"; chh.number_format = "#,##0"
        cl = ws.cell(row=r, column=6, value=f"=D{r}*E{r}"); cl.style = "td"; cl.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(TEAM) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL LABOR").style = "th"
    for c in (3, 4, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cl = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cl.style = "td"; cl.font = Font(bold=True, color=DANGER); cl.fill = fill(SURFACE); cl.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


def build_mileage(wb):
    ws, start, end = build_log(
        wb, "Mileage & Vehicle", "🚐", "MILEAGE & VEHICLE",
        "Every route — miles and cost, deductible and useful for pricing travel.",
        ["Day", "Route", "Miles", "Cost"],
        MILEAGE, [2, 12, 24, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)


def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  EXPENSES",
                "Monthly costs — supplies, labor, fuel and overhead, netted against revenue.")
    table_headers(ws, 4, ["Expense", "Amount"], start_col=2)
    start = L0
    for i, (item, amt) in enumerate(EXPENSES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EXPENSES) - 1
    nrange(wb, "ExpAmt", "Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Expenses", f"$C${tot}")
    nr = tot + 2
    ws.cell(row=nr, column=2, value="= MONTHLY PROFIT").style = "th"
    cn = ws.cell(row=nr, column=3, value="=MonthlyRevenue-ExpTotal"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetProfit", "Expenses", f"$C${nr}")
    ws.freeze_panes = "A5"


def build_reviews(wb):
    ws, start, end = build_log(
        wb, "Reviews & Referrals", "⭐", "REVIEWS & REFERRALS",
        "Every review and where it came from — referrals are the cheapest client you'll ever get.",
        ["Client", "Rating", "Source"],
        REVIEWS, [2, 24, 12, 18, 2], text_left={2, 4}, dec={3}, reserved=24, start_col=2)
    nrange(wb, "RevRating", "Reviews & Referrals", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVG RATING").style = "th"
    c = ws.cell(row=tot, column=3, value="=IFERROR(AVERAGE(RevRating),0)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = "0.0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 14, 2])
    luxe_header(ws, "E", "📈  MONTHLY SUMMARY",
                "Revenue, costs & profit by month — watch the recurring book compound.")
    table_headers(ws, 4, ["Month", "Revenue", "Costs", "Profit"], start_col=2)
    start = L0
    for i, (m, rev, exp) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ci = ws.cell(row=r, column=3, value=rev); ci.style = "input"; ci.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=4, value=exp); ce.style = "input"; ce.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthProfit", "Monthly Summary", "E", start, end)
    ws.add_chart(_barchart(ws, "Profit by Month", start, end, 5, 2), "G4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🧽  CLEANING & SERVICE BUSINESS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Profit per job & hour, recurring revenue & a Service Score — your whole business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("JOB PRICE", "=JobPrice", "money"),
        ("JOB COSTS", "=JobCosts", "money"),
        ("JOB PROFIT", "=JobProfit", "money"),
        ("JOB MARGIN", "=JobMargin", "pct"),
        ("PROFIT / HOUR", "=ProfitHour", "money2"),
        ("JOBS / MONTH", "=JobsMonth", "num"),
    ]
    row2 = [
        ("MONTHLY REVENUE", "=MonthlyRevenue", "money"),
        ("RECURRING (MRR)", "=MRR", "money"),
        ("ACTIVE CLIENTS", "=RecurringClients", "num"),
        ("MONTHLY PROFIT", "=NetProfit", "money"),
        ("AVG JOB", "=AvgJob", "money"),
        ("SERVICE SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "SERVICE HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "PROFIT BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Job margin healthy", "=IFERROR(MIN(JobMargin/MarginGoal,1),0)"),
        ("Profit/hour on target", "=IFERROR(MIN(ProfitHour/HourGoal,1),0)"),
        ("Services priced", "=IFERROR(COUNTIF(SvcPrice,\">0\")/COUNTA(SvcName),0)"),
        ("Profitable", "=IFERROR(MIN(NetProfit/ProfitGoal,1),0)"),
        ("Recurring revenue healthy", "=IFERROR(MIN(MRR/MRRGoal,1),0)"),
        ("Client base growing", "=IFERROR(MIN(RecurringClients/ClientGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Profit by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=5, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Cleaning & Service Business Command Center™ — price every job, grow recurring revenue.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_jobpnl(wb); build_services(wb)
    build_clients(wb); build_schedule(wb); build_leads(wb); build_supplies(wb)
    build_team(wb); build_mileage(wb); build_expenses(wb); build_reviews(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Job P&L", "Services & Pricing", "Clients", "Schedule",
             "Leads & Quotes", "Supplies", "Team & Labor", "Mileage & Vehicle", "Expenses",
             "Reviews & Referrals", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Cleaning_Business_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
