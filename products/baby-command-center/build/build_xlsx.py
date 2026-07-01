"""Build Baby Command Center™ — The Ultimate Baby Care & Family Organization System.

21 sheets + Welcome · a warm, premium baby operating system in Excel & Sheets.

Organization & record-keeping tool only — NOT a medical device or a substitute
for professional healthcare (see the Welcome disclaimer).

Run: python3 build_xlsx.py   ->  ../Baby_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

FEED_TYPES = ["Breast (L)", "Breast (R)", "Bottle - BM", "Bottle - Formula", "Pumping", "Solids"]
SLEEP_CATS = ["Nap", "Overnight"]
APPT_TYPES = ["Pediatrician", "Dentist", "Specialist", "Therapy", "Vaccination", "Parent"]
EXPENSE_CATS = ["Diapers", "Formula", "Food", "Clothing", "Toys", "Nursery",
                "Furniture", "Childcare", "Healthcare", "Insurance", "Baby Gear", "Miscellaneous"]
MILESTONE_CATS = ["Motor", "Language", "Social", "Cognitive", "Feeding", "Sleep", "Teeth", "Walking"]
ROUTINE_TYPES = ["Morning", "Nap", "Bedtime", "Feeding", "Play"]
CLOTHING_SIZES = ["Newborn", "0-3m", "3-6m", "6-9m", "9-12m", "12-18m", "18-24m"]
SUPPLY_CATS = ["Diapering", "Feeding", "Clothing", "Bath", "Health", "Nursery", "Toys", "Gear"]
DIAPER_TYPES = ["Wet", "Dirty", "Mixed", "Dry"]
STATUS = ["Done", "Due", "Upcoming", "Skipped"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set(); dates = dates or set(); pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula; vc.font = Font(size=18, bold=True, color=PRIMARY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pct": "0%", "days": "0", "dec": "0.0"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def day_summary(ws, wb, col_label, col_val, header_val, per_day_formula, name_prefix, first_row=5, days=7):
    """Write a 7-day summary block; returns (labels_range, vals_range)."""
    ws.cell(row=first_row - 1, column=col_label, value="LAST 7 DAYS").style = "section_gold"
    ws.cell(row=first_row, column=col_label, value="Date").style = "th"
    ws.cell(row=first_row, column=col_val, value=header_val).style = "th"
    for i in range(days):
        r = first_row + 1 + i
        dcell = ws.cell(row=r, column=col_label, value=dminus(days - 1 - i)); dcell.style = "td"; dcell.number_format = "mm/dd"
        vc = ws.cell(row=r, column=col_val, value=per_day_formula(get_column_letter(col_label) + str(r))); vc.style = "td"
    lab = f"{get_column_letter(col_label)}{first_row+1}:{get_column_letter(col_label)}{first_row+days}"
    val = f"{get_column_letter(col_val)}{first_row+1}:{get_column_letter(col_val)}{first_row+days}"
    wb.defined_names[name_prefix + "Labels"] = DefinedName(name_prefix + "Labels", attr_text=f"'{ws.title}'!{lab}")
    wb.defined_names[name_prefix + "Vals"] = DefinedName(name_prefix + "Vals", attr_text=f"'{ws.title}'!{val}")
    return first_row + days


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set baby's details once — every dashboard follows. Edit the lists to fit your family.")
    merge_set(ws, "B5:C5", "BABY & BUDGET", "section")
    controls = [
        ("Baby Name", "Baby Rose", None, "BabyName"),
        ("Birth Date", dminus(125), "mm/dd/yyyy", "BirthDate"),
        ("Monthly Baby Budget", 650, '"$"#,##0', "MonthlyBudget"),
        ("Feedings / Day Goal", 8, "0", "FeedGoal"),
        ("Sleep / Day Goal (hrs)", 15, "0", "SleepGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Feeding Type", FEED_TYPES, "FeedTypeList"), ("F", "Sleep Category", SLEEP_CATS, "SleepCatList"),
             ("G", "Appointment Type", APPT_TYPES, "ApptTypeList"), ("H", "Expense Category", EXPENSE_CATS, "ExpenseCatList"),
             ("I", "Milestone Category", MILESTONE_CATS, "MilestoneCatList"), ("J", "Clothing Size", CLOTHING_SIZES, "SizeList"),
             ("K", "Supply Category", SUPPLY_CATS, "SupplyCatList"), ("L", "Status / Diaper", STATUS + ["—"] + DIAPER_TYPES + ["—"] + YESNO, "MiscList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    wb.defined_names["DiaperTypeList"] = DefinedName("DiaperTypeList", attr_text="Settings!$L$12:$L$15")
    wb.defined_names["StatusList"] = DefinedName("StatusList", attr_text="Settings!$L$7:$L$10")
    wb.defined_names["YesNoList"] = DefinedName("YesNoList", attr_text="Settings!$L$17:$L$18")
    wb.defined_names["RoutineTypeList"] = DefinedName("RoutineTypeList", attr_text="Settings!$G$7:$G$11")


# ===========================================================================
# Welcome (with disclaimer)
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 76, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  👶  BABY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate baby care & family organization system — your whole day in one place.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "WELCOME, PARENTS & CAREGIVERS", "section_gold")
    ws.merge_cells("B6:B8")
    ws["B6"].value = ("The newborn season is the busiest — and most precious — time. Baby Command Center™ "
                      "brings feeding, sleep, diapers, growth, health, budget, routines and memories into one "
                      "warm, organized dashboard, so everyone caring for your little one is on the same page and "
                      "you can spend less time tracking and more time snuggling.")
    ws["B6"].style = "body"
    for r in (6, 7, 8):
        ws.row_dimensions[r].height = 22
    merge_set(ws, "B10:B10", "START HERE", "section")
    steps = ["1.  Open Settings and add baby's name, birth date & monthly budget.",
             "2.  Fill the Baby Profile (pediatrician, allergies, emergency contacts).",
             "3.  Log feedings, sleep & diapers on their tabs — daily totals update themselves.",
             "4.  Track growth, milestones, appointments & the baby budget.",
             "5.  Watch the Executive Baby Dashboard bring it all together automatically."]
    for i, s in enumerate(steps):
        r = 11 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 18
    merge_set(ws, f"B{dr}:B{dr}", "  IMPORTANT — PLEASE READ", "th")
    ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+4}")
    c = ws[f"B{dr+1}"]
    c.value = ("This workbook is designed as an organization and record-keeping tool for parents and caregivers. "
               "It is NOT a medical device and should not replace guidance from qualified healthcare professionals. "
               "Always follow your pediatrician's advice for feeding, sleep, growth and health decisions. In an "
               "emergency, contact your doctor or local emergency services right away.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 5):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+6}:B{dr+6}", "One organized home for your little one — you've got this.", "section_gold")


# ===========================================================================
# 2 — Baby Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 26, 6, 24, 26, 2])
    luxe_header(ws, "G", "👶  BABY PROFILE", "Everything about your little one in one secure place.")
    blocks = [
        ("BIRTH DETAILS", [("Baby Name", "=BabyName"), ("Birth Date", "=BirthDate"), ("Birth Time", "3:42 AM"),
                           ("Birth Weight", "7 lb 6 oz"), ("Birth Length", "20 in"), ("Blood Type", "O+")]),
        ("CARE TEAM", [("Pediatrician", "Dr. Ellis"), ("Clinic Phone", "(555) 210-7788"),
                       ("Dentist", "—"), ("Lactation Consultant", "Maria R."),
                       ("Insurance Provider", "____________"), ("Policy / Member #", "____________")]),
        ("HEALTH & EMERGENCY", [("Allergies", "None known"), ("Medical Notes", "Mild reflux — see pediatrician"),
                                ("Emergency Contact 1", "Mom — (555) 210-4521"), ("Emergency Contact 2", "Grandma — (555) 661-2048"),
                                ("Blood Type", "O+"), ("Preferred Hospital", "City Children's")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            cv = ws.cell(row=row, column=3, value=fields[i][1]); cv.style = "field_value"
            if fields[i][0] == "Birth Date":
                cv.number_format = "mm/dd/yyyy"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1


# ===========================================================================
# 4 — Feeding Command Center
# ===========================================================================
def build_feeding(wb):
    ws = wb.create_sheet("Feeding"); ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 11, 16, 11, 11, 22, 3, 12, 12])
    luxe_header(ws, "I", "🍼  FEEDING COMMAND CENTER", "Every feeding logged — daily counts & trends update automatically.")
    table_headers(ws, 4, ["Date", "Time", "Type", "Amount", "Duration", "Notes"])
    sample = []
    counts = {0: 7, 1: 8, 2: 7, 3: 6, 4: 8, 5: 7, 6: 8}
    times = ["6:00 AM", "9:00 AM", "12:00 PM", "3:00 PM", "6:00 PM", "8:30 PM", "11:00 PM", "2:00 AM"]
    for dd in range(0, 7):
        for k in range(counts[dd]):
            typ = FEED_TYPES[k % 4]
            amt = "4 oz" if "Bottle" in typ else ("15 min" if "Breast" in typ else "—")
            sample.append((dminus(dd), times[k % len(times)], typ, amt, "15 min", ""))
    start = L0; end = start + max(len(sample), 60) - 1
    for i, row in enumerate(sample):
        for ci, val in enumerate(row, 1):
            ws.cell(row=start + i, column=ci, value=val)
    style_rows(ws, start, end, 6, text_left={3, 6}, dates={1})
    add_dv(ws, f"C{start}:C{end}", "FeedTypeList")
    nrange(wb, "FeedDate", "Feeding", "A", start, end)
    nrange(wb, "FeedType", "Feeding", "C", start, end)
    day_summary(ws, wb, 8, 9, "Feedings", lambda dref: f'=COUNTIF(FeedDate,{dref})', "FeedDay")
    ws.freeze_panes = "A5"
    line = LineChart(); line.title = "Feedings per Day"; line.height = 7.5; line.width = 12
    line.add_data(Reference(ws, min_col=9, min_row=5, max_row=12), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=8, min_row=6, max_row=12))
    line.legend = None; ws.add_chart(line, "H14")


# ===========================================================================
# 5 — Sleep Tracker
# ===========================================================================
def build_sleep(wb):
    ws = wb.create_sheet("Sleep"); ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 11, 11, 13, 10, 11, 20, 3, 12, 12])
    luxe_header(ws, "J", "😴  SLEEP TRACKER", "Naps & overnight sleep — total daily hours & weekly average, automatic.")
    table_headers(ws, 4, ["Date", "Start", "End", "Category", "Hours", "Wake-Ups", "Notes"])
    sample = []
    sess = {0: [("Nap", 1.5), ("Nap", 1.0), ("Nap", 1.0), ("Overnight", 10.5)],
            1: [("Nap", 1.5), ("Nap", 1.5), ("Overnight", 10.0)],
            2: [("Nap", 2.0), ("Nap", 1.0), ("Overnight", 11.0)],
            3: [("Nap", 1.5), ("Nap", 1.5), ("Overnight", 10.0)],
            4: [("Nap", 1.0), ("Nap", 1.5), ("Nap", 1.0), ("Overnight", 10.5)],
            5: [("Nap", 2.0), ("Nap", 1.0), ("Overnight", 11.0)],
            6: [("Nap", 1.5), ("Nap", 1.5), ("Overnight", 10.5)]}
    for dd in range(0, 7):
        for cat, hrs in sess[dd]:
            sample.append((dminus(dd), "—", "—", cat, hrs, 1 if cat == "Overnight" else 0, ""))
    start = L0; end = start + max(len(sample), 50) - 1
    for i, row in enumerate(sample):
        for ci, val in enumerate(row, 1):
            ws.cell(row=start + i, column=ci, value=val)
    style_rows(ws, start, end, 7, text_left={7}, dates={1}, ints={6})
    for r in range(start, end + 1):
        ws.cell(row=r, column=5).number_format = "0.0"
    add_dv(ws, f"D{start}:D{end}", "SleepCatList")
    nrange(wb, "SleepDate", "Sleep", "A", start, end)
    nrange(wb, "SleepHrs", "Sleep", "E", start, end)
    day_summary(ws, wb, 9, 10, "Hours", lambda dref: f'=SUMIF(SleepDate,{dref},SleepHrs)', "SleepDay")
    ws.freeze_panes = "A5"
    line = LineChart(); line.title = "Sleep Hours per Day"; line.height = 7.5; line.width = 12
    line.add_data(Reference(ws, min_col=10, min_row=5, max_row=12), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=9, min_row=6, max_row=12))
    line.legend = None; ws.add_chart(line, "I14")


# ===========================================================================
# 6 — Diaper Tracker
# ===========================================================================
def build_diaper(wb):
    ws = wb.create_sheet("Diapers"); ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 11, 12, 22, 3, 12, 10, 3, 12, 10])
    luxe_header(ws, "J", "🧷  DIAPER TRACKER", "Wet, dirty & mixed — daily totals and type breakdown, automatic.")
    table_headers(ws, 4, ["Date", "Time", "Type", "Notes"])
    sample = []
    counts = {0: 6, 1: 7, 2: 6, 3: 8, 4: 7, 5: 6, 6: 7}
    times = ["6:15 AM", "9:15 AM", "12:15 PM", "3:15 PM", "6:15 PM", "9:00 PM", "12:00 AM", "3:00 AM"]
    for dd in range(0, 7):
        for k in range(counts[dd]):
            typ = DIAPER_TYPES[k % 3]
            sample.append((dminus(dd), times[k % len(times)], typ, ""))
    start = L0; end = start + max(len(sample), 60) - 1
    for i, row in enumerate(sample):
        for ci, val in enumerate(row, 1):
            ws.cell(row=start + i, column=ci, value=val)
    style_rows(ws, start, end, 4, text_left={4}, dates={1})
    add_dv(ws, f"C{start}:C{end}", "DiaperTypeList")
    nrange(wb, "DiaperDate", "Diapers", "A", start, end)
    nrange(wb, "DiaperType", "Diapers", "C", start, end)
    day_summary(ws, wb, 6, 7, "Changes", lambda dref: f'=COUNTIF(DiaperDate,{dref})', "DiaperDay")
    # type breakdown
    ws.cell(row=4, column=9, value="BY TYPE").style = "section_gold"
    ws.cell(row=5, column=9, value="Type").style = "th"; ws.cell(row=5, column=10, value="Count").style = "th"
    for i, t in enumerate(["Wet", "Dirty", "Mixed"]):
        r = 6 + i
        ws.cell(row=r, column=9, value=t).style = "td_left"
        ws.cell(row=r, column=10, value=f'=COUNTIF(DiaperType,I{r})').style = "td"
    wb.defined_names["DiaperTypeLabels"] = DefinedName("DiaperTypeLabels", attr_text="Diapers!$I$6:$I$8")
    wb.defined_names["DiaperTypeVals"] = DefinedName("DiaperTypeVals", attr_text="Diapers!$J$6:$J$8")
    ws.freeze_panes = "A5"


# ===========================================================================
# 7 — Growth Tracker
# ===========================================================================
def build_growth(wb):
    ws = wb.create_sheet("Growth"); ws.sheet_view.showGridLines = False
    set_widths(ws, [14, 12, 12, 14, 26, 3])
    luxe_header(ws, "E", "📏  GROWTH TRACKER", "Weight, length & head circumference over time — with a growth chart.")
    table_headers(ws, 4, ["Date", "Weight (lb)", "Length (in)", "Head (in)", "Visit Notes"])
    sample = [
        (dminus(125), 7.4, 20.0, 13.5, "Birth"),
        (dminus(118), 7.1, 20.0, 13.6, "Back to birth weight"),
        (dminus(95), 9.2, 21.5, 14.3, "2-week check — 55th %ile"),
        (dminus(65), 11.6, 22.8, 15.1, "2-month check — 60th %ile"),
        (dminus(10), 14.2, 24.4, 15.9, "4-month check soon — 62nd %ile"),
    ]
    start = L0; end = start + 24 - 1
    for i, row in enumerate(sample):
        for ci, val in enumerate(row, 1):
            ws.cell(row=start + i, column=ci, value=val)
    style_rows(ws, start, end, 5, text_left={5}, dates={1})
    for r in range(start, end + 1):
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).number_format = "0.0"
    nrange(wb, "GrowthDate", "Growth", "A", start, end)
    nrange(wb, "GrowthWeight", "Growth", "B", start, end)
    line = LineChart(); line.title = "Weight Over Time (lb)"; line.height = 8; line.width = 13
    line.add_data(Reference(ws, min_col=2, min_row=4, max_row=start + len(sample) - 1), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=start, max_row=start + len(sample) - 1))
    line.legend = None; ws.add_chart(line, "A14")
    ws.freeze_panes = "A5"


# ===========================================================================
# 8 — Milestone Tracker
# ===========================================================================
def build_milestones(wb):
    ws = wb.create_sheet("Milestones"); ws.sheet_view.showGridLines = False
    set_widths(ws, [14, 26, 14, 22, 12, 3, 16, 10])
    luxe_header(ws, "H", "⭐  MILESTONE TRACKER", "Every first, captured — progress by category rolls up automatically.")
    table_headers(ws, 4, ["Category", "Milestone", "Date Achieved", "Notes", "Photo?"])
    sample = [
        ("Social", "First real smile", dminus(90), "Melted our hearts", "Yes"),
        ("Motor", "Holds head up", dminus(85), "", "No"),
        ("Feeding", "Latches well", dminus(115), "", "No"),
        ("Social", "Laughs out loud", dminus(60), "", "Yes"),
        ("Motor", "Rolls tummy to back", dminus(35), "", "Yes"),
        ("Cognitive", "Tracks objects", dminus(70), "", "No"),
        ("Language", "Coos & babbles", dminus(55), "", "No"),
        ("Motor", "Grabs toys", dminus(20), "", "Yes"),
        ("Sleep", "6-hr stretch", dminus(40), "Hallelujah", "No"),
        ("Motor", "Sits with support", "", "Working on it", "No"),
        ("Motor", "Sits unassisted", "", "", "No"),
        ("Teeth", "First tooth", "", "", "No"),
        ("Language", "First word", "", "", "No"),
        ("Walking", "Pulls to stand", "", "", "No"),
    ]
    start = L0; end = start + 30 - 1
    for i, row in enumerate(sample):
        for ci, val in enumerate(row, 1):
            ws.cell(row=start + i, column=ci, value=val)
    style_rows(ws, start, end, 5, text_left={2, 4}, dates={3})
    add_dv(ws, f"A{start}:A{end}", "MilestoneCatList")
    add_dv(ws, f"E{start}:E{end}", "YesNoList")
    ws.conditional_formatting.add(f"A{start}:E{end}", FormulaRule(formula=[f'$C{start}<>""'], fill=fill(MINT_BG)))
    nrange(wb, "MileCat", "Milestones", "A", start, end)
    nrange(wb, "MileName", "Milestones", "B", start, end)
    nrange(wb, "MileDate", "Milestones", "C", start, end)
    # by-category summary
    ws.cell(row=4, column=7, value="BY CATEGORY").style = "section_gold"
    ws.cell(row=5, column=7, value="Category").style = "th"; ws.cell(row=5, column=8, value="Done").style = "th"
    for i, cat in enumerate(MILESTONE_CATS):
        r = 6 + i
        ws.cell(row=r, column=7, value=cat).style = "td_left"
        ws.cell(row=r, column=8, value=f'=COUNTIFS(MileCat,G{r},MileDate,"<>"&"")').style = "td"
    wb.defined_names["MileCatLabels"] = DefinedName("MileCatLabels", attr_text=f"Milestones!$G$6:$G${5+len(MILESTONE_CATS)}")
    wb.defined_names["MileCatVals"] = DefinedName("MileCatVals", attr_text=f"Milestones!$H$6:$H${5+len(MILESTONE_CATS)}")
    ws.freeze_panes = "A5"


# ===========================================================================
# 9 — Medical Center
# ===========================================================================
def build_medical(wb):
    ws, start, end = build_log(
        wb, "Medical", "🏥", "MEDICAL CENTER",
        "Visits, vaccinations & health history — organized and always at hand.",
        ["Date", "Type", "Reason / Vaccine", "Provider", "Status", "Follow-Up", "Notes"],
        [
            (dminus(125), "Pediatrician", "Newborn check", "Dr. Ellis", "Done", "—", "Healthy"),
            (dminus(115), "Vaccination", "Hep B (1st)", "Dr. Ellis", "Done", "—", ""),
            (dminus(95), "Pediatrician", "2-week check", "Dr. Ellis", "Done", "—", "55th %ile"),
            (dminus(65), "Pediatrician", "2-month check", "Dr. Ellis", "Done", "—", ""),
            (dminus(65), "Vaccination", "DTaP / IPV / Hib (2mo)", "Dr. Ellis", "Done", "—", "Slight fever"),
            (dminus(65), "Vaccination", "PCV13 / Rotavirus (2mo)", "Dr. Ellis", "Done", "—", ""),
            (dplus(9), "Pediatrician", "4-month check", "Dr. Ellis", "Upcoming", "Yes", "Bring questions"),
            (dplus(9), "Vaccination", "4-month vaccines", "Dr. Ellis", "Due", "Yes", ""),
        ],
        [13, 14, 22, 14, 12, 12, 20],
        text_left={3, 7}, dates={1, 6}, validations=[("B", "ApptTypeList"), ("E", "StatusList")], reserved=40)
    nrange(wb, "MedType", "Medical", "B", start, end)
    nrange(wb, "MedStatus", "Medical", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=['"Done"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=['"Due"'], fill=fill(WARN_BG)))


# ===========================================================================
# 10 — Appointment Calendar
# ===========================================================================
def build_appointments(wb):
    ws, start, end = build_log(
        wb, "Appointments", "📅", "APPOINTMENT CALENDAR",
        "Every appointment with automatic countdowns — never miss a check-up.",
        ["Appointment", "Type", "Date", "Time", "Provider / Location", "Notes"],
        [
            ("4-month well check", "Pediatrician", dplus(9), "10:00 AM", "Dr. Ellis — City Peds", "Growth + vaccines"),
            ("Lactation follow-up", "Specialist", dplus(3), "2:00 PM", "Maria R.", ""),
            ("Photo session (4mo)", "Parent", dplus(14), "11:00 AM", "Little Smiles Studio", ""),
            ("6-month well check", "Pediatrician", dplus(70), "10:00 AM", "Dr. Ellis", ""),
            ("Parent — pediatric CPR class", "Parent", dplus(20), "6:00 PM", "Community Center", ""),
        ],
        [24, 14, 13, 11, 22, 20],
        text_left={1, 5, 6}, dates={3}, validations=[("B", "ApptTypeList")], reserved=40)
    nrange(wb, "ApptDate", "Appointments", "C", start, end)
    ws.cell(row=4, column=7, value="In").style = "th"; ws.column_dimensions["G"].width = 8
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=7, value=f'=IF(C{r}="","",C{r}-TODAY())'); c.style = "td"; c.number_format = "0;[Red]-0"
        c.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
    ws.conditional_formatting.add(f"A{start}:G{end}",
                                  FormulaRule(formula=[f'AND($C{start}<>"",$C{start}>=TODAY(),$C{start}-TODAY()<=7)'], fill=fill(MINT_BG)))


# ===========================================================================
# 11 — Baby Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 14, 14, 14, 12, 3, 20, 14])
    luxe_header(ws, "H", "💰  BABY BUDGET", "Every baby cost in one place — plan vs actual & remaining, automatic.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining", "% Used"])
    planned = {"Diapers": 80, "Formula": 140, "Food": 40, "Clothing": 60, "Toys": 30,
               "Nursery": 40, "Furniture": 30, "Childcare": 0, "Healthcare": 60,
               "Insurance": 90, "Baby Gear": 50, "Miscellaneous": 30}
    actual = {"Diapers": 88, "Formula": 152, "Food": 22, "Clothing": 74, "Toys": 25,
              "Nursery": 18, "Furniture": 0, "Childcare": 0, "Healthcare": 45,
              "Insurance": 90, "Baby Gear": 66, "Miscellaneous": 20}
    start = L0; end = start + len(EXPENSE_CATS) - 1
    for i, cat in enumerate(EXPENSE_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=2, value=planned[cat]); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=actual[cat]); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in range(2, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})"); c.style = "td"
        c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=total, column=5, value=f"=IFERROR(C{total}/B{total},0)"); cu.style = "td"
    cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    ws.conditional_formatting.add(f"E{start}:E{end}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.2, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))
    nrange(wb, "BudgetCat", "Budget", "A", start, end)
    nrange(wb, "BudgetActual", "Budget", "C", start, end)
    wb.defined_names["BudgetTotalPlanned"] = DefinedName("BudgetTotalPlanned", attr_text=f"Budget!$B${total}")
    wb.defined_names["BudgetTotalActual"] = DefinedName("BudgetTotalActual", attr_text=f"Budget!$C${total}")
    ws.cell(row=4, column=7, value="THIS MONTH").style = "section_gold"
    for i, (lab, fml) in enumerate([("Planned", "=BudgetTotalPlanned"), ("Actual", "=BudgetTotalActual"),
                                    ("Remaining", "=BudgetTotalPlanned-BudgetTotalActual")]):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml); c.style = "field_value"; c.number_format = '"$"#,##0'
    donut = DoughnutChart(); donut.title = "Baby Spending"; donut.height = 8.5; donut.width = 12
    donut.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    donut.dataLabels = no_labels(); ws.add_chart(donut, "G9")
    ws.freeze_panes = "A5"


# ===========================================================================
# 3 — Daily Baby Log  (built after feeding/sleep/diaper for summaries)
# ===========================================================================
def build_daily(wb):
    ws = wb.create_sheet("Daily Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 6, 28, 18, 2])
    luxe_header(ws, "G", "📋  DAILY BABY LOG", "Today at a glance — pulled together from all your tracking tabs.")
    merge_set(ws, "B5:C5", "TODAY'S SUMMARY", "section_gold")
    rows = [
        ("Feedings today", "=COUNTIF(FeedDate,TODAY())", "0"),
        ("Sleep today (hrs)", "=SUMIF(SleepDate,TODAY(),SleepHrs)", "0.0"),
        ("Diaper changes today", "=COUNTIF(DiaperDate,TODAY())", "0"),
        ("Wet today", '=COUNTIFS(DiaperDate,TODAY(),DiaperType,"Wet")', "0"),
        ("Dirty today", '=COUNTIFS(DiaperDate,TODAY(),DiaperType,"Dirty")', "0"),
        ("Baby's age (days)", "=TODAY()-BirthDate", "0"),
    ]
    for i, (lab, fml, fmt) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
    merge_set(ws, "E5:F5", "TODAY'S CARE CHECKLIST", "section_gold")
    checks = ["Morning cuddle & feed", "Tummy time (3x)", "Bath", "Vitamin D drops",
              "Read a book", "Outdoor walk", "Bedtime routine", "Restock diaper bag"]
    for i, it in enumerate(checks):
        r = 6 + i
        cb = ws.cell(row=r, column=5, value="☐"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=12, color=ACCENT); cb.border = BOX
        ct = ws.cell(row=r, column=6, value=it); ct.style = "td_left"
        if i % 2:
            cb.fill = fill(MUTED_ROW); ct.fill = fill(MUTED_ROW)
    merge_set(ws, "B15:F15", "DAILY HIGHLIGHT & NOTES", "section")
    ws.merge_cells("B16:F19"); ws["B16"].value = "The little moment you don't want to forget from today..."
    ws["B16"].style = "body"; ws["B16"].fill = fill(IVORY)
    for r in range(16, 20):
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = BOX; ws.cell(row=r, column=c).fill = fill(IVORY)


# ===========================================================================
# Simple trackers (shopping, inventory, childcare, development)
# ===========================================================================
def build_simple(wb):
    ws, s, e = build_log(
        wb, "Shopping", "🛒", "SHOPPING & BABY SUPPLIES",
        "Never run out — flag reorders and track estimated vs actual cost.",
        ["Item", "Category", "Qty", "Est. Cost", "Purchased?", "Reorder?", "Notes"],
        [
            ("Diapers size 2 (box)", "Diapering", 1, 40, "Yes", "No", ""),
            ("Formula (tub)", "Feeding", 2, 60, "No", "Yes", "Running low"),
            ("Wipes (bulk)", "Diapering", 1, 22, "No", "Yes", ""),
            ("0-3m onesies", "Clothing", 5, 30, "Yes", "No", "Growing fast"),
            ("Baby wash", "Bath", 1, 8, "Yes", "No", ""),
            ("Bottles (pack)", "Feeding", 1, 24, "No", "No", ""),
            ("Vitamin D drops", "Health", 1, 12, "No", "Yes", "Almost out"),
            ("Teether", "Toys", 2, 10, "No", "No", ""),
        ],
        [22, 14, 8, 12, 12, 11, 20],
        text_left={1, 7}, money={4}, ints={3}, validations=[("B", "SupplyCatList"), ("E", "YesNoList"), ("F", "YesNoList")], reserved=40)
    ws.conditional_formatting.add(f"F{s}:F{e}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(WARN_BG)))

    build_log(
        wb, "Inventory", "🧸", "BABY INVENTORY",
        "What you have and what size — clothing, gear, bottles, toys & more.",
        ["Item", "Category", "Size", "Qty", "Status", "Notes"],
        [
            ("Onesies", "Clothing", "3-6m", 12, "In Use", "Moving up soon"),
            ("Sleepers", "Clothing", "3-6m", 8, "In Use", ""),
            ("Bottles", "Feeding", "8 oz", 6, "In Use", ""),
            ("Pacifiers", "Feeding", "0-6m", 4, "In Use", ""),
            ("Swaddle blankets", "Nursery", "—", 5, "In Use", ""),
            ("Car seat", "Gear", "Infant", 1, "In Use", "Expires 2032"),
            ("Stroller", "Gear", "—", 1, "In Use", ""),
            ("Carrier", "Gear", "—", 1, "In Use", ""),
            ("Board books", "Toys", "—", 10, "In Use", ""),
            ("Winter coat", "Clothing", "6-9m", 1, "Stored", "For later"),
        ],
        [20, 14, 12, 8, 12, 22],
        text_left={1, 6}, ints={4}, validations=[("B", "SupplyCatList"), ("C", "SizeList")], reserved=40)

    build_log(
        wb, "Childcare", "🧑‍🍼", "CHILDCARE ORGANIZER",
        "Everyone who helps care for baby — schedules, contacts & pickup authorization.",
        ["Caregiver", "Type", "Days / Hours", "Phone", "Rate", "Pickup Auth?", "Notes"],
        [
            ("Grandma Rose", "Family Care", "Tue, Thu", "(555) 661-2048", 0, "Yes", "Loves it"),
            ("Sunshine Daycare", "Daycare", "Mon, Wed, Fri 8-4", "(555) 778-3300", 55, "Yes", "Starts next month"),
            ("Emma (sitter)", "Babysitter", "As needed", "(555) 210-9090", 18, "Yes", "CPR certified"),
            ("Dad", "Family Care", "Weekends", "—", 0, "Yes", ""),
        ],
        [20, 14, 18, 16, 10, 12, 20],
        text_left={1, 7}, money={5}, validations=[("F", "YesNoList")], reserved=30)

    build_log(
        wb, "Development", "🎨", "DEVELOPMENT ACTIVITIES",
        "Play with a purpose — track the activities that help baby grow.",
        ["Activity", "Type", "Frequency", "Last Done", "Baby's Response", "Notes"],
        [
            ("Tummy time", "Gross Motor", "3x daily", dminus(0), "Getting stronger", ""),
            ("Reading board books", "Language", "Daily", dminus(0), "Loves the colors", ""),
            ("Music & singing", "Cognitive", "Daily", dminus(0), "Kicks & smiles", ""),
            ("Sensory play (textures)", "Sensory", "3x weekly", dminus(1), "Curious", ""),
            ("Outdoor walk", "Gross Motor", "Daily", dminus(0), "Naps after", ""),
            ("Grasping toys", "Fine Motor", "Daily", dminus(0), "Improving grip", ""),
            ("Mirror play", "Social", "Daily", dminus(1), "Laughs", ""),
        ],
        [22, 14, 13, 13, 20, 18],
        text_left={1, 5, 6}, dates={4}, reserved=30)


# ===========================================================================
# 15 — Routine Planner
# ===========================================================================
def build_routine(wb):
    ws = wb.create_sheet("Routine"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 4, 26, 16, 4, 4, 26, 16, 2])
    luxe_header(ws, "I", "🕒  ROUTINE PLANNER", "Gentle, predictable rhythms — check off the day's routine as you go.")
    blocks = [
        ("MORNING ROUTINE", 5, 2, ["Wake & cuddle", "Diaper change", "Morning feed", "Tummy time",
                                   "Get dressed", "Play & talk"]),
        ("NAP ROUTINE", 5, 6, ["Watch for sleepy cues", "Dim the room", "Swaddle / sleep sack",
                               "White noise on", "Down drowsy but awake"]),
        ("BEDTIME ROUTINE", 13, 2, ["Warm bath", "Baby massage", "Fresh diaper & PJs", "Final feed",
                                    "Book & lullaby", "Into crib"]),
        ("WEEKLY RHYTHM", 13, 6, ["Laundry (Mon/Thu)", "Restock diaper bag", "Bath nights (M/W/F)",
                                  "Grandma days (Tue/Thu)", "Family walk (weekend)"]),
    ]
    routine_cells = []
    for title, top, cbox, items in blocks:
        ctext = cbox + 1
        merge_set(ws, f"{get_column_letter(cbox)}{top}:{get_column_letter(ctext)}{top}", title, "section_gold")
        ws.row_dimensions[top].height = 22
        for i, it in enumerate(items):
            r = top + 1 + i
            cb = ws.cell(row=r, column=cbox, value="Yes" if len(routine_cells) < 15 else "No")
            cb.style = "input"
            add_dv(ws, f"{get_column_letter(cbox)}{r}", "YesNoList")
            ct = ws.cell(row=r, column=ctext, value=it); ct.style = "td_left"
            ws.merge_cells(f"{get_column_letter(ctext)}{r}:{get_column_letter(ctext)}{r}")
            routine_cells.append((cbox, r))
            if i % 2:
                ct.fill = fill(MUTED_ROW)
    # Build a helper column consolidating routine Yes/No for the completion KPI
    hcol = 10
    ws.column_dimensions[get_column_letter(hcol)].width = 2
    for idx, (cbox, r) in enumerate(routine_cells):
        ws.cell(row=5 + idx, column=hcol, value=f'={get_column_letter(cbox)}{r}')
    rstart, rend = 5, 5 + len(routine_cells) - 1
    nrange(wb, "RoutineDone", "Routine", get_column_letter(hcol), rstart, rend)
    # tidy: make check cells show Yes/No plainly
    for cbox, r in routine_cells:
        ws.cell(row=r, column=cbox).number_format = "General"


# ===========================================================================
# 17 — Memory Book (image placeholders)
# ===========================================================================
def build_memories(wb):
    ws = wb.create_sheet("Memories"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "📸  MEMORY BOOK", "Keep every precious first — paste photos with the date & the story.")
    sections = ["Birth Day", "First Smile", "First Bath", "First Tooth",
                "First Steps", "First Birthday", "Family Photo", "Holiday Memory"]
    top0 = 5; card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th"); ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+5}", "📷\nPaste photo here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 6):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 6, column=col, value="Date").style = "field_label"
        merge_set(ws, f"{M}{row+6}:{R}{row+6}", "", "field_value")
        ws.cell(row=row + 7, column=col, value="Caption").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")


# ===========================================================================
# 18 — Travel with Baby
# ===========================================================================
def build_travel(wb):
    ws = wb.create_sheet("Travel"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 4, 28, 4, 4, 28, 4, 4, 28, 2])
    luxe_header(ws, "J", "🧳  TRAVEL WITH BABY", "Ready-made packing lists so a trip with baby stays (mostly) calm.")
    lists = [
        ("DIAPER BAG", ["Diapers (1 / hr + extra)", "Wipes", "Changing pad", "Diaper cream",
                        "Bags for dirties", "Change of clothes x2", "Burp cloths", "Pacifiers"]),
        ("FEEDING KIT", ["Bottles + formula/BM", "Cooler + ice pack", "Bibs", "Nursing cover",
                         "Snacks (if solids)", "Extra water", "Cleaning brush", "Hand sanitizer"]),
        ("COMFORT & HEALTH", ["Favorite blanket", "2 favorite toys", "Sleep sack", "White-noise app",
                              "Infant meds (per doctor)", "Thermometer", "First-aid basics",
                              "Sunhat & sunscreen (6m+)"]),
    ]
    cols = [(2, 3), (5, 6), (8, 9)]
    for li, (title, items) in enumerate(lists):
        cbox, ctext = cols[li]
        merge_set(ws, f"{get_column_letter(cbox)}5:{get_column_letter(ctext)}5", title, "section_gold"); ws.row_dimensions[5].height = 22
        for i, it in enumerate(items):
            r = 6 + i
            cb = ws.cell(row=r, column=cbox, value="☐"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=12, color=ACCENT); cb.border = BOX
            ct = ws.cell(row=r, column=ctext, value=it); ct.style = "td_left"
            if i % 2:
                cb.fill = fill(MUTED_ROW); ct.fill = fill(MUTED_ROW)


# ===========================================================================
# 19 — Family Goals
# ===========================================================================
def build_goals(wb):
    ws, start, end = build_log(
        wb, "Family Goals", "🎯", "FAMILY GOALS",
        "Where your family is headed this year — progress bars keep it in view.",
        ["Goal", "Category", "Target Date", "Progress", "Status", "Notes"],
        [
            ("Build 3-month emergency fund", "Savings", dplus(150), 0.5, "On Track", ""),
            ("Start 529 college fund", "Savings", dplus(60), 0.3, "On Track", ""),
            ("Consistent bedtime routine", "Routine", dplus(30), 0.7, "On Track", ""),
            ("Read 100 board books", "Reading", dplus(120), 0.4, "On Track", "Currently 41"),
            ("Family walk 5x / week", "Health", dplus(30), 0.6, "On Track", ""),
            ("Baby-proof the house", "Health", dplus(45), 0.35, "On Track", "Before crawling"),
            ("Weekly date night", "Family", dplus(30), 0.5, "On Track", ""),
        ],
        [30, 14, 14, 12, 14, 22],
        text_left={1, 6}, dates={3}, pcts={4}, reserved=30)
    nrange(wb, "GoalName", "Family Goals", "A", start, end)
    nrange(wb, "GoalProgress", "Family Goals", "D", start, end)
    ws.conditional_formatting.add(f"D{start}:D{end}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))


# ===========================================================================
# 20 — Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 18, 3, 16, 12, 12, 12, 12, 12, 2])
    luxe_header(ws, "L", "📈  ANALYTICS DASHBOARD", "Baby & family by the numbers — trends, budget health & an organization score.")
    merge_set(ws, "B5:D5", "FAMILY ORGANIZATION SCORES", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Feeding on goal", '=IFERROR(MIN(COUNTIF(FeedDate,TODAY())/FeedGoal,1),0)'),
        ("Sleep on goal", '=IFERROR(MIN(SUMIF(SleepDate,TODAY(),SleepHrs)/SleepGoal,1),0)'),
        ("Budget health", '=IFERROR(1-BudgetTotalActual/BudgetTotalPlanned,0)'),
        ("Milestones logged", '=IFERROR(COUNTIF(MileDate,"<>"&"")/MAX(COUNTA(MileName),1),0)'),
        ("Vaccinations done", '=IFERROR(COUNTIFS(MedType,"Vaccination",MedStatus,"Done")/MAX(COUNTIF(MedType,"Vaccination"),1),0)'),
        ("Goals progress", '=IFERROR(AVERAGE(GoalProgress),0)'),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.75,"Great",IF(C{r}>=0.4,"On Track","Needs Love"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(f"C{start}:C{end}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "F5:H5", "FAMILY ORGANIZATION SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY); cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "A gentle blend of care, budget, milestones & goals.", "subtitle")
    ws["F10"].fill = fill(IVORY)
    bar = BarChart(); bar.type = "bar"; bar.title = "Organization by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None; ws.add_chart(bar, "B17")
    # sleep trend on the right
    sl = LineChart(); sl.title = "Sleep Hours (7 days)"; sl.height = 8; sl.width = 13
    sl.add_data(Reference(wb["Sleep"], min_col=10, min_row=5, max_row=12), titles_from_data=True)
    sl.set_categories(Reference(wb["Sleep"], min_col=9, min_row=6, max_row=12))
    sl.legend = None; ws.add_chart(sl, "F13")


# ===========================================================================
# 1 — Executive Baby Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  👶  BABY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Feeding, sleep, diapers, growth, health & memories — your whole day, gently organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("AGE (DAYS)", "=TODAY()-BirthDate", "days"),
        ("FEEDINGS TODAY", "=COUNTIF(FeedDate,TODAY())", "num"),
        ("SLEEP TODAY (H)", "=SUMIF(SleepDate,TODAY(),SleepHrs)", "dec"),
        ("DIAPERS TODAY", "=COUNTIF(DiaperDate,TODAY())", "num"),
        ("NEXT APPT (DAYS)", '=IF(COUNTIF(ApptDate,">="&TODAY())=0,0,MINIFS(ApptDate,ApptDate,">="&TODAY())-TODAY())', "days"),
        ("GROWTH ENTRIES", "=COUNTA(GrowthDate)", "num"),
    ]
    row2 = [
        ("MONTHLY BUDGET", "=MonthlyBudget", "money"),
        ("MILESTONES", '=IFERROR(COUNTIF(MileDate,"<>"&"")/MAX(COUNTA(MileName),1),0)', "pct"),
        ("VACCINES DONE", '=IFERROR(COUNTIFS(MedType,"Vaccination",MedStatus,"Done")/MAX(COUNTIF(MedType,"Vaccination"),1),0)', "pct"),
        ("BUDGET LEFT", "=BudgetTotalPlanned-BudgetTotalActual", "money"),
        ("GOALS PROGRESS", "=IFERROR(AVERAGE(GoalProgress),0)", "pct"),
        ("ROUTINE DONE", '=IFERROR(COUNTIF(RoutineDone,"Yes")/MAX(COUNTA(RoutineDone),1),0)', "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "DAILY RHYTHMS", "section_gold")
    f1 = LineChart(); f1.title = "Feedings per Day"; f1.height = 8.2; f1.width = 11.5
    f1.add_data(Reference(wb["Feeding"], min_col=9, min_row=5, max_row=12), titles_from_data=True)
    f1.set_categories(Reference(wb["Feeding"], min_col=8, min_row=6, max_row=12)); f1.legend = None
    ws.add_chart(f1, "B12")
    s1 = LineChart(); s1.title = "Sleep Hours per Day"; s1.height = 8.2; s1.width = 11.5
    s1.add_data(Reference(wb["Sleep"], min_col=10, min_row=5, max_row=12), titles_from_data=True)
    s1.set_categories(Reference(wb["Sleep"], min_col=9, min_row=6, max_row=12)); s1.legend = None
    ws.add_chart(s1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "GROWTH · BUDGET · MILESTONES", "section_gold")
    g1 = LineChart(); g1.title = "Weight Over Time"; g1.height = 8.2; g1.width = 11.5
    g1.add_data(Reference(wb["Growth"], min_col=2, min_row=4, max_row=9), titles_from_data=True)
    g1.set_categories(Reference(wb["Growth"], min_col=1, min_row=5, max_row=9)); g1.legend = None
    ws.add_chart(g1, "B30")
    bend = L0 + len(EXPENSE_CATS) - 1
    d1 = DoughnutChart(); d1.title = "Baby Spending"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=bend), titles_from_data=True)
    d1.set_categories(Reference(wb["Budget"], min_col=1, min_row=L0, max_row=bend)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47",
              "Organization & record-keeping tool only — not a medical device or a substitute for professional healthcare. See the Welcome tab.",
              "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb)
    build_feeding(wb); build_sleep(wb); build_diaper(wb)
    build_daily(wb); build_growth(wb); build_milestones(wb)
    build_medical(wb); build_appointments(wb); build_budget(wb)
    build_simple(wb)     # shopping, inventory, childcare, development
    build_routine(wb); build_memories(wb); build_travel(wb); build_goals(wb)
    build_analytics(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Profile", "Daily Log", "Feeding", "Sleep",
             "Diapers", "Growth", "Milestones", "Medical", "Appointments", "Budget",
             "Shopping", "Inventory", "Childcare", "Routine", "Development",
             "Memories", "Travel", "Family Goals", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Baby_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
