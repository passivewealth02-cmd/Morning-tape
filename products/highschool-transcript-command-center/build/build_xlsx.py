"""Build High School Transcript Command Center™ — GPA, credits & a print-ready transcript.

16 planning tabs (+ Settings) · a premium homeschool/high-school records system in
Google Sheets & Excel. Official transcript, course records with auto GPA (weighted
& unweighted), credit tracker vs grad requirements, 4-year plan, test scores,
activities, awards, service hours, course descriptions, reading list and a college-
ready score — one dashboard. Built for homeschool & college-bound families.

Run: python3 build_xlsx.py   ->  ../HS_Transcript_Command_Center.xlsx
"""
from __future__ import annotations
import datetime as dt
import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

AREAS = ["English", "Math", "Science", "Social Studies", "World Language",
         "Fine Arts", "PE/Health", "Elective", "Bible / Character"]
LEVELS = ["Regular", "Honors", "AP", "Dual Credit"]
GRADE_SCALE = [("A", 4.0), ("A-", 3.7), ("B+", 3.3), ("B", 3.0), ("B-", 2.7),
               ("C+", 2.3), ("C", 2.0), ("C-", 1.7), ("D+", 1.3), ("D", 1.0), ("F", 0.0)]
YESNO = ["Yes", "No"]

# (year, area, course, level, credits, grade)
COURSES = [
    (9, "English", "English 9", "Regular", 1.0, "A"), (9, "Math", "Algebra I", "Regular", 1.0, "A-"),
    (9, "Science", "Biology", "Regular", 1.0, "B+"), (9, "Social Studies", "World History", "Regular", 1.0, "A"),
    (9, "World Language", "Spanish I", "Regular", 1.0, "A"), (9, "PE/Health", "PE & Health", "Regular", 0.5, "A"),
    (9, "Fine Arts", "Art I", "Regular", 0.5, "A"),
    (10, "English", "English 10", "Regular", 1.0, "A-"), (10, "Math", "Geometry", "Regular", 1.0, "A"),
    (10, "Science", "Chemistry", "Honors", 1.0, "B+"), (10, "Social Studies", "US History", "Honors", 1.0, "A"),
    (10, "World Language", "Spanish II", "Regular", 1.0, "A-"), (10, "Fine Arts", "Music", "Regular", 0.5, "A"),
    (10, "Elective", "Speech & Debate", "Regular", 0.5, "A"),
    (11, "English", "English 11", "Honors", 1.0, "A"), (11, "Math", "Algebra II", "Regular", 1.0, "B+"),
    (11, "Science", "Physics", "Honors", 1.0, "A-"), (11, "Social Studies", "AP US Government", "AP", 1.0, "B+"),
    (11, "World Language", "Spanish III", "Regular", 1.0, "A"), (11, "Elective", "Economics", "Regular", 0.5, "A"),
    (11, "Elective", "Logic", "Regular", 0.5, "A"),
    (12, "English", "English 12", "Honors", 1.0, "A"), (12, "Math", "Pre-Calculus", "Honors", 1.0, "A-"),
    (12, "Science", "AP Biology", "AP", 1.0, "A-"), (12, "English", "AP Literature", "AP", 1.0, "A"),
    (12, "World Language", "Spanish IV", "Regular", 1.0, "A"), (12, "Elective", "Capstone Project", "Regular", 1.0, "A"),
]
TESTS = [
    ("SAT", "Mar 2026", "1380", "EBRW 690 · Math 690"), ("ACT", "Apr 2026", "30", "Composite"),
    ("PSAT/NMSQT", "Oct 2025", "1360", "Commended"), ("AP US Government", "May 2026", "4", "Qualified"),
    ("AP Biology", "May 2027", "—", "Scheduled"),
]
ACTIVITIES = [
    ("Homeschool Speech & Debate League", "9–12", "Team captain (12)", "4 yrs"),
    ("Church youth worship team", "9–12", "Keyboard", "4 yrs"),
    ("Volunteer tutor — co-op", "10–12", "Math & Spanish", "3 yrs"),
    ("Robotics club (co-op)", "11–12", "Programming lead", "2 yrs"),
    ("Part-time job — bakery", "11–12", "Shift lead", "2 yrs"),
    ("Community theater", "9–11", "Cast & crew", "3 yrs"),
    ("Youth soccer (rec league)", "9–12", "Player", "4 yrs"),
]
AWARDS = [
    ("PSAT Commended Scholar", "11", "College Board"), ("1st place — regional speech", "11", "NCFCA"),
    ("Presidential Volunteer Service Award", "12", "Bronze"), ("Spanish honor society", "11", "Co-op"),
    ("Perfect attendance (co-op)", "9–12", "Co-op"),
]
SERVICE = [
    (dt.date(2026, 9, 14), "Food bank sorting", 24), (dt.date(2026, 10, 5), "Church nursery", 18),
    (dt.date(2026, 11, 8), "Nursing-home visits", 20), (dt.date(2027, 1, 17), "Tutoring younger students", 30),
    (dt.date(2027, 3, 21), "Community garden", 16), (dt.date(2027, 4, 12), "5k race volunteer", 12),
]
GRADREQ = [
    ("English — 4 credits", True), ("Math — 3+ credits", True), ("Science — 3+ credits", True),
    ("Social Studies — 3 credits", True), ("World Language — 2+ credits", True), ("Fine Arts — 1 credit", True),
    ("PE/Health — 1 credit", True), ("Electives — 4+ credits", True), ("Total 24 credits", True),
    ("Course descriptions written", False),
]
DESCRIPTIONS = [
    ("English 11 (Honors)", "American literature survey with a research paper, rhetoric & weekly essays; texts included Gatsby, Scarlet Letter & Frederick Douglass."),
    ("AP US Government", "College-level study of the US Constitution, institutions & political behavior; passed the AP exam (score 4)."),
    ("Physics (Honors)", "Algebra-based mechanics, energy, waves & electricity with a hands-on lab component and lab reports."),
    ("Spanish III", "Intermediate Spanish — grammar, conversation & culture, using Breaking the Barrier + native-speaker practice."),
    ("Capstone Project", "Independent year-long project: researched, built & presented a small-business plan; 40+ hours documented."),
]
READING = [
    ("Ella", "Pride and Prejudice", "Austen", "Yes"), ("Ella", "1984", "Orwell", "Yes"),
    ("Ella", "The Great Gatsby", "Fitzgerald", "Yes"), ("Ella", "Till We Have Faces", "Lewis", "Yes"),
    ("Ella", "Crime and Punishment", "Dostoevsky", "In Progress"), ("Ella", "The Odyssey", "Homer", "Yes"),
]
PLAN = [
    ("English", "English 9", "English 10", "English 11 (H)", "English 12 (H) + AP Lit"),
    ("Math", "Algebra I", "Geometry", "Algebra II", "Pre-Calculus (H)"),
    ("Science", "Biology", "Chemistry (H)", "Physics (H)", "AP Biology"),
    ("Social Studies", "World History", "US History (H)", "AP US Gov", "Economics"),
    ("World Language", "Spanish I", "Spanish II", "Spanish III", "Spanish IV"),
    ("Electives / Arts", "Art I · PE", "Music · Speech", "Econ · Logic", "Capstone"),
]
CREDITS_REQUIRED = 24
GPA_TARGET = 3.5
SERVICE_GOAL = 100
CLASS_OF = "2027"

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "pct": "0%", "gpa": "0.00", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings (with the grade scale that powers GPA)
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 3, 14, 10, 3, 16, 16, 16])
    luxe_header(ws, "J", "⚙  SETTINGS", "Set the student, school & scale once — GPA and the transcript follow.")
    merge_set(ws, "B5:C5", "STUDENT & SCHOOL", "section")
    controls = [
        ("Student Name", "Ella Bennett", None, "StudentName"),
        ("School Name", "Bennett Family Academy", None, "SchoolName"),
        ("Class Of", CLASS_OF, None, "ClassOf"),
        ("Graduation Date", "May 28, 2027", None, "GradDate"),
        ("Credits Required", CREDITS_REQUIRED, "0.0", "CreditsReq"),
        ("GPA Target", GPA_TARGET, "0.00", "GPATarget"),
        ("Service-Hours Goal", SERVICE_GOAL, "#,##0", "ServiceGoal"),
        ("Rigor Goal (Honors+AP)", 8, "0", "RigorGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    # grade scale table (drives VLOOKUP)
    merge_set(ws, "E5:F5", "GRADE SCALE (4.0)", "section_gold")
    ws.cell(row=6, column=5, value="Grade").style = "th"; ws.cell(row=6, column=6, value="Points").style = "th"
    for i, (g, p) in enumerate(GRADE_SCALE):
        r = 7 + i
        ws.cell(row=r, column=5, value=g).style = "td"
        c = ws.cell(row=r, column=6, value=p); c.style = "td"; c.number_format = "0.0"
    wb.defined_names["GradeScale"] = DefinedName("GradeScale", attr_text=f"Settings!$E$7:$F${6+len(GRADE_SCALE)}")
    # lists
    merge_set(ws, "H5:J5", "DROPDOWN LISTS", "section_gold")
    banks = [("H", "Subject Area", AREAS, "AreaList"), ("I", "Level", LEVELS, "LevelList"),
             ("J", "Yes / No", YESNO, "YesNoList")]
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    # weighted-bump reference (Regular 0 / Honors 0.5 / AP 1 / Dual 1)
    ws.cell(row=20, column=5, value="Level").style = "th"; ws.cell(row=20, column=6, value="Bump").style = "th"
    for i, (lv, bp) in enumerate([("Regular", 0.0), ("Honors", 0.5), ("AP", 1.0), ("Dual Credit", 1.0)]):
        r = 21 + i
        ws.cell(row=r, column=5, value=lv).style = "td"
        c = ws.cell(row=r, column=6, value=bp); c.style = "td"; c.number_format = "0.0"
    wb.defined_names["LevelScale"] = DefinedName("LevelScale", attr_text="Settings!$E$21:$F$24")


# ===========================================================================
# Course Records — THE engine (auto GPA points)
# ===========================================================================
def build_courses(wb):
    ws = wb.create_sheet("Course Records"); ws.sheet_view.showGridLines = False
    headers = ["Grade Yr", "Subject Area", "Course Title", "Level", "Credits", "Grade", "Pts", "Qual", "Wtd"]
    set_widths(ws, [10, 18, 26, 13, 10, 9, 8, 9, 9])
    luxe_header(ws, "I", "📗  COURSE RECORDS",
                "Every course in one place — type the grade, GPA points calculate themselves.")
    table_headers(ws, 4, headers)
    start = L0; reserved = 45; end = start + reserved - 1
    for i, (yr, area, course, level, cr, grade) in enumerate(COURSES):
        r = start + i
        ws.cell(row=r, column=1, value=yr)
        ws.cell(row=r, column=2, value=area)
        ws.cell(row=r, column=3, value=course)
        ws.cell(row=r, column=4, value=level)
        ws.cell(row=r, column=5, value=cr)
        ws.cell(row=r, column=6, value=grade)
        ws.cell(row=r, column=7, value=f'=IFERROR(VLOOKUP(F{r},GradeScale,2,FALSE),"")')
        ws.cell(row=r, column=8, value=f'=IFERROR(G{r}*E{r},"")')
        ws.cell(row=r, column=9, value=f'=IFERROR((G{r}+IFERROR(VLOOKUP(D{r},LevelScale,2,FALSE),0))*E{r},"")')
    style_rows(ws, start, end, len(headers), text_left={2, 3}, ints={1}, dec={5, 7, 8, 9})
    add_dv(ws, f"B{start}:B{end}", "AreaList"); add_dv(ws, f"D{start}:D{end}", "LevelList")
    ws.freeze_panes = "A5"
    nrange(wb, "CrsYear", "Course Records", "A", start, end)
    nrange(wb, "CrsArea", "Course Records", "B", start, end)
    nrange(wb, "CrsLevel", "Course Records", "D", start, end)
    nrange(wb, "CrsCredits", "Course Records", "E", start, end)
    nrange(wb, "CrsQual", "Course Records", "H", start, end)
    nrange(wb, "CrsWtd", "Course Records", "I", start, end)
    # totals row
    tot = end + 1
    ws.cell(row=tot, column=4, value="TOTAL").style = "th"
    for col, fml, fmt in ((5, "=SUM(CrsCredits)", "0.0"), (8, "=SUM(CrsQual)", "0.0"), (9, "=SUM(CrsWtd)", "0.0")):
        c = ws.cell(row=tot, column=col, value=fml); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt
    for col in (1, 2, 3, 6, 7):
        ws.cell(row=tot, column=col).style = "td"; ws.cell(row=tot, column=col).fill = fill(SURFACE)
    cell_name(wb, "CreditsEarned", "Course Records", f"$E${tot}")
    cell_name(wb, "QualTotal", "Course Records", f"$H${tot}")
    cell_name(wb, "WtdTotal", "Course Records", f"$I${tot}")
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# GPA Calculator
# ===========================================================================
def build_gpa(wb):
    ws = wb.create_sheet("GPA Calculator"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "🧮  GPA CALCULATOR",
                "Weighted & unweighted, by year and cumulative — calculated from Course Records.")
    merge_set(ws, "B5:E5", "GPA BY YEAR", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Year", "Credits", "Unweighted", "Weighted"], start_col=2)
    for i, yr in enumerate([9, 10, 11, 12]):
        r = 7 + i
        ws.cell(row=r, column=2, value=f"Grade {yr}").style = "td_left"
        cc = ws.cell(row=r, column=3, value=f'=SUMIF(CrsYear,{yr},CrsCredits)'); cc.style = "td"; cc.number_format = "0.0"
        cu = ws.cell(row=r, column=4, value=f'=IFERROR(SUMIF(CrsYear,{yr},CrsQual)/SUMIF(CrsYear,{yr},CrsCredits),0)'); cu.style = "td"; cu.number_format = "0.00"
        cw = ws.cell(row=r, column=5, value=f'=IFERROR(SUMIF(CrsYear,{yr},CrsWtd)/SUMIF(CrsYear,{yr},CrsCredits),0)'); cw.style = "td"; cw.number_format = "0.00"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    merge_set(ws, "B12:C12", "CUMULATIVE", "section_gold"); ws.row_dimensions[12].height = 22
    rows2 = [("Total credits", "=CreditsEarned", "0.0", None),
             ("Unweighted GPA", "=IFERROR(QualTotal/CreditsEarned,0)", "0.00", "GPAUW"),
             ("Weighted GPA", "=IFERROR(WtdTotal/CreditsEarned,0)", "0.00", "GPAW"),
             ("Honors + AP courses", '=COUNTIF(CrsLevel,"Honors")+COUNTIF(CrsLevel,"AP")+COUNTIF(CrsLevel,"Dual Credit")', "0", "HonorsAP"),
             ("Total courses", "=COUNTA(CrsArea)", "0", "CourseCount")]
    for i, (lab, fml, fmt, nm) in enumerate(rows2):
        r = 13 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if nm:
            cell_name(wb, nm, "GPA Calculator", f"$C${r}")
        if lab in ("Unweighted GPA", "Weighted GPA"):
            c.fill = fill(MINT_BG)


# ===========================================================================
# Credit Tracker
# ===========================================================================
def build_credits(wb):
    ws = wb.create_sheet("Credit Tracker"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "🎓  CREDIT TRACKER",
                "Credits earned vs graduation requirements, by subject area — see what's left.")
    table_headers(ws, 4, ["Subject Area", "Required", "Earned", "Remaining"], start_col=2)
    req = {"English": 4, "Math": 3, "Science": 3, "Social Studies": 3, "World Language": 2,
           "Fine Arts": 1, "PE/Health": 1, "Elective": 4, "Bible / Character": 0}
    start = L0
    areas = list(req.keys())
    for i, area in enumerate(areas):
        r = start + i
        ws.cell(row=r, column=2, value=area).style = "td_left"
        cq = ws.cell(row=r, column=3, value=req[area]); cq.style = "input"; cq.number_format = "0.0"
        ce = ws.cell(row=r, column=4, value=f'=SUMIF(CrsArea,B{r},CrsCredits)'); ce.style = "td"; ce.number_format = "0.0"
        cr = ws.cell(row=r, column=5, value=f"=MAX(C{r}-D{r},0)"); cr.style = "td"; cr.number_format = "0.0"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(areas) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0.0"
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=fill(WARN_BG)))
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=["0"], fill=fill(MINT_BG)))
    merge_set(ws, "B15:E15", "TO GRADUATE", "section_gold")
    rows2 = [("Credits required", "=CreditsReq", "0.0"), ("Credits earned", "=CreditsEarned", "0.0"),
             ("Credits remaining", "=MAX(CreditsReq-CreditsEarned,0)", "0.0"),
             ("Graduation progress", "=IFERROR(MIN(CreditsEarned/CreditsReq,1),0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 16 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Graduation progress":
            c.fill = fill(MINT_BG)


# ===========================================================================
# Official Transcript (print-ready view)
# ===========================================================================
def build_transcript(wb):
    ws = wb.create_sheet("Official Transcript"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 10, 10, 2])
    ws.row_dimensions[1].height = 30
    merge_set(ws, "B1:E1", "=SchoolName", "section"); ws["B1"].font = Font(size=16, bold=True, color=PRIMARY); ws["B1"].alignment = Alignment(horizontal="center")
    merge_set(ws, "B2:E2", "OFFICIAL HIGH SCHOOL TRANSCRIPT", "section_gold"); ws["B2"].alignment = Alignment(horizontal="center")
    for c in range(2, 6):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    ws.row_dimensions[3].height = 4
    info = [("Student", "=StudentName", "Class Of", "=ClassOf"),
            ("School", "=SchoolName", "Graduation", "=GradDate")]
    r = 5
    for a, av, b, bv in info:
        ws.cell(row=r, column=2, value=a).style = "field_label"
        ws.cell(row=r, column=3, value=av).style = "field_value"
        ws.cell(row=r, column=4, value=b).style = "field_label"
        ws.cell(row=r, column=5, value=bv).style = "field_value"
        r += 1
    r += 1
    # per-year blocks
    for yr in (9, 10, 11, 12):
        merge_set(ws, f"B{r}:E{r}", f"GRADE {yr}", "section_gold"); ws.row_dimensions[r].height = 20; r += 1
        table_headers(ws, r, ["Course", "Area", "Cr", "Gr"], start_col=2); r += 1
        yr_courses = [c for c in COURSES if c[0] == yr]
        for (y, area, course, level, cr, grade) in yr_courses:
            title = course + (" (H)" if level == "Honors" else " (AP)" if level == "AP" else "")
            ws.cell(row=r, column=2, value=title).style = "td_left"
            ws.cell(row=r, column=3, value=area).style = "td"
            cc = ws.cell(row=r, column=4, value=cr); cc.style = "td"; cc.number_format = "0.0"
            ws.cell(row=r, column=5, value=grade).style = "td"
            r += 1
        cy = ws.cell(row=r, column=2, value=f"Year credits: {sum(c[4] for c in yr_courses):.1f}"); cy.style = "td_left"; cy.font = Font(bold=True, color=ACCENT)
        for c in range(3, 6):
            ws.cell(row=r, column=c).fill = fill(SOFT_BG)
        r += 2
    # summary box
    merge_set(ws, f"B{r}:E{r}", "SUMMARY", "section_gold"); r += 1
    summ = [("Cumulative GPA (Unweighted)", "=IFERROR(QualTotal/CreditsEarned,0)", "0.00"),
            ("Cumulative GPA (Weighted)", "=IFERROR(WtdTotal/CreditsEarned,0)", "0.00"),
            ("Total Credits", "=CreditsEarned", "0.0"),
            ("Honors / AP Courses", "=HonorsAP", "0")]
    for lab, fml, fmt in summ:
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        ws.merge_cells(f"C{r}:E{r}")
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if "GPA" in lab:
            c.fill = fill(MINT_BG)
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="Administrator signature").style = "field_label"
    ws.merge_cells(f"C{r}:E{r}"); ws.cell(row=r, column=3, value="__________________________  Date: __________").style = "field_value"


# ===========================================================================
# 4-Year Plan
# ===========================================================================
def build_plan(wb):
    ws = wb.create_sheet("4-Year Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 20, 20, 20, 2])
    luxe_header(ws, "F", "🗺  4-YEAR PLAN",
                "Map all four years at once — spot gaps in credits & requirements early.")
    table_headers(ws, 4, ["Subject", "Grade 9", "Grade 10", "Grade 11", "Grade 12"], start_col=2)
    start = L0
    for i, row in enumerate(PLAN):
        r = start + i
        ws.cell(row=r, column=2, value=row[0]).style = "td_left"
        for j in range(4):
            ws.cell(row=r, column=3 + j, value=row[1 + j]).style = "input"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


# ===========================================================================
# Test Scores  (defines TestName)
# ===========================================================================
def build_tests(wb):
    ws, start, end = build_log(
        wb, "Test Scores", "📊", "TEST SCORES",
        "SAT, ACT, PSAT, AP & more — the scores colleges ask for, in one place.",
        ["Test", "Date", "Score", "Detail"],
        TESTS, [22, 14, 12, 26], text_left={4}, reserved=24)
    nrange(wb, "TestName", "Test Scores", "A", start, end)


# ===========================================================================
# Activities & Leadership
# ===========================================================================
def build_activities(wb):
    ws, start, end = build_log(
        wb, "Activities", "🏅", "ACTIVITIES & LEADERSHIP",
        "The extracurricular story colleges read — activity, role & years involved.",
        ["Activity", "Grades", "Role / Detail", "Years"],
        ACTIVITIES, [34, 12, 24, 12], text_left={1, 3}, reserved=30)
    nrange(wb, "ActName", "Activities", "A", start, end)


# ===========================================================================
# Awards & Honors
# ===========================================================================
def build_awards(wb):
    ws, start, end = build_log(
        wb, "Awards", "🏆", "AWARDS & HONORS",
        "Every recognition, dated — ready to drop into applications.",
        ["Award / Honor", "Grade", "From"],
        AWARDS, [40, 12, 22], text_left={1, 3}, reserved=24)
    nrange(wb, "AwardName", "Awards", "A", start, end)


# ===========================================================================
# Community Service  (defines ServiceHours)
# ===========================================================================
def build_service(wb):
    ws, start, end = build_log(
        wb, "Service", "🤝", "COMMUNITY SERVICE HOURS",
        "Log every hour — many scholarships & honor societies require a total.",
        ["Date", "Activity", "Hours"],
        SERVICE, [14, 34, 12], dates={1}, ints={3}, reserved=30)
    nrange(wb, "SvcHours", "Service", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL HOURS").style = "th"
    c = ws.cell(row=tot, column=3, value="=SUM(SvcHours)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "#,##0"
    ws.cell(row=tot, column=1).style = "td"; ws.cell(row=tot, column=1).fill = fill(SURFACE)
    cell_name(wb, "ServiceHours", "Service", f"$C${tot}")


# ===========================================================================
# Course Descriptions
# ===========================================================================
def build_descriptions(wb):
    ws = wb.create_sheet("Descriptions"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 62, 2])
    luxe_header(ws, "C", "📝  COURSE DESCRIPTIONS",
                "Homeschool essential — a short paragraph per course for college admissions.")
    table_headers(ws, 4, ["Course", "Description"], start_col=2)
    start = L0
    for i, (course, desc) in enumerate(DESCRIPTIONS):
        r = start + i
        ws.cell(row=r, column=2, value=course).style = "td_left"
        c = ws.cell(row=r, column=3, value=desc); c.style = "td_left"
        ws.row_dimensions[r].height = 58
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


# ===========================================================================
# Reading List
# ===========================================================================
def build_reading(wb):
    ws, start, end = build_log(
        wb, "Reading List", "📚", "READING LIST",
        "Great books read — strengthens English credits & the homeschool narrative.",
        ["Student", "Title", "Author", "Finished?"],
        READING, [16, 32, 18, 14], text_left={2}, reserved=40,
        validations=[("D", "YesNoList")])
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# Grad Requirements  (defines ReqDone / ReqName)
# ===========================================================================
def build_gradreq(wb):
    rows = [(item, "Yes" if done else "No") for (item, done) in GRADREQ]
    ws, start, end = build_log(
        wb, "Grad Requirements", "✅", "GRADUATION REQUIREMENTS",
        "The checklist to a diploma — confirm your state / umbrella's specific rules.",
        ["Requirement", "Met?"],
        rows, [46, 12], text_left={1}, reserved=20,
        validations=[("B", "YesNoList")])
    nrange(wb, "ReqName", "Grad Requirements", "A", start, end)
    nrange(wb, "ReqDone", "Grad Requirements", "B", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# Portfolio
# ===========================================================================
def build_portfolio(wb):
    ws = wb.create_sheet("Portfolio"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "🖼  PORTFOLIO & WORK SAMPLES",
                "A few standout pieces per year — proof behind the transcript for applications.")
    merge_set(ws, "B5:D5", "HOW TO ADD SAMPLES", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or =IMAGE(\"link\"). Excel: Insert "
                      "▸ Pictures ▸ Place in Cell. Add a research paper, lab report, art piece or project photo & "
                      "caption it (course · year · what it shows).")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Research Paper", "Lab Report", "Capstone Project", "Art / Music", "Essay", "Award / Certificate"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6; cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add sample)"); ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Course · year · notes…"); cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 3 — Student Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Student Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 26, 6, 22, 22, 2])
    luxe_header(ws, "G", "🎓  STUDENT PROFILE", "The student behind the transcript — the details colleges ask for.")
    blocks = [
        ("THE STUDENT", [("Name", "=StudentName"), ("Class Of", "=ClassOf"),
                         ("Graduation", "=GradDate"), ("Date of Birth", "—"),
                         ("Diploma Type", "College-prep"), ("Counselor / Parent", "—")]),
        ("THE SCHOOL", [("School Name", "=SchoolName"), ("Type", "Homeschool"),
                        ("Address", "—"), ("Phone / Email", "—"),
                        ("Grading Scale", "4.0 (see Settings)"), ("Weighted?", "Yes — H +0.5, AP +1.0")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "COLLEGE LIST", "section_gold"); ws.row_dimensions[15].height = 22
    for i, (c, s) in enumerate([("Reach school", "—"), ("Target school", "—"), ("Safety school", "—"),
                                ("Applied", "—"), ("Deadlines", "—"), ("Scholarships", "—")]):
        r = 16 + (i // 2); col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=c).style = "field_label"
        ws.cell(row=r, column=col + 1, value=s).style = "field_value"


# ===========================================================================
# Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🎓  HIGH SCHOOL TRANSCRIPT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  GPA, credits & a print-ready transcript — done right, done once.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "A COLLEGE-READY TRANSCRIPT, WITHOUT THE STRESS", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Type your courses and grades once — GPA (weighted AND unweighted), credits and an official-looking "
                      "transcript build themselves. Track requirements, test scores, activities, awards, service hours "
                      "and course descriptions in ONE premium Google Sheets & Excel system, then print a clean "
                      "transcript for college applications. Built for homeschool & college-bound families.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add the student, school, class year & grade scale.",
             "2.  Enter every class in Course Records — GPA points auto-calculate.",
             "3.  Check the GPA Calculator & Credit Tracker — weighted, unweighted, by year.",
             "4.  Fill Test Scores, Activities, Awards & Service — your application file.",
             "5.  Write Course Descriptions (homeschool must-have) & the 4-Year Plan.",
             "6.  Print the Official Transcript — and watch the College-Ready score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("A full sample transcript (Ella Bennett, Class of 2027 — 24 credits, 3.81 unweighted / 4.06 weighted GPA) "
               "is included so you can see how it all connects — just type over it. GPA, credits, the transcript and a "
               "College-Ready score update automatically. A matching printable PDF pack (official transcript, 4-year "
               "plan, activities résumé, course-description & GPA sheets) is included. Requirements vary by state / "
               "college — confirm your own; this is a record-keeping tool, not legal or admissions advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One clean transcript, one calm application season — you've got this.", "section_gold")


# ===========================================================================
# 2 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🎓  HIGH SCHOOL TRANSCRIPT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  GPA, credits, tests & activities — a college-ready transcript, automatically calculated.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("GPA (WEIGHTED)", "=GPAW", "gpa"),
        ("GPA (UNWEIGHTED)", "=GPAUW", "gpa"),
        ("CREDITS EARNED", "=CreditsEarned", "dec"),
        ("GRAD PROGRESS", "=IFERROR(MIN(CreditsEarned/CreditsReq,1),0)", "pct"),
        ("COURSES", "=CourseCount", "num"),
        ("HONORS / AP", "=HonorsAP", "num"),
    ]
    row2 = [
        ("CLASS OF", "=ClassOf", "text"),
        ("BEST SAT", "1380", "num"),
        ("BEST ACT", "30", "num"),
        ("SERVICE HRS", "=ServiceHours", "num"),
        ("ACTIVITIES", "=COUNTA(ActName)", "num"),
        ("COLLEGE-READY", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "COLLEGE-READY CHECK", "section_gold")
    merge_set(ws, "H11:M11", "CREDITS BY SUBJECT", "section_gold")
    table_headers(ws, 12, ["Area", "Progress", "Status"], start_col=2)
    dims = [
        ("GPA vs target", "=IFERROR(MIN(GPAUW/GPATarget,1),0)"),
        ("Credits vs required", "=IFERROR(MIN(CreditsEarned/CreditsReq,1),0)"),
        ("Grad requirements met", '=IFERROR(COUNTIF(ReqDone,"Yes")/COUNTA(ReqName),0)'),
        ("Test scores on file", "=IFERROR(MIN(COUNTA(TestName)/3,1),0)"),
        ("Course rigor (H+AP)", "=IFERROR(MIN(HonorsAP/RigorGoal,1),0)"),
        ("Service & activities", "=IFERROR(MIN(ServiceHours/ServiceGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"Good","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    db = BarChart(); db.type = "bar"; db.title = "Credits Earned by Subject"; db.height = 7.4; db.width = 8.4
    db.add_data(Reference(wb["Credit Tracker"], min_col=4, min_row=4, max_row=13), titles_from_data=True)
    db.set_categories(Reference(wb["Credit Tracker"], min_col=2, min_row=5, max_row=13)); db.legend = None
    ws.add_chart(db, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "High School Transcript Command Center™ — enter grades once; GPA & the transcript build themselves.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profile(wb); build_transcript(wb)
    build_courses(wb); build_gpa(wb); build_credits(wb); build_plan(wb)
    build_tests(wb); build_activities(wb); build_awards(wb); build_service(wb)
    build_descriptions(wb); build_reading(wb); build_gradreq(wb); build_portfolio(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Student Profile", "Official Transcript", "Course Records",
             "GPA Calculator", "Credit Tracker", "4-Year Plan", "Test Scores", "Activities", "Awards",
             "Service", "Descriptions", "Reading List", "Grad Requirements", "Portfolio", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "HS_Transcript_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
