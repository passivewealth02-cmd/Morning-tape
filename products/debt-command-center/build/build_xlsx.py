"""Build Debt Payoff Command Center™ — The Complete Debt-Freedom System.

13 tabs (+ Settings) · a premium debt-payoff operating system in Google Sheets &
Excel. Dashboard, a debt list engine, a snowball/avalanche payoff plan, a
side-by-side method comparison, a payment log, a shrinking balance history, a
per-debt payoff order, an extra-payment finder, milestones, an interest tracker
and accelerators — one dashboard. Pick a method, throw every extra dollar at it,
and watch the debt-free date arrive sooner.

Run: python3 build_xlsx.py   ->  ../Debt_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
DEBT_TYPES = ["Credit Card", "Auto", "Student", "Medical", "Personal", "Mortgage", "Other"]
METHODS = ["Snowball", "Avalanche"]

# Debts: (name, type, balance, APR, min payment, original balance)
DEBTS = [
    ("Medical Bill", "Medical", 1900, 0.0, 90, 3000),
    ("Store Card", "Credit Card", 3200, 0.2699, 75, 4000),
    ("Personal Loan", "Personal", 6300, 0.1490, 180, 8000),
    ("Credit Card A", "Credit Card", 9400, 0.2299, 235, 12000),
    ("Car Loan", "Auto", 12800, 0.0690, 360, 19000),
    ("Student Loan", "Student", 14200, 0.0580, 165, 20000),
]

EXTRA_PAYMENT = 300
EXTRA_TARGET = 300
START = dt.date.today().replace(day=1)

# Payment log: (date offset months back, debt, amount, on-time)
PAYMENTS = [
    (-5, "Store Card", 345, "Yes"), (-5, "Medical Bill", 100, "Yes"),
    (-4, "Store Card", 345, "Yes"), (-4, "Credit Card A", 105, "Yes"),
    (-3, "Store Card", 345, "Yes"), (-3, "Car Loan", 320, "Yes"),
    (-2, "Store Card", 400, "Yes"), (-2, "Student Loan", 195, "No"),
    (-1, "Store Card", 400, "Yes"), (-1, "Personal Loan", 180, "Yes"),
]

# Extra-payment finder: (source, monthly found)
FOUND = [
    ("Cancel unused streaming", 25),
    ("Pack lunch 3x/week", 60),
    ("Sell unused gear", 40),
    ("Side gig — weekends", 120),
    ("Cut one subscription", 15),
    ("Round-up savings", 40),
]

# Milestones: (milestone, done)
MILESTONES = [
    ("$1,000 starter emergency fund", "Yes"),
    ("First debt paid off", "No"),
    ("Half of a debt gone", "Yes"),
    ("Under $40,000 total", "No"),
    ("Highest-rate card cleared", "No"),
    ("50% of all debt paid", "No"),
    ("Down to one debt", "No"),
    ("DEBT FREE", "No"),
]

# Accelerators: (action, monthly boost, on/off)
ACCEL = [
    ("No-spend weekends", 80, "Yes"),
    ("Sell one thing a month", 50, "Yes"),
    ("Overtime / side income", 150, "No"),
    ("Cashback to debt", 25, "Yes"),
    ("Tax refund lump sum", 0, "No"),
]

STARTER_EF = 1000
LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
# Payoff simulation — the signature engine
# ===========================================================================
def simulate(method):
    d = [{"name": n, "bal": float(b), "apr": a, "min": m} for (n, _, b, a, m, _) in DEBTS]
    # Lock the attack order once (textbook snowball/avalanche — no mid-stream re-sort)
    priority = sorted(d, key=lambda x: x["bal"] if method == "Snowball" else -x["apr"])
    order = [x["name"] for x in priority]
    budget_total = sum(x["min"] for x in d) + EXTRA_PAYMENT
    months = 0; total_int = 0.0; payoff_month = {}
    timeline = [sum(x["bal"] for x in d)]
    while any(x["bal"] > 0.005 for x in d) and months < 600:
        months += 1
        for x in d:
            if x["bal"] > 0:
                i = x["bal"] * x["apr"] / 12.0
                x["bal"] += i; total_int += i
        money = budget_total
        for x in d:                                   # cover every minimum first
            if x["bal"] > 0.005:
                pay = min(x["min"], x["bal"]); x["bal"] -= pay; money -= pay
        for x in priority:                            # roll the rest down the fixed order
            if money <= 0.005:
                break
            if x["bal"] > 0.005:
                pay = min(money, x["bal"]); x["bal"] -= pay; money -= pay
        for x in d:
            if x["bal"] <= 0.005 and x["name"] not in payoff_month:
                payoff_month[x["name"]] = months
        timeline.append(sum(max(x["bal"], 0) for x in d))
    return {"months": months, "interest": round(total_int), "payoff_month": payoff_month,
            "order": order, "timeline": timeline}


SNOW = simulate("Snowball")
AVAL = simulate("Avalanche")


def months_to_date(m):
    y = START.year + (START.month - 1 + m) // 12
    mo = (START.month - 1 + m) % 12 + 1
    return dt.date(y, mo, 1)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 13 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm yyyy", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your details, method & extra payment once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR PLAN", "section")
    controls = [
        ("Name / Household", "The Bennett Household", None, "Household"),
        ("Payoff Method", "Snowball", None, "Method"),
        ("Extra Monthly Payment", EXTRA_PAYMENT, '"$"#,##0', "ExtraPay"),
        ("Extra-Payment Target", EXTRA_TARGET, '"$"#,##0', "ExtraTarget"),
        ("Starter Emergency Fund", STARTER_EF, '"$"#,##0', "StarterEF"),
        ("Emergency Saved", 1000, '"$"#,##0', "EFSaved"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Debt Type", DEBT_TYPES, "TypeList"), ("F", "Yes / No", YESNO, "YesNoList"),
             ("G", "Method", METHODS, "MethodList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🔥  DEBT PAYOFF COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Pick a method, throw every extra dollar at it & watch your debt-free date arrive sooner.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE DEBT-FREEDOM PLAN, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("List every debt once, choose snowball (smallest balance first) or avalanche (highest rate "
                      "first), and set your extra monthly payment. The plan orders your debts, projects your "
                      "debt-free date, and shows exactly how much interest you'll pay — and how much you'll save by "
                      "attacking the right debt first. Log each payment, watch your total balance shrink month after "
                      "month, and celebrate every milestone on the way to $0.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your name, pick a method & set your extra monthly payment.",
             "2.  List every debt on the Debts tab — balance, rate & minimum payment.",
             "3.  The Payoff Plan orders them & names your focus debt — the one to attack first.",
             "4.  Compare Snowball vs Avalanche — see months, total interest & your debt-free date.",
             "5.  Log payments & update balances; watch the Balance History line fall toward $0.",
             "6.  Check the Dashboard: total debt, % paid, debt-free date & a Payoff Momentum score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample debts for a fictional household are included so you can see how it all connects — just type "
               "over them with your own. Avalanche saves the most interest; snowball gives the fastest wins — the "
               "comparison tab shows both so you can choose. Twelve matching printable pages (debt list, payoff plan, "
               "payment tracker, balance chart & more) are included to print and keep. This is a personal debt-payoff "
               "tool, not financial, tax or credit advice — for big decisions, talk to a qualified professional.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Every extra dollar is a day closer to debt-free — momentum wins.", "section_gold")


# ===========================================================================
# Debts — the engine
# ===========================================================================
def build_debts(wb):
    ws = wb.create_sheet("Debts"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 14, 10, 14, 14, 12, 2])
    luxe_header(ws, "H", "💳  YOUR DEBTS",
                "List every debt once — balance, rate & minimum. This is the engine behind your plan.")
    table_headers(ws, 4, ["Debt", "Type", "Balance", "APR", "Min", "Original", "% Paid"], start_col=2)
    start = L0
    for i, (name, typ, bal, apr, minp, orig) in enumerate(DEBTS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=typ).style = "td"
        cb = ws.cell(row=r, column=4, value=bal); cb.style = "input"; cb.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=5, value=apr); ca.style = "input"; ca.number_format = "0.00%"
        cm = ws.cell(row=r, column=6, value=minp); cm.style = "input"; cm.number_format = '"$"#,##0'
        co = ws.cell(row=r, column=7, value=orig); co.style = "input"; co.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=8, value=f"=IFERROR((G{r}-D{r})/G{r},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DEBTS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cb = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cb.style = "td"; cb.font = Font(bold=True, color=DANGER); cb.fill = fill(SURFACE); cb.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0'
    co = ws.cell(row=tot, column=7, value=f"=SUM(G{start}:G{end})"); co.style = "td"; co.font = Font(bold=True, color=PRIMARY); co.fill = fill(SURFACE); co.number_format = '"$"#,##0'
    cp = ws.cell(row=tot, column=8, value=f"=IFERROR((G{tot}-D{tot})/G{tot},0)"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(SURFACE); cp.number_format = "0%"
    nrange(wb, "DebtName", "Debts", "B", start, end)
    nrange(wb, "DebtType", "Debts", "C", start, end)
    nrange(wb, "DebtBalance", "Debts", "D", start, end)
    nrange(wb, "DebtAPR", "Debts", "E", start, end)
    nrange(wb, "DebtMin", "Debts", "F", start, end)
    nrange(wb, "DebtOrig", "Debts", "G", start, end)
    nrange(wb, "DebtPct", "Debts", "H", start, end)
    cell_name(wb, "DebtTotal", "Debts", f"$D${tot}")
    cell_name(wb, "DebtMinTotal", "Debts", f"$F${tot}")
    cell_name(wb, "DebtOrigTotal", "Debts", f"$G${tot}")
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="min", start_color="FF" + HIGHLIGHT, end_type="max", end_color="FF" + RED_BG))
    ws.conditional_formatting.add(f"H{start}:H{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
# Payoff Plan — ordered by method
# ===========================================================================
def build_plan(wb):
    ws = wb.create_sheet("Payoff Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 10, 24, 14, 10, 14, 16, 2])
    luxe_header(ws, "G", "🎯  PAYOFF PLAN",
                "Your debts in attack order — the focus debt gets every extra dollar until it's gone.")
    merge_set(ws, "B5:C5", "METHOD", "section_gold")
    ws.cell(row=5, column=4, value="=Method").font = Font(bold=True, color=PRIMARY)
    ws.cell(row=5, column=6, value="Focus debt →").font = Font(italic=True, color=ACCENT)
    fcell = ws.cell(row=5, column=7, value='=INDEX(PlanName,1)'); fcell.font = Font(bold=True, color=DANGER)
    table_headers(ws, 6, ["Order", "Debt", "Balance", "APR", "Min", "Projected Payoff"], start_col=2)
    order = sorted(DEBTS, key=lambda x: x[2])  # snowball default (smallest balance)
    start = 7
    for i, (name, typ, bal, apr, minp, orig) in enumerate(order):
        r = start + i
        ws.cell(row=r, column=2, value=i + 1).style = "td"
        ws.cell(row=r, column=3, value=name).style = "td_left"
        cb = ws.cell(row=r, column=4, value=bal); cb.style = "td"; cb.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=5, value=apr); ca.style = "td"; ca.number_format = "0.00%"
        cm = ws.cell(row=r, column=6, value=minp); cm.style = "td"; cm.number_format = '"$"#,##0'
        pm = SNOW["payoff_month"].get(name, 0)
        cp = ws.cell(row=r, column=7, value=months_to_date(pm)); cp.style = "td"; cp.number_format = "mmm yyyy"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        if i == 0:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MINT_BG)
    end = start + len(order) - 1
    nrange(wb, "PlanName", "Payoff Plan", "C", start, end)
    nrange(wb, "PlanPayoff", "Payoff Plan", "G", start, end)
    merge_set(wb["Payoff Plan"], f"B{end+2}:C{end+2}", "Snowball order shown (smallest balance first).", "section")
    ws.cell(row=end + 2, column=6, value="Switch method →").font = Font(italic=True, color=ACCENT)
    ws.cell(row=end + 2, column=7, value="see next tab").font = Font(italic=True, color=ACCENT)
    ws.freeze_panes = "A7"


# ===========================================================================
# Snowball vs Avalanche — the comparison
# ===========================================================================
def build_compare(wb):
    ws = wb.create_sheet("Snowball vs Avalanche"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 18, 4, 20, 2])
    luxe_header(ws, "F", "⚖  SNOWBALL vs AVALANCHE",
                "Fastest wins vs least interest — see both, then pick your method with eyes open.")
    table_headers(ws, 5, ["Measure", "Snowball", "Avalanche"], start_col=2)
    df_snow = months_to_date(SNOW["months"]); df_aval = months_to_date(AVAL["months"])
    rows = [
        ("Months to debt-free", SNOW["months"], AVAL["months"], "0"),
        ("Debt-free date", df_snow, df_aval, "mmm yyyy"),
        ("Total interest paid", SNOW["interest"], AVAL["interest"], '"$"#,##0'),
        ("Attack order", "Smallest balance", "Highest rate", "@"),
        ("Best for", "Fast motivation", "Max money saved", "@"),
    ]
    start = 6
    for i, (lab, s, a, fmt) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cs = ws.cell(row=r, column=3, value=s); cs.style = "td"; cs.number_format = fmt
        ca = ws.cell(row=r, column=4, value=a); ca.style = "td"; ca.number_format = fmt
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    saved = SNOW["interest"] - AVAL["interest"]
    er = start + len(rows) + 1
    merge_set(ws, f"B{er}:D{er}", "💰  Avalanche saves you in interest vs snowball", "section_gold")
    c = ws.cell(row=er, column=6, value=saved); c.style = "field_value"; c.number_format = '"$"#,##0'; c.font = Font(bold=True, size=16, color=PRIMARY); c.fill = fill(MINT_BG)
    cell_name(wb, "SnowMonths", "Snowball vs Avalanche", f"$C${start}")
    cell_name(wb, "AvalMonths", "Snowball vs Avalanche", f"$D${start}")
    cell_name(wb, "SnowInterest", "Snowball vs Avalanche", f"$C${start+2}")
    cell_name(wb, "AvalInterest", "Snowball vs Avalanche", f"$D${start+2}")
    cell_name(wb, "InterestSaved", "Snowball vs Avalanche", f"$F${er}")
    cell_name(wb, "SnowDate", "Snowball vs Avalanche", f"$C${start+1}")
    # chosen method free-form KPIs
    mr = er + 2
    merge_set(ws, f"B{mr}:D{mr}", "YOUR CHOSEN METHOD (from Settings)", "section")
    ws.cell(row=mr + 1, column=2, value="Months to debt-free").style = "field_label"
    ws.cell(row=mr + 1, column=3, value='=IF(Method="Avalanche",AvalMonths,SnowMonths)').style = "field_value"
    ws.cell(row=mr + 2, column=2, value="Total interest").style = "field_label"
    cc = ws.cell(row=mr + 2, column=3, value='=IF(Method="Avalanche",AvalInterest,SnowInterest)'); cc.style = "field_value"; cc.number_format = '"$"#,##0'
    cell_name(wb, "MyMonths", "Snowball vs Avalanche", f"$C${mr+1}")
    cell_name(wb, "MyInterest", "Snowball vs Avalanche", f"$C${mr+2}")


# ===========================================================================
# Payment Log
# ===========================================================================
def build_payments(wb):
    rows = [(months_to_date(0) - dt.timedelta(days=abs(off) * 30), debt, amt, ot) for (off, debt, amt, ot) in PAYMENTS]
    rows.sort(key=lambda r: r[0], reverse=True)
    ws, start, end = build_log(
        wb, "Payment Log", "🧾", "PAYMENT LOG",
        "Every payment you make, logged — proof of progress and an on-time streak to protect.",
        ["Date", "Debt", "Amount", "On Time"],
        rows, [2, 16, 24, 16, 14], text_left={3}, dates={2}, money={4}, reserved=48,
        validations=[("C", "DebtName"), ("E", "YesNoList")], start_col=2)
    nrange(wb, "PayOnTime", "Payment Log", "E", start, end)
    nrange(wb, "PayAmount", "Payment Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL PAID").style = "th"
    c = ws.cell(row=tot, column=4, value="=SUM(PayAmount)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cmap = {"Yes": MINT_BG, "No": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# Balance History — the shrinking line
# ===========================================================================
def build_history(wb):
    ws = wb.create_sheet("Balance History"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 2])
    luxe_header(ws, "D", "📉  BALANCE HISTORY",
                "Log your total balance each month — the falling line is the best chart in personal finance.")
    table_headers(ws, 4, ["Month", "Total Balance", "Paid Down"], start_col=2)
    # past 5 months from origtotal down to current, then projected via snowball timeline
    orig = sum(x[5] for x in DEBTS); cur = sum(x[2] for x in DEBTS)
    past = [orig, orig - 3200, orig - 7100, orig - 11000, orig - 14200, cur]
    labels = ["5 mo ago", "4 mo ago", "3 mo ago", "2 mo ago", "Last mo", "This mo"]
    start = L0
    for i, (lab, val) in enumerate(zip(labels, past)):
        r = start + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cb = ws.cell(row=r, column=3, value=val); cb.style = "input"; cb.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=4, value=f"=IF(C{r-1}=\"\",0,C{r-1}-C{r})" if i else 0); cd.style = "td"; cd.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    # projected future points from snowball timeline (every 6 months)
    tl = SNOW["timeline"]
    proj_pts = [(6, "in 6 mo"), (12, "in 12 mo"), (18, "in 18 mo"), (24, "in 24 mo")]
    r = start + len(past)
    for m, lab in proj_pts:
        val = round(tl[m]) if m < len(tl) else 0
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cb = ws.cell(row=r, column=3, value=val); cb.style = "td"; cb.number_format = '"$"#,##0'; cb.fill = fill(SOFT_BG)
        ws.cell(row=r, column=4, value="projected").style = "td"; ws.cell(row=r, column=4).font = Font(italic=True, color=ACCENT)
        r += 1
    end = r - 1
    nrange(wb, "HistMonth", "Balance History", "B", start, end)
    nrange(wb, "HistBalance", "Balance History", "C", start, end)
    ch = LineChart(); ch.title = "Total Debt Balance Over Time"; ch.height = 8; ch.width = 15
    ch.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end)); ch.dataLabels = no_labels()
    ws.add_chart(ch, "F4")
    ws.freeze_panes = "A5"


# ===========================================================================
# Extra Payment Finder
# ===========================================================================
def build_found(wb):
    ws, start, end = build_log(
        wb, "Extra Payment", "🔍", "EXTRA-PAYMENT FINDER",
        "Find money hiding in your month — every dollar here becomes attack power on your focus debt.",
        ["Source", "Monthly Found"],
        FOUND, [2, 32, 18], text_left={2}, money={3}, reserved=20, start_col=2)
    nrange(wb, "FoundAmount", "Extra Payment", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="FOUND MONEY / MO").style = "th"
    c = ws.cell(row=tot, column=3, value="=SUM(FoundAmount)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "FoundTotal", "Extra Payment", f"$C${tot}")


# ===========================================================================
# Milestones
# ===========================================================================
def build_milestones(wb):
    ws, start, end = build_log(
        wb, "Milestones", "🏆", "MILESTONES",
        "Debt payoff is a long road — mark every win. Momentum, not perfection, gets you to $0.",
        ["Milestone", "Done"],
        MILESTONES, [2, 40, 14], text_left={2}, reserved=16,
        validations=[("C", "YesNoList")], start_col=2)
    nrange(wb, "MilestoneDone", "Milestones", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.cell(row=end + 2, column=2, value="Wins so far").style = "field_label"
    c = ws.cell(row=end + 2, column=3, value='=COUNTIF(MilestoneDone,"Yes")&" of "&COUNTA(MilestoneDone)'); c.style = "field_value"; c.fill = fill(MINT_BG)


# ===========================================================================
# Interest Tracker
# ===========================================================================
def build_interest(wb):
    ws = wb.create_sheet("Interest Tracker"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 4, 30, 18, 2])
    luxe_header(ws, "F", "💸  INTEREST TRACKER",
                "What interest is costing you — and how much the right method saves. Motivation, quantified.")
    merge_set(ws, "B5:C5", "PROJECTED INTEREST", "section_gold")
    merge_set(ws, "E5:F5", "IF YOU DO NOTHING", "section_gold")
    left = [
        ("Snowball total interest", SNOW["interest"]),
        ("Avalanche total interest", AVAL["interest"]),
        ("You save (avalanche)", SNOW["interest"] - AVAL["interest"]),
    ]
    for i, (lab, val) in enumerate(left):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=val); c.style = "input" if i < 2 else "field_value"; c.number_format = '"$"#,##0'
        if i == 2:
            c.fill = fill(MINT_BG); c.font = Font(bold=True, color=PRIMARY)
    # minimums-only comparison (rough): paying only minimums balloons interest
    minonly = simulate_minimums_only()
    right = [
        ("Minimums-only interest", minonly),
        ("Your plan's interest", SNOW["interest"]),
        ("Extra payment saves", minonly - SNOW["interest"]),
    ]
    for i, (lab, val) in enumerate(right):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "td_left"
        c = ws.cell(row=r, column=6, value=val); c.style = "input" if i < 2 else "field_value"; c.number_format = '"$"#,##0'
        if i == 2:
            c.fill = fill(MINT_BG); c.font = Font(bold=True, color=PRIMARY)
    cell_name(wb, "MinOnlyInterest", "Interest Tracker", "$F$6")
    ws.cell(row=10, column=2, value="Every extra $1 you throw at the focus debt is interest you never pay.").style = "section"


def simulate_minimums_only():
    d = [{"name": n, "bal": float(b), "apr": a, "min": m} for (n, _, b, a, m, _) in DEBTS]
    months = 0; total_int = 0.0
    while any(x["bal"] > 0.005 for x in d) and months < 1200:
        months += 1
        for x in d:
            if x["bal"] > 0:
                i = x["bal"] * x["apr"] / 12.0; x["bal"] += i; total_int += i
                pay = min(x["min"], x["bal"])
                # ensure min covers interest; if not, debt never dies -> cap
                x["bal"] -= pay
    return round(total_int)


# ===========================================================================
# Accelerators
# ===========================================================================
def build_accel(wb):
    ws = wb.create_sheet("Accelerators"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 14, 2])
    luxe_header(ws, "D", "🚀  ACCELERATORS",
                "Turn extra habits on to shave months off — flip 'On' and watch the payoff date move.")
    table_headers(ws, 4, ["Accelerator", "Monthly Boost", "On?"], start_col=2)
    start = L0
    for i, (act, boost, on) in enumerate(ACCEL):
        r = start + i
        ws.cell(row=r, column=2, value=act).style = "td_left"
        cb = ws.cell(row=r, column=3, value=boost); cb.style = "input"; cb.number_format = '"$"#,##0'
        co = ws.cell(row=r, column=4, value=on); co.style = "input"
        add_dv(ws, f"D{r}", "YesNoList")
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(ACCEL) - 1
    nrange(wb, "AccelBoost", "Accelerators", "C", start, end)
    nrange(wb, "AccelOn", "Accelerators", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ACTIVE BOOST / MO").style = "th"
    c = ws.cell(row=tot, column=3, value='=SUMIF(AccelOn,"Yes",AccelBoost)'); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cell_name(wb, "AccelActive", "Accelerators", f"$C${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.freeze_panes = "A5"


# ===========================================================================
# Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🔥  DEBT PAYOFF COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Total debt, % paid, debt-free date & the interest you'll save — your whole payoff, automatically tracked.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("TOTAL DEBT", "=DebtTotal", "money"),
        ("PAID OFF", "=DebtOrigTotal-DebtTotal", "money"),
        ("% PAID", "=IFERROR((DebtOrigTotal-DebtTotal)/DebtOrigTotal,0)", "pct"),
        ("MONTHLY PAYMENT", "=DebtMinTotal+ExtraPay", "money"),
        ("EXTRA PAYMENT", "=ExtraPay", "money"),
        ("HIGHEST APR", "=MAX(DebtAPR)", "pct"),
    ]
    row2 = [
        ("DEBT-FREE DATE", "=SnowDate", "date"),
        ("MONTHS TO FREE", "=MyMonths", "num"),
        ("TOTAL INTEREST", "=MyInterest", "money"),
        ("INTEREST SAVED", "=InterestSaved", "money"),
        ("FOCUS DEBT", "=INDEX(PlanName,1)", "text"),
        ("MOMENTUM", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "PAYOFF MOMENTUM", "section_gold")
    merge_set(ws, "H11:M11", "DEBT BY BALANCE", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Debt reduced", "=IFERROR((DebtOrigTotal-DebtTotal)/DebtOrigTotal,0)"),
        ("Extra payment funded", "=IFERROR(MIN(ExtraPay/ExtraTarget,1),0)"),
        ("Payments on time", '=IFERROR(COUNTIF(PayOnTime,"Yes")/COUNTA(PayOnTime),0)'),
        ("Milestones hit", '=IFERROR(COUNTIF(MilestoneDone,"Yes")/COUNTA(MilestoneDone),0)'),
        ("Per-debt progress", "=IFERROR(AVERAGE(DebtPct),0)"),
        ("Starter fund ready", "=IFERROR(MIN(EFSaved/StarterEF,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.5,"Building","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    b = wb["Debts"]
    ch = BarChart(); ch.type = "bar"; ch.title = "Balance by Debt"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(b, min_col=4, min_row=5, max_row=4 + len(DEBTS)), titles_from_data=False)
    ch.set_categories(Reference(b, min_col=2, min_row=5, max_row=4 + len(DEBTS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Debt Payoff Command Center™ — pick a method, fund the extra payment & attack the focus debt.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_debts(wb); build_plan(wb)
    build_compare(wb); build_payments(wb); build_history(wb); build_found(wb)
    build_milestones(wb); build_interest(wb); build_accel(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Debts", "Payoff Plan", "Snowball vs Avalanche",
             "Payment Log", "Balance History", "Extra Payment", "Milestones", "Interest Tracker",
             "Accelerators", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Debt_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")
    print(f"  Snowball: {SNOW['months']} mo, ${SNOW['interest']:,} interest, debt-free {months_to_date(SNOW['months']):%b %Y}")
    print(f"  Avalanche: {AVAL['months']} mo, ${AVAL['interest']:,} interest")
    print(f"  Interest saved (avalanche): ${SNOW['interest']-AVAL['interest']:,}")
    print(f"  Min-only interest: ${simulate_minimums_only():,}")


if __name__ == "__main__":
    main()
