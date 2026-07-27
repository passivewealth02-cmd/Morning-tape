"""Build Notary & Loan Signing Agent Command Center™ — The Signing Agent's Operating System.

14 tabs · a premium mobile-notary & loan-signing-agent operating system in Google Sheets &
Excel. Dashboard, a per-signing profit engine (fee − printing − driving, over the hours it
ACTUALLY takes door to door), a signings log, a fee schedule, a mileage log, printing
costs, invoices & who hasn't paid, signing companies, a notarial journal, expenses, a tax
set-aside and a monthly summary — one dashboard. That $125 signing is not $138 an hour.

Run: python3 build_xlsx.py   ->  ../Notary_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
SIGNTYPE = ["Refinance", "Purchase / seller", "Reverse mortgage", "Loan modification",
            "Single notarization", "Jail / hospital", "Apostille", "HELOC"]
PAYSTATUS = ["Paid", "Invoiced", "Overdue", "Not paid \u2014 fell through"]
IDTYPE = ["Driver's licence", "Passport", "State ID", "Military ID", "Credible witness"]
EXPCAT = ["Insurance", "Commission", "Equipment", "Software", "Vehicle", "Supplies", "Education", "Other"]

# --- Per-signing profit engine ---
AVG_FEE = 125.00
PAGES_PER_SIGNING = 180
COST_PER_PAGE = 0.035
MILES_PER_SIGNING = 38
VEHICLE_COST_PER_MILE = 0.22
IRS_MILEAGE_RATE = 0.70
DRIVE_HOURS = 1.1
APPT_HOURS = 0.9
PREP_HOURS = 0.5
SIGNINGS_MONTH = 52
UNPAID = 2
RECEIVABLE = 1250
TAX_SET_ASIDE = 2100
TAX_RESERVE_GOAL = 5250

# --- Goals ---
HOURLY_GOAL = 40
MARGIN_GOAL = 0.60
COVER_GOAL = 4.0
PAID_GOAL = 0.95
SIGNING_GOAL = 52

# Fixed costs — what the business costs before a single signing: (line, category, monthly)
FIXED_LINES = [
    ("E&O insurance", "Insurance", 38), ("Notary bond & commission (amortized)", "Commission", 22),
    ("Printer lease & maintenance", "Equipment", 95), ("Signing platform & scheduling", "Software", 45),
    ("Phone & internet share", "Software", 60), ("Background check & NNA renewal (amortized)", "Education", 29),
]

# Fee schedule: (service, fee, notes)
FEES = [
    ("Refinance signing", 125, "Standard \u2014 150-200 pages"),
    ("Purchase / seller package", 150, "Two packages, often two locations"),
    ("Reverse mortgage", 175, "Longer appointment, counselling docs"),
    ("Loan modification", 90, "Short package"),
    ("HELOC", 100, "Short package, quick turnaround"),
    ("Single notarization", 15, "Plus travel \u2014 check your state maximum"),
    ("Jail / hospital signing", 85, "Includes wait time"),
    ("Apostille prep & courier", 95, "Per document set"),
    ("After-hours surcharge (after 7pm)", 35, "Add to base fee"),
    ("Printing surcharge (over 150 pages)", 25, "Add to base fee"),
    ("Scanbacks", 20, "Add to base fee"),
    ("Travel beyond 30 miles (per mile)", 1.25, "Add to base fee"),
    ("Trip fee \u2014 no-show or cancelled at door", 60, "Bill it. Every time."),
]

# Signings log (a fortnight's worth): (date, company, type, fee, pages, miles, status)
SIGNINGS = [
    ("07/01", "Summit Signing", "Refinance", 125, 178, 34, "Paid"),
    ("07/02", "Clearview Title", "Purchase / seller", 150, 212, 41, "Paid"),
    ("07/03", "Summit Signing", "Refinance", 125, 164, 22, "Paid"),
    ("07/05", "Anchor Notary Net", "HELOC", 100, 96, 48, "Paid"),
    ("07/07", "Clearview Title", "Refinance", 125, 186, 29, "Paid"),
    ("07/08", "Meridian Escrow", "Reverse mortgage", 175, 246, 52, "Invoiced"),
    ("07/09", "Summit Signing", "Refinance", 125, 172, 31, "Paid"),
    ("07/10", "Direct \u2014 Ortega", "Single notarization", 45, 4, 12, "Paid"),
    ("07/12", "Anchor Notary Net", "Loan modification", 90, 74, 44, "Overdue"),
    ("07/14", "Clearview Title", "Purchase / seller", 150, 204, 38, "Invoiced"),
    ("07/15", "Meridian Escrow", "Refinance", 125, 190, 27, "Paid"),
    ("07/16", "Summit Signing", "Refinance", 125, 168, 36, "Not paid \u2014 fell through"),
    ("07/17", "Direct \u2014 Hale", "Jail / hospital", 85, 6, 19, "Paid"),
    ("07/19", "Anchor Notary Net", "Refinance", 125, 182, 40, "Invoiced"),
]

# Mileage log: (date, trip, miles, purpose)
MILEAGE = [
    ("07/01", "Home \u2192 Cedar Heights \u2192 home", 34, "Refinance signing"),
    ("07/02", "Home \u2192 Westbrook \u2192 title office \u2192 home", 41, "Purchase signing + drop"),
    ("07/03", "Home \u2192 Old Mill \u2192 home", 22, "Refinance signing"),
    ("07/05", "Home \u2192 Fairhaven \u2192 home", 48, "HELOC signing"),
    ("07/07", "Home \u2192 Riverside \u2192 home", 29, "Refinance signing"),
    ("07/08", "Home \u2192 Lakeview \u2192 home", 52, "Reverse mortgage"),
    ("07/09", "Home \u2192 Northgate \u2192 home", 31, "Refinance signing"),
    ("07/10", "Home \u2192 downtown \u2192 home", 12, "Single notarization"),
    ("07/11", "Home \u2192 office supply \u2192 home", 9, "Toner & paper"),
    ("07/12", "Home \u2192 Brookside \u2192 home", 44, "Loan modification"),
]

# Printing & supplies: (item, cost, yields, unit)
PRINTING = [
    ("Toner \u2014 black, high yield", 118.00, 6000, "pages"),
    ("Paper \u2014 letter, case", 42.00, 5000, "sheets"),
    ("Paper \u2014 legal, case", 58.00, 5000, "sheets"),
    ("Printer maintenance kit", 149.00, 25000, "pages"),
    ("Pens, seals & thumbprint pads", 34.00, 200, "signings"),
]

# Invoices & payments: (company, signings, amount, status)
INVOICES = [
    ("Summit Signing", 18, 2250, "Paid"),
    ("Clearview Title", 11, 1525, "Paid"),
    ("Anchor Notary Net", 9, 990, "Invoiced"),
    ("Meridian Escrow", 8, 1150, "Invoiced"),
    ("Direct clients", 6, 585, "Paid"),
]

# Signing companies: (company, contact, pay terms, avg fee, notes)
COMPANIES = [
    ("Summit Signing", "scheduling@summit.example", "Net 15", 125, "Reliable, pays early"),
    ("Clearview Title", "orders@clearview.example", "Net 30", 139, "Best fees, occasional rush"),
    ("Anchor Notary Net", "dispatch@anchor.example", "Net 45", 105, "Slow payer \u2014 chase at day 45"),
    ("Meridian Escrow", "signings@meridian.example", "Net 30", 144, "Long packages, worth it"),
    ("Direct clients", "\u2014", "At signing", 98, "Best margin \u2014 build this"),
]

# Notarial journal: (date, document type, signer, ID type, fee, thumbprint)
JOURNAL = [
    ("07/01", "Deed of trust", "R. Alvarez", "Driver's licence", 15, "Yes"),
    ("07/01", "Promissory note", "R. Alvarez", "Driver's licence", 15, "Yes"),
    ("07/02", "Grant deed", "M. Whitfield", "Passport", 15, "Yes"),
    ("07/02", "Deed of trust", "M. Whitfield", "Passport", 15, "Yes"),
    ("07/03", "Deed of trust", "J. Okafor", "Driver's licence", 15, "Yes"),
    ("07/05", "HELOC agreement", "S. Nakamura", "State ID", 15, "Yes"),
    ("07/07", "Deed of trust", "L. Brennan", "Driver's licence", 15, "Yes"),
    ("07/08", "Reverse mortgage note", "E. Kowalski", "Driver's licence", 15, "Yes"),
    ("07/09", "Deed of trust", "P. Ramanathan", "Passport", 15, "Yes"),
    ("07/10", "Power of attorney", "D. Ortega", "Driver's licence", 15, "Yes"),
    ("07/12", "Modification agreement", "T. Feld", "Driver's licence", 15, "Yes"),
    ("07/17", "Advance directive", "C. Hale", "Military ID", 15, "Yes"),
]

# Expenses beyond the fixed lines: (item, category, monthly)
EXPENSES = [
    ("Toner & paper", "Supplies", 78), ("Vehicle fuel & wear", "Vehicle", 435),
    ("Parking & tolls", "Vehicle", 22), ("Continuing education", "Education", 25),
    ("Marketing & business cards", "Other", 30),
]

# Tax set-aside: (quarter, income, set aside, estimated due)
TAXES = [
    ("Q1", 15400, 600, 1300), ("Q2", 16800, 700, 1300),
    ("Q3", 17200, 800, 1325), ("Q4", 17600, 0, 1325),
]

# Monthly summary: (month, revenue)
MONTHS = [("Feb", 5125), ("Mar", 5500), ("Apr", 5875), ("May", 6125), ("Jun", 6375), ("Jul", 6500)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "money3": '"$"#,##0.000', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 40, 20, 3] + [24] * 5)
    luxe_header(ws, "I", "⚙  SETTINGS", "Set your rates & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR RATES & GOALS", "section")
    controls = [
        ("Business", "Quill & Seal Notary", None, "Business"),
        ("Signing agent", "Sloane", None, "Agent"),
        ("Average signing fee", AVG_FEE, '"$"#,##0.00', "AvgFee"),
        ("Cost per printed page", COST_PER_PAGE, '"$"#,##0.000', "CostPerPage"),
        ("Vehicle cost per mile (out of pocket)", VEHICLE_COST_PER_MILE, '"$"#,##0.00', "VehicleCPM"),
        ("IRS mileage rate (deduction)", IRS_MILEAGE_RATE, '"$"#,##0.00', "IRSRate"),
        ("Signings this month", SIGNINGS_MONTH, "0", "SigningsMonth"),
        ("Signings that fell through unpaid", UNPAID, "0", "Unpaid"),
        ("Your hourly goal", HOURLY_GOAL, '"$"#,##0', "HourlyGoal"),
        ("Margin goal", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Break-even cover goal (×)", COVER_GOAL, "0.0", "CoverGoal"),
        ("Getting-paid goal", PAID_GOAL, "0%", "PaidGoal"),
        ("Signings per month goal", SIGNING_GOAL, "0", "SigningGoal"),
        ("Tax reserve goal (year)", TAX_RESERVE_GOAL, '"$"#,##0', "TaxReserveGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Signing type", SIGNTYPE, "SignTypeList"), ("F", "Payment status", PAYSTATUS, "PayStatusList"),
             ("G", "ID type", IDTYPE, "IDTypeList"), ("H", "Expense category", EXPCAT, "ExpCatList"),
             ("I", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:I5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  \U0001f58b  NOTARY & LOAN SIGNING AGENT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  That $125 signing is not $138 an hour.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "THE WHOLE BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Here is the arithmetic every new signing agent gets wrong. A $125 refinance is a 45-minute "
                      "appointment, so it feels like well over a hundred dollars an hour. But you printed 180 pages "
                      "before you left, you drove 38 miles round trip, and between prep, drive time and the "
                      "appointment itself the job took two and a half hours door to door. The real number is $44 an "
                      "hour. That is still a good business \\u2014 it is just a completely different business from the one "
                      "in your head. This workbook does that maths on every signing, then runs your fee schedule, "
                      "mileage, printing, invoices, signing companies, your notarial journal and your tax set-aside.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — your average fee, page cost and mileage rate.",
             "2.  Signing Profit: enter pages, miles and how long it really takes.",
             "3.  Read your true net per signing and your real hourly rate.",
             "4.  Log signings, mileage and your notarial journal as you go.",
             "5.  Track invoices — and chase the companies that pay slowly.",
             "6.  Check the Dashboard: profit, per hour & a Signing Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+4}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional business (Quill & Seal Notary, agent Sloane) is included so you can see "
               "how it connects \\u2014 type over it with your own. Two things vary by state and you must check yours: "
               "the MAXIMUM FEE you may charge per notarial act, and whether a journal is required and what it must "
               "record. The Notarial Journal tab here is a convenience record, not a substitute for the bound "
               "sequential journal your state may require. Mileage rates change every year. Twelve matching "
               "printable pages (signing worksheet, mileage log, journal sheet, invoice & more) are included. This "
               "is a business & organizing tool, not legal, tax or accounting advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 5):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+6}:B{dr+6}", "Count the drive and the printer. They are the two costs nobody puts in the fee.", "section_gold")


# ===========================================================================
def build_profit(wb):
    ws = wb.create_sheet("Signing Profit"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 48, 20, 2])
    luxe_header(ws, "C", "\U0001f9ee  SIGNING PROFIT — THE ENGINE",
                "Fee minus printing minus driving, over the hours it ACTUALLY takes door to door.")
    ws.cell(row=5, column=2, value="WHAT IT LOOKS LIKE").style = "section_gold"
    ws.cell(row=6, column=2, value="The fee").style = "field_label"
    cf = ws.cell(row=6, column=3, value="=AvgFee"); cf.style = "field_value"; cf.number_format = '"$"#,##0.00'
    ws.cell(row=7, column=2, value="÷ length of the appointment (hours)").style = "field_label"
    ca = ws.cell(row=7, column=3, value=APPT_HOURS); ca.style = "input"; ca.number_format = "0.0"
    cell_name(wb, "ApptHours", "Signing Profit", "$C$7")
    ws.cell(row=8, column=2, value="= WHAT IT FEELS LIKE YOU EARN").style = "th"
    cl = ws.cell(row=8, column=3, value="=IFERROR(AvgFee/ApptHours,0)"); cl.style = "td"
    cl.font = Font(bold=True, size=15, color=PRIMARY); cl.fill = fill(SURFACE); cl.number_format = '"$"#,##0.00'
    cell_name(wb, "LooksLike", "Signing Profit", "$C$8")

    ws.cell(row=10, column=2, value="⚠ NOW COUNT WHAT IT ACTUALLY COST YOU").style = "section_gold"
    ws.cell(row=11, column=2, value="Pages printed (borrower + lender copies)").style = "field_label"
    cp = ws.cell(row=11, column=3, value=PAGES_PER_SIGNING); cp.style = "input"; cp.number_format = "#,##0"
    cell_name(wb, "Pages", "Signing Profit", "$C$11")
    ws.cell(row=12, column=2, value="− Printing (pages × cost per page)").style = "field_label"
    cpc = ws.cell(row=12, column=3, value="=Pages*CostPerPage"); cpc.style = "field_value"
    cpc.number_format = '"$"#,##0.00'; cpc.fill = fill(RED_BG)
    cell_name(wb, "PrintCost", "Signing Profit", "$C$12")
    ws.cell(row=13, column=2, value="Round-trip miles").style = "field_label"
    cm = ws.cell(row=13, column=3, value=MILES_PER_SIGNING); cm.style = "input"; cm.number_format = "#,##0"
    cell_name(wb, "Miles", "Signing Profit", "$C$13")
    ws.cell(row=14, column=2, value="− Driving (miles × vehicle cost per mile)").style = "field_label"
    cdc = ws.cell(row=14, column=3, value="=Miles*VehicleCPM"); cdc.style = "field_value"
    cdc.number_format = '"$"#,##0.00'; cdc.fill = fill(RED_BG)
    cell_name(wb, "DriveCost", "Signing Profit", "$C$14")
    ws.cell(row=15, column=2, value="= OUT OF POCKET, EVERY SIGNING").style = "th"
    cv = ws.cell(row=15, column=3, value="=PrintCost+DriveCost"); cv.style = "td"
    cv.font = Font(bold=True, size=13, color=PRIMARY); cv.fill = fill(WARN_BG); cv.number_format = '"$"#,##0.00'
    cell_name(wb, "CostPerSigning", "Signing Profit", "$C$15")
    ws.cell(row=16, column=2, value="= NET PER SIGNING").style = "th"
    cn = ws.cell(row=16, column=3, value="=AvgFee-CostPerSigning"); cn.style = "td"
    cn.font = Font(bold=True, size=15, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0.00'
    cell_name(wb, "NetPerSigning", "Signing Profit", "$C$16")

    ws.cell(row=18, column=2, value="AND HOW LONG IT REALLY TOOK").style = "section_gold"
    ws.cell(row=19, column=2, value="Prep, printing & confirming (hours)").style = "field_label"
    cpr = ws.cell(row=19, column=3, value=PREP_HOURS); cpr.style = "input"; cpr.number_format = "0.0"
    cell_name(wb, "PrepHours", "Signing Profit", "$C$19")
    ws.cell(row=20, column=2, value="+ Drive time, round trip (hours)").style = "field_label"
    cdr = ws.cell(row=20, column=3, value=DRIVE_HOURS); cdr.style = "input"; cdr.number_format = "0.0"
    cell_name(wb, "DriveHours", "Signing Profit", "$C$20")
    ws.cell(row=21, column=2, value="+ The appointment itself (hours)").style = "field_label"
    cah = ws.cell(row=21, column=3, value="=ApptHours"); cah.style = "field_value"; cah.number_format = "0.0"
    ws.cell(row=22, column=2, value="= HOURS DOOR TO DOOR").style = "th"
    cth = ws.cell(row=22, column=3, value="=PrepHours+DriveHours+ApptHours"); cth.style = "td"
    cth.font = Font(bold=True, size=13, color=PRIMARY); cth.fill = fill(SURFACE); cth.number_format = '0.0" hrs"'
    cell_name(wb, "TotalHours", "Signing Profit", "$C$22")

    ws.cell(row=24, column=2, value="= WHAT YOU ACTUALLY EARN PER HOUR").style = "th"
    cr = ws.cell(row=24, column=3, value="=IFERROR(NetPerSigning/TotalHours,0)"); cr.style = "td"
    cr.font = Font(bold=True, size=17, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0.00'
    cell_name(wb, "RealHourly", "Signing Profit", "$C$24")
    ws.cell(row=25, column=2, value="…which is this much less than it felt like").style = "field_label"
    cg = ws.cell(row=25, column=3, value="=LooksLike-RealHourly"); cg.style = "field_value"
    cg.number_format = '"$"#,##0.00'; cg.fill = fill(RED_BG)

    ws.cell(row=27, column=2, value="THE GOOD NEWS — YOUR MILEAGE DEDUCTION").style = "section_gold"
    ws.cell(row=28, column=2, value="Miles × IRS rate (per signing)").style = "field_label"
    cmd = ws.cell(row=28, column=3, value="=Miles*IRSRate"); cmd.style = "field_value"
    cmd.number_format = '"$"#,##0.00'; cmd.fill = fill(MINT_BG)
    cell_name(wb, "MileageDeduction", "Signing Profit", "$C$28")
    ws.cell(row=29, column=2, value="× signings this month = deductible").style = "field_label"
    cmm = ws.cell(row=29, column=3, value="=MileageDeduction*SigningsMonth"); cmm.style = "field_value"
    cmm.number_format = '"$"#,##0'; cmm.fill = fill(MINT_BG)
    cell_name(wb, "MileageMonth", "Signing Profit", "$C$29")
    ws.cell(row=31, column=2, value="Your mileage deduction is worth more than the gas costs you. Log every mile.").style = "section_gold"


def build_signings(wb):
    ws = wb.create_sheet("Signings Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 11, 22, 22, 13, 11, 11, 15, 14, 14, 2])
    luxe_header(ws, "J", "\U0001f4dd  SIGNINGS LOG",
                "Every signing — the fee, the pages, the miles, and what it actually netted.")
    table_headers(ws, 4, ["Date", "Company", "Type", "Fee", "Pages", "Miles", "Status", "Cost", "Net"], start_col=2)
    start = L0
    for i, (dt, co, typ, fee, pages, miles, status) in enumerate(SIGNINGS):
        r = start + i
        ws.cell(row=r, column=2, value=dt).style = "td"
        ws.cell(row=r, column=3, value=co).style = "td_left"
        ws.cell(row=r, column=4, value=typ).style = "td_left"
        cf = ws.cell(row=r, column=5, value=fee); cf.style = "input"; cf.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=6, value=pages); cp.style = "input"; cp.number_format = "#,##0"
        cm = ws.cell(row=r, column=7, value=miles); cm.style = "input"; cm.number_format = "#,##0"
        ws.cell(row=r, column=8, value=status).style = "td"
        cc = ws.cell(row=r, column=9, value=f"=F{r}*CostPerPage+G{r}*VehicleCPM"); cc.style = "td"; cc.number_format = '"$"#,##0.00'
        cn = ws.cell(row=r, column=10, value=f"=E{r}-I{r}"); cn.style = "td"; cn.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SIGNINGS) - 1
    nrange(wb, "SigFee", "Signings Log", "E", start, end)
    nrange(wb, "SigPages", "Signings Log", "F", start, end)
    nrange(wb, "SigMiles", "Signings Log", "G", start, end)
    nrange(wb, "SigStatus", "Signings Log", "H", start, end)
    nrange(wb, "SigNet", "Signings Log", "J", start, end)
    add_dv(ws, f"D{start}:D{end}", "SignTypeList"); add_dv(ws, f"H{start}:H{end}", "PayStatusList")
    ws.conditional_formatting.add(f"H{start}:H{end}", CellIsRule(operator="equal", formula=['"Overdue"'],
                                                                fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))
    ws.conditional_formatting.add(f"H{start}:H{end}", CellIsRule(operator="containsText", formula=['"fell through"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    ws.conditional_formatting.add(f"J{start}:J{end}", DataBarRule(start_type="min", end_type="max", color=HIGHLIGHT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="FORTNIGHT").style = "th"
    for c in (3, 4, 8):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cf = ws.cell(row=tot, column=5, value="=SUM(SigFee)"); cf.style = "td"
    cf.font = Font(bold=True, color=PRIMARY); cf.fill = fill(SURFACE); cf.number_format = '"$"#,##0'
    cp = ws.cell(row=tot, column=6, value="=SUM(SigPages)"); cp.style = "td"
    cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(SURFACE); cp.number_format = "#,##0"
    cm = ws.cell(row=tot, column=7, value="=SUM(SigMiles)"); cm.style = "td"
    cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = "#,##0"
    ws.cell(row=tot, column=9).style = "td"; ws.cell(row=tot, column=9).fill = fill(SURFACE)
    cn = ws.cell(row=tot, column=10, value="=SUM(SigNet)"); cn.style = "td"
    cn.font = Font(bold=True, size=12, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0.00'
    ws.cell(row=tot + 2, column=2, value="Signings that fell through unpaid").style = "field_label"
    cu = ws.cell(row=tot + 2, column=7, value="=Unpaid"); cu.style = "field_value"
    cu.number_format = "0"; cu.fill = fill(RED_BG)
    ws.cell(row=tot + 3, column=2, value="= GETTING-PAID RATE").style = "th"
    cpr = ws.cell(row=tot + 3, column=7, value="=IFERROR((SigningsMonth-Unpaid)/SigningsMonth,0)"); cpr.style = "td"
    cpr.font = Font(bold=True, size=13, color=PRIMARY); cpr.fill = fill(MINT_BG); cpr.number_format = "0.0%"
    cell_name(wb, "PaidRate", "Signings Log", f"$G${tot+3}")
    ws.cell(row=tot + 5, column=2, value="A signing that falls through at the door is still a printed package and a tank of gas. Bill the trip fee.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_fees(wb):
    ws = wb.create_sheet("Fee Schedule"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 44, 16, 46, 2])
    luxe_header(ws, "D", "\U0001f4b5  FEE SCHEDULE",
                "What you charge, written down — so you stop quoting from memory and undercharging.")
    table_headers(ws, 4, ["Service", "Fee", "Notes"], start_col=2)
    start = L0
    for i, (svc, fee, note) in enumerate(FEES):
        r = start + i
        ws.cell(row=r, column=2, value=svc).style = "td_left"
        cf = ws.cell(row=r, column=3, value=fee); cf.style = "input"; cf.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=4, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FEES) - 1
    nrange(wb, "FeeAmount", "Fee Schedule", "C", start, end)
    tot = end + 2
    ws.cell(row=tot, column=2, value="Highest fee on the schedule").style = "field_label"
    c1 = ws.cell(row=tot, column=3, value="=MAX(FeeAmount)"); c1.style = "field_value"
    c1.number_format = '"$"#,##0.00'; c1.fill = fill(MINT_BG)
    ws.cell(row=tot + 1, column=2, value="Your average signing fee").style = "field_label"
    c2 = ws.cell(row=tot + 1, column=3, value="=AvgFee"); c2.style = "field_value"; c2.number_format = '"$"#,##0.00'
    ws.cell(row=tot + 3, column=2, value="⚠  Check your state's MAXIMUM fee per notarial act. Travel and printing are usually separate — but confirm.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_mileage(wb):
    ws = wb.create_sheet("Mileage Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 46, 14, 28, 16, 2])
    luxe_header(ws, "F", "\U0001f697  MILEAGE LOG",
                "Every mile, logged — because the deduction is worth more than the gas costs you.")
    table_headers(ws, 4, ["Date", "Trip", "Miles", "Purpose", "Deduction"], start_col=2)
    start = L0
    for i, (dt, trip, miles, purpose) in enumerate(MILEAGE):
        r = start + i
        ws.cell(row=r, column=2, value=dt).style = "td"
        ws.cell(row=r, column=3, value=trip).style = "td_left"
        cm = ws.cell(row=r, column=4, value=miles); cm.style = "input"; cm.number_format = "#,##0"
        ws.cell(row=r, column=5, value=purpose).style = "td_left"
        cd = ws.cell(row=r, column=6, value=f"=D{r}*IRSRate"); cd.style = "td"; cd.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MILEAGE) - 1
    nrange(wb, "MileMiles", "Mileage Log", "D", start, end)
    nrange(wb, "MileDeduct", "Mileage Log", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="LOGGED").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=4, value="=SUM(MileMiles)"); cm.style = "td"
    cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = "#,##0"
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cd = ws.cell(row=tot, column=6, value="=SUM(MileDeduct)"); cd.style = "td"
    cd.font = Font(bold=True, size=12, color=PRIMARY); cd.fill = fill(MINT_BG); cd.number_format = '"$"#,##0.00'
    ws.cell(row=tot + 2, column=2, value="Projected for the full month").style = "field_label"
    cp = ws.cell(row=tot + 2, column=6, value="=MileageMonth"); cp.style = "field_value"
    cp.number_format = '"$"#,##0'; cp.fill = fill(MINT_BG)
    ws.cell(row=tot + 4, column=2, value="⚠  The IRS rate changes every year. Update it in Settings each January.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_printing(wb):
    ws = wb.create_sheet("Printing & Supplies"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 16, 16, 16, 18, 2])
    luxe_header(ws, "F", "\U0001f5a8  PRINTING & SUPPLIES",
                "What a page actually costs you — the number that makes 180-page packages sting.")
    table_headers(ws, 4, ["Item", "Cost", "Yields", "Unit", "Cost per unit"], start_col=2)
    start = L0
    for i, (item, cost, yields, unit) in enumerate(PRINTING):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cy = ws.cell(row=r, column=4, value=yields); cy.style = "input"; cy.number_format = "#,##0"
        ws.cell(row=r, column=5, value=unit).style = "td"
        cu = ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/D{r},0)"); cu.style = "td"; cu.number_format = '"$"#,##0.0000'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRINTING) - 1
    nrange(wb, "SupplyCost", "Printing & Supplies", "C", start, end)
    tot = end + 2
    ws.cell(row=tot, column=2, value="Your cost per printed page (Settings)").style = "field_label"
    c1 = ws.cell(row=tot, column=6, value="=CostPerPage"); c1.style = "field_value"; c1.number_format = '"$"#,##0.000'
    ws.cell(row=tot + 1, column=2, value="× pages per signing").style = "field_label"
    c2 = ws.cell(row=tot + 1, column=6, value="=Pages"); c2.style = "field_value"; c2.number_format = "#,##0"
    ws.cell(row=tot + 2, column=2, value="= PRINTING PER SIGNING").style = "th"
    c3 = ws.cell(row=tot + 2, column=6, value="=PrintCost"); c3.style = "td"
    c3.font = Font(bold=True, size=13, color=PRIMARY); c3.fill = fill(WARN_BG); c3.number_format = '"$"#,##0.00'
    ws.cell(row=tot + 3, column=2, value="× signings this month").style = "th"
    c4 = ws.cell(row=tot + 3, column=6, value="=PrintCost*SigningsMonth"); c4.style = "td"
    c4.font = Font(bold=True, size=13, color=PRIMARY); c4.fill = fill(RED_BG); c4.number_format = '"$"#,##0'
    cell_name(wb, "PrintMonth", "Printing & Supplies", f"$F${tot+3}")
    ws.cell(row=tot + 5, column=2, value="Charge a printing surcharge over 150 pages. Every established agent does.").style = "section_gold"


def build_invoices(wb):
    ws = wb.create_sheet("Invoices"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 18, 18, 18, 2])
    luxe_header(ws, "F", "\U0001f4b3  INVOICES & PAYMENTS",
                "Who owes you, how much, and how long they've had it.")
    table_headers(ws, 4, ["Company", "Signings", "Amount", "Status", "Outstanding"], start_col=2)
    start = L0
    for i, (co, n, amt, status) in enumerate(INVOICES):
        r = start + i
        ws.cell(row=r, column=2, value=co).style = "td_left"
        cn = ws.cell(row=r, column=3, value=n); cn.style = "input"; cn.number_format = "0"
        ca = ws.cell(row=r, column=4, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=status).style = "input"
        co_ = ws.cell(row=r, column=6, value=f'=IF(E{r}="Paid",0,D{r})'); co_.style = "td"; co_.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVOICES) - 1
    nrange(wb, "InvCount", "Invoices", "C", start, end)
    nrange(wb, "InvAmount", "Invoices", "D", start, end)
    nrange(wb, "InvStatus", "Invoices", "E", start, end)
    nrange(wb, "InvOutstanding", "Invoices", "F", start, end)
    add_dv(ws, f"E{start}:E{end}", "PayStatusList")
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="greaterThan", formula=["0"],
                                                                fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL COMPANIES").style = "th"
    cn = ws.cell(row=tot, column=3, value="=SUM(InvCount)"); cn.style = "td"
    cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = "0"
    ca = ws.cell(row=tot, column=4, value="=SUM(InvAmount)"); ca.style = "td"
    ca.font = Font(bold=True, size=12, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = '"$"#,##0'
    cell_name(wb, "Invoiced", "Invoices", f"$D${tot}")
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    co_ = ws.cell(row=tot, column=6, value="=SUM(InvOutstanding)"); co_.style = "td"
    co_.font = Font(bold=True, size=12, color=ACCENT); co_.fill = fill(WARN_BG); co_.number_format = '"$"#,##0'
    cell_name(wb, "Receivable", "Invoices", f"$F${tot}")
    ws.cell(row=tot + 2, column=2, value="Net 45 is a loan you gave them. Price it in, or stop taking their work.").style = "section_gold"


def build_companies(wb):
    ws, start, end = build_log(
        wb, "Signing Companies", "\U0001f3e2", "SIGNING COMPANIES", "Who pays well, who pays late, and who is worth saying yes to.",
        ["Company", "Contact", "Pay terms", "Avg fee", "Notes"], COMPANIES,
        [2, 24, 30, 14, 14, 34, 2], text_left={2, 3, 6}, money={5}, start_col=2)
    nrange(wb, "CoAvgFee", "Signing Companies", "E", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="Best average fee").style = "field_label"
    c1 = ws.cell(row=tr, column=5, value="=MAX(CoAvgFee)"); c1.style = "field_value"
    c1.number_format = '"$"#,##0'; c1.fill = fill(MINT_BG)
    ws.cell(row=tr + 1, column=2, value="Worst average fee").style = "field_label"
    c2 = ws.cell(row=tr + 1, column=5, value="=MIN(CoAvgFee)"); c2.style = "field_value"
    c2.number_format = '"$"#,##0'; c2.fill = fill(WARN_BG)
    ws.cell(row=tr + 3, column=2, value="Direct clients pay best and pay fastest. Every hour spent building them is worth two on a platform.").style = "section_gold"


def build_journal(wb):
    ws = wb.create_sheet("Notarial Journal"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 30, 22, 22, 12, 16, 2])
    luxe_header(ws, "G", "\U0001f4d5  NOTARIAL JOURNAL",
                "A convenience record of every notarial act — check what your state requires.")
    ws.cell(row=4, column=2, value="⚠  Many states require a BOUND, SEQUENTIAL journal. This is a convenience copy, not a substitute — check your state.").style = "section_gold"
    table_headers(ws, 5, ["Date", "Document type", "Signer", "ID presented", "Fee", "Thumbprint"], start_col=2)
    start = 6
    for i, (dt, doc, signer, idt, fee, thumb) in enumerate(JOURNAL):
        r = start + i
        ws.cell(row=r, column=2, value=dt).style = "td"
        ws.cell(row=r, column=3, value=doc).style = "td_left"
        ws.cell(row=r, column=4, value=signer).style = "td_left"
        ws.cell(row=r, column=5, value=idt).style = "input"
        cf = ws.cell(row=r, column=6, value=fee); cf.style = "input"; cf.number_format = '"$"#,##0'
        ws.cell(row=r, column=7, value=thumb).style = "input"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(JOURNAL) - 1
    nrange(wb, "JournalDate", "Notarial Journal", "B", start, end)
    nrange(wb, "JournalFee", "Notarial Journal", "F", start, end)
    add_dv(ws, f"E{start}:E{end}", "IDTypeList"); add_dv(ws, f"G{start}:G{end}", "YesNoList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="ACTS RECORDED").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cn = ws.cell(row=tot, column=5, value="=COUNTA(JournalDate)"); cn.style = "td"
    cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = "0"
    cell_name(wb, "ActsRecorded", "Notarial Journal", f"$E${tot}")
    cf = ws.cell(row=tot, column=6, value="=SUM(JournalFee)"); cf.style = "td"
    cf.font = Font(bold=True, size=12, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = '"$"#,##0'
    ws.cell(row=tot + 2, column=2, value="Your journal is the only thing that protects you if a signing is ever questioned. Fill it in at the table, not later.").style = "section_gold"
    ws.freeze_panes = "A6"


def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 42, 20, 16, 16, 2])
    luxe_header(ws, "E", "\U0001f9fe  EXPENSES",
                "Fixed costs that run whether you take a signing or not, plus everything else.")
    ws.cell(row=4, column=2, value="FIXED — RUNS WHETHER YOU WORK OR NOT").style = "section_gold"
    table_headers(ws, 5, ["Line", "Category", "Monthly", "Yearly"], start_col=2)
    fs_ = 6
    for i, (lab, cat, amt) in enumerate(FIXED_LINES):
        r = fs_ + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        c = ws.cell(row=r, column=4, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        cy = ws.cell(row=r, column=5, value=f"=D{r}*12"); cy.style = "td"; cy.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 6):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    fe = fs_ + len(FIXED_LINES) - 1
    nrange(wb, "FixedLines", "Expenses", "D", fs_, fe)
    tot = fe + 1
    ws.cell(row=tot, column=2, value="= FIXED COSTS / MONTH").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=4, value="=SUM(FixedLines)"); ct.style = "td"
    ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "FixedTotal", "Expenses", f"$D${tot}")
    cty = ws.cell(row=tot, column=5, value="=FixedTotal*12"); cty.style = "td"
    cty.font = Font(bold=True, color=PRIMARY); cty.fill = fill(MINT_BG); cty.number_format = '"$"#,##0'

    o = tot + 2
    ws.cell(row=o, column=2, value="OTHER RUNNING COSTS").style = "section_gold"
    table_headers(ws, o + 1, ["Line", "Category", "Monthly", "Yearly"], start_col=2)
    os_ = o + 2
    for i, (lab, cat, amt) in enumerate(EXPENSES):
        r = os_ + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        c = ws.cell(row=r, column=4, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        cy = ws.cell(row=r, column=5, value=f"=D{r}*12"); cy.style = "td"; cy.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 6):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    oe = os_ + len(EXPENSES) - 1
    nrange(wb, "OtherLines", "Expenses", "D", os_, oe)
    add_dv(ws, f"C{fs_}:C{oe}", "ExpCatList")
    ot = oe + 1
    ws.cell(row=ot, column=2, value="= OTHER / MONTH").style = "th"
    ws.cell(row=ot, column=3).style = "td"; ws.cell(row=ot, column=3).fill = fill(SURFACE)
    co = ws.cell(row=ot, column=4, value="=SUM(OtherLines)"); co.style = "td"
    co.font = Font(bold=True, size=12, color=PRIMARY); co.fill = fill(SURFACE); co.number_format = '"$"#,##0'
    cell_name(wb, "OtherTotal", "Expenses", f"$D${ot}")
    ws.cell(row=ot + 2, column=2, value="Your fixed costs are tiny. Three signings a month covers them — everything after that is yours.").style = "section_gold"


def build_tax(wb):
    ws = wb.create_sheet("Tax Set-Aside"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 18, 18, 18, 18, 2])
    luxe_header(ws, "F", "\U0001f3e6  TAX SET-ASIDE",
                "Nobody withholds anything from a 1099. That is entirely on you.")
    table_headers(ws, 4, ["Quarter", "Income", "Set aside", "Estimated due", "Short by"], start_col=2)
    start = L0
    for i, (q, inc, saved, due) in enumerate(TAXES):
        r = start + i
        ws.cell(row=r, column=2, value=q).style = "td_left"
        ci = ws.cell(row=r, column=3, value=inc); ci.style = "input"; ci.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=5, value=due); cd.style = "input"; cd.number_format = '"$"#,##0'
        csh = ws.cell(row=r, column=6, value=f"=MAX(E{r}-D{r},0)"); csh.style = "td"; csh.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(TAXES) - 1
    nrange(wb, "TaxIncome", "Tax Set-Aside", "C", start, end)
    nrange(wb, "TaxSaved", "Tax Set-Aside", "D", start, end)
    nrange(wb, "TaxDue", "Tax Set-Aside", "E", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="greaterThan", formula=["0"],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="THE YEAR").style = "th"
    ci = ws.cell(row=tot, column=3, value="=SUM(TaxIncome)"); ci.style = "td"
    ci.font = Font(bold=True, color=PRIMARY); ci.fill = fill(SURFACE); ci.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value="=SUM(TaxSaved)"); cs.style = "td"
    cs.font = Font(bold=True, size=12, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cell_name(wb, "TaxReserve", "Tax Set-Aside", f"$D${tot}")
    cd = ws.cell(row=tot, column=5, value="=SUM(TaxDue)"); cd.style = "td"
    cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = '"$"#,##0'
    csh = ws.cell(row=tot, column=6, value=f"=MAX(E{tot}-D{tot},0)"); csh.style = "td"
    csh.font = Font(bold=True, size=12, color=DANGER); csh.fill = fill(RED_BG); csh.number_format = '"$"#,##0'
    r = tot + 2
    ws.cell(row=r, column=2, value="Don't forget your mileage deduction").style = "field_label"
    cm = ws.cell(row=r, column=4, value="=MileageMonth*12"); cm.style = "field_value"
    cm.number_format = '"$"#,##0'; cm.fill = fill(MINT_BG)
    ws.cell(row=r + 1, column=2, value="Reserve funded against goal").style = "field_label"
    cf = ws.cell(row=r + 1, column=4, value="=IFERROR(TaxReserve/TaxReserveGoal,0)"); cf.style = "field_value"
    cf.number_format = "0%"; cf.fill = fill(WARN_BG)
    ws.cell(row=r + 3, column=2, value="Self-employment tax is the bill that surprises every new signing agent. Move it the week it lands.").style = "section_gold"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 40, 18, 2])
    luxe_header(ws, "C", "\U0001f4c8  MONTHLY SUMMARY",
                "Signings in, costs out, and what is actually yours at the end of it.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    ws.cell(row=6, column=2, value="Signings").style = "field_label"
    c0 = ws.cell(row=6, column=3, value="=SigningsMonth"); c0.style = "field_value"; c0.number_format = "0"
    ws.cell(row=7, column=2, value="× average fee = REVENUE").style = "th"
    c1 = ws.cell(row=7, column=3, value="=SigningsMonth*AvgFee"); c1.style = "td"
    c1.font = Font(bold=True, size=14, color=PRIMARY); c1.fill = fill(SURFACE); c1.number_format = '"$"#,##0'
    cell_name(wb, "Revenue", "Monthly Summary", "$C$7")
    ws.cell(row=8, column=2, value="− Printing & driving (per signing × signings)").style = "field_label"
    c2 = ws.cell(row=8, column=3, value="=CostPerSigning*SigningsMonth"); c2.style = "field_value"
    c2.number_format = '"$"#,##0.00'; c2.fill = fill(WARN_BG)
    cell_name(wb, "VariableCosts", "Monthly Summary", "$C$8")
    ws.cell(row=9, column=2, value="− Fixed costs").style = "field_label"
    c3 = ws.cell(row=9, column=3, value="=FixedTotal"); c3.style = "field_value"; c3.number_format = '"$"#,##0'
    ws.cell(row=10, column=2, value="= PROFIT").style = "th"
    c4 = ws.cell(row=10, column=3, value="=Revenue-VariableCosts-FixedTotal"); c4.style = "td"
    c4.font = Font(bold=True, size=16, color=PRIMARY); c4.fill = fill(MINT_BG); c4.number_format = '"$"#,##0'
    cell_name(wb, "Profit", "Monthly Summary", "$C$10")
    ws.cell(row=11, column=2, value="= MARGIN").style = "th"
    c5 = ws.cell(row=11, column=3, value="=IFERROR(Profit/Revenue,0)"); c5.style = "td"
    c5.font = Font(bold=True, size=13, color=PRIMARY); c5.fill = fill(MINT_BG); c5.number_format = "0.0%"
    cell_name(wb, "Margin", "Monthly Summary", "$C$11")
    ws.cell(row=12, column=2, value="= RUN-RATE YEAR").style = "th"
    c6 = ws.cell(row=12, column=3, value="=Profit*12"); c6.style = "td"
    c6.font = Font(bold=True, size=13, color=PRIMARY); c6.fill = fill(SURFACE); c6.number_format = '"$"#,##0'
    cell_name(wb, "RunRate", "Monthly Summary", "$C$12")
    ws.cell(row=14, column=2, value="BREAK-EVEN").style = "section_gold"
    ws.cell(row=15, column=2, value="Fixed costs ÷ net per signing").style = "field_label"
    c7 = ws.cell(row=15, column=3, value="=IFERROR(ROUNDUP(FixedTotal/NetPerSigning,0),0)"); c7.style = "td"
    c7.font = Font(bold=True, size=15, color=PRIMARY); c7.fill = fill(SURFACE); c7.number_format = '0" signings"'
    cell_name(wb, "BreakEven", "Monthly Summary", "$C$15")
    ws.cell(row=16, column=2, value="= YOU COVER IT THIS MANY TIMES OVER").style = "th"
    c8 = ws.cell(row=16, column=3, value="=IFERROR(SigningsMonth/BreakEven,0)"); c8.style = "td"
    c8.font = Font(bold=True, size=13, color=PRIMARY); c8.fill = fill(MINT_BG); c8.number_format = '0.0"×"'
    cell_name(wb, "CoverRatio", "Monthly Summary", "$C$16")
    ws.cell(row=18, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 19, ["Month", "Revenue"], start_col=2)
    ts = 20
    for i, (m, rev) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        crv = ws.cell(row=r, column=3, value=rev); crv.style = "input"; crv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "RevTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Revenue by Month", ts, te, 3, 2), "E5")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  \U0001f58b  NOTARY & LOAN SIGNING AGENT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  What a signing really pays, per hour — plus a Signing Score, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("AVERAGE FEE", "=AvgFee", "money2"),
        ("LOOKS LIKE / HOUR", "=LooksLike", "money2"),
        ("PRINTING", "=PrintCost", "money2"),
        ("DRIVING", "=DriveCost", "money2"),
        ("NET PER SIGNING", "=NetPerSigning", "money2"),
        ("REALLY EARNS / HOUR", "=RealHourly", "money2"),
    ]
    row2 = [
        ("SIGNINGS / MONTH", "=SigningsMonth", "num"),
        ("MONTHLY REVENUE", "=Revenue", "money"),
        ("MONTHLY PROFIT", "=Profit", "money"),
        ("BREAK-EVEN SIGNINGS", "=BreakEven", "num"),
        ("MILEAGE DEDUCTION", "=MileageMonth", "money"),
        ("SIGNING SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "BUSINESS HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Real hourly on target", "=IFERROR(MIN(RealHourly/HourlyGoal,1),0)"),
        ("Margin healthy", "=IFERROR(MIN(Margin/MarginGoal,1),0)"),
        ("Overhead more than covered", "=IFERROR(MIN(CoverRatio/CoverGoal,1),0)"),
        ("Actually getting paid", "=IFERROR(MIN(PaidRate/PaidGoal,1),0)"),
        ("Enough signings booked", "=IFERROR(MIN(SigningsMonth/SigningGoal,1),0)"),
        ("Tax reserve funded", "=IFERROR(MIN(TaxReserve/TaxReserveGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=20, max_row=19 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=20, max_row=19 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Notary & Loan Signing Agent Command Center™ — count the drive and the printer.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profit(wb); build_signings(wb)
    build_fees(wb); build_mileage(wb); build_printing(wb); build_invoices(wb)
    build_companies(wb); build_journal(wb); build_expenses(wb); build_tax(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Signing Profit", "Signings Log", "Fee Schedule", "Mileage Log",
             "Printing & Supplies", "Invoices", "Signing Companies", "Notarial Journal", "Expenses",
             "Tax Set-Aside", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Notary_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
