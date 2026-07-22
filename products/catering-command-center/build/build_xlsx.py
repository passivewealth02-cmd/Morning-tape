"""Build Catering Command Center™ — The Complete Catering Business System.

14 tabs · a premium catering operating system in Google Sheets & Excel. Dashboard,
a per-head plate-costing engine, a menu-package price list, an event quote & P&L
builder, staffing & labor, rentals, a bookings calendar, inventory, waste,
ordering, cash & deposits and a client CRM — one dashboard. Cost every head, quote
with confidence & book more profit.

Run: python3 build_xlsx.py   ->  ../Catering_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["each", "lb", "oz", "case", "box", "dozen", "gallon", "tray"]
PACKAGES_LIST = ["Plated Dinner", "Buffet Classic", "Cocktail Reception", "Boxed Lunch",
                 "Wedding Premium", "Breakfast Spread", "BBQ Buffet", "Grazing Table"]
STATUS = ["Confirmed", "Deposit", "Inquiry", "Recurring"]

TARGET_FC = 0.30
MARGIN_GOAL = 0.30
LABOR_LIMIT = 0.40
BOOKING_GOAL = 6

# Plate-costing engine — flagship Plated Dinner, cost per head: (component, cost/head)
PLATE = [
    ("Beef tenderloin (6 oz)", 6.50),
    ("Starch + seasonal veg", 2.20),
    ("Salad + artisan bread", 1.40),
    ("Plated dessert", 1.80),
    ("Disposables / rentals", 1.20),
    ("Kitchen labor allocation", 0.90),
]

# Menu packages: (package, cost/head, price/head). Plated Dinner = PlatedCostHead.
PACKAGES = [
    ("Plated Dinner", None, 48.00),
    ("Buffet Classic", 9.50, 32.00),
    ("Cocktail Reception", 8.00, 28.00),
    ("Boxed Lunch", 6.50, 18.00),
    ("Wedding Premium", 22.00, 75.00),
    ("Breakfast Spread", 5.50, 16.00),
    ("BBQ Buffet", 11.00, 34.00),
    ("Grazing Table", 7.50, 26.00),
]

# Event quotes / P&L: (event, package, guests, service fee, staff cost, rentals)
EVENTS = [
    ("Corporate Gala", "Plated Dinner", 120, 600, 1400, 900),
    ("Garden Wedding", "Wedding Premium", 90, 800, 1800, 1600),
    ("Birthday Buffet", "Buffet Classic", 60, 250, 600, 300),
    ("Office Lunch", "Boxed Lunch", 40, 120, 200, 0),
    ("Cocktail Mixer", "Cocktail Reception", 80, 400, 900, 500),
    ("Backyard BBQ", "BBQ Buffet", 50, 300, 500, 400),
]

# Staffing & labor rate card: (role, rate/hr, typical hrs/event)
STAFFING = [
    ("Event Lead", 28, 8), ("Chef", 32, 8), ("Prep Cook", 19, 6), ("Server", 18, 6),
    ("Bartender", 22, 6), ("Dishwasher", 16, 5), ("Driver / Setup", 18, 4),
]

# Rentals & equipment: (item, qty, unit cost)
RENTALS = [
    ("Round tables (60\")", 10, 12.00), ("Chiavari chairs", 100, 3.50), ("Linens", 40, 6.00),
    ("Chafing dishes", 12, 9.00), ("Glassware sets", 8, 14.00), ("Serving platters", 20, 4.50),
]

# Bookings calendar: (event, days out, guests, deposit, status)
BOOKINGS = [
    ("Corporate Gala", 12, 120, 1500, "Confirmed"),
    ("Garden Wedding", 26, 90, 2000, "Confirmed"),
    ("Birthday Buffet", 6, 60, 500, "Confirmed"),
    ("Office Lunch (weekly)", 3, 40, 0, "Recurring"),
    ("Cocktail Mixer", 18, 80, 800, "Deposit"),
    ("Backyard BBQ", 9, 50, 400, "Confirmed"),
]

# Inventory & par: (item, par, on hand, unit)
INVENTORY = [
    ("Beef tenderloin", 40, 18, "lb"), ("Chicken breast", 60, 25, "lb"), ("Seasonal produce", 30, 12, "box"),
    ("Artisan bread", 24, 10, "dozen"), ("Dairy & eggs", 20, 8, "case"), ("Dessert base", 15, 6, "tray"),
    ("Disposables", 30, 12, "case"), ("Beverages", 25, 10, "case"),
]

# Waste log: (item, reason, cost) — total $430
WASTE = [
    ("Over-prep / leftovers", "Batch surplus", 180.00),
    ("Spoilage", "Cold-chain break", 110.00),
    ("Broken rentals", "Handling", 85.00),
    ("Delivery loss", "In transit", 55.00),
]

# Ordering / suppliers: (item, supplier, par order, cost)
ORDERING = [
    ("Proteins (case)", "Restaurant Depot", 1, 480.00),
    ("Produce (box)", "Green Farms", 2, 220.00),
    ("Dairy & eggs", "Local Dairy", 1, 140.00),
    ("Disposables", "PartySupply", 1, 160.00),
    ("Beverages", "BevCo", 1, 190.00),
]

# Cash & deposits: (client / event, amount, method)
CASHDEP = [
    ("Corporate Gala — deposit", 1500, "Card"),
    ("Garden Wedding — deposit", 2000, "Transfer"),
    ("Birthday Buffet — deposit", 500, "Cash"),
    ("Cocktail Mixer — deposit", 800, "Card"),
    ("Backyard BBQ — deposit", 400, "Card"),
]

# Clients / CRM: (client, event type, contact, status)
CLIENTS = [
    ("Meridian Corp", "Corporate", "events@meridian.co", "Repeat"),
    ("The Hale Family", "Wedding", "hale@mail.com", "New"),
    ("Sofia R.", "Birthday", "sofia@mail.com", "New"),
    ("Downtown Law", "Office Lunch", "office@dtlaw.com", "Repeat"),
    ("Craft Collective", "Mixer", "hi@craftco.com", "New"),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your company, targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR COMPANY", "section")
    controls = [
        ("Company name", "Wildflower & Oak Catering Co.", None, "Company"),
        ("Owner", "Camille", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Event margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Labor limit %", LABOR_LIMIT, "0%", "LaborLimit"),
        ("Booking goal / month", BOOKING_GOAL, "#,##0", "BookingGoal"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Package", PACKAGES_LIST, "PackageList"),
             ("G", "Status", STATUS, "StatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🍽  CATERING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every head, quote with confidence & book more profit.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE CATERING BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Cost each plate by the head, then price menu packages with a healthy margin. Build an event quote "
                      "that becomes a full event P&L — guests × price per head, plus service, staff and rentals — so you "
                      "know the profit on every booking before you say yes. Track staffing, rentals, a bookings calendar, "
                      "inventory, waste, ordering, deposits and a client list — all in ONE premium Google Sheets & Excel "
                      "system built for caterers.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your company name & target food-cost %.",
             "2.  Cost a plate by the head; Menu Packages show margin per head.",
             "3.  Build an Event Quote — guests × package = a full event P&L.",
             "4.  Add staffing & rentals; each event's profit calculates live.",
             "5.  Track Bookings, deposits, inventory, waste & ordering.",
             "6.  Check the Dashboard: revenue, food cost, margin & a Catering Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional caterer (Wildflower & Oak) is included so you can see how it all connects — "
               "just type over it with your own packages and events. Cost per head and event margin are the two numbers "
               "that decide whether a catering job makes money, and they roll into a live Catering Score. Twelve matching "
               "printable pages (plate cost card, quote sheet, event order & run sheet, staffing sheet & more) are "
               "included. This is a business tool, not financial or legal advice — confirm figures with your own books.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Quote from real cost per head — never guess a per-plate price again.", "section_gold")


# ===========================================================================
def build_plate(wb):
    ws = wb.create_sheet("Plate Costing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 2])
    luxe_header(ws, "C", "🍴  PLATE COSTING",
                "Cost a plate by the head — protein, sides, dessert & overhead. The engine behind your packages.")
    ws.cell(row=5, column=2, value="PACKAGE").style = "section_gold"
    ws.cell(row=5, column=3, value="Plated Dinner").font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Component", "Cost / Head"], start_col=2)
    start = 7
    for i, (comp, cost) in enumerate(PLATE):
        r = start + i
        ws.cell(row=r, column=2, value=comp).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PLATE) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="COST PER HEAD").style = "th"
    ct = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); ct.style = "td"
    ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0.00'
    cell_name(wb, "PlatedCostHead", "Plate Costing", f"$C${tot}")
    ws.cell(row=tot + 2, column=2, value="Copy this build for every package — swap the protein, sides & dessert.").style = "section"


def build_packages(wb):
    ws = wb.create_sheet("Menu Packages"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 14, 12, 2])
    luxe_header(ws, "F", "📋  MENU PACKAGES",
                "Cost per head, price per head, margin & food-cost % on every package — price for profit.")
    table_headers(ws, 4, ["Package", "Cost/Head", "Price/Head", "Margin/Head", "Food %"], start_col=2)
    start = L0
    for i, (pkg, cost, price) in enumerate(PACKAGES):
        r = start + i
        ws.cell(row=r, column=2, value=pkg).style = "td_left"
        cc = ws.cell(row=r, column=3, value="=PlatedCostHead" if cost is None else cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=5, value=f"=D{r}-C{r}"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.number_format = '"$"#,##0.00'
        cf = ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/D{r},0)"); cf.style = "td"; cf.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PACKAGES) - 1
    nrange(wb, "PkgItem", "Menu Packages", "B", start, end)
    nrange(wb, "PkgCostHead", "Menu Packages", "C", start, end)
    nrange(wb, "PkgPriceHead", "Menu Packages", "D", start, end)
    nrange(wb, "PkgMargin", "Menu Packages", "E", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        ColorScaleRule(start_type="num", start_value=0.15, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.30, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.45, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


def build_quotes(wb):
    ws = wb.create_sheet("Event Quotes"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 20, 10, 12, 12, 12, 11, 11, 13, 13, 11, 2])
    luxe_header(ws, "L", "🧾  EVENT QUOTES & P&L",
                "Guests × package price, plus service, staff & rentals — a full profit picture on every event.")
    table_headers(ws, 4, ["Event", "Package", "Guests", "Price/Hd", "Cost/Hd", "Service",
                          "Staff", "Rentals", "Revenue", "Food Cost", "Margin $", "Margin %"], start_col=2)
    start = L0
    for i, (name, pkg, guests, fee, staff, rentals) in enumerate(EVENTS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=pkg).style = "td_left"
        cg = ws.cell(row=r, column=4, value=guests); cg.style = "input"; cg.number_format = "#,##0"
        cph = ws.cell(row=r, column=5, value=f"=IFERROR(INDEX(PkgPriceHead,MATCH(C{r},PkgItem,0)),0)"); cph.style = "td"; cph.number_format = '"$"#,##0.00'
        cch = ws.cell(row=r, column=6, value=f"=IFERROR(INDEX(PkgCostHead,MATCH(C{r},PkgItem,0)),0)"); cch.style = "td"; cch.number_format = '"$"#,##0.00'
        cf = ws.cell(row=r, column=7, value=fee); cf.style = "input"; cf.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=8, value=staff); cs.style = "input"; cs.number_format = '"$"#,##0'
        crn = ws.cell(row=r, column=9, value=rentals); crn.style = "input"; crn.number_format = '"$"#,##0'
        crev = ws.cell(row=r, column=10, value=f"=D{r}*E{r}+G{r}"); crev.style = "td"; crev.font = Font(bold=True, color=PRIMARY); crev.number_format = '"$"#,##0'
        cfood = ws.cell(row=r, column=11, value=f"=D{r}*F{r}"); cfood.style = "td"; cfood.number_format = '"$"#,##0'
        cmg = ws.cell(row=r, column=12, value=f"=J{r}-K{r}-H{r}-I{r}"); cmg.style = "td"; cmg.number_format = '"$"#,##0'
        cmp = ws.cell(row=r, column=13, value=f"=IFERROR(L{r}/J{r},0)"); cmp.style = "td"; cmp.font = Font(bold=True, color=PRIMARY); cmp.number_format = "0%"
        if i % 2:
            for c in range(2, 14):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "PackageList")
    end = start + len(EVENTS) - 1
    nrange(wb, "EventName", "Event Quotes", "B", start, end)
    nrange(wb, "EventPkg", "Event Quotes", "C", start, end)
    nrange(wb, "EventGuests", "Event Quotes", "D", start, end)
    nrange(wb, "EventStaff", "Event Quotes", "H", start, end)
    nrange(wb, "EventRev", "Event Quotes", "J", start, end)
    nrange(wb, "EventFood", "Event Quotes", "K", start, end)
    nrange(wb, "EventMarginPct", "Event Quotes", "M", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTALS").style = "th"
    for c in range(3, 9):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    crev = ws.cell(row=tot, column=10, value="=SUM(EventRev)"); crev.style = "td"; crev.font = Font(bold=True, color=PRIMARY); crev.fill = fill(SURFACE); crev.number_format = '"$"#,##0'
    cfood = ws.cell(row=tot, column=11, value="=SUM(EventFood)"); cfood.style = "td"; cfood.font = Font(bold=True, color=PRIMARY); cfood.fill = fill(SURFACE); cfood.number_format = '"$"#,##0'
    cmg = ws.cell(row=tot, column=12, value=f"=J{tot}-K{tot}-SUM(EventStaff)-SUM(I{start}:I{end})"); cmg.style = "td"; cmg.font = Font(bold=True, color=PRIMARY); cmg.fill = fill(SURFACE); cmg.number_format = '"$"#,##0'
    cmp = ws.cell(row=tot, column=13, value=f"=IFERROR(L{tot}/J{tot},0)"); cmp.style = "td"; cmp.font = Font(bold=True, color=PRIMARY); cmp.fill = fill(SURFACE); cmp.number_format = "0%"
    cell_name(wb, "TotalRevenue", "Event Quotes", f"$J${tot}")
    cell_name(wb, "TotalFood", "Event Quotes", f"$K${tot}")
    # named summary cells
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Food cost % of revenue").style = "field_label"
    c = ws.cell(row=sr, column=6, value="=IFERROR(TotalFood/TotalRevenue,0)"); c.style = "field_value"; c.number_format = "0%"; c.fill = fill(MINT_BG)
    cell_name(wb, "FoodCostPct", "Event Quotes", f"$F${sr}")
    ws.cell(row=sr + 1, column=2, value="Avg event margin %").style = "field_label"
    c2 = ws.cell(row=sr + 1, column=6, value="=IFERROR(AVERAGE(EventMarginPct),0)"); c2.style = "field_value"; c2.number_format = "0%"; c2.fill = fill(MINT_BG)
    cell_name(wb, "AvgEventMargin", "Event Quotes", f"$F${sr+1}")
    ws.cell(row=sr + 2, column=2, value="Labor % of revenue").style = "field_label"
    c3 = ws.cell(row=sr + 2, column=6, value="=IFERROR(SUM(EventStaff)/TotalRevenue,0)"); c3.style = "field_value"; c3.number_format = "0%"; c3.fill = fill(MINT_BG)
    cell_name(wb, "LaborPct", "Event Quotes", f"$F${sr+2}")
    ws.freeze_panes = "A5"


def build_staffing(wb):
    ws = wb.create_sheet("Staffing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 14, 2])
    luxe_header(ws, "E", "👥  STAFFING & LABOR",
                "Your crew rate card — role, hourly rate & typical hours, so event staff costs are never a guess.")
    table_headers(ws, 4, ["Role", "Rate / hr", "Hrs / event", "Cost / event"], start_col=2)
    start = L0
    for i, (role, rate, hrs) in enumerate(STAFFING):
        r = start + i
        ws.cell(row=r, column=2, value=role).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rate); cr.style = "input"; cr.number_format = '"$"#,##0'
        ch = ws.cell(row=r, column=4, value=hrs); ch.style = "input"; ch.number_format = "#,##0"
        cc = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(STAFFING) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="FULL CREW / EVENT").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot + 2, column=2, value="Pull the roles you need into each Event Quote's Staff column.").style = "section"
    ws.freeze_panes = "A5"


def build_rentals(wb):
    ws = wb.create_sheet("Rentals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 14, 14, 2])
    luxe_header(ws, "E", "🪑  RENTALS & EQUIPMENT",
                "Tables, chairs, linens & serviceware — quantity, unit cost & line total per event.")
    table_headers(ws, 4, ["Item", "Qty", "Unit Cost", "Total"], start_col=2)
    start = L0
    for i, (item, qty, unit) in enumerate(RENTALS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cq = ws.cell(row=r, column=3, value=qty); cq.style = "input"; cq.number_format = "#,##0"
        cu = ws.cell(row=r, column=4, value=unit); cu.style = "input"; cu.number_format = '"$"#,##0.00'
        ct = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(RENTALS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="RENTALS TOTAL").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


def build_bookings(wb):
    ws = wb.create_sheet("Bookings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 12, 14, 14, 2])
    luxe_header(ws, "F", "📅  BOOKINGS",
                "Your event calendar — date, guest count, deposit & status, so nothing slips.")
    table_headers(ws, 4, ["Event", "Date", "Guests", "Deposit", "Status"], start_col=2)
    start = L0
    import datetime as dt
    for i, (name, days, guests, dep, status) in enumerate(BOOKINGS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cd = ws.cell(row=r, column=3, value=dt.date.today() + dt.timedelta(days=days)); cd.style = "input"; cd.number_format = "mm/dd"
        cg = ws.cell(row=r, column=4, value=guests); cg.style = "input"; cg.number_format = "#,##0"
        cdp = ws.cell(row=r, column=5, value=dep); cdp.style = "input"; cdp.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=6, value=status); cs.style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"F{r}", "StatusList")
    end = start + len(BOOKINGS) - 1
    cmap = {"Confirmed": MINT_BG, "Deposit": WARN_BG, "Recurring": SURFACE, "Inquiry": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    ws.freeze_panes = "A5"


def build_inventory(wb):
    ws = wb.create_sheet("Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 12, 12, 14, 2])
    luxe_header(ws, "F", "📦  INVENTORY & PAR",
                "Par vs on hand — order before the event, not the morning of.")
    table_headers(ws, 4, ["Item", "Par", "On Hand", "Unit", "To Order"], start_col=2)
    start = L0
    for i, (item, par, oh, unit) in enumerate(INVENTORY):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=par); cp.style = "input"; cp.number_format = "#,##0"
        co = ws.cell(row=r, column=4, value=oh); co.style = "input"; co.number_format = "#,##0"
        ws.cell(row=r, column=5, value=unit).style = "td"
        cb = ws.cell(row=r, column=6, value=f"=MAX(C{r}-D{r},0)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.number_format = "#,##0"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVENTORY) - 1
    add_dv(ws, f"E{start}:E{end}", "UnitList")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=[f"C{start}*0.5"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Over-prep, spoilage & breakage — the quiet leaks that eat your margin.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_ordering(wb):
    ws, start, end = build_log(
        wb, "Ordering", "🚚", "ORDERING & SUPPLIERS",
        "Your standing order — par quantities & cost, by supplier.",
        ["Item", "Supplier", "Par Order", "Cost"],
        ORDERING, [2, 26, 20, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)
    nrange(wb, "OrderCost", "Ordering", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ORDER TOTAL").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(OrderCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_cash(wb):
    ws = wb.create_sheet("Cash & Deposits"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 14, 2])
    luxe_header(ws, "D", "💵  CASH & DEPOSITS",
                "Deposits & payments in — know what's collected and what's still owed.")
    table_headers(ws, 4, ["Client / Event", "Amount", "Method"], start_col=2)
    start = L0
    for i, (client, amt, method) in enumerate(CASHDEP):
        r = start + i
        ws.cell(row=r, column=2, value=client).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=method).style = "td"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CASHDEP) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="DEPOSITS COLLECTED").style = "th"
    c = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cell_name(wb, "DepositsTotal", "Cash & Deposits", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_clients(wb):
    ws, start, end = build_log(
        wb, "Clients", "🤝", "CLIENTS",
        "Your book of business — who they are, what they book & where they stand.",
        ["Client", "Event Type", "Contact", "Status"],
        CLIENTS, [2, 24, 18, 24, 14, 2], text_left={2, 3, 4}, reserved=24, start_col=2,
        validations=[("E", "StatusList")])
    cmap = {"Repeat": MINT_BG, "New": WARN_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🍽  CATERING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Revenue, food cost, event margin & a Catering Score — your whole business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("EVENTS", "=COUNTA(EventName)", "num"),
        ("AVG GUESTS", "=IFERROR(AVERAGE(EventGuests),0)", "num"),
        ("REVENUE", "=TotalRevenue", "money"),
        ("AVG PER HEAD", "=IFERROR(TotalRevenue/SUM(EventGuests),0)", "money2"),
        ("FOOD COST", "=FoodCostPct", "pct"),
        ("TOP PACKAGE", "=INDEX(EventPkg,MATCH(MAX(EventRev),EventRev,0))", "text"),
    ]
    row2 = [
        ("AVG EVENT", "=IFERROR(TotalRevenue/COUNTA(EventName),0)", "money"),
        ("AVG MARGIN", "=AvgEventMargin", "pct"),
        ("LABOR", "=LaborPct", "pct"),
        ("PACKAGES", "=COUNTA(PkgItem)", "num"),
        ("WASTE %", "=IFERROR(WasteTotal/TotalRevenue,0)", "pct"),
        ("CATERING SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "CATERING HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY EVENT", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Food cost on target", "=IFERROR(MIN(TargetFC/FoodCostPct,1),0)"),
        ("Margin per event", "=IFERROR(MIN(AvgEventMargin/MarginGoal,1),0)"),
        ("Packages fully costed", "=IFERROR(COUNTIF(PkgCostHead,\">0\")/COUNTA(PkgItem),0)"),
        ("Labor under control", "=IFERROR(1-MIN(LaborPct/LaborLimit,1),0)"),
        ("Bookings vs goal", "=IFERROR(MIN(COUNTA(EventName)/BookingGoal,1),0)"),
        ("Gross margin", "=IFERROR(MIN((1-FoodCostPct)/0.7,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    eq = wb["Event Quotes"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Event"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(eq, min_col=10, min_row=5, max_row=4 + len(EVENTS)), titles_from_data=False)
    ch.set_categories(Reference(eq, min_col=2, min_row=5, max_row=4 + len(EVENTS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Catering Command Center™ — cost every head, quote with confidence, book more profit.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_plate(wb); build_packages(wb)
    build_quotes(wb); build_staffing(wb); build_rentals(wb); build_bookings(wb)
    build_inventory(wb); build_waste(wb); build_ordering(wb); build_cash(wb)
    build_clients(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Plate Costing", "Menu Packages", "Event Quotes", "Staffing",
             "Rentals", "Bookings", "Inventory", "Waste Log", "Ordering", "Cash & Deposits",
             "Clients", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Catering_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
