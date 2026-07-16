"""Build Preschool Command Center™ — The Complete At-Home Preschool System.

16 planning tabs (+ Settings) · a premium play-based preschool operating system
in Google Sheets & Excel. Preschool dashboard, a profile for every child, a
developmental skills & milestones tracker, ABC & 123 mastery, weekly themes, a
daily rhythm, a play-based activity planner, read-aloud log, arts & sensory
bank, nature walks & field trips, attendance, portfolio, supplies budget, a
kindergarten-readiness checklist and goals — one dashboard. Ages 3–5, one child
or a houseful.

Run: python3 build_xlsx.py   ->  ../Preschool_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

DOMAINS = ["Fine Motor", "Gross Motor", "Language", "Pre-Math", "Pre-Reading",
           "Social-Emotional", "Self-Help"]
SKILL_STATUS = ["Mastered", "Emerging", "Not Yet"]
ACT_STATUS = ["Done", "Planned", "Skipped"]
THEME_STATUS = ["Done", "In Progress", "Upcoming"]
YESNO = ["Yes", "No"]
LOVED = ["Loved it", "Liked it", "Meh"]
SUPPLY_STATUS = ["Have", "Ordered", "Need"]

# ---- the sample preschool — the data that drives every KPI ----
CHILDREN = [
    ("Millie Bennett", "PreK (4)", 4, "Loves art, animals & books · chatty"),
    ("Owen Bennett", "Preschool (3)", 3, "Busy builder · lots of energy · hands-on"),
]

# Developmental skills: (domain, skill, child, status)
SKILLS = [
    ("Fine Motor", "Holds a crayon with a tripod grip", "Millie", "Mastered"),
    ("Fine Motor", "Cuts along a line with scissors", "Millie", "Emerging"),
    ("Fine Motor", "Buttons & zips clothing", "Millie", "Mastered"),
    ("Fine Motor", "Strings beads / laces cards", "Owen", "Mastered"),
    ("Fine Motor", "Builds a tower of 8+ blocks", "Owen", "Mastered"),
    ("Gross Motor", "Hops on one foot", "Millie", "Mastered"),
    ("Gross Motor", "Pedals a tricycle", "Owen", "Mastered"),
    ("Gross Motor", "Throws & catches a ball", "Millie", "Emerging"),
    ("Gross Motor", "Walks up stairs alternating feet", "Owen", "Emerging"),
    ("Gross Motor", "Balances on a beam", "Millie", "Mastered"),
    ("Language", "Speaks in full sentences", "Millie", "Mastered"),
    ("Language", "Follows 2-step directions", "Millie", "Mastered"),
    ("Language", "Names 8+ colors", "Millie", "Mastered"),
    ("Language", "Identifies rhyming words", "Millie", "Emerging"),
    ("Language", "Uses 3–4 word sentences", "Owen", "Mastered"),
    ("Pre-Math", "Counts to 20", "Millie", "Mastered"),
    ("Pre-Math", "Counts to 10", "Owen", "Emerging"),
    ("Pre-Math", "Sorts by color & size", "Millie", "Mastered"),
    ("Pre-Math", "Recognizes numerals 0–10", "Millie", "Emerging"),
    ("Pre-Math", "Copies an AB pattern", "Millie", "Mastered"),
    ("Pre-Math", "Names basic shapes", "Millie", "Mastered"),
    ("Pre-Reading", "Recognizes own name", "Millie", "Mastered"),
    ("Pre-Reading", "Knows most letters", "Millie", "Mastered"),
    ("Pre-Reading", "Retells a simple story", "Millie", "Emerging"),
    ("Pre-Reading", "Holds a book & turns pages", "Owen", "Mastered"),
    ("Pre-Reading", "Print awareness (left-to-right)", "Millie", "Emerging"),
    ("Social-Emotional", "Takes turns", "Millie", "Mastered"),
    ("Social-Emotional", "Shares with a reminder", "Owen", "Emerging"),
    ("Social-Emotional", "Names her feelings", "Millie", "Mastered"),
    ("Social-Emotional", "Separates from parent easily", "Millie", "Mastered"),
    ("Self-Help", "Dresses self", "Millie", "Mastered"),
    ("Self-Help", "Uses the toilet independently", "Millie", "Mastered"),
    ("Self-Help", "Washes hands", "Owen", "Mastered"),
    ("Self-Help", "Cleans up toys", "Owen", "Emerging"),
    ("Self-Help", "Uses a fork & spoon", "Owen", "Mastered"),
]

# ABC mastery for the focus child (Millie): recognizes / sound / writes (Yes/No)
_NO_RECOGNIZE = {"Q", "X"}
_NO_SOUND = {"Q", "U", "V", "X", "Y", "Z"}
_WRITES = set("MILLEABTOSHRP")  # the letters she can write
def _abc_rows():
    rows = []
    for i in range(26):
        L = chr(65 + i)
        rows.append((L, "No" if L in _NO_RECOGNIZE else "Yes",
                     "No" if L in _NO_SOUND else "Yes",
                     "Yes" if L in _WRITES else "No"))
    return rows
ABC = _abc_rows()

# Weekly themes: (theme, focus / letter, status)
THEMES = [
    ("All About Me", "Letter M · self & family", "Done"),
    ("My Family & Home", "Letter F · homes", "Done"),
    ("Fall & Leaves", "Letter L · seasons", "Done"),
    ("Farm Animals", "Letter A · animal sounds", "Done"),
    ("Community Helpers", "Letter H · jobs", "Done"),
    ("Pumpkins & Harvest", "Letter P · counting", "Done"),
    ("Colors & Shapes", "Sorting · patterns", "Done"),
    ("Weather & Seasons", "Letter W · calendar", "In Progress"),
    ("Under the Sea", "Letter S · ocean", "Upcoming"),
    ("Outer Space", "Letter R · rockets", "Upcoming"),
    ("Dinosaurs", "Letter D · big/small", "Upcoming"),
    ("Spring & Bugs", "Letter B · life cycles", "Upcoming"),
    ("Transportation", "Letter T · go/stop", "Upcoming"),
    ("Zoo Animals", "Letter Z · habitats", "Upcoming"),
]

# Activity planner: (theme / area, activity, domain, status)
ACTIVITIES = [
    ("All About Me", "Paint a self-portrait", "Fine Motor", "Done"),
    ("All About Me", "Name puzzle with letters", "Pre-Reading", "Done"),
    ("My Family", "Family photo sorting & counting", "Pre-Math", "Done"),
    ("Fall & Leaves", "Leaf-rubbing crayon art", "Fine Motor", "Done"),
    ("Fall & Leaves", "Nature walk & leaf hunt", "Gross Motor", "Done"),
    ("Farm Animals", "Animal-sound matching game", "Language", "Done"),
    ("Farm Animals", "Count the farm animals 1–10", "Pre-Math", "Done"),
    ("Community Helpers", "Dress-up dramatic play", "Social-Emotional", "Done"),
    ("Pumpkins", "Pumpkin sensory bin", "Fine Motor", "Done"),
    ("Pumpkins", "Pumpkin seed counting", "Pre-Math", "Done"),
    ("Colors & Shapes", "Shape hunt around the house", "Pre-Math", "Done"),
    ("Colors & Shapes", "Rainbow AB pattern beads", "Fine Motor", "Done"),
    ("Colors & Shapes", "Color-sorting pom-poms", "Pre-Math", "Done"),
    ("Weather", "Weather chart & calendar", "Pre-Math", "Done"),
    ("Weather", "Cloud-dough sensory play", "Fine Motor", "Planned"),
    ("Weather", "Rain-cloud science jar", "Pre-Reading", "Planned"),
    ("Under the Sea", "Ocean sensory small world", "Language", "Planned"),
    ("Under the Sea", "Fish counting clip cards", "Pre-Math", "Planned"),
    ("Outer Space", "Build a cardboard rocket", "Gross Motor", "Planned"),
    ("Dinosaurs", "Dino dig excavation bin", "Fine Motor", "Planned"),
    ("Spring & Bugs", "Butterfly life-cycle craft", "Pre-Reading", "Planned"),
    ("Transportation", "Ramp & car races", "Pre-Math", "Planned"),
    ("Zoo Animals", "Zoo animal habitat sort", "Language", "Planned"),
    ("Any", "Play-dough letter formation", "Fine Motor", "Done"),
]

# Read-aloud log: (title, author, theme tie-in, times read, loved)
READALOUD = [
    ("Brown Bear, Brown Bear", "Bill Martin Jr.", "Colors", 12, "Loved it"),
    ("The Very Hungry Caterpillar", "Eric Carle", "Bugs / counting", 9, "Loved it"),
    ("Chicka Chicka Boom Boom", "Bill Martin Jr.", "Letters", 14, "Loved it"),
    ("Goodnight Moon", "Margaret Wise Brown", "Bedtime", 20, "Loved it"),
    ("The Gruffalo", "Julia Donaldson", "Rhyming", 7, "Loved it"),
    ("Dragons Love Tacos", "Adam Rubin", "Silly", 8, "Liked it"),
    ("Press Here", "Hervé Tullet", "Colors / interactive", 6, "Loved it"),
    ("The Pout-Pout Fish", "Deborah Diesen", "Under the sea", 5, "Liked it"),
    ("Room on the Broom", "Julia Donaldson", "Rhyming", 6, "Loved it"),
    ("Little Blue Truck", "Alice Schertle", "Transportation", 10, "Loved it"),
    ("We're Going on a Bear Hunt", "Michael Rosen", "Movement", 8, "Loved it"),
    ("Dear Zoo", "Rod Campbell", "Zoo animals", 7, "Liked it"),
]

# Arts & sensory idea bank (no KPI): (activity, area, materials, builds)
ARTS = [
    ("Rainbow rice sensory bin", "Sensory", "Rice, food color, cups, scoops", "Fine motor · pouring"),
    ("Salt-tray letter writing", "Pre-Reading", "Tray, salt, letter cards", "Letter formation"),
    ("Cotton-ball painting", "Art", "Cotton balls, clips, paint", "Grip · color mixing"),
    ("Play-dough station", "Fine Motor", "Dough, cutters, rollers", "Hand strength"),
    ("Water-bead tweezing", "Fine Motor", "Water beads, tweezers, tray", "Pincer grasp"),
    ("Nature paintbrushes", "Art", "Leaves, twigs, paint", "Creativity · texture"),
    ("Ice excavation", "Science", "Frozen toys, warm water, droppers", "Cause & effect"),
    ("Pom-pom color sort", "Pre-Math", "Pom-poms, muffin tin, scoop", "Sorting · counting"),
    ("Sticker line-up", "Fine Motor", "Stickers, lined paper", "Placement · control"),
    ("Kinetic-sand molds", "Sensory", "Kinetic sand, molds", "Calming · shapes"),
    ("Contact-paper collage", "Art", "Contact paper, tissue scraps", "Sticky exploration"),
    ("Sink-or-float tub", "Science", "Tub, small objects, chart", "Predict · observe"),
]

# Nature walks & field trips: (date offset, place, theme tie-in, notes)
TRIPS = [
    (-42, "Pumpkin patch", "Pumpkins", "Picked & counted pumpkins"),
    (-30, "Children's library story time", "Letters", "Weekly · loves it"),
    (-22, "Local farm", "Farm animals", "Fed the goats"),
    (-14, "Nature-center leaf walk", "Fall", "Collected 6 leaf types"),
    (-6, "Fire station tour", "Community helpers", "Sat in the truck"),
    (7, "Aquarium", "Under the sea", "Booked for theme week"),
    (16, "Botanical garden bug hunt", "Spring & bugs", "Magnifying glasses"),
]

# Read-aloud goal & book count in Settings; days per week ~4
WEEKS = [(f"Week {i}", 4, "") for i in range(1, 25)]
DAYS_GOAL = 150
BOOKS_READ = 38
SUPPLIES_BUDGET = 400

# Kindergarten-readiness checklist: (item, done)
READINESS = [
    ("Writes first name", True), ("Recognizes 20+ letters", True),
    ("Says most letter sounds", False), ("Counts to 20", True),
    ("Recognizes numerals 0–10", False), ("Knows 8+ colors", True),
    ("Names basic shapes", True), ("Holds a pencil correctly", True),
    ("Cuts with scissors on a line", False), ("Uses the bathroom independently", True),
    ("Dresses self", True), ("Follows 2-step directions", True),
    ("Takes turns & shares", True), ("Separates from caregiver calmly", True),
    ("Sits & listens to a short lesson", True), ("Identifies rhyming words", False),
]

# Goals: (goal, child, progress)
GOALS = [
    ("Learn all letter sounds", "Millie", 0.77),
    ("Count to 30", "Millie", 0.66),
    ("Write her whole first name", "Millie", 0.90),
    ("Fully potty trained", "Owen", 0.80),
    ("Cut on a line with scissors", "Millie", 0.60),
    ("Share without a reminder", "Owen", 0.55),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your family details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "PRESCHOOL INPUTS", "section")
    controls = [
        ("Family Name", "The Bennett Family", None, "FamilyName"),
        ("School Year", "2026–2027", None, "SchoolYear"),
        ("Children", 2, "0", "Children"),
        ("Days Goal (year)", DAYS_GOAL, "0", "DaysGoal"),
        ("Books Read (year)", BOOKS_READ, "#,##0", "BooksRead"),
        ("Supplies Budget", SUPPLIES_BUDGET, '"$"#,##0', "SuppliesBudget"),
        ("Kindergarten Date", "Fall 2028", None, "KinderDate"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Domain", DOMAINS, "DomainList"), ("F", "Skill Status", SKILL_STATUS, "SkillStatusList"),
             ("G", "Activity", ACT_STATUS, "ActStatusList"), ("H", "Theme", THEME_STATUS, "ThemeStatusList"),
             ("I", "Supplies", SUPPLY_STATUS, "SupplyStatusList"), ("J", "Loved?", LOVED, "LovedList"),
             ("K", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🧸  PRESCHOOL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Play with a plan — themes, skills & sweet memories, all in one calm system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE PRESCHOOL, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Weekly themes, a play-based activity planner, a developmental skills & milestones tracker, "
                      "ABC & 123 mastery, read-alouds, arts & sensory, nature walks, a portfolio and a kindergarten-"
                      "readiness checklist — all in ONE premium Google Sheets & Excel system. A profile for every child "
                      "and a live dashboard that shows how they're growing. Play-based and gentle, it works for one "
                      "little one or a houseful — no teaching degree required.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your family, children & days goal.",
             "2.  Fill a Child Profile for each little one — age, interests, what they're working on.",
             "3.  Pick this week's Theme, then plan a few Activities from the play bank.",
             "4.  As they grow, mark Skills & Milestones and check off ABC & 123 mastery.",
             "5.  Log read-alouds, nature walks & save a few portfolio photos as you go.",
             "6.  Watch the Dashboard track letters, skills, themes & a Ready-for-K Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for the Bennett family (Millie, 4, and Owen, 3) is included so you can see how it all "
               "connects — just type over it with your own. Letters, skills, themes, activities and the readiness "
               "checklist roll up into a live Ready-for-Kindergarten Score. Twelve matching printable pages (weekly "
               "theme plan, skills checklist, ABC & 123 chart, readiness list, portfolio & more) are included to print "
               "and keep. This is a planning & keepsake tool — every child grows at their own pace, so use it as a gentle "
               "guide, not a test.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Follow the fun — the learning follows. You've got this.", "section_gold")


# ===========================================================================
# 3 — Child Profiles
# ===========================================================================
def build_profiles(wb):
    ws = wb.create_sheet("Child Profiles"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 22, 4, 18, 18, 2])
    luxe_header(ws, "G", "🧒  CHILD PROFILES",
                "A profile for every little one — age, interests, strengths & what they're working on.")
    row = 5
    for name, stage, age, note in CHILDREN:
        merge_set(ws, f"B{row}:F{row}", name, "section_gold"); ws.row_dimensions[row].height = 22
        row += 1
        fields = [("Stage", stage), ("Age", age), ("Personality", note), ("Loves", "—"),
                  ("Working on", "See Skills tab"), ("Favorite book", "—"), ("Favorite play", "—"),
                  ("Naps?", "—"), ("Allergies / notes", "—"), ("Big goal", "See Goals tab")]
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 22; i += 2; row += 1
        row += 1
    ws.freeze_panes = "A5"


# ===========================================================================
# 4 — Skills & Milestones  (defines SkillDomain / SkillName / SkillStatus)
# ===========================================================================
def build_skills(wb):
    ws, start, end = build_log(
        wb, "Skills & Milestones", "🌱", "SKILLS & MILESTONES",
        "The whole-child picture — fine & gross motor, language, pre-math, pre-reading, social & self-help.",
        ["Domain", "Skill / Milestone", "Child", "Status"],
        SKILLS, [18, 34, 14, 14], text_left={2}, reserved=45,
        validations=[("A", "DomainList"), ("D", "SkillStatusList")])
    nrange(wb, "SkillDomain", "Skills & Milestones", "A", start, end)
    nrange(wb, "SkillName", "Skills & Milestones", "B", start, end)
    nrange(wb, "SkillStatus", "Skills & Milestones", "D", start, end)
    cmap = {"Mastered": MINT_BG, "Emerging": WARN_BG, "Not Yet": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    # domain summary (mastered per domain) — feeds the dashboard bar chart
    sr = end + 3
    ws.cell(row=sr - 1, column=2, value="MASTERED BY DOMAIN").style = "section_gold"
    for i, dom in enumerate(DOMAINS):
        r = sr + i
        ws.cell(row=r, column=2, value=dom).style = "td_left"
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(SkillDomain,B{r},SkillStatus,"Mastered")')
        c.style = "td"; c.number_format = "#,##0"
    cell_name(wb, "DomSumLabels", "Skills & Milestones", f"$B${sr}:$B${sr + len(DOMAINS) - 1}")
    cell_name(wb, "DomSumVals", "Skills & Milestones", f"$C${sr}:$C${sr + len(DOMAINS) - 1}")


# ===========================================================================
# 5 — ABC & 123  (defines AbcRecognize)
# ===========================================================================
def build_abc(wb):
    ws, start, end = build_log(
        wb, "ABC & 123", "🔤", "ABC & 123 MASTERY",
        "Letter by letter — recognizes, says the sound & writes it. Set the focus child in the header note.",
        ["Letter", "Recognizes?", "Says Sound?", "Writes?"],
        ABC, [12, 16, 16, 14], reserved=26,
        validations=[("B", "YesNoList"), ("C", "YesNoList"), ("D", "YesNoList")])
    nrange(wb, "AbcRecognize", "ABC & 123", "B", start, end)
    nrange(wb, "AbcSound", "ABC & 123", "C", start, end)
    nrange(wb, "AbcWrite", "ABC & 123", "D", start, end)
    for col in ("B", "C", "D"):
        ws.conditional_formatting.add(f"{col}{start}:{col}{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    # 123 counting summary block to the right
    ws.cell(row=4, column=6, value="COUNTING & NUMBERS").style = "th"
    ws.column_dimensions["F"].width = 26; ws.column_dimensions["G"].width = 12
    rows123 = [("Counts aloud to", 20), ("Recognizes numerals 0–", 8),
               ("1:1 correspondence to", 12), ("Writes numerals 0–", 5)]
    for i, (lab, val) in enumerate(rows123):
        r = 5 + i
        ws.cell(row=r, column=6, value=lab).style = "field_label"
        c = ws.cell(row=r, column=7, value=val); c.style = "input"; c.number_format = "#,##0"
    ws.cell(row=10, column=6, value="LETTERS KNOWN (of 26)").style = "field_label"
    c = ws.cell(row=10, column=7, value='=COUNTIF(AbcRecognize,"Yes")'); c.style = "field_value"; c.number_format = "#,##0"; c.fill = fill(MINT_BG)


# ===========================================================================
# 6 — Weekly Themes  (defines ThemeStatus / ThemeName)
# ===========================================================================
def build_themes(wb):
    ws, start, end = build_log(
        wb, "Weekly Themes", "🗓", "WEEKLY THEMES",
        "A gentle theme a week — pair a letter, a few books and some play. Plan the whole year at a glance.",
        ["Theme", "Focus / Letter", "Status"],
        THEMES, [24, 28, 14], text_left={1, 2}, reserved=30,
        validations=[("C", "ThemeStatusList")])
    nrange(wb, "ThemeName", "Weekly Themes", "A", start, end)
    nrange(wb, "ThemeStatus", "Weekly Themes", "C", start, end)
    cmap = {"Done": MINT_BG, "In Progress": WARN_BG, "Upcoming": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 7 — Daily Rhythm
# ===========================================================================
def build_rhythm(wb):
    ws = wb.create_sheet("Daily Rhythm"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 26, 30, 2])
    luxe_header(ws, "D", "🌤  DAILY RHYTHM & CIRCLE TIME",
                "A loose, repeatable flow — little ones thrive on rhythm, not a rigid clock.")
    table_headers(ws, 4, ["When", "Block", "What we do"], start_col=2)
    rows = [
        ("Morning", "Welcome & breakfast", "Free play while the day gets going"),
        ("~9:00", "Circle time", "Calendar, weather, songs, letter & theme intro"),
        ("~9:30", "Theme activity", "The day's planned play-based activity"),
        ("~10:00", "Snack & story", "Snack + a theme read-aloud"),
        ("~10:30", "Outside / gross motor", "Nature walk, park or backyard play"),
        ("~11:15", "Table time", "ABC/123, fine-motor tray, art or sensory bin"),
        ("Noon", "Lunch & rest", "Lunch, then quiet time or nap"),
        ("Afternoon", "Open play", "Blocks, dramatic play, puzzles, books"),
        ("Late day", "Tidy & wind-down", "Clean-up song, one more story, reflect"),
    ]
    start = L0
    for i, (when, block, what) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=when).style = "td"
        ws.cell(row=r, column=3, value=block).style = "td_left"
        ws.cell(row=r, column=4, value=what).style = "td_left"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"
    merge_set(ws, f"B{start+10}:D{start+10}",
              "Follow your child: skip, swap or repeat blocks freely. Consistency of order matters more than the clock.",
              "section_gold")


# ===========================================================================
# 8 — Activity Planner  (defines ActName / ActStatus)
# ===========================================================================
def build_activities(wb):
    ws, start, end = build_log(
        wb, "Activity Planner", "🎨", "ACTIVITY PLANNER",
        "Play with a purpose — every activity tagged to a theme and a skill. Plan it, do it, check it off.",
        ["Theme / Area", "Activity", "Skill Domain", "Status"],
        ACTIVITIES, [18, 30, 18, 14], text_left={2}, reserved=45,
        validations=[("C", "DomainList"), ("D", "ActStatusList")])
    nrange(wb, "ActName", "Activity Planner", "B", start, end)
    nrange(wb, "ActStatus", "Activity Planner", "D", start, end)
    cmap = {"Done": MINT_BG, "Planned": WARN_BG, "Skipped": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 9 — Read-Aloud Log
# ===========================================================================
def build_readaloud(wb):
    ws, start, end = build_log(
        wb, "Read-Aloud Log", "📖", "READ-ALOUD LOG",
        "The heart of preschool — every favorite, how many times & who loved it. Set the year's count in Settings.",
        ["Title", "Author", "Theme Tie-in", "Times Read", "Loved?"],
        READALOUD, [30, 20, 18, 12, 14], text_left={1, 2, 3}, ints={4}, reserved=40,
        validations=[("E", "LovedList")])
    ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=['"Loved it"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=20, color=PRIMARY, showValue=True))


# ===========================================================================
# 10 — Arts & Sensory bank
# ===========================================================================
def build_arts(wb):
    build_log(wb, "Arts & Sensory", "🖐", "ARTS, CRAFTS & SENSORY BANK",
              "A grab-and-go idea bank — pick one when you need calm, busy hands. Materials you likely already have.",
              ["Activity", "Area", "Materials", "Builds"],
              ARTS, [26, 14, 30, 22], text_left={1, 3, 4}, reserved=30)


# ===========================================================================
# 11 — Nature Walks & Field Trips  (defines TripName)
# ===========================================================================
def build_trips(wb):
    rows = [(dplus(off) if off >= 0 else dminus(-off), place, theme, notes)
            for (off, place, theme, notes) in TRIPS]
    ws, start, end = build_log(
        wb, "Field Trips", "🌳", "NATURE WALKS & FIELD TRIPS",
        "Learning beyond the table — every outing, what it tied to & a memory to keep.",
        ["Date", "Place", "Theme Tie-in", "Notes / Memory"],
        rows, [14, 26, 18, 28], text_left={2, 4}, dates={1}, reserved=30)
    nrange(wb, "TripName", "Field Trips", "B", start, end)


# ===========================================================================
# 12 — Attendance & Days  (defines DaysDone)
# ===========================================================================
def build_attendance(wb):
    ws, start, end = build_log(
        wb, "Attendance", "📅", "ATTENDANCE & DAYS",
        "Preschool days by week — a light record of the rhythm you're keeping. Set your days goal in Settings.",
        ["Week", "Days", "Notes"],
        WEEKS, [16, 12, 40], text_left={3}, ints={2}, reserved=40)
    nrange(wb, "AttDays", "Attendance", "B", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL").style = "th"
    cd = ws.cell(row=tot, column=2, value="=SUM(AttDays)"); cd.style = "td"; cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = "#,##0"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cell_name(wb, "DaysDone", "Attendance", f"$B${tot}")


# ===========================================================================
# 13 — Portfolio
# ===========================================================================
def build_portfolio(wb):
    ws = wb.create_sheet("Portfolio"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "🖼  PORTFOLIO & KEEPSAKES",
                "Snap a few favorites each month — the art, the milestones, the sweet little wins.")
    merge_set(ws, "B5:D5", "HOW TO ADD PHOTOS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"link\"). "
                      "Excel: Insert ▸ Pictures ▸ Place in Cell. Photograph a painting, a first name written, a block "
                      "tower or a proud face and caption it — child, date & what it shows. A few a month is plenty.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["First Name Written", "Best Painting", "Block Tower", "Nature Find", "Dress-Up Play", "Milestone Moment"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add photo)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Child · what it shows · date…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 14 — Supplies & Budget  (defines SuppliesSpent)
# ===========================================================================
def build_supplies(wb):
    ws = wb.create_sheet("Supplies"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "🧺  SUPPLIES & BUDGET",
                "Art, sensory, books & printables — planned vs actual, so preschool stays low-cost & low-stress.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining"], start_col=2)
    rows = [
        ("Art supplies", 90, 74), ("Sensory bin fillers", 60, 48), ("Books & library", 70, 52),
        ("Printables & laminating", 40, 33), ("Manipulatives / toys", 80, 61),
        ("Craft & consumables", 40, 28), ("Field trips", 60, 42), ("Misc", 30, 12),
    ]
    start = L0
    for i, (cat, plan, actual) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=3, value=plan); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=actual); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "SuppliesPlanTotal", "Supplies", f"$C${tot}")
    cell_name(wb, "SuppliesSpent", "Supplies", f"$D${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=90, color=PRIMARY, showValue=True))
    merge_set(ws, "B15:E15", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Total budget", "=SuppliesPlanTotal", '"$"#,##0'), ("Spent so far", "=SuppliesSpent", '"$"#,##0'),
             ("Remaining", "=SuppliesPlanTotal-SuppliesSpent", '"$"#,##0'),
             ("Per child (avg)", "=IFERROR(SuppliesSpent/Children,0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 16 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Remaining", "Per child (avg)"):
            c.fill = fill(MINT_BG)


# ===========================================================================
# 15 — Kindergarten Readiness  (defines RdyName / RdyDone)
# ===========================================================================
def build_readiness(wb):
    rows = [(item, "Yes" if done else "No", "—") for (item, done) in READINESS]
    ws, start, end = build_log(
        wb, "Kindergarten Readiness", "🎒", "KINDERGARTEN READINESS",
        "The gentle checklist — the skills most kindergartens hope for. A guide, never a test; every child is different.",
        ["Readiness Skill", "Ready?", "Notes"],
        rows, [40, 12, 30], text_left={1, 3}, reserved=20,
        validations=[("B", "YesNoList")])
    nrange(wb, "RdyName", "Kindergarten Readiness", "A", start, end)
    nrange(wb, "RdyDone", "Kindergarten Readiness", "B", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 16 — Goals & Milestones  (defines GoalProgress)
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 14, 2])
    luxe_header(ws, "D", "🎯  GOALS & MILESTONES",
                "The little wins that matter — one meaningful goal per child, plus family goals.")
    table_headers(ws, 4, ["Goal", "Child", "Progress"], start_col=2)
    start = L0
    for i, (goal, who, prog) in enumerate(GOALS):
        r = start + i
        ws.cell(row=r, column=2, value=goal).style = "td_left"
        ws.cell(row=r, column=3, value=who).style = "td"
        cp = ws.cell(row=r, column=4, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(GOALS) - 1
    nrange(wb, "GoalProgress", "Goals", "D", start, end)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 2 — Preschool Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🧸  PRESCHOOL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Letters, skills, themes & sweet memories — your whole preschool year, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("CHILDREN", "=Children", "num"),
        ("PRESCHOOL DAYS", "=DaysDone", "num"),
        ("LETTERS KNOWN", '=COUNTIF(AbcRecognize,"Yes")', "num"),
        ("SKILLS MASTERED", '=COUNTIF(SkillStatus,"Mastered")', "num"),
        ("THEMES DONE", '=COUNTIF(ThemeStatus,"Done")', "num"),
        ("ACTIVITIES DONE", '=IFERROR(COUNTIF(ActStatus,"Done")/COUNTA(ActName),0)', "pct"),
    ]
    row2 = [
        ("BOOKS READ", "=BooksRead", "num"),
        ("FIELD TRIPS", "=COUNTA(TripName)", "num"),
        ("SUPPLIES SPENT", "=SuppliesSpent", "money"),
        ("READY FOR K", '=IFERROR(COUNTIF(RdyDone,"Yes")/COUNTA(RdyName),0)', "pct"),
        ("GOALS", "=IFERROR(AVERAGE(GoalProgress),0)", "pct"),
        ("READY SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "GROWING WELL?", "section_gold")
    merge_set(ws, "H11:M11", "SKILLS MASTERED BY DOMAIN", "section_gold")
    table_headers(ws, 12, ["Area", "Progress", "Status"], start_col=2)
    dims = [
        ("Letters known (of 26)", '=IFERROR(COUNTIF(AbcRecognize,"Yes")/26,0)'),
        ("Skills mastered", '=IFERROR(COUNTIF(SkillStatus,"Mastered")/COUNTA(SkillName),0)'),
        ("Themes done", '=IFERROR(COUNTIF(ThemeStatus,"Done")/COUNTA(ThemeName),0)'),
        ("Activities done", '=IFERROR(COUNTIF(ActStatus,"Done")/COUNTA(ActName),0)'),
        ("Kindergarten readiness", '=IFERROR(COUNTIF(RdyDone,"Yes")/COUNTA(RdyName),0)'),
        ("Goals & milestones", "=IFERROR(AVERAGE(GoalProgress),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Soaring",IF(C{r}>=0.6,"Growing","Emerging"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # Skills-by-domain bar chart (from Skills tab summary)
    sk = wb["Skills & Milestones"]
    dn = wb.defined_names["DomSumVals"].value  # e.g. 'Skills & Milestones'!$C$..:$C$..
    # derive rows from the named range
    import re
    m = re.search(r"\$C\$(\d+):\$C\$(\d+)", dn)
    r0, r1 = int(m.group(1)), int(m.group(2))
    bar = BarChart(); bar.type = "bar"; bar.title = "Skills Mastered by Domain"; bar.height = 7.4; bar.width = 8.6
    bar.add_data(Reference(sk, min_col=3, min_row=r0, max_row=r1), titles_from_data=False)
    bar.set_categories(Reference(sk, min_col=2, min_row=r0, max_row=r1))
    bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
    bar.legend = None
    ws.add_chart(bar, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Preschool Command Center™ — play with a plan, and watch them grow. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profiles(wb); build_skills(wb)
    build_abc(wb); build_themes(wb); build_rhythm(wb); build_activities(wb)
    build_readaloud(wb); build_arts(wb); build_trips(wb); build_attendance(wb)
    build_portfolio(wb); build_supplies(wb); build_readiness(wb); build_goals(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Child Profiles", "Skills & Milestones", "ABC & 123",
             "Weekly Themes", "Daily Rhythm", "Activity Planner", "Read-Aloud Log", "Arts & Sensory",
             "Field Trips", "Attendance", "Portfolio", "Supplies", "Kindergarten Readiness", "Goals", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Preschool_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
