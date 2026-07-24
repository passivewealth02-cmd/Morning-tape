"""Build Subscription & Bills Audit Command Center™ — The Recurring-Spend Operating System.

14 tabs · a premium subscription & bills audit operating system in Google Sheets &
Excel. Dashboard, an audit engine (every recurring charge → monthly & annualized, with a
cancel-savings finder), subscriptions, bills, a cancel finder, spend by category,
renewals, free trials, price hikes, a savings log, a negotiation list and a monthly
summary — one dashboard. See every recurring charge, and cut the ones you forgot.

Run: python3 build_xlsx.py   ->  ../Subscription_Audit_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
ACTION = ["Keep", "Cancel", "Review"]
BILLING = ["Monthly", "Yearly", "Quarterly"]
CATEGORY = ["Streaming", "Music", "Software", "Cloud", "Fitness", "News", "Shopping", "Food", "Books", "Other"]

SUB_BUDGET = 250
SAVE_GOAL = 40
ANNUAL_GOAL = 5

# Subscriptions: (name, category, billing, monthly, action)
SUBS = [
    ("Netflix", "Streaming", "Monthly", 15.49, "Keep"), ("Spotify", "Music", "Monthly", 11.99, "Keep"),
    ("Disney+", "Streaming", "Monthly", 13.99, "Keep"), ("Hulu", "Streaming", "Monthly", 17.99, "Cancel"),
    ("Amazon Prime", "Shopping", "Yearly", 12.42, "Keep"), ("iCloud 200GB", "Cloud", "Monthly", 2.99, "Keep"),
    ("Adobe CC", "Software", "Monthly", 22.99, "Cancel"), ("NYT", "News", "Monthly", 17.00, "Keep"),
    ("Gym membership", "Fitness", "Monthly", 40.00, "Cancel"), ("YouTube Premium", "Streaming", "Monthly", 13.99, "Keep"),
    ("Audible", "Books", "Monthly", 14.95, "Cancel"), ("Dropbox", "Cloud", "Yearly", 9.99, "Keep"),
    ("Peloton App", "Fitness", "Monthly", 12.99, "Keep"), ("DoorDash DashPass", "Food", "Monthly", 9.99, "Cancel"),
]

# Bills: (bill, category, monthly)
BILLS = [
    ("Electric", "Utilities", 120), ("Internet", "Utilities", 70), ("Phone", "Utilities", 85),
    ("Car insurance", "Insurance", 130), ("Renters insurance", "Insurance", 18), ("Water", "Utilities", 45),
]

# Categories to roll up (spend by category)
CAT_ROLLUP = ["Streaming", "Music", "Software", "Cloud", "Fitness", "News", "Shopping", "Food", "Books"]

# Renewals: (service, renewal date, amount)
RENEWALS = [
    ("Amazon Prime", "Feb 14", 149), ("Dropbox", "Mar 2", 119), ("NYT (annual roll)", "Apr 9", 204),
    ("Car insurance", "May 1", 780), ("Adobe CC", "Jun 20", 276),
]

# Free trials: (service, started, cancel by)
TRIALS = [
    ("Paramount+", "Jul 1", "Jul 8"), ("Masterclass", "Jul 3", "Jul 10"), ("Canva Pro", "Jul 5", "Aug 4"),
]

# Price hikes: (service, old, new)
HIKES = [
    ("Netflix", 13.49, 15.49), ("Disney+", 10.99, 13.99), ("Spotify", 10.99, 11.99),
    ("YouTube Premium", 11.99, 13.99), ("NYT", 15.00, 17.00),
]

# Savings log (what you've cut): (item, monthly saved, date)
SAVINGS = [
    ("Cable TV", 89, "Jan 2"), ("Duplicate cloud", 10, "Feb 8"), ("Unused magazine", 8, "Mar 3"),
    ("Old gym", 25, "Apr 1"),
]

# Negotiation list: (bill, provider, current, target, called?)
NEGOTIATION = [
    ("Internet", "Xfinity", 70, 50, "No"), ("Phone", "Verizon", 85, 65, "No"),
    ("Car insurance", "Geico", 130, 110, "Yes"), ("Electric", "Utility", 120, 105, "No"),
]

# Monthly summary: (month, recurring spend) — trending down as you cut
MONTHS = [("Feb", 322), ("Mar", 315), ("Apr", 300), ("May", 280), ("Jun", 250), ("Jul", 217)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your goals & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR GOALS", "section")
    controls = [
        ("Your name", "Devon", None, "Owner"),
        ("Audit name", "Clearing House", None, "AuditName"),
        ("Monthly subscription budget", SUB_BUDGET, '"$"#,##0', "SubBudget"),
        ("Monthly-savings goal", SAVE_GOAL, '"$"#,##0', "SaveGoal"),
        ("Annual-billing goal (count)", ANNUAL_GOAL, "0", "AnnualGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Category", CATEGORY, "CategoryList"), ("F", "Billing", BILLING, "BillingList"),
             ("G", "Action", ACTION, "ActionList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🔍  SUBSCRIPTION & BILLS AUDIT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  See every recurring charge, and cut the ones you forgot.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE RECURRING SPEND, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("The average household leaks hundreds a year to subscriptions it forgot it had. This finds them: an "
                      "audit engine lists every subscription and bill, totals your true monthly and annual recurring "
                      "spend, and a cancel-savings finder adds up exactly what you'd save by cutting the ones you flag. "
                      "Track price hikes, renewals and free trials before they charge you — all in ONE premium Google "
                      "Sheets & Excel system built to plug the leaks.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your subscription budget & goals.",
             "2.  List every Subscription and Bill you pay.",
             "3.  Flag the ones to cancel — the Cancel Finder totals your savings.",
             "4.  Log Price Hikes, Renewals and Free Trials.",
             "5.  Work the Negotiation list to lower the keepers.",
             "6.  Check the Dashboard: recurring spend, savings & an Audit Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional audit (Clearing House, Devon) is included so you can see how it all connects "
               "— just type over it with your own subscriptions and bills. Your monthly recurring spend and the annual "
               "savings you'd get from cancelling are the two numbers that decide whether an audit pays off, and they "
               "roll into a live Audit Score. Twelve matching printable pages (subscription audit, cancel list, renewal "
               "calendar, bill negotiation & more) are included. This is a budgeting & organizing tool, not financial "
               "advice — confirm figures with your own statements.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Every charge you forgot is money you're giving away. Find them all.", "section_gold")


# ===========================================================================
def build_subscriptions(wb):
    ws = wb.create_sheet("Subscriptions"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 14, 14, 14, 2])
    luxe_header(ws, "F", "📺  SUBSCRIPTIONS",
                "Every recurring subscription — category, billing, monthly cost and keep-or-cancel. The whole leak, listed.")
    table_headers(ws, 4, ["Subscription", "Category", "Billing", "Monthly", "Action"], start_col=2)
    start = L0
    for i, (name, cat, bill, monthly, action) in enumerate(SUBS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        ws.cell(row=r, column=4, value=bill).style = "td"
        cm = ws.cell(row=r, column=5, value=monthly); cm.style = "input"; cm.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6, value=action).style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SUBS) - 1
    nrange(wb, "SubName", "Subscriptions", "B", start, end)
    nrange(wb, "SubCategory", "Subscriptions", "C", start, end)
    nrange(wb, "SubBilling", "Subscriptions", "D", start, end)
    nrange(wb, "SubMonthlyRange", "Subscriptions", "E", start, end)
    nrange(wb, "SubAction", "Subscriptions", "F", start, end)
    add_dv(ws, f"C{start}:C{end}", "CategoryList")
    add_dv(ws, f"D{start}:D{end}", "BillingList")
    add_dv(ws, f"F{start}:F{end}", "ActionList")
    ws.conditional_formatting.add(f"F{start}:F{end}",
        __import__("openpyxl").formatting.rule.CellIsRule(operator="equal", formula=['"Cancel"'], fill=fill(RED_BG)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="MONTHLY SUBS").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=5, value="=SUM(SubMonthlyRange)"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0.00'
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)
    cell_name(wb, "SubMonthly", "Subscriptions", f"$E${tot}")
    ws.cell(row=tot + 1, column=2, value="Subscriptions counted").style = "field_label"
    cc = ws.cell(row=tot + 1, column=5, value="=COUNTA(SubName)"); cc.style = "field_value"; cc.number_format = "#,##0"
    cell_name(wb, "SubCount", "Subscriptions", f"$E${tot+1}")
    ws.cell(row=tot + 2, column=2, value="On annual billing").style = "field_label"
    ca = ws.cell(row=tot + 2, column=5, value='=COUNTIF(SubBilling,"Yearly")'); ca.style = "field_value"; ca.number_format = "#,##0"
    cell_name(wb, "AnnualBilled", "Subscriptions", f"$E${tot+2}")
    ws.freeze_panes = "A5"


def build_audit(wb):
    ws = wb.create_sheet("Subscription Audit"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 18, 2])
    luxe_header(ws, "C", "🔍  SUBSCRIPTION AUDIT",
                "Every recurring charge, totalled monthly and annually — and exactly what you'd save by cutting the flagged ones.")
    rows = [
        ("Monthly subscriptions", "=SubMonthly", '"$"#,##0.00', MINT_BG),
        ("× 12 months", "=SubMonthly*12", '"$"#,##0', None),
    ]
    for i, (lab, val, fmt, bg) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "field_value"; c.number_format = fmt
        if bg:
            c.fill = fill(bg)
    ws.cell(row=7, column=2, value="= ANNUAL SUBSCRIPTION SPEND").style = "th"
    ca = ws.cell(row=7, column=3, value="=SubMonthly*12"); ca.style = "td"; ca.font = Font(bold=True, size=13, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = '"$"#,##0'
    ws.cell(row=9, column=2, value="CANCEL-SAVINGS FINDER").style = "section_gold"
    ws.cell(row=10, column=2, value="Flagged to cancel (monthly)").style = "field_label"
    ccm = ws.cell(row=10, column=3, value='=SUMIF(SubAction,"Cancel",SubMonthlyRange)'); ccm.style = "field_value"; ccm.number_format = '"$"#,##0.00'
    cell_name(wb, "CancelMonthly", "Subscription Audit", "$C$10")
    ws.cell(row=11, column=2, value="Flagged to cancel (count)").style = "field_label"
    ccc = ws.cell(row=11, column=3, value='=COUNTIF(SubAction,"Cancel")'); ccc.style = "field_value"; ccc.number_format = "#,##0"
    cell_name(wb, "CancelCount", "Subscription Audit", "$C$11")
    ws.cell(row=12, column=2, value="= ANNUAL SAVINGS IF YOU CANCEL").style = "th"
    cas = ws.cell(row=12, column=3, value='=SUMIF(SubAction,"Cancel",SubMonthlyRange)*12'); cas.style = "td"; cas.font = Font(bold=True, size=14, color=PRIMARY); cas.fill = fill(MINT_BG); cas.number_format = '"$"#,##0'
    ws.cell(row=14, column=2, value="= YOU KEEP (monthly)").style = "th"
    ck = ws.cell(row=14, column=3, value='=SubMonthly-SUMIF(SubAction,"Cancel",SubMonthlyRange)'); ck.style = "td"; ck.font = Font(bold=True, size=13, color=PRIMARY); ck.fill = fill(MINT_BG); ck.number_format = '"$"#,##0.00'
    cell_name(wb, "KeepMonthly", "Subscription Audit", "$C$14")
    ws.cell(row=16, column=2, value="Flag a subscription 'Cancel' and watch your annual savings climb.").style = "section_gold"


def build_bills(wb):
    ws = wb.create_sheet("Bills"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 14, 2])
    luxe_header(ws, "D", "🧾  BILLS",
                "Your recurring bills — utilities, insurance and more. The other half of what leaves every month.")
    table_headers(ws, 4, ["Bill", "Category", "Monthly"], start_col=2)
    start = L0
    for i, (bill, cat, monthly) in enumerate(BILLS):
        r = start + i
        ws.cell(row=r, column=2, value=bill).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        cm = ws.cell(row=r, column=4, value=monthly); cm.style = "input"; cm.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BILLS) - 1
    nrange(wb, "BillMonthlyRange", "Bills", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="MONTHLY BILLS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=4, value="=SUM(BillMonthlyRange)"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0'
    cell_name(wb, "BillMonthly", "Bills", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_cancelfinder(wb):
    ws = wb.create_sheet("Cancel Finder"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 14, 16, 2])
    luxe_header(ws, "E", "✂  CANCEL FINDER",
                "Every subscription you flagged 'Cancel' — the monthly and annual savings, ranked. Your hit list.")
    table_headers(ws, 4, ["Subscription", "Category", "Monthly", "Annual Save"], start_col=2)
    start = L0
    for i, (name, cat, bill, monthly, action) in enumerate([s for s in SUBS if s[4] == "Cancel"]):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        cm = ws.cell(row=r, column=4, value=monthly); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        ca = ws.cell(row=r, column=5, value=f"=D{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'; ca.font = Font(bold=True, color=PRIMARY)
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len([s for s in SUBS if s[4] == "Cancel"]) - 1
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL SAVINGS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=4, value="=CancelMonthly"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0.00'
    ca = ws.cell(row=tot, column=5, value="=CancelMonthly*12"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = '"$"#,##0'
    ws.cell(row=tot + 2, column=2, value="This list is your fastest money back — cancel from the top down.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_categories(wb):
    ws = wb.create_sheet("Categories"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 16, 16, 2])
    luxe_header(ws, "D", "🗂  CATEGORIES",
                "Where your subscription money goes — spend by category, so you can see the biggest bucket at a glance.")
    table_headers(ws, 4, ["Category", "Monthly", "Annual"], start_col=2)
    start = L0
    for i, cat in enumerate(CAT_ROLLUP):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cm = ws.cell(row=r, column=3, value=f'=SUMIF(SubCategory,B{r},SubMonthlyRange)'); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        ca = ws.cell(row=r, column=4, value=f"=C{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CAT_ROLLUP) - 1
    ws.conditional_formatting.add(f"C{start}:C{end}", DataBarRule(start_type="min", end_type="max", color=HIGHLIGHT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL CATEGORIES").style = "th"
    cm = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0.00'
    ca = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


def build_renewals(wb):
    ws, start, end = build_log(
        wb, "Renewals", "📅", "RENEWALS",
        "The big annual charges and when they hit — so a $149 renewal is never a surprise on your statement.",
        ["Service", "Renews", "Amount"],
        RENEWALS, [2, 24, 16, 16, 2], text_left={2, 3}, money={4}, reserved=24, start_col=2)
    nrange(wb, "RenewAmt", "Renewals", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ANNUAL RENEWALS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(RenewAmt)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'


def build_trials(wb):
    ws = wb.create_sheet("Free Trials"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 16, 2])
    luxe_header(ws, "D", "⏳  FREE TRIALS",
                "Every trial and the date to cancel by — the calendar reminder that saves you from the auto-charge.")
    table_headers(ws, 4, ["Service", "Started", "Cancel By"], start_col=2)
    start = L0
    for i, (svc, started, cancelby) in enumerate(TRIALS):
        r = start + i
        ws.cell(row=r, column=2, value=svc).style = "td_left"
        ws.cell(row=r, column=3, value=started).style = "td"
        cc = ws.cell(row=r, column=4, value=cancelby); cc.style = "input"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(TRIALS) - 1
    nrange(wb, "TrialName", "Free Trials", "B", start, end)
    nrange(wb, "TrialCancelBy", "Free Trials", "D", start, end)
    ws.cell(row=end + 2, column=2, value="Trials tracked").style = "field_label"
    ct = ws.cell(row=end + 2, column=4, value="=COUNTA(TrialName)"); ct.style = "field_value"; ct.number_format = "#,##0"
    cell_name(wb, "TrialCount", "Free Trials", f"$D${end+2}")
    ws.freeze_panes = "A5"


def build_hikes(wb):
    ws = wb.create_sheet("Price Hikes"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 16, 2])
    luxe_header(ws, "E", "📈  PRICE HIKES",
                "Every stealth price increase — old vs new, and what the hikes add per year. The creep, caught.")
    table_headers(ws, 4, ["Service", "Was", "Now", "+ / Year"], start_col=2)
    start = L0
    for i, (svc, old, new) in enumerate(HIKES):
        r = start + i
        ws.cell(row=r, column=2, value=svc).style = "td_left"
        co = ws.cell(row=r, column=3, value=old); co.style = "input"; co.number_format = '"$"#,##0.00'
        cn = ws.cell(row=r, column=4, value=new); cn.style = "input"; cn.number_format = '"$"#,##0.00'
        cy = ws.cell(row=r, column=5, value=f"=(D{r}-C{r})*12"); cy.style = "td"; cy.number_format = '"$"#,##0'; cy.font = Font(bold=True, color=DANGER)
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(HIKES) - 1
    tot = end + 1
    ws.cell(row=tot, column=2, value="HIKES ADD / YEAR").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cy = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cy.style = "td"; cy.font = Font(bold=True, color=DANGER); cy.fill = fill(SURFACE); cy.number_format = '"$"#,##0'
    cell_name(wb, "HikeTotal", "Price Hikes", f"$E${tot}")
    ws.freeze_panes = "A5"


def build_savings(wb):
    ws, start, end = build_log(
        wb, "Savings", "🌿", "SAVINGS LOG",
        "What you've already cut and when — proof the audit pays off, month after month.",
        ["Cut item", "Monthly Saved", "Date"],
        SAVINGS, [2, 26, 16, 14, 2], text_left={2, 4}, money={3}, reserved=24, start_col=2)
    nrange(wb, "CutSaved", "Savings", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALREADY CUT (mo)").style = "th"
    c = ws.cell(row=tot, column=3, value="=SUM(CutSaved)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)


def build_negotiation(wb):
    ws, start, end = build_log(
        wb, "Negotiation", "📞", "NEGOTIATION LIST",
        "The bills worth a phone call — current vs target, and whether you've called. Ten minutes can save $20 a month.",
        ["Bill", "Provider", "Current", "Target", "Called?"],
        NEGOTIATION, [2, 20, 18, 14, 14, 14, 2], text_left={2, 3}, money={4, 5}, reserved=22, start_col=2,
        validations=[("F", "YesNoList")])
    nrange(wb, "NegCurrent", "Negotiation", "D", start, end)
    nrange(wb, "NegTarget", "Negotiation", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="POTENTIAL / MONTH").style = "th"
    for c in (3,):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cc = ws.cell(row=tot, column=4, value="=SUM(NegCurrent)-SUM(NegTarget)"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(MINT_BG); cc.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "📉  MONTHLY SUMMARY",
                "Your recurring spend, month by month — the line that should go DOWN as you cut.")
    ws.cell(row=5, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 6, ["Month", "Recurring"], start_col=2)
    ts = 7
    for i, (m, sp) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sp); cs.style = "input"; cs.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "SpendTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Recurring by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🔍  SUBSCRIPTION & BILLS AUDIT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Recurring spend, cancel savings & an Audit Score — every charge you pay, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MONTHLY SUBS", "=SubMonthly", "money2"),
        ("ANNUAL SUBS", "=SubMonthly*12", "money"),
        ("SUBSCRIPTIONS", "=SubCount", "num"),
        ("AVG / SUB", "=IFERROR(SubMonthly/SubCount,0)", "money2"),
        ("FLAGGED", "=CancelCount", "num"),
        ("MONTHLY SAVINGS", "=CancelMonthly", "money2"),
    ]
    row2 = [
        ("ANNUAL SAVINGS", "=CancelMonthly*12", "money"),
        ("MONTHLY BILLS", "=BillMonthly", "money"),
        ("KEEP MONTHLY", "=KeepMonthly", "money2"),
        ("HIKES ADD / YR", "=HikeTotal", "money"),
        ("TRIALS ENDING", "=TrialCount", "num"),
        ("AUDIT SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "AUDIT HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "RECURRING BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Everything reviewed", '=IFERROR((COUNTIF(SubAction,"Keep")+COUNTIF(SubAction,"Cancel"))/SubCount,0)'),
        ("Cutting waste", "=IFERROR(MIN(CancelMonthly/SaveGoal,1),0)"),
        ("Every sub categorized", "=IFERROR(COUNTA(SubCategory)/SubCount,0)"),
        ("Under budget", "=IF(SubMonthly<=SubBudget,1,IFERROR(SubBudget/SubMonthly,0))"),
        ("Free trials tracked", "=IFERROR(MIN(COUNTA(TrialCancelBy)/TrialCount,1),0)"),
        ("Billed annually where cheaper", "=IFERROR(MIN(AnnualBilled/AnnualGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Recurring by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=7, max_row=6 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=7, max_row=6 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Subscription & Bills Audit Command Center™ — see every charge, and cut the ones you forgot.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_subscriptions(wb); build_audit(wb)
    build_bills(wb); build_cancelfinder(wb); build_categories(wb); build_renewals(wb)
    build_trials(wb); build_hikes(wb); build_savings(wb); build_negotiation(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Subscription Audit", "Subscriptions", "Bills", "Cancel Finder", "Categories",
             "Renewals", "Free Trials", "Price Hikes", "Savings", "Negotiation", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Subscription_Audit_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
