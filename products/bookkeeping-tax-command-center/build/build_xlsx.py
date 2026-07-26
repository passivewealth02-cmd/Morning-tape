"""Build Small Business Bookkeeping & Tax Command Center™ — The Books & Tax Operating System.

14 tabs · a premium small-business bookkeeping & tax operating system in Google Sheets &
Excel. Dashboard, a Schedule C P&L engine (revenue − COGS − expenses → net profit, then
self-employment + income tax → your quarterly payment), income, expenses, COGS &
inventory, mileage, sales tax, quarterly taxes, categories, invoices, reconciliation and
a monthly summary — one dashboard. Know your profit, and never be surprised by a tax
bill again.

Run: python3 build_xlsx.py   ->  ../Bookkeeping_Tax_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
PAYSTATUS = ["Paid", "Sent", "Overdue", "Draft"]
INCOME_SRC = ["Product sales", "Services", "Wholesale", "Online", "Other"]
EXP_CAT = ["Advertising", "Supplies", "Software", "Insurance", "Rent", "Utilities",
           "Meals", "Travel", "Fees", "Professional", "Other"]

# Settings / assumptions
MILEAGE_RATE = 0.70
SE_TAX_RATE = 0.153
SE_BASE_PCT = 0.9235
INCOME_TAX_RATE = 0.12
PROFIT_GOAL = 30000
MARGIN_GOAL = 0.30
RECONCILE_GOAL = 6
RECEIPTS_PCT = 0.40
MONTHS_RECONCILED = 6

# Income: (month, source, amount)
INCOME = [
    ("Feb", "Product sales", 12000), ("Mar", "Product sales", 14000), ("Apr", "Services", 16000),
    ("May", "Product sales", 17000), ("Jun", "Wholesale", 18000), ("Jul", "Product sales", 19000),
]

# Expenses (annual, by category): (category, amount)
EXPENSES = [
    ("Advertising", 4200), ("Supplies", 2400), ("Software", 1800), ("Insurance", 1500),
    ("Rent / studio", 3600), ("Utilities & phone", 1200), ("Meals (50%)", 600),
    ("Mileage & travel", 2800), ("Fees & merchant", 700), ("Professional (CPA)", 400),
]

# COGS & inventory: (item, units, unit cost, total)
COGS_ITEMS = [
    ("Raw materials", 1200, 9.00, 10800), ("Packaging", 1200, 2.50, 3000),
    ("Shipping supplies", 1200, 1.50, 1800), ("Wholesale goods", 600, 15.00, 9000),
    ("Labels & inserts", 1200, 0.75, 900), ("Freight in", 1, 3300.00, 3300),
]

# Mileage log: (date, purpose, miles)
MILEAGE = [
    ("Feb 12", "Supply run", 320), ("Mar 8", "Client meeting", 480), ("Apr 3", "Craft fair", 640),
    ("May 19", "Post office runs", 560), ("Jun 7", "Wholesale delivery", 900), ("Jul 15", "Trade show", 1100),
]

# Sales tax: (period, collected, remitted)
SALES_TAX = [
    ("Q1", 600, 600), ("Q2", 900, 900), ("Q3", 600, 300), ("Q4", 300, 0),
]

# Quarterly taxes: (quarter, due date, paid?)
QUARTERS = [
    ("Q1", "Apr 15", "Yes"), ("Q2", "Jun 15", "Yes"), ("Q3", "Sep 15", "No"), ("Q4", "Jan 15", "No"),
]

# Categories / Schedule C line map: (category, schedule C line)
SCHED_C = [
    ("Advertising", "Line 8"), ("Car & truck (mileage)", "Line 9"), ("Insurance", "Line 15"),
    ("Legal & professional", "Line 17"), ("Office expense", "Line 18"), ("Rent", "Line 20b"),
    ("Supplies", "Line 22"), ("Travel", "Line 24a"), ("Meals (50%)", "Line 24b"),
    ("Utilities", "Line 25"), ("Other expenses", "Line 27a"),
]

# Invoices: (invoice, client, amount, status)
INVOICES = [
    ("#1041", "Rowan Retail", 4200, "Paid"), ("#1042", "Birch Boutique", 3600, "Paid"),
    ("#1043", "Cedar Co.", 2800, "Sent"), ("#1044", "Alder Shop", 1900, "Sent"),
    ("#1045", "Maple Market", 1500, "Overdue"),
]

# Reconciliation: (month, bank matched?, receipts?)
RECONCILE = [
    ("Feb", "Yes", "Yes"), ("Mar", "Yes", "Yes"), ("Apr", "Yes", "No"),
    ("May", "Yes", "No"), ("Jun", "Yes", "No"), ("Jul", "Yes", "No"),
]

# Monthly summary trend: (month, net profit)
MONTHS = [("Feb", 6000), ("Mar", 7000), ("Apr", 8000), ("May", 8500), ("Jun", 9000), ("Jul", 9500)]

GROSS_REVENUE = 96000
COGS_TOTAL = 28800
EXP_TOTAL = 19200

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 20, 3] + [18] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your rates & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR RATES & GOALS", "section")
    controls = [
        ("Business name", "Quill & Ledger", None, "BizName"),
        ("Owner", "Morgan", None, "Owner"),
        ("IRS mileage rate ($/mi)", MILEAGE_RATE, '"$"#,##0.00', "MileageRate"),
        ("Self-employment tax rate", SE_TAX_RATE, "0.0%", "SETaxRate"),
        ("SE taxable base %", SE_BASE_PCT, "0.00%", "SEBasePct"),
        ("Income tax rate (est.)", INCOME_TAX_RATE, "0%", "IncomeTaxRate"),
        ("Net-profit goal", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Net-margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Months-to-reconcile goal", RECONCILE_GOAL, "0", "ReconcileGoal"),
        ("Receipts attached %", RECEIPTS_PCT, "0%", "ReceiptsPct"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Income source", INCOME_SRC, "IncomeSrcList"), ("F", "Expense category", EXP_CAT, "ExpCatList"),
             ("G", "Invoice status", PAYSTATUS, "PayStatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  📗  SMALL BUSINESS BOOKKEEPING & TAX COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know your profit, and never be surprised by a tax bill again.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE BOOKS & TAX, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Most small businesses find out what they owe in April — far too late to do anything about it. "
                      "This fixes that. A Schedule C engine takes your revenue, subtracts your cost of goods and your "
                      "expenses to give your real net profit, then calculates your self-employment tax, your income tax "
                      "and the exact quarterly payment to send. Track income, expenses, COGS, mileage, sales tax and "
                      "invoices — all in ONE premium Google Sheets & Excel system your accountant will thank you for.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your mileage rate & tax rates.",
             "2.  Log your Income and your Expenses by category.",
             "3.  Enter COGS & Inventory — the cost of what you sold.",
             "4.  Read the Schedule C P&L: your real net profit.",
             "5.  Track Mileage, Sales Tax and your Quarterly payments.",
             "6.  Check the Dashboard: profit, tax owed & a Books Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional business (Quill & Ledger, owner Morgan) is included so you can see how it "
               "all connects — just type over it with your own numbers. Your net profit and the quarterly tax payment "
               "it implies are the two numbers that decide whether tax season is calm or brutal, and they roll into a "
               "live Books Score. Expense categories are mapped to real Schedule C lines so handing this to a CPA is "
               "painless. Twelve matching printable pages (P&L worksheet, expense log, mileage log, receipt tracker & "
               "more) are included. This is a bookkeeping & organizing tool, not tax, legal or accounting advice — "
               "confirm every figure with your own tax professional.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Books kept monthly beat books panicked in April. Ten minutes a week is all it takes.", "section_gold")


# ===========================================================================
def build_income(wb):
    ws = wb.create_sheet("Income"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 22, 18, 2])
    luxe_header(ws, "D", "💵  INCOME",
                "Every dollar in, by month and source — this is your gross revenue, the top of the Schedule C.")
    table_headers(ws, 4, ["Month", "Source", "Amount"], start_col=2)
    start = L0
    for i, (month, src, amt) in enumerate(INCOME):
        r = start + i
        ws.cell(row=r, column=2, value=month).style = "td_left"
        ws.cell(row=r, column=3, value=src).style = "td"
        ca = ws.cell(row=r, column=4, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INCOME) - 1
    nrange(wb, "IncomeAmt", "Income", "D", start, end)
    add_dv(ws, f"C{start}:C{end}", "IncomeSrcList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="GROSS REVENUE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cg = ws.cell(row=tot, column=4, value="=SUM(IncomeAmt)"); cg.style = "td"; cg.font = Font(bold=True, size=12, color=PRIMARY); cg.fill = fill(MINT_BG); cg.number_format = '"$"#,##0'
    cell_name(wb, "GrossRevenue", "Income", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 14, 2])
    luxe_header(ws, "D", "🧾  EXPENSES",
                "Every deductible cost by category — each one mapped to a real Schedule C line for your CPA.")
    table_headers(ws, 4, ["Category", "Amount", "% of Rev"], start_col=2)
    start = L0
    for i, (cat, amt) in enumerate(EXPENSES):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/GrossRevenue,0)"); cp.style = "td"; cp.number_format = "0.0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EXPENSES) - 1
    nrange(wb, "ExpCat", "Expenses", "B", start, end)
    nrange(wb, "ExpAmt", "Expenses", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", DataBarRule(start_type="min", end_type="max", color=GOLD_LT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, size=12, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cep = ws.cell(row=tot, column=4, value=f"=IFERROR(C{tot}/GrossRevenue,0)"); cep.style = "td"; cep.font = Font(bold=True, color=PRIMARY); cep.fill = fill(MINT_BG); cep.number_format = "0.0%"
    cell_name(wb, "ExpTotal", "Expenses", f"$C${tot}")
    ws.cell(row=tot + 2, column=2, value="Categories used").style = "field_label"
    cc = ws.cell(row=tot + 2, column=3, value="=COUNTA(ExpCat)"); cc.style = "field_value"; cc.number_format = "#,##0"
    cell_name(wb, "CategoryCount", "Expenses", f"$C${tot+2}")
    ws.freeze_panes = "A5"


def build_cogs(wb):
    ws = wb.create_sheet("COGS & Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 16, 2])
    luxe_header(ws, "E", "📦  COGS & INVENTORY",
                "The direct cost of what you actually sold — the line that separates revenue from real gross profit.")
    table_headers(ws, 4, ["Item", "Units", "Unit Cost", "Total"], start_col=2)
    start = L0
    for i, (item, units, unit_cost, total) in enumerate(COGS_ITEMS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cu = ws.cell(row=r, column=3, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cc = ws.cell(row=r, column=4, value=unit_cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        ct = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); ct.style = "td"; ct.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(COGS_ITEMS) - 1
    nrange(wb, "COGSRows", "COGS & Inventory", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL COGS").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=5, value="=SUM(COGSRows)"); ct.style = "td"; ct.font = Font(bold=True, size=12, color=DANGER); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "COGSTotal", "COGS & Inventory", f"$E${tot}")
    ws.cell(row=tot + 2, column=2, value="COGS % of revenue").style = "field_label"
    cp = ws.cell(row=tot + 2, column=5, value="=IFERROR(COGSTotal/GrossRevenue,0)"); cp.style = "field_value"; cp.number_format = "0%"; cp.fill = fill(MINT_BG)
    ws.freeze_panes = "A5"


def build_schedulec(wb):
    ws = wb.create_sheet("Schedule C P&L"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 18, 2])
    luxe_header(ws, "C", "🧮  SCHEDULE C P&L",
                "Revenue − COGS − expenses = your real net profit, then the exact tax it creates and what to send each quarter.")
    ws.cell(row=5, column=2, value="THE PROFIT & LOSS").style = "section_gold"
    ws.cell(row=6, column=2, value="Gross revenue (Line 1)").style = "field_label"
    cg = ws.cell(row=6, column=3, value="=GrossRevenue"); cg.style = "field_value"; cg.number_format = '"$"#,##0'; cg.fill = fill(MINT_BG)
    ws.cell(row=7, column=2, value="− Cost of goods sold (Line 4)").style = "field_label"
    cc = ws.cell(row=7, column=3, value="=COGSTotal"); cc.style = "field_value"; cc.number_format = '"$"#,##0'
    ws.cell(row=8, column=2, value="= GROSS PROFIT (Line 7)").style = "th"
    cgp = ws.cell(row=8, column=3, value="=GrossRevenue-COGSTotal"); cgp.style = "td"; cgp.font = Font(bold=True, size=12, color=PRIMARY); cgp.fill = fill(MINT_BG); cgp.number_format = '"$"#,##0'
    cell_name(wb, "GrossProfit", "Schedule C P&L", "$C$8")
    ws.cell(row=9, column=2, value="− Total expenses (Line 28)").style = "field_label"
    ce = ws.cell(row=9, column=3, value="=ExpTotal"); ce.style = "field_value"; ce.number_format = '"$"#,##0'
    ws.cell(row=10, column=2, value="= NET PROFIT (Line 31)").style = "th"
    cn = ws.cell(row=10, column=3, value="=GrossRevenue-COGSTotal-ExpTotal"); cn.style = "td"; cn.font = Font(bold=True, size=14, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetProfit", "Schedule C P&L", "$C$10")
    ws.cell(row=11, column=2, value="Net margin").style = "field_label"
    cm = ws.cell(row=11, column=3, value="=IFERROR(NetProfit/GrossRevenue,0)"); cm.style = "field_value"; cm.number_format = "0%"; cm.fill = fill(MINT_BG)
    cell_name(wb, "NetMargin", "Schedule C P&L", "$C$11")
    # Tax block
    ws.cell(row=13, column=2, value="THE TAX IT CREATES").style = "section_gold"
    ws.cell(row=14, column=2, value="SE taxable base (92.35% of net)").style = "field_label"
    cb = ws.cell(row=14, column=3, value="=NetProfit*SEBasePct"); cb.style = "field_value"; cb.number_format = '"$"#,##0'
    ws.cell(row=15, column=2, value="× Self-employment tax (15.3%)").style = "th"
    cs = ws.cell(row=15, column=3, value="=NetProfit*SEBasePct*SETaxRate"); cs.style = "td"; cs.font = Font(bold=True, size=12, color=DANGER); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cell_name(wb, "SETax", "Schedule C P&L", "$C$15")
    ws.cell(row=16, column=2, value="Taxable after ½ SE deduction").style = "field_label"
    ct = ws.cell(row=16, column=3, value="=NetProfit-SETax/2"); ct.style = "field_value"; ct.number_format = '"$"#,##0'
    ws.cell(row=17, column=2, value="× Income tax (your rate)").style = "th"
    ci = ws.cell(row=17, column=3, value="=(NetProfit-SETax/2)*IncomeTaxRate"); ci.style = "td"; ci.font = Font(bold=True, size=12, color=DANGER); ci.fill = fill(SURFACE); ci.number_format = '"$"#,##0'
    cell_name(wb, "IncomeTax", "Schedule C P&L", "$C$17")
    ws.cell(row=18, column=2, value="= TOTAL TAX OWED").style = "th"
    cto = ws.cell(row=18, column=3, value="=SETax+IncomeTax"); cto.style = "td"; cto.font = Font(bold=True, size=14, color=DANGER); cto.fill = fill(WARN_BG); cto.number_format = '"$"#,##0'
    cell_name(wb, "TotalTax", "Schedule C P&L", "$C$18")
    ws.cell(row=19, column=2, value="= SEND EACH QUARTER").style = "th"
    cq = ws.cell(row=19, column=3, value="=(SETax+IncomeTax)/4"); cq.style = "td"; cq.font = Font(bold=True, size=14, color=PRIMARY); cq.fill = fill(MINT_BG); cq.number_format = '"$"#,##0'
    cell_name(wb, "QuarterlyTax", "Schedule C P&L", "$C$19")
    ws.cell(row=20, column=2, value="Effective tax rate on profit").style = "field_label"
    cer = ws.cell(row=20, column=3, value="=IFERROR(TotalTax/NetProfit,0)"); cer.style = "field_value"; cer.number_format = "0.0%"; cer.fill = fill(MINT_BG)
    cell_name(wb, "EffectiveRate", "Schedule C P&L", "$C$20")
    ws.cell(row=22, column=2, value="Set aside the quarterly figure the day money lands — never scramble in April.").style = "section_gold"


def build_mileage(wb):
    ws = wb.create_sheet("Mileage"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 26, 14, 16, 2])
    luxe_header(ws, "E", "🚗  MILEAGE",
                "Every business mile — at the IRS rate this is often the biggest deduction a small business forgets.")
    table_headers(ws, 4, ["Date", "Purpose", "Miles", "Deduction"], start_col=2)
    start = L0
    for i, (date, purpose, miles) in enumerate(MILEAGE):
        r = start + i
        ws.cell(row=r, column=2, value=date).style = "td_left"
        ws.cell(row=r, column=3, value=purpose).style = "td_left"
        cm = ws.cell(row=r, column=4, value=miles); cm.style = "input"; cm.number_format = "#,##0"
        cd = ws.cell(row=r, column=5, value=f"=D{r}*MileageRate"); cd.style = "td"; cd.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MILEAGE) - 1
    nrange(wb, "MilesRange", "Mileage", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL MILES").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=4, value="=SUM(MilesRange)"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = "#,##0"
    cd = ws.cell(row=tot, column=5, value="=SUM(MilesRange)*MileageRate"); cd.style = "td"; cd.font = Font(bold=True, size=12, color=PRIMARY); cd.fill = fill(MINT_BG); cd.number_format = '"$"#,##0'
    cell_name(wb, "MilesDriven", "Mileage", f"$D${tot}")
    cell_name(wb, "MileageDeduction", "Mileage", f"$E${tot}")
    ws.freeze_panes = "A5"


def build_salestax(wb):
    ws = wb.create_sheet("Sales Tax"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 18, 18, 16, 2])
    luxe_header(ws, "E", "🏛  SALES TAX",
                "What you collected versus what you've remitted — the money that was never yours to keep.")
    table_headers(ws, 4, ["Period", "Collected", "Remitted", "Owed"], start_col=2)
    start = L0
    for i, (period, coll, rem) in enumerate(SALES_TAX):
        r = start + i
        ws.cell(row=r, column=2, value=period).style = "td_left"
        cc = ws.cell(row=r, column=3, value=coll); cc.style = "input"; cc.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=rem); cr.style = "input"; cr.number_format = '"$"#,##0'
        co = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); co.style = "td"; co.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SALES_TAX) - 1
    nrange(wb, "STCollected", "Sales Tax", "C", start, end)
    nrange(wb, "STRemitted", "Sales Tax", "D", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=fill(WARN_BG)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="YEAR TO DATE").style = "th"
    cc = ws.cell(row=tot, column=3, value="=SUM(STCollected)"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0'
    cr = ws.cell(row=tot, column=4, value="=SUM(STRemitted)"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = '"$"#,##0'
    co = ws.cell(row=tot, column=5, value="=SUM(STCollected)-SUM(STRemitted)"); co.style = "td"; co.font = Font(bold=True, size=12, color=DANGER); co.fill = fill(WARN_BG); co.number_format = '"$"#,##0'
    cell_name(wb, "SalesTaxCollected", "Sales Tax", f"$C${tot}")
    cell_name(wb, "SalesTaxOwed", "Sales Tax", f"$E${tot}")
    ws.cell(row=tot + 2, column=2, value="Sales tax you collected is held in trust — remit it, never spend it.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_quarterly(wb):
    ws = wb.create_sheet("Quarterly Taxes"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 18, 18, 14, 2])
    luxe_header(ws, "E", "📆  QUARTERLY TAXES",
                "The four dates that matter and what to send on each — pay these and April is a non-event.")
    ws.cell(row=5, column=2, value="Send each quarter").style = "field_label"
    cq = ws.cell(row=5, column=4, value="=QuarterlyTax"); cq.style = "field_value"; cq.number_format = '"$"#,##0'; cq.fill = fill(MINT_BG)
    table_headers(ws, 7, ["Quarter", "Due Date", "Amount", "Paid?"], start_col=2)
    start = 8
    for i, (q, due, paid) in enumerate(QUARTERS):
        r = start + i
        ws.cell(row=r, column=2, value=q).style = "td_left"
        ws.cell(row=r, column=3, value=due).style = "td"
        ca = ws.cell(row=r, column=4, value="=QuarterlyTax"); ca.style = "td"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=paid).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(QUARTERS) - 1
    nrange(wb, "QPaid", "Quarterly Taxes", "E", start, end)
    add_dv(ws, f"E{start}:E{end}", "YesNoList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="PAID SO FAR").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cp = ws.cell(row=tot, column=4, value='=COUNTIF(QPaid,"Yes")*QuarterlyTax'); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = '"$"#,##0'
    cn = ws.cell(row=tot, column=5, value='=COUNTIF(QPaid,"Yes")&" / 4"'); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE)
    ws.freeze_panes = "A8"


def build_categories(wb):
    ws = wb.create_sheet("Schedule C Lines"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "🗂  SCHEDULE C LINES",
                "Every expense category mapped to its real Schedule C line — hand this to a CPA and they'll smile.")
    table_headers(ws, 4, ["Category", "Schedule C Line"], start_col=2)
    start = L0
    for i, (cat, line) in enumerate(SCHED_C):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        ws.cell(row=r, column=3, value=line).style = "td"
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SCHED_C) - 1
    ws.cell(row=end + 2, column=2, value="Categories mapped").style = "field_label"
    cm = ws.cell(row=end + 2, column=3, value=f"=COUNTA(B{start}:B{end})"); cm.style = "field_value"; cm.number_format = "#,##0"
    ws.freeze_panes = "A5"


def build_invoices(wb):
    ws, start, end = build_log(
        wb, "Invoices", "📨", "INVOICES",
        "Who owes you what — every invoice, its status and the cash still outstanding.",
        ["Invoice", "Client", "Amount", "Status"],
        INVOICES, [2, 14, 22, 16, 14, 2], text_left={2, 3}, money={4}, reserved=28, start_col=2,
        validations=[("E", "PayStatusList")])
    nrange(wb, "InvAmt", "Invoices", "D", start, end)
    nrange(wb, "InvStatus", "Invoices", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="INVOICED").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ci = ws.cell(row=tot, column=4, value="=SUM(InvAmt)"); ci.style = "td"; ci.font = Font(bold=True, color=PRIMARY); ci.fill = fill(SURFACE); ci.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    ws.cell(row=tot + 1, column=2, value="Outstanding (not paid)").style = "field_label"
    co = ws.cell(row=tot + 1, column=4, value='=SUM(InvAmt)-SUMIF(InvStatus,"Paid",InvAmt)'); co.style = "field_value"; co.number_format = '"$"#,##0'; co.fill = fill(WARN_BG)
    cell_name(wb, "Outstanding", "Invoices", f"$D${tot+1}")


def build_reconcile(wb):
    ws = wb.create_sheet("Reconciliation"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 20, 20, 2])
    luxe_header(ws, "D", "✅  RECONCILIATION",
                "Month by month — bank matched and receipts attached. Ten minutes a month keeps an audit painless.")
    table_headers(ws, 4, ["Month", "Bank Matched?", "Receipts Filed?"], start_col=2)
    start = L0
    for i, (month, matched, receipts) in enumerate(RECONCILE):
        r = start + i
        ws.cell(row=r, column=2, value=month).style = "td_left"
        ws.cell(row=r, column=3, value=matched).style = "td"
        ws.cell(row=r, column=4, value=receipts).style = "td"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(RECONCILE) - 1
    nrange(wb, "BankMatched", "Reconciliation", "C", start, end)
    nrange(wb, "ReceiptsFiled", "Reconciliation", "D", start, end)
    add_dv(ws, f"C{start}:C{end}", "YesNoList")
    add_dv(ws, f"D{start}:D{end}", "YesNoList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="MONTHS DONE").style = "th"
    cm = ws.cell(row=tot, column=3, value='=COUNTIF(BankMatched,"Yes")'); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = "#,##0"
    cr = ws.cell(row=tot, column=4, value='=COUNTIF(ReceiptsFiled,"Yes")'); cr.style = "td"; cr.font = Font(bold=True, color=ACCENT); cr.fill = fill(SURFACE); cr.number_format = "#,##0"
    cell_name(wb, "MonthsReconciled", "Reconciliation", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "Net profit month by month — the line that tells you whether the business is actually working.")
    ws.cell(row=5, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 6, ["Month", "Net Profit"], start_col=2)
    ts = 7
    for i, (m, np_) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cn = ws.cell(row=r, column=3, value=np_); cn.style = "input"; cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "ProfitTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Net Profit by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  📗  SMALL BUSINESS BOOKKEEPING & TAX COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Profit, tax owed, quarterly payment & a Books Score — your whole business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("GROSS REVENUE", "=GrossRevenue", "money"),
        ("COGS", "=COGSTotal", "money"),
        ("GROSS PROFIT", "=GrossProfit", "money"),
        ("EXPENSES", "=ExpTotal", "money"),
        ("NET PROFIT", "=NetProfit", "money"),
        ("NET MARGIN", "=NetMargin", "pct"),
    ]
    row2 = [
        ("SE TAX", "=SETax", "money"),
        ("INCOME TAX", "=IncomeTax", "money"),
        ("TOTAL TAX", "=TotalTax", "money"),
        ("QUARTERLY", "=QuarterlyTax", "money"),
        ("MILEAGE DEDUCTION", "=MileageDeduction", "money"),
        ("BOOKS SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "BOOKS HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "NET PROFIT BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Profitable", "=IFERROR(MIN(NetProfit/ProfitGoal,1),0)"),
        ("Healthy net margin", "=IFERROR(MIN(NetMargin/MarginGoal,1),0)"),
        ("Tax set aside", "=IFERROR(MIN((QuarterlyTax*4)/TotalTax,1),0)"),
        ("Expenses categorized", "=IFERROR(MIN(CategoryCount/COUNTA(ExpCat),1),0)"),
        ("Books reconciled", "=IFERROR(MIN(MonthsReconciled/ReconcileGoal,1),0)"),
        ("Receipts attached", "=IFERROR(MIN(ReceiptsPct/1,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Net Profit by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=7, max_row=6 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=7, max_row=6 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Small Business Bookkeeping & Tax Command Center™ — know your profit, own your tax bill.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_income(wb); build_cogs(wb)
    build_expenses(wb); build_schedulec(wb); build_mileage(wb); build_salestax(wb)
    build_quarterly(wb); build_categories(wb); build_invoices(wb); build_reconcile(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Schedule C P&L", "Income", "COGS & Inventory", "Expenses",
             "Mileage", "Sales Tax", "Quarterly Taxes", "Schedule C Lines", "Invoices",
             "Reconciliation", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Bookkeeping_Tax_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
