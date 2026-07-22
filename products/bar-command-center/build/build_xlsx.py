"""Build Bar & Pub Command Center™ — The Complete Bar Business System.

14 tabs · a premium bar operating system in Google Sheets & Excel. Dashboard, a
pour-cost engine, a drink menu with pour-cost %, keg & draft yield, inventory
variance, liquor inventory & par, happy-hour pricing, weekly sales, a waste &
spill log, ordering, cash & tips and events & tabs — one dashboard. Cost every
pour, kill your shrinkage & pour more profit.

Run: python3 build_xlsx.py   ->  ../Bar_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["oz", "ml", "dash", "each", "pint", "bottle", "case", "keg"]
CATS = ["Beer", "Draft", "Wine", "Cocktail", "Spirit", "Shot", "N/A"]

TARGET_POUR = 0.22

# Pour cost engine — flagship Old Fashioned: (component, qty, unit, cost/unit)
OLDFASHIONED = [
    ("Bourbon", 2, "oz", 0.90),
    ("Simple syrup", 0.25, "oz", 0.10),
    ("Bitters", 3, "dash", 0.03),
    ("Orange garnish", 1, "each", 0.18),
    ("Ice / napkin", 1, "each", 0.105),
]

# Drink menu: (drink, category, pour cost, price, units/wk). Old Fashioned = OldFashionedCost.
DRINKS = [
    ("Draft Beer (pint)", "Draft", 0.95, 7.00, 420),
    ("Well Cocktail", "Cocktail", 1.40, 9.00, 260),
    ("Premium Cocktail", "Cocktail", 2.80, 14.00, 180),
    ("House Wine (glass)", "Wine", 1.60, 10.00, 150),
    ("Craft Draft", "Draft", 1.30, 8.00, 300),
    ("Well Shot", "Shot", 0.70, 5.00, 200),
    ("Old Fashioned", "Cocktail", None, 13.00, 140),
    ("Margarita", "Cocktail", 1.90, 11.00, 220),
]

# Keg & draft: (beer, keg cost, keg pints, price/pint)
KEGS = [
    ("House Lager (1/2 bbl)", 160, 124, 7.00),
    ("Craft IPA (1/2 bbl)", 220, 124, 8.00),
    ("Seasonal (1/6 bbl)", 95, 41, 8.00),
    ("Cider (1/2 bbl)", 180, 124, 8.00),
]

# Inventory variance: (category, theoretical usage, actual usage)
VARIANCE = [
    ("Beer / draft", 2400, 2520),
    ("Liquor / spirits", 1800, 1940),
    ("Wine", 900, 950),
]

# Liquor inventory & par: (item, par, on hand, unit)
INVENTORY = [
    ("Well vodka", 12, 5, "bottle"), ("Bourbon", 10, 4, "bottle"), ("Tequila", 8, 3, "bottle"),
    ("House lager keg", 4, 2, "keg"), ("Craft IPA keg", 3, 1, "keg"), ("House red wine", 24, 10, "bottle"),
    ("Triple sec", 6, 3, "bottle"), ("Garnishes", 20, 8, "each"),
]

# Happy hour: (drink, regular price, HH price, pour cost)
HAPPYHOUR = [
    ("Draft Beer", 7.00, 5.00, 0.95),
    ("Well Cocktail", 9.00, 6.00, 1.40),
    ("House Wine", 10.00, 7.00, 1.60),
    ("Well Shot", 5.00, 3.50, 0.70),
]

# Weekly sales: (day, sales, covers)
WEEKLY = [
    ("Monday", 1450, 95), ("Tuesday", 1680, 110), ("Wednesday", 2100, 140),
    ("Thursday", 2680, 178), ("Friday", 3850, 255), ("Saturday", 4180, 278), ("Sunday", 1000, 68),
]

# Waste & spill: (item, reason, cost)
WASTE = [
    ("Over-pours", "Free-pour drift", 145.00),
    ("Spilled / comped", "Service", 90.00),
    ("Broken bottles", "Handling", 62.00),
    ("Flat draft (line)", "Cleaning loss", 48.00),
]

# Ordering / suppliers: (item, supplier, par order, cost)
ORDERING = [
    ("House lager (kegs)", "Distributor A", 4, 640.00),
    ("Well spirits (case)", "Distributor B", 1, 210.00),
    ("House wine (case)", "Wine Co.", 2, 168.00),
    ("Bitters & mixers", "BarSupply", 1, 96.00),
    ("Garnishes & ice", "Local", 1, 84.00),
]

# Cash & tips (7 days): (day, cash, card, tips)
CASHTIPS = [
    ("Monday", 320, 1130, 210), ("Tuesday", 360, 1320, 245), ("Wednesday", 440, 1660, 310),
    ("Thursday", 560, 2120, 400), ("Friday", 820, 3030, 580), ("Saturday", 900, 3280, 630),
    ("Sunday", 220, 780, 150),
]

# Events & big tabs: (event, days out, deposit, min spend, status)
EVENTS = [
    ("Trivia Night (weekly)", 3, 0, 0, "Recurring"),
    ("Birthday — party of 20", 6, 100, 500, "Confirmed"),
    ("Corporate happy hour", 11, 200, 1200, "Confirmed"),
    ("Live band Saturday", 4, 300, 0, "Confirmed"),
    ("Private buyout", 18, 500, 3000, "Deposit"),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your bar, targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR BAR", "section")
    controls = [
        ("Bar name", "The Oak & Iron", None, "Bar"),
        ("Owner", "Reed", None, "Owner"),
        ("Target pour-cost %", TARGET_POUR, "0%", "TargetPour"),
        ("Margin goal ($/drink)", 6, '"$"#,##0', "MarginGoal"),
        ("Variance limit %", 0.10, "0%", "VarLimit"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Category", CATS, "CatList"),
             ("G", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🍺  BAR & PUB COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every pour, kill your shrinkage & pour more profit.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE BAR, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Cost every drink to the pour — liquor, mixers & garnish — and see each drink's pour-cost % and "
                      "margin. Track keg yield and profit per keg, catch shrinkage with an inventory-variance check "
                      "(theoretical vs actual usage), price happy hour so it still makes money, and run liquor "
                      "inventory, waste, sales, ordering, cash, tips and events — all in ONE premium Google Sheets & "
                      "Excel system built for bars & pubs.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your bar name & target pour-cost %.",
             "2.  Cost a drink to the pour; the Drink Menu shows pour-cost % & margin.",
             "3.  Enter your Kegs — yield & profit per keg calculate automatically.",
             "4.  Run the Inventory Variance check — theoretical vs actual usage.",
             "5.  Set Happy Hour prices that still profit; log Waste, Sales & Cash.",
             "6.  Check the Dashboard: sales, pour cost, variance & a Bar Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional pub (The Oak & Iron) is included so you can see how it all connects — just "
               "type over it with your own drinks and numbers. Pour cost and inventory variance (shrinkage) are the "
               "two numbers that decide whether a bar makes money, and they roll into a live Bar Score. Twelve "
               "matching printable pages (pour-cost card, keg log, variance sheet, waste log & more) are included for "
               "behind the bar. This is a business tool, not financial or legal advice — confirm figures with your books.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Pour cost & shrinkage make or break a bar — measure both, every week.", "section_gold")


# ===========================================================================
def build_pourcost(wb):
    ws = wb.create_sheet("Pour Cost"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 10, 10, 14, 14, 2])
    luxe_header(ws, "F", "🥃  POUR COST",
                "Cost a drink to the pour — spirit, mixer & garnish. The engine behind your menu.")
    merge_set(ws, "B5:C5", "DRINK", "section_gold")
    ws.cell(row=5, column=4, value="Old Fashioned").font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Component", "Qty", "Unit", "Cost/Unit", "Ext. Cost"], start_col=2)
    start = 7
    for i, (comp, qty, unit, cu) in enumerate(OLDFASHIONED):
        r = start + i
        ws.cell(row=r, column=2, value=comp).style = "td_left"
        cq = ws.cell(row=r, column=3, value=qty); cq.style = "input"; cq.number_format = "0.##"
        ws.cell(row=r, column=4, value=unit).style = "td"
        cc = ws.cell(row=r, column=5, value=cu); cc.style = "input"; cc.number_format = '"$"#,##0.000'
        ce = ws.cell(row=r, column=6, value=f"=C{r}*E{r}"); ce.style = "td"; ce.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(OLDFASHIONED) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="POUR COST").style = "th"
    for c in range(3, 6):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0.00'
    cell_name(wb, "OldFashionedCost", "Pour Cost", f"$F${tot}")
    ws.cell(row=tot + 2, column=2, value="Copy this build for every drink — swap the spirit, add mixers & garnish.").style = "section"


def build_menu(wb):
    ws = wb.create_sheet("Drink Menu"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 12, 12, 12, 10, 12, 12, 2])
    luxe_header(ws, "H", "🍸  DRINK MENU",
                "Every drink's pour cost, price, pour-cost % & margin — price for a healthy pour.")
    table_headers(ws, 4, ["Drink", "Category", "Pour Cost", "Price", "Units/wk", "Pour %", "Margin $"], start_col=2)
    start = L0
    for i, (item, cat, cost, price, units) in enumerate(DRINKS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        cc = ws.cell(row=r, column=4, value="=OldFashionedCost" if cost is None else cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cu = ws.cell(row=r, column=6, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cb = ws.cell(row=r, column=7, value=f"=IFERROR(D{r}/E{r},0)"); cb.style = "td"; cb.number_format = "0%"
        cm = ws.cell(row=r, column=8, value=f"=E{r}-D{r}"); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DRINKS) - 1
    nrange(wb, "DrinkItem", "Drink Menu", "B", start, end)
    nrange(wb, "DrinkCost", "Drink Menu", "D", start, end)
    nrange(wb, "DrinkPrice", "Drink Menu", "E", start, end)
    nrange(wb, "DrinkUnits", "Drink Menu", "F", start, end)
    nrange(wb, "DrinkMargin", "Drink Menu", "H", start, end)
    # weekly sales & revenue helper (col I) and pour% helper
    ws.cell(row=4, column=9, value="Wk Rev").style = "th"
    for i, (item, cat, cost, price, units) in enumerate(DRINKS):
        r = start + i
        cr = ws.cell(row=r, column=9, value=f"=E{r}*F{r}"); cr.style = "td"; cr.number_format = '"$"#,##0'
    nrange(wb, "DrinkRev", "Drink Menu", "I", start, end)
    br = end + 2
    ws.cell(row=br, column=2, value="Weekly drink sales").style = "field_label"
    c = ws.cell(row=br, column=5, value="=SUM(DrinkRev)"); c.style = "field_value"; c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    cell_name(wb, "WeeklySales", "Drink Menu", f"$E${br}")
    ws.cell(row=br + 1, column=2, value="Overall pour-cost %").style = "field_label"
    c2 = ws.cell(row=br + 1, column=5, value="=IFERROR(SUMPRODUCT(DrinkCost,DrinkUnits)/SUMPRODUCT(DrinkPrice,DrinkUnits),0)"); c2.style = "field_value"; c2.number_format = "0%"; c2.fill = fill(MINT_BG)
    cell_name(wb, "PourCostPct", "Drink Menu", f"$E${br+1}")
    ws.conditional_formatting.add(f"G{start}:G{end}",
        ColorScaleRule(start_type="num", start_value=0.10, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.22, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.35, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


def build_kegs(wb):
    ws = wb.create_sheet("Keg & Draft"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 12, 14, 14, 14, 2])
    luxe_header(ws, "G", "🛢  KEG & DRAFT",
                "Keg cost ÷ pints = cost per pint. Price it right and one keg is hundreds in profit.")
    table_headers(ws, 4, ["Keg", "Keg Cost", "Pints", "Cost/Pint", "Price/Pint", "Profit/Keg"], start_col=2)
    start = L0
    for i, (name, cost, pints, price) in enumerate(KEGS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=pints); cp.style = "input"; cp.number_format = "#,##0"
        ccp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); ccp.style = "td"; ccp.number_format = '"$"#,##0.00'
        cpp = ws.cell(row=r, column=6, value=price); cpp.style = "input"; cpp.number_format = '"$"#,##0.00'
        cpk = ws.cell(row=r, column=7, value=f"=(F{r}-E{r})*D{r}"); cpk.style = "td"; cpk.font = Font(bold=True, color=PRIMARY); cpk.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(KEGS) - 1
    nrange(wb, "KegProfit", "Keg & Draft", "G", start, end)
    ws.cell(row=end + 2, column=2, value="Avg profit / keg").style = "field_label"
    c = ws.cell(row=end + 2, column=7, value="=AVERAGE(KegProfit)"); c.style = "field_value"; c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    cell_name(wb, "AvgKegProfit", "Keg & Draft", f"$G${end+2}")
    ws.conditional_formatting.add(f"G{start}:G{end}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


def build_variance(wb):
    ws = wb.create_sheet("Inventory Variance"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 16, 14, 2])
    luxe_header(ws, "E", "🔍  INVENTORY VARIANCE",
                "Theoretical usage vs actual — the gap is shrinkage: over-pours, comps & walk-outs.")
    table_headers(ws, 4, ["Category", "Theoretical", "Actual", "Variance %"], start_col=2)
    start = L0
    for i, (cat, theo, actual) in enumerate(VARIANCE):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        ct = ws.cell(row=r, column=3, value=theo); ct.style = "input"; ct.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=actual); ca.style = "input"; ca.number_format = '"$"#,##0'
        cv = ws.cell(row=r, column=5, value=f"=IFERROR((D{r}-C{r})/C{r},0)"); cv.style = "td"; cv.number_format = "0.0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(VARIANCE) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    ct = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    ca = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = '"$"#,##0'
    cv = ws.cell(row=tot, column=5, value=f"=IFERROR((D{tot}-C{tot})/C{tot},0)"); cv.style = "td"; cv.font = Font(bold=True, color=DANGER); cv.fill = fill(SURFACE); cv.number_format = "0.0%"
    cell_name(wb, "VariancePct", "Inventory Variance", f"$E${tot}")
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="greaterThan", formula=["0.05"], fill=fill(RED_BG)))
    ws.cell(row=tot + 2, column=2, value="Every point of variance over 2-3% is profit walking out the door.").style = "section"
    ws.freeze_panes = "A5"


def build_inventory(wb):
    ws = wb.create_sheet("Liquor Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 12, 12, 14, 2])
    luxe_header(ws, "F", "📦  LIQUOR INVENTORY & PAR",
                "Par vs on hand — order before you run dry on a Friday night.")
    table_headers(ws, 4, ["Item", "Par", "On Hand", "Unit", "To Order"], start_col=2)
    start = L0
    for i, (item, par, oh, unit) in enumerate(INVENTORY):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=par); cp.style = "input"; cp.number_format = "#,##0"
        co = ws.cell(row=r, column=4, value=oh); co.style = "input"; co.number_format = "#,##0"
        ws.cell(row=r, column=5, value=unit).style = "td"
        cb = ws.cell(row=r, column=6, value=f"=MAX(C{r}-D{r},0)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.number_format = "#,##0"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVENTORY) - 1
    for c in range(2, 7):
        add_dv(ws, f"E{start}:E{end}", "UnitList")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=[f"C{start}*0.5"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


def build_happyhour(wb):
    ws = wb.create_sheet("Happy Hour"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 12, 12, 2])
    luxe_header(ws, "F", "🕔  HAPPY HOUR",
                "Discount that still profits — check the margin before you cut the price.")
    table_headers(ws, 4, ["Drink", "Regular", "HH Price", "Pour Cost", "HH Margin"], start_col=2)
    start = L0
    for i, (drink, reg, hh, cost) in enumerate(HAPPYHOUR):
        r = start + i
        ws.cell(row=r, column=2, value=drink).style = "td_left"
        cr = ws.cell(row=r, column=3, value=reg); cr.style = "input"; cr.number_format = '"$"#,##0.00'
        ch = ws.cell(row=r, column=4, value=hh); ch.style = "input"; ch.number_format = '"$"#,##0.00'
        cc = ws.cell(row=r, column=5, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=6, value=f"=IFERROR((D{r}-E{r})/D{r},0)"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(HAPPYHOUR) - 1
    nrange(wb, "HHMargin", "Happy Hour", "F", start, end)
    ws.cell(row=end + 2, column=2, value="Avg happy-hour margin").style = "field_label"
    c = ws.cell(row=end + 2, column=6, value="=AVERAGE(HHMargin)"); c.style = "field_value"; c.number_format = "0%"; c.fill = fill(MINT_BG)
    cell_name(wb, "HHAvgMargin", "Happy Hour", f"$F${end+2}")
    ws.freeze_panes = "A5"


def build_weekly(wb):
    ws = wb.create_sheet("Weekly Sales"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 2])
    luxe_header(ws, "E", "📈  WEEKLY SALES",
                "Sales & covers by day — see the weekend carry the week.")
    table_headers(ws, 4, ["Day", "Sales", "Covers", "Avg Check"], start_col=2)
    start = L0
    for i, (day, sales, covers) in enumerate(WEEKLY):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        cc = ws.cell(row=r, column=4, value=covers); cc.style = "input"; cc.number_format = "#,##0"
        ca = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); ca.style = "td"; ca.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(WEEKLY) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    cs = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cc = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = "#,##0"
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "WeekSalesTotal", "Weekly Sales", f"$C${tot}")
    ws.add_chart(_barchart(ws, "Sales by Day", start, end, 3, 2), "G4")
    ws.freeze_panes = "A5"


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste & Spill", "🗑", "WASTE & SPILL LOG",
        "Over-pours, comps & breakage — the quiet leaks that show up as variance.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste & Spill", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="WEEKLY WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste & Spill", f"$D${tot}")


def build_ordering(wb):
    ws, start, end = build_log(
        wb, "Ordering", "🚚", "ORDERING & SUPPLIERS",
        "Your standing order — par quantities & cost, by distributor.",
        ["Item", "Supplier", "Par Order", "Cost"],
        ORDERING, [2, 26, 20, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)
    nrange(wb, "OrderCost", "Ordering", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ORDER TOTAL").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(OrderCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_cash(wb):
    ws = wb.create_sheet("Cash & Tips"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 14, 2])
    luxe_header(ws, "F", "💵  CASH & TIPS",
                "Reconcile each night — cash, card & tips, so the drawer always balances.")
    table_headers(ws, 4, ["Day", "Cash", "Card", "Tips", "Total"], start_col=2)
    start = L0
    for i, (day, cash, card, tips) in enumerate(CASHTIPS):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cash); cc.style = "input"; cc.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=4, value=card); cd.style = "input"; cd.number_format = '"$"#,##0'
        ct = ws.cell(row=r, column=5, value=tips); ct.style = "input"; ct.number_format = '"$"#,##0'
        cto = ws.cell(row=r, column=6, value=f"=C{r}+D{r}+E{r}"); cto.style = "td"; cto.font = Font(bold=True, color=PRIMARY); cto.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CASHTIPS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    for col in (3, 4, 5, 6):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


def build_events(wb):
    ws = wb.create_sheet("Events & Tabs"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 12, 14, 14, 2])
    luxe_header(ws, "F", "🎤  EVENTS & TABS",
                "Trivia, private parties & big tabs — the bookings that pack slow nights.")
    table_headers(ws, 4, ["Event", "Date", "Deposit", "Min Spend", "Status"], start_col=2)
    start = L0
    import datetime as dt
    for i, (name, days, dep, minsp, status) in enumerate(EVENTS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cd = ws.cell(row=r, column=3, value=dt.date.today() + dt.timedelta(days=days)); cd.style = "input"; cd.number_format = "mm/dd"
        cdp = ws.cell(row=r, column=4, value=dep); cdp.style = "input"; cdp.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=5, value=minsp); cm.style = "input"; cm.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=6, value=status); cs.style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EVENTS) - 1
    nrange(wb, "EventStatus", "Events & Tabs", "F", start, end)
    cmap = {"Confirmed": MINT_BG, "Deposit": WARN_BG, "Recurring": SURFACE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🍺  BAR & PUB COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Pour cost, keg profit, inventory variance & a Bar Score — your whole bar, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("DRINKS", "=COUNTA(DrinkItem)", "num"),
        ("AVG POUR COST", "=AVERAGE(DrinkCost)", "money2"),
        ("AVG PRICE", "=AVERAGE(DrinkPrice)", "money2"),
        ("AVG MARGIN", "=AVERAGE(DrinkMargin)", "money2"),
        ("POUR COST", "=PourCostPct", "pct"),
        ("TOP SELLER", "=INDEX(DrinkItem,MATCH(MAX(DrinkRev),DrinkRev,0))", "text"),
    ]
    row2 = [
        ("WEEKLY SALES", "=WeeklySales", "money"),
        ("WEEKLY UNITS", "=SUM(DrinkUnits)", "num"),
        ("AVG PROFIT/KEG", "=AvgKegProfit", "money"),
        ("INV VARIANCE", "=VariancePct", "pct"),
        ("HAPPY HR MARGIN", "=HHAvgMargin", "pct"),
        ("BAR SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "BAR HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SALES BY DAY", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Pour cost on target", "=IFERROR(MIN(TargetPour/PourCostPct,1),0)"),
        ("Margin per drink", "=IFERROR(MIN(AVERAGE(DrinkMargin)/MarginGoal,1),0)"),
        ("Menu fully costed", "=IFERROR(COUNTIF(DrinkCost,\">0\")/COUNTA(DrinkItem),0)"),
        ("Inventory variance low", "=IFERROR(1-MIN(VariancePct/VarLimit,1),0)"),
        ("Happy-hour margin", "=IFERROR(MIN(HHAvgMargin/0.5,1),0)"),
        ("Gross margin", "=IFERROR(MIN((1-PourCostPct)/0.75,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    wk = wb["Weekly Sales"]
    ch = BarChart(); ch.type = "col"; ch.title = "Sales by Day"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(wk, min_col=3, min_row=5, max_row=4 + len(WEEKLY)), titles_from_data=False)
    ch.set_categories(Reference(wk, min_col=2, min_row=5, max_row=4 + len(WEEKLY))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Bar & Pub Command Center™ — cost every pour, kill your shrinkage, pour more profit.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_pourcost(wb); build_menu(wb)
    build_kegs(wb); build_variance(wb); build_inventory(wb); build_happyhour(wb)
    build_weekly(wb); build_waste(wb); build_ordering(wb); build_cash(wb)
    build_events(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Pour Cost", "Drink Menu", "Keg & Draft", "Inventory Variance",
             "Liquor Inventory", "Happy Hour", "Weekly Sales", "Waste & Spill", "Ordering",
             "Cash & Tips", "Events & Tabs", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Bar_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
