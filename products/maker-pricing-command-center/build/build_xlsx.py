"""Build Handmade & Maker Pricing Command Center™ — The Maker's Operating System.

14 tabs · a premium handmade / maker-business operating system in Google Sheets &
Excel. Dashboard, a price-for-profit engine (materials + labor + overhead → wholesale
& retail), a materials list, a product line, supplies, orders, overhead & fees, a
sales log, time & labor, expenses, reviews and a monthly summary — one dashboard.
Price for real profit, and finally pay yourself for your time.

Run: python3 build_xlsx.py   ->  ../Maker_Pricing_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
CHANNEL = ["Etsy", "Website", "Market", "Wholesale", "Custom"]
ORDSTATUS = ["New", "Making", "Shipped", "Delivered", "Refunded"]
UNIT = ["each", "oz", "yd", "g", "set", "pair"]

MARGIN_GOAL = 0.50
MARKUP_GOAL = 2.0
HOUR_GOAL = 40
PROFIT_GOAL = 1000
WHOLESALE_MARGIN_GOAL = 0.50

# Pricing calculator — flagship product & inputs
LABOR_RATE = 20
LABOR_HOURS = 0.25
OVERHEAD_ITEM = 1.00
MARKUP = 2.5
WHOLESALE_FACTOR = 0.5
# Materials for the flagship (Soy Candle 8oz): (material, cost)
MATERIALS = [("Soy wax", 1.80), ("Fragrance oil", 1.20), ("Wick", 0.15), ("Jar", 1.85), ("Label", 0.25), ("Box", 0.75)]

# Product line: (product, base cost, retail, wholesale)
PRODUCTS = [
    ("Soy Candle 8oz", 12.00, 30, 15), ("Soy Candle 4oz", 7.50, 18, 9), ("Wax Melts (6)", 5.00, 14, 7),
    ("Reed Diffuser", 9.00, 24, 12), ("Candle Trio Set", 30.00, 72, 36), ("Room Spray", 4.50, 12, 6),
]

# Supplies & inventory: (item, on hand, reorder at, unit cost)
SUPPLIES = [
    ("Soy wax (lb)", 40, 15, 2.90), ("Fragrance (oz)", 60, 20, 1.20), ("Jars 8oz", 120, 50, 1.85),
    ("Wicks", 300, 100, 0.15), ("Labels", 250, 100, 0.25), ("Boxes", 140, 60, 0.75),
]

# Orders: (order, channel, items, total, status)
ORDERS = [
    ("#1042 Bright", "Etsy", 3, 84, "Making"), ("#1043 Cole", "Website", 1, 30, "Shipped"),
    ("#1044 Rowe", "Etsy", 2, 42, "New"), ("Market — Sat", "Market", 22, 540, "Delivered"),
    ("Cedar Boutique", "Wholesale", 24, 360, "Making"),
]

# Overhead & fees (monthly): (item, amount)
OVERHEAD = [
    ("Etsy fees & ads", 140), ("Shipping supplies", 60), ("Studio & utilities", 90),
    ("Software & website", 30), ("Insurance", 20), ("Craft show fees", 20),
]

# Sales log (this month, sample): (date, product, qty, price)
SALES = [
    ("Jul 2", "Soy Candle 8oz", 4, 30), ("Jul 5", "Candle Trio Set", 2, 72), ("Jul 9", "Wax Melts (6)", 6, 14),
    ("Jul 14", "Reed Diffuser", 3, 24), ("Jul 20", "Soy Candle 4oz", 8, 18), ("Jul 26", "Room Spray", 5, 12),
]

# Time & labor: (task, minutes)
TASKS = [
    ("Pour & set", 8), ("Wick & prep", 4), ("Label & box", 3), ("Photograph & list", 2), ("Pack & ship", 3),
]

# Expenses (monthly): (item, amount)
EXPENSES = [
    ("Materials", 540), ("Overhead & fees", 360), ("Your labor (paid)", 450), ("Marketing", 90), ("Misc", 60),
]

# Reviews & repeat customers: (customer, rating, repeat?)
REVIEWS = [
    ("Bright", 5, "Yes"), ("Cole", 5, "No"), ("Rowe", 4, "Yes"), ("Marsh", 5, "Yes"), ("Vega", 5, "No"),
]

# Channels & markets: (channel, revenue, fee %)
CHANNELS_DATA = [
    ("Etsy", 1200, 0.105), ("Website", 480, 0.029), ("Market", 540, 0.0),
    ("Wholesale", 360, 0.0), ("Custom", 120, 0.0),
]

# Monthly summary: (month, revenue, profit)
MONTHS = [("Feb", 1800, 900), ("Mar", 2100, 1050), ("Apr", 2400, 1200),
          ("May", 2600, 1300), ("Jun", 2500, 1250), ("Jul", 2700, 1350)]

# This month figures
THIS_REVENUE = 2700
THIS_UNITS = 90
THIS_MATERIALS = 540
THIS_OVERHEAD = 360

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Shop name", "Ember & Oak", None, "Business"),
        ("Owner", "Wren", None, "Owner"),
        ("Your labor rate ($/hr)", LABOR_RATE, '"$"#,##0', "LaborRate"),
        ("Retail markup (×)", MARKUP, "0.0", "MarkupDefault"),
        ("Wholesale factor (of retail)", WHOLESALE_FACTOR, "0%", "WholesaleFactor"),
        ("Retail-margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Markup goal (×)", MARKUP_GOAL, "0.0", "MarkupGoal"),
        ("Hourly-wage goal ($/hr)", HOUR_GOAL, '"$"#,##0', "HourGoal"),
        ("Monthly profit goal", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Wholesale-margin goal %", WHOLESALE_MARGIN_GOAL, "0%", "WholesaleMarginGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Channel", CHANNEL, "ChannelList"), ("F", "Order status", ORDSTATUS, "OrdStatusList"),
             ("G", "Unit", UNIT, "UnitList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🧶  HANDMADE & MAKER PRICING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Price for real profit, and finally pay yourself for your time.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE HANDMADE SHOP, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A handmade shop lives or dies on one number most makers guess at: the real price a product must "
                      "sell for to cover materials, pay you for your time, and still make a profit. This makes it clear: "
                      "a price-for-profit engine that adds your materials, your labor and your overhead, then applies a "
                      "markup to give a wholesale and a retail price — and shows your true hourly wage. Build your product "
                      "line, run orders and supplies, track sales, overhead and expenses — all in ONE premium Google "
                      "Sheets & Excel system built for makers.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your labor rate, markup & margin goals.",
             "2.  Price a product in the Pricing Calculator — materials, labor, overhead.",
             "3.  Read your wholesale, retail & your true hourly wage.",
             "4.  Build your Product Line and run Orders & Supplies.",
             "5.  Log sales, track overhead, fees and expenses.",
             "6.  Check the Dashboard: profit, your wage & a Maker Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional candle shop (Ember & Oak) is included so you can see how it all connects — "
               "just type over it with your own materials and products. The retail price and your true hourly wage are "
               "the two numbers that decide whether a handmade shop pays you or costs you, and they roll into a live "
               "Maker Score. Twelve matching printable pages (pricing worksheet, product price list, order form, supply "
               "list & more) are included. This is a business tool, not financial, legal or tax advice — confirm figures "
               "with your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Add your time to the price — a handmade product that doesn't pay you isn't priced right.", "section_gold")


# ===========================================================================
def build_pricing(wb):
    ws = wb.create_sheet("Pricing Calculator"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 2])
    luxe_header(ws, "C", "🧮  PRICING CALCULATOR",
                "Materials + your labor + overhead → wholesale & retail price, and your true hourly wage.")
    ws.cell(row=5, column=2, value="PRODUCT").style = "section_gold"
    ws.cell(row=5, column=3, value="Soy Candle 8oz").font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Material", "Cost"], start_col=2)
    start = 7
    for i, (mat, cost) in enumerate(MATERIALS):
        r = start + i
        ws.cell(row=r, column=2, value=mat).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MATERIALS) - 1
    nrange(wb, "MatRows", "Pricing Calculator", "C", start, end)
    mt = end + 1
    ws.cell(row=mt, column=2, value="= MATERIALS").style = "th"
    cm = ws.cell(row=mt, column=3, value="=SUM(MatRows)"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0.00'
    cell_name(wb, "MatCost", "Pricing Calculator", f"$C${mt}")
    lr = mt + 1
    ws.cell(row=lr, column=2, value="Labor hours (your time)").style = "field_label"
    cl = ws.cell(row=lr, column=3, value=LABOR_HOURS); cl.style = "input"; cl.number_format = "0.00"
    cell_name(wb, "LaborHours", "Pricing Calculator", f"$C${lr}")
    ws.cell(row=lr + 1, column=2, value="+ Labor (hours × your rate)").style = "field_label"
    clc = ws.cell(row=lr + 1, column=3, value="=LaborHours*LaborRate"); clc.style = "field_value"; clc.number_format = '"$"#,##0.00'
    cell_name(wb, "LaborCost", "Pricing Calculator", f"$C${lr+1}")
    ws.cell(row=lr + 2, column=2, value="+ Overhead per item").style = "field_label"
    co = ws.cell(row=lr + 2, column=3, value=OVERHEAD_ITEM); co.style = "input"; co.number_format = '"$"#,##0.00'
    cell_name(wb, "OverheadItem", "Pricing Calculator", f"$C${lr+2}")
    bc = lr + 3
    ws.cell(row=bc, column=2, value="= BASE COST").style = "th"
    cb = ws.cell(row=bc, column=3, value="=MatCost+LaborCost+OverheadItem"); cb.style = "td"; cb.font = Font(bold=True, size=12, color=PRIMARY); cb.fill = fill(MINT_BG); cb.number_format = '"$"#,##0.00'
    cell_name(wb, "BaseCost", "Pricing Calculator", f"$C${bc}")
    mk = bc + 2
    ws.cell(row=mk, column=2, value="Retail markup (×)").style = "field_label"
    cmk = ws.cell(row=mk, column=3, value=MARKUP); cmk.style = "input"; cmk.number_format = "0.0"
    cell_name(wb, "Markup", "Pricing Calculator", f"$C${mk}")
    ws.cell(row=mk + 1, column=2, value="= RETAIL PRICE").style = "th"
    cr = ws.cell(row=mk + 1, column=3, value="=BaseCost*Markup"); cr.style = "td"; cr.font = Font(bold=True, size=13, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0.00'
    cell_name(wb, "RetailPrice", "Pricing Calculator", f"$C${mk+1}")
    ws.cell(row=mk + 2, column=2, value="= WHOLESALE PRICE").style = "th"
    cw = ws.cell(row=mk + 2, column=3, value="=RetailPrice*WholesaleFactor"); cw.style = "td"; cw.font = Font(bold=True, size=12, color=PRIMARY); cw.fill = fill(MINT_BG); cw.number_format = '"$"#,##0.00'
    cell_name(wb, "WholesalePrice", "Pricing Calculator", f"$C${mk+2}")
    mr = mk + 4
    ws.cell(row=mr, column=2, value="Profit per item (retail)").style = "field_label"
    cp = ws.cell(row=mr, column=3, value="=RetailPrice-BaseCost"); cp.style = "field_value"; cp.number_format = '"$"#,##0.00'; cp.fill = fill(MINT_BG)
    cell_name(wb, "RetailProfit", "Pricing Calculator", f"$C${mr}")
    ws.cell(row=mr + 1, column=2, value="Retail margin").style = "field_label"
    crm = ws.cell(row=mr + 1, column=3, value="=IFERROR(RetailProfit/RetailPrice,0)"); crm.style = "field_value"; crm.number_format = "0%"; crm.fill = fill(MINT_BG)
    cell_name(wb, "RetailMargin", "Pricing Calculator", f"$C${mr+1}")
    ws.cell(row=mr + 2, column=2, value="Wholesale margin").style = "field_label"
    cwm = ws.cell(row=mr + 2, column=3, value="=IFERROR((WholesalePrice-BaseCost)/WholesalePrice,0)"); cwm.style = "field_value"; cwm.number_format = "0%"; cwm.fill = fill(MINT_BG)
    cell_name(wb, "WholesaleMargin", "Pricing Calculator", f"$C${mr+2}")
    ws.cell(row=mr + 3, column=2, value="= YOUR TRUE HOURLY WAGE").style = "th"
    ce = ws.cell(row=mr + 3, column=3, value="=IFERROR((RetailPrice-MatCost-OverheadItem)/LaborHours,0)"); ce.style = "td"; ce.font = Font(bold=True, size=12, color=PRIMARY); ce.fill = fill(MINT_BG); ce.number_format = '"$"#,##0.00'
    cell_name(wb, "EffHourly", "Pricing Calculator", f"$C${mr+3}")


def build_productline(wb):
    ws = wb.create_sheet("Product Line"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 14, 12, 2])
    luxe_header(ws, "F", "🕯  PRODUCT LINE",
                "Every product — base cost, wholesale, retail and the retail margin. Price the whole line for profit.")
    table_headers(ws, 4, ["Product", "Base Cost", "Wholesale", "Retail", "Margin"], start_col=2)
    start = L0
    for i, (prod, base, retail, whole) in enumerate(PRODUCTS):
        r = start + i
        ws.cell(row=r, column=2, value=prod).style = "td_left"
        cb = ws.cell(row=r, column=3, value=base); cb.style = "input"; cb.number_format = '"$"#,##0.00'
        cw = ws.cell(row=r, column=4, value=whole); cw.style = "input"; cw.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=retail); cr.style = "input"; cr.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=6, value=f"=IFERROR((E{r}-C{r})/E{r},0)"); cm.style = "td"; cm.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRODUCTS) - 1
    nrange(wb, "ProdName", "Product Line", "B", start, end)
    nrange(wb, "ProdRetail", "Product Line", "E", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        ColorScaleRule(start_type="num", start_value=0.35, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=0.55, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.70, end_color="FF" + HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_supplies(wb):
    ws = wb.create_sheet("Supplies & Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 12, 12, 14, 14, 2])
    luxe_header(ws, "F", "📦  SUPPLIES & INVENTORY",
                "Your raw materials — on hand, reorder point and value, so you never run out mid-batch.")
    table_headers(ws, 4, ["Item", "On Hand", "Reorder At", "Unit Cost", "Value"], start_col=2)
    start = L0
    for i, (item, oh, ro, uc) in enumerate(SUPPLIES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        co = ws.cell(row=r, column=3, value=oh); co.style = "input"; co.number_format = "#,##0"
        crr = ws.cell(row=r, column=4, value=ro); crr.style = "input"; crr.number_format = "#,##0"
        cu = ws.cell(row=r, column=5, value=uc); cu.style = "input"; cu.number_format = '"$"#,##0.00'
        cv = ws.cell(row=r, column=6, value=f"=C{r}*E{r}"); cv.style = "td"; cv.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SUPPLIES) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="INVENTORY VALUE").style = "th"
    for c in (3, 4, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cv = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cv.style = "td"; cv.font = Font(bold=True, color=PRIMARY); cv.fill = fill(MINT_BG); cv.number_format = '"$"#,##0.00'
    ws.freeze_panes = "A5"


def build_orders(wb):
    ws, start, end = build_log(
        wb, "Orders", "🧾", "ORDERS",
        "Every order by channel — items, total and status, so nothing ships late.",
        ["Order", "Channel", "Items", "Total", "Status"],
        ORDERS, [2, 22, 14, 12, 14, 14, 2], text_left={2}, ints={4}, money={5}, reserved=28, start_col=2,
        validations=[("C", "ChannelList"), ("F", "OrdStatusList")])
    nrange(wb, "OrderTotal", "Orders", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="OPEN ORDERS VALUE").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(OrderTotal)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)


def build_overhead(wb):
    ws = wb.create_sheet("Overhead & Fees"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "🏷  OVERHEAD & FEES",
                "The monthly costs behind every sale — fees, shipping, studio. Spread across your units.")
    table_headers(ws, 4, ["Item", "Monthly"], start_col=2)
    start = L0
    for i, (item, amt) in enumerate(OVERHEAD):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(OVERHEAD) - 1
    nrange(wb, "OHAmt", "Overhead & Fees", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL OVERHEAD (mo)").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(OHAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "OverheadTotal", "Overhead & Fees", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_sales(wb):
    ws, start, end = build_log(
        wb, "Sales Log", "🛒", "SALES LOG",
        "Every sale this month — product, quantity and price, so revenue and units are always current.",
        ["Date", "Product", "Qty", "Price"],
        SALES, [2, 14, 24, 10, 14, 2], text_left={2, 3}, ints={4}, money={5}, reserved=30, start_col=2)
    nrange(wb, "SaleQty", "Sales Log", "D", start, end)


def build_time(wb):
    ws = wb.create_sheet("Time & Labor"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 14, 2])
    luxe_header(ws, "D", "⏱  TIME & LABOR",
                "Where your minutes go per item — add them up and pay yourself for every one.")
    table_headers(ws, 4, ["Task", "Minutes", "Labor $"], start_col=2)
    start = L0
    for i, (task, mins) in enumerate(TASKS):
        r = start + i
        ws.cell(row=r, column=2, value=task).style = "td_left"
        cm = ws.cell(row=r, column=3, value=mins); cm.style = "input"; cm.number_format = "#,##0"
        cl = ws.cell(row=r, column=4, value=f"=C{r}/60*LaborRate"); cl.style = "td"; cl.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(TASKS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="PER ITEM").style = "th"
    cm = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = "#,##0"
    cl = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cl.style = "td"; cl.font = Font(bold=True, color=PRIMARY); cl.fill = fill(MINT_BG); cl.number_format = '"$"#,##0.00'
    ws.freeze_panes = "A5"


def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  EXPENSES",
                "This month's costs — materials, overhead, your paid labor and more, netted against revenue.")
    table_headers(ws, 4, ["Expense", "Amount"], start_col=2)
    start = L0
    for i, (item, amt) in enumerate(EXPENSES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EXPENSES) - 1
    nrange(wb, "ExpAmt", "Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Expenses", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_reviews(wb):
    ws, start, end = build_log(
        wb, "Reviews & Repeat", "⭐", "REVIEWS & REPEAT",
        "Ratings and who came back — repeat customers are the cheapest sale you'll ever make.",
        ["Customer", "Rating", "Repeat?"],
        REVIEWS, [2, 24, 12, 14, 2], text_left={2}, dec={3}, reserved=24, start_col=2,
        validations=[("D", "YesNoList")])
    nrange(wb, "RevRating", "Reviews & Repeat", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVG RATING").style = "th"
    c = ws.cell(row=tot, column=3, value="=IFERROR(AVERAGE(RevRating),0)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = "0.0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)


def build_channels(wb):
    ws = wb.create_sheet("Channels & Markets"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 14, 12, 14, 14, 2])
    luxe_header(ws, "F", "🛍  CHANNELS & MARKETS",
                "Where your sales come from — revenue, the fees each channel takes and what you actually keep.")
    table_headers(ws, 4, ["Channel", "Revenue", "Fee %", "Fees", "Net Kept"], start_col=2)
    start = L0
    for i, (chan, rev, fee) in enumerate(CHANNELS_DATA):
        r = start + i
        ws.cell(row=r, column=2, value=chan).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rev); cr.style = "input"; cr.number_format = '"$"#,##0'
        cf = ws.cell(row=r, column=4, value=fee); cf.style = "input"; cf.number_format = "0.0%"
        cfe = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); cfe.style = "td"; cfe.number_format = '"$"#,##0'
        cn = ws.cell(row=r, column=6, value=f"=C{r}-E{r}"); cn.style = "td"; cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CHANNELS_DATA) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="ALL CHANNELS").style = "th"
    cr = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cfe = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cfe.style = "td"; cfe.font = Font(bold=True, color=DANGER); cfe.fill = fill(SURFACE); cfe.number_format = '"$"#,##0'
    cn = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    ws.conditional_formatting.add(f"F{start}:F{end}", DataBarRule(start_type="min", end_type="max", color=HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "This month's numbers, and the trend — watch the profit grow.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    rows = [
        ("Revenue", THIS_REVENUE, '"$"#,##0', "MonthlyRevenue", MINT_BG),
        ("Units sold", THIS_UNITS, "#,##0", "UnitsSold", None),
        ("Materials cost", THIS_MATERIALS, '"$"#,##0', "ThisMaterials", None),
        ("Overhead & fees", THIS_OVERHEAD, '"$"#,##0', "ThisOverhead", None),
    ]
    for i, (lab, val, fmt, nm, bg) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"; c.number_format = fmt
        if bg:
            c.fill = fill(bg)
        cell_name(wb, nm, "Monthly Summary", f"$C${r}")
    ws.cell(row=10, column=2, value="Your labor (units × per-item)").style = "field_label"
    cl = ws.cell(row=10, column=3, value="=UnitsSold*LaborCost"); cl.style = "field_value"; cl.number_format = '"$"#,##0'
    cell_name(wb, "ThisLabor", "Monthly Summary", "$C$10")
    ws.cell(row=11, column=2, value="= MONTHLY PROFIT (paid yourself first)").style = "th"
    cp = ws.cell(row=11, column=3, value="=MonthlyRevenue-ThisMaterials-ThisOverhead-ThisLabor"); cp.style = "td"; cp.font = Font(bold=True, size=13, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyProfit", "Monthly Summary", "$C$11")
    ws.cell(row=12, column=2, value="Materials % of revenue").style = "field_label"
    cmp = ws.cell(row=12, column=3, value="=IFERROR(ThisMaterials/MonthlyRevenue,0)"); cmp.style = "field_value"; cmp.number_format = "0%"; cmp.fill = fill(MINT_BG)
    cell_name(wb, "MaterialsPct", "Monthly Summary", "$C$12")
    # trend
    ws.cell(row=14, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 15, ["Month", "Revenue"], start_col=2)
    ts = 16
    for i, (m, rev, prof) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rev); cr.style = "input"; cr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "MonthRev", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Revenue by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🧶  HANDMADE & MAKER PRICING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Base cost, retail, your true hourly wage & a Maker Score — your whole shop, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MATERIALS", "=MatCost", "money2"),
        ("LABOR", "=LaborCost", "money2"),
        ("BASE COST", "=BaseCost", "money2"),
        ("WHOLESALE", "=WholesalePrice", "money2"),
        ("RETAIL", "=RetailPrice", "money2"),
        ("RETAIL MARGIN", "=RetailMargin", "pct"),
    ]
    row2 = [
        ("PROFIT / ITEM", "=RetailProfit", "money2"),
        ("YOUR HOURLY", "=EffHourly", "money2"),
        ("MONTHLY REVENUE", "=MonthlyRevenue", "money"),
        ("UNITS SOLD", "=UnitsSold", "num"),
        ("MONTHLY PROFIT", "=MonthlyProfit", "money"),
        ("MAKER SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "MAKER HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Retail margin healthy", "=IFERROR(MIN(RetailMargin/MarginGoal,1),0)"),
        ("Markup healthy", "=IFERROR(MIN(Markup/MarkupGoal,1),0)"),
        ("Paying yourself", "=IFERROR(MIN(EffHourly/HourGoal,1),0)"),
        ("Products priced", "=IFERROR(COUNTIF(ProdRetail,\">0\")/COUNTA(ProdName),0)"),
        ("Profitable", "=IFERROR(MIN(MonthlyProfit/ProfitGoal,1),0)"),
        ("Wholesale margin healthy", "=IFERROR(MIN(WholesaleMargin/WholesaleMarginGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=16, max_row=15 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=16, max_row=15 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Handmade & Maker Pricing Command Center™ — price for profit, and pay yourself.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_pricing(wb); build_productline(wb)
    build_supplies(wb); build_orders(wb); build_overhead(wb); build_sales(wb)
    build_time(wb); build_expenses(wb); build_reviews(wb); build_channels(wb); build_summary(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Pricing Calculator", "Product Line", "Supplies & Inventory", "Orders",
             "Overhead & Fees", "Sales Log", "Time & Labor", "Expenses", "Reviews & Repeat",
             "Channels & Markets", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Maker_Pricing_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
