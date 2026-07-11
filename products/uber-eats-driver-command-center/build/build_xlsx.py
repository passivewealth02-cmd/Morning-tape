"""Build Uber Eats Driver Command Center™ — The Ultimate Food-Delivery Earnings & Budget System.

18 sheets + Welcome · a premium delivery-driver operating system in Excel & Sheets.
Delivery log, earnings, mileage, fuel, vehicle & maintenance, business expenses, a
household budget, a tax center (mileage deduction), savings & goals, promos & quests,
hotspots, ratings and an Analytics driver-health score — one dashboard.
Built for Uber Eats; works for DoorDash, Grubhub, Instacart or any delivery gig.

Run: python3 build_xlsx.py   ->  ../Uber_Eats_Driver_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

PLATFORMS = ["Uber Eats", "DoorDash", "Grubhub", "Instacart", "Multi", "Other"]
EXP_CATS = ["Fuel", "Car Payment", "Insurance", "Maintenance & Repairs", "Phone & Data",
            "Bags & Supplies", "Tolls", "Car Wash & Cleaning", "Miscellaneous"]
BUDGET_CATS = ["Housing", "Utilities", "Groceries", "Health", "Debt", "Savings",
               "Personal", "Insurance", "Other"]
MAINT_TYPES = ["Oil Change", "Tires", "Brakes", "Inspection", "Rotation", "Battery", "Repair", "Other"]
GOAL_CATS = ["Earnings", "Hourly", "Deliveries", "Savings", "Miles", "Tax", "Debt"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your goals & vehicle details once — every dashboard follows.")
    merge_set(ws, "B5:C5", "DRIVER INPUTS", "section")
    controls = [
        ("Driver Name", "Alex Morgan", None, "DriverName"),
        ("Vehicle", "2019 Honda Civic", None, "VehicleName"),
        ("Home City", "Denver, CO", None, "HomeCity"),
        ("Monthly Net Goal", 1800, '"$"#,##0', "NetGoal"),
        ("Target Net $ / Hour", 16, '"$"#,##0.00', "HourlyTarget"),
        ("Monthly Deliveries Goal", 300, "#,##0", "DeliveryGoal"),
        ("Monthly Shifts Goal", 16, "0", "ShiftGoal"),
        ("Emergency Fund Goal", 4000, '"$"#,##0', "SavingsGoal"),
        ("Tax Reserve Goal (mo)", 1000, '"$"#,##0', "TaxReserveGoal"),
        ("IRS Mileage Rate", 0.70, '"$"#,##0.00', "MileageRate"),
        ("Tax Set-Aside %", 0.25, "0%", "TaxRate"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Platform", PLATFORMS, "PlatformList"), ("F", "Expense Category", EXP_CATS, "ExpCatList"),
             ("G", "Budget Category", BUDGET_CATS, "BudgetCatList"), ("H", "Maintenance", MAINT_TYPES, "MaintList"),
             ("I", "Goal Category", GOAL_CATS, "GoalCatList"), ("J", "Priority", PRIORITIES, "PriorityList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    ws.cell(row=17, column=5, value="Yes / No").style = "th"
    for ri, v in enumerate(YESNO):
        ws.cell(row=18 + ri, column=5, value=v).style = "td_left"
    wb.defined_names["YesNoList"] = DefinedName("YesNoList", attr_text="Settings!$E$18:$E$19")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🛵  UBER EATS DRIVER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate food-delivery earnings & budget system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "RUN YOUR WHOLE DELIVERY BUSINESS FROM ONE FILE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Every dash, mile and dollar in ONE premium Excel & Google Sheets system. Log your deliveries and "
                      "the Command Center instantly shows your real take-home — net earnings, true $/hour and $/mile "
                      "after gas and expenses, your mileage tax deduction, your budget and your savings. See which apps, "
                      "zones and hours actually pay, and stop driving for pennies. Built for Uber Eats — works for "
                      "DoorDash, Grubhub, Instacart or any delivery gig.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your vehicle, goals & the IRS mileage rate.",
             "2.  Log each shift in the Delivery Log — hours, orders, base, tips, miles & gas.",
             "3.  Track Mileage, Fuel & Vehicle Maintenance as you go.",
             "4.  Set your Business Expenses & the Monthly Budget — net income flows in live.",
             "5.  Use the Tax Center to bank your mileage deduction & set-aside.",
             "6.  Watch the Dashboard track net $/hour, savings & a Driver Health Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (Alex, a full-time Denver delivery driver doing ~$2,540/mo across 16 shifts, "
               "~1,500 miles and 321 deliveries) is included so you can see how everything connects — just type over it "
               "with your own numbers. Net earnings, true $/hour and $/mile, the mileage tax deduction, budget leftover "
               "and the Driver Health Score all update automatically. Every sheet is print-friendly and works in Excel "
               "and Google Sheets, on desktop and phone — perfect for logging between drop-offs.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, every mile counted — let's grow your take-home.", "section_gold")


# ===========================================================================
# Driver Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Driver Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "🪪  DRIVER PROFILE", "Your delivery business, defined — the details every number flows from.")
    blocks = [
        ("THE DRIVER", [("Driver Name", "=DriverName"), ("Home City", "=HomeCity"),
                        ("Apps", "Uber Eats + DoorDash"), ("Driving Since", "2020"),
                        ("Status", "Full-time"), ("Weekly Target", "5 shifts")]),
        ("THE VEHICLE", [("Vehicle", "=VehicleName"), ("MPG (avg)", "34 city"),
                         ("Odometer", "82,600 mi"), ("Plate / Unit", "DEL-2291"),
                         ("Insurance", "Delivery rider"), ("Est. Value", "$14,200")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "RATINGS & STANDING", "section_gold"); ws.row_dimensions[15].height = 22
    stand = [("Uber Eats Rating", "4.92 ★"), ("DoorDash Rating", "4.89 ★"), ("Acceptance", "62%"),
             ("Completion", "99%"), ("Lifetime Deliveries", "7,180"), ("Rewards Tier", "Diamond")]
    for i, (p, h) in enumerate(stand):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# Delivery Log — core earnings engine
# ===========================================================================
SHIFTS = [
    (23, "Uber Eats", 5.5, 19, 62, 78, 12, 88, 17),
    (22, "DoorDash", 6.0, 22, 71, 88, 20, 102, 20),
    (20, "Uber Eats", 5.0, 17, 55, 66, 10, 80, 16),
    (19, "Multi", 6.5, 24, 78, 95, 22, 112, 22),
    (17, "Grubhub", 4.5, 14, 46, 52, 6, 66, 13),
    (16, "DoorDash", 6.5, 25, 82, 98, 25, 118, 23),
    (15, "Uber Eats", 5.5, 20, 65, 80, 14, 94, 18),
    (13, "Multi", 6.0, 23, 74, 90, 18, 108, 21),
    (12, "Uber Eats", 5.0, 18, 58, 68, 10, 84, 16),
    (10, "DoorDash", 6.0, 22, 72, 89, 21, 102, 20),
    (9, "Grubhub", 5.0, 16, 52, 60, 8, 76, 15),
    (8, "Multi", 6.0, 21, 70, 85, 16, 98, 19),
    (6, "Uber Eats", 4.5, 15, 48, 55, 6, 70, 14),
    (5, "DoorDash", 6.5, 24, 80, 94, 24, 114, 22),
    (3, "Uber Eats", 5.5, 20, 64, 76, 12, 92, 18),
    (2, "Multi", 6.0, 21, 68, 84, 17, 98, 19),
]


def build_deliverylog(wb):
    ws = wb.create_sheet("Delivery Log"); ws.sheet_view.showGridLines = False
    headers = ["Date", "Platform", "Hours", "Orders", "Base Pay", "Tips", "Promo/Peak",
               "Earnings", "Miles", "Gas $"]
    set_widths(ws, [13, 13, 9, 9, 11, 10, 13, 12, 10, 10])
    luxe_header(ws, "J", "🛵  DELIVERY LOG",
                "Log every shift — earnings, miles & gas roll straight into your dashboard.")
    table_headers(ws, 4, headers)
    start = L0
    reserved = 60
    end = start + reserved - 1
    for i, (off, plat, hrs, orders, base, tips, promo, miles, gas) in enumerate(SHIFTS):
        r = start + i
        ws.cell(row=r, column=1, value=dminus(off))
        ws.cell(row=r, column=2, value=plat)
        ws.cell(row=r, column=3, value=hrs)
        ws.cell(row=r, column=4, value=orders)
        ws.cell(row=r, column=5, value=base)
        ws.cell(row=r, column=6, value=tips)
        ws.cell(row=r, column=7, value=promo)
        ws.cell(row=r, column=8, value=f"=E{r}+F{r}+G{r}")
        ws.cell(row=r, column=9, value=miles)
        ws.cell(row=r, column=10, value=gas)
    style_rows(ws, start, end, len(headers), text_left=set(), dates={1},
               money={5, 6, 7, 8, 10}, ints={4, 9}, dec={3})
    add_dv(ws, f"B{start}:B{end}", "PlatformList")
    ws.freeze_panes = "A5"
    trow = end + 1
    ws.cell(row=trow, column=1, value="TOTAL").style = "th"
    ws.cell(row=trow, column=2).style = "th"
    fmts = {3: "0.0", 4: "#,##0", 5: '"$"#,##0', 6: '"$"#,##0', 7: '"$"#,##0',
            8: '"$"#,##0', 9: "#,##0", 10: '"$"#,##0'}
    for col in range(3, 11):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmts[col]
    nrange(wb, "ShiftDate", "Delivery Log", "A", start, end)
    nrange(wb, "ShiftPlatform", "Delivery Log", "B", start, end)
    nrange(wb, "ShiftHours", "Delivery Log", "C", start, end)
    nrange(wb, "ShiftOrders", "Delivery Log", "D", start, end)
    nrange(wb, "ShiftBase", "Delivery Log", "E", start, end)
    nrange(wb, "ShiftTips", "Delivery Log", "F", start, end)
    nrange(wb, "ShiftPromo", "Delivery Log", "G", start, end)
    nrange(wb, "ShiftEarn", "Delivery Log", "H", start, end)
    nrange(wb, "ShiftMiles", "Delivery Log", "I", start, end)
    nrange(wb, "ShiftFuel", "Delivery Log", "J", start, end)
    cell_name(wb, "TotalHours", "Delivery Log", f"$C${trow}")
    cell_name(wb, "TotalOrders", "Delivery Log", f"$D${trow}")
    cell_name(wb, "TotalBase", "Delivery Log", f"$E${trow}")
    cell_name(wb, "TotalTips", "Delivery Log", f"$F${trow}")
    cell_name(wb, "TotalPromo", "Delivery Log", f"$G${trow}")
    cell_name(wb, "GrossEarn", "Delivery Log", f"$H${trow}")
    cell_name(wb, "TotalMiles", "Delivery Log", f"$I${trow}")
    cell_name(wb, "FuelTotal", "Delivery Log", f"$J${trow}")
    ws.conditional_formatting.add(f"H{start}:H{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=210, color=PRIMARY, showValue=True))


# ===========================================================================
# Earnings Breakdown
# ===========================================================================
def build_earnings(wb):
    ws = wb.create_sheet("Earnings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 14, 12, 3, 18, 14, 12, 2])
    luxe_header(ws, "H", "💵  EARNINGS BREAKDOWN",
                "Where the money comes from — earnings mix and app split, live from your log.")
    merge_set(ws, "B5:D5", "EARNINGS MIX (THIS MONTH)", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Source", "Amount", "Share"], start_col=2)
    mix = [("Base Pay", "=TotalBase"), ("Tips", "=TotalTips"), ("Promo / Peak", "=TotalPromo")]
    ms = 7
    for i, (lab, fml) in enumerate(mix):
        r = ms + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cv = ws.cell(row=r, column=3, value=fml); cv.style = "td"; cv.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/GrossEarn,0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    me = ms + len(mix) - 1
    mt = me + 1
    ws.cell(row=mt, column=2, value="TOTAL EARNINGS").style = "th"
    ct = ws.cell(row=mt, column=3, value="=GrossEarn"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    ws.cell(row=mt, column=4).style = "td"; ws.cell(row=mt, column=4).fill = fill(SURFACE)
    nrange(wb, "MixSource", "Earnings", "B", ms, me)
    nrange(wb, "MixVal", "Earnings", "C", ms, me)
    ws.conditional_formatting.add(f"C{ms}:C{me}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1300, color=GOLD_LT, showValue=True))
    merge_set(ws, "F5:H5", "BY APP", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["App", "Earnings", "Share"], start_col=6)
    plats = ["Uber Eats", "DoorDash", "Grubhub", "Multi"]
    ps = 7
    for i, p in enumerate(plats):
        r = ps + i
        ws.cell(row=r, column=6, value=p).style = "td_left"
        cv = ws.cell(row=r, column=7, value=f'=SUMIF(ShiftPlatform,"{p}",ShiftEarn)'); cv.style = "td"; cv.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/GrossEarn,0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(6, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    pe = ps + len(plats) - 1
    nrange(wb, "PlatName", "Earnings", "F", ps, pe)
    nrange(wb, "PlatEarn", "Earnings", "G", ps, pe)
    merge_set(ws, "B13:D13", "PER-SHIFT AVERAGES", "section_gold"); ws.row_dimensions[13].height = 22
    avgs = [("Avg / shift", "=IFERROR(GrossEarn/COUNT(ShiftEarn),0)", '"$"#,##0'),
            ("Avg / delivery", "=IFERROR(GrossEarn/TotalOrders,0)", '"$"#,##0.00'),
            ("Gross $ / hour", "=IFERROR(GrossEarn/TotalHours,0)", '"$"#,##0.00'),
            ("Tips % of earnings", "=IFERROR(TotalTips/GrossEarn,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(avgs):
        r = 14 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        ws.cell(row=r, column=4).style = "td"; ws.cell(row=r, column=4).fill = fill(WHITE)
    merge_set(ws, "F13:H13", "TAKE-HOME (AFTER EXPENSES)", "section_gold"); ws.row_dimensions[13].height = 22
    take = [("Net earnings", "=NetEarn", '"$"#,##0'), ("Net $ / hour", "=PerHour", '"$"#,##0.00'),
            ("Net $ / mile", "=PerMile", '"$"#,##0.00'), ("Expense ratio", "=IFERROR(ExpenseTotal/GrossEarn,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(take):
        r = 14 + i
        ws.cell(row=r, column=6, value=lab).style = "field_label"
        c = ws.cell(row=r, column=7, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net earnings", "Net $ / hour"):
            c.fill = fill(MINT_BG)
        ws.cell(row=r, column=8).style = "td"; ws.cell(row=r, column=8).fill = fill(WHITE)


# ===========================================================================
# Mileage Tracker
# ===========================================================================
def build_mileage(wb):
    rows = [
        (dminus(23), "Downtown dinner rush", "Business", 88, "Restaurant row"),
        (dminus(22), "Multi-restaurant loop", "Business", 102, "DoorDash peak pay"),
        (dminus(20), "University hill", "Business", 80, "Campus orders"),
        (dminus(19), "Multi-app day", "Business", 112, "Uber Eats + DoorDash"),
        (dminus(16), "Suburbs + downtown", "Business", 118, "Long dinner shift"),
        (dminus(13), "Cherry Creek", "Business", 108, "High-tip zone"),
        (dminus(10), "Downtown + Highlands", "Business", 102, "Steady"),
        (dminus(8), "Multi-app dinner", "Business", 98, "Stacked orders"),
        (dminus(5), "Weekend rush", "Business", 114, "Best day this week"),
        (dminus(2), "Downtown night", "Business", 98, "Late-night bars"),
        (dminus(11), "Personal errands", "Personal", 18, "Not deductible"),
    ]
    ws, start, end = build_log(
        wb, "Mileage", "🛣", "MILEAGE TRACKER",
        "Log business miles for the IRS deduction — the biggest tax break delivery drivers miss.",
        ["Date", "Route / Purpose", "Type", "Miles", "Notes"],
        rows, [13, 30, 14, 12, 26],
        text_left={2, 5}, dates={1}, ints={4}, reserved=40)
    nrange(wb, "MileType", "Mileage", "C", start, end)
    nrange(wb, "MileMiles", "Mileage", "D", start, end)
    for t, cc in {"Business": MINT_BG, "Personal": WARN_BG}.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{t}"'], fill=fill(cc)))


# ===========================================================================
# Fuel Log
# ===========================================================================
def build_fuel(wb):
    rows = [
        (dminus(23), "Costco", 8.9, 3.05, "Cheapest in town"),
        (dminus(20), "QuikTrip", 9.2, 3.19, "Topped off"),
        (dminus(17), "Shell", 8.7, 3.29, "Near restaurant row"),
        (dminus(14), "Costco", 9.1, 3.05, "Membership pays off"),
        (dminus(11), "7-Eleven", 8.4, 3.35, "Quick splash"),
        (dminus(8), "QuikTrip", 9.0, 3.15, "Weekend fill"),
        (dminus(5), "Costco", 9.3, 3.05, "Before big shift"),
        (dminus(2), "Shell", 8.6, 3.29, "End of month"),
    ]
    sample = [(d, st, g, ppg, f"=ROUND(C{i+L0}*D{i+L0},2)", note)
              for i, (d, st, g, ppg, note) in enumerate(rows)]
    ws, start, end = build_log(
        wb, "Fuel Log", "⛽", "FUEL LOG",
        "Track every fill-up — gallons, price & cost so your real gas spend is never a guess.",
        ["Date", "Station", "Gallons", "$ / Gal", "Cost", "Notes"],
        sample, [13, 16, 12, 12, 12, 24],
        text_left={2, 6}, dates={1}, dec={3}, money2={4, 5}, reserved=40)
    trow = end + 1
    ws.cell(row=trow, column=2, value="TOTAL").style = "th"
    for col, fmt in ((3, "0.0"), (5, '"$"#,##0.00')):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt
    for col in (1, 4, 6):
        ws.cell(row=trow, column=col).style = "td"; ws.cell(row=trow, column=col).fill = fill(SURFACE)


# ===========================================================================
# Vehicle & Maintenance
# ===========================================================================
def build_vehicle(wb):
    rows = [
        (dminus(18), "Oil Change", 55, 81400, "Synthetic — every 5k mi", dplus(60)),
        (dminus(40), "Rotation", 22, 80100, "With oil change", dplus(38)),
        (dminus(120), "Tires", 520, 76200, "4 new — all-season", dplus(600)),
        (dminus(70), "Brakes", 240, 78500, "Front pads + rotors", dplus(700)),
        (dminus(9), "Inspection", 35, 82200, "State + emissions", dplus(355)),
        (dminus(200), "Battery", 130, 72800, "3-yr warranty", dplus(900)),
        (dminus(30), "Repair", 95, 80900, "Cabin filter + wipers", "—"),
    ]
    ws, start, end = build_log(
        wb, "Vehicle", "🔧", "VEHICLE & MAINTENANCE",
        "Keep your money-maker running — service history, cost & what's due next.",
        ["Date", "Service", "Cost", "Odometer", "Notes", "Next Due"],
        rows, [13, 16, 12, 13, 26, 13],
        text_left={2, 5}, dates={1, 6}, money={3}, ints={4}, reserved=30,
        validations=[("B", "MaintList")])
    trow = end + 1
    ws.cell(row=trow, column=2, value="TOTAL SPENT").style = "th"
    c = ws.cell(row=trow, column=3, value=f"=SUM(C{start}:C{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    for col in (1, 4, 5, 6):
        ws.cell(row=trow, column=col).style = "td"; ws.cell(row=trow, column=col).fill = fill(SURFACE)


# ===========================================================================
# Business Expenses  — defines ExpenseTotal
# ===========================================================================
def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 16, 3, 30, 2])
    luxe_header(ws, "F", "🧾  BUSINESS EXPENSES",
                "Every cost of delivering — because deductions are money back in your pocket.")
    table_headers(ws, 4, ["Category", "This Month", "Annual (est.)"])
    exp = {"Fuel": "=FuelTotal", "Car Payment": 320, "Insurance": 165, "Maintenance & Repairs": 110,
           "Phone & Data": 45, "Bags & Supplies": 30, "Tolls": 25, "Car Wash & Cleaning": 30, "Miscellaneous": 0}
    start = L0; eend = start + len(EXP_CATS) - 1
    for i, cat in enumerate(EXP_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        val = exp[cat]
        cm = ws.cell(row=r, column=2, value=val)
        cm.style = "field_value" if cat == "Fuel" else "input"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    etot = eend + 1
    ws.cell(row=etot, column=1, value="TOTAL EXPENSES").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=etot, column=col, value=f"=SUM({L}{start}:{L}{eend})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "ExpCat", "Expenses", "A", start, eend)
    nrange(wb, "ExpMonthly", "Expenses", "B", start, eend)
    cell_name(wb, "ExpenseTotal", "Expenses", f"$B${etot}")
    ws.conditional_formatting.add(f"B{start}:B{eend}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=330, color=ACCENT, showValue=True))
    merge_set(ws, "E4:F4", "COST PER MILE", "section_gold")
    rows2 = [("Total business miles", "=TotalMiles", "#,##0"),
             ("Expenses this month", "=ExpenseTotal", '"$"#,##0'),
             ("Cost per mile", "=IFERROR(ExpenseTotal/TotalMiles,0)", '"$"#,##0.00'),
             ("Gas per mile", "=IFERROR(FuelTotal/TotalMiles,0)", '"$"#,##0.00'),
             ("Expense ratio", "=IFERROR(ExpenseTotal/GrossEarn,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 5 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Cost per mile":
            c.fill = fill(WARN_BG)


# ===========================================================================
# Monthly Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "📊  MONTHLY BUDGET",
                "Your take-home, put to work — delivery net income flows in; plan every dollar.")
    merge_set(ws, "B5:E5", "INCOME", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Source", "Planned", "Actual", ""], start_col=2)
    inc = [("Delivery (net)", "=NetGoal", "=NetEarn"), ("Side income", 400, 350), ("Other", 0, 0)]
    is_ = 7
    for i, (lab, pl, ac) in enumerate(inc):
        r = is_ + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cp = ws.cell(row=r, column=3, value=pl); cp.style = "td" if lab == "Delivery (net)" else "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=ac); ca.style = "td" if lab == "Delivery (net)" else "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=5).style = "td"; ws.cell(row=r, column=5).fill = fill(WHITE)
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ie = is_ + len(inc) - 1; itot = ie + 1
    ws.cell(row=itot, column=2, value="TOTAL INCOME").style = "th"
    for col in (3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=itot, column=col, value=f"=SUM({L}{is_}:{L}{ie})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=itot, column=5).style = "td"; ws.cell(row=itot, column=5).fill = fill(SURFACE)
    cell_name(wb, "BudgetIncome", "Budget", f"$D${itot}")
    merge_set(ws, "B12:E12", "HOUSEHOLD EXPENSES", "section_gold"); ws.row_dimensions[12].height = 22
    table_headers(ws, 13, ["Category", "Planned", "Actual", "Δ"], start_col=2)
    hh = [("Housing", 950, 950), ("Utilities", 160, 175), ("Groceries", 380, 400),
          ("Health", 200, 200), ("Debt", 220, 220), ("Savings", 250, 250),
          ("Personal", 200, 180), ("Insurance", 120, 120), ("Other", 100, 85)]
    hs = 14
    for i, (cat, pl, ac) in enumerate(hh):
        r = hs + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=3, value=pl); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=ac); ca.style = "input"; ca.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cd.style = "td"; cd.number_format = '"$"#,##0;[Red]-"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    he = hs + len(hh) - 1; htot = he + 1
    ws.cell(row=htot, column=2, value="TOTAL EXPENSES").style = "th"
    for col in (3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=htot, column=col, value=f"=SUM({L}{hs}:{L}{he})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=htot, column=5).style = "td"; ws.cell(row=htot, column=5).fill = fill(SURFACE)
    nrange(wb, "BudgetCat", "Budget", "B", hs, he)
    nrange(wb, "BudgetActual", "Budget", "D", hs, he)
    cell_name(wb, "BudgetExpense", "Budget", f"$D${htot}")
    lo = htot + 2
    merge_set(ws, f"B{lo}:C{lo}", "LEFTOVER (INCOME − EXPENSES)", "section_gold")
    c = ws.cell(row=lo, column=4, value="=BudgetIncome-BudgetExpense"); c.style = "field_value"; c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    ws.conditional_formatting.add(f"D{lo}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))


# ===========================================================================
# Tax Center
# ===========================================================================
def build_tax(wb):
    ws = wb.create_sheet("Tax Center"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 4, 26, 16, 2])
    luxe_header(ws, "G", "🧮  TAX CENTER",
                "Keep more of what you earn — the mileage deduction, set-aside & quarterly estimates.")
    merge_set(ws, "B5:C5", "MILEAGE VS ACTUAL METHOD", "section_gold"); ws.row_dimensions[5].height = 22
    left = [("Business miles", "=TotalMiles", "#,##0"),
            ("IRS mileage rate", "=MileageRate", '"$"#,##0.00'),
            ("Mileage deduction", "=TotalMiles*MileageRate", '"$"#,##0'),
            ("Actual costs (expenses)", "=ExpenseTotal", '"$"#,##0'),
            ("Better method", '=IF(TotalMiles*MileageRate>=ExpenseTotal,"Mileage","Actual")', "text"),
            ("Deduction (best)", "=MAX(TotalMiles*MileageRate,ExpenseTotal)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(left):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"
        if fmt != "text":
            c.number_format = fmt
        if lab in ("Mileage deduction", "Deduction (best)"):
            c.fill = fill(MINT_BG)
    cell_name(wb, "TaxDeduction", "Tax Center", "$C$8")
    merge_set(ws, "E5:F5", "SET-ASIDE & ESTIMATE", "section_gold"); ws.row_dimensions[5].height = 22
    right = [("Gross earnings", "=GrossEarn", '"$"#,##0'),
             ("− Mileage deduction", "=-TotalMiles*MileageRate", '"$"#,##0'),
             ("Taxable (est.)", "=MAX(GrossEarn-TotalMiles*MileageRate,0)", '"$"#,##0'),
             ("Set-aside rate", "=TaxRate", "0%"),
             ("Set aside this month", "=MAX(GrossEarn-TotalMiles*MileageRate,0)*TaxRate", '"$"#,##0'),
             ("Quarterly estimate", "=MAX(GrossEarn-TotalMiles*MileageRate,0)*TaxRate*3", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(right):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Set aside this month", "Quarterly estimate"):
            c.fill = fill(WARN_BG)
    merge_set(ws, "B13:F13", "DEDUCTION CHECKLIST — DON'T LEAVE MONEY ON THE TABLE", "section_gold"); ws.row_dimensions[13].height = 22
    checks = [("Business mileage (biggest one)", "Yes"), ("Phone & data (business %)", "Yes"),
              ("Hot bags & delivery supplies", "Yes"), ("Tolls & parking", "Yes"),
              ("Phone mount & chargers", "Yes"), ("Insulated cooler & drinks", "Yes"),
              ("Car wash & cleaning", "Yes"), ("Platform & bank fees", "Yes")]
    for i, (item, yn) in enumerate(checks):
        r = 14 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=item).style = "td_left"
        cc = ws.cell(row=r, column=col + 1, value=yn); cc.style = "td"
        add_dv(ws, f"{get_column_letter(col+1)}{r}", "YesNoList")
        ws.conditional_formatting.add(f"{get_column_letter(col+1)}{r}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# Savings & Goals  — defines SavingsSaved, TaxSetAside
# ===========================================================================
def build_savings(wb):
    ws = wb.create_sheet("Savings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 12, 2])
    luxe_header(ws, "E", "🐷  SAVINGS & GOALS",
                "Pay yourself first — build an emergency fund, tax reserve & drive toward big goals.")
    merge_set(ws, "B5:E5", "SAVINGS BUCKETS", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Bucket", "Goal", "Saved", "Progress"], start_col=2)
    buckets = [("Emergency fund", "=SavingsGoal", 2600), ("Tax reserve", "=TaxReserveGoal", 700),
               ("New(er) car fund", 6000, 1900), ("Vacation", 2000, 650)]
    bs = 7
    for i, (lab, goal, saved) in enumerate(buckets):
        r = bs + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cg = ws.cell(row=r, column=3, value=goal); cg.style = "field_value" if isinstance(goal, str) else "input"; cg.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(MIN(D{r}/C{r},1),0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    be = bs + len(buckets) - 1
    cell_name(wb, "SavingsSaved", "Savings", f"$D${bs}")
    cell_name(wb, "TaxSetAside", "Savings", f"$D${bs+1}")
    ws.conditional_formatting.add(f"E{bs}:E{be}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    merge_set(ws, "B13:E13", "DRIVER GOALS & OKRs", "section_gold"); ws.row_dimensions[13].height = 22
    table_headers(ws, 14, ["Goal", "Target", "Current", "Progress"], start_col=2)
    goals = [("$1,800 net / month", "$1,800", "$1,526", 0.85),
             ("$16 net / hour", "$16.00", "$16.96", 1.00),
             ("300 deliveries / month", "300", "321", 1.00),
             ("$4k emergency fund", "$4,000", "$2,600", 0.65),
             ("Pay off $3k card", "$3,000", "$1,800", 0.60),
             ("Lift tips to 50%", "50%", "49%", 0.99)]
    gs = 15
    for i, (g, tgt, cur, prog) in enumerate(goals):
        r = gs + i
        ws.cell(row=r, column=2, value=g).style = "td_left"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ge = gs + len(goals) - 1
    ws.conditional_formatting.add(f"E{gs}:E{ge}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GOLD_LT, showValue=True))


# ===========================================================================
# Promos & Quests
# ===========================================================================
def build_promos(wb):
    rows = [
        ("Uber Eats", "Quest", "50 trips / week", "Completed", 90, dminus(3), "Hit with room to spare"),
        ("DoorDash", "Peak Pay", "+$2 / order dinner", "Active", 48, dplus(1), "Dinner rush only"),
        ("Uber Eats", "Boosted Zone", "1.4x downtown", "Available", 0, dplus(1), "Watch the map"),
        ("DoorDash", "Challenge", "25 deliveries (Fri-Sun)", "In Progress", 60, dplus(2), "18 / 25 done"),
        ("Grubhub", "Bonus", "$5 / order lunch", "Available", 0, dplus(1), "Lunch shifts"),
        ("Uber Eats", "Consecutive", "3 trips in a row", "Completed", 15, dminus(5), "Stacked 5 streaks"),
        ("DoorDash", "Weekly Guarantee", "$900 / 60 dashes", "In Progress", 75, dplus(3), "On pace"),
    ]
    ws, start, end = build_log(
        wb, "Promos", "🎯", "PROMOS & QUESTS",
        "Never miss a promo — track every quest, peak pay & guarantee and the extra it pays.",
        ["App", "Type", "Requirement", "Status", "Reward", "Deadline", "Notes"],
        rows, [13, 15, 20, 14, 11, 13, 24],
        text_left={3, 7}, money={5}, dates={6}, reserved=30,
        validations=[("A", "PlatformList")])
    nrange(wb, "PromoReward", "Promos", "E", start, end)
    for st, cc in {"Completed": MINT_BG, "In Progress": WARN_BG, "Active": WARN_BG, "Available": SOFT_BG}.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# Hotspots & Hours
# ===========================================================================
def build_hotspots(wb):
    ws = wb.create_sheet("Hotspots"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 16, 14, 14, 26, 2])
    luxe_header(ws, "G", "📍  HOTSPOTS & HOURS",
                "Drive where it pays — the zones, days & hours that earn the most per hour.")
    merge_set(ws, "B5:F5", "BEST ZONES", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Zone", "Best Time", "Avg $/hr", "Demand", "Notes"], start_col=2)
    zones = [
        ("Downtown / LoDo", "Fri-Sat 6-9pm", 30, "High", "Restaurant density = short trips"),
        ("Cherry Creek", "Dinner rush", 32, "High", "High tips, longer drives"),
        ("University Hill", "11am-2pm, 6-9pm", 27, "Medium", "Steady campus orders"),
        ("Highlands", "Weekend brunch", 26, "Medium", "Brunch + dinner"),
        ("Stadium district", "Event nights", 34, "High", "Check the event calendar"),
        ("Suburbs", "Weekday dinner", 21, "Low", "Long drives — reposition"),
    ]
    zs = 7
    for i, (z, t, hr, dem, note) in enumerate(zones):
        r = zs + i
        ws.cell(row=r, column=2, value=z).style = "td_left"
        ws.cell(row=r, column=3, value=t).style = "td"
        ch = ws.cell(row=r, column=4, value=hr); ch.style = "td"; ch.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=dem).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ze = zs + len(zones) - 1
    nrange(wb, "ZoneName", "Hotspots", "B", zs, ze)
    nrange(wb, "ZoneRate", "Hotspots", "D", zs, ze)
    ws.conditional_formatting.add(f"D{zs}:D{ze}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=34, color=PRIMARY, showValue=True))
    for st, cc in {"High": MINT_BG, "Medium": WARN_BG, "Low": WHITE}.items():
        ws.conditional_formatting.add(f"E{zs}:E{ze}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    merge_set(ws, "B15:F15", "BEST HOURS & DAYS", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Window", "Avg $/hr", "Why", "Rating"], start_col=2)
    days = [("Fri dinner", 33, "Weekend rush + tips", "★★★★★"), ("Sat dinner", 34, "Peak demand", "★★★★★"),
            ("Lunch (11-2)", 26, "Office + campus", "★★★★☆"), ("Sun brunch", 25, "Late-morning orders", "★★★☆☆"),
            ("Late night", 28, "Bars 9pm-1am", "★★★★☆")]
    ds = 17
    for i, (d, hr, why, rate) in enumerate(days):
        r = ds + i
        ws.cell(row=r, column=2, value=d).style = "td_left"
        cr = ws.cell(row=r, column=3, value=hr); cr.style = "td"; cr.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=why).style = "td"
        ws.cell(row=r, column=5, value=rate).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)


# ===========================================================================
# Ratings & Feedback
# ===========================================================================
def build_ratings(wb):
    ws = wb.create_sheet("Ratings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 12, 12, 14, 30, 2])
    luxe_header(ws, "G", "⭐  RATINGS & FEEDBACK",
                "Protect your rating — it drives your access to the best orders & promos.")
    merge_set(ws, "B5:F5", "PLATFORM STANDING", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Metric", "Score", "Count", "Trend", "Customers Say"], start_col=2)
    rows = [
        ("Uber Eats Rating", 4.92, 3820, "▲ +0.01", "Fast, friendly, food sealed"),
        ("DoorDash Rating", 4.89, 2840, "► flat", "On-time, careful handling"),
        ("Completion Rate", 0.99, 0, "► flat", "Almost never unassigns"),
        ("On-Time / Early", 0.96, 0, "▲ +2%", "Knows the fast routes"),
    ]
    start = 7
    for i, (plat, rt, cnt, trend, note) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=plat).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rt); cr.style = "td"
        cr.number_format = "0.00" if rt > 3 else "0%"
        cc = ws.cell(row=r, column=4, value=cnt if cnt else "—"); cc.style = "td"
        if cnt:
            cc.number_format = "#,##0"
        ws.cell(row=r, column=5, value=trend).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    ws.conditional_formatting.add(f"C{start}:C{start+1}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=5, color=GOLD_LT, showValue=True))
    merge_set(ws, "B12:F12", "5-STAR CHECKLIST", "section_gold"); ws.row_dimensions[12].height = 22
    tips = ["Use an insulated hot bag — every order", "Double-check the order & the address",
            "Follow drop-off notes exactly (leave at door / hand it)", "Send a quick 'on my way' / 'arrived' text",
            "Keep drinks upright & bags sealed", "Be quick but never unsafe — accuracy over speed"]
    for i, t in enumerate(tips):
        r = 13 + (i // 2)
        col = 2 if i % 2 == 0 else 4
        merge_set(ws, f"{get_column_letter(col)}{r}:{get_column_letter(col+1)}{r}", "✓  " + t, "td_left")


# ===========================================================================
# Analytics — Driver Health Score + computed cells
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 3, 20, 12, 12, 2])
    luxe_header(ws, "G", "📈  ANALYTICS COMMAND CENTER",
                "Your delivery business by the numbers — take-home, health dimensions & the Driver Health Score.")
    merge_set(ws, "B5:C5", "TAKE-HOME SNAPSHOT", "section")
    snap = [("Gross earnings", "=GrossEarn", '"$"#,##0', None),
            ("Business expenses", "=ExpenseTotal", '"$"#,##0', None),
            ("Net earnings", "=GrossEarn-ExpenseTotal", '"$"#,##0', "NetEarn"),
            ("Online hours", "=TotalHours", "0.0", None),
            ("Net $ / hour", "=IFERROR((GrossEarn-ExpenseTotal)/TotalHours,0)", '"$"#,##0.00', "PerHour"),
            ("Net $ / mile", "=IFERROR((GrossEarn-ExpenseTotal)/TotalMiles,0)", '"$"#,##0.00', "PerMile"),
            ("Deliveries", "=TotalOrders", "#,##0", None),
            ("Mileage deduction", "=TotalMiles*MileageRate", '"$"#,##0', None)]
    for i, (lab, fml, fmt, nm) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if nm:
            cell_name(wb, nm, "Analytics", f"$C${r}")
        if lab in ("Net earnings", "Net $ / hour"):
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    merge_set(ws, "E5:G5", "DRIVER HEALTH", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=5)
    metrics = [
        ("Net earnings vs goal", "=IFERROR(MIN(NetEarn/NetGoal,1),0)"),
        ("Net $/hour vs target", "=IFERROR(MIN(PerHour/HourlyTarget,1),0)"),
        ("Deliveries vs goal", "=IFERROR(MIN(TotalOrders/DeliveryGoal,1),0)"),
        ("Emergency fund", "=IFERROR(MIN(SavingsSaved/SavingsGoal,1),0)"),
        ("Driving consistency", "=IFERROR(MIN(COUNT(ShiftMiles)/ShiftGoal,1),0)"),
        ("Tax reserve", "=IFERROR(MIN(TaxSetAside/TaxReserveGoal,1),0)"),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(metrics):
        r = hs + i
        ws.cell(row=r, column=5, value=dim).style = "td_left"
        c = ws.cell(row=r, column=6, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=7, value=f'=IF(F{r}>=0.75,"Strong",IF(F{r}>=0.5,"Growing","Focus"))').style = "td"
        if i % 2:
            for c2 in range(5, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(metrics) - 1
    cell_name(wb, "HealthRange", "Analytics", f"$F${hs}:$F${he}")
    ws.conditional_formatting.add(f"F{hs}:F{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "B15:G15", "BEST SHIFTS (BY EARNINGS)", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Shift", "Earnings", "Hours", "$/hr", "Miles"], start_col=2)
    best = [("Fri dinner rush", 205, 6.5, 118), ("Sat dinner rush", 198, 6.5, 114),
            ("Sat multi-app", 195, 6.5, 112), ("Wed DoorDash", 182, 6.0, 102),
            ("Weeknight multi", 182, 6.0, 108), ("Tue peak pay", 179, 6.0, 102)]
    vs = 17
    for i, (title, earn, hrs, miles) in enumerate(best):
        r = vs + i
        ws.cell(row=r, column=2, value=title).style = "td_left"
        ce = ws.cell(row=r, column=3, value=earn); ce.style = "td"; ce.number_format = '"$"#,##0'
        ch = ws.cell(row=r, column=4, value=hrs); ch.style = "td"; ch.number_format = "0.0"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cp.style = "td"; cp.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=6, value=miles); cm.style = "td"; cm.number_format = "#,##0"
        if i % 2:
            for c2 in range(2, 7):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    ve = vs + len(best) - 1
    nrange(wb, "BestShift", "Analytics", "B", vs, ve)
    nrange(wb, "BestEarn", "Analytics", "C", vs, ve)
    ws.conditional_formatting.add(f"C{vs}:C{ve}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=205, color=PRIMARY, showValue=True))
    merge_set(ws, "B25:C25", "NET EARNINGS — 6 MONTHS", "section")
    ws.cell(row=26, column=2, value="Month").style = "th"; ws.cell(row=26, column=3, value="Net ($)").style = "th"
    months = _recent_months(6); vals = [1280, 1360, 1440, 1410, 1490, 1526]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 27 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "TrendMonth", "Analytics", "$B$27:$B$32")
    cell_name(wb, "TrendVal", "Analytics", "$C$27:$C$32")


# ===========================================================================
# Weekly Planner
# ===========================================================================
def build_planner(wb):
    ws = wb.create_sheet("Planner"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 16, 14, 14, 26, 2])
    luxe_header(ws, "G", "🗓  WEEKLY PLANNER",
                "Plan the week that hits your goal — target hours, zones & the earnings you need.")
    table_headers(ws, 4, ["Day", "Plan", "Target Hrs", "$ Goal", "Zone / Focus"])
    days = [
        ("Monday", "Off", 0, 0, "Rest / car maintenance"),
        ("Tuesday", "Dinner", 5, 90, "Downtown restaurants"),
        ("Wednesday", "Lunch + dinner", 6, 110, "Campus + downtown"),
        ("Thursday", "Dinner", 5, 95, "Cherry Creek"),
        ("Friday", "Dinner rush", 6.5, 150, "LoDo + events"),
        ("Saturday", "Dinner rush", 6.5, 155, "Highlands + downtown"),
        ("Sunday", "Brunch + dinner", 5, 100, "Brunch spots"),
    ]
    start = L0
    for i, (d, plan, hrs, goal, zone) in enumerate(days):
        r = start + i
        ws.cell(row=r, column=1, value=d).style = "td_left"
        ws.cell(row=r, column=2, value=plan).style = "input"
        ch = ws.cell(row=r, column=3, value=hrs); ch.style = "input"; ch.number_format = "0.0"
        cg = ws.cell(row=r, column=4, value=goal); cg.style = "input"; cg.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=zone).style = "td_left"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(days) - 1
    trow = end + 1
    ws.cell(row=trow, column=1, value="WEEK TOTAL").style = "th"
    ws.cell(row=trow, column=2).style = "th"
    for col, fmt in ((3, "0.0"), (4, '"$"#,##0')):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt
    ws.cell(row=trow, column=5).style = "td"; ws.cell(row=trow, column=5).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


# ===========================================================================
# Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  RECEIPTS & VEHICLE GALLERY",
                "Your paper trail — snap receipts, odometer & maintenance photos for tax time.")
    merge_set(ws, "B5:D5", "HOW TO ADD PHOTOS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Excel: Insert ▸ Pictures ▸ Place in Cell (or drag an image) into any framed box below. "
                      "Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"paste-link-here\"). "
                      "Caption each one — date, amount & category — so tax season is a five-minute job.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Gas Receipt", "Maintenance", "Odometer", "Hot Bag / Supplies", "Repair Invoice", "Car Wash"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add photo)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Date · amount · category…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🛵  UBER EATS DRIVER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Earnings, miles, taxes & savings — your whole delivery business, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("NET EARNINGS", "=NetEarn", "money"),
        ("GROSS EARNINGS", "=GrossEarn", "money"),
        ("TIPS", "=TotalTips", "money"),
        ("NET $ / HOUR", "=PerHour", "money2"),
        ("NET $ / MILE", "=PerMile", "money2"),
        ("MILES DRIVEN", "=TotalMiles", "num"),
    ]
    row2 = [
        ("ONLINE HOURS", "=TotalHours", "dec"),
        ("DELIVERIES", "=TotalOrders", "num"),
        ("EXPENSES", "=ExpenseTotal", "money"),
        ("TAX DEDUCTION", "=TotalMiles*MileageRate", "money"),
        ("SAVINGS", "=IFERROR(MIN(SavingsSaved/SavingsGoal,1),0)", "pct"),
        ("DRIVER HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "EARNINGS & TAKE-HOME", "section_gold")
    ln = LineChart(); ln.title = "Net Earnings ($) — 6 mo"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=26, max_row=32), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=27, max_row=32)); ln.legend = None
    ws.add_chart(ln, "B12")
    d1 = DoughnutChart(); d1.title = "Earnings Mix"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Earnings"], min_col=3, min_row=6, max_row=9), titles_from_data=True)
    d1.set_categories(Reference(wb["Earnings"], min_col=2, min_row=7, max_row=9)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "BEST SHIFTS & EXPENSES", "section_gold")
    cb = BarChart(); cb.type = "bar"; cb.title = "Best Shifts by Earnings"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); cb.legend = None
    ws.add_chart(cb, "B30")
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Expenses"], min_col=2, min_row=4, max_row=13), titles_from_data=True)
    eb.set_categories(Reference(wb["Expenses"], min_col=1, min_row=5, max_row=13)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Uber Eats Driver Command Center™ — every dash, mile & dollar in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_deliverylog(wb)
    build_earnings(wb); build_mileage(wb); build_fuel(wb); build_vehicle(wb)
    build_expenses(wb); build_budget(wb); build_tax(wb); build_savings(wb)
    build_promos(wb); build_hotspots(wb); build_ratings(wb); build_analytics(wb)
    build_planner(wb); build_gallery(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Driver Profile", "Delivery Log", "Earnings", "Mileage", "Fuel Log",
             "Vehicle", "Expenses", "Budget", "Tax Center", "Savings", "Promos", "Hotspots", "Ratings",
             "Analytics", "Planner", "Gallery", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Uber_Eats_Driver_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
