"""Build Family Vacation Command Center™ — The Ultimate Family Travel Planning
& Memory Management System.

24 sheets + Welcome · a premium family travel operating system in Excel & Sheets.
Dashboard, family profile, trip planner, budget, savings, itinerary, transport,
hotels, packing, kids, meals, activities, reservations, documents, emergency,
road trip, responsibilities, rewards, photo vault, journal, review, wishlist,
analytics & settings — one beautiful command center.

Run: python3 build_xlsx.py   ->  ../Family_Vacation_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

TRIP_TYPES = ["Disney / Theme Park", "Beach", "Road Trip", "Cruise", "International",
              "City Break", "Camping", "Ski", "All-Inclusive", "Multi-Gen"]
EXP_CATS = ["Flights", "Hotels", "Rental Car", "Transportation", "Food & Dining",
            "Activities", "Shopping", "Insurance", "Emergency Fund", "Miscellaneous"]
PACK_CATS = ["Adults", "Kids", "Baby", "Toiletries", "Clothing", "Electronics",
             "Beach Gear", "Documents", "Medicine", "Entertainment", "Comfort"]
ACT_TYPES = ["Theme Park", "Beach", "Tour", "Dining", "Show", "Museum", "Outdoor", "Shopping", "Relax"]
DOC_TYPES = ["Passport", "Visa", "Insurance", "Tickets", "Hotel Confirmation", "Medical", "ID"]
TRANS_TYPES = ["Flight", "Rental Car", "Train", "Bus", "Transfer", "Parking", "Rideshare"]
RES_TYPES = ["Flight", "Hotel", "Rental Car", "Activity", "Restaurant", "Show", "Transfer", "Park Ticket"]
MEALS = ["Breakfast", "Lunch", "Dinner", "Snack"]
ROLES = ["Packing Leader", "Navigator", "Snack Manager", "Entertainment Manager",
         "Document Manager", "Photographer", "Budget Keeper"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]
PAY_STATUS = ["Paid", "Deposit", "Pending", "Refunded"]
DOC_STATUS = ["Ready", "In Progress", "Missing", "N/A"]

MEMBERS = ["Marcus (Dad)", "Elena (Mom)", "Sofia", "Diego", "Mateo"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def totals(ws, row, cols, start, end, fmt='"$"#,##0', label="TOTAL"):
    ws.cell(row=row, column=1, value=label).style = "th"
    for col in cols:
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt


# ===========================================================================
# 24 — Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 22, 3] + [16] * 7)
    luxe_header(ws, "K", "⚙  SETTINGS", "Set your trip details once — every dashboard follows.")
    merge_set(ws, "B5:C5", "TRIP INPUTS", "section")
    controls = [
        ("Family Name", "The Rivera Family", None, "FamilyName"),
        ("Trip Name", "Walt Disney World 2026", None, "TripName"),
        ("Destination", "Orlando, Florida", None, "Destination"),
        ("Trip Type", "Disney / Theme Park", None, "TripType"),
        ("Trip Start", dplus(45), "mm/dd/yyyy", "TripStart"),
        ("Trip End", dplus(51), "mm/dd/yyyy", "TripEnd"),
        ("Travelers", 5, "0", "TravelerCount"),
        ("Savings Goal", 8500, '"$"#,##0', "SavingsGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Trip Type", TRIP_TYPES, "TripTypeList"), ("F", "Expense Category", EXP_CATS, "ExpCatList"),
             ("G", "Family Member", MEMBERS, "MemberList"), ("H", "Activity Type", ACT_TYPES, "ActTypeList"),
             ("I", "Packing Category", PACK_CATS, "PackCatList"), ("J", "Transport Type", TRANS_TYPES, "TransTypeList"),
             ("K", "Document Type", DOC_TYPES, "DocTypeList")]
    merge_set(ws, "E5:K5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 18, "Priority", PRIORITIES, "PriorityList"), ("F", 18, "Yes / No", YESNO, "YesNoList"),
             ("G", 18, "Payment", PAY_STATUS, "PayStatusList"), ("H", 18, "Doc Status", DOC_STATUS, "DocStatusList"),
             ("I", 18, "Reservation", RES_TYPES, "ResTypeList"), ("J", 18, "Role", ROLES, "RoleList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🌎  FAMILY VACATION COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate family travel planning & memory management system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "PLAN EVERY TRIP FROM ONE BEAUTIFUL COMMAND CENTER", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("From the first idea to the final photo — Family Vacation Command Center™ manages budgeting, "
                      "scheduling, packing, transportation, accommodation, meals, kids' activities, documents, "
                      "emergencies and memories in ONE premium Excel & Google Sheets system. Remove the travel stress, "
                      "keep the whole family organized, and create unforgettable trips — all with app-level automation. "
                      "This isn't a vacation checklist — it's your complete Family Travel Operating System.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your family name, trip, dates, travelers & savings goal.",
             "2.  Fill the Family Profile — sizes, allergies & emergency contacts.",
             "3.  Build your Budget & Savings — plan every category and track your fund.",
             "4.  Plan the Itinerary, Hotels, Flights, Activities & Meals.",
             "5.  Work the Packing, Documents & Reservations checklists — watch % climb.",
             "6.  Travel, then capture it all in the Photo Vault, Journal & Post-Trip Review."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("A realistic sample trip (the Rivera family's 7-day, 5-person Walt Disney World vacation on an $8,500 "
               "budget) is included so you can see how everything connects — just type over it with your own. The "
               "countdown, budget, savings progress, packing %, booking %, documents % and the Family Trip Readiness "
               "Score all update automatically. Every sheet is print-friendly and works in Excel and Google Sheets, "
               "on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more time to make memories — let's plan your best trip yet.", "section_gold")


# ===========================================================================
# 2 — Family Profile
# ===========================================================================
def build_profile(wb):
    rows = [
        ("Marcus (Dad)", 41, "Mar 12", "L / 34x32", "10.5", "None", "Coffee, spicy", "Elena — 555-0142", "—"),
        ("Elena (Mom)", 39, "Aug 03", "M / 8", "8", "Shellfish", "Vegetarian", "Marcus — 555-0143", "EpiPen in bag"),
        ("Sofia", 12, "Jan 22", "Youth L", "6", "Peanuts", "No nuts", "Parents", "Carries inhaler"),
        ("Diego", 9, "Jun 15", "Youth M", "3", "None", "Picky — plain", "Parents", "Motion sickness"),
        ("Mateo", 4, "Nov 30", "5T", "11 (toddler)", "Dairy", "Lactose-free", "Parents", "Naps 1–3pm"),
    ]
    ws, start, end = build_log(
        wb, "Family Profile", "👨‍👩‍👧‍👦", "FAMILY PROFILE",
        "Everyone in one place — sizes, allergies, preferences & emergency contacts.",
        ["Family Member", "Age", "Birthday", "Clothing Size", "Shoe", "Allergies", "Food Prefs", "Emergency Contact", "Special Notes"],
        rows, [18, 7, 12, 15, 10, 14, 16, 20, 20],
        text_left={1, 6, 7, 8, 9}, ints={2}, reserved=14)
    nrange(wb, "MemberName", "Family Profile", "A", start, end)
    cell_name(wb, "MemberCount", "Family Profile", f"$A${start}:$A${end}")


# ===========================================================================
# 3 — Master Trip Planner
# ===========================================================================
def build_master(wb):
    ws = wb.create_sheet("Master Trip"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 4, 24, 30, 2])
    luxe_header(ws, "G", "🧭  MASTER TRIP PLANNER", "The big picture — one trip, one plan, from before to after.")
    blocks = [
        ("THE TRIP", [("Trip Name", "=TripName"), ("Destination", "=Destination"),
                      ("Trip Type", "=TripType"), ("Travelers", "=TravelerCount"),
                      ("Trip Start", "=TEXT(TripStart,\"ddd, mmm d\")"), ("Trip End", "=TEXT(TripEnd,\"ddd, mmm d\")")]),
        ("OVERVIEW", [("Days Traveling", "=TripEnd-TripStart+1"), ("Countdown (days)", "=TripStart-TODAY()"),
                      ("Theme", "Disney magic + rest days"), ("Weather (est.)", "Warm · 88°F · afternoon rain"),
                      ("Home Base", "On-site resort"), ("Trip Motto", "Slow mornings, big memories")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    # timeline
    merge_set(ws, "B15:F15", "TRIP TIMELINE", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Phase", "Key Tasks", "Owner", "Done?"], start_col=2)
    phases = [
        ("Before Trip", "Book flights & hotel · save fund · documents · packing", "Elena", "In progress"),
        ("Getting There", "Airport 3h early · car seats · snacks · entertainment", "Marcus", "Planned"),
        ("During Trip", "Follow itinerary · track budget · capture photos daily", "Family", "Planned"),
        ("After Trip", "Journal · review · sort photos · plan next adventure", "Sofia", "Planned"),
    ]
    for i, (ph, task, owner, done) in enumerate(phases):
        r = 17 + i
        ws.cell(row=r, column=2, value=ph).style = "field_label"
        ws.cell(row=r, column=3, value=task).style = "td_left"
        ws.cell(row=r, column=5, value=owner).style = "td"
        ws.cell(row=r, column=6, value=done).style = "td"
        ws.row_dimensions[r].height = 26


# ===========================================================================
# 4 — Budget Command Center
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 13, 13, 13, 3, 22, 14, 2])
    luxe_header(ws, "H", "💰  FAMILY VACATION BUDGET COMMAND CENTER",
                "Every dollar planned — budget vs actual, cost per person & per day, live.")
    # expenses
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining"])
    planned = {"Flights": 2200, "Hotels": 2400, "Rental Car": 450, "Transportation": 200,
               "Food & Dining": 1300, "Activities": 1200, "Shopping": 300, "Insurance": 250,
               "Emergency Fund": 200, "Miscellaneous": 0}
    actual = {"Flights": 2200, "Hotels": 2400, "Rental Car": 0, "Transportation": 0,
              "Food & Dining": 0, "Activities": 600, "Shopping": 0, "Insurance": 0,
              "Emergency Fund": 0, "Miscellaneous": 0}
    start = L0; end = start + len(EXP_CATS) - 1
    for i, cat in enumerate(EXP_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=2, value=planned[cat]); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=actual[cat]); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL BUDGET").style = "th"
    for col in (2, 3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG), font=Font(color=DANGER, bold=True)))
    nrange(wb, "ExpCat", "Budget", "A", start, end)
    nrange(wb, "ExpPlanned", "Budget", "B", start, end)
    nrange(wb, "ExpActual", "Budget", "C", start, end)
    cell_name(wb, "TripBudget", "Budget", f"$B${tot}")
    cell_name(wb, "SpentTotal", "Budget", f"$C${tot}")
    # savings sources (right)
    merge_set(ws, "G4:H4", "VACATION FUND", "section_gold")
    sources = [("Vacation Fund", 4200), ("Monthly Savings", 1800), ("Gift Contributions", 500),
               ("Points / Miles (value)", 200), ("Cash-Back Rewards", 100)]
    sstart = 5
    for i, (src, amt) in enumerate(sources):
        r = sstart + i
        ws.cell(row=r, column=7, value=src).style = "td_left"
        c = ws.cell(row=r, column=8, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=7).fill = fill(MUTED_ROW); ws.cell(row=r, column=8).fill = fill(MUTED_ROW)
    send = sstart + len(sources) - 1; stot = send + 1
    ws.cell(row=stot, column=7, value="MONEY SAVED").style = "th"
    c = ws.cell(row=stot, column=8, value=f"=SUM(H{sstart}:H{send})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "SavedTotal", "Budget", f"$H${stot}")
    # bottom line
    merge_set(ws, "A17:D17", "THE BOTTOM LINE", "section_gold")
    lines = [("Total budget", "=TripBudget", '"$"#,##0'), ("Money saved", "=SavedTotal", '"$"#,##0'),
             ("Still to save", "=MAX(TripBudget-SavedTotal,0)", '"$"#,##0'), ("Spent so far", "=SpentTotal", '"$"#,##0'),
             ("Budget remaining", "=TripBudget-SpentTotal", '"$"#,##0'), ("Cost / person", "=IFERROR(TripBudget/TravelerCount,0)", '"$"#,##0'),
             ("Cost / day", "=IFERROR(TripBudget/(TripEnd-TripStart+1),0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(lines):
        r = 18 + i
        ws.cell(row=r, column=1, value=lab).style = "field_label"
        c = ws.cell(row=r, column=2, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Budget remaining", "Money saved"):
            ws.cell(row=r, column=2).fill = fill(MINT_BG)


# ===========================================================================
# 5 — Savings Planner
# ===========================================================================
def build_savings(wb):
    ws = wb.create_sheet("Savings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 4, 16, 14, 14, 14, 2])
    luxe_header(ws, "H", "🏦  SAVINGS PLANNER", "Fund the fun — set a goal and watch the vacation fund grow.")
    merge_set(ws, "B5:C5", "SAVINGS GOAL", "section")
    goal = [("Vacation Goal", "=TripName"), ("Target Amount", "=SavingsGoal"), ("Current Savings", "=SavedTotal"),
            ("Still to Save", "=MAX(SavingsGoal-SavedTotal,0)"), ("Monthly Contribution", 800),
            ("Months to Goal", "=IFERROR(ROUNDUP(MAX(SavingsGoal-SavedTotal,0)/C10,0),0)"),
            ("Progress", "=IFERROR(MIN(SavedTotal/SavingsGoal,1),0)")]
    for i, (lab, val) in enumerate(goal):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val)
        c.style = "input" if lab == "Monthly Contribution" else "field_value"
        if lab in ("Target Amount", "Current Savings", "Still to Save", "Monthly Contribution"):
            c.number_format = '"$"#,##0'
        if lab == "Progress":
            c.number_format = "0%"; c.fill = fill(MINT_BG)
    ws.conditional_formatting.add("C12", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT))
    # monthly ledger
    merge_set(ws, "E5:H5", "SAVINGS LEDGER", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Month", "Planned", "Actual", "Running Total"], start_col=5)
    months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul"]; actuals = [1000, 1200, 1100, 1200, 1300, 1000]
    lstart = 7
    for i, (m, a) in enumerate(zip(months, actuals)):
        r = lstart + i
        ws.cell(row=r, column=5, value=m).style = "td_left"
        cp = ws.cell(row=r, column=6, value=800); cp.style = "td"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=7, value=a); ca.style = "input"; ca.number_format = '"$"#,##0'
        rt = ws.cell(row=r, column=8, value=f"=SUM($G${lstart}:G{r})"); rt.style = "td"; rt.number_format = '"$"#,##0'
        if i % 2:
            for c in range(5, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)


# ===========================================================================
# 6 — Itinerary
# ===========================================================================
def build_itinerary(wb):
    rows = [
        (0, "Travel & Check-In", "Fly MCO · resort check-in", "Airport / resort", "Disney Springs stroll", "Quick-service dinner", "Rest & pool", "Arrival day — keep it easy"),
        (1, "Magic Kingdom", "Rope drop · headliners", "Magic Kingdom", "Parade & lunch", "Fireworks", "Character dining 6pm", "FastPass the mountains"),
        (2, "EPCOT", "Test Track · Frozen", "EPCOT", "World Showcase snacks", "Ride Guardians", "Space 220 lunch", "Nap for Mateo 1–3pm"),
        (3, "Rest & Pool Day", "Sleep in · resort pool", "Resort", "Mini-golf", "Movie night in", "Grocery pickup", "Recharge — no park"),
        (4, "Hollywood Studios", "Rise of Resistance", "Hollywood Studios", "Toy Story Land", "Fantasmic!", "Sci-Fi Diner 12:30", "Single-rider for adults"),
        (5, "Animal Kingdom", "Flight of Passage", "Animal Kingdom", "Safari midday", "Kilimanjaro", "Tusker House 8am", "Water bottles a must"),
        (6, "Departure", "Souvenirs · pack", "Resort → MCO", "Last breakfast", "Fly home", "—", "Bags down by 9am"),
    ]
    sample = [(dplus(45 + d), title, morning, loc, aft, eve, meals, notes) for (d, title, morning, loc, aft, eve, meals, notes) in rows]
    ws, start, end = build_log(
        wb, "Itinerary", "🗓", "ITINERARY COMMAND CENTER",
        "Day by day — every morning, afternoon and evening planned, with reservations & notes.",
        ["Date", "Day / Theme", "Morning", "Location", "Afternoon", "Evening", "Meals / Res.", "Notes"],
        sample, [13, 20, 22, 18, 20, 18, 20, 24],
        text_left={2, 3, 4, 5, 6, 7, 8}, dates={1}, reserved=21)
    nrange(wb, "ItinDate", "Itinerary", "A", start, end)


# ===========================================================================
# 7 — Flights & Transportation
# ===========================================================================
def build_transport(wb):
    rows = [
        ("Flight", "Delta 1422", "MCO Orlando", "Confirmed", 5, "$2,200", "DL-8KX2P9", "Seats 22A–E · 8:05am"),
        ("Rental Car", "Enterprise SUV", "MCO pickup", "Confirmed", 5, "$450", "ENT-77341", "Full-size · car seat added"),
        ("Transfer", "Resort shuttle", "Airport → resort", "Included", 5, "$0", "RES-INCL", "Free with stay"),
        ("Parking", "Resort self-park", "Resort", "Pending", 1, "$120", "—", "$20/night x 6"),
        ("Rideshare", "Uber XL (budget)", "Disney Springs", "Planned", 5, "$80", "—", "Est. 2 trips"),
        ("Flight", "Delta 1889", "MCO → home", "Confirmed", 5, "$0", "DL-8KX2P9", "Return 5:40pm · same PNR"),
    ]
    ws, start, end = build_log(
        wb, "Transport", "✈", "FLIGHT & TRANSPORTATION MANAGER",
        "Every leg covered — flights, car, transfers & parking with confirmations.",
        ["Type", "Carrier / Vehicle", "Route", "Status", "Seats", "Cost", "Confirmation", "Notes"],
        rows, [13, 20, 20, 13, 8, 12, 16, 24],
        text_left={2, 3, 7, 8}, ints={5}, money={6},
        validations=[("A", "TransTypeList")], reserved=20)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Confirmed"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Pending"'], fill=fill(WARN_BG)))


# ===========================================================================
# 8 — Hotel & Accommodation
# ===========================================================================
def build_hotels(wb):
    rows = [
        ("Disney Caribbean Beach", "900 Cayman Way, Orlando", "WDW-CB-55219", dplus(45), dplus(51), "Preferred · 2 queens", "$2,400", "Pool, transport, dining plan", "407-934-3400"),
        ("Airport Hyatt (1 night)", "MCO Terminal, Orlando", "HY-99120", dplus(51), dplus(52), "2 doubles", "$0", "Late flight backup — hold only", "407-825-1234"),
    ]
    ws, start, end = build_log(
        wb, "Hotels", "🏨", "HOTEL & ACCOMMODATION MANAGER",
        "Where you'll rest — confirmations, dates, room details & amenities in one place.",
        ["Hotel / Rental", "Address", "Confirmation", "Check-In", "Check-Out", "Room", "Cost", "Amenities", "Phone"],
        rows, [24, 26, 16, 13, 13, 18, 12, 26, 15],
        text_left={1, 2, 3, 6, 8, 9}, dates={4, 5}, money={7}, reserved=12)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        CellIsRule(operator="notEqual", formula=['""'], fill=fill(MINT_BG)))


# ===========================================================================
# 9 — Packing Command Center
# ===========================================================================
def build_packing(wb):
    rows = [
        ("Passports & IDs", "Documents", "Marcus", 5, "Yes"), ("Park tickets (mobile)", "Documents", "Elena", 5, "Yes"),
        ("Phone chargers", "Electronics", "Marcus", 3, "Yes"), ("Portable battery", "Electronics", "Marcus", 2, "Yes"),
        ("Camera + SD cards", "Electronics", "Sofia", 1, "No"), ("Sunscreen SPF 50", "Toiletries", "Elena", 3, "Yes"),
        ("Toiletry bags", "Toiletries", "Elena", 5, "Yes"), ("Family medicine kit", "Medicine", "Elena", 1, "Yes"),
        ("Mateo's allergy meds", "Medicine", "Elena", 1, "Yes"), ("Motion-sickness bands", "Medicine", "Elena", 2, "No"),
        ("T-shirts", "Clothing", "Each", 21, "Yes"), ("Shorts", "Clothing", "Each", 14, "Yes"),
        ("Swimsuits", "Clothing", "Each", 10, "Yes"), ("Rain ponchos", "Clothing", "Marcus", 5, "No"),
        ("Comfortable shoes", "Clothing", "Each", 10, "Yes"), ("Light jackets", "Clothing", "Each", 5, "No"),
        ("Sun hats", "Beach Gear", "Each", 5, "Yes"), ("Refillable water bottles", "Beach Gear", "Diego", 5, "Yes"),
        ("Cooling towels", "Beach Gear", "Elena", 5, "No"), ("Stroller (Mateo)", "Kids", "Marcus", 1, "Yes"),
        ("Kids tablets + cases", "Kids", "Sofia", 2, "Yes"), ("Coloring books", "Kids", "Diego", 3, "Yes"),
        ("Snacks for travel", "Kids", "Elena", 1, "No"), ("Mateo comfort blanket", "Comfort", "Mateo", 1, "Yes"),
        ("Neck pillows", "Comfort", "Each", 5, "No"), ("Autograph books & pens", "Entertainment", "Sofia", 3, "Yes"),
        ("Glow sticks (fireworks)", "Entertainment", "Diego", 6, "No"), ("Laundry bag", "Adults", "Elena", 1, "Yes"),
        ("First-day outfit", "Baby", "Mateo", 1, "Yes"), ("Wet wipes", "Baby", "Elena", 3, "Yes"),
        ("Ziploc bags", "Adults", "Marcus", 10, "Yes"), ("Umbrella", "Adults", "Marcus", 2, "No"),
    ]
    ws, start, end = build_log(
        wb, "Packing", "🧳", "FAMILY PACKING COMMAND CENTER",
        "Nothing forgotten — assign items, track quantity and check them off as you pack.",
        ["Item", "Category", "Assigned To", "Qty", "Packed?"],
        rows, [28, 16, 15, 8, 12],
        text_left={1}, ints={4},
        validations=[("B", "PackCatList"), ("C", "MemberList"), ("E", "YesNoList")], reserved=60)
    nrange(wb, "PackStatus", "Packing", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 10 — Kids Travel Organizer
# ===========================================================================
def build_kids(wb):
    rows = [
        ("Sofia", "Autograph hunting, coasters", "Camera, journal", "Percy Jackson series", "Tablet + headphones", "Trail mix", "Star pillow", "Photographer of the day"),
        ("Diego", "Character meet & greets", "Star Wars figures", "Dog Man comics", "Nintendo Switch", "Fruit snacks, pretzels", "Dinosaur plush", "Snack manager"),
        ("Mateo", "Dumbo, splash pad", "Blocks, toy cars", "Board books", "Kids tablet", "Goldfish, cheese", "Blue blanket + bear", "Nap keeper"),
    ]
    ws, start, end = build_log(
        wb, "Kids", "🧒", "KIDS TRAVEL ORGANIZER",
        "Happy kids, happy trip — favorites, entertainment, snacks & little jobs.",
        ["Child", "Favorite Activities", "Toys", "Books", "Electronics", "Snacks", "Comfort Items", "Responsibility"],
        rows, [12, 24, 18, 18, 18, 20, 18, 20],
        text_left={2, 3, 4, 5, 6, 7, 8}, reserved=10)


# ===========================================================================
# 11 — Meal Planner
# ===========================================================================
def build_meals(wb):
    rows = [
        (dplus(45), "Dinner", "Disney Springs — quick service", "No", "$60", "Casual arrival meal", "Nut-free options"),
        (dplus(46), "Dinner", "Cinderella's Royal Table", "Yes 6:00pm", "$250", "Character dining", "Kids menu · dairy-free for Mateo"),
        (dplus(47), "Lunch", "Space 220", "Yes 12:30pm", "$180", "Bucket-list lunch", "Vegetarian for Elena"),
        (dplus(48), "Dinner", "Resort — grocery cook", "No", "$40", "Rest-day home meal", "Simple & cheap"),
        (dplus(49), "Lunch", "Sci-Fi Dine-In", "Yes 12:30pm", "$120", "Fun theming", "Diego = plain burger"),
        (dplus(50), "Breakfast", "Tusker House", "Yes 8:00am", "$140", "Pre-safari fuel", "Shellfish-free"),
        (dplus(51), "Breakfast", "Resort grab-and-go", "No", "$35", "Travel-day breakfast", "Fast checkout"),
    ]
    ws, start, end = build_log(
        wb, "Meals", "🍽", "MEAL PLANNER",
        "Fed & happy — reservations, budget and every diet requirement covered.",
        ["Date", "Meal", "Where", "Reservation", "Est. Cost", "Notes", "Diet Needs"],
        rows, [13, 12, 26, 15, 12, 20, 22],
        text_left={3, 6, 7}, dates={1}, money={5},
        validations=[("B", "ExpCatList")], reserved=30)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="beginsWith", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 12 — Activity Planner
# ===========================================================================
def build_activities(wb):
    rows = [
        ("Magic Kingdom day", "Theme Park", dplus(46), "Magic Kingdom", "$0", "Included", "All ages", "Everyone", "Rope drop 8am"),
        ("EPCOT day", "Theme Park", dplus(47), "EPCOT", "$0", "Included", "All ages", "Everyone", "Nap break midday"),
        ("Bibbidi Bobbidi Boutique", "Show", dplus(46), "Magic Kingdom", "$200", "Yes", "3–12", "Sofia, Mateo", "Princess makeover"),
        ("Pirates League", "Show", dplus(46), "Magic Kingdom", "$100", "Yes", "3+", "Diego", "Pirate transformation"),
        ("Hollywood Studios day", "Theme Park", dplus(49), "Hollywood Studios", "$0", "Included", "All ages", "Everyone", "Rise virtual queue"),
        ("Animal Kingdom day", "Theme Park", dplus(50), "Animal Kingdom", "$0", "Included", "All ages", "Everyone", "Safari before noon"),
        ("Mini-golf", "Outdoor", dplus(48), "Fantasia Gardens", "$70", "No", "All ages", "Everyone", "Rest-day fun"),
        ("Resort pool + slide", "Relax", dplus(48), "Resort", "$0", "No", "All ages", "Everyone", "Afternoon cooldown"),
        ("Disney Springs shopping", "Shopping", dplus(45), "Disney Springs", "$150", "No", "All ages", "Everyone", "Souvenir budget cap"),
        ("Fireworks dessert party", "Show", dplus(46), "Magic Kingdom", "$260", "Yes", "All ages", "Everyone", "Reserved viewing"),
        ("Safari (Kilimanjaro)", "Tour", dplus(50), "Animal Kingdom", "$0", "Included", "All ages", "Everyone", "Best light early"),
        ("Character breakfast", "Dining", dplus(50), "Tusker House", "$0", "Yes", "All ages", "Everyone", "Counts to food budget"),
        ("Movie night in", "Relax", dplus(48), "Resort room", "$0", "No", "All ages", "Everyone", "Pajamas + popcorn"),
        ("Splash pad", "Outdoor", dplus(47), "Resort", "$0", "No", "0–6", "Mateo", "Morning before EPCOT"),
    ]
    ws, start, end = build_log(
        wb, "Activities", "🎡", "ACTIVITY PLANNER",
        "The fun list — dates, cost, reservations, age ranges & who's joining.",
        ["Activity", "Type", "Date", "Location", "Cost", "Reserved?", "Age Range", "Who's Joining", "Notes"],
        rows, [24, 14, 13, 18, 11, 12, 12, 16, 22],
        text_left={1, 4, 8, 9}, dates={3}, money={5},
        validations=[("B", "ActTypeList"), ("F", "YesNoList")], reserved=40)
    nrange(wb, "ActName", "Activities", "A", start, end)
    nrange(wb, "ActCost", "Activities", "E", start, end)


# ===========================================================================
# 13 — Reservation Organizer
# ===========================================================================
def build_reservations(wb):
    rows = [
        ("Flight", "DL-8KX2P9", "delta.com", dplus(45), "Paid", "Yes", dplus(20), "Round trip · 5 seats"),
        ("Hotel", "WDW-CB-55219", "disneyworld.com", dplus(45), "Deposit", "Yes", dplus(40), "Balance due at check-in"),
        ("Rental Car", "ENT-77341", "enterprise.com", dplus(45), "Pending", "Yes", dplus(43), "Pay at pickup"),
        ("Park Ticket", "WDW-TIX-7742", "disneyworld.com", dplus(46), "Paid", "Yes", "—", "5-day park hopper"),
        ("Restaurant", "CRT-6PM", "disneyworld.com", dplus(46), "Paid", "Yes", dplus(44), "Cinderella's Royal Table"),
        ("Restaurant", "S220-1230", "disneyworld.com", dplus(47), "Paid", "Yes", dplus(45), "Space 220"),
        ("Activity", "BBB-0946", "disneyworld.com", dplus(46), "Paid", "Yes", dplus(44), "Bibbidi Bobbidi"),
        ("Show", "FDP-9PM", "disneyworld.com", dplus(46), "Paid", "Yes", dplus(44), "Fireworks dessert party"),
        ("Restaurant", "SCI-1230", "disneyworld.com", dplus(49), "Deposit", "Yes", dplus(46), "Sci-Fi Dine-In"),
        ("Restaurant", "TUS-8AM", "disneyworld.com", dplus(50), "Paid", "Yes", dplus(47), "Tusker House"),
        ("Transfer", "RES-INCL", "disneyworld.com", dplus(45), "Paid", "Yes", "—", "Resort shuttle"),
        ("Parking", "—", "resort", dplus(45), "Pending", "No", "—", "Confirm self-park rate"),
    ]
    ws, start, end = build_log(
        wb, "Reservations", "📑", "RESERVATION ORGANIZER",
        "Nothing double-booked — confirmations, payment status & cancellation deadlines.",
        ["Type", "Confirmation", "Website", "Date", "Payment", "Confirmed?", "Cancel By", "Notes"],
        rows, [14, 16, 18, 13, 12, 12, 13, 24],
        text_left={2, 3, 8}, dates={4}, money=set(),
        validations=[("A", "ResTypeList"), ("E", "PayStatusList"), ("F", "YesNoList")], reserved=30)
    nrange(wb, "ResConfirm", "Reservations", "F", start, end)
    for st, cc in {"Paid": MINT_BG, "Deposit": WARN_BG, "Pending": RED_BG}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 14 — Travel Document Center
# ===========================================================================
def build_documents(wb):
    rows = [
        ("Passports (x5)", "Passport", "Marcus", "2029–2031", "Home safe → travel wallet", "Ready", "All valid 6+ months"),
        ("Photo IDs", "ID", "Adults", "—", "Wallets", "Ready", "Driver's licenses"),
        ("Park tickets", "Tickets", "Elena", dplus(46), "Disney app (mobile)", "Ready", "Linked to MagicBands"),
        ("Hotel confirmation", "Hotel Confirmation", "Elena", "—", "Email + printout", "Ready", "Caribbean Beach"),
        ("Travel insurance", "Insurance", "Marcus", dplus(60), "Email + wallet card", "In Progress", "Family plan — finalizing"),
        ("Medical info sheet", "Medical", "Elena", "—", "Printed + phone photo", "Ready", "Allergies & meds listed"),
    ]
    ws, start, end = build_log(
        wb, "Documents", "🛂", "TRAVEL DOCUMENT CENTER",
        "Papers in order — every document, expiry, where it's stored & whether it's ready.",
        ["Document", "Type", "Owner", "Expiry / Date", "Stored Location", "Status", "Notes"],
        rows, [22, 18, 14, 16, 26, 14, 24],
        text_left={1, 5, 7}, reserved=16,
        validations=[("B", "DocTypeList"), ("F", "DocStatusList")])
    nrange(wb, "DocStatus", "Documents", "F", start, end)
    for st, cc in {"Ready": MINT_BG, "In Progress": WARN_BG, "Missing": RED_BG}.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 15 — Emergency Command Center
# ===========================================================================
def build_emergency(wb):
    ws = wb.create_sheet("Emergency"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 26, 4, 24, 28, 2])
    luxe_header(ws, "G", "🚨  EMERGENCY COMMAND CENTER",
                "Peace of mind — every contact and health note in one fast-access place.")
    merge_set(ws, "B5:C5", "EMERGENCY CONTACTS", "section_gold"); ws.row_dimensions[5].height = 22
    contacts = [("Local Emergency", "911"), ("Grandma (Rosa)", "555-0110"),
                ("Family Doctor", "555-0188"), ("Pediatrician", "555-0177"),
                ("Insurance 24/7 line", "800-555-0100"), ("Resort front desk", "407-934-3400"),
                ("Airline (Delta)", "800-221-1212"), ("Roadside assist", "800-555-0199")]
    for i, (name, num) in enumerate(contacts):
        r = 6 + i
        ws.cell(row=r, column=2, value=name).style = "field_label"
        ws.cell(row=r, column=3, value=num).style = "field_value"
        ws.row_dimensions[r].height = 24
    merge_set(ws, "E5:F5", "HEALTH INFORMATION", "section_gold"); ws.row_dimensions[5].height = 22
    health = [("Sofia", "Peanut allergy · carries inhaler"), ("Mateo", "Dairy allergy · EpiPen in mom's bag"),
              ("Elena", "Shellfish allergy"), ("Blood types", "On medical info sheet"),
              ("Medications", "Allergy meds, children's Tylenol"), ("Pharmacy", "Any CVS — insurance on file"),
              ("Nearest ER", "AdventHealth Orlando"), ("Notes", "Photo of insurance cards on all phones")]
    for i, (name, info) in enumerate(health):
        r = 6 + i
        ws.cell(row=r, column=5, value=name).style = "field_label"
        ws.cell(row=r, column=6, value=info).style = "td_left"
        ws.row_dimensions[r].height = 24


# ===========================================================================
# 16 — Road Trip Command Center
# ===========================================================================
def build_roadtrip(wb):
    rows = [
        (1, "Home → Savannah", "Savannah, GA", 250, 4.0, "$45", "Lunch stop · historic district", "$0"),
        (2, "Savannah → Jacksonville", "Jacksonville, FL", 140, 2.5, "$25", "Beach walk & stretch", "$0"),
        (3, "Jacksonville → Orlando", "Orlando, FL", 140, 2.5, "$25", "Arrive resort by 3pm", "$0"),
        (4, "Return leg 1", "Savannah, GA", 280, 4.5, "$50", "Overnight — same hotel", "$120"),
        (5, "Return leg 2 → Home", "Home", 250, 4.0, "$45", "Home by dinner", "$0"),
    ]
    ws, start, end = build_log(
        wb, "Road Trip", "🚗", "ROAD TRIP COMMAND CENTER",
        "For the drive option — route, stops, fuel, hours & mileage (ignore if flying).",
        ["Leg", "Route", "Overnight / Stop", "Miles", "Drive Hrs", "Fuel", "Attractions / Notes", "Lodging"],
        rows, [8, 24, 20, 10, 11, 12, 26, 12],
        text_left={2, 3, 7}, ints={4}, dec={5}, money={6, 8}, reserved=16)
    totals(ws, end + 1, [4], start, end, fmt="#,##0", label="TOTAL MILES")


# ===========================================================================
# 17 — Responsibility Board
# ===========================================================================
def build_responsibilities(wb):
    rows = [
        ("Packing Leader", "Elena", "Final bag check the night before", "Yes"),
        ("Navigator", "Marcus", "Directions, park maps & ride order", "Yes"),
        ("Snack Manager", "Diego", "Keep the day-bag stocked", "Yes"),
        ("Entertainment Manager", "Sofia", "Games & tablets for travel days", "Yes"),
        ("Document Manager", "Marcus", "Passports, tickets & confirmations", "Yes"),
        ("Photographer", "Sofia", "One family photo every day", "No"),
        ("Budget Keeper", "Elena", "Log spending each night", "No"),
    ]
    ws, start, end = build_log(
        wb, "Responsibilities", "📋", "FAMILY TRAVEL RESPONSIBILITY BOARD",
        "Everyone has a job — share the load so no one carries the whole trip.",
        ["Role", "Assigned To", "What It Covers", "On Track?"],
        rows, [22, 16, 34, 12],
        text_left={3}, reserved=12,
        validations=[("B", "MemberList"), ("D", "YesNoList")])
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 18 — Points & Rewards
# ===========================================================================
def build_points(wb):
    rows = [
        ("Delta SkyMiles", "Airline", 62000, dplus(400), "Free flight next trip", "Pooled family account"),
        ("Marriott Bonvoy", "Hotel", 48000, dplus(600), "2 free nights", "From work travel"),
        ("Chase Sapphire", "Credit Card", 85000, "None", "Transfer to Delta", "1.25x on travel portal"),
        ("Disney Rewards", "Credit Card", 320, "None", "Gift cards for souvenirs", "Redeemed in-app"),
        ("Hertz Gold", "Rental", 4, dplus(365), "1 free rental day", "Almost at reward"),
        ("Amex Membership", "Credit Card", 54000, "None", "Statement credit", "Hold for emergencies"),
    ]
    ws, start, end = build_log(
        wb, "Points", "🎁", "TRAVEL POINTS & REWARDS TRACKER",
        "Travel for less — every point, expiration & redemption goal in one ledger.",
        ["Program", "Type", "Balance", "Expires", "Redemption Goal", "Notes"],
        rows, [22, 14, 14, 14, 24, 24],
        text_left={5, 6}, ints={3}, reserved=20)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=90000, color=GOLD_LT, showValue=True))


# ===========================================================================
# 19 — Photo & Memory Vault
# ===========================================================================
def build_photovault(wb):
    ws = wb.create_sheet("Photo Vault"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  PHOTO & MEMORY VAULT",
                "Keep the magic — drop in photos and caption the moments you never want to forget.")
    merge_set(ws, "B5:D5", "HOW TO ADD YOUR PHOTOS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Excel: Insert ▸ Pictures ▸ Place in Cell (or drag a photo) into any framed box below. "
                      "Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"paste-link-here\"). "
                      "Caption each memory underneath — best moments, food, adventures, funny moments & kids' highlights.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Best Moment", "Family Photo", "Food Memory", "Adventure", "Funny Moment", "Kids' Highlight"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add photo)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Date · Location · caption…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 20 — Travel Journal
# ===========================================================================
def build_journal(wb):
    rows = [
        (dplus(45), "Made it! Kids gasped at the castle", "Airport pretzels", "Mateo fell asleep in stroller mid-parade", "First monorail ride", "Everyone together, finally on vacation"),
        (dplus(46), "Sofia met Cinderella", "Character dinner mac & cheese", "Diego's pirate voice all day", "Fireworks up close", "Mateo's face at the castle lighting"),
        (dplus(47), "Rode Guardians 3x", "Space 220 'floating' lunch", "Dad stuck on 'it's a small world'", "Tried sushi in Japan pavilion", "World Showcase snacking tour"),
        (dplus(48), "Slept till 9 — glorious", "Homemade tacos in room", "Pool cannonball contest", "Won mini-golf as a team", "A calm day to just be together"),
        (dplus(49), "Rise of the Resistance!", "Sci-Fi Dine-In burgers", "Diego 'directed' the whole day", "Single-rider strategy worked", "Kids brave on the big rides"),
    ]
    ws, start, end = build_log(
        wb, "Journal", "📔", "TRAVEL JOURNAL",
        "One line a day keeps the memories forever — capture it before you sleep.",
        ["Date", "Best Moment", "Favorite Meal", "Funniest Memory", "New Experience", "Family Highlight"],
        rows, [13, 26, 22, 26, 22, 26],
        text_left={2, 3, 4, 5, 6}, dates={1}, reserved=21)


# ===========================================================================
# 21 — Post-Trip Review
# ===========================================================================
def build_review(wb):
    ws = wb.create_sheet("Post-Trip"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 34, 4, 24, 16, 2])
    luxe_header(ws, "G", "⭐  POST-TRIP REVIEW",
                "Learn & improve — rate the trip, keep what worked and plan next time better.")
    merge_set(ws, "B5:C5", "THE VERDICT", "section_gold"); ws.row_dimensions[5].height = 22
    verdict = [("Overall Rating", "9 / 10"), ("Best Experience", "Fireworks dessert party"),
               ("Best Value", "Rest-day pool + grocery meals"), ("Biggest Splurge", "Character dining"),
               ("Would Return?", "Yes — in 2 years"), ("Kids' Favorite", "Bibbidi Bobbidi + safari"),
               ("Trip Pace", "Just right — rest day saved us"), ("Under / Over Budget", "$180 under")]
    for i, (lab, val) in enumerate(verdict):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        ws.cell(row=r, column=3, value=val).style = "field_value"
        ws.row_dimensions[r].height = 24
    merge_set(ws, "E5:F5", "LESSONS FOR NEXT TIME", "section_gold"); ws.row_dimensions[5].height = 22
    lessons = [("What worked", "Rope drop"), ("What worked", "Mid-trip rest day"), ("What worked", "Mobile order meals"),
               ("To improve", "Pack more ponchos"), ("To improve", "Book Space 220 earlier"),
               ("To improve", "Budget more for snacks"), ("Try next", "Add a beach day"), ("Try next", "Skip 5th park")]
    for i, (lab, val) in enumerate(lessons):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        ws.cell(row=r, column=6, value=val).style = "td_left"
        ws.row_dimensions[r].height = 24


# ===========================================================================
# 22 — Future Travel Wishlist
# ===========================================================================
def build_wishlist(wb):
    rows = [
        ("Hawaii (Maui)", "Beach", "$11,000", "Spring", "High", "$3,200", "Multi-gen with grandparents"),
        ("Yellowstone road trip", "Road Trip", "$5,500", "Summer", "High", "$1,000", "Wildlife + camping mix"),
        ("Disney Cruise (Bahamas)", "Cruise", "$7,800", "Winter", "Medium", "$500", "Kids' clubs sell it"),
        ("London & Paris", "International", "$14,000", "Summer", "Medium", "$0", "When Mateo is older"),
        ("San Diego (zoo + beach)", "City Break", "$6,200", "Fall", "High", "$800", "Easy flights, great weather"),
        ("Grand Canyon + Sedona", "Road Trip", "$4,800", "Spring", "Low", "$0", "Bucket-list views"),
        ("Costa Rica", "International", "$9,500", "Winter", "Low", "$0", "Adventure + wildlife"),
    ]
    ws, start, end = build_log(
        wb, "Wishlist", "🌟", "FUTURE TRAVEL WISHLIST",
        "Dream a little — rank the next adventures and start a fund for the front-runners.",
        ["Destination", "Type", "Est. Cost", "Best Season", "Priority", "Saved So Far", "Why We Want It"],
        rows, [24, 14, 13, 13, 11, 13, 28],
        text_left={1, 7}, money={6}, reserved=20,
        validations=[("B", "TripTypeList"), ("E", "PriorityList")])
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"High"'], fill=fill(MINT_BG)))


# ===========================================================================
# 23 — Travel Analytics
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 3, 22, 12, 12, 2])
    luxe_header(ws, "G", "📊  TRAVEL ANALYTICS DASHBOARD",
                "The bigger picture — this trip by the numbers, plus a Family Trip Readiness Score.")
    # readiness dims
    merge_set(ws, "B5:C5", "TRIP READINESS", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score"], start_col=2)
    dims = [
        ("Savings progress", "=IFERROR(MIN(SavedTotal/SavingsGoal,1),0)"),
        ("Bookings confirmed", '=IFERROR(COUNTIF(ResConfirm,"Yes")/COUNTA(ResConfirm),0)'),
        ("Packing complete", '=IFERROR(COUNTIF(PackStatus,"Yes")/COUNTA(PackStatus),0)'),
        ("Documents ready", '=IFERROR(COUNTIF(DocStatus,"Ready")/COUNTA(DocStatus),0)'),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "ReadinessRange", "Analytics", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # trip snapshot
    merge_set(ws, "E5:F5", "TRIP SNAPSHOT", "section")
    snap = [("Total budget", "=TripBudget", '"$"#,##0'), ("Money saved", "=SavedTotal", '"$"#,##0'),
            ("Spent so far", "=SpentTotal", '"$"#,##0'), ("Cost / person", "=IFERROR(TripBudget/TravelerCount,0)", '"$"#,##0'),
            ("Cost / day", "=IFERROR(TripBudget/(TripEnd-TripStart+1),0)", '"$"#,##0'), ("Activities planned", "=COUNTA(ActName)", "#,##0"),
            ("Readiness score", "=IFERROR(AVERAGE(ReadinessRange),0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Readiness score":
            c.fill = fill(MINT_BG)
    # savings trend for chart
    merge_set(ws, "B14:C14", "SAVINGS BY MONTH", "section"); ws.row_dimensions[14].height = 20
    ws.cell(row=15, column=2, value="Month").style = "th"; ws.cell(row=15, column=3, value="Saved").style = "th"
    months = ["Feb", "Mar", "Apr", "May", "Jun", "Jul"]; vals = [1000, 2200, 3300, 4500, 5800, 6800]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 16 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "SaveMonth", "Analytics", "$B$16:$B$21")
    cell_name(wb, "SaveVal", "Analytics", "$C$16:$C$21")


# ===========================================================================
# 1 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🌎  FAMILY VACATION COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Budget, packing, bookings & memories — your whole family trip, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("TRIP COUNTDOWN", "=MAX(TripStart-TODAY(),0)", "num"),
        ("TOTAL BUDGET", "=TripBudget", "money"),
        ("MONEY SAVED", "=SavedTotal", "money"),
        ("BUDGET REMAINING", "=TripBudget-SpentTotal", "money"),
        ("TRAVELERS", "=TravelerCount", "num"),
        ("DAYS TRAVELING", "=TripEnd-TripStart+1", "num"),
    ]
    row2 = [
        ("PACKING COMPLETE", '=IFERROR(COUNTIF(PackStatus,"Yes")/COUNTA(PackStatus),0)', "pct"),
        ("BOOKINGS DONE", '=IFERROR(COUNTIF(ResConfirm,"Yes")/COUNTA(ResConfirm),0)', "pct"),
        ("DOCUMENTS READY", '=IFERROR(COUNTIF(DocStatus,"Ready")/COUNTA(DocStatus),0)', "pct"),
        ("ACTIVITIES PLANNED", "=COUNTA(ActName)", "num"),
        ("COST / PERSON", "=IFERROR(TripBudget/TravelerCount,0)", "money"),
        ("READINESS SCORE", "=IFERROR(AVERAGE(ReadinessRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "BUDGET & SAVINGS", "section_gold")
    # budget donut
    d1 = DoughnutChart(); d1.title = "Budget by Category"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Budget"], min_col=2, min_row=4, max_row=14), titles_from_data=True)
    d1.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=14)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")
    # savings growth line
    ln = LineChart(); ln.title = "Vacation Fund Growth"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=15, max_row=21), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=16, max_row=21)); ln.legend = None
    ws.add_chart(ln, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "READINESS & SPENDING", "section_gold")
    # readiness bar
    cb = BarChart(); cb.type = "bar"; cb.title = "Trip Readiness"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=6, max_row=10), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=7, max_row=10)); cb.legend = None
    ws.add_chart(cb, "B30")
    # planned vs actual
    bc = BarChart(); bc.type = "col"; bc.title = "Planned vs Actual"; bc.height = 8.2; bc.width = 11.5
    bc.add_data(Reference(wb["Budget"], min_col=2, min_row=4, max_col=3, max_row=14), titles_from_data=True)
    bc.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=14))
    ws.add_chart(bc, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Family Vacation Command Center™ — from first idea to final photo, all in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_master(wb)
    build_budget(wb); build_savings(wb); build_itinerary(wb); build_transport(wb)
    build_hotels(wb); build_packing(wb); build_kids(wb); build_meals(wb)
    build_activities(wb); build_reservations(wb); build_documents(wb); build_emergency(wb)
    build_roadtrip(wb); build_responsibilities(wb); build_points(wb); build_photovault(wb)
    build_journal(wb); build_review(wb); build_wishlist(wb); build_analytics(wb)
    build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Family Profile", "Master Trip", "Budget", "Savings", "Itinerary",
             "Transport", "Hotels", "Packing", "Kids", "Meals", "Activities", "Reservations", "Documents",
             "Emergency", "Road Trip", "Responsibilities", "Points", "Photo Vault", "Journal", "Post-Trip",
             "Wishlist", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Family_Vacation_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
