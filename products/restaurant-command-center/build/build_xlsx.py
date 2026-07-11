"""Build Restaurant Command Center™ — The Ultimate Restaurant Operations, Cost
& Team Management System.

22 sheets + Welcome · a premium restaurant operating system in Excel & Sheets.
Dashboard, profile, menu & recipe costing, inventory, par levels, suppliers,
sales, labor, staff, P&L / prime cost, expenses, cash & tips, reservations,
marketing, reviews, compliance, checklists, waste, vendor payments, training,
analytics & settings — one elegant command center.

Run: python3 build_xlsx.py   ->  ../Restaurant_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

MENU_CATS = ["Starters", "Mains", "Sides", "Desserts", "Drinks", "Cocktails"]
EXP_CATS = ["Rent", "Utilities", "Insurance", "Marketing", "Repairs & Maint.", "POS / Software",
            "Cleaning & Supplies", "Licenses & Permits", "Accounting", "Card Fees", "Miscellaneous"]
INV_CATS = ["Produce", "Meat & Seafood", "Dairy", "Dry Goods", "Frozen", "Beverage", "Bar / Liquor", "Paper & Disposables"]
UNITS = ["lb", "kg", "each", "case", "bottle", "gal", "L", "dozen"]
STATIONS = ["Kitchen", "Front of House", "Bar", "Management", "Support"]
SHIFTS = ["Open", "Mid", "Close", "Double", "Off"]
PAY_STATUS = ["Paid", "Due", "Overdue", "Scheduled"]
RES_STATUS = ["Confirmed", "Pending", "Seated", "Cancelled", "No-Show"]
PROMO_STATUS = ["Live", "Planned", "Ended", "Paused"]
COMPLY_STATUS = ["Pass", "Action Needed", "Overdue", "N/A"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]
DAYPARTS = ["Breakfast", "Lunch", "Dinner", "Late Night", "Brunch"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# --- sample sales data (30 days), shared by Sales + Analytics ---------------
def _gen_sales():
    cov = {0: 95, 1: 105, 2: 120, 3: 135, 4: 180, 5: 220, 6: 150}
    chk = {0: 28, 1: 28, 2: 29, 3: 30, 4: 31, 5: 33, 6: 30}
    start = dt.date.today() - dt.timedelta(days=30)
    rows = []
    for i in range(30):
        d = start + dt.timedelta(days=i)
        wd = d.weekday()
        rows.append((d, cov[wd], cov[wd] * chk[wd]))
    return rows


SALES = _gen_sales()
REV_TOTAL = sum(s for _, _, s in SALES)
COV_TOTAL = sum(c for _, c, _ in SALES)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def totals(ws, row, cols, start, end, fmt='"$"#,##0', label="TOTAL"):
    ws.cell(row=row, column=1, value=label).style = "th"
    for col in cols:
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt


# ===========================================================================
# 22 — Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 22, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your restaurant details & targets once — every dashboard follows.")
    merge_set(ws, "B5:C5", "RESTAURANT INPUTS", "section")
    controls = [
        ("Restaurant Name", "Olive & Ember", None, "RestName"),
        ("Concept", "Modern American Bistro", None, "Concept"),
        ("Seats", 84, "0", "Seats"),
        ("Food Cost Target", 0.30, "0%", "FoodTarget"),
        ("Labor Cost Target", 0.32, "0%", "LaborTarget"),
        ("Prime Cost Target", 0.60, "0%", "PrimeTarget"),
        ("Net Margin Target", 0.15, "0%", "MarginTarget"),
        ("Avg Check Goal", 32, '"$"#,##0', "CheckGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Menu Category", MENU_CATS, "MenuCatList"), ("F", "Expense Category", EXP_CATS, "ExpCatList"),
             ("G", "Inventory Category", INV_CATS, "InvCatList"), ("H", "Unit", UNITS, "UnitList"),
             ("I", "Station", STATIONS, "StationList"), ("J", "Shift", SHIFTS, "ShiftList"),
             ("K", "Daypart", DAYPARTS, "DaypartList"), ("L", "Reservation", RES_STATUS, "ResStatusList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 18, "Pay Status", PAY_STATUS, "PayStatusList"), ("F", 18, "Promo", PROMO_STATUS, "PromoStatusList"),
             ("G", 18, "Compliance", COMPLY_STATUS, "ComplyStatusList"), ("H", 18, "Priority", PRIORITIES, "PriorityList"),
             ("I", 18, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🍽  RESTAURANT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate restaurant operations, cost & team management system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "RUN YOUR ENTIRE RESTAURANT FROM ONE FILE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("From menu costing to the bottom line — Restaurant Command Center™ manages sales, food & labor "
                      "cost, inventory, suppliers, scheduling, P&L, reservations, marketing, reviews, compliance and "
                      "your team in ONE premium Excel & Google Sheets system. Protect your margins, control prime cost, "
                      "cut waste and run a tighter, more profitable operation — all with restaurant-grade automation. "
                      "This isn't a spreadsheet — it's your complete Restaurant Operating System.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your restaurant, concept, seats & cost targets.",
             "2.  Build the Menu & Recipe Costing — food cost % & margin per item, live.",
             "3.  Load Inventory & Par Levels — value and low-stock flags calculate themselves.",
             "4.  Log daily Sales; set Labor & the P&L — prime cost & net profit update automatically.",
             "5.  Track Suppliers, Reservations, Marketing, Reviews & Compliance.",
             "6.  Watch the Dashboard track revenue, prime cost & an Operations Health Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (Olive & Ember, an 84-seat bistro doing ~$128k/month) is included so you can see "
               "how everything connects — just type over it with your own. Revenue, food & labor cost %, prime cost, "
               "net profit, inventory value, low-stock alerts, guest ratings and the Operations Health Score all update "
               "automatically. Every sheet is print-friendly and works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, healthier margins — let's run a tighter restaurant.", "section_gold")


# ===========================================================================
# 2 — Restaurant Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 4, 24, 30, 2])
    luxe_header(ws, "G", "🏛  RESTAURANT PROFILE", "Your operation, defined — the concept every number flows from.")
    blocks = [
        ("THE RESTAURANT", [("Name", "=RestName"), ("Concept", "=Concept"),
                            ("Seats", "=Seats"), ("Avg Check Goal", "=CheckGoal"),
                            ("Opened", "2019"), ("Service", "Lunch & Dinner · 7 days")]),
        ("THE TARGETS", [("Food Cost Target", "=FoodTarget"), ("Labor Cost Target", "=LaborTarget"),
                         ("Prime Cost Target", "=PrimeTarget"), ("Net Margin Target", "=MarginTarget"),
                         ("Manager", "Dana Reyes (GM)"), ("Chef", "Marco Ellis (Exec)")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "KEY CONTACTS", "section_gold"); ws.row_dimensions[15].height = 22
    contacts = [("Landlord", "555-0120"), ("POS Support", "800-555-0140"),
                ("Health Dept.", "555-0166"), ("Primary Supplier", "555-0188"),
                ("Bookkeeper", "555-0155"), ("Equipment Repair", "555-0177")]
    for i, (p, h) in enumerate(contacts):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 3 — Menu & Recipe Costing
# ===========================================================================
def build_menu(wb):
    rows = [
        ("Charred Octopus", "Starters", 16, 4.60, "Star"),
        ("Burrata & Heirloom", "Starters", 14, 4.20, "Star"),
        ("Steak Tartare", "Starters", 17, 5.80, "Plow"),
        ("Crispy Brussels", "Starters", 11, 2.30, "Star"),
        ("French Onion Soup", "Starters", 12, 2.80, "Star"),
        ("Dry-Aged Ribeye", "Mains", 46, 18.40, "Plow"),
        ("Pan-Roasted Chicken", "Mains", 28, 7.60, "Star"),
        ("Seared Salmon", "Mains", 32, 11.20, "Star"),
        ("Braised Short Rib", "Mains", 34, 12.60, "Puzzle"),
        ("Wild Mushroom Risotto", "Mains", 26, 6.40, "Star"),
        ("House Burger", "Mains", 19, 5.70, "Star"),
        ("Cacio e Pepe", "Mains", 22, 4.80, "Star"),
        ("Roasted Cauliflower", "Sides", 10, 2.10, "Star"),
        ("Truffle Fries", "Sides", 12, 3.10, "Star"),
        ("Garlic Green Beans", "Sides", 9, 1.90, "Puzzle"),
        ("Mac & Cheese", "Sides", 11, 2.80, "Star"),
        ("Warm Chocolate Cake", "Desserts", 12, 3.00, "Star"),
        ("Crème Brûlée", "Desserts", 11, 2.40, "Star"),
        ("Seasonal Sorbet", "Desserts", 9, 1.80, "Dog"),
        ("Cheesecake", "Desserts", 11, 3.20, "Puzzle"),
        ("House Red (glass)", "Drinks", 13, 3.40, "Star"),
        ("Craft Lager", "Drinks", 8, 1.80, "Star"),
        ("Sparkling Water", "Drinks", 5, 0.60, "Star"),
        ("Espresso", "Drinks", 4, 0.55, "Star"),
        ("Fresh Lemonade", "Drinks", 6, 0.90, "Star"),
        ("Old Fashioned", "Cocktails", 15, 3.10, "Star"),
        ("Ember Negroni", "Cocktails", 15, 3.40, "Star"),
        ("Spicy Margarita", "Cocktails", 14, 3.00, "Star"),
        ("Garden Gimlet", "Cocktails", 14, 2.80, "Puzzle"),
        ("Smoked Manhattan", "Cocktails", 16, 3.60, "Star"),
    ]
    ws = wb.create_sheet("Menu"); ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 14, 12, 12, 12, 12, 12])
    luxe_header(ws, "G", "📖  MENU & RECIPE COSTING",
                "Menu engineering — food cost %, contribution margin & item class, per dish.")
    table_headers(ws, 4, ["Item", "Category", "Menu Price", "Food Cost", "Food Cost %", "Margin", "Class"])
    start = L0
    for i, (item, cat, price, cost, cls) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=item).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        fp = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); fp.style = "td"; fp.number_format = "0%"
        mg = ws.cell(row=r, column=6, value=f"=C{r}-D{r}"); mg.style = "td"; mg.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7, value=cls).style = "td"
        if i % 2:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    # reserve a few blank rows
    style_rows(ws, end + 1, end + 6, 7, text_left={1})
    for r in range(end + 1, end + 7):
        ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)").number_format = "0%"
        ws.cell(row=r, column=6, value=f"=C{r}-D{r}").number_format = '"$"#,##0.00'
    add_dv(ws, f"B{start}:B{end+6}", "MenuCatList")
    ws.freeze_panes = "A5"
    nrange(wb, "MenuItem", "Menu", "A", start, end)
    nrange(wb, "MenuCat", "Menu", "B", start, end)
    nrange(wb, "MenuFCPct", "Menu", "E", start, end)
    cell_name(wb, "MenuCount", "Menu", f"$A${start}:$A${end+6}")
    ws.conditional_formatting.add(f"E{start}:E{end}", ColorScaleRule(
        start_type="num", start_value=0.2, start_color="FF" + HIGHLIGHT, mid_type="num", mid_value=0.32,
        mid_color="FFFFF3CD", end_type="num", end_value=0.45, end_color="FF" + RED_BG))


# ===========================================================================
# 4 — Inventory Master
# ===========================================================================
def build_inventory(wb):
    rows = [
        ("Heirloom Tomatoes", "Produce", "lb", 22, 15, 2.40),
        ("Mixed Greens", "Produce", "case", 4, 5, 28.00),
        ("Yukon Potatoes", "Produce", "lb", 60, 40, 0.90),
        ("Yellow Onions", "Produce", "lb", 45, 30, 0.70),
        ("Ribeye (dry-aged)", "Meat & Seafood", "lb", 18, 20, 16.50),
        ("Chicken Airline", "Meat & Seafood", "each", 40, 30, 3.80),
        ("Salmon Fillet", "Meat & Seafood", "lb", 12, 15, 11.00),
        ("Short Rib", "Meat & Seafood", "lb", 16, 12, 9.50),
        ("Octopus", "Meat & Seafood", "lb", 8, 6, 12.00),
        ("Burrata", "Dairy", "each", 24, 18, 3.20),
        ("Heavy Cream", "Dairy", "qt", 14, 12, 3.60),
        ("Butter", "Dairy", "lb", 30, 20, 3.90),
        ("Parmesan", "Dairy", "lb", 9, 8, 12.50),
        ("Arborio Rice", "Dry Goods", "lb", 90, 50, 2.10),
        ("00 Flour", "Dry Goods", "lb", 120, 70, 1.40),
        ("Olive Oil (EVOO)", "Dry Goods", "L", 10, 12, 18.00),
        ("Sea Salt", "Dry Goods", "lb", 40, 20, 1.80),
        ("Frozen Fries", "Frozen", "case", 6, 8, 22.00),
        ("Vanilla Ice Cream", "Frozen", "gal", 5, 6, 14.00),
        ("House Red Wine", "Bar / Liquor", "bottle", 96, 60, 14.00),
        ("Sparkling Wine", "Bar / Liquor", "bottle", 48, 30, 16.00),
        ("Craft Lager", "Beverage", "case", 40, 24, 26.00),
        ("Bourbon", "Bar / Liquor", "bottle", 24, 18, 32.00),
        ("Gin", "Bar / Liquor", "bottle", 20, 15, 28.00),
        ("Vodka", "Bar / Liquor", "bottle", 22, 15, 26.00),
        ("Tequila", "Bar / Liquor", "bottle", 14, 12, 34.00),
        ("To-Go Containers", "Paper & Disposables", "case", 5, 6, 34.00),
    ]
    ws = wb.create_sheet("Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 18, 8, 11, 10, 12, 12, 13])
    luxe_header(ws, "H", "📦  INVENTORY MASTER",
                "Every item valued — on-hand vs par, with automatic low-stock flags.")
    table_headers(ws, 4, ["Item", "Category", "Unit", "On Hand", "Par", "Unit Cost", "Value", "Status"])
    start = L0
    for i, (item, cat, unit, onhand, par, cost) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=item).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ws.cell(row=r, column=3, value=unit).style = "td"
        oh = ws.cell(row=r, column=4, value=onhand); oh.style = "input"; oh.number_format = "#,##0"
        pa = ws.cell(row=r, column=5, value=par); pa.style = "input"; pa.number_format = "#,##0"
        uc = ws.cell(row=r, column=6, value=cost); uc.style = "input"; uc.number_format = '"$"#,##0.00'
        vv = ws.cell(row=r, column=7, value=f"=D{r}*F{r}"); vv.style = "td"; vv.number_format = '"$"#,##0'
        st = ws.cell(row=r, column=8, value=f'=IF(D{r}<E{r},"Low","OK")'); st.style = "td"
        if i % 2:
            for c in range(1, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    tot = end + 1
    ws.cell(row=tot, column=1, value="INVENTORY VALUE").style = "th"
    for c in range(2, 7):
        ws.cell(row=tot, column=c).style = "th"
    cv = ws.cell(row=tot, column=7, value=f"=SUM(G{start}:G{end})")
    cv.style = "td"; cv.font = Font(bold=True, color=PRIMARY); cv.fill = fill(SURFACE); cv.number_format = '"$"#,##0'
    ws.cell(row=tot, column=8).style = "td"; ws.cell(row=tot, column=8).fill = fill(SURFACE)
    ws.freeze_panes = "A5"
    add_dv(ws, f"B{start}:B{end}", "InvCatList"); add_dv(ws, f"C{start}:C{end}", "UnitList")
    nrange(wb, "InvName", "Inventory", "A", start, end)
    nrange(wb, "InvStatus", "Inventory", "H", start, end)
    cell_name(wb, "InvValue", "Inventory", f"$G${tot}")
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"Low"'], fill=fill(RED_BG), font=Font(color=DANGER, bold=True)))
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=fill(MINT_BG)))


# ===========================================================================
# 5 — Par Levels & Order Guide
# ===========================================================================
def build_par(wb):
    rows = [
        ("Ribeye (dry-aged)", "Meat & Seafood", 18, 20, "=MAX(D5-C5,0)", "Prime Cuts Co.", "Tue / Fri"),
        ("Salmon Fillet", "Meat & Seafood", 12, 15, "=MAX(D6-C6,0)", "Harbor Fish", "Daily"),
        ("Mixed Greens", "Produce", 4, 5, "=MAX(D7-C7,0)", "Green Valley", "Mon/Wed/Fri"),
        ("House Red Wine", "Bar / Liquor", 36, 24, "=MAX(D8-C8,0)", "Vintners Direct", "Weekly"),
        ("Craft Lager", "Beverage", 10, 8, "=MAX(D9-C9,0)", "City Beverage", "Weekly"),
        ("Frozen Fries", "Frozen", 6, 8, "=MAX(D10-C10,0)", "US Foods", "Tue"),
        ("To-Go Containers", "Paper & Disposables", 5, 6, "=MAX(D11-C11,0)", "RestaurantDepot", "As needed"),
        ("00 Flour", "Dry Goods", 55, 30, "=MAX(D12-C12,0)", "Bulk Pantry", "Bi-weekly"),
    ]
    ws, start, end = build_log(
        wb, "Par Levels", "📋", "PAR LEVELS & ORDER GUIDE",
        "Never 86 a dish — on-hand vs par tells you exactly what to reorder.",
        ["Item", "Category", "On Hand", "Par", "Order Qty", "Supplier", "Order Days"],
        rows, [24, 18, 11, 10, 12, 18, 16],
        text_left={1, 6, 7}, ints={3, 4, 5}, reserved=24)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=fill(WARN_BG)))


# ===========================================================================
# 6 — Suppliers & Ordering
# ===========================================================================
def build_suppliers(wb):
    rows = [
        ("Prime Cuts Co.", "Meat & Seafood", "Sal M.", "555-0201", "Tue / Fri", "Net 14", 4.8, "Reliable, premium"),
        ("Harbor Fish", "Meat & Seafood", "Lena K.", "555-0202", "Daily", "Net 7", 4.6, "Freshest catch"),
        ("Green Valley", "Produce", "Omar R.", "555-0203", "Mon/Wed/Fri", "Net 14", 4.5, "Local & organic"),
        ("US Foods", "Broadline", "Rep line", "800-555-0204", "Tue", "Net 21", 4.2, "Everything else"),
        ("Vintners Direct", "Bar / Liquor", "Chris P.", "555-0205", "Weekly", "Net 30", 4.7, "Great wine list"),
        ("City Beverage", "Beverage", "Dana T.", "555-0206", "Weekly", "COD", 4.4, "Beer & NA"),
        ("Bulk Pantry", "Dry Goods", "Ivy S.", "555-0207", "Bi-weekly", "Net 14", 4.3, "Best dry-goods pricing"),
        ("RestaurantDepot", "Paper & Disposables", "Walk-in", "—", "As needed", "COD", 4.0, "Backup / emergencies"),
    ]
    ws, start, end = build_log(
        wb, "Suppliers", "🚚", "SUPPLIERS & ORDERING",
        "Your vendor book — contacts, order days, terms & a reliability rating.",
        ["Supplier", "Category", "Contact", "Phone", "Order Days", "Terms", "Rating", "Notes"],
        rows, [20, 18, 14, 14, 16, 12, 10, 24],
        text_left={1, 3, 8}, dec={7}, reserved=20)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=5, color=GOLD_LT, showValue=True))


# ===========================================================================
# 7 — Sales Tracker
# ===========================================================================
def build_sales(wb):
    ws = wb.create_sheet("Sales"); ws.sheet_view.showGridLines = False
    set_widths(ws, [14, 12, 12, 14, 12, 26])
    luxe_header(ws, "F", "💵  SALES TRACKER",
                "Every service, logged — covers, net sales & average check, totaled live.")
    table_headers(ws, 4, ["Date", "Day", "Covers", "Net Sales", "Avg Check", "Notes"])
    start = L0
    notes_map = {5: "Sat — busiest", 4: "Fri rush", 6: "Sun brunch"}
    for i, (d, cov, sales) in enumerate(SALES):
        r = start + i
        cd = ws.cell(row=r, column=1, value=d); cd.style = "td"; cd.number_format = "mm/dd/yyyy"
        ws.cell(row=r, column=2, value=d.strftime("%a")).style = "td"
        cc = ws.cell(row=r, column=3, value=cov); cc.style = "input"; cc.number_format = "#,##0"
        cs = ws.cell(row=r, column=4, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); ca.style = "td"; ca.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6, value=notes_map.get(d.weekday(), "")).style = "td_left"
        if i % 2:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SALES) - 1
    tot = end + 1
    ws.cell(row=tot, column=1, value="MONTH TOTAL").style = "th"
    ws.cell(row=tot, column=2).style = "th"
    cc = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = "#,##0"
    cs = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    ca = ws.cell(row=tot, column=5, value=f"=IFERROR(D{tot}/C{tot},0)"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = '"$"#,##0.00'
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)
    ws.freeze_panes = "A5"
    cell_name(wb, "RevenueTotal", "Sales", f"$D${tot}")
    cell_name(wb, "CoversTotal", "Sales", f"$C${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=8000, color=PRIMARY, showValue=True))


# ===========================================================================
# 8 — Labor & Scheduling
# ===========================================================================
def build_labor(wb):
    rows = [
        ("Executive Chef", "Kitchen", 1, 168, "Salary", 5800),
        ("Sous Chef", "Kitchen", 1, 176, 24.0, 4224),
        ("Line Cooks", "Kitchen", 3, 480, 19.0, 9120),
        ("Prep Cook", "Kitchen", 1, 160, 17.0, 2720),
        ("Dishwashers", "Support", 2, 300, 15.0, 4500),
        ("Servers", "Front of House", 6, 540, 12.0, 6480),
        ("Bartenders", "Bar", 2, 200, 15.0, 3000),
        ("Host", "Front of House", 1, 140, 15.0, 2100),
        ("General Manager", "Management", 1, 180, "Salary", 5200),
    ]
    ws = wb.create_sheet("Labor"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 16, 10, 12, 12, 14, 12])
    luxe_header(ws, "G", "👥  LABOR & SCHEDULING",
                "Control your second-biggest cost — hours, rate & monthly labor by role.")
    table_headers(ws, 4, ["Role", "Station", "Count", "Hrs / Mo", "Rate", "Monthly Cost", "% of Sales"])
    start = L0
    for i, (role, st, cnt, hrs, rate, cost) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=role).style = "td_left"
        ws.cell(row=r, column=2, value=st).style = "td"
        ws.cell(row=r, column=3, value=cnt).style = "td"
        ws.cell(row=r, column=4, value=hrs).style = "td"
        rt = ws.cell(row=r, column=5, value=rate); rt.style = "td"
        if isinstance(rate, (int, float)):
            rt.number_format = '"$"#,##0.00'
        cc = ws.cell(row=r, column=6, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0'
        pc = ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/RevenueTotal,0)"); pc.style = "td"; pc.number_format = "0.0%"
        if i % 2:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL LABOR").style = "th"
    for c in range(2, 6):
        ws.cell(row=tot, column=c).style = "th"
    cc = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0'
    pc = ws.cell(row=tot, column=7, value="=IFERROR(F" + str(tot) + "/RevenueTotal,0)"); pc.style = "td"; pc.font = Font(bold=True, color=PRIMARY); pc.fill = fill(SURFACE); pc.number_format = "0.0%"
    ws.freeze_panes = "A5"
    cell_name(wb, "LaborTotal", "Labor", f"$F${tot}")


# ===========================================================================
# 9 — Staff Roster
# ===========================================================================
def build_staff(wb):
    rows = [
        ("Marco Ellis", "Executive Chef", "Kitchen", "Full-time", dminus(1600), "ServSafe (2027)", "555-0301"),
        ("Priya Shah", "Sous Chef", "Kitchen", "Full-time", dminus(900), "ServSafe (2026)", "555-0302"),
        ("Diego Ramos", "Line Cook", "Kitchen", "Full-time", dminus(600), "Food Handler", "555-0303"),
        ("Aya Tanaka", "Line Cook", "Kitchen", "Full-time", dminus(420), "Food Handler", "555-0304"),
        ("Sam Okoye", "Line Cook", "Kitchen", "Part-time", dminus(200), "Food Handler", "555-0305"),
        ("Nora Bishop", "Server", "Front of House", "Full-time", dminus(1100), "Alcohol Cert", "555-0306"),
        ("Leo Martins", "Server", "Front of House", "Part-time", dminus(300), "Alcohol Cert", "555-0307"),
        ("Ivy Chen", "Bartender", "Bar", "Full-time", dminus(800), "Alcohol Cert", "555-0308"),
        ("Dana Reyes", "General Manager", "Management", "Full-time", dminus(1500), "ServSafe Mgr", "555-0309"),
        ("Owen Park", "Host", "Front of House", "Part-time", dminus(150), "—", "555-0310"),
    ]
    ws, start, end = build_log(
        wb, "Staff", "🧑‍🍳", "STAFF ROSTER",
        "Your team, organized — role, station, tenure, certifications & contact.",
        ["Name", "Role", "Station", "Type", "Since", "Certification", "Phone"],
        rows, [18, 18, 16, 12, 13, 16, 13],
        text_left={1, 6}, dates={5}, reserved=24,
        validations=[("C", "StationList")])
    nrange(wb, "StaffName", "Staff", "A", start, end)


# ===========================================================================
# 10 — P&L / Prime Cost
# ===========================================================================
def build_pnl(wb):
    ws = wb.create_sheet("P&L"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 14, 3, 26, 14, 2])
    luxe_header(ws, "G", "📈  P&L / PRIME COST",
                "The bottom line, live — prime cost, gross profit & net margin, automatically.")
    merge_set(ws, "B5:D5", "MONTHLY P&L", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Line", "Amount", "% of Sales"], start_col=2)
    lines = [
        ("Net Sales (Revenue)", "=RevenueTotal", "=1", False),
        ("Food Cost (COGS)", 42240, "=C8/RevenueTotal", "input"),
        ("Labor Cost", "=LaborTotal", "=C9/RevenueTotal", False),
        ("PRIME COST", "=C8+C9", "=C10/RevenueTotal", "bold"),
        ("Gross Profit", "=C7-C10", "=C11/RevenueTotal", False),
        ("Overhead / Opex", "=OpexTotal", "=C12/RevenueTotal", False),
        ("NET PROFIT", "=C7-C10-C12", "=C13/RevenueTotal", "bold"),
    ]
    for i, (lab, amt, pct, kind) in enumerate(lines):
        r = 7 + i
        cl = ws.cell(row=r, column=2, value=lab)
        cl.style = "field_label" if kind not in ("bold",) else "th"
        ca = ws.cell(row=r, column=3, value=amt)
        ca.style = "input" if kind == "input" else "field_value"
        ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=pct); cp.style = "td"; cp.number_format = "0.0%"
        if kind == "bold":
            for c in (2, 3, 4):
                ws.cell(row=r, column=c).fill = fill(SURFACE); ws.cell(row=r, column=c).font = Font(bold=True, color=PRIMARY)
        if lab == "NET PROFIT":
            ca.fill = fill(MINT_BG)
    cell_name(wb, "FoodCostTotal", "P&L", "$C$8")
    # targets vs actual (right)
    merge_set(ws, "F5:G5", "TARGET vs ACTUAL", "section_gold"); ws.row_dimensions[5].height = 22
    tv = [("Food Cost %", "=C8/RevenueTotal", "=FoodTarget"), ("Labor Cost %", "=LaborTotal/RevenueTotal", "=LaborTarget"),
          ("Prime Cost %", "=(C8+LaborTotal)/RevenueTotal", "=PrimeTarget"), ("Net Margin %", "=(C7-C10-C12)/RevenueTotal", "=MarginTarget")]
    ws.cell(row=6, column=6, value="Metric").style = "th"; ws.cell(row=6, column=7, value="Actual / Target").style = "th"
    for i, (lab, act, tgt) in enumerate(tv):
        r = 7 + i
        ws.cell(row=r, column=6, value=lab).style = "field_label"
        c = ws.cell(row=r, column=7, value=act); c.style = "field_value"; c.number_format = "0.0%"
    ws.cell(row=12, column=6, value="Targets: 30% food · 32% labor · 60% prime · 15% margin").font = Font(italic=True, color=ACCENT)


# ===========================================================================
# 11 — Expenses & Overhead
# ===========================================================================
def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 14, 14, 26])
    luxe_header(ws, "D", "🧾  EXPENSES & OVERHEAD",
                "Every fixed & variable cost beyond food and labor — your true overhead.")
    table_headers(ws, 4, ["Category", "Monthly", "Annual", "Notes"])
    vals = {"Rent": 9500, "Utilities": 3200, "Insurance": 1400, "Marketing": 2000, "Repairs & Maint.": 1500,
            "POS / Software": 900, "Cleaning & Supplies": 1600, "Licenses & Permits": 700, "Accounting": 1200,
            "Card Fees": 3600, "Miscellaneous": 1000}
    notes = {"Rent": "Base + CAM", "Card Fees": "~2.8% of card sales", "Marketing": "Ads + social"}
    start = L0; end = start + len(EXP_CATS) - 1
    for i, cat in enumerate(EXP_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cm = ws.cell(row=r, column=2, value=vals[cat]); cm.style = "input"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=notes.get(cat, "")).style = "td_left"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL OVERHEAD").style = "th"
    cm = ws.cell(row=tot, column=2, value=f"=SUM(B{start}:B{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0'
    ca = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    ws.freeze_panes = "A5"
    add_dv(ws, f"A{start}:A{end}", "ExpCatList")
    nrange(wb, "ExpCat", "Expenses", "A", start, end)
    nrange(wb, "ExpMonthly", "Expenses", "B", start, end)
    cell_name(wb, "OpexTotal", "Expenses", f"$B${tot}")


# ===========================================================================
# 12 — Cash, Tips & Deposits
# ===========================================================================
def build_cash(wb):
    rows = [
        (dminus(6), 6435, 4820, 1615, 965, 1050, "Balanced"),
        (dminus(5), 3800, 2760, 1040, 570, 780, "Balanced"),
        (dminus(4), 3480, 2510, 970, 522, 620, "Balanced"),
        (dminus(3), 4050, 2900, 1150, 608, 700, "-$4 over"),
        (dminus(2), 5580, 4180, 1400, 837, 900, "Balanced"),
        (dminus(1), 7260, 5560, 1700, 1089, 1180, "Balanced"),
    ]
    ws, start, end = build_log(
        wb, "Cash & Tips", "🧮", "CASH, TIPS & DEPOSITS",
        "Nightly reconciliation — card vs cash, tip pool & bank deposit, balanced.",
        ["Date", "Net Sales", "Card", "Cash", "Tips Pool", "Deposit", "Drawer"],
        rows, [13, 13, 12, 12, 12, 12, 16],
        text_left={7}, dates={1}, money={2, 3, 4, 5, 6}, reserved=20)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"Balanced"'], fill=fill(MINT_BG)))


# ===========================================================================
# 13 — Reservations & Events
# ===========================================================================
def build_reservations(wb):
    rows = [
        (dplus(1), "7:00p", "Whitman party", 8, "Confirmed", "Anniversary — window", "555-0410"),
        (dplus(1), "8:30p", "Corporate dinner", 12, "Confirmed", "Set menu · wine pairing", "555-0411"),
        (dplus(2), "6:30p", "Nguyen", 4, "Confirmed", "Highchair", "555-0412"),
        (dplus(3), "7:30p", "Birthday — Dana", 10, "Pending", "Cake at 9pm", "555-0413"),
        (dplus(5), "12:00p", "Rotary lunch", 20, "Confirmed", "Private room", "555-0414"),
        (dplus(7), "6:00p", "Rehearsal dinner", 30, "Pending", "Buyout inquiry", "555-0415"),
        (dplus(9), "7:00p", "Wine club", 16, "Confirmed", "5-course tasting", "555-0416"),
        (dplus(12), "5:30p", "Retirement party", 24, "Pending", "Deposit due", "555-0417"),
    ]
    ws, start, end = build_log(
        wb, "Reservations", "📆", "RESERVATIONS & EVENTS",
        "Big tables & private events — covers, status & every special request.",
        ["Date", "Time", "Party / Host", "Guests", "Status", "Notes", "Phone"],
        rows, [13, 10, 22, 10, 14, 26, 13],
        text_left={3, 6}, dates={1}, ints={4}, reserved=30,
        validations=[("E", "ResStatusList")])
    for st, cc in {"Confirmed": MINT_BG, "Pending": WARN_BG, "Cancelled": RED_BG}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 14 — Marketing & Promos
# ===========================================================================
def build_marketing(wb):
    rows = [
        ("Weekday Prix Fixe", "Dine-in", dminus(20), dplus(40), "Live", 600, "$4,800", "3-course $39"),
        ("Happy Hour 4–6", "Bar", dminus(90), dplus(90), "Live", 300, "$3,200", "Half-price apps"),
        ("Instagram Reels", "Social", dminus(14), dplus(16), "Live", 250, "$1,900", "Chef behind-the-scenes"),
        ("Loyalty Club", "CRM", dminus(120), dplus(240), "Live", 900, "$6,400", "9th meal free"),
        ("Restaurant Week", "Event", dplus(14), dplus(24), "Planned", 400, "$5,000", "City promo"),
        ("Google / Yelp Ads", "Paid", dminus(30), dplus(30), "Live", 500, "$4,100", "$2k spend"),
        ("Wine Dinner Series", "Event", dplus(20), dplus(21), "Planned", 60, "$3,900", "Ticketed $95"),
        ("Email Newsletter", "CRM", dminus(200), dplus(160), "Live", 350, "$2,000", "Monthly"),
    ]
    ws, start, end = build_log(
        wb, "Marketing", "📣", "MARKETING & PROMOS",
        "Fill more seats — campaigns, reach, attributed revenue & status.",
        ["Campaign", "Channel", "Start", "End", "Status", "Reach", "Est. Revenue", "Notes"],
        rows, [22, 12, 12, 12, 12, 10, 14, 22],
        text_left={1, 8}, dates={3, 4}, ints={6}, reserved=20,
        validations=[("E", "PromoStatusList")])
    for st, cc in {"Live": MINT_BG, "Planned": WARN_BG, "Ended": MUTED_ROW}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 15 — Reviews & Reputation
# ===========================================================================
def build_reviews(wb):
    ws = wb.create_sheet("Reviews"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 12, 12, 14, 30, 2])
    luxe_header(ws, "G", "⭐  REVIEWS & REPUTATION",
                "Your online reputation — ratings by platform and the themes guests mention.")
    merge_set(ws, "B5:F5", "PLATFORM RATINGS", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Platform", "Rating", "Reviews", "Trend", "Focus / Notes"], start_col=2)
    rows = [
        ("Google", 4.6, 320, "▲ +0.1", "Service & ambiance praised"),
        ("Yelp", 4.4, 180, "▲ +0.2", "Watch wait times"),
        ("TripAdvisor", 4.7, 95, "► flat", "Tourists love the ribeye"),
        ("OpenTable", 4.8, 140, "▲ +0.1", "Great for special occasions"),
        ("Facebook", 4.5, 60, "► flat", "Event bookings source"),
    ]
    start = 7
    for i, (plat, rt, cnt, trend, note) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=plat).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rt); cr.style = "td"; cr.number_format = "0.0"
        ws.cell(row=r, column=4, value=cnt).style = "td"
        ws.cell(row=r, column=5, value=trend).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    trow = end + 1
    ws.cell(row=trow, column=2, value="AVG RATING").style = "th"
    cr = ws.cell(row=trow, column=3, value=f"=AVERAGE(C{start}:C{end})"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = "0.0"
    ct = ws.cell(row=trow, column=4, value=f"=SUM(D{start}:D{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = "#,##0"
    ws.cell(row=trow, column=5).style = "td"; ws.cell(row=trow, column=5).fill = fill(SURFACE)
    ws.cell(row=trow, column=6).style = "td"; ws.cell(row=trow, column=6).fill = fill(SURFACE)
    cell_name(wb, "AvgRating", "Reviews", f"$C${trow}")
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=5, color=GOLD_LT, showValue=True))


# ===========================================================================
# 16 — Health & Safety Compliance
# ===========================================================================
def build_compliance(wb):
    rows = [
        ("Walk-in cooler temp (≤40°F)", "Daily", dminus(0), "Pass", "38°F logged AM & PM"),
        ("Freezer temp (≤0°F)", "Daily", dminus(0), "Pass", "-2°F"),
        ("Line hot-hold (≥135°F)", "Daily", dminus(0), "Pass", "Checked each service"),
        ("Handwash stations stocked", "Daily", dminus(0), "Pass", "Soap + towels"),
        ("Sanitizer buckets (200-400ppm)", "Daily", dminus(0), "Action Needed", "Re-test PM bucket"),
        ("Health inspection", "Annual", dminus(120), "Pass", "Score 96 — next in 8 mo"),
        ("Fire suppression service", "Semi-annual", dminus(70), "Pass", "Ansul certified"),
        ("Hood cleaning", "Quarterly", dminus(40), "Pass", "Next due in 2 mo"),
        ("Pest control", "Monthly", dminus(12), "Pass", "No activity"),
        ("Grease trap service", "Quarterly", dminus(95), "Action Needed", "Schedule this week"),
        ("Allergen matrix current", "Menu change", dminus(20), "Pass", "Updated with new menu"),
        ("Liquor license", "Annual", dminus(200), "Pass", "Renews in 5 mo"),
    ]
    ws, start, end = build_log(
        wb, "Compliance", "🛡", "HEALTH & SAFETY COMPLIANCE",
        "Inspection-ready every day — logs, cadence, status & corrective notes.",
        ["Check", "Cadence", "Last Done", "Status", "Notes"],
        rows, [30, 14, 13, 16, 30],
        text_left={1, 5}, dates={3}, reserved=24,
        validations=[("D", "ComplyStatusList")])
    for st, cc in {"Pass": MINT_BG, "Action Needed": WARN_BG, "Overdue": RED_BG}.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 17 — Cleaning & Maintenance Checklists
# ===========================================================================
def build_checklists(wb):
    ws = wb.create_sheet("Checklists"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 4, 30, 14, 2])
    luxe_header(ws, "G", "🧽  CLEANING & MAINTENANCE",
                "Open and close like clockwork — side-work & equipment upkeep, checked off.")
    merge_set(ws, "B5:C5", "OPENING CHECKLIST", "section_gold"); ws.row_dimensions[5].height = 22
    opening = [("Unlock & disarm", "Yes"), ("Turn on hood & line", "Yes"), ("Temp-check coolers", "Yes"),
               ("Prep par check", "Yes"), ("Sanitizer buckets set", "Yes"), ("FOH tables & menus", "Yes"),
               ("POS & float count", "Yes"), ("Restrooms stocked", "No"), ("Music & lights", "Yes"),
               ("Pre-shift meeting", "No")]
    for i, (task, done) in enumerate(opening):
        r = 6 + i
        ws.cell(row=r, column=2, value=task).style = "td_left"
        ws.cell(row=r, column=3, value=done).style = "td"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    merge_set(ws, "E5:F5", "CLOSING & EQUIPMENT", "section_gold"); ws.row_dimensions[5].height = 22
    closing = [("Break down & clean line", "Yes"), ("Wrap & label all prep", "Yes"), ("Log waste", "Yes"),
               ("Clean hood filters", "No"), ("Mop floors", "Yes"), ("Trash & recycling out", "Yes"),
               ("Count drawer & deposit", "Yes"), ("Descale espresso machine", "No"), ("Set alarm & lock", "Yes"),
               ("Equipment issue log", "No")]
    for i, (task, done) in enumerate(closing):
        r = 6 + i
        ws.cell(row=r, column=5, value=task).style = "td_left"
        ws.cell(row=r, column=6, value=done).style = "td"
        if i % 2:
            ws.cell(row=r, column=5).fill = fill(MUTED_ROW); ws.cell(row=r, column=6).fill = fill(MUTED_ROW)
    for rng in ("C6:C15", "F6:F15"):
        add_dv(ws, rng, "YesNoList")
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 18 — Waste & Loss Tracker
# ===========================================================================
def build_waste(wb):
    rows = [
        (dminus(6), "Salmon Fillet", "Meat & Seafood", 3.0, "lb", 33, "Over-portioned", "Spoilage"),
        (dminus(5), "Mixed Greens", "Produce", 1.0, "case", 28, "Wilted", "Spoilage"),
        (dminus(4), "Ribeye", "Meat & Seafood", 1.0, "lb", 17, "Sent back", "Comp / error"),
        (dminus(3), "Bread", "Dry Goods", 12, "each", 9, "Day-old", "Overproduction"),
        (dminus(2), "Milk", "Dairy", 1.0, "gal", 4, "Expired", "Spoilage"),
        (dminus(1), "House Red", "Bar / Liquor", 1.0, "bottle", 9, "Corked", "Quality"),
    ]
    ws, start, end = build_log(
        wb, "Waste", "🗑", "WASTE & LOSS TRACKER",
        "Plug the leaks — every dollar of waste logged by reason to protect food cost.",
        ["Date", "Item", "Category", "Qty", "Unit", "Cost", "Reason", "Type"],
        rows, [13, 20, 16, 8, 10, 11, 18, 14],
        text_left={2, 7}, dates={1}, dec={4}, money={6}, reserved=30)
    totals(ws, end + 1, [6], start, end, fmt='"$"#,##0', label="TOTAL WASTE")
    ws.conditional_formatting.add(f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=40, color=DANGER, showValue=True))


# ===========================================================================
# 19 — Vendor Payments (AP)
# ===========================================================================
def build_payments(wb):
    rows = [
        ("Prime Cuts Co.", "INV-8841", dminus(10), dplus(4), 2860, "Due", "Net 14"),
        ("Harbor Fish", "INV-2207", dminus(4), dplus(3), 1240, "Due", "Net 7"),
        ("Green Valley", "INV-5590", dminus(12), dplus(2), 980, "Due", "Net 14"),
        ("US Foods", "INV-7781", dminus(18), dplus(3), 3420, "Due", "Net 21"),
        ("Vintners Direct", "INV-3320", dminus(25), dplus(5), 2100, "Scheduled", "Net 30"),
        ("City Beverage", "INV-1188", dminus(2), dminus(2), 640, "Paid", "COD"),
        ("Bulk Pantry", "INV-4471", dminus(20), dminus(1), 720, "Overdue", "Net 14"),
        ("RestaurantDepot", "INV-9002", dminus(1), dminus(1), 310, "Paid", "COD"),
    ]
    ws, start, end = build_log(
        wb, "Payments", "💳", "VENDOR PAYMENTS (AP)",
        "Never miss a due date — invoices, amounts, terms & payment status.",
        ["Vendor", "Invoice", "Date", "Due", "Amount", "Status", "Terms"],
        rows, [20, 14, 13, 13, 12, 13, 12],
        dates={3, 4}, money={5}, reserved=24,
        validations=[("F", "PayStatusList")])
    totals(ws, end + 1, [5], start, end, fmt='"$"#,##0', label="TOTAL AP")
    for st, cc in {"Paid": MINT_BG, "Due": WARN_BG, "Overdue": RED_BG, "Scheduled": SOFT_BG}.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 20 — Training & Certifications
# ===========================================================================
def build_training(wb):
    rows = [
        ("Marco Ellis", "ServSafe Manager", dminus(400), dplus(700), "Current", "Renew online"),
        ("Priya Shah", "ServSafe Food Handler", dminus(300), dplus(65), "Renew Soon", "Book class"),
        ("Diego Ramos", "Food Handler", dminus(200), dplus(165), "Current", "—"),
        ("Nora Bishop", "Alcohol Service", dminus(500), dplus(20), "Renew Soon", "Expiring — schedule"),
        ("Ivy Chen", "Alcohol Service", dminus(150), dplus(215), "Current", "—"),
        ("All FOH", "POS & Upselling", dminus(30), dplus(150), "Current", "Quarterly refresher"),
        ("All Kitchen", "Knife Skills & Safety", dminus(60), dplus(120), "Current", "New hires pending"),
        ("Owen Park", "Food Handler", None, dplus(14), "Overdue", "New hire — must complete"),
    ]
    ws, start, end = build_log(
        wb, "Training", "🎓", "TRAINING & CERTIFICATIONS",
        "Keep everyone certified — courses, expiries & renewal alerts.",
        ["Who", "Course / Cert", "Completed", "Expires", "Status", "Action"],
        rows, [18, 22, 13, 13, 14, 22],
        text_left={2, 6}, dates={3, 4}, reserved=24)
    for st, cc in {"Current": MINT_BG, "Renew Soon": WARN_BG, "Overdue": RED_BG}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 21 — Analytics
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 3, 20, 14, 3, 18, 12, 2])
    luxe_header(ws, "I", "📊  OPERATIONS ANALYTICS",
                "The whole operation by the numbers — plus a live Operations Health Score.")
    # health dims
    merge_set(ws, "B5:C5", "OPERATIONS HEALTH", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score"], start_col=2)
    dims = [
        ("Food cost control", "=IFERROR(MIN(FoodTarget/(FoodCostTotal/RevenueTotal),1),0)"),
        ("Labor cost control", "=IFERROR(MIN(LaborTarget/(LaborTotal/RevenueTotal),1),0)"),
        ("Prime cost control", "=IFERROR(MIN(PrimeTarget/((FoodCostTotal+LaborTotal)/RevenueTotal),1),0)"),
        ("Profit margin", "=IFERROR(MIN(((RevenueTotal-FoodCostTotal-LaborTotal-OpexTotal)/RevenueTotal)/MarginTarget,1),0)"),
        ("Guest rating", "=IFERROR(AvgRating/5,0)"),
        ("Inventory in-stock", '=IFERROR(COUNTIF(InvStatus,"OK")/COUNTA(InvStatus),0)'),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Analytics", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # cost structure (donut source)
    merge_set(ws, "E5:F5", "COST STRUCTURE", "section")
    cost = [("Food Cost", "=FoodCostTotal"), ("Labor", "=LaborTotal"), ("Overhead", "=OpexTotal"),
            ("Net Profit", "=RevenueTotal-FoodCostTotal-LaborTotal-OpexTotal")]
    ws.cell(row=6, column=5, value="Component").style = "th"; ws.cell(row=6, column=6, value="Amount").style = "th"
    for i, (lab, fml) in enumerate(cost):
        r = 7 + i
        ws.cell(row=r, column=5, value=lab).style = "td_left"
        c = ws.cell(row=r, column=6, value=fml); c.style = "td"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=5).fill = fill(MUTED_ROW); ws.cell(row=r, column=6).fill = fill(MUTED_ROW)
    # snapshot
    merge_set(ws, "H5:I5", "SNAPSHOT", "section")
    snap = [("Revenue", "=RevenueTotal", '"$"#,##0'), ("Covers", "=CoversTotal", "#,##0"),
            ("Avg check", "=IFERROR(RevenueTotal/CoversTotal,0)", '"$"#,##0.00'),
            ("Inventory value", "=InvValue", '"$"#,##0'), ("Low-stock items", '=COUNTIF(InvStatus,"Low")', "#,##0"),
            ("Health score", "=IFERROR(AVERAGE(HealthRange),0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=8, value=lab).style = "field_label"
        c = ws.cell(row=r, column=9, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Health score":
            c.fill = fill(MINT_BG)
    # weekly sales for chart
    merge_set(ws, "B15:C15", "WEEKLY SALES", "section"); ws.row_dimensions[15].height = 20
    ws.cell(row=16, column=2, value="Week").style = "th"; ws.cell(row=16, column=3, value="Sales").style = "th"
    weeks = [sum(s for _, _, s in SALES[w * 7:(w + 1) * 7]) for w in range(4)]
    for i, wv in enumerate(weeks):
        r = 17 + i
        ws.cell(row=r, column=2, value=f"Week {i+1}").style = "td_left"
        c = ws.cell(row=r, column=3, value=wv); c.style = "td"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "WeekLabel", "Analytics", "$B$17:$B$20")
    cell_name(wb, "WeekSales", "Analytics", "$C$17:$C$20")


# ===========================================================================
# 1 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🍽  RESTAURANT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Sales, food & labor cost, inventory & team — your whole restaurant, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MONTHLY REVENUE", "=RevenueTotal", "money"),
        ("COVERS", "=CoversTotal", "num"),
        ("AVG CHECK", "=IFERROR(RevenueTotal/CoversTotal,0)", "money2"),
        ("FOOD COST %", "=IFERROR(FoodCostTotal/RevenueTotal,0)", "pct1"),
        ("LABOR COST %", "=IFERROR(LaborTotal/RevenueTotal,0)", "pct1"),
        ("PRIME COST %", "=IFERROR((FoodCostTotal+LaborTotal)/RevenueTotal,0)", "pct1"),
    ]
    row2 = [
        ("NET PROFIT", "=RevenueTotal-FoodCostTotal-LaborTotal-OpexTotal", "money"),
        ("PROFIT MARGIN", "=IFERROR((RevenueTotal-FoodCostTotal-LaborTotal-OpexTotal)/RevenueTotal,0)", "pct1"),
        ("INVENTORY VALUE", "=InvValue", "money"),
        ("LOW-STOCK ITEMS", '=COUNTIF(InvStatus,"Low")', "num"),
        ("AVG RATING", "=AvgRating", "dec"),
        ("HEALTH SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "COST STRUCTURE & SALES", "section_gold")
    # cost structure donut
    d1 = DoughnutChart(); d1.title = "Where Every Dollar Goes"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Analytics"], min_col=6, min_row=6, max_row=10), titles_from_data=True)
    d1.set_categories(Reference(wb["Analytics"], min_col=5, min_row=7, max_row=10)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")
    # weekly sales bar
    cb = BarChart(); cb.type = "col"; cb.title = "Weekly Sales"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=20), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=20)); cb.legend = None
    ws.add_chart(cb, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "HEALTH & MENU", "section_gold")
    # health bar
    rb = BarChart(); rb.type = "bar"; rb.title = "Operations Health"; rb.height = 8.2; rb.width = 11.5
    rb.add_data(Reference(wb["Analytics"], min_col=3, min_row=6, max_row=12), titles_from_data=True)
    rb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=7, max_row=12)); rb.legend = None
    ws.add_chart(rb, "B30")
    # menu food cost scatter-ish: use price vs food cost columns as bar of top items
    mb = BarChart(); mb.type = "bar"; mb.title = "Menu Food Cost %"; mb.height = 8.2; mb.width = 11.5
    mb.add_data(Reference(wb["Menu"], min_col=5, min_row=4, max_row=12), titles_from_data=True)
    mb.set_categories(Reference(wb["Menu"], min_col=1, min_row=5, max_row=12)); mb.legend = None
    ws.add_chart(mb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Restaurant Command Center™ — from menu costing to the bottom line, all in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_menu(wb)
    build_inventory(wb); build_par(wb); build_suppliers(wb); build_sales(wb)
    build_labor(wb); build_staff(wb); build_pnl(wb); build_expenses(wb)
    build_cash(wb); build_reservations(wb); build_marketing(wb); build_reviews(wb)
    build_compliance(wb); build_checklists(wb); build_waste(wb); build_payments(wb)
    build_training(wb); build_analytics(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Profile", "Menu", "Inventory", "Par Levels", "Suppliers", "Sales",
             "Labor", "Staff", "P&L", "Expenses", "Cash & Tips", "Reservations", "Marketing", "Reviews",
             "Compliance", "Checklists", "Waste", "Payments", "Training", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Restaurant_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
