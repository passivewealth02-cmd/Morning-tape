"""Build Next Chapter™ — Divorce Organization & Life Rebuild System (PREMIUM).

20 sheets · a calm, professional life-transition operating system.

Organizational & planning tool only — NOT legal, financial, or mental-health
advice (see the Welcome sheet disclaimer).

Run: python3 build_xlsx.py
Outputs: ../Next_Chapter.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"
ACCENT = "937356"
GOLD_LT = "C9A86A"
SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"
MINT_BG = "E3F8EF"
BG = "FFFFFF"
WHITE = "FFFFFF"
TEXT = "333333"
DANGER = "C94C4C"
RED_BG = "FBE6E6"
WARN_BG = "FBF0E2"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"
IVORY = "FBF8F2"

TASK_CATS = ["Legal", "Financial", "Documents", "Property", "Parenting",
             "Housing", "Insurance", "Personal"]
PRIORITIES = ["High", "Medium", "Low"]
STATUSES = ["Not Started", "In Progress", "Complete", "Blocked"]
DOC_STATUS = ["Needed", "In Progress", "Collected"]
EXPENSE_CATS = ["Housing", "Utilities", "Food", "Transportation", "Insurance",
                "Child Expenses", "Healthcare", "Entertainment", "Savings",
                "Miscellaneous"]
GOAL_CATS = ["Financial", "Career", "Home", "Parenting", "Health", "Personal Growth"]
ASSET_CATS = ["Cash & Savings", "Investments", "Retirement", "Property",
              "Vehicles", "Other"]
OWNER = ["Me", "Ex", "Joint"]
ASSET_TYPE = ["Asset", "Debt"]
PAY_METHODS = ["Card", "Bank Transfer", "Cash", "Check", "Auto-pay"]
CONTACT_ROLES = ["Lawyer", "Mediator", "Financial Advisor", "Therapist",
                 "School", "Doctor", "Insurance", "Utility", "Other"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40
L0 = 5
L1 = L0 + LOG_ROWS - 1

THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
# Styles & helpers (shared premium pattern)
# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"),
                            fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True),
                               fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY),
                              alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT),
                                   alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"),
                         fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT),
                              alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
                              border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY),
                            fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX),
        "prompt": NamedStyle(name="prompt", font=f(11, False, TEXT, italic=True),
                             alignment=Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True),
                             border=BOX, fill=PatternFill("solid", fgColor=IVORY)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT),
                           alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    from openpyxl.utils import column_index_from_string
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None,
               dates=None, pcts=None):
    text_left = text_left or set()
    money = money or set()
    ints = ints or set()
    dates = dates or set()
    pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else BG)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}")
    ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label
    lc.font = Font(size=10, bold=True, color=ACCENT)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vc.font = Font(size=23, bold=True, color=PRIMARY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pct": "0%",
                        "days": '0" days"'}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc)
            c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN,
                              top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 24
    ws.row_dimensions[row + 1].height = 50


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", tab=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers),
               text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 20, 4, 20, 18, 18, 18, 18, 18])
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your key inputs once — every dashboard follows.")

    merge_set(ws, "B5:C5", "YOUR INPUTS", "section")
    controls = [
        ("Process Start Date", dminus(95), "mm/dd/yyyy", "ProcessStart"),
        ("Monthly Budget", 4200, '"$"#,##0', "MonthlyBudget"),
        ("Your Name (optional)", "", None, "UserName"),
    ]
    for i, (lab, val, fmt, name) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val)
        c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Settings!$C${r}")

    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    lists = [
        ("E", "Task Category", TASK_CATS, "TaskCatList"),
        ("F", "Priority", PRIORITIES, "PriorityList"),
        ("G", "Status", STATUSES, "StatusList"),
        ("H", "Doc Status", DOC_STATUS, "DocStatusList"),
        ("I", "Expense Category", EXPENSE_CATS, "ExpenseCatList"),
        ("J", "Goal Category", GOAL_CATS, "GoalCatList"),
    ]
    for col, h, data, name in lists:
        ci = ord(col) - 64
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[name] = DefinedName(
            name, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")

    start2 = 22
    lists2 = [
        ("E", "Owner", OWNER, "OwnerList"),
        ("F", "Asset / Debt", ASSET_TYPE, "AssetTypeList"),
        ("G", "Asset Category", ASSET_CATS, "AssetCatList"),
        ("H", "Payment Method", PAY_METHODS, "PayMethodList"),
        ("I", "Contact Role", CONTACT_ROLES, "RoleList"),
        ("J", "Yes / No", YESNO, "YesNoList"),
    ]
    for col, h, data, name in lists2:
        ci = ord(col) - 64
        ws.cell(row=start2, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=start2 + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[name] = DefinedName(
            name, attr_text=f"Settings!${col}${start2+1}:${col}${start2 + len(data)}")


# ===========================================================================
# Welcome (disclaimer)
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 72, 3])
    ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ⚖  NEXT CHAPTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate divorce organization & life-rebuild system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    merge_set(ws, "B5:B5", "YOU'VE GOT THIS", "section_gold")
    ws.merge_cells("B6:B8")
    ws["B6"].value = (
        "Divorce is one of life's most complex administrative seasons. Next Chapter™ "
        "gives you one calm, organized command center for your documents, finances, "
        "schedules, tasks, parenting logistics, and the goals that will rebuild your "
        "life — so you can focus your energy where it matters most.")
    ws["B6"].style = "body"
    for r in (6, 7, 8):
        ws.row_dimensions[r].height = 22

    merge_set(ws, "B10:B10", "HOW TO USE IT", "section")
    steps = [
        "1.  Open Settings and add your process start date & monthly budget.",
        "2.  Work through the Task Manager and Document Vault at your own pace.",
        "3.  Log finances, expenses, property, and debts as you gather them.",
        "4.  Use Parenting, Appointments & Contacts to keep logistics in one place.",
        "5.  Check the Life Dashboard & Analytics to see your progress over time.",
    ]
    for i, s in enumerate(steps):
        r = 11 + i
        ws.merge_cells(f"B{r}:B{r}")
        ws[f"B{r}"].value = s
        ws[f"B{r}"].style = "body"
        ws.row_dimensions[r].height = 22

    dr = 18
    merge_set(ws, f"B{dr}:B{dr}", "  IMPORTANT — PLEASE READ", "th")
    ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+4}")
    c = ws[f"B{dr+1}"]
    c.value = (
        "This workbook is an organizational and planning tool only. It is NOT legal, "
        "financial, or mental health advice. Laws and procedures vary by jurisdiction. "
        "Please consider consulting qualified professionals (an attorney, financial "
        "advisor, and/or counselor) for advice specific to your situation.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT)
    c.fill = fill(WARN_BG)
    c.border = BOX
    for rr in range(dr + 1, dr + 5):
        ws.row_dimensions[rr].height = 22
        ws.cell(row=rr, column=2).fill = fill(WARN_BG)

    merge_set(ws, f"B{dr+6}:B{dr+6}",
              "One organized step at a time — your next chapter starts here.", "section_gold")


# ===========================================================================
# Tasks · Documents · Finances · Budget · Expenses · Appointments · Timeline · Goals
# ===========================================================================
def build_tasks(wb):
    today = dt.date.today()
    def d(o): return today + dt.timedelta(days=o)
    sample = [
        ("Legal", "Consult / retain attorney", "High", d(-60), "Complete", "Me", "Initial consult done"),
        ("Documents", "Gather last 3 yrs tax returns", "High", d(-30), "Complete", "Me", ""),
        ("Financial", "List all bank & investment accounts", "High", d(-20), "Complete", "Me", ""),
        ("Financial", "Open individual checking account", "High", d(3), "In Progress", "Me", ""),
        ("Documents", "Collect mortgage statements", "Medium", d(7), "In Progress", "Me", ""),
        ("Property", "Create home contents inventory", "Medium", d(14), "Not Started", "Me", "Photos + values"),
        ("Parenting", "Draft proposed parenting schedule", "High", d(10), "In Progress", "Me", ""),
        ("Insurance", "Review health insurance options", "Medium", d(21), "Not Started", "Me", ""),
        ("Housing", "Research new housing options", "Medium", d(30), "Not Started", "Me", ""),
        ("Financial", "Update beneficiaries & will", "High", d(25), "Not Started", "Me", "After filing"),
        ("Personal", "Change passwords & secure accounts", "High", d(-5), "Complete", "Me", ""),
        ("Legal", "Prepare financial affidavit", "High", d(18), "Not Started", "Me", ""),
    ]
    ws, start, end = build_log(
        wb, "Tasks", "✅", "MASTER TASK MANAGER",
        "Everything to do, in one place — completion tracks to your dashboard.",
        ["Category", "Task", "Priority", "Due Date", "Status", "Assigned To", "Notes"],
        sample, [16, 36, 12, 14, 14, 14, 26],
        text_left={2, 7}, dates={4},
        validations=[("A", "TaskCatList"), ("C", "PriorityList"), ("E", "StatusList")])
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Complete"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(
        f"A{start}:G{end}",
        FormulaRule(formula=[f'AND($D{start}<>"",$D{start}<TODAY(),$E{start}<>"Complete")'],
                    fill=fill(RED_BG)))
    wb.defined_names["TaskName"] = DefinedName("TaskName", attr_text=f"Tasks!$B${start}:$B${end}")
    wb.defined_names["TaskStatus"] = DefinedName("TaskStatus", attr_text=f"Tasks!$E${start}:$E${end}")
    wb.defined_names["TaskDue"] = DefinedName("TaskDue", attr_text=f"Tasks!$D${start}:$D${end}")


def build_documents(wb):
    sample = [
        ("Tax returns (3 yrs)", "Financial", "Home safe / Drive", "Yes", "Collected", dminus(30), ""),
        ("Bank statements (12 mo)", "Financial", "Drive folder", "Yes", "Collected", dminus(20), ""),
        ("Mortgage statement", "Property", "Lender portal", "No", "In Progress", "", "Download PDF"),
        ("Home deed / title", "Property", "Safe deposit box", "No", "Needed", "", ""),
        ("Auto titles", "Property", "Glovebox / file", "No", "Needed", "", ""),
        ("Retirement statements", "Financial", "Provider portal", "Yes", "Collected", dminus(15), ""),
        ("Investment statements", "Financial", "Drive folder", "Yes", "Collected", dminus(15), ""),
        ("Pay stubs (recent)", "Financial", "HR portal", "Yes", "In Progress", "", ""),
        ("Health insurance card/policy", "Insurance", "Wallet / portal", "Yes", "Collected", dminus(10), ""),
        ("Life insurance policy", "Insurance", "File cabinet", "No", "Needed", "", ""),
        ("Birth certificates", "Personal", "Safe", "No", "Needed", "", "Kids + self"),
        ("Marriage certificate", "Legal", "Safe", "Yes", "Collected", dminus(40), ""),
        ("Driver's license / ID", "Personal", "Wallet", "Yes", "Collected", dminus(40), ""),
        ("Will / Trust", "Legal", "Attorney", "No", "In Progress", "", "Updating"),
        ("Credit card statements", "Financial", "Issuer portal", "Yes", "Collected", dminus(12), ""),
    ]
    ws, start, end = build_log(
        wb, "Documents", "📁", "DOCUMENT VAULT INDEX",
        "Know exactly what you have and what's still needed.",
        ["Document", "Category", "Location", "Digital Copy", "Status", "Date Updated", "Notes"],
        sample, [30, 16, 22, 13, 14, 14, 22],
        text_left={1, 3, 7}, dates={6},
        validations=[("D", "YesNoList"), ("E", "DocStatusList")])
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Collected"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Needed"'], fill=fill(WARN_BG)))
    wb.defined_names["DocName"] = DefinedName("DocName", attr_text=f"Documents!$A${start}:$A${end}")
    wb.defined_names["DocStatus"] = DefinedName("DocStatus", attr_text=f"Documents!$E${start}:$E${end}")


def build_financial(wb):
    ws = wb.create_sheet("Finances")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 12, 18, 14, 4, 24, 16])
    luxe_header(ws, "G", "💰  FINANCIAL SNAPSHOT",
                "Assets, debts, and your net-worth picture — gathered in one place.")
    table_headers(ws, 4, ["Account / Item", "Type", "Category", "Value"])
    sample = [
        ("Joint checking", "Asset", "Cash & Savings", 8400),
        ("Personal savings", "Asset", "Cash & Savings", 12500),
        ("401(k)", "Asset", "Retirement", 86000),
        ("Roth IRA", "Asset", "Retirement", 24000),
        ("Brokerage account", "Asset", "Investments", 31000),
        ("Home (equity est.)", "Asset", "Property", 140000),
        ("Vehicle 1", "Asset", "Vehicles", 18000),
        ("Vehicle 2", "Asset", "Vehicles", 9000),
        ("Mortgage", "Debt", "Property", 215000),
        ("Auto loan", "Debt", "Vehicles", 11000),
        ("Credit card 1", "Debt", "Other", 4200),
        ("Credit card 2", "Debt", "Other", 1800),
        ("Student loan", "Debt", "Other", 9500),
    ]
    start = L0
    end = start + 30 - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, 4, text_left={1}, money={4})
    add_dv(ws, f"B{start}:B{end}", "AssetTypeList")
    add_dv(ws, f"C{start}:C{end}", "AssetCatList")
    wb.defined_names["SnapType"] = DefinedName("SnapType", attr_text=f"Finances!$B${start}:$B${end}")
    wb.defined_names["SnapCat"] = DefinedName("SnapCat", attr_text=f"Finances!$C${start}:$C${end}")
    wb.defined_names["SnapValue"] = DefinedName("SnapValue", attr_text=f"Finances!$D${start}:$D${end}")

    # Summary
    merge_set(ws, "F4:G4", "NET WORTH", "section_gold")
    rows = [
        ("Total Assets", '=SUMIF(SnapType,"Asset",SnapValue)', "TotalAssets"),
        ("Total Debts", '=SUMIF(SnapType,"Debt",SnapValue)', "TotalDebts"),
        ("Net Worth", "=TotalAssets-TotalDebts", "NetWorth"),
        ("Cash & Savings", '=SUMIFS(SnapValue,SnapType,"Asset",SnapCat,"Cash & Savings")', "SavingsBalance"),
    ]
    for i, (lab, fml, name) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=6, value=lab).style = "field_label"
        c = ws.cell(row=r, column=7, value=fml); c.style = "field_value"
        c.number_format = '"$"#,##0;[Red]-"$"#,##0'
        wb.defined_names[name] = DefinedName(name, attr_text=f"Finances!$G${r}")

    # Asset allocation summary (for donut)
    merge_set(ws, "F11:G11", "ASSET ALLOCATION", "section_gold")
    ws.cell(row=12, column=6, value="Category").style = "th"
    ws.cell(row=12, column=7, value="Value").style = "th"
    for i, cat in enumerate(ASSET_CATS):
        r = 13 + i
        ws.cell(row=r, column=6, value=cat).style = "td_left"
        c = ws.cell(row=r, column=7, value=f'=SUMIFS(SnapValue,SnapType,"Asset",SnapCat,F{r})')
        c.style = "td"; c.number_format = '"$"#,##0'
    aend = 12 + len(ASSET_CATS)
    wb.defined_names["AllocLabels"] = DefinedName("AllocLabels", attr_text=f"Finances!$F$13:$F${aend}")
    wb.defined_names["AllocValues"] = DefinedName("AllocValues", attr_text=f"Finances!$G$13:$G${aend}")

    # Net worth bar (assets vs debts)
    ws.cell(row=22, column=6, value="Assets").style = "td_left"
    ws.cell(row=22, column=7, value="=TotalAssets").number_format = '"$"#,##0'
    ws.cell(row=23, column=6, value="Debts").style = "td_left"
    ws.cell(row=23, column=7, value="=TotalDebts").number_format = '"$"#,##0'
    bar = BarChart(); bar.type = "col"; bar.title = "Assets vs Debts"; bar.height = 7.5; bar.width = 10
    bar.add_data(Reference(ws, min_col=7, min_row=22, max_row=23), titles_from_data=False)
    bar.set_categories(Reference(ws, min_col=6, min_row=22, max_row=23))
    bar.legend = None
    ws.add_chart(bar, "A38")
    ws.freeze_panes = "A5"


def build_budget(wb):
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 16, 16, 16, 16, 4, 22, 18])
    luxe_header(ws, "H", "📊  MONTHLY BUDGET",
                "Plan vs actual by category. Actuals pull from your Expense Tracker.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining", "% Used"])
    planned = {"Housing": 1500, "Utilities": 260, "Food": 600, "Transportation": 320,
               "Insurance": 280, "Child Expenses": 450, "Healthcare": 180,
               "Entertainment": 150, "Savings": 400, "Miscellaneous": 160}
    start = L0
    end = start + len(EXPENSE_CATS) - 1
    for i, cat in enumerate(EXPENSE_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=2, value=planned.get(cat, 0)); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f'=IFERROR(SUMIF(ExpCat,A{r},ExpAmount),0)')
        ca.style = "td"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in range(2, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=total, column=5, value=f"=IFERROR(C{total}/B{total},0)")
    cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    wb.defined_names["BudgetCat"] = DefinedName("BudgetCat", attr_text=f"Budget!$A${start}:$A${end}")
    wb.defined_names["BudgetActual"] = DefinedName("BudgetActual", attr_text=f"Budget!$C${start}:$C${end}")
    wb.defined_names["BudgetTotalPlanned"] = DefinedName("BudgetTotalPlanned", attr_text=f"Budget!$B${total}")
    wb.defined_names["BudgetTotalActual"] = DefinedName("BudgetTotalActual", attr_text=f"Budget!$C${total}")

    # KPI sidebar + donut
    ws.cell(row=4, column=7, value="THIS MONTH").style = "section_gold"
    kpis = [("Planned", "=BudgetTotalPlanned"), ("Actual", "=BudgetTotalActual"),
            ("Remaining", "=BudgetTotalPlanned-BudgetTotalActual")]
    for i, (lab, fml) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml); c.style = "field_value"; c.number_format = '"$"#,##0'
    donut = DoughnutChart(); donut.title = "Spending by Category"; donut.height = 8; donut.width = 13
    donut.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    donut.dataLabels = no_labels()
    ws.add_chart(donut, "G10")
    ws.freeze_panes = "A5"


def build_expenses(wb):
    today = dt.date.today()
    def d(o): return today + dt.timedelta(days=o)
    sample = [
        (d(-2), "Housing", 1500, "Auto-pay", "Rent / mortgage"),
        (d(-2), "Utilities", 95, "Card", "Electric"),
        (d(-5), "Food", 142, "Card", "Groceries"),
        (d(-6), "Child Expenses", 220, "Card", "After-school care"),
        (d(-8), "Transportation", 60, "Card", "Gas"),
        (d(-10), "Insurance", 280, "Auto-pay", "Health premium"),
        (d(-12), "Food", 38, "Card", "Dinner out"),
        (d(-14), "Healthcare", 45, "Card", "Copay"),
        (d(-16), "Entertainment", 35, "Card", "Streaming + activity"),
        (d(-18), "Savings", 400, "Bank Transfer", "Emergency fund"),
        (d(-20), "Miscellaneous", 60, "Card", "Household"),
        (d(-22), "Transportation", 55, "Card", "Gas"),
        (d(-24), "Food", 128, "Card", "Groceries"),
    ]
    ws, start, end = build_log(
        wb, "Expenses", "🧾", "EXPENSE TRACKER",
        "Log spending as it happens — your budget actuals update automatically.",
        ["Date", "Category", "Amount", "Payment Method", "Notes"],
        sample, [13, 18, 14, 18, 28],
        text_left={5}, dates={1}, money={3},
        validations=[("B", "ExpenseCatList"), ("D", "PayMethodList")], reserved=60)
    wb.defined_names["ExpDate"] = DefinedName("ExpDate", attr_text=f"Expenses!$A${start}:$A${end}")
    wb.defined_names["ExpCat"] = DefinedName("ExpCat", attr_text=f"Expenses!$B${start}:$B${end}")
    wb.defined_names["ExpAmount"] = DefinedName("ExpAmount", attr_text=f"Expenses!$C${start}:$C${end}")


def build_appointments(wb):
    today = dt.date.today()
    def d(o): return today + dt.timedelta(days=o)
    sample = [
        ("Attorney strategy meeting", "Legal", d(4), "10:00 AM", "Law office / Zoom", "Bring affidavit draft"),
        ("Mediation session 1", "Mediation", d(12), "1:00 PM", "Mediation center", ""),
        ("Financial planning review", "Financial", d(18), "9:30 AM", "Advisor office", "Bring net-worth sheet"),
        ("Counseling session", "Personal", d(6), "5:00 PM", "Therapist", ""),
        ("Kids' parent-teacher meeting", "School", d(9), "3:30 PM", "School", ""),
        ("Notary for documents", "Legal", d(15), "11:00 AM", "Bank", ""),
    ]
    ws, start, end = build_log(
        wb, "Appointments", "📅", "APPOINTMENTS",
        "Every meeting in one calm calendar — with automatic countdowns.",
        ["Appointment", "Type", "Date", "Time", "Location", "Notes"],
        sample, [28, 16, 14, 12, 22, 24],
        text_left={1, 5, 6}, dates={3}, reserved=40)
    # Days Until helper in col G
    ws.cell(row=4, column=7, value="Days Until").style = "th"
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=7, value=f'=IF(C{r}="","",C{r}-TODAY())')
        c.style = "td"; c.number_format = "0;[Red]-0"
        c.fill = fill(MUTED_ROW if (r - start) % 2 else BG)
    ws.conditional_formatting.add(
        f"A{start}:G{end}",
        FormulaRule(formula=[f'AND($C{start}<>"",$C{start}-TODAY()>=0,$C{start}-TODAY()<=7)'],
                    fill=fill(MINT_BG)))
    wb.defined_names["ApptDate"] = DefinedName("ApptDate", attr_text=f"Appointments!$C${start}:$C${end}")


def build_timeline(wb):
    today = dt.date.today()
    def d(o): return today + dt.timedelta(days=o)
    sample = [
        ("Process started", "Milestone", dminus(95), "First consultation"),
        ("Financial disclosures exchanged", "Filing", d(-40), ""),
        ("Petition filed", "Filing", d(-30), ""),
        ("Mediation session 1", "Mediation", d(12), ""),
        ("Mediation session 2", "Mediation", d(40), ""),
        ("Settlement review", "Meeting", d(60), ""),
        ("Court date (if needed)", "Court", d(90), "Tentative"),
        ("Finalization (est.)", "Milestone", d(120), "Estimate only"),
    ]
    ws, start, end = build_log(
        wb, "Timeline", "🗓", "MASTER TIMELINE",
        "Key dates and milestones with automatic countdowns. (Dates are examples.)",
        ["Event", "Type", "Date", "Notes"],
        sample, [30, 16, 14, 30],
        text_left={1, 4}, dates={3}, reserved=40)
    ws.cell(row=4, column=5, value="Days Until").style = "th"
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=5, value=f'=IF(C{r}="","",C{r}-TODAY())')
        c.style = "td"; c.number_format = "0;[Red]-0"
        c.fill = fill(MUTED_ROW if (r - start) % 2 else BG)


def build_goals(wb):
    sample = [
        ("Build 6-month emergency fund", "Financial", dminus(-180), 0.35, "In Progress", ""),
        ("Secure stable housing", "Home", dminus(-60), 0.5, "In Progress", ""),
        ("Update career / income plan", "Career", dminus(-90), 0.2, "In Progress", ""),
        ("Establish steady parenting routine", "Parenting", dminus(-45), 0.6, "In Progress", ""),
        ("Rebuild individual credit", "Financial", dminus(-365), 0.25, "In Progress", ""),
        ("Prioritize health & wellbeing", "Health", dminus(-120), 0.4, "In Progress", ""),
    ]
    ws, start, end = build_log(
        wb, "Goals", "🎯", "GOAL PLANNER",
        "Your forward-looking goals — progress over perfection.",
        ["Goal", "Category", "Target Date", "Progress", "Status", "Notes"],
        sample, [34, 16, 14, 12, 14, 26],
        text_left={1, 6}, dates={3},
        validations=[("B", "GoalCatList"), ("E", "StatusList")])
    for r in range(start, end + 1):
        ws.cell(row=r, column=4).number_format = "0%"
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))
    wb.defined_names["GoalStatus"] = DefinedName("GoalStatus", attr_text=f"Goals!$E${start}:$E${end}")
    wb.defined_names["GoalProgress"] = DefinedName("GoalProgress", attr_text=f"Goals!$D${start}:$D${end}")
    wb.defined_names["GoalName"] = DefinedName("GoalName", attr_text=f"Goals!$A${start}:$A${end}")


# ===========================================================================
# Parenting (with weekly schedule -> dashboard status)
# ===========================================================================
def build_parenting(wb):
    ws = wb.create_sheet("Parenting")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [14, 20, 34, 4, 22, 22])
    luxe_header(ws, "F", "👪  CHILD & PARENTING ORGANIZER",
                "Schedules, activities, and key contacts — calm logistics for the kids.")

    merge_set(ws, "A5:C5", "WEEKLY PARENTING SCHEDULE", "section_gold")
    ws.cell(row=6, column=1, value="Day").style = "th"
    ws.cell(row=6, column=2, value="With").style = "th"
    ws.cell(row=6, column=3, value="Notes / Hand-off").style = "th"
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    withp = ["Me", "Me", "Ex", "Ex", "Me", "Me", "Ex"]
    for i, day in enumerate(days):
        r = 7 + i
        ws.cell(row=r, column=1, value=day).style = "td_left"
        ws.cell(row=r, column=2, value=withp[i]).style = "td"
        ws.cell(row=r, column=3, value="").style = "td_left"
        if i % 2:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    add_dv(ws, "B7:B13", "OwnerList")
    wb.defined_names["ParentSchedule"] = DefinedName("ParentSchedule", attr_text="Parenting!$B$7:$B$13")

    # Activities / medical on the right
    merge_set(ws, "E5:F5", "ACTIVITIES & APPOINTMENTS", "section_gold")
    ws.cell(row=6, column=5, value="Item").style = "th"
    ws.cell(row=6, column=6, value="When / Notes").style = "th"
    items = [("Soccer practice", "Tue/Thu 4pm"), ("Piano lesson", "Sat 10am"),
             ("Pediatrician", "Schedule annual"), ("Dentist", "Due in spring"),
             ("School calendar", "See Contacts"), ("Childcare", "After-school")]
    for i, (a, b) in enumerate(items):
        r = 7 + i
        ws.cell(row=r, column=5, value=a).style = "td_left"
        ws.cell(row=r, column=6, value=b).style = "td_left"
        if i % 2:
            ws.cell(row=r, column=5).fill = fill(MUTED_ROW)
            ws.cell(row=r, column=6).fill = fill(MUTED_ROW)


# ===========================================================================
# Simpler sheets (profile, property, debts, contacts, rebuild, journal, checklist)
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Profile")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 30, 6, 26, 30, 2])
    luxe_header(ws, "G", "👤  PERSONAL PROFILE",
                "Your key people and reference details — all in one secure place.")
    blocks = [
        ("KEY DATES", [("Date of marriage", ""), ("Date of separation", ""),
                       ("Process start date", "=ProcessStart"), ("Target finalization", "")]),
        ("LEGAL & PROFESSIONAL", [("Attorney", ""), ("Attorney phone", ""),
                                  ("Mediator", ""), ("Mediator phone", ""),
                                  ("Financial advisor", ""), ("Therapist / counselor", "")]),
        ("REFERENCE NUMBERS", [("Case / file number", ""), ("Attorney client ID", ""),
                               ("Insurance member ID", ""), ("Emergency contact", "")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold")
        ws.row_dimensions[row].height = 22
        row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            cv = ws.cell(row=row, column=3, value=fields[i][1]); cv.style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24
            i += 2
            row += 1
        row += 1


def build_simple(wb):
    # Property & Asset Inventory
    build_log(
        wb, "Property", "🏠", "PROPERTY & ASSET INVENTORY",
        "Document what you own, its value, and who it belongs to.",
        ["Item / Description", "Category", "Estimated Value", "Owner", "Notes"],
        [
            ("Family home", "Property", 360000, "Joint", "Mortgage on Finances tab"),
            ("Vehicle 1 (SUV)", "Vehicle", 18000, "Me", ""),
            ("Vehicle 2 (Sedan)", "Vehicle", 9000, "Ex", ""),
            ("Living room furniture", "Furniture", 3500, "Joint", "Photos saved"),
            ("Electronics (TVs, laptops)", "Electronics", 2800, "Joint", ""),
            ("Jewelry", "Jewelry", 4200, "Me", "Appraisal pending"),
            ("Investment account", "Investments", 31000, "Joint", "See Finances"),
            ("Collectibles", "Collectibles", 1500, "Me", ""),
        ],
        [28, 16, 16, 12, 26],
        text_left={1, 5}, money={3}, validations=[("D", "OwnerList")], reserved=40)

    # Debt Tracker
    build_log(
        wb, "Debts", "💳", "DEBT TRACKER",
        "Track balances, payments, and who's responsible for each.",
        ["Creditor", "Balance", "Interest %", "Monthly Payment", "Due Date", "Responsibility", "Status"],
        [
            ("Mortgage lender", 215000, 0.0625, 1450, "1st", "Joint", "Current"),
            ("Auto loan", 11000, 0.049, 320, "15th", "Me", "Current"),
            ("Credit card 1", 4200, 0.219, 150, "20th", "Me", "Paying down"),
            ("Credit card 2", 1800, 0.245, 75, "5th", "Ex", "Disputed"),
            ("Student loan", 9500, 0.045, 110, "28th", "Me", "Current"),
        ],
        [22, 14, 12, 16, 12, 16, 16],
        text_left={1, 7}, money={2, 4}, validations=[("F", "OwnerList")], reserved=30)
    ws = wb["Debts"]
    for r in range(L0, L0 + 30):
        ws.cell(row=r, column=3).number_format = "0.0%"

    # Contact Directory
    build_log(
        wb, "Contacts", "📇", "CONTACT DIRECTORY",
        "Everyone you need to reach — lawyers, advisors, schools & more.",
        ["Name", "Role", "Phone", "Email", "Notes"],
        [
            ("(your attorney)", "Lawyer", "", "", "Primary contact"),
            ("(mediator)", "Mediator", "", "", ""),
            ("(financial advisor)", "Financial Advisor", "", "", ""),
            ("(therapist)", "Therapist", "", "", ""),
            ("(children's school)", "School", "", "", ""),
            ("(pediatrician)", "Doctor", "", "", ""),
            ("(health insurance)", "Insurance", "", "", ""),
            ("(utility company)", "Utility", "", "", ""),
        ],
        [24, 18, 18, 26, 24],
        text_left={1, 4, 5}, validations=[("B", "RoleList")], reserved=40)

    # Life Rebuild Planner
    build_log(
        wb, "Life Rebuild", "🌱", "LIFE REBUILD PLANNER",
        "Design the life you're moving toward — one area at a time.",
        ["Area", "Action / Goal", "Target", "Status", "Notes"],
        [
            ("New Routine", "Set a calm weekday morning routine", "This month", "In Progress", ""),
            ("Weekly Habits", "Meal prep on Sundays", "Ongoing", "Not Started", ""),
            ("Home Setup", "Make new space feel like home", "60 days", "In Progress", ""),
            ("Career", "Update resume & LinkedIn", "30 days", "Not Started", ""),
            ("Education", "Research a course / certification", "90 days", "Not Started", ""),
            ("Savings", "Automate monthly transfer", "This month", "In Progress", ""),
            ("Fitness", "Walk or workout 3x/week", "Ongoing", "In Progress", ""),
        ],
        [18, 34, 14, 14, 26],
        text_left={2, 5}, validations=[("D", "StatusList")], reserved=30)

    # Journal & Notes
    build_log(
        wb, "Journal", "📓", "JOURNAL & NOTES",
        "A private space to process the week — wins, challenges & what's next.",
        ["Week Of", "Wins", "Challenges", "Gratitude", "Future Plans"],
        [
            (dminus(7), "Opened my own account", "Hard conversation about schedule",
             "Support from a friend", "Finalize parenting plan"),
            (dminus(0), "Stayed organized with docs", "Felt overwhelmed mid-week",
             "A calm weekend", "Start emergency fund"),
        ],
        [13, 26, 26, 24, 26],
        text_left={2, 3, 4, 5}, dates={1}, reserved=40)

    # Document Checklist
    build_log(
        wb, "Doc Checklist", "📋", "DOCUMENT CHECKLIST",
        "A quick-tick list of the paperwork most processes need.",
        ["Document", "Category", "Have It?", "Notes"],
        [
            ("Tax returns (3 years)", "Financial", "Yes", ""),
            ("Bank statements (12 months)", "Financial", "Yes", ""),
            ("Pay stubs (recent)", "Financial", "No", ""),
            ("Mortgage / lease documents", "Property", "No", ""),
            ("Property deeds / titles", "Property", "No", ""),
            ("Vehicle titles", "Property", "No", ""),
            ("Retirement & investment statements", "Financial", "Yes", ""),
            ("Credit card statements", "Financial", "Yes", ""),
            ("Insurance policies", "Insurance", "No", ""),
            ("Wills / trusts", "Legal", "No", ""),
            ("Birth & marriage certificates", "Personal", "No", ""),
            ("IDs / passports", "Personal", "Yes", ""),
            ("Loan documents", "Financial", "No", ""),
            ("Monthly expense records", "Financial", "Yes", ""),
        ],
        [34, 16, 12, 28],
        text_left={1, 4}, validations=[("C", "YesNoList")], reserved=40)
    ws = wb["Doc Checklist"]
    ws.conditional_formatting.add(
        "C5:C44", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# Analytics
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 18, 4, 30, 18, 2])
    luxe_header(ws, "G", "📈  ANALYTICS",
                "Your progress at a glance across tasks, documents, budget & goals.")

    merge_set(ws, "B5:D5", "HEALTH SCORES", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Task Completion", '=IFERROR(COUNTIF(TaskStatus,"Complete")/MAX(COUNTA(TaskName),1),0)'),
        ("Documents Collected", '=IFERROR(COUNTIF(DocStatus,"Collected")/MAX(COUNTA(DocName),1),0)'),
        ("Budget Used", '=IFERROR(BudgetTotalActual/BudgetTotalPlanned,0)'),
        ("Goal Progress", '=IFERROR(AVERAGE(GoalProgress),0)'),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4,
                value=f'=IF(C{r}>=0.75,"On Track",IF(C{r}>=0.4,"In Progress","Getting Started"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(
        f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF" + WARN_BG,
                       mid_type="num", mid_value=0.5, mid_color="FFFFF3CD",
                       end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))

    # Overall readiness
    merge_set(ws, "F5:G5", "OVERALL PROGRESS", "section_gold")
    ws.merge_cells("F6:G8")
    cell = ws["F6"]; cell.value = f"=AVERAGE(C{start}:C{start+1},C{end})"
    cell.font = Font(size=44, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 9):
        for cc in (6, 7):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(
                top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)

    bar = BarChart(); bar.type = "bar"; bar.title = "Progress by Area"; bar.height = 8; bar.width = 14
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None
    ws.add_chart(bar, "F11")


# ===========================================================================
# Life Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [15] * 12 + [2])
    ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ⚖  NEXT CHAPTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2",
              "  Organize documents, finances, schedules & goals — and move forward with clarity and calm.",
              "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    row1 = [
        ("DAYS IN PROCESS", "=MAX(TODAY()-ProcessStart,0)", "days"),
        ("TASKS COMPLETED", '=COUNTIF(TaskStatus,"Complete")', "num"),
        ("DOCS COLLECTED", '=COUNTIF(DocStatus,"Collected")', "num"),
        ("UPCOMING APPTS", '=SUMPRODUCT((ApptDate<>"")*(ApptDate>=TODAY()))', "num"),
    ]
    row2 = [
        ("MONTHLY BUDGET", "=MonthlyBudget", "money"),
        ("SAVINGS BALANCE", "=SavingsBalance", "money"),
        ("PARENTING SET", '=IFERROR(COUNTA(ParentSchedule)/7,0)', "pct"),
        ("OVERALL PROGRESS",
         '=IFERROR((COUNTIF(TaskStatus,"Complete")/MAX(COUNTA(TaskName),1)'
         '+COUNTIF(DocStatus,"Collected")/MAX(COUNTA(DocName),1)'
         '+AVERAGE(GoalProgress))/3,0)', "pct"),
    ]
    cols = [2, 5, 8, 11]
    for (lab, fml, kind), col in zip(row1, cols):
        kpi_card(ws, 5, col, 3, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols):
        kpi_card(ws, 8, col, 3, lab, fml, kind)

    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "YOUR PROGRESS", "section_gold")

    # Budget donut
    bend = L0 + len(EXPENSE_CATS) - 1
    donut = DoughnutChart(); donut.title = "Spending by Category"; donut.height = 8.5; donut.width = 12
    donut.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=bend), titles_from_data=True)
    donut.set_categories(Reference(wb["Budget"], min_col=1, min_row=L0, max_row=bend))
    donut.dataLabels = no_labels()
    ws.add_chart(donut, "B12")

    # Asset allocation donut
    aend = 12 + len(ASSET_CATS)
    alloc = DoughnutChart(); alloc.title = "Asset Allocation"; alloc.height = 8.5; alloc.width = 12
    alloc.add_data(Reference(wb["Finances"], min_col=7, min_row=12, max_row=aend), titles_from_data=True)
    alloc.set_categories(Reference(wb["Finances"], min_col=6, min_row=13, max_row=aend))
    alloc.dataLabels = no_labels()
    ws.add_chart(alloc, "H12")

    ws.row_dimensions[30].height = 26
    merge_set(ws, "B30:M30", "READINESS", "section")

    # Analytics progress bar
    ab = BarChart(); ab.type = "bar"; ab.title = "Progress by Area"; ab.height = 8.5; ab.width = 12
    ab.add_data(Reference(wb["Analytics"], min_col=3, min_row=6, max_row=10), titles_from_data=True)
    ab.set_categories(Reference(wb["Analytics"], min_col=2, min_row=7, max_row=10))
    ab.legend = None
    ws.add_chart(ab, "B31")

    # Net worth bar
    nb = BarChart(); nb.type = "col"; nb.title = "Assets vs Debts"; nb.height = 8.5; nb.width = 12
    nb.add_data(Reference(wb["Finances"], min_col=7, min_row=22, max_row=23), titles_from_data=False)
    nb.set_categories(Reference(wb["Finances"], min_col=6, min_row=22, max_row=23))
    nb.legend = None
    ws.add_chart(nb, "H31")

    ws.row_dimensions[49].height = 26
    merge_set(ws, "B49:M49",
              "Next Chapter™ is an organizational tool — not legal, financial, or mental-health advice. See the Welcome tab.",
              "subtitle")


# shared no-labels helper
def no_labels():
    dl = DataLabelList()
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    dl.showVal = False
    dl.showPercent = False
    return dl


# ===========================================================================
# Build
# ===========================================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    build_settings(wb)
    build_welcome(wb)
    build_profile(wb)
    build_timeline(wb)
    build_tasks(wb)
    build_documents(wb)
    build_financial(wb)
    build_budget(wb)
    build_expenses(wb)
    build_parenting(wb)
    build_appointments(wb)
    build_goals(wb)
    build_simple(wb)        # property, debts, contacts, life rebuild, journal, doc checklist
    build_analytics(wb)
    build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Profile", "Timeline", "Tasks", "Documents",
             "Finances", "Budget", "Property", "Debts", "Parenting", "Expenses",
             "Appointments", "Contacts", "Goals", "Life Rebuild", "Journal",
             "Doc Checklist", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "Next_Chapter.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
