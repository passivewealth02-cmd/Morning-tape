"""Build Food Cost & Inventory Command Center™ — The Food-Cost Control System.

14 tabs · a premium food-cost & inventory operating system in Google Sheets &
Excel. Dashboard, a period food-cost engine (Beginning + Purchases − Ending ÷
Sales), inventory valuation, a purchases & sales log, usage vs theoretical
variance, par & ordering, vendors, a price tracker, menu costing, waste and a
category breakdown — one dashboard. Know your food cost to the point.

Run: python3 build_xlsx.py   ->  ../Food_Cost_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["each", "lb", "oz", "case", "box", "jug", "gal", "lot"]
CATS = ["Meat & seafood", "Produce", "Dairy & eggs", "Dry goods", "Beverages", "Paper/disposables"]
STATUS = ["Preferred", "Backup", "New", "Watch"]

TARGET_FC = 0.30
VAR_LIMIT = 0.05
TURN_GOAL = 1.5
WASTE_LIMIT = 0.03
BEGINNING = 12400

# Inventory count: (item, count, unit cost, unit) — extended value sums to Ending inventory
INVENTORY = [
    ("Beef", 120, 6.50, "lb"), ("Chicken", 180, 2.80, "lb"), ("Salmon", 60, 9.00, "lb"),
    ("Produce mix", 24, 32.00, "case"), ("Cheese", 18, 48.00, "case"), ("Dairy", 20, 28.00, "case"),
    ("Flour / dry", 40, 22.00, "case"), ("Canned goods", 60, 18.00, "case"), ("Beverages", 50, 16.00, "case"),
    ("Wine / beer", 30, 40.00, "case"), ("Frozen", 45, 35.00, "case"), ("Paper goods", 55, 15.00, "case"),
    ("Oil / fats", 24, 19.00, "jug"), ("Spices / misc", 1, 968.00, "lot"),
]

# Purchases log: (date-day, vendor, category, amount) — sums to Purchases
PURCHASES = [
    ("Meat Co.", "Meat & seafood", 6800),
    ("Green Farms", "Produce", 3600),
    ("Dairy Best", "Dairy & eggs", 2300),
    ("Broadline Foods", "Dry goods", 3200),
    ("BevDist", "Beverages", 1700),
    ("PaperPlus", "Paper/disposables", 1000),
]

# Sales log: (period, sales) — sums to Sales
SALES = [("Week 1", 15200), ("Week 2", 16400), ("Week 3", 16800), ("Week 4", 15600)]

# Usage & variance: (category, theoretical usage, actual usage). Actual sums to food-cost $.
USAGE = [
    ("Meat & seafood", 7000, 7200), ("Produce", 3800, 3900), ("Dairy & eggs", 2350, 2400),
    ("Dry goods", 2650, 2700), ("Beverages", 1750, 1800), ("Paper/disposables", 1100, 1200),
]

# Par & ordering: (item, par, on hand, unit cost)
PARORDER = [
    ("Beef", 150, 120, 6.50), ("Chicken", 220, 180, 2.80), ("Salmon", 90, 60, 9.00),
    ("Produce mix", 40, 24, 32.00), ("Cheese", 30, 18, 48.00), ("Dairy", 30, 20, 28.00),
    ("Dry goods", 60, 40, 22.00), ("Beverages", 70, 50, 16.00),
]

# Vendors: (vendor, category, contact, status)
VENDORS = [
    ("Meat Co.", "Meat & seafood", "orders@meatco.com", "Preferred"),
    ("Green Farms", "Produce", "sales@greenfarms.com", "Preferred"),
    ("Dairy Best", "Dairy & eggs", "hello@dairybest.com", "Preferred"),
    ("Broadline Foods", "Dry goods", "csr@broadline.com", "Backup"),
    ("BevDist", "Beverages", "orders@bevdist.com", "Preferred"),
    ("PaperPlus", "Paper/disposables", "team@paperplus.com", "Backup"),
]

# Price tracker: (item, vendor A, price A, vendor B, price B)
PRICES = [
    ("Beef (lb)", "Meat Co.", 6.50, "Butcher Bros", 6.85),
    ("Chicken (lb)", "Meat Co.", 2.80, "Poultry Direct", 2.65),
    ("Salmon (lb)", "SeaFresh", 9.00, "Ocean Co.", 9.40),
    ("Produce mix (case)", "Green Farms", 32.00, "City Produce", 34.50),
    ("Cheese (case)", "Dairy Best", 48.00, "Cheese Hub", 46.00),
    ("Cooking oil (jug)", "Broadline Foods", 19.00, "Restaurant Depot", 17.80),
]

# Menu costing: (item, plate cost, menu price)
MENU = [
    ("House Burger", 3.85, 14.00), ("Grilled Salmon", 6.20, 24.00), ("Caesar Salad", 2.10, 12.00),
    ("Pasta Primavera", 2.60, 16.00), ("Steak Frites", 7.40, 29.00), ("Fish Tacos", 3.30, 15.00),
]

# Waste log: (item, reason, cost) — total $380
WASTE = [
    ("Spoilage", "Produce past date", 140.00),
    ("Over-prep", "Batch surplus", 110.00),
    ("Expired", "Dairy", 80.00),
    ("Spill / drop", "Handling", 50.00),
]

# Categories: (category, actual food-cost $) — % of food cost computed
CATEGORIES = [
    ("Meat & seafood", 7200), ("Produce", 3900), ("Dairy & eggs", 2400),
    ("Dry goods", 2700), ("Beverages", 1800), ("Paper/disposables", 1200),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Business name", "The Harvest Table", None, "Business"),
        ("Owner", "Sam", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Variance limit %", VAR_LIMIT, "0%", "VarLimit"),
        ("Inventory-turns goal", TURN_GOAL, "0.0", "TurnGoal"),
        ("Waste limit %", WASTE_LIMIT, "0%", "WasteLimit"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Category", CATS, "CatList"),
             ("G", "Status", STATUS, "StatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  📋  FOOD COST & INVENTORY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know your food cost to the point — and stop the leaks.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE FOOD-COST PICTURE, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Food cost is the number that makes or breaks a kitchen — and most operators only guess at it. This "
                      "fixes that: a period food-cost engine that takes your beginning inventory, adds purchases, subtracts "
                      "ending inventory and divides by sales for a true food-cost %. Value your inventory to the dollar, "
                      "compare theoretical usage to actual to catch shrinkage, set par levels, track vendor prices, cost "
                      "your menu and log waste — all in ONE premium Google Sheets & Excel system. Everything connects: "
                      "your counts, purchases and sales flow straight into the food-cost %.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your target food-cost % & variance limit.",
             "2.  Count your inventory — the extended value is your ending inventory.",
             "3.  Log Purchases & Sales for the period — totals flow automatically.",
             "4.  The Food Cost Calc shows your true food-cost % to the point.",
             "5.  Check Usage & Variance for shrinkage; set Par & Ordering.",
             "6.  Check the Dashboard: food cost %, inventory value & a Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional restaurant (The Harvest Table) is included so you can see how it all connects "
               "— just type over it with your own counts and invoices. Food cost % and inventory variance are the two "
               "numbers that decide whether a kitchen makes money, and they roll into a live Inventory Score. Twelve "
               "matching printable pages (inventory count sheet, food-cost worksheet, order guide, waste log & more) are "
               "included. This is a business tool, not financial or legal advice — confirm figures with your own books.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Count it, cost it, control it — a point of food cost is real money.", "section_gold")


# ===========================================================================
def build_inventory(wb):
    ws = wb.create_sheet("Inventory Count"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 12, 14, 10, 14, 2])
    luxe_header(ws, "F", "📦  INVENTORY COUNT",
                "Count × unit cost = extended value. Add it up and that's your inventory, to the dollar.")
    table_headers(ws, 4, ["Item", "Count", "Unit Cost", "Unit", "Ext. Value"], start_col=2)
    start = L0
    for i, (item, count, cost, unit) in enumerate(INVENTORY):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=count); cc.style = "input"; cc.number_format = "#,##0"
        cu = ws.cell(row=r, column=4, value=cost); cu.style = "input"; cu.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5, value=unit).style = "td"
        ce = ws.cell(row=r, column=6, value=f"=C{r}*D{r}"); ce.style = "td"; ce.font = Font(bold=True, color=PRIMARY); ce.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVENTORY) - 1
    nrange(wb, "InvItem", "Inventory Count", "B", start, end)
    nrange(wb, "InvExt", "Inventory Count", "F", start, end)
    add_dv(ws, f"E{start}:E{end}", "UnitList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="INVENTORY VALUE").style = "th"
    for c in range(3, 6):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=6, value="=SUM(InvExt)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    cell_name(wb, "InvValue", "Inventory Count", f"$F${tot}")
    ws.freeze_panes = "A5"


def build_purchases(wb):
    ws = wb.create_sheet("Purchases Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 20, 14, 2])
    luxe_header(ws, "D", "🧾  PURCHASES LOG",
                "Every invoice in — by vendor & category. The total flows into your food cost.")
    table_headers(ws, 4, ["Vendor", "Category", "Amount"], start_col=2)
    start = L0
    for i, (vendor, cat, amt) in enumerate(PURCHASES):
        r = start + i
        ws.cell(row=r, column=2, value=vendor).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td_left"
        ca = ws.cell(row=r, column=4, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "CatList")
    end = start + len(PURCHASES) - 1
    nrange(wb, "PurchAmt", "Purchases Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="PURCHASES TOTAL").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(PurchAmt)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "PurchTotal", "Purchases Log", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_sales(wb):
    ws = wb.create_sheet("Sales Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 16, 2])
    luxe_header(ws, "C", "📈  SALES LOG",
                "Food sales for the period — the denominator of your food-cost %.")
    table_headers(ws, 4, ["Period", "Sales"], start_col=2)
    start = L0
    for i, (period, sales) in enumerate(SALES):
        r = start + i
        ws.cell(row=r, column=2, value=period).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SALES) - 1
    nrange(wb, "SalesAmt", "Sales Log", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="SALES TOTAL").style = "th"
    c = ws.cell(row=tot, column=3, value="=SUM(SalesAmt)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "SalesTotal", "Sales Log", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_foodcost(wb):
    ws = wb.create_sheet("Food Cost Calc"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "🧮  FOOD COST CALC",
                "Beginning + Purchases − Ending ÷ Sales = your true food-cost %. Everything flows in here.")
    table_headers(ws, 5, ["Line", "Amount"], start_col=2)
    rows = [
        ("Beginning inventory", BEGINNING, "input"),
        ("+ Purchases (from log)", "=PurchTotal", "calc"),
        ("− Ending inventory (from count)", "=InvValue", "calc"),
        ("= Food used (COGS)", "=C6+C7-C8", "bold"),
        ("÷ Food sales (from log)", "=SalesTotal", "calc"),
    ]
    r0 = 6
    for i, (lab, val, kind) in enumerate(rows):
        r = r0 + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=val)
        c.number_format = '"$"#,##0'
        if kind == "input":
            c.style = "input"
        elif kind == "bold":
            c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE)
        else:
            c.style = "td"
        if i % 2 and kind != "bold":
            for cc in range(2, 4):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    cell_name(wb, "Beginning", "Food Cost Calc", "$C$6")
    cell_name(wb, "FoodCostDollar", "Food Cost Calc", "$C$9")
    fcr = r0 + len(rows) + 1
    ws.cell(row=fcr, column=2, value="= FOOD COST %").style = "th"
    c = ws.cell(row=fcr, column=3, value="=IFERROR(FoodCostDollar/SalesTotal,0)"); c.style = "td"
    c.font = Font(bold=True, size=13, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = "0.0%"
    cell_name(wb, "FoodCostPct", "Food Cost Calc", f"$C${fcr}")
    # turns
    tr = fcr + 2
    ws.cell(row=tr, column=2, value="Avg inventory  ·  (Begin + End) ÷ 2").style = "field_label"
    ca = ws.cell(row=tr, column=3, value="=(Beginning+InvValue)/2"); ca.style = "field_value"; ca.number_format = '"$"#,##0'
    cell_name(wb, "AvgInv", "Food Cost Calc", f"$C${tr}")
    ws.cell(row=tr + 1, column=2, value="Inventory turns  ·  Food used ÷ Avg inventory").style = "field_label"
    ct = ws.cell(row=tr + 1, column=3, value="=IFERROR(FoodCostDollar/AvgInv,0)"); ct.style = "field_value"; ct.number_format = "0.00"; ct.fill = fill(MINT_BG)
    cell_name(wb, "Turns", "Food Cost Calc", f"$C${tr+1}")
    ws.cell(row=tr + 3, column=2, value="Change one count or invoice and your food-cost % updates instantly.").style = "section"


def build_usage(wb):
    ws = wb.create_sheet("Usage & Variance"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 16, 14, 2])
    luxe_header(ws, "E", "🔍  USAGE & VARIANCE",
                "Theoretical usage (what recipes say) vs actual (what you used) — the gap is shrinkage.")
    table_headers(ws, 4, ["Category", "Theoretical", "Actual", "Variance %"], start_col=2)
    start = L0
    for i, (cat, theo, actual) in enumerate(USAGE):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        ct = ws.cell(row=r, column=3, value=theo); ct.style = "input"; ct.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=actual); ca.style = "input"; ca.number_format = '"$"#,##0'
        cv = ws.cell(row=r, column=5, value=f"=IFERROR((D{r}-C{r})/C{r},0)"); cv.style = "td"; cv.number_format = "0.0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(USAGE) - 1; tot = end + 1
    nrange(wb, "CatName", "Usage & Variance", "B", start, end)
    nrange(wb, "CatActual", "Usage & Variance", "D", start, end)
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    ct = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    ca = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = '"$"#,##0'
    cv = ws.cell(row=tot, column=5, value=f"=IFERROR((D{tot}-C{tot})/C{tot},0)"); cv.style = "td"; cv.font = Font(bold=True, color=DANGER); cv.fill = fill(SURFACE); cv.number_format = "0.0%"
    cell_name(wb, "VariancePct", "Usage & Variance", f"$E${tot}")
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="greaterThan", formula=["0.05"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


def build_parorder(wb):
    ws = wb.create_sheet("Par & Ordering"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 10, 12, 12, 14, 2])
    luxe_header(ws, "F", "📥  PAR & ORDERING",
                "Par vs on hand × unit cost = what to order and what it costs — before you run out.")
    table_headers(ws, 4, ["Item", "Par", "On Hand", "Unit Cost", "Order $"], start_col=2)
    start = L0
    for i, (item, par, oh, cost) in enumerate(PARORDER):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=par); cp.style = "input"; cp.number_format = "#,##0"
        co = ws.cell(row=r, column=4, value=oh); co.style = "input"; co.number_format = "#,##0"
        cu = ws.cell(row=r, column=5, value=cost); cu.style = "input"; cu.number_format = '"$"#,##0.00'
        cb = ws.cell(row=r, column=6, value=f"=MAX(C{r}-D{r},0)*E{r}"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PARORDER) - 1
    nrange(wb, "ParItem", "Par & Ordering", "B", start, end)
    nrange(wb, "ParLevel", "Par & Ordering", "C", start, end)
    nrange(wb, "OrderVal", "Par & Ordering", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TO ORDER").style = "th"
    for c in range(3, 6):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=6, value="=SUM(OrderVal)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "ToOrderValue", "Par & Ordering", f"$F${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=[f"C{start}*0.5"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


def build_vendors(wb):
    ws, start, end = build_log(
        wb, "Vendors", "🤝", "VENDORS",
        "Who you buy from — category, contact & standing. Keep two per category.",
        ["Vendor", "Category", "Contact", "Status"],
        VENDORS, [2, 22, 20, 24, 14, 2], text_left={2, 3, 4}, reserved=24, start_col=2,
        validations=[("C", "CatList"), ("E", "StatusList")])
    nrange(wb, "VendorName", "Vendors", "B", start, end)
    cmap = {"Preferred": MINT_BG, "Backup": SURFACE, "Watch": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


def build_prices(wb):
    ws = wb.create_sheet("Price Tracker"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 18, 12, 18, 12, 14, 2])
    luxe_header(ws, "G", "🏷  PRICE TRACKER",
                "Compare two vendors on the same item — the cheaper one wins, and the gap adds up.")
    table_headers(ws, 4, ["Item", "Vendor A", "Price A", "Vendor B", "Price B", "Best"], start_col=2)
    start = L0
    for i, (item, va, pa, vb, pb) in enumerate(PRICES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ws.cell(row=r, column=3, value=va).style = "td_left"
        cpa = ws.cell(row=r, column=4, value=pa); cpa.style = "input"; cpa.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5, value=vb).style = "td_left"
        cpb = ws.cell(row=r, column=6, value=pb); cpb.style = "input"; cpb.number_format = '"$"#,##0.00'
        cbe = ws.cell(row=r, column=7, value=f'=IF(D{r}<=F{r},C{r},E{r})'); cbe.style = "td"; cbe.font = Font(bold=True, color=PRIMARY)
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


def build_menu(wb):
    ws = wb.create_sheet("Menu Costing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 12, 14, 2])
    luxe_header(ws, "F", "🍽  MENU COSTING",
                "Plate cost vs menu price — the food-cost % and margin on every dish.")
    table_headers(ws, 4, ["Item", "Plate Cost", "Menu Price", "Food %", "Margin"], start_col=2)
    start = L0
    for i, (item, cost, price) in enumerate(MENU):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cf = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cf.style = "td"; cf.number_format = "0%"
        cm = ws.cell(row=r, column=6, value=f"=D{r}-C{r}"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MENU) - 1
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=0.20, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.32, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.45, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Spoilage, over-prep & expired stock — the leaks that quietly raise your food cost.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_categories(wb):
    ws = wb.create_sheet("Categories"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 14, 2])
    luxe_header(ws, "D", "📊  CATEGORIES",
                "Where the food dollar goes — each category's spend and share of food cost.")
    table_headers(ws, 4, ["Category", "Food Cost $", "% of Total"], start_col=2)
    start = L0
    for i, (cat, amt) in enumerate(CATEGORIES):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/SUM(CatSpend),0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CATEGORIES) - 1
    nrange(wb, "CatLabel", "Categories", "B", start, end)
    nrange(wb, "CatSpend", "Categories", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    c = ws.cell(row=tot, column=3, value="=SUM(CatSpend)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  📋  FOOD COST & INVENTORY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Food cost %, inventory value, variance & a Score — your whole food-cost picture, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("FOOD COST %", "=FoodCostPct", "pct"),
        ("INVENTORY VALUE", "=InvValue", "money"),
        ("PURCHASES", "=PurchTotal", "money"),
        ("FOOD USED", "=FoodCostDollar", "money"),
        ("SALES", "=SalesTotal", "money"),
        ("TOP CATEGORY", "=INDEX(CatName,MATCH(MAX(CatActual),CatActual,0))", "text"),
    ]
    row2 = [
        ("VARIANCE", "=VariancePct", "pct"),
        ("ITEMS TRACKED", "=COUNTA(InvItem)", "num"),
        ("TO ORDER", "=ToOrderValue", "money"),
        ("INV TURNS", "=Turns", "dec"),
        ("VENDORS", "=COUNTA(VendorName)", "num"),
        ("INVENTORY SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "INVENTORY HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "FOOD COST BY CATEGORY", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Food cost on target", "=IFERROR(MIN(TargetFC/FoodCostPct,1),0)"),
        ("Inventory variance low", "=IFERROR(1-MIN(VariancePct/VarLimit,1),0)"),
        ("Inventory counted", "=IFERROR(COUNTIF(InvExt,\">0\")/COUNTA(InvItem),0)"),
        ("Pars set", "=IFERROR(COUNTIF(ParLevel,\">0\")/COUNTA(ParItem),0)"),
        ("Inventory turns healthy", "=IFERROR(MIN(Turns/TurnGoal,1),0)"),
        ("Gross margin", "=IFERROR(MIN((1-FoodCostPct)/0.7,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    cat = wb["Categories"]
    ch = BarChart(); ch.type = "bar"; ch.title = "Food Cost by Category"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(cat, min_col=3, min_row=5, max_row=4 + len(CATEGORIES)), titles_from_data=False)
    ch.set_categories(Reference(cat, min_col=2, min_row=5, max_row=4 + len(CATEGORIES))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Food Cost & Inventory Command Center™ — count it, cost it, control it.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_inventory(wb); build_purchases(wb)
    build_sales(wb); build_foodcost(wb); build_usage(wb); build_parorder(wb)
    build_vendors(wb); build_prices(wb); build_menu(wb); build_waste(wb)
    build_categories(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Food Cost Calc", "Inventory Count", "Purchases Log", "Sales Log",
             "Usage & Variance", "Par & Ordering", "Vendors", "Price Tracker", "Menu Costing", "Waste Log",
             "Categories", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Food_Cost_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
