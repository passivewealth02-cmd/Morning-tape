"""Build First-Time Home Buyer & Mortgage Command Center™ — The Home-Buying Operating System.

14 tabs · a premium first-time-home-buyer operating system in Google Sheets & Excel.
Dashboard, an affordability engine (price + down payment + rate → your true monthly PITI
and your debt-to-income), down-payment savings, closing costs, a home comparison scorer,
a lender/rate compare, amortization, life-after-buying budget, credit prep, a house
hunting log, a moving checklist and a monthly summary — one dashboard. Know what you can
truly afford, before an agent tells you.

Run: python3 build_xlsx.py   ->  ../Home_Buyer_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
LOANTYPE = ["Conventional", "FHA", "VA", "USDA"]
HOMESTATUS = ["Interested", "Toured", "Offer made", "Passed"]
PRIORITY = ["Must have", "Nice to have", "Deal breaker"]

# Affordability inputs
HOME_PRICE = 290000
DOWN_PCT = 0.08
RATE = 0.065
TERM_YEARS = 30
TAX_RATE = 0.011
INSURANCE_ANNUAL = 1800
PMI_RATE = 0.005
HOA_MONTHLY = 0
ANNUAL_INCOME = 95000
OTHER_DEBTS = 450
CLOSING_PCT = 0.03
SAVED = 32000
CREDIT_SCORE = 760
CREDIT_GOAL = 740
EF_AFTER = 12000
EF_GOAL = 12000
FRONT_GOAL = 0.28
BACK_GOAL = 0.36
DOWN_GOAL = 0.20

# Down payment savings: (month, deposit)
SAVINGS_LOG = [
    ("Feb", 1200), ("Mar", 1200), ("Apr", 1400), ("May", 1400), ("Jun", 1600), ("Jul", 1600),
]

# Closing costs: (item, amount)
CLOSING_ITEMS = [
    ("Loan origination", 2670), ("Appraisal", 650), ("Home inspection", 500),
    ("Title & escrow", 1900), ("Recording & transfer", 780), ("Prepaid taxes & insurance", 1500),
    ("Survey & misc", 700),
]

# Home comparison: (address, price, beds, baths, sqft, score)
HOMES = [
    ("12 Birch Lane", 289000, 3, 2, 1580, 92), ("48 Cedar Ct", 305000, 3, 2, 1720, 86),
    ("7 Maple Row", 275000, 3, 1.5, 1440, 78), ("91 Oak Ridge", 320000, 4, 2.5, 1980, 74),
    ("30 Alder St", 268000, 2, 2, 1290, 69),
]

# Lender / rate compare: (lender, rate, points, closing costs)
LENDERS = [
    ("Northbank", 0.065, 0.0, 8700), ("Cedar Credit Union", 0.0625, 0.5, 9400),
    ("Summit Mortgage", 0.0675, 0.0, 8100), ("Online Lender", 0.064, 0.25, 8900),
]

# Amortization snapshot: (year, remaining balance)
AMORT = [
    (1, 263700), (5, 249800), (10, 227400), (15, 197100), (20, 156100), (25, 100600), (30, 0),
]

# Budget after buying: (line, monthly)
AFTER_BUDGET = [
    ("Mortgage (PITI)", 2213), ("Utilities", 260), ("Maintenance fund", 240),
    ("Groceries", 600), ("Transportation", 420), ("Other debts", 450),
    ("Savings & investing", 700), ("Everything else", 800),
]

# Credit prep: (item, status)
CREDIT_PREP = [
    ("Pull all 3 reports", "Yes"), ("Dispute any errors", "Yes"), ("Pay cards under 30%", "Yes"),
    ("No new credit lines", "Yes"), ("Keep old cards open", "Yes"), ("Verify 2 yrs income docs", "No"),
]

# House hunting log: (date, address, status, notes)
HUNT_LOG = [
    ("Jun 8", "12 Birch Lane", "Offer made", "Best layout, great light"),
    ("Jun 8", "48 Cedar Ct", "Toured", "Bigger but busy road"),
    ("Jun 15", "7 Maple Row", "Toured", "Needs a new roof"),
    ("Jun 22", "91 Oak Ridge", "Passed", "Over budget"),
    ("Jun 29", "30 Alder St", "Interested", "Small, but cheapest"),
]

# Moving checklist: (task, weeks before)
MOVING = [
    ("Lock your rate", "6 wks"), ("Schedule inspection", "5 wks"), ("Book movers", "4 wks"),
    ("Transfer utilities", "2 wks"), ("Change address (USPS)", "2 wks"), ("Final walkthrough", "1 day"),
]

# Monthly summary: (month, total saved)
MONTHS = [("Feb", 24000), ("Mar", 25200), ("Apr", 26600), ("May", 28000), ("Jun", 30000), ("Jul", 32000)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 20, 3] + [18] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your assumptions & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR NUMBERS & GOALS", "section")
    controls = [
        ("Your name", "Avery", None, "Owner"),
        ("Plan name", "Keystone & Co.", None, "PlanName"),
        ("Annual household income", ANNUAL_INCOME, '"$"#,##0', "AnnualIncome"),
        ("Other monthly debts", OTHER_DEBTS, '"$"#,##0', "OtherDebts"),
        ("Property tax rate (of price)", TAX_RATE, "0.00%", "TaxRate"),
        ("Home insurance (annual)", INSURANCE_ANNUAL, '"$"#,##0', "InsuranceAnnual"),
        ("PMI rate (of loan)", PMI_RATE, "0.00%", "PMIRate"),
        ("Closing costs (% of price)", CLOSING_PCT, "0.0%", "ClosingPct"),
        ("Front-end DTI goal (max)", FRONT_GOAL, "0%", "FrontGoal"),
        ("Back-end DTI goal (max)", BACK_GOAL, "0%", "BackGoal"),
        ("Down-payment goal %", DOWN_GOAL, "0%", "DownGoal"),
        ("Credit-score goal", CREDIT_GOAL, "0", "CreditGoal"),
        ("Your credit score", CREDIT_SCORE, "0", "CreditScore"),
        ("Emergency fund after close", EF_AFTER, '"$"#,##0', "EFAfter"),
        ("Emergency fund goal", EF_GOAL, '"$"#,##0', "EFGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Loan type", LOANTYPE, "LoanTypeList"), ("F", "Home status", HOMESTATUS, "HomeStatusList"),
             ("G", "Priority", PRIORITY, "PriorityList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🔑  FIRST-TIME HOME BUYER & MORTGAGE COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know what you can truly afford — before an agent tells you.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE HOME PURCHASE, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A lender will tell you the biggest loan you can get. That is not the same as what you can "
                      "comfortably afford. This shows the difference: an affordability engine turns a home price, your "
                      "down payment and a rate into your true all-in monthly payment — principal, interest, taxes, "
                      "insurance and PMI — then checks it against the debt-to-income ratios lenders actually use. Save "
                      "your down payment, compare homes and lenders, and plan the life you'll live after closing — all "
                      "in ONE premium Google Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — enter your income, debts and goals.",
             "2.  Run the Affordability tab — see your true monthly PITI.",
             "3.  Check your DTI: is this payment actually comfortable?",
             "4.  Track your Down Payment savings and Closing Costs.",
             "5.  Score homes and compare lenders side by side.",
             "6.  Check the Dashboard: payment, cash to close & a Buyer Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional buyer (Keystone & Co., Avery) is included so you can see how it all "
               "connects — just type over it with your own numbers. Your all-in monthly payment and your debt-to-income "
               "ratio are the two numbers that decide whether a house is a home or a trap, and they roll into a live "
               "Buyer Score. Note the sample buyer puts 8% down, so PMI applies — the score is honest about that. "
               "Twelve matching printable pages (affordability worksheet, home comparison, closing-cost checklist & "
               "more) are included. This is a planning & organizing tool, not financial, lending, legal or real-estate "
               "advice — confirm every figure with your own lender and advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Buy the payment, not the price. The monthly number is the one you live with.", "section_gold")


# ===========================================================================
def build_affordability(wb):
    ws = wb.create_sheet("Affordability"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 18, 2])
    luxe_header(ws, "C", "🧮  AFFORDABILITY",
                "Price + down payment + rate → your true all-in monthly payment, and the DTI lenders actually check.")
    ws.cell(row=5, column=2, value="THE PURCHASE").style = "section_gold"
    rows = [
        ("Home price", HOME_PRICE, '"$"#,##0', "HomePrice", SURFACE),
        ("Down payment %", DOWN_PCT, "0%", "DownPct", SURFACE),
        ("Interest rate", RATE, "0.00%", "Rate", SURFACE),
        ("Term (years)", TERM_YEARS, "0", "TermYears", SURFACE),
    ]
    for i, (lab, val, fmt, nm, bg) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"; c.number_format = fmt
        cell_name(wb, nm, "Affordability", f"$C${r}")
    ws.cell(row=10, column=2, value="Down payment $").style = "field_label"
    cd = ws.cell(row=10, column=3, value="=HomePrice*DownPct"); cd.style = "field_value"; cd.number_format = '"$"#,##0'
    cell_name(wb, "DownPayment", "Affordability", "$C$10")
    ws.cell(row=11, column=2, value="= LOAN AMOUNT").style = "th"
    cl = ws.cell(row=11, column=3, value="=HomePrice-HomePrice*DownPct"); cl.style = "td"; cl.font = Font(bold=True, size=12, color=PRIMARY); cl.fill = fill(MINT_BG); cl.number_format = '"$"#,##0'
    cell_name(wb, "LoanAmount", "Affordability", "$C$11")
    # payment build-up
    ws.cell(row=13, column=2, value="THE MONTHLY PAYMENT (PITI)").style = "section_gold"
    ws.cell(row=14, column=2, value="Principal & interest").style = "field_label"
    cpi = ws.cell(row=14, column=3, value="=-PMT(Rate/12,TermYears*12,LoanAmount)"); cpi.style = "field_value"; cpi.number_format = '"$"#,##0'
    cell_name(wb, "PrincipalInterest", "Affordability", "$C$14")
    ws.cell(row=15, column=2, value="+ Property tax").style = "field_label"
    ct = ws.cell(row=15, column=3, value="=HomePrice*TaxRate/12"); ct.style = "field_value"; ct.number_format = '"$"#,##0'
    cell_name(wb, "PropertyTax", "Affordability", "$C$15")
    ws.cell(row=16, column=2, value="+ Home insurance").style = "field_label"
    ci = ws.cell(row=16, column=3, value="=InsuranceAnnual/12"); ci.style = "field_value"; ci.number_format = '"$"#,##0'
    cell_name(wb, "Insurance", "Affordability", "$C$16")
    ws.cell(row=17, column=2, value="+ PMI (if under 20% down)").style = "field_label"
    cp = ws.cell(row=17, column=3, value="=IF(DownPct<0.2,LoanAmount*PMIRate/12,0)"); cp.style = "field_value"; cp.number_format = '"$"#,##0'
    cell_name(wb, "PMI", "Affordability", "$C$17")
    ws.cell(row=18, column=2, value="+ HOA dues").style = "field_label"
    ch = ws.cell(row=18, column=3, value=HOA_MONTHLY); ch.style = "input"; ch.number_format = '"$"#,##0'
    cell_name(wb, "HOA", "Affordability", "$C$18")
    ws.cell(row=19, column=2, value="= YOUR MONTHLY PAYMENT").style = "th"
    cpiti = ws.cell(row=19, column=3, value="=PrincipalInterest+PropertyTax+Insurance+PMI+HOA")
    cpiti.style = "td"; cpiti.font = Font(bold=True, size=14, color=PRIMARY); cpiti.fill = fill(MINT_BG); cpiti.number_format = '"$"#,##0'
    cell_name(wb, "PITI", "Affordability", "$C$19")
    # DTI
    ws.cell(row=21, column=2, value="CAN YOU AFFORD IT?").style = "section_gold"
    ws.cell(row=22, column=2, value="Gross monthly income").style = "field_label"
    cg = ws.cell(row=22, column=3, value="=AnnualIncome/12"); cg.style = "field_value"; cg.number_format = '"$"#,##0'
    cell_name(wb, "GrossMonthly", "Affordability", "$C$22")
    ws.cell(row=23, column=2, value="= FRONT-END DTI (housing only)").style = "th"
    cf = ws.cell(row=23, column=3, value="=IFERROR(PITI/GrossMonthly,0)"); cf.style = "td"; cf.font = Font(bold=True, size=12, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = "0.0%"
    cell_name(wb, "FrontDTI", "Affordability", "$C$23")
    ws.cell(row=24, column=2, value="= BACK-END DTI (all debts)").style = "th"
    cb = ws.cell(row=24, column=3, value="=IFERROR((PITI+OtherDebts)/GrossMonthly,0)"); cb.style = "td"; cb.font = Font(bold=True, size=12, color=PRIMARY); cb.fill = fill(MINT_BG); cb.number_format = "0.0%"
    cell_name(wb, "BackDTI", "Affordability", "$C$24")
    ws.cell(row=25, column=2, value="Verdict").style = "field_label"
    cv = ws.cell(row=25, column=3, value='=IF(AND(FrontDTI<=FrontGoal,BackDTI<=BackGoal),"COMFORTABLE","STRETCHED")')
    cv.style = "field_value"
    ws.cell(row=27, column=2, value="Total interest paid over the life of the loan").style = "field_label"
    cti = ws.cell(row=27, column=3, value="=PrincipalInterest*TermYears*12-LoanAmount"); cti.style = "field_value"; cti.number_format = '"$"#,##0'; cti.fill = fill(WARN_BG)
    cell_name(wb, "TotalInterest", "Affordability", "$C$27")
    ws.cell(row=29, column=2, value="Lenders approve you for more than you should spend. Trust the DTI, not the approval.").style = "section_gold"


def build_downpayment(wb):
    ws = wb.create_sheet("Down Payment"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 18, 2])
    luxe_header(ws, "D", "🏦  DOWN PAYMENT",
                "What you've saved against what you need — the number that decides when you can actually buy.")
    ws.cell(row=5, column=2, value="Saved so far").style = "field_label"
    cs = ws.cell(row=5, column=3, value=SAVED); cs.style = "input"; cs.number_format = '"$"#,##0'
    cell_name(wb, "Saved", "Down Payment", "$C$5")
    ws.cell(row=6, column=2, value="Down payment needed").style = "field_label"
    cd = ws.cell(row=6, column=3, value="=DownPayment"); cd.style = "field_value"; cd.number_format = '"$"#,##0'
    ws.cell(row=7, column=2, value="+ Closing costs").style = "field_label"
    cc = ws.cell(row=7, column=3, value="=ClosingTotal"); cc.style = "field_value"; cc.number_format = '"$"#,##0'
    ws.cell(row=8, column=2, value="= CASH TO CLOSE").style = "th"
    ct = ws.cell(row=8, column=3, value="=DownPayment+ClosingTotal"); ct.style = "td"; ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0'
    cell_name(wb, "CashToClose", "Down Payment", "$C$8")
    ws.cell(row=9, column=2, value="= STILL TO SAVE").style = "th"
    cst = ws.cell(row=9, column=3, value="=MAX(CashToClose-Saved,0)"); cst.style = "td"; cst.font = Font(bold=True, size=12, color=DANGER); cst.fill = fill(WARN_BG); cst.number_format = '"$"#,##0'
    ws.cell(row=10, column=2, value="Progress to cash to close").style = "field_label"
    cpr = ws.cell(row=10, column=3, value="=IFERROR(MIN(Saved/CashToClose,1),0)"); cpr.style = "field_value"; cpr.number_format = "0%"; cpr.fill = fill(MINT_BG)
    cell_name(wb, "SaveProgress", "Down Payment", "$C$10")
    table_headers(ws, 12, ["Month", "Deposit", "Running Total"], start_col=2)
    start = 13
    running = 24000 - sum(x[1] for x in SAVINGS_LOG)
    for i, (m, dep) in enumerate(SAVINGS_LOG):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cdp = ws.cell(row=r, column=3, value=dep); cdp.style = "input"; cdp.number_format = '"$"#,##0'
        if i == 0:
            crt = ws.cell(row=r, column=4, value=f"=Saved-SUM(C{start}:C{start+len(SAVINGS_LOG)-1})+C{r}")
        else:
            crt = ws.cell(row=r, column=4, value=f"=D{r-1}+C{r}")
        crt.style = "td"; crt.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SAVINGS_LOG) - 1
    nrange(wb, "Deposits", "Down Payment", "C", start, end)
    ws.cell(row=end + 2, column=2, value="Avg monthly deposit").style = "field_label"
    ca = ws.cell(row=end + 2, column=3, value="=IFERROR(AVERAGE(Deposits),0)"); ca.style = "field_value"; ca.number_format = '"$"#,##0'
    ws.cell(row=end + 3, column=2, value="Months to fully funded").style = "field_label"
    cm = ws.cell(row=end + 3, column=3, value="=IFERROR(MAX(CEILING((CashToClose-Saved)/AVERAGE(Deposits),1),0),0)"); cm.style = "field_value"; cm.number_format = "#,##0"
    ws.freeze_panes = "A13"


def build_closing(wb):
    ws = wb.create_sheet("Closing Costs"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "📄  CLOSING COSTS",
                "The thousands nobody warns first-time buyers about — itemized, so there are no surprises at the table.")
    table_headers(ws, 4, ["Item", "Amount"], start_col=2)
    start = L0
    for i, (item, amt) in enumerate(CLOSING_ITEMS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CLOSING_ITEMS) - 1
    nrange(wb, "ClosingAmt", "Closing Costs", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL CLOSING COSTS").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(ClosingAmt)"); ct.style = "td"; ct.font = Font(bold=True, size=12, color=DANGER); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "ClosingTotal", "Closing Costs", f"$C${tot}")
    ws.cell(row=tot + 2, column=2, value="As % of home price").style = "field_label"
    cp = ws.cell(row=tot + 2, column=3, value="=IFERROR(ClosingTotal/HomePrice,0)"); cp.style = "field_value"; cp.number_format = "0.0%"; cp.fill = fill(MINT_BG)
    ws.cell(row=tot + 4, column=2, value="Ask every lender for a Loan Estimate — closing costs are negotiable.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_homes(wb):
    ws = wb.create_sheet("Home Comparison"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 14, 10, 10, 12, 12, 14, 2])
    luxe_header(ws, "H", "🏡  HOME COMPARISON",
                "Score every house side by side — price, size, cost per square foot and your own rating.")
    table_headers(ws, 4, ["Address", "Price", "Beds", "Baths", "Sq Ft", "$ / Sq Ft", "Score"], start_col=2)
    start = L0
    for i, (addr, price, beds, baths, sqft, score) in enumerate(HOMES):
        r = start + i
        ws.cell(row=r, column=2, value=addr).style = "td_left"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=beds).style = "input"
        ws.cell(row=r, column=5, value=baths).style = "input"
        csq = ws.cell(row=r, column=6, value=sqft); csq.style = "input"; csq.number_format = "#,##0"
        cps = ws.cell(row=r, column=7, value=f"=IFERROR(C{r}/F{r},0)"); cps.style = "td"; cps.number_format = '"$"#,##0'
        csc = ws.cell(row=r, column=8, value=score); csc.style = "input"; csc.number_format = "#,##0"
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(HOMES) - 1
    nrange(wb, "HomeScore", "Home Comparison", "H", start, end)
    nrange(wb, "HomePriceRange", "Home Comparison", "C", start, end)
    ws.conditional_formatting.add(f"H{start}:H{end}",
        ColorScaleRule(start_type="num", start_value=60, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=80, mid_color="FFFFF3CD",
                       end_type="num", end_value=95, end_color="FF" + HIGHLIGHT))
    ws.cell(row=end + 2, column=2, value="Homes toured").style = "field_label"
    ch = ws.cell(row=end + 2, column=3, value="=COUNTA(HomeScore)"); ch.style = "field_value"; ch.number_format = "#,##0"
    cell_name(wb, "HomesSeen", "Home Comparison", f"$C${end+2}")
    ws.cell(row=end + 3, column=2, value="Best score").style = "field_label"
    cb = ws.cell(row=end + 3, column=3, value="=IFERROR(MAX(HomeScore),0)"); cb.style = "field_value"; cb.number_format = "#,##0"; cb.fill = fill(MINT_BG)
    ws.freeze_panes = "A5"


def build_lenders(wb):
    ws = wb.create_sheet("Lender Compare"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 12, 16, 16, 2])
    luxe_header(ws, "F", "🏛  LENDER COMPARE",
                "The same loan from four lenders — a quarter point is thousands of dollars. Shop it.")
    table_headers(ws, 4, ["Lender", "Rate", "Points", "Closing Costs", "Monthly P&I"], start_col=2)
    start = L0
    for i, (lender, rate, points, closing) in enumerate(LENDERS):
        r = start + i
        ws.cell(row=r, column=2, value=lender).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rate); cr.style = "input"; cr.number_format = "0.000%"
        cpt = ws.cell(row=r, column=4, value=points); cpt.style = "input"; cpt.number_format = "0.00"
        cc = ws.cell(row=r, column=5, value=closing); cc.style = "input"; cc.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=6, value=f"=-PMT(C{r}/12,TermYears*12,LoanAmount)"); cm.style = "td"; cm.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LENDERS) - 1
    nrange(wb, "LenderRate", "Lender Compare", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", DataBarRule(start_type="min", end_type="max", color=GOLD_LT))
    ws.cell(row=end + 2, column=2, value="Best rate offered").style = "field_label"
    cbr = ws.cell(row=end + 2, column=3, value="=IFERROR(MIN(LenderRate),0)"); cbr.style = "field_value"; cbr.number_format = "0.000%"; cbr.fill = fill(MINT_BG)
    ws.cell(row=end + 3, column=2, value="Lenders compared").style = "field_label"
    clc = ws.cell(row=end + 3, column=3, value="=COUNTA(LenderRate)"); clc.style = "field_value"; clc.number_format = "#,##0"
    ws.freeze_panes = "A5"


def build_amort(wb):
    ws = wb.create_sheet("Amortization"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 20, 20, 2])
    luxe_header(ws, "D", "📉  AMORTIZATION",
                "Where your loan balance lands at each milestone — and how much of it is still interest.")
    ws.cell(row=5, column=2, value="Total interest over the loan").style = "field_label"
    cti = ws.cell(row=5, column=4, value="=TotalInterest"); cti.style = "field_value"; cti.number_format = '"$"#,##0'; cti.fill = fill(WARN_BG)
    table_headers(ws, 7, ["Year", "Balance", "Equity Built"], start_col=2)
    start = 8
    for i, (year, bal) in enumerate(AMORT):
        r = start + i
        ws.cell(row=r, column=2, value=year).style = "td"
        cb = ws.cell(row=r, column=3, value=bal); cb.style = "input"; cb.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=4, value=f"=LoanAmount-C{r}"); ce.style = "td"; ce.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(AMORT) - 1
    ws.cell(row=end + 2, column=2, value="Every extra $100 a month knocks years off this table.").style = "section_gold"
    ws.freeze_panes = "A8"


def build_afterbudget(wb):
    ws = wb.create_sheet("Life After Buying"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "🏠  LIFE AFTER BUYING",
                "The budget you'll actually live on once the keys are yours — including the costs renters never see.")
    table_headers(ws, 4, ["Line", "Monthly"], start_col=2)
    start = L0
    for i, (line, amt) in enumerate(AFTER_BUDGET):
        r = start + i
        ws.cell(row=r, column=2, value=line).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(AFTER_BUDGET) - 1
    nrange(wb, "AfterAmt", "Life After Buying", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL MONTHLY").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(AfterAmt)"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    ws.cell(row=tot + 1, column=2, value="Left over each month").style = "field_label"
    cl = ws.cell(row=tot + 1, column=3, value="=GrossMonthly-SUM(AfterAmt)"); cl.style = "field_value"; cl.number_format = '"$"#,##0'; cl.fill = fill(MINT_BG)
    ws.cell(row=tot + 3, column=2, value="Budget 1% of the home price a year for maintenance — it always comes.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_credit(wb):
    ws = wb.create_sheet("Credit Prep"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 2])
    luxe_header(ws, "C", "📊  CREDIT PREP",
                "The moves that raise your score before you apply — every 20 points can cut your rate.")
    ws.cell(row=5, column=2, value="Your credit score").style = "field_label"
    cs = ws.cell(row=5, column=3, value="=CreditScore"); cs.style = "field_value"; cs.number_format = "#,##0"; cs.fill = fill(MINT_BG)
    ws.cell(row=6, column=2, value="Score goal").style = "field_label"
    cg = ws.cell(row=6, column=3, value="=CreditGoal"); cg.style = "field_value"; cg.number_format = "#,##0"
    table_headers(ws, 8, ["Action", "Done?"], start_col=2)
    start = 9
    for i, (item, status) in enumerate(CREDIT_PREP):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ws.cell(row=r, column=3, value=status).style = "td"
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CREDIT_PREP) - 1
    nrange(wb, "CreditDone", "Credit Prep", "C", start, end)
    add_dv(ws, f"C{start}:C{end}", "YesNoList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="DONE").style = "th"
    ct = ws.cell(row=tot, column=3, value='=COUNTIF(CreditDone,"Yes")&" / "&COUNTA(CreditDone)'); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(MINT_BG)
    ws.freeze_panes = "A9"


def build_hunt(wb):
    ws, start, end = build_log(
        wb, "House Hunting", "🔍", "HOUSE HUNTING LOG",
        "Every showing, in order — what you saw, what you thought, and where it went.",
        ["Date", "Address", "Status", "Notes"],
        HUNT_LOG, [2, 14, 22, 16, 30, 2], text_left={2, 3, 5}, reserved=28, start_col=2,
        validations=[("D", "HomeStatusList")])


def build_moving(wb):
    ws, start, end = build_log(
        wb, "Moving Checklist", "📦", "MOVING CHECKLIST",
        "Everything to do between the offer and the first night in your new home.",
        ["Task", "When", "Done?"],
        [(t, w, "No") for t, w in MOVING], [2, 30, 16, 14, 2], text_left={2}, reserved=26, start_col=2,
        validations=[("D", "YesNoList")])


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "Your savings climbing toward cash-to-close — the line that ends with keys in your hand.")
    ws.cell(row=5, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 6, ["Month", "Total Saved"], start_col=2)
    ts = 7
    for i, (m, amt) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "SavedTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Saved by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🔑  FIRST-TIME HOME BUYER & MORTGAGE COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Payment, DTI, cash to close & a Buyer Score — your whole purchase, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("HOME PRICE", "=HomePrice", "money"),
        ("DOWN PAYMENT", "=DownPayment", "money"),
        ("LOAN AMOUNT", "=LoanAmount", "money"),
        ("PRINCIPAL & INT", "=PrincipalInterest", "money"),
        ("MONTHLY PAYMENT", "=PITI", "money"),
        ("FRONT-END DTI", "=FrontDTI", "pct1"),
    ]
    row2 = [
        ("BACK-END DTI", "=BackDTI", "pct1"),
        ("CLOSING COSTS", "=ClosingTotal", "money"),
        ("CASH TO CLOSE", "=CashToClose", "money"),
        ("SAVED", "=Saved", "money"),
        ("TOTAL INTEREST", "=TotalInterest", "money"),
        ("BUYER SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "BUYER HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SAVED BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Cash to close saved", "=IFERROR(MIN(Saved/CashToClose,1),0)"),
        ("Payment affordable", "=IFERROR(MIN(FrontGoal/FrontDTI,1),0)"),
        ("Total debt healthy", "=IFERROR(MIN(BackGoal/BackDTI,1),0)"),
        ("Credit ready", "=IFERROR(MIN(CreditScore/CreditGoal,1),0)"),
        ("Emergency fund after close", "=IFERROR(MIN(EFAfter/EFGoal,1),0)"),
        ("20% down (no PMI)", "=IFERROR(MIN(DownPct/DownGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Saved by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=7, max_row=6 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=7, max_row=6 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "First-Time Home Buyer & Mortgage Command Center™ — buy the payment, not the price.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_closing(wb); build_affordability(wb)
    build_downpayment(wb); build_homes(wb); build_lenders(wb); build_amort(wb)
    build_afterbudget(wb); build_credit(wb); build_hunt(wb); build_moving(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Affordability", "Down Payment", "Closing Costs", "Home Comparison",
             "Lender Compare", "Amortization", "Life After Buying", "Credit Prep", "House Hunting",
             "Moving Checklist", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Home_Buyer_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
