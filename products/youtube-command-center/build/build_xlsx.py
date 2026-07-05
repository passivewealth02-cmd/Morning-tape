"""Build YouTube Command Center™ — The Ultimate YouTube Business Operating System.

30 sheets + Welcome · a premium YouTube business OS in Excel & Sheets.
Ideas, calendar, production, SEO, analytics, shorts, live, sponsors, finance,
equipment, repurposing, brand kit, AI prompts, goals & more — one dashboard.

Run: python3 build_xlsx.py   ->  ../YouTube_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

PILLARS = ["Systems", "Money", "Growth", "Tools", "Mindset", "Behind-the-scenes"]
CONTENT_STATUS = ["Idea", "Scripting", "Filming", "Editing", "Review", "Scheduled", "Published", "Repurposed"]
VIDEO_TYPES = ["Long-form", "Short", "Live", "Podcast", "Tutorial", "Vlog"]
REV_CATS = ["AdSense", "Sponsors", "Affiliate", "Courses", "Digital Products", "Memberships", "Consulting", "Merchandise"]
EXP_CATS = ["Software", "Equipment", "Contractors", "Marketing", "Office", "Education", "Travel", "Miscellaneous"]
SPON_STAGES = ["Lead", "Outreach", "Negotiation", "Signed", "Delivered", "Paid"]
CAMPAIGN_TYPES = ["Sponsored", "Affiliate", "Product", "Organic", "Series", "Collab"]
GOAL_CATS = ["Subscribers", "Watch Hours", "Revenue", "Uploads", "Sponsors", "Products", "Engagement"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]
CURRENCIES = ["USD", "EUR", "GBP", "CAD", "AUD"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


def totals(ws, row, cols, start, end, fmt='"$"#,##0', label="TOTAL"):
    ws.cell(row=row, column=1, value=label).style = "th"
    for col in cols:
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt


# ===========================================================================
# 30 — Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your channel details once — every dashboard follows.")
    merge_set(ws, "B5:C5", "CHANNEL INPUTS", "section")
    controls = [
        ("Channel Name", "Vale Studio", None, "ChannelName"),
        ("Creator", "Jordan Vale", None, "CreatorName"),
        ("Niche", "Creator business & tech", None, "Niche"),
        ("Currency", "USD", None, "HomeCurr"),
        ("Monthly Revenue Goal", 12000, '"$"#,##0', "RevenueGoal"),
        ("Monthly Upload Goal", 7, "0", "UploadGoal"),
        ("Active Deal Target", 6, "0", "DealTarget"),
        ("Avg View Dur Target (min)", 6, "0", "AVDTarget"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Content Pillar", PILLARS, "PillarList"), ("F", "Video Type", VIDEO_TYPES, "VideoTypeList"),
             ("G", "Status", CONTENT_STATUS, "StatusList"), ("H", "Revenue Category", REV_CATS, "RevCatList"),
             ("I", "Expense Category", EXP_CATS, "ExpCatList"), ("J", "Campaign Type", CAMPAIGN_TYPES, "CampaignList"),
             ("K", "Goal Category", GOAL_CATS, "GoalCatList"), ("L", "Sponsor Stage", SPON_STAGES, "SponStageList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 16, "Priority", PRIORITIES, "PriorityList"), ("F", 16, "Currency", CURRENCIES, "CurrencyList"),
             ("G", 16, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ▶  YOUTUBE COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate YouTube business operating system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "RUN YOUR ENTIRE YOUTUBE BUSINESS FROM ONE FILE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("From idea to income — YouTube Command Center™ manages ideation, production, publishing, SEO, "
                      "analytics, sponsorships, finances and long-term growth in ONE premium Excel & Google Sheets "
                      "system. Stay consistent, publish higher-quality videos, track your growth, land brand deals, "
                      "grow revenue and scale a real business — all with academy-level automation. This isn't a "
                      "content calendar. It's your YouTube business OS.")
    ws["B6"].style = "body"
    for r in (6, 7, 8, 9):
        ws.row_dimensions[r].height = 22
    merge_set(ws, "B11:B11", "START HERE", "section")
    steps = ["1.  Open Settings and add your channel name, niche & monthly goals.",
             "2.  Fill in the Channel Profile & Brand Kit (pillars, colors, links).",
             "3.  Capture ideas in the Idea Vault, then move them through the Pipeline.",
             "4.  Plan uploads in the Content Calendar; optimize with the SEO Center.",
             "5.  Log analytics, sponsors, affiliates & finances — profit updates live.",
             "6.  Watch the Executive Dashboard track growth, revenue & channel health."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (an 84k-subscriber channel earning $9,840/mo) is included so you can see how "
               "everything connects — just type over it with your own. Revenue, net profit, subscriber growth, "
               "upload consistency, sponsorship pipeline and the Channel Health Score all update automatically. "
               "Every sheet is print-friendly and works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more time to create — let's grow your channel.", "section_gold")


# ===========================================================================
# 2 — Channel Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Channel Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "📺  CHANNEL PROFILE", "Your channel, defined — the identity every video flows from.")
    blocks = [
        ("THE CHANNEL", [("Channel Name", "=ChannelName"), ("Creator", "=CreatorName"),
                         ("Niche", "=Niche"), ("Upload Schedule", "Tue & Fri, 9am ET"),
                         ("Founded", "2021"), ("Business Email", "hi@valestudio.co")]),
        ("STRATEGY", [("Mission", "Turn creators into CEOs"), ("Target Audience", "Aspiring creators 24–38"),
                      ("Content Pillars", "Systems · Money · Growth"), ("Manager", "Self-managed"),
                      ("Tax Entity", "Vale Studio LLC"), ("Website", "valestudio.co")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "SOCIAL LINKS", "section_gold"); ws.row_dimensions[15].height = 22
    handles = [("YouTube", "@valestudio"), ("Instagram", "@vale.studio"), ("TikTok", "@valestudio"),
               ("X", "@valestudio"), ("Newsletter", "The Vale Note"), ("LinkedIn", "Jordan Vale")]
    for i, (p, h) in enumerate(handles):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 3 — Content Master Calendar
# ===========================================================================
def build_calendar(wb):
    rows = [
        (-18, "5 Systems That 10x Your Output", "Systems", "Long-form", "Published"),
        (-14, "The $0 editing workflow", "Tools", "Long-form", "Published"),
        (-11, "How I plan a month in 90 min", "Systems", "Long-form", "Published"),
        (-8, "I quit my job. Here's the math", "Money", "Long-form", "Published"),
        (-4, "3 tools I can't live without", "Tools", "Short", "Published"),
        (-2, "Reply: how much I actually make", "Money", "Short", "Published"),
        (1, "My full content OS (2024)", "Systems", "Long-form", "Scheduled"),
        (2, "Launch week: behind the build", "Behind-the-scenes", "Short", "Scheduled"),
        (5, "Sponsor integration: Riverside", "Tools", "Long-form", "Scheduled"),
        (8, "Faceless channel playbook", "Growth", "Long-form", "Scheduled"),
        (3, "Repurpose 1 video into 10", "Systems", "Long-form", "Editing"),
        (6, "AI tools tier list", "Tools", "Long-form", "Filming"),
        (10, "How creators actually get sponsors", "Money", "Long-form", "Scripting"),
        (12, "Studio tour + gear", "Behind-the-scenes", "Long-form", "Idea"),
    ]
    sample = []
    for doff, title, pillar, vtype, status in rows:
        d = dplus(doff) if doff >= 0 else dminus(-doff)
        sample.append((d, title, pillar, vtype, "High" if vtype == "Long-form" else "Medium", status))
    ws, start, end = build_log(
        wb, "Calendar", "🗓", "CONTENT MASTER CALENDAR",
        "Plan every upload in one view — publishing status calculates itself.",
        ["Publish Date", "Title", "Pillar", "Type", "Priority", "Status"],
        sample, [14, 36, 16, 13, 11, 14],
        text_left={2}, dates={1},
        validations=[("C", "PillarList"), ("D", "VideoTypeList"), ("E", "PriorityList"), ("F", "StatusList")], reserved=60)
    nrange(wb, "CalDate", "Calendar", "A", start, end)
    nrange(wb, "CalStatus", "Calendar", "F", start, end)
    cmap = {"Published": MINT_BG, "Scheduled": WARN_BG, "Editing": SOFT_BG, "Filming": SOFT_BG, "Idea": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 4 — Video Production Pipeline
# ===========================================================================
def build_pipeline(wb):
    rows = [
        ("My full content OS (2024)", "Editing", 0.82, "Jordan", 1, "Final cut"),
        ("Sponsor integration: Riverside", "Script", 0.40, "Jordan", 5, "Brief approved"),
        ("Repurpose 1 video into 10", "Editing", 0.70, "Editor — Jon", 3, "3 of 10 cut"),
        ("AI tools tier list", "Filming", 0.55, "Jordan", 6, "Recorded, editing"),
        ("Faceless channel playbook", "Outline", 0.30, "Jordan", 8, "Keyword map done"),
        ("How creators get sponsors", "Research", 0.20, "Writer — Mia", 10, "Interviews booked"),
        ("Launch week short", "Thumbnail", 0.90, "Designer", 2, "A/B thumbs ready"),
        ("Studio tour + gear", "Idea", 0.10, "Jordan", 12, "Shot list"),
    ]
    sample = [(t, st, prog, o, dplus(d), note) for (t, st, prog, o, d, note) in rows]
    ws, start, end = build_log(
        wb, "Pipeline", "🎬", "VIDEO PRODUCTION PIPELINE",
        "Every video, every stage — from idea to published, with live progress %.",
        ["Title", "Stage", "Progress", "Owner", "Due Date", "Notes"],
        sample, [30, 14, 12, 16, 13, 22],
        text_left={1, 4, 6}, dates={5}, pcts={3}, reserved=40)
    nrange(wb, "PipeTitle", "Pipeline", "A", start, end)
    nrange(wb, "PipeProgress", "Pipeline", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 5 — Idea Vault
# ===========================================================================
def build_ideas(wb):
    rows = [
        ("Creator tax masterclass", "Money", "youtube taxes", "The tax bill that shocked me", 120000, 8, 6, "Planned"),
        ("Notion vs Sheets for creators", "Tools", "notion vs sheets", "I switched back. Here's why", 90000, 7, 4, "Idea"),
        ("30-day upload challenge", "Growth", "posting consistency", "Day 1 of daily uploads", 60000, 6, 3, "Idea"),
        ("How I price digital products", "Money", "digital product pricing", "$9 vs $49: what sold more", 140000, 9, 2, "Planned"),
        ("$2k home studio tour", "Behind-the-scenes", "creator studio setup", "My entire setup", 80000, 6, 5, "Idea"),
        ("Faceless automation playbook", "Growth", "faceless youtube", "No face, 100k subs", 180000, 9, 4, "Idea"),
        ("Sponsor email templates", "Money", "sponsorship outreach", "The email that landed $3k", 110000, 8, 2, "Planned"),
        ("Batch a month in a weekend", "Systems", "batch content", "12 videos in 2 days", 130000, 8, 5, "Idea"),
        ("AI tools I actually use", "Tools", "ai for youtube", "3 tools that save hours", 95000, 7, 2, "Idea"),
        ("Best camera for YouTube 2024", "Tools", "best youtube camera", "Don't buy the wrong one", 160000, 8, 5, "Idea"),
    ]
    ws = wb.create_sheet("Ideas"); ws.sheet_view.showGridLines = False
    set_widths(ws, [28, 15, 20, 28, 14, 9, 11, 12, 12])
    luxe_header(ws, "I", "💡  IDEA VAULT",
                "Never run dry — capture ideas and let an AI Opportunity Score (views ÷ difficulty) rank them.")
    table_headers(ws, 4, ["Idea", "Category", "Keyword", "Hook", "Est. Views", "Diff.", "AI Score", "Priority", "Status"])
    start = L0
    for i, (idea, cat, kw, hook, views, diff, prio_score, status) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=idea)
        ws.cell(row=r, column=2, value=cat)
        ws.cell(row=r, column=3, value=kw)
        ws.cell(row=r, column=4, value=hook)
        ws.cell(row=r, column=5, value=views)
        ws.cell(row=r, column=6, value=diff)
        ws.cell(row=r, column=7, value=f"=IFERROR(ROUND(E{r}/1000/F{r},0),0)")
        ws.cell(row=r, column=8, value="High" if prio_score >= 8 else ("Medium" if prio_score >= 6 else "Low"))
        ws.cell(row=r, column=9, value=status)
    end = start + 40 - 1
    style_rows(ws, start, end, 9, text_left={1, 3, 4}, ints={5, 6, 7})
    for col_letter, lst in [("B", "PillarList"), ("H", "PriorityList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"G{start}:G{end}",
        ColorScaleRule(start_type="num", start_value=5, start_color="FF" + WARN_BG,
                       end_type="num", end_value=30, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 6 — Script Manager
# ===========================================================================
def build_scripts(wb):
    rows = [
        ("My full content OS (2024)", "You're wasting 10 hrs/week…", "Problem→System→Proof→CTA", "Yes", "Mid-roll", 14, "Final", 1.0),
        ("Sponsor integration: Riverside", "Recording just got easier…", "Hook→Value→Sponsor→CTA", "Yes", "60s @ 5:00", 12, "Draft", 0.5),
        ("Faceless channel playbook", "You don't need to show your face…", "Myth→Method→Results→CTA", "Yes", "Mid-roll", 16, "Outline", 0.3),
        ("AI tools tier list", "I tested 30 AI tools so you don't…", "List→Tiers→Winners→CTA", "No", "None", 11, "Draft", 0.6),
        ("How creators get sponsors", "The email that landed $3k…", "Story→Template→Steps→CTA", "No", "None", 13, "Idea", 0.15),
    ]
    ws, start, end = build_log(
        wb, "Scripts", "📝", "SCRIPT MANAGER",
        "Every script, structured — hook, story beats, CTA & sponsor placement with completion %.",
        ["Video", "Opening Hook", "Structure", "Sponsor?", "Sponsor Slot", "Est. Min", "Status", "Complete"],
        rows, [26, 26, 24, 11, 14, 10, 12, 11],
        text_left={1, 2, 3}, ints={6}, pcts={8},
        validations=[("D", "YesNoList")], reserved=30)
    ws.conditional_formatting.add(f"H{start}:H{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 7 — Thumbnail Studio
# ===========================================================================
def build_thumbnails(wb):
    ws = wb.create_sheet("Thumbnails"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 22, 22, 3, 18, 14, 2])
    luxe_header(ws, "H", "🖼  THUMBNAIL STUDIO",
                "Win the click — A/B three thumbnails per video, paste them here and track CTR.")
    videos = [("My full content OS", "8.4%", "V2"), ("Faceless channel playbook", "—", "testing"),
              ("$0 editing workflow", "7.1%", "V1"), ("How I plan a month", "6.8%", "V3")]
    top0 = 5
    for idx, (title, ctr, winner) in enumerate(videos):
        row = top0 + idx * 8
        merge_set(ws, f"B{row}:G{row}", f"  🎬  {title}", "th"); ws.row_dimensions[row].height = 24
        for j, ver in enumerate(["V1", "V2", "V3"]):
            col = 2 + j
            L = get_column_letter(col)
            merge_set(ws, f"{L}{row+1}:{L}{row+4}", f"📷\n{ver}\nPaste thumb", "imgbox")
            for rr in range(row + 1, row + 5):
                ws.row_dimensions[rr].height = 20
        ws.cell(row=row + 5, column=2, value="Winning CTR").style = "field_label"
        cc = ws.cell(row=row + 5, column=3, value=ctr); cc.style = "field_value"; cc.fill = fill(MINT_BG)
        ws.cell(row=row + 5, column=5, value="Winner").style = "field_label"
        ws.cell(row=row + 5, column=6, value=winner).style = "field_value"


# ===========================================================================
# 8 — SEO Command Center
# ===========================================================================
def build_seo(wb):
    rows = [
        ("My full content OS (2024)", "content system", "productivity, workflow", "How I Built a Content OS", 92, "Yes", "Yes", 0.95),
        ("Faceless channel playbook", "faceless youtube", "automation, passive", "Faceless YouTube Playbook", 88, "Yes", "Yes", 0.90),
        ("$0 editing workflow", "free video editing", "capcut, davinci", "The $0 Editing Workflow", 84, "Yes", "No", 0.75),
        ("AI tools tier list", "ai tools for youtube", "chatgpt, ai video", "AI Tools Tier List 2024", 90, "Yes", "Yes", 0.85),
        ("How creators get sponsors", "youtube sponsorships", "brand deals, outreach", "How Creators Get Sponsors", 86, "No", "No", 0.55),
        ("Best camera 2024", "best youtube camera", "sony, canon, gear", "Best Camera for YouTube 2024", 80, "No", "No", 0.40),
    ]
    ws, start, end = build_log(
        wb, "SEO", "🔍", "SEO COMMAND CENTER",
        "Get found — primary keyword, title/description scores, chapters & an SEO completion %.",
        ["Video", "Primary Keyword", "Tags", "Optimized Title", "Title Score", "Chapters", "End Screens", "SEO %"],
        rows, [26, 18, 20, 28, 12, 11, 12, 10],
        text_left={1, 2, 3, 4}, ints={5}, pcts={8},
        validations=[("F", "YesNoList"), ("G", "YesNoList")], reserved=40)
    ws.conditional_formatting.add(f"H{start}:H{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=60, start_color="FF" + WARN_BG,
                       end_type="num", end_value=100, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 9 — Analytics Command Center
# ===========================================================================
def build_analytics(wb):
    # per-video analytics (recent)
    rows = [
        ("I quit my job. Here's the math", 148000, 0.084, "6:42", 2100, 6.20),
        ("5 Systems That 10x Output", 112000, 0.072, "7:15", 1580, 8.40),
        ("The $0 editing workflow", 96000, 0.069, "5:48", 1120, 7.10),
        ("How I plan a month in 90 min", 84000, 0.078, "6:05", 1580, 9.20),
        ("3 tools I can't live without", 58000, 0.061, "0:42", 720, 3.10),
        ("Reply: how much I make", 42000, 0.055, "0:38", 640, 2.80),
    ]
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 18, 3, 16, 12, 12, 2])
    luxe_header(ws, "H", "📊  ANALYTICS COMMAND CENTER",
                "Your channel by the numbers — a live snapshot, health dimensions & the Channel Health Score.")
    # snapshot block
    merge_set(ws, "B5:C5", "CHANNEL SNAPSHOT (28 DAYS)", "section")
    snap = [("Subscribers", 84200, "#,##0", "SubNow"), ("Views", 412000, "#,##0", "Views28"),
            ("Watch hours", 14600, "#,##0", "WatchHrs"), ("Avg CTR", 0.068, "0.0%", "AvgCTR"),
            ("Avg view duration", "4:52", "General", "AvgViewDur"), ("Returning viewers", 0.38, "0%", "Returning"),
            ("RPM", 8.40, '"$"#,##0.00', "RPM"), ("CPM", 14.20, '"$"#,##0.00', "CPM")]
    for i, (lab, val, fmt, nm) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "field_value"; c.number_format = fmt
        cell_name(wb, nm, "Analytics", f"$C${r}")
        if lab in ("Subscribers", "RPM"):
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    # health dims
    merge_set(ws, "E5:H5", "CHANNEL HEALTH", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=5)
    metrics = [
        ("Revenue vs goal", "=IFERROR(MIN(RevenueTotal/RevenueGoal,1),0)"),
        ("Upload consistency", '=IFERROR(MIN(COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Published")/UploadGoal,1),0)'),
        ("Audience / CTR", "=IFERROR(MIN(AvgCTR/0.08,1),0)"),
        ("Retention", "=IFERROR(MIN((LEFT(AvgViewDur,FIND(\":\",AvgViewDur)-1)+MID(AvgViewDur,FIND(\":\",AvgViewDur)+1,2)/60)/AVDTarget,1),0)"),
        ("Sponsorship pipeline", '=IFERROR(MIN((COUNTIF(SponStage,"Signed")+COUNTIF(SponStage,"Delivered")+COUNTIF(SponStage,"Negotiation"))/DealTarget,1),0)'),
        ("Goal progress", "=IFERROR(AVERAGE(GoalProgress),0)"),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(metrics):
        r = hs + i
        ws.cell(row=r, column=5, value=dim).style = "td_left"
        c = ws.cell(row=r, column=6, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=7, value=f'=IF(F{r}>=0.75,"Strong",IF(F{r}>=0.5,"Growing","Focus"))').style = "td"
        if i % 2:
            for c2 in range(5, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(metrics) - 1
    cell_name(wb, "HealthRange", "Analytics", f"$F${hs}:$F${he}")
    ws.conditional_formatting.add(f"F{hs}:F{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # per-video table
    merge_set(ws, "B15:H15", "RECENT VIDEO PERFORMANCE", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Video", "Views", "CTR", "Avg Duration", "Subs +", "RPM"], start_col=2)
    vs = 17
    for i, (title, views, ctr, dur, subs, rpm) in enumerate(rows):
        r = vs + i
        ws.cell(row=r, column=2, value=title).style = "td_left"
        cv = ws.cell(row=r, column=3, value=views); cv.style = "td"; cv.number_format = "#,##0"
        cc = ws.cell(row=r, column=4, value=ctr); cc.style = "td"; cc.number_format = "0.0%"
        ws.cell(row=r, column=5, value=dur).style = "td"
        cs = ws.cell(row=r, column=6, value=subs); cs.style = "td"; cs.number_format = "#,##0"
        cr = ws.cell(row=r, column=7, value=rpm); cr.style = "td"; cr.number_format = '"$"#,##0.00'
        if i % 2:
            for c2 in range(2, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    ve = vs + len(rows) - 1
    nrange(wb, "VidTitle", "Analytics", "B", vs, ve)
    nrange(wb, "VidViews", "Analytics", "C", vs, ve)
    ws.conditional_formatting.add(f"C{vs}:C{ve}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=150000, color=PRIMARY, showValue=True))
    # subscriber trend (feeds dashboard line)
    merge_set(ws, "B25:C25", "SUBSCRIBER GROWTH (K) — 6 MONTHS", "section")
    ws.cell(row=26, column=2, value="Month").style = "th"; ws.cell(row=26, column=3, value="Subs (K)").style = "th"
    months = _recent_months(6); vals = [58.4, 64.1, 69.8, 74.5, 79.6, 84.2]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 27 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = "0.0"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "SubMonth", "Analytics", "$B$27:$B$32")
    cell_name(wb, "SubVal", "Analytics", "$C$27:$C$32")


# ===========================================================================
# 10 — Long Form Analytics
# ===========================================================================
def build_longform(wb):
    ws = wb.create_sheet("Long Form"); ws.sheet_view.showGridLines = False
    set_widths(ws, [30, 14, 13, 13, 14, 12, 12])
    luxe_header(ws, "G", "🎞  LONG-FORM VIDEO ANALYTICS",
                "How your videos age — first 24 hours vs 7, 30 days & lifetime.")
    table_headers(ws, 4, ["Video", "First 24h", "7 Days", "30 Days", "Lifetime", "CTR", "Avg Dur"])
    rows = [
        ("I quit my job. Here's the math", 22000, 68000, 121000, 148000, "8.4%", "6:42"),
        ("5 Systems That 10x Output", 18000, 54000, 92000, 112000, "7.2%", "7:15"),
        ("The $0 editing workflow", 14000, 46000, 78000, 96000, "6.9%", "5:48"),
        ("How I plan a month in 90 min", 12000, 41000, 70000, 84000, "7.8%", "6:05"),
        ("Notion setup for creators", 9000, 31000, 54000, 68000, "6.4%", "8:10"),
    ]
    start = L0
    for i, (t, h24, d7, d30, life, ctr, dur) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([t, h24, d7, d30, life, ctr, dur], 1):
            ws.cell(row=r, column=ci, value=v)
    end = start + 30 - 1
    style_rows(ws, start, end, 7, text_left={1}, ints={2, 3, 4, 5})
    ws.freeze_panes = "A5"
    for col in (2, 3, 4, 5):
        L = get_column_letter(col)
        ws.conditional_formatting.add(f"{L}{start}:{L}{end}",
            DataBarRule(start_type="num", start_value=0, end_type="num", end_value=150000, color=PRIMARY, showValue=True))


# ===========================================================================
# 11 — Shorts Dashboard
# ===========================================================================
def build_shorts(wb):
    rows = [
        (2, "Reply: how much I make", 420000, 0.78, 1820, "0:38"),
        (5, "3 tools I can't live without", 260000, 0.71, 1120, "0:42"),
        (9, "POV: launch day", 180000, 0.66, 640, "0:31"),
        (13, "1 tip that 2x'd my views", 320000, 0.74, 1450, "0:29"),
        (18, "Editing hack", 145000, 0.69, 520, "0:35"),
        (24, "Don't buy this camera", 98000, 0.62, 310, "0:44"),
    ]
    sample = [(dminus(d), t, v, cr, subs, wt) for (d, t, v, cr, subs, wt) in rows]
    ws, start, end = build_log(
        wb, "Shorts", "📱", "SHORTS DASHBOARD",
        "Shorts are your top-of-funnel — views, completion rate & subs gained per Short.",
        ["Date", "Short", "Views", "Completion", "Subs +", "Length"],
        sample, [12, 28, 13, 13, 12, 11],
        text_left={2}, dates={1}, ints={3, 5}, pcts={4}, reserved=40)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=420000, color=HIGHLIGHT, showValue=True))
    ws.conditional_formatting.add(f"D{start}:D{end}",
        ColorScaleRule(start_type="num", start_value=0.5, start_color="FF" + WARN_BG,
                       end_type="num", end_value=0.8, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 12 — Community Tab Planner
# ===========================================================================
def build_community(wb):
    rows = [
        (-3, "Poll", "Which video next? (3 options)", "Published", "1.2k votes"),
        (-1, "Image", "BTS of today's shoot", "Published", "High engagement"),
        (2, "Question", "What's your #1 editing struggle?", "Scheduled", "For content ideas"),
        (4, "Announcement", "New video Friday", "Scheduled", "Drive views"),
        (7, "GIF", "Reaction to hitting 84k", "Idea", "Celebrate milestone"),
        (10, "Poll", "Long-form vs Shorts this week?", "Idea", ""),
    ]
    sample = [(dplus(d) if d >= 0 else dminus(-d), typ, content, status, note) for (d, typ, content, status, note) in rows]
    ws, start, end = build_log(
        wb, "Community", "💬", "COMMUNITY TAB PLANNER",
        "Keep the algorithm warm between uploads — plan polls, images & questions.",
        ["Date", "Type", "Content", "Status", "Notes"],
        sample, [13, 16, 34, 14, 22],
        text_left={3, 5}, dates={1}, reserved=30)
    for st, cc in {"Published": MINT_BG, "Scheduled": WARN_BG, "Idea": WHITE}.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 13 — Live Stream Command Center
# ===========================================================================
def build_live(wb):
    rows = [
        ("Q&A: scaling to full-time", "Solo", -14, 4200, 320, 180, "Great turnout"),
        ("Building a content OS live", "Guest — Mia", -30, 3100, 210, 95, "Collab boost"),
        ("Launch day livestream", "Solo", 2, 0, 0, 0, "Upcoming"),
        ("Year in review + AMA", "Solo", 40, 0, 0, 0, "Planned"),
    ]
    sample = [(t, g, dplus(d) if d >= 0 else dminus(-d), rv, sc, mem, note) for (t, g, d, rv, sc, mem, note) in rows]
    ws, start, end = build_log(
        wb, "Live", "🔴", "LIVE STREAM COMMAND CENTER",
        "Run great streams — topic, guests, super chats, memberships & replay views.",
        ["Topic", "Guests", "Date", "Replay Views", "Super Chats $", "Memberships", "Notes"],
        sample, [26, 16, 13, 14, 14, 13, 20],
        text_left={1, 2, 7}, dates={3}, ints={4}, money={5}, reserved=20)


# ===========================================================================
# 14 — Playlist Manager
# ===========================================================================
def build_playlists(wb):
    rows = [
        ("Creator Systems", 12, 640000, 18200, "Yes", "High"),
        ("Make Money on YouTube", 9, 480000, 14100, "Yes", "High"),
        ("Editing & Tools", 14, 390000, 9800, "No", "Medium"),
        ("Faceless Channels", 6, 220000, 7400, "Yes", "High"),
        ("Behind the Scenes", 8, 110000, 3200, "No", "Low"),
    ]
    ws = wb.create_sheet("Playlists"); ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 10, 14, 14, 14, 12])
    luxe_header(ws, "F", "🎵  PLAYLIST MANAGER",
                "Playlists drive session time — track videos, views, watch time & optimization.")
    table_headers(ws, 4, ["Playlist", "Videos", "Views", "Watch Hours", "Optimized", "Priority"])
    start = L0
    for i, (nm, vids, views, wh, opt, prio) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([nm, vids, views, wh, opt, prio], 1):
            ws.cell(row=r, column=ci, value=v)
    end = start + 30 - 1
    style_rows(ws, start, end, 6, text_left={1}, ints={2, 3, 4})
    for col_letter, lst in [("E", "YesNoList"), ("F", "PriorityList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    totals(ws, end + 1, [3, 4], start, end, fmt="#,##0")
    ws.cell(row=end + 1, column=2, value=f"=SUM(B{start}:B{end})").number_format = "#,##0"
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=700000, color=PRIMARY, showValue=True))


# ===========================================================================
# 15 — Sponsorship CRM
# ===========================================================================
def build_sponsorship(wb):
    rows = [
        ("Riverside", "Dana K.", "Q3 integration", 3200, "Signed", "Sent", 12),
        ("Skillshare", "Marco P.", "Course promo", 2600, "Negotiation", "Not Sent", 30),
        ("Notion", "Lena R.", "Template collab", 2200, "Delivered", "Sent", -5),
        ("Squarespace", "Ivy T.", "Website build", 1800, "Paid", "Paid", -25),
        ("Adobe", "Priya S.", "Express feature", 2400, "Outreach", "Not Sent", 60),
        ("NordVPN", "Chris N.", "VPN spot", 1500, "Lead", "Not Sent", 75),
        ("HelloFresh", "Owen B.", "Meal kit", 1100, "Paid", "Paid", -40),
        ("Epidemic Sound", "Sam W.", "Music partner", 900, "Signed", "Sent", 45),
    ]
    sample = [(b, c, camp, rate, stage, inv, dplus(d) if d >= 0 else dminus(-d)) for (b, c, camp, rate, stage, inv, d) in rows]
    ws, start, end = build_log(
        wb, "Sponsors", "🤝", "SPONSORSHIP CRM",
        "Turn brand deals into a pipeline — lead to paid, with rates & invoice status.",
        ["Brand", "Contact", "Campaign", "Rate", "Stage", "Invoice", "Due / Paid"],
        sample, [16, 14, 20, 12, 14, 12, 13],
        text_left={2, 3}, money={4}, dates={7},
        validations=[("E", "SponStageList")], reserved=30)
    nrange(wb, "SponRate", "Sponsors", "D", start, end)
    nrange(wb, "SponStage", "Sponsors", "E", start, end)
    for st, cc in {"Paid": MINT_BG, "Signed": WARN_BG, "Delivered": SOFT_BG, "Lead": WHITE}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 16 — Affiliate Dashboard
# ===========================================================================
def build_affiliate(wb):
    rows = [
        ("Amazon", "Gear picks", 0.04, 4200, 128, 340, -10),
        ("Notion", "Template affiliate", 0.50, 1100, 58, 290, 15),
        ("Riverside", "Recording tool", 0.30, 680, 22, 180, 5),
        ("Epidemic Sound", "Music referrals", 0.25, 540, 26, 110, -8),
        ("ConvertKit", "Email tool", 0.30, 320, 12, 60, 20),
    ]
    ws = wb.create_sheet("Affiliate"); ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 20, 13, 12, 12, 13, 13])
    luxe_header(ws, "G", "🔗  AFFILIATE DASHBOARD",
                "Passive income, tracked — clicks, conversions & commission by program.")
    table_headers(ws, 4, ["Program", "Product", "Commission", "Clicks", "Sales", "Revenue", "Payout"])
    start = L0
    for i, (prog, prod, comm, clicks, sales, rev, doff) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([prog, prod, comm, clicks, sales, rev], 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=7, value=dplus(doff))
    end = start + 30 - 1
    style_rows(ws, start, end, 7, text_left={1, 2}, pcts={3}, ints={4, 5}, money={6}, dates={7})
    totals(ws, end + 1, [6], start, end)
    ws.freeze_panes = "A5"
    nrange(wb, "AffRevenue", "Affiliate", "F", start, end)


# ===========================================================================
# 17 — Digital Product Dashboard
# ===========================================================================
def build_products(wb):
    rows = [
        ("Content OS (Notion)", -120, 49, 186, 12, 4.9),
        ("Faceless Channel Course", -60, 149, 84, 4, 4.8),
        ("Thumbnail Pack", -200, 19, 240, 15, 4.7),
        ("Sponsorship Templates", -40, 24, 96, 3, 4.9),
        ("Creator Tax Kit", -80, 29, 142, 6, 4.8),
    ]
    ws = wb.create_sheet("Products"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 14, 10, 12, 13, 11, 13, 12])
    luxe_header(ws, "H", "📦  DIGITAL PRODUCT DASHBOARD",
                "Products that sell while you sleep — units, revenue, refunds & profit per product.")
    table_headers(ws, 4, ["Product", "Launched", "Price", "Units", "Revenue", "Refunds", "Net Profit", "Rating"])
    start = L0
    for i, (nm, dago, price, units, refunds, rating) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=nm)
        ws.cell(row=r, column=2, value=dminus(-dago) if dago < 0 else dplus(dago))
        ws.cell(row=r, column=3, value=price)
        ws.cell(row=r, column=4, value=units)
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
        ws.cell(row=r, column=6, value=refunds)
        ws.cell(row=r, column=7, value=f"=E{r}-C{r}*F{r}")
        ws.cell(row=r, column=8, value=rating)
    end = start + 30 - 1
    style_rows(ws, start, end, 8, text_left={1}, dates={2}, money={3, 5, 7}, ints={4, 6}, dec={8})
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in (4, 5, 7):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE)
        c.number_format = "#,##0" if col == 4 else '"$"#,##0'
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=13000, color=PRIMARY, showValue=True))


# ===========================================================================
# 18 — Business Finance Center
# ===========================================================================
def build_finance(wb):
    ws = wb.create_sheet("Finance"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 16, 12, 3, 22, 14])
    luxe_header(ws, "G", "💰  BUSINESS FINANCE CENTER",
                "Every income & expense in one place — monthly, run-rate & net profit, live.")
    # income
    table_headers(ws, 4, ["Income Source", "This Month", "Annual (est.)", "% of Rev"])
    income = {"AdSense": 3850, "Sponsors": 3200, "Affiliate": 980, "Courses": 1100,
              "Digital Products": 410, "Memberships": 220, "Consulting": 0, "Merchandise": 80}
    start = L0; iend = start + len(REV_CATS) - 1
    for i, cat in enumerate(REV_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cm = ws.cell(row=r, column=2, value=income[cat]); cm.style = "input"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/$B${iend+1},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    itot = iend + 1
    ws.cell(row=itot, column=1, value="TOTAL REVENUE").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=itot, column=col, value=f"=SUM({L}{start}:{L}{iend})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.conditional_formatting.add(f"D{start}:D{iend}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=0.45, color=PRIMARY, showValue=True))
    nrange(wb, "RevSource", "Finance", "A", start, iend)
    nrange(wb, "RevMonthly", "Finance", "B", start, iend)
    cell_name(wb, "RevenueTotal", "Finance", f"$B${itot}")
    cell_name(wb, "RevAdsense", "Finance", "$B$5")
    cell_name(wb, "RevSponsorCell", "Finance", "$B$6")
    # expenses (right)
    merge_set(ws, "F4:G4", "MONTHLY EXPENSES", "section_gold")
    expenses = {"Software": 260, "Equipment": 680, "Contractors": 1000, "Marketing": 180,
                "Office": 90, "Education": 99, "Travel": 0, "Miscellaneous": 101}
    estart = 6
    for i, cat in enumerate(EXP_CATS):
        r = estart + i
        ws.cell(row=r, column=6, value=cat).style = "td_left"
        c = ws.cell(row=r, column=7, value=expenses[cat]); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=6).fill = fill(MUTED_ROW); ws.cell(row=r, column=7).fill = fill(MUTED_ROW)
    eend = estart + len(EXP_CATS) - 1; etot = eend + 1
    ws.cell(row=etot, column=6, value="TOTAL EXPENSES").style = "th"
    c = ws.cell(row=etot, column=7, value=f"=SUM(G{estart}:G{eend})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "ExpCat", "Finance", "F", estart, eend)
    nrange(wb, "ExpMonthly", "Finance", "G", estart, eend)
    cell_name(wb, "ExpenseTotal", "Finance", f"$G${etot}")
    # bottom line
    merge_set(ws, "A16:D16", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Revenue (month)", "=RevenueTotal", '"$"#,##0'), ("Annual run-rate", "=RevenueTotal*12", '"$"#,##0'),
             ("Expenses (month)", "=ExpenseTotal", '"$"#,##0'), ("Net profit", "=RevenueTotal-ExpenseTotal", '"$"#,##0'),
             ("Profit margin", "=IFERROR((RevenueTotal-ExpenseTotal)/RevenueTotal,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 17 + i
        ws.cell(row=r, column=1, value=lab).style = "field_label"
        c = ws.cell(row=r, column=2, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net profit", "Profit margin"):
            ws.cell(row=r, column=2).fill = fill(MINT_BG)


# ===========================================================================
# 19 — Equipment Manager
# ===========================================================================
def build_equipment(wb):
    ws = wb.create_sheet("Equipment"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 13, 18, 12, 11, 14, 13, 13, 16])
    luxe_header(ws, "I", "🎥  EQUIPMENT MANAGER",
                "Gear you rely on — condition, warranty & automatic replacement reminders.")
    table_headers(ws, 4, ["Item", "Type", "Detail", "Purchased", "Cost", "Warranty", "Replace By", "Status", "Notes"])
    rows = [
        ("Main Camera", "Camera", "Sony ZV-E1", 400, 2200, "Yes", 400, "Primary"),
        ("Backup Camera", "Camera", "Sony a6700", 220, 1400, "Yes", 500, "B-cam + BTS"),
        ("Prime Lens", "Lens", "Sigma 24mm f1.4", 30, 680, "Yes", 900, "New pickup"),
        ("Key Light", "Lighting", "Aputure 120D", 300, 400, "No", 250, "Workhorse"),
        ("Shotgun Mic", "Microphone", "Rode NTG5", 200, 500, "No", 40, "Replace foam"),
        ("Editing PC", "Computer", "M3 Max", 500, 3200, "Yes", 700, "Fast exports"),
        ("SSD Storage", "Storage", "8TB RAID", 150, 600, "No", 35, "Filling up"),
        ("Tripod", "Accessory", "Manfrotto", 600, 220, "No", 200, "Reliable"),
    ]
    start = L0
    for i, (item, ty, det, pdago, cost, war, rdoff) in enumerate([(r[0], r[1], r[2], r[3], r[4], r[5], r[6]) for r in rows]):
        r = start + i
        note = rows[i][7]
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=ty)
        ws.cell(row=r, column=3, value=det)
        ws.cell(row=r, column=4, value=dminus(pdago))
        ws.cell(row=r, column=5, value=cost)
        ws.cell(row=r, column=6, value=war)
        ws.cell(row=r, column=7, value=dplus(rdoff))
        ws.cell(row=r, column=8, value=f'=IF(G{r}="","",IF(G{r}<=TODAY()+45,"DUE SOON","OK"))')
        ws.cell(row=r, column=9, value=note)
    end = start + 30 - 1
    style_rows(ws, start, end, 9, text_left={1, 3, 9}, dates={4, 7}, money={5})
    add_dv(ws, f"F{start}:F{end}", "YesNoList")
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL INVESTED").style = "th"
    c = ws.cell(row=total, column=5, value=f"=SUM(E{start}:E{end})")
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "EquipName", "Equipment", "A", start, end)
    nrange(wb, "EquipReplace", "Equipment", "G", start, end)
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"DUE SOON"'], fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=fill(MINT_BG)))


# ===========================================================================
# 20 — Content Repurposing Center
# ===========================================================================
def build_repurpose(wb):
    ws = wb.create_sheet("Repurposing"); ws.sheet_view.showGridLines = False
    formats = ["Shorts", "Reels", "TikTok", "Pins", "Newsletter", "Blog", "LinkedIn", "X", "Podcast", "Threads"]
    set_widths(ws, [30] + [10] * len(formats) + [12])
    luxe_header(ws, get_column_letter(1 + len(formats) + 1), "♻  CONTENT REPURPOSING CENTER",
                "One video → ten platforms — turn every upload into a week of content.")
    table_headers(ws, 4, ["Source Video"] + formats + ["Done"])
    sources = [
        ("I quit my job. Here's the math", [1, 1, 1, 0, 1, 1, 1, 1, 1, 1]),
        ("5 Systems That 10x Output", [1, 1, 1, 1, 1, 1, 1, 1, 0, 0]),
        ("The $0 editing workflow", [1, 1, 1, 0, 0, 0, 1, 1, 0, 1]),
        ("How I plan a month", [1, 0, 1, 1, 1, 1, 0, 0, 1, 0]),
        ("My full content OS (2024)", [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]),
    ]
    start = L0
    for i, (title, flags) in enumerate(sources):
        r = start + i
        ws.cell(row=r, column=1, value=title).style = "td_left"
        for j, f in enumerate(flags):
            c = ws.cell(row=r, column=2 + j, value="✓" if f else "○")
            c.style = "td"; c.font = Font(bold=True, color=PRIMARY if f else BORDER)
            c.fill = fill(MINT_BG if f else WHITE); c.border = BOX
        done = sum(flags)
        cd = ws.cell(row=r, column=2 + len(formats), value=f"={done}/{len(formats)}")
        cd.style = "td"; cd.number_format = "0%"; cd.font = Font(bold=True, color=PRIMARY)
    end = start + len(sources) - 1
    ws.freeze_panes = "B5"
    ws.conditional_formatting.add(f"{get_column_letter(2+len(formats))}{start}:{get_column_letter(2+len(formats))}{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 21 — Brand Kit
# ===========================================================================
def build_brandkit(wb):
    ws = wb.create_sheet("Brand Kit"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "🎨  BRAND KIT", "One source of truth for your channel's look, sound & voice.")
    blocks = [
        ("VISUAL", [("Logo", "Brand/Logos/vale.svg"), ("Fonts", "Sora + Inter"),
                    ("Thumbnail Style", "Bold text · face · 1 idea"), ("Intro", "3s animated wordmark"),
                    ("Outro", "20s end screen template"), ("LUTs", "Vale-Warm.cube")]),
        ("SOUND & VOICE", [("Intro Music", "Epidemic — 'Momentum'"), ("Background Music", "Lo-fi playlist"),
                           ("Brand Voice", "Warm, sharp, encouraging"), ("Tagline", "Turn creators into CEOs"),
                           ("Animations", "Lower-thirds pack v3"), ("Sound Effects", "SFX/Vale-pack")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "BRAND COLORS", "section_gold"); ws.row_dimensions[15].height = 22
    colors = [("Primary", PRIMARY), ("Accent", ACCENT), ("Surface", SURFACE), ("Highlight", HIGHLIGHT)]
    for i, (nm, hexv) in enumerate(colors):
        c = 2 + i
        ws.cell(row=16, column=c, value=nm).style = "field_label"
        sw = ws.cell(row=17, column=c, value=f"#{hexv}"); sw.fill = fill(hexv)
        sw.font = Font(bold=True, color="FFFFFF" if nm in ("Primary", "Accent") else PRIMARY)
        sw.alignment = Alignment(horizontal="center", vertical="center"); sw.border = BOX
        ws.row_dimensions[17].height = 26


# ===========================================================================
# 22 — Asset Library
# ===========================================================================
def build_assets(wb):
    rows = [
        ("B-roll — studio", "B-roll", "Drive/B-roll/Studio", "Yes", "4K, 60 clips"),
        ("B-roll — city", "B-roll", "Drive/B-roll/City", "Yes", "Establishing shots"),
        ("Music license — Epidemic", "Music", "Licenses/Epidemic", "Yes", "Active subscription"),
        ("SFX pack", "Sound FX", "SFX/Vale-pack", "Yes", "200 effects"),
        ("Motion graphics pack", "Motion", "Drive/Motion", "Yes", "Lower-thirds v3"),
        ("Thumbnail template (PSD)", "Template", "Templates/Thumbs", "Yes", "3 layouts"),
        ("Stock footage — Storyblocks", "Stock", "Licenses/Storyblocks", "Yes", "Unlimited plan"),
        ("Intro / outro stingers", "Motion", "Drive/Motion/Stingers", "Yes", "5s each"),
    ]
    ws, start, end = build_log(
        wb, "Assets", "🗂", "ASSET LIBRARY",
        "Find any file in seconds — a searchable index of every asset & where it lives.",
        ["Asset", "Type", "Storage Location", "Backed Up", "Notes"],
        rows, [26, 14, 24, 12, 24],
        text_left={1, 3, 5}, validations=[("D", "YesNoList")], reserved=40)
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(RED_BG)))


# ===========================================================================
# 23 — AI Prompt Library
# ===========================================================================
def build_prompts(wb):
    ws = wb.create_sheet("AI Prompts"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 78, 2])
    luxe_header(ws, "D", "🤖  AI PROMPT LIBRARY",
                "Your best prompts, reusable — ideas, hooks, titles, scripts, SEO & outreach.")
    prompts = [
        ("Video Ideas", "Act as a YouTube strategist for a {niche} channel. Give me 10 video ideas with a proven "
         "format (listicle, tutorial, story) and a 1-line hook for each. Prioritize search demand + shareability."),
        ("Hooks", "Write 8 opening hooks (first 10 seconds) for a video titled \"{title}\". Use curiosity, stakes "
         "or a bold claim. Keep each under 2 sentences, spoken aloud."),
        ("Titles", "Generate 12 high-CTR YouTube titles for \"{topic}\". Mix curiosity, numbers, and outcome. "
         "Keep under 60 characters. Flag the top 3 you'd test."),
        ("Descriptions", "Write an SEO description for \"{title}\": 2 keyword-rich sentences, then a value summary, "
         "timestamps placeholder, links section and 3 hashtags."),
        ("Scripts", "Outline a {length}-minute script for \"{title}\" using Hook → Problem → Steps → Proof → CTA. "
         "Note where to place a 60s mid-roll sponsor read."),
        ("SEO", "Give me a primary keyword, 8 secondary keywords, 15 tags and 5 chapter titles for a video about "
         "\"{topic}\". Base it on likely YouTube search intent."),
        ("Thumbnails", "Suggest 5 thumbnail concepts for \"{title}\": the single big idea, 3-word max text, facial "
         "expression, and color contrast. Make them scroll-stopping."),
        ("Sponsor Outreach", "Write a concise sponsorship pitch email to {brand}. Include my channel stats "
         "({subs} subs, {views}/mo views), a relevant integration idea, deliverables and a clear CTA."),
        ("Community Posts", "Write 5 community-tab posts for a {niche} channel: 2 polls, 1 question, 1 BTS image "
         "caption, 1 announcement. Keep them casual and engagement-driving."),
    ]
    row = 4
    ws.cell(row=row, column=2, value="Use Case").style = "th"
    ws.cell(row=row, column=3, value="Prompt (edit the {tokens})").style = "th"
    ws.row_dimensions[row].height = 28
    for i, (name, prompt) in enumerate(prompts):
        r = 5 + i
        ws.cell(row=r, column=2, value=name).style = "field_label"
        ws.cell(row=r, column=3, value=prompt).style = "msg"
        ws.row_dimensions[r].height = 58
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW)


# ===========================================================================
# 24 — Goals & OKRs
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 14, 16, 12, 12, 2])
    luxe_header(ws, "F", "🎯  GOALS & OKRs",
                "Run your channel on objectives — measurable key results with live progress.")
    table_headers(ws, 4, ["Objective / Key Result", "Category", "Target", "Current", "Progress"])
    goals = [
        ("Hit 100k subscribers", "Subscribers", "100,000", "84,200", 0.84),
        ("$12k/mo revenue", "Revenue", "$12,000", "$9,840", 0.82),
        ("Publish 7 videos / month", "Uploads", "7", "6", 0.86),
        ("Reach 20k watch hrs / mo", "Watch Hours", "20,000", "14,600", 0.73),
        ("Land 6 active brand deals", "Sponsors", "6", "4", 0.67),
        ("Launch faceless course", "Products", "Ship it", "In build", 0.55),
        ("Grow CTR to 8%", "Engagement", "8%", "6.8%", 0.68),
    ]
    start = L0
    for i, (g, cat, tgt, cur, prog) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalCatList")
    nrange(wb, "GoalCategory", "Goals", "B", start, end)
    nrange(wb, "GoalProgress", "Goals", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    merge_set(ws, "B14:E14", "THIS QUARTER'S FOCUS", "section_gold"); ws.row_dimensions[14].height = 22
    focus = ["Ship the faceless course (biggest revenue lever)",
             "Systematize sponsor outreach → 6 active deals",
             "Double down on long-form (best watch time + RPM)",
             "Improve thumbnails/titles → CTR to 8%"]
    for i, fx in enumerate(focus):
        r = 15 + i
        cb = ws.cell(row=r, column=2, value="◆"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=11, color=ACCENT); cb.border = BOX
        merge_set(ws, f"C{r}:E{r}", fx, "td_left")


# ===========================================================================
# 25 — Collaboration CRM
# ===========================================================================
def build_collab(wb):
    rows = [
        ("Leo Marín (@leomakes)", "Creator", "Podcast swap", 10, "50/50", "Confirmed"),
        ("Studio Kwan", "Designer", "Thumbnail redesign", -3, "Flat $400", "Delivered"),
        ("Mia (writer)", "Editor", "Blog SEO series", 12, "$250/post", "In Progress"),
        ("Jon (editor)", "Editor", "Shorts editing", 5, "$60/short", "In Progress"),
        ("Priya D. (@priyacodes)", "Creator", "Collab video", 8, "50/50", "Scheduled"),
        ("Peak Agency", "Agency", "Channel management", 30, "15% rev", "Lead"),
    ]
    sample = [(who, role, proj, dplus(d) if d >= 0 else dminus(-d), split, status) for (who, role, proj, d, split, status) in rows]
    ws, start, end = build_log(
        wb, "Collabs", "👥", "COLLABORATION CRM",
        "Creators, editors, designers & agencies — projects, splits & status in one place.",
        ["Name", "Role", "Project", "Deadline", "Revenue Split", "Status"],
        sample, [24, 14, 22, 13, 14, 15],
        text_left={1, 3, 5}, dates={4}, reserved=30)


# ===========================================================================
# 26 — Audience Insights
# ===========================================================================
def build_audience(wb):
    ws = wb.create_sheet("Audience"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 30, 4, 24, 24, 2])
    luxe_header(ws, "G", "👤  AUDIENCE INSIGHTS",
                "Know who you serve — personas, FAQs & the content they're asking for.")
    merge_set(ws, "B5:C5", "PRIMARY PERSONA", "section")
    persona = [("Name", "Side-Hustle Sam"), ("Age", "24–38"), ("Stage", "0–10k subs, part-time"),
               ("Goal", "Turn a channel into income"), ("Blocker", "Inconsistent, overwhelmed"),
               ("Watches on", "Desktop + mobile"), ("Spends on", "Tools & courses that save time")]
    for i, (lab, val) in enumerate(persona):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        ws.cell(row=r, column=3, value=val).style = "field_value"
    merge_set(ws, "E5:F5", "TOP VIEWER REQUESTS", "section_gold")
    reqs = [("Sponsorship outreach", 68), ("Editing workflows", 54), ("Faceless channels", 47),
            ("Pricing products", 39), ("Batch filming", 33)]
    ws.cell(row=6, column=5, value="Topic").style = "th"; ws.cell(row=6, column=6, value="Requests").style = "th"
    for i, (t, n) in enumerate(reqs):
        r = 7 + i
        ws.cell(row=r, column=5, value=t).style = "td_left"
        ws.cell(row=r, column=6, value=n).style = "td"
        if i % 2:
            ws.cell(row=r, column=5).fill = fill(MUTED_ROW); ws.cell(row=r, column=6).fill = fill(MUTED_ROW)
    merge_set(ws, "B14:C14", "FAQs TO ANSWER ON VIDEO", "section_gold"); ws.row_dimensions[14].height = 22
    faqs = ["\"What editing software do you use?\"", "\"How do you get sponsors this small?\"",
            "\"Do you script every video?\"", "\"What's your upload schedule?\""]
    for i, f in enumerate(faqs):
        r = 15 + i
        merge_set(ws, f"B{r}:C{r}", f"•  {f}", "td_left")
    merge_set(ws, "E14:F14", "CONTENT GAPS", "section_gold"); ws.row_dimensions[14].height = 22
    gaps = ["Beginner faceless setup (high demand)", "Live-stream monetization",
            "AI editing tutorial", "Analytics deep-dive series"]
    for i, g in enumerate(gaps):
        r = 15 + i
        merge_set(ws, f"E{r}:F{r}", f"•  {g}", "td_left")


# ===========================================================================
# 27 — Tax Preparation
# ===========================================================================
def build_taxes(wb):
    ws = wb.create_sheet("Taxes"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 16, 18, 3, 22, 14, 2])
    luxe_header(ws, "I", "🧾  TAX PREPARATION",
                "Deduction-ready all year — categorized expenses, quarterly estimates & receipts.")
    table_headers(ws, 4, ["Deductible Category", "This Month", "YTD", "Receipt Ref"])
    cats = [
        ("Software & Subscriptions", 260, 1820, "Card stmt"),
        ("Equipment", 680, 4760, "Receipts /gear"),
        ("Contractors (editors)", 1000, 7000, "1099s"),
        ("Marketing & Ads", 180, 1260, "Card stmt"),
        ("Home Office", 90, 630, "Sq-ft method"),
        ("Education & Courses", 99, 693, "Receipts"),
        ("Travel & Mileage", 0, 940, "Mileage log"),
        ("Depreciation (est.)", 320, 2240, "CPA schedule"),
        ("Platform / Payment Fees", 140, 980, "Payout reports"),
    ]
    start = L0
    for i, (cat, mo, ytd, rec) in enumerate(cats):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        c1 = ws.cell(row=r, column=2, value=mo); c1.style = "td"; c1.number_format = '"$"#,##0'
        c2 = ws.cell(row=r, column=3, value=ytd); c2.style = "td"; c2.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=rec).style = "td_left"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = start + len(cats)
    ws.cell(row=total, column=1, value="TOTAL DEDUCTIONS").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{total-1})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    merge_set(ws, "G4:H4", "QUARTERLY ESTIMATE", "section_gold")
    q = [("Est. taxable income (YTD)", "=RevenueTotal*6-C" + str(total), '"$"#,##0'),
         ("Est. tax rate", 0.25, "0%"),
         ("Quarterly set-aside", "=MAX(H5*H6/4,0)", '"$"#,##0'),
         ("Next payment due", _next_quarter(), "mm/dd/yyyy")]
    for i, (lab, val, fmt) in enumerate(q):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=val); c.style = "field_value"; c.number_format = fmt
        if lab == "Quarterly set-aside":
            ws.cell(row=r, column=8).fill = fill(WARN_BG)
    merge_set(ws, "G10:H10", "Not tax advice — confirm with your CPA.", "subtitle")


def _next_quarter():
    today = dt.date.today()
    for md in [(4, 15), (6, 15), (9, 15), (1, 15)]:
        y = today.year if md != (1, 15) else today.year + 1
        d = dt.date(y, md[0], md[1])
        if d >= today:
            return d
    return dt.date(today.year + 1, 1, 15)


# ===========================================================================
# 28 — Annual Strategic Planner
# ===========================================================================
def build_annual(wb):
    ws = wb.create_sheet("Annual Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 22, 22, 20, 16, 2])
    luxe_header(ws, "G", "📆  ANNUAL STRATEGIC PLANNER",
                "See the whole year — quarterly themes, launches, campaigns & rest.")
    table_headers(ws, 5, ["Quarter", "Theme", "Flagship Series", "Major Launch", "Objective"], start_col=2)
    quarters = [
        ("Q1", "Foundations", "Creator Systems", "Content OS template", "Hit 90k subs"),
        ("Q2", "Monetize", "Make Money on YT", "Faceless course", "$12k/mo revenue"),
        ("Q3", "Scale", "Faceless Channels", "Membership tier", "20k watch hrs/mo"),
        ("Q4", "Authority", "Year in Review", "Cohort program", "100k subs"),
    ]
    start = 6
    for i, (q, theme, series, launch, obj) in enumerate(quarters):
        r = start + i
        ws.cell(row=r, column=2, value=q).style = "input"
        ws.cell(row=r, column=3, value=theme).style = "td_left"
        ws.cell(row=r, column=4, value=series).style = "td_left"
        ws.cell(row=r, column=5, value=launch).style = "td_left"
        ws.cell(row=r, column=6, value=obj).style = "td_left"
        ws.row_dimensions[r].height = 26
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    merge_set(ws, "B11:F11", "TENTPOLE MOMENTS & REST", "section_gold"); ws.row_dimensions[11].height = 22
    moments = ["Jan — New-year 'systems' push (highest search)", "Apr — Faceless course launch week",
               "Jul — Mid-year deload + video-bank shoot", "Nov — Black Friday product bundle",
               "Dec — Year-in-review + 2 weeks off"]
    for i, m in enumerate(moments):
        r = 12 + i
        cb = ws.cell(row=r, column=2, value="◆"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=11, color=ACCENT); cb.border = BOX
        merge_set(ws, f"C{r}:F{r}", m, "td_left")


# ===========================================================================
# 29 — Photo Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "📸  PHOTO GALLERY",
                "Your channel's visual board — branding, studio, thumbnails, awards & vision.")
    sections = ["Channel Branding", "Studio Setup", "Thumbnail Board", "Milestones / Awards", "Creator Team", "Vision Board"]
    top0 = 5; card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th"); ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+5}", "📷\nPaste image here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 6):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 6, column=col, value="Notes").style = "field_label"
        merge_set(ws, f"{M}{row+6}:{R}{row+6}", "", "field_value")
        ws.cell(row=row + 7, column=col, value="Status").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")


# ===========================================================================
# 1 — Executive Command Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ▶  YOUTUBE COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Subscribers, revenue, sponsors & content — your whole YouTube business, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("SUBSCRIBERS", "=SubNow", "num"),
        ("VIEWS (28D)", "=Views28", "num"),
        ("WATCH HOURS", "=WatchHrs", "num"),
        ("PUBLISHED (28D)", '=COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Published")', "num"),
        ("MONTHLY REVENUE", "=RevenueTotal", "money"),
        ("NET PROFIT", "=RevenueTotal-ExpenseTotal", "money"),
    ]
    row2 = [
        ("RPM", "=RPM", "money2"),
        ("AVG CTR", "=AvgCTR", "pct1"),
        ("AVG VIEW DUR", "=AvgViewDur", "text"),
        ("BRAND DEALS", '=COUNTIF(SponStage,"Signed")+COUNTIF(SponStage,"Delivered")+COUNTIF(SponStage,"Negotiation")', "num"),
        ("UPLOAD CONSISTENCY", '=IFERROR(MIN(COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Published")/UploadGoal,1),0)', "pct"),
        ("CHANNEL HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "AUDIENCE & REVENUE", "section_gold")
    # subscriber growth line
    ln = LineChart(); ln.title = "Subscriber Growth (K)"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=26, max_row=32), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=27, max_row=32)); ln.legend = None
    ws.add_chart(ln, "B12")
    # revenue by source donut
    d1 = DoughnutChart(); d1.title = "Revenue by Source"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Finance"], min_col=2, min_row=4, max_row=12), titles_from_data=True)
    d1.set_categories(Reference(wb["Finance"], min_col=1, min_row=5, max_row=12)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "CONTENT & PROFIT", "section_gold")
    # top videos bar
    cb = BarChart(); cb.type = "bar"; cb.title = "Top Videos by Views"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); cb.legend = None
    ws.add_chart(cb, "B30")
    # expense breakdown donut
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Finance"], min_col=7, min_row=6, max_row=13), titles_from_data=True)
    eb.set_categories(Reference(wb["Finance"], min_col=6, min_row=6, max_row=13)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "YouTube Command Center™ — from idea to income, all in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_calendar(wb)
    build_pipeline(wb); build_ideas(wb); build_scripts(wb); build_thumbnails(wb)
    build_seo(wb); build_analytics(wb); build_longform(wb); build_shorts(wb)
    build_community(wb); build_live(wb); build_playlists(wb); build_sponsorship(wb)
    build_affiliate(wb); build_products(wb); build_finance(wb); build_equipment(wb)
    build_repurpose(wb); build_brandkit(wb); build_assets(wb); build_prompts(wb)
    build_goals(wb); build_collab(wb); build_audience(wb); build_taxes(wb)
    build_annual(wb); build_gallery(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Channel Profile", "Calendar", "Pipeline", "Ideas", "Scripts",
             "Thumbnails", "SEO", "Analytics", "Long Form", "Shorts", "Community", "Live", "Playlists",
             "Sponsors", "Affiliate", "Products", "Finance", "Equipment", "Repurposing", "Brand Kit",
             "Assets", "AI Prompts", "Goals", "Collabs", "Audience", "Taxes", "Annual Plan", "Gallery", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "YouTube_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
