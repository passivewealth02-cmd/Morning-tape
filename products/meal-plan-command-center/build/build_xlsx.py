"""Build Meal Planning & Grocery Budget Command Center™ — The Kitchen Operating System.

14 tabs · a premium meal-planning & grocery-budget operating system in Google Sheets &
Excel. Dashboard, a cost-per-meal engine (recipe ingredients + a price book → cost per
serving vs eating out), a weekly plan, recipes, a price book, a grocery list, a pantry,
a budget, an eating-out log, a savings roll-up, meal ideas and a monthly summary — one
dashboard. Plan the week, know the cost, and save on every meal.

Run: python3 build_xlsx.py   ->  ../Meal_Plan_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
CATEGORY = ["Dinner", "Lunch", "Breakfast", "Snack", "Dessert"]
STORE = ["Aldi", "Costco", "Kroger", "Walmart", "Farmers Mkt"]
MEALTYPE = ["Quick", "Batch", "One-pan", "Slow cooker", "No-cook"]

SAVE_GOAL = 5
SERVING_GOAL = 5
PRICE_GOAL = 12
EATOUT_GOAL = 4

# Cost-per-meal engine — flagship recipe & inputs
SERVINGS_N = 4
EATOUT_SERVING = 12.50
# Ingredients for the flagship (Sheet-Pan Chicken & Veg): (ingredient, cost)
INGREDIENTS = [("Chicken thighs", 6.00), ("Potatoes", 1.50), ("Broccoli", 1.50), ("Olive oil & seasoning", 1.00)]

# Recipes: (recipe, total cost, servings)
RECIPES = [
    ("Sheet-Pan Chicken & Veg", 10.00, 4), ("Pasta Primavera", 8.00, 4), ("Beef Tacos", 12.00, 4),
    ("Veggie Chili", 9.00, 6), ("Fried Rice", 7.00, 4), ("Salmon & Rice", 14.00, 4), ("Lentil Soup", 6.00, 6),
]

# Price book: (item, store, unit, price)
PRICE_BOOK = [
    ("Chicken thighs", "Aldi", "lb", 1.99), ("Ground beef", "Costco", "lb", 3.99), ("Potatoes", "Aldi", "5 lb", 2.99),
    ("Broccoli", "Aldi", "lb", 1.49), ("Rice", "Costco", "10 lb", 8.99), ("Pasta", "Aldi", "lb", 0.99),
    ("Eggs", "Aldi", "dozen", 2.29), ("Milk", "Aldi", "gal", 2.79), ("Onions", "Aldi", "3 lb", 2.49),
    ("Olive oil", "Costco", "2 L", 12.99), ("Canned beans", "Aldi", "can", 0.79), ("Cheese", "Costco", "2 lb", 6.99),
]

# Weekly plan: (day, meal, cost)
WEEKLY = [
    ("Mon", "Sheet-Pan Chicken & Veg", 10.00), ("Tue", "Pasta Primavera", 8.00), ("Wed", "Beef Tacos", 12.00),
    ("Thu", "Veggie Chili", 9.00), ("Fri", "Fried Rice", 7.00), ("Sat", "Salmon & Rice", 14.00), ("Sun", "Lentil Soup", 6.00),
]

# Grocery list: (item, qty, est cost, got?)
GROCERY = [
    ("Chicken thighs", "3 lb", 6.00, "No"), ("Ground beef", "1 lb", 4.00, "No"), ("Broccoli", "2 lb", 3.00, "Yes"),
    ("Potatoes", "5 lb", 3.00, "No"), ("Rice", "bag", 4.00, "Yes"), ("Pasta", "2 lb", 2.00, "No"),
    ("Eggs", "dozen", 2.29, "No"), ("Milk", "gal", 2.79, "No"),
]

# Pantry: (item, on hand, staple?)
PANTRY = [
    ("Rice", "8 lb", "Yes"), ("Pasta", "3 lb", "Yes"), ("Canned beans", "6 cans", "Yes"),
    ("Olive oil", "1.5 L", "Yes"), ("Onions", "4", "Yes"), ("Spices", "stocked", "Yes"),
]

# Budget: weekly grocery spend this month
WEEKS_SPEND = [("Week 1", 135), ("Week 2", 140), ("Week 3", 130), ("Week 4", 135)]

# Eating-out log: (date, place, cost)
EATOUT = [
    ("Jul 4", "Pizza night", 32), ("Jul 11", "Thai takeout", 38), ("Jul 18", "Burgers", 28), ("Jul 25", "Sushi", 44),
]
EATOUT_MEALS_N = 10

# Meal ideas: (idea, type, ~cost/serving)
IDEAS = [
    ("Sheet-pan fajitas", "One-pan", 2.75), ("Slow-cooker chili", "Slow cooker", 1.50), ("Egg fried rice", "Quick", 1.75),
    ("Big-batch bolognese", "Batch", 2.25), ("Grain bowls", "No-cook", 2.50), ("Breakfast-for-dinner", "Quick", 1.50),
]

# Monthly summary: (month, savings vs eating out)
MONTHS = [("Feb", 720), ("Mar", 800), ("Apr", 860), ("May", 900), ("Jun", 940), ("Jul", 960)]

SERVINGS_COOKED_N = 120
GROCERY_BUDGET_N = 600

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your goals & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR GOALS", "section")
    controls = [
        ("Household name", "Hearth & Harvest", None, "Household"),
        ("Cook", "Nora", None, "Cook"),
        ("Monthly grocery budget", GROCERY_BUDGET_N, '"$"#,##0', "GroceryBudget"),
        ("Save-per-serving goal", SAVE_GOAL, '"$"#,##0', "SaveGoal"),
        ("Cost-per-serving goal (under)", SERVING_GOAL, '"$"#,##0', "ServingGoal"),
        ("Price-book items goal", PRICE_GOAL, "0", "PriceGoal"),
        ("Eating-out meals goal (max)", EATOUT_GOAL, "0", "EatOutGoal"),
        ("Eat-out price / serving", EATOUT_SERVING, '"$"#,##0.00', "EatOutServing"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Category", CATEGORY, "CategoryList"), ("F", "Store", STORE, "StoreList"),
             ("G", "Meal type", MEALTYPE, "MealTypeList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🍽  MEAL PLANNING & GROCERY BUDGET COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Plan the week, know the cost, and save on every meal.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE KITCHEN, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A grocery budget lives or dies on one number most people never work out: what a home-cooked meal "
                      "actually costs per serving — and how much that beats eating out. This makes it clear: a "
                      "cost-per-meal engine adds a recipe's ingredients, divides by servings, and shows your cost per "
                      "plate and what you save versus takeout. Plan your week, build a price book, run a grocery list "
                      "and pantry, and track your budget — all in ONE premium Google Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your grocery budget & goals.",
             "2.  Cost a meal in the Cost Per Meal tab — ingredients & servings.",
             "3.  Read your cost per serving and what you save vs eating out.",
             "4.  Plan your week and build your grocery list & price book.",
             "5.  Track your budget, pantry and eating-out log.",
             "6.  Check the Dashboard: cost, savings & a Kitchen Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional household (Hearth & Harvest, cook Nora) is included so you can see how it "
               "all connects — just type over it with your own recipes and prices. Your cost per serving and your "
               "monthly savings versus eating out are the two numbers that decide whether your grocery budget works, "
               "and they roll into a live Kitchen Score. Twelve matching printable pages (weekly plan, grocery list, "
               "price book, meal planner & more) are included. This is a budgeting & organizing tool, not financial or "
               "nutrition advice — confirm figures with your own sources.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "A planned week is a cheaper week — and a calmer one.", "section_gold")


# ===========================================================================
def build_costmeal(wb):
    ws = wb.create_sheet("Cost Per Meal"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "🧮  COST PER MEAL",
                "A recipe's ingredients ÷ servings = your cost per plate, and what you save vs eating out.")
    ws.cell(row=5, column=2, value="RECIPE").style = "section_gold"
    ws.cell(row=5, column=3, value="Sheet-Pan Chicken & Veg").font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Ingredient", "Cost"], start_col=2)
    start = 7
    for i, (ing, cost) in enumerate(INGREDIENTS):
        r = start + i
        ws.cell(row=r, column=2, value=ing).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INGREDIENTS) - 1
    nrange(wb, "IngRows", "Cost Per Meal", "C", start, end)
    mt = end + 1
    ws.cell(row=mt, column=2, value="= RECIPE COST").style = "th"
    cm = ws.cell(row=mt, column=3, value="=SUM(IngRows)"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0.00'
    cell_name(wb, "RecipeCost", "Cost Per Meal", f"$C${mt}")
    sv = mt + 1
    ws.cell(row=sv, column=2, value="Servings").style = "field_label"
    cs = ws.cell(row=sv, column=3, value=SERVINGS_N); cs.style = "input"; cs.number_format = "#,##0"
    cell_name(wb, "Servings", "Cost Per Meal", f"$C${sv}")
    cp = sv + 1
    ws.cell(row=cp, column=2, value="= COST PER SERVING").style = "th"
    ccp = ws.cell(row=cp, column=3, value="=IFERROR(RecipeCost/Servings,0)"); ccp.style = "td"; ccp.font = Font(bold=True, size=13, color=PRIMARY); ccp.fill = fill(MINT_BG); ccp.number_format = '"$"#,##0.00'
    cell_name(wb, "CostPerServing", "Cost Per Meal", f"$C${cp}")
    eo = cp + 2
    ws.cell(row=eo, column=2, value="Eat-out price per serving").style = "field_label"
    ceo = ws.cell(row=eo, column=3, value="=EatOutServing"); ceo.style = "field_value"; ceo.number_format = '"$"#,##0.00'
    ws.cell(row=eo + 1, column=2, value="= YOU SAVE / SERVING").style = "th"
    csv = ws.cell(row=eo + 1, column=3, value="=EatOutServing-CostPerServing"); csv.style = "td"; csv.font = Font(bold=True, size=13, color=PRIMARY); csv.fill = fill(MINT_BG); csv.number_format = '"$"#,##0.00'
    cell_name(wb, "SavedServing", "Cost Per Meal", f"$C${eo+1}")
    ws.cell(row=eo + 3, column=2, value="Cook once, and every plate beats the drive-thru.").style = "section_gold"


def build_recipes(wb):
    ws = wb.create_sheet("Recipes"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 12, 16, 2])
    luxe_header(ws, "E", "📖  RECIPES",
                "Your go-to meals — total cost, servings and the cost per serving on each. Build your rotation.")
    table_headers(ws, 4, ["Recipe", "Cost", "Servings", "Per Serving"], start_col=2)
    start = L0
    for i, (rec, cost, serv) in enumerate(RECIPES):
        r = start + i
        ws.cell(row=r, column=2, value=rec).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cs = ws.cell(row=r, column=4, value=serv); cs.style = "input"; cs.number_format = "#,##0"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cp.style = "td"; cp.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(RECIPES) - 1
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=1.5, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=2.5, mid_color="FFFFF3CD",
                       end_type="num", end_value=4.0, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


def build_pricebook(wb):
    ws = wb.create_sheet("Price Book"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 12, 12, 2])
    luxe_header(ws, "E", "🏷  PRICE BOOK",
                "What each staple costs and where it's cheapest — the secret weapon of a low grocery bill.")
    table_headers(ws, 4, ["Item", "Store", "Unit", "Price"], start_col=2)
    start = L0
    for i, (item, store, unit, price) in enumerate(PRICE_BOOK):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ws.cell(row=r, column=3, value=store).style = "td"
        ws.cell(row=r, column=4, value=unit).style = "td"
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRICE_BOOK) - 1
    nrange(wb, "PriceItem", "Price Book", "B", start, end)
    add_dv(ws, f"C{start}:C{end}", "StoreList")
    ws.cell(row=end + 2, column=2, value="Items priced").style = "field_label"
    ci = ws.cell(row=end + 2, column=5, value="=COUNTA(PriceItem)"); ci.style = "field_value"; ci.number_format = "#,##0"
    cell_name(wb, "PricedItems", "Price Book", f"$E${end+2}")
    ws.freeze_panes = "A5"


def build_weekly(wb):
    ws = wb.create_sheet("Weekly Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 30, 14, 2])
    luxe_header(ws, "D", "🗓  WEEKLY PLAN",
                "Seven dinners, planned and priced — the whole point of a cheaper, calmer week.")
    table_headers(ws, 4, ["Day", "Dinner", "Cost"], start_col=2)
    start = L0
    for i, (day, meal, cost) in enumerate(WEEKLY):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        ws.cell(row=r, column=3, value=meal).style = "td_left"
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(WEEKLY) - 1
    nrange(wb, "PlanDay", "Weekly Plan", "B", start, end)
    nrange(wb, "PlanCost", "Weekly Plan", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    cd = ws.cell(row=tot, column=3, value="=COUNTA(PlanDay)"); cd.style = "td"; cd.fill = fill(SURFACE); cd.number_format = "#,##0"
    ct = ws.cell(row=tot, column=4, value="=SUM(PlanCost)"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0.00'
    cell_name(wb, "WeeklyPlanCost", "Weekly Plan", f"$D${tot}")
    cell_name(wb, "PlannedDinners", "Weekly Plan", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_grocery(wb):
    ws, start, end = build_log(
        wb, "Grocery List", "🛒", "GROCERY LIST",
        "This week's shopping — quantity, estimated cost and what's already in the cart.",
        ["Item", "Qty", "Est. Cost", "Got?"],
        GROCERY, [2, 24, 14, 14, 12, 2], text_left={2, 3}, money2={4}, reserved=30, start_col=2,
        validations=[("E", "YesNoList")])
    nrange(wb, "GrocCost", "Grocery List", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="EST. TOTAL").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(GrocCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0.00'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)


def build_pantry(wb):
    ws, start, end = build_log(
        wb, "Pantry", "🥫", "PANTRY",
        "What's on the shelf — shop from here first, and never buy what you already own.",
        ["Item", "On Hand", "Staple?"],
        PANTRY, [2, 24, 16, 14, 2], text_left={2, 3}, reserved=26, start_col=2,
        validations=[("D", "YesNoList")])


def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  BUDGET",
                "Your monthly grocery budget versus what you've spent — stay in the green.")
    ws.cell(row=5, column=2, value="Monthly grocery budget").style = "field_label"
    cb = ws.cell(row=5, column=3, value="=GroceryBudget"); cb.style = "field_value"; cb.number_format = '"$"#,##0'; cb.fill = fill(MINT_BG)
    table_headers(ws, 7, ["Week", "Spent"], start_col=2)
    start = 8
    for i, (wk, amt) in enumerate(WEEKS_SPEND):
        r = start + i
        ws.cell(row=r, column=2, value=wk).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(WEEKS_SPEND) - 1
    nrange(wb, "WeekSpend", "Budget", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="SPENT THIS MONTH").style = "th"
    cs = ws.cell(row=tot, column=3, value="=SUM(WeekSpend)"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cell_name(wb, "SpentMonth", "Budget", f"$C${tot}")
    ws.cell(row=tot + 1, column=2, value="Under / (over) budget").style = "field_label"
    cu = ws.cell(row=tot + 1, column=3, value="=GroceryBudget-SpentMonth"); cu.style = "field_value"; cu.number_format = '"$"#,##0'; cu.fill = fill(MINT_BG)
    ws.freeze_panes = "A8"


def build_eatout(wb):
    ws = wb.create_sheet("Eating Out"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 24, 14, 2])
    luxe_header(ws, "D", "🍔  EATING OUT",
                "Every takeout & restaurant meal — what it cost, and how many. Cooking is the cheaper habit.")
    ws.cell(row=5, column=2, value="Meals eaten out this month").style = "field_label"
    cm = ws.cell(row=5, column=3, value=EATOUT_MEALS_N); cm.style = "input"; cm.number_format = "#,##0"
    cell_name(wb, "EatOutMeals", "Eating Out", "$C$5")
    table_headers(ws, 7, ["Date", "Place", "Cost"], start_col=2)
    start = 8
    for i, (date, place, cost) in enumerate(EATOUT):
        r = start + i
        ws.cell(row=r, column=2, value=date).style = "td_left"
        ws.cell(row=r, column=3, value=place).style = "td_left"
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EATOUT) - 1
    nrange(wb, "EatOutCost", "Eating Out", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="SPENT EATING OUT").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(EatOutCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A8"


def build_savings(wb):
    ws = wb.create_sheet("Savings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 16, 2])
    luxe_header(ws, "C", "🌿  SAVINGS",
                "What cooking saved you this month — servings at home versus the eat-out price.")
    ws.cell(row=5, column=2, value="Servings cooked at home this month").style = "field_label"
    cs = ws.cell(row=5, column=3, value=SERVINGS_COOKED_N); cs.style = "input"; cs.number_format = "#,##0"
    cell_name(wb, "ServingsCooked", "Savings", "$C$5")
    ws.cell(row=6, column=2, value="× Eat-out price / serving").style = "field_label"
    ce = ws.cell(row=6, column=3, value="=EatOutServing"); ce.style = "field_value"; ce.number_format = '"$"#,##0.00'
    ws.cell(row=7, column=2, value="= Eat-out equivalent").style = "field_label"
    ceq = ws.cell(row=7, column=3, value="=ServingsCooked*EatOutServing"); ceq.style = "field_value"; ceq.number_format = '"$"#,##0'
    ws.cell(row=8, column=2, value="− Grocery spent this month").style = "field_label"
    cg = ws.cell(row=8, column=3, value="=SpentMonth"); cg.style = "field_value"; cg.number_format = '"$"#,##0'
    ws.cell(row=9, column=2, value="= MONTHLY SAVINGS").style = "th"
    cms = ws.cell(row=9, column=3, value="=ServingsCooked*EatOutServing-SpentMonth"); cms.style = "td"; cms.font = Font(bold=True, size=14, color=PRIMARY); cms.fill = fill(MINT_BG); cms.number_format = '"$"#,##0'
    cell_name(wb, "MonthlySavings", "Savings", "$C$9")
    ws.cell(row=11, column=2, value="Every home-cooked plate is money back in your pocket.").style = "section_gold"


def build_ideas(wb):
    ws = wb.create_sheet("Meal Ideas"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 16, 2])
    luxe_header(ws, "D", "💡  MEAL IDEAS",
                "A running list of cheap, easy wins — never stare into the fridge wondering what's for dinner again.")
    table_headers(ws, 4, ["Idea", "Type", "~ / Serving"], start_col=2)
    start = L0
    for i, (idea, typ, cost) in enumerate(IDEAS):
        r = start + i
        ws.cell(row=r, column=2, value=idea).style = "td_left"
        ws.cell(row=r, column=3, value=typ).style = "td"
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(IDEAS) - 1
    add_dv(ws, f"C{start}:C{end}", "MealTypeList")
    ws.freeze_panes = "A5"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "Your savings versus eating out, month by month — watch the habit pay off.")
    ws.cell(row=5, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 6, ["Month", "Saved"], start_col=2)
    ts = 7
    for i, (m, sv) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sv); cs.style = "input"; cs.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "SaveTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Saved by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🍽  MEAL PLANNING & GROCERY BUDGET COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Cost per plate, budget, savings & a Kitchen Score — your whole kitchen, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("RECIPE COST", "=RecipeCost", "money2"),
        ("SERVINGS", "=Servings", "num"),
        ("COST / SERVING", "=CostPerServing", "money2"),
        ("EAT-OUT / SERVING", "=EatOutServing", "money2"),
        ("SAVED / SERVING", "=SavedServing", "money2"),
        ("WEEKLY PLAN", "=WeeklyPlanCost", "money"),
    ]
    row2 = [
        ("GROCERY BUDGET", "=GroceryBudget", "money"),
        ("SPENT THIS MONTH", "=SpentMonth", "money"),
        ("BUDGET USED", "=IFERROR(SpentMonth/GroceryBudget,0)", "pct"),
        ("SERVINGS COOKED", "=ServingsCooked", "num"),
        ("MONTHLY SAVINGS", "=MonthlySavings", "money"),
        ("KITCHEN SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "KITCHEN HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SAVED BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Under budget", "=IF(SpentMonth<=GroceryBudget,1,IFERROR(GroceryBudget/SpentMonth,0))"),
        ("Cooking saves money", "=IFERROR(MIN(SavedServing/SaveGoal,1),0)"),
        ("Week is planned", "=IFERROR(MIN(PlannedDinners/7,1),0)"),
        ("Cheap per serving", "=IFERROR(MIN(ServingGoal/CostPerServing,1),0)"),
        ("Price book built", "=IFERROR(MIN(PricedItems/PriceGoal,1),0)"),
        ("Eating out rare", "=IFERROR(MIN(EatOutGoal/EatOutMeals,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Saved by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=7, max_row=6 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=7, max_row=6 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Meal Planning & Grocery Budget Command Center™ — plan it, cost it, save on every meal.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_costmeal(wb); build_weekly(wb)
    build_recipes(wb); build_pricebook(wb); build_grocery(wb); build_pantry(wb)
    build_budget(wb); build_eatout(wb); build_savings(wb); build_ideas(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Cost Per Meal", "Weekly Plan", "Recipes", "Price Book", "Grocery List",
             "Pantry", "Budget", "Eating Out", "Savings", "Meal Ideas", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Meal_Plan_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
