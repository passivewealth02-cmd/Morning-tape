"""Build Budget & Money Command Center™ — The Complete Personal-Finance System.

15 tabs (+ Settings) · a premium personal-finance operating system in Google
Sheets & Excel. Dashboard, a zero-based monthly budget (planned vs actual),
income, bill tracker, expense log, savings goals, sinking funds, debt snapshot,
net worth, subscriptions, a 12-month year view, a no-spend tracker and money
goals — one dashboard. Give every dollar a job & build real wealth.

Run: python3 build_xlsx.py   ->  ../Budget_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
CATEGORIES = ["Housing", "Utilities", "Groceries", "Transportation", "Insurance",
              "Dining & Fun", "Health", "Kids / Pets", "Giving", "Savings", "Debt", "Other"]
PAID = ["Paid", "Due", "Overdue"]
FREQ = ["Monthly", "Weekly", "Annual", "Quarterly"]

INCOME_SOURCES = [("Salary (net)", 4400), ("Side gig", 600), ("Other / interest", 200)]

# Monthly budget: (category, group, planned, actual)
BUDGET = [
    ("Rent / Mortgage", "Housing", 1500, 1500),
    ("Utilities", "Utilities", 250, 238),
    ("Groceries", "Groceries", 600, 642),
    ("Transportation & gas", "Transportation", 300, 280),
    ("Insurance", "Insurance", 200, 200),
    ("Dining & takeout", "Dining & Fun", 250, 310),
    ("Subscriptions", "Dining & Fun", 80, 80),
    ("Personal & fun", "Dining & Fun", 200, 175),
    ("Health & medical", "Health", 150, 120),
    ("Kids / pets", "Kids / Pets", 170, 155),
    ("Giving", "Giving", 200, 200),
    ("Savings & investing", "Savings", 800, 800),
    ("Extra debt payment", "Debt", 450, 450),
    ("Sinking funds", "Savings", 50, 50),
]

# Bills: (bill, amount, due day, status)
BILLS = [
    ("Rent / Mortgage", 1500, 1, "Paid"),
    ("Electric", 120, 8, "Paid"),
    ("Water / sewer", 45, 10, "Paid"),
    ("Internet", 70, 12, "Paid"),
    ("Phone", 90, 15, "Paid"),
    ("Car insurance", 140, 18, "Paid"),
    ("Streaming bundle", 25, 20, "Paid"),
    ("Credit card (min)", 80, 22, "Paid"),
    ("Gym", 20, 25, "Due"),
    ("Student loan", 210, 28, "Due"),
]

# Expense log: (date offset, merchant, category, amount)
EXPENSES = [
    (-1, "Grocery Market", "Groceries", 82), (-2, "Gas Station", "Transportation", 44),
    (-3, "Coffee & lunch", "Dining & Fun", 28), (-4, "Pharmacy", "Health", 36),
    (-5, "Online order", "Kids / Pets", 41), (-6, "Restaurant", "Dining & Fun", 64),
    (-7, "Grocery Market", "Groceries", 96), (-8, "Utilities autopay", "Utilities", 238),
    (-9, "Streaming", "Dining & Fun", 25), (-10, "Hardware store", "Housing", 34),
    (-12, "Grocery Market", "Groceries", 78), (-14, "Pet supplies", "Kids / Pets", 52),
    (-15, "Gas Station", "Transportation", 47), (-18, "Gift", "Giving", 40),
    (-20, "Grocery Market", "Groceries", 88),
]

# Savings goals: (goal, target, saved)
SAVEGOALS = [
    ("Emergency Fund (6 mo)", 15000, 9500),
    ("Vacation Fund", 3000, 1800),
    ("New Car Fund", 10000, 2200),
    ("Home Projects", 5000, 3400),
]

# Sinking funds: (fund, monthly set-aside, balance)
SINKING = [
    ("Car maintenance", 60, 600),
    ("Gifts & holidays", 50, 450),
    ("Annual subscriptions", 15, 180),
    ("Medical / dental", 30, 300),
    ("Pet care", 20, 120),
]

# Debt snapshot: (debt, balance, APR, min payment)
DEBTS = [
    ("Credit Card", 3200, 0.199, 80),
    ("Student Loan", 18000, 0.055, 210),
    ("Car Loan", 9000, 0.069, 290),
]

# Net worth — assets: (asset, value)
ASSETS = [
    ("Cash & checking", 12000),
    ("Emergency fund", 9500),
    ("Retirement (401k/IRA)", 45000),
    ("Home equity", 60000),
    ("Car value", 15000),
]

# Subscriptions: (subscription, amount, frequency)
SUBS = [
    ("Streaming bundle", 25, "Monthly"),
    ("Music", 12, "Monthly"),
    ("Cloud storage", 8, "Monthly"),
    ("Gym", 20, "Monthly"),
    ("News", 6, "Monthly"),
    ("Software", 9, "Monthly"),
]

# No-spend tracker: (day, no-spend?)
NOSPEND = [(f"Day {i}", "Yes" if i not in (2, 5, 6, 9, 12, 13, 16, 19, 20, 23, 26, 27) else "No") for i in range(1, 31)]

SAVINGS_TARGET = 0.20
EF_TARGET = 15000
LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR MONEY", "section")
    controls = [
        ("Name / Household", "The Bennett Household", None, "Household"),
        ("Budget Month", "This Month", None, "BudgetMonth"),
        ("Savings-Rate Target", SAVINGS_TARGET, "0%", "SavingsTarget"),
        ("Emergency-Fund Target", EF_TARGET, '"$"#,##0', "EFTarget"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Category", CATEGORIES, "CategoryList"), ("F", "Yes / No", YESNO, "YesNoList"),
             ("G", "Bill Status", PAID, "PaidList"), ("H", "Frequency", FREQ, "FreqList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  💵  BUDGET & MONEY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Give every dollar a job — budget, save, crush debt & watch your net worth grow.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE FINANCIAL LIFE, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A zero-based monthly budget (planned vs actual), income & bill trackers, an expense log, savings "
                      "goals, sinking funds, a debt snapshot, a net-worth tracker, subscriptions, a 12-month year view "
                      "and a no-spend challenge — all in ONE premium Google Sheets & Excel system. Tell every dollar "
                      "where to go before the month starts, catch overspending the day it happens, and watch your "
                      "savings rate and net worth climb month after month. Simple enough for beginners, powerful enough "
                      "for nerds.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open the Income tab — add every paycheck & source for the month.",
             "2.  Build your Monthly Budget — give every dollar a job until 'Left to Budget' hits $0.",
             "3.  Log expenses as you spend; planned-vs-actual flags any category going over.",
             "4.  Fund your Savings Goals & Sinking Funds, and track your Debt Snapshot.",
             "5.  Update Net Worth monthly — assets minus liabilities, watch it grow.",
             "6.  Check the Dashboard: income, spent, saved, savings rate & a Budget Health score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional household is included so you can see how it all connects — just type over "
               "it with your own numbers. Your budget, savings, debt and net worth roll up into a live Budget Health "
               "score. Twelve matching printable pages (monthly budget, bill tracker, savings-goal & debt trackers, "
               "net-worth sheet & more) are included to print and keep. This is a personal budgeting tool, not "
               "financial, tax or investment advice — for big decisions, talk to a qualified professional.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Every dollar with a job, every month — that's how wealth gets built.", "section_gold")


# ===========================================================================
# 3 — Income  (defines Income)
# ===========================================================================
def build_income(wb):
    ws = wb.create_sheet("Income"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  MONTHLY INCOME",
                "Every paycheck & source — this is the top of your zero-based budget.")
    table_headers(ws, 4, ["Source", "Amount"], start_col=2)
    start = L0
    for i, (src, amt) in enumerate(INCOME_SOURCES):
        r = start + i
        ws.cell(row=r, column=2, value=src).style = "td_left"
        c = ws.cell(row=r, column=3, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    end = start + len(INCOME_SOURCES) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL INCOME").style = "th"
    c = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "Income", "Income", f"$C${tot}")
    ws.freeze_panes = "A5"


# ===========================================================================
# 4 — Monthly Budget  (defines BudgetCat / BudgetPlanned / BudgetActual + totals)
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Monthly Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 14, 14, 14, 2])
    luxe_header(ws, "F", "📊  MONTHLY BUDGET (ZERO-BASED)",
                "Give every dollar a job — planned vs actual, until 'Left to Budget' hits $0.")
    merge_set(ws, "B5:C5", "INCOME", "section_gold")
    ws.cell(row=5, column=5, value="=Income").number_format = '"$"#,##0'
    ws.cell(row=5, column=5).font = Font(bold=True, color=PRIMARY)
    ws.cell(row=5, column=4, value="Left to budget →").font = Font(italic=True, color=ACCENT)
    table_headers(ws, 6, ["Category", "Group", "Planned", "Actual", "Remaining"], start_col=2)
    start = L0 + 2
    for i, (cat, grp, plan, actual) in enumerate(BUDGET):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        ws.cell(row=r, column=3, value=grp).style = "td"
        cp = ws.cell(row=r, column=4, value=plan); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=5, value=actual); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=6, value=f"=D{r}-E{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BUDGET) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    for col in (4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=6, value=f"=D{tot}-E{tot}").number_format = '"$"#,##0;[Red]-"$"#,##0'
    ws.cell(row=tot, column=6).font = Font(bold=True, color=PRIMARY); ws.cell(row=tot, column=6).fill = fill(SURFACE); ws.cell(row=tot, column=6).border = BOX
    nrange(wb, "BudgetCat", "Monthly Budget", "B", start, end)
    nrange(wb, "BudgetGroup", "Monthly Budget", "C", start, end)
    nrange(wb, "BudgetPlanned", "Monthly Budget", "D", start, end)
    nrange(wb, "BudgetActual", "Monthly Budget", "E", start, end)
    cell_name(wb, "BudgetPlanTotal", "Monthly Budget", f"$D${tot}")
    cell_name(wb, "BudgetActualTotal", "Monthly Budget", f"$E${tot}")
    # left-to-budget in the header
    ws.cell(row=5, column=5).value = "=Income-BudgetPlanTotal"
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1500, color=PRIMARY, showValue=True))
    ws.freeze_panes = "A7"


# ===========================================================================
# 5 — Bills  (defines BillStatus)
# ===========================================================================
def build_bills(wb):
    ws, start, end = build_log(
        wb, "Bills", "🧾", "BILL TRACKER",
        "Every recurring bill — amount, due day & paid status. Never a late fee again.",
        ["Bill", "Amount", "Due Day", "Status"],
        BILLS, [26, 14, 12, 14], text_left={1}, money={2}, ints={3}, reserved=24,
        validations=[("D", "PaidList")])
    nrange(wb, "BillStatus", "Bills", "D", start, end)
    nrange(wb, "BillAmount", "Bills", "B", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="MONTHLY BILLS").style = "th"
    c = ws.cell(row=tot, column=2, value="=SUM(BillAmount)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cmap = {"Paid": MINT_BG, "Due": WARN_BG, "Overdue": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 6 — Expense Log
# ===========================================================================
def build_expenses(wb):
    rows = [(dminus(-off), merch, cat, amt) for (off, merch, cat, amt) in EXPENSES]
    rows.sort(key=lambda r: r[0], reverse=True)
    build_log(wb, "Expense Log", "🧮", "EXPENSE LOG",
              "Every transaction, tagged to a category — the raw data behind your actual spending.",
              ["Date", "Merchant", "Category", "Amount"],
              rows, [14, 26, 18, 14], text_left={2}, dates={1}, money={4}, reserved=60,
              validations=[("C", "CategoryList")])


# ===========================================================================
# 7 — Savings Goals  (defines GoalTarget / GoalSaved / GoalPct)
# ===========================================================================
def build_savegoals(wb):
    ws = wb.create_sheet("Savings Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 16, 12, 2])
    luxe_header(ws, "E", "🎯  SAVINGS GOALS",
                "Every goal with its target & progress — a full bar is a very good day.")
    table_headers(ws, 4, ["Goal", "Target", "Saved", "Progress"], start_col=2)
    start = L0
    for i, (goal, target, saved) in enumerate(SAVEGOALS):
        r = start + i
        ws.cell(row=r, column=2, value=goal).style = "td_left"
        ct = ws.cell(row=r, column=3, value=target); ct.style = "input"; ct.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SAVEGOALS) - 1
    nrange(wb, "GoalTarget", "Savings Goals", "C", start, end)
    nrange(wb, "GoalSaved", "Savings Goals", "D", start, end)
    nrange(wb, "GoalPct", "Savings Goals", "E", start, end)
    cell_name(wb, "EFSaved", "Savings Goals", f"$D${start}")
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
# 8 — Sinking Funds  (defines SinkBalance)
# ===========================================================================
def build_sinking(wb):
    ws, start, end = build_log(
        wb, "Sinking Funds", "🪙", "SINKING FUNDS",
        "Save a little each month for the big irregular costs — so they never wreck a budget.",
        ["Fund", "Monthly", "Balance"],
        SINKING, [28, 14, 14], text_left={1}, money={2, 3}, reserved=20)
    nrange(wb, "SinkBalance", "Sinking Funds", "C", start, end)
    nrange(wb, "SinkMonthly", "Sinking Funds", "B", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL SET ASIDE").style = "th"
    ws.cell(row=tot, column=2).style = "td"; ws.cell(row=tot, column=2).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=3, value="=SUM(SinkBalance)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "SinkTotal", "Sinking Funds", f"$C${tot}")


# ===========================================================================
# 9 — Debt Snapshot  (defines DebtBalance)
# ===========================================================================
def build_debt(wb):
    ws = wb.create_sheet("Debt Snapshot"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 12, 14, 2])
    luxe_header(ws, "E", "💳  DEBT SNAPSHOT",
                "What you owe, at a glance — balance, rate & minimum. Attack the highest rate first.")
    table_headers(ws, 4, ["Debt", "Balance", "APR", "Min Payment"], start_col=2)
    start = L0
    for i, (name, bal, apr, minp) in enumerate(DEBTS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cb = ws.cell(row=r, column=3, value=bal); cb.style = "input"; cb.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=apr); cr.style = "input"; cr.number_format = "0.0%"
        cm = ws.cell(row=r, column=5, value=minp); cm.style = "input"; cm.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DEBTS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL DEBT").style = "th"
    c = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0'
    nrange(wb, "DebtBalance", "Debt Snapshot", "C", start, end)
    cell_name(wb, "DebtTotal", "Debt Snapshot", f"$C${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        ColorScaleRule(start_type="min", start_color="FF" + HIGHLIGHT, end_type="max", end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


# ===========================================================================
# 10 — Net Worth  (defines AssetsTotal / LiabTotal / NetWorth)
# ===========================================================================
def build_networth(wb):
    ws = wb.create_sheet("Net Worth"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 4, 28, 18, 2])
    luxe_header(ws, "F", "📈  NET WORTH",
                "Assets minus liabilities — the single number that tells the truth. Watch it grow.")
    merge_set(ws, "B5:C5", "ASSETS", "section_gold")
    merge_set(ws, "E5:F5", "LIABILITIES", "section_gold")
    for i, (asset, val) in enumerate(ASSETS):
        r = 6 + i
        ws.cell(row=r, column=2, value=asset).style = "td_left"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"; c.number_format = '"$"#,##0'
    for i, (name, bal, apr, minp) in enumerate(DEBTS):
        r = 6 + i
        ws.cell(row=r, column=5, value=name).style = "td_left"
        c = ws.cell(row=r, column=6, value="=DebtBalance" if False else bal); c.style = "input"; c.number_format = '"$"#,##0'
    ae = 6 + len(ASSETS) - 1; le = 6 + len(DEBTS) - 1
    at = max(ae, le) + 1
    ws.cell(row=at, column=2, value="TOTAL ASSETS").style = "th"
    ca = ws.cell(row=at, column=3, value=f"=SUM(C6:C{ae})"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = '"$"#,##0'
    ws.cell(row=at, column=5, value="TOTAL LIABILITIES").style = "th"
    cl = ws.cell(row=at, column=6, value=f"=SUM(F6:F{le})"); cl.style = "td"; cl.font = Font(bold=True, color=DANGER); cl.fill = fill(SURFACE); cl.number_format = '"$"#,##0'
    cell_name(wb, "AssetsTotal", "Net Worth", f"$C${at}")
    cell_name(wb, "LiabTotal", "Net Worth", f"$F${at}")
    nw = at + 2
    merge_set(ws, f"B{nw}:E{nw}", "NET WORTH  (assets − liabilities)", "section")
    c = ws.cell(row=nw, column=6, value="=AssetsTotal-LiabTotal"); c.style = "field_value"; c.font = Font(bold=True, size=14, color=PRIMARY); c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    cell_name(wb, "NetWorth", "Net Worth", f"$F${nw}")


# ===========================================================================
# 11 — Subscriptions
# ===========================================================================
def build_subs(wb):
    ws, start, end = build_log(
        wb, "Subscriptions", "🔁", "SUBSCRIPTIONS",
        "Every recurring charge in one place — the quiet budget-killer, finally visible.",
        ["Subscription", "Amount", "Frequency"],
        SUBS, [28, 14, 16], text_left={1}, money={2}, reserved=24,
        validations=[("C", "FreqList")])
    nrange(wb, "SubAmount", "Subscriptions", "B", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="MONTHLY TOTAL").style = "th"
    c = ws.cell(row=tot, column=2, value="=SUM(SubAmount)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=3, value="=SUM(SubAmount)*12").number_format = '"$"#,##0'
    ws.cell(row=tot, column=3).font = Font(italic=True, color=ACCENT); ws.cell(row=tot, column=3).fill = fill(SURFACE); ws.cell(row=tot, column=3).border = BOX
    cell_name(wb, "SubTotal", "Subscriptions", f"$B${tot}")


# ===========================================================================
# 12 — Year View
# ===========================================================================
def build_yearview(wb):
    ws = wb.create_sheet("Year View"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 14, 2])
    luxe_header(ws, "F", "🗓  12-MONTH VIEW",
                "The whole year at a glance — income, spending, savings & net worth month by month.")
    table_headers(ws, 4, ["Month", "Income", "Spent", "Saved", "Net Worth"], start_col=2)
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    nw = 104000
    start = L0
    for i, m in enumerate(months):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ws.cell(row=r, column=3, value=5200 if i < 7 else "").style = "input"
        ws.cell(row=r, column=4, value=[4400, 4520, 4380, 4610, 4400, 4550, 4400][i] if i < 7 else "").style = "input"
        ws.cell(row=r, column=5, value=[800, 680, 820, 590, 800, 650, 800][i] if i < 7 else "").style = "input"
        val = nw + i * 1100 if i < 7 else ""
        ws.cell(row=r, column=6, value=val).style = "input"
        for c in range(3, 7):
            ws.cell(row=r, column=c).number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    nrange(wb, "YearNW", "Year View", "F", start, start + 11)
    nrange(wb, "YearMonth", "Year View", "B", start, start + 11)
    ws.freeze_panes = "A5"


# ===========================================================================
# 13 — No-Spend Tracker
# ===========================================================================
def build_nospend(wb):
    ws = wb.create_sheet("No-Spend"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 14, 3, 14, 14, 2])
    luxe_header(ws, "F", "🚫  NO-SPEND CHALLENGE",
                "Color a day green for every no-spend day — small wins add up fast.")
    merge_set(ws, "B5:F5", "MARK EACH DAY: YES = NO-SPEND", "section_gold")
    half = 15
    for i in range(15):
        r = 6 + i
        ws.cell(row=r, column=2, value=NOSPEND[i][0]).style = "td_left"
        cv = ws.cell(row=r, column=3, value=NOSPEND[i][1]); cv.style = "input"
        ws.cell(row=r, column=5, value=NOSPEND[i + 15][0]).style = "td_left"
        cv2 = ws.cell(row=r, column=6, value=NOSPEND[i + 15][1]); cv2.style = "input"
        add_dv(ws, f"C{r}", "YesNoList"); add_dv(ws, f"F{r}", "YesNoList")
    for col in ("C", "F"):
        ws.conditional_formatting.add(f"{col}6:{col}20", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    nrange(wb, "NoSpend1", "No-Spend", "C", 6, 20)
    nrange(wb, "NoSpend2", "No-Spend", "F", 6, 20)
    ws.cell(row=22, column=2, value="No-spend days").style = "field_label"
    c = ws.cell(row=22, column=3, value='=COUNTIF(NoSpend1,"Yes")+COUNTIF(NoSpend2,"Yes")'); c.style = "field_value"; c.number_format = "0"; c.fill = fill(MINT_BG)


# ===========================================================================
# 2 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  💵  BUDGET & MONEY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Income, spending, saving, debt & net worth — your whole financial life, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("INCOME", "=Income", "money"),
        ("SPENT", "=BudgetActualTotal", "money"),
        ("LEFT TO BUDGET", "=Income-BudgetPlanTotal", "money"),
        ("SAVED", '=SUMIF(BudgetGroup,"Savings",BudgetActual)', "money"),
        ("SAVINGS RATE", '=IFERROR(SUMIF(BudgetGroup,"Savings",BudgetActual)/Income,0)', "pct"),
        ("BILLS PAID", '=IFERROR(COUNTIF(BillStatus,"Paid")/COUNTA(BillStatus),0)', "pct"),
    ]
    row2 = [
        ("NET WORTH", "=NetWorth", "money"),
        ("TOTAL DEBT", "=DebtTotal", "money"),
        ("SAVINGS GOALS", "=IFERROR(AVERAGE(GoalPct),0)", "pct"),
        ("SINKING FUNDS", "=SinkTotal", "money"),
        ("SUBSCRIPTIONS", "=SubTotal", "money"),
        ("HEALTH SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "FINANCIAL HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SPENDING BY GROUP", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("On budget", "=IFERROR(IF(BudgetActualTotal<=BudgetPlanTotal,1,BudgetPlanTotal/BudgetActualTotal),0)"),
        ("Savings rate", '=IFERROR(MIN((SUMIF(BudgetGroup,"Savings",BudgetActual)/Income)/SavingsTarget,1),0)'),
        ("Bills paid on time", '=IFERROR(COUNTIF(BillStatus,"Paid")/COUNTA(BillStatus),0)'),
        ("Emergency fund", "=IFERROR(MIN(EFSaved/EFTarget,1),0)"),
        ("Savings goals", "=IFERROR(AVERAGE(GoalPct),0)"),
        ("Sinking funds ready", "=IFERROR(MIN(SinkTotal/1650,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"On track","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    b = wb["Monthly Budget"]
    d = DoughnutChart(); d.title = "Spending by Category (actual)"; d.height = 7.4; d.width = 8.6
    d.add_data(Reference(b, min_col=5, min_row=7, max_row=6 + len(BUDGET)), titles_from_data=False)
    d.set_categories(Reference(b, min_col=2, min_row=7, max_row=6 + len(BUDGET))); d.dataLabels = no_labels()
    ws.add_chart(d, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Budget & Money Command Center™ — give every dollar a job. Edit income & categories anytime.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_income(wb); build_budget(wb)
    build_bills(wb); build_expenses(wb); build_savegoals(wb); build_sinking(wb)
    build_debt(wb); build_networth(wb); build_subs(wb); build_yearview(wb)
    build_nospend(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Income", "Monthly Budget", "Bills", "Expense Log",
             "Savings Goals", "Sinking Funds", "Debt Snapshot", "Net Worth", "Subscriptions",
             "Year View", "No-Spend", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Budget_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
