"""Build the Wedding Command Center (WCC) Excel workbook — flagship product.

32 sheets · luxury wedding operating system.

Run: python3 build_xlsx.py
Outputs: ../Wedding_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
)
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens (+ luxury extensions for the flagship)
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"      # Deep forest green
PRIMARY_DK = "133A35"   # Deeper green for layered headers
ACCENT = "937356"       # Rich gold
GOLD_LT = "C9A86A"      # Lighter gold for dividers/accents
SURFACE = "E5D3BA"      # Light tan
HIGHLIGHT = "75E6C1"    # Mint
BG = "FFFFFF"
TEXT = "333333"
SUCCESS = "75E6C1"
WARNING = "937356"
DANGER = "C94C4C"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"
BLUSH = "F3E4DD"        # Soft blush for luxury accent rows
IVORY = "FBF8F2"

THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD_BOTTOM = Border(bottom=Side(style="medium", color=GOLD_LT))


# ---------------------------------------------------------------------------
# Dropdown vocab
# ---------------------------------------------------------------------------
BUDGET_CATEGORIES = [
    "Venue", "Photography", "Videography", "Florist", "Cake", "Decor",
    "Dress", "Suit", "Jewelry", "Hair", "Makeup", "Invitations",
    "Transportation", "Accommodation", "Entertainment", "Music",
    "Catering", "Bar", "Flowers", "Rentals", "Miscellaneous",
]
VENDOR_TYPES = [
    "Venue", "Photographer", "Videographer", "DJ", "Band", "Florist",
    "Cake", "Officiant", "Transportation", "Hair", "Makeup", "Decor",
    "Rental Company", "Planner", "Caterer", "Stationery",
]
STATUSES = ["Not Started", "In Progress", "Booked", "Complete", "Overdue"]
RSVP_STATUS = ["Pending", "Accepted", "Declined"]
MEAL_TYPES = ["Beef", "Chicken", "Fish", "Vegetarian", "Vegan", "Kids", "None"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]
WEDDING_STYLES = ["Classic", "Modern", "Rustic", "Boho", "Glam",
                  "Garden", "Beach", "Vintage", "Minimalist", "Destination"]
CONTRACT_STATUS = ["Signed", "Pending", "Sent", "Not Started"]
PAYMENT_METHODS = ["Card", "Bank Transfer", "Check", "Cash", "PayPal", "Venmo"]
RELATIONSHIPS = ["Family", "Friend", "Work", "Partner Family", "Plus One", "Vendor"]
TIMELINE_PHASES = ["18 Months", "12 Months", "9 Months", "6 Months",
                   "3 Months", "1 Month", "2 Weeks", "Wedding Week",
                   "Wedding Day", "Post Wedding"]


# ===========================================================================
# Style registration
# ===========================================================================
def register_styles(wb: Workbook) -> None:
    def font(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    styles = {
        "title": NamedStyle(
            name="title",
            font=font(24, True, "FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center", indent=2),
        ),
        "subtitle": NamedStyle(
            name="subtitle",
            font=font(11, False, "E5D3BA", italic=True),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center", indent=2),
        ),
        "section": NamedStyle(
            name="section",
            font=font(12, True, PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "section_gold": NamedStyle(
            name="section_gold",
            font=font(12, True, ACCENT),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "th": NamedStyle(
            name="th",
            font=font(11, True, "FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=BOX,
        ),
        "td": NamedStyle(
            name="td",
            font=font(11, False, TEXT),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=BOX,
        ),
        "td_left": NamedStyle(
            name="td_left",
            font=font(11, False, TEXT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
            border=BOX,
        ),
        "kpi_label": NamedStyle(
            name="kpi_label",
            font=font(10, True, ACCENT),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
        ),
        "kpi_value": NamedStyle(
            name="kpi_value",
            font=font(20, True, PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
        ),
        "kpi_money": NamedStyle(
            name="kpi_money",
            font=font(20, True, PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
            number_format='"$"#,##0',
        ),
        "kpi_pct": NamedStyle(
            name="kpi_pct",
            font=font(20, True, PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
            number_format="0%",
        ),
        "input": NamedStyle(
            name="input",
            font=font(11, True, PRIMARY),
            fill=PatternFill("solid", fgColor=SURFACE),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=BOX,
        ),
        "field_label": NamedStyle(
            name="field_label",
            font=font(10, True, ACCENT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=BOX,
            fill=PatternFill("solid", fgColor=SOFT_BG),
        ),
        "field_value": NamedStyle(
            name="field_value",
            font=font(11, False, TEXT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=BOX,
        ),
        "monogram": NamedStyle(
            name="monogram",
            font=font(40, True, ACCENT),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=PRIMARY),
        ),
        "image_ph": NamedStyle(
            name="image_ph",
            font=font(11, False, ACCENT, italic=True),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            fill=PatternFill("solid", fgColor=BLUSH),
            border=Border(
                left=Side(style="dashed", color=GOLD_LT),
                right=Side(style="dashed", color=GOLD_LT),
                top=Side(style="dashed", color=GOLD_LT),
                bottom=Side(style="dashed", color=GOLD_LT),
            ),
        ),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def merge_set(ws, rng, value, style_name):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.style = style_name
    return cell


def luxe_header(ws, last_col: str, title: str, subtitle: str) -> None:
    """Two-row luxury header with a gold divider beneath."""
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 22
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    # Gold divider row 3
    ws.row_dimensions[3].height = 6
    last_idx = ord(last_col) - 64 if len(last_col) == 1 else None
    if last_idx is None:
        # handle 2-letter columns
        from openpyxl.utils import column_index_from_string
        last_idx = column_index_from_string(last_col)
    for c in range(1, last_idx + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng: str, list_name: str) -> None:
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def clean_labels(pct=False, val=False):
    """Concise data labels: never show series/category text (that collides on
    multi-slice pies). Percentages or values only, or nothing at all."""
    dl = DataLabelList()
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    dl.showVal = val
    dl.showPercent = pct
    return dl


def no_labels():
    """Explicitly suppress all data labels (legend carries the meaning).
    Used on the many-slice budget donuts where any on-chart text overlaps."""
    return clean_labels(pct=False, val=False)


def table_headers(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None,
               ints=None, dates=None, pcts=None, start_col=1):
    text_left = text_left or set()
    money = money or set()
    ints = ints or set()
    dates = dates or set()
    pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 == 1 else BG)
            if c in money:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


# ===========================================================================
# Sheet 32: Settings (first so names exist)
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 22, 4, 22, 22, 22, 22, 22, 22])
    luxe_header(ws, "J", "⚙  SETTINGS", "Edit once — drives the entire planning system.")

    merge_set(ws, "B5:C5", "WEDDING CONTROLS", "section")
    controls = [
        ("Wedding Date",        dt.date(2027, 6, 12), "mm/dd/yyyy",    "WeddingDate"),
        ("Total Budget",        45000,                '"$"#,##0',      "TotalBudget"),
        ("Estimated Guests",    140,                  "0",             "EstGuests"),
        ("Number of Tables",    16,                   "0",             "NumTables"),
        ("Seats Per Table",     10,                   "0",             "SeatsPerTable"),
        ("Monthly Savings",     1500,                 '"$"#,##0',      "MonthlySavings"),
        ("Tax Rate",            0.08,                 "0.00%",         "TaxRate"),
        ("Today (auto)",        "=TODAY()",           "mm/dd/yyyy",    "TodayDate"),
    ]
    for i, (label, default, fmt, name) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=label).style = "field_label"
        cell = ws.cell(row=r, column=3, value=default)
        cell.style = "input"
        if fmt:
            cell.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Settings!$C${r}")

    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    lists = [
        ("E", "Budget Categories", BUDGET_CATEGORIES, "BudgetCatList"),
        ("F", "Vendor Types",      VENDOR_TYPES,      "VendorTypeList"),
        ("G", "Statuses",          STATUSES,          "StatusList"),
        ("H", "Meal Types",        MEAL_TYPES,        "MealList"),
        ("I", "Priority",          PRIORITIES,        "PriorityList"),
        ("J", "Wedding Styles",    WEDDING_STYLES,    "StyleList"),
    ]
    for col, h, data, name in lists:
        ci = ord(col) - 64
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, val in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=val).style = "td_left"
        wb.defined_names[name] = DefinedName(
            name, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")

    # Second block of lists
    start2 = 26
    lists2 = [
        ("E", "RSVP Status",      RSVP_STATUS,      "RsvpList"),
        ("F", "Contract Status",  CONTRACT_STATUS,  "ContractList"),
        ("G", "Payment Methods",  PAYMENT_METHODS,  "PayMethodList"),
        ("H", "Relationships",    RELATIONSHIPS,    "RelationList"),
        ("I", "Yes / No",         YESNO,            "YesNoList"),
        ("J", "Timeline Phases",  TIMELINE_PHASES,  "PhaseList"),
    ]
    for col, h, data, name in lists2:
        ci = ord(col) - 64
        ws.cell(row=start2, column=ci, value=h).style = "th"
        for ri, val in enumerate(data):
            ws.cell(row=start2 + 1 + ri, column=ci, value=val).style = "td_left"
        wb.defined_names[name] = DefinedName(
            name, attr_text=f"Settings!${col}${start2+1}:${col}${start2 + len(data)}")


# ===========================================================================
# Sheet 2: Couple Profile
# ===========================================================================
def build_couple(wb):
    ws = wb.create_sheet("Couple")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 30, 6, 26, 30, 2])
    luxe_header(ws, "G", "💍  COUPLE PROFILE", "The heart of your celebration.")

    # Monogram block
    ws.row_dimensions[5].height = 70
    merge_set(ws, "B5:C6", "A  &  J", "monogram")
    merge_set(ws, "E5:F5", "Alexandra  &  James", "section")
    merge_set(ws, "E6:F6", "“Forever starts here.”", "section_gold")

    merge_set(ws, "B8:G8", "WEDDING DETAILS", "section")
    details = [
        ("Bride", "Alexandra Rose Bennett", "Groom", "James Michael Carter"),
        ("Wedding Date", "=TEXT(WeddingDate,\"dddd, mmmm d, yyyy\")", "Wedding Style", "Garden Glam"),
        ("Venue", "The Hartwell Estate", "Color Palette", "Emerald · Gold · Ivory"),
        ("Theme", "Enchanted Garden", "Estimated Guests", "=EstGuests"),
        ("Ceremony Time", "4:30 PM", "Reception Time", "6:00 PM"),
        ("Total Budget", "=TotalBudget", "Days Until", '=MAX(WeddingDate-TODAY(),0)&" days"'),
    ]
    r = 9
    for l1, v1, l2, v2 in details:
        ws.cell(row=r, column=2, value=l1).style = "field_label"
        c1 = ws.cell(row=r, column=3, value=v1); c1.style = "field_value"
        ws.cell(row=r, column=5, value=l2).style = "field_label"
        c2 = ws.cell(row=r, column=6, value=v2); c2.style = "field_value"
        if l1 == "Total Budget":
            c1.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 26
        r += 1

    r += 1
    merge_set(ws, f"B{r}:G{r}", "PLANNER & KEY CONTACTS", "section_gold")
    r += 1
    contacts = [
        ("Wedding Planner", "Sophia Lane — Lane & Co. Events", "Planner Phone", "(555) 200-8841"),
        ("Day-of Coordinator", "Marcus Reed", "Coordinator Phone", "(555) 200-2299"),
        ("Maid of Honor", "Emily Bennett", "Best Man", "Daniel Carter"),
        ("Emergency Contact", "Mom — Linda Bennett", "Emergency Phone", "(555) 412-7833"),
    ]
    for l1, v1, l2, v2 in contacts:
        ws.cell(row=r, column=2, value=l1).style = "field_label"
        ws.cell(row=r, column=3, value=v1).style = "field_value"
        ws.cell(row=r, column=5, value=l2).style = "field_label"
        ws.cell(row=r, column=6, value=v2).style = "field_value"
        ws.row_dimensions[r].height = 26
        r += 1


# ===========================================================================
# Sheet 3: Master Timeline
# ===========================================================================
def build_timeline(wb):
    ws = wb.create_sheet("Timeline")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [16, 40, 16, 14, 16, 14, 28])
    luxe_header(ws, "G", "🗓  MASTER WEDDING TIMELINE",
                "18-month plan — auto-anchored to your wedding date.")
    table_headers(ws, 4, ["Phase", "Milestone Task", "Owner",
                          "Target Date", "Status", "Done", "Notes"])

    # (phase, months_before, task, owner)
    tasks = [
        ("18 Months", 18, "Set the date & overall vision", "Couple"),
        ("18 Months", 18, "Determine total budget", "Couple"),
        ("18 Months", 18, "Draft preliminary guest list", "Couple"),
        ("12 Months", 12, "Book ceremony & reception venue", "Couple"),
        ("12 Months", 12, "Hire wedding planner", "Couple"),
        ("12 Months", 12, "Book photographer & videographer", "Couple"),
        ("9 Months", 9, "Choose wedding party", "Couple"),
        ("9 Months", 9, "Order wedding dress", "Bride"),
        ("9 Months", 9, "Book caterer & florist", "Planner"),
        ("6 Months", 6, "Send save-the-dates", "Couple"),
        ("6 Months", 6, "Book entertainment / DJ / band", "Planner"),
        ("6 Months", 6, "Reserve room block & transportation", "Couple"),
        ("3 Months", 3, "Order invitations", "Bride"),
        ("3 Months", 3, "Finalize menu & cake tasting", "Couple"),
        ("3 Months", 3, "Purchase wedding bands", "Couple"),
        ("1 Month", 1, "Mail invitations / track RSVPs", "Couple"),
        ("1 Month", 1, "Final dress fitting", "Bride"),
        ("1 Month", 1, "Confirm all vendor details", "Planner"),
        ("2 Weeks", 0.5, "Finalize seating chart", "Couple"),
        ("2 Weeks", 0.5, "Confirm final headcount with caterer", "Planner"),
        ("Wedding Week", 0.2, "Pick up attire & rings", "Couple"),
        ("Wedding Week", 0.2, "Rehearsal & rehearsal dinner", "Couple"),
        ("Wedding Day", 0, "Execute day-of timeline", "Coordinator"),
        ("Post Wedding", -1, "Send thank-you cards", "Couple"),
        ("Post Wedding", -1, "Review & tip vendors", "Couple"),
    ]
    start = 5
    end = start + 80 - 1
    for i, (phase, months, task, owner) in enumerate(tasks):
        r = start + i
        ws.cell(row=r, column=1, value=phase)
        ws.cell(row=r, column=2, value=task)
        ws.cell(row=r, column=3, value=owner)
        # Target date = WeddingDate - months*30
        ws.cell(row=r, column=4, value=f"=WeddingDate-{int(months*30)}")
        ws.cell(row=r, column=6, value="No")

    for r in range(start, end + 1):
        # Status auto from Done + date
        ws.cell(row=r, column=5,
                value=(f'=IF(B{r}="","",IF(F{r}="Yes","Complete",'
                       f'IF(D{r}<TODAY(),"Overdue","In Progress")))'))

    style_rows(ws, start, end, 7, text_left={2, 7}, dates={4})
    add_dv(ws, f"A{start}:A{end}", "PhaseList")
    add_dv(ws, f"F{start}:F{end}", "YesNoList")

    # Conditional formats
    ws.conditional_formatting.add(
        f"A{start}:G{end}",
        FormulaRule(formula=[f'$F{start}="Yes"'],
                    fill=fill("E3F8EF")))
    ws.conditional_formatting.add(
        f"A{start}:G{end}",
        FormulaRule(formula=[f'AND($B{start}<>"",$D{start}<TODAY(),$F{start}<>"Yes")'],
                    fill=fill("FBE6E6")))
    ws.conditional_formatting.add(
        f"A{start}:G{end}",
        FormulaRule(formula=[f'AND($D{start}<>"",$D{start}-TODAY()>=0,$D{start}-TODAY()<=30,$F{start}<>"Yes")'],
                    fill=fill("FBF0E2")))

    wb.defined_names["TLTask"] = DefinedName("TLTask", attr_text=f"Timeline!$B${start}:$B${end}")
    wb.defined_names["TLDone"] = DefinedName("TLDone", attr_text=f"Timeline!$F${start}:$F${end}")
    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 4: Master Checklist
# ===========================================================================
def build_checklist(wb):
    ws = wb.create_sheet("Checklist")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 42, 14, 12, 14, 10, 16, 26])
    luxe_header(ws, "H", "✅  MASTER CHECKLIST",
                "Every task in one place — completion auto-tallies to the dashboard.")
    table_headers(ws, 4, ["Category", "Task", "Owner", "Priority",
                          "Due Date", "Done", "Status", "Notes"])

    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)
    tasks = [
        ("Venue", "Sign venue contract", "Couple", "High", d(-60), "Yes"),
        ("Venue", "Confirm ceremony rehearsal slot", "Planner", "Medium", d(30), "No"),
        ("Attire", "Final dress fitting", "Bride", "High", d(45), "No"),
        ("Attire", "Groom suit fitting", "Groom", "Medium", d(50), "No"),
        ("Stationery", "Order invitations", "Bride", "High", d(-20), "Yes"),
        ("Stationery", "Mail invitations", "Couple", "High", d(10), "No"),
        ("Catering", "Finalize menu", "Couple", "High", d(25), "No"),
        ("Catering", "Confirm final headcount", "Planner", "High", d(40), "No"),
        ("Flowers", "Approve floral proposal", "Bride", "Medium", d(35), "No"),
        ("Photography", "Build shot list", "Couple", "Medium", d(20), "No"),
        ("Music", "Submit DJ playlist", "Couple", "Low", d(38), "No"),
        ("Beauty", "Hair & makeup trial", "Bride", "Medium", d(15), "No"),
        ("Transportation", "Book guest shuttle", "Couple", "Medium", d(28), "No"),
        ("Rings", "Purchase wedding bands", "Couple", "High", d(22), "No"),
        ("Registry", "Finalize gift registry", "Couple", "Low", d(-5), "Yes"),
        ("Guests", "Finalize seating chart", "Couple", "High", d(42), "No"),
        ("Vendors", "Confirm vendor arrival times", "Planner", "High", d(44), "No"),
        ("Day-of", "Assemble emergency kit", "MOH", "Medium", d(43), "No"),
    ]
    start = 5
    end = start + 120 - 1
    for i, (cat, task, owner, prio, due, done) in enumerate(tasks):
        r = start + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=task)
        ws.cell(row=r, column=3, value=owner)
        ws.cell(row=r, column=4, value=prio)
        ws.cell(row=r, column=5, value=due)
        ws.cell(row=r, column=6, value=done)
    for r in range(start, end + 1):
        ws.cell(row=r, column=7,
                value=(f'=IF(B{r}="","",IF(F{r}="Yes","Complete",'
                       f'IF(E{r}<TODAY(),"Overdue","In Progress")))'))
    style_rows(ws, start, end, 8, text_left={2, 8}, dates={5})
    add_dv(ws, f"D{start}:D{end}", "PriorityList")
    add_dv(ws, f"F{start}:F{end}", "YesNoList")

    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'$F{start}="Yes"'], fill=fill("E3F8EF")))
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'AND($B{start}<>"",$E{start}<TODAY(),$F{start}<>"Yes")'],
                    fill=fill("FBE6E6")))

    wb.defined_names["ChkTask"] = DefinedName("ChkTask", attr_text=f"Checklist!$B${start}:$B${end}")
    wb.defined_names["ChkDone"] = DefinedName("ChkDone", attr_text=f"Checklist!$F${start}:$F${end}")
    wb.defined_names["ChkDue"] = DefinedName("ChkDue", attr_text=f"Checklist!$E${start}:$E${end}")
    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 5: Budget Command Center
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 16, 16, 16, 16, 16, 4, 24, 18])
    luxe_header(ws, "I", "💰  BUDGET COMMAND CENTER",
                "Plan, spend, and track deposits across every category.")
    table_headers(ws, 4, ["Category", "Budget", "Actual", "Deposit",
                          "Paid", "Remaining"])

    planned = {
        "Venue": 12000, "Photography": 4500, "Videography": 3000, "Florist": 3500,
        "Cake": 900, "Decor": 2500, "Dress": 2800, "Suit": 800, "Jewelry": 1500,
        "Hair": 400, "Makeup": 400, "Invitations": 700, "Transportation": 1200,
        "Accommodation": 1800, "Entertainment": 2500, "Music": 1500, "Catering": 9000,
        "Bar": 3500, "Flowers": 1200, "Rentals": 2000, "Miscellaneous": 1000,
    }
    actuals = {
        "Venue": 12000, "Photography": 4500, "Videography": 3000, "Florist": 3300,
        "Cake": 850, "Decor": 2100, "Dress": 2800, "Suit": 0, "Jewelry": 1400,
        "Hair": 0, "Makeup": 0, "Invitations": 680, "Transportation": 0,
        "Accommodation": 1800, "Entertainment": 2500, "Music": 1500, "Catering": 4500,
        "Bar": 0, "Flowers": 0, "Rentals": 1000, "Miscellaneous": 300,
    }
    deposits = {
        "Venue": 6000, "Photography": 1500, "Videography": 1000, "Florist": 1000,
        "Cake": 200, "Decor": 500, "Dress": 1400, "Catering": 2000, "Entertainment": 800,
        "Music": 500, "Accommodation": 900, "Jewelry": 700, "Rentals": 500,
    }
    start = 5
    end = start + len(BUDGET_CATEGORIES) - 1
    for i, cat in enumerate(BUDGET_CATEGORIES):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cb = ws.cell(row=r, column=2, value=planned.get(cat, 0)); cb.style = "input"; cb.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=actuals.get(cat, 0)); ca.style = "td"; ca.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=4, value=deposits.get(cat, 0)); cd.style = "td"; cd.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=D{r}"); cp.style = "td"; cp.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=6, value=f"=C{r}-E{r}"); cr.style = "td"; cr.number_format = '"$"#,##0'
        ws.row_dimensions[r].height = 21
        if i % 2 == 1:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)

    total_row = end + 1
    ws.cell(row=total_row, column=1, value="TOTAL").style = "th"
    for col in range(2, 7):
        L = get_column_letter(col)
        c = ws.cell(row=total_row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE)
        c.number_format = '"$"#,##0'

    wb.defined_names["BudCat"] = DefinedName("BudCat", attr_text=f"Budget!$A${start}:$A${end}")
    wb.defined_names["BudPlanned"] = DefinedName("BudPlanned", attr_text=f"Budget!$B${start}:$B${end}")
    wb.defined_names["BudActual"] = DefinedName("BudActual", attr_text=f"Budget!$C${start}:$C${end}")
    wb.defined_names["BudPaid"] = DefinedName("BudPaid", attr_text=f"Budget!$E${start}:$E${end}")
    wb.defined_names["BudTotalPlanned"] = DefinedName("BudTotalPlanned", attr_text=f"Budget!$B${total_row}")
    wb.defined_names["BudTotalActual"] = DefinedName("BudTotalActual", attr_text=f"Budget!$C${total_row}")
    wb.defined_names["BudTotalPaid"] = DefinedName("BudTotalPaid", attr_text=f"Budget!$E${total_row}")

    # KPI sidebar
    merge_set(ws, "H4:I4", "BUDGET HEALTH", "section_gold")
    kpis = [
        ("Total Budget",   "=TotalBudget",                    '"$"#,##0'),
        ("Total Estimated","=BudTotalActual",                 '"$"#,##0'),
        ("Total Paid",     "=BudTotalPaid",                   '"$"#,##0'),
        ("Remaining",      "=TotalBudget-BudTotalPaid",       '"$"#,##0'),
        ("% of Budget",    "=IFERROR(BudTotalActual/TotalBudget,0)", "0%"),
        ("Over / Under",   "=TotalBudget-BudTotalActual",     '"$"#,##0;[Red]-"$"#,##0'),
    ]
    for i, (lab, fml, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=8, value=lab).style = "field_label"
        c = ws.cell(row=r, column=9, value=fml); c.style = "field_value"
        c.number_format = fmt; c.font = Font(bold=True, color=PRIMARY)

    # Charts
    pie = DoughnutChart(); pie.title = "Category Spending"; pie.height = 9; pie.width = 14
    pie.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    pie.dataLabels = no_labels()   # 21 slices — legend carries labels
    ws.add_chart(pie, "A28")
    bar = BarChart(); bar.type = "col"; bar.title = "Budget vs Actual"; bar.height = 9; bar.width = 18
    bar.add_data(Reference(ws, min_col=2, min_row=4, max_row=end), titles_from_data=True)
    bar.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    ws.add_chart(bar, "A47")


# ===========================================================================
# Sheet 6: Payment Schedule
# ===========================================================================
def build_payments(wb):
    ws = wb.create_sheet("Payments")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 14, 14, 14, 14, 10, 16, 16, 12])
    luxe_header(ws, "I", "💳  PAYMENT SCHEDULE",
                "Deposits, balances, due dates — with automatic late alerts.")
    table_headers(ws, 4, ["Vendor", "Total", "Deposit", "Balance Due",
                          "Due Date", "Paid", "Method", "Invoice #", "Days Out"])
    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)
    rows = [
        ("The Hartwell Estate (Venue)", 12000, 6000, 6000, d(30), "No", "Bank Transfer", "HE-1042"),
        ("Lumiere Photography", 4500, 1500, 3000, d(45), "No", "Card", "LP-2231"),
        ("Evergreen Films", 3000, 1000, 2000, d(50), "No", "Card", "EF-0099"),
        ("Petal & Stem Florals", 3300, 1000, 2300, d(20), "No", "Card", "PS-7781"),
        ("Sweet Layers Cake Co.", 850, 200, 650, d(35), "No", "Venmo", "SL-3310"),
        ("Grand Catering Group", 9000, 2000, 7000, d(40), "No", "Bank Transfer", "GC-5521"),
        ("Skyline Entertainment (DJ)", 2500, 800, 1700, d(-3), "No", "PayPal", "SE-1188"),
        ("Luxe Linens & Rentals", 2000, 500, 1500, d(28), "No", "Card", "LL-4402"),
        ("Belle Bridal (Dress)", 2800, 1400, 1400, d(15), "No", "Card", "BB-9931"),
        ("Radiant Beauty Team", 800, 0, 800, d(44), "No", "Cash", "RB-2020"),
    ]
    start = 5
    end = start + 40 - 1
    for i, row in enumerate(rows):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    for r in range(start, end + 1):
        ws.cell(row=r, column=9, value=f'=IF(E{r}="","",E{r}-TODAY())')
    style_rows(ws, start, end, 9, text_left={1}, money={2, 3, 4}, dates={5})
    for r in range(start, end + 1):
        ws.cell(row=r, column=9).number_format = "0;[Red]-0"
    add_dv(ws, f"F{start}:F{end}", "YesNoList")
    add_dv(ws, f"G{start}:G{end}", "PayMethodList")
    # Late alert: due date passed and not paid
    ws.conditional_formatting.add(
        f"A{start}:I{end}",
        FormulaRule(formula=[f'AND($E{start}<>"",$E{start}<TODAY(),$F{start}<>"Yes")'],
                    fill=fill("FBE6E6")))
    ws.conditional_formatting.add(
        f"A{start}:I{end}",
        FormulaRule(formula=[f'AND($E{start}<>"",$E{start}-TODAY()>=0,$E{start}-TODAY()<=14,$F{start}<>"Yes")'],
                    fill=fill("FBF0E2")))
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill("E3F8EF")))
    wb.defined_names["PayBalance"] = DefinedName("PayBalance", attr_text=f"Payments!$D${start}:$D${end}")
    wb.defined_names["PayPaid"] = DefinedName("PayPaid", attr_text=f"Payments!$F${start}:$F${end}")
    wb.defined_names["PayVendor"] = DefinedName("PayVendor", attr_text=f"Payments!$A${start}:$A${end}")
    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 7: Savings Planner
# ===========================================================================
def build_savings(wb):
    ws = wb.create_sheet("Savings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 18, 18, 18, 2])
    luxe_header(ws, "G", "🏦  SAVINGS PLANNER",
                "Track progress toward fully funding the celebration.")
    merge_set(ws, "B5:F5", "SAVINGS SUMMARY", "section")
    summary = [
        ("Wedding Goal (Total Budget)", "=TotalBudget", '"$"#,##0'),
        ("Monthly Contribution", "=MonthlySavings", '"$"#,##0'),
        ("Months Until Wedding", "=MAX(ROUND((WeddingDate-TODAY())/30,0),0)", "0"),
        ("Projected Saved by Date", "=MonthlySavings*MAX(ROUND((WeddingDate-TODAY())/30,0),0)", '"$"#,##0'),
        ("Funding Gap / Surplus", "=MonthlySavings*MAX(ROUND((WeddingDate-TODAY())/30,0),0)-TotalBudget", '"$"#,##0;[Red]-"$"#,##0'),
    ]
    r = 6
    for lab, fml, fmt in summary:
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "input"; c.number_format = fmt
        ws.row_dimensions[r].height = 24
        r += 1

    # Monthly contribution log
    merge_set(ws, "B13:F13", "MONTHLY CONTRIBUTION LOG", "section_gold")
    table_headers(ws, 14, ["Month", "Planned", "Actual", "Cumulative", "% to Goal"], start_col=2)
    start = 15
    months = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
              "Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    actual_saved = [1500, 1500, 2000, 1500, 1800, 1500,
                    1500, 1600, 1500, 2000, 1500, 1500]
    end = start + len(months) - 1
    for i, m in enumerate(months):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cp = ws.cell(row=r, column=3, value="=MonthlySavings"); cp.style = "td"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=actual_saved[i]); ca.style = "input"; ca.number_format = '"$"#,##0'
        cc = ws.cell(row=r, column=5, value=f"=SUM($D${start}:D{r})"); cc.style = "td"; cc.number_format = '"$"#,##0'
        cg = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/TotalBudget,0)"); cg.style = "td"; cg.number_format = "0%"
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color=HIGHLIGHT, showValue=True))
    # Progress line chart
    line = LineChart(); line.title = "Savings Progress to Goal"; line.height = 9; line.width = 16
    line.add_data(Reference(ws, min_col=5, min_row=14, max_row=end), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    ws.add_chart(line, "H14")


# ===========================================================================
# Sheet 8: Vendor CRM
# ===========================================================================
def build_vendor_crm(wb):
    ws = wb.create_sheet("Vendors")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 24, 22, 16, 22, 14, 14, 14, 12, 10, 22])
    luxe_header(ws, "K", "📇  VENDOR CRM",
                "Every vendor, contact, quote, and contract status in one record.")
    table_headers(ws, 4, ["Type", "Vendor", "Email", "Phone", "Website",
                          "Quote", "Deposit", "Balance", "Contract", "Rating", "Notes"])
    rows = [
        ("Venue", "The Hartwell Estate", "events@hartwell.com", "(555) 200-1000", "hartwellestate.com", 12000, 6000, 6000, "Signed", 5, "Includes tables & chairs"),
        ("Photographer", "Lumiere Photography", "hello@lumiere.co", "(555) 200-2231", "lumiere.co", 4500, 1500, 3000, "Signed", 5, "8hr + engagement shoot"),
        ("Videographer", "Evergreen Films", "book@evergreen.tv", "(555) 200-0099", "evergreen.tv", 3000, 1000, 2000, "Signed", 4, "Highlight + full film"),
        ("Florist", "Petal & Stem", "studio@petalstem.com", "(555) 200-7781", "petalstem.com", 3300, 1000, 2300, "Pending", 5, "Emerald + ivory palette"),
        ("Cake", "Sweet Layers", "order@sweetlayers.com", "(555) 200-3310", "sweetlayers.com", 850, 200, 650, "Signed", 4, "3-tier, lemon + vanilla"),
        ("Caterer", "Grand Catering Group", "events@grandcater.com", "(555) 200-5521", "grandcater.com", 9000, 2000, 7000, "Signed", 5, "Plated dinner, 140pax"),
        ("DJ", "Skyline Entertainment", "info@skylinedj.com", "(555) 200-1188", "skylinedj.com", 2500, 800, 1700, "Signed", 4, "Ceremony + reception"),
        ("Officiant", "Rev. Thomas Hale", "thale@example.com", "(555) 200-6644", "—", 500, 0, 500, "Pending", 5, "Custom ceremony script"),
        ("Hair", "Radiant Beauty", "team@radiant.com", "(555) 200-2020", "radiantbeauty.com", 400, 0, 400, "Pending", 4, "Trial booked"),
        ("Makeup", "Glow Artistry", "book@glow.com", "(555) 200-3030", "glowartistry.com", 400, 0, 400, "Pending", 5, "Airbrush"),
        ("Rental Company", "Luxe Linens & Rentals", "rentals@luxe.com", "(555) 200-4402", "luxerentals.com", 2000, 500, 1500, "Signed", 4, "Linens, lounge furniture"),
        ("Planner", "Lane & Co. Events", "sophia@laneco.com", "(555) 200-8841", "laneco.com", 5000, 2500, 2500, "Signed", 5, "Full-service planning"),
    ]
    start = 5
    end = start + 30 - 1
    for i, row in enumerate(rows):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, 11, text_left={2, 3, 5, 11}, money={6, 7, 8}, ints={10})
    add_dv(ws, f"A{start}:A{end}", "VendorTypeList")
    add_dv(ws, f"I{start}:I{end}", "ContractList")
    ws.conditional_formatting.add(
        f"I{start}:I{end}",
        CellIsRule(operator="equal", formula=['"Signed"'], fill=fill("E3F8EF")))
    ws.conditional_formatting.add(
        f"I{start}:I{end}",
        CellIsRule(operator="equal", formula=['"Pending"'], fill=fill("FBF0E2")))
    ws.conditional_formatting.add(
        f"J{start}:J{end}",
        ColorScaleRule(start_type="num", start_value=1, start_color="FFFBE6E6",
                       mid_type="num", mid_value=3, mid_color="FFFFF3CD",
                       end_type="num", end_value=5, end_color="FF75E6C1"))
    wb.defined_names["VenName"] = DefinedName("VenName", attr_text=f"Vendors!$B${start}:$B${end}")
    wb.defined_names["VenContract"] = DefinedName("VenContract", attr_text=f"Vendors!$I${start}:$I${end}")
    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 9: Vendor Comparison
# ===========================================================================
def build_comparison(wb):
    ws = wb.create_sheet("Comparison")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 22, 14, 12, 14, 22, 22, 14, 12])
    luxe_header(ws, "I", "⚖  VENDOR COMPARISON",
                "Score candidates side by side — auto-ranked by weighted total.")
    table_headers(ws, 4, ["Category", "Vendor", "Price", "Reviews",
                          "Availability", "Package", "Pros / Cons", "Score", "Rank"])
    rows = [
        ("Photographer", "Lumiere Photography", 4500, 4.9, "Available", "8hr + engagement", "Best portfolio / premium", 95),
        ("Photographer", "Bright Frame Studio", 3800, 4.6, "Available", "7hr coverage", "Great value / no 2nd shooter", 84),
        ("Photographer", "Aperture Co.", 5200, 4.8, "Waitlist", "10hr + album", "Luxe album / over budget", 80),
        ("Florist", "Petal & Stem", 3300, 4.9, "Available", "Full design", "Stunning / books fast", 93),
        ("Florist", "Wildwood Blooms", 2700, 4.4, "Available", "Standard pkg", "Affordable / smaller team", 79),
        ("DJ", "Skyline Entertainment", 2500, 4.7, "Available", "Ceremony+reception", "Great energy / add lighting", 88),
        ("DJ", "Beat Collective", 2100, 4.3, "Available", "Reception only", "Budget / no ceremony", 74),
        ("Caterer", "Grand Catering Group", 9000, 4.8, "Available", "Plated 140pax", "Excellent / premium", 90),
        ("Caterer", "Harvest Table", 7600, 4.5, "Available", "Buffet 140pax", "Good value / buffet only", 82),
    ]
    start = 5
    end = start + 30 - 1
    for i, row in enumerate(rows):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    for r in range(start, end + 1):
        # Rank within category by score
        ws.cell(row=r, column=9,
                value=(f'=IF(H{r}="","",SUMPRODUCT(($A${start}:$A${end}=A{r})*'
                       f'($H${start}:$H${end}>H{r}))+1)'))
    style_rows(ws, start, end, 9, text_left={2, 5, 6, 7}, money={3}, ints={8, 9})
    for r in range(start, end + 1):
        ws.cell(row=r, column=4).number_format = "0.0"
    ws.conditional_formatting.add(
        f"H{start}:H{end}",
        ColorScaleRule(start_type="num", start_value=60, start_color="FFFBE6E6",
                       mid_type="num", mid_value=80, mid_color="FFFFF3CD",
                       end_type="num", end_value=100, end_color="FF75E6C1"))
    # Highlight rank 1
    ws.conditional_formatting.add(
        f"A{start}:I{end}",
        FormulaRule(formula=[f'$I{start}=1'], fill=fill("E3F8EF")))
    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 10: Contract Tracker
# ===========================================================================
def build_contracts(wb):
    ws = wb.create_sheet("Contracts")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 16, 14, 16, 16, 28, 24])
    luxe_header(ws, "G", "📄  CONTRACT TRACKER",
                "Signed, pending, insurance, and where each document lives.")
    table_headers(ws, 4, ["Vendor", "Status", "Insurance", "Signed Date",
                          "Due Date", "Storage Location", "Notes"])
    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)
    rows = [
        ("The Hartwell Estate", "Signed", "Yes", d(-60), d(-65), "Drive / Contracts / Venue.pdf", "COI on file"),
        ("Lumiere Photography", "Signed", "Yes", d(-40), d(-45), "Drive / Contracts / Photo.pdf", ""),
        ("Evergreen Films", "Signed", "Yes", d(-38), d(-42), "Drive / Contracts / Video.pdf", ""),
        ("Petal & Stem", "Pending", "No", None, d(10), "—", "Awaiting signature"),
        ("Grand Catering Group", "Signed", "Yes", d(-30), d(-35), "Drive / Contracts / Catering.pdf", "Final count due 2wks"),
        ("Skyline Entertainment", "Signed", "Yes", d(-25), d(-30), "Drive / Contracts / DJ.pdf", ""),
        ("Rev. Thomas Hale", "Pending", "No", None, d(14), "—", "Reviewing script clause"),
        ("Luxe Linens & Rentals", "Signed", "Yes", d(-20), d(-25), "Drive / Contracts / Rentals.pdf", ""),
        ("Belle Bridal", "Sent", "No", None, d(7), "Email inbox", "Sent for e-sign"),
    ]
    start = 5
    end = start + 30 - 1
    for i, row in enumerate(rows):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, 7, text_left={1, 6, 7}, dates={4, 5})
    add_dv(ws, f"B{start}:B{end}", "ContractList")
    add_dv(ws, f"C{start}:C{end}", "YesNoList")
    ws.conditional_formatting.add(
        f"B{start}:B{end}",
        CellIsRule(operator="equal", formula=['"Signed"'], fill=fill("E3F8EF")))
    ws.conditional_formatting.add(
        f"B{start}:B{end}",
        CellIsRule(operator="equal", formula=['"Pending"'], fill=fill("FBE6E6")))
    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 11: Guest CRM
# ===========================================================================
def build_guest_crm(wb):
    ws = wb.create_sheet("Guests")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 16, 26, 22, 14, 10, 14, 12, 14, 10, 12, 14, 12, 12, 20])
    luxe_header(ws, "O", "👥  GUEST CRM",
                "The master guest database — RSVP, meals, gifts, logistics.")
    table_headers(ws, 4, ["Guest", "Family", "Address", "Email", "Phone",
                          "Invited", "RSVP", "# Seats", "Meal", "Kids",
                          "Gift", "Thank-You", "Hotel", "Transport", "Notes"])
    rows = [
        ("Linda Bennett", "Bennett", "12 Rose Ln, Springfield", "linda@ex.com", "(555) 412-7833", "Yes", "Accepted", 2, "Beef", "No", "Yes", "No", "No", "No", "Mother of bride"),
        ("Robert Bennett", "Bennett", "12 Rose Ln, Springfield", "rob@ex.com", "(555) 412-7834", "Yes", "Accepted", 2, "Fish", "No", "Yes", "No", "No", "No", "Father of bride"),
        ("Daniel Carter", "Carter", "88 Oak St, Rivertown", "dan@ex.com", "(555) 901-2271", "Yes", "Accepted", 1, "Chicken", "No", "No", "No", "Yes", "Yes", "Best man"),
        ("Emily Bennett", "Bennett", "5 Lake Dr, Springfield", "emily@ex.com", "(555) 412-7799", "Yes", "Accepted", 1, "Vegetarian", "No", "Yes", "No", "No", "No", "Maid of honor"),
        ("Sophia & Marc Lane", "Lane", "240 Hill Rd, Metro", "sophia@ex.com", "(555) 200-8841", "Yes", "Accepted", 2, "Beef", "No", "No", "No", "Yes", "No", "Planner + guest"),
        ("Grandma Rose", "Bennett", "12 Rose Ln, Springfield", "—", "(555) 412-1900", "Yes", "Accepted", 1, "Fish", "No", "Yes", "No", "No", "Yes", "Needs accessible seating"),
        ("The Hartleys", "Hartley", "31 Pine Ave, Metro", "hartley@ex.com", "(555) 388-2210", "Yes", "Pending", 4, "None", "Yes", "No", "No", "No", "No", "Family of 4, 2 kids"),
        ("Aiden Brooks", "Brooks", "210 North Ave, Rivertown", "aiden@ex.com", "(555) 901-2272", "Yes", "Declined", 0, "None", "No", "No", "No", "No", "No", "Out of town"),
        ("Maya Rivera", "Rivera", "1500 River Rd, Metro", "maya@ex.com", "(555) 412-7000", "Yes", "Accepted", 2, "Chicken", "No", "No", "No", "Yes", "No", "College friend"),
        ("The Patels", "Patel", "77 Cedar Blvd, Metro", "patel@ex.com", "(555) 412-9032", "Yes", "Accepted", 3, "Vegan", "Yes", "No", "No", "Yes", "Yes", "1 child, vegan family"),
        ("Olivia Reed", "Reed", "9 Maple Ct, Springfield", "olivia@ex.com", "(555) 200-2244", "Yes", "Pending", 1, "None", "No", "No", "No", "No", "No", "Work friend"),
        ("James Sr. & Carol Carter", "Carter", "88 Oak St, Rivertown", "carol@ex.com", "(555) 901-0000", "Yes", "Accepted", 2, "Beef", "No", "Yes", "No", "Yes", "Yes", "Parents of groom"),
    ]
    start = 5
    end = start + 200 - 1
    for i, row in enumerate(rows):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, 15, text_left={1, 2, 3, 4, 15}, ints={8})
    add_dv(ws, f"F{start}:F{end}", "YesNoList")
    add_dv(ws, f"G{start}:G{end}", "RsvpList")
    add_dv(ws, f"I{start}:I{end}", "MealList")
    for col in ("J", "K", "L", "M", "N"):
        add_dv(ws, f"{col}{start}:{col}{end}", "YesNoList")
    ws.conditional_formatting.add(
        f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"Accepted"'], fill=fill("E3F8EF")))
    ws.conditional_formatting.add(
        f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"Declined"'], fill=fill("FBE6E6")))
    ws.conditional_formatting.add(
        f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"Pending"'], fill=fill("FBF0E2")))
    wb.defined_names["GuestName"] = DefinedName("GuestName", attr_text=f"Guests!$A${start}:$A${end}")
    wb.defined_names["GuestRsvp"] = DefinedName("GuestRsvp", attr_text=f"Guests!$G${start}:$G${end}")
    wb.defined_names["GuestSeats"] = DefinedName("GuestSeats", attr_text=f"Guests!$H${start}:$H${end}")
    wb.defined_names["GuestInvited"] = DefinedName("GuestInvited", attr_text=f"Guests!$F${start}:$F${end}")
    wb.defined_names["GuestMeal"] = DefinedName("GuestMeal", attr_text=f"Guests!$I${start}:$I${end}")
    wb.defined_names["GuestThankYou"] = DefinedName("GuestThankYou", attr_text=f"Guests!$L${start}:$L${end}")
    ws.freeze_panes = "B5"


# ===========================================================================
# Sheet 12: RSVP Dashboard
# ===========================================================================
def build_rsvp(wb):
    ws = wb.create_sheet("RSVP")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 4, 22, 16, 16])
    luxe_header(ws, "G", "✉  RSVP DASHBOARD",
                "Live response tracking and meal counts.")
    merge_set(ws, "B5:C5", "RESPONSE SUMMARY", "section")
    summary = [
        ("Invited Guests (seats)", '=SUMPRODUCT((GuestInvited="Yes")*IFERROR(GuestSeats,0))', "0"),
        ("Accepted (seats)", '=SUMPRODUCT((GuestRsvp="Accepted")*IFERROR(GuestSeats,0))', "0"),
        ("Declined (count)", '=COUNTIF(GuestRsvp,"Declined")', "0"),
        ("Pending (count)", '=COUNTIF(GuestRsvp,"Pending")', "0"),
        ("RSVP Completion %", '=IFERROR(COUNTIF(GuestRsvp,"Accepted")+COUNTIF(GuestRsvp,"Declined"),0)/MAX(COUNTA(GuestName),1)', "0%"),
        ("Attendance %", '=IFERROR(SUMPRODUCT((GuestRsvp="Accepted")*IFERROR(GuestSeats,0))/MAX(SUMPRODUCT((GuestInvited="Yes")*IFERROR(GuestSeats,0)),1),0)', "0%"),
    ]
    r = 6
    for lab, fml, fmt in summary:
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "input"; c.number_format = fmt
        ws.row_dimensions[r].height = 24
        r += 1

    # Meal counts
    merge_set(ws, "E5:F5", "MEAL COUNTS", "section_gold")
    ws.cell(row=6, column=5, value="Meal").style = "th"
    ws.cell(row=6, column=6, value="Count").style = "th"
    for i, meal in enumerate(MEAL_TYPES):
        rr = 7 + i
        ws.cell(row=rr, column=5, value=meal).style = "td_left"
        c = ws.cell(row=rr, column=6,
                    value=f'=SUMPRODUCT((GuestMeal="{meal}")*(GuestRsvp="Accepted")*IFERROR(GuestSeats,0))')
        c.style = "td"; c.number_format = "0"
        if i % 2 == 1:
            ws.cell(row=rr, column=5).fill = fill(MUTED_ROW)
            ws.cell(row=rr, column=6).fill = fill(MUTED_ROW)

    # Helper table for the donut chart (Accepted / Declined / Pending)
    ws.cell(row=15, column=2, value="Status").style = "th"
    ws.cell(row=15, column=3, value="Count").style = "th"
    for i, (st, fml) in enumerate([
        ("Accepted", '=COUNTIF(GuestRsvp,"Accepted")'),
        ("Declined", '=COUNTIF(GuestRsvp,"Declined")'),
        ("Pending", '=COUNTIF(GuestRsvp,"Pending")'),
    ]):
        rr = 16 + i
        ws.cell(row=rr, column=2, value=st).style = "td_left"
        c = ws.cell(row=rr, column=3, value=fml); c.style = "td"; c.number_format = "0"

    # Two charts stacked with a clear gap (donut ~7cm ≈ 13 rows ends ~row 28).
    donut = DoughnutChart(); donut.title = "RSVP Progress"; donut.height = 7; donut.width = 12
    donut.add_data(Reference(ws, min_col=3, min_row=15, max_row=18), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=2, min_row=16, max_row=18))
    donut.dataLabels = clean_labels(pct=True)
    ws.add_chart(donut, "H4")

    bar = BarChart(); bar.type = "col"; bar.title = "Meal Selections"; bar.height = 7; bar.width = 12
    bar.add_data(Reference(ws, min_col=6, min_row=6, max_row=6 + len(MEAL_TYPES)), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=5, min_row=7, max_row=6 + len(MEAL_TYPES)))
    ws.add_chart(bar, "H20")


# ===========================================================================
# Sheet 13: Seating Planner
# ===========================================================================
def build_seating(wb):
    ws = wb.create_sheet("Seating")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [12, 26, 12, 12, 14, 28])
    luxe_header(ws, "F", "🪑  SEATING PLANNER",
                "Assign tables with live capacity warnings and VIP flags.")
    table_headers(ws, 4, ["Table", "Assigned Guests", "Seated", "Capacity",
                          "Open", "VIP / Notes"])
    sample = [
        ("Table 1", "Bride, Groom, MOH, Best Man, parents (8)", 8, 10, "Head table — VIP"),
        ("Table 2", "Bennett family", 9, 10, "VIP — grandparents"),
        ("Table 3", "Carter family", 10, 10, ""),
        ("Table 4", "College friends", 8, 10, ""),
        ("Table 5", "Work friends", 7, 10, ""),
        ("Table 6", "Patel + Rivera families", 10, 10, "1 highchair needed"),
        ("Table 7", "Neighbors & family friends", 9, 10, ""),
        ("Table 8", "Plus-ones & misc", 6, 10, ""),
    ]
    start = 5
    end = start + 30 - 1
    for i, (tbl, guests, seated, cap, notes) in enumerate(sample):
        r = start + i
        ws.cell(row=r, column=1, value=tbl)
        ws.cell(row=r, column=2, value=guests)
        ws.cell(row=r, column=3, value=seated)
        ws.cell(row=r, column=4, value=cap)
        ws.cell(row=r, column=6, value=notes)
    for r in range(start, end + 1):
        ws.cell(row=r, column=5, value=f'=IF(D{r}="","",D{r}-IFERROR(C{r},0))')
    style_rows(ws, start, end, 6, text_left={2, 6}, ints={3, 4, 5})
    # Over capacity warning
    ws.conditional_formatting.add(
        f"A{start}:F{end}",
        FormulaRule(formula=[f'AND($C{start}<>"",$C{start}>$D{start})'], fill=fill("FBE6E6")))
    # Full table mint
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=["0"], fill=fill("E3F8EF")))
    # VIP gold
    ws.conditional_formatting.add(
        f"A{start}:F{end}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("VIP",$F{start}))'], fill=fill("F6EFE0")))

    # Capacity summary
    merge_set(ws, "A38:F38", "CAPACITY SUMMARY", "section_gold")
    sums = [
        ("Total Seated", f"=SUM(C{start}:C{end})"),
        ("Total Capacity", f"=SUM(D{start}:D{end})"),
        ("Seats Open", f"=SUM(E{start}:E{end})"),
        ("Accepted Guests (need seats)", '=SUMPRODUCT((GuestRsvp="Accepted")*IFERROR(GuestSeats,0))'),
    ]
    for i, (lab, fml) in enumerate(sums):
        r = 39 + i
        ws.cell(row=r, column=1, value=lab).style = "field_label"
        c = ws.cell(row=r, column=2, value=fml); c.style = "field_value"
        c.number_format = "0"; c.font = Font(bold=True, color=PRIMARY)
    ws.freeze_panes = "A5"


# ===========================================================================
# Generic table-style sheet builder for the remaining planners
# ===========================================================================
def build_generic(wb, name, icon, title, subtitle, headers, rows,
                  text_left_cols=None, money_cols=None, date_cols=None,
                  reserved=40, validations=None, widths=None):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    n = len(headers)
    if widths is None:
        widths = [18] + [16] * (n - 1)
    set_widths(ws, widths)
    last_col = get_column_letter(n)
    luxe_header(ws, last_col, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start = 5
    end = start + reserved - 1
    for i, row in enumerate(rows):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, n,
               text_left=text_left_cols or {2},
               money=money_cols or set(),
               dates=date_cols or set())
    if validations:
        for col_letter, list_name in validations:
            add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", list_name)
    ws.freeze_panes = "A5"
    return ws, start, end


# ===========================================================================
# Sheet 28: Vision Board (image placeholders)
# ===========================================================================
def build_vision(wb):
    ws = wb.create_sheet("VisionBoard")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 26, 2])
    luxe_header(ws, "F", "🖼  VISION BOARD",
                "Drop inspiration images into each tile. Notes + cost + priority below.")

    tiles = [
        "Venue", "Dress", "Florals", "Cake", "Hair", "Makeup",
        "Decor", "Tablescapes", "Stationery", "Color Palette",
        "Bouquets", "Photo Poses", "Reception Setup", "Lighting",
        "Favors", "Travel Ideas",
    ]
    # 4 columns of tiles
    col_positions = [2, 3, 4, 5]
    tile_h = 8  # rows per tile (image + 3 meta)
    row = 5
    for i, label in enumerate(tiles):
        col = col_positions[i % 4]
        if i % 4 == 0 and i > 0:
            row += tile_h + 1
        base = row
        # Image placeholder (merged 5 rows tall in this single column)
        ws.merge_cells(start_row=base, start_column=col, end_row=base + 4, end_column=col)
        cell = ws.cell(row=base, column=col, value=f"⬆ {label}\nDrop image here")
        cell.style = "image_ph"
        for rr in range(base, base + 5):
            ws.row_dimensions[rr].height = 22
            ws.cell(row=rr, column=col).fill = fill(BLUSH)
        # Meta line
        meta = ws.cell(row=base + 5, column=col, value="Notes / cost / priority ↓")
        meta.style = "field_label"
        n1 = ws.cell(row=base + 6, column=col, value="")
        n1.style = "field_value"
        n2 = ws.cell(row=base + 7, column=col, value="")
        n2.style = "field_value"

    # Bottom helper note
    last = row + tile_h + 1
    merge_set(ws, f"B{last}:E{last}",
              "Tip: In Excel use Insert ▸ Pictures ▸ Place in Cell. "
              "In Google Sheets use Insert ▸ Image ▸ Image in cell.",
              "section_gold")


# ===========================================================================
# Sheet 30: Wedding Day Command Center
# ===========================================================================
def build_wedding_day(wb):
    ws = wb.create_sheet("WeddingDay")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [14, 34, 18, 16, 28])
    luxe_header(ws, "E", "💒  WEDDING DAY COMMAND CENTER",
                "Hour-by-hour run of show — the single source of truth for the big day.")
    table_headers(ws, 4, ["Time", "Event", "Owner / Vendor", "Location", "Notes"])
    schedule = [
        ("8:00 AM",  "Hair & makeup begins (bridal suite)", "Radiant Beauty", "Bridal Suite", "Bride last"),
        ("10:30 AM", "Photographer arrives — detail shots", "Lumiere Photography", "Bridal Suite", "Rings, dress, invites ready"),
        ("11:30 AM", "Groomsmen arrive & dress", "Best Man", "Groom Suite", ""),
        ("12:30 PM", "Florist delivers bouquets & boutonnières", "Petal & Stem", "Both Suites", ""),
        ("1:00 PM",  "First look", "Couple + Photo", "Garden", "Private moment"),
        ("1:30 PM",  "Wedding party photos", "Photo + Video", "Garden", ""),
        ("2:30 PM",  "Vendors complete ceremony setup", "Coordinator", "Ceremony Lawn", "Chairs, arch, sound"),
        ("3:00 PM",  "Catering & rentals arrive", "Grand Catering / Luxe", "Reception Hall", ""),
        ("3:45 PM",  "Guests begin arriving", "Ushers", "Ceremony Lawn", "Prelude music"),
        ("4:30 PM",  "CEREMONY BEGINS", "Officiant", "Ceremony Lawn", "Processional"),
        ("5:00 PM",  "Ceremony ends — cocktail hour", "DJ / Bar", "Terrace", "Signature cocktails"),
        ("5:15 PM",  "Family & couple portraits", "Photo", "Garden", "Shot list ready"),
        ("6:00 PM",  "Reception entrance & first dance", "DJ", "Reception Hall", ""),
        ("6:30 PM",  "Dinner service (plated)", "Grand Catering", "Reception Hall", "Meal counts confirmed"),
        ("7:30 PM",  "Toasts & speeches", "MOH / Best Man", "Reception Hall", ""),
        ("8:00 PM",  "Cake cutting", "Couple", "Reception Hall", "Sweet Layers"),
        ("8:15 PM",  "Open dancing", "DJ", "Dance Floor", "Do-not-play list given"),
        ("9:30 PM",  "Bouquet & garter toss", "DJ", "Dance Floor", ""),
        ("10:30 PM", "Last dance & sparkler exit", "Coordinator", "Front Drive", "Sparklers + getaway car"),
        ("11:00 PM", "Vendor breakdown & load-out", "Coordinator", "All areas", "Gifts to suite"),
    ]
    start = 5
    end = start + len(schedule) - 1
    for i, row in enumerate(schedule):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, 5, text_left={2, 3, 4, 5})
    # Bold the ceremony row
    for c in range(1, 6):
        cell = ws.cell(row=start + 9, column=c)
        cell.font = Font(bold=True, color=PRIMARY)
        cell.fill = fill(HIGHLIGHT)
    # Time column primary fill
    for r in range(start, end + 1):
        tc = ws.cell(row=r, column=1)
        if r != start + 9:
            tc.fill = fill(PRIMARY)
            tc.font = Font(bold=True, color="FFFFFF")

    # Emergency contacts block
    er = end + 2
    merge_set(ws, f"A{er}:E{er}", "EMERGENCY CONTACTS", "section_gold")
    contacts = [
        ("Coordinator", "Marcus Reed — (555) 200-2299"),
        ("Planner", "Sophia Lane — (555) 200-8841"),
        ("Venue Manager", "Hartwell Estate — (555) 200-1000"),
        ("Maid of Honor", "Emily Bennett — (555) 412-7799"),
        ("Best Man", "Daniel Carter — (555) 901-2271"),
    ]
    for i, (role, who) in enumerate(contacts):
        r = er + 1 + i
        ws.cell(row=r, column=1, value=role).style = "field_label"
        merge_set(ws, f"B{r}:E{r}", who, "field_value")
    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 31: Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 18, 4, 30, 18])
    luxe_header(ws, "G", "📊  ANALYTICS — READINESS SCORE",
                "Weighted health scores across every dimension of planning.")

    merge_set(ws, "B5:D5", "HEALTH SCORES", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Budget Health", '=IFERROR(1-MAX(BudTotalActual-TotalBudget,0)/TotalBudget,0)'),
        ("Planning Completion", '=IFERROR(COUNTIF(ChkDone,"Yes")/MAX(COUNTA(ChkTask),1),0)'),
        ("Timeline Health", '=IFERROR(COUNTIF(TLDone,"Yes")/MAX(COUNTA(TLTask),1),0)'),
        ("Vendor Completion", '=IFERROR(COUNTIF(VenContract,"Signed")/MAX(COUNTA(VenName),1),0)'),
        ("Guest Completion", '=IFERROR((COUNTIF(GuestRsvp,"Accepted")+COUNTIF(GuestRsvp,"Declined"))/MAX(COUNTA(GuestName),1),0)'),
        ("Payment Health", '=IFERROR(COUNTIF(PayPaid,"Yes")/MAX(COUNTA(PayVendor),1),0)'),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4,
                value=(f'=IF(C{r}>=0.75,"On Track",IF(C{r}>=0.5,"Watch","Behind"))')).style = "td"
        if i % 2 == 1:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(
        f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FFFBE6E6",
                       mid_type="num", mid_value=0.5, mid_color="FFFFF3CD",
                       end_type="num", end_value=1, end_color="FF75E6C1"))
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"On Track"'], fill=fill("E3F8EF")))
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Behind"'], fill=fill("FBE6E6")))

    # Overall readiness — big number
    merge_set(ws, "F5:G5", "OVERALL READINESS", "section_gold")
    ws.merge_cells("F6:G8")
    cell = ws.cell(row=6, column=6, value=f"=AVERAGE(C{start}:C{end})")
    cell.style = "kpi_pct"
    cell.font = Font(size=48, bold=True, color=PRIMARY)
    cell.number_format = "0%"
    cell.fill = fill(IVORY)
    ws.merge_cells("F9:G9")
    sub = ws.cell(row=9, column=6, value="Wedding Readiness Score")
    sub.style = "kpi_label"

    # Score bar chart
    bar = BarChart(); bar.type = "bar"; bar.title = "Health by Dimension"; bar.height = 9; bar.width = 14
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    ws.add_chart(bar, "F12")


# ===========================================================================
# Sheet 1: Executive Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 17, 17, 17, 17, 17, 17, 17, 17, 17, 17, 2])
    ws.row_dimensions[1].height = 60
    merge_set(ws, "A1:L1", "  💍  WEDDING COMMAND CENTER", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:L2",
              "  Alexandra & James  ·  A complete wedding operating system — plan every detail in one elegant place.",
              "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 13):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    # KPI row 1 (5 cards spanning 2 cols each)
    ws.row_dimensions[5].height = 22
    ws.row_dimensions[6].height = 54
    kpi1 = [
        ("⏳ Days Until Wedding", "=MAX(WeddingDate-TODAY(),0)", "kpi_value"),
        ("💰 Budget Remaining", "=TotalBudget-BudTotalPaid", "kpi_money"),
        ("💳 Budget Spent", "=BudTotalActual", "kpi_money"),
        ("📋 Planning Complete", '=IFERROR(COUNTIF(ChkDone,"Yes")/MAX(COUNTA(ChkTask),1),0)', "kpi_pct"),
        ("✉ RSVP Complete", '=IFERROR((COUNTIF(GuestRsvp,"Accepted")+COUNTIF(GuestRsvp,"Declined"))/MAX(COUNTA(GuestName),1),0)', "kpi_pct"),
    ]
    col = 2
    for label, formula, vstyle in kpi1:
        merge_set(ws, f"{get_column_letter(col)}5:{get_column_letter(col+1)}5", label, "kpi_label")
        merge_set(ws, f"{get_column_letter(col)}6:{get_column_letter(col+1)}6", formula, vstyle)
        for r in (5, 6):
            for cc in range(col, col + 2):
                ws.cell(row=r, column=cc).border = BOX
        col += 2

    # KPI row 2 (5 cards)
    ws.row_dimensions[8].height = 22
    ws.row_dimensions[9].height = 54
    kpi2 = [
        ("🤝 Vendors Booked", '=COUNTIF(VenContract,"Signed")', "kpi_value"),
        ("⏰ Payments Outstanding", '=SUMPRODUCT((PayPaid<>"Yes")*IFERROR(PayBalance,0))', "kpi_money"),
        ("👥 Guest Count (accepted)", '=SUMPRODUCT((GuestRsvp="Accepted")*IFERROR(GuestSeats,0))', "kpi_value"),
        ("🪑 Tables Filled", '=ROUNDUP(SUMPRODUCT((GuestRsvp="Accepted")*IFERROR(GuestSeats,0))/MAX(SeatsPerTable,1),0)&" / "&NumTables', "kpi_value"),
        ("🔔 Tasks Due This Week", '=SUMPRODUCT((ChkDue<>"")*(ChkDue-TODAY()>=0)*(ChkDue-TODAY()<=7)*(ChkDone<>"Yes"))', "kpi_value"),
    ]
    col = 2
    for label, formula, vstyle in kpi2:
        merge_set(ws, f"{get_column_letter(col)}8:{get_column_letter(col+1)}8", label, "kpi_label")
        merge_set(ws, f"{get_column_letter(col)}9:{get_column_letter(col+1)}9", formula, vstyle)
        for r in (8, 9):
            for cc in range(col, col + 2):
                ws.cell(row=r, column=cc).border = BOX
        col += 2

    # Quick navigation
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:K11", "QUICK NAVIGATION", "section_gold")
    ws.row_dimensions[12].height = 28
    nav = ["Budget", "Guests", "Vendors", "Timeline", "Seating",
           "Vision", "Payments", "Contracts", "Checklist", "Day"]
    for i, name in enumerate(nav):
        cell = ws.cell(row=12, column=2 + i, value=name)
        cell.fill = fill(PRIMARY); cell.font = Font(size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center"); cell.border = BOX

    # Analytics charts
    ws.row_dimensions[14].height = 26
    merge_set(ws, "B14:K14", "ANALYTICS", "section_gold")
    # Charts are laid out on a clear 2x2 grid: left column anchored at B,
    # right column at H (past the left chart's right edge); two row bands
    # (15 and 35) spaced so 9cm-tall charts never overlap.
    bend = 4 + len(BUDGET_CATEGORIES)
    donut = DoughnutChart(); donut.title = "Budget Breakdown"; donut.height = 8.5; donut.width = 12
    donut.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=bend), titles_from_data=True)
    donut.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=bend))
    donut.dataLabels = no_labels()   # 21 slices — legend carries labels
    ws.add_chart(donut, "B15")

    bar = BarChart(); bar.type = "col"; bar.title = "Budget vs Actual"; bar.height = 8.5; bar.width = 12
    bar.add_data(Reference(wb["Budget"], min_col=2, min_row=4, max_row=bend), titles_from_data=True)
    bar.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=bend), titles_from_data=True)
    bar.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=bend))
    ws.add_chart(bar, "H15")

    # RSVP donut from RSVP helper table
    donut2 = DoughnutChart(); donut2.title = "RSVP Progress"; donut2.height = 8.5; donut2.width = 12
    donut2.add_data(Reference(wb["RSVP"], min_col=3, min_row=15, max_row=18), titles_from_data=True)
    donut2.set_categories(Reference(wb["RSVP"], min_col=2, min_row=16, max_row=18))
    donut2.dataLabels = clean_labels(pct=True)
    ws.add_chart(donut2, "B35")

    # Readiness scores bar from Analytics
    bar2 = BarChart(); bar2.type = "bar"; bar2.title = "Readiness by Dimension"; bar2.height = 8.5; bar2.width = 12
    bar2.add_data(Reference(wb["Analytics"], min_col=3, min_row=6, max_row=12), titles_from_data=True)
    bar2.set_categories(Reference(wb["Analytics"], min_col=2, min_row=7, max_row=12))
    ws.add_chart(bar2, "H35")

    # Footer
    ws.row_dimensions[56].height = 26
    merge_set(ws, "B56:K56",
              "Wedding Command Center v1.0  ·  Edit Settings → Wedding Date · Budget · Guests  ·  Hyperlink nav chips to each sheet in Excel/Sheets",
              "subtitle")


# ===========================================================================
# Build remaining planner sheets via generic builder
# ===========================================================================
def build_remaining(wb):
    today = dt.date.today()
    def d(off): return today + dt.timedelta(days=off)

    # 14 Bridal Party
    build_generic(
        wb, "BridalParty", "👗", "BRIDAL PARTY",
        "Wedding party, roles, sizes, outfits, gifts, and tasks.",
        ["Role", "Name", "Phone", "Outfit / Size", "Gift", "Gift Status", "Tasks"],
        [
            ("Maid of Honor", "Emily Bennett", "(555) 412-7799", "Dress — Emerald, M", "Robe + jewelry", "Ordered", "Plan shower"),
            ("Bridesmaid", "Maya Rivera", "(555) 412-7000", "Dress — Emerald, S", "Robe", "Ordered", "Bouquet help"),
            ("Bridesmaid", "Olivia Reed", "(555) 200-2244", "Dress — Emerald, L", "Robe", "Pending", ""),
            ("Best Man", "Daniel Carter", "(555) 901-2271", "Suit — Charcoal, 42R", "Cufflinks", "Ordered", "Plan stag"),
            ("Groomsman", "Aiden Brooks", "(555) 901-2272", "Suit — Charcoal, 40R", "Cufflinks", "Pending", ""),
            ("Groomsman", "Marc Lane", "(555) 200-8842", "Suit — Charcoal, 44R", "Cufflinks", "Pending", ""),
            ("Flower Girl", "Lily Patel", "—", "Ivory dress, 6Y", "Charm bracelet", "Ordered", ""),
            ("Ring Bearer", "Noah Carter", "—", "Mini suit, 5Y", "Toy + medal", "Ordered", ""),
            ("Mother of Bride", "Linda Bennett", "(555) 412-7833", "Gown — Champagne", "Corsage", "Pending", ""),
            ("Mother of Groom", "Carol Carter", "(555) 901-0000", "Gown — Navy", "Corsage", "Pending", ""),
        ],
        text_left_cols={2, 4, 5, 7}, reserved=30,
        validations=[("F", "StatusList")],
        widths=[18, 22, 16, 24, 18, 14, 26])

    # 15 Dress & Attire
    build_generic(
        wb, "Attire", "👰", "DRESS & ATTIRE",
        "Dress, suit, accessories, alterations, and fitting appointments.",
        ["Item", "Vendor", "Size", "Cost", "Status", "Fitting Date", "Notes"],
        [
            ("Wedding Dress", "Belle Bridal", "6", 2800, "Ordered", d(15), "2nd fitting"),
            ("Veil", "Belle Bridal", "—", 250, "Ordered", d(15), "Cathedral length"),
            ("Bridal Shoes", "Glass Slipper Co.", "7", 180, "Purchased", None, "Block heel"),
            ("Bridal Jewelry", "Aurelia Fine", "—", 600, "Purchased", None, "Borrowed earrings"),
            ("Alterations", "Stitch Studio", "—", 350, "In Progress", d(30), "Hem + bustle"),
            ("Groom Suit", "Bespoke Tailors", "42R", 800, "Ordered", d(50), "Charcoal 3-piece"),
            ("Groom Shoes", "Oxford & Co.", "10", 160, "Purchased", None, "Black oxfords"),
            ("Tie / Pocket Square", "Bespoke Tailors", "—", 90, "Ordered", None, "Emerald silk"),
        ],
        text_left_cols={1, 2, 7}, money_cols={4}, date_cols={6}, reserved=25,
        validations=[("E", "StatusList")],
        widths=[22, 22, 12, 14, 16, 16, 26])

    # 16 Beauty Timeline
    build_generic(
        wb, "Beauty", "💄", "BEAUTY TIMELINE",
        "Countdown of hair, skin, nails, and wellness appointments.",
        ["Treatment", "Provider", "Appointment", "Cost", "Status", "Notes"],
        [
            ("Skincare facial #1", "Glow Spa", d(-30), 120, "Complete", "Monthly until wedding"),
            ("Hair color & cut", "Radiant Beauty", d(-14), 180, "Booked", "Pre-wedding refresh"),
            ("Hair & makeup trial", "Radiant + Glow", d(15), 200, "Booked", "Bring veil"),
            ("Spray tan trial", "Sunlit Studio", d(20), 45, "Booked", "Test shade"),
            ("Nails (gel)", "Polished", d(43), 60, "Pending", "Day before"),
            ("Spray tan (final)", "Sunlit Studio", d(44), 45, "Pending", "Day before"),
            ("Bridal hair & makeup", "Radiant + Glow", d(45), 350, "Booked", "Wedding morning 8am"),
            ("Massage (couple)", "Serenity Spa", d(42), 220, "Pending", "Relax pre-wedding"),
        ],
        text_left_cols={1, 2, 6}, money_cols={4}, date_cols={3}, reserved=25,
        validations=[("E", "StatusList")],
        widths=[24, 20, 16, 14, 16, 26])

    # 17 Ceremony Planner
    build_generic(
        wb, "Ceremony", "💐", "CEREMONY PLANNER",
        "Processional order, readings, music, and vow logistics.",
        ["Segment", "Detail", "Person / Vendor", "Music / Notes"],
        [
            ("Prelude", "Guests seated", "DJ — Skyline", "Acoustic instrumental"),
            ("Processional — Grandparents", "Seated first", "Ushers", "Canon in D"),
            ("Processional — Parents", "Mothers seated", "Ushers", "Canon in D"),
            ("Processional — Wedding Party", "Bridesmaids + groomsmen", "Coordinator", "A Thousand Years (instr.)"),
            ("Processional — Flower Girl/Ring Bearer", "Down the aisle", "Coordinator", ""),
            ("Bride's Entrance", "Escorted by father", "—", "Here Comes the Sun"),
            ("Welcome & Opening", "Officiant remarks", "Rev. Hale", ""),
            ("Reading 1", "Poem", "Emily Bennett", "‘The Art of Marriage’"),
            ("Reading 2", "Scripture / passage", "Daniel Carter", ""),
            ("Vows", "Personal vows", "Couple", "Handwritten"),
            ("Ring Exchange", "Rings from ring bearer", "Couple", ""),
            ("Pronouncement & Kiss", "—", "Rev. Hale", ""),
            ("Recessional", "Newlyweds exit", "DJ", "Signed, Sealed, Delivered"),
        ],
        text_left_cols={1, 2, 3, 4}, reserved=25,
        widths=[28, 26, 22, 28])

    # 18 Reception Planner
    build_generic(
        wb, "Reception", "🥂", "RECEPTION PLANNER",
        "Run of show for the reception — entrances, dances, speeches, exit.",
        ["Time", "Event", "Lead", "Music / Notes"],
        [
            ("6:00 PM", "Grand entrance", "DJ", "Can't Stop the Feeling"),
            ("6:05 PM", "First dance", "Couple", "Perfect — Ed Sheeran"),
            ("6:15 PM", "Welcome toast", "Father of bride", ""),
            ("6:30 PM", "Dinner service", "Grand Catering", "Plated, 3 courses"),
            ("7:30 PM", "Speeches", "MOH + Best Man", "5 min each"),
            ("7:50 PM", "Parent dances", "Couple", "Mother-son / Father-daughter"),
            ("8:00 PM", "Cake cutting", "Couple", "Sweet Layers"),
            ("8:15 PM", "Open dancing", "DJ", "Party playlist"),
            ("9:00 PM", "Bouquet toss", "Bride", ""),
            ("9:15 PM", "Garter toss", "Groom", ""),
            ("9:30 PM", "Late-night snack", "Catering", "Slider bar"),
            ("10:30 PM", "Last dance", "Couple", "At Last — Etta James"),
            ("10:45 PM", "Sparkler send-off", "Coordinator", "Getaway car ready"),
        ],
        text_left_cols={2, 3, 4}, reserved=25,
        widths=[14, 26, 20, 30])

    # 19 Menu Planner
    build_generic(
        wb, "Menu", "🍽", "MENU PLANNER",
        "Courses, bar, cocktails, dietary, and kids' meals.",
        ["Course", "Selection", "Notes / Dietary", "Per-Head Cost"],
        [
            ("Passed Hors d'oeuvres", "Caprese skewers · mini crab cakes · bruschetta", "GF option available", 12),
            ("Salad", "Mixed greens, candied pecans, goat cheese", "Vegan w/o cheese", 0),
            ("Entrée — Beef", "Filet mignon, truffle mash, asparagus", "", 0),
            ("Entrée — Chicken", "Herb chicken, risotto, seasonal veg", "", 0),
            ("Entrée — Fish", "Pan-seared salmon, quinoa, greens", "", 0),
            ("Entrée — Vegetarian", "Wild mushroom wellington", "Vegan available", 0),
            ("Kids Meal", "Chicken tenders, fruit, fries", "", 18),
            ("Cake", "3-tier lemon & vanilla", "Sweet Layers", 0),
            ("Dessert Table", "Macarons, mini tarts, chocolate truffles", "", 8),
            ("Bar", "Open bar — beer, wine, signature cocktails", "Cash for top-shelf", 35),
            ("Signature Cocktail 1", "‘The Emerald’ — gin, cucumber, mint", "", 0),
            ("Signature Cocktail 2", "‘Golden Hour’ — bourbon, honey, lemon", "", 0),
            ("Late-night Snack", "Slider bar + fries", "", 9),
        ],
        text_left_cols={1, 2, 3}, money_cols={4}, reserved=25,
        widths=[22, 38, 26, 16])

    # 20 Music Planner
    build_generic(
        wb, "Music", "🎵", "MUSIC PLANNER",
        "Ceremony, cocktail, reception playlist, and the do-not-play list.",
        ["Moment", "Song", "Artist", "Status"],
        [
            ("Processional", "Canon in D", "Pachelbel", "Confirmed"),
            ("Bride's Entrance", "Here Comes the Sun", "The Beatles (instr.)", "Confirmed"),
            ("Recessional", "Signed, Sealed, Delivered", "Stevie Wonder", "Confirmed"),
            ("Cocktail Hour", "Jazz standards playlist", "Various", "Confirmed"),
            ("Grand Entrance", "Can't Stop the Feeling", "Justin Timberlake", "Confirmed"),
            ("First Dance", "Perfect", "Ed Sheeran", "Confirmed"),
            ("Mother-Son Dance", "A Song for Mama", "Boyz II Men", "Confirmed"),
            ("Father-Daughter Dance", "My Girl", "The Temptations", "Confirmed"),
            ("Cake Cutting", "Sugar, Sugar", "The Archies", "Confirmed"),
            ("Last Dance", "At Last", "Etta James", "Confirmed"),
            ("DO NOT PLAY", "Chicken Dance", "—", "Banned"),
            ("DO NOT PLAY", "Macarena", "—", "Banned"),
        ],
        text_left_cols={1, 2, 3}, reserved=30,
        widths=[24, 28, 24, 14])

    # 21 Floral Planner
    build_generic(
        wb, "Floral", "🌸", "FLORAL PLANNER",
        "Bouquets, centerpieces, and ceremony / reception florals.",
        ["Item", "Description", "Quantity", "Unit Cost", "Total", "Notes"],
        [
            ("Bridal Bouquet", "Garden roses, ranunculus, eucalyptus", 1, 250, None, "Cascade style"),
            ("Bridesmaid Bouquets", "Smaller garden mix", 3, 85, None, ""),
            ("Boutonnières", "Rose + greenery", 6, 25, None, ""),
            ("Corsages", "Wrist — mothers/grandmothers", 4, 35, None, ""),
            ("Ceremony Arch", "Lush emerald + ivory installation", 1, 850, None, "Focal point"),
            ("Aisle Markers", "Petals + small arrangements", 8, 30, None, ""),
            ("Centerpieces — Tall", "Elevated arrangements", 8, 120, None, "Head + VIP tables"),
            ("Centerpieces — Low", "Compote bowls", 8, 75, None, ""),
            ("Cake Flowers", "Fresh blooms for cake", 1, 60, None, ""),
            ("Flower Girl Petals", "Loose petals basket", 1, 25, None, ""),
        ],
        text_left_cols={1, 2, 6}, money_cols={4, 5}, reserved=25,
        widths=[22, 32, 12, 14, 14, 24])
    # add Total formulas to Floral
    ws = wb["Floral"]
    for r in range(5, 5 + 10):
        ws.cell(row=r, column=5, value=f'=IF(C{r}="","",C{r}*D{r})')

    # 22 Decor Planner
    build_generic(
        wb, "Decor", "🕯", "DECOR PLANNER",
        "Decor, rentals, lighting, signage, and furniture.",
        ["Item", "Type", "Quantity", "Source", "Cost", "Status"],
        [
            ("String / bistro lighting", "Lighting", 1, "Luxe Rentals", 450, "Booked"),
            ("Uplighting (emerald)", "Lighting", 12, "Skyline", 360, "Booked"),
            ("Lounge furniture set", "Furniture", 1, "Luxe Rentals", 600, "Booked"),
            ("Welcome sign (acrylic)", "Signage", 1, "Etsy / Custom", 95, "Ordered"),
            ("Seating chart display", "Signage", 1, "Custom", 120, "Ordered"),
            ("Table numbers", "Signage", 16, "Custom", 80, "Ordered"),
            ("Candles & holders", "Decor", 60, "Bulk", 240, "Purchased"),
            ("Aisle runner", "Decor", 1, "Luxe Rentals", 75, "Booked"),
            ("Charger plates", "Rentals", 140, "Luxe Rentals", 280, "Booked"),
            ("Linens (emerald/ivory)", "Rentals", 16, "Luxe Rentals", 320, "Booked"),
            ("Arbor draping", "Decor", 1, "Petal & Stem", 150, "Pending"),
        ],
        text_left_cols={1, 4, 2}, money_cols={5}, reserved=30,
        validations=[("F", "StatusList")],
        widths=[24, 16, 12, 20, 14, 16])

    # 23 Photography Shot List
    build_generic(
        wb, "ShotList", "📸", "PHOTOGRAPHY SHOT LIST",
        "Every must-have shot, grouped and check-off ready.",
        ["Group", "Shot", "Priority", "Captured"],
        [
            ("Details", "Rings, invitation suite, shoes, dress", "High", "No"),
            ("Getting Ready", "Bride with bridesmaids", "High", "No"),
            ("Getting Ready", "Groom with groomsmen", "High", "No"),
            ("First Look", "Couple first look reaction", "High", "No"),
            ("Bride", "Solo bridal portraits", "High", "No"),
            ("Groom", "Solo groom portraits", "Medium", "No"),
            ("Family", "Bride + parents", "High", "No"),
            ("Family", "Groom + parents", "High", "No"),
            ("Family", "Full family — both sides", "High", "No"),
            ("Wedding Party", "Full party fun + formal", "High", "No"),
            ("Ceremony", "Processional, vows, first kiss", "High", "No"),
            ("Ceremony", "Ring exchange close-up", "High", "No"),
            ("Reception", "Grand entrance, first dance", "High", "No"),
            ("Reception", "Toasts, cake cutting", "Medium", "No"),
            ("Golden Hour", "Couple sunset portraits", "High", "No"),
            ("Detail", "Tablescape + centerpieces", "Medium", "No"),
            ("Exit", "Sparkler send-off", "High", "No"),
        ],
        text_left_cols={1, 2}, reserved=40,
        validations=[("C", "PriorityList"), ("D", "YesNoList")],
        widths=[18, 38, 14, 12])
    ws = wb["ShotList"]
    ws.conditional_formatting.add(
        "A5:D44",
        FormulaRule(formula=['$D5="Yes"'], fill=fill("E3F8EF")))

    # 24 Honeymoon Planner
    build_generic(
        wb, "Honeymoon", "🌴", "HONEYMOON PLANNER",
        "Flights, hotels, activities, and the travel checklist.",
        ["Category", "Item", "Date", "Cost", "Booked", "Confirmation / Notes"],
        [
            ("Destination", "Amalfi Coast, Italy", None, 0, "Yes", "10 nights"),
            ("Flights", "Round-trip — JFK ↔ Naples", d(60), 2400, "Yes", "Conf. AB12CD"),
            ("Hotel", "Cliffside resort, Positano", d(60), 4200, "Yes", "Conf. POS-9981"),
            ("Transfer", "Airport → hotel private car", d(60), 180, "Yes", ""),
            ("Activity", "Capri boat day", d(62), 350, "No", "Book 2 weeks ahead"),
            ("Activity", "Cooking class", d(63), 200, "No", ""),
            ("Activity", "Pompeii guided tour", d(65), 160, "No", ""),
            ("Dining", "Anniversary dinner reservation", d(64), 250, "No", "Michelin star"),
            ("Travel Docs", "Passports valid 6+ months", None, 0, "Yes", "Both confirmed"),
            ("Travel Docs", "Travel insurance", d(55), 220, "No", ""),
        ],
        text_left_cols={1, 2, 6}, money_cols={4}, date_cols={3}, reserved=25,
        validations=[("E", "YesNoList")],
        widths=[18, 30, 14, 14, 12, 26])

    # 25 Registry Tracker
    build_generic(
        wb, "Registry", "🎁", "REGISTRY TRACKER",
        "Registry items, store, price, priority, and purchase status.",
        ["Item", "Store", "Price", "Priority", "Purchased", "Link / Notes"],
        [
            ("Stand mixer", "Williams Sonoma", 450, "High", "Yes", "Emerald color"),
            ("Cookware set", "Williams Sonoma", 600, "High", "No", "12-piece"),
            ("Dinnerware (set of 8)", "Crate & Barrel", 320, "High", "No", ""),
            ("Flatware set", "Crate & Barrel", 180, "Medium", "Yes", ""),
            ("Bedding set (king)", "Pottery Barn", 280, "Medium", "No", ""),
            ("Bath towel bundle", "Pottery Barn", 140, "Medium", "Yes", ""),
            ("Espresso machine", "Breville", 700, "High", "No", "Most wanted"),
            ("Luggage set", "Away", 595, "Low", "No", "For honeymoon"),
            ("Smart speaker", "Amazon", 120, "Low", "Yes", ""),
            ("Wine fridge", "Wayfair", 350, "Low", "No", ""),
        ],
        text_left_cols={1, 2, 6}, money_cols={3}, reserved=40,
        validations=[("D", "PriorityList"), ("E", "YesNoList")],
        widths=[26, 20, 14, 14, 12, 26])
    ws = wb["Registry"]
    ws.conditional_formatting.add(
        "A5:F44",
        FormulaRule(formula=['$E5="Yes"'], fill=fill("E3F8EF")))

    # 26 Gift Tracker
    build_generic(
        wb, "Gifts", "💝", "GIFT TRACKER",
        "Gifts received, value, and thank-you status.",
        ["Gift", "From", "Value", "Received Date", "Thank-You Sent", "Notes"],
        [
            ("Stand mixer", "Linda & Robert Bennett", 450, d(-5), "No", "From registry"),
            ("Cash gift", "Carter Family", 500, d(-3), "No", ""),
            ("Crystal vase", "The Hartleys", 120, d(-2), "No", ""),
            ("Espresso machine", "Sophia & Marc Lane", 700, d(-1), "No", ""),
            ("Gift card — Crate & Barrel", "Olivia Reed", 100, d(-1), "No", ""),
            ("Hand towels set", "Grandma Rose", 60, d(-6), "Yes", "Handmade"),
        ],
        text_left_cols={1, 2, 6}, money_cols={3}, date_cols={4}, reserved=60,
        validations=[("E", "YesNoList")],
        widths=[26, 24, 14, 16, 16, 22])

    # 27 Thank-You Card Tracker
    build_generic(
        wb, "ThankYou", "💌", "THANK-YOU CARD TRACKER",
        "Track every card from written to mailed.",
        ["Recipient", "Gift", "Written", "Mailed", "Date Sent", "Notes"],
        [
            ("Linda & Robert Bennett", "Stand mixer", "Yes", "No", None, ""),
            ("Carter Family", "Cash gift", "No", "No", None, ""),
            ("The Hartleys", "Crystal vase", "No", "No", None, ""),
            ("Sophia & Marc Lane", "Espresso machine", "Yes", "Yes", d(-1), "Hand-delivered"),
            ("Olivia Reed", "Gift card", "No", "No", None, ""),
            ("Grandma Rose", "Hand towels", "Yes", "Yes", d(-2), ""),
        ],
        text_left_cols={1, 2, 6}, date_cols={5}, reserved=60,
        validations=[("C", "YesNoList"), ("D", "YesNoList")],
        widths=[26, 24, 12, 12, 16, 24])
    ws = wb["ThankYou"]
    ws.conditional_formatting.add(
        "A5:F64",
        FormulaRule(formula=['$D5="Yes"'], fill=fill("E3F8EF")))

    # 29 Packing Checklist (multi-list)
    ws = wb.create_sheet("Packing")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 6, 26, 6, 26, 6, 26, 6])
    luxe_header(ws, "H", "🧳  PACKING CHECKLISTS",
                "Wedding day, honeymoon, emergency kit, and vendor items.")
    lists = [
        ("WEDDING DAY", ["Dress + garment bag", "Veil", "Shoes", "Jewelry",
                         "Vows book", "Rings", "Something blue", "Emergency kit",
                         "Robe for getting ready", "Perfume", "Flat shoes for reception"]),
        ("HONEYMOON", ["Passports", "Tickets / confirmations", "Swimwear",
                       "Evening outfits", "Sunscreen", "Adapters", "Medications",
                       "Camera", "Chargers", "Travel insurance docs"]),
        ("EMERGENCY KIT", ["Sewing kit", "Safety pins", "Stain remover",
                           "Band-aids", "Pain reliever", "Bobby pins", "Mints",
                           "Tissues", "Deodorant", "Phone charger", "Snacks",
                           "Static spray", "Double-sided tape"]),
        ("VENDOR / RECEPTION ITEMS", ["Guest book + pen", "Card box", "Toasting flutes",
                                       "Cake server set", "Table numbers", "Place cards",
                                       "Signage", "Favors", "Tip envelopes",
                                       "Timeline printouts", "Vendor meals list"]),
    ]
    pairs = [(1, 2), (3, 4), (5, 6), (7, 8)]
    for (title, items), (cn, cc) in zip(lists, pairs):
        merge_set(ws, f"{get_column_letter(cn)}4:{get_column_letter(cc)}4", title, "th")
        ws.row_dimensions[4].height = 28
        for i, item in enumerate(items):
            r = 5 + i
            ws.cell(row=r, column=cn, value=item).style = "td_left"
            chk = ws.cell(row=r, column=cc, value="☐")
            chk.style = "td"; chk.font = Font(size=14, bold=True, color=PRIMARY)
            if i % 2 == 1:
                ws.cell(row=r, column=cn).fill = fill(MUTED_ROW)
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)


# ===========================================================================
# Build
# ===========================================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    build_settings(wb)
    build_couple(wb)
    build_timeline(wb)
    build_checklist(wb)
    build_budget(wb)
    build_payments(wb)
    build_savings(wb)
    build_vendor_crm(wb)
    build_comparison(wb)
    build_contracts(wb)
    build_guest_crm(wb)
    build_rsvp(wb)
    build_seating(wb)
    build_remaining(wb)       # bridal party, attire, beauty, ceremony, reception,
                              # menu, music, floral, decor, shotlist, honeymoon,
                              # registry, gifts, thankyou, packing
    build_vision(wb)
    build_wedding_day(wb)
    build_analytics(wb)
    build_dashboard(wb)       # last (references everything)

    # Final sheet order per PRD (1..32)
    order = [
        "Dashboard", "Couple", "Timeline", "Checklist", "Budget", "Payments",
        "Savings", "Vendors", "Comparison", "Contracts", "Guests", "RSVP",
        "Seating", "BridalParty", "Attire", "Beauty", "Ceremony", "Reception",
        "Menu", "Music", "Floral", "Decor", "ShotList", "Honeymoon",
        "Registry", "Gifts", "ThankYou", "VisionBoard", "Packing",
        "WeddingDay", "Analytics", "Settings",
    ]
    wb._sheets = [wb[name] for name in order]

    # Tab colors — alternate brand palette
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, name in enumerate(order):
        wb[name].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Analytics"].sheet_properties.tabColor = PRIMARY
    wb["WeddingDay"].sheet_properties.tabColor = DANGER
    wb["Settings"].sheet_properties.tabColor = SURFACE

    out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(out_dir, "Wedding_Command_Center.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
