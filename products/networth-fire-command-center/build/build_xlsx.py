"""Build Net Worth & FIRE Command Center™ — The Financial-Independence Operating System.

14 tabs · a premium net-worth & FIRE (financial independence, retire early) operating
system in Google Sheets & Excel. Dashboard, a FIRE-number engine (annual expenses →
your number, progress, coast FIRE & years to FI), a net-worth roll-up, assets,
liabilities, accounts, contributions, income & expenses, a savings-rate log, a coast &
projection, milestones and a net-worth trend — one dashboard. Know your number, and
your progress to it.

Run: python3 build_xlsx.py   ->  ../Networth_FIRE_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
ASSET_TYPE = ["Investment", "Cash", "Property", "Vehicle", "Other"]
LIAB_TYPE = ["Mortgage", "Loan", "Credit", "Other"]
ACCT_TYPE = ["Brokerage", "401(k)", "IRA", "HSA", "Bank", "Other"]

SAVINGS_GOAL = 0.50
EMERGENCY_FUND_GOAL = 20000
CONTRIB_GOAL = 23000
HALFWAY_GOAL = 0.50
WITHDRAWAL_RATE = 0.04
RETURN_RATE = 0.07
CURRENT_AGE = 35
RETIRE_AGE = 65

# Income & expenses
ANNUAL_INCOME = 80000
ANNUAL_EXPENSES = 40000
CONTRIB_ANNUAL = 23000

# Assets: (asset, type, value)
ASSETS = [
    ("Brokerage", "Investment", 120000), ("401(k)", "Investment", 60000), ("Roth IRA", "Investment", 20000),
    ("Cash & savings", "Cash", 20000), ("Home", "Property", 300000), ("Car", "Vehicle", 10000),
]

# Liabilities: (liability, type, balance)
LIABILITIES = [
    ("Mortgage", "Mortgage", 250000), ("Student loan", "Loan", 20000), ("Credit card", "Credit", 10000),
]

# Accounts: (account, institution, type, balance)
ACCOUNTS = [
    ("Brokerage", "Vanguard", "Brokerage", 120000), ("401(k)", "Fidelity", "401(k)", 60000),
    ("Roth IRA", "Vanguard", "IRA", 20000), ("Savings", "Ally", "Bank", 20000),
    ("Checking", "Local CU", "Bank", 4000),
]

# Contributions log: (date, account, amount)
CONTRIB = [
    ("Jul 1", "401(k)", 1500), ("Jul 1", "Roth IRA", 500), ("Jun 1", "401(k)", 1500),
    ("Jun 1", "Roth IRA", 500), ("May 1", "401(k)", 1500), ("May 1", "Roth IRA", 500),
]

# Savings-rate log: (month, income, expenses)
SAVINGS_LOG = [
    ("Feb", 6600, 3400), ("Mar", 6600, 3300), ("Apr", 6700, 3300),
    ("May", 6700, 3300), ("Jun", 6700, 3200), ("Jul", 6700, 3200),
]

# Milestones: (milestone, target, reached?)
MILESTONES = [
    ("First $10k invested", 10000, "Yes"), ("Emergency fund funded", 20000, "Yes"),
    ("$100k net worth", 100000, "Yes"), ("Coast FIRE reached", 131367, "Yes"),
    ("$250k net worth", 250000, "Yes"), ("Halfway to FI ($500k)", 500000, "No"),
    ("Debt-free (ex-mortgage)", 0, "No"), ("FIRE number ($1M)", 1000000, "No"),
]

# Net-worth trend: (month, net worth)
MONTHS = [("Feb", 214000), ("Mar", 222000), ("Apr", 231000),
          ("May", 238000), ("Jun", 244000), ("Jul", 250000)]

# Coast & projection: (age, note)
PROJ_AGES = [35, 40, 45, 50, 55, 60, 65]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your assumptions & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR ASSUMPTIONS & GOALS", "section")
    controls = [
        ("Your name", "Sage", None, "Owner"),
        ("Plan name", "North Star Finance", None, "PlanName"),
        ("Safe withdrawal rate", WITHDRAWAL_RATE, "0.0%", "WithdrawalRate"),
        ("Expected return (annual)", RETURN_RATE, "0.0%", "ReturnRate"),
        ("Current age", CURRENT_AGE, "0", "CurrentAge"),
        ("Target retire age", RETIRE_AGE, "0", "RetireAge"),
        ("Savings-rate goal %", SAVINGS_GOAL, "0%", "SavingsGoal"),
        ("Emergency-fund goal", EMERGENCY_FUND_GOAL, '"$"#,##0', "EmergencyFundGoal"),
        ("Annual contribution goal", CONTRIB_GOAL, '"$"#,##0', "ContribGoal"),
        ("Halfway-to-FI goal %", HALFWAY_GOAL, "0%", "HalfwayGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Asset type", ASSET_TYPE, "AssetTypeList"), ("F", "Liability type", LIAB_TYPE, "LiabTypeList"),
             ("G", "Account type", ACCT_TYPE, "AcctTypeList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🧭  NET WORTH & FIRE COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know your number, and your progress to it.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE FINANCIAL PICTURE, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Financial independence comes down to two numbers most people never calculate: the number you "
                      "need to never work again, and how far along you are. This makes both clear: a FIRE-number engine "
                      "turns your annual spending into the amount you need, then shows your progress, your coast-FIRE "
                      "number and your years to freedom. Track your net worth, assets, liabilities, accounts, "
                      "contributions and savings rate — all in ONE premium Google Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your withdrawal rate, return, ages & goals.",
             "2.  Enter your Assets and Liabilities — your Net Worth appears.",
             "3.  Enter your spending in the FIRE Number tab — see your number.",
             "4.  Log your Income & Expenses and your Contributions.",
             "5.  Read your progress, coast FIRE and years to FI.",
             "6.  Check the Dashboard: net worth, progress & a FIRE Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional saver (North Star Finance, Sage) is included so you can see how it all "
               "connects — just type over it with your own numbers. The FIRE number and your progress to it are the two "
               "figures that decide when work becomes optional, and they roll into a live FIRE Score. Twelve matching "
               "printable pages (net-worth worksheet, FIRE-number worksheet, asset & debt lists & more) are included. "
               "This is a personal-finance organizing tool, not financial, legal or tax advice — confirm figures with "
               "your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "The best day to know your number was years ago. The second best day is today.", "section_gold")


# ===========================================================================
def build_assets(wb):
    ws = wb.create_sheet("Assets"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 16, 2])
    luxe_header(ws, "D", "💎  ASSETS",
                "Everything you own — investments, cash, property. Your invested assets drive FIRE.")
    table_headers(ws, 4, ["Asset", "Type", "Value"], start_col=2)
    start = L0
    for i, (asset, typ, val) in enumerate(ASSETS):
        r = start + i
        ws.cell(row=r, column=2, value=asset).style = "td_left"
        ws.cell(row=r, column=3, value=typ).style = "td"
        cv = ws.cell(row=r, column=4, value=val); cv.style = "input"; cv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(ASSETS) - 1
    nrange(wb, "AssetType", "Assets", "C", start, end)
    nrange(wb, "AssetValue", "Assets", "D", start, end)
    add_dv(ws, f"C{start}:C{end}", "AssetTypeList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL ASSETS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ce = ws.cell(row=tot, column=4, value="=SUM(AssetValue)"); ce.style = "td"; ce.font = Font(bold=True, color=PRIMARY); ce.fill = fill(MINT_BG); ce.number_format = '"$"#,##0'
    cell_name(wb, "TotalAssets", "Assets", f"$D${tot}")
    ws.cell(row=tot + 2, column=2, value="Invested assets").style = "field_label"
    ci = ws.cell(row=tot + 2, column=4, value='=SUMIF(AssetType,"Investment",AssetValue)'); ci.style = "field_value"; ci.number_format = '"$"#,##0'
    cell_name(wb, "InvestedAssets", "Assets", f"$D${tot+2}")
    ws.cell(row=tot + 3, column=2, value="Cash & emergency fund").style = "field_label"
    cc = ws.cell(row=tot + 3, column=4, value='=SUMIF(AssetType,"Cash",AssetValue)'); cc.style = "field_value"; cc.number_format = '"$"#,##0'
    cell_name(wb, "CashAssets", "Assets", f"$D${tot+3}")
    ws.freeze_panes = "A5"


def build_liabilities(wb):
    ws = wb.create_sheet("Liabilities"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 16, 2])
    luxe_header(ws, "D", "🧾  LIABILITIES",
                "Everything you owe — mortgage, loans, cards. Net worth is what's left after these.")
    table_headers(ws, 4, ["Liability", "Type", "Balance"], start_col=2)
    start = L0
    for i, (liab, typ, bal) in enumerate(LIABILITIES):
        r = start + i
        ws.cell(row=r, column=2, value=liab).style = "td_left"
        ws.cell(row=r, column=3, value=typ).style = "td"
        cv = ws.cell(row=r, column=4, value=bal); cv.style = "input"; cv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LIABILITIES) - 1
    nrange(wb, "LiabType", "Liabilities", "C", start, end)
    nrange(wb, "LiabBalance", "Liabilities", "D", start, end)
    add_dv(ws, f"C{start}:C{end}", "LiabTypeList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL LIABILITIES").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ce = ws.cell(row=tot, column=4, value="=SUM(LiabBalance)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "TotalLiab", "Liabilities", f"$D${tot}")
    ws.cell(row=tot + 2, column=2, value="Non-mortgage debt").style = "field_label"
    cn = ws.cell(row=tot + 2, column=4, value='=TotalLiab-SUMIF(LiabType,"Mortgage",LiabBalance)'); cn.style = "field_value"; cn.number_format = '"$"#,##0'
    cell_name(wb, "NonMortgageDebt", "Liabilities", f"$D${tot+2}")
    ws.freeze_panes = "A5"


def build_networth(wb):
    ws = wb.create_sheet("Net Worth"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 18, 2])
    luxe_header(ws, "C", "📊  NET WORTH",
                "What you own minus what you owe — the one number that tracks your whole financial life.")
    rows = [
        ("Total assets", "=TotalAssets", '"$"#,##0', None, MINT_BG),
        ("− Total liabilities", "=TotalLiab", '"$"#,##0', None, None),
    ]
    for i, (lab, val, fmt, nm, bg) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "field_value"; c.number_format = fmt
        if bg:
            c.fill = fill(bg)
    ws.cell(row=7, column=2, value="= NET WORTH").style = "th"
    cn = ws.cell(row=7, column=3, value="=TotalAssets-TotalLiab"); cn.style = "td"; cn.font = Font(bold=True, size=14, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetWorth", "Net Worth", "$C$7")
    ws.cell(row=9, column=2, value="Liquid net worth (cash + investments − non-mortgage debt)").style = "field_label"
    cl = ws.cell(row=9, column=3, value="=CashAssets+InvestedAssets-NonMortgageDebt"); cl.style = "field_value"; cl.number_format = '"$"#,##0'; cl.fill = fill(MINT_BG)
    cell_name(wb, "LiquidNW", "Net Worth", "$C$9")
    ws.cell(row=10, column=2, value="Invested assets (your FIRE engine)").style = "field_label"
    ci = ws.cell(row=10, column=3, value="=InvestedAssets"); ci.style = "field_value"; ci.number_format = '"$"#,##0'
    ws.cell(row=12, column=2, value="Net worth rolls up your Assets and Liabilities tabs — keep them current.").style = "section_gold"


def build_fire(wb):
    ws = wb.create_sheet("FIRE Number"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 2])
    luxe_header(ws, "C", "🔥  FIRE NUMBER",
                "Your annual spending × 25 (the 4% rule) is the amount you need to never work again.")
    ws.cell(row=5, column=2, value="Annual spending (what your life costs)").style = "field_label"
    ce = ws.cell(row=5, column=3, value=ANNUAL_EXPENSES); ce.style = "input"; ce.number_format = '"$"#,##0'
    cell_name(wb, "AnnualExpenses", "FIRE Number", "$C$5")
    ws.cell(row=6, column=2, value="÷ Safe withdrawal rate").style = "field_label"
    cw = ws.cell(row=6, column=3, value="=WithdrawalRate"); cw.style = "field_value"; cw.number_format = "0.0%"
    ws.cell(row=7, column=2, value="= YOUR FIRE NUMBER").style = "th"
    cf = ws.cell(row=7, column=3, value="=AnnualExpenses/WithdrawalRate"); cf.style = "td"; cf.font = Font(bold=True, size=14, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = '"$"#,##0'
    cell_name(wb, "FIRENumber", "FIRE Number", "$C$7")
    ws.cell(row=9, column=2, value="Your net worth today").style = "field_label"
    cnw = ws.cell(row=9, column=3, value="=NetWorth"); cnw.style = "field_value"; cnw.number_format = '"$"#,##0'
    ws.cell(row=10, column=2, value="= FIRE PROGRESS").style = "th"
    cp = ws.cell(row=10, column=3, value="=IFERROR(NetWorth/FIRENumber,0)"); cp.style = "td"; cp.font = Font(bold=True, size=13, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = "0%"
    cell_name(wb, "FIREProgress", "FIRE Number", "$C$10")
    ws.cell(row=12, column=2, value="COAST FIRE — invest nothing more, coast to retirement").style = "section_gold"
    ws.cell(row=13, column=2, value="Years to target retirement").style = "field_label"
    cy = ws.cell(row=13, column=3, value="=RetireAge-CurrentAge"); cy.style = "field_value"; cy.number_format = "0"
    ws.cell(row=14, column=2, value="= COAST FIRE NUMBER").style = "th"
    cc = ws.cell(row=14, column=3, value="=FIRENumber/(1+ReturnRate)^(RetireAge-CurrentAge)"); cc.style = "td"; cc.font = Font(bold=True, size=12, color=PRIMARY); cc.fill = fill(MINT_BG); cc.number_format = '"$"#,##0'
    cell_name(wb, "CoastNumber", "FIRE Number", "$C$14")
    ws.cell(row=15, column=2, value="Coast FIRE reached?").style = "field_label"
    cr = ws.cell(row=15, column=3, value='=IF(InvestedAssets>=CoastNumber,"YES — you can coast","Not yet")'); cr.style = "field_value"
    ws.cell(row=17, column=2, value="= YEARS TO FI (at your savings & return)").style = "th"
    cnf = ws.cell(row=17, column=3, value="=IFERROR(NPER(ReturnRate,-AnnualSavings,-InvestedAssets,FIRENumber),0)"); cnf.style = "td"; cnf.font = Font(bold=True, size=13, color=PRIMARY); cnf.fill = fill(MINT_BG); cnf.number_format = "0.0"
    cell_name(wb, "YearsToFI", "FIRE Number", "$C$17")


def build_income(wb):
    ws = wb.create_sheet("Income & Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 2])
    luxe_header(ws, "C", "💵  INCOME & EXPENSES",
                "Take-home income minus spending is your savings — and your savings RATE is what builds wealth.")
    ws.cell(row=5, column=2, value="Annual take-home income").style = "field_label"
    ci = ws.cell(row=5, column=3, value=ANNUAL_INCOME); ci.style = "input"; ci.number_format = '"$"#,##0'
    cell_name(wb, "AnnualIncome", "Income & Expenses", "$C$5")
    ws.cell(row=6, column=2, value="− Annual spending (from FIRE Number)").style = "field_label"
    ce = ws.cell(row=6, column=3, value="=AnnualExpenses"); ce.style = "field_value"; ce.number_format = '"$"#,##0'
    ws.cell(row=7, column=2, value="= ANNUAL SAVINGS").style = "th"
    cs = ws.cell(row=7, column=3, value="=AnnualIncome-AnnualExpenses"); cs.style = "td"; cs.font = Font(bold=True, size=13, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cell_name(wb, "AnnualSavings", "Income & Expenses", "$C$7")
    ws.cell(row=8, column=2, value="= SAVINGS RATE").style = "th"
    cr = ws.cell(row=8, column=3, value="=IFERROR(AnnualSavings/AnnualIncome,0)"); cr.style = "td"; cr.font = Font(bold=True, size=13, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = "0%"
    cell_name(wb, "SavingsRate", "Income & Expenses", "$C$8")
    ws.cell(row=10, column=2, value="Your savings rate is the single biggest lever on your years to FI.").style = "section_gold"


def build_accounts(wb):
    ws, start, end = build_log(
        wb, "Accounts", "🏦", "ACCOUNTS",
        "Every account and where it lives — so your whole net worth is one glance, never a guess.",
        ["Account", "Institution", "Type", "Balance"],
        ACCOUNTS, [2, 22, 20, 16, 16, 2], text_left={2, 3}, money={5}, reserved=24, start_col=2,
        validations=[("D", "AcctTypeList")])
    nrange(wb, "AcctBalance", "Accounts", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL ACCOUNTS").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(AcctBalance)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'


def build_contributions(wb):
    ws = wb.create_sheet("Contributions"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 22, 16, 2])
    luxe_header(ws, "D", "📈  CONTRIBUTIONS",
                "What you invest, and when. Consistent contributions are how the FIRE number gets hit.")
    ws.cell(row=5, column=2, value="Planned annual contributions").style = "field_label"
    ca = ws.cell(row=5, column=3, value=CONTRIB_ANNUAL); ca.style = "input"; ca.number_format = '"$"#,##0'
    cell_name(wb, "ContribAnnual", "Contributions", "$C$5")
    table_headers(ws, 7, ["Date", "Account", "Amount"], start_col=2)
    start = 8
    for i, (date, acct, amt) in enumerate(CONTRIB):
        r = start + i
        ws.cell(row=r, column=2, value=date).style = "td_left"
        ws.cell(row=r, column=3, value=acct).style = "td_left"
        cm = ws.cell(row=r, column=4, value=amt); cm.style = "input"; cm.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CONTRIB) - 1
    nrange(wb, "ContribAmt", "Contributions", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="LOGGED").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(ContribAmt)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A8"


def build_savingsrate(wb):
    ws = wb.create_sheet("Savings Rate"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 16, 16, 16, 14, 2])
    luxe_header(ws, "F", "💰  SAVINGS RATE",
                "Month by month — income in, spending out, and the percentage you kept. Watch it climb.")
    table_headers(ws, 4, ["Month", "Income", "Expenses", "Saved", "Rate"], start_col=2)
    start = L0
    for i, (m, inc, exp) in enumerate(SAVINGS_LOG):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cinc = ws.cell(row=r, column=3, value=inc); cinc.style = "input"; cinc.number_format = '"$"#,##0'
        cexp = ws.cell(row=r, column=4, value=exp); cexp.style = "input"; cexp.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cs.style = "td"; cs.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=6, value=f"=IFERROR((C{r}-D{r})/C{r},0)"); cr.style = "td"; cr.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SAVINGS_LOG) - 1
    ws.conditional_formatting.add(f"F{start}:F{end}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVG").style = "th"
    for c in (3, 4, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cr = ws.cell(row=tot, column=6, value=f"=IFERROR(AVERAGE(F{start}:F{end}),0)"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = "0%"
    ws.freeze_panes = "A5"


def build_projection(wb):
    ws = wb.create_sheet("Coast & Projection"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 20, 20, 2])
    luxe_header(ws, "D", "🚀  COAST & PROJECTION",
                "Where your invested assets land at each age if you keep saving — and when you cross your FIRE number.")
    table_headers(ws, 4, ["Age", "Years", "Projected Invested"], start_col=2)
    start = L0
    for i, age in enumerate(PROJ_AGES):
        r = start + i
        yrs = age - CURRENT_AGE
        ws.cell(row=r, column=2, value=age).style = "td"
        ws.cell(row=r, column=3, value=yrs).style = "td"
        # FV of invested + annuity of annual savings
        cp = ws.cell(row=r, column=4, value=f"=InvestedAssets*(1+ReturnRate)^{yrs}+IFERROR(AnnualSavings*((1+ReturnRate)^{yrs}-1)/ReturnRate,0)")
        cp.style = "td"; cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PROJ_AGES) - 1
    ws.cell(row=end + 2, column=2, value="At your FIRE number, work becomes optional — everything after is a choice.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_milestones(wb):
    ws = wb.create_sheet("Milestones"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 18, 14, 2])
    luxe_header(ws, "D", "🏁  MILESTONES",
                "The markers on the road to freedom — check them off as your net worth climbs.")
    table_headers(ws, 4, ["Milestone", "Target", "Reached?"], start_col=2)
    start = L0
    for i, (name, tgt, reached) in enumerate(MILESTONES):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ct = ws.cell(row=r, column=3, value=tgt); ct.style = "input"; ct.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=reached).style = "td"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MILESTONES) - 1
    add_dv(ws, f"D{start}:D{end}", "YesNoList")
    nrange(wb, "MilestoneReached", "Milestones", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="REACHED").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value='=COUNTIF(MilestoneReached,"Yes")&" / "&COUNTA(MilestoneReached)'); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG)
    ws.freeze_panes = "A5"


def build_trend(wb):
    ws = wb.create_sheet("Net Worth Trend"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "📈  NET WORTH TREND",
                "Your net worth, month by month — the line that should only go up over time.")
    ws.cell(row=5, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 6, ["Month", "Net Worth"], start_col=2)
    ts = 7
    for i, (m, nw) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cn = ws.cell(row=r, column=3, value=nw); cn.style = "input"; cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "NWTrend", "Net Worth Trend", "C", ts, te)
    ws.add_chart(_barchart(ws, "Net Worth by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🧭  NET WORTH & FIRE COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Net worth, your FIRE number, your progress & a FIRE Score — your whole picture, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("NET WORTH", "=NetWorth", "money"),
        ("TOTAL ASSETS", "=TotalAssets", "money"),
        ("TOTAL LIABILITIES", "=TotalLiab", "money"),
        ("FIRE NUMBER", "=FIRENumber", "money"),
        ("FIRE PROGRESS", "=FIREProgress", "pct"),
        ("ANNUAL EXPENSES", "=AnnualExpenses", "money"),
    ]
    row2 = [
        ("SAVINGS RATE", "=SavingsRate", "pct"),
        ("ANNUAL SAVINGS", "=AnnualSavings", "money"),
        ("COAST NUMBER", "=CoastNumber", "money"),
        ("YEARS TO FI", "=YearsToFI", "dec"),
        ("INVESTED ASSETS", "=InvestedAssets", "money"),
        ("FIRE SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "FIRE HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "NET WORTH BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Saving enough", "=IFERROR(MIN(SavingsRate/SavingsGoal,1),0)"),
        ("Emergency fund funded", "=IFERROR(MIN(CashAssets/EmergencyFundGoal,1),0)"),
        ("Positive net worth", "=IF(NetWorth>0,1,0)"),
        ("Coast FIRE reached", "=IFERROR(MIN(InvestedAssets/CoastNumber,1),0)"),
        ("Investing to goal", "=IFERROR(MIN(ContribAnnual/ContribGoal,1),0)"),
        ("Halfway to FI", "=IFERROR(MIN((InvestedAssets/FIRENumber)/HalfwayGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    tr = wb["Net Worth Trend"]
    ch = BarChart(); ch.type = "col"; ch.title = "Net Worth by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(tr, min_col=3, min_row=7, max_row=6 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(tr, min_col=2, min_row=7, max_row=6 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Net Worth & FIRE Command Center™ — know your number, and your progress to it.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_networth(wb); build_fire(wb)
    build_assets(wb); build_liabilities(wb); build_accounts(wb); build_contributions(wb)
    build_income(wb); build_savingsrate(wb); build_projection(wb); build_milestones(wb)
    build_trend(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Net Worth", "FIRE Number", "Assets", "Liabilities", "Accounts",
             "Contributions", "Income & Expenses", "Savings Rate", "Coast & Projection", "Milestones",
             "Net Worth Trend", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Networth_FIRE_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
