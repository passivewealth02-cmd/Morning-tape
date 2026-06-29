"""Build the Dividend Wealth Builder System (DWBS) Excel workbook.

Run: python3 build_xlsx.py
Outputs: ../Dividend_Wealth_Builder.xlsx
"""
from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
)
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"      # Deep forest green
ACCENT = "937356"       # Gold
SURFACE = "E5D3BA"      # Soft tan
HIGHLIGHT = "75E6C1"    # Mint
BG = "FFFFFF"
TEXT = "333333"
SUCCESS = "75E6C1"
WARNING = "937356"
DANGER = "C94C4C"
MUTED_ROW = "F4ECDE"    # ~rgba(229,211,186,0.35) flattened on white
BORDER = "D6D2C8"

SECTORS = ["Tech", "Finance", "Energy", "REITs", "Consumer", "Healthcare"]
RISK_LEVELS = ["Low", "Medium", "High"]
FREQUENCIES = ["Monthly", "Quarterly", "Semi-Annual", "Annual"]
FREQ_PAYMENTS = {"Monthly": 12, "Quarterly": 4, "Semi-Annual": 2, "Annual": 1}

THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def clean_labels(pct=False, val=False):
    """Concise data labels: never show series/category text (it collides on
    multi-slice pies). Percentages or values only."""
    dl = DataLabelList()
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    dl.showVal = val
    dl.showPercent = pct
    return dl


def no_labels():
    """Suppress all on-chart labels (legend carries meaning)."""
    return clean_labels(pct=False, val=False)


def register_styles(wb: Workbook) -> None:
    styles = {
        "title": NamedStyle(
            name="brand_title",
            font=Font(name="Calibri", size=22, bold=True, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "subtitle": NamedStyle(
            name="brand_subtitle",
            font=Font(name="Calibri", size=11, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "section": NamedStyle(
            name="brand_section",
            font=Font(name="Calibri", size=12, bold=True, color=PRIMARY),
            fill=PatternFill("solid", fgColor=BG),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "th": NamedStyle(
            name="brand_th",
            font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=BOX,
        ),
        "td": NamedStyle(
            name="brand_td",
            font=Font(name="Calibri", size=11, color=TEXT),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=BOX,
        ),
        "td_left": NamedStyle(
            name="brand_td_left",
            font=Font(name="Calibri", size=11, color=TEXT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=BOX,
        ),
        "kpi_label": NamedStyle(
            name="kpi_label",
            font=Font(name="Calibri", size=10, bold=True, color=ACCENT),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
        ),
        "kpi_value": NamedStyle(
            name="kpi_value",
            font=Font(name="Calibri", size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
        ),
        "kpi_value_money": NamedStyle(
            name="kpi_value_money",
            font=Font(name="Calibri", size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
            number_format='"$"#,##0.00',
        ),
        "kpi_value_pct": NamedStyle(
            name="kpi_value_pct",
            font=Font(name="Calibri", size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
            number_format="0.00%",
        ),
        "input": NamedStyle(
            name="brand_input",
            font=Font(name="Calibri", size=11, bold=True, color=PRIMARY),
            fill=PatternFill("solid", fgColor=SURFACE),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=BOX,
        ),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def merge_and_set(ws, rng, value, style):
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    cell = ws[top_left]
    cell.value = value
    cell.style = style
    for row in ws[rng]:
        for c in row:
            c.fill = style.fill
    return cell


# ---------------------------------------------------------------------------
# Sample data
# ---------------------------------------------------------------------------
SAMPLE_HOLDINGS = [
    # ticker, company, sector, shares, buy_price, current_price, div_per_share, frequency, risk
    ("AAPL", "Apple Inc.", "Tech", 50, 145.30, 192.40, 0.96, "Quarterly", 2),
    ("MSFT", "Microsoft Corp.", "Tech", 40, 250.10, 415.20, 3.00, "Quarterly", 2),
    ("JNJ", "Johnson & Johnson", "Healthcare", 60, 162.45, 158.20, 4.76, "Quarterly", 1),
    ("KO", "Coca-Cola Co.", "Consumer", 120, 56.10, 63.40, 1.84, "Quarterly", 1),
    ("PG", "Procter & Gamble", "Consumer", 45, 142.20, 161.80, 4.03, "Quarterly", 1),
    ("JPM", "JPMorgan Chase", "Finance", 35, 132.50, 198.10, 4.60, "Quarterly", 3),
    ("BAC", "Bank of America", "Finance", 150, 32.40, 39.85, 1.04, "Quarterly", 3),
    ("XOM", "Exxon Mobil", "Energy", 80, 95.20, 116.30, 3.80, "Quarterly", 3),
    ("CVX", "Chevron Corp.", "Energy", 30, 158.10, 162.40, 6.52, "Quarterly", 3),
    ("O", "Realty Income", "REITs", 200, 62.40, 58.10, 3.16, "Monthly", 2),
    ("STAG", "Stag Industrial", "REITs", 180, 35.20, 36.85, 1.48, "Monthly", 3),
    ("VYM", "Vanguard High Div ETF", "Finance", 100, 105.40, 124.80, 3.45, "Quarterly", 2),
    ("SCHD", "Schwab US Div ETF", "Finance", 220, 72.30, 78.40, 2.66, "Quarterly", 2),
    ("MAIN", "Main Street Capital", "Finance", 90, 41.10, 49.80, 2.94, "Monthly", 3),
    ("MO", "Altria Group", "Consumer", 110, 45.60, 49.20, 3.92, "Quarterly", 3),
]


# ===========================================================================
# Sheets
# ===========================================================================
def build_settings(wb: Workbook) -> None:
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 2
    ws.column_dimensions["E"].width = 26
    ws.column_dimensions["F"].width = 22
    for r in range(1, 50):
        ws.row_dimensions[r].height = 22

    ws.row_dimensions[1].height = 48
    merge_and_set(ws, "A1:F1", "  ⚙  SETTINGS & INPUTS", wb._named_styles[wb.style_names.index("brand_title")])
    merge_and_set(ws, "A2:F2", "  Configure portfolio goals, dropdowns, and constants.",
                  wb._named_styles[wb.style_names.index("brand_subtitle")])

    # User controls
    merge_and_set(ws, "B4:C4", "USER CONTROLS",
                  wb._named_styles[wb.style_names.index("brand_section")])

    controls = [
        ("Monthly Contribution",            500,      "kpi_value_money",  "MonthlyContribution"),
        ("Target Monthly Dividend Income",  5000,     "kpi_value_money",  "TargetMonthlyIncome"),
        ("Expected Annual Growth Rate",     0.08,     "kpi_value_pct",    "AnnualGrowthRate"),
        ("Dividend Growth Rate",            0.06,     "kpi_value_pct",    "DividendGrowthRate"),
        ("Reinvestment (1 = YES, 0 = NO)",  1,        None,               "ReinvestToggle"),
        ("Tax Rate on Dividends",           0.15,     "kpi_value_pct",    "TaxRate"),
        ("Years to Project",                30,       None,               "YearsToProject"),
    ]
    start = 5
    for i, (label, default, num_fmt, name) in enumerate(controls):
        r = start + i
        cell_lbl = ws.cell(row=r, column=2, value=label)
        cell_lbl.style = "brand_td_left"
        cell_lbl.fill = fill(BG)
        cell_val = ws.cell(row=r, column=3, value=default)
        cell_val.style = "brand_input"
        if num_fmt == "kpi_value_money":
            cell_val.number_format = '"$"#,##0.00'
        elif num_fmt == "kpi_value_pct":
            cell_val.number_format = "0.00%"
        else:
            cell_val.number_format = "0"
        wb.defined_names[name] = DefinedName(name, attr_text=f"Settings!$C${r}")

    # Dropdown lists
    merge_and_set(ws, "E4:F4", "DROPDOWN LISTS",
                  wb._named_styles[wb.style_names.index("brand_section")])

    ws.cell(row=5, column=5, value="Sectors").style = "brand_th"
    ws.cell(row=5, column=6, value="Risk / Frequency").style = "brand_th"
    for i, sector in enumerate(SECTORS):
        c = ws.cell(row=6 + i, column=5, value=sector)
        c.style = "brand_td_left"
    for i, level in enumerate(RISK_LEVELS):
        c = ws.cell(row=6 + i, column=6, value=level)
        c.style = "brand_td_left"
    for i, freq in enumerate(FREQUENCIES):
        c = ws.cell(row=6 + len(RISK_LEVELS) + i, column=6, value=freq)
        c.style = "brand_td_left"

    wb.defined_names["SectorList"] = DefinedName(
        "SectorList", attr_text=f"Settings!$E$6:$E${5 + len(SECTORS)}"
    )
    wb.defined_names["RiskList"] = DefinedName(
        "RiskList", attr_text=f"Settings!$F$6:$F${5 + len(RISK_LEVELS)}"
    )
    wb.defined_names["FrequencyList"] = DefinedName(
        "FrequencyList",
        attr_text=f"Settings!$F${6 + len(RISK_LEVELS)}:$F${5 + len(RISK_LEVELS) + len(FREQUENCIES)}",
    )

    # Frequency-payments lookup table (used by calc engine)
    ws.cell(row=18, column=5, value="Frequency").style = "brand_th"
    ws.cell(row=18, column=6, value="Payments/Yr").style = "brand_th"
    for i, freq in enumerate(FREQUENCIES):
        ws.cell(row=19 + i, column=5, value=freq).style = "brand_td_left"
        c = ws.cell(row=19 + i, column=6, value=FREQ_PAYMENTS[freq])
        c.style = "brand_td"
    wb.defined_names["FreqTable"] = DefinedName(
        "FreqTable",
        attr_text=f"Settings!$E$19:$F${18 + len(FREQUENCIES)}",
    )


def build_holdings(wb: Workbook) -> None:
    ws = wb.create_sheet("Holdings")
    ws.sheet_view.showGridLines = False
    widths = [12, 28, 14, 10, 14, 14, 16, 14, 16, 16, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 48
    merge_and_set(ws, "A1:L1", "  📂  HOLDINGS DATABASE",
                  wb._named_styles[wb.style_names.index("brand_title")])
    ws.row_dimensions[2].height = 22
    merge_and_set(ws, "A2:L2", "  Add / edit holdings. Formulas auto-calculate.",
                  wb._named_styles[wb.style_names.index("brand_subtitle")])

    headers = [
        "Ticker", "Company", "Sector", "Shares", "Buy Price", "Current Price",
        "Total Value", "Div / Share", "Frequency", "Annual Income", "Yield %", "Risk (1-5)",
    ]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.style = "brand_th"
    ws.row_dimensions[4].height = 30

    # Reserve 60 rows for holdings (15 sample + room to grow)
    max_rows = 60
    start_row = 5
    end_row = start_row + max_rows - 1

    for i, h in enumerate(SAMPLE_HOLDINGS):
        r = start_row + i
        ticker, company, sector, shares, buy, current, dps, freq, risk = h
        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=company)
        ws.cell(row=r, column=3, value=sector)
        ws.cell(row=r, column=4, value=shares)
        ws.cell(row=r, column=5, value=buy)
        ws.cell(row=r, column=6, value=current)
        ws.cell(row=r, column=8, value=dps)
        ws.cell(row=r, column=9, value=freq)
        ws.cell(row=r, column=12, value=risk)

    # Formulas for every reserved row
    for r in range(start_row, end_row + 1):
        # Total Value = Shares * Current Price
        ws.cell(row=r, column=7,
                value=f'=IF(D{r}="","",D{r}*F{r})')
        # Annual Income = Shares * Div/Share * Payments per Year (via FreqTable)
        ws.cell(row=r, column=10,
                value=(f'=IF(OR(D{r}="",H{r}="",I{r}=""),"",'
                       f'D{r}*H{r}*VLOOKUP(I{r},FreqTable,2,FALSE))'))
        # Yield % = (Div/Share * Pmts) / Current Price
        ws.cell(row=r, column=11,
                value=(f'=IFERROR(IF(OR(F{r}="",H{r}="",I{r}=""),"",'
                       f'(H{r}*VLOOKUP(I{r},FreqTable,2,FALSE))/F{r}),"")'))

    # Style the data rows
    money_cols = {5, 6, 7, 8, 10}
    pct_cols = {11}
    int_cols = {4, 12}
    text_left_cols = {1, 2}
    for r in range(start_row, end_row + 1):
        for col in range(1, 13):
            cell = ws.cell(row=r, column=col)
            if col in text_left_cols:
                cell.style = "brand_td_left"
            else:
                cell.style = "brand_td"
            if (r - start_row) % 2 == 1:
                cell.fill = fill(MUTED_ROW)
            else:
                cell.fill = fill(BG)
            if col in money_cols:
                cell.number_format = '"$"#,##0.00'
            elif col in pct_cols:
                cell.number_format = "0.00%"
            elif col in int_cols:
                cell.number_format = "0"

    # Data validation: Sector, Frequency, Risk
    dv_sector = DataValidation(type="list", formula1="=SectorList", allow_blank=True)
    dv_freq = DataValidation(type="list", formula1="=FrequencyList", allow_blank=True)
    dv_risk = DataValidation(type="whole", operator="between", formula1=1, formula2=5, allow_blank=True)
    ws.add_data_validation(dv_sector)
    ws.add_data_validation(dv_freq)
    ws.add_data_validation(dv_risk)
    dv_sector.add(f"C{start_row}:C{end_row}")
    dv_freq.add(f"I{start_row}:I{end_row}")
    dv_risk.add(f"L{start_row}:L{end_row}")

    # Conditional formatting
    # Yield: 3-color scale on column K
    ws.conditional_formatting.add(
        f"K{start_row}:K{end_row}",
        ColorScaleRule(start_type="min", start_color="FFE5D3BA",
                       mid_type="percentile", mid_value=50, mid_color="FFFFE08C",
                       end_type="max", end_color="FF75E6C1"),
    )
    # Risk: red bar
    ws.conditional_formatting.add(
        f"L{start_row}:L{end_row}",
        DataBarRule(start_type="num", start_value=1, end_type="num", end_value=5,
                    color=DANGER, showValue=True),
    )
    # Highlight rows where current < buy (loss)
    ws.conditional_formatting.add(
        f"A{start_row}:L{end_row}",
        FormulaRule(formula=[f'AND($F{start_row}<>"",$F{start_row}<$E{start_row})'],
                    fill=PatternFill("solid", fgColor="FBE6E6")),
    )

    # Defined ranges used elsewhere
    wb.defined_names["TblTotalValue"] = DefinedName(
        "TblTotalValue", attr_text=f"Holdings!$G${start_row}:$G${end_row}"
    )
    wb.defined_names["TblAnnualIncome"] = DefinedName(
        "TblAnnualIncome", attr_text=f"Holdings!$J${start_row}:$J${end_row}"
    )
    wb.defined_names["TblSector"] = DefinedName(
        "TblSector", attr_text=f"Holdings!$C${start_row}:$C${end_row}"
    )
    wb.defined_names["TblFrequency"] = DefinedName(
        "TblFrequency", attr_text=f"Holdings!$I${start_row}:$I${end_row}"
    )
    wb.defined_names["TblRisk"] = DefinedName(
        "TblRisk", attr_text=f"Holdings!$L${start_row}:$L${end_row}"
    )
    wb.defined_names["TblYield"] = DefinedName(
        "TblYield", attr_text=f"Holdings!$K${start_row}:$K${end_row}"
    )
    wb.defined_names["TblTicker"] = DefinedName(
        "TblTicker", attr_text=f"Holdings!$A${start_row}:$A${end_row}"
    )

    ws.freeze_panes = "A5"


def build_calculations(wb: Workbook) -> None:
    ws = wb.create_sheet("Calculations")
    ws.sheet_view.showGridLines = False
    for col, w in enumerate([2, 30, 20, 4, 22, 18, 18, 18], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 48
    merge_and_set(ws, "A1:H1", "  🧮  CALCULATIONS ENGINE",
                  wb._named_styles[wb.style_names.index("brand_title")])
    merge_and_set(ws, "A2:H2", "  Centralized logic. Do not edit values — formulas only.",
                  wb._named_styles[wb.style_names.index("brand_subtitle")])

    # Portfolio aggregates
    merge_and_set(ws, "B4:C4", "PORTFOLIO AGGREGATES",
                  wb._named_styles[wb.style_names.index("brand_section")])

    rows = [
        ("Total Portfolio Value",        "=SUMPRODUCT((TblTotalValue<>\"\")*IFERROR(TblTotalValue,0))", '"$"#,##0.00', "TotalPortfolioValue"),
        ("Total Annual Dividend Income", "=SUMPRODUCT((TblAnnualIncome<>\"\")*IFERROR(TblAnnualIncome,0))", '"$"#,##0.00', "TotalAnnualIncome"),
        ("Total Monthly Dividend Income","=TotalAnnualIncome/12", '"$"#,##0.00', "TotalMonthlyIncome"),
        ("Portfolio Yield (Weighted)",   "=IFERROR(TotalAnnualIncome/TotalPortfolioValue,0)", "0.00%", "PortfolioYield"),
        ("After-Tax Monthly Income",     "=TotalMonthlyIncome*(1-TaxRate)", '"$"#,##0.00', "AfterTaxMonthly"),
        ("Holdings Count",               '=SUMPRODUCT((TblTicker<>"")*1)', "0", "HoldingsCount"),
        ("Progress to Income Goal",      "=IFERROR(TotalMonthlyIncome/TargetMonthlyIncome,0)", "0.00%", "ProgressToGoal"),
    ]
    for i, (label, formula, fmt, name) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=2, value=label).style = "brand_td_left"
        c = ws.cell(row=r, column=3, value=formula)
        c.style = "brand_td"
        c.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Calculations!$C${r}")

    # Sector breakdown
    merge_and_set(ws, "E4:H4", "SECTOR BREAKDOWN",
                  wb._named_styles[wb.style_names.index("brand_section")])
    ws.cell(row=5, column=5, value="Sector").style = "brand_th"
    ws.cell(row=5, column=6, value="Value").style = "brand_th"
    ws.cell(row=5, column=7, value="Annual Income").style = "brand_th"
    ws.cell(row=5, column=8, value="Allocation %").style = "brand_th"
    for i, sector in enumerate(SECTORS):
        r = 6 + i
        ws.cell(row=r, column=5, value=sector).style = "brand_td_left"
        v = ws.cell(row=r, column=6,
                    value=f'=SUMIFS(TblTotalValue,TblSector,E{r})')
        v.style = "brand_td"; v.number_format = '"$"#,##0.00'
        ai = ws.cell(row=r, column=7,
                     value=f'=SUMIFS(TblAnnualIncome,TblSector,E{r})')
        ai.style = "brand_td"; ai.number_format = '"$"#,##0.00'
        a = ws.cell(row=r, column=8,
                    value=f'=IFERROR(F{r}/TotalPortfolioValue,0)')
        a.style = "brand_td"; a.number_format = "0.00%"

    sector_end = 5 + len(SECTORS)
    wb.defined_names["SectorLabels"] = DefinedName(
        "SectorLabels", attr_text=f"Calculations!$E$6:$E${sector_end}"
    )
    wb.defined_names["SectorValues"] = DefinedName(
        "SectorValues", attr_text=f"Calculations!$F$6:$F${sector_end}"
    )
    wb.defined_names["SectorIncome"] = DefinedName(
        "SectorIncome", attr_text=f"Calculations!$G$6:$G${sector_end}"
    )

    # Monthly dividend pattern (income distribution across calendar months)
    month_start = sector_end + 3
    merge_and_set(ws, f"B{month_start}:H{month_start}", "MONTHLY DIVIDEND DISTRIBUTION",
                  wb._named_styles[wb.style_names.index("brand_section")])
    header_row = month_start + 1
    ws.cell(row=header_row, column=2, value="Month").style = "brand_th"
    ws.cell(row=header_row, column=3, value="Monthly $ (Even Spread)").style = "brand_th"
    ws.cell(row=header_row, column=4, value="Cumulative").style = "brand_th"
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    for i, m in enumerate(months):
        r = header_row + 1 + i
        ws.cell(row=r, column=2, value=m).style = "brand_td_left"
        c = ws.cell(row=r, column=3, value="=TotalAnnualIncome/12")
        c.style = "brand_td"; c.number_format = '"$"#,##0.00'
        cum = ws.cell(row=r, column=4,
                      value=f"=SUM($C${header_row+1}:C{r})")
        cum.style = "brand_td"; cum.number_format = '"$"#,##0.00'
    wb.defined_names["MonthLabels"] = DefinedName(
        "MonthLabels", attr_text=f"Calculations!$B${header_row+1}:$B${header_row+12}"
    )
    wb.defined_names["MonthValues"] = DefinedName(
        "MonthValues", attr_text=f"Calculations!$C${header_row+1}:$C${header_row+12}"
    )


def build_projections(wb: Workbook) -> None:
    ws = wb.create_sheet("Projections")
    ws.sheet_view.showGridLines = False
    widths = [2, 10, 22, 22, 22, 22, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 48
    merge_and_set(ws, "A1:G1", "  🚀  PROJECTIONS & FIRE MODEL",
                  wb._named_styles[wb.style_names.index("brand_title")])
    merge_and_set(ws, "A2:G2", "  Compounding projection driven by Settings inputs.",
                  wb._named_styles[wb.style_names.index("brand_subtitle")])

    # FIRE summary
    merge_and_set(ws, "B4:G4", "FIRE METRICS",
                  wb._named_styles[wb.style_names.index("brand_section")])

    fire_rows = [
        ("FIRE Number (25× Annual Spend Goal)",
         "=TargetMonthlyIncome*12*25", '"$"#,##0.00', "FireNumber"),
        ("Years to $1,000 / month Passive",
         '=IFERROR(IF(TotalMonthlyIncome>=1000,0,'
         'LOG((1000/IFERROR((TotalMonthlyIncome+0.0001),0.0001)))/LOG(1+DividendGrowthRate)),"—")',
         "0.0", "YearsTo1k"),
        ("Years to $5,000 / month Passive",
         '=IFERROR(IF(TotalMonthlyIncome>=5000,0,'
         'LOG((5000/IFERROR((TotalMonthlyIncome+0.0001),0.0001)))/LOG(1+DividendGrowthRate)),"—")',
         "0.0", "YearsTo5k"),
        ("Years to Target Income",
         '=IFERROR(IF(TotalMonthlyIncome>=TargetMonthlyIncome,0,'
         'LOG((TargetMonthlyIncome/IFERROR((TotalMonthlyIncome+0.0001),0.0001)))/LOG(1+DividendGrowthRate)),"—")',
         "0.0", "YearsToTarget"),
    ]
    for i, (label, f, fmt, name) in enumerate(fire_rows):
        r = 5 + i
        ws.cell(row=r, column=2, value=label).style = "brand_td_left"
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
        c = ws.cell(row=r, column=3, value=f)
        c.style = "brand_td"; c.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Projections!$C${r}")

    # Year-by-year projection table
    merge_and_set(ws, "B11:G11", "YEAR-BY-YEAR PROJECTION",
                  wb._named_styles[wb.style_names.index("brand_section")])
    headers = ["Year", "Portfolio Value", "Annual Income", "Monthly Income",
               "Cumulative Contributions", "% to FIRE"]
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=12, column=i, value=h)
        c.style = "brand_th"
    ws.row_dimensions[12].height = 30

    # Year 1 row uses current portfolio and current income; subsequent rows compound
    proj_start = 13
    proj_end = 13 + 30 - 1  # 30 years
    for i in range(30):
        r = proj_start + i
        year = i + 1
        ws.cell(row=r, column=2, value=year).style = "brand_td"
        if i == 0:
            ws.cell(row=r, column=3,
                    value="=TotalPortfolioValue*(1+AnnualGrowthRate)+MonthlyContribution*12")
            ws.cell(row=r, column=4,
                    value="=TotalAnnualIncome*(1+DividendGrowthRate)")
            ws.cell(row=r, column=6, value="=MonthlyContribution*12")
        else:
            ws.cell(row=r, column=3,
                    value=f"=C{r-1}*(1+AnnualGrowthRate)+MonthlyContribution*12+IF(ReinvestToggle=1,D{r-1},0)")
            ws.cell(row=r, column=4,
                    value=f"=D{r-1}*(1+DividendGrowthRate)+IF(ReinvestToggle=1,C{r-1}*PortfolioYield*0.05,0)")
            ws.cell(row=r, column=6,
                    value=f"=F{r-1}+MonthlyContribution*12")
        ws.cell(row=r, column=5, value=f"=D{r}/12")
        ws.cell(row=r, column=7, value=f"=IFERROR(C{r}/FireNumber,0)")

        for col in range(2, 8):
            cell = ws.cell(row=r, column=col)
            cell.border = BOX
            if cell.style != "brand_td":
                cell.style = "brand_td"
            if (r - proj_start) % 2 == 1:
                cell.fill = fill(MUTED_ROW)
            else:
                cell.fill = fill(BG)
        for col in (3, 4, 5, 6):
            ws.cell(row=r, column=col).number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7).number_format = "0.00%"
        ws.cell(row=r, column=2).number_format = "0"

    wb.defined_names["ProjYears"] = DefinedName(
        "ProjYears", attr_text=f"Projections!$B${proj_start}:$B${proj_end}"
    )
    wb.defined_names["ProjPortfolio"] = DefinedName(
        "ProjPortfolio", attr_text=f"Projections!$C${proj_start}:$C${proj_end}"
    )
    wb.defined_names["ProjAnnualIncome"] = DefinedName(
        "ProjAnnualIncome", attr_text=f"Projections!$D${proj_start}:$D${proj_end}"
    )
    wb.defined_names["ProjMonthlyIncome"] = DefinedName(
        "ProjMonthlyIncome", attr_text=f"Projections!$E${proj_start}:$E${proj_end}"
    )
    wb.defined_names["ProjContrib"] = DefinedName(
        "ProjContrib", attr_text=f"Projections!$F${proj_start}:$F${proj_end}"
    )

    # Charts: wealth growth + dividend income curve
    wealth = LineChart()
    wealth.title = "Wealth Growth Curve"
    wealth.style = 2
    wealth.height = 9
    wealth.width = 18
    data = Reference(ws, min_col=3, min_row=12, max_col=3, max_row=proj_end)
    cats = Reference(ws, min_col=2, min_row=13, max_row=proj_end)
    wealth.add_data(data, titles_from_data=True)
    wealth.set_categories(cats)
    ws.add_chart(wealth, "I4")

    income_chart = LineChart()
    income_chart.title = "Annual Dividend Income Curve"
    income_chart.style = 2
    income_chart.height = 9
    income_chart.width = 18
    data2 = Reference(ws, min_col=4, min_row=12, max_col=4, max_row=proj_end)
    income_chart.add_data(data2, titles_from_data=True)
    income_chart.set_categories(cats)
    ws.add_chart(income_chart, "I24")

    contrib_vs_income = BarChart()
    contrib_vs_income.type = "col"
    contrib_vs_income.title = "Contributions vs Passive Income"
    contrib_vs_income.style = 2
    contrib_vs_income.height = 9
    contrib_vs_income.width = 18
    data3 = Reference(ws, min_col=6, min_row=12, max_col=6, max_row=proj_end)
    data4 = Reference(ws, min_col=4, min_row=12, max_col=4, max_row=proj_end)
    contrib_vs_income.add_data(data3, titles_from_data=True)
    contrib_vs_income.add_data(data4, titles_from_data=True)
    contrib_vs_income.set_categories(cats)
    ws.add_chart(contrib_vs_income, "I44")


def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    for col, w in enumerate([2, 22, 22, 22, 22, 22, 22, 22, 22, 22, 22], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Brand navbar
    ws.row_dimensions[1].height = 56
    merge_and_set(ws, "A1:K1", "  💎  DIVIDEND WEALTH BUILDER",
                  wb._named_styles[wb.style_names.index("brand_title")])
    ws.row_dimensions[2].height = 22
    merge_and_set(ws, "A2:K2", "  Track. Optimize. Reach financial independence through dividends.",
                  wb._named_styles[wb.style_names.index("brand_subtitle")])

    # KPI row (6 cards across cols B..K, two cols per card)
    kpis = [
        ("💰 Portfolio Value",      "=TotalPortfolioValue",  "kpi_value_money"),
        ("📈 Monthly Income",       "=TotalMonthlyIncome",   "kpi_value_money"),
        ("📊 Annual Income",        "=TotalAnnualIncome",    "kpi_value_money"),
        ("📉 Average Yield",        "=PortfolioYield",       "kpi_value_pct"),
        ("🔁 Dividend Growth",      "=DividendGrowthRate",   "kpi_value_pct"),
        ("🎯 Progress to Goal",     "=ProgressToGoal",       "kpi_value_pct"),
    ]
    col = 2
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 50
    for label, formula, val_style in kpis:
        rng_label = f"{get_column_letter(col)}4:{get_column_letter(col+1)}4"
        rng_value = f"{get_column_letter(col)}5:{get_column_letter(col+1)}5"
        merge_and_set(ws, rng_label, label, wb._named_styles[wb.style_names.index("kpi_label")])
        merge_and_set(ws, rng_value, formula, wb._named_styles[wb.style_names.index(val_style)])
        # Box border for the card
        for r in (4, 5):
            for cc in range(col, col + 2):
                cell = ws.cell(row=r, column=cc)
                cell.border = BOX
        col += 2  # leave no gap, cards sit side-by-side

    # Conditional formatting on Progress card
    progress_cell = f"{get_column_letter(12)}5"  # not used; we keep simple
    # KPI sub-row: filters / status
    ws.row_dimensions[7].height = 22
    merge_and_set(ws, "B7:K7", "FILTERS — Use these dropdowns to focus the holdings view (data filters live on the Holdings sheet)",
                  wb._named_styles[wb.style_names.index("brand_section")])

    # Filter dropdowns (drive nothing programmatically — they are guide dropdowns
    # that mirror the Holdings AutoFilter; users apply via Holdings filter dropdowns).
    filters = [
        ("Asset Sector",       "B8", "SectorList"),
        ("Risk Level",         "D8", "RiskList"),
        ("Dividend Frequency", "F8", "FrequencyList"),
    ]
    for label, target, list_name in filters:
        col_letter = target[0]
        ws.cell(row=8, column=ord(col_letter) - 64, value=label).style = "brand_section"
        cell = ws.cell(row=9, column=ord(col_letter) - 64)
        cell.value = ""
        cell.style = "brand_input"
        ws.column_dimensions[col_letter].width = 24
        dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(cell.coordinate)

    # Charts pulled from Calculations
    # 1) Sector Allocation Pie
    pie = PieChart()
    pie.title = "Portfolio Allocation by Sector"
    labels = Reference(wb["Calculations"], min_col=5, min_row=6,
                       max_row=5 + len(SECTORS))
    data = Reference(wb["Calculations"], min_col=6, min_row=5,
                     max_row=5 + len(SECTORS))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.height = 9
    pie.width = 13
    pie.dataLabels = clean_labels(pct=True)
    ws.add_chart(pie, "B12")

    # 2) Monthly Dividend Bar
    bar = BarChart()
    bar.type = "col"
    bar.title = "Monthly Dividend Income"
    bar.height = 9
    bar.width = 13
    # MonthLabels begins on Calculations after sector breakdown
    bar_data = Reference(wb["Calculations"], min_col=3, min_row=12, max_row=24)
    bar_cats = Reference(wb["Calculations"], min_col=2, min_row=13, max_row=24)
    bar.add_data(bar_data, titles_from_data=True)
    bar.set_categories(bar_cats)
    ws.add_chart(bar, "G12")

    # 3) Wealth growth line (from Projections)
    line = LineChart()
    line.title = "Projected Wealth Growth"
    line.height = 9
    line.width = 13
    line_data = Reference(wb["Projections"], min_col=3, min_row=12,
                          max_col=3, max_row=42)
    line_cats = Reference(wb["Projections"], min_col=2, min_row=13, max_row=42)
    line.add_data(line_data, titles_from_data=True)
    line.set_categories(line_cats)
    ws.add_chart(line, "B30")

    # 4) Yield vs Risk scatter
    scatter = ScatterChart()
    scatter.title = "Yield vs Risk Score"
    scatter.style = 2
    scatter.height = 9
    scatter.width = 13
    scatter.x_axis.title = "Risk"
    scatter.y_axis.title = "Yield %"
    xvalues = Reference(wb["Holdings"], min_col=12, min_row=5, max_row=64)
    yvalues = Reference(wb["Holdings"], min_col=11, min_row=5, max_row=64)
    series = Series(yvalues, xvalues, title="Holdings")
    scatter.series.append(series)
    ws.add_chart(scatter, "G30")

    # Footer / signature row
    ws.row_dimensions[48].height = 30
    merge_and_set(ws, "B48:K48",
                  "  Dividend Wealth Builder System v1.0  •  Inputs live on Settings  •  Add holdings on Holdings",
                  wb._named_styles[wb.style_names.index("brand_subtitle")])


# ===========================================================================
# Build
# ===========================================================================
def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    # Build supporting sheets first so defined names exist when Dashboard
    # references them.
    build_settings(wb)
    build_holdings(wb)
    build_calculations(wb)
    build_projections(wb)
    build_dashboard(wb)

    # Reorder
    order = ["Dashboard", "Holdings", "Calculations", "Settings", "Projections"]
    wb._sheets = [wb[name] for name in order]

    # Set tab colors
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Holdings"].sheet_properties.tabColor = ACCENT
    wb["Calculations"].sheet_properties.tabColor = HIGHLIGHT
    wb["Settings"].sheet_properties.tabColor = SURFACE
    wb["Projections"].sheet_properties.tabColor = PRIMARY

    out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(out_dir, "Dividend_Wealth_Builder.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
