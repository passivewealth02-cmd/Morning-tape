"""Build Calm Compass™ — Anxiety & Social Confidence Planner (PREMIUM).

15 sheets · a calm, premium wellness operating system.
Welcome · Dashboard · Daily Check-In · Habits · Daily Planner · Social Prep ·
Reflection · Goals · Self-Care · Sleep · Exercise · Gratitude · Resources ·
Progress · Settings

Wellness & organization tool only — NOT medical advice (see Welcome sheet).

Run: python3 build_xlsx.py
Outputs: ../Calm_Compass.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens (calm + premium)
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"
ACCENT = "937356"
GOLD_LT = "C9A86A"
SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"
MINT_BG = "E3F8EF"
BG = "FFFFFF"
WHITE = "FFFFFF"
TEXT = "333333"
DANGER = "C94C4C"
RED_BG = "FBE6E6"
WARN_BG = "FBF0E2"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"
IVORY = "FBF8F2"

HABITS = ["Meditation", "Deep Breathing", "Exercise", "Reading", "Walking",
          "Healthy Meals", "Sleep Routine", "Gratitude", "Limit Screens"]
GOAL_CATS = ["Health", "Career", "Social", "Personal", "Financial",
             "Learning", "Relationships"]
INTENSITY = ["Low", "Moderate", "High"]
QUALITY = ["Poor", "Fair", "Good", "Great"]
SELFCARE_CATS = ["Physical", "Emotional", "Social", "Creative", "Rest", "Outdoor"]
STATUS = ["Not Started", "In Progress", "Complete"]
RES_TYPES = ["Book", "Podcast", "Video", "Article", "App", "Affirmation", "Contact"]
YESNO = ["Yes", "No"]
SCALE = list(range(1, 11))

LOG_ROWS = 60          # reserved rows for daily logs
L0 = 5                 # first data row
L1 = L0 + LOG_ROWS - 1

THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
# Styles & helpers
# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"),
                            fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True),
                               fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY),
                              alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT),
                                   alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"),
                         fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT),
                              alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
                              border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY),
                            fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX),
        "prompt": NamedStyle(name="prompt", font=f(11, False, TEXT, italic=True),
                             alignment=Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True),
                             border=BOX, fill=PatternFill("solid", fgColor=IVORY)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT),
                           alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    from openpyxl.utils import column_index_from_string
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def add_dv_whole(ws, rng, lo, hi):
    dv = DataValidation(type="whole", operator="between", formula1=lo, formula2=hi, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None,
               dates=None, pcts=None, start_col=1):
    text_left = text_left or set()
    money = money or set()
    ints = ints or set()
    dates = dates or set()
    pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else BG)
            if c in money:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}")
    ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]
    lc.value = label
    lc.font = Font(size=10, bold=True, color=ACCENT)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]
    vc.value = formula
    vc.font = Font(size=23, bold=True, color=PRIMARY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "dec": "0.0", "pct": "0%",
                        "hrs": '0.0" hrs"', "min": '0" min"'}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc)
            c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN,
                              top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 24
    ws.row_dimensions[row + 1].height = 52


def today_minus(n):
    return dt.date.today() - dt.timedelta(days=n)


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 4, 18, 18, 18, 18, 18])
    luxe_header(ws, "I", "⚙  SETTINGS", "Personalize your planner. Set once — the dashboards follow.")

    merge_set(ws, "B5:C5", "YOUR PREFERENCES", "section")
    controls = [
        ("Your Name", "Friend", None, "UserName"),
        ("Week Starts On", "Monday", None, "WeekStart"),
        ("Daily Habit Goal (count)", 7, "0", "HabitGoal"),
        ("Target Sleep Hours", 8, "0.0", "SleepTarget"),
        ("Mindfulness Goal (min/day)", 10, "0", "MindGoal"),
    ]
    for i, (lab, val, fmt, name) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val)
        c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Settings!$C${r}")

    merge_set(ws, "E5:I5", "DROPDOWN LISTS", "section_gold")
    lists = [
        ("E", "Scale 1–10", SCALE, "ScaleList"),
        ("F", "Yes / No", YESNO, "YesNoList"),
        ("G", "Goal Category", GOAL_CATS, "GoalCatList"),
        ("H", "Status", STATUS, "StatusList"),
        ("I", "Intensity", INTENSITY, "IntensityList"),
    ]
    for col, h, data, name in lists:
        ci = ord(col) - 64
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[name] = DefinedName(
            name, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")

    start2 = 20
    lists2 = [
        ("E", "Sleep Quality", QUALITY, "QualityList"),
        ("F", "Self-Care Type", SELFCARE_CATS, "SelfCareList"),
        ("G", "Resource Type", RES_TYPES, "ResTypeList"),
    ]
    for col, h, data, name in lists2:
        ci = ord(col) - 64
        ws.cell(row=start2, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=start2 + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[name] = DefinedName(
            name, attr_text=f"Settings!${col}${start2+1}:${col}${start2 + len(data)}")


# ===========================================================================
# Welcome / onboarding (with required disclaimer)
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 70, 3])
    ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🌿  CALM COMPASS™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Your calm, premium wellness operating system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    merge_set(ws, "B5:B5", "WELCOME", "section_gold")
    intro = (
        "Calm Compass™ is a gentle, organized space to build healthy routines, "
        "notice your patterns, prepare for stressful moments, and celebrate your "
        "progress over time. Take it one small step at a time — there's no wrong way "
        "to use it.")
    ws.merge_cells("B6:B8")
    ws["B6"].value = intro
    ws["B6"].style = "body"
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 22

    merge_set(ws, "B10:B10", "HOW TO USE IT", "section")
    steps = [
        "1.  Open Settings and add your name & simple goals (sleep, habits, mindfulness).",
        "2.  Each day, fill in the Daily Check-In (mood, energy, sleep, a few notes).",
        "3.  Tick off your Habits — streaks and your weekly routine score update for you.",
        "4.  Use the Daily Planner, Social Prep, and Reflection pages whenever they help.",
        "5.  Visit the Dashboard and Progress tabs to see your trends and wins over time.",
    ]
    for i, s in enumerate(steps):
        r = 11 + i
        ws.merge_cells(f"B{r}:B{r}")
        ws[f"B{r}"].value = s
        ws[f"B{r}"].style = "body"
        ws.row_dimensions[r].height = 22

    # Disclaimer box (prominent)
    dr = 18
    merge_set(ws, f"B{dr}:B{dr}", "  IMPORTANT — PLEASE READ", "th")
    ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+4}")
    disc = (
        "This planner is designed as a wellness and organization tool. It is not a "
        "medical device and is not a substitute for professional mental health care. "
        "If you are experiencing severe anxiety or emotional distress, please consider "
        "reaching out to a qualified mental health professional. If you are in crisis, "
        "contact your local emergency services or a crisis line in your country right away.")
    c = ws[f"B{dr+1}"]
    c.value = disc
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT)
    c.fill = fill(WARN_BG)
    c.border = BOX
    for rr in range(dr + 1, dr + 5):
        ws.row_dimensions[rr].height = 22
        ws.cell(row=rr, column=2).fill = fill(WARN_BG)

    merge_set(ws, f"B{dr+6}:B{dr+6}",
              "You've got this — one calm day at a time. 🌿", "section_gold")


# ===========================================================================
# Daily Check-In
# ===========================================================================
def build_checkin(wb):
    ws = wb.create_sheet("Check-In")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 9, 9, 9, 11, 9, 11, 12, 12, 30, 4, 12])
    luxe_header(ws, "L", "📝  DAILY CHECK-IN",
                "A 30-second daily snapshot. Be honest and kind to yourself.")
    headers = ["Date", "Mood", "Energy", "Stress", "Sleep Hrs", "Water",
               "Exercised", "Mindful Min", "Journaled", "Notes"]
    table_headers(ws, 4, headers)
    ws.cell(row=4, column=12, value="Streak~").style = "th"  # helper

    sample = [
        (13, 6, 5, 6, 6.5, 6, "Yes", 10, "Yes", "Felt rushed in the morning"),
        (12, 7, 7, 4, 7.5, 7, "Yes", 15, "Yes", "Good walk at lunch"),
        (11, 5, 4, 7, 6.0, 5, "No", 5, "Yes", "Tough meeting, breathed through it"),
        (10, 8, 8, 3, 8.0, 8, "Yes", 20, "Yes", "Rested day, felt calm"),
        (9, 6, 6, 5, 7.0, 6, "Yes", 10, "No", ""),
        (8, 7, 7, 4, 7.5, 7, "Yes", 10, "Yes", "Called a friend"),
        (7, 5, 5, 6, 6.5, 5, "No", 0, "Yes", "Low energy, early night"),
        (6, 8, 7, 3, 8.0, 8, "Yes", 15, "Yes", "Morning meditation helped"),
        (5, 7, 6, 4, 7.0, 7, "Yes", 10, "Yes", ""),
        (4, 6, 6, 5, 7.0, 6, "No", 5, "Yes", "Busy but steady"),
        (3, 8, 8, 2, 8.5, 8, "Yes", 20, "Yes", "Great day outdoors"),
        (2, 7, 7, 4, 7.5, 7, "Yes", 10, "Yes", ""),
        (1, 6, 5, 5, 6.5, 6, "Yes", 10, "No", "Slightly anxious before call"),
        (0, 7, 7, 4, 7.5, 7, "Yes", 15, "Yes", "Used Social Prep — went well!"),
    ]
    for i, row in enumerate(reversed(sample)):
        r = L0 + i
        days_ago, mood, energy, stress, sleep, water, ex, mind, jour, notes = row
        ws.cell(row=r, column=1, value=today_minus(days_ago))
        ws.cell(row=r, column=2, value=mood)
        ws.cell(row=r, column=3, value=energy)
        ws.cell(row=r, column=4, value=stress)
        ws.cell(row=r, column=5, value=sleep)
        ws.cell(row=r, column=6, value=water)
        ws.cell(row=r, column=7, value=ex)
        ws.cell(row=r, column=8, value=mind)
        ws.cell(row=r, column=9, value=jour)
        ws.cell(row=r, column=10, value=notes)

    # Journaling streak helper (col L): consecutive "Yes" up to each row
    for r in range(L0, L1 + 1):
        if r == L0:
            ws.cell(row=r, column=12, value=f'=IF($A{r}="","",IF($I{r}="Yes",1,0))')
        else:
            ws.cell(row=r, column=12,
                    value=f'=IF($A{r}="","",IF($I{r}="Yes",N({get_column_letter(12)}{r-1})+1,0))')

    style_rows(ws, L0, L1, 10, text_left={10}, dates={1})
    for r in range(L0, L1 + 1):
        ws.cell(row=r, column=5).number_format = "0.0"
        c = ws.cell(row=r, column=12)
        c.font = Font(size=9, color="BBBBBB")
        c.alignment = Alignment(horizontal="center")
    ws.column_dimensions["L"].hidden = True

    add_dv_whole(ws, f"B{L0}:D{L1}", 1, 10)
    add_dv(ws, f"G{L0}:G{L1}", "YesNoList")
    add_dv(ws, f"I{L0}:I{L1}", "YesNoList")

    # Mood color scale (low=soft red, high=mint); Stress reversed
    ws.conditional_formatting.add(
        f"B{L0}:B{L1}",
        ColorScaleRule(start_type="num", start_value=1, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=5, mid_color="FFFFF3CD",
                       end_type="num", end_value=10, end_color="FF" + HIGHLIGHT))
    ws.conditional_formatting.add(
        f"D{L0}:D{L1}",
        ColorScaleRule(start_type="num", start_value=1, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=5, mid_color="FFFFF3CD",
                       end_type="num", end_value=10, end_color="FF" + RED_BG))

    nm = {"CheckDate": "A", "CheckMood": "B", "CheckEnergy": "C", "CheckStress": "D",
          "CheckSleep": "E", "CheckWater": "F", "CheckExercise": "G",
          "CheckMindful": "H", "CheckJournaled": "I", "JournalStreak": "L"}
    for name, col in nm.items():
        wb.defined_names[name] = DefinedName(name, attr_text=f"'Check-In'!${col}${L0}:${col}${L1}")
    ws.freeze_panes = "A5"


# ===========================================================================
# Habits
# ===========================================================================
def build_habits(wb):
    ws = wb.create_sheet("Habits")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [13] + [11] * len(HABITS) + [11, 10, 11])
    luxe_header(ws, get_column_letter(1 + len(HABITS) + 3), "🌱  HABIT TRACKER",
                "Tiny habits, big calm. Tick what you did — streaks update automatically.")
    headers = ["Date"] + HABITS + ["Done", "Daily %", "Streak~"]
    table_headers(ws, 4, headers)

    import random
    rnd = random.Random(7)
    n_h = len(HABITS)
    done_col = 1 + n_h + 1          # col index for Done
    pct_col = done_col + 1
    streak_col = pct_col + 1

    # 14 days of sample completion
    for i in range(14):
        r = L0 + i
        ws.cell(row=r, column=1, value=today_minus(13 - i))
        for h in range(n_h):
            if rnd.random() < 0.68:
                ws.cell(row=r, column=2 + h, value="Yes")

    for r in range(L0, L1 + 1):
        first = get_column_letter(2)
        last = get_column_letter(1 + n_h)
        ws.cell(row=r, column=done_col,
                value=f'=IF($A{r}="","",COUNTIF({first}{r}:{last}{r},"Yes"))')
        ws.cell(row=r, column=pct_col,
                value=f'=IF($A{r}="","",{get_column_letter(done_col)}{r}/{n_h})')
        # routine streak: trailing days meeting the habit goal
        gl = get_column_letter(done_col)
        sl = get_column_letter(streak_col)
        if r == L0:
            ws.cell(row=r, column=streak_col,
                    value=f'=IF($A{r}="","",IF({gl}{r}>=HabitGoal,1,0))')
        else:
            ws.cell(row=r, column=streak_col,
                    value=f'=IF($A{r}="","",IF({gl}{r}>=HabitGoal,N({sl}{r-1})+1,0))')

    style_rows(ws, L0, L1, 1 + n_h + 3, dates={1})
    for r in range(L0, L1 + 1):
        ws.cell(row=r, column=pct_col).number_format = "0%"
        c = ws.cell(row=r, column=streak_col)
        c.font = Font(size=9, color="BBBBBB")

    add_dv(ws, f"{get_column_letter(2)}{L0}:{get_column_letter(1+n_h)}{L1}", "YesNoList")
    ws.conditional_formatting.add(
        f"{get_column_letter(2)}{L0}:{get_column_letter(1+n_h)}{L1}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(
        f"{get_column_letter(pct_col)}{L0}:{get_column_letter(pct_col)}{L1}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color=HIGHLIGHT, showValue=True))

    # Per-habit summary (for the dashboard bar chart)
    sr = L1 + 3
    merge_set(ws, f"A{sr}:C{sr}", "HABIT SUMMARY (last 60 days)", "section_gold")
    ws.cell(row=sr + 1, column=1, value="Habit").style = "th"
    ws.cell(row=sr + 1, column=2, value="Times Done").style = "th"
    ws.cell(row=sr + 1, column=3, value="Consistency %").style = "th"
    for i, h in enumerate(HABITS):
        r = sr + 2 + i
        col = get_column_letter(2 + i)
        ws.cell(row=r, column=1, value=h).style = "td_left"
        d = ws.cell(row=r, column=2, value=f'=COUNTIF({col}{L0}:{col}{L1},"Yes")')
        d.style = "td"
        p = ws.cell(row=r, column=3,
                    value=f'=IFERROR(B{r}/MAX(COUNT(CheckDate),1),0)')
        p.style = "td"; p.number_format = "0%"
    wb.defined_names["HabitSumNames"] = DefinedName(
        "HabitSumNames", attr_text=f"Habits!$A${sr+2}:$A${sr+1+len(HABITS)}")
    wb.defined_names["HabitSumPct"] = DefinedName(
        "HabitSumPct", attr_text=f"Habits!$C${sr+2}:$C${sr+1+len(HABITS)}")

    wb.defined_names["HabitDate"] = DefinedName("HabitDate", attr_text=f"Habits!$A${L0}:$A${L1}")
    wb.defined_names["HabitDone"] = DefinedName(
        "HabitDone", attr_text=f"Habits!${get_column_letter(done_col)}${L0}:${get_column_letter(done_col)}${L1}")
    wb.defined_names["HabitPct"] = DefinedName(
        "HabitPct", attr_text=f"Habits!${get_column_letter(pct_col)}${L0}:${get_column_letter(pct_col)}${L1}")
    wb.defined_names["HabitStreak"] = DefinedName(
        "HabitStreak", attr_text=f"Habits!${get_column_letter(streak_col)}${L0}:${get_column_letter(streak_col)}${L1}")
    ws.freeze_panes = "B5"


# ===========================================================================
# Daily Planner (single calm day)
# ===========================================================================
def build_planner(wb):
    ws = wb.create_sheet("Daily Planner")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 34, 4, 22, 22, 2])
    luxe_header(ws, "F", "📋  DAILY PLANNER", "Gentle structure for a calmer day. Print or duplicate as needed.")

    ws.cell(row=5, column=2, value="Date").style = "field_label"
    ws.cell(row=5, column=3, value=dt.date.today()).style = "input"
    ws.cell(row=5, column=3).number_format = "dddd, mmm d"

    merge_set(ws, "B7:C7", "TOP 3 PRIORITIES", "section")
    for i in range(3):
        r = 8 + i
        ws.cell(row=r, column=2, value=f"{i+1}.").style = "td"
        ws.cell(row=r, column=3, value="").style = "td_left"

    merge_set(ws, "B12:C12", "TIME BLOCKS", "section_gold")
    ws.cell(row=13, column=2, value="Time").style = "th"
    ws.cell(row=13, column=3, value="Plan").style = "th"
    times = ["6:00 AM", "7:00", "8:00", "9:00", "10:00", "11:00", "12:00 PM",
             "1:00", "2:00", "3:00", "4:00", "5:00", "6:00", "7:00", "8:00", "9:00"]
    for i, t in enumerate(times):
        r = 14 + i
        ws.cell(row=r, column=2, value=t).style = "td"
        ws.cell(row=r, column=3, value="").style = "td_left"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW)
            ws.cell(row=r, column=3).fill = fill(MUTED_ROW)

    # Right column: self-care, breaks, notes
    merge_set(ws, "E7:F7", "SELF-CARE TODAY", "section")
    care = ["Move my body", "Drink water", "Step outside", "Connect with someone",
            "Breathe / meditate", "Something I enjoy"]
    for i, c in enumerate(care):
        r = 8 + i
        ws.cell(row=r, column=5, value=c).style = "td_left"
        chk = ws.cell(row=r, column=6, value="☐")
        chk.style = "td"; chk.font = Font(size=14, bold=True, color=PRIMARY)

    merge_set(ws, "E15:F15", "REMEMBER TO BREAK", "section_gold")
    merge_set(ws, "E16:F18",
              "Pause every 60–90 minutes. A few slow breaths, a stretch, or a short walk "
              "keeps your nervous system calm.", "prompt")

    merge_set(ws, "E20:F20", "NOTES", "section")
    merge_set(ws, "E21:F28", "", "prompt")


# ===========================================================================
# Generic log builder for the simpler sheets
# ===========================================================================
def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, ints=None, pcts=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers),
               text_left=text_left or set(), dates=dates or set(),
               ints=ints or set(), pcts=pcts or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


# ===========================================================================
# Goals
# ===========================================================================
def build_goals(wb):
    sample = [
        ("Meditate 5 days a week", "Health", today_minus(-30), 0.6, "In Progress", "Mon/Wed/Fri done", ""),
        ("Attend one social event", "Social", today_minus(-14), 0.5, "In Progress", "RSVP sent", "Use Social Prep"),
        ("Read 2 calming books", "Learning", today_minus(-60), 0.25, "In Progress", "1 started", ""),
        ("Consistent 8-hr sleep", "Health", today_minus(-21), 0.7, "In Progress", "5/7 nights", ""),
        ("Daily gratitude for 30 days", "Personal", today_minus(-30), 1.0, "Complete", "Done!", "Felt a real shift"),
    ]
    ws, start, end = build_log(
        wb, "Goals", "🎯", "PERSONAL GOALS",
        "Small, kind goals. Progress over perfection.",
        ["Goal", "Category", "Target Date", "Progress", "Status", "Milestones", "Notes"],
        sample, [34, 16, 14, 12, 14, 26, 26],
        text_left={1, 6, 7}, dates={3},
        validations=[("B", "GoalCatList"), ("E", "StatusList")])
    for r in range(start, end + 1):
        ws.cell(row=r, column=4).number_format = "0%"
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color=HIGHLIGHT, showValue=True))
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Complete"'], fill=fill(MINT_BG)))
    wb.defined_names["GoalStatus"] = DefinedName("GoalStatus", attr_text=f"Goals!$E${start}:$E${end}")
    wb.defined_names["GoalProgress"] = DefinedName("GoalProgress", attr_text=f"Goals!$D${start}:$D${end}")


# ===========================================================================
# Sleep
# ===========================================================================
def build_sleep(wb):
    sample = [
        (13, "11:15 PM", "6:30 AM", 6.5, "Fair", "Woke once"),
        (12, "10:45 PM", "6:15 AM", 7.5, "Good", ""),
        (11, "11:30 PM", "5:30 AM", 6.0, "Poor", "Restless"),
        (10, "10:30 PM", "6:30 AM", 8.0, "Great", "Deep sleep"),
        (9, "11:00 PM", "6:00 AM", 7.0, "Good", ""),
        (8, "10:45 PM", "6:15 AM", 7.5, "Good", ""),
        (7, "11:45 PM", "6:15 AM", 6.5, "Fair", "Late night"),
        (6, "10:15 PM", "6:15 AM", 8.0, "Great", ""),
        (5, "10:45 PM", "5:45 AM", 7.0, "Good", ""),
        (4, "11:00 PM", "6:00 AM", 7.0, "Good", ""),
        (3, "10:15 PM", "6:45 AM", 8.5, "Great", "Best sleep this week"),
        (2, "10:45 PM", "6:15 AM", 7.5, "Good", ""),
        (1, "11:15 PM", "5:45 AM", 6.5, "Fair", ""),
        (0, "10:45 PM", "6:15 AM", 7.5, "Good", ""),
    ]
    rows = [(today_minus(d), bt, wt, ts, q, nt) for (d, bt, wt, ts, q, nt) in reversed(sample)]
    ws, start, end = build_log(
        wb, "Sleep", "🌙", "SLEEP TRACKER",
        "Rest is the foundation of calm. Notice your patterns gently.",
        ["Date", "Bedtime", "Wake Time", "Total Sleep", "Quality", "Notes"],
        rows, [13, 14, 14, 14, 14, 28],
        text_left={6}, dates={1}, validations=[("E", "QualityList")])
    for r in range(start, end + 1):
        ws.cell(row=r, column=4).number_format = '0.0" hrs"'
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        ColorScaleRule(start_type="num", start_value=4, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=7, mid_color="FFFFF3CD",
                       end_type="num", end_value=9, end_color="FF" + HIGHLIGHT))
    wb.defined_names["SleepDate"] = DefinedName("SleepDate", attr_text=f"Sleep!$A${start}:$A${end}")
    wb.defined_names["SleepTotal"] = DefinedName("SleepTotal", attr_text=f"Sleep!$D${start}:$D${end}")

    # Sleep trend chart
    line = LineChart(); line.title = "Weekly Sleep Trend"; line.height = 8; line.width = 16
    line.add_data(Reference(ws, min_col=4, min_row=4, max_row=start + 20), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=start, max_row=start + 20))
    line.legend = None
    ws.add_chart(line, "H5")


# ===========================================================================
# Build & deliver
# ===========================================================================
def build_simple_sheets(wb):
    # Social Prep
    build_log(
        wb, "Social Prep", "🤝", "SOCIAL PREPARATION PLANNER",
        "Prepare gently for social moments — then reflect with kindness afterward.",
        ["Event", "Date", "My Goal", "Things to Remember", "Conversation Topics",
         "Helpful Reminders", "Reflection Afterward"],
        [
            ("Team lunch", today_minus(-3), "Stay 30 min, ask 2 questions",
             "I can leave when I need to", "Weekend plans · new project",
             "Breathe before walking in", ""),
            ("Friend's birthday", today_minus(-9), "Arrive on time, enjoy one real chat",
             "People are glad I came", "Travel · shows · food",
             "It's okay to take a quiet break", ""),
            ("Work presentation", today_minus(-14), "Speak slowly, pause to breathe",
             "I am prepared", "Stick to my 3 key points",
             "Nerves are normal and pass", ""),
        ],
        [20, 12, 24, 26, 26, 24, 26],
        text_left={1, 3, 4, 5, 6, 7}, dates={2}, reserved=25)

    # Reflection Journal
    build_log(
        wb, "Reflection", "🪞", "REFLECTION JOURNAL",
        "A few honest lines a day. There are no wrong answers.",
        ["Date", "What went well?", "What challenged me?", "What helped?",
         "Grateful for", "Improve tomorrow"],
        [
            (today_minus(2), "Finished a hard task", "Felt anxious before a call",
             "Took 5 slow breaths", "A kind text from a friend", "Start earlier"),
            (today_minus(1), "Went for a walk", "Overthinking at night",
             "Wrote my worries down", "Warm coffee, quiet morning", "Phone off by 10pm"),
            (today_minus(0), "Used Social Prep & it worked", "Crowded room felt loud",
             "Stepped outside for a minute", "Progress I'm making", "Keep using breaks"),
        ],
        [13, 26, 26, 24, 24, 24],
        text_left={2, 3, 4, 5, 6}, dates={1}, reserved=40)

    # Self-Care
    build_log(
        wb, "Self-Care", "💗", "SELF-CARE PLANNER",
        "Care isn't selfish — it's maintenance. Mix a few categories each week.",
        ["Category", "Activity", "Goal / Week", "Done This Week", "Last Done", "Notes"],
        [
            ("Physical", "Stretch or yoga", 3, 2, today_minus(1), ""),
            ("Emotional", "Journal feelings", 4, 3, today_minus(0), ""),
            ("Social", "Reach out to someone", 2, 1, today_minus(3), ""),
            ("Creative", "Draw / play music", 2, 1, today_minus(5), ""),
            ("Rest", "Screen-free wind-down", 5, 4, today_minus(0), ""),
            ("Outdoor", "Walk in nature", 3, 2, today_minus(2), ""),
        ],
        [16, 24, 12, 14, 14, 26],
        text_left={2, 6}, dates={5}, ints={3, 4},
        validations=[("A", "SelfCareList")], reserved=25)

    # Exercise & Movement
    build_log(
        wb, "Exercise", "🏃", "EXERCISE & MOVEMENT",
        "Movement lifts mood. Track how you feel before and after.",
        ["Date", "Activity", "Duration (min)", "Intensity", "Mood Before", "Mood After", "Notes"],
        [
            (today_minus(3), "Walk", 30, "Low", 5, 7, "Felt clearer"),
            (today_minus(2), "Yoga", 25, "Low", 6, 8, "Very calming"),
            (today_minus(1), "Jog", 20, "Moderate", 5, 7, ""),
            (today_minus(0), "Stretching", 15, "Low", 6, 7, "Quick reset"),
        ],
        [13, 20, 14, 14, 13, 12, 24],
        text_left={2, 7}, dates={1}, ints={3, 5, 6},
        validations=[("D", "IntensityList")], reserved=40)

    # Gratitude Journal
    build_log(
        wb, "Gratitude", "🙏", "GRATITUDE JOURNAL",
        "Naming the good rewires the mind toward calm. Keep it simple.",
        ["Date", "Three things I'm grateful for", "One positive experience", "One act of kindness"],
        [
            (today_minus(2), "Sunshine · coffee · my pet", "A good laugh at work", "Held the door for someone"),
            (today_minus(1), "Health · friends · music", "Finished a book", "Texted a friend to check in"),
            (today_minus(0), "Rest · safety · progress", "Calm evening walk", "Complimented a coworker"),
        ],
        [13, 34, 28, 28],
        text_left={2, 3, 4}, dates={1}, reserved=40)

    # Resource Library
    build_log(
        wb, "Resources", "📚", "RESOURCE LIBRARY",
        "Your toolkit of calm — books, podcasts, affirmations & support contacts.",
        ["Type", "Title", "Link / Detail", "Why It Helps"],
        [
            ("Book", "Breathing & calm techniques", "(add link)", "Quick grounding tools"),
            ("Podcast", "Daily calm / mindfulness", "(add link)", "Morning reset"),
            ("App", "Meditation timer", "(add link)", "10-min sessions"),
            ("Affirmation", "“This feeling will pass.”", "—", "Repeat when anxious"),
            ("Affirmation", "“I am safe right now.”", "—", "Grounding"),
            ("Contact", "Trusted friend / family", "(add phone)", "Reach out when low"),
            ("Contact", "Therapist / counselor (optional)", "(add info)", "Professional support"),
            ("Contact", "Local crisis / helpline", "(look up in your country)", "If in crisis"),
        ],
        [16, 30, 24, 28],
        text_left={2, 3, 4}, validations=[("A", "ResTypeList")], reserved=30)


def build_progress(wb):
    ws = wb.create_sheet("Progress")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 16, 4, 30, 16, 2])
    luxe_header(ws, "G", "📈  PROGRESS", "Zoom out and see how far you've come. Be proud of small wins.")

    # Wellness score (big)
    merge_set(ws, "B5:C5", "YOUR WELLNESS SCORE", "section_gold")
    ws.merge_cells("B6:C8")
    score = (
        "=IFERROR(0.30*(AVERAGE(CheckMood)/10)"
        "+0.30*AVERAGE(HabitPct)"
        "+0.20*MIN(AVERAGE(CheckSleep)/SleepTarget,1)"
        "+0.20*(COUNTIF(CheckJournaled,\"Yes\")/MAX(COUNT(CheckDate),1)),0)")
    c = ws["B6"]; c.value = score
    c.style = "field_value"; c.number_format = "0%"
    c.font = Font(size=46, bold=True, color=PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = fill(IVORY)
    for rr in range(6, 9):
        for cc in (2, 3):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(
                top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "B9:C9",
              "Blend of mood, routine, sleep & journaling — gentle, not a grade.", "subtitle")
    ws["B9"].font = Font(size=10, italic=True, color=ACCENT)
    ws["B9"].fill = fill(WHITE)
    ws["C9"].fill = fill(WHITE)

    # Snapshot metrics
    merge_set(ws, "E5:G5", "SNAPSHOT", "section_gold")
    metrics = [
        ("Avg Mood (1–10)", "=IFERROR(AVERAGE(CheckMood),0)", "0.0"),
        ("Avg Sleep", "=IFERROR(AVERAGE(CheckSleep),0)", '0.0" hrs"'),
        ("Avg Routine %", "=IFERROR(AVERAGE(HabitPct),0)", "0%"),
        ("Best Habit Streak", "=IFERROR(MAX(HabitStreak),0)", "0"),
        ("Journaling Days", '=COUNTIF(CheckJournaled,"Yes")', "0"),
        ("Goals Completed", '=COUNTIF(GoalStatus,"Complete")', "0"),
        ("Mindful Minutes (total)", "=SUM(CheckMindful)", "0"),
        ("Exercise Sessions", '=COUNTIF(CheckExercise,"Yes")', "0"),
    ]
    for i, (lab, fml, fmt) in enumerate(metrics):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml)
        c.style = "field_value"; c.number_format = fmt
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)

    # Mood trend + habit consistency charts
    mood = LineChart(); mood.title = "Mood Trend"; mood.height = 8; mood.width = 15
    mood.add_data(Reference(wb["Check-In"], min_col=2, min_row=4, max_row=L0 + 19), titles_from_data=True)
    mood.set_categories(Reference(wb["Check-In"], min_col=1, min_row=L0, max_row=L0 + 19))
    mood.legend = None
    ws.add_chart(mood, "B16")

    # Habit consistency bar (from the Habits summary block)
    hs, he = L1 + 5, L1 + 4 + len(HABITS)   # data rows of the habit summary
    hb = BarChart(); hb.type = "bar"; hb.title = "Habit Consistency"; hb.height = 8; hb.width = 15
    hb.add_data(Reference(wb["Habits"], min_col=3, min_row=hs, max_row=he), titles_from_data=False)
    hb.set_categories(Reference(wb["Habits"], min_col=1, min_row=hs, max_row=he))
    hb.legend = None
    ws.add_chart(hb, "B33")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [15] * 12 + [2])
    ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🌿  CALM COMPASS™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2",
              "  Build calm routines, notice your patterns, and celebrate progress — one gentle day at a time.",
              "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    row1 = [
        ("AVG MOOD (1–10)", "=IFERROR(AVERAGE(CheckMood),0)", "dec"),
        ("HABITS COMPLETED", "=SUM(HabitDone)", "num"),
        ("ROUTINE SCORE", "=IFERROR(AVERAGE(HabitPct),0)", "pct"),
        ("AVG SLEEP", "=IFERROR(AVERAGE(CheckSleep),0)", "hrs"),
    ]
    row2 = [
        ("EXERCISE SESSIONS", '=COUNTIF(CheckExercise,"Yes")', "num"),
        ("JOURNALING STREAK", "=IFERROR(LOOKUP(2,1/(JournalStreak<>\"\"),JournalStreak),0)", "num"),
        ("MINDFUL MINUTES", "=SUM(CheckMindful)", "min"),
        ("GOALS COMPLETED", '=COUNTIF(GoalStatus,"Complete")', "num"),
    ]
    cols = [2, 5, 8, 11]
    for (lab, fml, kind), col in zip(row1, cols):
        kpi_card(ws, 5, col, 3, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols):
        kpi_card(ws, 8, col, 3, lab, fml, kind)

    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "YOUR TRENDS", "section_gold")

    mood = LineChart(); mood.title = "Mood Trend"; mood.height = 8.5; mood.width = 12
    mood.add_data(Reference(wb["Check-In"], min_col=2, min_row=4, max_row=L0 + 19), titles_from_data=True)
    mood.set_categories(Reference(wb["Check-In"], min_col=1, min_row=L0, max_row=L0 + 19))
    mood.legend = None
    ws.add_chart(mood, "B12")

    energy = LineChart(); energy.title = "Energy Trend"; energy.height = 8.5; energy.width = 12
    energy.add_data(Reference(wb["Check-In"], min_col=3, min_row=4, max_row=L0 + 19), titles_from_data=True)
    energy.set_categories(Reference(wb["Check-In"], min_col=1, min_row=L0, max_row=L0 + 19))
    energy.legend = None
    ws.add_chart(energy, "H12")

    ws.row_dimensions[30].height = 26
    merge_set(ws, "B30:M30", "ROUTINE & REST", "section")

    hs, he = L1 + 5, L1 + 4 + len(HABITS)   # data rows of the Habits summary block
    habit = BarChart(); habit.type = "bar"; habit.title = "Habit Consistency"; habit.height = 8.5; habit.width = 12
    habit.add_data(Reference(wb["Habits"], min_col=3, min_row=hs, max_row=he), titles_from_data=False)
    habit.set_categories(Reference(wb["Habits"], min_col=1, min_row=hs, max_row=he))
    habit.legend = None
    ws.add_chart(habit, "B31")

    sleep = LineChart(); sleep.title = "Sleep Trend"; sleep.height = 8.5; sleep.width = 12
    sleep.add_data(Reference(wb["Sleep"], min_col=4, min_row=4, max_row=L0 + 19), titles_from_data=True)
    sleep.set_categories(Reference(wb["Sleep"], min_col=1, min_row=L0, max_row=L0 + 19))
    sleep.legend = None
    ws.add_chart(sleep, "H31")

    ws.row_dimensions[49].height = 26
    merge_set(ws, "B49:M49",
              "Calm Compass™ is a wellness & organization tool — not medical advice. See the Welcome tab.",
              "subtitle")


# ===========================================================================
# Build
# ===========================================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    build_settings(wb)
    build_welcome(wb)
    build_checkin(wb)
    build_habits(wb)
    build_planner(wb)
    build_goals(wb)
    build_sleep(wb)
    build_simple_sheets(wb)     # social prep, reflection, self-care, exercise, gratitude, resources
    build_progress(wb)
    build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Check-In", "Habits", "Daily Planner",
             "Social Prep", "Reflection", "Goals", "Self-Care", "Sleep",
             "Exercise", "Gratitude", "Resources", "Progress", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "Calm_Compass.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
