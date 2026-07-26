"""Build Savings Challenge & Sinking Funds Command Center™ — The Save-On-Purpose Operating System.

14 tabs · a premium savings-challenge & sinking-funds operating system in Google Sheets
& Excel. Dashboard, a sinking-funds engine (target ÷ months → what to set aside every
month), the 100-envelope challenge, the 52-week challenge, custom challenges, savings
accounts, an emergency fund, a deposit log, a savings streak, goal countdowns, a
no-spend tracker and a monthly summary — one dashboard. Give every saved dollar a name,
and never be blindsided by a bill again.

Run: python3 build_xlsx.py   ->  ../Savings_Challenge_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
FUNDCAT = ["Annual bill", "Holiday", "Travel", "Home", "Car", "Health", "Gifts", "Other"]
ACCTTYPE = ["High-yield savings", "Checking", "Cash envelope", "CD", "Other"]
FREQ = ["Weekly", "Biweekly", "Monthly"]

# Settings / goals
MONTHS_IN = 8
MONTHLY_SAVE_GOAL = 500
STREAK_GOAL = 60
CHALLENGE_GOAL = 3
FUNDS_GOAL_COUNT = 8
EF_GOAL = 6000
EF_CURRENT = 6000
THIS_MONTH_SAVED = 520
STREAK_DAYS = 24

# Sinking funds: (fund, category, annual target, saved)
FUNDS = [
    ("Christmas & holidays", "Holiday", 1200, 800), ("Car maintenance", "Car", 900, 600),
    ("Vacation", "Travel", 2400, 1600), ("Home repairs", "Home", 1500, 1000),
    ("Insurance premiums", "Annual bill", 1200, 800), ("Gifts & birthdays", "Gifts", 600, 400),
    ("Medical & dental", "Health", 800, 533), ("Pet care", "Other", 400, 267),
]

# 100-envelope challenge — sample of envelopes pulled: (envelope #, amount, done?)
ENVELOPES = [
    (7, 7, "Yes"), (23, 23, "Yes"), (41, 41, "Yes"), (58, 58, "Yes"), (66, 66, "Yes"),
    (79, 79, "Yes"), (84, 84, "No"), (92, 92, "No"), (97, 97, "No"), (100, 100, "No"),
]

# 52-week challenge: (week, amount, done?)
WEEKS_52 = [
    (1, 1, "Yes"), (2, 2, "Yes"), (3, 3, "Yes"), (4, 4, "Yes"), (5, 5, "Yes"),
    (6, 6, "Yes"), (7, 7, "Yes"), (8, 8, "Yes"), (9, 9, "No"), (10, 10, "No"),
    (11, 11, "No"), (12, 12, "No"),
]

# Custom challenges: (challenge, target, saved, active?)
CHALLENGES = [
    ("100 Envelope Challenge", 5050, 274, "Yes"), ("52-Week Challenge", 1378, 36, "Yes"),
    ("No-Spend Month", 300, 300, "Yes"), ("$5 Bill Challenge", 500, 145, "No"),
    ("Round-Up Challenge", 400, 210, "No"),
]

# Savings accounts: (account, type, balance, APY)
ACCOUNTS = [
    ("Ally HYSA", "High-yield savings", 6000, 0.042), ("Sinking funds acct", "High-yield savings", 6000, 0.042),
    ("Cash envelopes", "Cash envelope", 274, 0.0), ("Vacation CD", "CD", 0, 0.045),
]

# Deposit log: (date, to, amount)
DEPOSITS = [
    ("Jul 1", "Sinking funds", 250), ("Jul 5", "100 Envelope", 66), ("Jul 8", "Emergency fund", 100),
    ("Jul 12", "52-Week", 8), ("Jul 15", "Sinking funds", 250), ("Jul 22", "100 Envelope", 79),
]

# Goal countdowns: (goal, target, saved, deadline)
GOALS = [
    ("Christmas 2026", 1200, 800, "Dec 1"), ("Summer trip", 2400, 1600, "Jun 1"),
    ("New laptop", 1400, 350, "Oct 1"), ("Car tires", 900, 600, "Nov 1"),
]

# No-spend tracker: (month, no-spend days, target)
NOSPEND = [
    ("Feb", 12, 12), ("Mar", 14, 12), ("Apr", 15, 12), ("May", 13, 12), ("Jun", 16, 12), ("Jul", 18, 12),
]

# Cash envelopes (cash stuffing): (envelope, loaded, spent)
CASH_ENVELOPES = [
    ("Groceries", 600, 480), ("Eating out", 200, 165), ("Fuel", 180, 140),
    ("Fun money", 150, 95), ("Household", 120, 70), ("Beauty & hair", 100, 60),
    ("Pet supplies", 80, 45), ("Misc", 100, 40),
]

# Monthly summary: (month, total saved)
MONTHS = [("Feb", 420), ("Mar", 450), ("Apr", 480), ("May", 500), ("Jun", 510), ("Jul", 520)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 20, 3] + [20] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your goals & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR GOALS", "section")
    controls = [
        ("Your name", "June", None, "Owner"),
        ("Plan name", "Copper & Clover", None, "PlanName"),
        ("Months into the year", MONTHS_IN, "0", "MonthsIn"),
        ("Monthly savings goal", MONTHLY_SAVE_GOAL, '"$"#,##0', "MonthlySaveGoal"),
        ("Saved this month", THIS_MONTH_SAVED, '"$"#,##0', "ThisMonthSaved"),
        ("Savings streak (days)", STREAK_DAYS, "0", "StreakDays"),
        ("Streak goal (days)", STREAK_GOAL, "0", "StreakGoal"),
        ("Active-challenges goal", CHALLENGE_GOAL, "0", "ChallengeGoal"),
        ("Sinking-funds count goal", FUNDS_GOAL_COUNT, "0", "FundsGoalCount"),
        ("Emergency fund goal", EF_GOAL, '"$"#,##0', "EFGoal"),
        ("Emergency fund now", EF_CURRENT, '"$"#,##0', "EFCurrent"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Fund category", FUNDCAT, "FundCatList"), ("F", "Account type", ACCTTYPE, "AcctTypeList"),
             ("G", "Frequency", FREQ, "FreqList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🫙  SAVINGS CHALLENGE & SINKING FUNDS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Give every saved dollar a name — and never be blindsided by a bill again.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE SAVINGS LIFE, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Most people don't have a spending problem — they have a timing problem. Christmas, car tyres and "
                      "the insurance renewal are not emergencies; they're appointments you forgot to save for. This "
                      "fixes that: a sinking-funds engine divides each yearly target by the months you have left and "
                      "tells you exactly what to set aside every month. Then run the challenges that make saving "
                      "genuinely fun — 100 envelopes, 52 weeks, no-spend months — track your streak and watch every "
                      "fund fill up, all in ONE premium Google Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your monthly goal and streak target.",
             "2.  List your Sinking Funds — every bill that isn't monthly.",
             "3.  Read the monthly set-aside each fund needs. That's your number.",
             "4.  Pick a challenge: 100 Envelope, 52-Week or your own.",
             "5.  Log every deposit and keep your streak alive.",
             "6.  Check the Dashboard: saved, on-pace & a Savings Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional saver (Copper & Clover, June) is included so you can see how it all "
               "connects — just type over it with your own funds and goals. The monthly set-aside your sinking funds "
               "need, and whether you're on pace, are the two numbers that decide whether a big bill is a crisis or a "
               "non-event, and they roll into a live Savings Score. The 100-Envelope Challenge totals $5,050 and the "
               "52-Week Challenge totals $1,378 — both are built in and tracked. Twelve matching printable pages "
               "(sinking-fund worksheet, envelope grid, 52-week tracker, savings thermometer & more) are included. "
               "This is a budgeting & organizing tool, not financial advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "A bill you saved for is just a purchase. A bill you didn't is an emergency.", "section_gold")


# ===========================================================================
def build_sinking(wb):
    ws = wb.create_sheet("Sinking Funds"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 14, 14, 14, 16, 2])
    luxe_header(ws, "G", "🫙  SINKING FUNDS",
                "Every non-monthly bill, divided by the months you have — this is the engine of a calm budget.")
    table_headers(ws, 4, ["Fund", "Category", "Target", "Saved", "% Full", "Monthly Set-Aside"], start_col=2)
    start = L0
    for i, (fund, cat, target, saved) in enumerate(FUNDS):
        r = start + i
        ws.cell(row=r, column=2, value=fund).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        ct = ws.cell(row=r, column=4, value=target); ct.style = "input"; ct.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=5, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/D{r},0)"); cp.style = "td"; cp.number_format = "0%"
        cm = ws.cell(row=r, column=7, value=f"=IFERROR(D{r}/12,0)"); cm.style = "td"; cm.number_format = '"$"#,##0'; cm.font = Font(bold=True, color=PRIMARY)
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FUNDS) - 1
    nrange(wb, "FundName", "Sinking Funds", "B", start, end)
    nrange(wb, "FundTarget", "Sinking Funds", "D", start, end)
    nrange(wb, "FundSaved", "Sinking Funds", "E", start, end)
    add_dv(ws, f"C{start}:C{end}", "FundCatList")
    ws.conditional_formatting.add(f"F{start}:F{end}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL FUNDS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=4, value="=SUM(FundTarget)"); ct.style = "td"; ct.font = Font(bold=True, size=12, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=5, value="=SUM(FundSaved)"); cs.style = "td"; cs.font = Font(bold=True, size=12, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cp = ws.cell(row=tot, column=6, value=f"=IFERROR(E{tot}/D{tot},0)"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = "0%"
    cm = ws.cell(row=tot, column=7, value=f"=D{tot}/12"); cm.style = "td"; cm.font = Font(bold=True, size=13, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0'
    cell_name(wb, "FundsTarget", "Sinking Funds", f"$D${tot}")
    cell_name(wb, "FundsSaved", "Sinking Funds", f"$E${tot}")
    cell_name(wb, "MonthlyNeed", "Sinking Funds", f"$G${tot}")
    ws.cell(row=tot + 2, column=2, value="Where you should be by now (months in ÷ 12)").style = "field_label"
    ce = ws.cell(row=tot + 2, column=4, value="=FundsTarget*MonthsIn/12"); ce.style = "field_value"; ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpectedToDate", "Sinking Funds", f"$D${tot+2}")
    ws.cell(row=tot + 3, column=2, value="= ON PACE?").style = "th"
    cop = ws.cell(row=tot + 3, column=4, value="=IFERROR(MIN(FundsSaved/ExpectedToDate,1),0)"); cop.style = "td"; cop.font = Font(bold=True, size=12, color=PRIMARY); cop.fill = fill(MINT_BG); cop.number_format = "0%"
    cell_name(wb, "OnPace", "Sinking Funds", f"$D${tot+3}")
    ws.cell(row=tot + 4, column=2, value="Funds set up").style = "field_label"
    cfc = ws.cell(row=tot + 4, column=4, value="=COUNTA(FundName)"); cfc.style = "field_value"; cfc.number_format = "#,##0"
    cell_name(wb, "FundCount", "Sinking Funds", f"$D${tot+4}")
    ws.freeze_panes = "A5"


def build_envelope(wb):
    ws = wb.create_sheet("100 Envelope"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 14, 2])
    luxe_header(ws, "D", "✉  100 ENVELOPE CHALLENGE",
                "Number 100 envelopes 1–100, pull one at random each day, and stuff it. Finish and you've saved $5,050.")
    ws.cell(row=5, column=2, value="Challenge total (1+2+…+100)").style = "field_label"
    ct = ws.cell(row=5, column=3, value=5050); ct.style = "field_value"; ct.number_format = '"$"#,##0'; ct.fill = fill(MINT_BG)
    cell_name(wb, "EnvelopeTotal", "100 Envelope", "$C$5")
    table_headers(ws, 7, ["Envelope #", "Amount", "Stuffed?"], start_col=2)
    start = 8
    for i, (num, amt, done) in enumerate(ENVELOPES):
        r = start + i
        ws.cell(row=r, column=2, value=num).style = "td"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=done).style = "td"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(ENVELOPES) - 1
    nrange(wb, "EnvAmt", "100 Envelope", "C", start, end)
    nrange(wb, "EnvDone", "100 Envelope", "D", start, end)
    add_dv(ws, f"D{start}:D{end}", "YesNoList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="STUFFED SO FAR").style = "th"
    cs = ws.cell(row=tot, column=3, value='=SUMIF(EnvDone,"Yes",EnvAmt)'); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cn = ws.cell(row=tot, column=4, value='=COUNTIF(EnvDone,"Yes")&" done"'); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE)
    cell_name(wb, "EnvelopeSaved", "100 Envelope", f"$C${tot}")
    ws.cell(row=tot + 2, column=2, value="Add rows 1–100 and pull one at random each day — the randomness is the fun.").style = "section_gold"
    ws.freeze_panes = "A8"


def build_52week(wb):
    ws = wb.create_sheet("52-Week"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 16, 14, 2])
    luxe_header(ws, "D", "📅  52-WEEK CHALLENGE",
                "Save $1 in week 1, $2 in week 2, all the way to $52. Finish the year up $1,378.")
    ws.cell(row=5, column=2, value="Challenge total (1+2+…+52)").style = "field_label"
    ct = ws.cell(row=5, column=3, value=1378); ct.style = "field_value"; ct.number_format = '"$"#,##0'; ct.fill = fill(MINT_BG)
    cell_name(wb, "Week52Total", "52-Week", "$C$5")
    table_headers(ws, 7, ["Week", "Amount", "Saved?"], start_col=2)
    start = 8
    for i, (wk, amt, done) in enumerate(WEEKS_52):
        r = start + i
        ws.cell(row=r, column=2, value=wk).style = "td"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=done).style = "td"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(WEEKS_52) - 1
    nrange(wb, "WkAmt", "52-Week", "C", start, end)
    nrange(wb, "WkDone", "52-Week", "D", start, end)
    add_dv(ws, f"D{start}:D{end}", "YesNoList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="SAVED SO FAR").style = "th"
    cs = ws.cell(row=tot, column=3, value='=SUMIF(WkDone,"Yes",WkAmt)'); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cn = ws.cell(row=tot, column=4, value='=COUNTIF(WkDone,"Yes")&" wks"'); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE)
    cell_name(wb, "Week52Saved", "52-Week", f"$C${tot}")
    ws.cell(row=tot + 2, column=2, value="Flip it — start at $52 in January while motivation is high, end on $1.").style = "section_gold"
    ws.freeze_panes = "A8"


def build_challenges(wb):
    ws = wb.create_sheet("Challenges"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 16, 14, 14, 2])
    luxe_header(ws, "F", "🏆  CHALLENGES",
                "Every challenge you're running — target, progress and whether it's live right now.")
    table_headers(ws, 4, ["Challenge", "Target", "Saved", "% Done", "Active?"], start_col=2)
    start = L0
    for i, (name, target, saved, active) in enumerate(CHALLENGES):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ct = ws.cell(row=r, column=3, value=target); ct.style = "input"; ct.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); cp.style = "td"; cp.number_format = "0%"
        ws.cell(row=r, column=6, value=active).style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CHALLENGES) - 1
    nrange(wb, "ChTarget", "Challenges", "C", start, end)
    nrange(wb, "ChSaved", "Challenges", "D", start, end)
    nrange(wb, "ChActive", "Challenges", "F", start, end)
    add_dv(ws, f"F{start}:F{end}", "YesNoList")
    ws.conditional_formatting.add(f"E{start}:E{end}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GOLD_LT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ACTIVE CHALLENGES").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(ChTarget)"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value="=SUM(ChSaved)"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    ca = ws.cell(row=tot, column=6, value='=COUNTIF(ChActive,"Yes")'); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = "#,##0"
    cell_name(wb, "ActiveChallenges", "Challenges", f"$F${tot}")
    ws.freeze_panes = "A5"


def build_accounts(wb):
    ws, start, end = build_log(
        wb, "Savings Accounts", "🏦", "SAVINGS ACCOUNTS",
        "Where the money actually lives — balance and interest rate, so nothing sits idle in checking.",
        ["Account", "Type", "Balance", "APY"],
        ACCOUNTS, [2, 24, 22, 16, 12, 2], text_left={2}, money={4}, reserved=24, start_col=2,
        validations=[("C", "AcctTypeList")])
    nrange(wb, "AcctBal", "Savings Accounts", "D", start, end)
    for r in range(start, end + 1):
        ws.cell(row=r, column=5).number_format = "0.00%"
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL SAVED").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cb = ws.cell(row=tot, column=4, value="=SUM(AcctBal)"); cb.style = "td"; cb.font = Font(bold=True, size=12, color=PRIMARY); cb.fill = fill(MINT_BG); cb.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "TotalSaved", "Savings Accounts", f"$D${tot}")


def build_emergency(wb):
    ws = wb.create_sheet("Emergency Fund"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 2])
    luxe_header(ws, "C", "🛟  EMERGENCY FUND",
                "The fund that is not a sinking fund — the one that catches you when life goes sideways.")
    ws.cell(row=5, column=2, value="Emergency fund now").style = "field_label"
    cn = ws.cell(row=5, column=3, value="=EFCurrent"); cn.style = "field_value"; cn.number_format = '"$"#,##0'; cn.fill = fill(MINT_BG)
    ws.cell(row=6, column=2, value="Emergency fund goal").style = "field_label"
    cg = ws.cell(row=6, column=3, value="=EFGoal"); cg.style = "field_value"; cg.number_format = '"$"#,##0'
    ws.cell(row=7, column=2, value="= FUNDED").style = "th"
    cf = ws.cell(row=7, column=3, value="=IFERROR(MIN(EFCurrent/EFGoal,1),0)"); cf.style = "td"; cf.font = Font(bold=True, size=13, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = "0%"
    cell_name(wb, "EFFunded", "Emergency Fund", "$C$7")
    ws.cell(row=8, column=2, value="Still to save").style = "field_label"
    cs = ws.cell(row=8, column=3, value="=MAX(EFGoal-EFCurrent,0)"); cs.style = "field_value"; cs.number_format = '"$"#,##0'
    ws.cell(row=10, column=2, value="WHAT COUNTS AS AN EMERGENCY").style = "section_gold"
    notes = ["A true emergency is urgent, necessary and unexpected.",
             "Christmas is none of those — that's a sinking fund.",
             "New tyres after a blowout: emergency. Scheduled tyres: sinking fund.",
             "Aim for 3–6 months of essential expenses over time.",
             "Keep it in a separate high-yield account so it's boring to touch."]
    for i, n in enumerate(notes):
        ws.cell(row=11 + i, column=2, value="•  " + n).style = "body"
        ws.row_dimensions[11 + i].height = 20
    ws.cell(row=18, column=2, value="Sinking funds stop emergencies from happening. This one handles the rest.").style = "section_gold"


def build_deposits(wb):
    ws, start, end = build_log(
        wb, "Deposit Log", "💵", "DEPOSIT LOG",
        "Every deposit, dated — proof that the habit is real, and the number your streak is built on.",
        ["Date", "Into", "Amount"],
        DEPOSITS, [2, 14, 26, 16, 2], text_left={2, 3}, money={4}, reserved=34, start_col=2)
    nrange(wb, "DepAmt", "Deposit Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="DEPOSITED").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cd = ws.cell(row=tot, column=4, value="=SUM(DepAmt)"); cd.style = "td"; cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(MINT_BG); cd.number_format = '"$"#,##0'


def build_streak(wb):
    ws = wb.create_sheet("Savings Streak"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 18, 2])
    luxe_header(ws, "C", "🔥  SAVINGS STREAK",
                "Consecutive days you moved money — however small. Streaks build habits faster than budgets do.")
    ws.cell(row=5, column=2, value="Current streak (days)").style = "field_label"
    cs = ws.cell(row=5, column=3, value="=StreakDays"); cs.style = "field_value"; cs.number_format = "#,##0"; cs.fill = fill(MINT_BG)
    ws.cell(row=6, column=2, value="Streak goal (days)").style = "field_label"
    cg = ws.cell(row=6, column=3, value="=StreakGoal"); cg.style = "field_value"; cg.number_format = "#,##0"
    ws.cell(row=7, column=2, value="= STREAK PROGRESS").style = "th"
    cp = ws.cell(row=7, column=3, value="=IFERROR(MIN(StreakDays/StreakGoal,1),0)"); cp.style = "td"; cp.font = Font(bold=True, size=13, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = "0%"
    cell_name(wb, "StreakProgress", "Savings Streak", "$C$7")
    ws.cell(row=9, column=2, value="NO-SPEND DAYS BY MONTH").style = "section_gold"
    table_headers(ws, 10, ["Month", "No-Spend Days", "Target"], start_col=2)
    start = 11
    for i, (m, days, target) in enumerate(NOSPEND):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cd = ws.cell(row=r, column=3, value=days); cd.style = "input"; cd.number_format = "#,##0"
        ct = ws.cell(row=r, column=4, value=target); ct.style = "td"; ct.number_format = "#,##0"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(NOSPEND) - 1
    nrange(wb, "NoSpendDays", "Savings Streak", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVG NO-SPEND DAYS").style = "th"
    ca = ws.cell(row=tot, column=3, value="=IFERROR(AVERAGE(NoSpendDays),0)"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = "0.0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    ws.freeze_panes = "A11"


def build_goals(wb):
    ws = wb.create_sheet("Goal Countdown"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 14, 14, 2])
    luxe_header(ws, "F", "🎯  GOAL COUNTDOWN",
                "The things you're saving toward — how far along, how much is left, and by when.")
    table_headers(ws, 4, ["Goal", "Target", "Saved", "To Go", "Deadline"], start_col=2)
    start = L0
    for i, (goal, target, saved, deadline) in enumerate(GOALS):
        r = start + i
        ws.cell(row=r, column=2, value=goal).style = "td_left"
        ct = ws.cell(row=r, column=3, value=target); ct.style = "input"; ct.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cg = ws.cell(row=r, column=5, value=f"=MAX(C{r}-D{r},0)"); cg.style = "td"; cg.number_format = '"$"#,##0'
        ws.cell(row=r, column=6, value=deadline).style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(GOALS) - 1
    nrange(wb, "GoalTarget", "Goal Countdown", "C", start, end)
    nrange(wb, "GoalSaved", "Goal Countdown", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL GOALS").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(GoalTarget)"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value="=SUM(GoalSaved)"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cg = ws.cell(row=tot, column=5, value=f"=C{tot}-D{tot}"); cg.style = "td"; cg.font = Font(bold=True, color=DANGER); cg.fill = fill(WARN_BG); cg.number_format = '"$"#,##0'
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


def build_cashenvelopes(wb):
    ws = wb.create_sheet("Cash Envelopes"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 14, 14, 2])
    luxe_header(ws, "F", "💌  CASH ENVELOPES",
                "Cash stuffing, tracked — what you loaded, what you spent and what's still in the wallet.")
    table_headers(ws, 4, ["Envelope", "Loaded", "Spent", "Left", "% Left"], start_col=2)
    start = L0
    for i, (name, loaded, spent) in enumerate(CASH_ENVELOPES):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cl = ws.cell(row=r, column=3, value=loaded); cl.style = "input"; cl.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=spent); cs.style = "input"; cs.number_format = '"$"#,##0'
        cle = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cle.style = "td"; cle.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=6, value=f"=IFERROR((C{r}-D{r})/C{r},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CASH_ENVELOPES) - 1
    nrange(wb, "CashLoaded", "Cash Envelopes", "C", start, end)
    nrange(wb, "CashSpent", "Cash Envelopes", "D", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="lessThan", formula=["0.15"], fill=fill(RED_BG)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL ENVELOPES").style = "th"
    cl = ws.cell(row=tot, column=3, value="=SUM(CashLoaded)"); cl.style = "td"; cl.font = Font(bold=True, color=PRIMARY); cl.fill = fill(SURFACE); cl.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value="=SUM(CashSpent)"); cs.style = "td"; cs.font = Font(bold=True, color=DANGER); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cle = ws.cell(row=tot, column=5, value=f"=C{tot}-D{tot}"); cle.style = "td"; cle.font = Font(bold=True, size=12, color=PRIMARY); cle.fill = fill(MINT_BG); cle.number_format = '"$"#,##0'
    cp = ws.cell(row=tot, column=6, value=f"=IFERROR(E{tot}/C{tot},0)"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = "0%"
    cell_name(wb, "CashLeft", "Cash Envelopes", f"$E${tot}")
    ws.cell(row=tot + 2, column=2, value="When an envelope is empty, that category is done for the month. That's the whole method.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "What you saved each month — the line that should keep climbing.")
    ws.cell(row=5, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 6, ["Month", "Saved"], start_col=2)
    ts = 7
    for i, (m, amt) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "SaveTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Saved by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🫙  SAVINGS CHALLENGE & SINKING FUNDS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Funds, challenges, streak & a Savings Score — every saved dollar, named.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("FUNDS TARGET", "=FundsTarget", "money"),
        ("FUNDS SAVED", "=FundsSaved", "money"),
        ("MONTHLY SET-ASIDE", "=MonthlyNeed", "money"),
        ("ON PACE", "=OnPace", "pct"),
        ("EMERGENCY FUND", "=EFCurrent", "money"),
        ("TOTAL SAVED", "=TotalSaved", "money"),
    ]
    row2 = [
        ("100 ENVELOPE", "=EnvelopeTotal", "money"),
        ("52-WEEK", "=Week52Total", "money"),
        ("ACTIVE CHALLENGES", "=ActiveChallenges", "num"),
        ("SAVED THIS MONTH", "=ThisMonthSaved", "money"),
        ("SAVINGS STREAK", "=StreakDays", "num"),
        ("SAVINGS SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "SAVINGS HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SAVED BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Sinking funds on pace", "=IFERROR(MIN(FundsSaved/ExpectedToDate,1),0)"),
        ("Emergency fund funded", "=IFERROR(MIN(EFCurrent/EFGoal,1),0)"),
        ("Challenges running", "=IFERROR(MIN(ActiveChallenges/ChallengeGoal,1),0)"),
        ("Hitting monthly goal", "=IFERROR(MIN(ThisMonthSaved/MonthlySaveGoal,1),0)"),
        ("Every bill has a fund", "=IFERROR(MIN(FundCount/FundsGoalCount,1),0)"),
        ("Savings streak", "=IFERROR(MIN(StreakDays/StreakGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Saved by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=7, max_row=6 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=7, max_row=6 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Savings Challenge & Sinking Funds Command Center™ — give every saved dollar a name.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_sinking(wb); build_envelope(wb)
    build_52week(wb); build_challenges(wb); build_accounts(wb); build_emergency(wb)
    build_deposits(wb); build_streak(wb); build_goals(wb); build_cashenvelopes(wb); build_summary(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Sinking Funds", "100 Envelope", "52-Week", "Challenges",
             "Goal Countdown", "Cash Envelopes", "Savings Accounts", "Emergency Fund", "Deposit Log",
             "Savings Streak", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Savings_Challenge_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
