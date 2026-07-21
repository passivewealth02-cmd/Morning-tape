"""Build IEP & Special-Needs Command Center™ — The Complete Advocacy & Progress System.

16 tabs (+ Settings) · a premium, parent-friendly system in Google Sheets & Excel
for organizing a child's IEP/504: goals & progress monitoring, services & minutes,
accommodations, therapy log, behavior tracking, meeting & communication logs,
health/meds, a strengths profile, records checklist and a wins log — one calm
dashboard. Built by-a-parent, for parents. Not medical or legal advice.

Run: python3 build_xlsx.py   ->  ../IEP_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

AREAS = ["Reading", "Math", "Writing", "Speech / Language", "Social / Emotional",
         "Behavior", "Fine Motor", "Gross Motor", "Life Skills", "Other"]
YESNO = ["Yes", "No"]
SETTINGS_LIST = ["All settings", "Classroom", "Testing", "Specials", "Home", "Recess"]
SERVICE_TYPE = ["Special Education", "Speech Therapy", "Occupational Therapy",
                "Physical Therapy", "Counseling", "Behavior Support", "Other"]
MEET_TYPE = ["Annual Review", "Eligibility", "Re-evaluation", "Amendment", "Progress", "Other"]
DONE_LIST = ["Yes", "In Progress", "No"]

# ---- sample child — the data that drives every KPI (not a real person) ----
STUDENT = {"name": "Sam Bennett", "grade": "3", "age": 9,
           "plan": "IEP — Specific Learning Disability + Speech", "case": "Ms. Rivera"}

# IEP goals: (area, goal, baseline, target, current, unit)
GOALS = [
    ("Reading", "Read grade-level text at fluency target", 40, 80, 66, "WPM"),
    ("Math", "Solve two-step word problems", 30, 80, 62, "% correct"),
    ("Writing", "Write on-topic sentences with support", 2, 6, 4, "sentences"),
    ("Speech / Language", "Produce /r/ in sentences", 40, 90, 78, "% accuracy"),
    ("Social / Emotional", "Use a calm-down strategy when frustrated", 20, 80, 58, "% of times"),
]

# Progress monitoring data points: (date offset, goal area, measure, value)
MONITORING = [
    (-56, "Reading", "WPM", 48), (-49, "Reading", "WPM", 52), (-42, "Reading", "WPM", 57),
    (-35, "Reading", "WPM", 60), (-28, "Reading", "WPM", 63), (-14, "Reading", "WPM", 66),
    (-52, "Math", "% correct", 38), (-38, "Math", "% correct", 50), (-24, "Math", "% correct", 56),
    (-10, "Math", "% correct", 62), (-45, "Speech / Language", "% accuracy", 52),
    (-31, "Speech / Language", "% accuracy", 64), (-17, "Speech / Language", "% accuracy", 72),
    (-6, "Speech / Language", "% accuracy", 78), (-40, "Writing", "sentences", 3),
    (-19, "Writing", "sentences", 4), (-33, "Social / Emotional", "% of times", 40),
    (-12, "Social / Emotional", "% of times", 58), (-25, "Reading", "WPM", 61),
    (-8, "Math", "% correct", 60), (-21, "Speech / Language", "% accuracy", 70),
    (-15, "Writing", "sentences", 4), (-9, "Social / Emotional", "% of times", 54),
    (-3, "Reading", "WPM", 65),
]

# Services & minutes: (service, provider, frequency, min/week scheduled, min delivered)
SERVICES = [
    ("Special Education", "Resource room · Ms. Rivera", "Daily", 300, 285),
    ("Speech Therapy", "SLP · Mr. Nguyen", "2× / week", 60, 60),
    ("Occupational Therapy", "OT · Ms. Patel", "1× / week", 30, 30),
    ("Counseling", "School counselor", "1× / week", 30, 20),
]

# Accommodations & modifications: (accommodation, setting, active?)
ACCOMMODATIONS = [
    ("Extended time (1.5×) on tests", "Testing", "Yes"),
    ("Preferential seating (front, near teacher)", "Classroom", "Yes"),
    ("Assignments chunked into smaller steps", "Classroom", "Yes"),
    ("Frequent movement / sensory breaks", "All settings", "Yes"),
    ("Text-to-speech for reading passages", "Classroom", "Yes"),
    ("Reduced written output / scribe", "Classroom", "Yes"),
    ("Visual schedule & checklists", "All settings", "Yes"),
    ("Quiet, low-distraction test space", "Testing", "Yes"),
    ("Directions repeated & checked for understanding", "All settings", "Yes"),
    ("Modified spelling list", "Classroom", "No"),
]

# Therapy log: (date offset, type, provider, focus / note)
THERAPY = [
    (-3, "Speech Therapy", "Mr. Nguyen", "/r/ in phrases — 78% accuracy, great session"),
    (-5, "Occupational Therapy", "Ms. Patel", "Handwriting grip & letter spacing"),
    (-7, "Speech Therapy", "Mr. Nguyen", "/r/ in words — carryover homework sent"),
    (-10, "Counseling", "Counselor", "Named feelings, practiced 'take 5' breathing"),
    (-12, "Occupational Therapy", "Ms. Patel", "Scissor skills & bilateral coordination"),
    (-14, "Speech Therapy", "Mr. Nguyen", "Story retell with /r/ words"),
    (-17, "Counseling", "Counselor", "Calm-corner routine reviewed"),
    (-19, "Occupational Therapy", "Ms. Patel", "Zones of Regulation check-in"),
]

# Behavior tracker (ABC): (date offset, situation/antecedent, behavior, support that helped)
BEHAVIOR = [
    (-2, "Long worksheet, no breaks", "Put head down, refused", "Chunked it + a 3-min break"),
    (-6, "Group work, loud room", "Covered ears, left seat", "Noise headphones + quiet spot"),
    (-9, "Timed test", "Rushed, then upset", "Extended time reminder helped"),
    (-13, "Transition from recess", "Slow to settle", "Visual timer + first/then"),
    (-16, "Read-aloud turn", "Refused to read", "Pre-taught passage privately"),
]

# Meeting log: (date offset, type, attendees, key decisions)
MEETINGS = [
    (-120, "Annual Review", "Parents, Ms. Rivera, SLP, OT, admin", "Goals updated; added counseling 1×/wk"),
    (-30, "Progress", "Parents, Ms. Rivera", "On track for reading; adjust math pacing"),
    (60, "Annual Review", "Full team", "Scheduled — bring updated data & requests"),
]

# Communication log: (date offset, with, topic, follow-up?)
COMMUNICATION = [
    (-2, "Ms. Rivera", "Weekly progress note on reading fluency", "None"),
    (-5, "Mr. Nguyen (SLP)", "Home practice for /r/ carryover", "Do 5 min/day"),
    (-8, "Counselor", "Calm-down strategy working at home too", "Keep reinforcing"),
    (-11, "Front office", "Requested copy of latest data", "Received"),
    (-15, "Ms. Rivera", "Question about extended-time on state test", "Confirmed in place"),
]

# Records & documents checklist: (item, done)
RECORDS = [
    ("Current IEP / 504 on file", "Yes"), ("Signed consent forms", "Yes"),
    ("Most recent evaluation report", "Yes"), ("Progress reports (each period)", "Yes"),
    ("Meeting notices & minutes", "Yes"), ("Work samples for the binder", "In Progress"),
    ("Prior written notices (PWN)", "Yes"), ("Data charts printed for meeting", "No"),
]

# Wins & milestones: (win, date offset)
WINS = [
    ("Read a chapter book start to finish", -8), ("Used 'take 5' without a reminder", -14),
    ("Hit 66 WPM — up from 40!", -3), ("Said /r/ clearly in a whole sentence", -6),
    ("Asked for a break instead of melting down", -11), ("Finished a math test in the time", -20),
]

DATA_TARGET = 30
LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your child's details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "CHILD & PLAN", "section")
    controls = [
        ("Child's Name", STUDENT["name"], None, "ChildName"),
        ("Grade", STUDENT["grade"], None, "Grade"),
        ("Plan Type", STUDENT["plan"], None, "PlanType"),
        ("Case Manager", STUDENT["case"], None, "CaseManager"),
        ("Next Annual Review", dt.date.today() + dt.timedelta(days=60), "mm/dd/yyyy", "NextReview"),
        ("Data Points Target", DATA_TARGET, "0", "DataTarget"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Goal Area", AREAS, "AreaList"), ("F", "Yes / No", YESNO, "YesNoList"),
             ("G", "Setting", SETTINGS_LIST, "SettingList"), ("H", "Service", SERVICE_TYPE, "ServiceList"),
             ("I", "Meeting", MEET_TYPE, "MeetList"), ("J", "Done?", DONE_LIST, "DoneList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  💛  IEP & SPECIAL-NEEDS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  One calm place for goals, services, data & meetings — walk in prepared, advocate with confidence.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR CHILD'S WHOLE PLAN, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("IEP/504 goals with real progress monitoring, services & minutes, accommodations, a therapy log, "
                      "a gentle behavior tracker, meeting & communication logs, health/medication notes, a strengths "
                      "profile, a records checklist and a wins log — all in ONE calm Google Sheets & Excel system. See "
                      "exactly how each goal is trending, keep every note and date in one binder, and walk into the next "
                      "meeting organized and confident. Built by-a-parent, for parents — your child is so much more than a plan.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your child, plan type, case manager & next review date.",
             "2.  Enter each IEP Goal with its baseline, target & current level — progress calculates itself.",
             "3.  List Services & minutes, and check off the Accommodations in place.",
             "4.  Log therapy sessions, data points, behavior notes & every communication as they happen.",
             "5.  Keep the Records checklist current so the binder is always meeting-ready.",
             "6.  Watch the Dashboard track goal progress, services & a Progress Score — and celebrate the Wins."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  PLEASE READ", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional child (Sam) is included so you can see how it all connects — just type over "
               "it with your own. This is a personal organizing & advocacy tool to help you track your child's plan and "
               "prepare for meetings. It is NOT medical, legal, psychological or educational advice, and it does not "
               "create or replace an IEP/504 — always work with your child's IEP team and qualified professionals, and "
               "follow your district's process. Keep this file private; it may contain sensitive information.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "You know your child best. Walk in prepared — you've got this.", "section_gold")


# ===========================================================================
# 3 — Student Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Student Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 24, 4, 22, 22, 2])
    luxe_header(ws, "F", "🧒  STUDENT PROFILE & TEAM",
                "The whole child & the whole team — the one-page snapshot you bring to every meeting.")
    merge_set(ws, "B5:F5", STUDENT["name"], "section_gold"); ws.row_dimensions[5].height = 22
    fields = [("Grade / School", "3 · Springfield Elem."), ("Date of Birth", "—"),
              ("Plan Type", "IEP — SLD + Speech"), ("Eligibility Category", "Specific Learning Disability"),
              ("Case Manager", "Ms. Rivera"), ("Next Annual Review", "See Settings"),
              ("Diagnoses (if shared)", "—"), ("Strengths", "Kind, curious, great memory"),
              ("What helps", "Breaks, visuals, praise"), ("What's hard", "Long reading, timed work"),
              ("Parent contacts", "—"), ("Notes", "Loves dinosaurs & Legos")]
    row = 6; i = 0
    while i < len(fields):
        ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
        ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
        if i + 1 < len(fields):
            ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
            ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
        ws.row_dimensions[row].height = 24; i += 2; row += 1


# ===========================================================================
# 4 — IEP Goals  (defines GoalArea / GoalName / GoalPct)
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("IEP Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 18, 34, 12, 12, 12, 12, 2])
    luxe_header(ws, "G", "🎯  IEP GOALS & PROGRESS",
                "Every goal with its baseline, target & current level — progress toward each goal calculates itself.")
    table_headers(ws, 4, ["Area", "Goal", "Baseline", "Target", "Current", "Progress"], start_col=2)
    start = L0
    for i, (area, goal, base, target, current, unit) in enumerate(GOALS):
        r = start + i
        ws.cell(row=r, column=2, value=area).style = "td_left"
        ws.cell(row=r, column=3, value=goal).style = "td_left"
        cb = ws.cell(row=r, column=4, value=base); cb.style = "input"; cb.number_format = "0"
        ct = ws.cell(row=r, column=5, value=target); ct.style = "input"; ct.number_format = "0"
        cc = ws.cell(row=r, column=6, value=current); cc.style = "input"; cc.number_format = "0"
        cp = ws.cell(row=r, column=7, value=f"=IFERROR(MIN(MAX((F{r}-D{r})/(E{r}-D{r}),0),1),0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(GOALS) - 1
    nrange(wb, "GoalArea", "IEP Goals", "B", start, end)
    nrange(wb, "GoalName", "IEP Goals", "C", start, end)
    nrange(wb, "GoalPct", "IEP Goals", "G", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"
    merge_set(ws, f"B{end+2}:G{end+2}",
              "Progress = (current − baseline) ÷ (target − baseline). Update 'Current' as new data comes in.",
              "section_gold")


# ===========================================================================
# 5 — Progress Monitoring  (defines DataDate)
# ===========================================================================
def build_monitoring(wb):
    rows = [(dminus(-off) if off < 0 else dplus(off), area, measure, val) for (off, area, measure, val) in MONITORING]
    rows.sort(key=lambda r: r[0])
    ws, start, end = build_log(
        wb, "Progress Monitoring", "📈", "PROGRESS MONITORING",
        "Every data point, dated — the evidence behind each goal. Bring these charts to the meeting.",
        ["Date", "Goal Area", "Measure", "Value"],
        rows, [14, 22, 18, 12], text_left={3}, dates={1}, ints={4}, reserved=60,
        validations=[("B", "AreaList")])
    nrange(wb, "DataDate", "Progress Monitoring", "A", start, end)


# ===========================================================================
# 6 — Services & Minutes  (defines ServiceName / ServiceMin / ServiceDelivered)
# ===========================================================================
def build_services(wb):
    ws = wb.create_sheet("Services"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 26, 14, 14, 14, 2])
    luxe_header(ws, "F", "🧩  SERVICES & MINUTES",
                "The supports your child is owed — service, provider, frequency & minutes scheduled vs delivered.")
    table_headers(ws, 4, ["Service", "Provider", "Frequency", "Min/Wk", "Delivered"], start_col=2)
    start = L0
    for i, (svc, prov, freq, sched, deliv) in enumerate(SERVICES):
        r = start + i
        ws.cell(row=r, column=2, value=svc).style = "td_left"
        ws.cell(row=r, column=3, value=prov).style = "td_left"
        ws.cell(row=r, column=4, value=freq).style = "td"
        cs = ws.cell(row=r, column=5, value=sched); cs.style = "input"; cs.number_format = "#,##0"
        cd = ws.cell(row=r, column=6, value=deliv); cd.style = "input"; cd.number_format = "#,##0"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SERVICES) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL / WEEK").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cs = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = "#,##0"
    cd = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cd.style = "td"; cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = "#,##0"
    nrange(wb, "ServiceName", "Services", "B", start, end)
    nrange(wb, "ServiceMin", "Services", "E", start, end)
    nrange(wb, "ServiceDeliv", "Services", "F", start, end)
    cell_name(wb, "ServiceMinTotal", "Services", f"$E${tot}")
    cell_name(wb, "ServiceDelivTotal", "Services", f"$F${tot}")
    ws.freeze_panes = "A5"


# ===========================================================================
# 7 — Accommodations  (defines AccomActive)
# ===========================================================================
def build_accommodations(wb):
    ws, start, end = build_log(
        wb, "Accommodations", "🛠", "ACCOMMODATIONS & MODIFICATIONS",
        "Every accommodation & where it applies — check that each one is actually in place.",
        ["Accommodation / Modification", "Setting", "In Place?"],
        ACCOMMODATIONS, [46, 16, 14], text_left={1}, reserved=24,
        validations=[("B", "SettingList"), ("C", "YesNoList")])
    nrange(wb, "AccomActive", "Accommodations", "C", start, end)
    nrange(wb, "AccomName", "Accommodations", "A", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 8 — Therapy Log  (defines TherapyDate)
# ===========================================================================
def build_therapy(wb):
    rows = [(dminus(-off), typ, prov, note) for (off, typ, prov, note) in THERAPY]
    rows.sort(key=lambda r: r[0], reverse=True)
    ws, start, end = build_log(
        wb, "Therapy Log", "🗣", "THERAPY & SESSION LOG",
        "Every OT, PT, speech & counseling session — what was worked on and any home carryover.",
        ["Date", "Type", "Provider", "Focus / Note"],
        rows, [14, 20, 16, 38], text_left={4}, dates={1}, reserved=40,
        validations=[("B", "ServiceList")])
    nrange(wb, "TherapyDate", "Therapy Log", "A", start, end)


# ===========================================================================
# 9 — Behavior Tracker
# ===========================================================================
def build_behavior(wb):
    rows = [(dminus(-off), sit, beh, sup) for (off, sit, beh, sup) in BEHAVIOR]
    rows.sort(key=lambda r: r[0], reverse=True)
    build_log(wb, "Behavior", "🌈", "BEHAVIOR TRACKER (ABC)",
              "Gentle ABC notes — the situation, what happened & the support that helped. Patterns become clear.",
              ["Date", "Situation (before)", "Behavior", "Support that helped"],
              rows, [14, 26, 22, 28], text_left={2, 3, 4}, dates={1}, reserved=30)


# ===========================================================================
# 10 — Meeting Log  (defines MeetDate)
# ===========================================================================
def build_meetings(wb):
    rows = [(dminus(-off) if off < 0 else dplus(off), typ, att, dec) for (off, typ, att, dec) in MEETINGS]
    ws, start, end = build_log(
        wb, "Meetings", "📋", "MEETING LOG",
        "Every IEP meeting — who was there and what was decided. Your written record of every agreement.",
        ["Date", "Type", "Attendees", "Key Decisions"],
        rows, [14, 18, 26, 30], text_left={3, 4}, dates={1}, reserved=20,
        validations=[("B", "MeetList")])
    nrange(wb, "MeetDate", "Meetings", "A", start, end)


# ===========================================================================
# 11 — Communication Log
# ===========================================================================
def build_communication(wb):
    rows = [(dminus(-off), who, topic, follow) for (off, who, topic, follow) in COMMUNICATION]
    rows.sort(key=lambda r: r[0], reverse=True)
    build_log(wb, "Communication", "✉", "COMMUNICATION LOG",
              "Every email, call & note — dated, with follow-ups. A paper trail protects your child.",
              ["Date", "With", "Topic", "Follow-up"],
              rows, [14, 22, 30, 22], text_left={3, 4}, dates={1}, reserved=40)


# ===========================================================================
# 12 — Health & Medications
# ===========================================================================
def build_health(wb):
    ws = wb.create_sheet("Health & Meds"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 4, 22, 22, 2])
    luxe_header(ws, "F", "🩺  HEALTH & MEDICATIONS",
                "The health picture the team should know — keep it current & private.")
    fields = [("Doctor / clinic", "—"), ("Diagnoses (as shared)", "—"),
              ("Medications", "—"), ("Dosage & timing", "—"),
              ("Allergies", "—"), ("Sensory needs", "Noise-sensitive; likes movement"),
              ("Emergency contact", "—"), ("Health plan at school", "—"),
              ("Sleep / diet notes", "—"), ("Other notes", "—")]
    row = 5; i = 0
    while i < len(fields):
        ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
        ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
        if i + 1 < len(fields):
            ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
            ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
        ws.row_dimensions[row].height = 24; i += 2; row += 1
    merge_set(ws, f"B{row+1}:F{row+1}",
              "Keep this file private — it may contain sensitive health information. Share only with those who need it.",
              "section_gold")


# ===========================================================================
# 13 — Strengths & Interests
# ===========================================================================
def build_strengths(wb):
    ws = wb.create_sheet("Strengths"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 42, 42, 2])
    luxe_header(ws, "C", "⭐  STRENGTHS & INTERESTS",
                "Start every meeting here — your child is so much more than a list of needs.")
    cols = [
        ("STRENGTHS", ["Kind & empathetic", "Amazing long-term memory", "Curious about how things work",
                       "Persistent when motivated", "Great with younger kids", "Creative builder & problem-solver"]),
        ("INTERESTS & MOTIVATORS", ["Dinosaurs & Legos", "Being the 'helper'", "Earning tablet time",
                                     "Hands-on projects", "Drawing & comics", "Praise & high-fives"]),
    ]
    for c, (head, items) in enumerate(cols):
        x = 2 + c
        merge_set(ws, f"{get_column_letter(x)}5:{get_column_letter(x)}5", head, "section_gold")
        for i, it in enumerate(items):
            cell = ws.cell(row=6 + i, column=x, value="•  " + it)
            cell.style = "td_left"; cell.fill = fill(MINT_BG if c == 0 else SOFT_BG)
            ws.row_dimensions[6 + i].height = 26
    merge_set(ws, "B14:C14",
              "Bring this to every meeting — a strengths-based start changes the whole conversation.", "section_gold")


# ===========================================================================
# 14 — Records & Documents  (defines RecDone)
# ===========================================================================
def build_records(wb):
    ws, start, end = build_log(
        wb, "Records", "🗂", "RECORDS & DOCUMENTS",
        "The binder checklist — everything you want on hand & meeting-ready.",
        ["Document / Record", "Have It?", "Location / Notes"],
        RECORDS, [40, 14, 26], text_left={1, 3}, reserved=20,
        validations=[("B", "DoneList")])
    nrange(wb, "RecDone", "Records", "B", start, end)
    nrange(wb, "RecName", "Records", "A", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"In Progress"'], fill=fill(WARN_BG)))
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(RED_BG)))


# ===========================================================================
# 15 — Wins & Milestones  (defines WinName)
# ===========================================================================
def build_wins(wb):
    rows = [(win, dminus(-off)) for (win, off) in WINS]
    rows.sort(key=lambda r: r[1], reverse=True)
    ws, start, end = build_log(
        wb, "Wins", "🎉", "WINS & MILESTONES",
        "Every step forward, big or small — the progress that gets lost between meetings. Celebrate it here.",
        ["Win / Milestone", "Date"],
        rows, [50, 16], text_left={1}, dates={2}, reserved=30)
    nrange(wb, "WinName", "Wins", "A", start, end)


# ===========================================================================
# 2 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  💛  IEP & SPECIAL-NEEDS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Goals, services, data & meetings — your child's whole plan, organized and meeting-ready.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("IEP GOALS", "=COUNTA(GoalName)", "num"),
        ("AVG PROGRESS", "=IFERROR(AVERAGE(GoalPct),0)", "pct"),
        ("GOALS ON PACE", '=COUNTIF(GoalPct,">=0.6")', "num"),
        ("SERVICES", "=COUNTA(ServiceName)", "num"),
        ("SERVICE MIN/WK", "=ServiceMinTotal", "num"),
        ("ACCOMMODATIONS", '=COUNTIF(AccomActive,"Yes")', "num"),
    ]
    row2 = [
        ("THERAPY LOGGED", "=COUNTA(TherapyDate)", "num"),
        ("DATA POINTS", "=COUNTA(DataDate)", "num"),
        ("MEETINGS", "=COUNTA(MeetDate)", "num"),
        ("WINS", "=COUNTA(WinName)", "num"),
        ("SUPPORTS IN PLACE", '=IFERROR(COUNTIF(AccomActive,"Yes")/COUNTA(AccomActive),0)', "pct"),
        ("PROGRESS SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "HOW WE'RE DOING", "section_gold")
    merge_set(ws, "H11:M11", "PROGRESS TOWARD EACH GOAL", "section_gold")
    table_headers(ws, 12, ["Area", "Progress", "Status"], start_col=2)
    dims = [
        ("Goal progress", "=IFERROR(AVERAGE(GoalPct),0)"),
        ("Goals on pace", '=IFERROR(COUNTIF(GoalPct,">=0.6")/COUNTA(GoalPct),0)'),
        ("Accommodations in place", '=IFERROR(COUNTIF(AccomActive,"Yes")/COUNTA(AccomActive),0)'),
        ("Services delivered", "=IFERROR(ServiceDelivTotal/ServiceMinTotal,0)"),
        ("Data collection", "=IFERROR(MIN(COUNTA(DataDate)/DataTarget,1),0)"),
        ("Records ready", '=IFERROR(COUNTIF(RecDone,"Yes")/COUNTA(RecName),0)'),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"On track","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    g = wb["IEP Goals"]
    bar = BarChart(); bar.type = "bar"; bar.title = "Progress Toward Each Goal"; bar.height = 7.4; bar.width = 8.6
    bar.add_data(Reference(g, min_col=7, min_row=5, max_row=4 + len(GOALS)), titles_from_data=False)
    bar.set_categories(Reference(g, min_col=2, min_row=5, max_row=4 + len(GOALS)))
    bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
    bar.legend = None
    ws.add_chart(bar, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "IEP & Special-Needs Command Center™ — walk in prepared, advocate with confidence. Not medical or legal advice.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profile(wb); build_goals(wb)
    build_monitoring(wb); build_services(wb); build_accommodations(wb); build_therapy(wb)
    build_behavior(wb); build_meetings(wb); build_communication(wb); build_health(wb)
    build_strengths(wb); build_records(wb); build_wins(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Student Profile", "IEP Goals", "Progress Monitoring",
             "Services", "Accommodations", "Therapy Log", "Behavior", "Meetings",
             "Communication", "Health & Meds", "Strengths", "Records", "Wins", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "IEP_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
