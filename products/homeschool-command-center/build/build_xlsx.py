"""Build Homeschool Command Center™ — The Complete Homeschool Planning & Records System.

18 planning tabs (+ Settings) · a premium homeschool operating system in Google
Sheets & Excel. Homeschool dashboard, a profile for every student, year planner,
curriculum & resources, subject/course plans, weekly lesson planner, attendance &
hours log, assignments & grades, reading log, field trips, co-op, portfolio,
curriculum budget, goals, report card / transcript, records & compliance and a
book list — one dashboard. Scales from 1 student to a full house.

Run: python3 build_xlsx.py   ->  ../Homeschool_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

GRADES = ["PreK", "K", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]
SUBJECTS = ["Math", "Language Arts", "Science", "History", "Geography", "Art", "Music",
            "PE", "Bible / Character", "Foreign Language", "Life Skills", "Other"]
STATUS = ["Not Started", "In Progress", "Done"]
ACQUIRED = ["Have", "Ordered", "Need"]
GRADED = ["Graded", "Turned In", "Not Yet"]
APPROACH = ["Classical", "Charlotte Mason", "Unit Study", "Eclectic", "Unschool", "Traditional", "Montessori"]
YESNO = ["Yes", "No"]

# ---- sample family — the data that drives every KPI ----
STUDENTS = [
    ("Ella Bennett", "8", 13, "Classical", "Strong reader · loves history"),
    ("Noah Bennett", "6", 11, "Eclectic", "Hands-on · needs movement breaks"),
    ("Grace Bennett", "3", 8, "Charlotte Mason", "Nature study · early reader"),
    ("Sam Bennett", "K", 5, "Play-based", "Letters & numbers · lots of read-alouds"),
]
# attendance weeks: (label, days, hours)
WEEKS = [(f"Week {i}", 5, 27 if i <= 16 else 26) for i in range(1, 25)]
# curriculum: (subject, student, resource, cost, status)
CURRICULUM = [
    ("Math", "Ella", "Saxon 8/7", 110, "Have"), ("Math", "Noah", "Singapore 6", 95, "Have"),
    ("Math", "Grace", "Math-U-See", 85, "Have"), ("Math", "Sam", "Math games (K)", 40, "Have"),
    ("Language Arts", "Ella", "IEW writing", 120, "Have"), ("Language Arts", "Noah", "Grammar + lit", 90, "Have"),
    ("Language Arts", "Grace", "Phonics + readers", 70, "Have"), ("Language Arts", "Sam", "Handwriting", 35, "Have"),
    ("Science", "All", "Apologia + lab kit", 130, "Have"), ("History", "All", "Story of the World", 95, "Have"),
    ("Art", "All", "Art supply box", 60, "Have"), ("Music", "All", "Piano books", 45, "Have"),
    ("Foreign Language", "Ella/Noah", "Spanish (online)", 0, "Need"), ("Bible / Character", "All", "Devotional set", 30, "Have"),
    ("Geography", "All", "Maps + atlas", 55, "Have"), ("PE", "All", "Co-op sports", 80, "Have"),
    ("Life Skills", "All", "Cooking & typing", 40, "Have"), ("Language Arts", "Ella", "Literature guides", 60, "Have"),
]
# lessons / units: (subject, student, unit, status)
LESSONS = [
    ("Math", "Ella", "Ratios & proportions", "Done"), ("Math", "Ella", "Percents", "In Progress"),
    ("Math", "Noah", "Fractions", "Done"), ("Math", "Grace", "Multiplication facts", "Done"),
    ("Math", "Sam", "Counting to 100", "Done"), ("Language Arts", "Ella", "Essay: persuasive", "Done"),
    ("Language Arts", "Noah", "Parts of speech", "Done"), ("Language Arts", "Grace", "Short vowels", "Done"),
    ("Language Arts", "Sam", "Letter sounds", "In Progress"), ("Science", "All", "Cells & life", "Done"),
    ("Science", "All", "Weather unit", "Done"), ("Science", "All", "Human body", "In Progress"),
    ("History", "All", "Ancient Egypt", "Done"), ("History", "All", "Ancient Greece", "Done"),
    ("History", "All", "Rome", "In Progress"), ("Geography", "All", "Continents & oceans", "Done"),
    ("Geography", "All", "US states", "Not Started"), ("Art", "All", "Watercolor basics", "Done"),
    ("Art", "All", "Famous artists", "Done"), ("Music", "All", "Note reading", "Done"),
    ("Bible / Character", "All", "Proverbs study", "Done"), ("Life Skills", "All", "Money & budgeting", "Done"),
    ("PE", "All", "Soccer skills", "Done"), ("Foreign Language", "Ella", "Spanish greetings", "Not Started"),
    ("Language Arts", "Ella", "Poetry unit", "Done"), ("Science", "All", "Simple machines", "Done"),
    ("History", "All", "Timeline project", "Done"),
]
# assignments: (date, student, subject, assignment, grade, graded)
ASSIGNMENTS = [
    (14, "Ella", "Math", "Ch. 8 test", "94%", "Graded"), (13, "Ella", "Lang Arts", "Persuasive essay", "A", "Graded"),
    (12, "Noah", "Math", "Fractions quiz", "88%", "Graded"), (11, "Grace", "Reading", "Oral reading check", "S+", "Graded"),
    (10, "Ella", "History", "Egypt project", "A-", "Graded"), (9, "Noah", "Science", "Weather lab", "B+", "Graded"),
    (8, "Ella", "Science", "Cells worksheet", "92%", "Graded"), (7, "Grace", "Math", "Facts drill", "S", "Graded"),
    (6, "Noah", "Lang Arts", "Grammar quiz", "85%", "Graded"), (5, "Ella", "Math", "Percents set", "90%", "Graded"),
    (4, "Sam", "Reading", "Letter sounds", "S", "Graded"), (3, "Grace", "Science", "Nature journal", "S+", "Graded"),
    (2, "Ella", "History", "Greece essay", "A", "Graded"), (2, "Noah", "History", "Map quiz", "A-", "Graded"),
    (1, "Ella", "Lang Arts", "Poetry recitation", "A", "Graded"), (1, "Grace", "Art", "Watercolor", "S", "Graded"),
    (1, "Noah", "Math", "Ch. 9 test", "91%", "Graded"), (0, "Ella", "Science", "Body systems", "", "Not Yet"),
    (0, "Noah", "Lang Arts", "Book report", "", "Turned In"), (0, "Grace", "Math", "Unit test", "", "Not Yet"),
]
# records & compliance: (item, done)
RECORDS = [
    ("Letter of intent filed", True), ("Immunization / exemption on file", True), ("Attendance log current", True),
    ("Portfolio started", True), ("Standardized test scheduled", False), ("Curriculum plan documented", True),
    ("Samples saved (each subject)", True), ("Report cards / transcript updated", True),
    ("Umbrella / evaluator contact", True), ("End-of-year evaluation booked", False),
]
# goals: (goal, student, progress)
GOALS = [
    ("Read 60 books this year", "Ella", 0.90), ("Master multiplication facts", "Grace", 0.80),
    ("Finish Story of the World Vol. 1", "All", 0.75), ("Learn to read (CVC words)", "Sam", 0.70),
    ("Spanish: 100 words", "Ella/Noah", 0.60), ("Complete science fair project", "Noah", 0.57),
]
# field trips: (date offset, place, subject, who, cost)
FIELDTRIPS = [
    (-40, "Natural History Museum", "Science", "All", 0), (-28, "Colonial farm", "History", "All", 24),
    (-20, "Symphony kids concert", "Music", "All", 30), (-12, "Tide pools", "Science", "All", 0),
    (-6, "Local bakery tour", "Life Skills", "Grace/Sam", 0), (5, "State capitol", "Geography", "Ella/Noah", 0),
    (12, "Botanical gardens", "Science", "All", 18),
]
# reading log sample: (student, title, pages, done)
READING = [
    ("Ella", "The Hobbit", 310, "Yes"), ("Ella", "Number the Stars", 137, "Yes"),
    ("Noah", "Hatchet", 195, "Yes"), ("Grace", "Charlotte's Web", 184, "Yes"),
    ("Sam", "Frog and Toad (read-aloud)", 64, "Yes"), ("Ella", "The Bronze Bow", 254, "In Progress"),
    ("Noah", "The Wingfeather Saga", 288, "Yes"), ("Grace", "Little House series", 240, "In Progress"),
]
DAYS_REQUIRED = 180
HOURS_GOAL = 900
CURRICULUM_BUDGET = 1500
BOOKS_READ = 46
SUBJECT_COUNT = 8

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your family details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "HOMESCHOOL INPUTS", "section")
    controls = [
        ("Family Name", "The Bennett Family", None, "FamilyName"),
        ("School Year", "2026–2027", None, "SchoolYear"),
        ("State", "Ohio", None, "HomeState"),
        ("Students", 4, "0", "Students"),
        ("Core Subjects", SUBJECT_COUNT, "0", "Subjects"),
        ("Required Days", DAYS_REQUIRED, "0", "DaysRequired"),
        ("Instructional-Hours Goal", HOURS_GOAL, "#,##0", "HoursGoal"),
        ("Curriculum Budget", CURRICULUM_BUDGET, '"$"#,##0', "CurriculumBudget"),
        ("Books Read (year)", BOOKS_READ, "#,##0", "BooksRead"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Grade", GRADES, "GradeList"), ("F", "Subject", SUBJECTS, "SubjectList"),
             ("G", "Status", STATUS, "StatusList"), ("H", "Curriculum", ACQUIRED, "AcquiredList"),
             ("I", "Graded", GRADED, "GradedList"), ("J", "Approach", APPROACH, "ApproachList"),
             ("K", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  📚  HOMESCHOOL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Plan the year, log the days, keep the records — one calm system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE HOMESCHOOL, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Curriculum, lesson plans, attendance & hours, grades, reading, field trips, portfolio and the "
                      "records your state may ask for — all in ONE premium Google Sheets & Excel system. A profile for "
                      "every student, a live dashboard that shows if you're on track, and a lesson planner that actually "
                      "keeps up. Whether you follow Classical, Charlotte Mason, unit studies or your own beautiful mix, "
                      "this bends to your family — and scales from 1 student to a full house.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your family, state, required days & hours goal.",
             "2.  Fill a Student Profile for each child — grade, approach, strengths.",
             "3.  Plan Curriculum & Subjects, then map the Lesson Planner week by week.",
             "4.  Each day, log Attendance & Hours and check off lessons & assignments.",
             "5.  Keep Reading, Field Trips, Portfolio & Records current as you go.",
             "6.  Watch the Dashboard track days, hours, grades & an On-Track Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for the Bennett family (4 students, grades K–8) is included so you can see how it all "
               "connects — just type over it with your own. Attendance, hours, lessons, grades, goals and records roll "
               "up into a live On-Track Score. Twelve matching printable pages (student-at-a-glance, weekly plan, daily "
               "schedule, attendance, reading log, transcript & more) are included to print and keep. Requirements vary "
               "by state — always confirm your own state's rules; this is a record-keeping tool, not legal advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more time to actually teach — you've got this.", "section_gold")


# ===========================================================================
# 3 — Student Profiles
# ===========================================================================
def build_profiles(wb):
    ws = wb.create_sheet("Student Profiles"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 22, 4, 18, 18, 2])
    luxe_header(ws, "G", "🧑‍🎓  STUDENT PROFILES",
                "A profile for every student — grade, approach, strengths & goals.")
    row = 5
    for name, grade, age, approach, note in STUDENTS:
        merge_set(ws, f"B{row}:F{row}", name, "section_gold"); ws.row_dimensions[row].height = 22
        row += 1
        fields = [("Grade", grade), ("Age", age), ("Approach", approach), ("Learning style", "—"),
                  ("Strengths", note), ("Working on", "—"), ("Curriculum", "See Curriculum tab"),
                  ("Reading level", "—"), ("Interests", "—"), ("Big goal", "See Goals tab")]
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 22; i += 2; row += 1
        row += 1
    ws.freeze_panes = "A5"


# ===========================================================================
# 4 — Year Planner
# ===========================================================================
def build_year(wb):
    rows = [
        ("Term 1 begins", dt.date(2026, 8, 24), "Start Term 1", "6 weeks on"),
        ("Fall break", dt.date(2026, 10, 5), "Break", "1 week off"),
        ("Term 2 begins", dt.date(2026, 10, 12), "Start Term 2", "6 weeks on"),
        ("Thanksgiving break", dt.date(2026, 11, 23), "Break", "Nov 23–27"),
        ("Term 3 begins", dt.date(2026, 11, 30), "Start Term 3", "Portfolio review"),
        ("Winter break", dt.date(2026, 12, 21), "Break", "Return Jan 5"),
        ("Term 4 begins", dt.date(2027, 1, 5), "Start Term 4", "Standardized test window"),
        ("Spring break", dt.date(2027, 3, 22), "Break", "1 week off"),
        ("Term 5 begins", dt.date(2027, 3, 29), "Start Term 5", "Science fair"),
        ("Co-op ends", dt.date(2027, 5, 7), "Milestone", "Last co-op Friday"),
        ("Year-end evaluation", dt.date(2027, 5, 21), "Records", "Evaluator meeting"),
        ("Last day / Day 180", dt.date(2027, 5, 28), "Milestone", "Portfolio complete"),
    ]
    build_log(wb, "Year Planner", "🗓", "YEAR AT A GLANCE — 2026–2027",
              "Terms, breaks & big milestones — map the whole year, then work the plan.",
              ["Milestone", "Date", "Type", "Notes"],
              rows, [26, 14, 16, 28], text_left={1, 4}, dates={2}, reserved=40)


# ===========================================================================
# 5 — Curriculum & Resources  (defines CurricSpent)
# ===========================================================================
def build_curriculum(wb):
    ws, start, end = build_log(
        wb, "Curriculum", "📖", "CURRICULUM & RESOURCES",
        "Every subject, every student — what you're using, what it cost & what's still needed.",
        ["Subject", "Student", "Resource", "Cost", "Status"],
        CURRICULUM, [18, 14, 26, 12, 12], text_left={3}, money={4}, reserved=40,
        validations=[("A", "SubjectList"), ("E", "AcquiredList")])
    nrange(wb, "CurricCost", "Curriculum", "D", start, end)
    nrange(wb, "CurricStatus", "Curriculum", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=3, value="ACQUIRED TO DATE").style = "th"
    c = ws.cell(row=tot, column=4, value='=SUMIF(CurricStatus,"Have",CurricCost)')
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "CurricSpent", "Curriculum", f"$D${tot}")
    cmap = {"Have": MINT_BG, "Ordered": WARN_BG, "Need": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 6 — Subject Plan
# ===========================================================================
def build_subjects(wb):
    ws = wb.create_sheet("Subject Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 30, 16, 16, 2])
    luxe_header(ws, "E", "🧭  SUBJECT / COURSE PLAN",
                "Your course of study — the scope for each subject and how you'll cover it.")
    table_headers(ws, 4, ["Subject", "Scope this year", "Students", "Cadence"], start_col=2)
    rows = [
        ("Math", "Grade-level curriculum + facts practice", "All (own level)", "Daily"),
        ("Language Arts", "Reading, writing, grammar, spelling", "All (own level)", "Daily"),
        ("Science", "Life science + weekly labs", "Together", "3× / week"),
        ("History", "Ancients (Story of the World Vol. 1)", "Together", "3× / week"),
        ("Geography", "Continents, maps, US states", "Together", "Weekly"),
        ("Art", "Techniques + artist study", "Together", "Weekly"),
        ("Music", "Piano + composer study", "Together", "2× / week"),
        ("Bible / Character", "Proverbs + character focus", "Together", "Daily"),
    ]
    start = L0
    for i, (subj, scope, who, cad) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=subj).style = "td_left"
        ws.cell(row=r, column=3, value=scope).style = "td_left"
        ws.cell(row=r, column=4, value=who).style = "td"
        ws.cell(row=r, column=5, value=cad).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


# ===========================================================================
# 7 — Lesson Planner  (defines LessonStatus / LessonUnit)
# ===========================================================================
def build_lessons(wb):
    ws, start, end = build_log(
        wb, "Lesson Planner", "✏", "LESSON PLANNER",
        "Unit by unit, per subject — plan the scope and check it off as you go.",
        ["Subject", "Student", "Unit / Lesson", "Status"],
        LESSONS, [18, 14, 30, 14], text_left={3}, reserved=60,
        validations=[("A", "SubjectList"), ("D", "StatusList")])
    nrange(wb, "LessonUnit", "Lesson Planner", "C", start, end)
    nrange(wb, "LessonStatus", "Lesson Planner", "D", start, end)
    cmap = {"Done": MINT_BG, "In Progress": WARN_BG, "Not Started": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 8 — Attendance & Hours  (defines DaysDone / HoursLogged)
# ===========================================================================
def build_attendance(wb):
    rows = [(lbl, d, h, "") for (lbl, d, h) in WEEKS]
    ws, start, end = build_log(
        wb, "Attendance", "📅", "ATTENDANCE & HOURS LOG",
        "Days and instructional hours by week — the records your state may require.",
        ["Week", "Days", "Hours", "Notes"],
        rows, [16, 12, 12, 34], text_left={4}, ints={2}, dec={3}, reserved=40)
    nrange(wb, "AttDays", "Attendance", "B", start, end)
    nrange(wb, "AttHours", "Attendance", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL").style = "th"
    cd = ws.cell(row=tot, column=2, value="=SUM(AttDays)"); cd.style = "td"; cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = "#,##0"
    ch = ws.cell(row=tot, column=3, value="=SUM(AttHours)"); ch.style = "td"; ch.font = Font(bold=True, color=PRIMARY); ch.fill = fill(SURFACE); ch.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cell_name(wb, "DaysDone", "Attendance", f"$B${tot}")
    cell_name(wb, "HoursLogged", "Attendance", f"$C${tot}")
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=30, color=PRIMARY, showValue=True))


# ===========================================================================
# 9 — Assignments & Grades  (defines AssignGraded / AssignName)
# ===========================================================================
def build_assignments(wb):
    rows = [(dminus(off), student, subj, name, grade, gstat) for (off, student, subj, name, grade, gstat) in ASSIGNMENTS]
    ws, start, end = build_log(
        wb, "Assignments", "📝", "ASSIGNMENTS & GRADES",
        "Every graded assignment, per student — the raw material for report cards.",
        ["Date", "Student", "Subject", "Assignment", "Grade", "Status"],
        rows, [13, 14, 14, 24, 10, 14], text_left={4}, dates={1}, reserved=60,
        validations=[("C", "SubjectList"), ("F", "GradedList")])
    nrange(wb, "AssignName", "Assignments", "D", start, end)
    nrange(wb, "AssignGraded", "Assignments", "F", start, end)
    cmap = {"Graded": MINT_BG, "Turned In": WARN_BG, "Not Yet": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 10 — Reading Log
# ===========================================================================
def build_reading(wb):
    ws, start, end = build_log(
        wb, "Reading Log", "📕", "READING LOG",
        "Every book, per child — reading is the heart of it. Set the year's total in Settings.",
        ["Student", "Title", "Pages", "Finished?"],
        READING, [16, 32, 12, 14], text_left={2}, ints={3}, reserved=60,
        validations=[("D", "StatusList")])
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 11 — Field Trips & Activities  (defines TripName)
# ===========================================================================
def build_trips(wb):
    rows = [(dplus(off) if off >= 0 else dminus(-off), place, subj, who, cost)
            for (off, place, subj, who, cost) in FIELDTRIPS]
    ws, start, end = build_log(
        wb, "Field Trips", "🚌", "FIELD TRIPS & ACTIVITIES",
        "Learning beyond the table — every trip, what it covered & what it cost.",
        ["Date", "Place", "Subject Tie-in", "Who", "Cost"],
        rows, [14, 26, 18, 16, 12], text_left={2}, dates={1}, money={5}, reserved=30)
    nrange(wb, "TripName", "Field Trips", "B", start, end)


# ===========================================================================
# 12 — Co-op & Extracurriculars
# ===========================================================================
def build_coop(wb):
    rows = [
        ("Ella", "Writing co-op class", "Fri", "9:00–10:00", "$60/sem", "Homeschool co-op"),
        ("Noah", "STEM club", "Fri", "10:15–11:15", "$45/sem", "Homeschool co-op"),
        ("Grace", "Nature study group", "Wed", "1:00–2:30", "Free", "Park meetup"),
        ("Ella", "Piano lessons", "Tue", "3:30–4:15", "$120/mo", "Private teacher"),
        ("Noah", "Rec soccer", "Sat", "9:00–10:00", "$80/season", "City league"),
        ("Grace", "Ballet", "Mon", "4:00–5:00", "$65/mo", "Studio"),
        ("Sam", "Library story time", "Thu", "10:30–11:00", "Free", "Public library"),
    ]
    build_log(wb, "Co-op", "🤝", "CO-OP & EXTRACURRICULARS",
              "Co-op classes, lessons & activities — the master schedule & carpool view.",
              ["Student", "Activity", "Day", "Time", "Cost", "Where"],
              rows, [14, 22, 10, 14, 14, 20], text_left={6}, reserved=30)


# ===========================================================================
# 13 — Portfolio
# ===========================================================================
def build_portfolio(wb):
    ws = wb.create_sheet("Portfolio"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "🖼  PORTFOLIO & WORK SAMPLES",
                "Save a few samples per subject — your proof of a year well taught.")
    merge_set(ws, "B5:D5", "HOW TO ADD SAMPLES", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"link\"). "
                      "Excel: Insert ▸ Pictures ▸ Place in Cell. Snap a photo of a great worksheet, art project or "
                      "writing sample and caption it — subject, date & what it shows. A few per subject is plenty.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Writing Sample", "Math Work", "Art Project", "Science / Nature", "History Project", "Reading Milestone"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add sample)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Student · subject · date…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 14 — Curriculum Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "💰  HOMESCHOOL BUDGET",
                "Curriculum, books, co-op & trips — planned vs actual, so there are no surprises.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining"], start_col=2)
    rows = [
        ("Curriculum", 900, "=CurricSpent"), ("Books & Library", 120, 96), ("Art & Science Supplies", 150, 128),
        ("Co-op Fees", 220, 210), ("Lessons (music/sport)", 300, 285), ("Field Trips", 120, 72),
        ("Technology / Printing", 150, 84), ("Standardized Testing", 60, 0), ("Misc", 80, 45),
    ]
    start = L0
    for i, (cat, plan, actual) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=3, value=plan); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=actual)
        ca.style = "field_value" if cat == "Curriculum" else "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "BudgetPlanTotal", "Budget", f"$C${tot}")
    cell_name(wb, "BudgetSpent", "Budget", f"$D${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=900, color=PRIMARY, showValue=True))
    merge_set(ws, "B15:E15", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Total budget", "=BudgetPlanTotal", '"$"#,##0'), ("Spent so far", "=BudgetSpent", '"$"#,##0'),
             ("Remaining", "=BudgetPlanTotal-BudgetSpent", '"$"#,##0'),
             ("Per student (avg)", "=IFERROR(BudgetSpent/Students,0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 16 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Remaining", "Per student (avg)"):
            c.fill = fill(MINT_BG)


# ===========================================================================
# 15 — Goals & Milestones  (defines GoalProgress)
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 14, 2])
    luxe_header(ws, "D", "🎯  GOALS & MILESTONES",
                "The wins that matter — one meaningful goal per child, plus family goals.")
    table_headers(ws, 4, ["Goal", "Student", "Progress"], start_col=2)
    start = L0
    for i, (goal, who, prog) in enumerate(GOALS):
        r = start + i
        ws.cell(row=r, column=2, value=goal).style = "td_left"
        ws.cell(row=r, column=3, value=who).style = "td"
        cp = ws.cell(row=r, column=4, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(GOALS) - 1
    nrange(wb, "GoalProgress", "Goals", "D", start, end)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 16 — Report Card / Transcript
# ===========================================================================
def build_report(wb):
    ws = wb.create_sheet("Report Card"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 18, 13, 13, 13, 13, 16, 2])
    luxe_header(ws, "G", "🎓  REPORT CARD / TRANSCRIPT",
                "A clean term-by-term record per student — print-ready for your files or high school.")
    table_headers(ws, 4, ["Student — Subject", "T1", "T2", "T3", "T4", "Final"], start_col=2)
    rows = [
        ("Ella — Math", "A-", "", "", "", ""), ("Ella — Language Arts", "A", "", "", "", ""),
        ("Ella — Science", "A-", "", "", "", ""), ("Ella — History", "A", "", "", "", ""),
        ("Noah — Math", "B+", "", "", "", ""), ("Noah — Language Arts", "B", "", "", "", ""),
        ("Grace — Reading", "S+", "", "", "", ""), ("Grace — Math", "S", "", "", "", ""),
        ("Sam — Readiness", "S", "", "", "", ""),
    ]
    start = L0
    for i, row in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=row[0]).style = "td_left"
        for j in range(5):
            ws.cell(row=r, column=3 + j, value=row[1 + j]).style = "input"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"
    merge_set(ws, "B16:F16", "GRADING KEY", "section_gold"); ws.row_dimensions[16].height = 22
    keys = [("A / 90–100", "Excellent"), ("B / 80–89", "Good"), ("S / S+", "Satisfactory (younger)"),
            ("I", "In progress"), ("Credits", "0.5 per semester (HS)")]
    for i, (k, v) in enumerate(keys):
        r = 17 + i
        ws.cell(row=r, column=2, value=k).style = "field_label"
        ws.cell(row=r, column=3, value=v).style = "field_value"


# ===========================================================================
# 17 — Records & Compliance  (defines RecName / RecDone)
# ===========================================================================
def build_records(wb):
    rows = [(item, "Yes" if done else "No", "—") for (item, done) in RECORDS]
    ws, start, end = build_log(
        wb, "Records", "🗂", "RECORDS & COMPLIANCE",
        "The paperwork trail — check off what your state asks for. Confirm your own state's rules.",
        ["Requirement / Record", "Done?", "Link / Location"],
        rows, [40, 12, 30], text_left={1, 3}, reserved=24,
        validations=[("B", "YesNoList")])
    nrange(wb, "RecName", "Records", "A", start, end)
    nrange(wb, "RecDone", "Records", "B", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 18 — Book List / Library
# ===========================================================================
def build_booklist(wb):
    rows = [
        ("The Chronicles of Narnia", "C.S. Lewis", "Read-aloud", "Have", "Family favorite"),
        ("D'Aulaires' Greek Myths", "D'Aulaire", "History", "Have", "Pairs with Ancients"),
        ("The Wingfeather Saga", "Peterson", "Noah", "Have", "Noah loves it"),
        ("Little House series", "Wilder", "Grace", "Have", "History + read-aloud"),
        ("The Bronze Bow", "Speare", "Ella", "Have", "Newbery"),
        ("Handbook of Nature Study", "Comstock", "Reference", "Have", "Charlotte Mason staple"),
        ("Poetry for Young People", "various", "Language Arts", "Need", "For poetry unit"),
        ("Life of Fred", "Schmidt", "Math (fun)", "Need", "Requested by kids"),
    ]
    build_log(wb, "Book List", "📚", "BOOK LIST & WISH LIST",
              "The year's reading & the wish list — what to borrow, buy or save for later.",
              ["Title", "Author", "For", "Status", "Notes"],
              rows, [30, 18, 16, 12, 22], text_left={1, 5}, reserved=40,
              validations=[("D", "AcquiredList")])


# ===========================================================================
# 2 — Homeschool Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  📚  HOMESCHOOL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Days, hours, lessons, grades & records — your whole homeschool year, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("STUDENTS", "=Students", "num"),
        ("SCHOOL DAYS", "=DaysDone", "num"),
        ("HOURS LOGGED", "=HoursLogged", "num"),
        ("DAYS LEFT", "=DaysRequired-DaysDone", "num"),
        ("SUBJECTS", "=Subjects", "num"),
        ("LESSONS DONE", '=IFERROR(COUNTIF(LessonStatus,"Done")/COUNTA(LessonUnit),0)', "pct"),
    ]
    row2 = [
        ("GRADED", '=IFERROR(COUNTIF(AssignGraded,"Graded")/COUNTA(AssignName),0)', "pct"),
        ("BOOKS READ", "=BooksRead", "num"),
        ("CURRICULUM SPENT", "=CurricSpent", "money"),
        ("FIELD TRIPS", "=COUNTA(TripName)", "num"),
        ("GOALS", "=IFERROR(AVERAGE(GoalProgress),0)", "pct"),
        ("ON-TRACK", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "ARE WE ON TRACK?", "section_gold")
    merge_set(ws, "H11:M11", "BUDGET BY CATEGORY", "section_gold")
    table_headers(ws, 12, ["Area", "Progress", "Status"], start_col=2)
    dims = [
        ("School days (of 180)", "=IFERROR(DaysDone/DaysRequired,0)"),
        ("Instructional hours", "=IFERROR(HoursLogged/HoursGoal,0)"),
        ("Lessons done", '=IFERROR(COUNTIF(LessonStatus,"Done")/COUNTA(LessonUnit),0)'),
        ("Assignments graded", '=IFERROR(COUNTIF(AssignGraded,"Graded")/COUNTA(AssignName),0)'),
        ("Goals & milestones", "=IFERROR(AVERAGE(GoalProgress),0)"),
        ("Records & compliance", '=IFERROR(COUNTIF(RecDone,"Yes")/COUNTA(RecName),0)'),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"On track",IF(C{r}>=0.6,"Going well","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    db = DoughnutChart(); db.title = "Budget by Category"; db.height = 7.4; db.width = 8.4
    db.add_data(Reference(wb["Budget"], min_col=4, min_row=4, max_row=13), titles_from_data=True)
    db.set_categories(Reference(wb["Budget"], min_col=2, min_row=5, max_row=13)); db.dataLabels = no_labels()
    ws.add_chart(db, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Homeschool Command Center™ — plan the year, log the days, keep the records. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profiles(wb); build_year(wb)
    build_curriculum(wb); build_subjects(wb); build_lessons(wb); build_attendance(wb)
    build_assignments(wb); build_reading(wb); build_trips(wb); build_coop(wb)
    build_portfolio(wb); build_budget(wb); build_goals(wb); build_report(wb)
    build_records(wb); build_booklist(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Student Profiles", "Year Planner", "Curriculum", "Subject Plan",
             "Lesson Planner", "Attendance", "Assignments", "Reading Log", "Field Trips", "Co-op",
             "Portfolio", "Budget", "Goals", "Report Card", "Records", "Book List", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Homeschool_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
