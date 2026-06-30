"""Build Home Command Center™ — The Ultimate Household Management System.

28 sheets + Welcome · a premium household operating system in one workbook.
Finances, schedules, cleaning, meals, maintenance, inventories, routines and
long-term goals — all wired into one Executive Home Dashboard.

Organizational & planning tool only. Run: python3 build_xlsx.py
Outputs: ../Home_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"
ACCENT = "937356"
GOLD_LT = "C9A86A"
SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"
MINT_BG = "E3F8EF"
BG = "FFFFFF"
WHITE = "FFFFFF"
TEXT = "333333"
DANGER = "C94C4C"
RED_BG = "FBE6E6"
WARN_BG = "FBF0E2"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"
IVORY = "FBF8F2"

# ---- dropdown vocabularies ----
FAMILY = ["Mom", "Dad", "Emma", "Liam", "Grandma"]
EXPENSE_CATS = ["Mortgage/Rent", "Utilities", "Internet", "Phones", "Insurance",
                "Groceries", "Fuel", "Transportation", "Childcare", "Clothing",
                "Entertainment", "Medical", "Savings", "Subscriptions",
                "Dining Out", "Home Improvement", "Pets", "Miscellaneous"]
GROCERY_CATS = ["Produce", "Meat & Seafood", "Dairy", "Bakery", "Frozen",
                "Pantry", "Cleaning", "Toiletries", "Baby", "Pet Supplies"]
PANTRY_CATS = ["Grains & Pasta", "Canned Goods", "Baking", "Snacks",
               "Breakfast", "Spices & Oils", "Beverages", "Other"]
CLEAN_FREQ = ["Daily", "Weekly", "Monthly", "Seasonal", "Deep Clean"]
CHORE_FREQ = ["Daily", "Weekly", "Monthly", "As Needed"]
MAINT_CATS = ["HVAC", "Plumbing", "Electrical", "Appliances", "Exterior",
              "Lawn & Garden", "Pest Control", "Vehicle", "Safety"]
GOAL_CATS = ["Financial", "Home", "Travel", "Education", "Health", "Personal"]
MEAL_TYPES = ["Breakfast", "Lunch", "Dinner", "Snack", "Dessert"]
STATUSES = ["Not Started", "In Progress", "Done"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]
STORES = ["Costco", "Target", "Kroger", "Walmart", "Trader Joe's", "Amazon", "Local"]
CAL_CATS = ["School", "Sports", "Appointment", "Holiday", "Vacation", "Work",
            "Birthday", "Bill Due", "Maintenance", "Family"]

LOG_ROWS = 40
L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


# ===========================================================================
# Styles & shared helpers (premium pattern)
# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"),
                            fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True),
                               fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY),
                              alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT),
                                   alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"),
                         fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT),
                              alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
                              border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY),
                            fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX),
        "prompt": NamedStyle(name="prompt", font=f(11, False, TEXT, italic=True),
                             alignment=Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True),
                             border=BOX, fill=PatternFill("solid", fgColor=IVORY)),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True),
                             fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT),
                           alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None,
               dates=None, pcts=None, start_col=1):
    text_left = text_left or set()
    money = money or set()
    ints = ints or set()
    dates = dates or set()
    pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else BG)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}")
    ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label
    lc.font = Font(size=9, bold=True, color=ACCENT)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vc.font = Font(size=20, bold=True, color=PRIMARY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pct": "0%",
                        "days": "0"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc)
            c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN,
                              top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 20
    ws.row_dimensions[row + 1].height = 44


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers),
               text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def name_range(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def no_labels():
    dl = DataLabelList()
    dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# 28 — Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 3] + [17] * 7)
    luxe_header(ws, "K", "⚙  SETTINGS", "Set your inputs once — every dashboard follows. Edit the lists to fit your home.")

    merge_set(ws, "B5:C5", "HOUSEHOLD INPUTS", "section")
    controls = [
        ("Household Name", "The Anderson Home", None, "HouseholdName"),
        ("Monthly Income", 7200, '"$"#,##0', "MonthlyIncome"),
        ("Monthly Budget", 6400, '"$"#,##0', "MonthlyBudget"),
        ("Week Starts", dminus(dt.date.today().weekday()), "mm/dd/yyyy", "WeekStart"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")

    bank1 = [("E", "Family Members", FAMILY, "FamilyList"),
             ("F", "Expense Category", EXPENSE_CATS, "ExpenseCatList"),
             ("G", "Grocery Category", GROCERY_CATS, "GroceryCatList"),
             ("H", "Pantry Category", PANTRY_CATS, "PantryCatList"),
             ("I", "Cleaning Frequency", CLEAN_FREQ, "CleanFreqList"),
             ("J", "Chore Frequency", CHORE_FREQ, "ChoreFreqList"),
             ("K", "Maintenance Category", MAINT_CATS, "MaintCatList")]
    _emit_lists(wb, ws, bank1, header_row=5, data_row=6)

    bank2 = [("E", "Goal Category", GOAL_CATS, "GoalCatList"),
             ("F", "Meal Type", MEAL_TYPES, "MealTypeList"),
             ("G", "Status", STATUSES, "StatusList"),
             ("H", "Priority", PRIORITIES, "PriorityList"),
             ("I", "Calendar Category", CAL_CATS, "CalCatList"),
             ("J", "Store", STORES, "StoreList"),
             ("K", "Yes / No", YESNO, "YesNoList")]
    _emit_lists(wb, ws, bank2, header_row=24, data_row=25)


def _emit_lists(wb, ws, lists, header_row, data_row):
    merge_set(ws, f"E{header_row}:K{header_row}", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in lists:
        ci = column_index_from_string(col)
        ws.cell(row=header_row + 1, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=data_row + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(
            nm, attr_text=f"Settings!${col}${data_row + 1}:${col}${data_row + len(data)}")
    # shift header label up one (lists header above the column header)
    # data starts data_row+1


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 74, 3])
    ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🏡  HOME COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate household operating system — your whole home in one place.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    merge_set(ws, "B5:B5", "WELCOME HOME", "section_gold")
    ws.merge_cells("B6:B8")
    ws["B6"].value = (
        "Running a household means juggling finances, schedules, meals, cleaning, "
        "maintenance, and a hundred small details. Home Command Center™ brings all of "
        "it into one elegant dashboard so your family runs on a system — not on memory. "
        "Less mental load, more calm, every single day.")
    ws["B6"].style = "body"
    for r in (6, 7, 8):
        ws.row_dimensions[r].height = 22

    merge_set(ws, "B10:B10", "HOW TO USE IT", "section")
    steps = [
        "1.  Open Settings and add your household name, income & monthly budget.",
        "2.  Fill the Family Directory, Budget, Bills, and the Master Calendar.",
        "3.  Plan meals, build your grocery list, and track the pantry & fridge.",
        "4.  Assign chores, schedule cleaning, and log home maintenance.",
        "5.  Watch the Executive Home Dashboard update itself automatically.",
    ]
    for i, s in enumerate(steps):
        r = 11 + i
        ws.merge_cells(f"B{r}:B{r}")
        ws[f"B{r}"].value = s
        ws[f"B{r}"].style = "body"
        ws.row_dimensions[r].height = 22

    dr = 18
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th")
    ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = (
        "Home Command Center™ is an organizational and planning tool. Sample data is "
        "included to show you how everything connects — just type over it with your own. "
        "Every sheet is print-friendly and works in Excel and Google Sheets, on desktop "
        "and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT)
    c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22
        ws.cell(row=rr, column=2).fill = fill(WARN_BG)

    merge_set(ws, f"B{dr+5}:B{dr+5}",
              "One organized home, one calm command center — let's begin.", "section_gold")


# ===========================================================================
# 2 — Family Directory
# ===========================================================================
def build_family_directory(wb):
    ws = wb.create_sheet("Family Directory")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 16, 14, 12, 20, 22, 22, 24])
    luxe_header(ws, "H", "👨‍👩‍👧‍👦  FAMILY DIRECTORY",
                "Everyone who matters, all in one secure place — people, providers & pets.")

    merge_set(ws, "A4:H4", "FAMILY MEMBERS", "section_gold")
    table_headers(ws, 5, ["Name", "Relationship", "Birthday", "Blood Type",
                          "Allergies", "Medications", "Doctor", "Notes"])
    fam = [
        ("Sarah Anderson", "Mom", dt.date(1988, 4, 12), "O+", "None", "—", "Dr. Lee", "Primary calendar owner"),
        ("Mike Anderson", "Dad", dt.date(1986, 9, 3), "A+", "Penicillin", "—", "Dr. Lee", ""),
        ("Emma Anderson", "Daughter", dt.date(2014, 6, 21), "O+", "Peanuts", "EpiPen", "Dr. Patel", "Grade 4"),
        ("Liam Anderson", "Son", dt.date(2017, 1, 9), "A+", "None", "Inhaler", "Dr. Patel", "Grade 1"),
        ("Grace Miller", "Grandma", dt.date(1957, 11, 30), "B+", "Sulfa", "BP meds", "Dr. Cohen", "Emergency contact"),
    ]
    style_rows(ws, 6, 13, 8, text_left={1, 5, 6, 7, 8}, dates={3})
    for i, row in enumerate(fam):
        for ci, val in enumerate(row, 1):
            ws.cell(row=6 + i, column=ci, value=val)

    merge_set(ws, "A15:H15", "KEY PROVIDERS & EMERGENCY CONTACTS", "section_gold")
    table_headers(ws, 16, ["Type", "Name / Practice", "Phone", "For Whom",
                           "Account / Policy #", "Address", "Email", "Notes"])
    prov = [
        ("Pediatrician", "Bright Kids Peds — Dr. Patel", "(555) 210-7788", "Emma, Liam", "—", "12 Oak Ave", "", ""),
        ("Family Doctor", "Anderson Family Med — Dr. Lee", "(555) 210-4521", "Mom, Dad", "—", "4 Main St", "", ""),
        ("Dentist", "Smile Dental", "(555) 332-1190", "All", "—", "88 Elm St", "", "6-mo cleanings"),
        ("School", "Maple Elementary", "(555) 778-3300", "Emma, Liam", "Student IDs in Children's Hub", "", "", ""),
        ("Home Insurance", "SafeGuard Insurance", "(555) 909-2020", "Household", "POL-44821", "", "", "Renews Mar"),
        ("Auto Insurance", "DriveWell", "(555) 909-4545", "Vehicles", "AUT-91022", "", "", ""),
        ("Emergency", "Grace Miller (Grandma)", "(555) 661-2048", "Children", "—", "", "", "If parents unreachable"),
        ("Vet", "Paws & Claws Vet", "(555) 443-7766", "Pets", "—", "", "", "See Pet Care"),
    ]
    style_rows(ws, 17, 26, 8, text_left={2, 5, 6, 7, 8})
    for i, row in enumerate(prov):
        for ci, val in enumerate(row, 1):
            ws.cell(row=17 + i, column=ci, value=val)
    add_dv(ws, "B6:B13", "FamilyList")
    ws.freeze_panes = "A6"


# ===========================================================================
# 3 — Master Family Calendar
# ===========================================================================
def build_calendar(wb):
    today = dt.date.today()
    sample = [
        ("Soccer practice", "Sports", dplus(1), "4:00 PM", "Emma", "Field 3", ""),
        ("Mortgage due", "Bill Due", dplus(2), "—", "Household", "Auto-pay", ""),
        ("Dentist — all", "Appointment", dplus(3), "9:00 AM", "All", "Smile Dental", "Cleanings"),
        ("Liam reading night", "School", dplus(4), "6:00 PM", "Liam", "School", ""),
        ("Grandma's birthday", "Birthday", dplus(6), "—", "Grace", "Home", "Cake + gift"),
        ("HVAC filter change", "Maintenance", dplus(9), "—", "Dad", "Home", ""),
        ("Family movie night", "Family", dplus(5), "7:30 PM", "All", "Living room", ""),
        ("Parent-teacher conf.", "School", dplus(12), "3:30 PM", "Emma", "Maple Elem.", ""),
        ("Internet bill due", "Bill Due", dplus(8), "—", "Household", "Auto-pay", ""),
        ("Weekend trip", "Vacation", dplus(20), "—", "All", "Lakeside", "Pack Fri"),
        ("Soccer game", "Sports", dplus(15), "10:00 AM", "Emma", "Central Park", ""),
        ("Car oil change", "Maintenance", dplus(18), "8:00 AM", "Dad", "DriveWell", ""),
    ]
    ws, start, end = build_log(
        wb, "Calendar", "📆", "MASTER FAMILY CALENDAR",
        "Every event in one shared calendar — with automatic week & month summaries.",
        ["Event", "Category", "Date", "Time", "Who", "Location", "Notes"],
        sample, [28, 16, 14, 12, 14, 18, 22],
        text_left={1, 6, 7}, dates={3},
        validations=[("B", "CalCatList"), ("E", "FamilyList")], reserved=60)
    ws.cell(row=4, column=8, value="In").style = "th"
    ws.column_dimensions["H"].width = 8
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=8, value=f'=IF(C{r}="","",C{r}-TODAY())')
        c.style = "td"; c.number_format = "0;[Red]-0"
        c.fill = fill(MUTED_ROW if (r - start) % 2 else BG)
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'AND($C{start}<>"",$C{start}-TODAY()>=0,$C{start}-TODAY()<=7)'],
                    fill=fill(MINT_BG)))
    name_range(wb, "CalDate", "Calendar", "C", start, end)
    name_range(wb, "CalEvent", "Calendar", "A", start, end)
    # summaries
    ws.cell(row=4, column=10, value="THIS WEEK").style = "section_gold"
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 12
    summ = [
        ("Events this week", '=SUMPRODUCT((CalDate<>"")*(CalDate>=TODAY())*(CalDate<=TODAY()+7))'),
        ("Events this month", '=SUMPRODUCT((CalDate<>"")*(CalDate>=TODAY())*(CalDate<=TODAY()+30))'),
        ("Total upcoming", '=SUMPRODUCT((CalDate<>"")*(CalDate>=TODAY()))'),
        ("Next event in (days)", '=IFERROR(SMALL(IF(CalDate>=TODAY(),CalDate-TODAY()),1),0)'),
    ]
    for i, (lab, fml) in enumerate(summ):
        r = 5 + i
        ws.cell(row=r, column=10, value=lab).style = "field_label"
        c = ws.cell(row=r, column=11, value=fml); c.style = "field_value"; c.number_format = "0"


# ===========================================================================
# 4 — Daily Command Center
# ===========================================================================
def build_daily(wb):
    ws = wb.create_sheet("Daily")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 4, 30, 18, 3, 4, 30, 18, 2])
    luxe_header(ws, "I", "☀  DAILY COMMAND CENTER",
                "Your one-page daily driver — routines, priorities, and wins.")

    def checklist(col0, title, items, top):
        merge_set(ws, f"{get_column_letter(col0+1)}{top}:{get_column_letter(col0+2)}{top}", title, "section_gold")
        ws.row_dimensions[top].height = 22
        for i, it in enumerate(items):
            r = top + 1 + i
            ws.cell(row=r, column=col0, value="☐").alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=col0).font = Font(size=12, color=ACCENT)
            ws.cell(row=r, column=col0).border = BOX
            c = ws.cell(row=r, column=col0 + 1, value=it); c.style = "td_left"
            ws.merge_cells(f"{get_column_letter(col0+1)}{r}:{get_column_letter(col0+2)}{r}")
            ws.cell(row=r, column=col0 + 2).border = BOX

    checklist(2, "MORNING ROUTINE",
              ["Make beds", "Breakfast & lunches", "Start a load of laundry",
               "Quick kitchen reset", "Check today's calendar", "Vitamins & meds"], 5)
    checklist(6, "EVENING ROUTINE",
              ["Tidy main living areas", "Prep tomorrow's bags", "Dishes & wipe counters",
               "Set out clothes", "Plan tomorrow's dinner", "Lights & doors check"], 5)

    merge_set(ws, "B13:D13", "TOP 3 PRIORITIES", "section")
    for i in range(3):
        r = 14 + i
        ws.cell(row=r, column=2, value=i + 1).style = "input"
        c = ws.cell(row=r, column=3, value="").style = "field_value"
        ws.merge_cells(f"C{r}:D{r}")
        ws.cell(row=r, column=3).border = BOX
        ws.cell(row=r, column=4).border = BOX

    merge_set(ws, "F13:H13", "APPOINTMENTS & ERRANDS", "section")
    appt = ["8:30  School drop-off", "12:00  Grocery pickup", "3:30  Emma soccer", "6:00  Family dinner"]
    for i, a in enumerate(appt):
        r = 14 + i
        c = ws.cell(row=r, column=6, value=a); c.style = "td_left"
        ws.merge_cells(f"F{r}:H{r}")
        ws.cell(row=r, column=7).border = BOX
        ws.cell(row=r, column=8).border = BOX

    merge_set(ws, "B19:D19", "WATER  (8 glasses)", "section_gold")
    for i in range(8):
        c = ws.cell(row=20, column=2 + i if i < 3 else 2 + i, value="○")
    for i in range(8):
        cell = ws.cell(row=20, column=2 + i)
        cell.value = "○"; cell.alignment = Alignment(horizontal="center")
        cell.font = Font(size=14, color=HIGHLIGHT); cell.border = BOX

    merge_set(ws, "F19:H19", "MOVEMENT / EXERCISE", "section_gold")
    ws.merge_cells("F20:H20"); ws["F20"].style = "field_value"; ws["F20"].value = "30-min walk + stretch"
    ws["F20"].border = BOX; ws["G20"].border = BOX; ws["H20"].border = BOX

    merge_set(ws, "B22:D22", "NOTES", "section")
    ws.merge_cells("B23:D27")
    ws["B23"].style = "prompt"; ws["B23"].value = "Anything on your mind..."
    for r in range(23, 28):
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).border = BOX

    merge_set(ws, "F22:H22", "TODAY'S WINS  ★", "section_gold")
    ws.merge_cells("F23:H27")
    ws["F23"].style = "prompt"; ws["F23"].value = "Three things that went right today..."
    for r in range(23, 28):
        for c in (6, 7, 8):
            ws.cell(row=r, column=c).border = BOX


# ===========================================================================
# 5 — Household Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 14, 14, 12, 3, 22, 16])
    luxe_header(ws, "H", "💵  HOUSEHOLD BUDGET",
                "Plan vs actual by category — cash flow, savings rate & remaining, automatically.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining", "% Used"])
    planned = {"Mortgage/Rent": 1850, "Utilities": 300, "Internet": 75, "Phones": 140,
               "Insurance": 360, "Groceries": 850, "Fuel": 220, "Transportation": 180,
               "Childcare": 600, "Clothing": 150, "Entertainment": 200, "Medical": 160,
               "Savings": 600, "Subscriptions": 90, "Dining Out": 240,
               "Home Improvement": 200, "Pets": 95, "Miscellaneous": 140}
    actual = {"Mortgage/Rent": 1850, "Utilities": 318, "Internet": 75, "Phones": 140,
              "Insurance": 360, "Groceries": 902, "Fuel": 205, "Transportation": 165,
              "Childcare": 600, "Clothing": 88, "Entertainment": 175, "Medical": 60,
              "Savings": 600, "Subscriptions": 96, "Dining Out": 268,
              "Home Improvement": 120, "Pets": 110, "Miscellaneous": 95}
    start = L0
    end = start + len(EXPENSE_CATS) - 1
    for i, cat in enumerate(EXPENSE_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=2, value=planned[cat]); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=actual[cat]); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in range(2, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=total, column=5, value=f"=IFERROR(C{total}/B{total},0)")
    cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.2, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))
    name_range(wb, "BudgetCat", "Budget", "A", start, end)
    name_range(wb, "BudgetActual", "Budget", "C", start, end)
    wb.defined_names["BudgetTotalPlanned"] = DefinedName("BudgetTotalPlanned", attr_text=f"Budget!$B${total}")
    wb.defined_names["BudgetTotalActual"] = DefinedName("BudgetTotalActual", attr_text=f"Budget!$C${total}")

    ws.cell(row=4, column=7, value="THIS MONTH").style = "section_gold"
    kpis = [("Income", "=MonthlyIncome", '"$"#,##0'),
            ("Total Spent", "=BudgetTotalActual", '"$"#,##0'),
            ("Remaining", "=BudgetTotalPlanned-BudgetTotalActual", '"$"#,##0;[Red]-"$"#,##0'),
            ("Cash Flow", "=MonthlyIncome-BudgetTotalActual", '"$"#,##0;[Red]-"$"#,##0'),
            ("Savings Rate", '=IFERROR(SUMIF(BudgetCat,"Savings",BudgetActual)/MonthlyIncome,0)', "0%")]
    for i, (lab, fml, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml); c.style = "field_value"; c.number_format = fmt
    donut = DoughnutChart(); donut.title = "Monthly Spending"; donut.height = 8.4; donut.width = 12.5
    donut.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    donut.dataLabels = no_labels()
    ws.add_chart(donut, "G12")
    ws.freeze_panes = "A5"


# ===========================================================================
# 6 — Bill Command Center
# ===========================================================================
def build_bills(wb):
    sample = [
        ("Mortgage", "Mortgage/Rent", 1850, dplus(2), "Yes", "No", "", dplus(335)),
        ("Electric", "Utilities", 165, dplus(6), "No", "No", "", ""),
        ("Water/Sewer", "Utilities", 78, dplus(11), "No", "No", "", ""),
        ("Internet", "Internet", 75, dplus(8), "Yes", "No", "", ""),
        ("Cell Phones", "Phones", 140, dplus(14), "Yes", "No", "", ""),
        ("Home Insurance", "Insurance", 180, dplus(20), "Yes", "No", "", dplus(120)),
        ("Auto Insurance", "Insurance", 180, dplus(20), "Yes", "No", "", dplus(95)),
        ("Childcare", "Childcare", 600, dplus(1), "No", "No", "", ""),
        ("Trash Service", "Utilities", 38, dplus(-2), "No", "No", "", ""),
        ("Gym Membership", "Subscriptions", 45, dplus(16), "Yes", "Yes", "AUTH-7781", dplus(196)),
        ("Streaming Bundle", "Subscriptions", 38, dplus(9), "Yes", "Yes", "", ""),
        ("Credit Card", "Miscellaneous", 420, dplus(4), "No", "No", "", ""),
    ]
    ws, start, end = build_log(
        wb, "Bills", "📬", "BILL COMMAND CENTER",
        "Never miss a due date — overdue bills flag themselves automatically.",
        ["Company", "Category", "Amount", "Due Date", "Auto Pay", "Paid",
         "Confirmation #", "Renewal Date"],
        sample, [20, 16, 12, 14, 11, 10, 16, 14],
        text_left={1, 7}, dates={4, 8}, money={3},
        validations=[("B", "ExpenseCatList"), ("E", "YesNoList"), ("F", "YesNoList")],
        reserved=40)
    name_range(wb, "BillName", "Bills", "A", start, end)
    name_range(wb, "BillAmount", "Bills", "C", start, end)
    name_range(wb, "BillDue", "Bills", "D", start, end)
    name_range(wb, "BillPaid", "Bills", "F", start, end)
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'AND($D{start}<>"",$D{start}<TODAY(),$F{start}<>"Yes")'],
                    fill=fill(RED_BG)))
    # paid/unpaid summary for donut
    ws.cell(row=4, column=10, value="STATUS").style = "section_gold"
    ws.column_dimensions["J"].width = 16
    ws.column_dimensions["K"].width = 12
    ws.cell(row=5, column=10, value="Paid").style = "td_left"
    ws.cell(row=5, column=11, value='=COUNTIF(BillPaid,"Yes")').style = "td"
    ws.cell(row=6, column=10, value="Unpaid").style = "td_left"
    ws.cell(row=6, column=11, value='=COUNTIF(BillPaid,"No")').style = "td"
    ws.cell(row=8, column=10, value="Due this month").style = "field_label"
    ws.cell(row=8, column=11, value='=SUMPRODUCT((BillDue<>"")*(BillDue<=TODAY()+30)*BillAmount)').number_format = '"$"#,##0'
    ws.cell(row=8, column=11).style = "field_value"


# ===========================================================================
# 7 — Grocery Planner
# ===========================================================================
def build_grocery(wb):
    sample = [
        ("Bananas", "Produce", 2, "Kroger", 2.0, "No", ""),
        ("Spinach", "Produce", 1, "Kroger", 3.5, "No", ""),
        ("Chicken breast", "Meat & Seafood", 3, "Costco", 18.0, "No", ""),
        ("Ground beef", "Meat & Seafood", 2, "Costco", 12.0, "No", ""),
        ("Milk", "Dairy", 2, "Kroger", 6.0, "Yes", "3.05"),
        ("Eggs", "Dairy", 1, "Kroger", 4.5, "Yes", "4.20"),
        ("Bread", "Bakery", 2, "Trader Joe's", 5.0, "No", ""),
        ("Frozen veggies", "Frozen", 3, "Costco", 9.0, "No", ""),
        ("Pasta", "Pantry", 4, "Kroger", 6.0, "No", ""),
        ("Dish soap", "Cleaning", 1, "Target", 4.0, "No", ""),
        ("Toothpaste", "Toiletries", 2, "Target", 6.5, "No", ""),
        ("Diapers", "Baby", 1, "Amazon", 24.0, "Yes", "23.10"),
        ("Dog food", "Pet Supplies", 1, "Amazon", 32.0, "No", ""),
    ]
    ws, start, end = build_log(
        wb, "Grocery", "🛒", "GROCERY PLANNER",
        "Build your list by category & store — estimated vs actual cost totals itself.",
        ["Item", "Category", "Qty", "Store", "Est. Cost", "Bought?", "Actual Cost"],
        sample, [22, 16, 8, 16, 13, 11, 13],
        text_left={1}, money={5, 7}, ints={3},
        validations=[("B", "GroceryCatList"), ("D", "StoreList"), ("F", "YesNoList")],
        reserved=60)
    name_range(wb, "GrocCat", "Grocery", "B", start, end)
    name_range(wb, "GrocEst", "Grocery", "E", start, end)
    name_range(wb, "GrocActual", "Grocery", "G", start, end)
    name_range(wb, "GrocBought", "Grocery", "F", start, end)
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.cell(row=4, column=9, value="TOTALS").style = "section_gold"
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 12
    tot = [("Est. total", "=SUM(GrocEst)", '"$"#,##0'),
           ("Actual total", "=SUM(GrocActual)", '"$"#,##0'),
           ("Items bought", '=COUNTIF(GrocBought,"Yes")', "0"),
           ("Items left", '=COUNTIF(GrocBought,"No")', "0")]
    for i, (lab, fml, fmt) in enumerate(tot):
        r = 5 + i
        ws.cell(row=r, column=9, value=lab).style = "field_label"
        c = ws.cell(row=r, column=10, value=fml); c.style = "field_value"; c.number_format = fmt


# ===========================================================================
# 8 — Pantry Inventory
# ===========================================================================
def build_pantry(wb):
    sample = [
        ("Rice (5 lb)", "Grains & Pasta", 2, 1, dplus(300), "Pantry shelf 1", ""),
        ("Pasta", "Grains & Pasta", 1, 3, dplus(280), "Pantry shelf 1", "Low — reorder"),
        ("Canned tomatoes", "Canned Goods", 6, 4, dplus(400), "Pantry shelf 2", ""),
        ("Black beans", "Canned Goods", 2, 4, dplus(380), "Pantry shelf 2", "Low — reorder"),
        ("Flour", "Baking", 1, 1, dplus(120), "Baking bin", ""),
        ("Sugar", "Baking", 2, 1, dplus(200), "Baking bin", ""),
        ("Granola bars", "Snacks", 1, 2, dplus(90), "Snack drawer", "Low — reorder"),
        ("Cereal", "Breakfast", 3, 2, dplus(150), "Pantry shelf 3", ""),
        ("Olive oil", "Spices & Oils", 1, 1, dplus(220), "Counter", ""),
        ("Coffee", "Beverages", 2, 1, dplus(140), "Counter", ""),
        ("Peanut butter", "Other", 1, 1, dplus(160), "Pantry shelf 2", ""),
        ("Chicken broth", "Canned Goods", 5, 3, dplus(260), "Pantry shelf 2", ""),
    ]
    ws, start, end = build_log(
        wb, "Pantry", "🥫", "PANTRY INVENTORY",
        "Track stock & expiration — items at or below their reorder point flag in red.",
        ["Food Item", "Category", "Qty", "Reorder At", "Expires", "Location", "Notes"],
        sample, [22, 16, 8, 12, 13, 18, 22],
        text_left={1, 6, 7}, dates={5}, ints={3, 4},
        validations=[("B", "PantryCatList")], reserved=60)
    name_range(wb, "PantryName", "Pantry", "A", start, end)
    name_range(wb, "PantryCat", "Pantry", "B", start, end)
    name_range(wb, "PantryQty", "Pantry", "C", start, end)
    name_range(wb, "PantryReorder", "Pantry", "D", start, end)
    ws.conditional_formatting.add(
        f"A{start}:G{end}",
        FormulaRule(formula=[f'AND($A{start}<>"",$C{start}<=$D{start})'], fill=fill(RED_BG)))
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        FormulaRule(formula=[f'AND($E{start}<>"",$E{start}<=TODAY()+14)'], fill=fill(WARN_BG)))
    # category summary for donut
    ws.cell(row=4, column=9, value="BY CATEGORY").style = "section_gold"
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 10
    ws.cell(row=5, column=9, value="Category").style = "th"
    ws.cell(row=5, column=10, value="Items").style = "th"
    for i, cat in enumerate(PANTRY_CATS):
        r = 6 + i
        ws.cell(row=r, column=9, value=cat).style = "td_left"
        c = ws.cell(row=r, column=10, value=f'=COUNTIF(PantryCat,I{r})'); c.style = "td"
    wb.defined_names["PantryCatLabels"] = DefinedName("PantryCatLabels", attr_text=f"Pantry!$I$6:$I${5+len(PANTRY_CATS)}")
    wb.defined_names["PantryCatVals"] = DefinedName("PantryCatVals", attr_text=f"Pantry!$J$6:$J${5+len(PANTRY_CATS)}")


# ===========================================================================
# 9 — Refrigerator & Freezer Inventory
# ===========================================================================
def build_fridge(wb):
    build_log(
        wb, "Fridge & Freezer", "🧊", "REFRIGERATOR & FREEZER INVENTORY",
        "Know what's chilled or frozen — and use it before it's wasted.",
        ["Item", "Location", "Qty", "Purchased", "Use / Freeze By", "Planned Meal", "Notes"],
        [
            ("Leftover chili", "Fridge", 1, dminus(2), dplus(2), "Lunch", ""),
            ("Chicken breast", "Freezer", 4, dminus(5), dplus(60), "Tue dinner", ""),
            ("Ground beef", "Freezer", 2, dminus(5), dplus(60), "Tacos", ""),
            ("Milk", "Fridge", 2, dminus(1), dplus(7), "—", ""),
            ("Yogurt", "Fridge", 6, dminus(3), dplus(10), "Snacks", ""),
            ("Frozen berries", "Freezer", 2, dminus(20), dplus(180), "Smoothies", ""),
            ("Salmon fillets", "Freezer", 3, dminus(7), dplus(45), "Fri dinner", ""),
            ("Shredded cheese", "Fridge", 1, dminus(4), dplus(20), "Pizza", ""),
            ("Bread (backup)", "Freezer", 1, dminus(10), dplus(60), "Toast", ""),
        ],
        [22, 12, 8, 14, 16, 16, 20],
        text_left={1, 6, 7}, dates={4, 5}, ints={3},
        validations=[("B", "YesNoList")], reserved=50)
    ws = wb["Fridge & Freezer"]
    ws.conditional_formatting.add(
        "E5:E54",
        FormulaRule(formula=['AND($E5<>"",$E5<=TODAY()+3)'], fill=fill(WARN_BG)))


# ===========================================================================
# 10 — Meal Planner
# ===========================================================================
def build_meal_planner(wb):
    ws = wb.create_sheet("Meal Planner")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [14] + [18] * 7 + [3, 22, 12])
    luxe_header(ws, "K", "🍽  MEAL PLANNER",
                "Plan the week's menu — completion feeds your dashboard, leftovers feed your grocery list.")
    merge_set(ws, "A4:H4", "WEEKLY MENU", "section_gold")
    ws.cell(row=5, column=1, value="Meal").style = "th"
    for i, day in enumerate(DAYS):
        ws.cell(row=5, column=2 + i, value=day).style = "th"
    menu = {
        "Breakfast": ["Oatmeal", "Eggs & toast", "Smoothies", "Pancakes", "Yogurt bowls", "Waffles", "Bagels"],
        "Lunch": ["Turkey wraps", "Leftover chili", "Pasta salad", "Grilled cheese", "Bento boxes", "Soup", "Tacos"],
        "Dinner": ["Sheet-pan chicken", "Taco night", "Spaghetti", "Salmon & rice", "Homemade pizza", "Stir-fry", "Pot roast"],
        "Snack": ["Apples", "Granola bars", "Cheese & crackers", "Veggies & dip", "Popcorn", "Fruit", "Trail mix"],
        "Dessert": ["—", "Ice cream", "—", "Cookies", "Movie treats", "Brownies", "—"],
    }
    for ri, meal in enumerate(MEAL_TYPES):
        r = 6 + ri
        ws.cell(row=r, column=1, value=meal).style = "field_label"
        for ci, val in enumerate(menu[meal]):
            c = ws.cell(row=r, column=2 + ci, value=val); c.style = "td_left"
            c.fill = fill(MUTED_ROW if ri % 2 else BG)
        ws.row_dimensions[r].height = 22
    wb.defined_names["MealWeek"] = DefinedName("MealWeek", attr_text="'Meal Planner'!$B$6:$H$10")

    merge_set(ws, "J4:K4", "PLANNING", "section_gold")
    ws.cell(row=5, column=10, value="Meals planned").style = "field_label"
    ws.cell(row=5, column=11, value='=COUNTIF(MealWeek,"<>"&"")-COUNTIF(MealWeek,"—")').style = "field_value"
    ws.cell(row=6, column=10, value="Plan complete").style = "field_label"
    c = ws.cell(row=6, column=11, value='=IFERROR((COUNTIF(MealWeek,"<>"&"")-COUNTIF(MealWeek,"—"))/35,0)')
    c.style = "field_value"; c.number_format = "0%"

    merge_set(ws, "A12:H12", "FAMILY FAVORITES", "section")
    favs = ["Taco night", "Homemade pizza", "Sheet-pan chicken", "Spaghetti & meatballs",
            "Pancake breakfast-for-dinner", "Grandma's pot roast"]
    for i, fav in enumerate(favs):
        r = 13 + i // 2
        col = 1 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value="★").font = Font(color=GOLD_LT, bold=True)
        ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=col + 1, value=fav); c.style = "td_left"
        ws.merge_cells(f"{get_column_letter(col+1)}{r}:{get_column_letter(col+3)}{r}")


# ===========================================================================
# 11 — Recipe Library
# ===========================================================================
def build_recipes(wb):
    build_log(
        wb, "Recipes", "📖", "RECIPE LIBRARY",
        "Your family cookbook — searchable, rated, and ready for meal planning.",
        ["Recipe", "Category", "Prep", "Cook", "Serves", "Rating", "Favorite", "Link / Source", "Notes"],
        [
            ("Sheet-Pan Chicken", "Dinner", "15 min", "35 min", 5, 5, "Yes", "Family recipe", "Kid-approved"),
            ("Taco Night", "Dinner", "20 min", "15 min", 5, 5, "Yes", "—", "Build-your-own"),
            ("Homemade Pizza", "Dinner", "30 min", "12 min", 6, 5, "Yes", "—", "Friday tradition"),
            ("Banana Pancakes", "Breakfast", "10 min", "15 min", 4, 4, "No", "—", ""),
            ("Veggie Stir-Fry", "Dinner", "15 min", "12 min", 4, 4, "No", "—", "Uses leftovers"),
            ("Pot Roast", "Dinner", "20 min", "4 hr", 6, 5, "Yes", "Grandma", "Sunday dinner"),
            ("Overnight Oats", "Breakfast", "5 min", "—", 2, 4, "No", "—", "Make ahead"),
            ("Chocolate Chip Cookies", "Dessert", "15 min", "11 min", 24, 5, "Yes", "—", ""),
        ],
        [22, 14, 9, 9, 8, 9, 11, 18, 20],
        text_left={1, 8, 9}, ints={5, 6},
        validations=[("B", "MealTypeList"), ("G", "YesNoList")], reserved=50)
    ws = wb["Recipes"]
    ws.conditional_formatting.add(
        "G5:G54", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 12 — Cleaning Command Center
# ===========================================================================
def build_cleaning(wb):
    sample = [
        ("Make beds", "Bedrooms", "Daily", "All", "Done", dminus(0)),
        ("Kitchen reset", "Kitchen", "Daily", "Mom", "Done", dminus(0)),
        ("Wipe counters", "Kitchen", "Daily", "Dad", "Done", dminus(0)),
        ("Quick tidy living room", "Living Room", "Daily", "Kids", "In Progress", dminus(1)),
        ("Vacuum main floor", "Whole Home", "Weekly", "Dad", "Done", dminus(3)),
        ("Bathrooms deep wipe", "Bathrooms", "Weekly", "Mom", "Not Started", dminus(8)),
        ("Mop floors", "Kitchen", "Weekly", "Dad", "Not Started", dminus(9)),
        ("Change bed sheets", "Bedrooms", "Weekly", "All", "Done", dminus(2)),
        ("Dust surfaces", "Whole Home", "Weekly", "Emma", "Not Started", dminus(10)),
        ("Clean fridge", "Kitchen", "Monthly", "Mom", "Not Started", dminus(20)),
        ("Wipe baseboards", "Whole Home", "Monthly", "Dad", "Not Started", dminus(25)),
        ("Wash windows", "Whole Home", "Seasonal", "Dad", "Not Started", dminus(80)),
        ("Clean gutters", "Exterior", "Seasonal", "Dad", "Not Started", dminus(95)),
        ("Carpet shampoo", "Whole Home", "Deep Clean", "Hired", "Not Started", dminus(180)),
        ("Garage clear-out", "Garage", "Deep Clean", "All", "Not Started", dminus(160)),
    ]
    ws, start, end = build_log(
        wb, "Cleaning", "🧹", "CLEANING COMMAND CENTER",
        "Daily, weekly, monthly, seasonal & deep cleaning — completion auto-calculates.",
        ["Task", "Area / Room", "Frequency", "Assigned To", "Status", "Last Done"],
        sample, [26, 16, 14, 14, 14, 14],
        text_left={1, 2}, dates={6},
        validations=[("C", "CleanFreqList"), ("D", "FamilyList"), ("E", "StatusList")],
        reserved=50)
    name_range(wb, "CleanTask", "Cleaning", "A", start, end)
    name_range(wb, "CleanFreq", "Cleaning", "C", start, end)
    name_range(wb, "CleanStatus", "Cleaning", "E", start, end)
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Done"'], fill=fill(MINT_BG)))
    # frequency summary
    ws.cell(row=4, column=8, value="COMPLETION").style = "section_gold"
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 9
    ws.column_dimensions["J"].width = 10
    ws.cell(row=5, column=8, value="Frequency").style = "th"
    ws.cell(row=5, column=9, value="Done").style = "th"
    ws.cell(row=5, column=10, value="%").style = "th"
    for i, fr in enumerate(CLEAN_FREQ):
        r = 6 + i
        ws.cell(row=r, column=8, value=fr).style = "td_left"
        ws.cell(row=r, column=9, value=f'=COUNTIFS(CleanFreq,H{r},CleanStatus,"Done")').style = "td"
        c = ws.cell(row=r, column=10, value=f'=IFERROR(COUNTIFS(CleanFreq,H{r},CleanStatus,"Done")/MAX(COUNTIF(CleanFreq,H{r}),1),0)')
        c.style = "td"; c.number_format = "0%"
    wb.defined_names["CleanFreqLabels"] = DefinedName("CleanFreqLabels", attr_text=f"Cleaning!$H$6:$H${5+len(CLEAN_FREQ)}")
    wb.defined_names["CleanFreqPct"] = DefinedName("CleanFreqPct", attr_text=f"Cleaning!$J$6:$J${5+len(CLEAN_FREQ)}")


# ===========================================================================
# 13 — Chore Manager
# ===========================================================================
def build_chores(wb):
    sample = [
        ("Set / clear table", "Emma", "Daily", "In Progress", "Screen time", 0.7),
        ("Feed the dog", "Liam", "Daily", "Done", "Sticker", 1.0),
        ("Take out trash", "Dad", "Weekly", "Done", "—", 1.0),
        ("Tidy bedroom", "Emma", "Daily", "Done", "Allowance", 0.85),
        ("Load dishwasher", "Mom", "Daily", "Done", "—", 1.0),
        ("Water plants", "Liam", "Weekly", "Not Started", "Sticker", 0.4),
        ("Fold laundry", "Emma", "Weekly", "In Progress", "Allowance", 0.6),
        ("Wipe bathroom sink", "Liam", "Weekly", "Not Started", "Screen time", 0.3),
        ("Vacuum room", "Emma", "Weekly", "Done", "Allowance", 0.9),
        ("Help with groceries", "Dad", "Weekly", "Done", "—", 1.0),
    ]
    ws, start, end = build_log(
        wb, "Chores", "🧽", "CHORE MANAGER",
        "Assign chores by family member — completion rolls into your Family Routine Score.",
        ["Chore", "Assigned To", "Frequency", "Status", "Reward", "Completion %"],
        sample, [24, 16, 14, 14, 16, 14],
        text_left={1, 5}, pcts={6},
        validations=[("B", "FamilyList"), ("C", "ChoreFreqList"), ("D", "StatusList")],
        reserved=40)
    name_range(wb, "ChoreAssigned", "Chores", "B", start, end)
    name_range(wb, "ChoreStatus", "Chores", "D", start, end)
    name_range(wb, "ChorePct", "Chores", "F", start, end)
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Done"'], fill=fill(MINT_BG)))
    # by-member summary
    ws.cell(row=4, column=8, value="BY MEMBER").style = "section_gold"
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10
    ws.cell(row=5, column=8, value="Member").style = "th"
    ws.cell(row=5, column=9, value="Avg %").style = "th"
    for i, m in enumerate(FAMILY):
        r = 6 + i
        ws.cell(row=r, column=8, value=m).style = "td_left"
        c = ws.cell(row=r, column=9, value=f'=IFERROR(AVERAGEIFS(ChorePct,ChoreAssigned,H{r}),0)')
        c.style = "td"; c.number_format = "0%"
    wb.defined_names["ChoreMembers"] = DefinedName("ChoreMembers", attr_text=f"Chores!$H$6:$H${5+len(FAMILY)}")
    wb.defined_names["ChoreMemberPct"] = DefinedName("ChoreMemberPct", attr_text=f"Chores!$I$6:$I${5+len(FAMILY)}")


# ===========================================================================
# 14 — Home Maintenance
# ===========================================================================
def build_maintenance(wb):
    sample = [
        ("HVAC service", "HVAC", "Twice a year", dminus(120), dplus(60), 120, "Spring & fall"),
        ("Replace HVAC filter", "HVAC", "Monthly", dminus(5), dplus(25), 18, "20x25x1"),
        ("Test smoke detectors", "Safety", "Monthly", dminus(10), dplus(20), 0, "Replace batteries 2x/yr"),
        ("Flush water heater", "Plumbing", "Yearly", dminus(300), dplus(65), 0, ""),
        ("Roof inspection", "Exterior", "Yearly", dminus(280), dplus(85), 0, "After winter"),
        ("Gutter cleaning", "Exterior", "Twice a year", dminus(100), dplus(80), 90, ""),
        ("Lawn fertilize", "Lawn & Garden", "Seasonal", dminus(40), dplus(50), 45, ""),
        ("Pest control", "Pest Control", "Quarterly", dminus(70), dplus(20), 95, "Service plan"),
        ("Dryer vent clean", "Appliances", "Yearly", dminus(200), dplus(5), 0, "Overdue soon"),
        ("Car oil change", "Vehicle", "Every 5k mi", dminus(80), dplus(15), 60, "Both cars"),
        ("Garbage disposal", "Plumbing", "As needed", dminus(150), "", 0, ""),
        ("Caulk bathrooms", "Plumbing", "Yearly", dminus(330), dplus(35), 15, ""),
    ]
    ws, start, end = build_log(
        wb, "Maintenance", "🛠", "HOME MAINTENANCE",
        "Every system on a schedule — anything due in the next 30 days flags automatically.",
        ["Task", "Category", "Frequency", "Last Done", "Next Due", "Est. Cost", "Notes"],
        sample, [22, 16, 14, 14, 14, 12, 22],
        text_left={1, 3, 7}, dates={4, 5}, money={6},
        validations=[("B", "MaintCatList")], reserved=50)
    name_range(wb, "MaintTask", "Maintenance", "A", start, end)
    name_range(wb, "MaintNext", "Maintenance", "E", start, end)
    ws.conditional_formatting.add(
        f"A{start}:G{end}",
        FormulaRule(formula=[f'AND($E{start}<>"",$E{start}<=TODAY()+30,$E{start}>=TODAY())'],
                    fill=fill(WARN_BG)))
    ws.conditional_formatting.add(
        f"A{start}:G{end}",
        FormulaRule(formula=[f'AND($E{start}<>"",$E{start}<TODAY())'], fill=fill(RED_BG)))


# ===========================================================================
# 15 — Home Inventory
# ===========================================================================
def build_home_inventory(wb):
    build_log(
        wb, "Home Inventory", "📦", "HOME INVENTORY",
        "Document valuables for insurance — purchase, warranty, serial & replacement value.",
        ["Item", "Category", "Room", "Purchase Price", "Purchase Date",
         "Warranty Until", "Serial #", "Insured Value"],
        [
            ("65\" Smart TV", "Electronics", "Living Room", 1200, dminus(400), dplus(330), "SN-TV-88210", 1200),
            ("Refrigerator", "Appliances", "Kitchen", 2100, dminus(900), dplus(190), "SN-RF-44120", 1800),
            ("Washer", "Appliances", "Laundry", 850, dminus(700), dplus(60), "SN-WS-90021", 700),
            ("Dryer", "Appliances", "Laundry", 800, dminus(700), dplus(60), "SN-DR-90022", 650),
            ("Sectional sofa", "Furniture", "Living Room", 1800, dminus(500), "", "—", 1500),
            ("Dining set", "Furniture", "Dining Room", 1400, dminus(1100), "", "—", 1000),
            ("MacBook", "Electronics", "Office", 1600, dminus(300), dplus(430), "SN-MB-77410", 1500),
            ("Wedding ring", "Jewelry", "Safe", 4500, dminus(3000), "", "Appraised", 5200),
            ("Lawn mower", "Tools", "Garage", 450, dminus(600), dplus(120), "SN-LM-22019", 350),
            ("Camera kit", "Electronics", "Office", 950, dminus(800), "", "SN-CM-31002", 800),
        ],
        [22, 14, 14, 14, 14, 14, 16, 14],
        text_left={1, 7}, money={4, 8}, dates={5, 6}, reserved=50)


# ===========================================================================
# 16 — Subscription Tracker
# ===========================================================================
def build_subscriptions(wb):
    sample = [
        ("Netflix", "Streaming", "Monthly", 18, 216, dplus(12), "Active"),
        ("Disney+ Bundle", "Streaming", "Monthly", 20, 240, dplus(20), "Active"),
        ("Spotify Family", "Streaming", "Monthly", 17, 204, dplus(8), "Active"),
        ("Amazon Prime", "Memberships", "Yearly", 12, 139, dplus(150), "Active"),
        ("Microsoft 365", "Software", "Yearly", 8, 99, dplus(200), "Active"),
        ("Gym Family", "Gym", "Monthly", 45, 540, dplus(16), "Active"),
        ("Kids Art Class", "Kids Activities", "Monthly", 60, 720, dplus(9), "Active"),
        ("Cloud Storage", "Software", "Monthly", 10, 120, dplus(5), "Review"),
        ("Meal Kit", "Memberships", "Monthly", 0, 0, "", "Paused"),
        ("News Subscription", "Memberships", "Monthly", 15, 180, dplus(22), "Consider cancel"),
    ]
    ws, start, end = build_log(
        wb, "Subscriptions", "🔁", "SUBSCRIPTION TRACKER",
        "Find money leaks — every recurring charge, monthly & annual cost, renewal dates.",
        ["Service", "Type", "Billing", "Monthly", "Annual", "Renews", "Status"],
        sample, [22, 16, 12, 12, 12, 14, 18],
        text_left={1, 7}, money={4, 5}, dates={6}, reserved=40)
    name_range(wb, "SubMonthly", "Subscriptions", "D", start, end)
    name_range(wb, "SubAnnual", "Subscriptions", "E", start, end)
    ws.cell(row=4, column=9, value="TOTALS").style = "section_gold"
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 12
    ws.cell(row=5, column=9, value="Monthly total").style = "field_label"
    c = ws.cell(row=5, column=10, value="=SUM(SubMonthly)"); c.style = "field_value"; c.number_format = '"$"#,##0'
    ws.cell(row=6, column=9, value="Annual total").style = "field_label"
    c = ws.cell(row=6, column=10, value="=SUM(SubAnnual)"); c.style = "field_value"; c.number_format = '"$"#,##0'


# ===========================================================================
# 17 — Family Goals
# ===========================================================================
def build_goals(wb):
    sample = [
        ("Build 6-month emergency fund", "Financial", dplus(300), 0.55, "In Progress", ""),
        ("Pay off credit card", "Financial", dplus(150), 0.4, "In Progress", ""),
        ("Renovate the kitchen", "Home", dplus(365), 0.2, "In Progress", "See Projects"),
        ("Family trip to the coast", "Travel", dplus(120), 0.6, "In Progress", "See Travel"),
        ("Emma reads 20 books", "Education", dplus(180), 0.45, "In Progress", "Reading log"),
        ("Family 5k in spring", "Health", dplus(90), 0.35, "In Progress", ""),
        ("Declutter whole house", "Home", dplus(60), 0.5, "In Progress", ""),
        ("Date night 2x / month", "Personal", dplus(30), 0.7, "In Progress", ""),
    ]
    ws, start, end = build_log(
        wb, "Family Goals", "🎯", "FAMILY GOALS",
        "Where your family is headed — milestones with automatic progress bars.",
        ["Goal", "Category", "Target Date", "Progress", "Status", "Notes"],
        sample, [32, 16, 14, 12, 14, 24],
        text_left={1, 6}, dates={3}, pcts={4},
        validations=[("B", "GoalCatList"), ("E", "StatusList")], reserved=40)
    name_range(wb, "GoalName", "Family Goals", "A", start, end)
    name_range(wb, "GoalProgress", "Family Goals", "D", start, end)
    name_range(wb, "GoalStatus", "Family Goals", "E", start, end)
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))


# ===========================================================================
# 18 — Savings Planner
# ===========================================================================
def build_savings(wb):
    ws = wb.create_sheet("Savings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 14, 16, 14, 12, 22])
    luxe_header(ws, "G", "🏦  SAVINGS PLANNER",
                "Every savings goal in one view — funded progress updates automatically.")
    table_headers(ws, 4, ["Fund", "Target", "Saved", "Monthly Add", "Target Date", "Funded", "Notes"])
    funds = [
        ("Emergency Fund", 18000, 9900, 600, dplus(300), "6-month buffer"),
        ("Vacation", 4000, 2400, 250, dplus(120), "Coast trip"),
        ("Christmas", 1500, 600, 150, dplus(150), "Gifts + travel"),
        ("Home Renovation", 25000, 5000, 500, dplus(365), "Kitchen"),
        ("New Vehicle", 12000, 3000, 300, dplus(540), "Down payment"),
        ("Education / 529", 30000, 8200, 200, dplus(2000), "Kids college"),
    ]
    start = L0
    end = start + len(funds) - 1
    for i, (fund, tgt, cur, add, tdate, note) in enumerate(funds):
        r = start + i
        ws.cell(row=r, column=1, value=fund).style = "td_left"
        c = ws.cell(row=r, column=2, value=tgt); c.style = "input"; c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=3, value=cur); c.style = "input"; c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=4, value=add); c.style = "td"; c.number_format = '"$"#,##0'
        c = ws.cell(row=r, column=5, value=tdate); c.style = "td"; c.number_format = "mm/dd/yyyy"
        c = ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/B{r},0)"); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=7, value=note).style = "td_left"
        if i % 2:
            for cc in range(1, 8):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in (2, 3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    c = ws.cell(row=total, column=6, value=f"=IFERROR(C{total}/B{total},0)")
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0%"
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    name_range(wb, "SavName", "Savings", "A", start, end)
    name_range(wb, "SavTarget", "Savings", "B", start, end)
    name_range(wb, "SavCurrent", "Savings", "C", start, end)
    name_range(wb, "SavFunded", "Savings", "F", start, end)
    bar = BarChart(); bar.type = "bar"; bar.title = "Savings Progress"; bar.height = 8.5; bar.width = 13
    bar.add_data(Reference(ws, min_col=6, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    bar.legend = None
    ws.add_chart(bar, "A15")
    ws.freeze_panes = "A5"


# ===========================================================================
# 19 — Holiday & Event Planner
# ===========================================================================
def build_holiday(wb):
    build_log(
        wb, "Holidays & Events", "🎁", "HOLIDAY & EVENT PLANNER",
        "Plan every celebration without the stress — gifts, budgets, menus & to-dos.",
        ["Occasion", "Date", "Budget", "Spent", "For / Guests", "Gift / Plan", "Status", "Notes"],
        [
            ("Grandma's Birthday", dplus(6), 120, 0, "Grace", "Cake + spa gift card", "Planning", ""),
            ("Emma's Birthday", dplus(75), 300, 40, "Emma + 8 friends", "Art party", "Planning", "Book venue"),
            ("Thanksgiving", dplus(150), 250, 0, "12 guests", "Host dinner", "Not Started", "Menu in Meals"),
            ("Christmas", dplus(178), 1200, 200, "Family", "Gift list below", "Shopping", "Save in Savings"),
            ("Liam's Birthday", dplus(220), 250, 0, "Liam + class", "Dinosaur theme", "Not Started", ""),
            ("Easter", dplus(280), 80, 0, "Family", "Baskets + brunch", "Not Started", ""),
            ("Halloween", dplus(120), 150, 0, "Kids", "Costumes + candy", "Not Started", ""),
            ("Anniversary", dplus(95), 200, 0, "Mom & Dad", "Dinner out", "Planning", ""),
        ],
        [22, 14, 12, 12, 18, 22, 14, 18],
        text_left={1, 5, 6, 8}, dates={2}, money={3, 4},
        validations=[("G", "StatusList")], reserved=40)


# ===========================================================================
# 20 — Travel Planner
# ===========================================================================
def build_travel(wb):
    ws = wb.create_sheet("Travel")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 16, 16, 14, 14, 3, 22, 14])
    luxe_header(ws, "H", "✈  TRAVEL PLANNER",
                "From packing lists to itineraries — plan the trip and stay on budget.")

    merge_set(ws, "A4:E4", "TRIP DETAILS & RESERVATIONS", "section_gold")
    table_headers(ws, 5, ["Item", "Type", "Details / Confirmation", "Date", "Cost"])
    res = [
        ("Lakeside Resort", "Hotel", "Conf #LK-88210, 2 nights", dplus(20), 420),
        ("Flights (x4)", "Flight", "AA 1180 / AA 1181", dplus(20), 960),
        ("Rental car", "Car", "Conf #RC-4421, SUV", dplus(20), 180),
        ("Dinner reservation", "Dining", "Lakeview Grill, 7pm", dplus(21), 0),
        ("Museum tickets", "Activity", "Family 4-pack", dplus(22), 64),
        ("Travel insurance", "Insurance", "Policy TRV-9981", dplus(15), 85),
    ]
    style_rows(ws, 6, 13, 5, text_left={1, 3}, dates={4}, money={5})
    for i, row in enumerate(res):
        for ci, val in enumerate(row, 1):
            ws.cell(row=6 + i, column=ci, value=val)
    ws.cell(row=14, column=1, value="TOTAL").style = "th"
    ws.cell(row=14, column=4, value="Trip cost").style = "field_label"
    c = ws.cell(row=14, column=5, value="=SUM(E6:E13)"); c.style = "field_value"; c.number_format = '"$"#,##0'

    merge_set(ws, "G4:H4", "PACKING LIST", "section_gold")
    pack = ["Clothes (per day)", "Toiletries", "Chargers & cords", "Medications",
            "Swimsuits", "Snacks & water", "Kids' activities", "Documents & IDs",
            "Sunscreen", "First-aid kit", "Camera", "Phone + wallet"]
    for i, it in enumerate(pack):
        r = 5 + i
        ws.cell(row=r, column=7, value=it).style = "td_left"
        c = ws.cell(row=r, column=8, value="☐"); c.alignment = Alignment(horizontal="center")
        c.font = Font(size=12, color=ACCENT); c.border = BOX
        if i % 2:
            ws.cell(row=r, column=7).fill = fill(MUTED_ROW)
            ws.cell(row=r, column=8).fill = fill(MUTED_ROW)

    merge_set(ws, "A16:E16", "DAILY ITINERARY", "section")
    table_headers(ws, 17, ["Day", "Morning", "Afternoon", "Evening", "Notes"])
    itin = [
        ("Day 1", "Drive / fly in", "Check in + pool", "Welcome dinner", ""),
        ("Day 2", "Lake activities", "Museum", "Lakeview Grill", ""),
        ("Day 3", "Hike trail", "Town shopping", "Movie night", ""),
        ("Day 4", "Pack up", "Travel home", "Unpack + rest", ""),
    ]
    style_rows(ws, 18, 21, 5, text_left={2, 3, 4, 5})
    for i, row in enumerate(itin):
        for ci, val in enumerate(row, 1):
            ws.cell(row=18 + i, column=ci, value=val)


# ===========================================================================
# 21 — Children's Hub
# ===========================================================================
def build_children(wb):
    ws = wb.create_sheet("Children's Hub")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 18, 16, 16, 18, 16, 22])
    luxe_header(ws, "G", "🎒  CHILDREN'S HUB",
                "School, activities, health & milestones — everything about the kids in one place.")

    merge_set(ws, "A4:G4", "SCHOOL & ACTIVITIES", "section_gold")
    table_headers(ws, 5, ["Child", "School / Grade", "Teacher", "Activity", "Day / Time", "Coach / Lead", "Notes"])
    sch = [
        ("Emma", "Maple Elem. — Gr 4", "Ms. Rivera", "Soccer", "Tue/Thu 4pm", "Coach Dan", "Room 12"),
        ("Emma", "Maple Elem. — Gr 4", "Ms. Rivera", "Piano", "Sat 10am", "Ms. Kim", ""),
        ("Liam", "Maple Elem. — Gr 1", "Mr. Olsen", "Swim", "Wed 5pm", "Coach Amy", "Room 4"),
        ("Liam", "Maple Elem. — Gr 1", "Mr. Olsen", "Art Club", "Fri 3:30pm", "Ms. Lee", ""),
    ]
    style_rows(ws, 6, 11, 7, text_left={2, 3, 4, 6, 7})
    for i, row in enumerate(sch):
        for ci, val in enumerate(row, 1):
            ws.cell(row=6 + i, column=ci, value=val)
    add_dv(ws, "A6:A11", "FamilyList")

    merge_set(ws, "A13:G13", "HOMEWORK & READING LOG", "section_gold")
    table_headers(ws, 14, ["Child", "Subject / Book", "Assigned", "Due", "Minutes", "Done?", "Notes"])
    hw = [
        ("Emma", "Math worksheet", dminus(1), dplus(1), 20, "No", ""),
        ("Emma", "Reading: 'Wonder'", dminus(3), dplus(4), 100, "No", "Ch 5-8"),
        ("Liam", "Sight words", dminus(1), dplus(1), 15, "Yes", ""),
        ("Liam", "Reading: picture book", dminus(0), dplus(0), 10, "Yes", ""),
    ]
    style_rows(ws, 15, 20, 7, text_left={2, 7}, dates={3, 4}, ints={5},
               )
    for i, row in enumerate(hw):
        for ci, val in enumerate(row, 1):
            ws.cell(row=15 + i, column=ci, value=val)
    add_dv(ws, "A15:A20", "FamilyList")
    add_dv(ws, "F15:F20", "YesNoList")
    ws.conditional_formatting.add(
        "F15:F20", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))

    merge_set(ws, "A22:G22", "GROWTH & MILESTONES", "section")
    table_headers(ws, 23, ["Child", "Date", "Height", "Weight", "Milestone / Achievement", "Doctor Visit", "Notes"])
    gr = [
        ("Emma", dminus(60), "4'4\"", "62 lb", "Lost first molar", "Annual check ✓", ""),
        ("Liam", dminus(60), "3'9\"", "44 lb", "Read first chapter book", "Annual check ✓", ""),
    ]
    style_rows(ws, 24, 27, 7, text_left={5, 7})
    for i, row in enumerate(gr):
        for ci, val in enumerate(row, 1):
            ws.cell(row=24 + i, column=ci, value=val)


# ===========================================================================
# 22 — Pet Care Center
# ===========================================================================
def build_pets(wb):
    ws = wb.create_sheet("Pet Care")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 16, 16, 16, 16, 16, 20])
    luxe_header(ws, "G", "🐾  PET CARE CENTER",
                "Food, vet visits, vaccinations & routines — happy, healthy pets on schedule.")

    merge_set(ws, "A4:G4", "PETS", "section_gold")
    table_headers(ws, 5, ["Pet", "Type / Breed", "Birthday", "Vet", "Microchip #", "Food", "Notes"])
    pets = [
        ("Buddy", "Dog / Labrador", dt.date(2019, 5, 2), "Paws & Claws", "MC-77120", "2 cups 2x/day", "Friendly"),
        ("Whiskers", "Cat / Tabby", dt.date(2021, 8, 14), "Paws & Claws", "MC-77121", "1/2 cup 2x/day", "Indoor"),
    ]
    style_rows(ws, 6, 8, 7, text_left={2, 6, 7}, dates={3})
    for i, row in enumerate(pets):
        for ci, val in enumerate(row, 1):
            ws.cell(row=6 + i, column=ci, value=val)

    merge_set(ws, "A10:G10", "HEALTH & VET SCHEDULE", "section_gold")
    table_headers(ws, 11, ["Pet", "Service", "Last Done", "Next Due", "Provider", "Cost", "Notes"])
    health = [
        ("Buddy", "Rabies vaccine", dminus(200), dplus(165), "Paws & Claws", 35, ""),
        ("Buddy", "Heartworm meds", dminus(20), dplus(10), "Auto-ship", 18, "Monthly"),
        ("Buddy", "Grooming", dminus(40), dplus(20), "Pampered Paws", 60, ""),
        ("Whiskers", "Annual checkup", dminus(150), dplus(215), "Paws & Claws", 75, ""),
        ("Whiskers", "Flea treatment", dminus(25), dplus(5), "Auto-ship", 15, "Monthly"),
        ("Buddy", "Dental cleaning", dminus(300), dplus(60), "Paws & Claws", 220, ""),
    ]
    style_rows(ws, 12, 19, 7, text_left={7}, dates={3, 4}, money={6})
    for i, row in enumerate(health):
        for ci, val in enumerate(row, 1):
            ws.cell(row=12 + i, column=ci, value=val)
    ws.conditional_formatting.add(
        "A12:G19", FormulaRule(formula=['AND($D12<>"",$D12<=TODAY()+14)'], fill=fill(WARN_BG)))

    merge_set(ws, "A21:G21", "WALKING / CARE SCHEDULE", "section")
    table_headers(ws, 22, ["Day", "Morning", "Midday", "Evening", "Walk?", "Fed?", "Notes"])
    for i, day in enumerate(DAYS):
        r = 23 + i
        ws.cell(row=r, column=1, value=day).style = "td_left"
        for c in (2, 3, 4, 5, 6, 7):
            cell = ws.cell(row=r, column=c, value="")
            cell.style = "td"
            cell.fill = fill(MUTED_ROW if i % 2 else BG)


# ===========================================================================
# 23 — Wellness & Self-Care
# ===========================================================================
def build_wellness(wb):
    ws = wb.create_sheet("Wellness")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [12, 10, 10, 10, 10, 10, 10, 12, 4, 22, 12])
    luxe_header(ws, "K", "🌿  WELLNESS & SELF-CARE",
                "Don't run on empty — track sleep, movement, water & the habits that recharge you.")
    merge_set(ws, "A4:H4", "WEEKLY HABIT TRACKER", "section_gold")
    headers = ["Day", "Sleep (hr)", "Water (oz)", "Exercise", "Read", "Meditate", "Outside", "Self-Care ★"]
    table_headers(ws, 5, headers)
    today = dt.date.today()
    for i, day in enumerate(DAYS):
        r = 6 + i
        ws.cell(row=r, column=1, value=day).style = "td_left"
        vals = [[7.5, 64, "✓", "✓", "✓", "✓", "✓"], [6.5, 48, "✓", "", "", "✓", ""],
                [8, 72, "✓", "✓", "✓", "", "✓"], [7, 56, "", "✓", "✓", "✓", ""],
                [7.5, 64, "✓", "", "", "", "✓"], [8.5, 80, "✓", "✓", "✓", "✓", "✓"],
                [8, 60, "", "✓", "", "✓", ""]][i]
        for ci, v in enumerate(vals, 2):
            c = ws.cell(row=r, column=ci, value=v); c.style = "td"
            c.fill = fill(MUTED_ROW if i % 2 else BG)
    name_range(wb, "WellSleep", "Wellness", "B", 6, 12)
    name_range(wb, "WellWater", "Wellness", "C", 6, 12)

    merge_set(ws, "J4:K4", "WEEK AT A GLANCE", "section_gold")
    summ = [("Avg sleep", "=IFERROR(AVERAGE(WellSleep),0)", "0.0"),
            ("Avg water (oz)", "=IFERROR(AVERAGE(WellWater),0)", "0"),
            ("Exercise days", '=COUNTIF(E6:E12,"✓")', "0"),
            ("Self-care days", '=COUNTIF(H6:H12,"✓")', "0")]
    for i, (lab, fml, fmt) in enumerate(summ):
        r = 5 + i
        ws.cell(row=r, column=10, value=lab).style = "field_label"
        c = ws.cell(row=r, column=11, value=fml); c.style = "field_value"; c.number_format = fmt

    merge_set(ws, "A14:H14", "SELF-CARE MENU & PERSONAL GOALS", "section")
    ideas = ["10-min morning quiet", "Walk outside daily", "Read before bed", "Call a friend",
             "Bath / spa night", "Hobby time (1 hr)", "Stretch / yoga", "Digital sunset 9pm",
             "Journal 3 gratitudes", "Meal-prep Sunday", "Date night", "Say no to one thing"]
    for i, it in enumerate(ideas):
        r = 15 + i // 2
        col = 1 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value="○").alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=col).font = Font(color=HIGHLIGHT, bold=True)
        c = ws.cell(row=r, column=col + 1, value=it); c.style = "td_left"
        ws.merge_cells(f"{get_column_letter(col+1)}{r}:{get_column_letter(col+3)}{r}")


# ===========================================================================
# 24 — Home Project Planner (with image placeholders)
# ===========================================================================
def build_projects(wb):
    ws = wb.create_sheet("Projects")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 14, 14, 16, 14, 14, 13, 22])
    luxe_header(ws, "H", "🏗  HOME PROJECT PLANNER",
                "Plan renovations & DIY — budget, contractors, timeline & before/after photos.")
    table_headers(ws, 4, ["Project / Room", "Budget", "Spent", "Contractor",
                          "Start", "Target", "% Done", "Status"])
    projs = [
        ("Kitchen remodel", 25000, 5200, "Stellar Kitchens", dminus(20), dplus(160), 0.2, "In Progress"),
        ("Master bath refresh", 6000, 0, "TBD", "", dplus(240), 0.0, "Not Started"),
        ("Backyard deck", 8500, 1200, "Outdoor Pros", dminus(5), dplus(90), 0.15, "In Progress"),
        ("Garage organization", 1200, 800, "DIY", dminus(40), dplus(10), 0.7, "In Progress"),
        ("Repaint living room", 600, 0, "DIY", "", dplus(45), 0.0, "Not Started"),
        ("Kids' playroom", 2000, 350, "DIY", dminus(15), dplus(60), 0.25, "In Progress"),
    ]
    start, end = L0, L0 + 11
    style_rows(ws, start, end, 8, text_left={1, 4}, money={2, 3}, dates={5, 6}, pcts={7})
    for i, row in enumerate(projs):
        for ci, val in enumerate(row, 1):
            ws.cell(row=start + i, column=ci, value=val)
    add_dv(ws, f"H{start}:H{end}", "StatusList")
    ws.conditional_formatting.add(
        f"G{start}:G{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    name_range(wb, "ProjName", "Projects", "A", start, end)
    name_range(wb, "ProjStatus", "Projects", "H", start, end)
    name_range(wb, "ProjBudget", "Projects", "B", start, end)
    name_range(wb, "ProjSpent", "Projects", "C", start, end)

    # before/after image placeholders for the active project
    merge_set(ws, f"A{end+2}:H{end+2}", "FEATURED PROJECT — BEFORE / AFTER", "section_gold")
    br = end + 3
    merge_set(ws, f"A{br}:D{br}", "BEFORE", "th")
    merge_set(ws, f"E{br}:H{br}", "AFTER", "th")
    merge_set(ws, f"A{br+1}:D{br+6}", "📷\nPaste your BEFORE photo here\n(Insert ▸ Picture)", "imgbox")
    merge_set(ws, f"E{br+1}:H{br+6}", "📷\nPaste your AFTER photo here\n(Insert ▸ Picture)", "imgbox")
    for r in range(br + 1, br + 7):
        ws.row_dimensions[r].height = 26


# ===========================================================================
# 25 — Document Vault Index
# ===========================================================================
def build_documents(wb):
    build_log(
        wb, "Documents", "🗄", "DOCUMENT VAULT INDEX",
        "Know where every important document lives — physical & digital, in one index.",
        ["Document", "Category", "Physical Location", "Digital Copy", "Expires / Renews", "Owner", "Notes"],
        [
            ("Home insurance policy", "Insurance", "File cabinet A", "Yes", dplus(120), "Household", ""),
            ("Mortgage documents", "Mortgage", "File cabinet A", "Yes", "", "Household", ""),
            ("Vehicle titles", "Vehicle", "Safe", "No", "", "Household", "2 cars"),
            ("Passports", "Identity", "Safe", "Yes", dplus(800), "All", "Renew Emma's"),
            ("Birth certificates", "Identity", "Safe", "Yes", "", "All", ""),
            ("Social Security cards", "Identity", "Safe", "No", "", "All", ""),
            ("Appliance warranties", "Warranty", "Kitchen drawer", "Yes", dplus(190), "Household", "See Inventory"),
            ("Tax returns (3 yrs)", "Tax", "File cabinet B", "Yes", "", "Household", ""),
            ("Wills / Estate", "Legal", "Attorney + safe", "Yes", "", "Mom & Dad", "Review yearly"),
            ("Medical records", "Medical", "Binder", "Yes", "", "All", ""),
            ("School records", "School", "Binder", "Yes", "", "Kids", ""),
            ("Pet records", "Pet", "Binder", "Yes", "", "Pets", "See Pet Care"),
        ],
        [24, 14, 18, 12, 16, 14, 20],
        text_left={1, 3, 7}, dates={5},
        validations=[("D", "YesNoList")], reserved=40)
    ws = wb["Documents"]
    ws.conditional_formatting.add(
        "D5:D44", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(
        "E5:E44", FormulaRule(formula=['AND($E5<>"",$E5<=TODAY()+60)'], fill=fill(WARN_BG)))


# ===========================================================================
# 26 — Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 18, 3, 16, 12, 12, 12, 12, 12, 12, 2])
    luxe_header(ws, "M", "📊  ANALYTICS DASHBOARD",
                "The numbers behind your home — health scores, spending trends & a Household Health Score.")

    merge_set(ws, "B5:D5", "HOUSEHOLD HEALTH SCORES", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Bills Paid", '=IFERROR(COUNTIF(BillPaid,"Yes")/MAX(COUNTA(BillName),1),0)'),
        ("Budget Used", '=IFERROR(BudgetTotalActual/BudgetTotalPlanned,0)'),
        ("Cleaning Done", '=IFERROR(COUNTIF(CleanStatus,"Done")/MAX(COUNTA(CleanTask),1),0)'),
        ("Chore Completion", '=IFERROR(AVERAGE(ChorePct),0)'),
        ("Pantry Stocked", '=IFERROR(SUMPRODUCT((PantryName<>"")*(PantryQty>PantryReorder))/MAX(COUNTA(PantryName),1),0)'),
        ("Savings Funded", '=IFERROR(SUM(SavCurrent)/SUM(SavTarget),0)'),
        ("Goal Progress", '=IFERROR(AVERAGE(GoalProgress),0)'),
        ("Meal Plan", '=IFERROR((COUNTIF(MealWeek,"<>"&"")-COUNTIF(MealWeek,"—"))/35,0)'),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4,
                value=f'=IF(C{r}>=0.75,"Great",IF(C{r}>=0.4,"On Track","Needs Love"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(
        f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF" + WARN_BG,
                       mid_type="num", mid_value=0.5, mid_color="FFFFF3CD",
                       end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    name_range(wb, "HealthScores", "Analytics", "C", start, end)

    # Household Health Score
    merge_set(ws, "F5:H5", "HOUSEHOLD HEALTH SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=48, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(
                top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "A blend of money, meals, cleaning, chores & goals.", "subtitle")
    ws["F10"].fill = fill(IVORY)

    # grocery spending trend
    merge_set(ws, "F12:H12", "GROCERY SPENDING TREND", "section")
    ws.cell(row=13, column=6, value="Week").style = "th"
    ws.cell(row=13, column=7, value="Spent").style = "th"
    trend = [("Wk 1", 210), ("Wk 2", 185), ("Wk 3", 232), ("Wk 4", 198),
             ("Wk 5", 176), ("Wk 6", 205)]
    for i, (wk, amt) in enumerate(trend):
        r = 14 + i
        ws.cell(row=r, column=6, value=wk).style = "td_left"
        c = ws.cell(row=r, column=7, value=amt); c.style = "td"; c.number_format = '"$"#,##0'
    line = LineChart(); line.title = "Grocery Spending (6 wks)"; line.height = 7.5; line.width = 13
    line.add_data(Reference(ws, min_col=7, min_row=13, max_row=13 + len(trend)), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=6, min_row=14, max_row=13 + len(trend)))
    line.legend = None
    ws.add_chart(line, "F21")

    bar = BarChart(); bar.type = "bar"; bar.title = "Health by Dimension"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None
    ws.add_chart(bar, "B17")


# ===========================================================================
# 27 — Home Inspiration Board (image placeholders)
# ===========================================================================
def build_inspiration(wb):
    ws = wb.create_sheet("Inspiration")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "💡  HOME INSPIRATION BOARD",
                "Your visual wishlist — paste inspiration photos, set a budget, and prioritize.")
    sections = ["Kitchen", "Living Room", "Bedroom", "Bathroom Remodel",
                "Organization", "Garden", "Holiday Decor", "Meal Presentation",
                "Cleaning Hacks", "DIY Projects", "Dream Home", "Storage Solutions"]
    # 2 columns of cards; each card = 8 rows
    top0 = 5
    card_h = 8
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6  # left block B.., right block F..
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th")
        ws.row_dimensions[row].height = 22
        # image placeholder
        merge_set(ws, f"{L}{row+1}:{R}{row+4}",
                  "📷\nPaste inspiration\nhere", "imgbox")
        for rr in range(row + 1, row + 5):
            ws.row_dimensions[rr].height = 22
        # fields
        ws.cell(row=row + 5, column=col, value="Budget").style = "field_label"
        c = ws.cell(row=row + 5, column=col + 1, value=""); c.style = "field_value"; c.number_format = '"$"#,##0'
        ws.cell(row=row + 5, column=col + 2, value="").style = "field_value"
        ws.cell(row=row + 6, column=col, value="Priority").style = "field_label"
        pc = ws.cell(row=row + 6, column=col + 1, value="Medium"); pc.style = "field_value"
        sc = ws.cell(row=row + 6, column=col + 2, value="Idea"); sc.style = "field_value"
        ws.cell(row=row + 7, column=col, value="Notes").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")
        add_dv(ws, f"{M}{row+6}", "PriorityList")
        add_dv(ws, f"{R}{row+6}", "StatusList")


# ===========================================================================
# 1 — Executive Home Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2])
    ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🏡  HOME COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2",
              "  Finances, schedules, meals, cleaning, maintenance & goals — your whole home, automatically organized.",
              "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    row1 = [
        ("BUDGET LEFT", "=BudgetTotalPlanned-BudgetTotalActual", "money"),
        ("BILLS PAID", '=COUNTIF(BillPaid,"Yes")', "num"),
        ("CLEANING DONE", '=COUNTIF(CleanStatus,"Done")', "num"),
        ("PANTRY STOCK", '=IFERROR(SUMPRODUCT((PantryName<>"")*(PantryQty>PantryReorder))/MAX(COUNTA(PantryName),1),0)', "pct"),
        ("EVENTS / WEEK", '=SUMPRODUCT((CalDate<>"")*(CalDate>=TODAY())*(CalDate<=TODAY()+7))', "num"),
    ]
    row2 = [
        ("MEAL PLAN", '=IFERROR((COUNTIF(MealWeek,"<>"&"")-COUNTIF(MealWeek,"—"))/35,0)', "pct"),
        ("SAVINGS GOAL", "=IFERROR(SUM(SavCurrent)/SUM(SavTarget),0)", "pct"),
        ("PROJECTS ACTIVE", '=COUNTIF(ProjStatus,"In Progress")', "num"),
        ("MAINT. DUE", '=SUMPRODUCT((MaintNext<>"")*(MaintNext<=TODAY()+30))', "num"),
        ("ROUTINE SCORE", "=IFERROR(AVERAGE(ChorePct),0)", "pct"),
    ]
    cols5 = [2, 4, 6, 8, 10]
    for (lab, fml, kind), col in zip(row1, cols5):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols5):
        kpi_card(ws, 8, col, 2, lab, fml, kind)

    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "MONEY & HOME AT A GLANCE", "section_gold")

    # Monthly Spending donut (Budget actuals)
    bend = L0 + len(EXPENSE_CATS) - 1
    d1 = DoughnutChart(); d1.title = "Monthly Spending"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=bend), titles_from_data=True)
    d1.set_categories(Reference(wb["Budget"], min_col=1, min_row=L0, max_row=bend))
    d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")

    # Bill payment status donut
    d2 = DoughnutChart(); d2.title = "Bill Payment Status"; d2.height = 8.2; d2.width = 11.5
    d2.add_data(Reference(wb["Bills"], min_col=11, min_row=4, max_row=6), titles_from_data=True)
    d2.set_categories(Reference(wb["Bills"], min_col=10, min_row=5, max_row=6))
    d2.dataLabels = no_labels()
    ws.add_chart(d2, "H12")

    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "ROUTINES & PROGRESS", "section_gold")

    # Cleaning progress by frequency (bar)
    cl_end = 5 + len(CLEAN_FREQ)
    c1 = BarChart(); c1.type = "col"; c1.title = "Cleaning Progress"; c1.height = 8.2; c1.width = 11.5
    c1.add_data(Reference(wb["Cleaning"], min_col=10, min_row=5, max_row=cl_end), titles_from_data=True)
    c1.set_categories(Reference(wb["Cleaning"], min_col=8, min_row=6, max_row=cl_end))
    c1.legend = None
    ws.add_chart(c1, "B30")

    # Chore completion by member (bar)
    ch_end = 5 + len(FAMILY)
    c2 = BarChart(); c2.type = "bar"; c2.title = "Chore Completion"; c2.height = 8.2; c2.width = 11.5
    c2.add_data(Reference(wb["Chores"], min_col=9, min_row=5, max_row=ch_end), titles_from_data=True)
    c2.set_categories(Reference(wb["Chores"], min_col=8, min_row=6, max_row=ch_end))
    c2.legend = None
    ws.add_chart(c2, "H30")

    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "PANTRY & GOALS", "section_gold")

    # Pantry by category donut
    pc_end = 5 + len(PANTRY_CATS)
    d3 = DoughnutChart(); d3.title = "Pantry by Category"; d3.height = 8.2; d3.width = 11.5
    d3.add_data(Reference(wb["Pantry"], min_col=10, min_row=5, max_row=pc_end), titles_from_data=True)
    d3.set_categories(Reference(wb["Pantry"], min_col=9, min_row=6, max_row=pc_end))
    d3.dataLabels = no_labels()
    ws.add_chart(d3, "B48")

    # Household goals progress (bar)
    g_end = L0 + 8 - 1
    g1 = BarChart(); g1.type = "bar"; g1.title = "Household Goals"; g1.height = 8.2; g1.width = 11.5
    g1.add_data(Reference(wb["Family Goals"], min_col=4, min_row=4, max_row=g_end), titles_from_data=True)
    g1.set_categories(Reference(wb["Family Goals"], min_col=1, min_row=L0, max_row=g_end))
    g1.legend = None
    ws.add_chart(g1, "H48")

    ws.row_dimensions[66].height = 26
    merge_set(ws, "B66:M66",
              "Home Command Center™ — your whole household, organized in one place. Edit anything in Settings.",
              "subtitle")


# ===========================================================================
# Build
# ===========================================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    build_settings(wb)
    build_welcome(wb)
    build_family_directory(wb)
    build_calendar(wb)
    build_daily(wb)
    build_budget(wb)
    build_bills(wb)
    build_grocery(wb)
    build_pantry(wb)
    build_fridge(wb)
    build_meal_planner(wb)
    build_recipes(wb)
    build_cleaning(wb)
    build_chores(wb)
    build_maintenance(wb)
    build_home_inventory(wb)
    build_subscriptions(wb)
    build_goals(wb)
    build_savings(wb)
    build_holiday(wb)
    build_travel(wb)
    build_children(wb)
    build_pets(wb)
    build_wellness(wb)
    build_projects(wb)
    build_documents(wb)
    build_analytics(wb)
    build_inspiration(wb)
    build_dashboard(wb)   # inserted at index 0

    order = ["Welcome", "Dashboard", "Family Directory", "Calendar", "Daily",
             "Budget", "Bills", "Grocery", "Pantry", "Fridge & Freezer",
             "Meal Planner", "Recipes", "Cleaning", "Chores", "Maintenance",
             "Home Inventory", "Subscriptions", "Family Goals", "Savings",
             "Holidays & Events", "Travel", "Children's Hub", "Pet Care",
             "Wellness", "Projects", "Documents", "Analytics", "Inspiration",
             "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "Home_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
