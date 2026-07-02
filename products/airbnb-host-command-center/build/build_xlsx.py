"""Build Airbnb Host Command Center™ — The Ultimate Airbnb Business Management System.

19 sheets + Welcome · a premium vacation-rental operating system in Excel & Sheets.
Reservations, finances, pricing, guests, cleaning, maintenance, inventory,
reviews, taxes, multi-property analytics — one elegant dashboard.

Run: python3 build_xlsx.py   ->  ../Airbnb_Host_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

PROPERTIES = ["The Pine Cabin", "Beachside Bungalow", "Downtown Loft"]
PROPERTY_TYPES = ["Cabin", "Beach House", "Loft", "Apartment", "Tiny Home", "Villa", "Cottage", "Condo"]
PLATFORMS = ["Airbnb", "VRBO", "Booking.com", "Direct", "Expedia"]
EXPENSE_CATS = ["Mortgage", "Utilities", "Internet", "Insurance", "Cleaning", "Supplies",
                "Repairs", "Property Taxes", "HOA Fees", "Software", "Marketing",
                "Platform Fees", "Miscellaneous"]
MAINT_CATS = ["Plumbing", "Electrical", "HVAC", "Appliance", "Structural", "Landscaping", "Safety", "Cosmetic"]
CLEAN_STATUS = ["Scheduled", "In Progress", "Done", "Skipped"]
SUPPLIER_CATS = ["Cleaner", "Handyman", "Electrician", "Plumber", "Landscaper", "Laundry", "Contractor", "Other"]
GOAL_CATS = ["Revenue", "Occupancy", "Reviews", "Profit", "Expansion", "Operations"]
PAY_STATUS = ["Paid", "Deposit", "Pending", "Refunded"]
BOOK_STATUS = ["Confirmed", "In Progress", "Completed", "Cancelled", "Inquiry"]
YESNO = ["Yes", "No"]
CURRENCIES = ["USD", "EUR", "GBP", "CAD", "AUD"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula; vc.font = Font(size=18, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pct": "0%", "days": "0", "dec": "0.0", "rate": '"$"#,##0'}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 3] + [17] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your business details once — every dashboard follows. Edit the lists to fit any portfolio.")
    merge_set(ws, "B5:C5", "BUSINESS INPUTS", "section")
    controls = [
        ("Business Name", "Summit Stays Co.", None, "BizName"),
        ("Primary Property", "The Pine Cabin", None, "PrimaryProp"),
        ("# of Properties", 3, "0", "NumProps"),
        ("Currency", "USD", None, "HomeCurr"),
        ("Occupancy Target", 0.75, "0%", "OccTarget"),
        ("Margin Target", 0.40, "0%", "MarginTarget"),
        ("Nightly Rate Target", 175, '"$"#,##0', "RateTarget"),
        ("Reporting Month", dt.date.today().replace(day=1), "mmmm yyyy", "ReportMonth"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Property", PROPERTIES, "PropertyList"), ("F", "Property Type", PROPERTY_TYPES, "PropTypeList"),
             ("G", "Platform", PLATFORMS, "PlatformList"), ("H", "Expense Category", EXPENSE_CATS, "ExpenseCatList"),
             ("I", "Maintenance", MAINT_CATS, "MaintCatList"), ("J", "Supplier", SUPPLIER_CATS, "SupplierCatList"),
             ("K", "Goal Category", GOAL_CATS, "GoalCatList"), ("L", "Currency", CURRENCIES, "CurrencyList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    # small lists lower down
    small = [("E", 18, "Booking Status", BOOK_STATUS, "BookStatusList"),
             ("F", 18, "Cleaning Status", CLEAN_STATUS, "CleanStatusList"),
             ("G", 18, "Payment Status", PAY_STATUS, "PayStatusList"),
             ("H", 18, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 82, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🏡  AIRBNB HOST COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate Airbnb business management system — run your rental like a business.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "WELCOME, HOST — THIS IS YOUR OPERATING SYSTEM", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Most Airbnb spreadsheets track only bookings or income. Airbnb Host Command Center™ replaces "
                      "your tangle of disconnected files with ONE professional system that manages reservations, "
                      "finances, pricing, guest communication, cleaning, maintenance, supplies, reviews, taxes and "
                      "business performance — all from a single elegant dashboard. Run your short-term rental like "
                      "the real business it is: save time, improve organization, and grow your profit.")
    ws["B6"].style = "body"
    for r in (6, 7, 8, 9):
        ws.row_dimensions[r].height = 22
    merge_set(ws, "B11:B11", "START HERE", "section")
    steps = ["1.  Open Settings and add your business name, properties, currency & targets.",
             "2.  Fill in Property Profile for each listing (Wi-Fi, codes, contacts).",
             "3.  Add bookings to the Master Booking Calendar — occupancy calculates itself.",
             "4.  Log income & expenses in the Financial Command Center for live profit.",
             "5.  Use Cleaning, Maintenance & Inventory to keep every turnover flawless.",
             "6.  Watch the Executive Host Dashboard track revenue, occupancy & health."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (a 3-property portfolio) is included so you can see how everything connects — "
               "just type over it with your own. Revenue, occupancy, profit margin, cleaning schedules, reorder "
               "alerts and the Business Health Score all update automatically. Every sheet is print-friendly and "
               "works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more five-star stays — welcome to professional hosting.", "section_gold")


# ===========================================================================
# 2 — Property Profile
# ===========================================================================
def build_property(wb):
    ws = wb.create_sheet("Property Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 28, 6, 24, 24, 2])
    luxe_header(ws, "G", "🏠  PROPERTY PROFILE", "Every detail of your flagship listing — the facts guests and cleaners need, in one place.")
    blocks = [
        ("THE LISTING", [("Property Name", "The Pine Cabin"), ("Property Type", "Cabin"),
                         ("Address", "412 Ridgeline Rd, Aspen CO"), ("Bedrooms", 3), ("Bathrooms", 2),
                         ("Maximum Guests", 8), ("Check-In Time", "4:00 PM"), ("Check-Out Time", "11:00 AM")]),
        ("ACCESS & SERVICES", [("Wi-Fi Network", "PineCabin_5G"), ("Wi-Fi Password", "cozy-nights-2024"),
                               ("Smart Lock Code Ref", "See Message Templates"), ("Parking", "2 spots, driveway"),
                               ("Trash Day", "Tuesday AM"), ("Cleaner", "Sparkle Co. — Maria"),
                               ("Handyman", "Dave's Repairs"), ("Insurance", "Proper #STR-88120")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B16:F16", "EMERGENCY CONTACTS", "section_gold"); ws.row_dimensions[16].height = 22
    ec = [("Property Manager", "You — (555) 210-4521"), ("Emergency Maintenance", "Dave — (555) 887-1200"),
          ("Nearest Hospital", "Aspen Valley — (555) 544-1000"), ("Local Police (non-emerg.)", "(555) 920-3311")]
    for i, (lab, val) in enumerate(ec):
        r = 17 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=lab).style = "field_label"
        ws.cell(row=r, column=col + 1, value=val).style = "field_value"
        ws.row_dimensions[r].height = 24


# ===========================================================================
# 3 — Master Booking Calendar
# ===========================================================================
def build_calendar(wb):
    # (guest, platform, checkin_offset(+future/-past), nights, property, status, cleaning, payment)
    bookings = [
        ("Sarah Mitchell", "Airbnb", -22, 5, "The Pine Cabin", "Completed", "Done", "Paid"),
        ("James Carter", "Airbnb", -18, 3, "Downtown Loft", "Completed", "Done", "Paid"),
        ("The Nguyen Family", "VRBO", -14, 6, "Beachside Bungalow", "Completed", "Done", "Paid"),
        ("Emily Rodriguez", "Airbnb", -9, 4, "The Pine Cabin", "Completed", "Done", "Paid"),
        ("Michael Brooks", "Direct", -6, 2, "Downtown Loft", "Completed", "Done", "Paid"),
        ("Olivia Bennett", "Airbnb", -4, 3, "Beachside Bungalow", "Completed", "Done", "Paid"),
        ("David Kim", "Airbnb", -2, 4, "The Pine Cabin", "In Progress", "In Progress", "Paid"),
        ("Priya Patel", "VRBO", -1, 5, "Beachside Bungalow", "In Progress", "Scheduled", "Paid"),
        ("Tom & Lisa Hayes", "Airbnb", 2, 3, "Downtown Loft", "Confirmed", "Scheduled", "Paid"),
        ("Grace Liu", "Airbnb", 4, 6, "The Pine Cabin", "Confirmed", "Scheduled", "Paid"),
        ("Robert Fox", "Direct", 6, 2, "Beachside Bungalow", "Confirmed", "Scheduled", "Deposit"),
        ("Hannah Scott", "Airbnb", 10, 4, "Downtown Loft", "Confirmed", "Scheduled", "Paid"),
        ("Marcus Webb", "VRBO", 14, 5, "The Pine Cabin", "Confirmed", "Scheduled", "Paid"),
        ("Ana Duarte", "Airbnb", 20, 3, "Beachside Bungalow", "Confirmed", "Scheduled", "Deposit"),
        ("Chris Allen", "Airbnb", 24, 4, "Downtown Loft", "Confirmed", "Scheduled", "Pending"),
    ]
    sample = []
    for g, plat, off, nights, prop, status, clean, pay in bookings:
        ci = dplus(off) if off >= 0 else dminus(-off)
        co = ci + dt.timedelta(days=nights)
        sample.append((g, plat, ci, co, nights, status, prop, clean, pay))
    ws, start, end = build_log(
        wb, "Calendar", "📅", "MASTER BOOKING CALENDAR",
        "Every stay across every property — occupancy, check-ins & check-outs calculate automatically.",
        ["Guest", "Platform", "Check-In", "Check-Out", "Nights", "Status", "Property", "Cleaning", "Payment"],
        sample, [22, 13, 13, 13, 9, 14, 20, 13, 12],
        text_left={1, 7}, dates={3, 4}, ints={5},
        validations=[("B", "PlatformList"), ("F", "BookStatusList"), ("G", "PropertyList"),
                     ("H", "CleanStatusList"), ("I", "PayStatusList")], reserved=40)
    nrange(wb, "CalGuest", "Calendar", "A", start, end)
    nrange(wb, "CalPlatform", "Calendar", "B", start, end)
    nrange(wb, "CalCheckIn", "Calendar", "C", start, end)
    nrange(wb, "CalCheckOut", "Calendar", "D", start, end)
    nrange(wb, "CalNights", "Calendar", "E", start, end)
    nrange(wb, "CalStatus", "Calendar", "F", start, end)
    nrange(wb, "CalProperty", "Calendar", "G", start, end)
    # booking-sources mini table (feeds dashboard donut)
    kcol = 11
    ws.cell(row=4, column=kcol, value="Source").style = "th"
    ws.cell(row=4, column=kcol + 1, value="Bookings").style = "th"
    for i, plat in enumerate(["Airbnb", "VRBO", "Direct"]):
        r = 5 + i
        ws.cell(row=r, column=kcol, value=plat).style = "td_left"
        c = ws.cell(row=r, column=kcol + 1, value=f'=COUNTIF(CalPlatform,"{plat}")'); c.style = "td"
    set_widths(ws, [22, 13, 13, 13, 9, 14, 20, 13, 12, 3, 14, 12])
    cell_name(wb, "SrcLabel", "Calendar", "$K$5:$K$7")
    cell_name(wb, "SrcCount", "Calendar", "$L$5:$L$7")
    # occupancy conditional accent on nights
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF" + WHITE,
                       end_type="num", end_value=7, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 4 — Reservation Manager
# ===========================================================================
def build_reservations(wb):
    # (id, guest, property, nights, gross, cleaning, taxes, discount, platform_fee)
    rows = [
        ("RES-1041", "Sarah Mitchell", "The Pine Cabin", 5, 1150, 120, 92, 0, 35),
        ("RES-1042", "James Carter", "Downtown Loft", 3, 480, 75, 38, 0, 15),
        ("RES-1043", "The Nguyen Family", "Beachside Bungalow", 6, 1260, 150, 101, 60, 38),
        ("RES-1044", "Emily Rodriguez", "The Pine Cabin", 4, 940, 120, 75, 0, 29),
        ("RES-1045", "Michael Brooks", "Downtown Loft", 2, 320, 75, 26, 0, 0),
        ("RES-1046", "Olivia Bennett", "Beachside Bungalow", 3, 660, 150, 53, 0, 20),
        ("RES-1047", "David Kim", "The Pine Cabin", 4, 940, 120, 75, 40, 29),
        ("RES-1048", "Priya Patel", "Beachside Bungalow", 5, 1150, 150, 92, 0, 35),
        ("RES-1049", "Tom & Lisa Hayes", "Downtown Loft", 3, 480, 75, 38, 0, 15),
        ("RES-1050", "Grace Liu", "The Pine Cabin", 6, 1440, 120, 115, 0, 44),
        ("RES-1051", "Robert Fox", "Beachside Bungalow", 2, 460, 150, 37, 0, 0),
        ("RES-1052", "Hannah Scott", "Downtown Loft", 4, 640, 75, 51, 30, 20),
    ]
    ws = wb.create_sheet("Reservations"); ws.sheet_view.showGridLines = False
    set_widths(ws, [12, 20, 20, 8, 12, 12, 10, 11, 12, 13, 22])
    luxe_header(ws, "K", "🧾  RESERVATION MANAGER",
                "Per-booking money, decoded — gross, fees, taxes & net revenue for every stay.")
    table_headers(ws, 4, ["Res ID", "Guest", "Property", "Nights", "Gross", "Clean Fee",
                          "Taxes", "Discount", "Platform Fee", "Net Revenue", "Special Requests"])
    reqs = ["Early check-in", "—", "Crib + high chair", "Late checkout", "—", "Ground floor",
            "Pet (dog)", "Extra towels", "—", "Anniversary setup", "—", "Airport pickup info"]
    start = L0
    for i, (rid, g, prop, n, gross, clean, tax, disc, pf) in enumerate(rows):
        r = start + i
        vals = [rid, g, prop, n, gross, clean, tax, disc, pf]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=r, column=ci, value=v)
        net = ws.cell(row=r, column=10, value=f"=E{r}+F{r}-G{r}-H{r}-I{r}")
        ws.cell(row=r, column=11, value=reqs[i])
    end = start + 40 - 1
    style_rows(ws, start, end, 11, text_left={2, 3, 11}, ints={4},
               money={5, 6, 7, 8, 9, 10})
    add_dv(ws, f"C{start}:C{end}", "PropertyList")
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in (5, 6, 7, 8, 9, 10):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "ResProperty", "Reservations", "C", start, end)
    nrange(wb, "ResNet", "Reservations", "J", start, end)
    ws.conditional_formatting.add(f"J{start}:J{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1500, color=PRIMARY, showValue=True))


# ===========================================================================
# 5 — Financial Command Center
# ===========================================================================
def build_financial(wb):
    ws = wb.create_sheet("Financial"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 3, 22, 14, 3, 20, 14, 3, 12, 12, 11])
    luxe_header(ws, "M", "💰  FINANCIAL COMMAND CENTER",
                "This month's P&L — revenue, expenses, net profit & margin, computed live.")
    # INCOME
    merge_set(ws, "B5:C5", "MONTHLY INCOME", "section")
    income = [("Nightly Revenue", 10030), ("Cleaning Fees", 1200), ("Extra Services", 370)]
    for i, (lab, val) in enumerate(income):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"; c.number_format = '"$"#,##0'
    ws.cell(row=9, column=2, value="TOTAL REVENUE").style = "th"
    c = ws.cell(row=9, column=3, value="=SUM(C6:C8)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=10, column=2, value="Security Deposits (held)").style = "field_label"
    c = ws.cell(row=10, column=3, value=500); c.style = "field_value"; c.number_format = '"$"#,##0'
    # EXPENSES
    merge_set(ws, "E5:F5", "MONTHLY EXPENSES", "section_gold")
    expenses = {"Mortgage": 3200, "Utilities": 420, "Internet": 95, "Insurance": 180, "Cleaning": 960,
                "Supplies": 340, "Repairs": 250, "Property Taxes": 520, "HOA Fees": 150, "Software": 85,
                "Marketing": 120, "Platform Fees": 348, "Miscellaneous": 82}
    for i, cat in enumerate(EXPENSE_CATS):
        r = 6 + i
        ws.cell(row=r, column=5, value=cat).style = "td_left"
        c = ws.cell(row=r, column=6, value=expenses[cat]); c.style = "input"; c.number_format = '"$"#,##0'
    etot = 6 + len(EXPENSE_CATS)
    ws.cell(row=etot, column=5, value="TOTAL EXPENSES").style = "th"
    c = ws.cell(row=etot, column=6, value=f"=SUM(F6:F{etot-1})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    # SUMMARY
    merge_set(ws, "H5:I5", "THE BOTTOM LINE", "section_gold")
    summ = [("Revenue", "=C9", '"$"#,##0'), ("Expenses", f"=F{etot}", '"$"#,##0'),
            ("Net Profit", "=I6-I7", '"$"#,##0'), ("Profit Margin", "=IFERROR(I8/I6,0)", "0%"),
            ("Cash Flow", "=I8", '"$"#,##0'), ("Revenue / Property", "=IFERROR(I6/NumProps,0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(summ):
        r = 6 + i
        ws.cell(row=r, column=8, value=lab).style = "field_label"
        c = ws.cell(row=r, column=9, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net Profit", "Profit Margin"):
            ws.cell(row=r, column=9).fill = fill(MINT_BG)
    cell_name(wb, "FinRevenue", "Financial", "$C$9")
    cell_name(wb, "FinNightly", "Financial", "$C$6")
    cell_name(wb, "FinExpenses", "Financial", f"$F${etot}")
    cell_name(wb, "FinNetProfit", "Financial", "$I$8")
    cell_name(wb, "FinMargin", "Financial", "$I$9")
    nrange(wb, "ExpItem", "Financial", "E", 6, 6 + len(EXPENSE_CATS) - 1)
    nrange(wb, "ExpAmt", "Financial", "F", 6, 6 + len(EXPENSE_CATS) - 1)
    # 6-MONTH TREND (feeds dashboard line chart)
    merge_set(ws, "K5:M5", "6-MONTH TREND", "section")
    ws.cell(row=6, column=11, value="Month").style = "th"
    ws.cell(row=6, column=12, value="Revenue").style = "th"
    ws.cell(row=6, column=13, value="Occ %").style = "th"
    months = list(_recent_months(6))
    revs = [8200, 9400, 10100, 11200, 10800, 11600]
    occs = [0.58, 0.64, 0.68, 0.72, 0.70, 0.66]
    for i, (m, rv, oc) in enumerate(zip(months, revs, occs)):
        r = 7 + i
        ws.cell(row=r, column=11, value=m).style = "td_left"
        c = ws.cell(row=r, column=12, value=rv); c.style = "td"; c.number_format = '"$"#,##0'
        c2 = ws.cell(row=r, column=13, value=oc); c2.style = "td"; c2.number_format = "0%"
    cell_name(wb, "TrendMonth", "Financial", "$K$7:$K$12")
    cell_name(wb, "TrendRev", "Financial", "$L$7:$L$12")


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    out = []
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1))
        m -= 1
        if m == 0:
            m = 12; y -= 1
    for d in reversed(seq):
        out.append(d.strftime("%b"))
    return out


# ===========================================================================
# 6 — Pricing Strategy
# ===========================================================================
def build_pricing(wb):
    ws = wb.create_sheet("Pricing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 14, 3, 22, 14, 2])
    luxe_header(ws, "I", "🏷  PRICING STRATEGY", "Set rates by property & season — then model what a price change does to your month.")
    table_headers(ws, 4, ["Property", "Base / Night", "Weekend", "Holiday", "Min Stay"])
    prices = [("The Pine Cabin", 235, 285, 340, 2), ("Beachside Bungalow", 195, 240, 300, 3),
              ("Downtown Loft", 145, 175, 210, 2)]
    for i, (p, base, wknd, hol, mn) in enumerate(prices):
        r = 5 + i
        ws.cell(row=r, column=1, value=p).style = "td_left"
        for ci, v, fmt in [(2, base, '"$"#,##0'), (3, wknd, '"$"#,##0'), (4, hol, '"$"#,##0'), (5, mn, "0")]:
            c = ws.cell(row=r, column=ci, value=v); c.style = "input" if ci in (2, 3, 4) else "td"; c.number_format = fmt
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    # scenario modeler
    merge_set(ws, "G4:H4", "SCENARIO MODELER", "section_gold")
    scen = [("Avg nightly rate", 175, '"$"#,##0', "input"),
            ("Occupancy assumption", 0.70, "0%", "input"),
            ("Nights available / mo", 90, "0", "input"),
            ("Projected revenue", "=H5*H6*H7", '"$"#,##0', "field_value"),
            ("vs current revenue", "=H8-FinRevenue", '"$"#,##0;[Red]-\"$\"#,##0', "field_value"),
            ("% change", "=IFERROR(H8/FinRevenue-1,0)", "0%", "field_value")]
    for i, (lab, val, fmt, st) in enumerate(scen):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=val); c.style = st; c.number_format = fmt
        if lab == "Projected revenue":
            ws.cell(row=r, column=8).fill = fill(MINT_BG)
    merge_set(ws, "B10:E10", "SEASONAL GUIDANCE", "section")
    tips = ["Raise weekend & holiday rates 20–45% above base — demand supports it.",
            "Set a 2–3 night minimum to cut costly one-night turnovers.",
            "Model occupancy vs rate here before changing prices in your listing.",
            "Small rate lifts in peak season compound into big monthly gains."]
    for i, t in enumerate(tips):
        r = 11 + i
        merge_set(ws, f"B{r}:E{r}", "•  " + t, "body"); ws.row_dimensions[r].height = 22


# ===========================================================================
# 7 — Guest CRM
# ===========================================================================
def build_guests(wb):
    sample = [
        ("Sarah Mitchell", "sarah.m@email.com", "(555) 201-8841", 3, "Yes", "USA", "Quiet, early check-in", 5),
        ("James Carter", "jcarter@email.com", "(555) 442-1190", 1, "No", "USA", "Business traveler", 5),
        ("The Nguyen Family", "nguyen.fam@email.com", "(555) 771-3320", 2, "Yes", "USA", "Kids, needs crib", 5),
        ("Emily Rodriguez", "emily.r@email.com", "(555) 660-2245", 1, "No", "USA", "Dog-friendly stay", 4),
        ("Michael Brooks", "mbrooks@email.com", "(555) 118-9007", 4, "Yes", "USA", "VIP repeat guest", 5),
        ("Olivia Bennett", "obennett@email.com", "(555) 233-6612", 1, "No", "UK", "Ground floor only", 5),
        ("David Kim", "dkim@email.com", "(555) 909-4471", 2, "Yes", "USA", "Late checkout usual", 5),
        ("Priya Patel", "priya.p@email.com", "(555) 556-3388", 1, "No", "Canada", "Extra towels", 4),
        ("Grace Liu", "grace.liu@email.com", "(555) 640-1120", 2, "Yes", "USA", "Anniversary trips", 5),
        ("Hannah Scott", "hscott@email.com", "(555) 812-7754", 1, "No", "Australia", "First-time guest", 5),
    ]
    ws, start, end = build_log(
        wb, "Guests", "🤝", "GUEST CRM",
        "Know your guests — history, preferences & repeat status turn one stay into many.",
        ["Guest", "Email", "Phone", "Stays", "Repeat?", "Country", "Preferences", "Rating Given"],
        sample, [22, 26, 15, 8, 10, 12, 26, 12],
        text_left={1, 2, 7}, ints={4}, dec={8},
        validations=[("E", "YesNoList")], reserved=40)
    nrange(wb, "GuestName", "Guests", "A", start, end)
    nrange(wb, "GuestRepeat", "Guests", "E", start, end)


# ===========================================================================
# 8 — Cleaning Command Center
# ===========================================================================
def build_cleaning(wb):
    sample = [
        (dminus(17), "Sparkle Co. — Maria", "The Pine Cabin", "Done", 1.0, "Yes", "Yes", "Deep clean"),
        (dminus(15), "Sparkle Co. — Maria", "Downtown Loft", "Done", 1.0, "Yes", "Yes", "—"),
        (dminus(8), "FreshTurn — Ana", "Beachside Bungalow", "Done", 1.0, "Yes", "Yes", "Sand everywhere"),
        (dminus(5), "Sparkle Co. — Maria", "The Pine Cabin", "Done", 1.0, "Yes", "Yes", "—"),
        (dminus(4), "FreshTurn — Ana", "Downtown Loft", "Done", 1.0, "Yes", "Yes", "—"),
        (dplus(2), "FreshTurn — Ana", "The Pine Cabin", "In Progress", 0.5, "No", "No", "Turnover for Hayes"),
        (dplus(4), "Sparkle Co. — Maria", "Beachside Bungalow", "Scheduled", 0.0, "No", "No", "Same-day turn"),
        (dplus(5), "FreshTurn — Ana", "Downtown Loft", "Scheduled", 0.0, "No", "No", "—"),
        (dplus(6), "Sparkle Co. — Maria", "The Pine Cabin", "Scheduled", 0.0, "No", "No", "Restock coffee"),
        (dplus(9), "FreshTurn — Ana", "Beachside Bungalow", "Scheduled", 0.0, "No", "No", "—"),
    ]
    ws, start, end = build_log(
        wb, "Cleaning", "🧽", "CLEANING COMMAND CENTER",
        "Never miss a turnover — every clean tracked, with checklist, linens & restock at a glance.",
        ["Date", "Cleaner", "Property", "Status", "Checklist %", "Linens", "Restocked", "Notes"],
        sample, [13, 22, 20, 14, 12, 10, 12, 22],
        text_left={2, 3, 8}, dates={1}, pcts={5},
        validations=[("C", "PropertyList"), ("D", "CleanStatusList"), ("F", "YesNoList"), ("G", "YesNoList")], reserved=40)
    nrange(wb, "CleanStatus", "Cleaning", "D", start, end)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Scheduled"'], fill=fill(WARN_BG)))
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="equal", formula=['"Done"'], fill=fill(MINT_BG)))


# ===========================================================================
# 9 — Maintenance Manager
# ===========================================================================
def build_maintenance(wb):
    sample = [
        ("Fix leaky kitchen faucet", "High", "The Pine Cabin", "Dave's Repairs", 85, dplus(3), "Open", "No", "Guest reported drip"),
        ("Service HVAC / filters", "Medium", "Beachside Bungalow", "CoolAir Inc.", 140, dplus(6), "In Progress", "Yes", "Annual service"),
        ("Replace smoke detector battery", "High", "Downtown Loft", "Self", 12, dplus(1), "Open", "No", "Safety check"),
        ("Repaint deck railing", "Low", "The Pine Cabin", "Dave's Repairs", 220, dplus(20), "Open", "No", "Cosmetic"),
        ("Unclog shower drain", "Medium", "Beachside Bungalow", "QuickPlumb", 95, dminus(4), "Done", "No", "Resolved"),
        ("Replace microwave", "Medium", "Downtown Loft", "Self", 130, dminus(9), "Done", "Yes", "New unit installed"),
        ("Reseal bathroom grout", "Low", "The Pine Cabin", "Dave's Repairs", 60, dminus(14), "Done", "No", "Done last turn"),
    ]
    ws, start, end = build_log(
        wb, "Maintenance", "🔧", "MAINTENANCE MANAGER",
        "Fix small things before they cost you five-star reviews — priority, cost & due date tracked.",
        ["Task", "Priority", "Property", "Assigned To", "Cost", "Due Date", "Status", "Warranty", "Notes"],
        sample, [28, 11, 20, 18, 10, 13, 14, 11, 22],
        text_left={1, 3, 4, 9}, money={5}, dates={6},
        validations=[("C", "PropertyList"), ("H", "YesNoList")], reserved=40)
    nrange(wb, "MaintTask", "Maintenance", "A", start, end)
    nrange(wb, "MaintStatus", "Maintenance", "G", start, end)
    nrange(wb, "MaintCost", "Maintenance", "E", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}",
        CellIsRule(operator="equal", formula=['"High"'], fill=fill(RED_BG)))
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"Open"'], fill=fill(WARN_BG)))


# ===========================================================================
# 10 — Inventory Manager
# ===========================================================================
def build_inventory(wb):
    # (category, item, qty, reorder, supplier, cost)
    items = [
        ("Kitchen", "Coffee pods", 18, 24, "Costco", 32),
        ("Kitchen", "Dish soap", 6, 4, "Costco", 8),
        ("Kitchen", "Paper towels", 9, 12, "Costco", 22),
        ("Coffee & Tea", "Assorted tea", 30, 20, "Amazon", 14),
        ("Bathroom", "Toilet paper", 30, 24, "Costco", 26),
        ("Bathroom", "Hand soap", 8, 6, "Amazon", 12),
        ("Toiletries", "Shampoo (bulk)", 3, 5, "Amazon", 28),
        ("Toiletries", "Conditioner (bulk)", 7, 5, "Amazon", 28),
        ("Cleaning Products", "All-purpose spray", 8, 6, "Costco", 18),
        ("Cleaning Products", "Laundry detergent", 2, 4, "Costco", 24),
        ("Linens", "Queen sheet sets", 8, 6, "Bed Supply Co.", 45),
        ("Linens", "Pillowcases", 16, 12, "Bed Supply Co.", 9),
        ("Towels", "Bath towels", 14, 12, "Bed Supply Co.", 12),
        ("Towels", "Beach towels", 10, 8, "Amazon", 15),
        ("Emergency Supplies", "First-aid kits", 4, 3, "Amazon", 20),
        ("Emergency Supplies", "Batteries (AA)", 20, 16, "Costco", 14),
        ("Appliances", "Spare coffee maker", 2, 1, "Amazon", 60),
        ("Furniture", "Patio chair cushions", 8, 6, "Wayfair", 22),
    ]
    ws = wb.create_sheet("Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 24, 9, 11, 18, 10, 14])
    luxe_header(ws, "G", "📦  INVENTORY MANAGER",
                "Never run out mid-stay — quantities, reorder levels & automatic low-stock alerts.")
    table_headers(ws, 4, ["Category", "Item", "Qty", "Reorder At", "Supplier", "Cost", "Status"])
    start = L0
    for i, (cat, item, qty, ro, sup, cost) in enumerate(items):
        r = start + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value=qty)
        ws.cell(row=r, column=4, value=ro)
        ws.cell(row=r, column=5, value=sup)
        ws.cell(row=r, column=6, value=cost)
        ws.cell(row=r, column=7, value=f'=IF(C{r}="","",IF(C{r}<=D{r},"REORDER","OK"))')  # <= matches KPI
    end = start + 40 - 1
    style_rows(ws, start, end, 7, text_left={1, 2, 5}, ints={3, 4}, money={6})
    ws.freeze_panes = "A5"
    nrange(wb, "InvCategory", "Inventory", "A", start, end)
    nrange(wb, "InvItem", "Inventory", "B", start, end)
    nrange(wb, "InvQty", "Inventory", "C", start, end)
    nrange(wb, "InvReorder", "Inventory", "D", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"REORDER"'], fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=fill(MINT_BG)))


# ===========================================================================
# 11 — Review Tracker
# ===========================================================================
def build_reviews(wb):
    # (guest, overall, clean, comm, checkin, accuracy, value, comment, responded)
    rows = [
        ("Sarah Mitchell", 5, 5, 5, 5, 5, 5, "Immaculate cabin, will return!", "Yes"),
        ("James Carter", 5, 5, 5, 5, 5, 4, "Perfect for a work trip.", "Yes"),
        ("The Nguyen Family", 5, 5, 5, 5, 5, 5, "Kids loved the beach house.", "Yes"),
        ("Emily Rodriguez", 4, 4, 5, 4, 4, 4, "Great, minor faucet drip.", "Yes"),
        ("Michael Brooks", 5, 5, 5, 5, 5, 5, "My go-to loft downtown.", "Yes"),
        ("Olivia Bennett", 5, 5, 5, 5, 5, 5, "Spotless and stylish.", "Yes"),
        ("David Kim", 5, 5, 5, 4, 5, 5, "Smooth check-in, cozy.", "Yes"),
        ("Priya Patel", 4, 4, 5, 5, 4, 4, "Lovely, towels ran low.", "Yes"),
        ("Grace Liu", 5, 5, 5, 5, 5, 5, "Romantic anniversary stay.", "Yes"),
        ("Hannah Scott", 5, 5, 5, 5, 5, 5, "Exceeded expectations!", "No"),
    ]
    ws = wb.create_sheet("Reviews"); ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 10, 11, 13, 10, 11, 9, 30, 12])
    luxe_header(ws, "I", "⭐  REVIEW TRACKER",
                "Protect your rating — every review scored by category, with response tracking.")
    table_headers(ws, 4, ["Guest", "Overall", "Cleanliness", "Communication", "Check-In", "Accuracy", "Value", "Comment", "Responded"])
    start = L0
    for i, row in enumerate(rows):
        r = start + i
        for ci, v in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=v)
    end = start + 30 - 1
    style_rows(ws, start, end, 9, text_left={1, 8}, dec={2, 3, 4, 5, 6, 7})
    add_dv(ws, f"I{start}:I{end}", "YesNoList")
    # averages row
    avg = end + 1
    ws.cell(row=avg, column=1, value="AVERAGE").style = "th"
    for col in range(2, 8):
        L = get_column_letter(col)
        c = ws.cell(row=avg, column=col, value=f"=IFERROR(AVERAGE({L}{start}:{L}{end}),0)")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0.0"
    ws.freeze_panes = "A5"
    nrange(wb, "RevOverall", "Reviews", "B", start, end)
    nrange(wb, "RevResponded", "Reviews", "I", start, end)
    cell_name(wb, "RevAvgRow", "Reviews", f"$B${avg}:$G${avg}")
    ws.conditional_formatting.add(f"B{start}:B{end}",
        ColorScaleRule(start_type="num", start_value=3, start_color="FF" + WARN_BG,
                       end_type="num", end_value=5, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 12 — Message Template Library
# ===========================================================================
def build_messages(wb):
    ws = wb.create_sheet("Messages"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 76, 2])
    luxe_header(ws, "D", "💬  MESSAGE TEMPLATE LIBRARY",
                "Reply in seconds, not minutes — polished, reusable templates for every touchpoint.")
    templates = [
        ("Booking Confirmation", "Hi {Guest}! Your booking at {Property} is confirmed for {CheckIn}–{CheckOut}. "
         "We're so glad to host you. I'll send check-in details 24 hours before arrival. — {Host}"),
        ("Welcome Message", "Welcome, {Guest}! Everything you need is in the house guide. Wi-Fi: {WiFi}. "
         "Reach me anytime at {Phone}. Enjoy your stay at {Property}!"),
        ("Check-In Instructions", "Check-in is 4:00 PM. Door code: {Code} (active from 4 PM). Parking: driveway, 2 spots. "
         "The house guide on the counter has everything — text me if anything's unclear."),
        ("Mid-Stay Check-In", "Hi {Guest}, just checking in — is everything comfortable so far? "
         "Happy to help with local recommendations or anything you need."),
        ("Checkout Instructions", "Checkout is 11:00 AM. Before you go: start the dishwasher, bag the trash, "
         "and leave keys on the counter. No need to strip beds. Safe travels, and thank you!"),
        ("Thank You Message", "Thank you for staying at {Property}, {Guest}! It was a pleasure hosting you. "
         "You're welcome back anytime — returning guests get a little something extra. 😊"),
        ("Review Request", "Hi {Guest}, we hope you loved your stay! If you have a moment, a review means the "
         "world to a small host like us. I've already left you a glowing one!"),
        ("Maintenance Notification", "Hi {Guest}, thanks for flagging the {Issue}. I've scheduled {Vendor} to fix it "
         "on {Date}. Apologies for the inconvenience — please let me know if it affects your stay."),
    ]
    row = 4
    ws.cell(row=row, column=2, value="Template").style = "th"
    ws.cell(row=row, column=3, value="Message (edit the {tokens})").style = "th"
    ws.row_dimensions[row].height = 28
    for i, (name, msg) in enumerate(templates):
        r = 5 + i
        ws.cell(row=r, column=2, value=name).style = "field_label"
        ws.cell(row=r, column=3, value=msg).style = "msg"
        ws.row_dimensions[r].height = 54
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW)


# ===========================================================================
# 13 — Supplier Directory
# ===========================================================================
def build_suppliers(wb):
    sample = [
        ("Sparkle Co. — Maria", "Cleaner", "(555) 771-2200", "maria@sparkleco.com", "The Pine Cabin, Loft", "Reliable, same-day turns", 5),
        ("FreshTurn — Ana", "Cleaner", "(555) 660-1188", "ana@freshturn.com", "Beachside, Loft", "Great with beach sand", 5),
        ("Dave's Repairs", "Handyman", "(555) 887-1200", "dave@repairs.com", "All", "Fast, fair pricing", 5),
        ("CoolAir Inc.", "HVAC", "(555) 340-9910", "service@coolair.com", "All", "Annual HVAC contract", 4),
        ("QuickPlumb", "Plumber", "(555) 220-4471", "help@quickplumb.com", "All", "24/7 emergencies", 4),
        ("BrightSpark Electric", "Electrician", "(555) 118-3390", "info@brightspark.com", "All", "Licensed & insured", 5),
        ("GreenScape", "Landscaper", "(555) 909-7712", "hello@greenscape.com", "Cabin, Beachside", "Bi-weekly service", 4),
        ("CleanLinen Svc", "Laundry", "(555) 556-0043", "orders@cleanlinen.com", "All", "Linen pickup Mondays", 5),
    ]
    ws, start, end = build_log(
        wb, "Suppliers", "📇", "SUPPLIER DIRECTORY",
        "Your trusted team on speed-dial — cleaners, trades & emergency contacts, all in one place.",
        ["Name", "Type", "Phone", "Email", "Serves", "Notes", "Rating"],
        sample, [24, 14, 15, 26, 20, 26, 10],
        text_left={1, 4, 5, 6}, dec={7},
        validations=[("B", "SupplierCatList")], reserved=30)
    nrange(wb, "SupName", "Suppliers", "A", start, end)


# ===========================================================================
# 14 — Tax & Expense Preparation
# ===========================================================================
def build_taxes(wb):
    ws = wb.create_sheet("Taxes"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 16, 18, 3, 22, 14, 2])
    luxe_header(ws, "I", "🧾  TAX & EXPENSE PREPARATION",
                "Deduction-ready all year — categorized expenses, quarterly estimates & annual totals.")
    table_headers(ws, 4, ["Deductible Category", "This Month", "YTD", "Receipt Ref"])
    # (category, month, ytd, receipt)
    cats = [
        ("Mortgage Interest", 1180, 8260, "Bank stmt"),
        ("Property Taxes", 520, 3640, "County bill"),
        ("Insurance", 180, 1260, "Policy STR-88120"),
        ("Utilities", 420, 2940, "Folder /util"),
        ("Internet", 95, 665, "ISP invoices"),
        ("Cleaning & Turnover", 960, 6720, "Vendor invoices"),
        ("Supplies & Consumables", 340, 2380, "Receipts /supp"),
        ("Repairs & Maintenance", 250, 1750, "Vendor invoices"),
        ("Software & Subscriptions", 85, 595, "Card stmt"),
        ("Marketing & Photography", 120, 840, "Card stmt"),
        ("Platform / Host Fees", 348, 2436, "Payout reports"),
        ("Depreciation (est.)", 900, 6300, "CPA schedule"),
    ]
    start = L0
    for i, (cat, mo, ytd, rec) in enumerate(cats):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        c1 = ws.cell(row=r, column=2, value=mo); c1.style = "td"; c1.number_format = '"$"#,##0'
        c2 = ws.cell(row=r, column=3, value=ytd); c2.style = "td"; c2.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=rec).style = "td_left"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = start + len(cats)
    ws.cell(row=total, column=1, value="TOTAL DEDUCTIONS").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{total-1})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    # quarterly estimate box
    merge_set(ws, "G4:H4", "QUARTERLY ESTIMATE", "section_gold")
    q = [("Est. taxable income (YTD)", "=FinRevenue*6-C" + str(total), '"$"#,##0'),
         ("Est. tax rate", 0.22, "0%"),
         ("Est. quarterly set-aside", "=MAX(H5*H6/4,0)", '"$"#,##0'),
         ("Next payment due", _next_quarter(), "mm/dd/yyyy")]
    for i, (lab, val, fmt) in enumerate(q):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=val); c.style = "field_value"; c.number_format = fmt
        if lab == "Est. quarterly set-aside":
            ws.cell(row=r, column=8).fill = fill(WARN_BG)
    merge_set(ws, "G10:H10", "Not tax advice — confirm categories with your CPA.", "subtitle")


def _next_quarter():
    today = dt.date.today()
    for md in [(4, 15), (6, 15), (9, 15), (1, 15)]:
        y = today.year if md != (1, 15) else today.year + 1
        d = dt.date(y, md[0], md[1])
        if d >= today:
            return d
    return dt.date(today.year + 1, 1, 15)


# ===========================================================================
# 15 — Multi-Property Dashboard
# ===========================================================================
def build_multiproperty(wb):
    ws = wb.create_sheet("Multi-Property"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 13, 14, 14, 14, 12, 14, 2])
    luxe_header(ws, "I", "🏘  MULTI-PROPERTY DASHBOARD",
                "Compare every listing side by side — occupancy, revenue, profit & rating at a glance.")
    table_headers(ws, 4, ["Property", "Occupancy", "Revenue", "Expenses", "Net Profit", "Rating", "Status"])
    # (name, occ, revenue, expenses, net, rating)
    props = [
        ("The Pine Cabin", 0.80, 5200, 2900, 2300, 4.8),
        ("Beachside Bungalow", 0.63, 3900, 2300, 1600, 4.9),
        ("Downtown Loft", 0.53, 2500, 1550, 950, 4.9),
    ]
    start = L0
    for i, (p, occ, rev, exp, net, rate) in enumerate(props):
        r = start + i
        ws.cell(row=r, column=1, value=p).style = "td_left"
        c = ws.cell(row=r, column=2, value=occ); c.style = "td"; c.number_format = "0%"
        for ci, v in [(3, rev), (4, exp), (5, net)]:
            cc = ws.cell(row=r, column=ci, value=v); cc.style = "td"; cc.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=6, value=rate); cr.style = "td"; cr.number_format = "0.0"
        ws.cell(row=r, column=7, value=f'=IF(B{r}>=0.7,"Strong",IF(B{r}>=0.5,"Steady","Boost"))').style = "td"
        if i % 2:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = start + len(props)
    ws.cell(row=total, column=1, value="PORTFOLIO").style = "th"
    ca = ws.cell(row=total, column=2, value=f"=AVERAGE(B{start}:B{total-1})"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = "0%"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{total-1})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cr = ws.cell(row=total, column=6, value=f"=AVERAGE(F{start}:F{total-1})"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = "0.0"
    nrange(wb, "PropName", "Multi-Property", "B", start, total - 1)
    nrange(wb, "PropOcc", "Multi-Property", "B", start, total - 1)
    nrange(wb, "PropRevenue", "Multi-Property", "C", start, total - 1)
    ws.conditional_formatting.add(f"B{start}:B{total-1}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))


# ===========================================================================
# 16 — Goal Planner
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 14, 12, 14, 14, 2])
    luxe_header(ws, "H", "🎯  GOAL PLANNER",
                "Turn hosting into a business plan — targets, progress bars & milestone tracking.")
    table_headers(ws, 4, ["Goal", "Category", "Target", "Current", "Progress", "Deadline"])
    # (goal, category, target, current, deadline_offset) progress computed
    goals = [
        ("Monthly revenue", "Revenue", 14000, 11600, 120),
        ("Portfolio occupancy", "Occupancy", 80, 66, 90),
        ("Average review score", "Reviews", 50, 48, 60),
        ("Net profit margin", "Profit", 50, 42, 150),
        ("Add 4th property", "Expansion", 100, 45, 300),
        ("Automate messaging", "Operations", 100, 80, 45),
    ]
    start = L0
    for i, (g, cat, tgt, cur, off) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ct = ws.cell(row=r, column=3, value=tgt); ct.style = "input"
        cc = ws.cell(row=r, column=4, value=cur); cc.style = "input"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(MIN(D{r}/C{r},1),0)"); cp.style = "td"; cp.number_format = "0%"
        cd = ws.cell(row=r, column=6, value=dplus(off)); cd.style = "td"; cd.number_format = "mm/dd/yyyy"
        if i % 2:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalCatList")
    nrange(wb, "GoalProgress", "Goals", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    # milestones
    merge_set(ws, "B13:F13", "MONTHLY MILESTONES", "section_gold"); ws.row_dimensions[13].height = 22
    ms = ["Hit 70% portfolio occupancy", "Reach $12k monthly revenue", "Maintain 4.8+ average rating",
          "Zero same-day cancellations", "Photograph refreshed listings"]
    for i, m in enumerate(ms):
        r = 14 + i
        cb = ws.cell(row=r, column=2, value="☐"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=12, color=ACCENT); cb.border = BOX
        merge_set(ws, f"C{r}:F{r}", m, "td_left")


# ===========================================================================
# 17 — Photo & Property Improvement Planner
# ===========================================================================
def build_improvements(wb):
    ws = wb.create_sheet("Improvements"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 20, 14, 3, 22, 20, 14, 2])
    luxe_header(ws, "I", "📸  PHOTO & IMPROVEMENT PLANNER",
                "Better photos book more nights — plan upgrades room by room with budget & priority.")
    sections = ["Living Room", "Primary Bedroom", "Kitchen", "Bathroom", "Outdoor Space", "Seasonal Decor"]
    notes = ["New sofa + throw pillows", "Upgrade to king bed", "Modern faucet & backsplash",
             "Rainfall showerhead", "String lights + fire pit", "Fall wreath & throws"]
    budgets = [1200, 1600, 650, 320, 480, 140]
    prio = ["High", "High", "Medium", "Medium", "Low", "Low"]
    top0 = 5; card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th"); ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+4}", "📷\nPaste photo here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 5):
            ws.row_dimensions[rr].height = 22
        ws.cell(row=row + 5, column=col, value="Improvement").style = "field_label"
        merge_set(ws, f"{M}{row+5}:{R}{row+5}", notes[idx], "field_value")
        ws.cell(row=row + 6, column=col, value="Est. Budget").style = "field_label"
        c = ws.cell(row=row + 6, column=col + 1, value=budgets[idx]); c.style = "field_value"; c.number_format = '"$"#,##0'
        ws.cell(row=row + 6, column=col + 2, value=prio[idx]).style = "field_value"
        ws.cell(row=row + 7, column=col, value="Status").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "Planned", "field_value")


# ===========================================================================
# 18 — Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 18, 3, 16, 12, 12, 2])
    luxe_header(ws, "H", "📈  ANALYTICS DASHBOARD",
                "Your business by the numbers — health across every area, and a single Business Health Score.")
    merge_set(ws, "B5:D5", "BUSINESS HEALTH DIMENSIONS", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Occupancy vs target", "=IFERROR(MIN((SUM(CalNights)/(COUNTA(PropName)*30))/OccTarget,1),0)"),
        ("Profit margin vs target", "=IFERROR(MIN(FinMargin/MarginTarget,1),0)"),
        ("Guest satisfaction", "=IFERROR(AVERAGE(RevOverall)/5,0)"),
        ("Maintenance handled", '=IFERROR(1-(COUNTIF(MaintStatus,"Open")+COUNTIF(MaintStatus,"In Progress"))/MAX(COUNTA(MaintTask),1),0)'),
        ("Inventory stocked", '=IFERROR(1-SUMPRODUCT((InvQty<=InvReorder)*(InvItem<>""))/MAX(COUNTA(InvItem),1),0)'),
        ("Goal progress", "=IFERROR(AVERAGE(GoalProgress),0)"),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.75,"Healthy",IF(C{r}>=0.5,"Watch","Act Now"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(f"C{start}:C{end}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "F5:H5", "BUSINESS HEALTH SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY); cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "Occupancy · margin · guests · maintenance · inventory · goals.", "subtitle")
    ws["F10"].fill = fill(IVORY)
    cell_name(wb, "HealthRange", "Analytics", f"$C${start}:$C${end}")
    bar = BarChart(); bar.type = "bar"; bar.title = "Health by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None; ws.add_chart(bar, "B15")


# ===========================================================================
# 1 — Executive Host Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🏡  AIRBNB HOST COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Revenue, occupancy, bookings, cleaning & business health — your whole rental, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MONTHLY REVENUE", "=FinRevenue", "money"),
        ("NET PROFIT", "=FinNetProfit", "money"),
        ("OCCUPANCY RATE", "=IFERROR(SUM(CalNights)/(COUNTA(PropName)*30),0)", "pct"),
        ("AVG NIGHTLY RATE", "=IFERROR(FinNightly/SUM(CalNights),0)", "rate"),
        ("UPCOMING CHECK-INS", '=COUNTIFS(CalCheckIn,">="&TODAY(),CalCheckIn,"<="&TODAY()+7)', "num"),
        ("UPCOMING CHECK-OUTS", '=COUNTIFS(CalCheckOut,">="&TODAY(),CalCheckOut,"<="&TODAY()+7)', "num"),
    ]
    row2 = [
        ("AVG LENGTH OF STAY", "=IFERROR(SUM(CalNights)/COUNT(CalNights),0)", "dec"),
        ("GUEST RATING", "=IFERROR(AVERAGE(RevOverall),0)", "dec"),
        ("CLEANING PENDING", '=COUNTIF(CleanStatus,"Scheduled")', "num"),
        ("MAINTENANCE DUE", '=COUNTIF(MaintStatus,"Open")+COUNTIF(MaintStatus,"In Progress")', "num"),
        ("LOW-STOCK ITEMS", '=SUMPRODUCT((InvQty<=InvReorder)*(InvItem<>""))', "num"),
        ("BUSINESS HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "PERFORMANCE", "section_gold")
    # revenue by property
    rp = BarChart(); rp.type = "col"; rp.title = "Revenue by Property"; rp.height = 8.2; rp.width = 11.5
    rp.add_data(Reference(wb["Multi-Property"], min_col=3, min_row=4, max_row=7), titles_from_data=True)
    rp.set_categories(Reference(wb["Multi-Property"], min_col=2, min_row=5, max_row=7)); rp.legend = None
    ws.add_chart(rp, "B12")
    # 6-month revenue trend
    tr = LineChart(); tr.title = "6-Month Revenue Trend"; tr.height = 8.2; tr.width = 11.5
    tr.add_data(Reference(wb["Financial"], min_col=12, min_row=6, max_row=12), titles_from_data=True)
    tr.set_categories(Reference(wb["Financial"], min_col=11, min_row=7, max_row=12)); tr.legend = None
    ws.add_chart(tr, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "BREAKDOWN", "section_gold")
    # expense breakdown donut
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Financial"], min_col=6, min_row=6, max_row=6 + len(EXPENSE_CATS) - 1), titles_from_data=False)
    eb.set_categories(Reference(wb["Financial"], min_col=5, min_row=6, max_row=6 + len(EXPENSE_CATS) - 1)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "B30")
    # booking sources donut
    bs = DoughnutChart(); bs.title = "Booking Sources"; bs.height = 8.2; bs.width = 11.5
    bs.add_data(Reference(wb["Calendar"], min_col=12, min_row=5, max_row=7), titles_from_data=False)
    bs.set_categories(Reference(wb["Calendar"], min_col=11, min_row=5, max_row=7)); bs.dataLabels = no_labels()
    ws.add_chart(bs, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Airbnb Host Command Center™ — run your rental like a business. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_property(wb); build_calendar(wb)
    build_reservations(wb); build_financial(wb); build_pricing(wb); build_guests(wb)
    build_cleaning(wb); build_maintenance(wb); build_inventory(wb); build_reviews(wb)
    build_messages(wb); build_suppliers(wb); build_taxes(wb); build_multiproperty(wb)
    build_goals(wb); build_improvements(wb); build_analytics(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Property Profile", "Calendar", "Reservations",
             "Financial", "Pricing", "Guests", "Cleaning", "Maintenance", "Inventory",
             "Reviews", "Messages", "Suppliers", "Taxes", "Multi-Property", "Goals",
             "Improvements", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Airbnb_Host_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
