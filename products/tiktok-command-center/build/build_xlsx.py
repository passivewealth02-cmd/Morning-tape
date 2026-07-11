"""Build TikTok Command Center™ — The Ultimate TikTok Creator Business System.

24 sheets + Welcome · a premium TikTok creator OS in Excel & Sheets.
Ideas, hooks, calendar, trends & sounds, hashtags, analytics, LIVE, TikTok Shop,
affiliate, brand deals (UGC CRM), finance, repurposing, goals & more — one
dashboard.

Run: python3 build_xlsx.py   ->  ../TikTok_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

PILLARS = ["DIY & Crafts", "Home", "Lifestyle", "Hauls", "Tutorials", "Storytime", "Behind-the-scenes"]
CONTENT_STATUS = ["Idea", "Filming", "Editing", "Scheduled", "Posted", "Viral", "Repurposed"]
VIDEO_TYPES = ["Talking Head", "Tutorial", "Trend", "Storytime", "Haul", "GRWM", "Vlog", "LIVE"]
REV_CATS = ["Brand Deals", "TikTok Shop", "Creator Rewards", "Affiliate", "LIVE Gifts", "Digital Products", "Consulting"]
EXP_CATS = ["Software", "Equipment", "Props & Materials", "Contractors", "Marketing", "Education", "Miscellaneous"]
SPON_STAGES = ["Lead", "Outreach", "Negotiation", "Signed", "Delivered", "Paid"]
HOOK_TYPES = ["Question", "Bold Claim", "POV", "Listicle", "Story", "Controversy", "Relatable"]
TREND_STATUS = ["Rising", "Peaking", "Fading", "Evergreen"]
GOAL_CATS = ["Followers", "Views", "Revenue", "Posting", "Brand Deals", "Shop", "Engagement"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


def totals(ws, row, cols, start, end, fmt='"$"#,##0', label="TOTAL"):
    ws.cell(row=row, column=1, value=label).style = "th"
    for col in cols:
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt


# ===========================================================================
# 24 — Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your account details & targets once — every dashboard follows.")
    merge_set(ws, "B5:C5", "CREATOR INPUTS", "section")
    controls = [
        ("Creator / Handle", "@vellacreates", None, "Handle"),
        ("Creator Name", "Vella Rowe", None, "CreatorName"),
        ("Niche", "DIY, home & lifestyle", None, "Niche"),
        ("Monthly Revenue Goal", 10000, '"$"#,##0', "RevenueGoal"),
        ("Monthly Post Goal", 24, "0", "PostGoal"),
        ("Active Deal Target", 5, "0", "DealTarget"),
        ("Completion Target", 0.55, "0%", "AVDTarget"),
        ("Follower Growth Goal (mo)", 20000, "#,##0", "GrowthGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Content Pillar", PILLARS, "PillarList"), ("F", "Video Type", VIDEO_TYPES, "VideoTypeList"),
             ("G", "Status", CONTENT_STATUS, "StatusList"), ("H", "Revenue Category", REV_CATS, "RevCatList"),
             ("I", "Expense Category", EXP_CATS, "ExpCatList"), ("J", "Hook Type", HOOK_TYPES, "HookTypeList"),
             ("K", "Goal Category", GOAL_CATS, "GoalCatList"), ("L", "Deal Stage", SPON_STAGES, "SponStageList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 16, "Trend Status", TREND_STATUS, "TrendStatusList"), ("F", 16, "Priority", PRIORITIES, "PriorityList"),
             ("G", 16, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🎵  TIKTOK COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate TikTok creator business system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "RUN YOUR ENTIRE TIKTOK BUSINESS FROM ONE FILE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("From hook to payout — TikTok Command Center™ manages ideas, hooks, your posting calendar, trends "
                      "& sounds, hashtags, analytics, LIVE, TikTok Shop, affiliate, brand deals and finances in ONE "
                      "premium Excel & Google Sheets system. Post more consistently, ride trends earlier, grow faster, "
                      "land brand deals and turn views into real income — all with creator-grade automation. This isn't "
                      "a content calendar — it's your complete TikTok Operating System.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your handle, niche & monthly goals.",
             "2.  Fill the Content Calendar & Pipeline — plan and move videos idea → posted.",
             "3.  Bank Ideas & Hooks; track Trends & Sounds while they're still rising.",
             "4.  Log Analytics, LIVE, Shop & Affiliate — revenue & net profit update live.",
             "5.  Work Brand Deals from lead → paid in the UGC CRM.",
             "6.  Watch the Dashboard track followers, revenue & a Creator Health Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (@vellacreates, a 215k-follower DIY & lifestyle creator earning ~$8,600/mo) is "
               "included so you can see how everything connects — just type over it with your own. Followers, revenue, "
               "net profit, engagement, posting consistency, the brand-deal pipeline and the Creator Health Score all "
               "update automatically. Every sheet is print-friendly and works in Excel and Google Sheets, on desktop "
               "and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more time to create — let's grow your account.", "section_gold")


# ===========================================================================
# 2 — Creator Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "👤  CREATOR PROFILE", "Your account, defined — the identity every video flows from.")
    blocks = [
        ("THE ACCOUNT", [("Handle", "=Handle"), ("Creator", "=CreatorName"),
                         ("Niche", "=Niche"), ("Posting Schedule", "Daily, 6pm & 8pm ET"),
                         ("Started", "2022"), ("Business Email", "hi@vellacreates.co")]),
        ("STRATEGY", [("Mission", "Make DIY feel doable"), ("Target Audience", "Women 22-38, first homes"),
                      ("Content Pillars", "DIY · Home · Lifestyle"), ("Manager", "Self-managed"),
                      ("Media Kit", "vellacreates.co/kit"), ("Rate Card", "$1.8k / integrated video")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "PLATFORMS & LINKS", "section_gold"); ws.row_dimensions[15].height = 22
    handles = [("TikTok", "@vellacreates"), ("Instagram", "@vella.creates"), ("YouTube", "Vella Creates"),
               ("Pinterest", "@vellacreates"), ("Newsletter", "The Weekend Project"), ("LinkedIn", "Vella Rowe")]
    for i, (p, h) in enumerate(handles):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 3 — Content Calendar
# ===========================================================================
def build_calendar(wb):
    titles = [
        "$12 thrift flip glow-up", "5 renter-friendly kitchen hacks", "IKEA dresser makeover",
        "POV: your first apartment", "DIY floating shelves under $30", "Sunday reset routine",
        "Thrifted mirror restoration", "Amazon home finds I actually use", "Peel-and-stick tile test",
        "Small-space storage wins", "Painting my kitchen cabinets", "Cozy corner makeover",
        "Dollar-store DIY vase", "How I style open shelves", "Renter hacks landlords hate",
        "Closet organization on a budget", "DIY headboard from a pallet", "My cleaning caddy tour",
        "Fixing a wobbly chair", "Gallery wall the easy way", "Thrift with me: goodwill run",
        "Kitchen drawer reorganize", "Faux built-ins hack", "Balcony makeover reveal",
        "3 trending sounds this week", "Weekend project: bench", "Before/after: living room",
    ]
    posted_off = [1, 2, 3, 5, 6, 7, 8, 10, 11, 12, 13, 15, 16, 17, 19, 20, 21, 22, 24, 25, 26, 27]
    viral_idx = {0, 7, 17}
    future = [(-1, "Scheduled"), (-2, "Scheduled"), (-3, "Scheduled"), (-5, "Filming"), (-6, "Editing"), (-8, "Idea")]
    rows = []
    ti = 0
    for k, off in enumerate(posted_off):
        status = "Viral" if k in viral_idx else "Posted"
        rows.append((dminus(off), titles[ti % len(titles)], PILLARS[ti % 4], VIDEO_TYPES[ti % len(VIDEO_TYPES)], status)); ti += 1
    for foff, status in future:
        rows.append((dplus(-foff), titles[ti % len(titles)], PILLARS[ti % 4], VIDEO_TYPES[ti % len(VIDEO_TYPES)], status)); ti += 1
    sample = [(d, t, p, vt, "High" if st in ("Posted", "Viral", "Scheduled") else "Medium", st) for (d, t, p, vt, st) in rows]
    ws, start, end = build_log(
        wb, "Calendar", "🗓", "CONTENT CALENDAR",
        "Plan every post in one view — posting status calculates itself.",
        ["Post Date", "Title / Concept", "Pillar", "Type", "Priority", "Status"],
        sample, [13, 32, 16, 14, 11, 13],
        text_left={2}, dates={1},
        validations=[("C", "PillarList"), ("D", "VideoTypeList"), ("E", "PriorityList"), ("F", "StatusList")], reserved=60)
    nrange(wb, "CalDate", "Calendar", "A", start, end)
    nrange(wb, "CalStatus", "Calendar", "F", start, end)
    cmap = {"Viral": HIGHLIGHT, "Posted": MINT_BG, "Scheduled": WARN_BG, "Editing": SOFT_BG, "Filming": SOFT_BG, "Idea": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 4 — Video Pipeline
# ===========================================================================
def build_pipeline(wb):
    rows = [
        ("Painting my kitchen cabinets", "Editing", 0.80, "Vella", 1, "Final cut + captions"),
        ("Faux built-ins hack", "Filming", 0.55, "Vella", 3, "B-roll left to shoot"),
        ("Balcony makeover reveal", "Script", 0.35, "Vella", 5, "Hook drafted"),
        ("Weekend project: bench", "Outline", 0.20, "Vella", 6, "Sourcing wood"),
        ("Thrift with me #12", "Idea", 0.10, "Vella", 8, "Wait for goodwill restock"),
        ("Renter hacks part 3", "Editing", 0.70, "Editor — Jo", 2, "Trim to 34s"),
        ("Amazon home finds v4", "Thumbnail", 0.90, "Vella", 1, "Cover frame chosen"),
        ("DIY headboard pallet", "Filming", 0.50, "Vella", 4, "Half filmed"),
    ]
    sample = [(t, st, prog, o, dplus(d), note) for (t, st, prog, o, d, note) in rows]
    ws, start, end = build_log(
        wb, "Pipeline", "🎬", "VIDEO PIPELINE",
        "Every video, every stage — from idea to posted, with live progress %.",
        ["Title", "Stage", "Progress", "Owner", "Due", "Notes"],
        sample, [30, 14, 12, 16, 13, 22],
        text_left={1, 4, 6}, dates={5}, pcts={3}, reserved=40)
    nrange(wb, "PipeProgress", "Pipeline", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 5 — Idea Vault
# ===========================================================================
def build_ideas(wb):
    rows = [
        ("Renovating my rental kitchen", "Home", "renter kitchen", "My landlord said yes to THIS", 900000, 6, "Planned"),
        ("$0 decor from my recycling", "DIY & Crafts", "upcycle decor", "I made this from trash", 1400000, 4, "Idea"),
        ("Everything wrong with my first apartment", "Storytime", "first apartment", "Do NOT do what I did", 700000, 3, "Idea"),
        ("Testing viral cleaning hacks", "Lifestyle", "cleaning hacks", "TikTok lied to me", 1100000, 5, "Planned"),
        ("Full closet makeover under $50", "Home", "closet makeover", "$50 closet glow-up", 800000, 6, "Idea"),
        ("DIY vs Store: same look, less $", "Hauls", "diy vs store", "Save $200 doing this", 1300000, 4, "Planned"),
        ("Painting tips nobody tells you", "Tutorials", "painting tips", "Stop painting like this", 600000, 5, "Idea"),
        ("A week of weekend projects", "DIY & Crafts", "weekend project", "7 projects, 7 days", 950000, 7, "Idea"),
        ("Reacting to my old videos", "Behind-the-scenes", "creator journey", "My first video was rough", 500000, 3, "Idea"),
        ("Best Amazon finds under $25", "Hauls", "amazon home", "Cart these before they sell out", 1500000, 5, "Idea"),
    ]
    ws = wb.create_sheet("Ideas"); ws.sheet_view.showGridLines = False
    set_widths(ws, [30, 16, 18, 30, 14, 9, 12, 12])
    luxe_header(ws, "H", "💡  IDEA VAULT",
                "Never run dry — capture ideas and let an AI Opportunity Score (views ÷ difficulty) rank them.")
    table_headers(ws, 4, ["Idea", "Pillar", "Keyword", "Hook", "Est. Views", "Diff.", "AI Score", "Status"])
    start = L0
    for i, (idea, pil, kw, hook, views, diff, status) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=idea).style = "td_left"
        ws.cell(row=r, column=2, value=pil).style = "td"
        ws.cell(row=r, column=3, value=kw).style = "td_left"
        ws.cell(row=r, column=4, value=hook).style = "td_left"
        cv = ws.cell(row=r, column=5, value=views); cv.style = "td"; cv.number_format = "#,##0"
        ws.cell(row=r, column=6, value=diff).style = "td"
        sc = ws.cell(row=r, column=7, value=f"=IFERROR(ROUND(E{r}/1000/F{r},0),0)"); sc.style = "td"; sc.number_format = "#,##0"
        ws.cell(row=r, column=8, value=status).style = "td"
        if i % 2:
            for c in range(1, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    style_rows(ws, end + 1, end + 40, 8, text_left={1, 3, 4})
    for r in range(end + 1, end + 41):
        ws.cell(row=r, column=7, value=f"=IFERROR(ROUND(E{r}/1000/F{r},0),0)").number_format = "#,##0"
    add_dv(ws, f"B{start}:B{end+40}", "PillarList"); add_dv(ws, f"H{start}:H{end+40}", "StatusList")
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"G{start}:G{end}", ColorScaleRule(
        start_type="num", start_value=100, start_color="FF" + WARN_BG,
        end_type="num", end_value=400, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 6 — Hook & Script Bank
# ===========================================================================
def build_hooks(wb):
    rows = [
        ("I made this for $0", "Bold Claim", "$0 decor from recycling", "Yes", 9.4, "Stops the scroll cold"),
        ("Do NOT do what I did", "Controversy", "first apartment mistakes", "Yes", 9.1, "Mistake hooks over-perform"),
        ("POV: your landlord said yes", "POV", "renter kitchen", "Yes", 8.8, "Aspirational + relatable"),
        ("3 hacks that saved my deposit", "Listicle", "renter hacks", "Yes", 8.6, "Numbers = clear promise"),
        ("TikTok lied to me about this", "Controversy", "cleaning hacks", "No", 8.9, "Curiosity + test format"),
        ("Save $200 doing this instead", "Bold Claim", "diy vs store", "Yes", 8.7, "Specific $ beats vague"),
        ("Would you keep or toss this?", "Question", "thrift flip", "No", 7.8, "Drives comments"),
        ("My first video was this bad", "Story", "creator journey", "No", 7.5, "Vulnerability builds trust"),
    ]
    ws, start, end = build_log(
        wb, "Hooks", "🪝", "HOOK & SCRIPT BANK",
        "The first 2 seconds win — bank your best hooks, rated by how they performed.",
        ["Hook (first line)", "Type", "Best For", "Reusable?", "Score", "Why It Works"],
        rows, [30, 14, 22, 12, 10, 26],
        text_left={1, 3, 6}, dec={5},
        validations=[("B", "HookTypeList"), ("D", "YesNoList")], reserved=40)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color=GOLD_LT, showValue=True))


# ===========================================================================
# 7 — Trends & Sounds
# ===========================================================================
def build_trends(wb):
    rows = [
        ("'aesthetic morning' sound", "Sound", "Rising", dminus(2), "Yes", "DIY morning routine", "Use in next 3 days"),
        ("#renterfriendly", "Hashtag", "Peaking", dminus(10), "Yes", "Renter hacks", "70M+ views"),
        ("Slow-zoom reveal transition", "Effect", "Evergreen", dminus(60), "Yes", "Before/after reveals", "Always works"),
        ("'oddly satisfying' remix", "Sound", "Peaking", dminus(5), "No", "Cleaning / organizing", "Jump on this week"),
        ("#diyonabudget", "Hashtag", "Rising", dminus(3), "Yes", "Budget projects", "Low competition"),
        ("Green-screen price tag", "Effect", "Rising", dminus(4), "No", "Haul / price reveals", "Test on Amazon finds"),
        ("'that girl' voiceover trend", "Sound", "Fading", dminus(20), "No", "Lifestyle", "Skip — past peak"),
        ("#hometok", "Hashtag", "Evergreen", dminus(90), "Yes", "All home content", "Core niche tag"),
    ]
    ws, start, end = build_log(
        wb, "Trends", "🔥", "TRENDS & SOUNDS",
        "Ride the wave early — track sounds, hashtags & effects while they're still rising.",
        ["Trend", "Kind", "Status", "Spotted", "Used?", "Best For", "Notes"],
        rows, [26, 12, 12, 13, 10, 22, 22],
        text_left={1, 6, 7}, dates={4},
        validations=[("C", "TrendStatusList"), ("E", "YesNoList")], reserved=30)
    for st, cc in {"Rising": MINT_BG, "Peaking": WARN_BG, "Fading": RED_BG, "Evergreen": SOFT_BG}.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 8 — Hashtag & SEO
# ===========================================================================
def build_hashtags(wb):
    rows = [
        ("DIY set", "#diy #diyhome #diyproject #upcycle #handmade", "DIY & Crafts", "Medium", "Mix big + niche"),
        ("Renter set", "#renterfriendly #renthack #apartmenttherapy #firstapartment", "Home", "Low", "High intent, low comp"),
        ("Budget set", "#diyonabudget #budgetdecor #savemoney #thrifted", "Hauls", "Low", "Great reach:comp ratio"),
        ("Home set", "#hometok #homedecor #smallspace #homeinspo", "Home", "High", "Core but competitive"),
        ("Cleaning set", "#cleantok #cleaninghacks #satisfying #organize", "Lifestyle", "High", "Broad discovery"),
        ("Tutorial set", "#howto #diytutorial #tutorial #learnontiktok", "Tutorials", "Medium", "Boosts saves"),
    ]
    ws, start, end = build_log(
        wb, "Hashtags", "#️⃣", "HASHTAG & SEO",
        "Get found — pre-built hashtag sets by pillar, balancing reach and competition.",
        ["Set", "Hashtags", "Pillar", "Competition", "Notes"],
        rows, [16, 42, 16, 14, 24],
        text_left={2, 5}, reserved=20,
        validations=[("C", "PillarList")])


# ===========================================================================
# 9 — Analytics
# ===========================================================================
def build_analytics(wb):
    rows = [
        ("$12 thrift flip glow-up", 3200000, 0.128, 0.61, 9200, 4200),
        ("Amazon home finds I use", 1900000, 0.104, 0.54, 6100, 8800),
        ("5 renter kitchen hacks", 1450000, 0.112, 0.58, 5400, 2100),
        ("IKEA dresser makeover", 980000, 0.096, 0.49, 3300, 1600),
        ("Peel-and-stick tile test", 620000, 0.081, 0.44, 1800, 900),
        ("Sunday reset routine", 410000, 0.073, 0.41, 1100, 300),
    ]
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 3, 18, 12, 12, 2])
    luxe_header(ws, "G", "📊  ANALYTICS COMMAND CENTER",
                "Your account by the numbers — a live snapshot, health dimensions & the Creator Health Score.")
    # snapshot
    merge_set(ws, "B5:C5", "SNAPSHOT (28 DAYS)", "section")
    snap = [("Followers", 215000, "#,##0", "FollowerNow"), ("Followers gained", 18400, "#,##0", "FollowerGrowth"),
            ("Views", 3200000, "#,##0", "Views28"), ("Likes", 410000, "#,##0", "Likes28"),
            ("Engagement rate", 0.092, "0.0%", "EngRate"), ("Avg completion", 0.48, "0%", "Completion"),
            ("Profile visits", 142000, "#,##0", "ProfileVisits"), ("Shares", 88000, "#,##0", "Shares28")]
    for i, (lab, val, fmt, nm) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "field_value"; c.number_format = fmt
        cell_name(wb, nm, "Analytics", f"$C${r}")
        if lab in ("Followers", "Engagement rate"):
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    # health dims
    merge_set(ws, "E5:G5", "CREATOR HEALTH", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=5)
    metrics = [
        ("Revenue vs goal", "=IFERROR(MIN(RevenueTotal/RevenueGoal,1),0)"),
        ("Posting consistency", '=IFERROR(MIN((COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Posted")+COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Viral"))/PostGoal,1),0)'),
        ("Engagement", "=IFERROR(MIN(EngRate/0.1,1),0)"),
        ("Retention", "=IFERROR(MIN(Completion/AVDTarget,1),0)"),
        ("Brand pipeline", '=IFERROR(MIN((COUNTIF(SponStage,"Signed")+COUNTIF(SponStage,"Delivered")+COUNTIF(SponStage,"Negotiation"))/DealTarget,1),0)'),
        ("Follower growth", "=IFERROR(MIN(FollowerGrowth/GrowthGoal,1),0)"),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(metrics):
        r = hs + i
        ws.cell(row=r, column=5, value=dim).style = "td_left"
        c = ws.cell(row=r, column=6, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=7, value=f'=IF(F{r}>=0.75,"Strong",IF(F{r}>=0.5,"Growing","Focus"))').style = "td"
        if i % 2:
            for c2 in range(5, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(metrics) - 1
    cell_name(wb, "HealthRange", "Analytics", f"$F${hs}:$F${he}")
    ws.conditional_formatting.add(f"F{hs}:F{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # per-video table
    merge_set(ws, "B15:G15", "TOP VIDEOS (28 DAYS)", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Video", "Views", "Eng. Rate", "Completion", "Follows +", "Shop Clicks"], start_col=2)
    vs = 17
    for i, (title, views, eng, comp, foll, clicks) in enumerate(rows):
        r = vs + i
        ws.cell(row=r, column=2, value=title).style = "td_left"
        cv = ws.cell(row=r, column=3, value=views); cv.style = "td"; cv.number_format = "#,##0"
        ce = ws.cell(row=r, column=4, value=eng); ce.style = "td"; ce.number_format = "0.0%"
        cc = ws.cell(row=r, column=5, value=comp); cc.style = "td"; cc.number_format = "0%"
        cf = ws.cell(row=r, column=6, value=foll); cf.style = "td"; cf.number_format = "#,##0"
        ck = ws.cell(row=r, column=7, value=clicks); ck.style = "td"; ck.number_format = "#,##0"
        if i % 2:
            for c2 in range(2, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    ve = vs + len(rows) - 1
    nrange(wb, "VidTitle", "Analytics", "B", vs, ve)
    nrange(wb, "VidViews", "Analytics", "C", vs, ve)
    ws.conditional_formatting.add(f"C{vs}:C{ve}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=3200000, color=PRIMARY, showValue=True))
    # follower growth (feeds dashboard line)
    merge_set(ws, "B25:C25", "FOLLOWER GROWTH (K) — 6 MONTHS", "section")
    ws.cell(row=26, column=2, value="Month").style = "th"; ws.cell(row=26, column=3, value="Followers (K)").style = "th"
    months = _recent_months(6); vals = [128, 149, 168, 184, 201, 215]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 27 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = "0"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "FollowMonth", "Analytics", "$B$27:$B$32")
    cell_name(wb, "FollowVal", "Analytics", "$C$27:$C$32")


# ===========================================================================
# 10 — LIVE Sessions
# ===========================================================================
def build_live(wb):
    rows = [
        (dminus(3), "DIY Q&A + thrift haul", 62, 1840, 5200, 128000, 210),
        (dminus(10), "Painting cabinets live", 95, 2600, 8100, 194000, 340),
        (dminus(17), "Sunday reset co-work", 48, 1120, 2900, 76000, 96),
        (dminus(24), "Amazon cart party", 74, 2100, 6400, 152000, 280),
    ]
    ws, start, end = build_log(
        wb, "LIVE", "🔴", "LIVE SESSIONS",
        "Go live with a plan — track duration, viewers, gifts & new followers.",
        ["Date", "Theme", "Mins", "Peak Viewers", "Diamonds", "Total Views", "Follows +"],
        rows, [13, 26, 10, 14, 12, 14, 12],
        text_left={2}, dates={1}, ints={3, 4, 5, 6, 7}, reserved=20)
    ws.cell(row=end + 2, column=2, value="Tip: ~$0.005 per diamond — LIVE gifts add up fast at scale.").font = Font(italic=True, color=ACCENT)


# ===========================================================================
# 11 — TikTok Shop
# ===========================================================================
def build_shop(wb):
    rows = [
        ("Cordless glue gun", "Tools", 24.99, 0.15, 88, "=C5*E5", "=F5*D5"),
        ("Peel-and-stick tile (10pk)", "Home", 32.00, 0.12, 58, "=C6*E6", "=F6*D6"),
        ("Command strip mega-pack", "Home", 18.50, 0.10, 130, "=C7*E7", "=F7*D7"),
        ("Mini paint sprayer", "Tools", 89.00, 0.18, 19, "=C8*E8", "=F8*D8"),
        ("Drawer organizer set", "Home", 27.99, 0.15, 54, "=C9*E9", "=F9*D9"),
        ("LED strip lights", "Lighting", 21.00, 0.12, 72, "=C10*E10", "=F10*D10"),
        ("Label maker", "Tools", 39.99, 0.15, 36, "=C11*E11", "=F11*D11"),
        ("Microfiber cloth (24pk)", "Cleaning", 14.99, 0.10, 108, "=C12*E12", "=F12*D12"),
    ]
    ws = wb.create_sheet("Shop"); ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 14, 12, 12, 11, 13, 13])
    luxe_header(ws, "G", "🛍  TIKTOK SHOP",
                "Turn views into sales — units, GMV and your commission, per product.")
    table_headers(ws, 4, ["Product", "Category", "Price", "Comm. %", "Units", "GMV", "Your Earnings"])
    start = L0
    for i, (prod, cat, price, comm, units, gmv, earn) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=prod).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=4, value=comm); cm.style = "input"; cm.number_format = "0%"
        cu = ws.cell(row=r, column=5, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cg = ws.cell(row=r, column=6, value=gmv); cg.style = "td"; cg.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=7, value=earn); ce.style = "td"; ce.number_format = '"$"#,##0'
        if i % 2:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL").style = "th"
    for c in range(2, 6):
        ws.cell(row=tot, column=c).style = "th"
    cg = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(SURFACE); cg.number_format = '"$"#,##0'
    ce = ws.cell(row=tot, column=7, value=f"=SUM(G{start}:G{end})"); ce.style = "td"; ce.font = Font(bold=True, color=PRIMARY); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    cell_name(wb, "ShopGMV", "Shop", f"$F${tot}")
    cell_name(wb, "ShopEarn", "Shop", f"$G${tot}")
    ws.conditional_formatting.add(f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=6000, color=PRIMARY, showValue=True))


# ===========================================================================
# 12 — Affiliate Tracker
# ===========================================================================
def build_affiliate(wb):
    rows = [
        ("Amazon Storefront", "Amazon", "amzn.to/vella", 12400, 640, "$980", "Home & tools bundle"),
        ("LTK", "LTK", "ltk.to/vella", 5200, 210, "$420", "Outfit + decor links"),
        ("Home Depot Affiliate", "Impact", "hd.co/vella", 3100, 96, "$310", "Big-ticket DIY"),
        ("Etsy Creator", "Etsy", "etsy.me/vella", 1800, 74, "$180", "Handmade props"),
        ("Skillshare", "Impact", "skl.sh/vella", 2400, 38, "$260", "DIY classes"),
    ]
    ws, start, end = build_log(
        wb, "Affiliate", "🔗", "AFFILIATE TRACKER",
        "Every link working for you — clicks, conversions & payouts by program.",
        ["Program", "Network", "Link", "Clicks", "Sales", "Payout", "Notes"],
        rows, [22, 12, 18, 12, 10, 12, 24],
        text_left={3, 7}, ints={4, 5},
        reserved=20)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=13000, color=GOLD_LT, showValue=True))


# ===========================================================================
# 13 — Brand Deals (UGC CRM)
# ===========================================================================
def build_deals(wb):
    rows = [
        ("Glossier", "Mara D.", "Spring GRWM", 2800, "Signed", "1 video + 3 stories", 12),
        ("Amazon Home", "Rep", "Prime Day haul", 3500, "Negotiation", "2 videos", 30),
        ("Notion", "Lena R.", "Creator template", 2200, "Delivered", "1 tutorial", -5),
        ("Chomps", "Ivy T.", "Snack integration", 1500, "Paid", "1 video", -25),
        ("Djerf Avenue", "Priya S.", "Home edit", 2600, "Outreach", "TBD", 60),
        ("Ruggable", "Owen B.", "Rug reveal", 1900, "Signed", "1 before/after", 18),
        ("HelloFresh", "Sam W.", "Weeknight dinner", 1100, "Lead", "TBD", 75),
        ("Dyson", "Chris N.", "Cleaning day", 3200, "Paid", "1 video + LIVE", -40),
    ]
    sample = [(b, c, camp, rate, stage, deliv, dplus(d) if d >= 0 else dminus(-d)) for (b, c, camp, rate, stage, deliv, d) in rows]
    ws, start, end = build_log(
        wb, "Brand Deals", "🤝", "BRAND DEALS · UGC CRM",
        "Turn your audience into income — lead to paid, with rates & deliverables.",
        ["Brand", "Contact", "Campaign", "Rate", "Stage", "Deliverables", "Due / Paid"],
        sample, [16, 14, 20, 12, 14, 20, 13],
        text_left={2, 3, 6}, money={4}, dates={7},
        validations=[("E", "SponStageList")], reserved=30)
    nrange(wb, "SponRate", "Brand Deals", "D", start, end)
    nrange(wb, "SponStage", "Brand Deals", "E", start, end)
    for st, cc in {"Paid": MINT_BG, "Signed": WARN_BG, "Delivered": SOFT_BG, "Lead": WHITE}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 14 — Series & Playlists
# ===========================================================================
def build_series(wb):
    rows = [
        ("Renter Hacks", 14, 8900000, 0.58, "Weekly", "Signature series — pin it"),
        ("Thrift Flips", 11, 6200000, 0.61, "Weekly", "Highest completion"),
        ("$50 Makeovers", 8, 4100000, 0.55, "Bi-weekly", "Great for brand deals"),
        ("Weekend Projects", 9, 3300000, 0.52, "Weekly", "Drives saves"),
        ("Amazon Finds", 12, 7400000, 0.49, "Weekly", "Best for Shop"),
        ("Storytime", 6, 2100000, 0.47, "Monthly", "Builds connection"),
    ]
    ws, start, end = build_log(
        wb, "Series", "📺", "SERIES & PLAYLISTS",
        "Turn one-off hits into a habit — group videos into series that keep viewers coming back.",
        ["Series", "Videos", "Total Views", "Avg Completion", "Cadence", "Notes"],
        rows, [22, 10, 14, 14, 13, 26],
        text_left={1, 6}, ints={2, 3}, pcts={4}, reserved=20)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=9000000, color=PRIMARY, showValue=True))


# ===========================================================================
# 15 — Finance Center
# ===========================================================================
def build_finance(wb):
    ws = wb.create_sheet("Finance"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 16, 12, 3, 22, 14])
    luxe_header(ws, "G", "💰  CREATOR FINANCE CENTER",
                "Every income & expense in one place — monthly, run-rate & net profit, live.")
    table_headers(ws, 4, ["Income Source", "This Month", "Annual (est.)", "% of Rev"])
    income = {"Brand Deals": 3800, "TikTok Shop": 1900, "Creator Rewards": 1200, "Affiliate": 780,
              "LIVE Gifts": 420, "Digital Products": 400, "Consulting": 100}
    start = L0; iend = start + len(REV_CATS) - 1
    for i, cat in enumerate(REV_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        val = "=ShopEarn" if cat == "TikTok Shop" else income[cat]
        cm = ws.cell(row=r, column=2, value=val); cm.style = "input" if cat != "TikTok Shop" else "field_value"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/$B${iend+1},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    itot = iend + 1
    ws.cell(row=itot, column=1, value="TOTAL REVENUE").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=itot, column=col, value=f"=SUM({L}{start}:{L}{iend})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "RevSource", "Finance", "A", start, iend)
    nrange(wb, "RevMonthly", "Finance", "B", start, iend)
    cell_name(wb, "RevenueTotal", "Finance", f"$B${itot}")
    # expenses
    merge_set(ws, "F4:G4", "MONTHLY EXPENSES", "section_gold")
    expenses = {"Software": 180, "Equipment": 520, "Props & Materials": 400, "Contractors": 500,
                "Marketing": 120, "Education": 90, "Miscellaneous": 140}
    estart = 5
    for i, cat in enumerate(EXP_CATS):
        r = estart + i
        ws.cell(row=r, column=6, value=cat).style = "td_left"
        c = ws.cell(row=r, column=7, value=expenses[cat]); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=6).fill = fill(MUTED_ROW); ws.cell(row=r, column=7).fill = fill(MUTED_ROW)
    eend = estart + len(EXP_CATS) - 1; etot = eend + 1
    ws.cell(row=etot, column=6, value="TOTAL EXPENSES").style = "th"
    c = ws.cell(row=etot, column=7, value=f"=SUM(G{estart}:G{eend})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "ExpCat", "Finance", "F", estart, eend)
    nrange(wb, "ExpMonthly", "Finance", "G", estart, eend)
    cell_name(wb, "ExpenseTotal", "Finance", f"$G${etot}")
    # bottom line
    merge_set(ws, "A16:D16", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Revenue (month)", "=RevenueTotal", '"$"#,##0'), ("Annual run-rate", "=RevenueTotal*12", '"$"#,##0'),
             ("Expenses (month)", "=ExpenseTotal", '"$"#,##0'), ("Net profit", "=RevenueTotal-ExpenseTotal", '"$"#,##0'),
             ("Profit margin", "=IFERROR((RevenueTotal-ExpenseTotal)/RevenueTotal,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 17 + i
        ws.cell(row=r, column=1, value=lab).style = "field_label"
        c = ws.cell(row=r, column=2, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net profit", "Profit margin"):
            ws.cell(row=r, column=2).fill = fill(MINT_BG)


# ===========================================================================
# 16 — Expenses
# ===========================================================================
def build_expenses(wb):
    rows = [
        (dminus(20), "CapCut Pro", "Software", "$18", "Editing", "Monthly"),
        (dminus(16), "Ring light + tripod", "Equipment", "$140", "Gear", "One-time"),
        (dminus(12), "Paint & supplies", "Props & Materials", "$180", "DIY content", "Per project"),
        (dminus(9), "Editor — Jo", "Contractors", "$500", "5 videos", "Monthly"),
        (dminus(6), "Thrift hauls (content)", "Props & Materials", "$120", "B-roll", "Ongoing"),
        (dminus(4), "Boost — top video", "Marketing", "$120", "Promote", "Test"),
        (dminus(2), "DIY masterclass", "Education", "$90", "Skill-up", "One-time"),
    ]
    ws, start, end = build_log(
        wb, "Expenses", "🧾", "EXPENSES",
        "Every business cost tracked — because creator income is a business.",
        ["Date", "Item", "Category", "Amount", "For", "Frequency"],
        rows, [13, 22, 18, 12, 18, 14],
        text_left={2, 5}, dates={1}, reserved=30,
        validations=[("C", "ExpCatList")])


# ===========================================================================
# 17 — Equipment
# ===========================================================================
def build_equipment(wb):
    rows = [
        ("iPhone 15 Pro", "Camera", dminus(300), "Good", "Main camera", "Yes"),
        ("Ring light 18\"", "Lighting", dminus(200), "Good", "Key light", "Yes"),
        ("Rode mic", "Audio", dminus(150), "Good", "Voiceover", "Yes"),
        ("Tripod + phone mount", "Support", dminus(400), "Fair", "Replace mount soon", "No"),
        ("Softbox kit", "Lighting", dminus(90), "New", "B-roll", "Yes"),
        ("MacBook Air", "Editing", dminus(500), "Good", "CapCut / edits", "Yes"),
        ("Backdrop stand", "Set", dminus(120), "Good", "Talking head", "Yes"),
    ]
    ws, start, end = build_log(
        wb, "Equipment", "🎥", "EQUIPMENT",
        "Your kit, tracked — condition & what to upgrade next.",
        ["Item", "Type", "Bought", "Condition", "Notes", "Working?"],
        rows, [22, 14, 13, 12, 22, 12],
        text_left={5}, dates={3}, reserved=20,
        validations=[("F", "YesNoList")])
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 18 — Repurposing
# ===========================================================================
def build_repurpose(wb):
    rows = [
        ("$12 thrift flip glow-up", "Yes", "Yes", "Yes", "Yes", "No", "Top performer — push everywhere"),
        ("Amazon home finds I use", "Yes", "Yes", "No", "Yes", "Yes", "Blog post + email"),
        ("5 renter kitchen hacks", "Yes", "No", "Yes", "No", "No", "Carousel next"),
        ("IKEA dresser makeover", "Yes", "Yes", "Yes", "No", "Yes", "Pinterest pin ready"),
        ("Painting cabinets live", "No", "No", "No", "Yes", "No", "Cut LIVE into clips"),
    ]
    ws, start, end = build_log(
        wb, "Repurposing", "♻", "REPURPOSING ENGINE",
        "One video → ten posts — track where each piece of content has been repurposed.",
        ["Original Video", "Reels", "Shorts", "Pinterest", "Story", "Newsletter", "Notes"],
        rows, [26, 10, 10, 12, 10, 13, 26],
        text_left={1, 7}, reserved=30,
        validations=[(c, "YesNoList") for c in ("B", "C", "D", "E", "F")])
    for col in ("B", "C", "D", "E", "F"):
        ws.conditional_formatting.add(f"{col}{start}:{col}{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 19 — Brand Kit
# ===========================================================================
def build_brandkit(wb):
    ws = wb.create_sheet("Brand Kit"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 4, 24, 24, 2])
    luxe_header(ws, "G", "🎨  BRAND KIT", "Stay recognizable — your colors, fonts, voice & links in one place.")
    blocks = [
        ("VISUAL", [("Primary Color", "#1B4F48 Forest"), ("Accent Color", "#C9A86A Gold"),
                    ("Font — Headline", "Serif display"), ("Font — Body", "Clean sans"),
                    ("Logo File", "brand/vella-logo.png"), ("Watermark", "@vellacreates")]),
        ("VOICE", [("Tone", "Warm, encouraging, doable"), ("We say", "'You've got this'"),
                   ("We avoid", "Gatekeeping, jargon"), ("Signature sign-off", "'See you this weekend'"),
                   ("Emoji set", "🛠 🏡 ✨"), ("Caption style", "Short + one clear CTA")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "LINK IN BIO", "section_gold"); ws.row_dimensions[15].height = 22
    links = [("Shop", "vellacreates.co/shop"), ("Amazon", "amzn.to/vella"), ("Newsletter", "vellacreates.co/note"),
             ("YouTube", "Vella Creates"), ("Media Kit", "vellacreates.co/kit"), ("Email", "hi@vellacreates.co")]
    for i, (p, h) in enumerate(links):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 20 — Content Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  CONTENT GALLERY",
                "Your visual archive — drop in covers & stills and caption your best posts.")
    merge_set(ws, "B5:D5", "HOW TO ADD YOUR COVERS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Excel: Insert ▸ Pictures ▸ Place in Cell (or drag an image) into any framed box below. "
                      "Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"paste-link-here\"). "
                      "Caption each one underneath — hook, views, and what made it work.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Top Post", "Viral Hit", "Brand Deal", "Series Cover", "Before/After", "Trend Win"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add cover)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Hook · views · why it worked…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 21 — Goals & OKRs
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 14, 16, 12, 12, 2])
    luxe_header(ws, "F", "🎯  GOALS & OKRs",
                "Run your account on objectives — measurable key results with live progress.")
    table_headers(ws, 4, ["Objective / Key Result", "Category", "Target", "Current", "Progress"])
    goals = [
        ("Hit 300k followers", "Followers", "300,000", "215,000", 0.72),
        ("$10k/mo revenue", "Revenue", "$10,000", "$8,600", 0.86),
        ("Post 24 videos / month", "Posting", "24", "22", 0.92),
        ("Grow 20k followers / mo", "Followers", "20,000", "18,400", 0.92),
        ("Land 5 active brand deals", "Brand Deals", "5", "4", 0.80),
        ("$2k/mo TikTok Shop income", "Shop", "$2,000", "$1,883", 0.94),
        ("Lift completion to 55%", "Engagement", "55%", "48%", 0.87),
    ]
    start = L0
    for i, (g, cat, tgt, cur, prog) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalCatList")
    nrange(wb, "GoalCategory", "Goals", "B", start, end)
    nrange(wb, "GoalProgress", "Goals", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 22 — Audience Insights
# ===========================================================================
def build_audience(wb):
    ws = wb.create_sheet("Audience"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 4, 24, 16, 2])
    luxe_header(ws, "G", "🌍  AUDIENCE INSIGHTS",
                "Know who's watching — demographics, timing & what your audience loves.")
    merge_set(ws, "B5:C5", "WHO'S WATCHING", "section_gold"); ws.row_dimensions[5].height = 22
    demo = [("Women", "78%"), ("Age 18-24", "31%"), ("Age 25-34", "44%"), ("Age 35-44", "18%"),
            ("Top country", "US 62%"), ("Then", "UK · CA · AU"), ("Returning viewers", "41%"), ("Followers who saw last post", "22%")]
    for i, (k, v) in enumerate(demo):
        r = 6 + i
        ws.cell(row=r, column=2, value=k).style = "field_label"
        ws.cell(row=r, column=3, value=v).style = "field_value"
    merge_set(ws, "E5:F5", "WHEN & WHAT", "section_gold"); ws.row_dimensions[5].height = 22
    when = [("Best days", "Sun, Tue, Thu"), ("Best times", "6pm & 8pm ET"), ("Peak activity", "Sun 7pm"),
            ("Top pillar", "Renter Hacks"), ("Most saved", "$50 Makeovers"), ("Most shared", "Thrift Flips"),
            ("Avg watch time", "18s"), ("Comment driver", "'Keep or toss?'")]
    for i, (k, v) in enumerate(when):
        r = 6 + i
        ws.cell(row=r, column=5, value=k).style = "field_label"
        ws.cell(row=r, column=6, value=v).style = "field_value"


# ===========================================================================
# 23 — Collabs & Duets
# ===========================================================================
def build_collabs(wb):
    rows = [
        ("@homewithnat", 340000, "Duet", "Planned", dplus(6), "Renter hacks crossover"),
        ("@thriftedhome", 210000, "Collab video", "Confirmed", dplus(12), "Thrift haul together"),
        ("@budgetdiy", 520000, "Stitch", "Posted", dminus(8), "Stitched their hack"),
        ("@firstplace", 95000, "Shoutout swap", "Confirmed", dplus(3), "Story swap"),
        ("@paintpro", 680000, "Duet", "Idea", dplus(20), "Cabinet painting tips"),
        ("@cleanwithme", 410000, "Collab LIVE", "Planned", dplus(15), "Sunday reset co-host"),
    ]
    ws, start, end = build_log(
        wb, "Collabs", "👯", "COLLABS & DUETS",
        "Grow through others — track duets, stitches & collabs from idea to posted.",
        ["Creator", "Their Reach", "Type", "Status", "Date", "Idea"],
        rows, [20, 14, 14, 13, 13, 26],
        text_left={6}, ints={2}, dates={5}, reserved=24)
    for st, cc in {"Posted": MINT_BG, "Confirmed": WARN_BG, "Planned": SOFT_BG}.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 1 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🎵  TIKTOK COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Followers, revenue, trends & brand deals — your whole TikTok business, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("FOLLOWERS", "=FollowerNow", "num"),
        ("VIEWS (28D)", "=Views28", "num"),
        ("ENGAGEMENT RATE", "=EngRate", "pct1"),
        ("AVG COMPLETION", "=Completion", "pct"),
        ("POSTED (28D)", '=COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Posted")+COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Viral")', "num"),
        ("MONTHLY REVENUE", "=RevenueTotal", "money"),
    ]
    row2 = [
        ("NET PROFIT", "=RevenueTotal-ExpenseTotal", "money"),
        ("SHOP SALES (GMV)", "=ShopGMV", "money"),
        ("BRAND DEALS", '=COUNTIF(SponStage,"Signed")+COUNTIF(SponStage,"Delivered")+COUNTIF(SponStage,"Negotiation")', "num"),
        ("FOLLOWER GROWTH", "=FollowerGrowth", "num"),
        ("POSTING CONSISTENCY", '=IFERROR(MIN((COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Posted")+COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Viral"))/PostGoal,1),0)', "pct"),
        ("CREATOR HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "GROWTH & REVENUE", "section_gold")
    # follower growth line
    ln = LineChart(); ln.title = "Follower Growth (K)"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=26, max_row=32), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=27, max_row=32)); ln.legend = None
    ws.add_chart(ln, "B12")
    # revenue by source donut
    d1 = DoughnutChart(); d1.title = "Revenue by Source"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Finance"], min_col=2, min_row=4, max_row=11), titles_from_data=True)
    d1.set_categories(Reference(wb["Finance"], min_col=1, min_row=5, max_row=11)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "CONTENT & PROFIT", "section_gold")
    # top videos bar
    cb = BarChart(); cb.type = "bar"; cb.title = "Top Videos by Views"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); cb.legend = None
    ws.add_chart(cb, "B30")
    # expense donut
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Finance"], min_col=7, min_row=4, max_row=11), titles_from_data=True)
    eb.set_categories(Reference(wb["Finance"], min_col=6, min_row=5, max_row=11)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "TikTok Command Center™ — from hook to payout, all in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_calendar(wb)
    build_pipeline(wb); build_ideas(wb); build_hooks(wb); build_trends(wb)
    build_hashtags(wb); build_analytics(wb); build_live(wb); build_shop(wb)
    build_affiliate(wb); build_deals(wb); build_series(wb); build_finance(wb)
    build_expenses(wb); build_equipment(wb); build_repurpose(wb); build_brandkit(wb)
    build_gallery(wb); build_goals(wb); build_audience(wb); build_collabs(wb)
    build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Profile", "Calendar", "Pipeline", "Ideas", "Hooks", "Trends",
             "Hashtags", "Analytics", "LIVE", "Shop", "Affiliate", "Brand Deals", "Series", "Finance",
             "Expenses", "Equipment", "Repurposing", "Brand Kit", "Gallery", "Goals", "Audience",
             "Collabs", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "TikTok_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
