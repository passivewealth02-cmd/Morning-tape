"""Build Podcast Command Center™ — The Ultimate Podcast Production & Business System.

24 sheets + Welcome · a premium podcast operating system in Excel & Sheets.
Episodes, guests, recording, show notes, downloads analytics, platforms,
sponsors (CPM), memberships, clips, finance, goals & more — one dashboard.

Run: python3 build_xlsx.py   ->  ../Podcast_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

FORMATS = ["Interview", "Solo", "Panel", "Q&A", "Deep Dive", "Bonus", "Trailer"]
EP_STATUS = ["Idea", "Booked", "Recording", "Editing", "Scheduled", "Published", "Repurposed"]
REV_CATS = ["Sponsorships", "Memberships", "YouTube", "Affiliate", "Merch", "Courses", "Live & Other"]
EXP_CATS = ["Hosting", "Editing", "Software", "Equipment", "Marketing", "Guests & Travel", "Music & Licensing", "Miscellaneous"]
SPON_STAGES = ["Lead", "Pitched", "Negotiation", "Booked", "Live", "Completed"]
GUEST_STATUS = ["Idea", "Invited", "Booked", "Recorded", "Published", "Declined"]
GOAL_CATS = ["Subscribers", "Downloads", "Revenue", "Episodes", "Sponsors", "Members", "Consumption"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


def totals(ws, row, cols, start, end, fmt='"$"#,##0', label="TOTAL"):
    ws.cell(row=row, column=1, value=label).style = "th"
    for col in cols:
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt


# ===========================================================================
# 24 — Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your show details & targets once — every dashboard follows.")
    merge_set(ws, "B5:C5", "SHOW INPUTS", "section")
    controls = [
        ("Show Name", "Make It Work", None, "ShowName"),
        ("Host", "Priya Anand", None, "HostName"),
        ("Category", "Business & creativity", None, "Category"),
        ("Monthly Revenue Goal", 10000, '"$"#,##0', "RevenueGoal"),
        ("Monthly Episode Goal", 5, "0", "EpGoal"),
        ("Monthly Download Goal", 50000, "#,##0", "DownloadGoal"),
        ("Consumption Target", 0.80, "0%", "ConsumptionTarget"),
        ("Active Sponsor Target", 4, "0", "SponsorTarget"),
        ("Subscriber Growth Goal (mo)", 2000, "#,##0", "GrowthGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Episode Format", FORMATS, "FormatList"), ("F", "Episode Status", EP_STATUS, "StatusList"),
             ("G", "Revenue Category", REV_CATS, "RevCatList"), ("H", "Expense Category", EXP_CATS, "ExpCatList"),
             ("I", "Sponsor Stage", SPON_STAGES, "SponStageList"), ("J", "Guest Status", GUEST_STATUS, "GuestStatusList"),
             ("K", "Goal Category", GOAL_CATS, "GoalCatList"), ("L", "Priority", PRIORITIES, "PriorityList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    ws.cell(row=16, column=5, value="Yes / No").style = "th"
    for ri, v in enumerate(YESNO):
        ws.cell(row=17 + ri, column=5, value=v).style = "td_left"
    wb.defined_names["YesNoList"] = DefinedName("YesNoList", attr_text="Settings!$E$17:$E$18")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🎙  PODCAST COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate podcast production & business system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "RUN YOUR ENTIRE PODCAST FROM ONE FILE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("From idea to episode to invoice — Podcast Command Center™ manages your episode calendar, guest "
                      "outreach, recording, show notes, downloads analytics, platforms, sponsors, memberships, clips and "
                      "finances in ONE premium Excel & Google Sheets system. Publish consistently, grow downloads, land "
                      "sponsors and turn listens into real income — all with producer-grade automation. This isn't an "
                      "episode tracker — it's your complete Podcast Operating System.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your show, host & monthly goals.",
             "2.  Plan the Episode Calendar & Pipeline — idea → recording → published.",
             "3.  Work the Guest CRM & Recording Log — book, record, track.",
             "4.  Log Analytics, Platforms & Sponsors — downloads & revenue update live.",
             "5.  Grow income with Memberships, Clips & the Repurposing engine.",
             "6.  Watch the Dashboard track downloads, revenue & a Show Health Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (Make It Work, an 18.5k-subscriber business & creativity show doing ~42k "
               "downloads/mo and ~$9,100/mo) is included so you can see how everything connects — just type over it "
               "with your own. Downloads, revenue, net profit, consumption, publishing consistency, the sponsor "
               "pipeline and the Show Health Score all update automatically. Every sheet is print-friendly and works "
               "in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more time to record — let's grow your show.", "section_gold")


# ===========================================================================
# 2 — Show Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Show Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "🎧  SHOW PROFILE", "Your show, defined — the identity every episode flows from.")
    blocks = [
        ("THE SHOW", [("Show Name", "=ShowName"), ("Host", "=HostName"),
                      ("Category", "=Category"), ("Cadence", "Weekly · Tuesdays"),
                      ("Launched", "2022"), ("Format", "Interview + solo")]),
        ("STRATEGY", [("Tagline", "Building a life around your work"), ("Audience", "Founders & creatives 28-45"),
                      ("Avg length", "48 min"), ("Host read CPM", "$28"),
                      ("Media Kit", "makeitwork.fm/sponsor"), ("Booking", "makeitwork.fm/guest")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "WHERE TO LISTEN", "section_gold"); ws.row_dimensions[15].height = 22
    handles = [("Apple Podcasts", "Make It Work"), ("Spotify", "Make It Work"), ("YouTube", "Make It Work Pod"),
               ("Website", "makeitwork.fm"), ("Newsletter", "The Debrief"), ("Community", "makeitwork.fm/circle")]
    for i, (p, h) in enumerate(handles):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 3 — Episode Calendar
# ===========================================================================
def build_calendar(wb):
    eps = [
        ("How to price your work", "Interview", "Dana Kim"), ("Quitting the 9-5: the real math", "Solo", "—"),
        ("Building in public", "Interview", "Marco Ellis"), ("The burnout trap", "Interview", "Dr. Lena Ross"),
        ("Your first $10k month", "Interview", "Ivy Chen"), ("Say no to grow", "Solo", "—"),
        ("Finding your niche", "Interview", "Owen Blake"), ("The creator middle class", "Panel", "3 guests"),
        ("Systems > motivation", "Interview", "Sam Park"), ("Pricing for retainers", "Q&A", "—"),
        ("Launch without an audience", "Interview", "Priya S."), ("The 4-hour content week", "Solo", "—"),
        ("From freelancer to agency", "Interview", "Chris N."), ("Money mindset reset", "Interview", "Nora B."),
    ]
    pub_off = [2, 9, 16, 23]
    future = [(-4, "Scheduled"), (-11, "Editing"), (-7, "Recording"), (-14, "Booked"), (-21, "Idea")]
    rows = []
    ei = 0
    for off in pub_off:
        t, fm, g = eps[ei % len(eps)]
        rows.append((dminus(off), t, fm, g, "Published")); ei += 1
    for foff, status in future:
        t, fm, g = eps[ei % len(eps)]
        rows.append((dplus(-foff), t, fm, g, status)); ei += 1
    sample = [(d, f"Ep {40 + i}: {t}", fm, g, "High" if st in ("Published", "Scheduled", "Recording") else "Medium", st)
              for i, (d, t, fm, g, st) in enumerate(rows)]
    ws, start, end = build_log(
        wb, "Calendar", "🗓", "EPISODE CALENDAR",
        "Plan every episode in one view — publishing status calculates itself.",
        ["Publish Date", "Episode", "Format", "Guest", "Priority", "Status"],
        sample, [13, 32, 14, 16, 11, 13],
        text_left={2, 4}, dates={1},
        validations=[("C", "FormatList"), ("E", "PriorityList"), ("F", "StatusList")], reserved=60)
    nrange(wb, "CalDate", "Calendar", "A", start, end)
    nrange(wb, "CalStatus", "Calendar", "F", start, end)
    cmap = {"Published": MINT_BG, "Scheduled": WARN_BG, "Editing": SOFT_BG, "Recording": SOFT_BG, "Booked": WHITE, "Idea": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 4 — Production Pipeline
# ===========================================================================
def build_pipeline(wb):
    rows = [
        ("Ep 44: Money mindset reset", "Editing", 0.80, "Editor — Jo", 2, "Cut to 46min + chapters"),
        ("Ep 45: From freelancer to agency", "Recording", 0.55, "Priya", 7, "Recorded, needs edit"),
        ("Ep 46: The 4-hour content week", "Script", 0.40, "Priya", 11, "Solo outline drafting"),
        ("Ep 47: Launch without an audience", "Booked", 0.20, "Priya", 14, "Guest confirmed"),
        ("Ep 48: Pricing for retainers", "Idea", 0.10, "Priya", 21, "Q&A — collect questions"),
        ("Ep 43: Systems > motivation", "Show notes", 0.90, "VA — Mia", 1, "Notes + clips left"),
        ("Ep 45 clips", "Clips", 0.60, "Editor — Jo", 3, "3 of 6 cut"),
        ("Ep 46: solo audio cleanup", "Editing", 0.50, "Editor — Jo", 5, "Noise reduction"),
    ]
    sample = [(t, st, prog, o, dplus(d), note) for (t, st, prog, o, d, note) in rows]
    ws, start, end = build_log(
        wb, "Pipeline", "🎬", "PRODUCTION PIPELINE",
        "Every episode, every stage — from idea to published, with live progress %.",
        ["Episode", "Stage", "Progress", "Owner", "Due", "Notes"],
        sample, [30, 14, 12, 16, 13, 24],
        text_left={1, 4, 6}, dates={5}, pcts={3}, reserved=40)
    nrange(wb, "PipeProgress", "Pipeline", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 5 — Episode Planner
# ===========================================================================
def build_episodes(wb):
    rows = [
        ("The creator middle class", "Panel", "Is a stable creator income real?", "Money · Creators", 5, "Planned"),
        ("Selling without being cringe", "Solo", "Ethical selling frameworks", "Sales · Mindset", 4, "Idea"),
        ("Your email list is your business", "Interview", "Newsletter growth tactics", "Email · Growth", 5, "Booked"),
        ("The anti-hustle roadmap", "Solo", "Sustainable pace + systems", "Productivity", 4, "Idea"),
        ("Raising your rates (a playbook)", "Q&A", "Scripts for the awkward talk", "Pricing", 5, "Planned"),
        ("What I'd do with 0 followers", "Solo", "Cold-start growth", "Growth", 4, "Idea"),
        ("Building a tiny team", "Interview", "First hires & delegation", "Ops", 4, "Booked"),
        ("The one-person media company", "Interview", "Repurposing at scale", "Content", 5, "Idea"),
    ]
    ws = wb.create_sheet("Episodes"); ws.sheet_view.showGridLines = False
    set_widths(ws, [28, 14, 30, 18, 12, 12])
    luxe_header(ws, "F", "💡  EPISODE PLANNER",
                "Never run dry — bank episode ideas & angles, ranked by an interest score.")
    table_headers(ws, 4, ["Working Title", "Format", "Angle / Promise", "Keywords", "Interest", "Status"])
    start = L0
    for i, (t, fm, angle, kw, interest, status) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=t).style = "td_left"
        ws.cell(row=r, column=2, value=fm).style = "td"
        ws.cell(row=r, column=3, value=angle).style = "td_left"
        ws.cell(row=r, column=4, value=kw).style = "td_left"
        ci = ws.cell(row=r, column=5, value=interest); ci.style = "td"; ci.number_format = "0"
        ws.cell(row=r, column=6, value=status).style = "td"
        if i % 2:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    style_rows(ws, end + 1, end + 30, 6, text_left={1, 3, 4})
    add_dv(ws, f"B{start}:B{end+30}", "FormatList"); add_dv(ws, f"F{start}:F{end+30}", "StatusList")
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"E{start}:E{end}", ColorScaleRule(
        start_type="num", start_value=1, start_color="FF" + WARN_BG, end_type="num", end_value=5, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 6 — Guest CRM
# ===========================================================================
def build_guests(wb):
    rows = [
        ("Dana Kim", "Pricing coach", "Booked", "@danakim", 84000, dplus(6), "Warm intro from Marco"),
        ("Marco Ellis", "Build-in-public founder", "Recorded", "@marcoe", 210000, dminus(9), "Great story, big reach"),
        ("Dr. Lena Ross", "Burnout researcher", "Published", "@drlena", 45000, dminus(16), "Evergreen topic"),
        ("Ivy Chen", "6-figure freelancer", "Invited", "@ivychen", 62000, "—", "Waiting on reply"),
        ("Owen Blake", "Niche strategist", "Booked", "@owenb", 38000, dplus(20), "Confirmed for Ep 47"),
        ("Sam Park", "Systems consultant", "Idea", "@sampark", 120000, "—", "Reach out via mutual"),
        ("Priya S.", "Launch expert", "Booked", "@priyas", 156000, dplus(13), "Cross-promo potential"),
        ("Chris N.", "Agency owner", "Idea", "@chrisn", 74000, "—", "Ep 45 topic fit"),
    ]
    ws, start, end = build_log(
        wb, "Guests", "🤝", "GUEST CRM",
        "Book better guests — outreach status, reach & the pipeline from idea to published.",
        ["Guest", "Who They Are", "Status", "Handle", "Their Reach", "Record Date", "Notes"],
        rows, [18, 22, 14, 14, 13, 14, 24],
        text_left={2, 4, 7}, ints={5}, reserved=30,
        validations=[("C", "GuestStatusList")])
    nrange(wb, "GuestName", "Guests", "A", start, end)
    for st, cc in {"Published": MINT_BG, "Recorded": WARN_BG, "Booked": SOFT_BG, "Invited": WHITE}.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 7 — Recording Log
# ===========================================================================
def build_recording(wb):
    rows = [
        (dminus(9), "Ep 42: Building in public", "Marco Ellis", "Riverside", 62, "Great", "Backup on SD too"),
        (dminus(16), "Ep 41: The burnout trap", "Dr. Lena Ross", "Riverside", 54, "Great", "Levels perfect"),
        (dminus(2), "Ep 44: Money mindset reset", "Nora B.", "Riverside", 58, "Good", "Minor room echo — fixed"),
        (dplus(6), "Ep 45: Freelancer to agency", "Chris N.", "In-person", 0, "Scheduled", "Bring travel kit"),
        (dplus(13), "Ep 47: Launch w/o audience", "Priya S.", "Riverside", 0, "Scheduled", "Send prep doc"),
    ]
    ws, start, end = build_log(
        wb, "Recording", "🎚", "RECORDING LOG",
        "Every session tracked — platform, length, quality & notes so nothing gets lost.",
        ["Date", "Episode", "Guest", "Platform", "Mins", "Quality", "Notes"],
        rows, [13, 28, 16, 13, 9, 12, 24],
        text_left={2, 3, 7}, dates={1}, ints={5}, reserved=30)
    for q, cc in {"Great": MINT_BG, "Good": SOFT_BG, "Scheduled": WARN_BG}.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{q}"'], fill=fill(cc)))


# ===========================================================================
# 8 — Show Notes & Assets
# ===========================================================================
def build_notes(wb):
    rows = [
        ("Ep 43: Systems > motivation", "Yes", "Yes", "Yes", "Yes", "No", "Chapters + 3 links"),
        ("Ep 42: Building in public", "Yes", "Yes", "Yes", "Yes", "Yes", "Full transcript live"),
        ("Ep 41: The burnout trap", "Yes", "Yes", "No", "Yes", "Yes", "Add pull-quotes"),
        ("Ep 44: Money mindset reset", "No", "No", "No", "No", "No", "In production"),
        ("Ep 40: Your first $10k month", "Yes", "Yes", "Yes", "Yes", "Yes", "Top performer — refresh SEO"),
    ]
    ws, start, end = build_log(
        wb, "Show Notes", "📝", "SHOW NOTES & ASSETS",
        "Ship complete episodes — notes, transcript, timestamps, links & SEO title, checked off.",
        ["Episode", "Notes", "Transcript", "Timestamps", "Links", "SEO Title", "Status"],
        rows, [28, 10, 12, 12, 10, 11, 24],
        text_left={1, 7}, reserved=30,
        validations=[(c, "YesNoList") for c in ("B", "C", "D", "E", "F")])
    for col in ("B", "C", "D", "E", "F"):
        ws.conditional_formatting.add(f"{col}{start}:{col}{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 9 — Analytics
# ===========================================================================
def build_analytics(wb):
    rows = [
        ("Ep 40: Your first $10k month", 18400, 0.84, 620, 4.9),
        ("Ep 42: Building in public", 14200, 0.81, 480, 4.8),
        ("Ep 39: Finding your niche", 12600, 0.79, 410, 4.8),
        ("Ep 41: The burnout trap", 11800, 0.82, 360, 4.9),
        ("Ep 43: Systems > motivation", 9800, 0.77, 300, 4.7),
        ("Ep 38: Say no to grow", 7600, 0.74, 210, 4.6),
    ]
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 3, 18, 12, 12, 2])
    luxe_header(ws, "G", "📊  ANALYTICS COMMAND CENTER",
                "Your show by the numbers — a live snapshot, health dimensions & the Show Health Score.")
    merge_set(ws, "B5:C5", "SNAPSHOT (28 DAYS)", "section")
    snap = [("Subscribers", 18500, "#,##0", "Subscribers"), ("Subs gained", 1400, "#,##0", "SubGrowth"),
            ("Downloads", 42000, "#,##0", "Downloads28"), ("Avg / episode", 11000, "#,##0", "AvgPerEp"),
            ("Unique listeners", 24000, "#,##0", "UniqueListeners"), ("Consumption", 0.78, "0%", "Consumption"),
            ("Avg rating", 4.8, "0.0", "AvgRating"), ("Ad CPM", 28, '"$"#,##0.00', "CPM")]
    for i, (lab, val, fmt, nm) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "field_value"; c.number_format = fmt
        cell_name(wb, nm, "Analytics", f"$C${r}")
        if lab in ("Subscribers", "Downloads"):
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    merge_set(ws, "E5:G5", "SHOW HEALTH", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=5)
    metrics = [
        ("Revenue vs goal", "=IFERROR(MIN(RevenueTotal/RevenueGoal,1),0)"),
        ("Publishing consistency", '=IFERROR(MIN(COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Published")/EpGoal,1),0)'),
        ("Downloads vs goal", "=IFERROR(MIN(Downloads28/DownloadGoal,1),0)"),
        ("Consumption", "=IFERROR(MIN(Consumption/ConsumptionTarget,1),0)"),
        ("Sponsor pipeline", '=IFERROR(MIN((COUNTIF(SponStage,"Booked")+COUNTIF(SponStage,"Live"))/SponsorTarget,1),0)'),
        ("Audience growth", "=IFERROR(MIN(SubGrowth/GrowthGoal,1),0)"),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(metrics):
        r = hs + i
        ws.cell(row=r, column=5, value=dim).style = "td_left"
        c = ws.cell(row=r, column=6, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=7, value=f'=IF(F{r}>=0.75,"Strong",IF(F{r}>=0.5,"Growing","Focus"))').style = "td"
        if i % 2:
            for c2 in range(5, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(metrics) - 1
    cell_name(wb, "HealthRange", "Analytics", f"$F${hs}:$F${he}")
    ws.conditional_formatting.add(f"F{hs}:F{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "B15:G15", "TOP EPISODES (30-DAY DOWNLOADS)", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Episode", "Downloads", "Consumption", "New Subs", "Rating"], start_col=2)
    vs = 17
    for i, (title, dl, cons, subs, rating) in enumerate(rows):
        r = vs + i
        ws.cell(row=r, column=2, value=title).style = "td_left"
        cd = ws.cell(row=r, column=3, value=dl); cd.style = "td"; cd.number_format = "#,##0"
        cc = ws.cell(row=r, column=4, value=cons); cc.style = "td"; cc.number_format = "0%"
        cs = ws.cell(row=r, column=5, value=subs); cs.style = "td"; cs.number_format = "#,##0"
        cr = ws.cell(row=r, column=6, value=rating); cr.style = "td"; cr.number_format = "0.0"
        if i % 2:
            for c2 in range(2, 7):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    ve = vs + len(rows) - 1
    nrange(wb, "EpTitle", "Analytics", "B", vs, ve)
    nrange(wb, "EpDownloads", "Analytics", "C", vs, ve)
    ws.conditional_formatting.add(f"C{vs}:C{ve}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=18400, color=PRIMARY, showValue=True))
    merge_set(ws, "B25:C25", "DOWNLOADS (K) — 6 MONTHS", "section")
    ws.cell(row=26, column=2, value="Month").style = "th"; ws.cell(row=26, column=3, value="Downloads (K)").style = "th"
    months = _recent_months(6); vals = [22, 27, 31, 35, 38, 42]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 27 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = "0"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "DlMonth", "Analytics", "$B$27:$B$32")
    cell_name(wb, "DlVal", "Analytics", "$C$27:$C$32")


# ===========================================================================
# 10 — Platform Analytics
# ===========================================================================
def build_platforms(wb):
    rows = [
        ("Apple Podcasts", 0.41, 17200, 8600, 4.8, "Strongest subscriber base"),
        ("Spotify", 0.34, 14300, 6900, 4.7, "Fastest-growing"),
        ("YouTube", 0.13, 5500, 3800, 0.0, "Video clips drive discovery"),
        ("Overcast", 0.05, 2100, 900, 0.0, "Loyal power listeners"),
        ("Pocket Casts", 0.04, 1700, 700, 0.0, "Steady"),
        ("Web / RSS", 0.03, 1200, 500, 0.0, "Direct + embeds"),
    ]
    ws, start, end = build_log(
        wb, "Platforms", "📡", "PLATFORM ANALYTICS",
        "Where your listeners are — download share, subscribers & ratings by platform.",
        ["Platform", "Share", "Downloads", "Subscribers", "Rating", "Notes"],
        rows, [20, 12, 14, 14, 12, 26],
        text_left={6}, pcts={2}, ints={3, 4}, dec={5}, reserved=16)
    ws.conditional_formatting.add(f"B{start}:B{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=0.45, color=PRIMARY, showValue=True))


# ===========================================================================
# 11 — Sponsors & Ads
# ===========================================================================
def build_sponsors(wb):
    rows = [
        ("Riverside", "Recording tool", "Live", 28, 3, 2400, "Host-read · 3 eps", 12),
        ("Notion", "Productivity", "Booked", 30, 2, 1800, "Mid-roll", 20),
        ("Squarespace", "Websites", "Live", 26, 2, 1300, "Pre + mid", -5),
        ("Athletic Greens", "Wellness", "Negotiation", 32, 4, 2600, "Whole flight", 30),
        ("HelloFresh", "Meal kits", "Pitched", 24, 2, 1100, "Awaiting reply", 45),
        ("Shopify", "E-commerce", "Lead", 30, 3, 2000, "Warm intro needed", 60),
        ("Fathom", "Analytics", "Completed", 25, 2, 900, "Paid — great CTR", -30),
    ]
    sample = [(b, cat, stage, cpm, spots, val, note, dplus(d) if d >= 0 else dminus(-d)) for (b, cat, stage, cpm, spots, val, note, d) in rows]
    ws, start, end = build_log(
        wb, "Sponsors", "💵", "SPONSORS & ADS",
        "Turn downloads into deals — pipeline, CPM, spots & value from lead to paid.",
        ["Sponsor", "Category", "Stage", "CPM", "Spots", "Deal Value", "Notes", "Flight / Paid"],
        sample, [18, 16, 14, 10, 9, 12, 22, 13],
        text_left={2, 7}, money2={4}, money={6}, ints={5}, dates={8},
        validations=[("C", "SponStageList")], reserved=30)
    nrange(wb, "SponStage", "Sponsors", "C", start, end)
    nrange(wb, "SponValue", "Sponsors", "F", start, end)
    for st, cc in {"Live": MINT_BG, "Booked": WARN_BG, "Completed": SOFT_BG, "Lead": WHITE}.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 12 — Memberships
# ===========================================================================
def build_members(wb):
    rows = [
        ("Ad-free feed", 5, 148, "Ad-free episodes early"),
        ("Bonus episodes", 10, 74, "2 bonus eps / month"),
        ("Inner Circle", 25, 30, "Bonus + monthly call + Discord"),
        ("Founding member", 50, 8, "All perks + name in credits"),
    ]
    ws = wb.create_sheet("Members"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 12, 12, 14, 30])
    luxe_header(ws, "E", "⭐  MEMBERSHIPS",
                "Recurring income from your biggest fans — tiers, members & monthly revenue.")
    table_headers(ws, 4, ["Tier", "Price", "Members", "Monthly", "Perks"])
    start = L0
    for i, (tier, price, mem, perks) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=tier).style = "td_left"
        cp = ws.cell(row=r, column=2, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=3, value=mem); cm.style = "input"; cm.number_format = "#,##0"
        cr = ws.cell(row=r, column=4, value=f"=B{r}*C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=perks).style = "td_left"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL").style = "th"
    ws.cell(row=tot, column=2).style = "th"
    cm = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = "#,##0"
    cr = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    ws.freeze_panes = "A5"
    cell_name(wb, "MemberCount", "Members", f"$C${tot}")
    cell_name(wb, "MemberRev", "Members", f"$D${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=800, color=PRIMARY, showValue=True))


# ===========================================================================
# 13 — Clips & Promo
# ===========================================================================
def build_clips(wb):
    rows = [
        ("Ep 40: '$10k month' moment", "Reel", "Yes", 92000, 3200, "Top clip — repost quarterly"),
        ("Ep 42: build-in-public tip", "Short", "Yes", 54000, 1800, "Drove Apple follows"),
        ("Ep 41: burnout warning signs", "Carousel", "Yes", 38000, 2100, "High saves"),
        ("Ep 43: systems > motivation", "Reel", "No", 0, 0, "Cutting this week"),
        ("Ep 39: niche framework", "Short", "Yes", 41000, 1400, "Evergreen"),
        ("Ep 40: pricing hot take", "Quote card", "Yes", 22000, 900, "Good for Threads"),
    ]
    ws, start, end = build_log(
        wb, "Clips", "✂", "CLIPS & PROMO",
        "One episode → many clips — track promo assets, where they ran & what they drove.",
        ["Source Clip", "Format", "Posted?", "Views", "New Follows", "Notes"],
        rows, [28, 13, 11, 13, 13, 26],
        text_left={1, 6}, ints={4, 5}, reserved=30,
        validations=[("C", "YesNoList")])
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=92000, color=GOLD_LT, showValue=True))


# ===========================================================================
# 14 — Repurposing
# ===========================================================================
def build_repurpose(wb):
    rows = [
        ("Ep 40: Your first $10k month", "Yes", "Yes", "Yes", "Yes", "Yes", "Best-performer everywhere"),
        ("Ep 42: Building in public", "Yes", "Yes", "No", "Yes", "Yes", "YouTube full ep live"),
        ("Ep 41: The burnout trap", "Yes", "No", "Yes", "Yes", "No", "Newsletter feature"),
        ("Ep 43: Systems > motivation", "No", "No", "Yes", "No", "No", "Clips in progress"),
        ("Ep 39: Finding your niche", "Yes", "Yes", "Yes", "Yes", "Yes", "Turned into a lead magnet"),
    ]
    ws, start, end = build_log(
        wb, "Repurposing", "♻", "REPURPOSING ENGINE",
        "One episode → ten pieces — track where each episode has been repurposed.",
        ["Episode", "Clips", "YouTube", "Newsletter", "Blog", "Social", "Notes"],
        rows, [26, 10, 11, 13, 10, 11, 26],
        text_left={1, 7}, reserved=30,
        validations=[(c, "YesNoList") for c in ("B", "C", "D", "E", "F")])
    for col in ("B", "C", "D", "E", "F"):
        ws.conditional_formatting.add(f"{col}{start}:{col}{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 15 — Finance Center
# ===========================================================================
def build_finance(wb):
    ws = wb.create_sheet("Finance"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 16, 12, 3, 22, 14])
    luxe_header(ws, "G", "💰  PODCAST FINANCE CENTER",
                "Every income & expense in one place — monthly, run-rate & net profit, live.")
    table_headers(ws, 4, ["Income Source", "This Month", "Annual (est.)", "% of Rev"])
    income = {"Sponsorships": 4200, "Memberships": 2150, "YouTube": 700, "Affiliate": 500,
              "Merch": 300, "Courses": 600, "Live & Other": 200}
    start = L0; iend = start + len(REV_CATS) - 1
    for i, cat in enumerate(REV_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        val = "=MemberRev" if cat == "Memberships" else income[cat]
        cm = ws.cell(row=r, column=2, value=val); cm.style = "input" if cat != "Memberships" else "field_value"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/$B${iend+1},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    itot = iend + 1
    ws.cell(row=itot, column=1, value="TOTAL REVENUE").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=itot, column=col, value=f"=SUM({L}{start}:{L}{iend})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "RevSource", "Finance", "A", start, iend)
    nrange(wb, "RevMonthly", "Finance", "B", start, iend)
    cell_name(wb, "RevenueTotal", "Finance", f"$B${itot}")
    merge_set(ws, "F4:G4", "MONTHLY EXPENSES", "section_gold")
    expenses = {"Hosting": 40, "Editing": 800, "Software": 120, "Equipment": 350,
                "Marketing": 200, "Guests & Travel": 150, "Music & Licensing": 90, "Miscellaneous": 150}
    estart = 5
    for i, cat in enumerate(EXP_CATS):
        r = estart + i
        ws.cell(row=r, column=6, value=cat).style = "td_left"
        c = ws.cell(row=r, column=7, value=expenses[cat]); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=6).fill = fill(MUTED_ROW); ws.cell(row=r, column=7).fill = fill(MUTED_ROW)
    eend = estart + len(EXP_CATS) - 1; etot = eend + 1
    ws.cell(row=etot, column=6, value="TOTAL EXPENSES").style = "th"
    c = ws.cell(row=etot, column=7, value=f"=SUM(G{estart}:G{eend})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "ExpCat", "Finance", "F", estart, eend)
    nrange(wb, "ExpMonthly", "Finance", "G", estart, eend)
    cell_name(wb, "ExpenseTotal", "Finance", f"$G${etot}")
    merge_set(ws, "A17:D17", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Revenue (month)", "=RevenueTotal", '"$"#,##0'), ("Annual run-rate", "=RevenueTotal*12", '"$"#,##0'),
             ("Expenses (month)", "=ExpenseTotal", '"$"#,##0'), ("Net profit", "=RevenueTotal-ExpenseTotal", '"$"#,##0'),
             ("Profit margin", "=IFERROR((RevenueTotal-ExpenseTotal)/RevenueTotal,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 18 + i
        ws.cell(row=r, column=1, value=lab).style = "field_label"
        c = ws.cell(row=r, column=2, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net profit", "Profit margin"):
            ws.cell(row=r, column=2).fill = fill(MINT_BG)


# ===========================================================================
# 16 — Expenses
# ===========================================================================
def build_expenses(wb):
    rows = [
        (dminus(22), "Transistor hosting", "Hosting", "$40", "RSS + analytics", "Monthly"),
        (dminus(18), "Editor — Jo", "Editing", "$800", "4 episodes", "Monthly"),
        (dminus(15), "Riverside Pro", "Software", "$24", "Recording", "Monthly"),
        (dminus(11), "New shotgun mic", "Equipment", "$220", "In-person eps", "One-time"),
        (dminus(8), "Ad campaign", "Marketing", "$200", "Promote Ep 40", "Test"),
        (dminus(5), "Guest travel", "Guests & Travel", "$150", "In-studio guest", "Per episode"),
        (dminus(2), "Music license", "Music & Licensing", "$90", "Intro/outro", "Annual"),
    ]
    ws, start, end = build_log(
        wb, "Expenses", "🧾", "EXPENSES",
        "Every production cost tracked — because a podcast is a business.",
        ["Date", "Item", "Category", "Amount", "For", "Frequency"],
        rows, [13, 22, 18, 12, 18, 14],
        text_left={2, 5}, dates={1}, reserved=30,
        validations=[("C", "ExpCatList")])


# ===========================================================================
# 17 — Equipment
# ===========================================================================
def build_equipment(wb):
    rows = [
        ("Shure SM7B", "Microphone", dminus(500), "Good", "Host mic", "Yes"),
        ("Cloudlifter", "Preamp", dminus(500), "Good", "Boosts SM7B", "Yes"),
        ("RodeCaster Pro", "Interface", dminus(400), "Good", "Mixer / recorder", "Yes"),
        ("Backup SD recorder", "Recorder", dminus(300), "Fair", "Replace card soon", "No"),
        ("Guest mic (2nd SM7B)", "Microphone", dminus(120), "New", "In-person", "Yes"),
        ("Closed-back headphones", "Monitoring", dminus(200), "Good", "x2", "Yes"),
        ("Acoustic panels", "Treatment", dminus(150), "Good", "Room treatment", "Yes"),
    ]
    ws, start, end = build_log(
        wb, "Equipment", "🎛", "EQUIPMENT",
        "Your studio kit, tracked — condition & what to upgrade next.",
        ["Item", "Type", "Bought", "Condition", "Notes", "Working?"],
        rows, [22, 14, 13, 12, 22, 12],
        text_left={5}, dates={3}, reserved=20,
        validations=[("F", "YesNoList")])
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 18 — Reviews & Ratings
# ===========================================================================
def build_reviews(wb):
    ws = wb.create_sheet("Reviews"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 12, 12, 14, 30, 2])
    luxe_header(ws, "G", "⭐  REVIEWS & RATINGS",
                "Your reputation — ratings by platform and the themes listeners mention.")
    merge_set(ws, "B5:F5", "PLATFORM RATINGS", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Platform", "Rating", "Reviews", "Trend", "Listeners Love"], start_col=2)
    rows = [
        ("Apple Podcasts", 4.8, 640, "▲ +0.1", "Actionable & honest"),
        ("Spotify", 4.7, 210, "▲ +0.2", "Great guests"),
        ("Podchaser", 4.9, 48, "► flat", "Best in the niche"),
        ("Goodpods", 4.8, 62, "▲ +0.1", "Underrated gem"),
    ]
    start = 7
    for i, (plat, rt, cnt, trend, note) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=plat).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rt); cr.style = "td"; cr.number_format = "0.0"
        ws.cell(row=r, column=4, value=cnt).style = "td"
        ws.cell(row=r, column=5, value=trend).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    trow = end + 1
    ws.cell(row=trow, column=2, value="AVG RATING").style = "th"
    cr = ws.cell(row=trow, column=3, value=f"=AVERAGE(C{start}:C{end})"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = "0.0"
    ct = ws.cell(row=trow, column=4, value=f"=SUM(D{start}:D{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = "#,##0"
    for c in (5, 6):
        ws.cell(row=trow, column=c).style = "td"; ws.cell(row=trow, column=c).fill = fill(SURFACE)
    cell_name(wb, "ReviewAvg", "Reviews", f"$C${trow}")
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=5, color=GOLD_LT, showValue=True))


# ===========================================================================
# 19 — Brand Kit
# ===========================================================================
def build_brandkit(wb):
    ws = wb.create_sheet("Brand Kit"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 4, 24, 24, 2])
    luxe_header(ws, "G", "🎨  BRAND KIT", "Stay recognizable — cover art, sound, voice & links in one place.")
    blocks = [
        ("SOUND & VISUAL", [("Primary Color", "#1B4F48 Forest"), ("Accent Color", "#C9A86A Gold"),
                            ("Cover Art", "brand/miw-cover.png"), ("Intro Music", "'Open Road' — 8s"),
                            ("Outro CTA", "Rate, review & subscribe"), ("Ad Slot", "Host-read, 60-90s")]),
        ("VOICE", [("Tone", "Warm, candid, practical"), ("We say", "'Let's make it work'"),
                   ("We avoid", "Hustle-porn & hype"), ("Episode title style", "Clear promise, no clickbait"),
                   ("Guest ask", "One story + one tactic"), ("Sign-off", "'See you next Tuesday'")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "LINKS", "section_gold"); ws.row_dimensions[15].height = 22
    links = [("Website", "makeitwork.fm"), ("Sponsor kit", "makeitwork.fm/sponsor"), ("Guest form", "makeitwork.fm/guest"),
             ("Newsletter", "The Debrief"), ("Membership", "makeitwork.fm/circle"), ("Email", "hi@makeitwork.fm")]
    for i, (p, h) in enumerate(links):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 20 — Content Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  COVER & CLIP GALLERY",
                "Your visual archive — episode art, clip covers & quote cards, captioned.")
    merge_set(ws, "B5:D5", "HOW TO ADD YOUR ART", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Excel: Insert ▸ Pictures ▸ Place in Cell (or drag an image) into any framed box below. "
                      "Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"paste-link-here\"). "
                      "Caption each one underneath — episode, downloads & what made it work.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Episode Art", "Top Clip", "Guest Promo", "Quote Card", "Season Cover", "Milestone"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add art)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Episode · downloads · notes…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 21 — Goals & OKRs
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 14, 16, 12, 12, 2])
    luxe_header(ws, "F", "🎯  GOALS & OKRs",
                "Run your show on objectives — measurable key results with live progress.")
    table_headers(ws, 4, ["Objective / Key Result", "Category", "Target", "Current", "Progress"])
    goals = [
        ("Hit 25k subscribers", "Subscribers", "25,000", "18,500", 0.74),
        ("$10k/mo revenue", "Revenue", "$10,000", "$9,130", 0.91),
        ("50k downloads / month", "Downloads", "50,000", "42,000", 0.84),
        ("Publish 5 episodes / month", "Episodes", "5", "4", 0.80),
        ("Land 4 active sponsors", "Sponsors", "4", "3", 0.75),
        ("300 paying members", "Members", "300", "260", 0.87),
        ("Lift consumption to 80%", "Consumption", "80%", "78%", 0.97),
    ]
    start = L0
    for i, (g, cat, tgt, cur, prog) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalCatList")
    nrange(wb, "GoalCategory", "Goals", "B", start, end)
    nrange(wb, "GoalProgress", "Goals", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 22 — Audience Insights
# ===========================================================================
def build_audience(wb):
    ws = wb.create_sheet("Audience"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 4, 24, 16, 2])
    luxe_header(ws, "G", "🌍  AUDIENCE INSIGHTS",
                "Know who's listening — demographics, devices & what your audience wants.")
    merge_set(ws, "B5:C5", "WHO'S LISTENING", "section_gold"); ws.row_dimensions[5].height = 22
    demo = [("Age 25-34", "38%"), ("Age 35-44", "34%"), ("Age 45-54", "16%"), ("Men / Women", "54 / 46"),
            ("Top country", "US 61%"), ("Then", "UK · CA · AU"), ("Mobile listens", "82%"), ("Finish rate", "78%")]
    for i, (k, v) in enumerate(demo):
        r = 6 + i
        ws.cell(row=r, column=2, value=k).style = "field_label"
        ws.cell(row=r, column=3, value=v).style = "field_value"
    merge_set(ws, "E5:F5", "WHEN & WHAT", "section_gold"); ws.row_dimensions[5].height = 22
    when = [("Best drop day", "Tuesday"), ("Peak listens", "Commute 7-9am"), ("2nd peak", "Gym 5-7pm"),
            ("Top format", "Interview"), ("Most-shared", "Money episodes"), ("Most-requested", "Solo deep dives"),
            ("Avg listen", "38 of 48 min"), ("Review driver", "End-of-ep ask")]
    for i, (k, v) in enumerate(when):
        r = 6 + i
        ws.cell(row=r, column=5, value=k).style = "field_label"
        ws.cell(row=r, column=6, value=v).style = "field_value"


# ===========================================================================
# 23 — Collabs & Cross-Promo
# ===========================================================================
def build_collabs(wb):
    rows = [
        ("The Founder Diaries", 92000, "Feed swap", "Confirmed", dplus(9), "Trade promo episodes"),
        ("Creative Cents", 54000, "Guest swap", "Recorded", dminus(8), "Both host on each other"),
        ("The Solo Show", 130000, "Feed drop", "Idea", dplus(20), "Ask for a feed drop"),
        ("Money Talks", 210000, "Shout-out", "Confirmed", dplus(4), "Cross-mention"),
        ("Build Something", 68000, "Bonus collab", "Planned", dplus(15), "Joint bonus ep"),
        ("The Weekly Review", 45000, "Newsletter swap", "Idea", dplus(25), "Trade newsletter blurbs"),
    ]
    ws, start, end = build_log(
        wb, "Collabs", "👥", "COLLABS & CROSS-PROMO",
        "Grow through other shows — feed swaps, guest trades & shout-outs, idea to done.",
        ["Show", "Their Reach", "Type", "Status", "Date", "Idea"],
        rows, [22, 14, 15, 13, 13, 26],
        text_left={6}, ints={2}, dates={5}, reserved=24)
    for st, cc in {"Recorded": MINT_BG, "Confirmed": WARN_BG, "Planned": SOFT_BG}.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 1 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🎙  PODCAST COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Downloads, revenue, guests & sponsors — your whole podcast, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("SUBSCRIBERS", "=Subscribers", "num"),
        ("DOWNLOADS (28D)", "=Downloads28", "num"),
        ("AVG / EPISODE", "=AvgPerEp", "num"),
        ("EPISODES (28D)", '=COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Published")', "num"),
        ("CONSUMPTION", "=Consumption", "pct"),
        ("MONTHLY REVENUE", "=RevenueTotal", "money"),
    ]
    row2 = [
        ("NET PROFIT", "=RevenueTotal-ExpenseTotal", "money"),
        ("AD RATE (CPM)", "=CPM", "money2"),
        ("ACTIVE SPONSORS", '=COUNTIF(SponStage,"Booked")+COUNTIF(SponStage,"Live")', "num"),
        ("MEMBERS", "=MemberCount", "num"),
        ("PUBLISHING", '=IFERROR(MIN(COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Published")/EpGoal,1),0)', "pct"),
        ("SHOW HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "DOWNLOADS & REVENUE", "section_gold")
    ln = LineChart(); ln.title = "Downloads (K)"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=26, max_row=32), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=27, max_row=32)); ln.legend = None
    ws.add_chart(ln, "B12")
    d1 = DoughnutChart(); d1.title = "Revenue by Source"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Finance"], min_col=2, min_row=4, max_row=11), titles_from_data=True)
    d1.set_categories(Reference(wb["Finance"], min_col=1, min_row=5, max_row=11)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "EPISODES & PROFIT", "section_gold")
    cb = BarChart(); cb.type = "bar"; cb.title = "Top Episodes by Downloads"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); cb.legend = None
    ws.add_chart(cb, "B30")
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Finance"], min_col=7, min_row=4, max_row=12), titles_from_data=True)
    eb.set_categories(Reference(wb["Finance"], min_col=6, min_row=5, max_row=12)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Podcast Command Center™ — from idea to episode to invoice, all in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_calendar(wb)
    build_pipeline(wb); build_episodes(wb); build_guests(wb); build_recording(wb)
    build_notes(wb); build_analytics(wb); build_platforms(wb); build_sponsors(wb)
    build_members(wb); build_clips(wb); build_repurpose(wb); build_finance(wb)
    build_expenses(wb); build_equipment(wb); build_reviews(wb); build_brandkit(wb)
    build_gallery(wb); build_goals(wb); build_audience(wb); build_collabs(wb)
    build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Show Profile", "Calendar", "Pipeline", "Episodes", "Guests", "Recording",
             "Show Notes", "Analytics", "Platforms", "Sponsors", "Members", "Clips", "Repurposing", "Finance",
             "Expenses", "Equipment", "Reviews", "Brand Kit", "Gallery", "Goals", "Audience", "Collabs", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Podcast_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
