"""Build Home Renovation & Flip Command Center™ — The Ultimate House-Flipping System.

16 tabs (+ Settings) · a premium real-estate deal & renovation operating system in
Google Sheets & Excel. Deal analyzer (70% rule, all-in cost, cash-on-cash ROI &
profit), rehab budget (planned vs actual by category), scope of work, contractor
directory, draws & payments, materials, timeline & phases, holding costs,
financing, comps & ARV, selling & exit, punch list and a before/after photo log —
one dashboard. Analyze it, fund it, flip it.

Run: python3 build_xlsx.py   ->  ../Flip_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

TASK_STATUS = ["Done", "In Progress", "Not Started"]
PHASE_STATUS = ["Done", "In Progress", "Not Started"]
PAY_STATUS = ["Paid", "Scheduled", "Unpaid"]
MAT_STATUS = ["Bought", "Ordered", "Need"]
BID_STATUS = ["Hired", "Bidding", "Passed"]
TRADES = ["General", "Demo", "Roofing", "Kitchen", "Bath", "Flooring", "Paint",
          "HVAC", "Electrical", "Plumbing", "Landscaping", "Other"]
ROOMS = ["Exterior", "Kitchen", "Primary Bath", "Hall Bath", "Living", "Bedrooms",
         "Basement", "Whole House", "Yard", "Other"]

# ---- the sample flip — the deal that drives every KPI ----
DEAL = {
    "ARV": 340000, "PurchasePrice": 185000, "BuyClosing": 3500, "HoldMonths": 5,
    "SellCostPct": 0.07, "LoanLTV": 0.80, "LoanRate": 0.10, "Rule70": 0.70,
}
PROJECT = "Maplewood Flip — 214 Maple St"

# Rehab budget: (category, planned, actual)
REHAB = [
    ("Demo & dumpster", 2800, 3100),
    ("Roof", 6500, 6500),
    ("Kitchen", 10500, 5500),
    ("Bathrooms (2)", 6500, 3900),
    ("Flooring", 4500, 4500),
    ("Interior & exterior paint", 3200, 900),
    ("HVAC", 4000, 4000),
    ("Electrical", 2500, 2650),
    ("Plumbing", 2000, 1900),
    ("Landscaping & curb appeal", 1500, 0),
    ("Permits & fees", 1000, 1050),
]

# Scope of work: (room / area, task, trade, status)
SCOPE = [
    ("Exterior", "Tear off & replace roof", "Roofing", "Done"),
    ("Exterior", "Power-wash & paint siding", "Paint", "In Progress"),
    ("Exterior", "New front door & hardware", "General", "Not Started"),
    ("Whole House", "Full demo of dated finishes", "Demo", "Done"),
    ("Whole House", "Replace HVAC system", "HVAC", "Done"),
    ("Whole House", "Rewire & new panel", "Electrical", "Done"),
    ("Whole House", "Re-pipe & fixtures", "Plumbing", "Done"),
    ("Kitchen", "Cabinets & quartz counters", "Kitchen", "In Progress"),
    ("Kitchen", "Tile backsplash", "Kitchen", "Not Started"),
    ("Kitchen", "Stainless appliance package", "Kitchen", "Not Started"),
    ("Primary Bath", "New vanity, tile & tub", "Bath", "In Progress"),
    ("Hall Bath", "New vanity, tile & toilet", "Bath", "Done"),
    ("Whole House", "LVP flooring throughout", "Flooring", "Done"),
    ("Whole House", "Interior paint (walls/trim)", "Paint", "In Progress"),
    ("Living", "New trim & lighting", "General", "Done"),
    ("Bedrooms", "Closets & doors", "General", "Done"),
    ("Yard", "Sod, mulch & fresh landscaping", "Landscaping", "Not Started"),
    ("Whole House", "Deep clean & stage", "General", "Done"),
]

# Contractors & vendors: (trade, company, contact, phone, bid, status)
CONTRACTORS = [
    ("Roofing", "Summit Roofing Co.", "Dave R.", "(555) 201-8834", 6500, "Hired"),
    ("HVAC", "TempPro Heating & Air", "Maria L.", "(555) 442-1190", 4000, "Hired"),
    ("Electrical", "Bright Spark Electric", "Tomas K.", "(555) 778-2213", 2500, "Hired"),
    ("Plumbing", "FlowRight Plumbing", "Angela D.", "(555) 330-4471", 2000, "Hired"),
    ("Kitchen", "Craft Cabinet Works", "Sam P.", "(555) 610-9902", 10500, "Hired"),
    ("Flooring", "Statewide Floors", "Nick B.", "(555) 118-3320", 4500, "Hired"),
    ("Paint", "TrueColor Painting", "Rosa M.", "(555) 905-6612", 3200, "Bidding"),
    ("Landscaping", "GreenScape Curb", "Leo V.", "(555) 274-8890", 1500, "Bidding"),
]

# Draws & payments: (date offset, payee, category, amount, status)
PAYMENTS = [
    (-70, "Title company", "Buy-side closing", 3500, "Paid"),
    (-64, "Summit Roofing Co.", "Roof", 6500, "Paid"),
    (-58, "Junk Bros (dumpster)", "Demo & dumpster", 3100, "Paid"),
    (-50, "TempPro Heating & Air", "HVAC", 4000, "Paid"),
    (-44, "Bright Spark Electric", "Electrical", 2650, "Paid"),
    (-38, "FlowRight Plumbing", "Plumbing", 1900, "Paid"),
    (-30, "Statewide Floors", "Flooring", 4500, "Paid"),
    (-20, "Craft Cabinet Works", "Kitchen (deposit)", 5500, "Paid"),
    (-8, "Hall-bath materials", "Bathrooms (2)", 3900, "Paid"),
    (6, "Craft Cabinet Works", "Kitchen (final)", 5000, "Scheduled"),
    (12, "TrueColor Painting", "Paint", 3200, "Scheduled"),
    (20, "GreenScape Curb", "Landscaping", 1500, "Unpaid"),
]

# Materials & purchases: (item, room, store, cost, status)
MATERIALS = [
    ("LVP flooring (1,600 sf)", "Whole House", "Floor & Decor", 3400, "Bought"),
    ("Quartz counters", "Kitchen", "Local fabricator", 2600, "Ordered"),
    ("Stainless appliances", "Kitchen", "Home Depot", 2800, "Need"),
    ("Vanities (2)", "Baths", "Wayfair", 900, "Bought"),
    ("Interior paint (12 gal)", "Whole House", "Sherwin-Williams", 480, "Bought"),
    ("Light fixtures", "Whole House", "Amazon", 640, "Ordered"),
    ("Front door", "Exterior", "Home Depot", 420, "Need"),
    ("Cabinet hardware", "Kitchen", "Amazon", 180, "Bought"),
    ("Bathroom tile", "Baths", "Floor & Decor", 560, "Bought"),
    ("Backsplash tile", "Kitchen", "Floor & Decor", 240, "Need"),
    ("Sod & mulch", "Yard", "Local nursery", 380, "Need"),
    ("Faucets & fixtures", "Baths/Kitchen", "Ferguson", 720, "Ordered"),
]

# Timeline & phases: (phase, start offset, end offset, status)
PHASES = [
    ("Acquisition & permits", -75, -60, "Done"),
    ("Demo", -60, -50, "Done"),
    ("Rough-in (MEP)", -50, -32, "Done"),
    ("Kitchen & baths", -32, 10, "In Progress"),
    ("Finishes (floor/paint/trim)", -20, 18, "In Progress"),
    ("Punch list & staging", 18, 26, "Not Started"),
    ("List & sell", 26, 60, "Not Started"),
]

# Holding cost monthly line items (loan interest is a formula): (label, monthly, formula?)
HOLDING = [
    ("Loan interest", "=LoanAmount*LoanRate/12", True),
    ("Property taxes", 300, False),
    ("Insurance (builder's risk)", 120, False),
    ("Utilities", 180, False),
    ("Lawn / security / misc", 67, False),
]

# Comps & ARV: (address, sqft, beds/baths, sold price, notes)
COMPS = [
    ("221 Maple St", 1580, "3 / 2", 336000, "Sold 30 days ago · updated"),
    ("18 Birch Ln", 1620, "3 / 2", 344000, "Similar finishes"),
    ("305 Oak Ave", 1550, "3 / 2", 331000, "Slightly smaller lot"),
    ("42 Cedar Ct", 1660, "4 / 2", 349000, "Extra bedroom"),
    ("119 Maple St", 1600, "3 / 2", 338000, "Best comp — same street"),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your project details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "PROJECT INPUTS", "section")
    controls = [
        ("Project Name", PROJECT, None, "ProjectName"),
        ("Investor / Flipper", "Your Name", None, "Investor"),
        ("Purchase Date", dt.date.today() - dt.timedelta(days=75), "mm/dd/yyyy", "PurchaseDate"),
        ("Target Sell Date", dt.date.today() + dt.timedelta(days=45), "mm/dd/yyyy", "TargetSell"),
        ("Profit Target (of ARV)", 0.15, "0%", "ProfitTarget"),
        ("ROI Target (cash-on-cash)", 0.30, "0%", "ROITarget"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Task", TASK_STATUS, "TaskStatusList"), ("F", "Phase", PHASE_STATUS, "PhaseStatusList"),
             ("G", "Payment", PAY_STATUS, "PayStatusList"), ("H", "Materials", MAT_STATUS, "MatStatusList"),
             ("I", "Bid", BID_STATUS, "BidStatusList"), ("J", "Trade", TRADES, "TradeList"),
             ("K", "Room / Area", ROOMS, "RoomList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🏚  HOME RENOVATION & FLIP COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Analyze the deal, run the rehab, hit your number — one system from offer to sold.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE FLIP, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A deal analyzer that runs the 70% rule, all-in cost, cash-on-cash ROI and projected profit in "
                      "seconds — plus a rehab budget (planned vs actual by category), scope of work, contractor "
                      "directory, draw schedule, materials list, timeline, holding costs, financing, comps & ARV, a "
                      "selling/exit calculator and a before/after photo log. Whether it's your first flip or your "
                      "fiftieth, know your number before you buy — and protect it all the way to the closing table.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open the Deal Analyzer — enter ARV, purchase price & rehab budget.",
             "2.  It runs the 70% rule, all-in cost, profit & ROI, and gives you a BUY / PASS verdict.",
             "3.  Build the Rehab Budget by category and the Scope of Work room by room.",
             "4.  Load contractors, the draw schedule, materials & the project timeline.",
             "5.  As work happens, log actual costs & payments and check off tasks.",
             "6.  Watch the Dashboard track profit, budget-used, ROI & a Deal Score live."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("A sample deal (the Maplewood Flip) is included so you can see how it all connects — just type over it "
               "with your own numbers. The Deal Analyzer, rehab budget, holding costs and selling costs all roll into "
               "one projected profit, ROI and Deal Score. Twelve matching printable pages (deal analyzer, rehab budget, "
               "scope of work, draw schedule, punch list & more) are included for the job-site binder. This is a "
               "planning & analysis tool, not financial, tax or investment advice — run your own numbers and confirm "
               "local costs, permits and market comps.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Know your number before you buy — then protect it to the closing table.", "section_gold")


# ===========================================================================
# 3 — Deal Analyzer  (the engine)
# ===========================================================================
def build_deal(wb):
    ws = wb.create_sheet("Deal Analyzer"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 4, 30, 16, 2])
    luxe_header(ws, "F", "🧮  DEAL ANALYZER",
                "Enter the numbers once — the 70% rule, all-in cost, profit & ROI compute instantly.")
    merge_set(ws, "B5:C5", "THE NUMBERS  (enter these)", "section_gold"); ws.row_dimensions[5].height = 22
    inputs = [
        ("After-Repair Value (ARV)", DEAL["ARV"], '"$"#,##0', "ARV"),
        ("Purchase Price", DEAL["PurchasePrice"], '"$"#,##0', "PurchasePrice"),
        ("Rehab Budget", "=RehabBudget", '"$"#,##0', None),
        ("Buy-Side Closing Costs", DEAL["BuyClosing"], '"$"#,##0', "BuyClosing"),
        ("Holding Period (months)", DEAL["HoldMonths"], "0", "HoldMonths"),
        ("Selling Cost % (agent + closing)", DEAL["SellCostPct"], "0%", "SellCostPct"),
        ("Loan-to-Value (of purchase)", DEAL["LoanLTV"], "0%", "LoanLTV"),
        ("Loan Rate (annual)", DEAL["LoanRate"], "0.0%", "LoanRate"),
        ("70% Rule Factor", DEAL["Rule70"], "0%", "Rule70"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(inputs):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val)
        c.style = "field_value" if isinstance(val, str) and val.startswith("=") else "input"
        c.number_format = fmt
        if nm:
            wb.defined_names[nm] = DefinedName(nm, attr_text=f"'Deal Analyzer'!$C${r}")
    merge_set(ws, "E5:F5", "THE DEAL  (auto-calculated)", "section_gold")
    outs = [
        ("Loan Amount", "=PurchasePrice*LoanLTV", '"$"#,##0', "LoanAmount"),
        ("Down Payment (cash)", "=PurchasePrice-LoanAmount", '"$"#,##0', "DownPayment"),
        ("Holding Costs (total)", "=HoldingTotal", '"$"#,##0', None),
        ("Selling Costs", "=ARV*SellCostPct", '"$"#,##0', "SellingCosts"),
        ("All-In Cost", "=PurchasePrice+RehabBudget+BuyClosing+HoldingTotal+SellingCosts", '"$"#,##0', "AllInCost"),
        ("Cash Invested", "=DownPayment+RehabBudget+BuyClosing+HoldingTotal", '"$"#,##0', "CashInvested"),
        ("Projected Profit", "=ARV-AllInCost", '"$"#,##0', "ProjectedProfit"),
        ("Cash-on-Cash ROI", "=IFERROR(ProjectedProfit/CashInvested,0)", "0%", "CashOnCash"),
        ("Return on Cost", "=IFERROR(ProjectedProfit/AllInCost,0)", "0%", "ReturnOnCost"),
        ("70% Rule — Max Offer (MAO)", "=Rule70*ARV-RehabBudget", '"$"#,##0', "MAO70"),
    ]
    for i, (lab, fml, fmt, nm) in enumerate(outs):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if nm:
            wb.defined_names[nm] = DefinedName(nm, attr_text=f"'Deal Analyzer'!$F${r}")
        if lab in ("Projected Profit", "Cash-on-Cash ROI"):
            c.fill = fill(MINT_BG)
    # Verdict banner (rows 17-18, spanning both columns)
    ws.merge_cells("B17:F18")
    v = ws["B17"]
    v.value = ('=IF(PurchasePrice<=MAO70,"✓  BUY — $"&TEXT(ProjectedProfit,"#,##0")&" projected profit ('
               '"&TEXT(CashOnCash,"0%")&" ROI), under the 70% rule",'
               '"✗  PASS / RENEGOTIATE — offer is above your 70% max of $"&TEXT(MAO70,"#,##0"))')
    v.font = Font(size=13, bold=True, color=PRIMARY)
    v.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for rr in (17, 18):
        for cc in range(2, 7):
            ws.cell(row=rr, column=cc).fill = fill(MINT_BG); ws.cell(row=rr, column=cc).border = BOX
    ws.row_dimensions[17].height = 24; ws.row_dimensions[18].height = 24


# ===========================================================================
# 4 — Property Details
# ===========================================================================
def build_property(wb):
    ws = wb.create_sheet("Property Details"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 22, 4, 22, 22, 2])
    luxe_header(ws, "F", "🏠  PROPERTY DETAILS",
                "The subject property at a glance — specs, dates & the story of the deal.")
    merge_set(ws, "B5:F5", PROJECT, "section_gold"); ws.row_dimensions[5].height = 22
    fields = [("Address", "214 Maple St"), ("City / State", "Springfield, USA"),
              ("Beds / Baths", "3 / 2"), ("Square Feet", "1,600"),
              ("Year Built", "1974"), ("Lot Size", "0.19 acre"),
              ("Property Type", "Single-family"), ("Purchase Date", "See Settings"),
              ("Condition at Buy", "Dated, cosmetic + roof"), ("Exit Strategy", "Fix & flip"),
              ("MLS # (on sale)", "—"), ("Notes", "Great school zone; strong comps")]
    row = 6; i = 0
    while i < len(fields):
        ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
        ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
        if i + 1 < len(fields):
            ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
            ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
        ws.row_dimensions[row].height = 24; i += 2; row += 1


# ===========================================================================
# 5 — Rehab Budget  (defines RehabBudget / RehabSpent + chart source)
# ===========================================================================
def build_rehab(wb):
    ws = wb.create_sheet("Rehab Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 14, 14, 12, 2])
    luxe_header(ws, "F", "🔨  REHAB BUDGET",
                "Every category — planned vs actual, so overruns show up the day they happen.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining", "% Used"], start_col=2)
    start = L0
    for i, (cat, plan, actual) in enumerate(REHAB):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=3, value=plan); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=actual); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/C{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(REHAB) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=tot, column=6, value=f"=IFERROR(D{tot}/C{tot},0)"); cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    nrange(wb, "RehabCat", "Rehab Budget", "B", start, end)
    nrange(wb, "RehabPlanned", "Rehab Budget", "C", start, end)
    nrange(wb, "RehabActual", "Rehab Budget", "D", start, end)
    cell_name(wb, "RehabBudget", "Rehab Budget", f"$C${tot}")
    cell_name(wb, "RehabSpent", "Rehab Budget", f"$D${tot}")
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="greaterThan", formula=["1"], fill=fill(RED_BG)))
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10500, color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
# 6 — Scope of Work  (defines TaskName / TaskStatus)
# ===========================================================================
def build_scope(wb):
    ws, start, end = build_log(
        wb, "Scope of Work", "📋", "SCOPE OF WORK",
        "Room by room, task by task — the full punch of the rehab, checked off as it's done.",
        ["Room / Area", "Task", "Trade", "Status"],
        SCOPE, [18, 34, 16, 14], text_left={2}, reserved=45,
        validations=[("A", "RoomList"), ("C", "TradeList"), ("D", "TaskStatusList")])
    nrange(wb, "TaskName", "Scope of Work", "B", start, end)
    nrange(wb, "TaskStatus", "Scope of Work", "D", start, end)
    cmap = {"Done": MINT_BG, "In Progress": WARN_BG, "Not Started": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 7 — Contractors & Vendors
# ===========================================================================
def build_contractors(wb):
    ws, start, end = build_log(
        wb, "Contractors", "👷", "CONTRACTORS & VENDORS",
        "Your crew & bids in one place — trade, contact, bid amount and who's hired.",
        ["Trade", "Company", "Contact", "Phone", "Bid", "Status"],
        CONTRACTORS, [16, 24, 14, 18, 12, 12], text_left={2, 3, 4}, money={5}, reserved=30,
        validations=[("A", "TradeList"), ("F", "BidStatusList")])
    cmap = {"Hired": MINT_BG, "Bidding": WARN_BG, "Passed": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 8 — Draws & Payments  (defines PayAmount / PayStatus)
# ===========================================================================
def build_payments(wb):
    rows = [(dplus(off) if off >= 0 else dminus(-off), payee, cat, amt, st) for (off, payee, cat, amt, st) in PAYMENTS]
    ws, start, end = build_log(
        wb, "Draws & Payments", "💵", "DRAWS & PAYMENTS",
        "Every dollar out the door — draws, deposits & final payments, with what's still scheduled.",
        ["Date", "Payee", "Category", "Amount", "Status"],
        rows, [14, 24, 22, 14, 14], text_left={2, 3}, dates={1}, money={4}, reserved=40,
        validations=[("E", "PayStatusList")])
    nrange(wb, "PayAmount", "Draws & Payments", "D", start, end)
    nrange(wb, "PayStatus", "Draws & Payments", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=3, value="PAID TO DATE").style = "th"
    c = ws.cell(row=tot, column=4, value='=SUMIF(PayStatus,"Paid",PayAmount)')
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "PaidToDate", "Draws & Payments", f"$D${tot}")
    cmap = {"Paid": MINT_BG, "Scheduled": WARN_BG, "Unpaid": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 9 — Materials & Purchases
# ===========================================================================
def build_materials(wb):
    ws, start, end = build_log(
        wb, "Materials", "🧱", "MATERIALS & PURCHASES",
        "The shopping list — every material, where it's from, what it costs & what's still needed.",
        ["Item", "Room", "Store", "Cost", "Status"],
        MATERIALS, [28, 16, 20, 12, 12], text_left={1, 3}, money={4}, reserved=40,
        validations=[("B", "RoomList"), ("E", "MatStatusList")])
    cmap = {"Bought": MINT_BG, "Ordered": WARN_BG, "Need": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 10 — Timeline & Phases  (defines PhaseName / PhaseStatus)
# ===========================================================================
def build_timeline(wb):
    rows = [(ph, dplus(s) if s >= 0 else dminus(-s), dplus(e) if e >= 0 else dminus(-e), st) for (ph, s, e, st) in PHASES]
    ws, start, end = build_log(
        wb, "Timeline", "🗓", "TIMELINE & PHASES",
        "The whole job on a calendar — each phase, its dates and where it stands right now.",
        ["Phase", "Start", "Target End", "Status"],
        rows, [30, 16, 16, 16], text_left={1}, dates={2, 3}, reserved=20,
        validations=[("D", "PhaseStatusList")])
    nrange(wb, "PhaseName", "Timeline", "A", start, end)
    nrange(wb, "PhaseStatus", "Timeline", "D", start, end)
    cmap = {"Done": MINT_BG, "In Progress": WARN_BG, "Not Started": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 11 — Holding Costs  (defines HoldingTotal)
# ===========================================================================
def build_holding(wb):
    ws = wb.create_sheet("Holding Costs"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 16, 2])
    luxe_header(ws, "D", "🕰  HOLDING COSTS",
                "The silent profit-killer — every month you hold, itemized. Fewer months = more profit.")
    table_headers(ws, 4, ["Line Item", "Per Month", "× Months"], start_col=2)
    start = L0
    for i, (lab, val, is_formula) in enumerate(HOLDING):
        r = start + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cm = ws.cell(row=r, column=3, value=val)
        cm.style = "field_value" if is_formula else "input"; cm.number_format = '"$"#,##0'
        ct = ws.cell(row=r, column=4, value=f"=C{r}*HoldMonths"); ct.style = "td"; ct.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(HOLDING) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL HOLDING").style = "th"
    cm = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0'
    ct = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "HoldingMonthly", "Holding Costs", f"$C${tot}")
    cell_name(wb, "HoldingTotal", "Holding Costs", f"$D${tot}")
    merge_set(ws, f"B{tot+2}:D{tot+2}",
              "Every extra month of holding eats directly into profit — this is why speed matters on a flip.",
              "section_gold")


# ===========================================================================
# 12 — Financing
# ===========================================================================
def build_financing(wb):
    ws = wb.create_sheet("Financing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "🏦  FINANCING",
                "How the deal is funded — loan, cash in, and what the money costs you each month.")
    rows = [
        ("Purchase Price", "=PurchasePrice", '"$"#,##0'),
        ("Loan-to-Value", "=LoanLTV", "0%"),
        ("Loan Amount", "=LoanAmount", '"$"#,##0'),
        ("Down Payment (cash)", "=DownPayment", '"$"#,##0'),
        ("Loan Rate (annual)", "=LoanRate", "0.0%"),
        ("Interest / Month", "=LoanAmount*LoanRate/12", '"$"#,##0'),
        ("Interest Over Hold", "=LoanAmount*LoanRate/12*HoldMonths", '"$"#,##0'),
        ("Rehab (out of pocket)", "=RehabBudget", '"$"#,##0'),
        ("Buy-Side Closing", "=BuyClosing", '"$"#,##0'),
        ("Total Cash Invested", "=CashInvested", '"$"#,##0'),
    ]
    r = L0
    merge_set(ws, f"B{r}:C{r}", "THE CAPITAL STACK", "section_gold"); ws.row_dimensions[r].height = 22
    for i, (lab, fml, fmt) in enumerate(rows):
        rr = r + 1 + i
        ws.cell(row=rr, column=2, value=lab).style = "field_label"
        c = ws.cell(row=rr, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Total Cash Invested":
            c.fill = fill(MINT_BG)


# ===========================================================================
# 13 — Comps & ARV
# ===========================================================================
def build_comps(wb):
    ws, start, end = build_log(
        wb, "Comps & ARV", "📊", "COMPS & ARV",
        "The sold comparables behind your ARV — the most important number in the whole deal.",
        ["Comp Address", "Sq Ft", "Beds / Baths", "Sold Price", "Notes"],
        COMPS, [22, 12, 14, 16, 30], text_left={1, 3, 5}, ints={2}, money={4}, reserved=12)
    nrange(wb, "CompPrice", "Comps & ARV", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="AVERAGE COMP").style = "th"
    ws.cell(row=tot, column=2).style = "td"; ws.cell(row=tot, column=2).fill = fill(SURFACE)
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=IFERROR(AVERAGE(CompPrice),0)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5, value="← use this to set your ARV in the Deal Analyzer").style = "td_left"; ws.cell(row=tot, column=5).fill = fill(SURFACE)


# ===========================================================================
# 14 — Selling & Exit
# ===========================================================================
def build_selling(wb):
    ws = wb.create_sheet("Selling & Exit"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "🏷  SELLING & EXIT",
                "The closing table — sale price minus selling costs, and the profit you actually pocket.")
    rows = [
        ("Sale Price (target = ARV)", "=ARV", '"$"#,##0'),
        ("Agent Commission + Closing", "=ARV*SellCostPct", '"$"#,##0'),
        ("Net Sale Proceeds", "=ARV-ARV*SellCostPct", '"$"#,##0'),
        ("Less: Loan Payoff", "=LoanAmount", '"$"#,##0'),
        ("Less: Cash Invested", "=CashInvested", '"$"#,##0'),
        ("Projected Profit", "=ProjectedProfit", '"$"#,##0'),
        ("Cash-on-Cash ROI", "=CashOnCash", "0%"),
        ("Profit Margin (of ARV)", "=IFERROR(ProjectedProfit/ARV,0)", "0%"),
    ]
    r = L0
    merge_set(ws, f"B{r}:C{r}", "THE PAYDAY", "section_gold"); ws.row_dimensions[r].height = 22
    for i, (lab, fml, fmt) in enumerate(rows):
        rr = r + 1 + i
        ws.cell(row=rr, column=2, value=lab).style = "field_label"
        c = ws.cell(row=rr, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Projected Profit", "Cash-on-Cash ROI"):
            c.fill = fill(MINT_BG)


# ===========================================================================
# 15 — Punch List
# ===========================================================================
def build_punch(wb):
    rows = [
        ("Kitchen", "Touch-up paint on cabinets", "Not Started"),
        ("Primary Bath", "Caulk tub & re-seat toilet", "Not Started"),
        ("Whole House", "Replace 2 outlet covers", "Not Started"),
        ("Exterior", "Fix sticking front door", "Not Started"),
        ("Living", "Adjust closet door alignment", "Not Started"),
        ("Whole House", "Final deep clean", "Not Started"),
        ("Yard", "Blow off driveway & walk", "Not Started"),
        ("Kitchen", "Test all appliances", "Not Started"),
    ]
    ws, start, end = build_log(
        wb, "Punch List", "✅", "PUNCH LIST & FINAL WALKTHROUGH",
        "The last 2% that sells the house — every little fix before photos & showings.",
        ["Room / Area", "Item", "Status"],
        rows, [18, 40, 16], text_left={2}, reserved=20,
        validations=[("A", "RoomList"), ("C", "TaskStatusList")])
    cmap = {"Done": MINT_BG, "In Progress": WARN_BG, "Not Started": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 16 — Photo Log (Before & After)
# ===========================================================================
def build_photos(wb):
    ws = wb.create_sheet("Photo Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  BEFORE & AFTER PHOTO LOG",
                "Document the transformation — for the listing, the portfolio & the next lender.")
    merge_set(ws, "B5:D5", "HOW TO ADD PHOTOS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"link\"). "
                      "Excel: Insert ▸ Pictures ▸ Place in Cell. Shoot each room before demo and again when staged — "
                      "before/after pairs sell the listing, win the next lender and build your flipping brand.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Kitchen — Before", "Kitchen — After", "Primary Bath", "Living Room", "Exterior / Curb", "Whole-House Reveal"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"📸\n{captions[idx]}\n(add photo)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Room · before / after · date…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 2 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🏚  HOME RENOVATION & FLIP COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Profit, ROI, budget & timeline — your whole flip, from offer to sold, on one screen.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("ARV", "=ARV", "money"),
        ("PURCHASE", "=PurchasePrice", "money"),
        ("REHAB BUDGET", "=RehabBudget", "money"),
        ("ALL-IN COST", "=AllInCost", "money"),
        ("PROJECTED PROFIT", "=ProjectedProfit", "money"),
        ("CASH-ON-CASH ROI", "=CashOnCash", "pct"),
    ]
    row2 = [
        ("70% RULE MAO", "=MAO70", "money"),
        ("VERDICT", '=IF(PurchasePrice<=MAO70,"BUY","PASS")', "text"),
        ("BUDGET USED", "=IFERROR(RehabSpent/RehabBudget,0)", "pct"),
        ("SPENT TO DATE", "=RehabSpent", "money"),
        ("TASKS DONE", '=IFERROR(COUNTIF(TaskStatus,"Done")/COUNTA(TaskName),0)', "pct"),
        ("DEAL SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "DEAL & PROJECT HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REHAB BUDGET — PLANNED vs ACTUAL", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Profit margin (of ARV)", "=IFERROR(MIN((ProjectedProfit/ARV)/ProfitTarget,1),0)"),
        ("Cash-on-cash ROI", "=IFERROR(MIN(CashOnCash/ROITarget,1),0)"),
        ("Meets 70% rule", "=IF(PurchasePrice<=MAO70,1,0.6)"),
        ("Rehab on budget", "=IFERROR(IF(RehabSpent<=RehabBudget,1,MAX(0,2-RehabSpent/RehabBudget)),0)"),
        ("Scope complete", '=IFERROR(COUNTIF(TaskStatus,"Done")/COUNTA(TaskName),0)'),
        ("Timeline progress", '=IFERROR(COUNTIF(PhaseStatus,"Done")/COUNTA(PhaseName),0)'),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"On track","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    rb = wb["Rehab Budget"]
    bar = BarChart(); bar.type = "col"; bar.grouping = "clustered"; bar.title = "Rehab: Planned vs Actual"
    bar.height = 7.4; bar.width = 8.6
    data = Reference(rb, min_col=3, max_col=4, min_row=4, max_row=4 + len(REHAB))
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(Reference(rb, min_col=2, min_row=5, max_row=4 + len(REHAB)))
    bar.dataLabels = no_labels()
    ws.add_chart(bar, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Home Renovation & Flip Command Center™ — know your number before you buy. Edit the deal in the Deal Analyzer.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_deal(wb); build_property(wb)
    build_rehab(wb); build_scope(wb); build_contractors(wb); build_payments(wb)
    build_materials(wb); build_timeline(wb); build_holding(wb); build_financing(wb)
    build_comps(wb); build_selling(wb); build_punch(wb); build_photos(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Deal Analyzer", "Property Details", "Rehab Budget",
             "Scope of Work", "Contractors", "Draws & Payments", "Materials", "Timeline",
             "Holding Costs", "Financing", "Comps & ARV", "Selling & Exit", "Punch List",
             "Photo Log", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Flip_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
