"""Build Dating Life Command Center™ — See It Clearly.

14 tabs · a premium dating operating system in Google Sheets & Excel. Dashboard, a dating
funnel (matches → conversations → first dates → second dates, with what each one really
costs in time and money), an effort & reciprocity scorecard, green & red flags, the people
you're seeing, a date log, time & money, non-negotiables, conversations, a safety plan,
reflection & patterns and a monthly summary — one dashboard. Not to judge anyone. To see
clearly what you already feel.

Run: python3 build_xlsx.py   ->  ../Dating_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
STATUS = ["Talking", "First date booked", "Seeing", "Slow fade", "Ended it", "They ended it", "Paused"]
METON = ["Hinge", "Bumble", "Tinder", "Through friends", "At work", "Out somewhere", "Event", "Other"]
FELT = ["Great", "Good", "Fine", "Flat", "Uncomfortable", "Relieved it ended"]
DEALBREAKER = ["Dealbreaker", "Strong preference", "Nice to have"]

# --- Funnel engine ---
MATCHES = 240
CONVERSATIONS = 68
FIRST_DATES = 14
SECOND_DATES = 5
STILL_SEEING = 2
DAYS_TO_FIRST_DATE = 9

# --- Time & money ---
SPEND_BUDGET = 520
SWIPE_HOURS = 9
MESSAGE_HOURS = 11
HOURS_PER_DATE = 2.5
PREP_HOURS_PER_DATE = 1.0

# --- Goals ---
DAYS_GOAL = 14
SECOND_DATE_GOAL = 0.30
FLAG_GOAL = 6
EFFORT_GOAL = 1.2

# Time & money: (item, monthly)
SPENDING = [
    ("Dating apps & boosts", 39.98), ("Dates \u2014 your share", 308.00),
    ("Getting ready \u2014 hair, nails, outfits", 85.00), ("Rideshare, parking & transit", 64.00),
]

# The funnel, month by month: (month, matches, conversations, first dates, second dates)
FUNNEL_MONTHS = [
    ("Feb", 190, 44, 8, 2), ("Mar", 205, 51, 9, 3), ("Apr", 218, 55, 11, 3),
    ("May", 226, 61, 12, 4), ("Jun", 233, 64, 13, 4), ("Jul", 240, 68, 14, 5),
]

# Effort — you: (what, score out of 10)
YOUR_EFFORT = [
    ("Initiating contact", 9), ("Planning the dates", 9), ("Paying / splitting fairly", 6),
    ("Being flexible on timing", 8), ("Following through on plans", 9), ("Asking about their life", 7),
]

# Effort — them: (what, score out of 10)
THEIR_EFFORT = [
    ("Initiating contact", 2), ("Planning the dates", 2), ("Paying / splitting fairly", 4),
    ("Being flexible on timing", 3), ("Following through on plans", 3), ("Asking about your life", 2),
]

# Green flags: (flag, seen?)
GREEN_FLAGS = [
    ("Texts back in a normal amount of time", "Yes"), ("Asks questions about your life", "Yes"),
    ("Plans an actual date, with a time and a place", "No"), ("Is kind to waiting staff", "Yes"),
    ("Talks about their friends warmly", "Yes"), ("Handles a small disagreement calmly", "Yes"),
    ("Respects a no the first time", "Yes"), ("Is where they said they'd be", "Yes"),
    ("Introduces you to people in their life", "No"), ("Says what they want out loud", "No"),
    ("Remembers what you told them", "Yes"), ("You feel calm afterwards, not anxious", "Yes"),
]

# Red flags: (flag, seen?)
RED_FLAGS = [
    ("Only texts late at night", "Yes"), ("Vague about plans until the last minute", "Yes"),
    ("Speaks badly about every ex", "No"), ("Pushes past a no", "No"),
    ("Love-bombs, then goes quiet", "No"), ("Won't be seen with you in public", "No"),
    ("Makes you feel like you're too much", "No"), ("Rude to staff or strangers", "No"),
    ("Story details keep changing", "No"), ("Won't discuss what this is", "No"),
    ("Only reaches out when they need something", "No"), ("You feel anxious after seeing them", "No"),
]

# People: (name, met on, first date, status, feeling)
PEOPLE = [
    ("J.", "Hinge", "07/06", "Seeing", "Good"),
    ("M.", "Through friends", "07/13", "Seeing", "Great"),
    ("R.", "Bumble", "07/02", "Slow fade", "Flat"),
    ("T.", "Hinge", "06/28", "Ended it", "Relieved it ended"),
    ("D.", "Out somewhere", "07/19", "First date booked", "Fine"),
    ("A.", "Tinder", "06/21", "They ended it", "Fine"),
    ("K.", "Hinge", "07/09", "Talking", "Good"),
]

# Date log: (date, who, what, cost, hours, felt)
DATE_LOG = [
    ("07/02", "R.", "Coffee, 40 minutes", 8, 1.5, "Flat"),
    ("07/06", "J.", "Dinner \u2014 they picked it", 0, 3.0, "Great"),
    ("07/09", "K.", "Drinks after work", 26, 2.0, "Good"),
    ("07/11", "J.", "Walk and the market", 14, 2.5, "Great"),
    ("07/13", "M.", "Friend's birthday, then a bar", 22, 4.0, "Great"),
    ("07/16", "R.", "They cancelled at 6pm", 0, 0.5, "Flat"),
    ("07/19", "D.", "Coffee", 9, 1.5, "Fine"),
    ("07/21", "M.", "Dinner \u2014 split it", 31, 2.5, "Great"),
    ("07/24", "J.", "Their place, film", 12, 3.5, "Good"),
    ("07/26", "M.", "Gallery then lunch", 27, 3.0, "Great"),
]

# Conversations: (who, days talking, messages you, messages them, met yet?)
CONVOS = [
    ("J.", 24, 210, 186, "Yes"), ("M.", 17, 164, 171, "Yes"), ("K.", 12, 88, 41, "Yes"),
    ("D.", 6, 34, 29, "Yes"), ("S.", 19, 62, 11, "No"), ("N.", 26, 47, 8, "No"),
    ("B.", 9, 22, 4, "No"), ("L.", 31, 38, 6, "No"),
]

# Non-negotiables: (what, weight, why)
NON_NEGOTIABLES = [
    ("Wants children", "Dealbreaker", "I do, and I'm not negotiating on it"),
    ("Actually single", "Dealbreaker", "Not 'separated'. Single."),
    ("Kind when things go wrong", "Dealbreaker", "Watch how they handle a cancelled train"),
    ("Doesn't drink heavily", "Dealbreaker", "I know what that costs me"),
    ("Has friends they've kept", "Strong preference", "Tells you almost everything"),
    ("Same city, or moving here", "Strong preference", "I've done long distance"),
    ("Financially steady", "Strong preference", "Not rich. Steady."),
    ("Reads something, anything", "Nice to have", "I just like it"),
    ("Likes dogs", "Nice to have", "Biscuit gets a vote"),
]

# Safety plan: (step, done?)
SAFETY = [
    ("First date is somewhere public, always", "Yes"),
    ("A friend has the name, photo and location", "Yes"),
    ("Share live location for the first few dates", "Yes"),
    ("Arrange your own transport there and back", "Yes"),
    ("Video-call before meeting in person", "Yes"),
    ("Reverse-image-search the photos", "Yes"),
    ("Don't share your home address early", "Yes"),
    ("Have a check-in text time agreed", "Yes"),
    ("Trust the feeling and leave if you want to", "Yes"),
]

# Reflection & patterns: (question, answer)
REFLECTION = [
    ("Who did I feel most like myself with?", "M. \u2014 I wasn't performing"),
    ("Who did I feel anxious around, and why?", "R. \u2014 never knew where I stood"),
    ("What pattern keeps repeating?", "I keep choosing people who need convincing"),
    ("What did I do differently this month?", "Said no to a third R. reschedule"),
    ("What am I tolerating that I shouldn't?", "Vague plans. It's not mysterious, it's rude."),
    ("What went better than last month?", "I stopped texting first every single time"),
    ("What do I want more of?", "The easy, warm kind. Like M."),
    ("What would I tell a friend in my position?", "You already know. Believe yourself."),
]

# Monthly summary: (month, spend)
MONTHS = [("Feb", 392), ("Mar", 421), ("Apr", 448), ("May", 465), ("Jun", 482), ("Jul", 497)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "money3": '"$"#,##0.000', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 40, 20, 3] + [24] * 5)
    luxe_header(ws, "I", "⚙  SETTINGS", "Set your numbers & what you're aiming for — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR NUMBERS & GOALS", "section")
    controls = [
        ("This is for", "Elise", None, "Owner"),
        ("Monthly dating budget", SPEND_BUDGET, '"$"#,##0', "Budget"),
        ("Hours per date (including travel)", HOURS_PER_DATE, "0.0", "HoursPerDate"),
        ("Hours getting ready per date", PREP_HOURS_PER_DATE, "0.0", "PrepPerDate"),
        ("Hours swiping this month", SWIPE_HOURS, "0", "SwipeHours"),
        ("Hours messaging this month", MESSAGE_HOURS, "0", "MessageHours"),
        ("Days from match to first date", DAYS_TO_FIRST_DATE, "0", "DaysToDate"),
        ("Goal: meet within this many days", DAYS_GOAL, "0", "DaysGoal"),
        ("Goal: second-date rate", SECOND_DATE_GOAL, "0%", "SecondDateGoal"),
        ("Goal: green flags minus red, at least", FLAG_GOAL, "0", "FlagGoal"),
        ("Goal: effort ratio no worse than", EFFORT_GOAL, "0.0", "EffortGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Status", STATUS, "StatusList"), ("F", "Met on", METON, "MetOnList"),
             ("G", "How it felt", FELT, "FeltList"), ("H", "Weight", DEALBREAKER, "WeightList"),
             ("I", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:I5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ✷  DATING LIFE COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Not to judge anyone. To see clearly what you already feel.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "WHAT THIS IS FOR", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Dating is the only part of life we're told not to think about clearly. Everywhere else, you'd "
                      "look at the numbers. So: this month you matched with 240 people, had 68 conversations, went on "
                      "14 first dates and 5 second dates. That's 69 hours and around $497 \\u2014 which means every second "
                      "date cost you about 14 hours and $99. Not to make you feel bad. To let you decide, on purpose, "
                      "whether that's how you want to spend a Tuesday. And then the part that actually matters: an "
                      "effort scorecard that shows, in a number, whether the person you're seeing is meeting you "
                      "halfway.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Non-Negotiables first. Write them before you meet anyone.",
             "2.  Dating Funnel: matches, conversations, dates. Just the counts.",
             "3.  Log dates as you go — what it cost, how long, how it felt.",
             "4.  Effort & Reciprocity: score you, then score them. Honestly.",
             "5.  Green & Red Flags: tick what you've actually SEEN, not hoped.",
             "6.  Reflection at month end. That page is the whole point."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  A FEW HONEST NOTES", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+4}")
    c = ws[f"B{dr+1}"]
    c.value = ("This is a personal organizing and reflection tool. It is not relationship, psychological or medical "
               "advice, and no spreadsheet can tell you whether to love someone. The scores are a mirror, not a "
               "verdict \\u2014 a low effort ratio doesn't mean a person is bad, it means you're doing more of the work, "
               "and only you can decide what to do about that. Please fill in the Safety Plan tab before your next "
               "first date. Sample data for a fictional person is included so you can see how it connects; type over "
               "it with your own. Keep this file private \\u2014 it's yours.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 5):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+6}:B{dr+6}", "You already know. This just puts a number on the thing you already know.", "section_gold")


# ===========================================================================
def build_funnel(wb):
    ws = wb.create_sheet("Dating Funnel"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 46, 18, 3, 12, 14, 14, 14, 14, 2])
    luxe_header(ws, "I", "\U0001f4ca  DATING FUNNEL — THE ENGINE",
                "Matches to conversations to dates to second dates — and what each one really costs.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    rows = [
        ("Matches", MATCHES, "#,##0", "Matches", "input"),
        ("→ Conversations that actually started", CONVERSATIONS, "#,##0", "Conversations", "input"),
        ("→ First dates", FIRST_DATES, "#,##0", "FirstDates", "input"),
        ("→ Second dates", SECOND_DATES, "#,##0", "SecondDates", "input"),
        ("→ Still seeing", STILL_SEEING, "#,##0", "StillSeeing", "input"),
    ]
    for i, (lab, val, fmt, nm, kind) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"; c.number_format = fmt
        cell_name(wb, nm, "Dating Funnel", f"$C${r}")
    ws.cell(row=12, column=2, value="WHAT SURVIVES EACH STEP").style = "section_gold"
    ws.cell(row=13, column=2, value="Matches that become a conversation").style = "field_label"
    c1 = ws.cell(row=13, column=3, value="=IFERROR(Conversations/Matches,0)"); c1.style = "field_value"; c1.number_format = "0.0%"
    ws.cell(row=14, column=2, value="Conversations that become a date").style = "field_label"
    c2 = ws.cell(row=14, column=3, value="=IFERROR(FirstDates/Conversations,0)"); c2.style = "field_value"; c2.number_format = "0.0%"
    ws.cell(row=15, column=2, value="= MATCHES THAT BECOME A DATE").style = "th"
    c3 = ws.cell(row=15, column=3, value="=IFERROR(FirstDates/Matches,0)"); c3.style = "td"
    c3.font = Font(bold=True, size=14, color=PRIMARY); c3.fill = fill(SURFACE); c3.number_format = "0.0%"
    cell_name(wb, "MatchToDate", "Dating Funnel", "$C$15")
    ws.cell(row=16, column=2, value="= FIRST DATES THAT BECOME A SECOND").style = "th"
    c4 = ws.cell(row=16, column=3, value="=IFERROR(SecondDates/FirstDates,0)"); c4.style = "td"
    c4.font = Font(bold=True, size=14, color=PRIMARY); c4.fill = fill(MINT_BG); c4.number_format = "0.0%"
    cell_name(wb, "SecondDateRate", "Dating Funnel", "$C$16")

    ws.cell(row=18, column=2, value="WHAT IT COST YOU").style = "section_gold"
    ws.cell(row=19, column=2, value="Hours spent this month").style = "field_label"
    ch = ws.cell(row=19, column=3, value="=SwipeHours+MessageHours+FirstDates*(HoursPerDate+PrepPerDate)")
    ch.style = "field_value"; ch.number_format = '0" hrs"'
    cell_name(wb, "HoursSpent", "Dating Funnel", "$C$19")
    ws.cell(row=20, column=2, value="Money spent this month").style = "field_label"
    cm = ws.cell(row=20, column=3, value="=SpendTotal"); cm.style = "field_value"; cm.number_format = '"$"#,##0.00'
    ws.cell(row=21, column=2, value="= PER FIRST DATE").style = "th"
    cf = ws.cell(row=21, column=3, value="=IFERROR(SpendTotal/FirstDates,0)"); cf.style = "td"
    cf.font = Font(bold=True, size=13, color=PRIMARY); cf.fill = fill(SURFACE); cf.number_format = '"$"#,##0.00'
    cell_name(wb, "CostPerFirstDate", "Dating Funnel", "$C$21")
    ws.cell(row=22, column=2, value="= PER SECOND DATE").style = "th"
    cs = ws.cell(row=22, column=3, value="=IFERROR(SpendTotal/SecondDates,0)"); cs.style = "td"
    cs.font = Font(bold=True, size=16, color=PRIMARY); cs.fill = fill(WARN_BG); cs.number_format = '"$"#,##0.00'
    cell_name(wb, "CostPerSecondDate", "Dating Funnel", "$C$22")
    ws.cell(row=23, column=2, value="= HOURS PER SECOND DATE").style = "th"
    chs = ws.cell(row=23, column=3, value="=IFERROR(HoursSpent/SecondDates,0)"); chs.style = "td"
    chs.font = Font(bold=True, size=16, color=PRIMARY); chs.fill = fill(WARN_BG); chs.number_format = '0.0" hrs"'
    cell_name(wb, "HoursPerSecondDate", "Dating Funnel", "$C$23")
    ws.cell(row=25, column=2, value="Not to make you feel bad. To let you decide, on purpose, how you want to spend a Tuesday.").style = "section_gold"

    merge_set(ws, "E5:I5", "THE FUNNEL, MONTH BY MONTH", "section_gold")
    table_headers(ws, 6, ["Month", "Matches", "Convos", "1st dates", "2nd dates"], start_col=5)
    ts = 7
    for i, (m, ma, co, fd, sd) in enumerate(FUNNEL_MONTHS):
        r = ts + i
        ws.cell(row=r, column=5, value=m).style = "td_left"
        for ci, v in enumerate((ma, co, fd, sd), 6):
            c = ws.cell(row=r, column=ci, value=v); c.style = "input"; c.number_format = "#,##0"
        if i % 2:
            for c in range(5, 10):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(FUNNEL_MONTHS) - 1
    nrange(wb, "FunnelDates", "Dating Funnel", "H", ts, te)
    ws.add_chart(_barchart(ws, "First Dates by Month", ts, te, 8, 5), "E15")


def build_effort(wb):
    ws = wb.create_sheet("Effort & Reciprocity"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 14, 14, 16, 2])
    luxe_header(ws, "E", "⚖  EFFORT & RECIPROCITY",
                "Score yourself, then score them. Out of ten, on what actually happened.")
    ws.cell(row=4, column=2, value="Be fair to both of you. Score what happened this month, not what you hope happens next month.").style = "section_gold"
    table_headers(ws, 5, ["What", "You", "Them", "Difference"], start_col=2)
    start = 6
    for i, ((lab, you), (_, them)) in enumerate(zip(YOUR_EFFORT, THEIR_EFFORT)):
        r = start + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cy = ws.cell(row=r, column=3, value=you); cy.style = "input"; cy.number_format = "0"
        ct = ws.cell(row=r, column=4, value=them); ct.style = "input"; ct.number_format = "0"
        cd = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cd.style = "td"; cd.number_format = "0"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(YOUR_EFFORT) - 1
    nrange(wb, "YouScores", "Effort & Reciprocity", "C", start, end)
    nrange(wb, "ThemScores", "Effort & Reciprocity", "D", start, end)
    nrange(wb, "EffortGap", "Effort & Reciprocity", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="greaterThanOrEqual", formula=["5"],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    cy = ws.cell(row=tot, column=3, value="=SUM(YouScores)"); cy.style = "td"
    cy.font = Font(bold=True, size=13, color=PRIMARY); cy.fill = fill(SURFACE); cy.number_format = "0"
    cell_name(wb, "YourEffort", "Effort & Reciprocity", f"$C${tot}")
    ct = ws.cell(row=tot, column=4, value="=SUM(ThemScores)"); ct.style = "td"
    ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = "0"
    cell_name(wb, "TheirEffort", "Effort & Reciprocity", f"$D${tot}")
    cg = ws.cell(row=tot, column=5, value="=YourEffort-TheirEffort"); cg.style = "td"
    cg.font = Font(bold=True, size=13, color=DANGER); cg.fill = fill(RED_BG); cg.number_format = "0"

    r = tot + 2
    ws.cell(row=r, column=2, value="= YOU ARE DOING THIS MUCH OF THE WORK").style = "th"
    cr = ws.cell(row=r, column=3, value="=IFERROR(YourEffort/TheirEffort,0)"); cr.style = "td"
    cr.font = Font(bold=True, size=18, color=PRIMARY); cr.fill = fill(WARN_BG); cr.number_format = '0.0"\\u00d7"'
    cell_name(wb, "EffortRatio", "Effort & Reciprocity", f"$C${r}")
    ws.cell(row=r + 1, column=2, value="Anything at 1.0 is even. Your goal").style = "field_label"
    cg2 = ws.cell(row=r + 1, column=3, value="=EffortGoal"); cg2.style = "field_value"; cg2.number_format = '0.0"\\u00d7"'
    ws.cell(row=r + 2, column=2, value="= WHERE THAT LEAVES YOU").style = "th"
    cw = ws.cell(row=r + 2, column=3, value='=IF(EffortRatio<=EffortGoal,"MET HALFWAY",IF(EffortRatio<=2,"CARRYING IT","CARRYING ALL OF IT"))')
    cw.style = "td"; cw.font = Font(bold=True, size=12, color=PRIMARY); cw.fill = fill(MINT_BG)
    ws.conditional_formatting.add(f"C{r+2}", CellIsRule(operator="equal", formula=['"CARRYING ALL OF IT"'],
                                                       fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    ws.cell(row=r + 4, column=2, value="A low score here doesn't make someone a bad person. It just means you're doing more of the work.").style = "section_gold"
    ws.cell(row=r + 5, column=2, value="What you do with that is entirely yours to decide.").style = "field_label"


def build_flags(wb):
    ws = wb.create_sheet("Green & Red Flags"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 48, 14, 3, 48, 14, 2])
    luxe_header(ws, "F", "\U0001f6a9  GREEN & RED FLAGS",
                "Tick what you have actually SEEN. Not what you hope. Not what they said once.")
    merge_set(ws, "B5:C5", "GREEN — THE ONES THAT MATTER", "section_gold")
    merge_set(ws, "E5:F5", "RED — THE ONES WE TALK OURSELVES OUT OF", "section_gold")
    table_headers(ws, 6, ["Green flag", "Seen it?"], start_col=2)
    table_headers(ws, 6, ["Red flag", "Seen it?"], start_col=5)
    start = 7
    for i, (g, seen) in enumerate(GREEN_FLAGS):
        r = start + i
        ws.cell(row=r, column=2, value=g).style = "td_left"
        ws.cell(row=r, column=3, value=seen).style = "input"
        if i % 2:
            for c in (2, 3):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    for i, (rf, seen) in enumerate(RED_FLAGS):
        r = start + i
        ws.cell(row=r, column=5, value=rf).style = "td_left"
        ws.cell(row=r, column=6, value=seen).style = "input"
        if i % 2:
            for c in (5, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(GREEN_FLAGS) - 1
    nrange(wb, "GreenSeen", "Green & Red Flags", "C", start, end)
    nrange(wb, "RedSeen", "Green & Red Flags", "F", start, end)
    add_dv(ws, f"C{start}:C{end}", "YesNoList"); add_dv(ws, f"F{start}:F{end}", "YesNoList")
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"Yes"'],
                                                                fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=['"Yes"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="GREEN FLAGS SEEN").style = "th"
    cg = ws.cell(row=tot, column=3, value='=COUNTIF(GreenSeen,"Yes")'); cg.style = "td"
    cg.font = Font(bold=True, size=14, color=PRIMARY); cg.fill = fill(MINT_BG); cg.number_format = "0"
    cell_name(wb, "GreenCount", "Green & Red Flags", f"$C${tot}")
    ws.cell(row=tot, column=5, value="RED FLAGS SEEN").style = "th"
    crd = ws.cell(row=tot, column=6, value='=COUNTIF(RedSeen,"Yes")'); crd.style = "td"
    crd.font = Font(bold=True, size=14, color=DANGER); crd.fill = fill(RED_BG); crd.number_format = "0"
    cell_name(wb, "RedCount", "Green & Red Flags", f"$F${tot}")
    ws.cell(row=tot + 2, column=2, value="= GREEN MINUS RED").style = "th"
    cn = ws.cell(row=tot + 2, column=3, value="=GreenCount-RedCount"); cn.style = "td"
    cn.font = Font(bold=True, size=16, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = "0"
    cell_name(wb, "FlagNet", "Green & Red Flags", f"$C${tot+2}")
    ws.cell(row=tot + 2, column=5, value="One red flag you keep explaining away outweighs five green ones.").style = "section_gold"
    ws.freeze_panes = "A7"


def build_people(wb):
    ws, start, end = build_log(
        wb, "People", "\U0001f465", "PEOPLE", "Who you're talking to, where you met, and where it actually stands.",
        ["Name", "Met on", "First date", "Status", "How it felt"], PEOPLE,
        [2, 18, 22, 16, 22, 22, 2], text_left={2},
        validations=[("C", "MetOnList"), ("E", "StatusList"), ("F", "FeltList")], start_col=2)
    nrange(wb, "PersonName", "People", "B", start, end)
    nrange(wb, "PersonStatus", "People", "E", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="People in the picture").style = "field_label"
    c1 = ws.cell(row=tr, column=5, value="=COUNTA(PersonName)"); c1.style = "field_value"; c1.number_format = "0"
    ws.cell(row=tr + 1, column=2, value="Actually seeing").style = "field_label"
    c2 = ws.cell(row=tr + 1, column=5, value='=COUNTIF(PersonStatus,"Seeing")'); c2.style = "field_value"
    c2.number_format = "0"; c2.fill = fill(MINT_BG)
    ws.cell(row=tr + 2, column=2, value="Slow fading").style = "field_label"
    c3 = ws.cell(row=tr + 2, column=5, value='=COUNTIF(PersonStatus,"Slow fade")'); c3.style = "field_value"
    c3.number_format = "0"; c3.fill = fill(WARN_BG)
    ws.cell(row=tr + 4, column=2, value="A slow fade is an answer. You're allowed to call it.").style = "section_gold"


def build_dates(wb):
    ws = wb.create_sheet("Date Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 14, 34, 14, 14, 20, 2])
    luxe_header(ws, "G", "\U0001f4c5  DATE LOG",
                "What you did, what it cost, how long it took — and honestly, how it felt.")
    table_headers(ws, 4, ["Date", "Who", "What", "Cost", "Hours", "How it felt"], start_col=2)
    start = L0
    for i, (dt, who, what, cost, hours, felt) in enumerate(DATE_LOG):
        r = start + i
        ws.cell(row=r, column=2, value=dt).style = "td"
        ws.cell(row=r, column=3, value=who).style = "td"
        ws.cell(row=r, column=4, value=what).style = "td_left"
        cc = ws.cell(row=r, column=5, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0'
        ch = ws.cell(row=r, column=6, value=hours); ch.style = "input"; ch.number_format = "0.0"
        ws.cell(row=r, column=7, value=felt).style = "input"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DATE_LOG) - 1
    nrange(wb, "DateCost", "Date Log", "E", start, end)
    nrange(wb, "DateHours", "Date Log", "F", start, end)
    nrange(wb, "DateFelt", "Date Log", "G", start, end)
    add_dv(ws, f"G{start}:G{end}", "FeltList")
    ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=['"Great"'],
                                                                fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=['"Uncomfortable"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="THIS MONTH").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cc = ws.cell(row=tot, column=5, value="=SUM(DateCost)"); cc.style = "td"
    cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0'
    ch = ws.cell(row=tot, column=6, value="=SUM(DateHours)"); ch.style = "td"
    ch.font = Font(bold=True, color=PRIMARY); ch.fill = fill(SURFACE); ch.number_format = "0.0"
    cgr = ws.cell(row=tot, column=7, value='=COUNTIF(DateFelt,"Great")&" great"'); cgr.style = "td"
    cgr.font = Font(bold=True, color=PRIMARY); cgr.fill = fill(MINT_BG)
    ws.cell(row=tot + 2, column=2, value="Look down the 'how it felt' column. That column is the answer.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_money(wb):
    ws = wb.create_sheet("Time & Money"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 44, 18, 18, 2])
    luxe_header(ws, "D", "\U0001f4b8  TIME & MONEY",
                "What dating actually costs you a month — in both currencies.")
    ws.cell(row=4, column=2, value="MONEY").style = "section_gold"
    table_headers(ws, 5, ["Where it goes", "Monthly", "Yearly"], start_col=2)
    start = 6
    for i, (item, amt) in enumerate(SPENDING):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        c = ws.cell(row=r, column=3, value=amt); c.style = "input"; c.number_format = '"$"#,##0.00'
        cy = ws.cell(row=r, column=4, value=f"=C{r}*12"); cy.style = "td"; cy.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 5):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    end = start + len(SPENDING) - 1
    nrange(wb, "SpendLines", "Time & Money", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="= TOTAL / MONTH").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(SpendLines)"); ct.style = "td"
    ct.font = Font(bold=True, size=14, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0.00'
    cell_name(wb, "SpendTotal", "Time & Money", f"$C${tot}")
    cty = ws.cell(row=tot, column=4, value="=SpendTotal*12"); cty.style = "td"
    cty.font = Font(bold=True, size=12, color=PRIMARY); cty.fill = fill(WARN_BG); cty.number_format = '"$"#,##0'
    ws.cell(row=tot + 1, column=2, value="Your monthly budget").style = "field_label"
    cb = ws.cell(row=tot + 1, column=3, value="=Budget"); cb.style = "field_value"; cb.number_format = '"$"#,##0'
    ws.cell(row=tot + 2, column=2, value="= UNDER OR OVER").style = "th"
    cu = ws.cell(row=tot + 2, column=3, value="=Budget-SpendTotal"); cu.style = "td"
    cu.font = Font(bold=True, size=13, color=PRIMARY); cu.fill = fill(MINT_BG); cu.number_format = '"$"#,##0.00'
    cell_name(wb, "BudgetLeft", "Time & Money", f"$C${tot+2}")

    t = tot + 4
    ws.cell(row=t, column=2, value="TIME").style = "section_gold"
    trows = [("Swiping", "=SwipeHours"), ("Messaging", "=MessageHours"),
             ("On dates", "=FirstDates*HoursPerDate"), ("Getting ready", "=FirstDates*PrepPerDate")]
    for i, (lab, fml) in enumerate(trows):
        r = t + 1 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = '0.0" hrs"'
    tt = t + 1 + len(trows)
    ws.cell(row=tt, column=2, value="= TOTAL HOURS THIS MONTH").style = "th"
    cth = ws.cell(row=tt, column=3, value="=HoursSpent"); cth.style = "td"
    cth.font = Font(bold=True, size=14, color=PRIMARY); cth.fill = fill(SURFACE); cth.number_format = '0.0" hrs"'
    ws.cell(row=tt + 1, column=2, value="…which is this many full days").style = "field_label"
    cd = ws.cell(row=tt + 1, column=3, value="=HoursSpent/24"); cd.style = "field_value"; cd.number_format = '0.0" days"'
    ws.cell(row=tt + 3, column=2, value="You are allowed to spend this. You're just not allowed to spend it by accident.").style = "section_gold"


def build_convos(wb):
    ws = wb.create_sheet("Conversations"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 18, 18, 18, 16, 18, 2])
    luxe_header(ws, "G", "\U0001f4ac  CONVERSATIONS",
                "Who's actually talking to you — and who's been texting for a month without asking you out.")
    table_headers(ws, 4, ["Who", "Days talking", "You sent", "They sent", "Met yet?", "Their share"], start_col=2)
    start = L0
    for i, (who, days, you, them, met) in enumerate(CONVOS):
        r = start + i
        ws.cell(row=r, column=2, value=who).style = "td_left"
        cd = ws.cell(row=r, column=3, value=days); cd.style = "input"; cd.number_format = "0"
        cy = ws.cell(row=r, column=4, value=you); cy.style = "input"; cy.number_format = "#,##0"
        ct = ws.cell(row=r, column=5, value=them); ct.style = "input"; ct.number_format = "#,##0"
        ws.cell(row=r, column=6, value=met).style = "input"
        cs = ws.cell(row=r, column=7, value=f"=IFERROR(E{r}/(D{r}+E{r}),0)"); cs.style = "td"; cs.number_format = "0%"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CONVOS) - 1
    nrange(wb, "ConvoDays", "Conversations", "C", start, end)
    nrange(wb, "ConvoMet", "Conversations", "F", start, end)
    nrange(wb, "ConvoShare", "Conversations", "G", start, end)
    add_dv(ws, f"F{start}:F{end}", "YesNoList")
    ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="lessThan", formula=["0.35"],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 2
    ws.cell(row=tot, column=2, value="Talking but never met").style = "field_label"
    c1 = ws.cell(row=tot, column=6, value='=COUNTIF(ConvoMet,"No")'); c1.style = "field_value"
    c1.number_format = "0"; c1.fill = fill(WARN_BG)
    ws.cell(row=tot + 1, column=2, value="Longest one that's gone nowhere (days)").style = "field_label"
    c2 = ws.cell(row=tot + 1, column=6, value="=MAX(ConvoDays)"); c2.style = "field_value"; c2.number_format = "0"
    ws.cell(row=tot + 3, column=2, value="Anyone under 35% is texting you, not dating you. Red rows are pen pals.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_nonneg(wb):
    ws, start, end = build_log(
        wb, "Non-Negotiables", "\U0001f512", "NON-NEGOTIABLES", "Write these before you meet anyone. Read them when you're tempted to bend one.",
        ["What", "Weight", "Why it matters to me"], NON_NEGOTIABLES,
        [2, 36, 22, 54, 2], text_left={2, 4},
        validations=[("C", "WeightList")], start_col=2)
    nrange(wb, "NonNegWeight", "Non-Negotiables", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"Dealbreaker"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tr = end + 2
    ws.cell(row=tr, column=2, value="Actual dealbreakers").style = "field_label"
    c1 = ws.cell(row=tr, column=3, value='=COUNTIF(NonNegWeight,"Dealbreaker")'); c1.style = "field_value"; c1.number_format = "0"
    ws.cell(row=tr + 2, column=2, value="If it's a dealbreaker, it's a dealbreaker on date one, not date twelve.").style = "section_gold"


def build_safety(wb):
    ws = wb.create_sheet("Safety Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 62, 18, 2])
    luxe_header(ws, "C", "\U0001f6e1  SAFETY PLAN",
                "Please fill this in before your next first date. All of it. Every time.")
    table_headers(ws, 4, ["Step", "Doing it?"], start_col=2)
    start = L0
    for i, (step, done) in enumerate(SAFETY):
        r = start + i
        ws.cell(row=r, column=2, value=step).style = "td_left"
        ws.cell(row=r, column=3, value=done).style = "input"
        if i % 2:
            for c in (2, 3):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SAFETY) - 1
    nrange(wb, "SafetyDone", "Safety Plan", "C", start, end)
    add_dv(ws, f"C{start}:C{end}", "YesNoList")
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"No"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="= STEPS YOU'RE ACTUALLY DOING").style = "th"
    cs = ws.cell(row=tot, column=3, value='=COUNTIF(SafetyDone,"Yes")'); cs.style = "td"
    cs.font = Font(bold=True, size=14, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = "0"
    cell_name(wb, "SafetyHeld", "Safety Plan", f"$C${tot}")
    r = tot + 2
    ws.cell(row=r, column=2, value="WHO KNOWS WHERE YOU ARE TONIGHT").style = "section_gold"
    for i, lab in enumerate(["Friend's name & number", "What time you'll check in", "Where you're going",
                             "Their name, photo & profile link", "How you're getting home"]):
        rr = r + 1 + i
        ws.cell(row=rr, column=2, value=lab).style = "field_label"
        ws.cell(row=rr, column=3, value="").style = "input"
    ws.cell(row=r + 7, column=2, value="You never owe anyone a second more of your evening. Leaving early is always allowed.").style = "section_gold"


def build_reflection(wb):
    ws, start, end = build_log(
        wb, "Reflection", "\U0001f4ad", "REFLECTION & PATTERNS", "The page that actually changes things. Do it at the end of every month.",
        ["Ask yourself", "Honestly"], REFLECTION,
        [2, 46, 62, 2], text_left={2, 3}, start_col=2)
    ws.cell(row=end + 2, column=2, value="Patterns don't break because you noticed them once. They break because you wrote them down twice.").style = "section_gold"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 42, 18, 2])
    luxe_header(ws, "C", "\U0001f4c8  MONTHLY SUMMARY",
                "The whole month in one place — and what you want the next one to look like.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    rows = [
        ("Matches", "=Matches", "#,##0"),
        ("Conversations", "=Conversations", "#,##0"),
        ("First dates", "=FirstDates", "#,##0"),
        ("Second dates", "=SecondDates", "#,##0"),
        ("Still seeing", "=StillSeeing", "#,##0"),
        ("Hours spent", "=HoursSpent", '0" hrs"'),
        ("Money spent", "=SpendTotal", '"$"#,##0.00'),
    ]
    for i, (lab, fml, fmt) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
    ws.cell(row=13, column=2, value="= COST OF EVERY SECOND DATE").style = "th"
    c1 = ws.cell(row=13, column=3, value="=CostPerSecondDate"); c1.style = "td"
    c1.font = Font(bold=True, size=15, color=PRIMARY); c1.fill = fill(WARN_BG); c1.number_format = '"$"#,##0.00'
    ws.cell(row=14, column=2, value="= HOURS FOR EVERY SECOND DATE").style = "th"
    c2 = ws.cell(row=14, column=3, value="=HoursPerSecondDate"); c2.style = "td"
    c2.font = Font(bold=True, size=15, color=PRIMARY); c2.fill = fill(WARN_BG); c2.number_format = '0.0" hrs"'
    ws.cell(row=15, column=2, value="= YOU ARE DOING THIS MUCH OF THE WORK").style = "th"
    c3 = ws.cell(row=15, column=3, value="=EffortRatio"); c3.style = "td"
    c3.font = Font(bold=True, size=15, color=PRIMARY); c3.fill = fill(MINT_BG); c3.number_format = '0.0"\\u00d7"'
    ws.cell(row=17, column=2, value="SPEND, MONTH BY MONTH").style = "section_gold"
    table_headers(ws, 18, ["Month", "Spend"], start_col=2)
    ts = 19
    for i, (m, sp) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=sp); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 4):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "SpendTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Spend by Month", ts, te, 3, 2), "E5")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ✷  DATING LIFE COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Your month, honestly — and whether they're meeting you halfway.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MATCHES", "=Matches", "num"),
        ("CONVERSATIONS", "=Conversations", "num"),
        ("FIRST DATES", "=FirstDates", "num"),
        ("SECOND DATES", "=SecondDates", "num"),
        ("SECOND-DATE RATE", "=SecondDateRate", "pct1"),
        ("STILL SEEING", "=StillSeeing", "num"),
    ]
    row2 = [
        ("HOURS THIS MONTH", "=HoursSpent", "num"),
        ("SPENT THIS MONTH", "=SpendTotal", "money"),
        ("PER SECOND DATE", "=CostPerSecondDate", "money2"),
        ("HOURS PER SECOND DATE", "=HoursPerSecondDate", "dec"),
        ("YOU'RE DOING", "=EffortRatio", "dec"),
        ("DATING SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "HOW IT'S ACTUALLY GOING", "section_gold")
    merge_set(ws, "H11:M11", "FIRST DATES BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("You're meeting people, not texting them", "=IFERROR(MIN(DaysGoal/DaysToDate,1),0)"),
        ("Second dates happening", "=IFERROR(MIN(SecondDateRate/SecondDateGoal,1),0)"),
        ("Green flags outweigh red", "=IFERROR(MIN(FlagNet/FlagGoal,1),0)"),
        ("Spending inside your budget", "=IFERROR(MIN(Budget/SpendTotal,1),0)"),
        ("Safety steps held", "=IFERROR(MIN(SafetyHeld/9,1),0)"),
        ("Effort is reciprocal", "=IFERROR(MIN(EffortGoal/EffortRatio,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Good",IF(C{r}>=0.6,"OK","Look at this"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    fw = wb["Dating Funnel"]
    ch = BarChart(); ch.type = "col"; ch.title = "First Dates by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(fw, min_col=8, min_row=7, max_row=6 + len(FUNNEL_MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(fw, min_col=5, min_row=7, max_row=6 + len(FUNNEL_MONTHS)))
    ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Dating Life Command Center™ — a mirror, not a verdict.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_money(wb); build_funnel(wb)
    build_effort(wb); build_flags(wb); build_people(wb); build_dates(wb)
    build_convos(wb); build_nonneg(wb); build_safety(wb); build_reflection(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Dating Funnel", "Effort & Reciprocity", "Green & Red Flags",
             "People", "Date Log", "Conversations", "Time & Money", "Non-Negotiables", "Safety Plan",
             "Reflection", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Dating_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
