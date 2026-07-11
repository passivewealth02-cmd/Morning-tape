"""Build Instagram Command Center™ — The Ultimate Instagram Creator Business System.

24 sheets + Welcome · a premium Instagram creator OS in Excel & Sheets.
Ideas, captions, calendar, grid planner, reels, stories, hashtags, analytics,
community, IG Shop, affiliate, brand deals (UGC CRM), finance, goals & more —
one dashboard.

Run: python3 build_xlsx.py   ->  ../Instagram_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

PILLARS = ["Home", "Plants", "Styling", "DIY", "Lifestyle", "Behind-the-scenes", "Shopping"]
CONTENT_STATUS = ["Idea", "Filming", "Editing", "Scheduled", "Posted", "Boosted", "Repurposed"]
POST_TYPES = ["Reel", "Carousel", "Story", "Static", "LIVE", "Guide", "Collab"]
REV_CATS = ["Brand Deals", "Affiliate", "Digital Products", "Instagram Shop", "Coaching", "Subscriptions", "Other"]
EXP_CATS = ["Software", "Equipment", "Props & Materials", "Contractors", "Marketing", "Education", "Miscellaneous"]
SPON_STAGES = ["Lead", "Outreach", "Negotiation", "Signed", "Delivered", "Paid"]
CAPTION_TYPES = ["Question", "Bold Claim", "Listicle", "Story", "Tutorial", "Relatable", "CTA"]
GOAL_CATS = ["Followers", "Reach", "Revenue", "Posting", "Brand Deals", "Shop", "Engagement"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


def totals(ws, row, cols, start, end, fmt='"$"#,##0', label="TOTAL"):
    ws.cell(row=row, column=1, value=label).style = "th"
    for col in cols:
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt


# ===========================================================================
# 24 — Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your account details & targets once — every dashboard follows.")
    merge_set(ws, "B5:C5", "CREATOR INPUTS", "section")
    controls = [
        ("Creator / Handle", "@studiofern", None, "Handle"),
        ("Creator Name", "Fern Alvarez", None, "CreatorName"),
        ("Niche", "Plants, home & styling", None, "Niche"),
        ("Monthly Revenue Goal", 9000, '"$"#,##0', "RevenueGoal"),
        ("Monthly Post Goal", 20, "0", "PostGoal"),
        ("Active Deal Target", 5, "0", "DealTarget"),
        ("Reel Retention Target", 0.55, "0%", "RetTarget"),
        ("Follower Growth Goal (mo)", 12000, "#,##0", "GrowthGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Content Pillar", PILLARS, "PillarList"), ("F", "Post Type", POST_TYPES, "PostTypeList"),
             ("G", "Status", CONTENT_STATUS, "StatusList"), ("H", "Revenue Category", REV_CATS, "RevCatList"),
             ("I", "Expense Category", EXP_CATS, "ExpCatList"), ("J", "Caption Type", CAPTION_TYPES, "CaptionTypeList"),
             ("K", "Goal Category", GOAL_CATS, "GoalCatList"), ("L", "Deal Stage", SPON_STAGES, "SponStageList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 16, "Priority", PRIORITIES, "PriorityList"), ("F", 16, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  📸  INSTAGRAM COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate Instagram creator business system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "RUN YOUR ENTIRE INSTAGRAM BUSINESS FROM ONE FILE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("From grid to growth — Instagram Command Center™ manages ideas, captions, your posting calendar, "
                      "feed grid, reels, stories, hashtags, analytics, community, IG Shop, affiliate, brand deals and "
                      "finances in ONE premium Excel & Google Sheets system. Post a cohesive feed, grow reach, land "
                      "brand deals and turn saves into real income — all with creator-grade automation. This isn't a "
                      "content planner — it's your complete Instagram Operating System.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your handle, niche & monthly goals.",
             "2.  Plan the Content Calendar & Grid — see your feed before you post.",
             "3.  Bank Ideas & Captions; track Reels & Stories performance.",
             "4.  Log Analytics, IG Shop & Affiliate — revenue & net profit update live.",
             "5.  Work Brand Deals from lead → paid in the UGC CRM.",
             "6.  Watch the Dashboard track followers, revenue & a Creator Health Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (@studiofern, a 148k-follower plant & home creator earning ~$8,100/mo) is "
               "included so you can see how everything connects — just type over it with your own. Followers, revenue, "
               "net profit, engagement, posting consistency, the brand-deal pipeline and the Creator Health Score all "
               "update automatically. Every sheet is print-friendly and works in Excel and Google Sheets, on desktop "
               "and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more time to create — let's grow your account.", "section_gold")


# ===========================================================================
# 2 — Creator Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "👤  CREATOR PROFILE", "Your account, defined — the identity every post flows from.")
    blocks = [
        ("THE ACCOUNT", [("Handle", "=Handle"), ("Creator", "=CreatorName"),
                         ("Niche", "=Niche"), ("Posting Schedule", "Reels M/W/F · Stories daily"),
                         ("Started", "2021"), ("Business Email", "hi@studiofern.co")]),
        ("STRATEGY", [("Mission", "Make plant parenthood easy"), ("Target Audience", "Women 25-40, renters & homeowners"),
                      ("Content Pillars", "Plants · Home · Styling"), ("Manager", "Self-managed"),
                      ("Media Kit", "studiofern.co/kit"), ("Rate Card", "$1.5k / in-feed reel")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "PLATFORMS & LINK IN BIO", "section_gold"); ws.row_dimensions[15].height = 22
    handles = [("Instagram", "@studiofern"), ("TikTok", "@studiofern"), ("Pinterest", "@studiofern"),
               ("YouTube", "Studio Fern"), ("Shop", "studiofern.co/shop"), ("Newsletter", "The Repot")]
    for i, (p, h) in enumerate(handles):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 3 — Content Calendar
# ===========================================================================
def build_calendar(wb):
    titles = [
        "5 unkillable plants for renters", "Styling a shelfie in 60s", "Repotting a monstera",
        "My $40 plant corner", "Carousel: plant care basics", "GRWM: plant shop haul",
        "Propagation station tour", "Thrifted pot makeover", "Reel: watering myths",
        "Small-space plant styling", "Behind my content setup", "3 trending home finds",
        "How I style open shelves", "Reader Q&A: yellow leaves", "Cozy corner reveal",
        "Plant-shopping with me", "Macrame hanger DIY", "My grow-light setup",
        "Before/after: living room", "Fiddle leaf survival guide", "Weekend plant reset",
        "Carousel: 7 low-light plants", "Faux vs real plants", "Balcony jungle reveal",
    ]
    posted_off = [1, 2, 4, 5, 7, 8, 10, 11, 13, 14, 15, 17, 18, 20, 21, 22, 24, 25]
    boosted_idx = {0, 8}
    future = [(-1, "Scheduled"), (-2, "Scheduled"), (-3, "Editing"), (-5, "Idea"), (-6, "Idea")]
    rows = []
    ti = 0
    for k, off in enumerate(posted_off):
        status = "Boosted" if k in boosted_idx else "Posted"
        rows.append((dminus(off), titles[ti % len(titles)], PILLARS[ti % 5], POST_TYPES[ti % len(POST_TYPES)], status)); ti += 1
    for foff, status in future:
        rows.append((dplus(-foff), titles[ti % len(titles)], PILLARS[ti % 5], POST_TYPES[ti % len(POST_TYPES)], status)); ti += 1
    sample = [(d, t, p, pt, "High" if st in ("Posted", "Boosted", "Scheduled") else "Medium", st) for (d, t, p, pt, st) in rows]
    ws, start, end = build_log(
        wb, "Calendar", "🗓", "CONTENT CALENDAR",
        "Plan every post in one view — posting status calculates itself.",
        ["Date", "Caption / Concept", "Pillar", "Type", "Priority", "Status"],
        sample, [13, 32, 16, 12, 11, 13],
        text_left={2}, dates={1},
        validations=[("C", "PillarList"), ("D", "PostTypeList"), ("E", "PriorityList"), ("F", "StatusList")], reserved=60)
    nrange(wb, "CalDate", "Calendar", "A", start, end)
    nrange(wb, "CalStatus", "Calendar", "F", start, end)
    cmap = {"Boosted": HIGHLIGHT, "Posted": MINT_BG, "Scheduled": WARN_BG, "Editing": SOFT_BG, "Filming": SOFT_BG, "Idea": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 4 — Grid Planner
# ===========================================================================
def build_grid(wb):
    ws = wb.create_sheet("Grid Planner"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "🔲  GRID PLANNER",
                "See your feed before you post — plan the next 9 for a cohesive aesthetic.")
    merge_set(ws, "B5:D5", "HOW TO PLAN YOUR GRID", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Drop a cover into each cell (Excel: Insert ▸ Pictures ▸ Place in Cell · Sheets: Insert ▸ Image "
                      "in cell or =IMAGE(\"link\")) to preview your 3×3 grid. Note the pillar & post type under each so "
                      "colours, subjects and types alternate for a balanced feed.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    grid = [("Reel", "Plants"), ("Carousel", "Styling"), ("Static", "Home"),
            ("Reel", "DIY"), ("Static", "Plants"), ("Carousel", "Shopping"),
            ("Reel", "Lifestyle"), ("Static", "Home"), ("Carousel", "Plants")]
    idx = 0
    for band in range(3):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 118
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            pt, pil = grid[idx]
            ic = ws.cell(row=img_row, column=col, value=f"🖼\nPost {idx+1}\n(add cover)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value=f"{pt} · {pil}")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 28
            idx += 1


# ===========================================================================
# 5 — Video Pipeline
# ===========================================================================
def build_pipeline(wb):
    rows = [
        ("Repotting a monstera", "Editing", 0.80, "Fern", 1, "Add captions + trending audio"),
        ("Carousel: 7 low-light plants", "Design", 0.55, "Fern", 3, "3 of 7 slides done"),
        ("Balcony jungle reveal", "Script", 0.35, "Fern", 5, "Hook drafted"),
        ("Weekend plant reset", "Outline", 0.20, "Fern", 6, "Shot list"),
        ("Thrifted pot makeover", "Idea", 0.10, "Fern", 8, "Sourcing pots"),
        ("Reel: watering myths", "Editing", 0.70, "Editor — Rae", 2, "Trim to 22s"),
        ("Cozy corner reveal", "Cover", 0.90, "Fern", 1, "Grid cover chosen"),
        ("My grow-light setup", "Filming", 0.50, "Fern", 4, "Half filmed"),
    ]
    sample = [(t, st, prog, o, dplus(d), note) for (t, st, prog, o, d, note) in rows]
    ws, start, end = build_log(
        wb, "Pipeline", "🎬", "CONTENT PIPELINE",
        "Every post, every stage — from idea to posted, with live progress %.",
        ["Title", "Stage", "Progress", "Owner", "Due", "Notes"],
        sample, [30, 14, 12, 16, 13, 24],
        text_left={1, 4, 6}, dates={5}, pcts={3}, reserved=40)
    nrange(wb, "PipeProgress", "Pipeline", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 6 — Idea Vault
# ===========================================================================
def build_ideas(wb):
    rows = [
        ("Renter-friendly plant wall", "Home", "renter plants", "My landlord approved THIS", 420000, 6, "Planned"),
        ("Plants that survived me", "Plants", "hard to kill", "I neglected these for a month", 680000, 3, "Idea"),
        ("$50 plant corner makeover", "Styling", "plant corner", "$50 glow-up", 550000, 5, "Planned"),
        ("Propagation for beginners", "Plants", "propagation", "Free plants forever", 610000, 4, "Idea"),
        ("Styling books + plants", "Styling", "shelf styling", "Shelfie in 60 seconds", 480000, 5, "Idea"),
        ("Thrift with me: pots", "Shopping", "thrift pots", "$3 pots that look $30", 520000, 4, "Planned"),
        ("Low-light plant guide", "Plants", "low light plants", "No window? No problem", 460000, 5, "Idea"),
        ("Behind my content days", "Behind-the-scenes", "creator day", "A realistic filming day", 300000, 3, "Idea"),
        ("Faux vs real debate", "Home", "faux plants", "Would you notice?", 720000, 4, "Idea"),
        ("Best plant shops (haul)", "Shopping", "plant haul", "I spent way too much", 640000, 5, "Idea"),
    ]
    ws = wb.create_sheet("Ideas"); ws.sheet_view.showGridLines = False
    set_widths(ws, [30, 16, 18, 30, 14, 9, 12, 12])
    luxe_header(ws, "H", "💡  IDEA VAULT",
                "Never run dry — capture ideas and let an AI Opportunity Score (reach ÷ difficulty) rank them.")
    table_headers(ws, 4, ["Idea", "Pillar", "Keyword", "Hook", "Est. Reach", "Diff.", "AI Score", "Status"])
    start = L0
    for i, (idea, pil, kw, hook, reach, diff, status) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=idea).style = "td_left"
        ws.cell(row=r, column=2, value=pil).style = "td"
        ws.cell(row=r, column=3, value=kw).style = "td_left"
        ws.cell(row=r, column=4, value=hook).style = "td_left"
        cv = ws.cell(row=r, column=5, value=reach); cv.style = "td"; cv.number_format = "#,##0"
        ws.cell(row=r, column=6, value=diff).style = "td"
        sc = ws.cell(row=r, column=7, value=f"=IFERROR(ROUND(E{r}/1000/F{r},0),0)"); sc.style = "td"; sc.number_format = "#,##0"
        ws.cell(row=r, column=8, value=status).style = "td"
        if i % 2:
            for c in range(1, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    style_rows(ws, end + 1, end + 40, 8, text_left={1, 3, 4})
    for r in range(end + 1, end + 41):
        ws.cell(row=r, column=7, value=f"=IFERROR(ROUND(E{r}/1000/F{r},0),0)").number_format = "#,##0"
    add_dv(ws, f"B{start}:B{end+40}", "PillarList"); add_dv(ws, f"H{start}:H{end+40}", "StatusList")
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"G{start}:G{end}", ColorScaleRule(
        start_type="num", start_value=60, start_color="FF" + WARN_BG,
        end_type="num", end_value=200, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 7 — Captions & Hooks
# ===========================================================================
def build_captions(wb):
    rows = [
        ("Save this for your next repot 🌱", "CTA", "Repotting reels", "Yes", 9.2, "Drives saves — the IG signal"),
        ("You're overwatering. Here's proof.", "Bold Claim", "Care myths", "Yes", 9.0, "Curiosity + correction"),
        ("POV: your first plant corner", "Story", "Styling", "Yes", 8.7, "Aspirational + relatable"),
        ("7 plants that survive low light 👇", "Listicle", "Carousels", "Yes", 8.9, "Number + clear promise"),
        ("Which one would you pick?", "Question", "Hauls", "No", 8.1, "Drives comments"),
        ("The $3 pot that looks $30", "Bold Claim", "Thrift", "Yes", 8.6, "Specific $ beats vague"),
        ("Comment PLANT for the list", "CTA", "Lead magnet", "Yes", 8.8, "DM automation trigger"),
        ("How I styled this in 60 seconds", "Tutorial", "Reels", "No", 8.3, "Time promise = watch"),
    ]
    ws, start, end = build_log(
        wb, "Captions", "✍", "CAPTIONS & HOOKS",
        "Win the first line & the CTA — bank captions that earn saves, shares & comments.",
        ["Caption / Hook", "Type", "Best For", "Reusable?", "Score", "Why It Works"],
        rows, [32, 14, 20, 12, 10, 26],
        text_left={1, 3, 6}, dec={5},
        validations=[("B", "CaptionTypeList"), ("D", "YesNoList")], reserved=40)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color=GOLD_LT, showValue=True))


# ===========================================================================
# 8 — Reels
# ===========================================================================
def build_reels(wb):
    rows = [
        ("5 unkillable plants for renters", 620000, 540000, 0.61, 24000, 12000, 3200),
        ("Repotting a monstera", 410000, 360000, 0.58, 15000, 8100, 2100),
        ("Styling a shelfie in 60s", 350000, 310000, 0.55, 13000, 6400, 1800),
        ("$40 plant corner", 280000, 250000, 0.52, 9800, 4200, 1400),
        ("Watering myths", 190000, 170000, 0.47, 6100, 2600, 700),
        ("Propagation station tour", 140000, 120000, 0.44, 4300, 1500, 420),
    ]
    ws = wb.create_sheet("Reels"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 13, 13, 13, 12, 12, 12, 2])
    luxe_header(ws, "I", "🎬  REELS ANALYTICS",
                "Your growth engine — plays, reach, watch-through, saves, shares & follows per reel.")
    table_headers(ws, 4, ["Reel", "Plays", "Reach", "Watch %", "Saves", "Shares", "Follows +"], start_col=2)
    start = L0
    for i, (title, plays, reach, watch, saves, shares, foll) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=title).style = "td_left"
        for ci, (v, fmt) in enumerate([(plays, "#,##0"), (reach, "#,##0"), (watch, "0%"), (saves, "#,##0"), (shares, "#,##0"), (foll, "#,##0")], 3):
            cc = ws.cell(row=r, column=ci, value=v); cc.style = "td"; cc.number_format = fmt
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    nrange(wb, "ReelTitle", "Reels", "B", start, end)
    nrange(wb, "ReelPlays", "Reels", "C", start, end)
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=620000, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(f"E{start}:E{end}", ColorScaleRule(
        start_type="num", start_value=0.4, start_color="FF" + WARN_BG, end_type="num", end_value=0.65, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 9 — Stories Planner
# ===========================================================================
def build_stories(wb):
    rows = [
        (dminus(0), "Morning plant check", "Behind-the-scenes", "Poll: over/underwater?", "None", "Warm-up + engagement"),
        (dminus(0), "New arrival unboxing", "Shopping", "Link sticker", "Shop", "Drives clicks"),
        (dminus(1), "Repot in progress", "Plants", "Question box", "None", "Save the how-to for a reel"),
        (dminus(1), "Shelf styling before/after", "Styling", "Slider: rate it", "None", "Fun + feedback"),
        (dminus(2), "This or that plant pots", "Home", "Poll", "Shop", "Product research"),
        (dminus(2), "Reader question answered", "Plants", "Question box", "None", "Repurpose top Qs"),
        (dminus(3), "Weekend haul recap", "Shopping", "Link sticker", "Affiliate", "LTK link"),
    ]
    ws, start, end = build_log(
        wb, "Stories", "📱", "STORIES PLANNER",
        "Plan daily stories with intent — frames, interactive stickers, links & purpose.",
        ["Date", "Frame / Idea", "Pillar", "Sticker", "Link", "Purpose"],
        rows, [13, 26, 16, 20, 12, 26],
        text_left={2, 4, 6}, dates={1}, reserved=30,
        validations=[("C", "PillarList")])
    for lk, cc in {"Shop": MINT_BG, "Affiliate": WARN_BG}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{lk}"'], fill=fill(cc)))


# ===========================================================================
# 10 — Hashtags & SEO
# ===========================================================================
def build_hashtags(wb):
    rows = [
        ("Plant set", "#plantsofinstagram #houseplants #plantcare #plantmom #urbanjungle", "Plants", "High", "Core niche"),
        ("Renter set", "#renterfriendly #rentersofinstagram #smallspaces #apartmenttherapy", "Home", "Low", "High intent"),
        ("Styling set", "#shelfie #homestyling #cozyhome #homedecor", "Styling", "Medium", "Aesthetic reach"),
        ("Reels set", "#reels #reelsinstagram #plantreels #satisfying", "Plants", "High", "Boosts discovery"),
        ("Budget set", "#thrifted #budgetdecor #diyhome #upcycle", "Shopping", "Low", "Great reach:comp"),
        ("Keyword SEO", "plant care · low light plants · renter decor · shelf styling", "Home", "—", "Put keywords in caption + name"),
    ]
    ws, start, end = build_log(
        wb, "Hashtags", "#️⃣", "HASHTAGS & SEO",
        "Get discovered — pre-built hashtag sets + keyword SEO for captions, alt text & your name field.",
        ["Set", "Hashtags / Keywords", "Pillar", "Competition", "Notes"],
        rows, [16, 44, 14, 14, 24],
        text_left={2, 5}, reserved=20,
        validations=[("C", "PillarList")])


# ===========================================================================
# 11 — Analytics
# ===========================================================================
def build_analytics(wb):
    rows = [
        ("5 unkillable plants for renters", 540000, 0.128, 0.61, 3200, 24000),
        ("Repotting a monstera", 360000, 0.104, 0.58, 2100, 15000),
        ("Styling a shelfie in 60s", 310000, 0.112, 0.55, 1800, 13000),
        ("$40 plant corner", 250000, 0.096, 0.52, 1400, 9800),
        ("Carousel: plant care basics", 180000, 0.142, 0.00, 1200, 21000),
        ("Watering myths", 170000, 0.081, 0.47, 700, 6100),
    ]
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 3, 18, 12, 12, 2])
    luxe_header(ws, "G", "📊  ANALYTICS COMMAND CENTER",
                "Your account by the numbers — a live snapshot, health dimensions & the Creator Health Score.")
    merge_set(ws, "B5:C5", "SNAPSHOT (28 DAYS)", "section")
    snap = [("Followers", 148000, "#,##0", "FollowerNow"), ("Followers gained", 9200, "#,##0", "FollowerGrowth"),
            ("Reach", 1150000, "#,##0", "Reach28"), ("Engagement rate", 0.064, "0.0%", "EngRate"),
            ("Saves", 42000, "#,##0", "Saves28"), ("Shares", 28000, "#,##0", "Shares28"),
            ("Reel retention", 0.52, "0%", "Completion"), ("Link clicks", 14000, "#,##0", "LinkClicks")]
    for i, (lab, val, fmt, nm) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "field_value"; c.number_format = fmt
        cell_name(wb, nm, "Analytics", f"$C${r}")
        if lab in ("Followers", "Saves"):
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    merge_set(ws, "E5:G5", "CREATOR HEALTH", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=5)
    metrics = [
        ("Revenue vs goal", "=IFERROR(MIN(RevenueTotal/RevenueGoal,1),0)"),
        ("Posting consistency", '=IFERROR(MIN((COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Posted")+COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Boosted"))/PostGoal,1),0)'),
        ("Engagement", "=IFERROR(MIN(EngRate/0.07,1),0)"),
        ("Reel retention", "=IFERROR(MIN(Completion/RetTarget,1),0)"),
        ("Brand pipeline", '=IFERROR(MIN((COUNTIF(SponStage,"Signed")+COUNTIF(SponStage,"Delivered")+COUNTIF(SponStage,"Negotiation"))/DealTarget,1),0)'),
        ("Follower growth", "=IFERROR(MIN(FollowerGrowth/GrowthGoal,1),0)"),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(metrics):
        r = hs + i
        ws.cell(row=r, column=5, value=dim).style = "td_left"
        c = ws.cell(row=r, column=6, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=7, value=f'=IF(F{r}>=0.75,"Strong",IF(F{r}>=0.5,"Growing","Focus"))').style = "td"
        if i % 2:
            for c2 in range(5, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(metrics) - 1
    cell_name(wb, "HealthRange", "Analytics", f"$F${hs}:$F${he}")
    ws.conditional_formatting.add(f"F{hs}:F{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "B15:G15", "TOP POSTS (28 DAYS)", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Post", "Reach", "Eng. Rate", "Watch %", "Follows +", "Saves"], start_col=2)
    vs = 17
    for i, (title, reach, eng, watch, foll, saves) in enumerate(rows):
        r = vs + i
        ws.cell(row=r, column=2, value=title).style = "td_left"
        cv = ws.cell(row=r, column=3, value=reach); cv.style = "td"; cv.number_format = "#,##0"
        ce = ws.cell(row=r, column=4, value=eng); ce.style = "td"; ce.number_format = "0.0%"
        cw = ws.cell(row=r, column=5, value=watch); cw.style = "td"; cw.number_format = "0%"
        cf = ws.cell(row=r, column=6, value=foll); cf.style = "td"; cf.number_format = "#,##0"
        cs = ws.cell(row=r, column=7, value=saves); cs.style = "td"; cs.number_format = "#,##0"
        if i % 2:
            for c2 in range(2, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    ve = vs + len(rows) - 1
    nrange(wb, "PostTitle", "Analytics", "B", vs, ve)
    nrange(wb, "PostReach", "Analytics", "C", vs, ve)
    ws.conditional_formatting.add(f"C{vs}:C{ve}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=540000, color=PRIMARY, showValue=True))
    merge_set(ws, "B25:C25", "FOLLOWER GROWTH (K) — 6 MONTHS", "section")
    ws.cell(row=26, column=2, value="Month").style = "th"; ws.cell(row=26, column=3, value="Followers (K)").style = "th"
    months = _recent_months(6); vals = [98, 110, 121, 131, 139, 148]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 27 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = "0"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "FollowMonth", "Analytics", "$B$27:$B$32")
    cell_name(wb, "FollowVal", "Analytics", "$C$27:$C$32")


# ===========================================================================
# 12 — Community
# ===========================================================================
def build_community(wb):
    rows = [
        ("Yellow leaves = ?", "FAQ", "Usually overwatering — let it dry out & check drainage.", "Saved reply", "High"),
        ("Where's that pot from?", "FAQ", "Linked in my shop — studiofern.co/shop 🌱", "Saved reply", "High"),
        ("Comment PLANT", "DM automation", "Auto-DM the low-light plant guide link", "Automated", "High"),
        ("Collab request", "DM", "Send media kit + rate card", "Manual", "Medium"),
        ("Negative comment", "Comment", "Kill with kindness or hide — never argue", "Policy", "Low"),
        ("Superfan shout-out", "Engagement", "Reply + save to Close Friends list", "Manual", "Medium"),
        ("Product question", "DM", "Answer + soft CTA to shop", "Manual", "High"),
    ]
    ws, start, end = build_log(
        wb, "Community", "💬", "COMMUNITY & DMs",
        "Turn comments & DMs into connection — saved replies, FAQs & automation triggers.",
        ["Trigger / Question", "Type", "Response", "Handling", "Priority"],
        rows, [24, 16, 34, 14, 12],
        text_left={1, 3}, reserved=24,
        validations=[("E", "PriorityList")])
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"High"'], fill=fill(MINT_BG)))


# ===========================================================================
# 13 — Instagram Shop
# ===========================================================================
def build_shop(wb):
    rows = [
        ("Ceramic planter set", "Home", 48.00, 0.12, 40, "=C5*E5", "=F5*D5"),
        ("Full-spectrum grow light", "Plants", 65.00, 0.15, 18, "=C6*E6", "=F6*D6"),
        ("Macrame plant hanger", "Home", 28.00, 0.10, 55, "=C7*E7", "=F7*D7"),
        ("Copper watering can", "Home", 34.00, 0.12, 30, "=C8*E8", "=F8*D8"),
        ("Linen curtains", "Home", 80.00, 0.10, 12, "=C9*E9", "=F9*D9"),
        ("Brass shelf brackets", "Home", 22.00, 0.10, 44, "=C10*E10", "=F10*D10"),
        ("Moisture meter", "Plants", 16.00, 0.15, 68, "=C11*E11", "=F11*D11"),
        ("Woven basket set", "Home", 42.00, 0.12, 26, "=C12*E12", "=F12*D12"),
    ]
    ws = wb.create_sheet("Shop"); ws.sheet_view.showGridLines = False
    set_widths(ws, [26, 12, 12, 12, 11, 13, 13])
    luxe_header(ws, "G", "🛍  INSTAGRAM SHOP",
                "Turn saves into sales — units, GMV and your commission, per product.")
    table_headers(ws, 4, ["Product", "Pillar", "Price", "Comm. %", "Units", "GMV", "Your Earnings"])
    start = L0
    for i, (prod, pil, price, comm, units, gmv, earn) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=prod).style = "td_left"
        ws.cell(row=r, column=2, value=pil).style = "td"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=4, value=comm); cm.style = "input"; cm.number_format = "0%"
        cu = ws.cell(row=r, column=5, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cg = ws.cell(row=r, column=6, value=gmv); cg.style = "td"; cg.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=7, value=earn); ce.style = "td"; ce.number_format = '"$"#,##0'
        if i % 2:
            for c in range(1, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL").style = "th"
    for c in range(2, 6):
        ws.cell(row=tot, column=c).style = "th"
    cg = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(SURFACE); cg.number_format = '"$"#,##0'
    ce = ws.cell(row=tot, column=7, value=f"=SUM(G{start}:G{end})"); ce.style = "td"; ce.font = Font(bold=True, color=PRIMARY); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    cell_name(wb, "ShopGMV", "Shop", f"$F${tot}")
    cell_name(wb, "ShopEarn", "Shop", f"$G${tot}")
    ws.conditional_formatting.add(f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=2000, color=PRIMARY, showValue=True))


# ===========================================================================
# 14 — Affiliate Tracker
# ===========================================================================
def build_affiliate(wb):
    rows = [
        ("LTK", "LTK", "ltk.to/studiofern", 9800, 480, "$780", "Home + decor links"),
        ("Amazon Storefront", "Amazon", "amzn.to/fern", 6400, 320, "$540", "Plant tools bundle"),
        ("Etsy Creator", "Etsy", "etsy.me/fern", 2200, 88, "$210", "Handmade pots"),
        ("Bloomscape", "Impact", "bloom.co/fern", 1600, 42, "$180", "Live plants"),
        ("Skillshare", "Impact", "skl.sh/fern", 2000, 30, "$220", "Plant styling class"),
    ]
    ws, start, end = build_log(
        wb, "Affiliate", "🔗", "AFFILIATE TRACKER",
        "Every link working for you — clicks, conversions & payouts by program.",
        ["Program", "Network", "Link", "Clicks", "Sales", "Payout", "Notes"],
        rows, [22, 12, 18, 12, 10, 12, 24],
        text_left={3, 7}, ints={4, 5}, reserved=20)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10000, color=GOLD_LT, showValue=True))


# ===========================================================================
# 15 — Brand Deals (UGC CRM)
# ===========================================================================
def build_deals(wb):
    rows = [
        ("Bloomscape", "Mara D.", "Spring plant drop", 2400, "Signed", "1 reel + 3 stories", 12),
        ("West Elm", "Rep", "Home refresh", 3200, "Negotiation", "1 reel + carousel", 30),
        ("The Sill", "Lena R.", "Care guide collab", 2000, "Delivered", "1 carousel", -5),
        ("Our Place", "Ivy T.", "Kitchen styling", 1600, "Paid", "1 reel", -25),
        ("Ruggable", "Priya S.", "Rug reveal", 2200, "Outreach", "TBD", 60),
        ("Article", "Owen B.", "Shelf styling", 1900, "Signed", "1 carousel", 18),
        ("HelloFresh", "Sam W.", "Weeknight dinner", 1100, "Lead", "TBD", 75),
        ("Anthropologie", "Chris N.", "Home haul", 2800, "Paid", "1 reel + stories", -40),
    ]
    sample = [(b, c, camp, rate, stage, deliv, dplus(d) if d >= 0 else dminus(-d)) for (b, c, camp, rate, stage, deliv, d) in rows]
    ws, start, end = build_log(
        wb, "Brand Deals", "🤝", "BRAND DEALS · UGC CRM",
        "Turn your audience into income — lead to paid, with rates & deliverables.",
        ["Brand", "Contact", "Campaign", "Rate", "Stage", "Deliverables", "Due / Paid"],
        sample, [16, 14, 20, 12, 14, 20, 13],
        text_left={2, 3, 6}, money={4}, dates={7},
        validations=[("E", "SponStageList")], reserved=30)
    nrange(wb, "SponRate", "Brand Deals", "D", start, end)
    nrange(wb, "SponStage", "Brand Deals", "E", start, end)
    for st, cc in {"Paid": MINT_BG, "Signed": WARN_BG, "Delivered": SOFT_BG, "Lead": WHITE}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 16 — Finance Center
# ===========================================================================
def build_finance(wb):
    ws = wb.create_sheet("Finance"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 16, 12, 3, 22, 14])
    luxe_header(ws, "G", "💰  CREATOR FINANCE CENTER",
                "Every income & expense in one place — monthly, run-rate & net profit, live.")
    table_headers(ws, 4, ["Income Source", "This Month", "Annual (est.)", "% of Rev"])
    income = {"Brand Deals": 3600, "Affiliate": 1400, "Digital Products": 1100, "Instagram Shop": 875,
              "Coaching": 500, "Subscriptions": 200, "Other": 100}
    start = L0; iend = start + len(REV_CATS) - 1
    for i, cat in enumerate(REV_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        val = "=ShopEarn" if cat == "Instagram Shop" else income[cat]
        cm = ws.cell(row=r, column=2, value=val); cm.style = "input" if cat != "Instagram Shop" else "field_value"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/$B${iend+1},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    itot = iend + 1
    ws.cell(row=itot, column=1, value="TOTAL REVENUE").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=itot, column=col, value=f"=SUM({L}{start}:{L}{iend})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "RevSource", "Finance", "A", start, iend)
    nrange(wb, "RevMonthly", "Finance", "B", start, iend)
    cell_name(wb, "RevenueTotal", "Finance", f"$B${itot}")
    merge_set(ws, "F4:G4", "MONTHLY EXPENSES", "section_gold")
    expenses = {"Software": 160, "Equipment": 450, "Props & Materials": 380, "Contractors": 450,
                "Marketing": 100, "Education": 60, "Miscellaneous": 50}
    estart = 5
    for i, cat in enumerate(EXP_CATS):
        r = estart + i
        ws.cell(row=r, column=6, value=cat).style = "td_left"
        c = ws.cell(row=r, column=7, value=expenses[cat]); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=6).fill = fill(MUTED_ROW); ws.cell(row=r, column=7).fill = fill(MUTED_ROW)
    eend = estart + len(EXP_CATS) - 1; etot = eend + 1
    ws.cell(row=etot, column=6, value="TOTAL EXPENSES").style = "th"
    c = ws.cell(row=etot, column=7, value=f"=SUM(G{estart}:G{eend})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "ExpCat", "Finance", "F", estart, eend)
    nrange(wb, "ExpMonthly", "Finance", "G", estart, eend)
    cell_name(wb, "ExpenseTotal", "Finance", f"$G${etot}")
    merge_set(ws, "A16:D16", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Revenue (month)", "=RevenueTotal", '"$"#,##0'), ("Annual run-rate", "=RevenueTotal*12", '"$"#,##0'),
             ("Expenses (month)", "=ExpenseTotal", '"$"#,##0'), ("Net profit", "=RevenueTotal-ExpenseTotal", '"$"#,##0'),
             ("Profit margin", "=IFERROR((RevenueTotal-ExpenseTotal)/RevenueTotal,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 17 + i
        ws.cell(row=r, column=1, value=lab).style = "field_label"
        c = ws.cell(row=r, column=2, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net profit", "Profit margin"):
            ws.cell(row=r, column=2).fill = fill(MINT_BG)


# ===========================================================================
# 17 — Expenses
# ===========================================================================
def build_expenses(wb):
    rows = [
        (dminus(20), "Later + Canva Pro", "Software", "$40", "Scheduling & design", "Monthly"),
        (dminus(16), "Lens + tripod", "Equipment", "$220", "Reels gear", "One-time"),
        (dminus(12), "Plants & pots (content)", "Props & Materials", "$180", "B-roll", "Ongoing"),
        (dminus(9), "Editor — Rae", "Contractors", "$450", "8 reels", "Monthly"),
        (dminus(6), "Props haul", "Props & Materials", "$120", "Styling", "Ongoing"),
        (dminus(4), "Boost — top reel", "Marketing", "$100", "Promote", "Test"),
        (dminus(2), "Styling course", "Education", "$60", "Skill-up", "One-time"),
    ]
    ws, start, end = build_log(
        wb, "Expenses", "🧾", "EXPENSES",
        "Every business cost tracked — because creator income is a business.",
        ["Date", "Item", "Category", "Amount", "For", "Frequency"],
        rows, [13, 22, 18, 12, 18, 14],
        text_left={2, 5}, dates={1}, reserved=30,
        validations=[("C", "ExpCatList")])


# ===========================================================================
# 18 — Equipment
# ===========================================================================
def build_equipment(wb):
    rows = [
        ("iPhone 15 Pro", "Camera", dminus(300), "Good", "Main camera", "Yes"),
        ("Ring light 18\"", "Lighting", dminus(200), "Good", "Key light", "Yes"),
        ("Clip-on lens", "Camera", dminus(90), "New", "Wide reels", "Yes"),
        ("Tripod + mount", "Support", dminus(400), "Fair", "Replace mount", "No"),
        ("Softbox kit", "Lighting", dminus(120), "Good", "Product shots", "Yes"),
        ("MacBook Air", "Editing", dminus(500), "Good", "Editing", "Yes"),
        ("Backdrop rolls", "Set", dminus(60), "New", "Flat lays", "Yes"),
    ]
    ws, start, end = build_log(
        wb, "Equipment", "🎥", "EQUIPMENT",
        "Your kit, tracked — condition & what to upgrade next.",
        ["Item", "Type", "Bought", "Condition", "Notes", "Working?"],
        rows, [22, 14, 13, 12, 22, 12],
        text_left={5}, dates={3}, reserved=20,
        validations=[("F", "YesNoList")])
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 19 — Repurposing
# ===========================================================================
def build_repurpose(wb):
    rows = [
        ("5 unkillable plants for renters", "Yes", "Yes", "Yes", "Yes", "No", "Top reel — push everywhere"),
        ("Repotting a monstera", "Yes", "Yes", "No", "Yes", "Yes", "Blog + email how-to"),
        ("Styling a shelfie in 60s", "Yes", "No", "Yes", "No", "No", "Carousel next"),
        ("$40 plant corner", "Yes", "Yes", "Yes", "No", "Yes", "Pinterest pin ready"),
        ("Carousel: plant care basics", "No", "No", "Yes", "Yes", "Yes", "Turn into a Reel"),
    ]
    ws, start, end = build_log(
        wb, "Repurposing", "♻", "REPURPOSING ENGINE",
        "One post → ten — track where each piece of content has been repurposed.",
        ["Original Post", "TikTok", "YouTube", "Pinterest", "Story", "Newsletter", "Notes"],
        rows, [26, 10, 11, 12, 10, 13, 26],
        text_left={1, 7}, reserved=30,
        validations=[(c, "YesNoList") for c in ("B", "C", "D", "E", "F")])
    for col in ("B", "C", "D", "E", "F"):
        ws.conditional_formatting.add(f"{col}{start}:{col}{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 20 — Brand Kit
# ===========================================================================
def build_brandkit(wb):
    ws = wb.create_sheet("Brand Kit"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 4, 24, 24, 2])
    luxe_header(ws, "G", "🎨  BRAND KIT", "Stay recognizable — your colours, fonts, voice & feed aesthetic in one place.")
    blocks = [
        ("VISUAL", [("Primary Color", "#1B4F48 Forest"), ("Accent Color", "#C9A86A Gold"),
                    ("Feed Aesthetic", "Warm, earthy, light-filled"), ("Preset", "SF-01 (warm matte)"),
                    ("Logo File", "brand/studiofern.png"), ("Watermark", "@studiofern")]),
        ("VOICE", [("Tone", "Warm, calm, encouraging"), ("We say", "'You can keep this alive'"),
                   ("We avoid", "Plant-snob gatekeeping"), ("Signature sign-off", "'Go water something'"),
                   ("Emoji set", "🌱 🪴 ✨"), ("Caption style", "Hook · value · save CTA")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "LINK IN BIO", "section_gold"); ws.row_dimensions[15].height = 22
    links = [("Shop", "studiofern.co/shop"), ("LTK", "ltk.to/studiofern"), ("Newsletter", "The Repot"),
             ("Care guide", "studiofern.co/guide"), ("Media Kit", "studiofern.co/kit"), ("Email", "hi@studiofern.co")]
    for i, (p, h) in enumerate(links):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 21 — Content Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  CONTENT GALLERY",
                "Your visual archive — drop in covers & stills and caption your best posts.")
    merge_set(ws, "B5:D5", "HOW TO ADD YOUR COVERS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Excel: Insert ▸ Pictures ▸ Place in Cell (or drag an image) into any framed box below. "
                      "Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"paste-link-here\"). "
                      "Caption each one underneath — hook, reach, and what made it work.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Top Reel", "Viral Carousel", "Brand Deal", "Grid Highlight", "Before/After", "Save Magnet"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add cover)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Hook · reach · why it worked…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 22 — Goals & OKRs
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 14, 16, 12, 12, 2])
    luxe_header(ws, "F", "🎯  GOALS & OKRs",
                "Run your account on objectives — measurable key results with live progress.")
    table_headers(ws, 4, ["Objective / Key Result", "Category", "Target", "Current", "Progress"])
    goals = [
        ("Hit 200k followers", "Followers", "200,000", "148,000", 0.74),
        ("$9k/mo revenue", "Revenue", "$9,000", "$7,800", 0.87),
        ("Post 20 times / month", "Posting", "20", "18", 0.90),
        ("Grow 12k followers / mo", "Followers", "12,000", "9,200", 0.77),
        ("Land 5 active brand deals", "Brand Deals", "5", "4", 0.80),
        ("$1.5k/mo Shop + affiliate", "Shop", "$1,500", "$1,275", 0.85),
        ("Lift reel retention to 55%", "Engagement", "55%", "52%", 0.95),
    ]
    start = L0
    for i, (g, cat, tgt, cur, prog) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalCatList")
    nrange(wb, "GoalCategory", "Goals", "B", start, end)
    nrange(wb, "GoalProgress", "Goals", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 23 — Audience Insights
# ===========================================================================
def build_audience(wb):
    ws = wb.create_sheet("Audience"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 4, 24, 16, 2])
    luxe_header(ws, "G", "🌍  AUDIENCE INSIGHTS",
                "Know who's watching — demographics, timing & what your audience loves.")
    merge_set(ws, "B5:C5", "WHO'S FOLLOWING", "section_gold"); ws.row_dimensions[5].height = 22
    demo = [("Women", "82%"), ("Age 25-34", "46%"), ("Age 35-44", "27%"), ("Age 18-24", "17%"),
            ("Top city", "Los Angeles"), ("Top country", "US 58%"), ("Followers reached", "68%"), ("Non-followers reached", "54%")]
    for i, (k, v) in enumerate(demo):
        r = 6 + i
        ws.cell(row=r, column=2, value=k).style = "field_label"
        ws.cell(row=r, column=3, value=v).style = "field_value"
    merge_set(ws, "E5:F5", "WHEN & WHAT", "section_gold"); ws.row_dimensions[5].height = 22
    when = [("Best days", "Sun, Tue, Thu"), ("Best times", "11am & 7pm"), ("Peak activity", "Sun 7pm"),
            ("Top pillar", "Plants"), ("Most saved", "Care carousels"), ("Most shared", "Renter reels"),
            ("Top format", "Reels"), ("Save driver", "'Save this for later'")]
    for i, (k, v) in enumerate(when):
        r = 6 + i
        ws.cell(row=r, column=5, value=k).style = "field_label"
        ws.cell(row=r, column=6, value=v).style = "field_value"


# ===========================================================================
# 1 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  📸  INSTAGRAM COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Followers, reach, revenue & brand deals — your whole Instagram business, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("FOLLOWERS", "=FollowerNow", "num"),
        ("REACH (28D)", "=Reach28", "num"),
        ("ENGAGEMENT RATE", "=EngRate", "pct1"),
        ("SAVES (28D)", "=Saves28", "num"),
        ("POSTED (28D)", '=COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Posted")+COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Boosted")', "num"),
        ("MONTHLY REVENUE", "=RevenueTotal", "money"),
    ]
    row2 = [
        ("NET PROFIT", "=RevenueTotal-ExpenseTotal", "money"),
        ("SHOP SALES (GMV)", "=ShopGMV", "money"),
        ("BRAND DEALS", '=COUNTIF(SponStage,"Signed")+COUNTIF(SponStage,"Delivered")+COUNTIF(SponStage,"Negotiation")', "num"),
        ("FOLLOWER GROWTH", "=FollowerGrowth", "num"),
        ("POSTING CONSISTENCY", '=IFERROR(MIN((COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Posted")+COUNTIFS(CalDate,">="&TODAY()-28,CalStatus,"Boosted"))/PostGoal,1),0)', "pct"),
        ("CREATOR HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "GROWTH & REVENUE", "section_gold")
    ln = LineChart(); ln.title = "Follower Growth (K)"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=26, max_row=32), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=27, max_row=32)); ln.legend = None
    ws.add_chart(ln, "B12")
    d1 = DoughnutChart(); d1.title = "Revenue by Source"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Finance"], min_col=2, min_row=4, max_row=11), titles_from_data=True)
    d1.set_categories(Reference(wb["Finance"], min_col=1, min_row=5, max_row=11)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "CONTENT & PROFIT", "section_gold")
    cb = BarChart(); cb.type = "bar"; cb.title = "Top Posts by Reach"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); cb.legend = None
    ws.add_chart(cb, "B30")
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Finance"], min_col=7, min_row=4, max_row=11), titles_from_data=True)
    eb.set_categories(Reference(wb["Finance"], min_col=6, min_row=5, max_row=11)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Instagram Command Center™ — from grid to growth, all in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_calendar(wb)
    build_grid(wb); build_pipeline(wb); build_ideas(wb); build_captions(wb)
    build_reels(wb); build_stories(wb); build_hashtags(wb); build_analytics(wb)
    build_community(wb); build_shop(wb); build_affiliate(wb); build_deals(wb)
    build_finance(wb); build_expenses(wb); build_equipment(wb); build_repurpose(wb)
    build_brandkit(wb); build_gallery(wb); build_goals(wb); build_audience(wb)
    build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Profile", "Calendar", "Grid Planner", "Pipeline", "Ideas", "Captions",
             "Reels", "Stories", "Hashtags", "Analytics", "Community", "Shop", "Affiliate", "Brand Deals",
             "Finance", "Expenses", "Equipment", "Repurposing", "Brand Kit", "Gallery", "Goals", "Audience",
             "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Instagram_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
