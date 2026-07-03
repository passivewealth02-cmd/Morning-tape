"""Build Tennis Command Center™ — The Ultimate Tennis Performance, Training &
Tournament Management System.

19 sheets + Welcome · a premium tennis operating system in Excel & Sheets.
Matches, analytics, tournaments, practice, skills, fitness, equipment, budget,
partners, coaching, goals, travel, nutrition, season & analytics.

Run: python3 build_xlsx.py   ->  ../Tennis_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

SKILLS = ["Forehand", "Backhand", "Slice", "Volley", "Overhead", "First Serve", "Second Serve",
          "Return of Serve", "Footwork", "Speed", "Court Positioning", "Tactical Awareness", "Mental Toughness"]
SURFACES = ["Hard", "Clay", "Grass", "Carpet"]
FORMATS = ["Singles", "Doubles"]
TOURN_LEVELS = ["Level 1", "Level 2", "Level 3", "Sectional", "National", "ITF Junior", "Club"]
EXPENSE_CATS = ["Coaching", "Court Fees", "Club Membership", "Equipment", "Tournament Fees",
                "Travel", "Hotels", "Food", "Apparel", "Recovery", "Miscellaneous"]
EQUIP_TYPES = ["Racquet", "Strings", "Shoes", "Bag", "Balls", "Overgrips", "Dampener", "Apparel", "Accessory"]
GOAL_CATS = ["Ranking", "Tournament", "Fitness", "Skill", "Mental", "Match", "Practice"]
FOCUS_AREAS = ["Serves", "Forehands", "Backhands", "Volleys", "Overheads", "Footwork", "Match Play", "Mental"]
RESULTS = ["W", "L"]
CONDITIONS = ["New", "Excellent", "Good", "Worn", "Replace Soon"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 16 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "dec": "0.0",
                        "rate": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [17] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your player details once — every dashboard follows. Works for any level.")
    merge_set(ws, "B5:C5", "PLAYER INPUTS", "section")
    controls = [
        ("Player Name", "Maya Chen", None, "PlayerName"),
        ("Age Division", "U16", None, "AgeDiv"),
        ("Current Rating (UTR)", 7.5, "0.0", "PlayerRating"),
        ("Current Ranking", 14, "0", "CurrentRank"),
        ("Home Club", "Cedar Park Tennis Club", None, "HomeClub"),
        ("Monthly Budget", 2000, '"$"#,##0', "MonthlyBudget"),
        ("Season Match Target", 40, "0", "SeasonTarget"),
        ("Practice Hours Target", 40, "0", "HoursTarget"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Surface", SURFACES, "SurfaceList"), ("F", "Format", FORMATS, "FormatList"),
             ("G", "Tournament Level", TOURN_LEVELS, "TournLevelList"), ("H", "Skill", SKILLS, "SkillList"),
             ("I", "Expense Category", EXPENSE_CATS, "ExpenseCatList"), ("J", "Equipment Type", EQUIP_TYPES, "EquipTypeList"),
             ("K", "Goal Category", GOAL_CATS, "GoalCatList"), ("L", "Practice Focus", FOCUS_AREAS, "FocusList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 21, "Result", RESULTS, "ResultList"),
             ("F", 21, "Condition", CONDITIONS, "CondList"),
             ("G", 21, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 82, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🎾  TENNIS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate tennis performance, training & tournament management system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "TRAIN LIKE A PRO — ORGANIZE LIKE AN ACADEMY", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Most tennis planners only track scores. Tennis Command Center™ brings match tracking, match "
                      "analytics, tournaments, practice, skills, fitness, equipment, finances, coaching and long-term "
                      "development into ONE premium Excel & Google Sheets dashboard — with academy-level reporting "
                      "and automation. Whether you're a junior chasing a ranking, a league player, or a coach "
                      "managing athletes, this is your tennis operating system.")
    ws["B6"].style = "body"
    for r in (6, 7, 8, 9):
        ws.row_dimensions[r].height = 22
    merge_set(ws, "B11:B11", "START HERE", "section")
    steps = ["1.  Open Settings and add your name, rating, ranking & season target.",
             "2.  Fill in your Player Profile (division, backhand style, coach, goals).",
             "3.  Log matches — win %, sets, games & surface stats calculate themselves.",
             "4.  Track serve & rally stats in Match Analytics; rate your skills.",
             "5.  Plan practice & tournaments, set budgets, log fitness — all rolls up.",
             "6.  Watch the Executive Dashboard track wins, rating & your performance score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (a competitive U16 junior season) is included so you can see how everything "
               "connects — just type over it with your own. Win %, sets & games won, surface records, serve stats, "
               "skill progress, budget and the Player Performance Score all update automatically. Every sheet is "
               "print-friendly and works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, better tennis — game, set, match!", "section_gold")


# ===========================================================================
# 2 — Player Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Player Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 28, 6, 24, 24, 2])
    luxe_header(ws, "G", "🎾  PLAYER PROFILE", "Who you are on court — the essentials, all in one place.")
    blocks = [
        ("THE PLAYER", [("Player Name", "=PlayerName"), ("Age Division", "=AgeDiv"),
                        ("Current Rating (UTR)", "=PlayerRating"), ("Current Ranking", "=CurrentRank"),
                        ("Dominant Hand", "Right"), ("Backhand Style", "Two-handed")]),
        ("TEAM & SUPPORT", [("Home Club", "=HomeClub"), ("Head Coach", "Coach Petrov"),
                            ("Fitness Coach", "Coach Reyes"), ("Playing Since", "2016"),
                            ("Emergency Contact", "Mom — (555) 210-4521"), ("Medical Notes", "Mild asthma — carry inhaler")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            cv = ws.cell(row=row, column=3, value=fields[i][1]); cv.style = "field_value"
            if fields[i][0] == "Current Rating (UTR)":
                cv.number_format = "0.0"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "LONG-TERM GOALS", "section_gold"); ws.row_dimensions[15].height = 22
    goals = ["Reach a UTR of 9.0", "Win a sectional Level 2 tournament",
             "Crack the top 10 in the section", "Earn a college team spot by senior year"]
    for i, g in enumerate(goals):
        r = 16 + i
        cb = ws.cell(row=r, column=2, value="◆"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=11, color=ACCENT); cb.border = BOX
        merge_set(ws, f"C{r}:F{r}", g, "td_left")


# ===========================================================================
# 3 — Match Tracker
# ===========================================================================
def build_matches(wb):
    # (dago, opponent, event, surface, format, [ (myGames, oppGames) per set ])
    M = [
        (78, "A. Novak", "Cedar Open", "Hard", "Singles", [(6, 2), (6, 3)]),
        (75, "R. Silva", "Cedar Open", "Clay", "Singles", [(6, 4), (3, 6), (6, 3)]),
        (71, "K. Adeyemi", "Spring Sectional", "Hard", "Singles", [(6, 4), (6, 4)]),
        (70, "Team Foster", "Spring Sectional", "Hard", "Doubles", [(6, 3), (6, 4)]),
        (66, "L. Marchetti", "Spring Sectional", "Clay", "Singles", [(7, 5), (4, 6), (6, 3)]),
        (63, "D. Okafor", "League Wk 3", "Hard", "Singles", [(6, 1), (6, 2)]),
        (60, "P. Andersson", "Grass Invitational", "Grass", "Singles", [(6, 4), (7, 5)]),
        (56, "J. Rivera", "Grass Invitational", "Hard", "Singles", [(7, 6), (6, 4)]),
        (53, "Team Lin", "Club Doubles", "Clay", "Doubles", [(6, 3), (6, 2)]),
        (49, "S. Kowalski", "Summer Level 2", "Hard", "Singles", [(6, 2), (6, 4)]),
        (48, "M. Bianchi", "Summer Level 2", "Clay", "Singles", [(7, 5), (4, 6), (6, 3)]),
        (44, "T. Nakamura", "Summer Level 2", "Hard", "Singles", [(6, 3), (7, 5)]),
        (40, "H. Costa", "Grass Cup", "Grass", "Singles", [(6, 2), (6, 1)]),
        (37, "Team Reyes", "Club Doubles", "Hard", "Doubles", [(6, 4), (4, 6), (6, 4)]),
        (33, "E. Volkov", "Fall Sectional", "Clay", "Singles", [(6, 3), (6, 4)]),
        (30, "N. Haddad", "Fall Sectional", "Hard", "Singles", [(6, 2), (6, 3)]),
        (28, "C. Dubois", "Fall Sectional", "Hard", "Singles", [(7, 5), (6, 4)]),
        (24, "V. Petrova", "League Wk 8", "Clay", "Singles", [(2, 6), (4, 6)]),
        (21, "F. Moreau", "State Level 1", "Hard", "Singles", [(6, 3), (4, 6), (5, 7)]),
        (18, "G. Schmidt", "State Level 1", "Grass", "Singles", [(3, 6), (4, 6)]),
        (15, "O. Ivanov", "State Level 1", "Hard", "Singles", [(6, 7), (6, 7)]),
        (12, "B. Yamada", "Autumn Clay", "Clay", "Singles", [(1, 6), (4, 6)]),
        (9, "Team Ali", "Club Doubles", "Hard", "Doubles", [(6, 4), (3, 6), (4, 6)]),
        (6, "W. Fischer", "League Wk 10", "Hard", "Singles", [(6, 7), (4, 6)]),
        (3, "Z. Popov", "Winter Open", "Clay", "Singles", [(2, 6), (3, 6)]),
        (1, "Q. Tanaka", "Winter Open", "Hard", "Singles", [(7, 5), (3, 6), (5, 7)]),
    ]
    ws = wb.create_sheet("Matches"); ws.sheet_view.showGridLines = False
    set_widths(ws, [12, 16, 20, 10, 11, 16, 7, 7, 8, 8, 9, 16])
    luxe_header(ws, "L", "🎾  MATCH TRACKER",
                "Every match logged — win %, sets, games & surface records all calculate automatically.")
    table_headers(ws, 4, ["Date", "Opponent", "Event", "Surface", "Format", "Score",
                          "Sets W", "Sets L", "Games W", "Games L", "Result", "Notes"])
    notes = ["Clean start", "Battled back", "Solid", "Great chemistry", "3-set grind", "Rolled",
             "Grass debut", "Tight tiebreak", "", "Confident", "Comeback", "Serve on point",
             "Fast start", "Doubles W", "", "Clutch", "3-setter", "Tough draw", "Close loss",
             "Grass slip", "Two tiebreaks", "Off day", "", "Cramped late", "Slow start", "Learning match"]
    start = L0
    for i, (dago, opp, ev, surf, fmt, sets) in enumerate(M):
        r = start + i
        sw = sum(1 for a, b in sets if a > b); sl = sum(1 for a, b in sets if a < b)
        gw = sum(a for a, b in sets); gl = sum(b for a, b in sets)
        score = ", ".join(f"{a}-{b}" for a, b in sets)
        ws.cell(row=r, column=1, value=dminus(dago))
        ws.cell(row=r, column=2, value=opp)
        ws.cell(row=r, column=3, value=ev)
        ws.cell(row=r, column=4, value=surf)
        ws.cell(row=r, column=5, value=fmt)
        ws.cell(row=r, column=6, value=score)
        ws.cell(row=r, column=7, value=sw)
        ws.cell(row=r, column=8, value=sl)
        ws.cell(row=r, column=9, value=gw)
        ws.cell(row=r, column=10, value=gl)
        ws.cell(row=r, column=11, value=f'=IF(G{r}="","",IF(G{r}>H{r},"W","L"))')
        ws.cell(row=r, column=12, value=notes[i])
    end = start + 60 - 1
    style_rows(ws, start, end, 12, text_left={2, 3, 12}, dates={1}, ints={7, 8, 9, 10})
    for col_letter, lst in [("D", "SurfaceList"), ("E", "FormatList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTALS").style = "th"
    for col in (7, 8, 9, 10):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "#,##0"
    ws.freeze_panes = "A5"
    nrange(wb, "MatchDate", "Matches", "A", start, end)
    nrange(wb, "MatchSurface", "Matches", "D", start, end)
    nrange(wb, "MatchSetsW", "Matches", "G", start, end)
    nrange(wb, "MatchSetsL", "Matches", "H", start, end)
    nrange(wb, "MatchGamesW", "Matches", "I", start, end)
    nrange(wb, "MatchResult", "Matches", "K", start, end)
    ws.conditional_formatting.add(f"K{start}:K{end}",
        CellIsRule(operator="equal", formula=['"W"'], fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    ws.conditional_formatting.add(f"K{start}:K{end}",
        CellIsRule(operator="equal", formula=['"L"'], fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    # results + surface summaries (dashboard donut)
    ws.cell(row=4, column=14, value="Result").style = "th"
    ws.cell(row=4, column=15, value="Count").style = "th"
    ws.cell(row=5, column=14, value="Wins").style = "td_left"
    ws.cell(row=5, column=15, value='=COUNTIF(MatchResult,"W")').style = "td"
    ws.cell(row=6, column=14, value="Losses").style = "td_left"
    ws.cell(row=6, column=15, value='=COUNTIF(MatchResult,"L")').style = "td"
    ws.cell(row=8, column=14, value="Surface").style = "th"
    ws.cell(row=8, column=15, value="Wins").style = "th"
    for i, s in enumerate(["Hard", "Clay", "Grass"]):
        r = 9 + i
        ws.cell(row=r, column=14, value=s).style = "td_left"
        ws.cell(row=r, column=15, value=f'=COUNTIFS(MatchSurface,N{r},MatchResult,"W")').style = "td"
    set_widths(ws, [12, 16, 20, 10, 11, 13, 7, 7, 8, 8, 9, 16, 3, 12, 10])
    cell_name(wb, "ResultLabel", "Matches", "$N$5:$N$6")
    cell_name(wb, "ResultCount", "Matches", "$O$5:$O$6")


# ===========================================================================
# 4 — Match Analytics
# ===========================================================================
def build_analytics_match(wb):
    # (dago, opponent, firstServe, aces, dfs, winners, ues, bpWon, bpSaved, rating)
    rows = [
        (49, "S. Kowalski", 0.66, 6, 2, 24, 15, 5, 7, 8.5),
        (44, "T. Nakamura", 0.62, 4, 3, 21, 18, 4, 5, 7.5),
        (40, "H. Costa", 0.70, 7, 1, 26, 12, 6, 8, 9.0),
        (33, "E. Volkov", 0.64, 5, 2, 22, 16, 5, 6, 8.0),
        (30, "N. Haddad", 0.68, 6, 2, 25, 14, 5, 7, 8.5),
        (28, "C. Dubois", 0.60, 3, 4, 20, 19, 4, 5, 7.0),
        (21, "F. Moreau", 0.58, 4, 5, 19, 24, 3, 4, 6.5),
        (18, "G. Schmidt", 0.55, 2, 6, 16, 26, 2, 3, 6.0),
        (12, "B. Yamada", 0.57, 3, 4, 17, 23, 3, 3, 6.5),
        (6, "W. Fischer", 0.61, 4, 3, 20, 20, 4, 5, 7.0),
    ]
    ws = wb.create_sheet("Match Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [12, 16, 13, 8, 9, 11, 12, 11, 12, 11])
    luxe_header(ws, "J", "📊  MATCH ANALYTICS",
                "The story behind the score — serve, rally & break-point stats with match ratings.")
    table_headers(ws, 4, ["Date", "Opponent", "1st Serve %", "Aces", "DFs", "Winners",
                          "Unf. Errors", "BP Won", "BP Saved", "Rating"])
    start = L0
    for i, (dago, opp, fs, aces, dfs, win, ue, bpw, bps, rating) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=dminus(dago))
        ws.cell(row=r, column=2, value=opp)
        ws.cell(row=r, column=3, value=fs)
        ws.cell(row=r, column=4, value=aces)
        ws.cell(row=r, column=5, value=dfs)
        ws.cell(row=r, column=6, value=win)
        ws.cell(row=r, column=7, value=ue)
        ws.cell(row=r, column=8, value=bpw)
        ws.cell(row=r, column=9, value=bps)
        ws.cell(row=r, column=10, value=rating)
    end = start + 40 - 1
    style_rows(ws, start, end, 10, text_left={2}, dates={1}, pcts={3}, ints={4, 5, 6, 7, 8, 9}, dec={10})
    avg = end + 1
    ws.cell(row=avg, column=1, value="AVERAGE").style = "th"
    ws.cell(row=avg, column=2, value="").style = "th"
    for col, fmt in [(3, "0%"), (4, "0.0"), (5, "0.0"), (6, "0.0"), (7, "0.0"), (8, "0.0"), (9, "0.0"), (10, "0.0")]:
        L = get_column_letter(col)
        c = ws.cell(row=avg, column=col, value=f"=IFERROR(AVERAGE({L}{start}:{L}{end}),0)")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt
    ws.freeze_panes = "A5"
    nrange(wb, "AnaDate", "Match Analytics", "A", start, end)
    nrange(wb, "Ana1stServe", "Match Analytics", "C", start, end)
    nrange(wb, "AnaAces", "Match Analytics", "D", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=0.5, start_color="FF" + WARN_BG,
                       end_type="num", end_value=0.72, end_color="FF" + HIGHLIGHT))
    ws.conditional_formatting.add(f"J{start}:J{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=10, color=PRIMARY, showValue=True))
    # serve stats chart (aces vs DFs)
    bar = BarChart(); bar.type = "col"; bar.title = "Serve Stats — Aces vs Double Faults"; bar.height = 9; bar.width = 15
    bar.add_data(Reference(ws, min_col=4, min_row=4, max_row=end), titles_from_data=True)
    bar.add_data(Reference(ws, min_col=5, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    ws.add_chart(bar, "L5")


# ===========================================================================
# 5 — Tournament Command Center
# ===========================================================================
def build_tournaments(wb):
    ws = wb.create_sheet("Tournaments"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 16, 13, 10, 11, 10, 13, 11, 12, 16])
    luxe_header(ws, "J", "🏆  TOURNAMENT COMMAND CENTER",
                "Every event tracked — entry, travel, placement, prize & ranking points.")
    table_headers(ws, 4, ["Tournament", "Location", "Date", "Level", "Entry Fee", "Travel",
                          "Placement", "Prize", "Points", "Notes"])
    # (name, loc, doff, level, fee, travel, placement, prize, points, note)
    rows = [
        ("Summer Level 2", "Riverside", -46, "Level 2", 90, 180, "1st", 0, 250, "Gold — great run!"),
        ("Spring Sectional", "Metro Center", -70, "Sectional", 75, 120, "3rd", 0, 150, "Bronze"),
        ("Grass Cup", "Lakeside", -40, "Level 3", 60, 90, "1st", 0, 180, "Undefeated"),
        ("Fall Sectional", "State Center", -30, "Sectional", 85, 200, "5th", 0, 90, "Tough draw"),
        ("State Level 1", "Capital Courts", -18, "Level 1", 120, 320, "Quarterfinal", 0, 120, "Lost to #3 seed"),
        ("Winter Open", "Cedar Park", -1, "Level 2", 90, 60, "Round 2", 0, 40, "Building form"),
        ("Spring Championships", "Metro Center", 24, "Level 1", 130, 220, "Registered", 0, 0, "Season goal"),
        ("Clay Court Classic", "Riverside", 45, "Level 2", 95, 140, "Registered", 0, 0, "Best surface"),
        ("Sectional Masters", "State Center", 68, "Sectional", 110, 260, "Registered", 0, 0, "Big points"),
    ]
    start = L0
    for i, (nm, loc, doff, lvl, fee, travel, place, prize, pts, note) in enumerate(rows):
        r = start + i
        d = dplus(doff) if doff >= 0 else dminus(-doff)
        for ci, v in enumerate([nm, loc, d, lvl, fee, travel, place, prize, pts, note], 1):
            ws.cell(row=r, column=ci, value=v)
    end = start + 30 - 1
    style_rows(ws, start, end, 10, text_left={1, 2, 10}, dates={3}, money={5, 6, 8}, ints={9})
    add_dv(ws, f"D{start}:D{end}", "TournLevelList")
    total = end + 1
    ws.cell(row=total, column=1, value="TOTALS").style = "th"
    for col in (5, 6, 8):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    c = ws.cell(row=total, column=9, value=f"=SUM(I{start}:I{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "#,##0"
    ws.freeze_panes = "A5"
    nrange(wb, "TournName", "Tournaments", "A", start, end)
    nrange(wb, "TournDate", "Tournaments", "C", start, end)
    nrange(wb, "TournPlace", "Tournaments", "G", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"1st"'], fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"Registered"'], fill=fill(WARN_BG)))


# ===========================================================================
# 6 — Practice Planner
# ===========================================================================
def build_practice(wb):
    rows = [
        (2, 2.0, "Coach Petrov", "Serves", "Kick serve, targets, 100 serves", "High", "Second serve improving"),
        (3, 1.5, "Coach Petrov", "Backhands", "Down-the-line, cross-court", "Medium", "Cleaner contact"),
        (5, 2.5, "Solo", "Match Play", "Practice sets vs hitting partner", "High", "Closed out 6-3"),
        (7, 2.0, "Coach Reyes", "Footwork", "Ladder, split-step, recovery", "Medium", "Faster to wide balls"),
        (9, 1.5, "Coach Petrov", "Volleys", "Approach + volley, overheads", "Medium", "Net game sharper"),
        (12, 2.5, "Coach Petrov", "Serves", "First serve %, placement", "High", "72% first serves"),
        (15, 2.0, "Solo", "Forehands", "Inside-out, heavy topspin", "Medium", "More spin"),
        (18, 2.0, "Coach Petrov", "Match Play", "Tiebreak scenarios", "High", "Clutch practice"),
        (22, 1.5, "Coach Reyes", "Mental", "Routines, breathing, focus", "Low", "Calmer between points"),
        (26, 2.5, "Coach Petrov", "Backhands", "Slice + drive mix", "High", "Variety working"),
        (30, 2.0, "Solo", "Overheads", "Smash under pressure", "Medium", "Timing better"),
        (34, 2.0, "Coach Petrov", "Match Play", "Full practice match", "High", "Won 6-4, 6-2"),
    ]
    sample = [(dminus(d), dur, coach, foc, dr, inten, note) for (d, dur, coach, foc, dr, inten, note) in rows]
    ws, start, end = build_log(
        wb, "Practice", "🎯", "PRACTICE PLANNER",
        "Train with intent — focus, drills, intensity & coach notes, session by session.",
        ["Date", "Hours", "Coach", "Focus", "Drills", "Intensity", "Notes"],
        sample, [12, 9, 15, 14, 30, 12, 24],
        text_left={3, 5, 7}, dates={1}, dec={2},
        validations=[("D", "FocusList")], reserved=40)
    nrange(wb, "PracticeDate", "Practice", "A", start, end)
    nrange(wb, "PracticeHrs", "Practice", "B", start, end)


# ===========================================================================
# 7 — Skill Development Center
# ===========================================================================
def build_skills(wb):
    ws = wb.create_sheet("Skills"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 12, 12, 12, 12, 24])
    luxe_header(ws, "F", "📈  SKILL DEVELOPMENT CENTER",
                "Rate every stroke & quality — watch each climb from start to now.")
    table_headers(ws, 4, ["Skill", "Start", "Current", "Target", "Gain", "Focus Notes"])
    data = [
        ("Forehand", 7, 9, 10, "Weapon — inside-out"),
        ("Backhand", 6, 8, 9, "Down-the-line better"),
        ("Slice", 5, 6, 8, "Use more on approach"),
        ("Volley", 6, 7, 9, "Punch volleys solid"),
        ("Overhead", 6, 7, 8, "Under pressure next"),
        ("First Serve", 6, 8, 9, "Placement + kick"),
        ("Second Serve", 5, 6, 8, "Priority — add spin"),
        ("Return of Serve", 6, 7, 9, "Step in earlier"),
        ("Footwork", 7, 8, 9, "Split-step timing"),
        ("Speed", 7, 8, 9, "Recovery quicker"),
        ("Court Positioning", 6, 7, 9, "Take the middle"),
        ("Tactical Awareness", 5, 7, 9, "Pattern play"),
        ("Mental Toughness", 5, 6, 8, "Reset routines"),
    ]
    start = L0
    for i, (sk, st, cur, tgt, note) in enumerate(data):
        r = start + i
        ws.cell(row=r, column=1, value=sk).style = "td_left"
        for ci, v in [(2, st), (3, cur), (4, tgt)]:
            c = ws.cell(row=r, column=ci, value=v); c.style = "input" if ci == 3 else "td"; c.number_format = "0"
        ws.cell(row=r, column=5, value=f"=C{r}-B{r}").style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for cc in range(1, 7):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    end = start + len(data) - 1
    avg = end + 1
    ws.cell(row=avg, column=1, value="AVERAGE").style = "th"
    for col in (2, 3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=avg, column=col, value=f"=AVERAGE({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0.0"
    ws.freeze_panes = "A5"
    nrange(wb, "SkillName", "Skills", "A", start, end)
    nrange(wb, "SkillStart", "Skills", "B", start, end)
    nrange(wb, "SkillCurrent", "Skills", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=3, start_color="FF" + WARN_BG,
                       end_type="num", end_value=10, end_color="FF" + HIGHLIGHT))
    bar = BarChart(); bar.type = "bar"; bar.title = "Skill Progress — Start vs Now"; bar.height = 11; bar.width = 15
    bar.add_data(Reference(ws, min_col=2, min_row=4, max_row=end), titles_from_data=True)
    bar.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    ws.add_chart(bar, "H5")


# ===========================================================================
# 8 — Fitness & Conditioning
# ===========================================================================
def build_fitness(wb):
    rows = [
        (2, "Strength — legs & core", "Strength", 60, 8.0, "Squats, lunges, planks"),
        (3, "Court sprints", "Agility", 40, 7.5, "Suicides, spider drill"),
        (4, "Mobility & stretch", "Flexibility", 25, 8.0, "Hips, shoulders"),
        (6, "Recovery day", "Recovery", 0, 8.5, "Foam roll, ice bath"),
        (8, "Strength — upper", "Strength", 50, 7.5, "Rows, presses, bands"),
        (10, "Interval cardio", "Cardio", 35, 7.0, "Bike intervals"),
        (12, "Agility ladder", "Agility", 30, 8.0, "Footwork patterns"),
        (15, "Full-body strength", "Strength", 55, 7.5, "Compound lifts"),
        (18, "Yoga / balance", "Flexibility", 40, 8.0, "Single-leg balance"),
        (22, "Recovery + massage", "Recovery", 0, 8.5, "Deep tissue"),
    ]
    sample = [(dminus(d), act, ty, mins, sleep, note) for (d, act, ty, mins, sleep, note) in rows]
    ws, start, end = build_log(
        wb, "Fitness", "💪", "FITNESS & CONDITIONING",
        "Off-court work wins matches — strength, cardio, agility & recovery, all logged.",
        ["Date", "Activity", "Type", "Minutes", "Sleep (hrs)", "Notes"],
        sample, [12, 26, 14, 11, 12, 26],
        text_left={2, 6}, dates={1}, ints={4}, dec={5}, reserved=40)
    nrange(wb, "FitDate", "Fitness", "A", start, end)
    nrange(wb, "FitType", "Fitness", "C", start, end)


# ===========================================================================
# 9 — Equipment Command Center
# ===========================================================================
def build_equipment(wb):
    ws = wb.create_sheet("Equipment"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 13, 18, 12, 11, 14, 13, 13, 16])
    luxe_header(ws, "I", "🎽  EQUIPMENT COMMAND CENTER",
                "Gear you can trust — condition tracking with automatic replacement reminders.")
    table_headers(ws, 4, ["Item", "Type", "Detail", "Purchased", "Cost", "Condition", "Replace By", "Status", "Notes"])
    # (item, type, detail, purch_dago, cost, condition, replace_doff, note)
    rows = [
        ("Match Racquet #1", "Racquet", "Wilson Blade 98", 210, 250, "Good", 120, "Main stick"),
        ("Match Racquet #2", "Racquet", "Wilson Blade 98", 210, 250, "Good", 130, "Identical backup"),
        ("String Job", "Strings", "Luxilon 4G @ 52 lbs", 12, 22, "Worn", 20, "Restring soon"),
        ("Court Shoes", "Shoes", "Nike Vapor", 150, 130, "Replace Soon", 35, "Tread low"),
        ("Practice Balls (case)", "Balls", "Penn Championship", 40, 40, "Good", 90, "Rotate"),
        ("Racquet Bag", "Bag", "Babolat 12-pack", 300, 120, "Excellent", 700, "Holds all gear"),
        ("Overgrips (x12)", "Overgrips", "Tourna Grip", 20, 18, "Good", 80, "Fresh stock"),
        ("Dampeners (x3)", "Dampener", "Wilson", 60, 10, "Good", 200, "Spares"),
        ("Match Kit (x3)", "Apparel", "Nike Dri-FIT", 120, 90, "Good", 300, "Tournament kit"),
    ]
    start = L0
    for i, (item, ty, det, pdago, cost, cond, rdoff, note) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=ty)
        ws.cell(row=r, column=3, value=det)
        ws.cell(row=r, column=4, value=dminus(pdago))
        ws.cell(row=r, column=5, value=cost)
        ws.cell(row=r, column=6, value=cond)
        ws.cell(row=r, column=7, value=dplus(rdoff))
        ws.cell(row=r, column=8, value=f'=IF(G{r}="","",IF(G{r}<=TODAY()+45,"DUE SOON","OK"))')
        ws.cell(row=r, column=9, value=note)
    end = start + 30 - 1
    style_rows(ws, start, end, 9, text_left={1, 3, 9}, dates={4, 7}, money={5})
    for col_letter, lst in [("B", "EquipTypeList"), ("F", "CondList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL INVESTED").style = "th"
    c = ws.cell(row=total, column=5, value=f"=SUM(E{start}:E{end})")
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "EquipName", "Equipment", "A", start, end)
    nrange(wb, "EquipReplace", "Equipment", "G", start, end)
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"DUE SOON"'], fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"OK"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Replace Soon"'], fill=fill(RED_BG)))


# ===========================================================================
# 10 — Tennis Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 14, 14, 12, 3, 20, 14])
    luxe_header(ws, "H", "💰  TENNIS BUDGET",
                "Know what the game costs — plan vs actual across every category, monthly.")
    table_headers(ws, 4, ["Category", "Monthly Budget", "Spent", "Remaining", "% Used"])
    # (cat, planned, actual)
    data = [
        ("Coaching", 700, 640), ("Court Fees", 140, 120), ("Club Membership", 95, 95),
        ("Equipment", 180, 180), ("Tournament Fees", 250, 210), ("Travel", 180, 160),
        ("Hotels", 130, 120), ("Food", 100, 90), ("Apparel", 60, 60),
        ("Recovery", 80, 75), ("Miscellaneous", 60, 50),
    ]
    start = L0; end = start + len(data) - 1
    for i, (cat, pl, ac) in enumerate(data):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=2, value=pl); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=ac); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in range(2, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})"); c.style = "td"
        c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=total, column=5, value=f"=IFERROR(C{total}/B{total},0)"); cu.style = "td"
    cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.1, color=PRIMARY, showValue=True))
    nrange(wb, "BudgetCat", "Budget", "A", start, end)
    nrange(wb, "BudgetActual", "Budget", "C", start, end)
    cell_name(wb, "BudgetTotalPlanned", "Budget", f"$B${total}")
    cell_name(wb, "BudgetTotalActual", "Budget", f"$C${total}")
    merge_set(ws, "G5:H5", "AT A GLANCE", "section_gold")
    rows2 = [("Monthly budget", f"=B{total}", '"$"#,##0'), ("Spent so far", f"=C{total}", '"$"#,##0'),
             ("Remaining", f"=B{total}-C{total}", '"$"#,##0'), ("Annual (est.)", f"=C{total}*12", '"$"#,##0'),
             ("Prize money won", "=SUM(Tournaments!H5:H34)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 6 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml); c.style = "field_value"; c.number_format = fmt


# ===========================================================================
# 11 — Doubles Partner Tracker
# ===========================================================================
def build_partners(wb):
    ws = wb.create_sheet("Partners"); ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 12, 9, 9, 10, 10, 26, 24])
    luxe_header(ws, "H", "🤝  DOUBLES PARTNER TRACKER",
                "Find your best team — records, chemistry notes & win % for every partner.")
    table_headers(ws, 4, ["Partner", "Matches", "Wins", "Losses", "Win %", "Record", "Team Strengths", "Areas to Improve"])
    data = [
        ("Jamie Foster", 5, 4, 1, "Poaches well, big serve", "Communication mid-game"),
        ("Riley Lin", 4, 3, 1, "Steady baseline, smart", "Net aggression"),
        ("Casey Reyes", 3, 1, 2, "Fast, great reflexes", "Serve consistency"),
        ("Alex Ali", 2, 1, 1, "Lefty advantage", "Return positioning"),
    ]
    start = L0
    for i, (p, matches, w, l, strong, improve) in enumerate(data):
        r = start + i
        ws.cell(row=r, column=1, value=p).style = "td_left"
        ws.cell(row=r, column=2, value=matches).style = "td"
        cw = ws.cell(row=r, column=3, value=w); cw.style = "td"; cw.number_format = "0"
        cl = ws.cell(row=r, column=4, value=l); cl.style = "td"; cl.number_format = "0"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)"); cp.style = "td"; cp.number_format = "0%"
        cr = ws.cell(row=r, column=6, value=f'=C{r}&"-"&D{r}'); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY)
        ws.cell(row=r, column=7, value=strong).style = "td_left"
        ws.cell(row=r, column=8, value=improve).style = "td_left"
        for cc in range(1, 9):
            ws.cell(row=r, column=cc).border = BOX
        if i % 2:
            for cc in range(1, 9):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    end = start + len(data) - 1
    ws.freeze_panes = "A5"
    nrange(wb, "PartnerName", "Partners", "A", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))


# ===========================================================================
# 12 — Coaching Notes
# ===========================================================================
def build_coaching(wb):
    rows = [
        (34, "Coach Petrov", "Serve toss drifting left", "Fix toss placement", "50 serves daily, film it", "Toss steadier"),
        (26, "Coach Petrov", "Backhand too flat under pressure", "Add margin + spin", "Cross-court rally, 4 ft over net", "More consistent"),
        (18, "Coach Reyes", "Slow first-step to wide balls", "Split-step timing", "Ladder + reaction drills", "Quicker starts"),
        (12, "Coach Petrov", "Rushing points when ahead", "Reset routine at 30-0", "Breathing between points", "Calmer closing"),
        (6, "Coach Petrov", "Second serve too passive", "Commit to kick serve", "Target cones, 60% pace", "Fewer DFs"),
        (2, "Coach Petrov", "Great match temperament vs Fischer", "Keep pre-serve routine", "Same routine every match", "Locked in"),
    ]
    sample = [(dminus(d), coach, fb, act, hw, rev) for (d, coach, fb, act, hw, rev) in rows]
    ws, start, end = build_log(
        wb, "Coaching", "📋", "COACHING NOTES",
        "Turn feedback into progress — key notes, action items & homework, lesson by lesson.",
        ["Date", "Coach", "Key Feedback", "Action Item", "Homework Drills", "Progress Review"],
        sample, [12, 15, 28, 22, 26, 20],
        text_left={2, 3, 4, 5, 6}, dates={1}, reserved=30)


# ===========================================================================
# 13 — Goal Command Center
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 14, 16, 12, 12, 2])
    luxe_header(ws, "F", "🎯  GOAL COMMAND CENTER",
                "Turn ambition into a plan — quarterly & yearly goals with live progress bars.")
    table_headers(ws, 4, ["Goal", "Category", "Target", "Current", "Progress"])
    goals = [
        ("Reach UTR 8.5 this year", "Ranking", "8.5", "7.5", 0.72),
        ("Win a Level 2 tournament", "Tournament", "Title", "1st @ Summer L2", 0.60),
        ("3× strength sessions / week", "Fitness", "12 / mo", "9 / mo", 0.78),
        ("Reliable kick second serve", "Skill", "8 / 10", "6 / 10", 0.70),
        ("Reset routine every match", "Mental", "100%", "65%", 0.65),
        ("Break 65% win rate", "Match", "65%", "65%", 0.82),
        ("40 practice hours / month", "Practice", "40", "28", 0.71),
    ]
    start = L0
    for i, (g, cat, tgt, cur, prog) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalCatList")
    nrange(wb, "GoalCategory", "Goals", "B", start, end)
    nrange(wb, "GoalProgress", "Goals", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    merge_set(ws, "B14:E14", "SEASON MILESTONES", "section_gold"); ws.row_dimensions[14].height = 22
    ms = ["First Level 2 title ✓ (Summer)", "Broke into top 15 ✓",
          "Reach UTR 8.5", "Win a Level 1 match", "Injury-free full season"]
    for i, m in enumerate(ms):
        r = 15 + i
        done = "✓" in m
        cb = ws.cell(row=r, column=2, value="☑" if done else "☐")
        cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=12, color=PRIMARY if done else ACCENT); cb.border = BOX
        merge_set(ws, f"C{r}:E{r}", m.replace(" ✓", ""), "td_left")
        if done:
            ws.cell(row=r, column=2).fill = fill(MINT_BG)


# ===========================================================================
# 14 — Travel Planner
# ===========================================================================
def build_travel(wb):
    rows = [
        ("Spring Championships", 24, "Metro Center", "Grand Hotel (2 nts)", "Drive · 140 mi", 340, "Book early rate"),
        ("Clay Court Classic", 45, "Riverside", "Home", "Drive · 30 mi", 25, "Day trips"),
        ("Sectional Masters", 68, "State Center", "Marriott (3 nts)", "Flight + rental", 620, "Big event"),
        ("Winter Open", -1, "Cedar Park", "Home", "Drive · 10 mi", 12, "Home tournament"),
    ]
    sample = [(nm, dplus(doff) if doff >= 0 else dminus(-doff), loc, hotel, transp, cost, note) for (nm, doff, loc, hotel, transp, cost, note) in rows]
    ws, start, end = build_log(
        wb, "Travel", "✈", "TRAVEL PLANNER",
        "Away tournaments made easy — hotels, transport & costs in one place.",
        ["Event", "Date", "Venue", "Lodging", "Transport", "Est. Cost", "Notes"],
        sample, [24, 13, 18, 20, 18, 12, 22],
        text_left={1, 3, 4, 5, 7}, dates={2}, money={6}, reserved=20)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    c = ws.cell(row=total, column=6, value=f"=SUM(F{start}:F{end})")
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'


# ===========================================================================
# 15 — Nutrition & Recovery
# ===========================================================================
def build_nutrition(wb):
    ws = wb.create_sheet("Nutrition"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 30, 4, 22, 26, 2])
    luxe_header(ws, "G", "🥗  NUTRITION & RECOVERY",
                "Fuel & recover like a pro — match-day nutrition, hydration & recovery.")
    blocks = [
        ("MATCH-DAY FUEL", 5, [("3 hrs before", "Oatmeal, banana, eggs, honey"),
                               ("60 min before", "Banana + electrolyte drink"),
                               ("Changeovers", "Banana, dates, sports drink"),
                               ("Post-match", "Protein shake + fruit + carbs")]),
        ("HYDRATION", 11, [("Daily target", "3–4 L water"), ("On court", "500 ml + electrolytes / hr"),
                           ("Hot days", "Add sodium tabs"), ("Recovery", "16 oz per lb lost")]),
    ]
    for title, top, items in blocks:
        merge_set(ws, f"B{top}:C{top}", title, "section_gold"); ws.row_dimensions[top].height = 22
        for i, (lab, val) in enumerate(items):
            r = top + 1 + i
            ws.cell(row=r, column=2, value=lab).style = "field_label"
            ws.cell(row=r, column=3, value=val).style = "field_value"
            ws.row_dimensions[r].height = 22
    blocks2 = [
        ("SUPPLEMENTS", 5, [("Morning", "Multivitamin, vitamin D"), ("Pre-match", "Electrolytes, caffeine"),
                            ("Post-match", "Whey protein, magnesium"), ("Joint care", "Fish oil, collagen")]),
        ("RECOVERY", 11, [("Sleep target", "8–9 hrs / night"), ("After matches", "Ice bath / cold shower"),
                          ("Weekly", "Massage or foam roll x2"), ("Rest days", "1 full day / week")]),
    ]
    for title, top, items in blocks2:
        merge_set(ws, f"E{top}:F{top}", title, "section_gold"); ws.row_dimensions[top].height = 22
        for i, (lab, val) in enumerate(items):
            r = top + 1 + i
            ws.cell(row=r, column=5, value=lab).style = "field_label"
            ws.cell(row=r, column=6, value=val).style = "field_value"


# ===========================================================================
# 16 — Photo & Video Library
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Media"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "📸  PHOTO & VIDEO LIBRARY",
                "Your season in pictures — paste photos & link technique videos for review.")
    sections = ["Trophy / Podium", "Match Action", "Serve — Technique", "Backhand — Technique", "Training Session", "New Equipment"]
    top0 = 5; card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th"); ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+5}", "📷\nPaste photo here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 6):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 6, column=col, value="Video link").style = "field_label"
        merge_set(ws, f"{M}{row+6}:{R}{row+6}", "", "field_value")
        ws.cell(row=row + 7, column=col, value="Notes").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")


# ===========================================================================
# 17 — Season Planner
# ===========================================================================
def build_season(wb):
    ws = wb.create_sheet("Season"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 22, 22, 16, 2])
    luxe_header(ws, "F", "📅  SEASON PLANNER",
                "Periodize the year — training blocks, tournaments & rest, at a glance.")
    table_headers(ws, 5, ["Month", "Training Block", "Key Tournaments", "Focus"], start_col=2)
    months = _recent_months(6) + [(_recent_months(1)[0])]  # placeholder; build 12 below
    plan = [
        ("Base", "Off-season prep", "—", "Strength + technique"),
        ("Build", "Volume training", "Winter Open", "Match reps"),
        ("Peak", "Competition block", "Spring Champs", "Ranking points"),
        ("Peak", "Clay swing", "Clay Court Classic", "Best surface"),
        ("Recover", "Deload week", "Rest", "Recovery + video review"),
        ("Build", "Summer prep", "Summer Level 2", "Serve + fitness"),
        ("Peak", "Grass block", "Grass Cup", "Net game"),
        ("Peak", "Fall majors", "Sectional Masters", "Big points"),
        ("Recover", "Post-season", "Rest", "Injury check"),
    ]
    labels = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep"]
    start = 6
    for i, ((blk, block, tourn, focus)) in enumerate(plan):
        r = start + i
        ws.cell(row=r, column=2, value=labels[i]).style = "td_left"
        ws.cell(row=r, column=3, value=block).style = "td_left"
        ws.cell(row=r, column=4, value=tourn).style = "td_left"
        ws.cell(row=r, column=5, value=focus).style = "td_left"
        cc = {"Base": SOFT_BG, "Build": WARN_BG, "Peak": MINT_BG, "Recover": SURFACE}[blk]
        for col in range(2, 6):
            ws.cell(row=r, column=col).fill = fill(cc)
    # legend
    merge_set(ws, "B16:E16", "BLOCK KEY", "section_gold"); ws.row_dimensions[16].height = 22
    keys = [("Base", SOFT_BG), ("Build", WARN_BG), ("Peak", MINT_BG), ("Recover", SURFACE)]
    for i, (nm, cc) in enumerate(keys):
        r = 17 + i
        c = ws.cell(row=r, column=2, value=nm); c.style = "td"; c.fill = fill(cc)
        merge_set(ws, f"C{r}:E{r}", {"Base": "Foundation — strength & technique",
                                     "Build": "Volume — match reps & fitness",
                                     "Peak": "Compete — chase ranking points",
                                     "Recover": "Rest — deload & video review"}[nm], "td_left")


# ===========================================================================
# 18 — Analytics Command Center
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 18, 3, 16, 12, 12, 2])
    luxe_header(ws, "H", "📊  ANALYTICS COMMAND CENTER",
                "Your season by the numbers — every area scored into one Player Performance Score.")
    merge_set(ws, "B5:D5", "PERFORMANCE DIMENSIONS", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Win rate", '=IFERROR(COUNTIF(MatchResult,"W")/COUNTA(MatchResult),0)'),
        ("Skill level", "=IFERROR(AVERAGE(SkillCurrent)/10,0)"),
        ("Fitness", '=IFERROR(AVERAGEIF(GoalCategory,"Fitness",GoalProgress),0)'),
        ("Tournament success", '=IFERROR(COUNTIF(TournPlace,"1st")/MAX(COUNTA(TournPlace),1),0)'),
        ("Goal progress", "=IFERROR(AVERAGE(GoalProgress),0)"),
        ("Practice consistency", "=IFERROR(MIN(SUM(PracticeHrs)/HoursTarget,1),0)"),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.75,"Elite",IF(C{r}>=0.5,"Solid","Develop"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(f"C{start}:C{end}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "F5:H5", "PLAYER PERFORMANCE SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY); cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "Wins · skills · fitness · tournaments · goals · practice.", "subtitle")
    ws["F10"].fill = fill(IVORY)
    cell_name(wb, "HealthRange", "Analytics", f"$C${start}:$C${end}")
    # rating trend table (feeds dashboard line chart)
    merge_set(ws, "B15:D15", "RATING PROGRESS (UTR) — 6 MONTHS", "section")
    ws.cell(row=16, column=2, value="Month").style = "th"; ws.cell(row=16, column=3, value="UTR").style = "th"
    months = _recent_months(6); vals = [6.8, 7.0, 7.1, 7.3, 7.4, 7.5]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 17 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = "0.0"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "RateMonth", "Analytics", "$B$17:$B$22")
    cell_name(wb, "RateVal", "Analytics", "$C$17:$C$22")
    bar = BarChart(); bar.type = "bar"; bar.title = "Performance by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None; ws.add_chart(bar, "F15")


# ===========================================================================
# 1 — Executive Tennis Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🎾  TENNIS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Matches, analytics, tournaments, fitness & finances — your whole game, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MATCHES PLAYED", "=COUNTA(MatchResult)", "num"),
        ("WIN PERCENTAGE", '=IFERROR(COUNTIF(MatchResult,"W")/COUNTA(MatchResult),0)', "pct"),
        ("CURRENT RANKING", '="#"&CurrentRank', "text"),
        ("CURRENT RATING", "=PlayerRating", "rate"),
        ("PRACTICE HOURS", "=SUM(PracticeHrs)", "dec"),
        ("TOURNAMENT WINS", '=COUNTIF(TournPlace,"1st")', "num"),
    ]
    row2 = [
        ("SETS WON", "=SUM(MatchSetsW)", "num"),
        ("GAMES WON", "=SUM(MatchGamesW)", "num"),
        ("FITNESS SCORE", '=IFERROR(AVERAGEIF(GoalCategory,"Fitness",GoalProgress),0)', "pct"),
        ("GEAR TO REPLACE", '=SUMPRODUCT((EquipReplace<=TODAY()+45)*(EquipReplace<>"")*(EquipName<>""))', "num"),
        ("MONTHLY BUDGET", "=MonthlyBudget", "money"),
        ("PERFORMANCE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "MATCHES & RATING", "section_gold")
    # match results donut
    d1 = DoughnutChart(); d1.title = "Match Results (W/L)"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Matches"], min_col=15, min_row=5, max_row=6), titles_from_data=False)
    d1.set_categories(Reference(wb["Matches"], min_col=14, min_row=5, max_row=6)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")
    # rating progress line
    ln = LineChart(); ln.title = "Rating Progress (UTR)"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); ln.legend = None
    ws.add_chart(ln, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "SKILLS & SPENDING", "section_gold")
    # skill progress bar
    sk = BarChart(); sk.type = "bar"; sk.title = "Skill Progress — Start vs Now"; sk.height = 8.2; sk.width = 11.5
    sk.add_data(Reference(wb["Skills"], min_col=2, min_row=4, max_row=17), titles_from_data=True)
    sk.add_data(Reference(wb["Skills"], min_col=3, min_row=4, max_row=17), titles_from_data=True)
    sk.set_categories(Reference(wb["Skills"], min_col=1, min_row=5, max_row=17))
    ws.add_chart(sk, "B30")
    # spending donut
    sp = DoughnutChart(); sp.title = "Budget Breakdown"; sp.height = 8.2; sp.width = 11.5
    sp.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=15), titles_from_data=True)
    sp.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=15)); sp.dataLabels = no_labels()
    ws.add_chart(sp, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Tennis Command Center™ — train, compete & improve, one organized season at a time. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_matches(wb)
    build_analytics_match(wb); build_tournaments(wb); build_practice(wb); build_skills(wb)
    build_fitness(wb); build_equipment(wb); build_budget(wb); build_partners(wb)
    build_coaching(wb); build_goals(wb); build_travel(wb); build_nutrition(wb)
    build_gallery(wb); build_season(wb); build_analytics(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Player Profile", "Matches", "Match Analytics", "Tournaments",
             "Practice", "Skills", "Fitness", "Equipment", "Budget", "Partners", "Coaching",
             "Goals", "Travel", "Nutrition", "Media", "Season", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Tennis_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
