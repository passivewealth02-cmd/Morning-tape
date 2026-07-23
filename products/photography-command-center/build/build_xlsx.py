"""Build Photography Business Command Center™ — The Photographer's Operating System.

14 tabs · a premium photography-business operating system in Google Sheets & Excel.
Dashboard, a CODB / break-even engine (what you must charge to pay yourself), a
per-shoot P&L, packages & pricing, bookings, clients & leads, an editing queue, gear,
expenses, mileage, reviews and a monthly summary — one dashboard. Price to pay
yourself, and know the profit on every shoot.

Run: python3 build_xlsx.py   ->  ../Photography_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
SHOOTTYPE = ["Wedding", "Portrait", "Event", "Branding", "Mini", "Product"]
BOOKSTATUS = ["Inquiry", "Scheduled", "Completed", "Delivered", "Cancelled"]
EDITSTATUS = ["Culling", "Editing", "Review", "Delivered"]

MARGIN_GOAL = 0.60
RATE_GOAL = 75
PACE_GOAL = 24
PROFIT_GOAL = 15000
BOOKED_AHEAD_GOAL = 24000

# CODB / break-even
OVERHEAD_ANNUAL = 24000
SALARY_TARGET = 48000
TARGET_SHOOTS = 48

# Shoot P&L — flagship wedding
PACKAGE_PRICE = 3000
SHOOT_COSTS = [("Second shooter", 300), ("Album", 200), ("Prints", 50), ("Travel", 50), ("Editing (outsourced)", 200)]
SHOOT_HOURS = 25

# Packages & pricing: (package, price, hours, includes)
PACKAGES = [
    ("Wedding — Full Day", 3000, 25, "8h coverage + album"), ("Wedding — Elopement", 1800, 12, "4h + gallery"),
    ("Portrait Session", 500, 4, "1h + 30 edits"), ("Family / Branding", 1000, 6, "2h + 50 edits"),
    ("Event Coverage", 1200, 8, "4h + gallery"), ("Mini Session", 300, 2, "20 min + 10 edits"),
]

# Bookings: (client, type, price, status)
BOOKINGS = [
    ("Rivera Wedding", "Wedding", 3000, "Scheduled"), ("Dey Wedding", "Wedding", 3000, "Scheduled"),
    ("Cho Family", "Portrait", 500, "Scheduled"), ("Lin Event", "Event", 1200, "Scheduled"),
    ("TechCo Branding", "Branding", 1000, "Scheduled"), ("Fall Minis", "Mini", 900, "Scheduled"),
    ("Harper Wedding", "Wedding", 3000, "Completed"), ("Vye Portrait", "Portrait", 500, "Inquiry"),
]

# Clients & leads: (client, type, source, stage)
CLIENTS = [
    ("Rivera", "Wedding", "Referral", "Booked"), ("Dey", "Wedding", "Instagram", "Booked"),
    ("Cho", "Portrait", "Google", "Booked"), ("Vye", "Portrait", "Website", "Inquiry"),
    ("Marsh Co.", "Branding", "Referral", "Quoted"),
]

# Editing queue: (client, shoot date, images, status)
EDITING = [
    ("Harper Wedding", "Jun 08", 620, "Delivered"), ("Cho Family", "Jun 21", 45, "Editing"),
    ("Lin Event", "Jun 28", 180, "Culling"), ("TechCo Branding", "Jul 02", 60, "Review"),
]

# Gear & inventory: (item, value, insured?)
GEAR = [
    ("Camera body ×2", 4800, "Yes"), ("Lenses (4)", 6200, "Yes"), ("Lighting kit", 1400, "Yes"),
    ("Computer & storage", 2600, "Yes"), ("Backup drives", 700, "No"), ("Props & backdrops", 900, "No"),
]

# Expenses (YTD): (item, amount)
EXPENSES = [
    ("Gear & lenses", 5000), ("Editing software", 900), ("Insurance", 1200), ("Studio rent", 6000),
    ("Marketing & website", 2400), ("Second shooters", 3600), ("Albums & prints", 2500), ("Education", 2400),
]

# Mileage & travel: (date, shoot, miles, cost)
MILEAGE = [
    ("Jun 08", "Harper Wedding", 48, 32.16), ("Jun 21", "Cho Family", 16, 10.72),
    ("Jun 28", "Lin Event", 34, 22.78), ("Jul 02", "TechCo Branding", 22, 14.74),
]

# Reviews & referrals: (client, rating, source)
REVIEWS = [
    ("Harper", 5, "Google"), ("Cho", 5, "Instagram"), ("Lin", 4, "Referral"),
    ("Marsh Co.", 5, "Website"), ("Rivera", 5, "Referral"),
]

# Monthly summary: (month, revenue, bookings)
MONTHS = [("Jan", 5000, 4), ("Feb", 6000, 5), ("Mar", 8000, 5),
          ("Apr", 7000, 5), ("May", 9000, 6), ("Jun", 10000, 5)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Studio name", "Amberlight Photo", None, "Business"),
        ("Owner", "Robin", None, "Owner"),
        ("Shoot-margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Effective-rate goal ($/hr)", RATE_GOAL, '"$"#,##0', "RateGoal"),
        ("Bookings pace goal (YTD)", PACE_GOAL, "#,##0", "PaceGoal"),
        ("Net profit goal (YTD)", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Booked-ahead goal ($)", BOOKED_AHEAD_GOAL, '"$"#,##0', "BookedAheadGoal"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Shoot type", SHOOTTYPE, "TypeList"), ("F", "Booking status", BOOKSTATUS, "BookStatusList"),
             ("G", "Edit status", EDITSTATUS, "EditStatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  📷  PHOTOGRAPHY BUSINESS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Price to pay yourself, and know the profit on every shoot.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE PHOTO BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A photography business lives or dies on two numbers: the price you must charge to cover your costs "
                      "and pay yourself, and the profit each shoot actually makes. This makes both visible: a cost-of-"
                      "doing-business engine that turns your overhead and desired salary into a break-even price per "
                      "shoot, and a per-shoot P&L that shows your net and your real hourly rate. Set your packages, run "
                      "bookings, manage clients and the editing queue, track gear, expenses and mileage — all in ONE "
                      "premium Google Sheets & Excel system built for photographers.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your margin, rate & booked-ahead goals.",
             "2.  Fill CODB & Break-Even — your minimum price per shoot.",
             "3.  Cost a shoot in Shoot P&L — net and your real hourly rate.",
             "4.  Set your Packages & Pricing above break-even.",
             "5.  Run Bookings, clients, the editing queue and gear.",
             "6.  Check the Dashboard: profit, break-even & a Studio Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional studio (Amberlight Photo) is included so you can see how it all connects — "
               "just type over it with your own shoots and numbers. Your break-even price and the profit per shoot are "
               "the two numbers that decide whether photography pays, and they roll into a live Studio Score. Twelve "
               "matching printable pages (CODB worksheet, shoot P&L, package menu, shot list & more) are included. This "
               "is a business tool, not financial, legal or tax advice — confirm figures with your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Know your break-even, then price above it — that's how photographers stay in business.", "section_gold")


# ===========================================================================
def build_codb(wb):
    ws = wb.create_sheet("CODB & Break-Even"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 16, 2])
    luxe_header(ws, "C", "🧮  CODB & BREAK-EVEN",
                "Your overhead + the salary you want ÷ the shoots you'll book = what you must charge to pay yourself.")
    ws.cell(row=5, column=2, value="Annual overhead & business costs").style = "field_label"
    co = ws.cell(row=5, column=3, value=OVERHEAD_ANNUAL); co.style = "input"; co.number_format = '"$"#,##0'
    cell_name(wb, "OverheadAnnual", "CODB & Break-Even", "$C$5")
    ws.cell(row=6, column=2, value="Salary you want to pay yourself").style = "field_label"
    cs = ws.cell(row=6, column=3, value=SALARY_TARGET); cs.style = "input"; cs.number_format = '"$"#,##0'
    cell_name(wb, "SalaryTarget", "CODB & Break-Even", "$C$6")
    ws.cell(row=7, column=2, value="= TOTAL YOU NEED (year)").style = "th"
    ct = ws.cell(row=7, column=3, value="=OverheadAnnual+SalaryTarget"); ct.style = "td"; ct.font = Font(bold=True, size=12, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0'
    cell_name(wb, "TotalNeeded", "CODB & Break-Even", "$C$7")
    ws.cell(row=9, column=2, value="Shoots you'll book (year)").style = "field_label"
    csh = ws.cell(row=9, column=3, value=TARGET_SHOOTS); csh.style = "input"; csh.number_format = "#,##0"
    cell_name(wb, "TargetShoots", "CODB & Break-Even", "$C$9")
    ws.cell(row=10, column=2, value="= CODB / SHOOT (break-even price)").style = "th"
    cc = ws.cell(row=10, column=3, value="=IFERROR(TotalNeeded/TargetShoots,0)"); cc.style = "td"; cc.font = Font(bold=True, size=13, color=PRIMARY); cc.fill = fill(MINT_BG); cc.number_format = '"$"#,##0'
    cell_name(wb, "CODBShoot", "CODB & Break-Even", "$C$10")
    ws.cell(row=12, column=2, value="Every package you sell must beat this number — or you're paying to work.").style = "section"


def build_shootpnl(wb):
    ws = wb.create_sheet("Shoot P&L"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "💵  SHOOT P&L",
                "Price a shoot to the profit — costs out, and your real hourly rate for the shoot and the edit.")
    ws.cell(row=5, column=2, value="SHOOT").style = "section_gold"
    ws.cell(row=5, column=3, value="Wedding — Full Day").font = Font(bold=True, color=PRIMARY)
    ws.cell(row=6, column=2, value="Package price").style = "field_label"
    cp = ws.cell(row=6, column=3, value=PACKAGE_PRICE); cp.style = "input"; cp.number_format = '"$"#,##0'
    cell_name(wb, "PackagePrice", "Shoot P&L", "$C$6")
    table_headers(ws, 8, ["Cost", "Amount"], start_col=2)
    start = 9
    for i, (item, amt) in enumerate(SHOOT_COSTS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=amt); cc.style = "input"; cc.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SHOOT_COSTS) - 1
    nrange(wb, "ShootCostAmt", "Shoot P&L", "C", start, end)
    ctot = end + 1
    ws.cell(row=ctot, column=2, value="= SHOOT COSTS").style = "th"
    cst = ws.cell(row=ctot, column=3, value="=SUM(ShootCostAmt)"); cst.style = "td"; cst.font = Font(bold=True, color=DANGER); cst.fill = fill(SURFACE); cst.number_format = '"$"#,##0'
    cell_name(wb, "ShootCosts", "Shoot P&L", f"$C${ctot}")
    nr = ctot + 1
    ws.cell(row=nr, column=2, value="= NET PER SHOOT").style = "th"
    cn = ws.cell(row=nr, column=3, value="=PackagePrice-ShootCosts"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetShoot", "Shoot P&L", f"$C${nr}")
    ws.cell(row=nr + 2, column=2, value="Hours (shoot + culling + editing)").style = "field_label"
    ch = ws.cell(row=nr + 2, column=3, value=SHOOT_HOURS); ch.style = "input"; ch.number_format = "#,##0"
    cell_name(wb, "ShootHours", "Shoot P&L", f"$C${nr+2}")
    ws.cell(row=nr + 3, column=2, value="= EFFECTIVE RATE ($/hr)").style = "th"
    ce = ws.cell(row=nr + 3, column=3, value="=IFERROR(NetShoot/ShootHours,0)"); ce.style = "td"; ce.font = Font(bold=True, size=12, color=PRIMARY); ce.fill = fill(MINT_BG); ce.number_format = '"$"#,##0.00'
    cell_name(wb, "EffRate", "Shoot P&L", f"$C${nr+3}")
    ws.cell(row=nr + 4, column=2, value="Shoot margin").style = "field_label"
    cm = ws.cell(row=nr + 4, column=3, value="=IFERROR(NetShoot/PackagePrice,0)"); cm.style = "field_value"; cm.number_format = "0%"; cm.fill = fill(MINT_BG)
    cell_name(wb, "ShootMargin", "Shoot P&L", f"$C${nr+4}")


def build_packages(wb):
    ws = wb.create_sheet("Packages & Pricing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 10, 24, 12, 2])
    luxe_header(ws, "F", "📦  PACKAGES & PRICING",
                "Your menu — price, hours, what's included and the rate per hour. Keep every package above break-even.")
    table_headers(ws, 4, ["Package", "Price", "Hours", "Includes", "$/hr"], start_col=2)
    start = L0
    for i, (pkg, price, hrs, inc) in enumerate(PACKAGES):
        r = start + i
        ws.cell(row=r, column=2, value=pkg).style = "td_left"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        ch = ws.cell(row=r, column=4, value=hrs); ch.style = "input"; ch.number_format = "#,##0"
        ws.cell(row=r, column=5, value=inc).style = "td_left"
        cr = ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/D{r},0)"); cr.style = "td"; cr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PACKAGES) - 1
    nrange(wb, "PkgName", "Packages & Pricing", "B", start, end)
    nrange(wb, "PkgPrice", "Packages & Pricing", "C", start, end)
    ws.freeze_panes = "A5"


def build_bookings(wb):
    ws = wb.create_sheet("Bookings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 14, 16, 2])
    luxe_header(ws, "E", "📅  BOOKINGS",
                "Your calendar — type, price and status. Scheduled shoots become your booked-ahead revenue.")
    table_headers(ws, 4, ["Client / Shoot", "Type", "Price", "Status"], start_col=2)
    start = L0
    for i, (client, typ, price, status) in enumerate(BOOKINGS):
        r = start + i
        ws.cell(row=r, column=2, value=client).style = "td_left"
        ws.cell(row=r, column=3, value=typ).style = "td"
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=status).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "TypeList")
        add_dv(ws, f"E{r}", "BookStatusList")
    end = start + len(BOOKINGS) - 1
    nrange(wb, "BookPrice", "Bookings", "D", start, end)
    nrange(wb, "BookStatus", "Bookings", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="BOOKED AHEAD (scheduled)").style = "th"
    for c in (3,):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cb = ws.cell(row=tot, column=4, value='=SUMIF(BookStatus,"Scheduled",BookPrice)'); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.fill = fill(MINT_BG); cb.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "UpcomingBookings", "Bookings", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_clients(wb):
    ws, start, end = build_log(
        wb, "Clients & Leads", "👥", "CLIENTS & LEADS",
        "Your inquiries and booked clients — type, source and stage, so no lead is lost.",
        ["Client", "Type", "Source", "Stage"],
        CLIENTS, [2, 22, 14, 16, 14, 2], text_left={2}, reserved=24, start_col=2,
        validations=[("C", "TypeList")])


def build_editing(wb):
    ws, start, end = build_log(
        wb, "Editing Queue", "🖥", "EDITING QUEUE",
        "What's in post — shoot date, image count and status, so galleries go out on time.",
        ["Client", "Shoot Date", "Images", "Status"],
        EDITING, [2, 24, 16, 12, 14, 2], text_left={2, 3}, ints={4}, reserved=24, start_col=2,
        validations=[("E", "EditStatusList")])


def build_gear(wb):
    ws = wb.create_sheet("Gear & Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 2])
    luxe_header(ws, "D", "🎒  GEAR & INVENTORY",
                "Your kit and its value — for insurance, depreciation and knowing what to replace.")
    table_headers(ws, 4, ["Item", "Value", "Insured?"], start_col=2)
    start = L0
    for i, (item, val, ins) in enumerate(GEAR):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cv = ws.cell(row=r, column=3, value=val); cv.style = "input"; cv.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=ins).style = "td"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"D{r}", "YesNoList")
    end = start + len(GEAR) - 1
    nrange(wb, "GearVal", "Gear & Inventory", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL KIT VALUE").style = "th"
    cv = ws.cell(row=tot, column=3, value="=SUM(GearVal)"); cv.style = "td"; cv.font = Font(bold=True, color=PRIMARY); cv.fill = fill(MINT_BG); cv.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  EXPENSES",
                "Your business costs (YTD) — netted against revenue to show your real profit.")
    table_headers(ws, 4, ["Expense", "Amount"], start_col=2)
    start = L0
    for i, (item, amt) in enumerate(EXPENSES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EXPENSES) - 1
    nrange(wb, "ExpAmt", "Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES (YTD)").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Expenses", f"$C${tot}")
    nr = tot + 2
    ws.cell(row=nr, column=2, value="= NET PROFIT (revenue − expenses)").style = "th"
    cn = ws.cell(row=nr, column=3, value="=RevenueYTD-ExpTotal"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetProfit", "Expenses", f"$C${nr}")
    ws.freeze_panes = "A5"


def build_mileage(wb):
    ws, start, end = build_log(
        wb, "Mileage & Travel", "🚗", "MILEAGE & TRAVEL",
        "Every drive to a shoot — miles and cost, deductible at tax time.",
        ["Date", "Shoot", "Miles", "Cost"],
        MILEAGE, [2, 14, 24, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)


def build_reviews(wb):
    ws, start, end = build_log(
        wb, "Reviews & Referrals", "⭐", "REVIEWS & REFERRALS",
        "Every review and where it came from — referrals book the calendar for free.",
        ["Client", "Rating", "Source"],
        REVIEWS, [2, 24, 12, 18, 2], text_left={2, 4}, dec={3}, reserved=24, start_col=2)
    nrange(wb, "RevRating", "Reviews & Referrals", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVG RATING").style = "th"
    c = ws.cell(row=tot, column=3, value="=IFERROR(AVERAGE(RevRating),0)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = "0.0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 2])
    luxe_header(ws, "D", "📈  MONTHLY SUMMARY",
                "Revenue & bookings by month — watch the year build toward your goal.")
    table_headers(ws, 4, ["Month", "Revenue", "Bookings"], start_col=2)
    start = L0
    for i, (m, rev, bk) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rev); cr.style = "input"; cr.number_format = '"$"#,##0'
        cb = ws.cell(row=r, column=4, value=bk); cb.style = "input"; cb.number_format = "#,##0"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthRev", "Monthly Summary", "C", start, end)
    nrange(wb, "MonthBook", "Monthly Summary", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="YTD").style = "th"
    cr = ws.cell(row=tot, column=3, value="=SUM(MonthRev)"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0'
    cb = ws.cell(row=tot, column=4, value="=SUM(MonthBook)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.fill = fill(SURFACE); cb.number_format = "#,##0"
    cell_name(wb, "RevenueYTD", "Monthly Summary", f"$C${tot}")
    cell_name(wb, "BookingsYTD", "Monthly Summary", f"$D${tot}")
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Average booking value").style = "field_label"
    ca = ws.cell(row=sr, column=3, value="=IFERROR(RevenueYTD/BookingsYTD,0)"); ca.style = "field_value"; ca.number_format = '"$"#,##0'; ca.fill = fill(MINT_BG)
    cell_name(wb, "AvgBooking", "Monthly Summary", f"$C${sr}")
    ws.add_chart(_barchart(ws, "Revenue by Month", start, end, 3, 2), "F4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  📷  PHOTOGRAPHY BUSINESS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Break-even, profit per shoot, revenue & a Studio Score — your whole business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("PACKAGE PRICE", "=PackagePrice", "money"),
        ("SHOOT COSTS", "=ShootCosts", "money"),
        ("NET / SHOOT", "=NetShoot", "money"),
        ("SHOOT MARGIN", "=ShootMargin", "pct"),
        ("CODB / SHOOT", "=CODBShoot", "money"),
        ("EFFECTIVE RATE", "=EffRate", "money2"),
    ]
    row2 = [
        ("BOOKINGS YTD", "=BookingsYTD", "num"),
        ("REVENUE YTD", "=RevenueYTD", "money"),
        ("AVG BOOKING", "=AvgBooking", "money"),
        ("NET PROFIT", "=NetProfit", "money"),
        ("BOOKED AHEAD", "=UpcomingBookings", "money"),
        ("STUDIO SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "STUDIO HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Shoot margin healthy", "=IFERROR(MIN(ShootMargin/MarginGoal,1),0)"),
        ("Net per shoot beats break-even", "=IFERROR(MIN(NetShoot/CODBShoot,1),0)"),
        ("Effective rate on target", "=IFERROR(MIN(EffRate/RateGoal,1),0)"),
        ("Bookings on pace", "=IFERROR(MIN(BookingsYTD/PaceGoal,1),0)"),
        ("Profitable", "=IFERROR(MIN(NetProfit/ProfitGoal,1),0)"),
        ("Booked ahead", "=IFERROR(MIN(UpcomingBookings/BookedAheadGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Photography Business Command Center™ — price to pay yourself, profit on every shoot.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_codb(wb); build_shootpnl(wb)
    build_packages(wb); build_bookings(wb); build_clients(wb); build_editing(wb)
    build_gear(wb); build_expenses(wb); build_mileage(wb); build_reviews(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "CODB & Break-Even", "Shoot P&L", "Packages & Pricing", "Bookings",
             "Clients & Leads", "Editing Queue", "Gear & Inventory", "Expenses", "Mileage & Travel",
             "Reviews & Referrals", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Photography_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
