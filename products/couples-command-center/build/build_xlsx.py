"""Build Relationship & Couples Command Center™ — Both of You, Fairly.

14 tabs · a premium shared-life operating system in Google Sheets & Excel. Dashboard, a
fair-share engine (split the bills by what each of you earns, not down the middle), shared
bills, the invisible labour split in hours, money goals, savings, date nights, a weekly
check-in, the big conversations, household admin, individual money and a monthly summary
— one dashboard. 50/50 is not the same as fair.

Run: python3 build_xlsx.py   ->  ../Couples_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
WHOPAYS = ["Partner A", "Partner B", "Split proportionally", "Split 50/50", "Joint account"]
BILLCAT = ["Home", "Food", "Utilities", "Transport", "Insurance", "Fun", "Other"]
GOALSTATUS = ["On track", "Ahead", "Behind", "Done", "Paused"]
FREQ = ["Weekly", "Fortnightly", "Monthly", "Quarterly", "Yearly"]

# --- Fair-share engine ---
A_NAME = "Nadia"
B_NAME = "Sam"
A_INCOME = 4200.00
B_INCOME = 6300.00

# --- Goals ---
FAIRNESS_GOAL = 0.95
DATE_GOAL = 4
CHECKIN_GOAL = 4
SAVINGS_GOAL = 800
CHORE_RATIO_GOAL = 1.15

DATE_NIGHTS = 4
CHECKINS = 4
SAVED_THIS_MONTH = 940

# Shared bills: (bill, category, monthly, who pays it)
BILLS = [
    ("Rent", "Home", 2150, "Joint account"), ("Utilities", "Utilities", 245, "Joint account"),
    ("Groceries", "Food", 680, "Joint account"), ("Internet & phones", "Utilities", 145, "Joint account"),
    ("Insurance", "Insurance", 190, "Joint account"), ("Car & transit", "Transport", 310, "Joint account"),
    ("Subscriptions", "Fun", 68, "Joint account"), ("Household & repairs", "Home", 120, "Joint account"),
]

# Invisible labour, hours per week: (task, A hours, B hours)
CHORES = [
    ("Cooking", 5.5, 0.5),
    ("Cleaning", 4.0, 1.0),
    ("Laundry", 2.5, 0.5),
    ("Shopping & errands", 3.0, 0.5),
    ("Admin, bills & appointments", 3.0, 0.5),
    ("Planning & remembering \u2014 the mental load", 4.0, 0.5),
    ("Pets", 1.0, 1.5),
    ("Yard, car & repairs", 0.0, 3.0),
]

# Money goals: (goal, target, saved, by when, status)
GOALS = [
    ("Emergency fund \u2014 6 months", 18000, 11400, "Jun 2027", "On track"),
    ("House deposit", 45000, 16800, "2029", "On track"),
    ("Trip \u2014 Portugal", 3600, 2150, "May 2027", "Ahead"),
    ("New car fund", 9000, 3400, "2028", "On track"),
    ("Christmas & birthdays", 1800, 900, "Dec 2026", "On track"),
]

# Savings & sinking funds: (fund, monthly, balance)
SAVINGS = [
    ("Emergency fund", 400, 11400), ("House deposit", 300, 16800), ("Trip fund", 120, 2150),
    ("Car fund", 80, 3400), ("Gifts & Christmas", 40, 900),
]

# Date nights: (date, what, who planned, cost, how it was)
DATES = [
    ("07/04", "Dinner at the Portuguese place", "Sam", 78, "Lovely"),
    ("07/12", "Cinema and a walk home", "Nadia", 34, "Easy"),
    ("07/19", "Cooked together, no phones", "Both", 22, "Best one"),
    ("07/26", "Gallery then coffee", "Nadia", 28, "Lovely"),
]

# Weekly check-in: (week, done?, what came up)
CHECKINS_LOG = [
    ("Week 1", "Yes", "Money felt tight \u2014 moved the trip fund down $40"),
    ("Week 2", "Yes", "Sam's work trip in Sept. Diarised."),
    ("Week 3", "Yes", "Talked about the chore split. Properly, this time."),
    ("Week 4", "Yes", "Booked the Portugal flights. Both excited."),
]

# The big conversations: (topic, discussed?, where we landed)
BIG_TALKS = [
    ("Do we both want children, and when", "Yes", "Yes, both. Trying from 2028."),
    ("Where we want to live in five years", "Yes", "Here, but a house not a flat"),
    ("How we handle a big unexpected bill", "Yes", "Emergency fund first, never a card"),
    ("What we each need when we're stressed", "Yes", "N: talk it out. S: an hour alone first."),
    ("Money we each spend without asking", "Yes", "Under $150, no questions"),
    ("How we split things if incomes change", "Yes", "Always proportional. Re-run this tab."),
    ("What we want the next ten years to feel like", "No", "Keep meaning to. Book an evening."),
    ("Wills, beneficiaries & who decides if", "No", "Neither of us has done this. Do it."),
]

# Household admin: (item, whose job, when, notes)
ADMIN = [
    ("Rent & bills paid", "Joint account", "Monthly", "Auto \u2014 both can see it"),
    ("Insurance renewals", "Sam", "Yearly", "Shop around, don't auto-renew"),
    ("Car service & MOT", "Sam", "Yearly", "March"),
    ("Doctor & dentist bookings", "Nadia", "Quarterly", "For both of us"),
    ("Birthdays & gifts", "Nadia", "Monthly", "This is invisible work \u2014 see the labour tab"),
    ("Deep clean", "Both", "Monthly", "Together, first Saturday"),
    ("Fire alarms & filters", "Sam", "Quarterly", "Phone reminder set"),
    ("Review this whole workbook", "Both", "Monthly", "Last Sunday of the month"),
]

# Individual money — what's yours is yours: (item, A, B)
INDIVIDUAL = [
    ("Income", 4200, 6300),
    ("Your share of the bills", 1563.20, 2344.80),
    ("Personal savings", 250, 400),
    ("Guilt-free spending", 400, 400),
]

# Monthly summary: (month, saved)
MONTHS = [("Feb", 780), ("Mar", 810), ("Apr", 845), ("May", 880), ("Jun", 910), ("Jul", 940)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "money3": '"$"#,##0.000', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 40, 20, 3] + [24] * 5)
    luxe_header(ws, "I", "⚙  SETTINGS", "Your names, your incomes, your goals — every tab follows.")
    merge_set(ws, "B5:C5", "THE TWO OF YOU", "section")
    controls = [
        ("Partner A", A_NAME, None, "PartnerA"),
        ("Partner B", B_NAME, None, "PartnerB"),
        ("A \\u2014 monthly income after tax", A_INCOME, '"$"#,##0.00', "IncomeA"),
        ("B \\u2014 monthly income after tax", B_INCOME, '"$"#,##0.00', "IncomeB"),
        ("Goal: date nights a month", DATE_GOAL, "0", "DateGoal"),
        ("Goal: check-ins a month", CHECKIN_GOAL, "0", "CheckinGoal"),
        ("Goal: saved together each month", SAVINGS_GOAL, '"$"#,##0', "SavingsGoal"),
        ("Goal: chore split no worse than", CHORE_RATIO_GOAL, "0.00", "ChoreGoal"),
        ("Goal: fair-split accuracy", FAIRNESS_GOAL, "0%", "FairnessGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Who pays", WHOPAYS, "WhoPaysList"), ("F", "Bill category", BILLCAT, "BillCatList"),
             ("G", "Goal status", GOALSTATUS, "GoalStatusList"), ("H", "How often", FREQ, "FreqList"),
             ("I", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:I5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ◎  RELATIONSHIP & COUPLES COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  50/50 is not the same as fair.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "WHY THIS EXISTS", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Two people earning different amounts, splitting the bills down the middle, is not equality \\u2014 "
                      "it just looks like it. On the numbers in this file, a 50/50 split leaves one of you with $2,246 "
                      "a month and the other with $4,346. Split the same bills in proportion to what you each earn and "
                      "you both keep exactly 62.8% of your own income. Same house, same bills, completely different "
                      "life. Then the second half: the hours. Cooking, laundry, admin, remembering birthdays \\u2014 add "
                      "them up honestly and see what the week actually looks like. Both of these are conversations "
                      "that go much better with a number in front of you.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Settings: your names and what you each actually bring home.",
             "2.  Shared Bills: every bill you pay together, in one list.",
             "3.  Fair Share does the maths. Read it together, not alone.",
             "4.  Invisible Labour: both of you fill it in. Separately, then compare.",
             "5.  Money goals, savings, date nights, the weekly check-in.",
             "6.  Big Conversations. The two unticked ones are the point."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  HOW TO USE THIS WITHOUT A FIGHT", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+4}")
    c = ws[f"B{dr+1}"]
    c.value = ("Fill in the Invisible Labour tab separately, then compare. Almost every couple finds the two columns "
               "don't match \\u2014 not because anyone is lying, but because the work you don't do is genuinely hard to "
               "see. That gap is the conversation, and it is a conversation, not a verdict. This is an organizing tool, "
               "not relationship or financial advice, and it can't tell you what a fair life together looks like \\u2014 "
               "only the two of you can. If money is being controlled rather than shared, a spreadsheet is not the "
               "right help; please talk to someone. Sample data for a fictional couple is included; type over it.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 5):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+6}:B{dr+6}", "These conversations go much better with a number in front of you.", "section_gold")


# ===========================================================================
def build_bills(wb):
    ws = wb.create_sheet("Shared Bills"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 18, 18, 26, 18, 2])
    luxe_header(ws, "F", "\U0001f9fe  SHARED BILLS",
                "Everything the two of you pay for together. Everything.")
    table_headers(ws, 4, ["Bill", "Category", "Monthly", "Who pays it", "Yearly"], start_col=2)
    start = L0
    for i, (bill, cat, amt, who) in enumerate(BILLS):
        r = start + i
        ws.cell(row=r, column=2, value=bill).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        c = ws.cell(row=r, column=4, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=who).style = "input"
        cy = ws.cell(row=r, column=6, value=f"=D{r}*12"); cy.style = "td"; cy.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 7):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    end = start + len(BILLS) - 1
    nrange(wb, "BillAmount", "Shared Bills", "D", start, end)
    add_dv(ws, f"C{start}:C{end}", "BillCatList"); add_dv(ws, f"E{start}:E{end}", "WhoPaysList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="= SHARED BILLS / MONTH").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=4, value="=SUM(BillAmount)"); ct.style = "td"
    ct.font = Font(bold=True, size=14, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "SharedBills", "Shared Bills", f"$D${tot}")
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cty = ws.cell(row=tot, column=6, value="=SharedBills*12"); cty.style = "td"
    cty.font = Font(bold=True, size=12, color=PRIMARY); cty.fill = fill(MINT_BG); cty.number_format = '"$"#,##0'
    ws.cell(row=tot + 2, column=2, value="If one of you doesn't know what a bill costs, that's the first thing to fix.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_fairshare(wb):
    ws = wb.create_sheet("Fair Share"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 48, 20, 20, 2])
    luxe_header(ws, "D", "⚖  FAIR SHARE — THE ENGINE",
                "Split the bills by what you each earn, not down the middle. Watch what changes.")
    ws.cell(row=5, column=2, value="WHAT YOU EACH BRING HOME").style = "section_gold"
    ws.cell(row=6, column=2, value="").style = "field_label"
    ws.cell(row=6, column=3, value="=PartnerA").style = "th"
    ws.cell(row=6, column=4, value="=PartnerB").style = "th"
    rows = [
        ("Monthly income after tax", "=IncomeA", "=IncomeB", '"$"#,##0.00', None, None),
        ("Share of your combined income", "=IFERROR(IncomeA/(IncomeA+IncomeB),0)",
         "=IFERROR(IncomeB/(IncomeA+IncomeB),0)", "0.0%", "ShareA", "ShareB"),
    ]
    for i, (lab, fa, fb, fmt, na, nb) in enumerate(rows):
        r = 7 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        ca = ws.cell(row=r, column=3, value=fa); ca.style = "field_value"; ca.number_format = fmt
        cb = ws.cell(row=r, column=4, value=fb); cb.style = "field_value"; cb.number_format = fmt
        if na:
            cell_name(wb, na, "Fair Share", f"$C${r}"); cell_name(wb, nb, "Fair Share", f"$D${r}")

    ws.cell(row=10, column=2, value="THE 50/50 SPLIT \\u2014 WHAT IT LOOKS LIKE").style = "section_gold"
    ws.cell(row=11, column=2, value="Each of you pays half the shared bills").style = "field_label"
    ha = ws.cell(row=11, column=3, value="=SharedBills/2"); ha.style = "field_value"; ha.number_format = '"$"#,##0.00'
    hb = ws.cell(row=11, column=4, value="=SharedBills/2"); hb.style = "field_value"; hb.number_format = '"$"#,##0.00'
    ws.cell(row=12, column=2, value="= WHAT'S LEFT FOR EACH OF YOU").style = "th"
    la = ws.cell(row=12, column=3, value="=IncomeA-SharedBills/2"); la.style = "td"
    la.font = Font(bold=True, size=13, color=DANGER); la.fill = fill(RED_BG); la.number_format = '"$"#,##0'
    cell_name(wb, "LeftHalfA", "Fair Share", "$C$12")
    lb = ws.cell(row=12, column=4, value="=IncomeB-SharedBills/2"); lb.style = "td"
    lb.font = Font(bold=True, size=13, color=PRIMARY); lb.fill = fill(SURFACE); lb.number_format = '"$"#,##0'
    cell_name(wb, "LeftHalfB", "Fair Share", "$D$12")
    ws.cell(row=13, column=2, value="…as a share of your own income").style = "field_label"
    pa = ws.cell(row=13, column=3, value="=IFERROR((IncomeA-SharedBills/2)/IncomeA,0)"); pa.style = "field_value"; pa.number_format = "0.0%"
    pb = ws.cell(row=13, column=4, value="=IFERROR((IncomeB-SharedBills/2)/IncomeB,0)"); pb.style = "field_value"; pb.number_format = "0.0%"
    ws.cell(row=14, column=2, value="= THE GAP BETWEEN YOU, EVERY MONTH").style = "th"
    cg = ws.cell(row=14, column=3, value="=ABS(LeftHalfB-LeftHalfA)"); cg.style = "td"
    cg.font = Font(bold=True, size=16, color=DANGER); cg.fill = fill(RED_BG); cg.number_format = '"$"#,##0'
    cell_name(wb, "HalfGap", "Fair Share", "$C$14")

    ws.cell(row=16, column=2, value="THE PROPORTIONAL SPLIT \\u2014 THE FAIR ONE").style = "section_gold"
    ws.cell(row=17, column=2, value="Each of you pays your share of the bills").style = "field_label"
    fa = ws.cell(row=17, column=3, value="=SharedBills*ShareA"); fa.style = "td"
    fa.font = Font(bold=True, size=13, color=PRIMARY); fa.fill = fill(SURFACE); fa.number_format = '"$"#,##0.00'
    cell_name(wb, "FairA", "Fair Share", "$C$17")
    fb = ws.cell(row=17, column=4, value="=SharedBills*ShareB"); fb.style = "td"
    fb.font = Font(bold=True, size=13, color=PRIMARY); fb.fill = fill(SURFACE); fb.number_format = '"$"#,##0.00'
    cell_name(wb, "FairB", "Fair Share", "$D$17")
    ws.cell(row=18, column=2, value="= WHAT'S LEFT FOR EACH OF YOU").style = "th"
    ra = ws.cell(row=18, column=3, value="=IncomeA-FairA"); ra.style = "td"
    ra.font = Font(bold=True, size=13, color=PRIMARY); ra.fill = fill(MINT_BG); ra.number_format = '"$"#,##0'
    rb = ws.cell(row=18, column=4, value="=IncomeB-FairB"); rb.style = "td"
    rb.font = Font(bold=True, size=13, color=PRIMARY); rb.fill = fill(MINT_BG); rb.number_format = '"$"#,##0'
    ws.cell(row=19, column=2, value="= …AS A SHARE OF YOUR OWN INCOME").style = "th"
    qa = ws.cell(row=19, column=3, value="=IFERROR((IncomeA-FairA)/IncomeA,0)"); qa.style = "td"
    qa.font = Font(bold=True, size=15, color=PRIMARY); qa.fill = fill(MINT_BG); qa.number_format = "0.0%"
    cell_name(wb, "KeepPctA", "Fair Share", "$C$19")
    qb = ws.cell(row=19, column=4, value="=IFERROR((IncomeB-FairB)/IncomeB,0)"); qb.style = "td"
    qb.font = Font(bold=True, size=15, color=PRIMARY); qb.fill = fill(MINT_BG); qb.number_format = "0.0%"
    cell_name(wb, "KeepPctB", "Fair Share", "$D$19")
    ws.cell(row=20, column=2, value="Identical. That is the entire point of this page.").style = "section_gold"

    ws.cell(row=22, column=2, value="WHAT YOU'RE ACTUALLY PAYING RIGHT NOW").style = "section_gold"
    ws.cell(row=23, column=2, value="Enter what each of you actually transfers").style = "field_label"
    aa = ws.cell(row=23, column=3, value=1563.20); aa.style = "input"; aa.number_format = '"$"#,##0.00'
    cell_name(wb, "PaysA", "Fair Share", "$C$23")
    ab = ws.cell(row=23, column=4, value=2344.80); ab.style = "input"; ab.number_format = '"$"#,##0.00'
    cell_name(wb, "PaysB", "Fair Share", "$D$23")
    ws.cell(row=24, column=2, value="Difference from your fair share").style = "field_label"
    da = ws.cell(row=24, column=3, value="=PaysA-FairA"); da.style = "field_value"; da.number_format = '"$"#,##0.00'
    db = ws.cell(row=24, column=4, value="=PaysB-FairB"); db.style = "field_value"; db.number_format = '"$"#,##0.00'
    ws.cell(row=25, column=2, value="= HOW CLOSE YOU ARE TO FAIR").style = "th"
    cf = ws.cell(row=25, column=3, value="=IFERROR(1-(ABS(PaysA-FairA)+ABS(PaysB-FairB))/SharedBills,0)"); cf.style = "td"
    cf.font = Font(bold=True, size=15, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = "0.0%"
    cell_name(wb, "Fairness", "Fair Share", "$C$25")
    ws.cell(row=27, column=2, value="If one income changes, come back to this page. It is a five-minute conversation, once.").style = "section_gold"


def build_labour(wb):
    ws = wb.create_sheet("Invisible Labour"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 44, 16, 16, 16, 18, 2])
    luxe_header(ws, "F", "\U0001f9fa  INVISIBLE LABOUR",
                "Hours a week. Fill it in separately, then compare. The gap is the conversation.")
    ws.cell(row=4, column=2, value="Include the remembering, not just the doing. Knowing the dog is due at the vet IS the work.").style = "section_gold"
    table_headers(ws, 5, ["Task", "=PartnerA", "=PartnerB", "Gap", "Who it falls on"], start_col=2)
    ws.cell(row=5, column=3, value="=PartnerA").style = "th"
    ws.cell(row=5, column=4, value="=PartnerB").style = "th"
    start = 6
    for i, (task, a, b) in enumerate(CHORES):
        r = start + i
        ws.cell(row=r, column=2, value=task).style = "td_left"
        ca = ws.cell(row=r, column=3, value=a); ca.style = "input"; ca.number_format = "0.0"
        cb = ws.cell(row=r, column=4, value=b); cb.style = "input"; cb.number_format = "0.0"
        cg = ws.cell(row=r, column=5, value=f"=ABS(C{r}-D{r})"); cg.style = "td"; cg.number_format = "0.0"
        cw = ws.cell(row=r, column=6, value=f'=IF(C{r}>D{r},PartnerA,IF(D{r}>C{r},PartnerB,"Even"))'); cw.style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CHORES) - 1
    nrange(wb, "ChoreA", "Invisible Labour", "C", start, end)
    nrange(wb, "ChoreB", "Invisible Labour", "D", start, end)
    nrange(wb, "ChoreGap", "Invisible Labour", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}", DataBarRule(start_type="min", end_type="max", color=GOLD_LT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="HOURS A WEEK").style = "th"
    ca = ws.cell(row=tot, column=3, value="=SUM(ChoreA)"); ca.style = "td"
    ca.font = Font(bold=True, size=14, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = "0.0"
    cell_name(wb, "HoursA", "Invisible Labour", f"$C${tot}")
    cb = ws.cell(row=tot, column=4, value="=SUM(ChoreB)"); cb.style = "td"
    cb.font = Font(bold=True, size=14, color=PRIMARY); cb.fill = fill(SURFACE); cb.number_format = "0.0"
    cell_name(wb, "HoursB", "Invisible Labour", f"$D${tot}")
    cg = ws.cell(row=tot, column=5, value="=ABS(HoursA-HoursB)"); cg.style = "td"
    cg.font = Font(bold=True, size=14, color=DANGER); cg.fill = fill(RED_BG); cg.number_format = "0.0"

    r = tot + 2
    ws.cell(row=r, column=2, value="= ONE OF YOU IS DOING THIS MUCH MORE").style = "th"
    cr = ws.cell(row=r, column=3, value="=IFERROR(MAX(HoursA,HoursB)/MIN(HoursA,HoursB),0)"); cr.style = "td"
    cr.font = Font(bold=True, size=18, color=PRIMARY); cr.fill = fill(WARN_BG); cr.number_format = '0.00"\\u00d7"'
    cell_name(wb, "ChoreRatio", "Invisible Labour", f"$C${r}")
    ws.cell(row=r, column=4, value='=IF(HoursA>HoursB,PartnerA,PartnerB)').style = "td"
    ws.cell(row=r + 1, column=2, value="= EXTRA HOURS A YEAR").style = "th"
    cy = ws.cell(row=r + 1, column=3, value="=ABS(HoursA-HoursB)*52"); cy.style = "td"
    cy.font = Font(bold=True, size=16, color=PRIMARY); cy.fill = fill(RED_BG); cy.number_format = '#,##0" hrs"'
    cell_name(wb, "ExtraHoursYear", "Invisible Labour", f"$C${r+1}")
    ws.cell(row=r + 2, column=2, value="…which is this many full working weeks").style = "field_label"
    cw2 = ws.cell(row=r + 2, column=3, value="=ExtraHoursYear/40"); cw2.style = "field_value"; cw2.number_format = '0.0" weeks"'
    ws.cell(row=r + 4, column=2, value="Nobody is lying when the columns don't match. The work you don't do is genuinely hard to see.").style = "section_gold"
    ws.cell(row=r + 5, column=2, value="Pick ONE row to move this month. Not all of them. One.").style = "field_label"
    ws.freeze_panes = "A6"


def build_goals(wb):
    ws = wb.create_sheet("Money Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 18, 16, 16, 16, 2])
    luxe_header(ws, "G", "\U0001f3af  MONEY GOALS",
                "What you're building together — and whether it's actually happening.")
    table_headers(ws, 4, ["Goal", "Target", "Saved", "By when", "Status", "Progress"], start_col=2)
    start = L0
    for i, (g, tgt, saved, when, status) in enumerate(GOALS):
        r = start + i
        ws.cell(row=r, column=2, value=g).style = "td_left"
        ct = ws.cell(row=r, column=3, value=tgt); ct.style = "input"; ct.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=when).style = "td"
        ws.cell(row=r, column=6, value=status).style = "input"
        cp = ws.cell(row=r, column=7, value=f"=IFERROR(D{r}/C{r},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(GOALS) - 1
    nrange(wb, "GoalTarget", "Money Goals", "C", start, end)
    nrange(wb, "GoalSaved", "Money Goals", "D", start, end)
    nrange(wb, "GoalStatusCol", "Money Goals", "F", start, end)
    nrange(wb, "GoalProgress", "Money Goals", "G", start, end)
    add_dv(ws, f"F{start}:F{end}", "GoalStatusList")
    ws.conditional_formatting.add(f"G{start}:G{end}", DataBarRule(start_type="num", start_value=0,
                                                                 end_type="num", end_value=1, color=HIGHLIGHT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL GOALS").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(GoalTarget)"); ct.style = "td"
    ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value="=SUM(GoalSaved)"); cs.style = "td"
    cs.font = Font(bold=True, size=12, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cell_name(wb, "GoalsSaved", "Money Goals", f"$D${tot}")
    ws.cell(row=tot + 2, column=2, value="Goals on track or better").style = "field_label"
    c1 = ws.cell(row=tot + 2, column=4, value='=COUNTIF(GoalStatusCol,"On track")+COUNTIF(GoalStatusCol,"Ahead")+COUNTIF(GoalStatusCol,"Done")')
    c1.style = "field_value"; c1.number_format = "0"; c1.fill = fill(MINT_BG)
    cell_name(wb, "GoalsOnTrack", "Money Goals", f"$D${tot+2}")
    ws.cell(row=tot + 3, column=2, value="Goals in total").style = "field_label"
    c2 = ws.cell(row=tot + 3, column=4, value="=COUNTA(GoalTarget)"); c2.style = "field_value"; c2.number_format = "0"
    cell_name(wb, "GoalsTotal", "Money Goals", f"$D${tot+3}")
    ws.freeze_panes = "A5"


def build_savings(wb):
    ws, start, end = build_log(
        wb, "Savings", "\U0001f3e6", "SAVINGS & SINKING FUNDS", "Small amounts, every month, into named pots.",
        ["Fund", "Monthly", "Balance", "Yearly"], [(f, m, b) for (f, m, b) in SAVINGS],
        [2, 32, 18, 18, 18, 2], text_left={2}, money={3, 4, 5}, start_col=2)
    for r in range(start, start + len(SAVINGS)):
        ws.cell(row=r, column=5, value=f"=C{r}*12").number_format = '"$"#,##0'
    nrange(wb, "SaveMonthly", "Savings", "C", start, end)
    nrange(wb, "SaveBalance", "Savings", "D", start, end)
    tr = end + 1
    ws.cell(row=tr, column=2, value="= SAVED EACH MONTH").style = "th"
    c1 = ws.cell(row=tr, column=3, value="=SUM(SaveMonthly)"); c1.style = "td"
    c1.font = Font(bold=True, size=14, color=PRIMARY); c1.fill = fill(MINT_BG); c1.number_format = '"$"#,##0'
    cell_name(wb, "SavedMonth", "Savings", f"$C${tr}")
    c2 = ws.cell(row=tr, column=4, value="=SUM(SaveBalance)"); c2.style = "td"
    c2.font = Font(bold=True, size=12, color=PRIMARY); c2.fill = fill(SURFACE); c2.number_format = '"$"#,##0'
    ws.cell(row=tr + 2, column=2, value="Your goal").style = "field_label"
    c3 = ws.cell(row=tr + 2, column=3, value="=SavingsGoal"); c3.style = "field_value"; c3.number_format = '"$"#,##0'


def build_dates(wb):
    ws, start, end = build_log(
        wb, "Date Nights", "✦", "DATE NIGHTS", "Because it stops happening by accident and starts needing a plan.",
        ["Date", "What you did", "Who planned it", "Cost", "How it was"], DATES,
        [2, 12, 38, 18, 14, 20, 2], text_left={3}, money={5}, start_col=2)
    nrange(wb, "DatePlanner", "Date Nights", "D", start, end)
    nrange(wb, "DateCost", "Date Nights", "E", start, end)
    tr = end + 1
    ws.cell(row=tr, column=2, value="THIS MONTH").style = "th"
    ws.cell(row=tr, column=3).style = "td"; ws.cell(row=tr, column=3).fill = fill(SURFACE)
    c0 = ws.cell(row=tr, column=4, value="=COUNTA(DatePlanner)"); c0.style = "td"
    c0.font = Font(bold=True, size=13, color=PRIMARY); c0.fill = fill(MINT_BG); c0.number_format = '0" nights"'
    cell_name(wb, "DateNights", "Date Nights", f"$D${tr}")
    c1 = ws.cell(row=tr, column=5, value="=SUM(DateCost)"); c1.style = "td"
    c1.font = Font(bold=True, color=PRIMARY); c1.fill = fill(SURFACE); c1.number_format = '"$"#,##0'
    ws.cell(row=tr + 2, column=2, value="Look at the 'who planned it' column. If it's always the same name, that's the labour tab again.").style = "section_gold"


def build_checkin(wb):
    ws, start, end = build_log(
        wb, "Weekly Check-In", "\U0001f5e3", "WEEKLY CHECK-IN", "Fifteen minutes, same time every week. Money, diary, and how are you actually.",
        ["Week", "Done?", "What came up"], CHECKINS_LOG,
        [2, 16, 14, 66, 2], text_left={4}, validations=[("C", "YesNoList")], start_col=2)
    nrange(wb, "CheckinDone", "Weekly Check-In", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"Yes"'],
                                                                fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    tr = end + 1
    ws.cell(row=tr, column=2, value="= CHECK-INS DONE").style = "th"
    c1 = ws.cell(row=tr, column=3, value='=COUNTIF(CheckinDone,"Yes")'); c1.style = "td"
    c1.font = Font(bold=True, size=14, color=PRIMARY); c1.fill = fill(MINT_BG); c1.number_format = "0"
    cell_name(wb, "CheckinsDone", "Weekly Check-In", f"$C${tr}")
    r = tr + 2
    ws.cell(row=r, column=2, value="THE FOUR QUESTIONS").style = "section_gold"
    for i, q in enumerate(["1.  Anything on your mind about money this week?",
                           "2.  What's in the diary that I should know about?",
                           "3.  Is anything feeling uneven right now?",
                           "4.  What do you need from me this week?"]):
        ws.cell(row=r + 1 + i, column=2, value=q).style = "body"
    ws.cell(row=r + 6, column=2, value="Fifteen minutes a week prevents most of the arguments. It really is that boring, and it really does work.").style = "section_gold"


def build_bigtalks(wb):
    ws, start, end = build_log(
        wb, "Big Conversations", "\U0001f4ac", "THE BIG CONVERSATIONS", "The ones couples avoid until they can't. Tick them off, one evening at a time.",
        ["Topic", "Had it?", "Where we landed"], BIG_TALKS,
        [2, 46, 14, 56, 2], text_left={2, 4}, validations=[("C", "YesNoList")], start_col=2)
    nrange(wb, "TalkDone", "Big Conversations", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"No"'],
                                                                fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))
    tr = end + 1
    ws.cell(row=tr, column=2, value="= HAD").style = "th"
    c1 = ws.cell(row=tr, column=3, value='=COUNTIF(TalkDone,"Yes")'); c1.style = "td"
    c1.font = Font(bold=True, size=13, color=PRIMARY); c1.fill = fill(MINT_BG); c1.number_format = "0"
    cell_name(wb, "TalksHad", "Big Conversations", f"$C${tr}")
    ws.cell(row=tr + 1, column=2, value="Still to have").style = "field_label"
    c2 = ws.cell(row=tr + 1, column=3, value='=COUNTIF(TalkDone,"No")'); c2.style = "field_value"
    c2.number_format = "0"; c2.fill = fill(WARN_BG)
    ws.cell(row=tr + 3, column=2, value="The unticked ones are the point of this page. Book one evening. Not all of them at once.").style = "section_gold"


def build_admin(wb):
    ws, start, end = build_log(
        wb, "Household Admin", "\U0001f4cb", "HOUSEHOLD ADMIN", "Who actually does what, and how often — written down so it stops being assumed.",
        ["Item", "Whose job", "How often", "Notes"], ADMIN,
        [2, 34, 20, 18, 44, 2], text_left={2, 5}, validations=[("D", "FreqList")], start_col=2)
    ws.cell(row=end + 2, column=2, value="Anything with nobody's name on it is being done by whoever notices first. That is not nothing.").style = "section_gold"


def build_individual(wb):
    ws = wb.create_sheet("Individual Money"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 40, 20, 20, 2])
    luxe_header(ws, "D", "\U0001f45b  INDIVIDUAL MONEY",
                "What's shared is shared. What's yours stays yours — and you don't have to justify it.")
    table_headers(ws, 4, ["", "A", "B"], start_col=2)
    ws.cell(row=4, column=3, value="=PartnerA").style = "th"
    ws.cell(row=4, column=4, value="=PartnerB").style = "th"
    start = L0
    for i, (item, a, b) in enumerate(INDIVIDUAL):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=a); ca.style = "input"; ca.number_format = '"$"#,##0.00'
        cb = ws.cell(row=r, column=4, value=b); cb.style = "input"; cb.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INDIVIDUAL) - 1
    tot = end + 2
    ws.cell(row=tot, column=2, value="GUILT-FREE SPENDING — THE RULE").style = "section_gold"
    ws.cell(row=tot + 1, column=2, value="Under this amount, no questions asked").style = "field_label"
    cl = ws.cell(row=tot + 1, column=3, value=150); cl.style = "input"; cl.number_format = '"$"#,##0'
    cell_name(wb, "NoQuestionsLimit", "Individual Money", f"$C${tot+1}")
    ws.cell(row=tot + 3, column=2, value="Every couple who has this number written down argues about money less. It doesn't matter what the number is.").style = "section_gold"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 42, 18, 2])
    luxe_header(ws, "C", "\U0001f4c8  MONTHLY SUMMARY",
                "The month in one place — money, hours, and each other.")
    ws.cell(row=5, column=2, value="MONEY").style = "section_gold"
    rows = [
        ("Shared bills", "=SharedBills", '"$"#,##0'),
        ("A's fair share", "=FairA", '"$"#,##0.00'),
        ("B's fair share", "=FairB", '"$"#,##0.00'),
        ("How close to fair you are", "=Fairness", "0.0%"),
        ("Saved together", "=SavedMonth", '"$"#,##0'),
        ("Total in your goals", "=GoalsSaved", '"$"#,##0'),
    ]
    for i, (lab, fml, fmt) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
    ws.cell(row=13, column=2, value="HOURS & EACH OTHER").style = "section_gold"
    rows2 = [
        ("Chore split", "=ChoreRatio", '0.00"\\u00d7"'),
        ("Extra hours a year", "=ExtraHoursYear", '#,##0" hrs"'),
        ("Date nights", "=DateNights", "0"),
        ("Check-ins done", "=CheckinsDone", "0"),
        ("Big conversations had", "=TalksHad", "0"),
    ]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 14 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
    ws.cell(row=20, column=2, value="SAVED, MONTH BY MONTH").style = "section_gold"
    table_headers(ws, 21, ["Month", "Saved"], start_col=2)
    ts = 22
    for i, (m, sv) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=sv); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 4):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "SaveTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Saved by Month", ts, te, 3, 2), "E5")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ◎  RELATIONSHIP & COUPLES COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Money split fairly, hours counted honestly — the two of you, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("SHARED BILLS", "=SharedBills", "money"),
        ("50/50 LEAVES A", "=LeftHalfA", "money"),
        ("50/50 LEAVES B", "=LeftHalfB", "money"),
        ("THE 50/50 GAP", "=HalfGap", "money"),
        ("A'S FAIR SHARE", "=FairA", "money"),
        ("B'S FAIR SHARE", "=FairB", "money"),
    ]
    row2 = [
        ("YOU EACH KEEP", "=KeepPctA", "pct1"),
        ("HOW CLOSE TO FAIR", "=Fairness", "pct1"),
        ("CHORE SPLIT", "=ChoreRatio", "dec"),
        ("EXTRA HOURS A YEAR", "=ExtraHoursYear", "num"),
        ("SAVED THIS MONTH", "=SavedMonth", "money"),
        ("TOGETHER SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "HOW THE TWO OF YOU ARE DOING", "section_gold")
    merge_set(ws, "H11:M11", "SAVED BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Bills split fairly", "=IFERROR(MIN(Fairness/FairnessGoal,1),0)"),
        ("Weekly check-ins happening", "=IFERROR(MIN(CheckinsDone/CheckinGoal,1),0)"),
        ("Saving together", "=IFERROR(MIN(SavedMonth/SavingsGoal,1),0)"),
        ("Goals on track", "=IFERROR(MIN(GoalsOnTrack/GoalsTotal,1),0)"),
        ("Date nights happening", "=IFERROR(MIN(DateNights/DateGoal,1),0)"),
        ("Housework shared evenly", "=IFERROR(MIN(ChoreGoal/ChoreRatio,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Good",IF(C{r}>=0.6,"OK","Talk about it"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Saved by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=22, max_row=21 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=22, max_row=21 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Relationship & Couples Command Center™ — 50/50 is not the same as fair.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_bills(wb); build_fairshare(wb)
    build_labour(wb); build_goals(wb); build_savings(wb); build_dates(wb)
    build_checkin(wb); build_bigtalks(wb); build_admin(wb); build_individual(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Fair Share", "Shared Bills", "Invisible Labour", "Money Goals",
             "Savings", "Date Nights", "Weekly Check-In", "Big Conversations", "Household Admin",
             "Individual Money", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Couples_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
