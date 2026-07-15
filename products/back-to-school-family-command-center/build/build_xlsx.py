"""Build Back-to-School Command Center™ — The Large-Family Back-to-School System.

17 planning tabs (+ Settings) · a premium family operating system in Google Sheets
& Excel. Family dashboard, a profile for every child, school & teacher contacts,
the 2026–2027 calendar, events & deadlines, supply shopping, clothing & uniform
inventory, budget vs actual, fees & payments, extracurriculars, lunch & grocery
planning, homework & reading, parent-teacher comms, absences, grades and important
documents — one dashboard. Built by a mom of six for families of 1–8 kids.

Run: python3 build_xlsx.py   ->  ../Back_to_School_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

GRADES = ["PreK", "K", "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12"]
EVENT_TYPES = ["Deadline", "Event", "Field Trip", "Meeting", "Holiday", "Early Release", "Picture Day", "Other"]
SUPPLY_CATS = ["Supplies", "Backpack", "Binders", "Devices", "Art", "PE / Sports", "Clothing", "Other"]
STATUS = ["To Do", "In Progress", "Done"]
PAID = ["Paid", "Partial", "Unpaid"]
READY = ["Ready", "Need", "Ordered"]
SUBJECTS = ["Math", "Reading", "Science", "Writing", "History", "Spanish", "Other"]
YESNO = ["Yes", "No"]

# ---- sample family (mom of six) — the data that drives every KPI ----
CHILDREN = [
    ("Mateo Rivera", "9", "Lincoln High", 14, "Navy blazer + khaki"),
    ("Sofia Rivera", "7", "Jefferson Middle", 12, "No uniform"),
    ("Liam Rivera", "5", "Oakwood Elementary", 10, "Polo + navy"),
    ("Ava Rivera", "3", "Oakwood Elementary", 8, "Polo + navy"),
    ("Noah Rivera", "1", "Oakwood Elementary", 6, "Polo + navy"),
    ("Emma Rivera", "PreK", "Little Sprouts", 4, "Play clothes"),
]
# supplies: (item, child, category, qty, cost, bought)
SUPPLIES = [
    ("Graphing calculator", "Mateo", "Devices", 1, 95, True), ("Binders (5)", "Mateo", "Binders", 5, 22, True),
    ("Composition books", "Mateo", "Supplies", 6, 9, True), ("Backpack", "Mateo", "Backpack", 1, 45, True),
    ("Loose-leaf paper", "Sofia", "Supplies", 4, 8, True), ("Colored pencils", "Sofia", "Art", 1, 6, True),
    ("Binders (4)", "Sofia", "Binders", 4, 18, True), ("Backpack", "Sofia", "Backpack", 1, 40, False),
    ("Pencils (24)", "Liam", "Supplies", 2, 7, True), ("Glue sticks", "Liam", "Supplies", 8, 5, True),
    ("Crayons", "Liam", "Art", 2, 4, True), ("Backpack", "Liam", "Backpack", 1, 30, True),
    ("Markers", "Ava", "Art", 1, 6, True), ("Folders", "Ava", "Supplies", 6, 6, True),
    ("Scissors", "Ava", "Supplies", 1, 4, True), ("Backpack", "Ava", "Backpack", 1, 28, False),
    ("Pencils (24)", "Noah", "Supplies", 1, 4, True), ("Watercolors", "Noah", "Art", 1, 5, True),
    ("Rest mat", "Noah", "Other", 1, 15, True), ("Backpack", "Noah", "Backpack", 1, 24, True),
    ("Nap mat", "Emma", "Other", 1, 18, True), ("Crayons (jumbo)", "Emma", "Art", 2, 4, True),
    ("Change of clothes", "Emma", "Clothing", 1, 12, False), ("Backpack (mini)", "Emma", "Backpack", 1, 20, True),
    ("Family label pack", "Family", "Other", 1, 14, True), ("Headphones x3", "Family", "Devices", 3, 15, False),
    ("Hand sanitizer", "Family", "Supplies", 6, 3, True),
]
# clothing / uniforms: (child, item, size, status)
CLOTHING = [
    ("Mateo", "Blazer", "16", "Ready"), ("Mateo", "Khaki pants x3", "16", "Ready"), ("Mateo", "Dress shoes", "8", "Need"),
    ("Sofia", "Fall wardrobe", "12", "Ready"), ("Sofia", "Sneakers", "6", "Ready"),
    ("Liam", "Polos x5", "10", "Ready"), ("Liam", "Navy pants x3", "10", "Need"), ("Liam", "Sneakers", "3", "Ready"),
    ("Ava", "Polos x5", "8", "Ready"), ("Ava", "Navy skirts x2", "8", "Ready"), ("Ava", "Sneakers", "1", "Need"),
    ("Noah", "Polos x5", "6", "Ready"), ("Noah", "Navy pants x3", "6", "Ready"), ("Noah", "Sneakers", "12", "Ready"),
    ("Emma", "Play outfits", "5", "Ready"), ("Emma", "Rain boots", "9", "Ordered"),
]
# fees: (child, fee, amount, paid)
FEES = [
    ("Mateo", "Registration", 75, 75), ("Mateo", "Athletics", 120, 120), ("Sofia", "Registration", 75, 75),
    ("Sofia", "Band", 90, 45), ("Liam", "Registration", 50, 50), ("Ava", "Registration", 50, 50),
    ("Noah", "Registration", 50, 0), ("Emma", "Preschool deposit", 150, 150), ("Family", "PTA", 40, 40),
    ("Family", "Yearbooks x3", 90, 90),
]
# documents / forms: (name, done)
DOCS = [
    ("Enrollment forms (all)", True), ("Immunization records", True), ("Emergency contacts", True),
    ("Free/reduced lunch app", True), ("Media consent forms", True), ("Transportation forms", True),
    ("Physical / sports forms", False), ("Medication authorizations", True), ("Photo order forms", False),
    ("Volunteer background check", True),
]
# budget: (category, planned, actual)
BUDGET = [
    ("School Supplies", 320, 300), ("Clothing", 480, 415), ("Shoes", 300, 255), ("Uniforms", 250, 210),
    ("Backpacks & Lunchboxes", 180, 165), ("Devices & Tech", 280, 0), ("School Fees", 380, 300),
    ("Extracurriculars", 200, 120), ("Haircuts & Misc", 110, 85),
]
FIRST_DAY = dt.date(2026, 8, 25)

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your family details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "FAMILY INPUTS", "section")
    controls = [
        ("Family Name", "The Rivera Family", None, "FamilyName"),
        ("School Year", "2026–2027", None, "SchoolYear"),
        ("First Day", FIRST_DAY, "mmm d, yyyy", "FirstDay"),
        ("Number of Children", 6, "0", "Children"),
        ("Number of Schools", 4, "0", "Schools"),
        ("Back-to-School Budget", "=BudgetPlanTotal", '"$"#,##0', "BudgetTotal"),
        ("Parent 1", "Elena Rivera", None, "Parent1"),
        ("Parent 2", "Marco Rivera", None, "Parent2"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Grade", GRADES, "GradeList"), ("F", "Event Type", EVENT_TYPES, "EventTypeList"),
             ("G", "Supply Category", SUPPLY_CATS, "SupCatList"), ("H", "Status", STATUS, "StatusList"),
             ("I", "Payment", PAID, "PaidList"), ("J", "Ready?", READY, "ReadyList"),
             ("K", "Subject", SUBJECTS, "SubjectList"), ("L", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🎒  BACK-TO-SCHOOL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The large-family back-to-school system — for 1 to 8 kids.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "FROM ONE MOM OF SIX TO YOU", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("I built this because no planner could keep up with SIX kids, four schools and one budget. This is "
                      "the whole back-to-school season — every child, contact, form, fee, supply, uniform, lunch and "
                      "deadline — in ONE calm command center. A profile for every child, a live Family Dashboard, a "
                      "budget that actually adds up, and matching printable pages for the fridge. Works in Google "
                      "Sheets and Excel, and prints beautifully. Scales from 1 child to a full house.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your family name, school year, first day & budget.",
             "2.  Fill a Child Profile for each kid — grade, school, sizes, allergies, teacher.",
             "3.  Add School & Teacher Contacts and the 2026–2027 Calendar & Events.",
             "4.  Work the Supply, Clothing & Fees trackers — the Budget updates live.",
             "5.  Use the daily tabs — lunch, homework, comms, absences & grades — all year.",
             "6.  Watch the Family Dashboard track readiness, budget & what's still due."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for the Rivera family (6 kids across 4 schools) is included so you can see how it all "
               "connects — just type over it with your own. Supplies, fees, uniforms, forms and budget roll up into a "
               "live Readiness Score. Twelve matching printable PDF pages (child info, backpack & supply checklists, "
               "weekly schedule, lunchbox planner, memory pages & more) are included to print and post. Everything "
               "works on desktop and phone.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, a calmer September — you've got this.", "section_gold")


# ===========================================================================
# 3 — Child Profiles
# ===========================================================================
def build_profiles(wb):
    ws = wb.create_sheet("Child Profiles"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 18, 4, 18, 18, 2])
    luxe_header(ws, "G", "🧒  CHILD PROFILES",
                "A profile for every child — grade, school, sizes, allergies & who to call.")
    row = 5
    for name, grade, school, age, uniform in CHILDREN:
        merge_set(ws, f"B{row}:F{row}", f"{name}", "section_gold"); ws.row_dimensions[row].height = 22
        row += 1
        fields = [("Grade", grade), ("School", school), ("Age", age), ("Teacher", "TBD"),
                  ("Shirt / Dress", "—"), ("Shoe size", "—"), ("Allergies", "None"), ("Bus / Pickup", "Pickup"),
                  ("Uniform", uniform), ("Emergency #", "(555) 010-0000")]
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 22; i += 2; row += 1
        row += 1
    nrange(wb, "ChildName", "Child Profiles", "B", 5, row)
    ws.freeze_panes = "A5"


# ===========================================================================
# 4 — School & Teacher Contacts
# ===========================================================================
def build_contacts(wb):
    rows = [
        ("Lincoln High", "Mateo", "Front Office", "(555) 201-4000", "office@lincoln.edu", "Attendance line ext. 2"),
        ("Lincoln High", "Mateo", "Ms. Alvarez (advisor)", "(555) 201-4021", "alvarez@lincoln.edu", "Homeroom + counselor"),
        ("Jefferson Middle", "Sofia", "Front Office", "(555) 201-3000", "office@jeffms.edu", "Bell 8:10"),
        ("Jefferson Middle", "Sofia", "Mr. Doyle (band)", "(555) 201-3044", "doyle@jeffms.edu", "Band fees due Sep 5"),
        ("Oakwood Elementary", "Liam/Ava/Noah", "Front Office", "(555) 201-2000", "office@oakwood.edu", "Car line 3:05"),
        ("Oakwood Elementary", "Liam", "Mrs. Kim (Rm 12)", "(555) 201-2012", "kim@oakwood.edu", "Reading log Fri"),
        ("Oakwood Elementary", "Ava", "Ms. Patel (Rm 7)", "(555) 201-2007", "patel@oakwood.edu", "Snack helper list"),
        ("Oakwood Elementary", "Noah", "Mr. Ross (Rm 3)", "(555) 201-2003", "ross@oakwood.edu", "Show & tell Wed"),
        ("Little Sprouts", "Emma", "Ms. Dana", "(555) 201-1000", "hello@littlesprouts.com", "Half day T/Th"),
        ("District", "All", "Transportation", "(555) 201-9000", "bus@district.org", "Route Q — 7:42 am"),
        ("District", "All", "Nurse (Oakwood)", "(555) 201-2050", "nurse@oakwood.edu", "Med forms on file"),
    ]
    build_log(wb, "Contacts", "☎", "SCHOOL & TEACHER CONTACTS",
              "Every school, teacher & office number — no more digging through emails.",
              ["School", "Child", "Contact", "Phone", "Email", "Notes"],
              rows, [20, 16, 22, 16, 22, 22], text_left={1, 3, 5, 6}, reserved=30)


# ===========================================================================
# 5 — Calendar 2026–2027
# ===========================================================================
def build_calendar(wb):
    rows = [
        (dt.date(2026, 8, 25), "First Day of School", "Event", "All", "Bus 7:42 · pics ready"),
        (dt.date(2026, 9, 4), "Band fees due", "Deadline", "Sofia", "$90"),
        (dt.date(2026, 9, 7), "Labor Day — no school", "Holiday", "All", ""),
        (dt.date(2026, 9, 11), "Picture Day", "Picture Day", "All", "Order forms in backpacks"),
        (dt.date(2026, 9, 18), "Back-to-School Night", "Meeting", "All", "6–8 pm, all schools"),
        (dt.date(2026, 10, 9), "Zoo Field Trip", "Field Trip", "Ava", "$12 + permission slip"),
        (dt.date(2026, 10, 16), "Early Release", "Early Release", "All", "Dismiss 12:30"),
        (dt.date(2026, 10, 30), "Fall Festival", "Event", "All", "Volunteer 5–7"),
        (dt.date(2026, 11, 6), "Report Cards", "Deadline", "All", "Q1 grades"),
        (dt.date(2026, 11, 12), "Parent-Teacher Conferences", "Meeting", "All", "Sign up online"),
        (dt.date(2026, 11, 25), "Thanksgiving Break", "Holiday", "All", "Nov 25–27"),
        (dt.date(2026, 12, 19), "Winter Break begins", "Holiday", "All", "Return Jan 5"),
    ]
    build_log(wb, "Calendar", "🗓", "2026–2027 SCHOOL CALENDAR",
              "The whole year at a glance — holidays, deadlines, trips & meetings.",
              ["Date", "What", "Type", "Who", "Notes"],
              rows, [14, 30, 14, 14, 26], text_left={2, 5}, dates={1}, reserved=60,
              validations=[("C", "EventTypeList")])


# ===========================================================================
# 6 — Events & Deadlines
# ===========================================================================
def build_events(wb):
    rows = [
        (dplus(2), "Return media consent forms", "Deadline", "All", "To Do"),
        (dplus(4), "Buy Sofia's backpack", "Deadline", "Sofia", "To Do"),
        (dplus(6), "Band fee $90", "Deadline", "Sofia", "In Progress"),
        (dplus(9), "Sports physical — Mateo", "Deadline", "Mateo", "To Do"),
        (dplus(11), "Order Ava's sneakers", "Deadline", "Ava", "To Do"),
        (dplus(14), "Picture Day forms", "Deadline", "All", "To Do"),
        (dplus(17), "Zoo trip permission + $12", "Field Trip", "Ava", "To Do"),
        (dplus(21), "Back-to-School Night", "Meeting", "All", "To Do"),
        (dplus(25), "Headphones x3", "Deadline", "Family", "To Do"),
        (dminus(3), "Immunization records", "Deadline", "All", "Done"),
        (dminus(6), "Enrollment forms", "Deadline", "All", "Done"),
        (dminus(1), "Noah registration fee", "Deadline", "Noah", "In Progress"),
    ]
    ws, start, end = build_log(
        wb, "Events", "⏰", "EVENTS & DEADLINES",
        "Everything with a due date — flagged before it sneaks up on you.",
        ["Due", "Task", "Type", "Child", "Status"],
        rows, [14, 32, 14, 14, 14], text_left={2}, dates={1}, reserved=50,
        validations=[("C", "EventTypeList"), ("E", "StatusList")])
    nrange(wb, "EventDate", "Events", "A", start, end)
    nrange(wb, "EventStatus", "Events", "E", start, end)
    cmap = {"Done": MINT_BG, "In Progress": WARN_BG, "To Do": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 7 — Supply Shopping Tracker
# ===========================================================================
def build_supplies(wb):
    rows = [(item, child, cat, qty, cost, "Yes" if bought else "No") for (item, child, cat, qty, cost, bought) in SUPPLIES]
    ws, start, end = build_log(
        wb, "Supplies", "✏", "SCHOOL SUPPLY SHOPPING TRACKER",
        "Every item, for every child — check it off and the budget & readiness update.",
        ["Item", "Child", "Category", "Qty", "Est. Cost", "Bought?"],
        rows, [24, 14, 14, 8, 12, 12], text_left={1}, ints={4}, money={5}, reserved=60,
        validations=[("C", "SupCatList"), ("F", "YesNoList")])
    nrange(wb, "SupItem", "Supplies", "A", start, end)
    nrange(wb, "SupBought", "Supplies", "F", start, end)
    nrange(wb, "SupCost", "Supplies", "E", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 8 — Clothing & Uniform Inventory
# ===========================================================================
def build_clothing(wb):
    rows = [(child, item, size, status) for (child, item, size, status) in CLOTHING]
    ws, start, end = build_log(
        wb, "Clothing", "👕", "CLOTHING, SHOES & UNIFORM INVENTORY",
        "Who has what, in which size — and what still needs ordering before day one.",
        ["Child", "Item", "Size", "Status"],
        rows, [16, 26, 12, 14], text_left={2}, reserved=40,
        validations=[("D", "ReadyList")])
    nrange(wb, "ClothItem", "Clothing", "B", start, end)
    nrange(wb, "ClothReady", "Clothing", "D", start, end)
    cmap = {"Ready": MINT_BG, "Ordered": WARN_BG, "Need": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 9 — Budget vs Actual
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "💰  BUDGET VS ACTUAL",
                "Set a number, track every dollar — see exactly where back-to-school money goes.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining"], start_col=2)
    start = L0
    for i, (cat, plan, actual) in enumerate(BUDGET):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=3, value=plan); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=actual); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BUDGET) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "BudgetPlan", "Budget", "C", start, end)
    nrange(wb, "BudgetActual", "Budget", "D", start, end)
    cell_name(wb, "BudgetPlanTotal", "Budget", f"$C${tot}")
    cell_name(wb, "BudgetSpent", "Budget", f"$D${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=480, color=PRIMARY, showValue=True))
    merge_set(ws, "B15:E15", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Total budget", "=BudgetPlanTotal", '"$"#,##0'), ("Spent so far", "=BudgetSpent", '"$"#,##0'),
             ("Remaining", "=BudgetPlanTotal-BudgetSpent", '"$"#,##0'),
             ("% of budget used", "=IFERROR(BudgetSpent/BudgetPlanTotal,0)", "0%"),
             ("Per child (avg)", "=IFERROR(BudgetSpent/Children,0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 16 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Remaining", "Per child (avg)"):
            c.fill = fill(MINT_BG)


# ===========================================================================
# 10 — Fees & Payments
# ===========================================================================
def build_fees(wb):
    ws = wb.create_sheet("Fees"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 22, 12, 12, 12, 2])
    luxe_header(ws, "F", "🧾  SCHOOL FEES & PAYMENT TRACKER",
                "Every registration, activity & fee — paid, partial or still owed.")
    table_headers(ws, 4, ["Child", "Fee", "Amount", "Paid", "Owed"], start_col=2)
    start = L0
    for i, (child, fee, amt, paid) in enumerate(FEES):
        r = start + i
        ws.cell(row=r, column=2, value=child).style = "td_left"
        ws.cell(row=r, column=3, value=fee).style = "td_left"
        ca = ws.cell(row=r, column=4, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=paid); cp.style = "input"; cp.number_format = '"$"#,##0'
        co = ws.cell(row=r, column=6, value=f"=D{r}-E{r}"); co.style = "td"; co.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FEES) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"; ws.cell(row=tot, column=3).style = "th"
    for col in (4, 5, 6):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "FeeAmt", "Fees", "D", start, end)
    nrange(wb, "FeePaid", "Fees", "E", start, end)
    cell_name(wb, "FeeAmtTotal", "Fees", f"$D${tot}")
    cell_name(wb, "FeePaidTotal", "Fees", f"$E${tot}")
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=fill(WARN_BG)))


# ===========================================================================
# 11 — Extracurriculars
# ===========================================================================
def build_extra(wb):
    rows = [
        ("Mateo", "JV Soccer", "Mon/Wed/Fri", "3:30–5:30", "Coach Ruiz", "$120 · cleats size 8"),
        ("Sofia", "Concert Band", "Tue/Thu", "3:15–4:30", "Mr. Doyle", "$90 · clarinet reeds"),
        ("Sofia", "Art Club", "Wed", "3:15–4:15", "Ms. Lane", "Free"),
        ("Liam", "Chess Club", "Thu", "3:15–4:15", "Mrs. Kim", "Free · rides needed"),
        ("Ava", "Ballet", "Sat", "10:00–11:00", "Miss Tara", "$65/mo · shoes 12.5"),
        ("Noah", "Soccer Tots", "Sat", "9:00–9:45", "Coach Ben", "$50 · shin guards"),
        ("Emma", "Story Time", "Fri", "10:30–11:00", "Library", "Free"),
    ]
    build_log(wb, "Extracurriculars", "⚽", "EXTRACURRICULAR SCHEDULE",
              "Every practice, class & club — the master carpool & conflict view.",
              ["Child", "Activity", "Days", "Time", "Coach / Lead", "Notes"],
              rows, [14, 20, 16, 14, 16, 24], text_left={5, 6}, reserved=30)


# ===========================================================================
# 12 — Lunch & Grocery Planner
# ===========================================================================
def build_lunch(wb):
    rows = [
        ("Monday", "Turkey wraps", "Apple slices", "Cheese stick", "Water", "Buy: wraps, turkey"),
        ("Tuesday", "PB&J (Mateo: SunButter)", "Grapes", "Pretzels", "Milk", "Emma: no nuts nearby"),
        ("Wednesday", "Pasta salad", "Carrots", "Yogurt", "Water", "Pizza day for K-2"),
        ("Thursday", "Ham & cheese", "Orange", "Crackers", "Water", "Buy: deli ham"),
        ("Friday", "Bagel & cream cheese", "Berries", "Granola bar", "Juice", "Hot lunch option"),
    ]
    ws, start, end = build_log(
        wb, "Lunch & Grocery", "🍎", "LUNCH ROTATION & GROCERY PLANNER",
        "A week of lunches everyone will actually eat — plus the grocery list it makes.",
        ["Day", "Main", "Fruit / Veg", "Snack", "Drink", "Grocery Notes"],
        rows, [12, 24, 16, 14, 12, 24], text_left={2, 6}, reserved=12)
    merge_set(ws, "A19:F19", "GROCERY LIST — TAP TO CHECK OFF", "section_gold"); ws.row_dimensions[19].height = 22
    groc = ["☐ Sandwich wraps & bread", "☐ Turkey, ham, SunButter", "☐ Apples, grapes, oranges, berries",
            "☐ Carrots & cucumbers", "☐ Cheese sticks & yogurt", "☐ Pretzels, crackers, granola bars",
            "☐ Milk & juice boxes", "☐ Pasta & pizza night items"]
    for i, g in enumerate(groc):
        r = 20 + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.merge_cells(f"A{r}:C{r}")


# ===========================================================================
# 13 — Homework & Reading Log
# ===========================================================================
def build_homework(wb):
    rows = [
        (dminus(1), "Liam", "Reading", "20 min — Magic Tree House", "Yes", "Signed log"),
        (dminus(1), "Ava", "Math", "Worksheet p.12", "Yes", ""),
        (dminus(1), "Noah", "Reading", "Sight words", "Yes", "Great job!"),
        (dminus(1), "Sofia", "Science", "Lab writeup", "No", "Due Fri"),
        (dminus(1), "Mateo", "History", "Ch. 3 notes", "Yes", ""),
        (dt.date.today(), "Liam", "Reading", "20 min", "No", ""),
        (dt.date.today(), "Ava", "Spanish", "Vocab", "No", ""),
    ]
    ws, start, end = build_log(
        wb, "Homework & Reading", "📚", "HOMEWORK & READING LOG",
        "Nightly reading minutes & homework, per child — no more forgotten worksheets.",
        ["Date", "Child", "Subject", "Assignment", "Done?", "Notes"],
        rows, [13, 14, 14, 26, 10, 20], text_left={4, 6}, dates={1}, reserved=50,
        validations=[("C", "SubjectList"), ("E", "YesNoList")])
    ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 14 — Parent-Teacher Comms Log
# ===========================================================================
def build_comms(wb):
    rows = [
        (dminus(4), "Mrs. Kim", "Liam", "Email", "Reading group placement", "Follow up in 2 wks"),
        (dminus(2), "Mr. Doyle", "Sofia", "Note home", "Band fee reminder", "Paid partial"),
        (dminus(1), "Ms. Patel", "Ava", "App message", "Snack helper sign-up", "Signed up 9/20"),
        (dt.date.today(), "Coach Ruiz", "Mateo", "Text", "Practice moved to 4pm", "Update calendar"),
        (dplus(3), "Ms. Alvarez", "Mateo", "Meeting", "Course schedule check", "Scheduled 3:15"),
    ]
    build_log(wb, "PT Comms", "✉", "PARENT-TEACHER COMMUNICATION LOG",
              "Every email, note & conversation — dated, so nothing falls through.",
              ["Date", "Teacher / Staff", "Child", "Method", "Topic", "Follow-up"],
              rows, [13, 18, 12, 13, 24, 22], text_left={5, 6}, dates={1}, reserved=40)


# ===========================================================================
# 15 — Absence & Late Tracker
# ===========================================================================
def build_absence(wb):
    rows = [
        (dminus(8), "Noah", "Absent", "Fever", "Yes", "Called office"),
        (dminus(5), "Ava", "Late", "Dentist", "Yes", "Note sent"),
        (dminus(2), "Mateo", "Late", "Traffic", "No", "5 min"),
    ]
    build_log(wb, "Absences", "🗒", "ABSENCE & LATE-DAY TRACKER",
              "A dated record for every school — handy when the office asks.",
              ["Date", "Child", "Type", "Reason", "Excused?", "Notes"],
              rows, [13, 14, 12, 18, 12, 22], text_left={4, 6}, dates={1}, reserved=40,
              validations=[("E", "YesNoList")])


# ===========================================================================
# 16 — Grades & Report Cards
# ===========================================================================
def build_grades(wb):
    ws = wb.create_sheet("Grades"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 12, 12, 12, 12, 22, 2])
    luxe_header(ws, "G", "🎓  GRADES & REPORT-CARD OVERVIEW",
                "A simple term-by-term snapshot per child — celebrate wins, catch dips early.")
    table_headers(ws, 4, ["Child", "Q1", "Q2", "Q3", "Q4", "Notes"], start_col=2)
    kids = ["Mateo", "Sofia", "Liam", "Ava", "Noah", "Emma"]
    seed = {"Mateo": ("A-", "", "", ""), "Sofia": ("B+", "", "", ""), "Liam": ("S", "", "", ""),
            "Ava": ("S+", "", "", ""), "Noah": ("S", "", "", ""), "Emma": ("😊", "", "", "")}
    start = L0
    for i, k in enumerate(kids):
        r = start + i
        ws.cell(row=r, column=2, value=k).style = "td_left"
        for j in range(4):
            ws.cell(row=r, column=3 + j, value=seed[k][j]).style = "input"
        ws.cell(row=r, column=7, value="—").style = "td_left"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"
    merge_set(ws, "B12:F12", "REPORT-CARD DATES", "section_gold"); ws.row_dimensions[12].height = 22
    dates = [("Q1 report cards", "Nov 6, 2026"), ("Q2 report cards", "Jan 29, 2027"),
             ("Q3 report cards", "Apr 9, 2027"), ("Q4 / final", "Jun 11, 2027")]
    for i, (lab, val) in enumerate(dates):
        r = 13 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        ws.cell(row=r, column=3, value=val).style = "field_value"


# ===========================================================================
# 17 — Important Documents
# ===========================================================================
def build_docs(wb):
    rows = [(name, "Yes" if done else "No", "—") for (name, done) in DOCS]
    ws, start, end = build_log(
        wb, "Documents", "🔗", "IMPORTANT DOCUMENTS & FORMS",
        "Every form & where it lives — paste the link and check it off when it's done.",
        ["Document / Form", "Done?", "Link / Location"],
        rows, [34, 12, 34], text_left={1, 3}, reserved=30,
        validations=[("B", "YesNoList")])
    nrange(wb, "DocName", "Documents", "A", start, end)
    nrange(wb, "DocDone", "Documents", "B", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 2 — Family Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Family Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🎒  BACK-TO-SCHOOL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Every child, form, fee & deadline — your whole back-to-school season, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("CHILDREN", "=Children", "num"),
        ("SCHOOLS", "=Schools", "num"),
        ("FIRST DAY", "=FirstDay", "date"),
        ("SUPPLIES BOUGHT", '=IFERROR(COUNTIF(SupBought,"Yes")/COUNTA(SupItem),0)', "pct"),
        ("BUDGET SPENT", "=BudgetSpent", "money"),
        ("BUDGET LEFT", "=BudgetPlanTotal-BudgetSpent", "money"),
    ]
    row2 = [
        ("FEES PAID", "=IFERROR(FeePaidTotal/FeeAmtTotal,0)", "pct"),
        ("FORMS DONE", '=IFERROR(COUNTIF(DocDone,"Yes")/COUNTA(DocName),0)', "pct"),
        ("UNIFORMS READY", '=IFERROR(COUNTIF(ClothReady,"Ready")/COUNTA(ClothItem),0)', "pct"),
        ("EVENTS (30D)", '=COUNTIFS(EventDate,">="&TODAY(),EventDate,"<="&TODAY()+30)', "num"),
        ("TO-DO OPEN", '=COUNTIF(EventStatus,"To Do")+COUNTIF(EventStatus,"In Progress")', "num"),
        ("READINESS", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    # readiness dimensions
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "BACK-TO-SCHOOL READINESS", "section_gold")
    merge_set(ws, "H11:M11", "BUDGET BY CATEGORY", "section_gold")
    table_headers(ws, 12, ["Area", "Ready", "Status"], start_col=2)
    dims = [
        ("Supplies bought", '=IFERROR(COUNTIF(SupBought,"Yes")/COUNTA(SupItem),0)'),
        ("Clothing & uniforms", '=IFERROR(COUNTIF(ClothReady,"Ready")/COUNTA(ClothItem),0)'),
        ("Fees paid", "=IFERROR(FeePaidTotal/FeeAmtTotal,0)"),
        ("Forms & documents", '=IFERROR(COUNTIF(DocDone,"Yes")/COUNTA(DocName),0)'),
        ("Overdue caught up", '=IFERROR(COUNTIFS(EventDate,"<"&TODAY(),EventStatus,"Done")/COUNTIF(EventDate,"<"&TODAY()),0)'),
        ("Budget on track", "=IFERROR(1-MAX(BudgetSpent-BudgetPlanTotal,0)/BudgetPlanTotal,0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"On track",IF(C{r}>=0.6,"Almost","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Family Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # charts
    db = DoughnutChart(); db.title = "Budget by Category"; db.height = 7.4; db.width = 8.4
    db.add_data(Reference(wb["Budget"], min_col=4, min_row=4, max_row=13), titles_from_data=True)
    db.set_categories(Reference(wb["Budget"], min_col=2, min_row=5, max_row=13)); db.dataLabels = no_labels()
    ws.add_chart(db, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:G21", "WHAT'S DUE NEXT — see the Events tab", "section")
    merge_set(ws, "B23:M23", "Back-to-School Command Center™ — from one mom of six. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profiles(wb); build_contacts(wb)
    build_calendar(wb); build_events(wb); build_supplies(wb); build_clothing(wb)
    build_budget(wb); build_fees(wb); build_extra(wb); build_lunch(wb)
    build_homework(wb); build_comms(wb); build_absence(wb); build_grades(wb)
    build_docs(wb); build_dashboard(wb)

    order = ["Start Here", "Family Dashboard", "Child Profiles", "Contacts", "Calendar", "Events",
             "Supplies", "Clothing", "Budget", "Fees", "Extracurriculars", "Lunch & Grocery",
             "Homework & Reading", "PT Comms", "Absences", "Grades", "Documents", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Back_to_School_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
