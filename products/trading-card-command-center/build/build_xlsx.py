"""Build Trading Card Collection Command Center™ — The Ultimate Trading Card
Collection, Value & Inventory System.

13 sheets + Welcome · a premium collection management system in Excel & Sheets.
Collection, purchases, sales, trades, grading, wishlist, duplicates, decks,
gallery & analytics — one elegant dashboard.

Run: python3 build_xlsx.py   ->  ../Trading_Card_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

GAMES = ["Pokémon", "Magic: The Gathering", "Yu-Gi-Oh!", "Lorcana", "One Piece", "Sports Cards", "Digimon", "Other"]
RARITIES = ["Common", "Uncommon", "Rare", "Rare Holo", "Ultra Rare", "Secret Rare", "Promo"]
LANGS = ["EN", "JP", "DE", "FR", "ES", "IT", "KR", "ZH"]
CONDITIONS = ["Gem Mint", "Mint", "Near Mint", "Lightly Played", "Moderately Played", "Heavily Played", "Damaged"]
LOCATIONS = ["Vault Case", "Binder A", "Binder B", "Display Shelf", "Bulk Box 1", "Bulk Box 2", "Deck Box", "Safe"]
GRADERS = ["PSA", "BGS", "CGC", "SGC", "TAG"]
MARKETS = ["eBay", "TCG Marketplace", "Card Show", "Local Game Store", "Facebook", "Mercari", "Whatnot"]
PRIORITIES = ["Grail", "High", "Medium", "Low"]
PLANS = ["Trade", "Sell", "Gift", "Keep"]
GRADE_STATUS = ["Returned", "Pending"]
YESNO = ["Yes", "No"]
CURRENCIES = ["USD", "EUR", "GBP", "CAD", "AUD", "JPY"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula; vc.font = Font(size=18, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pmoney": '+"$"#,##0;-"$"#,##0', "pct": "0%",
                        "ppct": "+0%;-0%", "days": "0", "dec": "0.0"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1))
        m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [18] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your collection details once — every dashboard follows. Works for any card game.")
    merge_set(ws, "B5:C5", "COLLECTION INPUTS", "section")
    controls = [
        ("Collection Name", "The Vaulted Collection", None, "CollName"),
        ("Collector", "You", None, "Collector"),
        ("Primary Game", "Pokémon", None, "PrimaryGame"),
        ("Currency", "USD", None, "HomeCurr"),
        ("Collection Value Goal", 5000, '"$"#,##0', "ValueGoal"),
        ("Growth Target (value ÷ cost)", 2.0, "0.0", "GrowthTarget"),
        ("Monthly Card Budget", 150, '"$"#,##0', "CardBudget"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Card Game", GAMES, "GameList"), ("F", "Rarity", RARITIES, "RarityList"),
             ("G", "Language", LANGS, "LangList"), ("H", "Condition", CONDITIONS, "CondList"),
             ("I", "Storage Location", LOCATIONS, "LocList"), ("J", "Grading Company", GRADERS, "GraderList"),
             ("K", "Marketplace", MARKETS, "MarketList"), ("L", "Currency", CURRENCIES, "CurrencyList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 17, "Priority", PRIORITIES, "PriorityList"),
             ("F", 17, "Duplicate Plan", PLANS, "PlanList"),
             ("G", 17, "Grading Status", GRADE_STATUS, "GradeStatusList"),
             ("H", 17, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 82, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🎴  TRADING CARD COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate trading card collection, value & inventory system — for every card game.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR COLLECTION IS AN ASSET — TREAT IT LIKE ONE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Serious collections deserve more than a shoebox and a guess. Trading Card Command Center™ "
                      "organizes every card, tracks what you paid and what it's worth, logs purchases, sales and "
                      "trades, manages grading submissions and your wishlist, and rolls everything into one elegant "
                      "dashboard — so you always know your collection's value, growth and health. Works for Pokémon, "
                      "Magic, Yu-Gi-Oh!, sports cards and any other game.")
    ws["B6"].style = "body"
    for r in (6, 7, 8, 9):
        ws.row_dimensions[r].height = 22
    merge_set(ws, "B11:B11", "START HERE", "section")
    steps = ["1.  Open Settings and set your collection name, game, currency & goals.",
             "2.  Add your cards to the Master Collection — value totals compute instantly.",
             "3.  Log purchases, sales & trades as they happen — profit/loss updates live.",
             "4.  Track grading submissions in the Grading Center (before vs after value).",
             "5.  Keep your Wishlist & Duplicates current — the dashboard scores your progress.",
             "6.  Watch the Collection Dashboard track value, growth & health automatically."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (a 146-card collection worth $2,840) is included so you can see how "
               "everything connects — just type over it with your own cards. Values, profit/loss, grading gains, "
               "wishlist progress and the Collection Health Score all update automatically. Estimated values are "
               "your own entries — update them from your favorite price source. Every sheet is print-friendly and "
               "works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized vault, zero mystery — happy collecting!", "section_gold")


# ===========================================================================
# 2 — Master Collection
# ===========================================================================
def build_collection(wb):
    ws = wb.create_sheet("Collection"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 17, 10, 8, 13, 15, 7, 13, 12, 10, 11, 12, 9, 20, 3, 13, 11, 3, 16, 11])
    luxe_header(ws, "T", "🎴  MASTER COLLECTION",
                "Every card, valued — quantity, condition, location, photo status & live profit/loss per card.")
    table_headers(ws, 4, ["Card Name", "Set", "No.", "Lang", "Rarity", "Condition", "Qty",
                          "Location", "Purchased", "Paid", "Value (ea)", "Total Value", "Photo?", "Notes"])
    rows = [
        ("Ancient Colossus", "Genesis", "004/120", "EN", "Secret Rare", "Gem Mint", 1, "Vault Case", 55, 210, 620, "BGS 9.5 slab"),
        ("Golden Phoenix (Alt Art)", "Obsidian Flames", "231/210", "EN", "Secret Rare", "Gem Mint", 1, "Vault Case", 80, 150, 480, "PSA 10 slab"),
        ("Emberfang Dragon (Holo)", "Obsidian Flames", "044/210", "EN", "Ultra Rare", "Near Mint", 1, "Binder A", 120, 120, 340, "At PSA now"),
        ("Twin Flames Duo", "Obsidian Flames", "199/210", "EN", "Ultra Rare", "Mint", 1, "Vault Case", 75, 88, 265, "PSA 9 slab"),
        ("Celestial Guardian", "Astral Dawn", "102/180", "JP", "Secret Rare", "Near Mint", 1, "Binder A", 90, 95, 210, "JP exclusive"),
        ("Royal Kraken", "Deep Currents", "145/140", "EN", "Secret Rare", "Mint", 1, "Vault Case", 100, 70, 150, "PSA 8 slab"),
        ("Shadow Assassin", "Midnight Veil", "077/160", "EN", "Ultra Rare", "Near Mint", 1, "Binder A", 70, 75, 145, ""),
        ("Thunderclap Roc", "Tempest Rising", "088/150", "EN", "Ultra Rare", "Near Mint", 1, "Binder A", 40, 55, 96, "From trade"),
        ("Tidecaller Leviathan", "Deep Currents", "031/140", "EN", "Ultra Rare", "Lightly Played", 1, "Binder A", 110, 60, 88, ""),
        ("Lunar Priestess", "Midnight Veil", "059/160", "EN", "Rare Holo", "Near Mint", 1, "Binder B", 35, 28, 42, ""),
        ("Storm Herald", "Tempest Rising", "041/150", "EN", "Rare Holo", "Near Mint", 2, "Binder B", 60, 30, 26, ""),
        ("Inferno Titan", "Obsidian Flames", "027/210", "EN", "Rare Holo", "Moderately Played", 2, "Binder B", 85, 22, 15, "Deck copies"),
        ("Dragon's Hoard", "Tempest Rising", "112/150", "EN", "Rare", "Near Mint", 2, "Binder B", 50, 14, 11, ""),
        ("Verdant Sprite", "Astral Dawn", "015/180", "EN", "Rare", "Near Mint", 4, "Binder B", 65, 24, 12, "3 for trade"),
        ("Frostbite Wolf", "Glacial Peak", "022/130", "EN", "Rare", "Near Mint", 6, "Binder B", 45, 20, 9, "4 to sell"),
        ("Arcane Scholar", "Astral Dawn", "066/180", "EN", "Uncommon", "Near Mint", 24, "Bulk Box 1", 30, 20, 3, ""),
        ("Mystic Fountain", "Deep Currents", "090/140", "EN", "Uncommon", "Near Mint", 30, "Bulk Box 1", 30, 18, 2, ""),
        ("Starlight Fairy", "Astral Dawn", "008/180", "EN", "Common", "Near Mint", 66, "Bulk Box 1", 30, 25, 1, "Bulk"),
    ]
    photo_yes = {0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 13}  # 12 of 18 cards photographed
    start = L0
    for i, (nm, st, no, lg, ra, cond, q, loc, dago, paid, each, note) in enumerate(rows):
        r = start + i
        vals = [nm, st, no, lg, ra, cond, q, loc, dminus(dago), paid, each]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=12, value=f'=IF(A{r}="","",G{r}*K{r})')
        ws.cell(row=r, column=13, value="Yes" if i in photo_yes else "No")
        ws.cell(row=r, column=14, value=note)
    end = start + LOG_ROWS - 1
    style_rows(ws, start, end, 14, text_left={1, 2, 8, 14}, ints={7}, dates={9}, money={10, 11, 12})
    for col_letter, lst in [("D", "LangList"), ("E", "RarityList"), ("F", "CondList"), ("H", "LocList"), ("M", "YesNoList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    tq = ws.cell(row=total, column=7, value=f"=SUM(G{start}:G{end})"); tq.style = "td"
    tq.font = Font(bold=True, color=PRIMARY); tq.fill = fill(SURFACE); tq.number_format = "0"
    for col in (10, 12):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "B5"
    nrange(wb, "ColName", "Collection", "A", start, end)
    nrange(wb, "ColSet", "Collection", "B", start, end)
    nrange(wb, "ColRarity", "Collection", "E", start, end)
    nrange(wb, "ColQty", "Collection", "G", start, end)
    nrange(wb, "ColLocation", "Collection", "H", start, end)
    nrange(wb, "ColPaid", "Collection", "J", start, end)
    nrange(wb, "ColEach", "Collection", "K", start, end)
    nrange(wb, "ColTotal", "Collection", "L", start, end)
    nrange(wb, "ColPhoto", "Collection", "M", start, end)
    ws.conditional_formatting.add(f"L{start}:L{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=650, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(f"K{start}:K{end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF" + WHITE,
                       end_type="num", end_value=600, end_color="FF" + HIGHLIGHT))
    ws.conditional_formatting.add(f"M{start}:M{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    # helper summaries for dashboard charts
    ws.cell(row=4, column=16, value="Rarity").style = "th"
    ws.cell(row=4, column=17, value="Value").style = "th"
    for i, ra in enumerate(["Secret Rare", "Ultra Rare", "Rare Holo", "Rare", "Uncommon", "Common"]):
        r = 5 + i
        ws.cell(row=r, column=16, value=ra).style = "td_left"
        c = ws.cell(row=r, column=17, value=f"=SUMIF(ColRarity,P{r},ColTotal)"); c.style = "td"; c.number_format = '"$"#,##0'
    ws.cell(row=4, column=19, value="Set").style = "th"
    ws.cell(row=4, column=20, value="Value").style = "th"
    for i, st in enumerate(["Obsidian Flames", "Genesis", "Astral Dawn", "Deep Currents", "Midnight Veil", "Tempest Rising", "Glacial Peak"]):
        r = 5 + i
        ws.cell(row=r, column=19, value=st).style = "td_left"
        c = ws.cell(row=r, column=20, value=f"=SUMIF(ColSet,S{r},ColTotal)"); c.style = "td"; c.number_format = '"$"#,##0'


# ===========================================================================
# 3 — Purchase Tracker
# ===========================================================================
def build_purchases(wb):
    rows = [
        ("CardVault Shop", 120, "Emberfang Dragon (Holo)", 1, 112, 8, 0),
        ("mint_pulls (eBay)", 80, "Golden Phoenix Alt Art — PSA 10", 1, 138, 12, 0),
        ("Dragon's Den LGS", 90, "Celestial Guardian (JP)", 1, 95, 0, 0),
        ("TCG Marketplace", 55, "Ancient Colossus — BGS 9.5", 1, 198, 12, 0),
        ("mint_pulls (eBay)", 75, "Twin Flames Duo — PSA 9", 1, 80, 8, 0),
        ("TCG Marketplace", 70, "Shadow Assassin", 1, 70, 5, 0),
        ("Card Show", 100, "Royal Kraken — PSA 8", 1, 70, 0, 0),
        ("Dragon's Den LGS", 45, "UR / Holo lot (Roc, Leviathan, Priestess)", 3, 143, 0, 0),
        ("Card Show", 60, "Rares lot (Herald, Titan, Hoard, Sprite, Wolf)", 16, 110, 0, 0),
        ("Dragon's Den LGS", 30, "Bulk commons & uncommons", 120, 63, 0, 0),
    ]
    ws = wb.create_sheet("Purchases"); ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 13, 34, 8, 11, 11, 9, 12])
    luxe_header(ws, "H", "🛒  PURCHASE TRACKER",
                "Every pickup logged — price, shipping & taxes roll into your true all-in cost.")
    table_headers(ws, 4, ["Seller", "Date", "Item", "Qty", "Price", "Shipping", "Taxes", "Total"])
    start = L0
    for i, (seller, dago, item, q, price, ship, tax) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([seller, dminus(dago), item, q, price, ship, tax], 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=8, value=f'=IF(A{r}="","",E{r}+F{r}+G{r})')
    end = start + LOG_ROWS - 1
    style_rows(ws, start, end, 8, text_left={1, 3}, dates={2}, ints={4}, money={5, 6, 7, 8})
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in (5, 6, 7, 8):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "PurchTotal", "Purchases", "H", start, end)
    nrange(wb, "PurchDate", "Purchases", "B", start, end)


# ===========================================================================
# 4 — Sales Tracker
# ===========================================================================
def build_sales(wb):
    rows = [
        ("kai_collects", 20, "Storm Herald (extra)", "eBay", 34, 5, 4),
        ("deckmaster99", 14, "Verdant Sprite playset", "TCG Marketplace", 18, 2, 3),
        ("Local pickup — Sam", 9, "Frostbite Wolf ×2", "Facebook", 16, 0, 0),
        ("bulk_buyer_tcg", 4, "Inferno Titan (MP)", "eBay", 12, 2, 4),
    ]
    ws = wb.create_sheet("Sales"); ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 13, 28, 17, 11, 9, 11, 12])
    luxe_header(ws, "H", "💵  SALES TRACKER",
                "Turn extras into cash — sale price, fees & shipping net out automatically.")
    table_headers(ws, 4, ["Buyer", "Date", "Item", "Marketplace", "Sale Price", "Fees", "Shipping", "Net Proceeds"])
    start = L0
    for i, (buyer, dago, item, mkt, price, fees, ship) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([buyer, dminus(dago), item, mkt, price, fees, ship], 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=8, value=f'=IF(A{r}="","",E{r}-F{r}-G{r})')
    end = start + 30 - 1
    style_rows(ws, start, end, 8, text_left={1, 3}, dates={2}, money={5, 6, 7, 8})
    add_dv(ws, f"D{start}:D{end}", "MarketList")
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in (5, 6, 7, 8):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "SaleNet", "Sales", "H", start, end)


# ===========================================================================
# 5 — Trade Tracker
# ===========================================================================
def build_trades(wb):
    rows = [
        ("Alex M.", 30, "Storm Herald + Dragon's Hoard", "Lunar Priestess + bulk lot", 45, 52),
        ("Sam K.", 24, "Tidecaller Leviathan", "Thunderclap Roc", 88, 96),
        ("League night — Riley", 17, "Uncommon lot (×6)", "Rare Holo (Storm Herald)", 12, 15),
        ("Jordan P.", 11, "Inferno Titan ×2", "Verdant Sprite ×2", 30, 26),
        ("Riley T.", 5, "Frostbite Wolf ×2", "Mystic Fountain lot", 20, 24),
    ]
    ws = wb.create_sheet("Trades"); ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 13, 28, 28, 12, 12, 12, 18])
    luxe_header(ws, "H", "🔁  TRADE TRACKER",
                "Win every swap — value out vs value in, with running profit/loss per trade.")
    table_headers(ws, 4, ["Trade Partner", "Date", "Cards Given", "Cards Received", "Value Out", "Value In", "Trade +/-", "Notes"])
    start = L0
    notes = ["Fair upgrade", "Even UR swap", "League trade", "Deck copies", "Bulk consolidation"]
    for i, (p, dago, gave, got, vout, vin) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([p, dminus(dago), gave, got, vout, vin], 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=7, value=f'=IF(A{r}="","",F{r}-E{r})')
        ws.cell(row=r, column=8, value=notes[i])
    end = start + 30 - 1
    style_rows(ws, start, end, 8, text_left={1, 3, 4, 8}, dates={2}, money={5, 6, 7})
    for r in range(start, end + 1):
        ws.cell(row=r, column=7).number_format = '+"$"#,##0;[Red]-"$"#,##0'
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in (5, 6, 7):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE)
        c.number_format = '+"$"#,##0;[Red]-"$"#,##0' if col == 7 else '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "TradePartner", "Trades", "A", start, end)
    nrange(wb, "TradeNet", "Trades", "G", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))


# ===========================================================================
# 6 — Grading Center
# ===========================================================================
def build_grading(wb):
    rows = [
        ("PSA", "Golden Phoenix (Alt Art)", 60, 25, "Returned", 10, 25, 150, 480),
        ("BGS", "Ancient Colossus", 55, 18, "Returned", 9.5, 30, 210, 620),
        ("PSA", "Twin Flames Duo", 60, 25, "Returned", 9, 25, 88, 265),
        ("PSA", "Royal Kraken", 90, 55, "Returned", 8, 25, 70, 150),
        ("PSA", "Emberfang Dragon (Holo)", 12, None, "Pending", None, 25, 340, None),
        ("PSA", "Celestial Guardian", 12, None, "Pending", None, 25, 210, None),
    ]
    ws = wb.create_sheet("Grading"); ws.sheet_view.showGridLines = False
    set_widths(ws, [11, 26, 13, 13, 12, 9, 9, 13, 12, 12])
    luxe_header(ws, "J", "💎  GRADING CENTER",
                "Track every submission — turnaround, grades & the value a slab adds.")
    table_headers(ws, 4, ["Company", "Card", "Submitted", "Returned", "Status", "Grade", "Fee", "Value Before", "Value After", "Value Added"])
    start = L0
    for i, (co, card, sub, ret, status, grade, fee, before, after) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=co)
        ws.cell(row=r, column=2, value=card)
        ws.cell(row=r, column=3, value=dminus(sub))
        if ret is not None:
            ws.cell(row=r, column=4, value=dminus(ret))
        ws.cell(row=r, column=5, value=status)
        if grade is not None:
            ws.cell(row=r, column=6, value=grade)
        ws.cell(row=r, column=7, value=fee)
        ws.cell(row=r, column=8, value=before)
        if after is not None:
            ws.cell(row=r, column=9, value=after)
        ws.cell(row=r, column=10, value=f'=IF(I{r}="","",I{r}-H{r})')
    end = start + 30 - 1
    style_rows(ws, start, end, 10, text_left={2}, dates={3, 4}, dec={6}, money={7, 8, 9, 10})
    for col_letter, lst in [("A", "GraderList"), ("E", "GradeStatusList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in (7, 10):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "GradeStatus", "Grading", "E", start, end)
    nrange(wb, "GradeCard", "Grading", "B", start, end)
    nrange(wb, "GradeAdded", "Grading", "J", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Pending"'], fill=fill(WARN_BG)))
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Returned"'], fill=fill(MINT_BG)))


# ===========================================================================
# 7 — Collection Value Analytics
# ===========================================================================
def build_value_analytics(wb):
    ws = wb.create_sheet("Value Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 14, 14, 3, 26, 13, 2])
    luxe_header(ws, "H", "📈  COLLECTION VALUE ANALYTICS",
                "Growth over time, purchase vs value, and your most valuable cards.")
    merge_set(ws, "B5:D5", "6-MONTH VALUE TREND", "section")
    ws.cell(row=6, column=2, value="Month").style = "th"
    ws.cell(row=6, column=3, value="Est. Value").style = "th"
    ws.cell(row=6, column=4, value="Total Cost").style = "th"
    months = _recent_months(6)
    vals = [1450, 1720, 1980, 2260, 2540, 2840]
    costs = [690, 800, 890, 990, 1060, 1124]
    for i, (m, v, cst) in enumerate(zip(months, vals, costs)):
        r = 7 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = '"$"#,##0'
        c2 = ws.cell(row=r, column=4, value=cst); c2.style = "td"; c2.number_format = '"$"#,##0'
        if i % 2:
            for cc in (2, 3, 4):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    cell_name(wb, "TrendMonth", "Value Analytics", "$B$7:$B$12")
    cell_name(wb, "TrendVal", "Value Analytics", "$C$7:$C$12")
    # top cards helper
    merge_set(ws, "F5:G5", "TOP 5 MOST VALUABLE", "section_gold")
    ws.cell(row=6, column=6, value="Card").style = "th"
    ws.cell(row=6, column=7, value="Value").style = "th"
    tops = [("Ancient Colossus (BGS 9.5)", 620), ("Golden Phoenix (PSA 10)", 480),
            ("Emberfang Dragon (Holo)", 340), ("Twin Flames Duo (PSA 9)", 265), ("Celestial Guardian (JP)", 210)]
    for i, (nm, v) in enumerate(tops):
        r = 7 + i
        ws.cell(row=r, column=6, value=nm).style = "td_left"
        c = ws.cell(row=r, column=7, value=v); c.style = "td"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=6).fill = fill(MUTED_ROW); ws.cell(row=r, column=7).fill = fill(MUTED_ROW)
    # charts
    ln = LineChart(); ln.title = "Collection Value Over Time"; ln.height = 8.5; ln.width = 13
    ln.add_data(Reference(ws, min_col=3, min_row=6, max_row=12), titles_from_data=True)
    ln.set_categories(Reference(ws, min_col=2, min_row=7, max_row=12)); ln.legend = None
    ws.add_chart(ln, "B15")
    bar = BarChart(); bar.type = "col"; bar.title = "Purchase Cost vs Value"; bar.height = 8.5; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=12), titles_from_data=True)
    bar.add_data(Reference(ws, min_col=4, min_row=6, max_row=12), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=7, max_row=12))
    ws.add_chart(bar, "F15")
    top = BarChart(); top.type = "bar"; top.title = "Top Valuable Cards"; top.height = 8.5; top.width = 13
    top.add_data(Reference(ws, min_col=7, min_row=6, max_row=11), titles_from_data=True)
    top.set_categories(Reference(ws, min_col=6, min_row=7, max_row=11)); top.legend = None
    ws.add_chart(top, "B33")


# ===========================================================================
# 8 — Wishlist
# ===========================================================================
def build_wishlist(wb):
    rows = [
        ("Prism Serpent (Alt Art)", "Obsidian Flames", "Grail", 350, 410, "No", "Waiting for a dip"),
        ("Solar Empress", "Genesis", "Grail", 280, 295, "No", "Watch weekly"),
        ("Ancient Colossus", "Genesis", "High", 220, 210, "Yes", "Got it graded!"),
        ("Twin Flames Duo", "Obsidian Flames", "High", 90, 88, "Yes", "PSA 9 pickup"),
        ("Nebula Djinn", "Astral Dawn", "High", 120, 135, "No", "Card show hunt"),
        ("Royal Kraken (graded)", "Deep Currents", "Medium", 75, 70, "Yes", "PSA 8 deal"),
        ("Gale Dancer", "Tempest Rising", "Medium", 40, 44, "No", ""),
        ("Obsidian Golem", "Obsidian Flames", "Medium", 55, 60, "No", "Trade target"),
        ("Lunar Priestess", "Midnight Veil", "Low", 30, 28, "Yes", "From Alex trade"),
        ("Coral Sentinel", "Deep Currents", "Low", 25, 27, "No", ""),
    ]
    ws, start, end = build_log(
        wb, "Wishlist", "⭐", "WISHLIST",
        "Hunt with a plan — targets, priorities & live progress toward every grail.",
        ["Card", "Set", "Priority", "Target Price", "Current Price", "Purchased?", "Notes"],
        rows, [26, 18, 11, 13, 13, 12, 24],
        text_left={1, 2, 7}, money={4, 5},
        validations=[("C", "PriorityList"), ("F", "YesNoList")], reserved=30)
    nrange(wb, "WishCard", "Wishlist", "A", start, end)
    nrange(wb, "WishBought", "Wishlist", "F", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"C{start}:C{end}",
        CellIsRule(operator="equal", formula=['"Grail"'], fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))


# ===========================================================================
# 9 — Duplicates Manager
# ===========================================================================
def build_duplicates(wb):
    rows = [
        ("Verdant Sprite", "Astral Dawn", 3, "Near Mint", "Trade", 12),
        ("Frostbite Wolf", "Glacial Peak", 4, "Near Mint", "Sell", 9),
        ("Storm Herald", "Tempest Rising", 1, "Near Mint", "Trade", 26),
        ("Arcane Scholar", "Astral Dawn", 18, "Near Mint", "Gift", 3),
        ("Mystic Fountain", "Deep Currents", 22, "Near Mint", "", 2),
    ]
    ws = wb.create_sheet("Duplicates"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 18, 12, 15, 12, 13, 13])
    luxe_header(ws, "G", "🃏  DUPLICATES MANAGER",
                "Extra copies are opportunities — flag each for trade, sale or gifting.")
    table_headers(ws, 4, ["Card", "Set", "Extra Copies", "Condition", "Plan", "Value (ea)", "Total Value"])
    start = L0
    for i, (nm, st, q, cond, plan, each) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([nm, st, q, cond, plan, each], 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=7, value=f'=IF(A{r}="","",C{r}*F{r})')
    end = start + 30 - 1
    style_rows(ws, start, end, 7, text_left={1, 2}, ints={3}, money={6, 7})
    for col_letter, lst in [("D", "CondList"), ("E", "PlanList")]:
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    c = ws.cell(row=total, column=7, value=f"=SUM(G{start}:G{end})")
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "DupCard", "Duplicates", "A", start, end)
    nrange(wb, "DupPlan", "Duplicates", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['""'], fill=fill(WARN_BG)))


# ===========================================================================
# 10 — Tournament Deck Builder
# ===========================================================================
def build_decks(wb):
    ws = wb.create_sheet("Deck Builder"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 8, 14, 26, 3, 22, 14, 2])
    luxe_header(ws, "I", "🏆  TOURNAMENT DECK BUILDER",
                "Build, tune & track — deck lists with roles, sideboard and a live win rate.")
    merge_set(ws, "B5:E5", "DECK — EMBERSTORM AGGRO", "section")
    table_headers(ws, 6, ["Card", "Qty", "Role", "Strategy Notes"], start_col=2)
    deck = [
        ("Emberfang Dragon (Holo)", 3, "Finisher", "Main win condition"),
        ("Inferno Titan", 4, "Core", "Early pressure"),
        ("Storm Herald", 4, "Support", "Energy acceleration"),
        ("Dragon's Hoard", 3, "Draw", "Card advantage"),
        ("Verdant Sprite", 2, "Tech", "Anti-control"),
        ("Frostbite Wolf", 2, "Tech", "Mirror breaker"),
        ("Energy / Resources", 20, "Engine", ""),
        ("Trainers / Supporters", 22, "Engine", "Draw + search"),
    ]
    start = 7
    for i, (nm, q, role, note) in enumerate(deck):
        r = start + i
        ws.cell(row=r, column=2, value=nm).style = "td_left"
        c = ws.cell(row=r, column=3, value=q); c.style = "td"; c.number_format = "0"
        ws.cell(row=r, column=4, value=role).style = "td"
        ws.cell(row=r, column=5, value=note).style = "td_left"
        if i % 2:
            for cc in range(2, 6):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    tr = start + len(deck)
    ws.cell(row=tr, column=2, value="TOTAL (60)").style = "th"
    c = ws.cell(row=tr, column=3, value=f"=SUM(C{start}:C{tr-1})"); c.style = "td"
    c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0"
    # record box
    merge_set(ws, "G5:H5", "SEASON RECORD", "section_gold")
    rec = [("Wins", 12, "DeckWins"), ("Losses", 4, "DeckLosses")]
    for i, (lab, v, nm) in enumerate(rec):
        r = 6 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=v); c.style = "input"; c.number_format = "0"
        cell_name(wb, nm, "Deck Builder", f"$H${r}")
    ws.cell(row=8, column=7, value="Win Rate").style = "field_label"
    c = ws.cell(row=8, column=8, value="=IFERROR(DeckWins/MAX(DeckWins+DeckLosses,1),0)")
    c.style = "field_value"; c.number_format = "0%"; c.fill = fill(MINT_BG)
    ws.cell(row=9, column=7, value="Format").style = "field_label"
    ws.cell(row=9, column=8, value="Standard").style = "field_value"
    # sideboard
    merge_set(ws, "G11:H11", "SIDEBOARD", "section_gold")
    for i, sb in enumerate(["Shadow Assassin ×2", "Mystic Fountain ×2", "Gale Dancer ×1"]):
        r = 12 + i
        merge_set(ws, f"G{r}:H{r}", sb, "td_left")
        if i % 2:
            ws.cell(row=r, column=7).fill = fill(MUTED_ROW)


# ===========================================================================
# 11 — Card Image Vault (photo gallery + per-card uploads)
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Card Vault"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "🖼  CARD IMAGE VAULT",
                "Upload a photo of every key card — visual proof of condition, grades & your best pulls.")
    # how-to-upload banner
    merge_set(ws, "B5:H5", "  HOW TO UPLOAD YOUR CARD PHOTOS", "th"); ws.row_dimensions[5].height = 24
    ws.merge_cells("B6:H8")
    c = ws["B6"]
    c.value = ("EXCEL:  click a photo box, then Insert ▸ Pictures ▸ This Device (Excel 365: right-click the image "
               "▸ Place in Cell to lock it to the slot).   GOOGLE SHEETS:  click the box, then Insert ▸ Image ▸ "
               "Insert image in cell — or type  =IMAGE(\"paste-your-photo-link\")  to pull one from the web. "
               "Then mark the card's Photo? column \"Yes\" in the Master Collection — the dashboard tracks your coverage.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in (6, 7, 8):
        ws.row_dimensions[rr].height = 22
        ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    # coverage stat
    ws.cell(row=9, column=2, value="Photos on file").style = "field_label"
    cc = ws.cell(row=9, column=3, value='=COUNTIF(ColPhoto,"Yes")&" of "&COUNTA(ColName)&" cards"')
    cc.style = "field_value"; cc.fill = fill(MINT_BG)
    ws.cell(row=9, column=6, value="Coverage").style = "field_label"
    cv = ws.cell(row=9, column=7, value='=IFERROR(COUNTIF(ColPhoto,"Yes")/MAX(COUNTA(ColName),1),0)')
    cv.style = "field_value"; cv.number_format = "0%"; cv.fill = fill(MINT_BG)
    # 8 upload slots with card details
    slots = [("Crown Jewel", "Ancient Colossus", "BGS 9.5", 620),
             ("Grail #2", "Golden Phoenix (Alt Art)", "PSA 10", 480),
             ("Newest Slab", "Twin Flames Duo", "PSA 9", 265),
             ("At The Graders", "Emberfang Dragon (Holo)", "Pending @ PSA", 340),
             ("Binder Spread", "Binder A — hits page", "—", None),
             ("Display Shelf", "Vault case lineup", "—", None),
             ("Tournament Win", "Emberstorm Aggro — 1st place", "—", None),
             ("Latest Pickup", "Celestial Guardian (JP)", "Near Mint", 210)]
    top0 = 11; card_h = 9
    for idx, (name, card, grade, val) in enumerate(slots):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th"); ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+4}", "📷\nUpload card photo here\n(see guide above)", "imgbox")
        for rr in range(row + 1, row + 5):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 5, column=col, value="Card").style = "field_label"
        merge_set(ws, f"{M}{row+5}:{R}{row+5}", card, "field_value")
        ws.cell(row=row + 6, column=col, value="Grade / Cond.").style = "field_label"
        ws.cell(row=row + 6, column=col + 1, value=grade).style = "field_value"
        cval = ws.cell(row=row + 6, column=col + 2, value=val if val is not None else "")
        cval.style = "field_value"; cval.number_format = '"$"#,##0'


# ===========================================================================
# 12 — Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 18, 3, 16, 12, 12, 2])
    luxe_header(ws, "H", "📊  ANALYTICS DASHBOARD",
                "Your collection by the numbers — health across every area, and one Collection Health Score.")
    merge_set(ws, "B5:D5", "COLLECTION HEALTH DIMENSIONS", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Value growth vs target", "=IFERROR(MIN((SUM(ColTotal)/SUM(ColPaid))/GrowthTarget,1),0)"),
        ("Grading pipeline returned", '=IFERROR(COUNTIF(GradeStatus,"Returned")/MAX(COUNTA(GradeStatus),1),0)'),
        ("Wishlist progress", '=IFERROR(COUNTIF(WishBought,"Yes")/MAX(COUNTA(WishCard),1),0)'),
        ("Storage documented", '=IFERROR(SUMPRODUCT((ColLocation<>"")*(ColName<>""))/MAX(COUNTA(ColName),1),0)'),
        ("Duplicates planned", '=IFERROR(SUMPRODUCT((DupPlan<>"")*(DupCard<>""))/MAX(COUNTA(DupCard),1),0)'),
        ("Tournament win rate", "=IFERROR(DeckWins/MAX(DeckWins+DeckLosses,1),0)"),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.75,"Strong",IF(C{r}>=0.5,"Growing","Focus Here"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(f"C{start}:C{end}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "F5:H5", "COLLECTION HEALTH SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY); cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "Growth · grading · wishlist · storage · duplicates · play.", "subtitle")
    ws["F10"].fill = fill(IVORY)
    cell_name(wb, "HealthRange", "Analytics", f"$C${start}:$C${end}")
    bar = BarChart(); bar.type = "bar"; bar.title = "Health by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None; ws.add_chart(bar, "B15")


# ===========================================================================
# 1 — Collection Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🎴  TRADING CARD COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Collection, value, grading, trades & wishlist — your whole vault, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("TOTAL CARDS", "=SUM(ColQty)", "num"),
        ("COLLECTION VALUE", "=SUM(ColTotal)", "money"),
        ("PURCHASE COST", "=SUM(ColPaid)", "money"),
        ("PROFIT / LOSS", "=SUM(ColTotal)-SUM(ColPaid)", "pmoney"),
        ("GROWTH", "=IFERROR(SUM(ColTotal)/SUM(ColPaid)-1,0)", "ppct"),
        ("TOP CARD VALUE", "=MAX(ColEach)", "money"),
    ]
    row2 = [
        ("PHOTOS ON FILE", '=IFERROR(COUNTIF(ColPhoto,"Yes")/MAX(COUNTA(ColName),1),0)', "pct"),
        ("CARDS GRADED", '=COUNTIF(GradeStatus,"Returned")', "num"),
        ("AWAITING GRADING", '=COUNTIF(GradeStatus,"Pending")', "num"),
        ("TRADES COMPLETED", "=COUNTA(TradePartner)", "num"),
        ("WISHLIST PROGRESS", '=IFERROR(COUNTIF(WishBought,"Yes")/MAX(COUNTA(WishCard),1),0)', "pct"),
        ("COLLECTION HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "VALUE & GROWTH", "section_gold")
    # value by rarity donut
    d1 = DoughnutChart(); d1.title = "Value by Rarity"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Collection"], min_col=17, min_row=4, max_row=10), titles_from_data=True)
    d1.set_categories(Reference(wb["Collection"], min_col=16, min_row=5, max_row=10)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")
    # collection value over time
    ln = LineChart(); ln.title = "Collection Value Over Time"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Value Analytics"], min_col=3, min_row=6, max_row=12), titles_from_data=True)
    ln.set_categories(Reference(wb["Value Analytics"], min_col=2, min_row=7, max_row=12)); ln.legend = None
    ws.add_chart(ln, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "SETS & GRADING", "section_gold")
    # value by set bar
    b1 = BarChart(); b1.type = "col"; b1.title = "Value by Set"; b1.height = 8.2; b1.width = 11.5
    b1.add_data(Reference(wb["Collection"], min_col=20, min_row=4, max_row=11), titles_from_data=True)
    b1.set_categories(Reference(wb["Collection"], min_col=19, min_row=5, max_row=11)); b1.legend = None
    ws.add_chart(b1, "B30")
    # grading before vs after
    g1 = BarChart(); g1.type = "col"; g1.title = "Grading: Value Before vs After"; g1.height = 8.2; g1.width = 11.5
    g1.add_data(Reference(wb["Grading"], min_col=8, min_row=4, max_row=8), titles_from_data=True)
    g1.add_data(Reference(wb["Grading"], min_col=9, min_row=4, max_row=8), titles_from_data=True)
    g1.set_categories(Reference(wb["Grading"], min_col=2, min_row=5, max_row=8))
    ws.add_chart(g1, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Trading Card Command Center™ — know your collection's value, always. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_collection(wb); build_purchases(wb)
    build_sales(wb); build_trades(wb); build_grading(wb); build_value_analytics(wb)
    build_wishlist(wb); build_duplicates(wb); build_decks(wb); build_gallery(wb)
    build_analytics(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Collection", "Purchases", "Sales", "Trades",
             "Grading", "Value Analytics", "Wishlist", "Duplicates", "Deck Builder",
             "Card Vault", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Trading_Card_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
