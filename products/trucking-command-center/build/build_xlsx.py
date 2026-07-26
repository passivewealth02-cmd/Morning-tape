"""Build Trucking Owner-Operator Command Center™ — The Truck's Operating System.

14 tabs · a premium owner-operator cost-per-mile operating system in Google Sheets &
Excel. Dashboard, a cost-per-mile engine (fixed ÷ miles + variable per mile → your true
cost, and the rate you cannot go below), fixed costs, variable costs, a load log with
deadhead, settlements, a fuel log with live MPG, maintenance, truck & trailer, IFTA
miles by state, a maintenance reserve fund and a monthly summary — one dashboard. Know
your cost per mile or the load board sets your pay for you.

Run: python3 build_xlsx.py   ->  ../Trucking_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
LOADSTATUS = ["Booked", "In transit", "Delivered", "Invoiced", "Paid"]
EQUIPTYPE = ["Tractor", "Trailer", "ELD", "APU", "Tools", "Other"]
MAINTTYPE = ["PM service", "Tires", "Brakes", "Engine", "DOT inspection", "Repair", "Other"]

# --- Cost-per-mile engine ---
MPG = 6.5
DIESEL_PRICE = 3.90
MAINT_CPM = 0.18
TIRE_CPM = 0.045
TOLL_CPM = 0.035
LOADED_MILES = 10000
DEADHEAD_MILES = 1200
RATE_PER_MILE = 2.35

# --- Goals ---
PPM_GOAL = 0.60
MARGIN_GOAL = 0.30
DEADHEAD_GOAL = 0.12
COVER_GOAL = 1.40
MPG_GOAL = 6.2
RESERVE_GOAL = 10000

# Fixed costs — what the truck costs before it turns a wheel: (line, monthly)
FIXED_LINES = [
    ("Tractor payment", 2200), ("Trailer payment", 600), ("Insurance (liability + cargo + physical)", 1100),
    ("Permits, plates, IFTA & IRP", 250), ("ELD & dispatch software", 65),
    ("Parking & yard", 150), ("Accounting & compliance", 125),
]

# Variable costs — what every mile costs: (line, per-mile, is_formula)
VARIABLE_LINES = [
    ("Fuel (diesel price \u00f7 MPG)", None, True),
    ("Maintenance & repair reserve", MAINT_CPM, False),
    ("Tires", TIRE_CPM, False),
    ("Tolls, scales & permits", TOLL_CPM, False),
]

# Loads this month: (date, broker, lane, loaded mi, deadhead mi, rate/mi)
LOADS = [
    ("07/01", "TQL", "Dallas TX \u2192 Memphis TN", 850, 110, 2.54),
    ("07/03", "CH Robinson", "Memphis TN \u2192 Atlanta GA", 620, 45, 2.55),
    ("07/05", "Landstar", "Atlanta GA \u2192 Miami FL", 1100, 160, 2.15),
    ("07/08", "Coyote", "Miami FL \u2192 Tampa FL", 480, 60, 2.70),
    ("07/10", "TQL", "Tampa FL \u2192 Birmingham AL", 940, 95, 2.30),
    ("07/12", "Echo", "Birmingham AL \u2192 Denver CO", 1250, 180, 2.10),
    ("07/15", "CH Robinson", "Denver CO \u2192 Kansas City MO", 720, 70, 2.45),
    ("07/17", "Convoy", "Kansas City MO \u2192 St Louis MO", 560, 55, 2.60),
    ("07/19", "Landstar", "St Louis MO \u2192 Houston TX", 1020, 140, 2.25),
    ("07/22", "TQL", "Houston TX \u2192 New Orleans LA", 800, 90, 2.40),
    ("07/24", "Echo", "New Orleans LA \u2192 Oklahoma City OK", 1160, 145, 2.20),
    ("07/27", "Coyote", "Oklahoma City OK \u2192 Dallas TX", 500, 50, 2.65),
]

# Settlements: (week, gross, deductions, note)
SETTLEMENTS = [
    ("Week 1", 6060, 420, "Fuel advance"), ("Week 2", 5754, 385, "Trailer rent"),
    ("Week 3", 5515, 410, "Fuel advance"), ("Week 4", 6171, 395, "Insurance draw"),
]

# Fuel log: (date, state, gallons, price/gal, odometer)
FUEL = [
    ("07/01", "TX", 215, 3.79, 412500), ("07/04", "TN", 220, 3.95, 413890),
    ("07/07", "FL", 210, 4.05, 415320), ("07/11", "AL", 218, 3.88, 416730),
    ("07/14", "CO", 212, 3.99, 418120), ("07/17", "MO", 225, 3.82, 419580),
    ("07/21", "TX", 208, 3.79, 420930), ("07/26", "OK", 215, 3.85, 422300),
]

# Maintenance: (date, type, item, cost, odometer)
MAINTENANCE = [
    ("07/02", "PM service", "Full service — oil, filters, grease", 480, 412700),
    ("07/09", "Tires", "2 steer tires replaced", 1240, 415400),
    ("07/16", "Brakes", "Trailer brake adjustment", 320, 418200),
    ("07/23", "DOT inspection", "Annual DOT inspection", 185, 421000),
    ("07/28", "Repair", "APU belt & alternator", 640, 422400),
]

# Truck & trailer: (item, type, value, monthly payment)
EQUIPMENT = [
    ("2021 Freightliner Cascadia", "Tractor", 92000, 2200),
    ("2019 Utility 53' dry van", "Trailer", 34000, 600),
    ("Motive ELD", "ELD", 0, 45),
    ("Thermo King APU", "APU", 8500, 0),
    ("Tool box & road kit", "Tools", 1800, 0),
]

# IFTA miles by state: (state, miles)
IFTA = [
    ("TX", 2400), ("OK", 900), ("AR", 1100), ("TN", 1400), ("GA", 1300),
    ("FL", 1600), ("AL", 1000), ("MS", 800), ("LA", 700),
]

# Reserve fund: (fund, target, saved)
RESERVE = [
    ("Engine & transmission", 6000, 2400), ("Tires", 2500, 900), ("Insurance deductible", 1500, 700),
]

# Monthly summary: (month, revenue)
MONTHS = [("Feb", 19800), ("Mar", 21200), ("Apr", 20400), ("May", 22600), ("Jun", 23100), ("Jul", 23500)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "money3": '"$"#,##0.000', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 20, 3] + [20] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your rates & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR RATES & GOALS", "section")
    controls = [
        ("Carrier", "Redline Freight Co.", None, "Carrier"),
        ("Owner-operator", "Wes", None, "Owner"),
        ("Truck MPG (loaded average)", MPG, "0.0", "MPG"),
        ("Diesel price / gallon", DIESEL_PRICE, '"$"#,##0.00', "DieselPrice"),
        ("Profit-per-mile goal", PPM_GOAL, '"$"#,##0.00', "PPMGoal"),
        ("Net-margin goal", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Deadhead goal (max)", DEADHEAD_GOAL, "0%", "DeadheadGoal"),
        ("Rate-vs-cost cover goal (×)", COVER_GOAL, "0.00", "CoverGoal"),
        ("MPG goal", MPG_GOAL, "0.0", "MPGGoal"),
        ("Maintenance reserve goal", RESERVE_GOAL, '"$"#,##0', "ReserveGoal"),
        ("Loaded miles this month", LOADED_MILES, "#,##0", "LoadedMiles"),
        ("Deadhead miles this month", DEADHEAD_MILES, "#,##0", "DeadheadMiles"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Load status", LOADSTATUS, "LoadStatusList"), ("F", "Equipment type", EQUIPTYPE, "EquipTypeList"),
             ("G", "Maintenance type", MAINTTYPE, "MaintTypeList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  \U0001f69b  TRUCKING OWNER-OPERATOR COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know your cost per mile — or the load board sets your pay for you.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE TRUCK, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("There is exactly one number that decides whether an owner-operator makes money: cost per mile. "
                      "Not the rate. Not the miles. The cost. And it has two halves — the fixed costs the truck "
                      "racks up sitting still, spread across the miles you actually run, plus the variable cost every "
                      "single mile burns. This workbook computes both, then does the part most cost-per-mile "
                      "calculators skip: it charges your DEADHEAD miles too, because nobody paid you for them. Then "
                      "run loads, settlements, fuel, maintenance, IFTA and a reserve fund — all in ONE premium "
                      "Google Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your MPG, diesel price and goals.",
             "2.  Fixed Costs: list what the truck costs standing still.",
             "3.  Variable Costs: fuel, maintenance, tires and tolls per mile.",
             "4.  Enter your loaded and deadhead miles for the month.",
             "5.  Read your TRUE cost per loaded mile — your rate floor.",
             "6.  Check the Dashboard: profit per mile & a Road Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional carrier (Redline Freight Co., owner-operator Wes) is included so you can "
               "see how it all connects — just type over it with your own numbers. The detail that catches "
               "everyone: your rate is paid on LOADED miles, but your costs run on TOTAL miles. Deadhead is not free "
               "— it burns fuel, wears tires and uses up the month. This workbook divides total cost by loaded "
               "miles, which is the only honest way to get a rate floor. Twelve matching printable pages (trip sheet, "
               "load worksheet, fuel log, PM schedule & more) are included. This is a business & organizing tool, not "
               "financial, tax, legal or DOT-compliance advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Deadhead is not free. Charge it, or the load board keeps the difference.", "section_gold")


# ===========================================================================
def build_fixed(wb):
    ws = wb.create_sheet("Fixed Costs"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 44, 18, 18, 2])
    luxe_header(ws, "D", "\U0001f3e6  FIXED COSTS",
                "What the truck costs you every month before it turns a wheel.")
    table_headers(ws, 4, ["Line", "Monthly", "Yearly"], start_col=2)
    fs_ = L0
    for i, (lab, amt) in enumerate(FIXED_LINES):
        r = fs_ + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        cy = ws.cell(row=r, column=4, value=f"=C{r}*12"); cy.style = "td"; cy.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 5):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    fe = fs_ + len(FIXED_LINES) - 1
    nrange(wb, "FixedLines", "Fixed Costs", "C", fs_, fe)
    tot = fe + 1
    ws.cell(row=tot, column=2, value="= FIXED COSTS / MONTH").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(FixedLines)"); ct.style = "td"
    ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "FixedTotal", "Fixed Costs", f"$C${tot}")
    cty = ws.cell(row=tot, column=4, value="=FixedTotal*12"); cty.style = "td"
    cty.font = Font(bold=True, size=12, color=PRIMARY); cty.fill = fill(MINT_BG); cty.number_format = '"$"#,##0'
    ws.cell(row=tot + 2, column=2, value="÷ Total miles run this month").style = "field_label"
    cm = ws.cell(row=tot + 2, column=3, value="=TotalMiles"); cm.style = "field_value"; cm.number_format = "#,##0"
    ws.cell(row=tot + 3, column=2, value="= FIXED COST PER MILE").style = "th"
    cf = ws.cell(row=tot + 3, column=3, value="=IFERROR(FixedTotal/TotalMiles,0)"); cf.style = "td"
    cf.font = Font(bold=True, size=14, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = '"$"#,##0.000'
    cell_name(wb, "FixedCPM", "Fixed Costs", f"$C${tot+3}")
    ws.cell(row=tot + 5, column=2, value="Park the truck for a week and this bill arrives anyway. Spread it over real miles.").style = "section_gold"


def build_variable(wb):
    ws = wb.create_sheet("Variable Costs"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 44, 18, 18, 2])
    luxe_header(ws, "D", "⛽  VARIABLE COSTS",
                "What every single mile burns — loaded or empty, it costs the same.")
    ws.cell(row=5, column=2, value="FUEL — THE BIGGEST LINE").style = "section_gold"
    ws.cell(row=6, column=2, value="Diesel price / gallon").style = "field_label"
    cd = ws.cell(row=6, column=3, value="=DieselPrice"); cd.style = "field_value"; cd.number_format = '"$"#,##0.00'
    ws.cell(row=7, column=2, value="÷ Miles per gallon").style = "field_label"
    cg = ws.cell(row=7, column=3, value="=MPG"); cg.style = "field_value"; cg.number_format = "0.0"
    ws.cell(row=8, column=2, value="= FUEL COST PER MILE").style = "th"
    cfu = ws.cell(row=8, column=3, value="=IFERROR(DieselPrice/MPG,0)"); cfu.style = "td"
    cfu.font = Font(bold=True, size=13, color=PRIMARY); cfu.fill = fill(MINT_BG); cfu.number_format = '"$"#,##0.000'
    cell_name(wb, "FuelCPM", "Variable Costs", "$C$8")

    ws.cell(row=10, column=2, value="EVERY OTHER PER-MILE COST").style = "section_gold"
    table_headers(ws, 11, ["Line", "Per mile", "Per month"], start_col=2)
    vs = 12
    ws.cell(row=vs, column=2, value="Fuel (diesel price ÷ MPG)").style = "td_left"
    cv0 = ws.cell(row=vs, column=3, value="=FuelCPM"); cv0.style = "td"; cv0.number_format = '"$"#,##0.000'
    cp0 = ws.cell(row=vs, column=4, value=f"=C{vs}*TotalMiles"); cp0.style = "td"; cp0.number_format = '"$"#,##0'
    for i, (lab, val, _) in enumerate(VARIABLE_LINES[1:], 1):
        r = vs + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"; c.number_format = '"$"#,##0.000'
        cp = ws.cell(row=r, column=4, value=f"=C{r}*TotalMiles"); cp.style = "td"; cp.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 5):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    ve = vs + len(VARIABLE_LINES) - 1
    nrange(wb, "VarLines", "Variable Costs", "C", vs, ve)
    tot = ve + 1
    ws.cell(row=tot, column=2, value="= VARIABLE COST PER MILE").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(VarLines)"); ct.style = "td"
    ct.font = Font(bold=True, size=14, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0.000'
    cell_name(wb, "VarCPM", "Variable Costs", f"$C${tot}")
    cm = ws.cell(row=tot, column=4, value="=VarCPM*TotalMiles"); cm.style = "td"
    cm.font = Font(bold=True, size=12, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0'
    cell_name(wb, "VarTotal", "Variable Costs", f"$D${tot}")
    ws.cell(row=tot + 2, column=2, value="Half a mile per gallon is worth thousands a year. Check the Fuel Log.").style = "section_gold"


def build_cpm(wb):
    ws = wb.create_sheet("Cost Per Mile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 44, 18, 2])
    luxe_header(ws, "C", "\U0001f4cf  COST PER MILE — THE ENGINE",
                "Fixed ÷ miles + variable per mile = your true cost, and the rate you must never go below.")
    ws.cell(row=5, column=2, value="THE MILES").style = "section_gold"
    ws.cell(row=6, column=2, value="Loaded miles (someone paid you)").style = "field_label"
    cl = ws.cell(row=6, column=3, value="=LoadedMiles"); cl.style = "field_value"; cl.number_format = "#,##0"
    ws.cell(row=7, column=2, value="+ Deadhead miles (nobody paid you)").style = "field_label"
    cdh = ws.cell(row=7, column=3, value="=DeadheadMiles"); cdh.style = "field_value"
    cdh.number_format = "#,##0"; cdh.fill = fill(RED_BG)
    ws.cell(row=8, column=2, value="= TOTAL MILES RUN").style = "th"
    ctm = ws.cell(row=8, column=3, value="=LoadedMiles+DeadheadMiles"); ctm.style = "td"
    ctm.font = Font(bold=True, size=13, color=PRIMARY); ctm.fill = fill(SURFACE); ctm.number_format = "#,##0"
    cell_name(wb, "TotalMiles", "Cost Per Mile", "$C$8")
    ws.cell(row=9, column=2, value="= DEADHEAD %").style = "th"
    cdp = ws.cell(row=9, column=3, value="=IFERROR(DeadheadMiles/TotalMiles,0)"); cdp.style = "td"
    cdp.font = Font(bold=True, size=12, color=PRIMARY); cdp.fill = fill(WARN_BG); cdp.number_format = "0.0%"
    cell_name(wb, "DeadheadPct", "Cost Per Mile", "$C$9")

    ws.cell(row=11, column=2, value="THE COST").style = "section_gold"
    ws.cell(row=12, column=2, value="Fixed cost per mile").style = "field_label"
    cf = ws.cell(row=12, column=3, value="=FixedCPM"); cf.style = "field_value"; cf.number_format = '"$"#,##0.000'
    ws.cell(row=13, column=2, value="+ Variable cost per mile").style = "field_label"
    cv = ws.cell(row=13, column=3, value="=VarCPM"); cv.style = "field_value"; cv.number_format = '"$"#,##0.000'
    ws.cell(row=14, column=2, value="= COST PER MILE RUN").style = "th"
    cc = ws.cell(row=14, column=3, value="=FixedCPM+VarCPM"); cc.style = "td"
    cc.font = Font(bold=True, size=14, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0.000'
    cell_name(wb, "TotalCPM", "Cost Per Mile", "$C$14")
    ws.cell(row=15, column=2, value="× Total miles = TOTAL COST THIS MONTH").style = "th"
    ctc = ws.cell(row=15, column=3, value="=TotalCPM*TotalMiles"); ctc.style = "td"
    ctc.font = Font(bold=True, size=13, color=PRIMARY); ctc.fill = fill(SURFACE); ctc.number_format = '"$"#,##0'
    cell_name(wb, "TotalCostMonth", "Cost Per Mile", "$C$15")

    ws.cell(row=17, column=2, value="⚠ THE NUMBER THAT ACTUALLY MATTERS").style = "section_gold"
    ws.cell(row=18, column=2, value="Total cost ÷ LOADED miles").style = "field_label"
    ws.cell(row=19, column=2, value="= COST PER LOADED MILE").style = "th"
    ccl = ws.cell(row=19, column=3, value="=IFERROR(TotalCostMonth/LoadedMiles,0)"); ccl.style = "td"
    ccl.font = Font(bold=True, size=16, color=PRIMARY); ccl.fill = fill(MINT_BG); ccl.number_format = '"$"#,##0.00'
    cell_name(wb, "CostPerLoaded", "Cost Per Mile", "$C$19")
    ws.cell(row=20, column=2, value="= YOUR RATE FLOOR — NEVER BOOK BELOW THIS").style = "th"
    cfl = ws.cell(row=20, column=3, value="=CostPerLoaded"); cfl.style = "td"
    cfl.font = Font(bold=True, size=14, color=PRIMARY); cfl.fill = fill(RED_BG); cfl.number_format = '"$"#,##0.00'

    ws.cell(row=22, column=2, value="THE LOAD").style = "section_gold"
    ws.cell(row=23, column=2, value="Your average rate per loaded mile").style = "field_label"
    cr = ws.cell(row=23, column=3, value=RATE_PER_MILE); cr.style = "input"; cr.number_format = '"$"#,##0.00'
    cell_name(wb, "RatePerMile", "Cost Per Mile", "$C$23")
    ws.cell(row=24, column=2, value="= PROFIT PER LOADED MILE").style = "th"
    cp = ws.cell(row=24, column=3, value="=RatePerMile-CostPerLoaded"); cp.style = "td"
    cp.font = Font(bold=True, size=16, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = '"$"#,##0.000'
    cell_name(wb, "ProfitPerMile", "Cost Per Mile", "$C$24")
    ws.cell(row=25, column=2, value="= RATE COVERS COST THIS MANY TIMES").style = "th"
    cv2 = ws.cell(row=25, column=3, value="=IFERROR(RatePerMile/CostPerLoaded,0)"); cv2.style = "td"
    cv2.font = Font(bold=True, size=13, color=PRIMARY); cv2.fill = fill(MINT_BG); cv2.number_format = '0.00"×"'
    cell_name(wb, "CoverRatio", "Cost Per Mile", "$C$25")

    ws.cell(row=27, column=2, value="⚠ WHAT DEADHEAD REALLY COSTS YOU").style = "section_gold"
    ws.cell(row=28, column=2, value="You booked at this rate per loaded mile").style = "field_label"
    c1 = ws.cell(row=28, column=3, value="=RatePerMile"); c1.style = "field_value"; c1.number_format = '"$"#,##0.00'
    ws.cell(row=29, column=2, value="But you got paid this per mile you ACTUALLY drove").style = "field_label"
    c2 = ws.cell(row=29, column=3, value="=IFERROR(RatePerMile*LoadedMiles/TotalMiles,0)"); c2.style = "field_value"
    c2.number_format = '"$"#,##0.00'; c2.fill = fill(RED_BG)
    ws.cell(row=30, column=2, value="Deadhead quietly took this much off the load").style = "field_label"
    c3 = ws.cell(row=30, column=3, value="=RatePerMile-IFERROR(RatePerMile*LoadedMiles/TotalMiles,0)")
    c3.style = "field_value"; c3.number_format = '"$"#,##0.00'; c3.fill = fill(RED_BG)
    ws.cell(row=31, column=2, value="…which is this much a month").style = "field_label"
    c4 = ws.cell(row=31, column=3, value="=(RatePerMile-IFERROR(RatePerMile*LoadedMiles/TotalMiles,0))*LoadedMiles")
    c4.style = "field_value"; c4.number_format = '"$"#,##0'; c4.fill = fill(WARN_BG)
    ws.cell(row=33, column=2, value="A $2.35 load with 11% deadhead is a $2.10 load. Book it knowing that.").style = "section_gold"


def build_loads(wb):
    ws = wb.create_sheet("Loads"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 11, 17, 34, 13, 13, 12, 14, 14, 2])
    luxe_header(ws, "I", "\U0001f4e6  LOADS",
                "Every load — the lane, the deadhead to get there, and what it actually paid.")
    table_headers(ws, 4, ["Date", "Broker", "Lane", "Loaded mi", "Deadhead", "Rate/mi", "Revenue", "All-in $/mi"], start_col=2)
    start = L0
    for i, (dt, brk, lane, lm, dh, rate) in enumerate(LOADS):
        r = start + i
        ws.cell(row=r, column=2, value=dt).style = "td"
        ws.cell(row=r, column=3, value=brk).style = "td"
        ws.cell(row=r, column=4, value=lane).style = "td_left"
        cl = ws.cell(row=r, column=5, value=lm); cl.style = "input"; cl.number_format = "#,##0"
        cd = ws.cell(row=r, column=6, value=dh); cd.style = "input"; cd.number_format = "#,##0"
        cr = ws.cell(row=r, column=7, value=rate); cr.style = "input"; cr.number_format = '"$"#,##0.00'
        cv = ws.cell(row=r, column=8, value=f"=E{r}*G{r}"); cv.style = "td"; cv.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=9, value=f"=IFERROR(H{r}/(E{r}+F{r}),0)"); ca.style = "td"; ca.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 10):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LOADS) - 1
    nrange(wb, "LoadLoaded", "Loads", "E", start, end)
    nrange(wb, "LoadDead", "Loads", "F", start, end)
    nrange(wb, "LoadRev", "Loads", "H", start, end)
    nrange(wb, "LoadAllIn", "Loads", "I", start, end)
    ws.conditional_formatting.add(f"I{start}:I{end}", DataBarRule(start_type="min", end_type="max", color=HIGHLIGHT))
    ws.conditional_formatting.add(f"I{start}:I{end}", CellIsRule(operator="lessThan", formula=["CostPerLoaded"],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL LOADS").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cl = ws.cell(row=tot, column=5, value="=SUM(LoadLoaded)"); cl.style = "td"
    cl.font = Font(bold=True, color=PRIMARY); cl.fill = fill(SURFACE); cl.number_format = "#,##0"
    cd = ws.cell(row=tot, column=6, value="=SUM(LoadDead)"); cd.style = "td"
    cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = "#,##0"
    ws.cell(row=tot, column=7).style = "td"; ws.cell(row=tot, column=7).fill = fill(SURFACE)
    cv = ws.cell(row=tot, column=8, value="=SUM(LoadRev)"); cv.style = "td"
    cv.font = Font(bold=True, size=12, color=PRIMARY); cv.fill = fill(MINT_BG); cv.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyRevenue", "Loads", f"$H${tot}")
    ca = ws.cell(row=tot, column=9, value=f"=IFERROR(H{tot}/(E{tot}+F{tot}),0)"); ca.style = "td"
    ca.font = Font(bold=True, size=12, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = '"$"#,##0.00'
    ws.cell(row=tot + 2, column=2, value="Average rate per LOADED mile").style = "field_label"
    cavg = ws.cell(row=tot + 2, column=8, value=f"=IFERROR(H{tot}/E{tot},0)"); cavg.style = "field_value"; cavg.number_format = '"$"#,##0.00'
    ws.cell(row=tot + 3, column=2, value="Any load whose all-in rate is under your cost per loaded mile is flagged red.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_settlements(wb):
    ws, start, end = build_log(
        wb, "Settlements", "\U0001f4b5", "SETTLEMENTS", "What the broker paid, what they took out, and what actually landed.",
        ["Week", "Gross", "Deductions", "Net", "Note"], SETTLEMENTS,
        [2, 16, 16, 16, 16, 28, 2], text_left={6}, money={3, 4, 5}, start_col=2)
    for i, (wk, gross, ded, note) in enumerate(SETTLEMENTS):
        r = start + i
        ws.cell(row=r, column=6, value=note).style = "td_left"
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}").number_format = '"$"#,##0'
    nrange(wb, "SetGross", "Settlements", "C", start, end)
    nrange(wb, "SetDed", "Settlements", "D", start, end)
    nrange(wb, "SetNet", "Settlements", "E", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="TOTAL SETTLED").style = "th"
    c1 = ws.cell(row=tr, column=3, value="=SUM(SetGross)"); c1.style = "td"
    c1.font = Font(bold=True, color=PRIMARY); c1.fill = fill(SURFACE); c1.number_format = '"$"#,##0'
    c2 = ws.cell(row=tr, column=4, value="=SUM(SetDed)"); c2.style = "td"
    c2.font = Font(bold=True, color=DANGER); c2.fill = fill(RED_BG); c2.number_format = '"$"#,##0'
    c3 = ws.cell(row=tr, column=5, value="=SUM(SetNet)"); c3.style = "td"
    c3.font = Font(bold=True, size=12, color=PRIMARY); c3.fill = fill(MINT_BG); c3.number_format = '"$"#,##0'
    cell_name(wb, "SettledNet", "Settlements", f"$E${tr}")
    ws.cell(row=tr + 2, column=2, value="Deductions are real money. Track every advance, escrow and trailer rent.").style = "section_gold"


def build_fuel(wb):
    ws = wb.create_sheet("Fuel Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 12, 14, 14, 15, 15, 14, 2])
    luxe_header(ws, "H", "⛽  FUEL LOG",
                "Every fill-up — and the live MPG that decides your biggest cost per mile.")
    table_headers(ws, 4, ["Date", "State", "Gallons", "$/gal", "Cost", "Odometer", "Miles run"], start_col=2)
    start = L0
    for i, (dt, st, gal, price, odo) in enumerate(FUEL):
        r = start + i
        ws.cell(row=r, column=2, value=dt).style = "td"
        ws.cell(row=r, column=3, value=st).style = "td"
        cg = ws.cell(row=r, column=4, value=gal); cg.style = "input"; cg.number_format = "#,##0"
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cc = ws.cell(row=r, column=6, value=f"=D{r}*E{r}"); cc.style = "td"; cc.number_format = '"$"#,##0.00'
        co = ws.cell(row=r, column=7, value=odo); co.style = "input"; co.number_format = "#,##0"
        if i == 0:
            cm = ws.cell(row=r, column=8, value=""); cm.style = "td"
        else:
            cm = ws.cell(row=r, column=8, value=f"=G{r}-G{r-1}"); cm.style = "td"; cm.number_format = "#,##0"
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FUEL) - 1
    nrange(wb, "FuelGal", "Fuel Log", "D", start, end)
    nrange(wb, "FuelCost", "Fuel Log", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="THIS MONTH").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cg = ws.cell(row=tot, column=4, value="=SUM(FuelGal)"); cg.style = "td"
    cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(SURFACE); cg.number_format = "#,##0"
    cell_name(wb, "GallonsMonth", "Fuel Log", f"$D${tot}")
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cc = ws.cell(row=tot, column=6, value="=SUM(FuelCost)"); cc.style = "td"
    cc.font = Font(bold=True, size=12, color=PRIMARY); cc.fill = fill(MINT_BG); cc.number_format = '"$"#,##0'
    cell_name(wb, "FuelSpend", "Fuel Log", f"$F${tot}")
    ws.cell(row=tot + 2, column=2, value="= ACTUAL MPG (total miles ÷ gallons)").style = "th"
    cm = ws.cell(row=tot + 2, column=6, value="=IFERROR(TotalMiles/GallonsMonth,0)"); cm.style = "td"
    cm.font = Font(bold=True, size=14, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = "0.00"
    cell_name(wb, "ActualMPG", "Fuel Log", f"$F${tot+2}")
    ws.cell(row=tot + 3, column=2, value="Average price per gallon paid").style = "field_label"
    cap = ws.cell(row=tot + 3, column=6, value="=IFERROR(FuelSpend/GallonsMonth,0)"); cap.style = "field_value"; cap.number_format = '"$"#,##0.00'
    ws.cell(row=tot + 5, column=2, value="Gain half an MPG and you cut about five cents a mile. That is thousands a year.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_maintenance(wb):
    ws, start, end = build_log(
        wb, "Maintenance", "\U0001f527", "MAINTENANCE", "Every service and repair — with the odometer, so PM intervals are real.",
        ["Date", "Type", "Item", "Cost", "Odometer"], MAINTENANCE,
        [2, 12, 17, 38, 14, 15, 2], text_left={4}, money={5}, ints={6},
        validations=[("C", "MaintTypeList")], start_col=2)
    nrange(wb, "MaintCost", "Maintenance", "E", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="SPENT THIS MONTH").style = "th"
    c1 = ws.cell(row=tr, column=5, value="=SUM(MaintCost)"); c1.style = "td"
    c1.font = Font(bold=True, size=12, color=PRIMARY); c1.fill = fill(SURFACE); c1.number_format = '"$"#,##0'
    cell_name(wb, "MaintSpend", "Maintenance", f"$E${tr}")
    ws.cell(row=tr + 1, column=2, value="Per mile run").style = "field_label"
    c2 = ws.cell(row=tr + 1, column=5, value="=IFERROR(MaintSpend/TotalMiles,0)"); c2.style = "field_value"
    c2.number_format = '"$"#,##0.000'; c2.fill = fill(MINT_BG)
    ws.cell(row=tr + 3, column=2, value="If this is consistently above your reserve rate, raise the reserve — don't hope.").style = "section_gold"


def build_equipment(wb):
    ws, start, end = build_log(
        wb, "Truck & Trailer", "\U0001f69b", "TRUCK & TRAILER", "What you own, what it's worth, and what it costs you every month.",
        ["Item", "Type", "Value", "Monthly payment", "Yearly"], [(i, t, v, p) for (i, t, v, p) in EQUIPMENT],
        [2, 34, 16, 16, 18, 16, 2], text_left={2}, money={4, 5, 6},
        validations=[("C", "EquipTypeList")], start_col=2)
    for r in range(start, start + len(EQUIPMENT)):
        ws.cell(row=r, column=6, value=f"=E{r}*12").number_format = '"$"#,##0'
    nrange(wb, "EquipValue", "Truck & Trailer", "D", start, end)
    nrange(wb, "EquipPay", "Truck & Trailer", "E", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="TOTAL").style = "th"
    c1 = ws.cell(row=tr, column=4, value="=SUM(EquipValue)"); c1.style = "td"
    c1.font = Font(bold=True, color=PRIMARY); c1.fill = fill(SURFACE); c1.number_format = '"$"#,##0'
    c2 = ws.cell(row=tr, column=5, value="=SUM(EquipPay)"); c2.style = "td"
    c2.font = Font(bold=True, size=12, color=PRIMARY); c2.fill = fill(MINT_BG); c2.number_format = '"$"#,##0'
    cell_name(wb, "EquipPayTotal", "Truck & Trailer", f"$E${tr}")


def build_ifta(wb):
    ws = wb.create_sheet("IFTA & Miles"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 18, 16, 16, 16, 2])
    luxe_header(ws, "E", "\U0001f5fa  IFTA & MILES BY STATE",
                "Miles per state and the gallons that go with them — quarterly filing made painless.")
    table_headers(ws, 4, ["State", "Miles", "Gallons", "Share"], start_col=2)
    start = L0
    for i, (st, mi) in enumerate(IFTA):
        r = start + i
        ws.cell(row=r, column=2, value=st).style = "td"
        cm = ws.cell(row=r, column=3, value=mi); cm.style = "input"; cm.number_format = "#,##0"
        cg = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/MPG,0)"); cg.style = "td"; cg.number_format = "#,##0.0"
        cs = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/TotalMiles,0)"); cs.style = "td"; cs.number_format = "0.0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(IFTA) - 1
    nrange(wb, "IftaMiles", "IFTA & Miles", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", DataBarRule(start_type="min", end_type="max", color=HIGHLIGHT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL STATES").style = "th"
    cm = ws.cell(row=tot, column=3, value="=SUM(IftaMiles)"); cm.style = "td"
    cm.font = Font(bold=True, size=12, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = "#,##0"
    cell_name(wb, "IftaTotal", "IFTA & Miles", f"$C${tot}")
    cg = ws.cell(row=tot, column=4, value="=IFERROR(IftaTotal/MPG,0)"); cg.style = "td"
    cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(MINT_BG); cg.number_format = "#,##0.0"
    ws.cell(row=tot + 2, column=2, value="State miles should tie back to your total miles run. If they don't, a trip sheet is missing.").style = "section_gold"
    ws.add_chart(_barchart(ws, "Miles by State", start, end, 3, 2), "G5")


def build_reserve(wb):
    ws = wb.create_sheet("Reserve Fund"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 16, 16, 2])
    luxe_header(ws, "E", "\U0001f6e1  MAINTENANCE RESERVE FUND",
                "The fund that decides whether a blown engine is a bad week or the end of the business.")
    table_headers(ws, 4, ["Fund", "Target", "Saved", "Funded"], start_col=2)
    start = L0
    for i, (fnd, tgt, sav) in enumerate(RESERVE):
        r = start + i
        ws.cell(row=r, column=2, value=fnd).style = "td_left"
        ct = ws.cell(row=r, column=3, value=tgt); ct.style = "input"; ct.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=sav); cs.style = "input"; cs.number_format = '"$"#,##0'
        cf = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); cf.style = "td"; cf.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(RESERVE) - 1
    nrange(wb, "ResTarget", "Reserve Fund", "C", start, end)
    nrange(wb, "ResSaved", "Reserve Fund", "D", start, end)
    nrange(wb, "ResPct", "Reserve Fund", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + RED_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL RESERVES").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(ResTarget)"); ct.style = "td"
    ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value="=SUM(ResSaved)"); cs.style = "td"
    cs.font = Font(bold=True, size=12, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cell_name(wb, "Reserve", "Reserve Fund", f"$D${tot}")
    cf = ws.cell(row=tot, column=5, value="=IFERROR(Reserve/ReserveGoal,0)"); cf.style = "td"
    cf.font = Font(bold=True, size=12, color=PRIMARY); cf.fill = fill(WARN_BG); cf.number_format = "0%"
    ws.cell(row=tot + 2, column=2, value="Set aside per mile run to fully fund it").style = "field_label"
    cpm = ws.cell(row=tot + 2, column=4, value="=IFERROR((ReserveGoal-Reserve)/TotalMiles/6,0)"); cpm.style = "field_value"
    cpm.number_format = '"$"#,##0.000'; cpm.fill = fill(MINT_BG)
    ws.cell(row=tot + 3, column=2, value="(Fully funded in six months at that rate.)").style = "field_label"
    ws.cell(row=tot + 5, column=2, value="An engine costs about the same whether or not you saved for it. One of those is survivable.").style = "section_gold"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 2])
    luxe_header(ws, "C", "\U0001f4c8  MONTHLY SUMMARY",
                "Revenue month by month — and the profit per mile that says whether the year is working.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    ws.cell(row=6, column=2, value="Revenue").style = "field_label"
    c1 = ws.cell(row=6, column=3, value="=MonthlyRevenue"); c1.style = "field_value"; c1.number_format = '"$"#,##0'
    ws.cell(row=7, column=2, value="− Total cost (fixed + variable)").style = "field_label"
    c2 = ws.cell(row=7, column=3, value="=TotalCostMonth"); c2.style = "field_value"; c2.number_format = '"$"#,##0'
    ws.cell(row=8, column=2, value="= MONTHLY PROFIT").style = "th"
    c3 = ws.cell(row=8, column=3, value="=MonthlyRevenue-TotalCostMonth"); c3.style = "td"
    c3.font = Font(bold=True, size=16, color=PRIMARY); c3.fill = fill(MINT_BG); c3.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyProfit", "Monthly Summary", "$C$8")
    ws.cell(row=9, column=2, value="= NET MARGIN").style = "th"
    c4 = ws.cell(row=9, column=3, value="=IFERROR(MonthlyProfit/MonthlyRevenue,0)"); c4.style = "td"
    c4.font = Font(bold=True, size=13, color=PRIMARY); c4.fill = fill(MINT_BG); c4.number_format = "0.0%"
    cell_name(wb, "NetMargin", "Monthly Summary", "$C$9")
    ws.cell(row=10, column=2, value="= RUN-RATE YEAR").style = "th"
    c5 = ws.cell(row=10, column=3, value="=MonthlyProfit*12"); c5.style = "td"
    c5.font = Font(bold=True, size=14, color=PRIMARY); c5.fill = fill(SURFACE); c5.number_format = '"$"#,##0'
    cell_name(wb, "RunRate", "Monthly Summary", "$C$10")
    ws.cell(row=12, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 13, ["Month", "Revenue"], start_col=2)
    ts = 14
    for i, (m, rev) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        crv = ws.cell(row=r, column=3, value=rev); crv.style = "input"; crv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "RevTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Revenue by Month", ts, te, 3, 2), "E5")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  \U0001f69b  TRUCKING OWNER-OPERATOR COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Cost per mile, profit per mile & a Road Score — your whole truck, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("FIXED / MONTH", "=FixedTotal", "money"),
        ("FIXED COST / MILE", "=FixedCPM", "money3"),
        ("VARIABLE / MILE", "=VarCPM", "money3"),
        ("COST PER MILE RUN", "=TotalCPM", "money3"),
        ("COST / LOADED MILE", "=CostPerLoaded", "money2"),
        ("RATE / LOADED MILE", "=RatePerMile", "money2"),
    ]
    row2 = [
        ("PROFIT / LOADED MILE", "=ProfitPerMile", "money3"),
        ("LOADED MILES", "=LoadedMiles", "num"),
        ("DEADHEAD", "=DeadheadPct", "pct1"),
        ("MONTHLY REVENUE", "=MonthlyRevenue", "money"),
        ("MONTHLY PROFIT", "=MonthlyProfit", "money"),
        ("ROAD SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "ROAD HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Profit per mile healthy", "=IFERROR(MIN(ProfitPerMile/PPMGoal,1),0)"),
        ("Net margin healthy", "=IFERROR(MIN(NetMargin/MarginGoal,1),0)"),
        ("Deadhead under control", "=IF(DeadheadPct<=DeadheadGoal,1,IFERROR(DeadheadGoal/DeadheadPct,0))"),
        ("Rate well above cost", "=IFERROR(MIN(CoverRatio/CoverGoal,1),0)"),
        ("Fuel economy on target", "=IFERROR(MIN(MPG/MPGGoal,1),0)"),
        ("Maintenance reserve funded", "=IFERROR(MIN(Reserve/ReserveGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=14, max_row=13 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=14, max_row=13 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Trucking Owner-Operator Command Center™ — know your cost per mile.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_cpm(wb); build_fixed(wb)
    build_variable(wb); build_loads(wb); build_settlements(wb); build_fuel(wb)
    build_maintenance(wb); build_equipment(wb); build_ifta(wb); build_reserve(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Cost Per Mile", "Fixed Costs", "Variable Costs", "Loads",
             "Settlements", "Fuel Log", "Maintenance", "Truck & Trailer", "IFTA & Miles",
             "Reserve Fund", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Trucking_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
