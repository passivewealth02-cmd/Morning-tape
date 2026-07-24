"""Build Home Renovation & Remodel Budget Command Center™ — The Remodel Operating System.

14 tabs · a premium home-renovation & remodel-budget operating system in Google Sheets
& Excel. Dashboard, a budget-vs-actual engine (by room, with a contingency reserve),
rooms, line items, contractors, payments, change orders, materials, a timeline,
financing, decisions and a monthly summary — one dashboard. Know what it costs, what
you've spent, and what's left — before it's a surprise.

Run: python3 build_xlsx.py   ->  ../Reno_Budget_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
ROOM_LIST = ["Kitchen", "Primary Bath", "Living Room", "Exterior & Paint", "Landscaping", "Whole Home"]
CATEGORY = ["Materials", "Labor", "Permits", "Fixtures", "Appliances", "Other"]
STATUS = ["Not started", "In progress", "Done", "On hold"]

ROOM_GOAL = 5
OUTSTANDING_GOAL = 5000
CHANGE_GOAL = 1200
CONTINGENCY_RATE = 0.15

CONTINGENCY_USED_N = 2000
PCT_COMPLETE_N = 0.65

# Rooms: (room, budget, spent)
ROOMS = [
    ("Kitchen", 30000, 22000), ("Primary Bath", 15000, 11000), ("Living Room", 8000, 5000),
    ("Exterior & Paint", 5000, 3000), ("Landscaping", 2000, 1000),
]

# Line items: (item, room, category, budget, actual)
LINE_ITEMS = [
    ("Cabinets", "Kitchen", "Materials", 9000, 8800), ("Countertops", "Kitchen", "Materials", 4500, 4500),
    ("Appliances", "Kitchen", "Appliances", 6000, 5200), ("Tile & floors", "Primary Bath", "Materials", 3500, 3200),
    ("Vanity", "Primary Bath", "Fixtures", 2200, 2100), ("Paint", "Living Room", "Labor", 1800, 1700),
    ("Lighting", "Kitchen", "Fixtures", 2000, 1900), ("Windows", "Exterior & Paint", "Materials", 3000, 2900),
]

# Contractors: (contractor, trade, bid, phone)
CONTRACTORS = [
    ("Cedar & Stone GC", "General", 42000, "555-0100"), ("Bright Spark Electric", "Electrical", 6000, "555-0110"),
    ("FlowRight Plumbing", "Plumbing", 5500, "555-0120"), ("TrueLine Painting", "Painting", 4200, "555-0130"),
    ("Stoneworks Tile", "Tile", 3800, "555-0140"),
]

# Payments: (date, paid to, amount)
PAYMENTS = [
    ("May 1", "GC deposit", 10000), ("Jun 1", "Cabinets", 8000), ("Jun 15", "Plumber", 4000),
    ("Jul 1", "Electrician", 3000), ("Jul 10", "GC draw 2", 9000), ("Jul 20", "Painter", 4000),
]

# Change orders: (item, room, amount)
CHANGE_ORDERS = [
    ("Move gas line", "Kitchen", 1200), ("Extra tile", "Primary Bath", 800),
    ("Add outlet", "Living Room", 400), ("Upgrade faucet", "Kitchen", 600),
]

# Materials & finishes: (item, room, selection, cost)
MATERIALS = [
    ("Cabinets", "Kitchen", "Shaker white", 8800), ("Countertop", "Kitchen", "Quartz", 4500),
    ("Backsplash", "Kitchen", "Subway tile", 700), ("Floor tile", "Primary Bath", "Hex marble", 1200),
    ("Vanity", "Primary Bath", "Walnut 48in", 2100), ("Paint", "Living Room", "Sage green", 300),
]

# Timeline: (phase, status, % done)
TIMELINE = [
    ("Demo", "Done", 1.0), ("Rough-in (elec/plumb)", "Done", 1.0), ("Cabinets & counters", "In progress", 0.7),
    ("Tile & floors", "In progress", 0.5), ("Paint & finish", "Not started", 0.0), ("Final & punch list", "Not started", 0.0),
]

# Financing: (source, amount)
FINANCING = [
    ("Cash / savings", 25000), ("HELOC", 30000), ("0% card", 5000),
]

# Decisions: (decision, room, choice, decided?)
DECISIONS = [
    ("Cabinet color", "Kitchen", "Shaker white", "Yes"), ("Countertop", "Kitchen", "Quartz", "Yes"),
    ("Tile", "Primary Bath", "Hex marble", "Yes"), ("Paint", "Living Room", "Sage green", "No"),
    ("Lighting", "Kitchen", "Pendants", "No"),
]

# Monthly summary: (month, spent)
MONTHS = [("Feb", 4000), ("Mar", 9000), ("Apr", 7000), ("May", 8000), ("Jun", 8000), ("Jul", 6000)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your goals & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR GOALS", "section")
    controls = [
        ("Project name", "Cedar & Stone Reno", None, "Project"),
        ("Homeowner", "Reese", None, "Owner"),
        ("Contingency rate", CONTINGENCY_RATE, "0%", "ContingencyRate"),
        ("Rooms-scoped goal", ROOM_GOAL, "0", "RoomGoal"),
        ("Outstanding goal (max)", OUTSTANDING_GOAL, '"$"#,##0', "OutstandingGoal"),
        ("Change-orders goal (max)", CHANGE_GOAL, '"$"#,##0', "ChangeGoal"),
        ("Contingency used", CONTINGENCY_USED_N, '"$"#,##0', "ContingencyUsed"),
        ("Project % complete", PCT_COMPLETE_N, "0%", "PctComplete"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Room", ROOM_LIST, "RoomBank"), ("F", "Category", CATEGORY, "CategoryList"),
             ("G", "Status", STATUS, "StatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🔨  HOME RENOVATION & REMODEL BUDGET COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know what it costs, what you've spent, and what's left — before it's a surprise.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE REMODEL, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A renovation goes over budget in the gaps — the change order nobody tracked, the room that crept, "
                      "the contingency that vanished. This closes them: a budget-vs-actual engine tracks every room "
                      "against its budget, holds a contingency reserve, and shows your remaining money and your variance "
                      "live. Track line items, contractors, payments, change orders, materials and your timeline — all "
                      "in ONE premium Google Sheets & Excel system built for a remodel.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your contingency rate & goals.",
             "2.  Enter each Room's budget in the Rooms tab.",
             "3.  Log Line Items, Contractors and Payments as they come.",
             "4.  Track Change Orders — the #1 cause of overruns.",
             "5.  Pick your Materials and run the Timeline & Financing.",
             "6.  Check the Dashboard: spent, remaining & a Reno Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional remodel (Cedar & Stone, homeowner Reese) is included so you can see how it "
               "all connects — just type over it with your own rooms and numbers. Your remaining budget and your "
               "contingency reserve are the two numbers that decide whether a remodel finishes on budget, and they roll "
               "into a live Reno Score. Twelve matching printable pages (budget worksheet, room list, contractor list, "
               "payment log & more) are included. This is a budgeting & organizing tool, not construction, financial or "
               "legal advice — confirm figures with your own contractors and advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "The budget you track is the budget you keep. Log the change orders.", "section_gold")


# ===========================================================================
def build_rooms(wb):
    ws = wb.create_sheet("Rooms"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 16, 16, 12, 2])
    luxe_header(ws, "F", "🏠  ROOMS",
                "Every room against its budget — spent, remaining and % used. This is where overruns hide.")
    table_headers(ws, 4, ["Room", "Budget", "Spent", "Remaining", "% Used"], start_col=2)
    start = L0
    for i, (room, budget, spent) in enumerate(ROOMS):
        r = start + i
        ws.cell(row=r, column=2, value=room).style = "td_left"
        cb = ws.cell(row=r, column=3, value=budget); cb.style = "input"; cb.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=spent); cs.style = "input"; cs.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0'
        cu = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/C{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(ROOMS) - 1
    nrange(wb, "RoomName", "Rooms", "B", start, end)
    nrange(wb, "RoomBudget", "Rooms", "C", start, end)
    nrange(wb, "RoomSpent", "Rooms", "D", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        ColorScaleRule(start_type="num", start_value=0.5, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.85, mid_color="FFFFF3CD",
                       end_type="num", end_value=1.0, end_color="FF" + RED_BG))
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL PROJECT").style = "th"
    cb = ws.cell(row=tot, column=3, value="=SUM(RoomBudget)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.fill = fill(SURFACE); cb.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value="=SUM(RoomSpent)"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cr = ws.cell(row=tot, column=5, value="=C" + str(tot) + "-D" + str(tot)); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0'
    cu = ws.cell(row=tot, column=6, value=f"=IFERROR(D{tot}/C{tot},0)"); cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(MINT_BG); cu.number_format = "0%"
    cell_name(wb, "TotalBudget", "Rooms", f"$C${tot}")
    cell_name(wb, "TotalSpent", "Rooms", f"$D${tot}")
    cell_name(wb, "Remaining", "Rooms", f"$E${tot}")
    ws.cell(row=tot + 2, column=2, value="Rooms on budget").style = "field_label"
    cru = ws.cell(row=tot + 2, column=3, value="=SUMPRODUCT((RoomSpent<=RoomBudget)*1)"); cru.style = "field_value"; cru.number_format = "#,##0"
    cell_name(wb, "RoomsUnder", "Rooms", f"$C${tot+2}")
    ws.cell(row=tot + 3, column=2, value="Rooms scoped").style = "field_label"
    crc = ws.cell(row=tot + 3, column=3, value="=COUNTA(RoomName)"); crc.style = "field_value"; crc.number_format = "#,##0"
    cell_name(wb, "RoomCount", "Rooms", f"$C${tot+3}")
    ws.freeze_panes = "A5"


def build_budgetactual(wb):
    ws = wb.create_sheet("Budget vs Actual"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 2])
    luxe_header(ws, "C", "📊  BUDGET VS ACTUAL",
                "Your whole project — budget, spent, remaining, and a contingency reserve so surprises don't sink you.")
    rows = [
        ("Total project budget", "=TotalBudget", '"$"#,##0', MINT_BG),
        ("− Spent to date", "=TotalSpent", '"$"#,##0', None),
    ]
    for i, (lab, val, fmt, bg) in enumerate(rows):
        r = 5 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "field_value"; c.number_format = fmt
        if bg:
            c.fill = fill(bg)
    ws.cell(row=7, column=2, value="= REMAINING BUDGET").style = "th"
    cr = ws.cell(row=7, column=3, value="=TotalBudget-TotalSpent"); cr.style = "td"; cr.font = Font(bold=True, size=14, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0'
    ws.cell(row=8, column=2, value="= BUDGET USED").style = "th"
    cu = ws.cell(row=8, column=3, value="=IFERROR(TotalSpent/TotalBudget,0)"); cu.style = "td"; cu.font = Font(bold=True, size=13, color=PRIMARY); cu.fill = fill(MINT_BG); cu.number_format = "0%"
    cell_name(wb, "BudgetUsed", "Budget vs Actual", "$C$8")
    ws.cell(row=10, column=2, value="CONTINGENCY RESERVE", ).style = "section_gold"
    ws.cell(row=11, column=2, value="Contingency reserve (rate × budget)").style = "field_label"
    ccr = ws.cell(row=11, column=3, value="=ContingencyRate*TotalBudget"); ccr.style = "field_value"; ccr.number_format = '"$"#,##0'
    cell_name(wb, "ContingencyReserve", "Budget vs Actual", "$C$11")
    ws.cell(row=12, column=2, value="− Contingency used").style = "field_label"
    ccu = ws.cell(row=12, column=3, value="=ContingencyUsed"); ccu.style = "field_value"; ccu.number_format = '"$"#,##0'
    ws.cell(row=13, column=2, value="= CONTINGENCY LEFT").style = "th"
    ccl = ws.cell(row=13, column=3, value="=ContingencyRate*TotalBudget-ContingencyUsed"); ccl.style = "td"; ccl.font = Font(bold=True, size=13, color=PRIMARY); ccl.fill = fill(MINT_BG); ccl.number_format = '"$"#,##0'
    cell_name(wb, "ContingencyLeft", "Budget vs Actual", "$C$13")
    ws.cell(row=15, column=2, value="FLAGSHIP ROOM — KITCHEN").style = "section_gold"
    ws.cell(row=16, column=2, value="Kitchen budget").style = "field_label"
    ck = ws.cell(row=16, column=3, value="=INDEX(RoomBudget,1)"); ck.style = "field_value"; ck.number_format = '"$"#,##0'
    ws.cell(row=17, column=2, value="Kitchen spent").style = "field_label"
    cks = ws.cell(row=17, column=3, value="=INDEX(RoomSpent,1)"); cks.style = "field_value"; cks.number_format = '"$"#,##0'
    ws.cell(row=18, column=2, value="= Kitchen variance (remaining)").style = "th"
    ckv = ws.cell(row=18, column=3, value="=INDEX(RoomBudget,1)-INDEX(RoomSpent,1)"); ckv.style = "td"; ckv.font = Font(bold=True, color=PRIMARY); ckv.fill = fill(MINT_BG); ckv.number_format = '"$"#,##0'
    ws.cell(row=20, column=2, value="Remaining budget and contingency left are the two numbers to watch.").style = "section_gold"


def build_lineitems(wb):
    ws, start, end = build_log(
        wb, "Line Items", "📋", "LINE ITEMS",
        "Every line of the job — by room and category, budget vs actual, so nothing slips through.",
        ["Item", "Room", "Category", "Budget", "Actual"],
        LINE_ITEMS, [2, 22, 18, 16, 14, 14, 2], text_left={2}, money={5, 6}, reserved=34, start_col=2,
        validations=[("C", "RoomBank"), ("D", "CategoryList")])
    nrange(wb, "LineBudget", "Line Items", "E", start, end)
    nrange(wb, "LineActual", "Line Items", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTALS").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cb = ws.cell(row=tot, column=5, value="=SUM(LineBudget)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.fill = fill(SURFACE); cb.number_format = '"$"#,##0'
    ca = ws.cell(row=tot, column=6, value="=SUM(LineActual)"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = '"$"#,##0'


def build_contractors(wb):
    ws, start, end = build_log(
        wb, "Contractors", "👷", "CONTRACTORS",
        "Who's doing what, for how much — bids and contacts in one place.",
        ["Contractor", "Trade", "Bid", "Phone"],
        CONTRACTORS, [2, 24, 16, 14, 16, 2], text_left={2, 3, 5}, money={4}, reserved=24, start_col=2)
    nrange(wb, "Bid", "Contractors", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL BIDS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(Bid)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)


def build_payments(wb):
    ws, start, end = build_log(
        wb, "Payments", "💳", "PAYMENTS",
        "Every draw and payment made — so you always know what's paid and what's still owed.",
        ["Date", "Paid To", "Amount"],
        PAYMENTS, [2, 14, 26, 16, 2], text_left={2, 3}, money={4}, reserved=30, start_col=2)
    nrange(wb, "PayAmt", "Payments", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="PAID TO DATE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(PayAmt)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    cell_name(wb, "PaidToDate", "Payments", f"$D${tot}")
    ws.cell(row=tot + 1, column=2, value="Outstanding (committed − paid)").style = "field_label"
    co = ws.cell(row=tot + 1, column=4, value="=TotalSpent-PaidToDate"); co.style = "field_value"; co.number_format = '"$"#,##0'; co.fill = fill(WARN_BG)
    cell_name(wb, "Outstanding", "Payments", f"$D${tot+1}")


def build_changeorders(wb):
    ws, start, end = build_log(
        wb, "Change Orders", "🔧", "CHANGE ORDERS",
        "Every mid-project change and what it added — the #1 reason remodels blow the budget.",
        ["Change", "Room", "Amount"],
        CHANGE_ORDERS, [2, 26, 18, 16, 2], text_left={2}, money={4}, reserved=26, start_col=2,
        validations=[("C", "RoomBank")])
    nrange(wb, "ChangeAmt", "Change Orders", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL CHANGE ORDERS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(ChangeAmt)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "ChangeOrders", "Change Orders", f"$D${tot}")


def build_materials(wb):
    ws, start, end = build_log(
        wb, "Materials", "🧱", "MATERIALS & FINISHES",
        "Every selection — the finish, the room and the cost, so your choices and your budget stay in sync.",
        ["Item", "Room", "Selection", "Cost"],
        MATERIALS, [2, 20, 16, 22, 14, 2], text_left={2, 4}, money={5}, reserved=26, start_col=2,
        validations=[("C", "RoomBank")])
    nrange(wb, "MatCost", "Materials", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL SELECTIONS").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(MatCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'


def build_timeline(wb):
    ws = wb.create_sheet("Timeline"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 14, 2])
    luxe_header(ws, "D", "📅  TIMELINE",
                "Every phase and how far along it is — so you always know how close the finish line really is.")
    ws.cell(row=5, column=2, value="Project % complete").style = "field_label"
    cp = ws.cell(row=5, column=3, value="=PctComplete"); cp.style = "field_value"; cp.number_format = "0%"; cp.fill = fill(MINT_BG)
    table_headers(ws, 7, ["Phase", "Status", "% Done"], start_col=2)
    start = 8
    for i, (phase, status, done) in enumerate(TIMELINE):
        r = start + i
        ws.cell(row=r, column=2, value=phase).style = "td_left"
        ws.cell(row=r, column=3, value=status).style = "td"
        cd = ws.cell(row=r, column=4, value=done); cd.style = "input"; cd.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(TIMELINE) - 1
    add_dv(ws, f"C{start}:C{end}", "StatusList")
    ws.conditional_formatting.add(f"D{start}:D{end}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT))
    ws.freeze_panes = "A8"


def build_financing(wb):
    ws = wb.create_sheet("Financing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 2])
    luxe_header(ws, "C", "🏦  FINANCING",
                "How the project is paid for — cash, HELOC, cards — so your funding covers the budget.")
    table_headers(ws, 4, ["Source", "Amount"], start_col=2)
    start = L0
    for i, (src, amt) in enumerate(FINANCING):
        r = start + i
        ws.cell(row=r, column=2, value=src).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FINANCING) - 1
    nrange(wb, "FundAmt", "Financing", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL FUNDING").style = "th"
    cf = ws.cell(row=tot, column=3, value="=SUM(FundAmt)"); cf.style = "td"; cf.font = Font(bold=True, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = '"$"#,##0'
    ws.cell(row=tot + 2, column=2, value="Funding vs budget").style = "field_label"
    cv = ws.cell(row=tot + 2, column=3, value="=SUM(FundAmt)-TotalBudget"); cv.style = "field_value"; cv.number_format = '"$"#,##0'; cv.fill = fill(MINT_BG)
    ws.freeze_panes = "A5"


def build_decisions(wb):
    ws, start, end = build_log(
        wb, "Decisions", "✅", "DECISIONS",
        "Every choice to make — the room, the pick and whether it's locked. Decision paralysis stalls remodels.",
        ["Decision", "Room", "Choice", "Decided?"],
        DECISIONS, [2, 22, 16, 20, 14, 2], text_left={2, 4}, reserved=26, start_col=2,
        validations=[("C", "RoomBank"), ("E", "YesNoList")])
    nrange(wb, "Decided", "Decisions", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="DECIDED").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value='=COUNTIF(Decided,"Yes")&" / "&COUNTA(Decided)'); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG)


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "What you spent each month — watch the burn rate against your remaining budget.")
    ws.cell(row=5, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 6, ["Month", "Spent"], start_col=2)
    ts = 7
    for i, (m, sp) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sp); cs.style = "input"; cs.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "SpendTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Spent by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🔨  HOME RENOVATION & REMODEL BUDGET COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Budget, spent, remaining, contingency & a Reno Score — your whole remodel, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("TOTAL BUDGET", "=TotalBudget", "money"),
        ("TOTAL SPENT", "=TotalSpent", "money"),
        ("REMAINING", "=Remaining", "money"),
        ("BUDGET USED", "=BudgetUsed", "pct"),
        ("CONTINGENCY", "=ContingencyReserve", "money"),
        ("CONTINGENCY LEFT", "=ContingencyLeft", "money"),
    ]
    row2 = [
        ("ROOMS", "=RoomCount", "num"),
        ("PAID TO DATE", "=PaidToDate", "money"),
        ("OUTSTANDING", "=Outstanding", "money"),
        ("CHANGE ORDERS", "=ChangeOrders", "money"),
        ("% COMPLETE", "=PctComplete", "pct"),
        ("RENO SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "RENO HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SPENT BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Under total budget", "=IF(TotalSpent<=TotalBudget,1,IFERROR(TotalBudget/TotalSpent,0))"),
        ("Every room on budget", "=IFERROR(RoomsUnder/RoomCount,0)"),
        ("Contingency healthy", "=IFERROR(MIN((ContingencyLeft/ContingencyReserve)/0.5,1),0)"),
        ("Scope defined", "=IFERROR(MIN(RoomCount/RoomGoal,1),0)"),
        ("Payments current", "=IF(Outstanding<=OutstandingGoal,1,IFERROR(OutstandingGoal/Outstanding,0))"),
        ("Change orders in check", "=IFERROR(MIN(ChangeGoal/ChangeOrders,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Spent by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=7, max_row=6 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=7, max_row=6 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Home Renovation & Remodel Budget Command Center™ — on budget, on schedule, no surprises.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_rooms(wb); build_budgetactual(wb)
    build_lineitems(wb); build_contractors(wb); build_payments(wb); build_changeorders(wb)
    build_materials(wb); build_timeline(wb); build_financing(wb); build_decisions(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Budget vs Actual", "Rooms", "Line Items", "Contractors", "Payments",
             "Change Orders", "Materials", "Timeline", "Financing", "Decisions", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Reno_Budget_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
