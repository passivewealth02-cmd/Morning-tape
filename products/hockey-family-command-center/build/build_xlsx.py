"""Build Hockey Family Command Center™ — The Ultimate Hockey Season OS.

25 sheets + Welcome · a premium operating system for hockey families:
schedule, budget, equipment, travel, stats, development, nutrition & team.

Run: python3 build_xlsx.py   ->  ../Hockey_Family_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"
ACCENT = "937356"
GOLD_LT = "C9A86A"
SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"
MINT_BG = "E3F8EF"
WHITE = "FFFFFF"
TEXT = "333333"
DANGER = "C94C4C"
RED_BG = "FBE6E6"
WARN_BG = "FBF0E2"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"
IVORY = "FBF8F2"

# ---- dropdown vocabularies ----
EVENT_TYPES = ["Practice", "Game", "Tournament", "Dryland", "Skills",
               "Team Meeting", "School", "Family", "Holiday"]
LEAGUES = ["House League", "Rep / Travel", "AA", "AAA", "Spring", "Summer"]
DIVISIONS = ["U7", "U9", "U11", "U13", "U15", "U18"]
POSITIONS = ["Center", "Left Wing", "Right Wing", "Defense", "Goalie"]
EXPENSE_CATS = ["Registration", "Team Fees", "Ice Fees", "Tournaments", "Hotels",
                "Flights", "Fuel", "Meals", "Skates", "Helmets", "Sticks",
                "Gloves", "Pants", "Shoulder Pads", "Elbow Pads", "Shin Guards",
                "Neck Guard", "Mouth Guard", "Skate Sharpening", "Tape", "Wax",
                "Training", "Private Lessons", "Photography", "Team Apparel",
                "Miscellaneous"]
EXPENSE_GROUPS = ["Fees & Registration", "Travel", "Equipment", "Maintenance",
                  "Training", "Extras"]
EQUIP_STATUS = ["New", "Good", "Fair", "Replace Soon", "Replace Now"]
SKILL_LEVELS = ["Developing", "Average", "Strong", "Elite"]
TOURNEY_TYPES = ["League", "Showcase", "Championship", "Friendly", "Spring"]
TRAVEL_STATUS = ["Planned", "Booked", "Paid", "Completed"]
PRACTICE_TYPES = ["On-Ice", "Dryland", "Skills", "Power Skating", "Goalie", "Team"]
ATTEND = ["Present", "Absent", "Excused"]
SHOOTS = ["Left", "Right"]
YESNO = ["Yes", "No"]

# map each expense category -> group (for dashboard roll-ups)
CAT_GROUP = {
    "Registration": "Fees & Registration", "Team Fees": "Fees & Registration",
    "Ice Fees": "Fees & Registration", "Tournaments": "Fees & Registration",
    "Hotels": "Travel", "Flights": "Travel", "Fuel": "Travel", "Meals": "Travel",
    "Skates": "Equipment", "Helmets": "Equipment", "Sticks": "Equipment",
    "Gloves": "Equipment", "Pants": "Equipment", "Shoulder Pads": "Equipment",
    "Elbow Pads": "Equipment", "Shin Guards": "Equipment", "Neck Guard": "Equipment",
    "Mouth Guard": "Equipment", "Skate Sharpening": "Maintenance", "Tape": "Maintenance",
    "Wax": "Maintenance", "Training": "Training", "Private Lessons": "Training",
    "Photography": "Extras", "Team Apparel": "Extras", "Miscellaneous": "Extras",
}

LOG_ROWS = 40
L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]


# ===========================================================================
# Styles & shared helpers (premium pattern)
# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"),
                            fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True),
                               fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY),
                              alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT),
                                   alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"),
                         fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT),
                              alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
                              border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY),
                            fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True),
                             fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "prompt": NamedStyle(name="prompt", font=f(11, False, TEXT, italic=True),
                             alignment=Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True),
                             border=BOX, fill=PatternFill("solid", fgColor=IVORY)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT),
                           alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None,
               dates=None, pcts=None, start_col=1):
    text_left = text_left or set()
    money = money or set()
    ints = ints or set()
    dates = dates or set()
    pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}")
    ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label
    lc.font = Font(size=9, bold=True, color=ACCENT)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vc.font = Font(size=18, bold=True, color=PRIMARY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pct": "0%",
                        "days": "0", "dec": "0.0"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc)
            c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN,
                              top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers),
               text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def no_labels():
    dl = DataLabelList()
    dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 3] + [17] * 7)
    luxe_header(ws, "K", "⚙  SETTINGS", "Set your season inputs once — every dashboard follows. Edit the lists to fit your program.")

    merge_set(ws, "B5:C5", "SEASON INPUTS", "section")
    controls = [
        ("Player Name", "Jordan Miller", None, "PlayerName"),
        ("Team", "Lightning U13 AA", None, "TeamName"),
        ("Monthly Hockey Budget", 1200, '"$"#,##0', "MonthlyBudget"),
        ("Season Start", dminus(60), "mm/dd/yyyy", "SeasonStart"),
        ("Season End", dplus(150), "mm/dd/yyyy", "SeasonEnd"),
        ("Home Arena", "Riverside Ice Centre", None, "HomeArena"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")

    bank1 = [("E", "Event Type", EVENT_TYPES, "EventTypeList"),
             ("F", "League", LEAGUES, "LeagueList"),
             ("G", "Age Division", DIVISIONS, "DivisionList"),
             ("H", "Position", POSITIONS, "PositionList"),
             ("I", "Expense Category", EXPENSE_CATS, "ExpenseCatList"),
             ("J", "Equipment Status", EQUIP_STATUS, "EquipStatusList"),
             ("K", "Skill Level", SKILL_LEVELS, "SkillLevelList")]
    _emit_lists(wb, ws, bank1, 5, 6)

    bank2 = [("E", "Tournament Type", TOURNEY_TYPES, "TourneyTypeList"),
             ("F", "Travel Status", TRAVEL_STATUS, "TravelStatusList"),
             ("G", "Practice Type", PRACTICE_TYPES, "PracticeTypeList"),
             ("H", "Attendance", ATTEND, "AttendList"),
             ("I", "Shoots", SHOOTS, "ShootsList"),
             ("J", "Yes / No", YESNO, "YesNoList"),
             ("K", "Expense Group", EXPENSE_GROUPS, "ExpenseGroupList")]
    _emit_lists(wb, ws, bank2, 30, 31)


def _emit_lists(wb, ws, lists, header_row, data_row):
    merge_set(ws, f"E{header_row}:K{header_row}", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in lists:
        ci = column_index_from_string(col)
        ws.cell(row=header_row + 1, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=data_row + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(
            nm, attr_text=f"Settings!${col}${data_row + 1}:${col}${data_row + len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 74, 3])
    ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🏒  HOCKEY FAMILY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate hockey season operating system — your whole season in one place.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    merge_set(ws, "B5:B5", "WELCOME TO THE SEASON", "section_gold")
    ws.merge_cells("B6:B8")
    ws["B6"].value = (
        "Hockey season means practices, games, tournaments, travel, equipment, and a "
        "budget that adds up fast. Hockey Family Command Center™ brings all of it into "
        "one elegant dashboard so nothing gets missed, costs stay under control, and the "
        "whole family runs on a system instead of from memory.")
    ws["B6"].style = "body"
    for r in (6, 7, 8):
        ws.row_dimensions[r].height = 22

    merge_set(ws, "B10:B10", "HOW TO USE IT", "section")
    steps = [
        "1.  Open Settings and add your player, team, budget & season dates.",
        "2.  Fill the Player Profile, Season Calendar, and Hockey Budget.",
        "3.  Log practices, games & stats; track equipment and sharpening.",
        "4.  Plan tournaments, travel, hotels & carpools in advance.",
        "5.  Watch the Executive Hockey Dashboard update itself automatically.",
    ]
    for i, s in enumerate(steps):
        r = 11 + i
        ws.merge_cells(f"B{r}:B{r}")
        ws[f"B{r}"].value = s
        ws[f"B{r}"].style = "body"
        ws.row_dimensions[r].height = 22

    dr = 18
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th")
    ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = (
        "Sample data is included to show how everything connects — just type over it with "
        "your own. Every sheet is print-friendly and works in Excel and Google Sheets, on "
        "desktop and mobile. Countdowns, alerts and totals all update automatically.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT)
    c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22
        ws.cell(row=rr, column=2).fill = fill(WARN_BG)

    merge_set(ws, f"B{dr+5}:B{dr+5}",
              "One organized season, less stress, more hockey — let's drop the puck.", "section_gold")


# ===========================================================================
# 2 — Player Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Player Profile")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 26, 6, 24, 26, 2])
    luxe_header(ws, "G", "🏒  PLAYER PROFILE",
                "Everything about your player in one place — stats card, contacts & medical.")
    blocks = [
        ("PLAYER", [("Player Name", "=PlayerName"), ("Date of Birth", dt.date(2012, 3, 14)),
                    ("Age Division", "U13"), ("Team", "=TeamName"), ("League", "AA"),
                    ("Jersey Number", 19), ("Position", "Center"), ("Shoots", "Left"),
                    ("Height", "5'2\""), ("Weight", "105 lb")]),
        ("TEAM & STAFF", [("Head Coach", "Coach Daniels"), ("Assistant Coach", "Coach Roy"),
                          ("Team Manager", "Sarah P."), ("Home Arena", "=HomeArena"),
                          ("Practice Night", "Tue / Thu"), ("Team Colors", "Navy / Gold")]),
        ("MEDICAL & EMERGENCY", [("Emergency Contact", "Mom — (555) 210-7788"),
                                 ("Secondary Contact", "Dad — (555) 210-4521"),
                                 ("Health Card #", "____________"), ("Allergies", "None"),
                                 ("Medical Notes", "Mild asthma — inhaler in bag"),
                                 ("Insurance Provider", "____________")]),
        ("SEASON GOALS", [("Primary Goal", "Make AA top line"), ("Skating", "Improve edges & speed"),
                          ("Scoring", "30+ points"), ("Off-Ice", "3x dryland / week")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold")
        ws.row_dimensions[row].height = 22
        row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            cv = ws.cell(row=row, column=3, value=fields[i][1]); cv.style = "field_value"
            if isinstance(fields[i][1], dt.date):
                cv.number_format = "mm/dd/yyyy"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24
            i += 2
            row += 1
        row += 1


# ===========================================================================
# 3 — Master Season Calendar
# ===========================================================================
def build_calendar(wb):
    sample = [
        ("Practice", "On-ice skills", dplus(1), "6:00 PM", "Riverside Ice", "Bring full gear"),
        ("Game", "vs Thunder", dplus(3), "10:30 AM", "Centennial Arena", "Away — white jersey"),
        ("Dryland", "Speed & agility", dplus(4), "5:00 PM", "Field House", ""),
        ("Practice", "Power skating", dplus(6), "6:00 PM", "Riverside Ice", ""),
        ("Game", "vs Ice Hawks", dplus(8), "1:15 PM", "Riverside Ice", "Home — navy jersey"),
        ("Tournament", "Fall Classic", dplus(15), "All weekend", "Capital City", "3 nights — see Tournaments"),
        ("School", "Parent-teacher", dplus(8), "4:00 PM", "Maple School", "Conflict w/ game!"),
        ("Practice", "Systems", dplus(10), "6:00 PM", "Riverside Ice", ""),
        ("Game", "vs Rangers", dplus(12), "9:00 AM", "North Arena", "Away"),
        ("Skills", "Private lesson", dplus(5), "7:30 PM", "Skills Centre", "Coach Roy"),
        ("Team Meeting", "Tournament prep", dplus(13), "7:00 PM", "Zoom", ""),
        ("Family", "Cousin's birthday", dplus(15), "2:00 PM", "Home", "Conflict w/ tournament"),
        ("Practice", "Goalie + skaters", dplus(17), "6:00 PM", "Riverside Ice", ""),
        ("Game", "vs Flyers", dplus(19), "11:00 AM", "Riverside Ice", "Home"),
        ("Tournament", "Winter Showcase", dplus(45), "All weekend", "Lakeside", ""),
        ("Game", "vs Wolves (W 4-2)", dminus(4), "2:00 PM", "Riverside Ice", "Home — won"),
        ("Game", "vs Kings (L 1-3)", dminus(11), "5:30 PM", "East Arena", "Away"),
    ]
    ws, start, end = build_log(
        wb, "Calendar", "📅", "MASTER SEASON CALENDAR",
        "Every practice, game & tournament — with countdowns and conflict detection.",
        ["Type", "Event / Opponent", "Date", "Time", "Location", "Notes"],
        sample, [15, 26, 13, 13, 20, 24],
        text_left={2, 5, 6}, dates={3},
        validations=[("A", "EventTypeList")], reserved=70)
    nrange(wb, "CalType", "Calendar", "A", start, end)
    nrange(wb, "CalEvent", "Calendar", "B", start, end)
    nrange(wb, "CalDate", "Calendar", "C", start, end)
    # countdown + conflict helpers
    ws.cell(row=4, column=7, value="In").style = "th"
    ws.cell(row=4, column=8, value="Clash").style = "th"
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 8
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=7, value=f'=IF(C{r}="","",C{r}-TODAY())')
        c.style = "td"; c.number_format = "0;[Red]-0"
        c.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
        cf = ws.cell(row=r, column=8,
                     value=f'=IF(C{r}="","",IF(AND(COUNTIF($C${start}:C{r},C{r})=1,COUNTIF(CalDate,C{r})>1),1,0))')
        cf.style = "td"; cf.number_format = "0;;"
        cf.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
    nrange(wb, "CalConflict", "Calendar", "H", start, end)
    ws.conditional_formatting.add(
        f"A{start}:F{end}",
        FormulaRule(formula=[f'AND($C{start}<>"",COUNTIF(CalDate,$C{start})>1)'], fill=fill(WARN_BG)))
    ws.conditional_formatting.add(
        f"A{start}:F{end}",
        FormulaRule(formula=[f'AND($A{start}="Game",$C{start}>=TODAY(),$C{start}-TODAY()<=3)'], fill=fill(MINT_BG)))
    # summary block
    ws.cell(row=4, column=10, value="SEASON COUNTS").style = "section_gold"
    ws.column_dimensions["J"].width = 18
    ws.column_dimensions["K"].width = 10
    counts = [("Practices", '=COUNTIF(CalType,"Practice")'), ("Games", '=COUNTIF(CalType,"Game")'),
              ("Tournaments", '=COUNTIF(CalType,"Tournament")'), ("Other events",
               '=SUMPRODUCT((CalType<>"")*(CalType<>"Practice")*(CalType<>"Game")*(CalType<>"Tournament"))')]
    for i, (lab, fml) in enumerate(counts):
        r = 5 + i
        ws.cell(row=r, column=10, value=lab).style = "td_left"
        ws.cell(row=r, column=11, value=fml).style = "td"
    wb.defined_names["CalCountLabels"] = DefinedName("CalCountLabels", attr_text="Calendar!$J$5:$J$8")
    wb.defined_names["CalCountVals"] = DefinedName("CalCountVals", attr_text="Calendar!$K$5:$K$8")


# ===========================================================================
# 4 — Game Day Command Center
# ===========================================================================
def build_gameday(wb):
    build_log(
        wb, "Game Day", "🥅", "GAME DAY COMMAND CENTER",
        "Everything for game day — arena, times, jersey, refs & the final score.",
        ["Opponent", "Date", "Arena / Address", "Game Time", "Arrive By", "Warm-Up",
         "Jersey", "Result", "Score", "Player Notes"],
        [
            ("Thunder", dplus(3), "Centennial Arena, 12 Rink Rd", "10:30 AM", "9:30 AM", "10:00 AM", "White", "—", "", ""),
            ("Ice Hawks", dplus(8), "Riverside Ice (home)", "1:15 PM", "12:15 PM", "12:45 PM", "Navy", "—", "", ""),
            ("Rangers", dplus(12), "North Arena, 88 Cold St", "9:00 AM", "8:00 AM", "8:30 AM", "White", "—", "", ""),
            ("Flyers", dplus(19), "Riverside Ice (home)", "11:00 AM", "10:00 AM", "10:30 AM", "Navy", "—", "", ""),
            ("Wolves", dminus(4), "Riverside Ice (home)", "2:00 PM", "1:00 PM", "1:30 PM", "Navy", "W", "4-2", "1G 1A"),
            ("Kings", dminus(11), "East Arena, 5 Glacier Ave", "5:30 PM", "4:30 PM", "5:00 PM", "White", "L", "1-3", "1A"),
        ],
        [16, 13, 24, 12, 11, 11, 10, 9, 9, 18],
        text_left={1, 3, 10}, dates={2}, reserved=40)


# ===========================================================================
# 5 — Practice Tracker
# ===========================================================================
def build_practice(wb):
    sample = [
        (dminus(2), "On-Ice", "Riverside Ice", "Coach Daniels", "Edges & breakouts", "Present", 1.5, "Great compete"),
        (dminus(4), "Dryland", "Field House", "Coach Roy", "Speed & agility", "Present", 1.0, ""),
        (dminus(6), "On-Ice", "Riverside Ice", "Coach Daniels", "Power play", "Present", 1.5, ""),
        (dminus(9), "Skills", "Skills Centre", "Coach Roy", "Shooting", "Present", 1.0, "Improved release"),
        (dminus(11), "On-Ice", "Riverside Ice", "Coach Daniels", "Defensive zone", "Excused", 0, "Sick"),
        (dminus(13), "Power Skating", "Riverside Ice", "Ms. Lee", "Edge work", "Present", 1.0, ""),
        (dminus(16), "On-Ice", "Riverside Ice", "Coach Daniels", "Systems", "Present", 1.5, ""),
        (dminus(18), "Dryland", "Field House", "Coach Roy", "Strength", "Present", 1.0, ""),
        (dminus(20), "On-Ice", "Riverside Ice", "Coach Daniels", "Scrimmage", "Present", 1.5, "2 goals"),
        (dminus(23), "Goalie", "Riverside Ice", "Coach G", "—", "Absent", 0, "Skater"),
    ]
    ws, start, end = build_log(
        wb, "Practice", "🏒", "PRACTICE TRACKER",
        "Log every session — ice time and attendance roll up to your dashboard.",
        ["Date", "Type", "Arena", "Coach", "Focus Area", "Attendance", "Ice (hrs)", "Notes"],
        sample, [13, 14, 18, 16, 20, 13, 11, 22],
        text_left={3, 4, 5, 8}, dates={1},
        validations=[("B", "PracticeTypeList"), ("F", "AttendList")], reserved=60)
    nrange(wb, "PracDate", "Practice", "A", start, end)
    nrange(wb, "PracType", "Practice", "B", start, end)
    nrange(wb, "PracAttend", "Practice", "F", start, end)
    nrange(wb, "PracIce", "Practice", "G", start, end)
    for r in range(start, end + 1):
        ws.cell(row=r, column=7).number_format = "0.0"
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Present"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Absent"'], fill=fill(RED_BG)))


# ===========================================================================
# 6 — Hockey Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 18, 14, 14, 14, 12, 3, 20, 14])
    luxe_header(ws, "I", "💰  HOCKEY BUDGET",
                "Every season cost in one place — spending, remaining & cost-per-game, automatic.")
    table_headers(ws, 4, ["Category", "Group", "Season Budget", "Spent", "Remaining", "% Used"])
    planned = {
        "Registration": 650, "Team Fees": 1800, "Ice Fees": 900, "Tournaments": 1200,
        "Hotels": 1400, "Flights": 600, "Fuel": 700, "Meals": 800, "Skates": 450,
        "Helmets": 180, "Sticks": 600, "Gloves": 140, "Pants": 160, "Shoulder Pads": 130,
        "Elbow Pads": 70, "Shin Guards": 110, "Neck Guard": 30, "Mouth Guard": 25,
        "Skate Sharpening": 180, "Tape": 90, "Wax": 25, "Training": 700,
        "Private Lessons": 900, "Photography": 150, "Team Apparel": 220, "Miscellaneous": 200,
    }
    spent = {
        "Registration": 650, "Team Fees": 1800, "Ice Fees": 540, "Tournaments": 800,
        "Hotels": 820, "Flights": 0, "Fuel": 430, "Meals": 410, "Skates": 450,
        "Helmets": 180, "Sticks": 380, "Gloves": 140, "Pants": 0, "Shoulder Pads": 0,
        "Elbow Pads": 70, "Shin Guards": 0, "Neck Guard": 30, "Mouth Guard": 25,
        "Skate Sharpening": 95, "Tape": 55, "Wax": 25, "Training": 420,
        "Private Lessons": 480, "Photography": 0, "Team Apparel": 220, "Miscellaneous": 110,
    }
    start = L0
    end = start + len(EXPENSE_CATS) - 1
    for i, cat in enumerate(EXPENSE_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        ws.cell(row=r, column=2, value=CAT_GROUP[cat]).style = "td_left"
        cp = ws.cell(row=r, column=3, value=planned[cat]); cp.style = "input"; cp.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=spent[cat]); cs.style = "input"; cs.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/C{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    ws.cell(row=total, column=2, value="").style = "th"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=total, column=6, value=f"=IFERROR(D{total}/C{total},0)")
    cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.2, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))
    nrange(wb, "BudgetCat", "Budget", "A", start, end)
    nrange(wb, "BudgetGroup", "Budget", "B", start, end)
    nrange(wb, "BudgetPlanned", "Budget", "C", start, end)
    nrange(wb, "BudgetActual", "Budget", "D", start, end)
    wb.defined_names["BudgetTotalPlanned"] = DefinedName("BudgetTotalPlanned", attr_text=f"Budget!$C${total}")
    wb.defined_names["BudgetTotalActual"] = DefinedName("BudgetTotalActual", attr_text=f"Budget!$D${total}")

    # KPI sidebar
    ws.cell(row=4, column=8, value="SEASON TO DATE").style = "section_gold"
    kpis = [("Total Budget", "=BudgetTotalPlanned", '"$"#,##0'),
            ("Total Spent", "=BudgetTotalActual", '"$"#,##0'),
            ("Remaining", "=BudgetTotalPlanned-BudgetTotalActual", '"$"#,##0;[Red]-"$"#,##0'),
            ("Cost / Game", '=IFERROR(BudgetTotalActual/MAX(COUNTIF(CalType,"Game"),1),0)', '"$"#,##0'),
            ("Cost / Practice", '=IFERROR(BudgetTotalActual/MAX(COUNTIF(CalType,"Practice"),1),0)', '"$"#,##0'),
            ("Cost / Tournament", '=IFERROR(BudgetTotalActual/MAX(COUNTIF(CalType,"Tournament"),1),0)', '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=8, value=lab).style = "field_label"
        c = ws.cell(row=r, column=9, value=fml); c.style = "field_value"; c.number_format = fmt

    # group summary (for charts)
    gstart = total + 3
    merge_set(ws, f"A{gstart-1}:F{gstart-1}", "SPENDING BY GROUP", "section_gold")
    ws.cell(row=gstart, column=1, value="Group").style = "th"
    ws.cell(row=gstart, column=2, value="Spent").style = "th"
    for i, g in enumerate(EXPENSE_GROUPS):
        r = gstart + 1 + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        c = ws.cell(row=r, column=2, value=f'=SUMIF(BudgetGroup,A{r},BudgetActual)')
        c.style = "td"; c.number_format = '"$"#,##0'
    gend = gstart + len(EXPENSE_GROUPS)
    wb.defined_names["GroupLabels"] = DefinedName("GroupLabels", attr_text=f"Budget!$A${gstart+1}:$A${gend}")
    wb.defined_names["GroupVals"] = DefinedName("GroupVals", attr_text=f"Budget!$B${gstart+1}:$B${gend}")
    donut = DoughnutChart(); donut.title = "Spending by Group"; donut.height = 8; donut.width = 12
    donut.add_data(Reference(ws, min_col=2, min_row=gstart, max_row=gend), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=1, min_row=gstart + 1, max_row=gend))
    donut.dataLabels = no_labels()
    ws.add_chart(donut, "H13")
    ws.freeze_panes = "A5"


# ===========================================================================
# 7 — Equipment Command Center
# ===========================================================================
def build_equipment(wb):
    sample = [
        ("Helmet", "Bauer", "M", dminus(220), 180, "Good", dplus(500), "Recert next yr"),
        ("Cage", "Bauer", "M", dminus(220), 45, "Good", dplus(500), ""),
        ("Skates", "CCM", "5.5", dminus(150), 450, "Replace Soon", dplus(40), "Feet growing"),
        ("Shoulder Pads", "CCM", "M", dminus(400), 130, "Good", dplus(200), ""),
        ("Elbow Pads", "Warrior", "M", dminus(400), 70, "Fair", dplus(60), ""),
        ("Gloves", "Bauer", "12\"", dminus(180), 140, "Good", dplus(300), ""),
        ("Pants", "CCM", "Jr L", dminus(400), 160, "Good", dplus(220), ""),
        ("Shin Guards", "Bauer", "13\"", dminus(150), 110, "Replace Soon", dplus(30), "Outgrowing"),
        ("Neck Guard", "Bauer", "Jr", dminus(150), 30, "Good", dplus(400), ""),
        ("Jock/Jill", "Shock Doctor", "M", dminus(300), 35, "Good", dplus(200), ""),
        ("Mouth Guard", "Shock Doctor", "Jr", dminus(60), 25, "Replace Soon", dplus(20), "Replace each season"),
        ("Game Jersey", "Team", "M", dminus(60), 90, "New", dplus(700), ""),
        ("Practice Jersey", "Team", "M", dminus(60), 40, "Good", dplus(500), ""),
        ("Water Bottle", "—", "—", dminus(60), 15, "Good", dplus(365), ""),
    ]
    ws, start, end = build_log(
        wb, "Equipment", "🎽", "EQUIPMENT COMMAND CENTER",
        "Track every piece — gear due for replacement in 45 days flags automatically.",
        ["Equipment", "Brand", "Size", "Purchased", "Price", "Condition", "Replace By", "Notes"],
        sample, [16, 14, 9, 13, 11, 14, 13, 22],
        text_left={1, 8}, dates={4, 7}, money={5},
        validations=[("F", "EquipStatusList")], reserved=40)
    nrange(wb, "EquipName", "Equipment", "A", start, end)
    nrange(wb, "EquipPrice", "Equipment", "E", start, end)
    nrange(wb, "EquipCond", "Equipment", "F", start, end)
    nrange(wb, "EquipReplace", "Equipment", "G", start, end)
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'AND($G{start}<>"",$G{start}<=TODAY()+45)'], fill=fill(WARN_BG)))
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Replace Now"'], fill=fill(RED_BG)))
    ws.conditional_formatting.add(
        f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"New"'], fill=fill(MINT_BG)))


# ===========================================================================
# 8 — Skate Sharpening Log
# ===========================================================================
def build_sharpening(wb):
    sample = [
        (dminus(4), "1/2\"", "Pro Shop", 8, 6.0, "Felt great"),
        (dminus(18), "1/2\"", "Pro Shop", 8, 7.5, ""),
        (dminus(32), "1/2\"", "Riverside Pro", 9, 7.0, "Bit deep"),
        (dminus(48), "5/8\"", "Pro Shop", 8, 8.0, "Tried shallower"),
        (dminus(63), "1/2\"", "Pro Shop", 8, 7.0, ""),
    ]
    ws, start, end = build_log(
        wb, "Sharpening", "⛸", "SKATE SHARPENING LOG",
        "Track hollow & ice hours — get a reminder when it's time to sharpen again.",
        ["Date", "Hollow", "Shop", "Cost", "Ice Hrs Since", "Notes"],
        sample, [13, 12, 18, 11, 16, 26],
        text_left={6}, dates={1}, money={4}, reserved=40)
    nrange(wb, "SharpDate", "Sharpening", "A", start, end)
    for r in range(start, end + 1):
        ws.cell(row=r, column=5).number_format = "0.0"
    # reminder: most recent sharpening older than 14 days
    ws.cell(row=4, column=8, value="SHARPENING STATUS").style = "section_gold"
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 12
    ws.cell(row=5, column=8, value="Last sharpened").style = "field_label"
    c = ws.cell(row=5, column=9, value="=IFERROR(MAX(SharpDate),0)"); c.style = "field_value"; c.number_format = "mm/dd/yyyy"
    ws.cell(row=6, column=8, value="Days ago").style = "field_label"
    ws.cell(row=6, column=9, value="=IFERROR(TODAY()-MAX(SharpDate),0)").style = "field_value"
    ws.cell(row=7, column=8, value="Reminder").style = "field_label"
    ws.cell(row=7, column=9, value='=IF(TODAY()-MAX(SharpDate)>=14,"⚠ Sharpen soon","✓ Good")').style = "field_value"


# ===========================================================================
# 9 — Stick Inventory
# ===========================================================================
def build_sticks(wb):
    build_log(
        wb, "Sticks", "🏑", "STICK INVENTORY",
        "Track your sticks — flex, curve, usage hours and which ones are broken.",
        ["Brand / Model", "Flex", "Curve", "Length", "Purchased", "Cost", "Broken?", "Usage Hrs"],
        [
            ("Bauer Vapor", "40", "P92", "Jr", dminus(40), 180, "No", 22),
            ("CCM Jetspeed", "40", "P28", "Jr", dminus(120), 170, "No", 48),
            ("Warrior Alpha", "40", "P92", "Jr", dminus(200), 160, "Yes", 60),
            ("True Catalyst", "45", "P28", "Int", dminus(20), 200, "No", 8),
            ("Bauer Nexus (backup)", "40", "P92", "Jr", dminus(260), 90, "No", 70),
        ],
        [22, 10, 10, 10, 13, 11, 11, 12],
        text_left={1}, dates={5}, money={6}, ints={8},
        validations=[("7", None)] and [("G", "YesNoList")], reserved=30)
    ws = wb["Sticks"]
    ws.conditional_formatting.add(
        "G5:G34", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(RED_BG)))


# ===========================================================================
# 10 — Tournament Command Center
# ===========================================================================
def build_tournaments(wb):
    build_log(
        wb, "Tournaments", "🏆", "TOURNAMENT COMMAND CENTER",
        "Plan every tournament — location, hotel, fees, schedule & checklists.",
        ["Tournament", "Type", "Location", "Start", "End", "Hotel", "Entry Fee", "Status", "Notes"],
        [
            ("Fall Classic", "Showcase", "Capital City", dplus(15), dplus(17), "Marriott Downtown", 350, "Booked", "3 nights, 4 games"),
            ("Winter Showcase", "Showcase", "Lakeside", dplus(45), dplus(47), "Lakeside Inn", 400, "Planned", "Need hotel"),
            ("Holiday Cup", "Championship", "Metro City", dplus(75), dplus(78), "TBD", 425, "Planned", ""),
            ("Spring Thaw", "Spring", "Riverside (home)", dplus(120), dplus(122), "— (home)", 300, "Planned", "No travel"),
        ],
        [20, 14, 16, 13, 13, 18, 11, 12, 20],
        text_left={1, 3, 6, 9}, dates={4, 5}, money={7},
        validations=[("B", "TourneyTypeList"), ("H", "TravelStatusList")], reserved=30)


# ===========================================================================
# 11 — Travel Planner
# ===========================================================================
def build_travel(wb):
    sample = [
        ("Fall Classic", "Drive", dplus(15), "Capital City", 220, 95, 40, 25, 160, "Booked"),
        ("Rangers @ North", "Drive", dplus(12), "North Arena", 65, 28, 10, 0, 0, "Planned"),
        ("Winter Showcase", "Drive", dplus(45), "Lakeside", 180, 78, 30, 15, 320, "Planned"),
        ("Holiday Cup", "Fly", dplus(75), "Metro City", 0, 0, 60, 0, 600, "Planned"),
    ]
    ws, start, end = build_log(
        wb, "Travel", "✈", "TRAVEL PLANNER",
        "Track miles, fuel, tolls & lodging for every road trip — costs total themselves.",
        ["Trip", "Mode", "Date", "Destination", "Miles", "Fuel", "Tolls", "Parking", "Lodging", "Status"],
        sample, [20, 10, 13, 18, 11, 11, 11, 11, 12, 12],
        text_left={1, 4}, dates={3}, money={6, 7, 8, 9}, ints={5},
        validations=[("J", "TravelStatusList")], reserved=40)
    nrange(wb, "TravelMiles", "Travel", "E", start, end)
    nrange(wb, "TravelFuel", "Travel", "F", start, end)
    nrange(wb, "TravelLodging", "Travel", "I", start, end)
    # totals
    ws.cell(row=4, column=12, value="TOTALS").style = "section_gold"
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 12
    tot = [("Total miles", "=SUM(TravelMiles)", "0"),
           ("Fuel + tolls", "=SUM(TravelFuel)+SUM(Travel!G5:G44)", '"$"#,##0'),
           ("Lodging", "=SUM(TravelLodging)", '"$"#,##0'),
           ("Trip cost", "=SUM(Travel!F5:I44)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(tot):
        r = 5 + i
        ws.cell(row=r, column=12, value=lab).style = "field_label"
        c = ws.cell(row=r, column=13, value=fml); c.style = "field_value"; c.number_format = fmt


# ===========================================================================
# 12 — Hotel Planner
# ===========================================================================
def build_hotels(wb):
    build_log(
        wb, "Hotels", "🏨", "HOTEL PLANNER",
        "Every booking in one place — confirmation, cost, dates & amenities.",
        ["Tournament", "Hotel", "Confirmation #", "Address", "Check-In", "Check-Out", "Nights", "Cost", "Amenities"],
        [
            ("Fall Classic", "Marriott Downtown", "MAR-88231", "10 Center St, Capital City", dplus(15), dplus(17), 2, 410, "Pool, breakfast, fridge"),
            ("Winter Showcase", "Lakeside Inn", "TBD", "5 Shore Rd, Lakeside", dplus(45), dplus(47), 2, 320, "Breakfast, parking"),
            ("Holiday Cup", "TBD", "—", "Metro City", dplus(75), dplus(78), 3, 540, "Team block"),
        ],
        [18, 20, 16, 24, 13, 13, 9, 11, 22],
        text_left={1, 2, 4, 9}, dates={5, 6}, money={8}, ints={7}, reserved=30)


# ===========================================================================
# 13 — Carpool Manager
# ===========================================================================
def build_carpool(wb):
    build_log(
        wb, "Carpool", "🚗", "CARPOOL MANAGER",
        "Coordinate rides — who drives, who's riding, pickups and fuel splits.",
        ["Date", "Event", "Driver", "Players", "Pickup", "Return", "Seats", "Fuel Split", "Notes"],
        [
            (dplus(3), "Game @ Centennial", "Miller", "Jordan, Alex, Sam", "9:00 AM", "1:00 PM", 4, 12, ""),
            (dplus(8), "Home game", "—", "—", "—", "—", 0, 0, "Home — no carpool"),
            (dplus(12), "Game @ North", "Patel", "Jordan, Riley, Max", "7:30 AM", "12:00 PM", 4, 14, ""),
            (dplus(15), "Fall Classic", "Miller", "Jordan, Alex", "8:00 AM", "—", 5, 40, "Tournament — drive up"),
            (dplus(1), "Practice", "Nguyen", "Jordan, Sam", "5:30 PM", "7:45 PM", 3, 6, "Rotation week 2"),
        ],
        [13, 20, 14, 24, 11, 11, 9, 11, 18],
        text_left={2, 3, 4, 9}, dates={1}, money={8}, ints={7}, reserved=40)


# ===========================================================================
# 14 — Team Roster
# ===========================================================================
def build_roster(wb):
    build_log(
        wb, "Roster", "📋", "TEAM ROSTER",
        "The whole team in one directory — players, parents & emergency contacts.",
        ["#", "Player", "Position", "Parent", "Phone", "Email", "Birthday", "Emergency Contact"],
        [
            (19, "Jordan Miller", "Center", "Sarah Miller", "(555) 210-7788", "smiller@email.com", dt.date(2012, 3, 14), "Dad (555) 210-4521"),
            (7, "Alex Patel", "Left Wing", "Raj Patel", "(555) 332-1190", "rpatel@email.com", dt.date(2012, 7, 2), "Mom (555) 332-1191"),
            (4, "Sam Nguyen", "Defense", "Kim Nguyen", "(555) 778-3300", "knguyen@email.com", dt.date(2011, 11, 9), "Dad (555) 778-3301"),
            (30, "Riley Cohen", "Goalie", "Dana Cohen", "(555) 909-2020", "dcohen@email.com", dt.date(2012, 1, 22), "—"),
            (12, "Max Roy", "Right Wing", "Luc Roy", "(555) 661-2048", "lroy@email.com", dt.date(2012, 5, 30), "Mom (555) 661-2049"),
            (8, "Charlie Lee", "Defense", "Pat Lee", "(555) 443-7766", "plee@email.com", dt.date(2011, 9, 18), "—"),
        ],
        [6, 18, 13, 18, 16, 22, 13, 22],
        text_left={2, 4, 6, 8}, dates={7}, ints={1},
        validations=[("C", "PositionList")], reserved=40)


# ===========================================================================
# 15 — Player Development
# ===========================================================================
def build_development(wb):
    sample = [
        ("Skating", "Strong", 0.75, "Edges much improved", "Add backward crossovers"),
        ("Shooting", "Average", 0.55, "Quicker release", "Work on one-timers"),
        ("Passing", "Strong", 0.7, "Good vision", "Saucer passes"),
        ("Stickhandling", "Average", 0.6, "Comfortable in traffic", "Toe drags"),
        ("Speed", "Developing", 0.45, "Building first-step", "Dryland sprints"),
        ("Defense", "Average", 0.6, "Better gap control", "Angling drills"),
        ("Offense", "Strong", 0.72, "Net-front presence", "Off-puck movement"),
        ("Hockey IQ", "Strong", 0.7, "Reads plays well", "Video review"),
    ]
    ws, start, end = build_log(
        wb, "Development", "📈", "PLAYER DEVELOPMENT",
        "Rate each skill, track progress and capture coach feedback all season.",
        ["Skill", "Level", "Progress", "Coach Feedback", "Next Focus"],
        sample, [18, 14, 12, 30, 28],
        text_left={4, 5}, pcts={3},
        validations=[("B", "SkillLevelList")], reserved=20)
    nrange(wb, "DevSkill", "Development", "A", start, end)
    nrange(wb, "DevProgress", "Development", "C", start, end)
    ws.conditional_formatting.add(
        f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))
    bar = BarChart(); bar.type = "bar"; bar.title = "Skill Progress"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    bar.legend = None
    ws.add_chart(bar, "G5")


# ===========================================================================
# 16 — Stats Tracker
# ===========================================================================
def build_stats(wb):
    ws = wb.create_sheet("Stats")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [13, 18, 8, 8, 8, 8, 9, 8, 4, 18, 12])
    luxe_header(ws, "K", "📊  STATS TRACKER",
                "Log every game — points, shots & plus/minus total themselves all season.")
    table_headers(ws, 4, ["Date", "Opponent", "G", "A", "PTS", "Shots", "+/-", "PIM"])
    games = [
        (dminus(25), "Sharks", 1, 0, "=C5+D5", 4, 1, 0),
        (dminus(18), "Wolves", 2, 1, "=C6+D6", 6, 2, 2),
        (dminus(11), "Kings", 0, 1, "=C7+D7", 3, -1, 0),
        (dminus(4), "Wolves", 1, 1, "=C8+D8", 5, 1, 0),
    ]
    start = L0
    end = start + 25 - 1
    for i, row in enumerate(games):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, 8, text_left={2}, dates={1}, ints={3, 4, 5, 6, 7, 8})
    for r in range(start + len(games), end + 1):
        ws.cell(row=r, column=5, value=f"=IF(AND(C{r}=\"\",D{r}=\"\"),\"\",N(C{r})+N(D{r}))")
    total = end + 1
    ws.cell(row=total, column=1, value="SEASON").style = "th"
    ws.cell(row=total, column=2, value="Totals").style = "th"
    for col in (3, 4, 5, 6, 7, 8):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE)
    nrange(wb, "StatG", "Stats", "C", start, end)
    nrange(wb, "StatA", "Stats", "D", start, end)
    nrange(wb, "StatPTS", "Stats", "E", start, end)
    wb.defined_names["StatGamesPlayed"] = DefinedName("StatGamesPlayed", attr_text=f"Stats!$A${start}:$A${end}")

    # season totals card block (for bar chart G/A/PTS)
    ws.cell(row=4, column=10, value="SEASON TOTALS").style = "section_gold"
    labels = [("Goals", f"=SUM(StatG)"), ("Assists", f"=SUM(StatA)"), ("Points", f"=SUM(StatPTS)")]
    for i, (lab, fml) in enumerate(labels):
        r = 5 + i
        ws.cell(row=r, column=10, value=lab).style = "td_left"
        ws.cell(row=r, column=11, value=fml).style = "td"
    wb.defined_names["StatTotLabels"] = DefinedName("StatTotLabels", attr_text="Stats!$J$5:$J$7")
    wb.defined_names["StatTotVals"] = DefinedName("StatTotVals", attr_text="Stats!$K$5:$K$7")
    line = LineChart(); line.title = "Points per Game"; line.height = 8; line.width = 13
    line.add_data(Reference(ws, min_col=5, min_row=4, max_row=start + len(games) - 1), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=2, min_row=start, max_row=start + len(games) - 1))
    line.legend = None
    ws.add_chart(line, "J9")
    ws.freeze_panes = "A5"


# ===========================================================================
# 17 — Nutrition Planner
# ===========================================================================
def build_nutrition(wb):
    ws = wb.create_sheet("Nutrition")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 34, 4, 22, 30, 2])
    luxe_header(ws, "G", "🥗  NUTRITION PLANNER",
                "Fuel for performance — game-day, practice & tournament meals + hydration.")
    blocks = [
        ("GAME-DAY FUEL", [
            ("3–4 hrs before", "Pasta + lean protein + veg"),
            ("1–2 hrs before", "Banana, toast, water"),
            ("Between periods", "Water + orange slices"),
            ("After game", "Protein + carbs within 45 min")]),
        ("PRACTICE DAYS", [
            ("Pre-practice snack", "Apple + peanut butter"),
            ("Hydration target", "Water through the day"),
            ("Post-practice", "Chocolate milk / smoothie"),
            ("Dinner", "Protein + rice + veg")]),
        ("TOURNAMENT WEEKEND", [
            ("Cooler packing", "Sandwiches, fruit, water, bars"),
            ("Hotel breakfast", "Eggs, oatmeal, fruit"),
            ("Avoid", "Heavy fried food before games"),
            ("Treat", "One after the last game!")]),
        ("HYDRATION & EXTRAS", [
            ("Water bottle", "Refill x3+ on game day"),
            ("Electrolytes", "For tournaments / hot rinks"),
            ("Supplements", "Per pediatrician only"),
            ("Shopping list", "See Budget / Groceries")]),
    ]
    row = 5
    col_pairs = [(2, 3), (5, 6)]
    for bi, (title, items) in enumerate(blocks):
        cl, cr = col_pairs[bi % 2]
        if bi % 2 == 0 and bi > 0:
            row += 1
        base = 5 if bi < 2 else 13
        merge_set(ws, f"{get_column_letter(cl)}{base}:{get_column_letter(cr)}{base}", title, "section_gold")
        for i, (a, b) in enumerate(items):
            r = base + 1 + i
            ws.cell(row=r, column=cl, value=a).style = "field_label"
            ws.cell(row=r, column=cr, value=b).style = "field_value"
            ws.row_dimensions[r].height = 22


# ===========================================================================
# 18 — Medical Center
# ===========================================================================
def build_medical(wb):
    build_log(
        wb, "Medical", "🏥", "MEDICAL CENTER",
        "Injuries, appointments, recovery & insurance claims — organized and private.",
        ["Date", "Type", "Issue / Visit", "Provider", "Status", "Follow-Up", "Cost", "Claim #", "Notes"],
        [
            (dminus(40), "Injury", "Rolled ankle", "Dr. Lee", "Recovered", "—", 0, "—", "RICE protocol"),
            (dminus(30), "Physio", "Ankle rehab", "PeakPhysio", "Complete", "—", 240, "CLM-2210", "6 sessions"),
            (dminus(10), "Checkup", "Sports physical", "Dr. Lee", "Cleared", "Next yr", 0, "—", "Cleared to play"),
            (dminus(5), "Dental", "Mouthguard fit", "Smile Dental", "Complete", "—", 60, "—", "Custom guard"),
        ],
        [13, 12, 20, 16, 13, 12, 11, 12, 22],
        text_left={3, 9}, dates={1}, money={7}, reserved=40)


# ===========================================================================
# 19 — Packing Checklist
# ===========================================================================
def build_packing(wb):
    ws = wb.create_sheet("Packing")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 4, 30, 4, 4, 30, 4, 4, 30, 2])
    luxe_header(ws, "J", "🎒  PACKING CHECKLIST",
                "Never forget the gear again — ready-made lists for every kind of trip.")
    lists = [
        ("PRACTICE BAG", ["Helmet + cage", "Shoulder pads", "Elbow pads", "Gloves",
                          "Pants", "Shin guards", "Skates", "Jock/Jill", "Socks",
                          "Practice jersey", "Stick", "Water bottle", "Tape", "Towel"]),
        ("GAME-DAY BAG", ["All practice gear", "Game jersey", "Neck guard", "Mouth guard",
                          "Backup laces", "Spare stick", "Snacks", "Extra water",
                          "Game sheet / pass", "Hand warmers", "Phone charger", "Cash"]),
        ("TOURNAMENT WEEKEND", ["Full game bag", "2 jerseys", "Extra socks x3", "Dry base layers",
                                "Skate sharpening done", "Cooler + meals", "Toiletries",
                                "Hotel confirmation", "Chargers", "First-aid kit",
                                "Schedule printout", "Team apparel"]),
    ]
    cols = [(2, 3), (5, 6), (8, 9)]
    for li, (title, items) in enumerate(lists):
        cbox, ctext = cols[li]
        merge_set(ws, f"{get_column_letter(cbox)}5:{get_column_letter(ctext)}5", title, "section_gold")
        ws.row_dimensions[5].height = 22
        for i, it in enumerate(items):
            r = 6 + i
            cb = ws.cell(row=r, column=cbox, value="☐"); cb.alignment = Alignment(horizontal="center")
            cb.font = Font(size=12, color=ACCENT); cb.border = BOX
            ct = ws.cell(row=r, column=ctext, value=it); ct.style = "td_left"
            if i % 2:
                cb.fill = fill(MUTED_ROW); ct.fill = fill(MUTED_ROW)


# ===========================================================================
# 20 — Fundraising Tracker
# ===========================================================================
def build_fundraising(wb):
    sample = [
        ("Bottle drive", "Fundraiser", dminus(20), 0, 640, "Team", "Great turnout"),
        ("Local sponsor — Joe's Pizza", "Sponsor", dminus(35), 0, 500, "Manager", "Logo on banner"),
        ("Chocolate sales", "Sales", dminus(10), 200, 820, "All families", "Sold 410 bars"),
        ("Raffle night", "Fundraiser", dplus(20), 0, 0, "Committee", "Tickets printing"),
        ("Rink board sponsor", "Sponsor", dminus(50), 0, 750, "Manager", "Season-long"),
        ("Volunteer hours (50/2)", "Volunteer", dminus(5), 0, 0, "Miller", "12 / 25 hrs done"),
    ]
    ws, start, end = build_log(
        wb, "Fundraising", "💵", "FUNDRAISING TRACKER",
        "Offset the season's cost — track sponsors, sales, donations & volunteer hours.",
        ["Activity", "Type", "Date", "Cost", "Raised", "Lead", "Notes"],
        sample, [26, 14, 13, 11, 11, 16, 22],
        text_left={1, 6, 7}, dates={3}, money={4, 5}, reserved=40)
    nrange(wb, "FundRaised", "Fundraising", "E", start, end)
    ws.cell(row=4, column=9, value="TOTALS").style = "section_gold"
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 12
    ws.cell(row=5, column=9, value="Total raised").style = "field_label"
    c = ws.cell(row=5, column=10, value="=SUM(FundRaised)"); c.style = "field_value"; c.number_format = '"$"#,##0'
    ws.cell(row=6, column=9, value="Offsets budget").style = "field_label"
    c = ws.cell(row=6, column=10, value="=IFERROR(SUM(FundRaised)/BudgetTotalPlanned,0)"); c.style = "field_value"; c.number_format = "0%"


# ===========================================================================
# 21 — Team Communication
# ===========================================================================
def build_comms(wb):
    build_log(
        wb, "Team Comms", "📣", "TEAM COMMUNICATION",
        "One log for coach messages, volunteer duties, announcements & parent notes.",
        ["Date", "From", "Type", "Message / Item", "Owner", "Done?", "Notes"],
        [
            (dminus(2), "Coach Daniels", "Announcement", "Wear navy for home games", "All", "Yes", ""),
            (dminus(1), "Team Manager", "Volunteer", "Need timekeeper Sat game", "Open", "No", "Sign up!"),
            (dplus(0), "Coach Roy", "Message", "Extra skills session Fri", "All", "No", "Optional"),
            (dplus(1), "Team Manager", "Event", "Team photos before game", "All", "No", "Arrive early"),
            (dminus(4), "Treasurer", "Announcement", "Tournament fees due", "Parents", "Yes", ""),
            (dplus(3), "Coach Daniels", "Volunteer", "Snack parent for tourney", "Open", "No", ""),
        ],
        [13, 16, 14, 28, 14, 9, 18],
        text_left={4, 7}, dates={1},
        validations=[("F", "YesNoList")], reserved=40)
    ws = wb["Team Comms"]
    ws.conditional_formatting.add(
        "F5:F44", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 22 — Photo & Memory Gallery (image placeholders)
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "📸  PHOTO & MEMORY GALLERY",
                "Keep the season's best moments — paste photos with captions & dates.")
    sections = ["Team Photo", "Tournament Win", "Player Award", "Action Shot",
                "Championship", "Family Memory"]
    top0 = 5
    card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th")
        ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+5}", "📷\nPaste photo here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 6):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 6, column=col, value="Caption").style = "field_label"
        merge_set(ws, f"{M}{row+6}:{R}{row+6}", "", "field_value")
        ws.cell(row=row + 7, column=col, value="Date / Event").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")


# ===========================================================================
# 23 — Hockey Goals
# ===========================================================================
def build_goals(wb):
    sample = [
        ("Make AA top line", "Season", dplus(120), 0.5, "On Track", ""),
        ("Score 30+ points", "Season", dplus(150), 0.4, "On Track", "13 pts so far"),
        ("Backward crossovers clean", "Skill", dplus(45), 0.6, "On Track", ""),
        ("Quicker shot release", "Skill", dplus(60), 0.55, "On Track", ""),
        ("3x dryland / week", "Fitness", dplus(30), 0.7, "On Track", ""),
        ("Keep B+ average", "Academic", dplus(120), 0.8, "On Track", "Balance school"),
        ("Be a good teammate", "Personal", dplus(150), 0.85, "On Track", ""),
    ]
    ws, start, end = build_log(
        wb, "Goals", "🎯", "HOCKEY GOALS",
        "Season, skill, fitness & personal goals — progress bars keep them in view.",
        ["Goal", "Category", "Target Date", "Progress", "Status", "Notes"],
        sample, [30, 14, 14, 12, 14, 24],
        text_left={1, 6}, dates={3}, pcts={4}, reserved=30)
    nrange(wb, "GoalName", "Goals", "A", start, end)
    nrange(wb, "GoalProgress", "Goals", "D", start, end)
    ws.conditional_formatting.add(
        f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))


# ===========================================================================
# 24 — Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 18, 3, 16, 12, 12, 12, 12, 12, 2])
    luxe_header(ws, "L", "📈  ANALYTICS DASHBOARD",
                "The season by the numbers — readiness scores, costs & a Hockey Readiness Score.")

    merge_set(ws, "B5:D5", "FAMILY READINESS SCORES", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Budget Health", '=IFERROR(1-BudgetTotalActual/BudgetTotalPlanned,0)'),
        ("Attendance", '=IFERROR(COUNTIF(PracAttend,"Present")/MAX(COUNTA(PracDate),1),0)'),
        ("Equipment Ready", '=IFERROR(1-SUMPRODUCT((EquipReplace<>"")*(EquipReplace<=TODAY()+45))/MAX(COUNTA(EquipName),1),0)'),
        ("Skill Progress", '=IFERROR(AVERAGE(DevProgress),0)'),
        ("Goal Progress", '=IFERROR(AVERAGE(GoalProgress),0)'),
        ("Season Complete", '=IFERROR(COUNTIFS(CalType,"Game",CalDate,"<"&TODAY())/MAX(COUNTIF(CalType,"Game"),1),0)'),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4,
                value=f'=IF(C{r}>=0.75,"Great",IF(C{r}>=0.4,"On Track","Needs Work"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(
        f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF" + WARN_BG,
                       mid_type="num", mid_value=0.5, mid_color="FFFFF3CD",
                       end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))

    merge_set(ws, "F5:H5", "HOCKEY READINESS SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(
                top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "A blend of budget, attendance, gear, skills & schedule.", "subtitle")
    ws["F10"].fill = fill(IVORY)

    # ice time by month trend
    merge_set(ws, "F12:H12", "ICE TIME BY MONTH (HRS)", "section")
    ws.cell(row=13, column=6, value="Month").style = "th"
    ws.cell(row=13, column=7, value="Hours").style = "th"
    trend = [("Sep", 14), ("Oct", 22), ("Nov", 26), ("Dec", 20), ("Jan", 24), ("Feb", 18)]
    for i, (m, h) in enumerate(trend):
        r = 14 + i
        ws.cell(row=r, column=6, value=m).style = "td_left"
        ws.cell(row=r, column=7, value=h).style = "td"
    line = LineChart(); line.title = "Ice Time by Month"; line.height = 7.5; line.width = 13
    line.add_data(Reference(ws, min_col=7, min_row=13, max_row=13 + len(trend)), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=6, min_row=14, max_row=13 + len(trend)))
    line.legend = None
    ws.add_chart(line, "F21")

    bar = BarChart(); bar.type = "bar"; bar.title = "Readiness by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None
    ws.add_chart(bar, "B17")


# ===========================================================================
# 1 — Executive Hockey Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2])
    ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🏒  HOCKEY FAMILY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2",
              "  Schedule, budget, equipment, travel, stats & development — your whole season, automatically organized.",
              "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    row1 = [
        ("DAYS TO NEXT GAME", '=IF(COUNTIFS(CalType,"Game",CalDate,">="&TODAY())=0,0,MINIFS(CalDate,CalType,"Game",CalDate,">="&TODAY())-TODAY())', "days"),
        ("PRACTICES / WEEK", '=COUNTIFS(CalType,"Practice",CalDate,">="&TODAY(),CalDate,"<="&TODAY()+7)', "num"),
        ("GAMES / MONTH", '=COUNTIFS(CalType,"Game",CalDate,">="&TODAY(),CalDate,"<="&TODAY()+30)', "num"),
        ("NEXT TOURNAMENT", '=IF(COUNTIFS(CalType,"Tournament",CalDate,">="&TODAY())=0,0,MINIFS(CalDate,CalType,"Tournament",CalDate,">="&TODAY())-TODAY())', "days"),
        ("MONTHLY BUDGET", "=MonthlyBudget", "money"),
        ("BUDGET LEFT", "=BudgetTotalPlanned-BudgetTotalActual", "money"),
    ]
    row2 = [
        ("EQUIP. ALERTS", '=SUMPRODUCT((EquipReplace<>"")*(EquipReplace<=TODAY()+45))', "num"),
        ("ICE TIME (30D)", '=SUMIFS(PracIce,PracDate,">="&TODAY()-30,PracDate,"<="&TODAY())', "dec"),
        ("CONFLICTS", "=SUM(CalConflict)", "num"),
        ("ATTENDANCE", '=IFERROR(COUNTIF(PracAttend,"Present")/MAX(COUNTA(PracDate),1),0)', "pct"),
        ("TRAVEL MILES", "=SUM(TravelMiles)", "num"),
        ("SEASON PROGRESS", '=IFERROR(COUNTIFS(CalType,"Game",CalDate,"<"&TODAY())/MAX(COUNTIF(CalType,"Game"),1),0)', "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)

    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "MONEY & SEASON AT A GLANCE", "section_gold")

    # Spending by group donut (group summary table lives at Budget rows 35-40)
    d1 = DoughnutChart(); d1.title = "Hockey Spending"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Budget"], min_col=2, min_row=35, max_row=40), titles_from_data=False)
    d1.set_categories(Reference(wb["Budget"], min_col=1, min_row=35, max_row=40))
    d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")

    # Practice vs Games donut (counts at Calendar J5:K8)
    d2 = DoughnutChart(); d2.title = "Practices vs Games"; d2.height = 8.2; d2.width = 11.5
    d2.add_data(Reference(wb["Calendar"], min_col=11, min_row=5, max_row=8), titles_from_data=False)
    d2.set_categories(Reference(wb["Calendar"], min_col=10, min_row=5, max_row=8))
    d2.dataLabels = no_labels()
    ws.add_chart(d2, "H12")

    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "DEVELOPMENT & RESULTS", "section_gold")

    # Skill progress bar
    s1 = BarChart(); s1.type = "bar"; s1.title = "Skill Progress"; s1.height = 8.2; s1.width = 11.5
    s1.add_data(Reference(wb["Development"], min_col=3, min_row=4, max_row=12), titles_from_data=True)
    s1.set_categories(Reference(wb["Development"], min_col=1, min_row=5, max_row=12))
    s1.legend = None
    ws.add_chart(s1, "B30")

    # Season totals bar (G/A/PTS at Stats J5:K7)
    g1 = BarChart(); g1.type = "col"; g1.title = "Season Totals"; g1.height = 8.2; g1.width = 11.5
    g1.add_data(Reference(wb["Stats"], min_col=11, min_row=5, max_row=7), titles_from_data=False)
    g1.set_categories(Reference(wb["Stats"], min_col=10, min_row=5, max_row=7))
    g1.legend = None
    ws.add_chart(g1, "H30")

    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47",
              "Hockey Family Command Center™ — your whole season, organized in one place. Edit anything in Settings.",
              "subtitle")


# ===========================================================================
# Build
# ===========================================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    build_settings(wb)
    build_welcome(wb)
    build_profile(wb)
    build_calendar(wb)
    build_gameday(wb)
    build_practice(wb)
    build_budget(wb)
    build_equipment(wb)
    build_sharpening(wb)
    build_sticks(wb)
    build_tournaments(wb)
    build_travel(wb)
    build_hotels(wb)
    build_carpool(wb)
    build_roster(wb)
    build_development(wb)
    build_stats(wb)
    build_nutrition(wb)
    build_medical(wb)
    build_packing(wb)
    build_fundraising(wb)
    build_comms(wb)
    build_gallery(wb)
    build_goals(wb)
    build_analytics(wb)
    build_dashboard(wb)   # index 0

    order = ["Welcome", "Dashboard", "Player Profile", "Calendar", "Game Day",
             "Practice", "Budget", "Equipment", "Sharpening", "Sticks",
             "Tournaments", "Travel", "Hotels", "Carpool", "Roster",
             "Development", "Stats", "Nutrition", "Medical", "Packing",
             "Fundraising", "Team Comms", "Gallery", "Goals", "Analytics",
             "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "Hockey_Family_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
