"""Build Freelancer Cashflow & Tax Command Center™ — The Self-Employed Operating System.

14 tabs · a premium freelancer / self-employed operating system in Google Sheets &
Excel. Dashboard, a cashflow-&-tax engine (income → net → auto tax set-aside → what's
really yours), invoices, clients, time & rates, business expenses, a tax vault,
mileage & home office, a pipeline, savings & runway, subscriptions and a monthly
summary — one dashboard. Get paid, set aside tax, and know your real hourly rate.

Run: python3 build_xlsx.py   ->  ../Freelancer_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
INVSTATUS = ["Paid", "Sent", "Unpaid", "Overdue", "Draft"]
EXPCAT = ["Software", "Equipment", "Marketing", "Contractor", "Office", "Travel", "Other"]
PIPESTAGE = ["Lead", "Proposal", "Negotiating", "Won", "Lost"]

INCOME_GOAL = 7500
MARGIN_GOAL = 0.70
RATE_GOAL = 60
TAKE_GOAL = 4000
RUNWAY_GOAL_MONTHS = 3
TAX_RATE = 0.30

# Invoices this month: (client, invoice #, amount, status)
INVOICES = [
    ("Northwind Co.", "INV-101", 1800, "Paid"), ("Acme Studio", "INV-102", 1500, "Paid"),
    ("Lumen Labs", "INV-103", 2200, "Paid"), ("Bright Agency", "INV-104", 2500, "Paid"),
    ("Cobalt Inc.", "INV-105", 1400, "Unpaid"), ("Vertex Media", "INV-106", 1000, "Sent"),
]

# Clients: (client, retainer?, since, rate)
CLIENTS = [
    ("Northwind Co.", "Retainer", "2024", 85), ("Acme Studio", "Project", "2025", 75),
    ("Lumen Labs", "Retainer", "2023", 95), ("Bright Agency", "Project", "2025", 80),
    ("Cobalt Inc.", "Project", "2026", 70),
]

# Business expenses this month: (item, category, amount)
EXPENSES = [
    ("Design software", "Software", 250), ("Cloud & subscriptions", "Software", 180),
    ("New monitor", "Equipment", 150), ("Ads & portfolio", "Marketing", 120),
    ("Contractor help", "Contractor", 300), ("Misc & office", "Office", 200),
]

# Time & rates
BILLABLE_HOURS = 100
AVAILABLE_HOURS = 160
RATECARD = [
    ("Design — hourly", 85), ("Strategy call", 150), ("Retainer (mo)", 2200), ("Rush surcharge", 40),
]

# Tax vault — quarterly estimates: (quarter, due, estimate)
QUARTERS = [
    ("Q1", "Apr 15", 5100), ("Q2", "Jun 15", 5400), ("Q3", "Sep 15", 6200), ("Q4", "Jan 15", 6800),
]
TAX_SAVED = 2040

# Mileage & home office: (item, detail, amount)
DEDUCTIONS = [
    ("Business miles (July)", "312 mi × $0.67", 209.04), ("Home office %", "12% of rent/utilities", 264.00),
    ("Phone & internet", "60% business use", 78.00), ("Professional development", "1 course", 49.00),
]

# Pipeline: (prospect, stage, value, next step)
PIPELINE = [
    ("Delta Corp", "Proposal", 3200, "Follow up Fri"), ("Ember Co.", "Negotiating", 2500, "Send contract"),
    ("Fable Studio", "Lead", 1800, "Discovery call"), ("Grove Media", "Proposal", 4000, "Await sign-off"),
]

# Savings & runway
EMERGENCY_FUND = 6000
MONTHLY_NEED = 5000
RETIREMENT_MO = 400
RETIREMENT_GOAL = 1200

# Subscriptions & tools: (tool, monthly, annual?)
SUBS = [
    ("Design suite", 55.00, "Monthly"), ("Cloud storage", 12.00, "Monthly"), ("Accounting app", 25.00, "Monthly"),
    ("Domain & hosting", 18.00, "Monthly"), ("Email & CRM", 29.00, "Monthly"), ("Stock assets", 41.00, "Monthly"),
]

# Monthly summary: (month, income, take-home)
MONTHS = [("Jul", 8000, 4760), ("Aug", 7400, 4400), ("Sep", 9200, 5460),
          ("Oct", 6800, 4050), ("Nov", 8600, 5100), ("Dec", 10200, 6050)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Your name / business", "Studio Fern", None, "Business"),
        ("Owner", "Sasha", None, "Owner"),
        ("Monthly income goal", INCOME_GOAL, '"$"#,##0', "IncomeGoal"),
        ("Profit-margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Effective-rate goal ($/hr)", RATE_GOAL, '"$"#,##0', "RateGoal"),
        ("Monthly take-home goal", TAKE_GOAL, '"$"#,##0', "TakeGoal"),
        ("Runway goal (months)", RUNWAY_GOAL_MONTHS, "#,##0", "RunwayGoalMonths"),
        ("Tax set-aside %", TAX_RATE, "0%", "TaxRate"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Inv status", INVSTATUS, "InvStatusList"), ("F", "Exp category", EXPCAT, "ExpCatList"),
             ("G", "Pipe stage", PIPESTAGE, "PipeStageList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  💼  FREELANCER CASHFLOW & TAX COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Get paid, set aside tax, and know your real hourly rate.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE FREELANCE BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Freelancing lives or dies on two numbers: what you actually keep after expenses and tax, and what "
                      "you truly earn per hour. This makes both visible: a cashflow-and-tax engine that takes your income, "
                      "subtracts business expenses, sets aside tax automatically and shows your real take-home. Send and "
                      "track invoices, keep clients and a pipeline, log deductible expenses and mileage, build a tax vault "
                      "and a runway, and watch your monthly take-home — all in ONE premium Google Sheets & Excel system "
                      "built for freelancers and the self-employed.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your income, rate & tax set-aside %.",
             "2.  Add Invoices; paid ones become your income, unpaid your outstanding.",
             "3.  Log Business Expenses by category — every deduction counts.",
             "4.  The Cashflow & Tax engine sets aside tax and shows your take-home.",
             "5.  Track time & rates, the tax vault, mileage and your runway.",
             "6.  Check the Dashboard: take-home, rate & a Freelance Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional freelancer (Studio Fern) is included so you can see how it all connects — "
               "just type over it with your own invoices and expenses. Take-home after tax and your effective hourly "
               "rate are the two numbers that decide whether freelancing is really working, and they roll into a live "
               "Freelance Score. Twelve matching printable pages (invoice, tax set-aside worksheet, mileage log, expense "
               "sheet & more) are included. This is a business tool, not financial, legal or tax advice — the tax "
               "set-aside is an estimate; confirm your rate and payments with your own tax advisor.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Set the tax aside before you spend it — that's how the self-employed sleep at night.", "section_gold")


# ===========================================================================
def build_invoices(wb):
    ws = wb.create_sheet("Invoices"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 16, 14, 14, 2])
    luxe_header(ws, "E", "🧾  INVOICES",
                "What you've billed — paid invoices become income, unpaid become your outstanding.")
    table_headers(ws, 4, ["Client", "Invoice #", "Amount", "Status"], start_col=2)
    start = L0
    for i, (client, num, amt, status) in enumerate(INVOICES):
        r = start + i
        ws.cell(row=r, column=2, value=client).style = "td_left"
        ws.cell(row=r, column=3, value=num).style = "td"
        ca = ws.cell(row=r, column=4, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=status).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"E{r}", "InvStatusList")
    end = start + len(INVOICES) - 1
    nrange(wb, "InvAmount", "Invoices", "D", start, end)
    nrange(wb, "InvStatus", "Invoices", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="INCOME (paid)").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ci = ws.cell(row=tot, column=4, value='=SUMIF(InvStatus,"Paid",InvAmount)'); ci.style = "td"; ci.font = Font(bold=True, color=PRIMARY); ci.fill = fill(MINT_BG); ci.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "MonthlyIncome", "Invoices", f"$D${tot}")
    sr = tot + 1
    ws.cell(row=sr, column=2, value="OUTSTANDING (unpaid + sent)").style = "th"
    ws.cell(row=sr, column=3).style = "td"; ws.cell(row=sr, column=3).fill = fill(SURFACE)
    co = ws.cell(row=sr, column=4, value='=SUMIF(InvStatus,"Unpaid",InvAmount)+SUMIF(InvStatus,"Sent",InvAmount)+SUMIF(InvStatus,"Overdue",InvAmount)'); co.style = "td"; co.font = Font(bold=True, color=DANGER); co.fill = fill(SURFACE); co.number_format = '"$"#,##0'
    ws.cell(row=sr, column=5).style = "td"; ws.cell(row=sr, column=5).fill = fill(SURFACE)
    cell_name(wb, "Outstanding", "Invoices", f"$D${sr}")
    ws.cell(row=sr + 1, column=2, value="Invoices this month").style = "field_label"
    cn = ws.cell(row=sr + 1, column=4, value="=COUNTA(InvStatus)"); cn.style = "field_value"; cn.number_format = "#,##0"; cn.fill = fill(MINT_BG)
    cell_name(wb, "InvCount", "Invoices", f"$D${sr+1}")
    ws.freeze_panes = "A5"


def build_expenses(wb):
    ws = wb.create_sheet("Business Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 14, 2])
    luxe_header(ws, "D", "💳  BUSINESS EXPENSES",
                "Every deductible cost by category — lowers your net and your tax, and feeds the engine.")
    table_headers(ws, 4, ["Item", "Category", "Amount"], start_col=2)
    start = L0
    for i, (item, cat, amt) in enumerate(EXPENSES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        ca = ws.cell(row=r, column=4, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "ExpCatList")
    end = start + len(EXPENSES) - 1
    nrange(wb, "ExpAmt", "Business Expenses", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ce = ws.cell(row=tot, column=4, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "BizExp", "Business Expenses", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_cashflow(wb):
    ws = wb.create_sheet("Cashflow & Tax"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 16, 2])
    luxe_header(ws, "C", "💰  CASHFLOW & TAX",
                "Income in, expenses and tax out — the number that's actually yours to keep.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    ws.cell(row=6, column=2, value="Income (paid invoices)").style = "field_label"
    ci = ws.cell(row=6, column=3, value="=MonthlyIncome"); ci.style = "field_value"; ci.number_format = '"$"#,##0'; ci.fill = fill(MINT_BG)
    ws.cell(row=7, column=2, value="− Business expenses").style = "field_label"
    ce = ws.cell(row=7, column=3, value="=BizExp"); ce.style = "field_value"; ce.number_format = '"$"#,##0'
    ws.cell(row=8, column=2, value="= NET INCOME").style = "th"
    cn = ws.cell(row=8, column=3, value="=MonthlyIncome-BizExp"); cn.style = "td"; cn.font = Font(bold=True, size=12, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetIncome", "Cashflow & Tax", "$C$8")
    ws.cell(row=10, column=2, value="THE TAX VAULT").style = "section_gold"
    ws.cell(row=11, column=2, value="Tax set-aside % (from Settings)").style = "field_label"
    ct = ws.cell(row=11, column=3, value="=TaxRate"); ct.style = "field_value"; ct.number_format = "0%"; ct.fill = fill(WARN_BG)
    ws.cell(row=12, column=2, value="− Tax to set aside").style = "field_label"
    cts = ws.cell(row=12, column=3, value="=ROUND(NetIncome*TaxRate,0)"); cts.style = "field_value"; cts.number_format = '"$"#,##0'; cts.fill = fill(WARN_BG)
    cell_name(wb, "TaxSetAside", "Cashflow & Tax", "$C$12")
    ws.cell(row=13, column=2, value="= YOUR TAKE-HOME (what's really yours)").style = "th"
    cth = ws.cell(row=13, column=3, value="=NetIncome-TaxSetAside"); cth.style = "td"; cth.font = Font(bold=True, size=13, color=PRIMARY); cth.fill = fill(MINT_BG); cth.number_format = '"$"#,##0'
    cell_name(wb, "TakeHome", "Cashflow & Tax", "$C$13")
    ws.cell(row=15, column=2, value="Tax actually moved to the vault").style = "field_label"
    cv = ws.cell(row=15, column=3, value=TAX_SAVED); cv.style = "input"; cv.number_format = '"$"#,##0'
    cell_name(wb, "TaxSaved", "Cashflow & Tax", "$C$15")
    ws.cell(row=16, column=2, value="Profit margin (net ÷ income)").style = "field_label"
    cm = ws.cell(row=16, column=3, value="=IFERROR(NetIncome/MonthlyIncome,0)"); cm.style = "field_value"; cm.number_format = "0%"; cm.fill = fill(MINT_BG)
    cell_name(wb, "ProfitMargin", "Cashflow & Tax", "$C$16")
    ws.cell(row=18, column=2, value="Set the tax aside the day you get paid — don't touch the vault.").style = "section"


def build_clients(wb):
    ws, start, end = build_log(
        wb, "Clients", "🤝", "CLIENTS",
        "Your roster — retainer or project, since when, and their rate.",
        ["Client", "Type", "Since", "Rate/hr"],
        CLIENTS, [2, 24, 16, 12, 14, 2], text_left={2, 3}, ints={4}, money={5}, reserved=24, start_col=2)


def build_time(wb):
    ws = wb.create_sheet("Time & Rates"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "⏱  TIME & RATES",
                "Billable hours and your rate card — see what you truly earn per hour.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    ws.cell(row=6, column=2, value="Billable hours").style = "field_label"
    cb = ws.cell(row=6, column=3, value=BILLABLE_HOURS); cb.style = "input"; cb.number_format = "#,##0"
    cell_name(wb, "BillableHours", "Time & Rates", "$C$6")
    ws.cell(row=7, column=2, value="Available hours").style = "field_label"
    cav = ws.cell(row=7, column=3, value=AVAILABLE_HOURS); cav.style = "input"; cav.number_format = "#,##0"
    cell_name(wb, "AvailableHours", "Time & Rates", "$C$7")
    ws.cell(row=8, column=2, value="Utilization (billable ÷ available)").style = "field_label"
    cu = ws.cell(row=8, column=3, value="=IFERROR(BillableHours/AvailableHours,0)"); cu.style = "field_value"; cu.number_format = "0%"; cu.fill = fill(MINT_BG)
    cell_name(wb, "Utilization", "Time & Rates", "$C$8")
    ws.cell(row=9, column=2, value="= EFFECTIVE RATE (net ÷ billable hrs)").style = "th"
    ce = ws.cell(row=9, column=3, value="=IFERROR(NetIncome/BillableHours,0)"); ce.style = "td"; ce.font = Font(bold=True, size=12, color=PRIMARY); ce.fill = fill(MINT_BG); ce.number_format = '"$"#,##0.00'
    cell_name(wb, "EffRate", "Time & Rates", "$C$9")
    ws.cell(row=11, column=2, value="RATE CARD").style = "section_gold"
    table_headers(ws, 12, ["Service", "Rate"], start_col=2)
    rs = 13
    for i, (svc, rate) in enumerate(RATECARD):
        r = rs + i
        ws.cell(row=r, column=2, value=svc).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rate); cr.style = "input"; cr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)


def build_taxvault(wb):
    ws = wb.create_sheet("Tax Vault"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 16, 16, 2])
    luxe_header(ws, "D", "🏦  TAX VAULT",
                "Quarterly estimated taxes — what's due and when, so a bill never surprises you.")
    table_headers(ws, 4, ["Quarter", "Due", "Estimate"], start_col=2)
    start = L0
    for i, (q, due, est) in enumerate(QUARTERS):
        r = start + i
        ws.cell(row=r, column=2, value=q).style = "td_left"
        ws.cell(row=r, column=3, value=due).style = "td"
        ce = ws.cell(row=r, column=4, value=est); ce.style = "input"; ce.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(QUARTERS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="ANNUAL ESTIMATE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ce = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    ws.cell(row=tot + 2, column=2, value="This month's set-aside").style = "field_label"
    cm = ws.cell(row=tot + 2, column=4, value="=TaxSetAside"); cm.style = "field_value"; cm.number_format = '"$"#,##0'; cm.fill = fill(MINT_BG)
    ws.freeze_panes = "A5"


def build_deductions(wb):
    ws, start, end = build_log(
        wb, "Mileage & Home Office", "🚗", "MILEAGE & HOME OFFICE",
        "The deductions freelancers forget — miles, home office, phone and learning.",
        ["Deduction", "Detail", "Amount"],
        DEDUCTIONS, [2, 26, 26, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "DeductAmt", "Mileage & Home Office", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL DEDUCTIONS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(DeductAmt)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0.00'


def build_pipeline(wb):
    ws, start, end = build_log(
        wb, "Pipeline", "🔭", "PIPELINE",
        "What's coming — prospects, stage, value and the next step, so work never dries up.",
        ["Prospect", "Stage", "Value", "Next step"],
        PIPELINE, [2, 20, 16, 14, 24, 2], text_left={2, 5}, money={4}, reserved=24, start_col=2,
        validations=[("C", "PipeStageList")])
    nrange(wb, "PipeValue", "Pipeline", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="PIPELINE VALUE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(PipeValue)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)


def build_runway(wb):
    ws = wb.create_sheet("Savings & Runway"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 2])
    luxe_header(ws, "C", "🛟  SAVINGS & RUNWAY",
                "Your cushion — how many months you could cover, and your retirement set-aside.")
    ws.cell(row=5, column=2, value="Emergency fund balance").style = "field_label"
    ce = ws.cell(row=5, column=3, value=EMERGENCY_FUND); ce.style = "input"; ce.number_format = '"$"#,##0'
    cell_name(wb, "EmergencyFund", "Savings & Runway", "$C$5")
    ws.cell(row=6, column=2, value="Monthly living need").style = "field_label"
    cm = ws.cell(row=6, column=3, value=MONTHLY_NEED); cm.style = "input"; cm.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyNeed", "Savings & Runway", "$C$6")
    ws.cell(row=7, column=2, value="= RUNWAY (months covered)").style = "th"
    cr = ws.cell(row=7, column=3, value="=IFERROR(EmergencyFund/MonthlyNeed,0)"); cr.style = "td"; cr.font = Font(bold=True, size=12, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = "0.0"
    ws.cell(row=8, column=2, value="Runway funded (of goal months)").style = "field_label"
    cp = ws.cell(row=8, column=3, value="=IFERROR(EmergencyFund/(RunwayGoalMonths*MonthlyNeed),0)"); cp.style = "field_value"; cp.number_format = "0%"; cp.fill = fill(MINT_BG)
    cell_name(wb, "RunwayPct", "Savings & Runway", "$C$8")
    ws.cell(row=10, column=2, value="Retirement set-aside (this month)").style = "field_label"
    crt = ws.cell(row=10, column=3, value=RETIREMENT_MO); crt.style = "input"; crt.number_format = '"$"#,##0'
    ws.cell(row=11, column=2, value="Retirement goal (monthly)").style = "field_label"
    crg = ws.cell(row=11, column=3, value=RETIREMENT_GOAL); crg.style = "input"; crg.number_format = '"$"#,##0'
    ws.cell(row=13, column=2, value="Three to six months of runway is the freelancer's safety net.").style = "section"


def build_subs(wb):
    ws, start, end = build_log(
        wb, "Subscriptions", "🔁", "SUBSCRIPTIONS & TOOLS",
        "Every recurring tool — the quiet monthly drain on a freelance business.",
        ["Tool", "Monthly", "Billing"],
        SUBS, [2, 24, 14, 16, 2], text_left={2, 4}, money2={3}, reserved=24, start_col=2)
    nrange(wb, "SubMonthly", "Subscriptions", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="MONTHLY / ANNUAL").style = "th"
    cm = ws.cell(row=tot, column=3, value="=SUM(SubMonthly)"); cm.style = "td"; cm.font = Font(bold=True, color=DANGER); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0.00'
    ca = ws.cell(row=tot, column=4, value="=SUM(SubMonthly)*12"); ca.style = "td"; ca.font = Font(bold=True, color=DANGER); ca.fill = fill(SURFACE); ca.number_format = '"$"#,##0'


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 14, 2])
    luxe_header(ws, "E", "📈  MONTHLY SUMMARY",
                "Income & take-home by month — watch the real, after-tax number grow.")
    table_headers(ws, 4, ["Month", "Income", "Take-home", "Tax set aside"], start_col=2)
    start = L0
    for i, (m, inc, take) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ci = ws.cell(row=r, column=3, value=inc); ci.style = "input"; ci.number_format = '"$"#,##0'
        ct = ws.cell(row=r, column=4, value=take); ct.style = "input"; ct.number_format = '"$"#,##0'
        cx = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cx.style = "td"; cx.font = Font(bold=True, color=ACCENT); cx.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthTake", "Monthly Summary", "D", start, end)
    ch = BarChart(); ch.type = "col"; ch.title = "Take-home by Month"; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=4, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "G4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  💼  FREELANCER CASHFLOW & TAX COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Take-home, effective rate, tax set-aside & a Freelance Score — your whole business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MONTHLY INCOME", "=MonthlyIncome", "money"),
        ("EXPENSES", "=BizExp", "money"),
        ("NET INCOME", "=NetIncome", "money"),
        ("TAX SET-ASIDE", "=TaxSetAside", "money"),
        ("TAKE-HOME", "=TakeHome", "money"),
        ("EFFECTIVE RATE", "=EffRate", "money2"),
    ]
    row2 = [
        ("BILLABLE HOURS", "=BillableHours", "num"),
        ("UTILIZATION", "=Utilization", "pct"),
        ("INVOICES", "=InvCount", "num"),
        ("OUTSTANDING", "=Outstanding", "money"),
        ("YEAR PACE", "=MonthlyIncome*12", "money"),
        ("FREELANCE SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "FREELANCE HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "TAKE-HOME BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Income vs goal", "=IFERROR(MIN(MonthlyIncome/IncomeGoal,1),0)"),
        ("Margin healthy", "=IFERROR(MIN(ProfitMargin/MarginGoal,1),0)"),
        ("Effective rate on target", "=IFERROR(MIN(EffRate/RateGoal,1),0)"),
        ("Take-home vs goal", "=IFERROR(MIN(TakeHome/TakeGoal,1),0)"),
        ("Tax set aside", "=IFERROR(MIN(TaxSaved/TaxSetAside,1),0)"),
        ("Runway built", "=IFERROR(MIN(RunwayPct,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Take-home by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=4, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Freelancer Cashflow & Tax Command Center™ — get paid, set aside tax, keep what's yours.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_invoices(wb); build_expenses(wb)
    build_cashflow(wb); build_clients(wb); build_time(wb); build_taxvault(wb)
    build_deductions(wb); build_pipeline(wb); build_runway(wb); build_subs(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Cashflow & Tax", "Invoices", "Clients", "Time & Rates",
             "Business Expenses", "Tax Vault", "Mileage & Home Office", "Pipeline", "Savings & Runway",
             "Subscriptions", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Freelancer_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
