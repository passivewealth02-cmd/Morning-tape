"""Build Contractor Job Costing & Bidding Command Center™ — The Builder's Operating System.

14 tabs · a premium contractor job-costing & bidding operating system in Google Sheets &
Excel. Dashboard, a bid builder (materials + labor + subs + equipment + overhead ÷
(1 − margin) → the price to quote), job costing (estimate vs actual), a job pipeline,
labor & crew, materials, subcontractors, change orders, equipment, invoices &
receivables, a bid log with your win rate, and a monthly summary — one dashboard. Bid it
right, cost it honestly, and finish every job in profit.

Run: python3 build_xlsx.py   ->  ../Contractor_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
JOBSTATUS = ["Bidding", "Won", "In progress", "Complete", "Lost"]
TRADE = ["General", "Electrical", "Plumbing", "HVAC", "Framing", "Drywall", "Tile", "Paint", "Roofing"]
BIDRESULT = ["Won", "Lost", "Pending"]

# Bid builder — flagship job
JOB_MATERIALS = 14000
LABOR_HOURS = 320
LABOR_RATE = 45
JOB_SUBS = 8500
JOB_EQUIP = 1200
OVERHEAD_RATE = 0.12
MARGIN_TARGET = 0.20
ACTUAL_COST = 41178

# Goals
MARGIN_GOAL = 0.18
BACKLOG_GOAL = 150000
NET_MARGIN_GOAL = 0.15
RECEIVABLE_GOAL = 35000
JOB_GOAL = 9
WINRATE_GOAL = 1.00

BACKLOG = 186000
RECEIVABLE = 31000
BIDS_WON = 9
BIDS_TOTAL = 22

# Jobs & pipeline: (job, client, status, bid, est. cost)
JOBS = [
    ("Kitchen remodel — Hale", "Hale", "In progress", 53340, 42672),
    ("Bath gut — Ortiz", "Ortiz", "Won", 28500, 22400),
    ("Deck & pergola — Nunes", "Nunes", "Complete", 19800, 15200),
    ("Basement finish — Park", "Park", "In progress", 61200, 48900),
    ("Garage conversion — Reid", "Reid", "Won", 44500, 35800),
    ("Whole-home paint — Vance", "Vance", "Complete", 12600, 9400),
    ("Roof replace — Cobb", "Cobb", "Complete", 24800, 19100),
    ("Addition — Marsh", "Marsh", "Bidding", 96000, 76500),
    ("Siding — Doyle", "Doyle", "Bidding", 31500, 25200),
]

# Labor & crew: (name, trade, hourly, hours this job)
CREW = [
    ("Cal (owner)", "General", 45, 60), ("Dev", "Framing", 38, 80), ("Rosa", "Drywall", 36, 72),
    ("Tomas", "Tile", 42, 56), ("Jae", "Paint", 32, 52),
]

# Materials: (item, qty, unit cost, total)
MATERIALS = [
    ("Cabinets", 1, 6200, 6200), ("Countertop — quartz", 1, 3400, 3400),
    ("Tile & backsplash", 1, 1450, 1450), ("Lumber & framing", 1, 1250, 1250),
    ("Fixtures & hardware", 1, 980, 980), ("Paint & finishes", 1, 420, 420),
    ("Fasteners & misc", 1, 300, 300),
]

# Subcontractors: (sub, trade, quoted, invoiced)
SUBS = [
    ("Bright Spark Electric", "Electrical", 3400, 3400), ("FlowRight Plumbing", "Plumbing", 2900, 2900),
    ("CoolAir HVAC", "HVAC", 1600, 1450), ("Stoneworks Tile", "Tile", 600, 600),
]

# Change orders: (job, description, amount, approved?)
CHANGES = [
    ("Kitchen remodel — Hale", "Move gas line", 1400, "Yes"),
    ("Kitchen remodel — Hale", "Upgrade to quartz", 2200, "Yes"),
    ("Basement finish — Park", "Add egress window", 1800, "Yes"),
    ("Garage conversion — Reid", "Extra insulation", 800, "Yes"),
]

# Equipment: (item, owned/rented, cost this job)
EQUIPMENT = [
    ("Mini excavator (rented)", "Rented", 480), ("Scaffolding (rented)", "Rented", 320),
    ("Dumpster", "Rented", 400), ("Tool consumables", "Owned", 0),
]

# Invoices: (job, amount, status)
INVOICES = [
    ("Kitchen remodel — Hale", 26670, "Paid"), ("Basement finish — Park", 30600, "Paid"),
    ("Garage conversion — Reid", 22250, "Sent"), ("Bath gut — Ortiz", 8750, "Sent"),
    ("Siding — Doyle", 0, "Draft"),
]

# Bid log: (job, bid amount, result)
BIDLOG = [
    ("Kitchen remodel — Hale", 53340, "Won"), ("Bath gut — Ortiz", 28500, "Won"),
    ("Deck & pergola — Nunes", 19800, "Won"), ("Basement finish — Park", 61200, "Won"),
    ("Garage conversion — Reid", 44500, "Won"), ("Whole-home paint — Vance", 12600, "Won"),
    ("Roof replace — Cobb", 24800, "Won"), ("Kitchen — Alder St", 41000, "Lost"),
    ("Bath — Cedar Ct", 18500, "Lost"), ("Deck — Rowe", 14200, "Lost"),
    ("Addition — Marsh", 96000, "Pending"), ("Siding — Doyle", 31500, "Pending"),
]

# Monthly summary: (month, revenue)
MONTHS = [("Feb", 52000), ("Mar", 61000), ("Apr", 68000), ("May", 74000), ("Jun", 81000), ("Jul", 92000)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 20, 3] + [20] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your rates & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR RATES & GOALS", "section")
    controls = [
        ("Company", "Ironwood Builders", None, "Company"),
        ("Owner", "Cal", None, "Owner"),
        ("Base labor rate ($/hr)", LABOR_RATE, '"$"#,##0', "LaborRate"),
        ("Overhead rate (of direct cost)", OVERHEAD_RATE, "0%", "OverheadRate"),
        ("Target margin on bids", MARGIN_TARGET, "0%", "MarginTarget"),
        ("Job-margin goal (min)", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Net-margin goal", NET_MARGIN_GOAL, "0%", "NetMarginGoal"),
        ("Backlog goal", BACKLOG_GOAL, '"$"#,##0', "BacklogGoal"),
        ("Receivables goal (max)", RECEIVABLE_GOAL, '"$"#,##0', "ReceivableGoal"),
        ("Jobs-this-year goal", JOB_GOAL, "0", "JobGoal"),
        ("Bids submitted", BIDS_TOTAL, "0", "BidsTotal"),
        ("Bids won", BIDS_WON, "0", "BidsWon"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Job status", JOBSTATUS, "JobStatusList"), ("F", "Trade", TRADE, "TradeList"),
             ("G", "Bid result", BIDRESULT, "BidResultList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  📐  CONTRACTOR JOB COSTING & BIDDING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Bid it right, cost it honestly, and finish every job in profit.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE OPERATION, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Most contractors don't lose money on bad jobs — they lose it on jobs they bid wrong. Forget the "
                      "overhead, guess the hours, add a margin on top of a number that was already short, and the job "
                      "finishes at break-even. This fixes that: a bid builder adds materials, labor, subs and equipment, "
                      "applies your real overhead rate, then divides by your margin to give the price you must quote. "
                      "Then cost every job estimate-versus-actual, run your crew, subs, change orders and receivables, "
                      "and watch your win rate — all in ONE premium Google Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your labor rate, overhead % and target margin.",
             "2.  Build a bid: materials, labor hours, subs, equipment.",
             "3.  Read the price to quote — overhead and margin already in it.",
             "4.  Log actual costs as the job runs. Estimate vs actual, live.",
             "5.  Track change orders, subs, invoices and your bid win rate.",
             "6.  Check the Dashboard: margin, backlog & a Builder Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional builder (Ironwood Builders, owner Cal) is included so you can see how it "
               "all connects — just type over it with your own jobs and rates. The critical detail most bid sheets get "
               "wrong: margin is a DIVISOR, not a multiplier. Cost ÷ (1 − 20%) gives a real 20% margin; cost × 1.20 "
               "only gives you 16.7%. This workbook does it correctly. Twelve matching printable pages (bid worksheet, "
               "job cost sheet, change order form, daily log & more) are included. This is a business & organizing "
               "tool, not financial, legal, tax or construction advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Markup is not margin. Divide by (1 − margin), or you're working for less than you think.", "section_gold")


# ===========================================================================
def build_bid(wb):
    ws = wb.create_sheet("Bid Builder"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 18, 2])
    luxe_header(ws, "C", "🧮  BID BUILDER",
                "Materials + labor + subs + equipment, plus real overhead, divided by your margin = the price to quote.")
    ws.cell(row=5, column=2, value="DIRECT COST").style = "section_gold"
    ws.cell(row=6, column=2, value="Materials").style = "field_label"
    cm = ws.cell(row=6, column=3, value=JOB_MATERIALS); cm.style = "input"; cm.number_format = '"$"#,##0'
    cell_name(wb, "BidMaterials", "Bid Builder", "$C$6")
    ws.cell(row=7, column=2, value="Labor hours").style = "field_label"
    ch = ws.cell(row=7, column=3, value=LABOR_HOURS); ch.style = "input"; ch.number_format = "#,##0"
    cell_name(wb, "BidHours", "Bid Builder", "$C$7")
    ws.cell(row=8, column=2, value="+ Labor (hours × rate)").style = "field_label"
    cl = ws.cell(row=8, column=3, value="=BidHours*LaborRate"); cl.style = "field_value"; cl.number_format = '"$"#,##0'
    cell_name(wb, "BidLabor", "Bid Builder", "$C$8")
    ws.cell(row=9, column=2, value="+ Subcontractors").style = "field_label"
    cs = ws.cell(row=9, column=3, value=JOB_SUBS); cs.style = "input"; cs.number_format = '"$"#,##0'
    cell_name(wb, "BidSubs", "Bid Builder", "$C$9")
    ws.cell(row=10, column=2, value="+ Equipment & rentals").style = "field_label"
    ce = ws.cell(row=10, column=3, value=JOB_EQUIP); ce.style = "input"; ce.number_format = '"$"#,##0'
    cell_name(wb, "BidEquip", "Bid Builder", "$C$10")
    ws.cell(row=11, column=2, value="= DIRECT COST").style = "th"
    cd = ws.cell(row=11, column=3, value="=BidMaterials+BidLabor+BidSubs+BidEquip"); cd.style = "td"; cd.font = Font(bold=True, size=12, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = '"$"#,##0'
    cell_name(wb, "DirectCost", "Bid Builder", "$C$11")
    ws.cell(row=13, column=2, value="OVERHEAD & MARGIN").style = "section_gold"
    ws.cell(row=14, column=2, value="+ Overhead (rate × direct cost)").style = "field_label"
    co = ws.cell(row=14, column=3, value="=DirectCost*OverheadRate"); co.style = "field_value"; co.number_format = '"$"#,##0'
    cell_name(wb, "Overhead", "Bid Builder", "$C$14")
    ws.cell(row=15, column=2, value="= TOTAL COST").style = "th"
    ct = ws.cell(row=15, column=3, value="=DirectCost+Overhead"); ct.style = "td"; ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "TotalCost", "Bid Builder", "$C$15")
    ws.cell(row=16, column=2, value="÷ (1 − target margin)").style = "field_label"
    cdv = ws.cell(row=16, column=3, value="=1-MarginTarget"); cdv.style = "field_value"; cdv.number_format = "0%"
    ws.cell(row=17, column=2, value="= BID THIS JOB AT").style = "th"
    cb = ws.cell(row=17, column=3, value="=IFERROR(TotalCost/(1-MarginTarget),0)"); cb.style = "td"; cb.font = Font(bold=True, size=15, color=PRIMARY); cb.fill = fill(MINT_BG); cb.number_format = '"$"#,##0'
    cell_name(wb, "BidPrice", "Bid Builder", "$C$17")
    ws.cell(row=18, column=2, value="= PLANNED PROFIT").style = "th"
    cp = ws.cell(row=18, column=3, value="=BidPrice-TotalCost"); cp.style = "td"; cp.font = Font(bold=True, size=13, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = '"$"#,##0'
    cell_name(wb, "PlannedProfit", "Bid Builder", "$C$18")
    ws.cell(row=20, column=2, value="⚠ MARKUP IS NOT MARGIN").style = "section_gold"
    ws.cell(row=21, column=2, value="If you multiplied by (1 + margin) instead...").style = "field_label"
    cw = ws.cell(row=21, column=3, value="=TotalCost*(1+MarginTarget)"); cw.style = "field_value"; cw.number_format = '"$"#,##0'; cw.fill = fill(RED_BG)
    ws.cell(row=22, column=2, value="...you'd actually earn only this margin").style = "field_label"
    cwm = ws.cell(row=22, column=3, value="=IFERROR((TotalCost*(1+MarginTarget)-TotalCost)/(TotalCost*(1+MarginTarget)),0)")
    cwm.style = "field_value"; cwm.number_format = "0.0%"; cwm.fill = fill(RED_BG)
    ws.cell(row=23, column=2, value="...and leave this much on the table").style = "field_label"
    cwl = ws.cell(row=23, column=3, value="=BidPrice-TotalCost*(1+MarginTarget)"); cwl.style = "field_value"; cwl.number_format = '"$"#,##0'; cwl.fill = fill(WARN_BG)
    ws.cell(row=25, column=2, value="Every bid you multiply instead of divide quietly hands money back to the client.").style = "section_gold"


def build_jobcost(wb):
    ws = wb.create_sheet("Job Costing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 18, 2])
    luxe_header(ws, "D", "📊  JOB COSTING",
                "What you estimated against what it actually cost — the only way to bid the next one better.")
    table_headers(ws, 5, ["Line", "Estimated", "Actual"], start_col=2)
    rows = [
        ("Materials", "=BidMaterials", "=MaterialTotal"),
        ("Labor", "=BidLabor", "=CrewCost"),
        ("Subcontractors", "=BidSubs", "=SubInvoiced"),
        ("Equipment & rentals", "=BidEquip", "=EquipTotal"),
        ("Overhead", "=Overhead", "=Overhead"),
    ]
    start = 6
    for i, (lab, est, act) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        ce = ws.cell(row=r, column=3, value=est); ce.style = "td"; ce.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=act); ca.style = "td"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL COST").style = "th"
    ce = ws.cell(row=tot, column=3, value="=TotalCost"); ce.style = "td"; ce.font = Font(bold=True, color=PRIMARY); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    ca = ws.cell(row=tot, column=4, value=ACTUAL_COST); ca.style = "input"; ca.font = Font(bold=True, color=PRIMARY); ca.number_format = '"$"#,##0'
    cell_name(wb, "ActualCost", "Job Costing", f"$D${tot}")
    ws.cell(row=tot + 2, column=2, value="Contract price (the bid)").style = "field_label"
    cb = ws.cell(row=tot + 2, column=4, value="=BidPrice"); cb.style = "field_value"; cb.number_format = '"$"#,##0'
    ws.cell(row=tot + 3, column=2, value="= ACTUAL PROFIT").style = "th"
    cap = ws.cell(row=tot + 3, column=4, value="=BidPrice-ActualCost"); cap.style = "td"; cap.font = Font(bold=True, size=14, color=PRIMARY); cap.fill = fill(MINT_BG); cap.number_format = '"$"#,##0'
    cell_name(wb, "ActualProfit", "Job Costing", f"$D${tot+3}")
    ws.cell(row=tot + 4, column=2, value="= ACTUAL MARGIN").style = "th"
    cam = ws.cell(row=tot + 4, column=4, value="=IFERROR(ActualProfit/BidPrice,0)"); cam.style = "td"; cam.font = Font(bold=True, size=13, color=PRIMARY); cam.fill = fill(MINT_BG); cam.number_format = "0.0%"
    cell_name(wb, "ActualMargin", "Job Costing", f"$D${tot+4}")
    ws.cell(row=tot + 5, column=2, value="Variance vs estimate (under / over)").style = "field_label"
    cv = ws.cell(row=tot + 5, column=4, value="=TotalCost-ActualCost"); cv.style = "field_value"; cv.number_format = '"$"#,##0'
    ws.cell(row=tot + 7, column=2, value="A job that beats its estimate tells you to keep bidding that way. One that misses tells you why.").style = "section_gold"


def build_jobs(wb):
    ws = wb.create_sheet("Jobs & Pipeline"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 16, 16, 16, 14, 2])
    luxe_header(ws, "G", "🏗  JOBS & PIPELINE",
                "Every job — bid, estimated cost, expected profit and where it stands right now.")
    table_headers(ws, 4, ["Job", "Client", "Status", "Bid", "Est. Cost", "Est. Profit"], start_col=2)
    start = L0
    for i, (job, client, status, bid, cost) in enumerate(JOBS):
        r = start + i
        ws.cell(row=r, column=2, value=job).style = "td_left"
        ws.cell(row=r, column=3, value=client).style = "td"
        ws.cell(row=r, column=4, value=status).style = "td"
        cb = ws.cell(row=r, column=5, value=bid); cb.style = "input"; cb.number_format = '"$"#,##0'
        cc = ws.cell(row=r, column=6, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=7, value=f"=E{r}-F{r}"); cp.style = "td"; cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(JOBS) - 1
    nrange(wb, "JobName", "Jobs & Pipeline", "B", start, end)
    nrange(wb, "JobStatus", "Jobs & Pipeline", "D", start, end)
    nrange(wb, "JobBid", "Jobs & Pipeline", "E", start, end)
    nrange(wb, "JobCost", "Jobs & Pipeline", "F", start, end)
    add_dv(ws, f"D{start}:D{end}", "JobStatusList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL JOBS").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cb = ws.cell(row=tot, column=5, value="=SUM(JobBid)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.fill = fill(SURFACE); cb.number_format = '"$"#,##0'
    cc = ws.cell(row=tot, column=6, value="=SUM(JobCost)"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0'
    cp = ws.cell(row=tot, column=7, value=f"=E{tot}-F{tot}"); cp.style = "td"; cp.font = Font(bold=True, size=12, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = '"$"#,##0'
    cell_name(wb, "AllBids", "Jobs & Pipeline", f"$E${tot}")
    cell_name(wb, "AllJobProfit", "Jobs & Pipeline", f"$G${tot}")
    ws.cell(row=tot + 2, column=2, value="Backlog (won + in progress)").style = "field_label"
    cbl = ws.cell(row=tot + 2, column=5, value='=SUMIF(JobStatus,"Won",JobBid)+SUMIF(JobStatus,"In progress",JobBid)')
    cbl.style = "field_value"; cbl.number_format = '"$"#,##0'; cbl.fill = fill(MINT_BG)
    cell_name(wb, "Backlog", "Jobs & Pipeline", f"$E${tot+2}")
    ws.cell(row=tot + 3, column=2, value="Jobs this year").style = "field_label"
    cjc = ws.cell(row=tot + 3, column=5, value="=COUNTA(JobName)"); cjc.style = "field_value"; cjc.number_format = "#,##0"
    cell_name(wb, "JobCount", "Jobs & Pipeline", f"$E${tot+3}")
    ws.freeze_panes = "A5"


def build_crew(wb):
    ws = wb.create_sheet("Labor & Crew"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 18, 14, 14, 16, 2])
    luxe_header(ws, "F", "👷  LABOR & CREW",
                "Who's on the job, at what rate, for how many hours — the biggest and least-tracked cost you have.")
    table_headers(ws, 4, ["Name", "Trade", "$ / Hour", "Hours", "Cost"], start_col=2)
    start = L0
    for i, (name, trade, rate, hours) in enumerate(CREW):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=trade).style = "td"
        cr = ws.cell(row=r, column=4, value=rate); cr.style = "input"; cr.number_format = '"$"#,##0'
        ch = ws.cell(row=r, column=5, value=hours); ch.style = "input"; ch.number_format = "#,##0"
        cc = ws.cell(row=r, column=6, value=f"=D{r}*E{r}"); cc.style = "td"; cc.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CREW) - 1
    nrange(wb, "CrewHours", "Labor & Crew", "E", start, end)
    nrange(wb, "CrewLine", "Labor & Crew", "F", start, end)
    add_dv(ws, f"C{start}:C{end}", "TradeList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="CREW TOTAL").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    ch = ws.cell(row=tot, column=5, value="=SUM(CrewHours)"); ch.style = "td"; ch.font = Font(bold=True, color=PRIMARY); ch.fill = fill(SURFACE); ch.number_format = "#,##0"
    cc = ws.cell(row=tot, column=6, value="=SUM(CrewLine)"); cc.style = "td"; cc.font = Font(bold=True, size=12, color=PRIMARY); cc.fill = fill(MINT_BG); cc.number_format = '"$"#,##0'
    cell_name(wb, "CrewCost", "Labor & Crew", f"$F${tot}")
    cell_name(wb, "CrewHoursTotal", "Labor & Crew", f"$E${tot}")
    ws.freeze_panes = "A5"


def build_materials(wb):
    ws = wb.create_sheet("Materials"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 12, 16, 16, 2])
    luxe_header(ws, "E", "🧱  MATERIALS",
                "Every material line on the job — what you allowed for and what it actually came to.")
    table_headers(ws, 4, ["Item", "Qty", "Unit Cost", "Total"], start_col=2)
    start = L0
    for i, (item, qty, unit, total) in enumerate(MATERIALS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cq = ws.cell(row=r, column=3, value=qty); cq.style = "input"; cq.number_format = "#,##0"
        cu = ws.cell(row=r, column=4, value=unit); cu.style = "input"; cu.number_format = '"$"#,##0'
        ct = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); ct.style = "td"; ct.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MATERIALS) - 1
    nrange(wb, "MatLine", "Materials", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="MATERIALS TOTAL").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=5, value="=SUM(MatLine)"); ct.style = "td"; ct.font = Font(bold=True, size=12, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0'
    cell_name(wb, "MaterialTotal", "Materials", f"$E${tot}")
    ws.freeze_panes = "A5"


def build_subs(wb):
    ws = wb.create_sheet("Subcontractors"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 16, 16, 2])
    luxe_header(ws, "E", "🔧  SUBCONTRACTORS",
                "What each sub quoted against what they actually invoiced — where jobs quietly go over.")
    table_headers(ws, 4, ["Subcontractor", "Trade", "Quoted", "Invoiced"], start_col=2)
    start = L0
    for i, (sub, trade, quoted, invoiced) in enumerate(SUBS):
        r = start + i
        ws.cell(row=r, column=2, value=sub).style = "td_left"
        ws.cell(row=r, column=3, value=trade).style = "td"
        cq = ws.cell(row=r, column=4, value=quoted); cq.style = "input"; cq.number_format = '"$"#,##0'
        ci = ws.cell(row=r, column=5, value=invoiced); ci.style = "input"; ci.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SUBS) - 1
    nrange(wb, "SubQuoted", "Subcontractors", "D", start, end)
    nrange(wb, "SubInv", "Subcontractors", "E", start, end)
    add_dv(ws, f"C{start}:C{end}", "TradeList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="SUBS TOTAL").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cq = ws.cell(row=tot, column=4, value="=SUM(SubQuoted)"); cq.style = "td"; cq.font = Font(bold=True, color=PRIMARY); cq.fill = fill(SURFACE); cq.number_format = '"$"#,##0'
    ci = ws.cell(row=tot, column=5, value="=SUM(SubInv)"); ci.style = "td"; ci.font = Font(bold=True, size=12, color=PRIMARY); ci.fill = fill(MINT_BG); ci.number_format = '"$"#,##0'
    cell_name(wb, "SubInvoiced", "Subcontractors", f"$E${tot}")
    ws.freeze_panes = "A5"


def build_changes(wb):
    ws, start, end = build_log(
        wb, "Change Orders", "📝", "CHANGE ORDERS",
        "Every scope change, priced and approved in writing — unbilled change orders are pure lost profit.",
        ["Job", "Change", "Amount", "Approved?"],
        CHANGES, [2, 26, 24, 14, 14, 2], text_left={2, 3}, money={4}, reserved=26, start_col=2,
        validations=[("E", "YesNoList")])
    nrange(wb, "ChangeAmt", "Change Orders", "D", start, end)
    nrange(wb, "ChangeOK", "Change Orders", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="APPROVED CHANGE ORDERS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value='=SUMIF(ChangeOK,"Yes",ChangeAmt)'); c.style = "td"; c.font = Font(bold=True, size=12, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "ChangeOrders", "Change Orders", f"$D${tot}")
    ws.cell(row=tot + 2, column=2, value="No signature, no work. A verbal change order is a donation.").style = "section_gold"


def build_equipment(wb):
    ws = wb.create_sheet("Equipment"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 16, 2])
    luxe_header(ws, "D", "🚜  EQUIPMENT",
                "Rentals and machine time charged to this job — small lines that add up fast.")
    table_headers(ws, 4, ["Item", "Owned / Rented", "Cost"], start_col=2)
    start = L0
    for i, (item, own, cost) in enumerate(EQUIPMENT):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ws.cell(row=r, column=3, value=own).style = "td"
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EQUIPMENT) - 1
    nrange(wb, "EquipCost", "Equipment", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="EQUIPMENT TOTAL").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=4, value="=SUM(EquipCost)"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0'
    cell_name(wb, "EquipTotal", "Equipment", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_invoices(wb):
    ws, start, end = build_log(
        wb, "Invoices", "💳", "INVOICES & RECEIVABLES",
        "What you've billed and what's still outstanding — cash flow kills more contractors than bad jobs.",
        ["Job", "Amount", "Status"],
        INVOICES, [2, 30, 16, 16, 2], text_left={2}, money={3}, reserved=26, start_col=2)
    nrange(wb, "InvAmt", "Invoices", "C", start, end)
    nrange(wb, "InvStatus", "Invoices", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="INVOICED").style = "th"
    ci = ws.cell(row=tot, column=3, value="=SUM(InvAmt)"); ci.style = "td"; ci.font = Font(bold=True, color=PRIMARY); ci.fill = fill(SURFACE); ci.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    ws.cell(row=tot + 1, column=2, value="Outstanding (sent, unpaid)").style = "field_label"
    co = ws.cell(row=tot + 1, column=3, value='=SUMIF(InvStatus,"Sent",InvAmt)'); co.style = "field_value"; co.number_format = '"$"#,##0'; co.fill = fill(WARN_BG)
    cell_name(wb, "Receivable", "Invoices", f"$C${tot+1}")


def build_bidlog(wb):
    ws = wb.create_sheet("Bid Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 16, 2])
    luxe_header(ws, "D", "🎯  BID LOG",
                "Every bid you submitted and how it landed — your win rate tells you if you're priced right.")
    table_headers(ws, 4, ["Job", "Bid Amount", "Result"], start_col=2)
    start = L0
    for i, (job, amt, result) in enumerate(BIDLOG):
        r = start + i
        ws.cell(row=r, column=2, value=job).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=result).style = "td"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BIDLOG) - 1
    nrange(wb, "BidAmt", "Bid Log", "C", start, end)
    nrange(wb, "BidResult", "Bid Log", "D", start, end)
    add_dv(ws, f"D{start}:D{end}", "BidResultList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="BID TOTAL").style = "th"
    cb = ws.cell(row=tot, column=3, value="=SUM(BidAmt)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.fill = fill(SURFACE); cb.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    ws.cell(row=tot + 2, column=2, value="Bids won / submitted").style = "field_label"
    cw = ws.cell(row=tot + 2, column=3, value='=COUNTIF(BidResult,"Won")&" / "&BidsTotal'); cw.style = "field_value"
    ws.cell(row=tot + 3, column=2, value="= WIN RATE").style = "th"
    cwr = ws.cell(row=tot + 3, column=3, value="=IFERROR(BidsWon/BidsTotal,0)"); cwr.style = "td"; cwr.font = Font(bold=True, size=13, color=PRIMARY); cwr.fill = fill(MINT_BG); cwr.number_format = "0.0%"
    cell_name(wb, "WinRate", "Bid Log", f"$C${tot+3}")
    ws.cell(row=tot + 5, column=2, value="Winning every bid means you're too cheap. Winning none means you're too dear.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "Revenue month by month — and the net margin that says whether the year is working.")
    ws.cell(row=5, column=2, value="THE YEAR").style = "section_gold"
    ws.cell(row=6, column=2, value="Revenue (all jobs bid & won)").style = "field_label"
    cr = ws.cell(row=6, column=3, value=428000); cr.style = "input"; cr.number_format = '"$"#,##0'
    cell_name(wb, "RevenueYTD", "Monthly Summary", "$C$6")
    ws.cell(row=7, column=2, value="Net profit").style = "field_label"
    cn = ws.cell(row=7, column=3, value=79000); cn.style = "input"; cn.number_format = '"$"#,##0'
    cell_name(wb, "NetProfit", "Monthly Summary", "$C$7")
    ws.cell(row=8, column=2, value="= NET MARGIN").style = "th"
    cm = ws.cell(row=8, column=3, value="=IFERROR(NetProfit/RevenueYTD,0)"); cm.style = "td"; cm.font = Font(bold=True, size=13, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = "0.0%"
    cell_name(wb, "NetMargin", "Monthly Summary", "$C$8")
    ws.cell(row=10, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 11, ["Month", "Revenue"], start_col=2)
    ts = 12
    for i, (m, rev) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        crv = ws.cell(row=r, column=3, value=rev); crv.style = "input"; crv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "RevTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Revenue by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  📐  CONTRACTOR JOB COSTING & BIDDING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Bid price, job margin, backlog & a Builder Score — your whole operation, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("DIRECT COST", "=DirectCost", "money"),
        ("OVERHEAD", "=Overhead", "money"),
        ("TOTAL COST", "=TotalCost", "money"),
        ("BID THIS JOB AT", "=BidPrice", "money"),
        ("PLANNED PROFIT", "=PlannedProfit", "money"),
        ("ACTUAL MARGIN", "=ActualMargin", "pct1"),
    ]
    row2 = [
        ("ACTUAL PROFIT", "=ActualProfit", "money"),
        ("JOBS THIS YEAR", "=JobCount", "num"),
        ("BACKLOG", "=Backlog", "money"),
        ("RECEIVABLES", "=Receivable", "money"),
        ("WIN RATE", "=WinRate", "pct1"),
        ("BUILDER SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "BUILDER HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Job margin healthy", "=IFERROR(MIN(ActualMargin/MarginGoal,1),0)"),
        ("Backlog full", "=IFERROR(MIN(Backlog/BacklogGoal,1),0)"),
        ("Net margin healthy", "=IFERROR(MIN(NetMargin/NetMarginGoal,1),0)"),
        ("Getting paid on time", "=IF(Receivable<=ReceivableGoal,1,IFERROR(ReceivableGoal/Receivable,0))"),
        ("Jobs booked", "=IFERROR(MIN(JobCount/JobGoal,1),0)"),
        ("Bid win rate", "=IFERROR(MIN(WinRate/WinRateGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=12, max_row=11 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=12, max_row=11 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Contractor Job Costing & Bidding Command Center™ — markup is not margin.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    wb.defined_names["WinRateGoal"] = DefinedName("WinRateGoal", attr_text=str(WINRATE_GOAL))
    build_settings(wb); build_start(wb); build_bid(wb); build_jobcost(wb)
    build_jobs(wb); build_crew(wb); build_materials(wb); build_subs(wb)
    build_changes(wb); build_equipment(wb); build_invoices(wb); build_bidlog(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Bid Builder", "Job Costing", "Jobs & Pipeline", "Labor & Crew",
             "Materials", "Subcontractors", "Change Orders", "Equipment", "Invoices", "Bid Log",
             "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Contractor_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
