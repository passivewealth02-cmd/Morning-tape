"""Build College Application Command Center™ — The Complete Admissions System.

15 tabs (+ Settings) · a premium college-admissions operating system in Google
Sheets & Excel. Dashboard, applicant profile, a college list (reach/match/safety
with per-school progress), essay & supplement tracker, recommendations, test
scores, activities & awards, scholarships, financial-aid / net-price compare,
visits & interviews, decisions & comparison, a to-do list and a deadlines
calendar — one dashboard. Apply smart, hit every deadline, compare offers.

Run: python3 build_xlsx.py   ->  ../College_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

TYPES = ["Reach", "Match", "Safety"]
YESNO = ["Yes", "No"]
ESSAY_STATUS = ["Not Started", "Drafting", "Revising", "Final"]
APP_TYPE = ["Common App", "Coalition", "Direct", "UC App"]
DECISION = ["Waiting", "Accepted", "Waitlist", "Denied", "Deferred"]
SCH_STATUS = ["Researching", "Applied", "Awarded", "Denied"]
REC_ROLE = ["Counselor", "Teacher — Core", "Teacher — Elective", "Coach / Mentor", "Other"]

# ---- sample applicant (Ella Bennett, Class of 2027 — continuity w/ the transcript) ----
STUDENT = {"name": "Ella Bennett", "class": "2027", "gpa_w": 4.06, "gpa_uw": 3.81,
           "sat": 1380, "act": 30, "major": "Environmental Science"}

# College list: (name, type, deadline offset, app fee, essays?, recs?, form?, submitted?)
COLLEGES = [
    ("Riverside State", "Safety", -18, 50, "Yes", "Yes", "Yes", "Yes"),
    ("State Flagship (EA)", "Match", -5, 60, "Yes", "Yes", "Yes", "Yes"),
    ("Community Honors", "Safety", 40, 0, "Yes", "Yes", "Yes", "Yes"),
    ("Lakeside College", "Match", 10, 65, "Yes", "Yes", "Yes", "Yes"),
    ("Hillcrest University", "Match", 20, 70, "Yes", "Yes", "Yes", "No"),
    ("Coastal U", "Reach", 24, 75, "Yes", "Yes", "No", "No"),
    ("Northwood University", "Reach", 30, 75, "Yes", "No", "No", "No"),
    ("Valley Tech", "Reach", 30, 80, "No", "No", "No", "No"),
]

# Essays & supplements: (school, prompt, word limit, status)
ESSAYS = [
    ("Common App", "Personal statement", 650, "Final"),
    ("State Flagship", "Why us? (supplement)", 300, "Final"),
    ("Lakeside College", "Community & you", 250, "Final"),
    ("Hillcrest University", "Academic interest", 400, "Final"),
    ("Coastal U", "Why this major?", 350, "Revising"),
    ("Coastal U", "Short answer: activity", 150, "Final"),
    ("Northwood University", "Intellectual curiosity", 500, "Revising"),
    ("Northwood University", "Community contribution", 250, "Drafting"),
    ("Valley Tech", "Engineering interest", 400, "Drafting"),
    ("Valley Tech", "Short answer: challenge", 200, "Not Started"),
]

# Recommendations: (recommender, role, requested?, submitted?)
RECS = [
    ("Ms. Alvarez", "Counselor", "Yes", "Yes"),
    ("Mr. Chen (Biology)", "Teacher — Core", "Yes", "Yes"),
    ("Ms. Patel (English)", "Teacher — Core", "Yes", "Yes"),
    ("Coach Rivera", "Coach / Mentor", "Yes", "No"),
]

# Test scores: (test, date, score, detail)
TESTS = [
    ("SAT", "Mar 2026", "1380", "EBRW 690 · Math 690"),
    ("ACT", "Apr 2026", "30", "Composite"),
    ("AP Environmental Sci", "May 2026", "4", "Qualified"),
    ("AP US Government", "May 2026", "4", "Qualified"),
    ("TOEFL / other", "—", "—", "N/A"),
]

# Activities & awards: (activity / award, role, years)
ACTIVITIES = [
    ("Environmental Club", "President", "3 yrs"),
    ("Varsity Cross Country", "Captain", "4 yrs"),
    ("Science Olympiad", "Team lead", "2 yrs"),
    ("Local River Cleanup", "Volunteer organizer", "3 yrs"),
    ("National Honor Society", "Member", "2 yrs"),
    ("Part-time job — nursery", "Associate", "2 yrs"),
    ("AP Scholar Award", "Honor", "2026"),
    ("Regional Science Fair — 1st", "Award", "2026"),
]

# Scholarships: (name, amount, deadline offset, status)
SCHOLARSHIPS = [
    ("State Merit Scholarship", 8000, -20, "Awarded"),
    ("Green Future Environmental", 4000, -12, "Awarded"),
    ("Community Foundation", 2500, 15, "Applied"),
    ("Rotary Club Local", 1500, 22, "Applied"),
    ("First-Gen STEM Grant", 5000, 30, "Applied"),
    ("Coca-Cola Scholars", 20000, -40, "Denied"),
    ("Athletic Booster Award", 1000, 18, "Applied"),
    ("National Merit", 2500, 45, "Researching"),
]

# Financial aid / net price: (school, sticker price, grants & aid, )
NETPRICE = [
    ("Riverside State", 24000, 12000),
    ("State Flagship", 28000, 9500),
    ("Community Honors", 8000, 6000),
    ("Lakeside College", 54000, 34000),
    ("Hillcrest University", 48000, 26000),
    ("Coastal U", 62000, 30000),
]

# Visits & interviews: (date offset, school, type, notes)
VISITS = [
    (-60, "State Flagship", "Campus tour", "Loved the sustainability program"),
    (-45, "Lakeside College", "Info session", "Small classes, strong advising"),
    (-30, "Riverside State", "Virtual tour", "Great value, close to home"),
    (12, "Coastal U", "Alumni interview", "Prep: why this major"),
    (18, "Hillcrest University", "Campus visit", "Scheduled"),
]

# Decisions: (school, decision, date offset)
DECISIONS = [
    ("Riverside State", "Accepted", -10),
    ("Community Honors", "Accepted", -6),
    ("State Flagship (EA)", "Deferred", -3),
    ("Lakeside College", "Waiting", 30),
    ("Hillcrest University", "Waiting", 45),
]

# To-do tasks: (task, done?)
TASKS = [
    ("Finalize Common App personal statement", "Yes"),
    ("Submit State Flagship EA", "Yes"),
    ("Request all recommendation letters", "Yes"),
    ("Send official test scores", "Yes"),
    ("Submit FAFSA", "Yes"),
    ("Finish Coastal U 'why major' essay", "Yes"),
    ("Submit Riverside State", "Yes"),
    ("Submit Community Honors", "Yes"),
    ("Draft Northwood community essay", "Yes"),
    ("Submit Lakeside College", "Yes"),
    ("Follow up with Coach Rivera on rec", "No"),
    ("Submit Hillcrest University", "No"),
    ("Finish Valley Tech essays", "No"),
    ("Submit Coastal U", "No"),
    ("Apply: Community Foundation scholarship", "Yes"),
    ("Apply: Rotary scholarship", "Yes"),
    ("Compare aid offers (net price)", "No"),
    ("Submit Northwood University", "No"),
    ("Submit Valley Tech", "No"),
    ("Send thank-you notes to recommenders", "No"),
]

SCHOLARSHIP_GOAL = 8
LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your applicant details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "APPLICANT", "section")
    controls = [
        ("Applicant Name", STUDENT["name"], None, "Applicant"),
        ("Class Of", STUDENT["class"], None, "ClassOf"),
        ("Weighted GPA", STUDENT["gpa_w"], "0.00", "GPAW"),
        ("Best SAT", STUDENT["sat"], "0", "BestSAT"),
        ("Best ACT", STUDENT["act"], "0", "BestACT"),
        ("Intended Major", STUDENT["major"], None, "Major"),
        ("Scholarship Goal (# applied)", SCHOLARSHIP_GOAL, "0", "SchGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Type", TYPES, "TypeList"), ("F", "Yes / No", YESNO, "YesNoList"),
             ("G", "Essay", ESSAY_STATUS, "EssayList"), ("H", "Decision", DECISION, "DecisionList"),
             ("I", "Scholarship", SCH_STATUS, "SchList"), ("J", "Rec Role", REC_ROLE, "RecList"),
             ("K", "App Type", APP_TYPE, "AppList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🎓  COLLEGE APPLICATION COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Apply smart, hit every deadline, compare offers — the whole admissions season, organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE APPLICATION SEASON, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A balanced college list (reach / match / safety) with per-school progress, an essay & supplement "
                      "tracker, recommendations, test scores, an activities résumé, scholarships, a financial-aid & "
                      "net-price comparison, visits & interviews, a decisions tracker, a master to-do list and a "
                      "deadlines calendar — all in ONE premium Google Sheets & Excel system. See exactly what's left on "
                      "every application, never miss a deadline, and compare offers apples-to-apples when they roll in.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your name, class year, GPA & test scores.",
             "2.  Build your College List — mark each Reach, Match or Safety with its deadline.",
             "3.  For each school, check off essays, recs, the form & submission — progress calculates itself.",
             "4.  Track essays & supplements, secure recommendations, and apply for scholarships.",
             "5.  Log visits & interviews, and record each decision as it arrives.",
             "6.  Watch the Dashboard track apps submitted, deadlines, aid & a Ready Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional applicant (Ella, Class of 2027) is included so you can see how it all "
               "connects — just type over it with your own. Your college list, essays, recs, scholarships and decisions "
               "roll up into a live Ready Score and a deadlines view. Twelve matching printable pages (college list, "
               "essay tracker, rec tracker, scholarship log, net-price compare & more) are included to print and keep. "
               "This is an organizing tool — confirm every deadline, fee and requirement directly with each college.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Apply smart, breathe easier — you've got a plan for every school.", "section_gold")


# ===========================================================================
# 3 — Applicant Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Applicant Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 22, 4, 22, 22, 2])
    luxe_header(ws, "F", "🧑‍🎓  APPLICANT PROFILE",
                "Your stats at a glance — the snapshot that shapes your whole college list.")
    merge_set(ws, "B5:F5", STUDENT["name"] + " · Class of " + STUDENT["class"], "section_gold"); ws.row_dimensions[5].height = 22
    fields = [("Weighted GPA", "4.06"), ("Unweighted GPA", "3.81"),
              ("Best SAT", "1380"), ("Best ACT", "30"),
              ("Class Rank", "Top 10%"), ("Intended Major", "Environmental Science"),
              ("Second Major / Interest", "Policy"), ("Recommenders", "See Recs tab"),
              ("Hooks / spikes", "Environmental leadership"), ("Fee waiver eligible?", "—"),
              ("FAFSA submitted?", "Yes"), ("Notes", "Loves field research")]
    row = 6; i = 0
    while i < len(fields):
        ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
        ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
        if i + 1 < len(fields):
            ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
            ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
        ws.row_dimensions[row].height = 24; i += 2; row += 1


# ===========================================================================
# 4 — College List  (defines CollegeName / CollegeType / CollegeProg / CollegeSubmitted / CollegeDeadline)
# ===========================================================================
def build_collegelist(wb):
    ws = wb.create_sheet("College List"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 12, 14, 8, 8, 8, 12, 12, 2])
    luxe_header(ws, "I", "🏫  COLLEGE LIST",
                "A balanced list — reach, match & safety — with per-school progress that calculates itself.")
    table_headers(ws, 4, ["College", "Type", "Deadline", "Essays", "Recs", "Form", "Submitted", "Progress"], start_col=2)
    start = L0
    for i, (name, typ, doff, fee, ess, rec, form, sub) in enumerate(COLLEGES):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=typ).style = "td"
        cd = ws.cell(row=r, column=4, value=dplus(doff) if doff >= 0 else dminus(-doff)); cd.style = "td"; cd.number_format = "mm/dd/yyyy"
        for j, val in enumerate([ess, rec, form, sub]):
            cv = ws.cell(row=r, column=5 + j, value=val); cv.style = "input"
        cp = ws.cell(row=r, column=9, value=f"=IFERROR(COUNTIF(E{r}:H{r},\"Yes\")/4,0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 10):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(COLLEGES) - 1
    nrange(wb, "CollegeName", "College List", "B", start, end)
    nrange(wb, "CollegeType", "College List", "C", start, end)
    nrange(wb, "CollegeDeadline", "College List", "D", start, end)
    nrange(wb, "CollegeSubmitted", "College List", "H", start, end)
    nrange(wb, "CollegeProg", "College List", "I", start, end)
    cmap = {"Reach": WARN_BG, "Match": MINT_BG, "Safety": SOFT_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    for col in ("E", "F", "G", "H"):
        ws.conditional_formatting.add(f"{col}{start}:{col}{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"I{start}:I{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
# 5 — Essays & Supplements  (defines EssayStatus)
# ===========================================================================
def build_essays(wb):
    ws, start, end = build_log(
        wb, "Essays", "✍", "ESSAYS & SUPPLEMENTS",
        "Every prompt, its word limit & where the draft stands — from blank page to final.",
        ["School", "Prompt", "Word Limit", "Status"],
        ESSAYS, [22, 30, 14, 14], text_left={2}, ints={3}, reserved=40,
        validations=[("D", "EssayList")])
    nrange(wb, "EssayStatus", "Essays", "D", start, end)
    cmap = {"Final": MINT_BG, "Revising": WARN_BG, "Drafting": SOFT_BG, "Not Started": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 6 — Recommendations  (defines RecSubmitted)
# ===========================================================================
def build_recs(wb):
    ws, start, end = build_log(
        wb, "Recommendations", "📨", "RECOMMENDATIONS",
        "Who's writing, whether you've asked & whether it's in — chase the stragglers early.",
        ["Recommender", "Role", "Requested?", "Submitted?"],
        RECS, [24, 20, 14, 14], text_left={1}, reserved=16,
        validations=[("B", "RecList"), ("C", "YesNoList"), ("D", "YesNoList")])
    nrange(wb, "RecSubmitted", "Recommendations", "D", start, end)
    for col in ("C", "D"):
        ws.conditional_formatting.add(f"{col}{start}:{col}{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
        ws.conditional_formatting.add(f"{col}{start}:{col}{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 7 — Test Scores
# ===========================================================================
def build_tests(wb):
    build_log(wb, "Test Scores", "📊", "TEST SCORES",
              "SAT, ACT, AP & more — the scores you'll send, in one place.",
              ["Test", "Date", "Score", "Detail"],
              TESTS, [24, 16, 14, 26], text_left={4}, reserved=16)


# ===========================================================================
# 8 — Activities & Awards
# ===========================================================================
def build_activities(wb):
    build_log(wb, "Activities", "🏅", "ACTIVITIES & AWARDS",
              "Your résumé for the application — activity, role & years, plus honors.",
              ["Activity / Award", "Role", "Years"],
              ACTIVITIES, [40, 22, 14], text_left={1, 2}, reserved=24)


# ===========================================================================
# 9 — Scholarships  (defines SchStatus / SchAmount)
# ===========================================================================
def build_scholarships(wb):
    rows = [(name, amt, dplus(off) if off >= 0 else dminus(-off), st) for (name, amt, off, st) in SCHOLARSHIPS]
    ws, start, end = build_log(
        wb, "Scholarships", "💰", "SCHOLARSHIPS & AID",
        "Every scholarship — amount, deadline & status. Free money is worth the extra essays.",
        ["Scholarship", "Amount", "Deadline", "Status"],
        rows, [30, 14, 14, 16], text_left={1}, money={2}, dates={3}, reserved=30,
        validations=[("D", "SchList")])
    nrange(wb, "SchName", "Scholarships", "A", start, end)
    nrange(wb, "SchAmount", "Scholarships", "B", start, end)
    nrange(wb, "SchStatus", "Scholarships", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="AWARDED TO DATE").style = "th"
    c = ws.cell(row=tot, column=2, value='=SUMIF(SchStatus,"Awarded",SchAmount)')
    c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cell_name(wb, "AidAwarded", "Scholarships", f"$B${tot}")
    cmap = {"Awarded": MINT_BG, "Applied": WARN_BG, "Researching": SOFT_BG, "Denied": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 10 — Financial Aid / Net Price  (defines NetPrice)
# ===========================================================================
def build_netprice(wb):
    ws = wb.create_sheet("Net Price"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 16, 16, 2])
    luxe_header(ws, "E", "🧾  FINANCIAL AID — NET PRICE",
                "Sticker price is a myth — compare what each school actually costs after grants & aid.")
    table_headers(ws, 4, ["College", "Sticker / Year", "Grants & Aid", "Net Price"], start_col=2)
    start = L0
    for i, (name, sticker, aid) in enumerate(NETPRICE):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sticker); cs.style = "input"; cs.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=aid); ca.style = "input"; ca.number_format = '"$"#,##0'
        cn = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(NETPRICE) - 1
    nrange(wb, "NetPrice", "Net Price", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="min", start_color="FF" + HIGHLIGHT, end_type="max", end_color="FF" + RED_BG))
    merge_set(ws, f"B{end+2}:E{end+2}",
              "Lowest net price is highlighted green. Compare true cost — not the sticker — when you decide.",
              "section_gold")
    ws.freeze_panes = "A5"


# ===========================================================================
# 11 — Visits & Interviews
# ===========================================================================
def build_visits(wb):
    rows = [(dplus(off) if off >= 0 else dminus(-off), sch, typ, note) for (off, sch, typ, note) in VISITS]
    build_log(wb, "Visits", "🚗", "VISITS & INTERVIEWS",
              "Tours, info sessions & interviews — with your impressions while they're fresh.",
              ["Date", "School", "Type", "Notes / Impression"],
              rows, [14, 22, 18, 32], text_left={2, 4}, dates={1}, reserved=24)


# ===========================================================================
# 12 — Decisions & Compare  (defines Decision)
# ===========================================================================
def build_decisions(wb):
    rows = [(sch, dec, dplus(off) if off >= 0 else dminus(-off)) for (sch, dec, off) in DECISIONS]
    ws, start, end = build_log(
        wb, "Decisions", "🎉", "DECISIONS & COMPARISON",
        "Every result as it arrives — accepted, waitlist or deferred — so you can compare & choose.",
        ["College", "Decision", "Date / Notify"],
        rows, [26, 18, 18], text_left={1}, dates={3}, reserved=20,
        validations=[("B", "DecisionList")])
    nrange(wb, "Decision", "Decisions", "B", start, end)
    cmap = {"Accepted": MINT_BG, "Waitlist": WARN_BG, "Deferred": SOFT_BG, "Denied": RED_BG, "Waiting": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 13 — To-Do  (defines TaskName / TaskDone)
# ===========================================================================
def build_tasks(wb):
    ws, start, end = build_log(
        wb, "To-Do", "✅", "MASTER TO-DO LIST",
        "Everything that has to happen this season — check it off & watch the finish line get closer.",
        ["Task", "Done?"],
        TASKS, [56, 14], text_left={1}, reserved=40,
        validations=[("B", "YesNoList")])
    nrange(wb, "TaskName", "To-Do", "A", start, end)
    nrange(wb, "TaskDone", "To-Do", "B", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 14 — Deadlines Calendar
# ===========================================================================
def build_deadlines(wb):
    ws = wb.create_sheet("Deadlines"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 20, 2])
    luxe_header(ws, "D", "🗓  DEADLINES CALENDAR",
                "The dates that matter, in order — applications, scholarships, aid & decisions.")
    table_headers(ws, 4, ["What's Due", "Date", "Type"], start_col=2)
    items = [
        ("State Flagship — EA", dminus(5), "Application"),
        ("Riverside State", dminus(18), "Application"),
        ("Lakeside College", dplus(10), "Application"),
        ("Community Foundation scholarship", dplus(15), "Scholarship"),
        ("Athletic Booster scholarship", dplus(18), "Scholarship"),
        ("Hillcrest University", dplus(20), "Application"),
        ("Rotary scholarship", dplus(22), "Scholarship"),
        ("Coastal U", dplus(24), "Application"),
        ("Northwood University", dplus(30), "Application"),
        ("Valley Tech", dplus(30), "Application"),
        ("First-Gen STEM grant", dplus(30), "Scholarship"),
        ("Community Honors (rolling)", dplus(40), "Application"),
        ("National Merit", dplus(45), "Scholarship"),
    ]
    items.sort(key=lambda x: x[1])
    start = L0
    for i, (what, date, typ) in enumerate(items):
        r = start + i
        ws.cell(row=r, column=2, value=what).style = "td_left"
        cd = ws.cell(row=r, column=3, value=date); cd.style = "td"; cd.number_format = "mm/dd/yyyy"
        ws.cell(row=r, column=4, value=typ).style = "td"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


# ===========================================================================
# 2 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🎓  COLLEGE APPLICATION COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Apps, essays, deadlines & aid — your whole admissions season, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("COLLEGES", "=COUNTA(CollegeName)", "num"),
        ("APPS SUBMITTED", '=COUNTIF(CollegeSubmitted,"Yes")', "num"),
        ("AVG PROGRESS", "=IFERROR(AVERAGE(CollegeProg),0)", "pct"),
        ("ESSAYS DONE", '=IFERROR(COUNTIF(EssayStatus,"Final")/COUNTA(EssayStatus),0)', "pct"),
        ("RECS SECURED", '=COUNTIF(RecSubmitted,"Yes")', "num"),
        ("SCHOLARSHIPS", '=COUNTIF(SchStatus,"Applied")+COUNTIF(SchStatus,"Awarded")', "num"),
    ]
    row2 = [
        ("AID AWARDED", "=AidAwarded", "money"),
        ("NEXT DEADLINE", '=IFERROR(MINIFS(CollegeDeadline,CollegeSubmitted,"No"),"")', "date"),
        ("ACCEPTANCES", '=COUNTIF(Decision,"Accepted")', "num"),
        ("BEST NET PRICE", "=IFERROR(MIN(NetPrice),0)", "money"),
        ("TASKS DONE", '=IFERROR(COUNTIF(TaskDone,"Yes")/COUNTA(TaskName),0)', "pct"),
        ("READY SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "ARE WE ON TRACK?", "section_gold")
    merge_set(ws, "H11:M11", "COLLEGE LIST BALANCE", "section_gold")
    table_headers(ws, 12, ["Area", "Progress", "Status"], start_col=2)
    dims = [
        ("Application progress", "=IFERROR(AVERAGE(CollegeProg),0)"),
        ("Apps submitted", '=IFERROR(COUNTIF(CollegeSubmitted,"Yes")/COUNTA(CollegeName),0)'),
        ("Essays final", '=IFERROR(COUNTIF(EssayStatus,"Final")/COUNTA(EssayStatus),0)'),
        ("Recs secured", '=IFERROR(COUNTIF(RecSubmitted,"Yes")/COUNTA(RecSubmitted),0)'),
        ("Scholarships applied", '=IFERROR((COUNTIF(SchStatus,"Applied")+COUNTIF(SchStatus,"Awarded"))/SchGoal,0)'),
        ("Tasks done", '=IFERROR(COUNTIF(TaskDone,"Yes")/COUNTA(TaskName),0)'),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"On track",IF(C{r}>=0.6,"Going well","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # reach/match/safety balance donut from a small summary
    ws.cell(row=20, column=8, value="Reach").style = "td_left"
    ws.cell(row=20, column=9, value='=COUNTIF(CollegeType,"Reach")').number_format = "0"
    ws.cell(row=21, column=8, value="Match").style = "td_left"
    ws.cell(row=21, column=9, value='=COUNTIF(CollegeType,"Match")').number_format = "0"
    ws.cell(row=22, column=8, value="Safety").style = "td_left"
    ws.cell(row=22, column=9, value='=COUNTIF(CollegeType,"Safety")').number_format = "0"
    d = DoughnutChart(); d.title = "Reach · Match · Safety"; d.height = 7.2; d.width = 8.4
    d.add_data(Reference(ws, min_col=9, min_row=20, max_row=22), titles_from_data=False)
    d.set_categories(Reference(ws, min_col=8, min_row=20, max_row=22)); d.dataLabels = no_labels()
    ws.add_chart(d, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "College Application Command Center™ — apply smart, hit every deadline, compare offers. Confirm details with each college.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profile(wb); build_collegelist(wb)
    build_essays(wb); build_recs(wb); build_tests(wb); build_activities(wb)
    build_scholarships(wb); build_netprice(wb); build_visits(wb); build_decisions(wb)
    build_tasks(wb); build_deadlines(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Applicant Profile", "College List", "Essays",
             "Recommendations", "Test Scores", "Activities", "Scholarships", "Net Price",
             "Visits", "Decisions", "To-Do", "Deadlines", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "College_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
