"""Build Recipe Costing & Menu Engineering Command Center™ — The Complete Menu-Profit System.

14 tabs · a premium recipe-costing & menu-engineering system in Google Sheets &
Excel. Dashboard, an ingredient price library, a recipe-costing engine, menu
items with food-cost % & margin, a menu-engineering quadrant (star/plowhorse/
puzzle/dog), a price calculator, sales mix, portion & yield, specials, batch
scaling, a vendor price log and a waste log — one dashboard. Cost every plate,
price with intent & engineer a more profitable menu.

Run: python3 build_xlsx.py   ->  ../Recipe_Costing_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["oz", "lb", "each", "cup", "tbsp", "tsp", "g", "kg", "ml", "L", "slice"]
STATIONS = ["Grill", "Sauté", "Fry", "Cold", "Pizza", "Pantry", "Pastry"]
CATS = ["Appetizer", "Salad", "Entrée", "Side", "Pizza", "Pasta", "Dessert", "Drink"]

TARGET_FC = 0.30

# Ingredient library: (item, pack size, unit, pack price)  -> cost/unit = price/size
INGREDIENTS = [
    ("Ground beef 80/20", 160, "oz", 44.80),
    ("Brioche bun", 12, "each", 6.60),
    ("Cheddar cheese", 80, "oz", 28.00),
    ("Lettuce", 100, "oz", 8.00),
    ("Tomato", 120, "oz", 12.00),
    ("Special sauce", 64, "oz", 7.68),
    ("Pickles", 128, "oz", 11.52),
    ("Onion", 160, "oz", 9.60),
    ("Pizza dough ball", 20, "each", 11.00),
    ("Mozzarella", 80, "oz", 24.00),
    ("Salmon fillet", 40, "oz", 320.00),
    ("Ribeye steak", 32, "oz", 288.00),
    ("Dry pasta", 160, "oz", 16.00),
    ("Romaine", 100, "oz", 9.00),
]

# Flagship recipe — Classic Burger: (ingredient, qty, unit, cost/unit)
BURGER = [
    ("Ground beef 80/20", 6, "oz", 0.28),
    ("Brioche bun", 1, "each", 0.55),
    ("Cheddar cheese", 1, "oz", 0.35),
    ("Lettuce", 0.5, "oz", 0.08),
    ("Tomato", 1, "oz", 0.10),
    ("Special sauce", 1, "oz", 0.12),
    ("Pickles", 0.5, "oz", 0.09),
    ("Onion", 0.5, "oz", 0.06),
]
BURGER_YIELD = 1

# Menu items: (item, category, plate cost, menu price, units sold / mo)
# Classic Burger cost is linked to the recipe engine (=BurgerCost)
MENU = [
    ("Classic Burger", "Entrée", None, 15.00, 340),
    ("Margherita Pizza", "Pizza", 3.10, 14.00, 240),
    ("Caesar Salad", "Salad", 2.40, 11.00, 150),
    ("Ribeye Steak", "Entrée", 11.50, 33.00, 110),
    ("Salmon Fillet", "Entrée", 8.20, 27.00, 200),
    ("Truffle Fries", "Side", 1.80, 9.00, 320),
    ("Pasta Carbonara", "Pasta", 3.60, 18.00, 250),
    ("Fish Tacos", "Entrée", 4.50, 17.00, 230),
]

# Price calculator targets: (item, plate cost, target food-cost %)
PRICECALC = [
    ("New Wings App", 2.80, 0.30),
    ("Veggie Bowl", 3.20, 0.28),
    ("Steak Sandwich", 5.40, 0.32),
    ("Dessert Special", 1.90, 0.25),
]

# Portion & yield: (item, as-purchased cost, yield %)
YIELD = [
    ("Whole salmon side", 8.00, 0.72),
    ("Beef tenderloin", 12.00, 0.68),
    ("Romaine case", 0.09, 0.85),
    ("Whole chicken", 4.20, 0.65),
]

# Specials / LTO: (special, plate cost, price, projected units)
SPECIALS = [
    ("Lobster Roll (summer)", 9.50, 26.00, 80),
    ("Pumpkin Ravioli (fall)", 4.10, 19.00, 120),
    ("Valentine Prix Fixe", 14.00, 55.00, 60),
]

# Batch scaling: (component, batch yield servings, batch cost)
BATCH = [
    ("Marinara (gal)", 96, 14.40),
    ("Caesar dressing (qt)", 32, 6.80),
    ("Burger patties", 40, 67.20),
    ("Pizza dough (batch)", 20, 11.00),
]

# Vendor price log: (ingredient, last price, this price)
VENDOR = [
    ("Ground beef 80/20 (case)", 42.00, 44.80),
    ("Salmon fillet (case)", 300.00, 320.00),
    ("Mozzarella (case)", 25.00, 24.00),
    ("Brioche bun (dozen)", 6.00, 6.60),
    ("Olive oil (jug)", 28.00, 31.00),
]

# Waste log: (item, reason, cost)
WASTE = [
    ("Spoiled produce", "Over-ordered", 42.00),
    ("Overcooked steaks", "Kitchen error", 34.00),
    ("Expired dairy", "Rotation miss", 18.00),
    ("Prep trim", "Normal yield loss", 26.00),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your restaurant, target food-cost % & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR RESTAURANT", "section")
    controls = [
        ("Restaurant name", "The Copper Skillet", None, "Restaurant"),
        ("Owner / chef", "Marco", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Margin goal ($/plate)", 12, '"$"#,##0', "MarginGoal"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Station", STATIONS, "StationList"),
             ("G", "Category", CATS, "CatList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🍽  RECIPE COSTING & MENU ENGINEERING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every plate, price with intent & engineer a more profitable menu.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE MENU'S PROFIT, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Build an ingredient price library, cost every recipe down to the plate, and see each menu item's "
                      "food-cost % and margin at a glance. Then engineer your menu: the star/plowhorse/puzzle/dog matrix "
                      "shows exactly which items to promote, reprice, rework or cut. A price calculator turns any plate "
                      "cost into a target-margin price, and vendor, waste, yield and batch tools keep your real costs "
                      "honest — all in ONE premium Google Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your restaurant name & target food-cost %.",
             "2.  Fill the Ingredients library — pack size & price gives cost per unit.",
             "3.  Cost a recipe on Recipe Costing — quantities × unit cost = plate cost.",
             "4.  Add each plate to Menu Items — food-cost % and margin calculate live.",
             "5.  Read Menu Engineering — promote Stars, fix Plowhorses, reprice Puzzles, cut Dogs.",
             "6.  Check the Dashboard: food-cost %, margins, the matrix & a Menu Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional bistro (The Copper Skillet) is included so you can see how it all connects — "
               "just type over it with your own recipes and prices. Menu engineering classifies every item by "
               "popularity and profitability so you know where the money is. Twelve matching printable pages (recipe "
               "cost cards, prep sheets, a menu-engineering worksheet & more) are included to print for the line. This "
               "is a business tool, not financial or accounting advice — confirm figures with your own books.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "A menu isn't a list of dishes — it's your most powerful profit lever.", "section_gold")


# ===========================================================================
# Ingredients — the price library
# ===========================================================================
def build_ingredients(wb):
    ws = wb.create_sheet("Ingredients"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 10, 14, 16, 2])
    luxe_header(ws, "F", "🧺  INGREDIENT LIBRARY",
                "Pack size & price for every ingredient — this gives the cost per unit behind every recipe.")
    table_headers(ws, 4, ["Ingredient", "Pack Size", "Unit", "Pack Price", "Cost / Unit"], start_col=2)
    start = L0
    for i, (item, size, unit, price) in enumerate(INGREDIENTS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cs = ws.cell(row=r, column=3, value=size); cs.style = "input"; cs.number_format = "#,##0"
        ws.cell(row=r, column=4, value=unit).style = "td"
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cu = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/C{r},0)"); cu.style = "td"; cu.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INGREDIENTS) - 1
    nrange(wb, "IngName", "Ingredients", "B", start, end)
    nrange(wb, "IngCostUnit", "Ingredients", "F", start, end)
    for c in range(2, 7):
        add_dv(ws, f"D{start}:D{end}", "UnitList")
    ws.freeze_panes = "A5"


# ===========================================================================
# Recipe Costing — the engine (flagship Classic Burger)
# ===========================================================================
def build_recipecost(wb):
    ws = wb.create_sheet("Recipe Costing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 10, 10, 14, 14, 2])
    luxe_header(ws, "F", "🧮  RECIPE COSTING",
                "Cost a plate line by line — quantity × unit cost. Change a price and the plate re-costs itself.")
    merge_set(ws, "B5:C5", "RECIPE", "section_gold")
    ws.cell(row=5, column=4, value="Classic Burger").font = Font(bold=True, color=PRIMARY)
    ws.cell(row=5, column=5, value="Yield →").font = Font(italic=True, color=ACCENT)
    ws.cell(row=5, column=6, value=BURGER_YIELD).font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Ingredient", "Qty", "Unit", "Cost/Unit", "Ext. Cost"], start_col=2)
    start = 7
    for i, (ing, qty, unit, cu) in enumerate(BURGER):
        r = start + i
        ws.cell(row=r, column=2, value=ing).style = "td_left"
        cq = ws.cell(row=r, column=3, value=qty); cq.style = "input"; cq.number_format = "0.##"
        ws.cell(row=r, column=4, value=unit).style = "td"
        ccu = ws.cell(row=r, column=5, value=cu); ccu.style = "input"; ccu.number_format = '"$"#,##0.00'
        ce = ws.cell(row=r, column=6, value=f"=C{r}*E{r}"); ce.style = "td"; ce.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BURGER) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL RECIPE COST").style = "th"
    for c in range(3, 6):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0.00'
    cs = tot + 1
    merge_set(ws, f"B{cs}:E{cs}", "COST PER SERVING  (total ÷ yield)", "section")
    cps = ws.cell(row=cs, column=6, value=f"=IFERROR(F{tot}/F5,0)"); cps.style = "field_value"; cps.font = Font(bold=True, size=13, color=PRIMARY); cps.number_format = '"$"#,##0.00'; cps.fill = fill(MINT_BG)
    cell_name(wb, "BurgerCost", "Recipe Costing", f"$F${cs}")
    ws.cell(row=cs + 2, column=2, value="Cost/Unit values come from your Ingredient Library — restate or link them here.").style = "section"


# ===========================================================================
# Menu Items — food cost %, margin, class
# ===========================================================================
def build_menu(wb):
    ws = wb.create_sheet("Menu Items"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 12, 12, 10, 12, 12, 14, 2])
    luxe_header(ws, "I", "📋  MENU ITEMS",
                "Every plate's cost, price, food-cost % & margin — and its menu-engineering class, live.")
    table_headers(ws, 4, ["Item", "Category", "Plate Cost", "Price", "Units/mo", "Food Cost %", "Margin $", "Class"], start_col=2)
    start = L0
    end = start + len(MENU) - 1
    # helper cells for averages (place below table)
    au_row = end + 3; am_row = end + 4
    for i, (item, cat, cost, price, units) in enumerate(MENU):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        if cost is None:
            cc = ws.cell(row=r, column=4, value="=BurgerCost")
        else:
            cc = ws.cell(row=r, column=4, value=cost)
        cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cu = ws.cell(row=r, column=6, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cf = ws.cell(row=r, column=7, value=f"=IFERROR(D{r}/E{r},0)"); cf.style = "td"; cf.number_format = "0%"
        cm = ws.cell(row=r, column=8, value=f"=E{r}-D{r}"); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        cl = ws.cell(row=r, column=9,
                     value=f'=IF(AND(F{r}>=$C${au_row},H{r}>=$C${am_row}),"Star",'
                           f'IF(F{r}>=$C${au_row},"Plowhorse",IF(H{r}>=$C${am_row},"Puzzle","Dog")))')
        cl.style = "td"
        if i % 2:
            for c in range(2, 10):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    nrange(wb, "MenuItem", "Menu Items", "B", start, end)
    nrange(wb, "MenuCost", "Menu Items", "D", start, end)
    nrange(wb, "MenuPrice", "Menu Items", "E", start, end)
    nrange(wb, "MenuUnits", "Menu Items", "F", start, end)
    nrange(wb, "MenuFC", "Menu Items", "G", start, end)
    nrange(wb, "MenuMargin", "Menu Items", "H", start, end)
    nrange(wb, "MenuClass", "Menu Items", "I", start, end)
    # average helper cells
    ws.cell(row=au_row, column=2, value="Avg units (popularity line)").style = "field_label"
    ca = ws.cell(row=au_row, column=3, value="=AVERAGE(MenuUnits)"); ca.style = "field_value"; ca.number_format = "#,##0"
    ws.cell(row=am_row, column=2, value="Avg margin (profit line)").style = "field_label"
    cm = ws.cell(row=am_row, column=3, value="=AVERAGE(MenuMargin)"); cm.style = "field_value"; cm.number_format = '"$"#,##0.00'
    cell_name(wb, "AvgUnits", "Menu Items", f"$C${au_row}")
    cell_name(wb, "AvgMargin", "Menu Items", f"$C${am_row}")
    # totals row
    tr = end + 1
    ws.cell(row=tr, column=2, value="MENU").style = "th"
    ws.cell(row=tr, column=3).style = "td"; ws.cell(row=tr, column=3).fill = fill(SURFACE)
    cavc = ws.cell(row=tr, column=4, value="=AVERAGE(MenuCost)"); cavc.style = "td"; cavc.font = Font(bold=True, color=PRIMARY); cavc.fill = fill(SURFACE); cavc.number_format = '"$"#,##0.00'
    cavp = ws.cell(row=tr, column=5, value="=AVERAGE(MenuPrice)"); cavp.style = "td"; cavp.font = Font(bold=True, color=PRIMARY); cavp.fill = fill(SURFACE); cavp.number_format = '"$"#,##0.00'
    ws.cell(row=tr, column=6, value="=SUM(MenuUnits)").number_format = "#,##0"
    ws.cell(row=tr, column=6).font = Font(bold=True, color=PRIMARY); ws.cell(row=tr, column=6).fill = fill(SURFACE); ws.cell(row=tr, column=6).border = BOX
    cavf = ws.cell(row=tr, column=7, value="=AVERAGE(MenuFC)"); cavf.style = "td"; cavf.font = Font(bold=True, color=PRIMARY); cavf.fill = fill(SURFACE); cavf.number_format = "0%"
    cavm = ws.cell(row=tr, column=8, value="=AVERAGE(MenuMargin)"); cavm.style = "td"; cavm.font = Font(bold=True, color=PRIMARY); cavm.fill = fill(SURFACE); cavm.number_format = '"$"#,##0.00'
    ws.cell(row=tr, column=9).style = "td"; ws.cell(row=tr, column=9).fill = fill(SURFACE)
    cell_name(wb, "AvgFC", "Menu Items", f"$G${tr}")
    cell_name(wb, "AvgCost", "Menu Items", f"$D${tr}")
    cell_name(wb, "AvgPrice", "Menu Items", f"$E${tr}")
    ws.conditional_formatting.add(f"G{start}:G{end}",
        ColorScaleRule(start_type="num", start_value=0.15, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.30, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.45, end_color="FF" + RED_BG))
    cmap = {"Star": MINT_BG, "Plowhorse": WARN_BG, "Puzzle": SURFACE, "Dog": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"I{start}:I{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    ws.freeze_panes = "A5"


# ===========================================================================
# Menu Engineering — the quadrant
# ===========================================================================
def build_engineering(wb):
    ws = wb.create_sheet("Menu Engineering"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 30, 2])
    luxe_header(ws, "D", "⭐  MENU ENGINEERING",
                "Every item by popularity × profitability. Promote Stars, rework Plowhorses, reprice Puzzles, cut Dogs.")
    table_headers(ws, 4, ["Class", "Count", "What to do"], start_col=2)
    rows = [
        ("Star", '=COUNTIF(MenuClass,"Star")', "High profit + popular → feature it, protect it, never discount."),
        ("Plowhorse", '=COUNTIF(MenuClass,"Plowhorse")', "Popular but low profit → trim cost or nudge price up."),
        ("Puzzle", '=COUNTIF(MenuClass,"Puzzle")', "High profit but slow → reposition, rename, upsell it."),
        ("Dog", '=COUNTIF(MenuClass,"Dog")', "Low profit + slow → rework or cut it from the menu."),
    ]
    cmap = {"Star": MINT_BG, "Plowhorse": WARN_BG, "Puzzle": SURFACE, "Dog": RED_BG}
    start = L0
    for i, (cls, fml, action) in enumerate(rows):
        r = start + i
        cc = ws.cell(row=r, column=2, value=cls); cc.style = "td_left"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(cmap[cls])
        cn = ws.cell(row=r, column=3, value=fml); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = "0"
        ws.cell(row=r, column=4, value=action).style = "td_left"
    end = start + len(rows) - 1
    tr = end + 1
    ws.cell(row=tr, column=2, value="TOTAL ITEMS").style = "th"
    ct = ws.cell(row=tr, column=3, value="=COUNTA(MenuItem)"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = "0"
    ws.cell(row=tr, column=4).style = "td"; ws.cell(row=tr, column=4).fill = fill(SURFACE)
    merge_set(ws, f"B{tr+2}:D{tr+2}", "The matrix uses your menu's own average popularity & margin as the dividing lines.", "section")
    # class doughnut
    d = DoughnutChart(); d.title = "Menu Mix by Class"; d.height = 7.4; d.width = 9
    d.add_data(Reference(ws, min_col=3, min_row=start, max_row=end), titles_from_data=False)
    d.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end)); d.dataLabels = no_labels()
    ws.add_chart(d, "F5")


# ===========================================================================
# Price Calculator
# ===========================================================================
def build_pricecalc(wb):
    ws = wb.create_sheet("Price Calculator"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 16, 16, 16, 2])
    luxe_header(ws, "F", "💲  PRICE CALCULATOR",
                "Turn any plate cost into a target-margin price — no more pricing by gut feel.")
    table_headers(ws, 4, ["Item", "Plate Cost", "Target FC %", "Suggested Price", "Margin $"], start_col=2)
    start = L0
    for i, (item, cost, tfc) in enumerate(PRICECALC):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        ct = ws.cell(row=r, column=4, value=tfc); ct.style = "input"; ct.number_format = "0%"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=6, value=f"=E{r}-C{r}"); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRICECALC) - 1
    merge_set(ws, f"B{end+2}:F{end+2}", "Suggested price = plate cost ÷ target food-cost %. Round to a menu-friendly number.", "section_gold")
    ws.freeze_panes = "A5"


# ===========================================================================
# Sales Mix
# ===========================================================================
def build_salesmix(wb):
    ws = wb.create_sheet("Sales Mix"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 12, 16, 16, 2])
    luxe_header(ws, "E", "📊  SALES MIX",
                "Where the money actually comes from — revenue & profit contribution per item.")
    table_headers(ws, 4, ["Item", "Units/mo", "Revenue", "Profit"], start_col=2)
    start = L0
    for i, (item, cat, cost, price, units) in enumerate(MENU):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cu = ws.cell(row=r, column=3, value=f"=INDEX(MenuUnits,{i+1})"); cu.style = "td"; cu.number_format = "#,##0"
        cr = ws.cell(row=r, column=4, value=f"=INDEX(MenuPrice,{i+1})*C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=INDEX(MenuMargin,{i+1})*C{r}"); cp.style = "td"; cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MENU) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL / MONTH").style = "th"
    ws.cell(row=tot, column=3, value="=SUM(MenuUnits)").number_format = "#,##0"
    ws.cell(row=tot, column=3).font = Font(bold=True, color=PRIMARY); ws.cell(row=tot, column=3).fill = fill(SURFACE); ws.cell(row=tot, column=3).border = BOX
    cr = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = '"$"#,##0'
    cp = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(SURFACE); cp.number_format = '"$"#,##0'
    cell_name(wb, "TotalRevenue", "Sales Mix", f"$D${tot}")
    cell_name(wb, "TotalProfit", "Sales Mix", f"$E${tot}")
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
# Portion & Yield
# ===========================================================================
def build_yield(wb):
    ws = wb.create_sheet("Portion & Yield"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 12, 18, 2])
    luxe_header(ws, "E", "🔪  PORTION & YIELD",
                "Trim, bones & shrinkage cost money — edible-portion cost is your true cost.")
    table_headers(ws, 4, ["Item", "As-Purchased Cost", "Yield %", "True (EP) Cost"], start_col=2)
    start = L0
    for i, (item, ap, y) in enumerate(YIELD):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=ap); ca.style = "input"; ca.number_format = '"$"#,##0.00'
        cy = ws.cell(row=r, column=4, value=y); cy.style = "input"; cy.number_format = "0%"
        ce = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); ce.style = "td"; ce.font = Font(bold=True, color=PRIMARY); ce.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(YIELD) - 1
    merge_set(ws, f"B{end+2}:E{end+2}", "True cost = as-purchased cost ÷ yield %. Always cost recipes at edible-portion cost.", "section_gold")
    ws.freeze_panes = "A5"


# ===========================================================================
# Specials / LTO
# ===========================================================================
def build_specials(wb):
    ws = wb.create_sheet("Specials & LTO"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 12, 14, 14, 2])
    luxe_header(ws, "F", "✨  SPECIALS & LTO",
                "Cost and price limited-time offers before you run them — protect the margin.")
    table_headers(ws, 4, ["Special", "Plate Cost", "Price", "Proj. Units", "Proj. Profit"], start_col=2)
    start = L0
    for i, (name, cost, price, units) in enumerate(SPECIALS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cu = ws.cell(row=r, column=5, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cpr = ws.cell(row=r, column=6, value=f"=(D{r}-C{r})*E{r}"); cpr.style = "td"; cpr.font = Font(bold=True, color=PRIMARY); cpr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


# ===========================================================================
# Batch Scaling
# ===========================================================================
def build_batch(wb):
    ws = wb.create_sheet("Batch & Prep"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 16, 16, 2])
    luxe_header(ws, "E", "🍲  BATCH & PREP",
                "Cost your prep batches — batch cost ÷ yield gives the true cost per portion.")
    table_headers(ws, 4, ["Component", "Batch Yield", "Batch Cost", "Cost / Serving"], start_col=2)
    start = L0
    for i, (comp, y, cost) in enumerate(BATCH):
        r = start + i
        ws.cell(row=r, column=2, value=comp).style = "td_left"
        cy = ws.cell(row=r, column=3, value=y); cy.style = "input"; cy.number_format = "#,##0"
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cs = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


# ===========================================================================
# Vendor Price Log
# ===========================================================================
def build_vendor(wb):
    ws = wb.create_sheet("Vendor Prices"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 16, 14, 2])
    luxe_header(ws, "E", "🚚  VENDOR PRICES",
                "Track what your suppliers charge — spot creep before it eats your margin.")
    table_headers(ws, 4, ["Ingredient (pack)", "Last Price", "This Price", "Change"], start_col=2)
    start = L0
    for i, (item, last, now) in enumerate(VENDOR):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cl = ws.cell(row=r, column=3, value=last); cl.style = "input"; cl.number_format = '"$"#,##0.00'
        cn = ws.cell(row=r, column=4, value=now); cn.style = "input"; cn.number_format = '"$"#,##0.00'
        cc = ws.cell(row=r, column=5, value=f"=IFERROR((D{r}-C{r})/C{r},0)"); cc.style = "td"; cc.number_format = "+0%;[Red]-0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(VENDOR) - 1
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="greaterThan", formula=["0.05"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


# ===========================================================================
# Waste Log
# ===========================================================================
def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Every dollar thrown away, logged — waste is pure profit leaving the building.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL WASTE / MO").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


# ===========================================================================
# Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🍽  RECIPE COSTING & MENU ENGINEERING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Food cost, margins, the star/dog matrix & a Menu Score — your whole menu's profit, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MENU ITEMS", "=COUNTA(MenuItem)", "num"),
        ("AVG FOOD COST", "=AvgFC", "pct"),
        ("TARGET FOOD COST", "=TargetFC", "pct"),
        ("AVG PLATE COST", "=AvgCost", "money2"),
        ("AVG MENU PRICE", "=AvgPrice", "money2"),
        ("AVG MARGIN", "=AvgMargin", "money2"),
    ]
    row2 = [
        ("STARS", '=COUNTIF(MenuClass,"Star")', "num"),
        ("PLOWHORSES", '=COUNTIF(MenuClass,"Plowhorse")', "num"),
        ("PUZZLES", '=COUNTIF(MenuClass,"Puzzle")', "num"),
        ("DOGS", '=COUNTIF(MenuClass,"Dog")', "num"),
        ("TOP MARGIN", '=INDEX(MenuItem,MATCH(MAX(MenuMargin),MenuMargin,0))', "text"),
        ("MENU SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "MENU HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "MENU MIX BY CLASS", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Food cost on target", "=IFERROR(MIN(TargetFC/AvgFC,1),0)"),
        ("Margin vs goal", "=IFERROR(MIN(AvgMargin/MarginGoal,1),0)"),
        ("Menu fully costed", "=IFERROR(COUNTIF(MenuCost,\">0\")/COUNTA(MenuItem),0)"),
        ("Items above margin goal", "=IFERROR(COUNTIF(MenuMargin,\">=\"&MarginGoal)/COUNTA(MenuItem),0)"),
        ("Menu balance (few dogs)", '=IFERROR(1-COUNTIF(MenuClass,"Dog")/COUNTA(MenuItem),0)'),
        ("Contribution margin", "=IFERROR(MIN((1-AvgFC)/0.7,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Fix"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    eng = wb["Menu Engineering"]
    d = DoughnutChart(); d.title = "Menu Mix by Class"; d.height = 7.4; d.width = 8.6
    d.add_data(Reference(eng, min_col=3, min_row=5, max_row=8), titles_from_data=False)
    d.set_categories(Reference(eng, min_col=2, min_row=5, max_row=8)); d.dataLabels = no_labels()
    ws.add_chart(d, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Recipe Costing & Menu Engineering Command Center™ — cost every plate, price with intent.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_ingredients(wb); build_recipecost(wb)
    build_menu(wb); build_engineering(wb); build_pricecalc(wb); build_salesmix(wb)
    build_yield(wb); build_specials(wb); build_batch(wb); build_vendor(wb); build_waste(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Ingredients", "Recipe Costing", "Menu Items",
             "Menu Engineering", "Price Calculator", "Sales Mix", "Portion & Yield",
             "Specials & LTO", "Batch & Prep", "Vendor Prices", "Waste Log", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Recipe_Costing_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
