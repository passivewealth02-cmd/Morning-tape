"""Build the Back-to-School Command Center (BTSCC) Excel workbook.

12 sheets · dashboard + 10 functional planners + settings.

Run: python3 build_xlsx.py
Outputs: ../Back_To_School_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import (
    CellIsRule,
    ColorScaleRule,
    DataBarRule,
    FormulaRule,
)
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"
ACCENT = "937356"
SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"
BG = "FFFFFF"
TEXT = "333333"
SUCCESS = "75E6C1"
WARNING = "937356"
DANGER = "C94C4C"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"

CATEGORIES = ["Supplies", "Clothing", "Shoes", "Electronics", "Sports",
              "Lunch Supplies", "Fees", "Transportation", "Miscellaneous"]
PRIORITIES = ["High", "Medium", "Low"]
STATUSES = ["Not Started", "In Progress", "Complete", "Overdue"]
SUBJECTS = ["Math", "English", "Science", "History", "Art",
            "PE", "Music", "Tech", "Spanish", "Other"]
STORES = ["Target", "Walmart", "Amazon", "Staples", "Local", "Other"]
GRADES = ["Pre-K", "K", "1", "2", "3", "4", "5", "6", "7",
          "8", "9", "10", "11", "12", "College"]
YESNO = ["Yes", "No"]
TRANSPORT = ["Bus", "Carpool", "Walk", "Bike", "Parent Drop-off"]

THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ---------------------------------------------------------------------------
# Style registration
# ---------------------------------------------------------------------------

def clean_labels(pct=False, val=False):
    """Concise data labels: never show series/category text (it collides on
    multi-slice pies). Percentages or values only."""
    dl = DataLabelList()
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showBubbleSize = False
    dl.showVal = val
    dl.showPercent = pct
    return dl


def no_labels():
    """Suppress all on-chart labels (legend carries meaning)."""
    return clean_labels(pct=False, val=False)


def register_styles(wb: Workbook) -> None:
    styles = {
        "title": NamedStyle(
            name="title",
            font=Font(size=22, bold=True, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "subtitle": NamedStyle(
            name="subtitle",
            font=Font(size=11, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
        ),
        "section": NamedStyle(
            name="section",
            font=Font(size=12, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        "th": NamedStyle(
            name="th",
            font=Font(size=11, bold=True, color="FFFFFF"),
            fill=PatternFill("solid", fgColor=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=BOX,
        ),
        "td": NamedStyle(
            name="td",
            font=Font(size=11, color=TEXT),
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=BOX,
        ),
        "td_left": NamedStyle(
            name="td_left",
            font=Font(size=11, color=TEXT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
            border=BOX,
        ),
        "kpi_label": NamedStyle(
            name="kpi_label",
            font=Font(size=10, bold=True, color=ACCENT),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
        ),
        "kpi_value": NamedStyle(
            name="kpi_value",
            font=Font(size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
        ),
        "kpi_money": NamedStyle(
            name="kpi_money",
            font=Font(size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
            number_format='"$"#,##0.00',
        ),
        "kpi_pct": NamedStyle(
            name="kpi_pct",
            font=Font(size=20, bold=True, color=PRIMARY),
            alignment=Alignment(horizontal="center", vertical="center"),
            fill=PatternFill("solid", fgColor=BG),
            number_format="0%",
        ),
        "input": NamedStyle(
            name="input",
            font=Font(size=11, bold=True, color=PRIMARY),
            fill=PatternFill("solid", fgColor=SURFACE),
            alignment=Alignment(horizontal="center", vertical="center"),
            border=BOX,
        ),
        "field_label": NamedStyle(
            name="field_label",
            font=Font(size=10, bold=True, color=ACCENT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=BOX,
            fill=PatternFill("solid", fgColor=SOFT_BG),
        ),
        "field_value": NamedStyle(
            name="field_value",
            font=Font(size=11, color=TEXT),
            alignment=Alignment(horizontal="left", vertical="center", indent=1),
            border=BOX,
        ),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def merge_set(ws, rng, value, style_name):
    ws.merge_cells(rng)
    top = rng.split(":")[0]
    cell = ws[top]
    cell.value = value
    cell.style = style_name
    return cell


def page_header(ws, last_col_letter: str, title: str, subtitle: str) -> None:
    ws.row_dimensions[1].height = 44
    ws.row_dimensions[2].height = 22
    merge_set(ws, f"A1:{last_col_letter}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col_letter}2", "  " + subtitle, "subtitle")


def set_column_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def band_row(ws, row: int, start_col: int, end_col: int, even: bool) -> None:
    color = MUTED_ROW if even else BG
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = fill(color)


def add_dv(ws, rng: str, list_name: str) -> None:
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


# ===========================================================================
# Sheet 12: Settings (built first so defined names exist)
# ===========================================================================
def build_settings(wb: Workbook) -> None:
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [2, 30, 24, 4, 22, 22, 22, 22])
    page_header(ws, "H", "⚙  SETTINGS", "Configure school year, goals, and dropdowns.")

    merge_set(ws, "B4:C4", "USER CONTROLS", "section")
    controls = [
        ("School Year",              "2026-2027",      None,                 "SchoolYear"),
        ("First Day of School",      dt.date(2026, 9, 2), "mm/dd/yyyy",      "FirstDayOfSchool"),
        ("Total Back-to-School Budget", 1200,          '"$"#,##0.00',        "TotalBudget"),
        ("Family Size",              2,                "0",                  "FamilySize"),
        ("Tax Rate",                 0.075,            "0.00%",              "TaxRate"),
        ("Today (auto)",             "=TODAY()",       "mm/dd/yyyy",         "TodayDate"),
    ]
    for i, (label, default, fmt, name) in enumerate(controls):
        r = 5 + i
        ws.cell(row=r, column=2, value=label).style = "field_label"
        cell = ws.cell(row=r, column=3, value=default)
        cell.style = "input"
        if fmt:
            cell.number_format = fmt
        wb.defined_names[name] = DefinedName(name, attr_text=f"Settings!$C${r}")

    # Dropdown lists arranged in adjacent columns
    merge_set(ws, "E4:H4", "DROPDOWN LISTS", "section")
    headers = ["Categories", "Priority", "Status", "Subjects"]
    cols_data = [CATEGORIES, PRIORITIES, STATUSES, SUBJECTS]
    for ci, (h, data) in enumerate(zip(headers, cols_data)):
        col = 5 + ci  # E, F, G, H
        ws.cell(row=5, column=col, value=h).style = "th"
        for ri, val in enumerate(data):
            cell = ws.cell(row=6 + ri, column=col, value=val)
            cell.style = "td_left"

    # More lists farther down
    extra_lists = [
        ("Stores",     STORES,    "E",  20),
        ("Grades",     GRADES,    "F",  20),
        ("Yes / No",   YESNO,     "G",  20),
        ("Transport",  TRANSPORT, "H",  20),
    ]
    for label, data, col, start_row in extra_lists:
        c_idx = ord(col) - 64
        ws.cell(row=start_row, column=c_idx, value=label).style = "th"
        for ri, val in enumerate(data):
            ws.cell(row=start_row + 1 + ri, column=c_idx, value=val).style = "td_left"

    # Defined-name ranges for every list
    wb.defined_names["CategoryList"] = DefinedName(
        "CategoryList", attr_text=f"Settings!$E$6:$E${5 + len(CATEGORIES)}"
    )
    wb.defined_names["PriorityList"] = DefinedName(
        "PriorityList", attr_text=f"Settings!$F$6:$F${5 + len(PRIORITIES)}"
    )
    wb.defined_names["StatusList"] = DefinedName(
        "StatusList", attr_text=f"Settings!$G$6:$G${5 + len(STATUSES)}"
    )
    wb.defined_names["SubjectList"] = DefinedName(
        "SubjectList", attr_text=f"Settings!$H$6:$H${5 + len(SUBJECTS)}"
    )
    wb.defined_names["StoreList"] = DefinedName(
        "StoreList", attr_text=f"Settings!$E$21:$E${20 + len(STORES)}"
    )
    wb.defined_names["GradeList"] = DefinedName(
        "GradeList", attr_text=f"Settings!$F$21:$F${20 + len(GRADES)}"
    )
    wb.defined_names["YesNoList"] = DefinedName(
        "YesNoList", attr_text=f"Settings!$G$21:$G${20 + len(YESNO)}"
    )
    wb.defined_names["TransportList"] = DefinedName(
        "TransportList", attr_text=f"Settings!$H$21:$H${20 + len(TRANSPORT)}"
    )


# ===========================================================================
# Sheet 2: Student Profiles
# ===========================================================================
def build_students(wb: Workbook) -> None:
    ws = wb.create_sheet("Students")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [2, 22, 26, 4, 22, 26])
    page_header(ws, "F", "👨‍🎓  STUDENT PROFILES",
                "One section per student. Duplicate the block for additional kids.")

    sample_students = [
        {
            "Name": "Emma Carter", "Grade": "4", "School": "Lincoln Elementary",
            "Teacher": "Ms. Hayes", "Student ID": "LE-2024-118",
            "Start Date": dt.date(2026, 9, 2), "Allergies": "Peanuts",
            "Emergency Contact": "Mom: (555) 412-7833",
            "Transportation": "Bus", "Locker #": "—",
            "Bus Route": "Route 14",  "Homeroom": "Rm 104",
        },
        {
            "Name": "Liam Carter", "Grade": "8", "School": "Roosevelt Middle",
            "Teacher": "Mr. Patel", "Student ID": "RM-2024-441",
            "Start Date": dt.date(2026, 9, 2), "Allergies": "None",
            "Emergency Contact": "Dad: (555) 412-7901",
            "Transportation": "Carpool", "Locker #": "212",
            "Bus Route": "—", "Homeroom": "Rm 8B",
        },
    ]

    row = 4
    for student in sample_students:
        # Card header bar
        merge_set(ws, f"B{row}:F{row}", "  STUDENT", "section")
        ws.row_dimensions[row].height = 22
        row += 1

        fields = list(student.items())
        # Render fields in two columns (label/value, label/value)
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24
            i += 2
            row += 1
        # Spacer
        row += 2

    # Validation on Grade / Transportation across both blocks
    add_dv(ws, "C5:C16", "GradeList")
    add_dv(ws, "F5:F16", "TransportList")


# ===========================================================================
# Sheet 3: School Supply Tracker
# ===========================================================================
def build_supplies(wb: Workbook) -> None:
    ws = wb.create_sheet("Supplies")
    ws.sheet_view.showGridLines = False
    widths = [14, 28, 12, 12, 12, 14, 14, 14, 12, 12, 26]
    set_column_widths(ws, widths)
    page_header(ws, "K", "📚  SCHOOL SUPPLY TRACKER",
                "Track every item from list to checkout. Formulas auto-tally.")
    headers = ["Category", "Item", "Req Qty", "Bought Qty", "Remaining",
               "Store", "Est. Price", "Actual Price", "Purchased",
               "Priority", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 32

    samples = [
        ("Supplies", "Backpack",              1, 1, "Target",  39.99, 34.99, "Yes", "High",   "Black with side pocket"),
        ("Supplies", "Pencil case",           2, 0, "Amazon",  8.99,  None,  "No",  "Medium", "Need two: one per child"),
        ("Supplies", "Pencils #2 (12pk)",     4, 3, "Walmart", 3.49,  3.49,  "No",  "High",   ""),
        ("Supplies", "Erasers",               6, 4, "Walmart", 1.99,  1.99,  "No",  "Low",    ""),
        ("Supplies", "Glue sticks",          12, 6, "Staples", 0.99,  0.99,  "No",  "Medium", ""),
        ("Supplies", "Notebook (wide-rule)",  8, 8, "Target",  1.49,  1.29,  "Yes", "High",   "All purchased"),
        ("Supplies", "Composition books",     4, 0, "Staples", 2.49,  None,  "No",  "Medium", ""),
        ("Supplies", "Folders",               6, 3, "Target",  0.79,  0.79,  "No",  "Medium", ""),
        ("Supplies", "Binders 1\"",           3, 1, "Staples", 4.99,  4.99,  "No",  "High",   ""),
        ("Supplies", "Highlighters (4pk)",    2, 2, "Amazon",  4.49,  4.49,  "Yes", "Low",    ""),
        ("Electronics", "Laptop bag",         1, 0, "Amazon", 24.99,  None,  "No",  "High",   "13\" sleeve"),
        ("Electronics", "USB-C charger",      2, 1, "Amazon", 19.99, 19.99,  "No",  "High",   ""),
        ("Electronics", "Headphones (wired)", 2, 2, "Target", 12.99, 12.99,  "Yes", "Medium", ""),
        ("Clothing", "School polo",           5, 2, "Target", 14.99, 14.99,  "No",  "High",   "Navy"),
        ("Clothing", "Khaki pants",           3, 1, "Target", 19.99, 19.99,  "No",  "High",   ""),
        ("Shoes",    "Sneakers",              2, 1, "Local",  44.99, 39.99,  "No",  "High",   "Size 4 & 7"),
        ("Sports",   "Soccer cleats",         1, 0, "Local",  54.99,  None,  "No",  "Medium", "Try-on 8/24"),
        ("Lunch Supplies", "Lunch box",       2, 2, "Target", 12.99, 12.99,  "Yes", "High",   ""),
        ("Lunch Supplies", "Water bottle",    2, 1, "Target",  9.99,  9.99,  "No",  "High",   ""),
        ("Fees",     "Registration fee",      1, 0, "Other", 125.00,  None,  "No",  "High",   "Due 8/20"),
        ("Fees",     "Field trip deposit",    1, 0, "Other",  35.00,  None,  "No",  "Medium", ""),
        ("Miscellaneous", "Locker organizer", 1, 0, "Amazon", 16.99,  None,  "No",  "Low",    ""),
    ]
    start = 5
    rows_reserved = 60
    end = start + rows_reserved - 1

    for i, item in enumerate(samples):
        r = start + i
        cat, name, req, bought, store, est, actual, purchased, prio, notes = item
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=req)
        ws.cell(row=r, column=4, value=bought)
        ws.cell(row=r, column=6, value=store)
        ws.cell(row=r, column=7, value=est)
        ws.cell(row=r, column=8, value=actual)
        ws.cell(row=r, column=9, value=purchased)
        ws.cell(row=r, column=10, value=prio)
        ws.cell(row=r, column=11, value=notes)

    # Formulas for every reserved row
    for r in range(start, end + 1):
        ws.cell(row=r, column=5,
                value=f'=IF(C{r}="","",MAX(C{r}-IFERROR(D{r},0),0))')

    # Style and number formats
    money_cols = {7, 8}
    int_cols = {3, 4, 5}
    text_left_cols = {2, 11}
    for r in range(start, end + 1):
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left_cols else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 == 1 else BG)
            if c in money_cols:
                cell.number_format = '"$"#,##0.00'
            elif c in int_cols:
                cell.number_format = "0"

    # Validation
    add_dv(ws, f"A{start}:A{end}", "CategoryList")
    add_dv(ws, f"F{start}:F{end}", "StoreList")
    add_dv(ws, f"I{start}:I{end}", "YesNoList")
    add_dv(ws, f"J{start}:J{end}", "PriorityList")

    # Conditional formatting
    # Purchased Yes → soft mint
    ws.conditional_formatting.add(
        f"A{start}:K{end}",
        FormulaRule(formula=[f'$I{start}="Yes"'],
                    fill=PatternFill("solid", fgColor="E3F8EF")),
    )
    # Priority High → gold left tint
    ws.conditional_formatting.add(
        f"J{start}:J{end}",
        CellIsRule(operator="equal", formula=['"High"'],
                   fill=PatternFill("solid", fgColor="EFE0CC")),
    )
    # Remaining > 0 → soft warning
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor="FBF0E2")),
    )

    # Defined ranges used by Dashboard
    wb.defined_names["SupActual"] = DefinedName(
        "SupActual", attr_text=f"Supplies!$H${start}:$H${end}"
    )
    wb.defined_names["SupEst"] = DefinedName(
        "SupEst", attr_text=f"Supplies!$G${start}:$G${end}"
    )
    wb.defined_names["SupPurchased"] = DefinedName(
        "SupPurchased", attr_text=f"Supplies!$I${start}:$I${end}"
    )
    wb.defined_names["SupItem"] = DefinedName(
        "SupItem", attr_text=f"Supplies!$B${start}:$B${end}"
    )
    wb.defined_names["SupReq"] = DefinedName(
        "SupReq", attr_text=f"Supplies!$C${start}:$C${end}"
    )
    wb.defined_names["SupBought"] = DefinedName(
        "SupBought", attr_text=f"Supplies!$D${start}:$D${end}"
    )

    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 4: Back-to-School Budget
# ===========================================================================
def build_budget(wb: Workbook) -> None:
    ws = wb.create_sheet("Budget")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [24, 20, 20, 20, 18, 4, 24, 20])
    page_header(ws, "H", "💰  BACK-TO-SCHOOL BUDGET",
                "Plan, track, and check variance across categories.")

    headers = ["Category", "Planned", "Actual", "Variance", "% Spent"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    sample_planned = {
        "Supplies":        200,
        "Clothing":        260,
        "Shoes":           120,
        "Electronics":     180,
        "Sports":           90,
        "Lunch Supplies":   60,
        "Fees":            180,
        "Transportation":   60,
        "Miscellaneous":    50,
    }
    start = 5
    end = start + len(CATEGORIES) - 1
    for i, cat in enumerate(CATEGORIES):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        c_planned = ws.cell(row=r, column=2, value=sample_planned.get(cat, 0))
        c_planned.style = "input"
        c_planned.number_format = '"$"#,##0.00'
        c_actual = ws.cell(row=r, column=3,
                           value=f'=IFERROR(SUMIFS(SupActual,Supplies!$A$5:$A$64,A{r}),0)')
        c_actual.style = "td"
        c_actual.number_format = '"$"#,##0.00'
        c_var = ws.cell(row=r, column=4, value=f"=B{r}-C{r}")
        c_var.style = "td"
        c_var.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
        c_pct = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)")
        c_pct.style = "td"
        c_pct.number_format = "0%"
        ws.row_dimensions[r].height = 22
        band_row(ws, r, 1, 5, even=(i % 2 == 1))

    total_row = end + 1
    ws.cell(row=total_row, column=1, value="TOTAL").style = "th"
    for col, fml in (
        (2, f"=SUM(B{start}:B{end})"),
        (3, f"=SUM(C{start}:C{end})"),
        (4, f"=SUM(D{start}:D{end})"),
        (5, f"=IFERROR(C{total_row}/B{total_row},0)"),
    ):
        c = ws.cell(row=total_row, column=col, value=fml)
        c.style = "td"
        c.font = Font(size=11, bold=True, color=PRIMARY)
        c.fill = fill(SURFACE)
        if col in (2, 3, 4):
            c.number_format = '"$"#,##0.00'
        else:
            c.number_format = "0%"

    wb.defined_names["BudgetCategories"] = DefinedName(
        "BudgetCategories", attr_text=f"Budget!$A${start}:$A${end}"
    )
    wb.defined_names["BudgetPlanned"] = DefinedName(
        "BudgetPlanned", attr_text=f"Budget!$B${start}:$B${end}"
    )
    wb.defined_names["BudgetActual"] = DefinedName(
        "BudgetActual", attr_text=f"Budget!$C${start}:$C${end}"
    )
    wb.defined_names["BudgetTotalPlanned"] = DefinedName(
        "BudgetTotalPlanned", attr_text=f"Budget!$B${total_row}"
    )
    wb.defined_names["BudgetTotalActual"] = DefinedName(
        "BudgetTotalActual", attr_text=f"Budget!$C${total_row}"
    )

    # Conditional formatting: variance < 0 red text already; % spent data bar
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color=PRIMARY, showValue=True),
    )

    # KPI side panel
    ws.cell(row=4, column=7, value="BUDGET KPIs").style = "section"
    kpi_rows = [
        ("Planned Budget",  "=BudgetTotalPlanned", '"$"#,##0.00'),
        ("Actual Spending", "=BudgetTotalActual",  '"$"#,##0.00'),
        ("Remaining",       "=TotalBudget-BudgetTotalActual", '"$"#,##0.00'),
        ("% Spent",         "=IFERROR(BudgetTotalActual/TotalBudget,0)", "0%"),
    ]
    for i, (lab, fml, fmt) in enumerate(kpi_rows):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml)
        c.style = "field_value"
        c.number_format = fmt
        c.font = Font(size=12, bold=True, color=PRIMARY)

    # Charts
    pie = DoughnutChart()
    pie.title = "Spending Breakdown"
    pie.height = 9
    pie.width = 14
    data = Reference(ws, min_col=3, min_row=4, max_row=end)
    labels = Reference(ws, min_col=1, min_row=start, max_row=end)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = no_labels()
    ws.add_chart(pie, "A18")

    bar = BarChart()
    bar.type = "col"
    bar.title = "Budget vs Actual"
    bar.height = 9
    bar.width = 14
    bar.add_data(Reference(ws, min_col=2, min_row=4, max_row=end),
                 titles_from_data=True)
    bar.add_data(Reference(ws, min_col=3, min_row=4, max_row=end),
                 titles_from_data=True)
    bar.set_categories(labels)
    ws.add_chart(bar, "G18")


# ===========================================================================
# Sheet 5: Clothing Inventory
# ===========================================================================
def build_clothing(wb: Workbook) -> None:
    ws = wb.create_sheet("Clothing")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [18, 26, 12, 16, 16, 14, 16, 24])
    page_header(ws, "H", "👕  CLOTHING INVENTORY",
                "Sizes, quantities, and what's left to buy.")
    headers = ["Category", "Item", "Size", "Qty Needed", "Qty Owned",
               "Cost / Unit", "Status", "Notes"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    rows = [
        ("Shirts",      "Polo (Navy)",            "S",  5, 2, 14.99, ""),
        ("Shirts",      "Polo (White)",           "S",  3, 1, 14.99, ""),
        ("Shirts",      "Long sleeve tee",        "S",  4, 4, 9.99,  ""),
        ("Pants",       "Khaki",                  "8",  3, 1, 19.99, ""),
        ("Pants",       "Jeans",                  "8",  2, 2, 24.99, ""),
        ("Shoes",       "Sneakers",               "4",  1, 0, 39.99, "Order online"),
        ("Shoes",       "Dress shoes",            "4",  1, 1, 29.99, ""),
        ("Jackets",     "Light jacket",           "S",  1, 0, 34.99, ""),
        ("Jackets",     "Winter coat",            "S",  1, 1, 59.99, "Last year still fits"),
        ("Uniforms",    "School sweater",         "S",  2, 0, 24.99, ""),
        ("Gym Clothes", "Athletic shorts",        "S",  3, 2, 11.99, ""),
        ("Gym Clothes", "Performance tee",        "S",  3, 1, 9.99,  ""),
        ("Accessories", "Belt",                   "S",  2, 1, 7.99,  ""),
        ("Accessories", "Socks (6pk)",            "M",  2, 1, 8.99,  ""),
        ("Accessories", "Hat",                    "S",  1, 0, 12.99, ""),
    ]
    start = 5
    end = start + 30 - 1
    for i, r_data in enumerate(rows):
        r = start + i
        cat, item, size, needed, owned, cost, notes = r_data
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=item)
        ws.cell(row=r, column=3, value=size)
        ws.cell(row=r, column=4, value=needed)
        ws.cell(row=r, column=5, value=owned)
        ws.cell(row=r, column=6, value=cost)
        ws.cell(row=r, column=8, value=notes)

    for r in range(start, end + 1):
        ws.cell(row=r, column=7,
                value=f'=IF(D{r}="","",IF(E{r}>=D{r},"Complete",IF(E{r}=0,"Not Started","In Progress")))')
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in (2, 8) else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 == 1 else BG)
            if c == 6:
                cell.number_format = '"$"#,##0.00'

    ws.conditional_formatting.add(
        f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"Complete"'],
                   fill=PatternFill("solid", fgColor="E3F8EF")),
    )
    ws.conditional_formatting.add(
        f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"Not Started"'],
                   fill=PatternFill("solid", fgColor="FBE6E6")),
    )


# ===========================================================================
# Sheet 6: Class Schedule
# ===========================================================================
def build_schedule(wb: Workbook) -> None:
    ws = wb.create_sheet("Schedule")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [14, 28, 28, 28, 28, 28])
    page_header(ws, "F", "📅  WEEKLY CLASS SCHEDULE",
                "Color-coded by subject. Drag in homework notes per slot.")
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    ws.cell(row=4, column=1, value="TIME").style = "th"
    for i, d in enumerate(days):
        ws.cell(row=4, column=2 + i, value=d).style = "th"
    ws.row_dimensions[4].height = 30

    times = ["8:00 – 8:55", "9:00 – 9:55", "10:00 – 10:55",
             "11:00 – 11:55", "12:00 – 12:45 (Lunch)", "12:50 – 1:45",
             "1:50 – 2:45", "2:50 – 3:30"]

    sample_grid = [
        ["Math · Hayes · R104",  "Math · Hayes · R104",  "Math · Hayes · R104",  "Math · Hayes · R104",  "Math · Hayes · R104"],
        ["English · Brooks · R208","Science · Patel · R301","English · Brooks · R208","Science · Patel · R301","English · Brooks · R208"],
        ["Science · Patel · R301","English · Brooks · R208","Science · Patel · R301","English · Brooks · R208","Science · Patel · R301"],
        ["History · Reed · R112", "Art · Lin · R150",      "History · Reed · R112", "Tech · Reyes · Lab2",   "Music · Park · R220"],
        ["LUNCH","LUNCH","LUNCH","LUNCH","LUNCH"],
        ["PE · Gomez · Gym",      "PE · Gomez · Gym",      "PE · Gomez · Gym",      "PE · Gomez · Gym",      "PE · Gomez · Gym"],
        ["Spanish · Ortega · R141","Spanish · Ortega · R141","Spanish · Ortega · R141","Spanish · Ortega · R141","Spanish · Ortega · R141"],
        ["Study Hall · R104",     "Study Hall · R104",     "Study Hall · R104",     "Club: Robotics · Lab2", "Study Hall · R104"],
    ]
    subject_colors = {
        "Math":     "D9E8E6",
        "English":  "EDE0CE",
        "Science":  "DCF5EC",
        "History":  "F0E4D2",
        "Art":      "F5E0D6",
        "Tech":     "DDE5F2",
        "Music":    "EAD9F0",
        "PE":       "F5E2E2",
        "Spanish":  "E0EBE1",
        "LUNCH":    SURFACE,
        "Study Hall": "EDE9E0",
        "Club":     "DCF5EC",
    }
    for ri, time_label in enumerate(times):
        r = 5 + ri
        ws.row_dimensions[r].height = 40
        c = ws.cell(row=r, column=1, value=time_label)
        c.style = "td"
        c.fill = fill(PRIMARY); c.font = Font(size=10, bold=True, color="FFFFFF")
        for ci, val in enumerate(sample_grid[ri]):
            cell = ws.cell(row=r, column=2 + ci, value=val)
            cell.style = "td_left"
            color = SOFT_BG
            for key, hex_ in subject_colors.items():
                if val.startswith(key):
                    color = hex_
                    break
            cell.fill = fill(color)


# ===========================================================================
# Sheet 7: Assignment Tracker
# ===========================================================================
def build_assignments(wb: Workbook) -> None:
    ws = wb.create_sheet("Assignments")
    ws.sheet_view.showGridLines = False
    widths = [14, 32, 14, 12, 14, 10, 28, 14]
    set_column_widths(ws, widths)
    page_header(ws, "H", "📝  ASSIGNMENT TRACKER",
                "Due dates, priorities, status — and live overdue flags.")
    headers = ["Subject", "Assignment", "Due Date", "Priority",
               "Status", "Grade", "Notes", "Days Until"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 32

    today = dt.date.today()

    def d(offset: int) -> dt.date:
        return today + dt.timedelta(days=offset)

    samples = [
        ("Math",     "Ch 3 problem set",            d(-2), "High",   "Complete",    92, "Showed all work"),
        ("English",  "Book report: Hatchet",        d(2),  "High",   "In Progress", "", "Outline done"),
        ("Science",  "Lab write-up",                d(-1), "Medium", "Not Started", "", "Need lab partner notes"),
        ("History",  "Civil War timeline",          d(5),  "Medium", "Not Started", "", ""),
        ("Art",      "Self portrait",               d(9),  "Low",    "In Progress", "", "Sketch stage"),
        ("Tech",     "Robotics build log",          d(0),  "High",   "In Progress", "", "Due end of day"),
        ("Spanish",  "Vocab quiz prep",             d(1),  "Medium", "Not Started", "", ""),
        ("Math",     "Quiz: fractions",             d(7),  "High",   "Not Started", "", "Study Mon evening"),
        ("English",  "Vocabulary list",             d(-3), "Low",    "Complete",   100, ""),
        ("Science",  "Reading: chapters 4–5",       d(3),  "Medium", "Not Started", "", ""),
    ]
    start = 5
    end = start + 50 - 1

    for i, s in enumerate(samples):
        r = start + i
        for ci, val in enumerate(s, 1):
            ws.cell(row=r, column=ci, value=val)

    for r in range(start, end + 1):
        ws.cell(row=r, column=3).number_format = "mm/dd/yyyy"
        ws.cell(row=r, column=8,
                value=f'=IF(C{r}="","",C{r}-TODAY())')
        for c in range(1, 9):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in (2, 7) else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 == 1 else BG)
            if c == 6:
                cell.number_format = "0"
            if c == 8:
                cell.number_format = "0;[Red]-0"

    add_dv(ws, f"A{start}:A{end}", "SubjectList")
    add_dv(ws, f"D{start}:D{end}", "PriorityList")
    add_dv(ws, f"E{start}:E{end}", "StatusList")

    # Overdue rows: due date < today AND status != Complete
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'AND($C{start}<>"",$C{start}<TODAY(),$E{start}<>"Complete")'],
                    fill=PatternFill("solid", fgColor="FBE6E6")),
    )
    # Complete: soft mint
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'$E{start}="Complete"'],
                    fill=PatternFill("solid", fgColor="E3F8EF")),
    )
    # Due this week
    ws.conditional_formatting.add(
        f"A{start}:H{end}",
        FormulaRule(formula=[f'AND($C{start}<>"",$C{start}-TODAY()>=0,$C{start}-TODAY()<=7,$E{start}<>"Complete")'],
                    fill=PatternFill("solid", fgColor="FBF0E2")),
    )

    wb.defined_names["AsgStatus"] = DefinedName(
        "AsgStatus", attr_text=f"Assignments!$E${start}:$E${end}"
    )
    wb.defined_names["AsgDue"] = DefinedName(
        "AsgDue", attr_text=f"Assignments!$C${start}:$C${end}"
    )
    wb.defined_names["AsgSubject"] = DefinedName(
        "AsgSubject", attr_text=f"Assignments!$A${start}:$A${end}"
    )

    ws.freeze_panes = "A5"


# ===========================================================================
# Sheet 8: Extracurricular Activities
# ===========================================================================
def build_extracurricular(wb: Workbook) -> None:
    ws = wb.create_sheet("Extracurricular")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [20, 22, 16, 18, 14, 28, 24])
    page_header(ws, "G", "🎯  EXTRACURRICULAR ACTIVITIES",
                "Practices, fees, equipment, and contacts in one view.")
    headers = ["Activity", "Coach / Instructor", "Practice Day",
               "Practice Time", "Fee", "Equipment Needed", "Contact"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 32

    samples = [
        ("Soccer (U10)",      "Coach Reyes",   "Mon / Wed", "4:30 – 6:00pm",  120, "Cleats, shin guards, ball", "reyes@league.org"),
        ("Robotics Club",     "Mr. Patel",     "Thursday",  "3:30 – 5:00pm",   45, "Laptop",                    "patel@school.edu"),
        ("Piano Lessons",     "Ms. Park",      "Saturday",  "10:00 – 10:30am", 35, "Practice book",             "(555) 412-8841"),
        ("Art Workshop",      "Ms. Lin",       "Friday",    "3:30 – 4:30pm",   25, "Sketchbook, pencils",       "lin@studio.com"),
        ("Math Olympiad",     "Mr. Hayes",     "Tuesday",   "3:30 – 4:30pm",    0, "—",                         "hayes@school.edu"),
        ("Swimming",          "Coach Diaz",    "Sun / Wed", "8:00 – 9:00am",   80, "Swimsuit, goggles, towel",  "(555) 901-2271"),
    ]
    start = 5
    end = start + 20 - 1
    for i, s in enumerate(samples):
        r = start + i
        for ci, val in enumerate(s, 1):
            ws.cell(row=r, column=ci, value=val)
    for r in range(start, end + 1):
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 == 1 else BG)
            if c == 5:
                cell.number_format = '"$"#,##0.00'


# ===========================================================================
# Sheet 9: School Calendar
# ===========================================================================
def build_calendar(wb: Workbook) -> None:
    ws = wb.create_sheet("Calendar")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [14, 14, 24, 28, 18, 24])
    page_header(ws, "F", "🗓  SCHOOL-YEAR CALENDAR",
                "Logged events: holidays, exams, meetings, deadlines.")
    headers = ["Date", "Day", "Event", "Notes", "Type", "Student"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h).style = "th"
    ws.row_dimensions[4].height = 30

    samples = [
        (dt.date(2026, 9,  2), "Wed", "First day of school",        "Bus pickup 7:35am", "Milestone",        "All"),
        (dt.date(2026, 9, 14), "Mon", "Picture day",                "Wear school colors","Picture Day",      "All"),
        (dt.date(2026, 9, 21), "Mon", "Back-to-school night",       "6:30pm – 8:30pm",  "Parent-Teacher",   "All"),
        (dt.date(2026,10,  9), "Fri", "Teacher work day (no school)","",                "Holiday",          "All"),
        (dt.date(2026,10, 26), "Mon", "Midterm exam — Math",        "Bring calculator", "Exam",             "Liam"),
        (dt.date(2026,11, 13), "Fri", "Field trip — Museum",        "Permission slip due 11/6","Field Trip","Emma"),
        (dt.date(2026,11, 26), "Thu", "Thanksgiving break begins",  "Return 12/1",      "Holiday",          "All"),
        (dt.date(2026,12, 18), "Fri", "Winter break begins",        "Return 1/5",       "Holiday",          "All"),
        (dt.date(2027, 1, 15), "Fri", "Report cards issued",        "",                 "Milestone",        "All"),
        (dt.date(2027, 2, 12), "Fri", "Parent-teacher conferences", "Sign up online",   "Parent-Teacher",   "All"),
        (dt.date(2027, 3, 26), "Fri", "Spring break begins",        "Return 4/5",       "Holiday",          "All"),
        (dt.date(2027, 5, 24), "Mon", "Final exams week",           "",                 "Exam",             "Liam"),
        (dt.date(2027, 6, 12), "Fri", "Last day of school",         "Early dismissal 12pm","Milestone",     "All"),
    ]
    start = 5
    end = start + 40 - 1
    for i, s in enumerate(samples):
        r = start + i
        for ci, val in enumerate(s, 1):
            ws.cell(row=r, column=ci, value=val)
    for r in range(start, end + 1):
        ws.cell(row=r, column=1).number_format = "mm/dd/yyyy"
        for c in range(1, 7):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in (3, 4, 6) else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 == 1 else BG)

    # Highlight upcoming events (next 14 days)
    ws.conditional_formatting.add(
        f"A{start}:F{end}",
        FormulaRule(formula=[f'AND($A{start}<>"",$A{start}-TODAY()>=0,$A{start}-TODAY()<=14)'],
                    fill=PatternFill("solid", fgColor="DCF5EC")),
    )
    # Past events: muted
    ws.conditional_formatting.add(
        f"A{start}:F{end}",
        FormulaRule(formula=[f'AND($A{start}<>"",$A{start}<TODAY())'],
                    fill=PatternFill("solid", fgColor="F1F1F1")),
    )


# ===========================================================================
# Sheet 10: Lunch & Meal Planner
# ===========================================================================
def build_meals(wb: Workbook) -> None:
    ws = wb.create_sheet("Meals")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [14, 26, 26, 26, 26, 26, 6, 22, 14])
    page_header(ws, "I", "🍎  LUNCH & MEAL PLANNER",
                "Weekly meals + auto-build grocery list.")
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    meal_types = ["Breakfast", "Lunch", "Snack"]

    ws.cell(row=4, column=1, value="MEAL").style = "th"
    for i, d in enumerate(days):
        ws.cell(row=4, column=2 + i, value=d).style = "th"
    ws.row_dimensions[4].height = 30

    plans = {
        "Breakfast": ["Oatmeal + berries", "Yogurt parfait", "Scrambled eggs + toast",
                     "Smoothie + granola", "Pancakes"],
        "Lunch":     ["Turkey wrap + apple", "PB&J + carrots", "Pasta salad + grapes",
                     "Chicken nuggets + cucumber", "Cheese pizza slice + orange"],
        "Snack":     ["Trail mix", "Cheese stick + crackers", "Apple slices + PB",
                     "Granola bar", "Veggie chips"],
    }
    for ri, meal in enumerate(meal_types):
        r = 5 + ri
        ws.row_dimensions[r].height = 42
        c = ws.cell(row=r, column=1, value=meal)
        c.style = "td"; c.fill = fill(PRIMARY); c.font = Font(size=11, bold=True, color="FFFFFF")
        for ci, plan in enumerate(plans[meal]):
            cell = ws.cell(row=r, column=2 + ci, value=plan)
            cell.style = "td_left"
            cell.fill = fill(SOFT_BG if ci % 2 == 0 else BG)

    # Dietary notes block
    note_row = 5 + len(meal_types) + 2
    merge_set(ws, f"A{note_row}:F{note_row}", "DIETARY NOTES", "section")
    merge_set(ws, f"A{note_row+1}:F{note_row+3}",
              "Emma: peanut allergy — strictly no peanuts in lunch packing.\n"
              "Liam: vegetarian Mondays.\n"
              "Family: gluten-free option preferred Wednesdays.",
              "field_value")
    ws.row_dimensions[note_row + 1].height = 22
    ws.row_dimensions[note_row + 2].height = 22
    ws.row_dimensions[note_row + 3].height = 22

    # Grocery list (right side)
    merge_set(ws, "H4:I4", "GROCERY LIST", "th")
    ws.row_dimensions[4].height = 30
    items = [
        ("Bread (whole grain)",    2),
        ("Turkey slices",          1),
        ("Cheese sticks",          8),
        ("Apples",                10),
        ("Carrots (baby)",         1),
        ("Cucumbers",              2),
        ("Yogurt cups",            8),
        ("Granola",                1),
        ("Trail mix",              1),
        ("Eggs (dozen)",           1),
        ("Oats",                   1),
        ("Berries (mixed)",        2),
        ("Pasta",                  1),
        ("Pizza slices (frozen)",  5),
        ("Veggie chips",           1),
    ]
    for i, (item, qty) in enumerate(items):
        r = 5 + i
        ws.cell(row=r, column=8, value=item).style = "td_left"
        c = ws.cell(row=r, column=9, value=qty); c.style = "td"; c.number_format = "0"
        if i % 2 == 1:
            ws.cell(row=r, column=8).fill = fill(MUTED_ROW)
            ws.cell(row=r, column=9).fill = fill(MUTED_ROW)


# ===========================================================================
# Sheet 11: Emergency Contacts
# ===========================================================================
def build_emergency(wb: Workbook) -> None:
    ws = wb.create_sheet("Emergency")
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [22, 22, 22, 28, 30])
    page_header(ws, "E", "🚨  EMERGENCY CONTACTS & INFO",
                "Keep this printable and accessible.")

    sections = [
        ("FAMILY", [
            ("Mom",    "Sarah Carter",    "(555) 412-7833", "sarah@example.com",  "Primary"),
            ("Dad",    "James Carter",    "(555) 412-7901", "james@example.com",  "Secondary"),
            ("Grandma","Linda Reyes",     "(555) 904-1320", "linda@example.com",  "Pickup authorized"),
        ]),
        ("MEDICAL", [
            ("Pediatrician",   "Dr. Mira Singh",     "(555) 388-2210", "Westside Pediatrics", "Mon–Fri 8–5"),
            ("Dentist",        "Dr. Aaron Park",     "(555) 388-9911", "Park Family Dental",  ""),
            ("Allergist",      "Dr. Eleanor Wu",     "(555) 388-3344", "Allergy Partners",    "EpiPen: see backpack"),
        ]),
        ("SCHOOL", [
            ("Lincoln Elementary office", "—",  "(555) 200-1188", "lincoln@district.edu", "Emma · Rm 104"),
            ("Roosevelt Middle office",   "—",  "(555) 200-2244", "rms@district.edu",     "Liam · Rm 8B"),
            ("After-school program",      "Ms. Hall", "(555) 200-7711", "afterschool@district.edu", ""),
        ]),
        ("INSURANCE", [
            ("Health insurance",  "BlueShield",     "1-800-555-0000", "Member ID: 8821-44A", "Group #14422"),
            ("Dental insurance",  "DeltaDental",    "1-800-555-7700", "Member ID: 8821-44D", ""),
        ]),
    ]
    row = 4
    for title, rows in sections:
        merge_set(ws, f"A{row}:E{row}", title, "section")
        ws.row_dimensions[row].height = 22
        row += 1
        for h in ["Role", "Name", "Phone", "Email / Provider", "Notes"]:
            ws.cell(row=row, column=ord(h[0]), value=h)
        # actually set headers cleanly:
        for i, h in enumerate(["Role", "Name", "Phone", "Email / Provider", "Notes"], 1):
            c = ws.cell(row=row, column=i, value=h)
            c.style = "th"
        ws.row_dimensions[row].height = 26
        row += 1
        for i, contact in enumerate(rows):
            for ci, val in enumerate(contact, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.style = "td_left"
                cell.fill = fill(MUTED_ROW if i % 2 == 1 else BG)
            row += 1
        row += 1  # spacer


# ===========================================================================
# Sheet 1: Dashboard
# ===========================================================================
def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_column_widths(ws, [2, 18, 18, 18, 18, 18, 18, 18, 18, 2])
    ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:J1", "  🎒  BACK-TO-SCHOOL COMMAND CENTER", "title")
    ws.row_dimensions[2].height = 22
    merge_set(ws, "A2:J2",
              "  Plan · Budget · Organize · Prepare — one premium dashboard for the whole school year.",
              "subtitle")

    # Top KPI row (4 cards)
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 56
    kpi_row1 = [
        ("💵 Total Budget",      "=TotalBudget",                     "kpi_money"),
        ("💳 Total Spent",       "=BudgetTotalActual",               "kpi_money"),
        ("🧮 Remaining",         "=TotalBudget-BudgetTotalActual",   "kpi_money"),
        ("⏳ Days Until School",
         "=MAX(FirstDayOfSchool-TODAY(),0)",                          "kpi_value"),
    ]
    col = 2
    for label, formula, val_style in kpi_row1:
        rng_label = f"{get_column_letter(col)}4:{get_column_letter(col+1)}4"
        rng_val = f"{get_column_letter(col)}5:{get_column_letter(col+1)}5"
        merge_set(ws, rng_label, label, "kpi_label")
        merge_set(ws, rng_val, formula, val_style)
        for r in (4, 5):
            for cc in range(col, col + 2):
                ws.cell(row=r, column=cc).border = BOX
        col += 2

    # Second KPI row
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[8].height = 56
    kpi_row2 = [
        ("🛒 Items Purchased",
         '=SUMPRODUCT((SupPurchased="Yes")*1)',           "kpi_value"),
        ("📋 Items Remaining",
         '=SUMPRODUCT((SupItem<>"")*(SupPurchased<>"Yes"))', "kpi_value"),
        ("🎒 Supply Completion",
         '=IFERROR(SUMPRODUCT((SupPurchased="Yes")*1)/SUMPRODUCT((SupItem<>"")*1),0)',
         "kpi_pct"),
        ("✅ Assignments Done",
         '=IFERROR(SUMPRODUCT((AsgStatus="Complete")*1)/SUMPRODUCT((AsgStatus<>"")*1),0)',
         "kpi_pct"),
    ]
    col = 2
    for label, formula, val_style in kpi_row2:
        rng_label = f"{get_column_letter(col)}7:{get_column_letter(col+1)}7"
        rng_val = f"{get_column_letter(col)}8:{get_column_letter(col+1)}8"
        merge_set(ws, rng_label, label, "kpi_label")
        merge_set(ws, rng_val, formula, val_style)
        for r in (7, 8):
            for cc in range(col, col + 2):
                ws.cell(row=r, column=cc).border = BOX
        col += 2

    # Quick navigation pills (row 10)
    ws.row_dimensions[10].height = 26
    merge_set(ws, "B10:I10", "QUICK NAVIGATION", "section")
    nav = ["Shopping List", "Budget", "Calendar",
           "Students", "Assignments", "Meals", "Emergency"]
    ws.row_dimensions[11].height = 30
    span_cols = 8 // 1  # we'll just lay 7 chips into 7 cells from B..H
    for i, name in enumerate(nav):
        cell = ws.cell(row=11, column=2 + i, value=name)
        cell.fill = fill(PRIMARY)
        cell.font = Font(size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BOX

    # Charts row label
    ws.row_dimensions[13].height = 26
    merge_set(ws, "B13:I13", "ANALYTICS", "section")

    # Charts
    # Spending by category (donut)
    donut = DoughnutChart()
    donut.title = "Spending by Category"
    donut.height = 9; donut.width = 12
    data = Reference(wb["Budget"], min_col=3, min_row=4, max_row=4 + len(CATEGORIES))
    labels = Reference(wb["Budget"], min_col=1, min_row=5, max_row=4 + len(CATEGORIES))
    donut.add_data(data, titles_from_data=True)
    donut.set_categories(labels)
    donut.dataLabels = no_labels()
    ws.add_chart(donut, "B14")

    # Budget vs Actual (bar)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Budget vs Actual"
    bar.height = 9; bar.width = 12
    bar.add_data(Reference(wb["Budget"], min_col=2, min_row=4, max_row=4 + len(CATEGORIES)),
                 titles_from_data=True)
    bar.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=4 + len(CATEGORIES)),
                 titles_from_data=True)
    bar.set_categories(labels)
    ws.add_chart(bar, "H14")

    # Footer
    ws.row_dimensions[34].height = 26
    merge_set(ws, "B34:I34",
              "Back-to-School Command Center v1.0  ·  Edit Settings → Total Budget · First Day  ·  Add items on Supplies",
              "subtitle")


# ===========================================================================
# Build
# ===========================================================================
def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    build_settings(wb)
    build_students(wb)
    build_supplies(wb)
    build_budget(wb)
    build_clothing(wb)
    build_schedule(wb)
    build_assignments(wb)
    build_extracurricular(wb)
    build_calendar(wb)
    build_meals(wb)
    build_emergency(wb)
    build_dashboard(wb)

    order = ["Dashboard", "Students", "Supplies", "Budget", "Clothing",
             "Schedule", "Assignments", "Extracurricular", "Calendar",
             "Meals", "Emergency", "Settings"]
    wb._sheets = [wb[name] for name in order]

    tab_colors = {
        "Dashboard": PRIMARY,
        "Students":  ACCENT,
        "Supplies":  HIGHLIGHT,
        "Budget":    PRIMARY,
        "Clothing":  ACCENT,
        "Schedule":  HIGHLIGHT,
        "Assignments": PRIMARY,
        "Extracurricular": ACCENT,
        "Calendar":  HIGHLIGHT,
        "Meals":     ACCENT,
        "Emergency": DANGER,
        "Settings":  SURFACE,
    }
    for name, color in tab_colors.items():
        wb[name].sheet_properties.tabColor = color

    out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    out_path = os.path.join(out_dir, "Back_To_School_Command_Center.xlsx")
    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
