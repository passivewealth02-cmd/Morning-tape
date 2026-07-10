"""Build Road Trip Command Center™ — The Ultimate Road Trip Planning, Budget &
Adventure Management System.

19 sheets + Welcome · a premium road trip operating system in Excel & Sheets.
Dashboard, trip profile, route, itinerary, budget, fuel, vehicle, stays,
camping, attractions, food, packing, emergency, journal, gallery, parks,
rewards, analytics & settings — one elegant command center.

Run: python3 build_xlsx.py   ->  ../Road_Trip_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

EXP_CATS = ["Fuel", "Hotels", "Campgrounds", "Food & Dining", "Attractions",
            "Vehicle Maintenance", "Parking", "Tolls", "Insurance", "Souvenirs",
            "Emergency Fund", "Miscellaneous"]
STAY_TYPES = ["Hotel", "Cabin", "Campground", "RV Park", "Vacation Rental", "Motel"]
ACT_TYPES = ["National Park", "Scenic Drive", "Hike", "Landmark", "Museum", "Town", "Outdoor", "Food"]
PACK_CATS = ["Clothing", "Camping Gear", "Electronics", "Vehicle Supplies", "Emergency Kit",
             "Cooking Equipment", "Hiking Gear", "Documents", "Pet Supplies", "Kids Items"]
FUEL_TYPES = ["Regular", "Mid-grade", "Premium", "Diesel", "EV"]
VEH_TYPES = ["SUV", "Camper Van", "RV", "Truck", "Sedan", "Motorcycle", "Crossover"]
ROAD_COND = ["Clear", "Construction", "Mountain", "Gravel", "Weather", "Remote"]
VEH_STATUS = ["OK", "Due Soon", "Overdue"]
RES_STATUS = ["Reserved", "Requested", "Open", "Waitlist"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]
CURRENCIES = ["USD", "CAD", "EUR", "GBP", "AUD"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def totals(ws, row, cols, start, end, fmt='"$"#,##0', label="TOTAL"):
    ws.cell(row=row, column=1, value=label).style = "th"
    for col in cols:
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt


# ===========================================================================
# 19 — Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 22, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your trip & vehicle details once — every dashboard follows.")
    merge_set(ws, "B5:C5", "TRIP INPUTS", "section")
    controls = [
        ("Trip Name", "Great Southwest Loop", None, "TripName"),
        ("Start Location", "Las Vegas, NV", None, "StartLoc"),
        ("Destination", "5 National Parks", None, "Destination"),
        ("Departure Date", dplus(30), "mm/dd/yyyy", "TripStart"),
        ("Return Date", dplus(41), "mm/dd/yyyy", "TripEnd"),
        ("Travelers", 2, "0", "TravelerCount"),
        ("Vehicle", "2020 Ford Transit Camper", None, "VehicleName"),
        ("Avg MPG", 16, "0.0", "AvgMPG"),
        ("Fuel Price ($/gal)", 3.60, '"$"#,##0.00', "FuelPrice"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Expense Category", EXP_CATS, "ExpCatList"), ("F", "Accommodation", STAY_TYPES, "StayTypeList"),
             ("G", "Activity Type", ACT_TYPES, "ActTypeList"), ("H", "Packing Category", PACK_CATS, "PackCatList"),
             ("I", "Fuel Type", FUEL_TYPES, "FuelTypeList"), ("J", "Vehicle Type", VEH_TYPES, "VehTypeList"),
             ("K", "Road Condition", ROAD_COND, "RoadCondList"), ("L", "Reservation", RES_STATUS, "ResStatusList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 20, "Priority", PRIORITIES, "PriorityList"), ("F", 20, "Yes / No", YESNO, "YesNoList"),
             ("G", 20, "Veh. Status", VEH_STATUS, "VehStatusList"), ("H", 20, "Currency", CURRENCIES, "CurrencyList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🚗  ROAD TRIP COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate road trip planning, budget & adventure management system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "PLAN EVERY MILE FROM ONE ELEGANT COMMAND CENTER", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("From the first pin on the map to the last sunset photo — Road Trip Command Center™ manages your "
                      "route, budget, fuel, vehicle, campgrounds, hotels, attractions, packing, safety and memories in "
                      "ONE premium Excel & Google Sheets system. Drive further, spend smarter, break down less and "
                      "remember more — all with app-level automation. This isn't a road trip checklist — it's your "
                      "complete Road Trip Operating System.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your trip, dates, travelers, vehicle & MPG.",
             "2.  Map your Route & Daily Itinerary — miles and drive time total themselves.",
             "3.  Build your Budget & log Fuel — cost per mile & per day calculate live.",
             "4.  Run the Vehicle checklist — readiness score & maintenance reminders.",
             "5.  Book Stays, Campgrounds & Attractions; work the Packing list.",
             "6.  Hit the road, then capture it all in the Journal, Gallery & Parks checklist."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("A realistic sample trip (a 12-day, 2-person, 1,555-mile Southwest national-park loop from Las Vegas on a "
               "$4,800 budget) is included so you can see how everything connects — just type over it with your own. "
               "The countdown, total distance, fuel economy, budget remaining, packing %, vehicle readiness and the "
               "Trip Readiness Score all update automatically. Every sheet is print-friendly and works in Excel and "
               "Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more miles of memories — let's plan the adventure.", "section_gold")


# ===========================================================================
# 2 — Trip Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Trip Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 4, 24, 30, 2])
    luxe_header(ws, "G", "🧭  TRIP PROFILE", "Your adventure at a glance — the who, what, where and how.")
    blocks = [
        ("THE TRIP", [("Trip Name", "=TripName"), ("Start", "=StartLoc"),
                      ("Destination", "=Destination"), ("Travelers", "=TravelerCount"),
                      ("Departure", "=TEXT(TripStart,\"ddd, mmm d\")"), ("Return", "=TEXT(TripEnd,\"ddd, mmm d\")")]),
        ("THE RIG", [("Vehicle", "=VehicleName"), ("Avg MPG", "=AvgMPG"),
                     ("Days on the Road", "=TripEnd-TripStart+1"), ("Trip Theme", "Parks, scenic drives & camping"),
                     ("Sleeping Setup", "Van + campgrounds"), ("Trip Motto", "The journey is the destination")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "EMERGENCY CONTACTS", "section_gold"); ws.row_dimensions[15].height = 22
    contacts = [("Emergency", "911"), ("Roadside Assist", "800-555-0199"),
                ("Home Contact (Mom)", "555-0110"), ("Insurance 24/7", "800-555-0100"),
                ("Vehicle Roadside", "800-555-0177"), ("Trip Buddy", "555-0142")]
    for i, (p, h) in enumerate(contacts):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 3 — Route Planner
# ===========================================================================
def build_route(wb):
    rows = [
        ("Day 1", "Las Vegas → Zion NP", "Valley of Fire detour", 165, 2.8, "Fill up before the park"),
        ("Day 2", "Zion NP", "Angels Landing / Narrows", 25, 0.6, "Shuttle-only in the canyon"),
        ("Day 3", "Zion → Bryce Canyon", "Red Canyon arches", 85, 1.6, "Cooler — pack layers"),
        ("Day 4", "Bryce → Capitol Reef", "Scenic Byway 12", 120, 2.4, "One of the best drives in the US"),
        ("Day 5", "Capitol Reef → Moab", "Goblin Valley", 150, 2.6, "Long open stretch — fuel up"),
        ("Day 6", "Arches + Canyonlands", "Delicate Arch sunset", 70, 1.8, "Timed entry for Arches"),
        ("Day 7", "Moab → Mesa Verde", "Colorado border", 150, 2.7, "Check tire pressure"),
        ("Day 8", "Mesa Verde → Monument Valley", "Four Corners", 165, 3.0, "Fuel scarce — top off"),
        ("Day 9", "Monument Valley → Grand Canyon", "Cameron Trading Post", 180, 3.2, "Long day — early start"),
        ("Day 10", "Grand Canyon South Rim", "Desert View Drive", 50, 1.2, "Sunset at Hopi Point"),
        ("Day 11", "Grand Canyon → Sedona", "Oak Creek Canyon", 115, 2.3, "Winding — take it slow"),
        ("Day 12", "Sedona → Las Vegas", "Hoover Dam stop", 280, 4.4, "Home stretch"),
    ]
    ws, start, end = build_log(
        wb, "Route", "🗺", "ROUTE PLANNER",
        "The whole drive, mapped — daily miles & drive time total themselves.",
        ["Day", "Route (From → To)", "Scenic Stops / Parks", "Miles", "Drive Time (h)", "Road Notes"],
        rows, [9, 30, 26, 10, 13, 30],
        text_left={2, 3, 6}, ints={4}, dec={5}, reserved=24)
    nrange(wb, "DailyMiles", "Route", "D", start, end)
    nrange(wb, "DriveTime", "Route", "E", start, end)
    totals(ws, end + 1, [4], start, end, fmt="#,##0", label="TOTAL MILES")
    ws.cell(row=end + 1, column=5, value=f"=SUM(E{start}:E{end})").number_format = "0.0"
    ws.cell(row=end + 1, column=5).font = Font(bold=True, color=PRIMARY); ws.cell(row=end + 1, column=5).fill = fill(SURFACE)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=300, color=PRIMARY, showValue=True))


# ===========================================================================
# 4 — Daily Itinerary
# ===========================================================================
def build_itinerary(wb):
    rows = [
        (0, "8:00a", "11:00a", 2.8, "Valley of Fire", "Picnic lunch", "Zion Canyon scenic drive", "Watchman CG"),
        (1, "7:00a", "3:00p", 0.6, "In-park shuttle", "Zion Lodge cafe", "Angels Landing hike", "Watchman CG"),
        (2, "8:30a", "10:30a", 1.6, "Red Canyon", "Bryce general store", "Sunset & Navajo Loop", "North CG (Bryce)"),
        (3, "8:00a", "12:00p", 2.4, "Byway 12 overlooks", "Boulder farm cafe", "Capitol Reef orchard walk", "Cabin"),
        (4, "8:00a", "1:00p", 2.6, "Goblin Valley", "Cook in van", "Arrive Moab, resupply", "Moab Inn"),
        (5, "6:00a", "9:00p", 1.8, "Canyonlands sunrise", "Moab BBQ", "Delicate Arch sunset", "Devils Garden CG"),
        (6, "8:00a", "12:30p", 2.7, "Wilson Arch", "Roadside diner", "Mesa Verde cliff dwellings", "Mesa Verde"),
        (7, "8:00a", "1:00p", 3.0, "Four Corners", "Trading post", "Monument Valley loop drive", "MV Cabin"),
    ]
    sample = [(dplus(30 + d), dep, arr, hrs, stop, meal, act, stay) for (d, dep, arr, hrs, stop, meal, act, stay) in rows]
    ws, start, end = build_log(
        wb, "Itinerary", "🗓", "DAILY ITINERARY",
        "Every day mapped — departure, arrival, drive time, stops, meals & where you sleep.",
        ["Date", "Depart", "Arrive", "Drive (h)", "Stops", "Meals", "Activities", "Accommodation"],
        sample, [13, 10, 10, 11, 18, 16, 24, 18],
        text_left={5, 6, 7, 8}, dates={1}, dec={4}, reserved=24)


# ===========================================================================
# 5 — Budget Command Center
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 13, 13, 13, 3, 22, 14, 2])
    luxe_header(ws, "H", "💰  ROAD TRIP BUDGET COMMAND CENTER",
                "Every dollar planned — budget vs actual, cost per traveler & per day, live.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining"])
    planned = {"Fuel": 520, "Hotels": 1400, "Campgrounds": 240, "Food & Dining": 820, "Attractions": 400,
               "Vehicle Maintenance": 300, "Parking": 60, "Tolls": 40, "Insurance": 120, "Souvenirs": 200,
               "Emergency Fund": 400, "Miscellaneous": 300}
    actual = {"Fuel": 0, "Hotels": 1050, "Campgrounds": 180, "Food & Dining": 0, "Attractions": 120,
              "Vehicle Maintenance": 300, "Parking": 0, "Tolls": 0, "Insurance": 0, "Souvenirs": 0,
              "Emergency Fund": 0, "Miscellaneous": 0}
    start = L0; end = start + len(EXP_CATS) - 1
    for i, cat in enumerate(EXP_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=2, value=planned[cat]); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=actual[cat]); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL BUDGET").style = "th"
    for col in (2, 3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG), font=Font(color=DANGER, bold=True)))
    nrange(wb, "ExpCat", "Budget", "A", start, end)
    nrange(wb, "ExpPlanned", "Budget", "B", start, end)
    nrange(wb, "ExpActual", "Budget", "C", start, end)
    cell_name(wb, "TripBudget", "Budget", f"$B${tot}")
    cell_name(wb, "SpentTotal", "Budget", f"$C${tot}")
    # savings sources
    merge_set(ws, "G4:H4", "TRAVEL FUND", "section_gold")
    sources = [("Travel Savings", 4000), ("Reward Points (value)", 300), ("Gift Contributions", 200)]
    sstart = 5
    for i, (src, amt) in enumerate(sources):
        r = sstart + i
        ws.cell(row=r, column=7, value=src).style = "td_left"
        c = ws.cell(row=r, column=8, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=7).fill = fill(MUTED_ROW); ws.cell(row=r, column=8).fill = fill(MUTED_ROW)
    send = sstart + len(sources) - 1; stot = send + 1
    ws.cell(row=stot, column=7, value="MONEY SAVED").style = "th"
    c = ws.cell(row=stot, column=8, value=f"=SUM(H{sstart}:H{send})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "SavedTotal", "Budget", f"$H${stot}")
    # bottom line
    merge_set(ws, "G10:H10", "THE BOTTOM LINE", "section_gold")
    lines = [("Total budget", "=TripBudget", '"$"#,##0'), ("Money saved", "=SavedTotal", '"$"#,##0'),
             ("Spent so far", "=SpentTotal", '"$"#,##0'), ("Budget remaining", "=TripBudget-SpentTotal", '"$"#,##0'),
             ("Cost / traveler", "=IFERROR(TripBudget/TravelerCount,0)", '"$"#,##0'),
             ("Cost / day", "=IFERROR(TripBudget/(TripEnd-TripStart+1),0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(lines):
        r = 11 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Budget remaining", "Money saved"):
            ws.cell(row=r, column=8).fill = fill(MINT_BG)


# ===========================================================================
# 6 — Fuel Tracker
# ===========================================================================
def build_fuel(wb):
    rows = [
        (dminus(3), "Costco — Las Vegas", 14.2, 3.42, "=C5*D5", 41200, "", "Pre-trip top-off"),
        (dminus(1), "Chevron — Vegas", 6.0, 3.58, "=C6*D6", 41360, "=IFERROR(ROUND((F6-F5)/C6,1),0)", "Full before departure"),
    ]
    ws, start, end = build_log(
        wb, "Fuel", "⛽", "FUEL TRACKER",
        "Every fill-up logged — average MPG and fuel cost per mile calculate themselves.",
        ["Date", "Station", "Gallons", "$/Gal", "Total", "Odometer", "MPG", "Notes"],
        rows, [13, 22, 12, 11, 12, 13, 11, 22],
        text_left={2, 8}, dates={1}, money={5}, ints={6}, dec={3, 7}, reserved=30)
    ws.cell(row=start, column=4).number_format = '"$"#,##0.00'
    ws.cell(row=start + 1, column=4).number_format = '"$"#,##0.00'
    for r in range(start, start + 2):
        ws.cell(row=r, column=5).number_format = '"$"#,##0.00'
    nrange(wb, "FuelGallons", "Fuel", "C", start, end)
    nrange(wb, "FuelTotal", "Fuel", "E", start, end)
    # summary block
    merge_set(ws, "B" + str(end + 2) + ":C" + str(end + 2), "FUEL SUMMARY", "section_gold")
    sr = end + 3
    summ = [("Trip distance (mi)", "=TotalDistance", "#,##0"), ("Avg MPG (planned)", "=AvgMPG", "0.0"),
            ("Est. fuel needed (gal)", "=IFERROR(TotalDistance/AvgMPG,0)", "0.0"),
            ("Est. fuel cost", "=IFERROR(TotalDistance/AvgMPG*FuelPrice,0)", '"$"#,##0'),
            ("Fuel cost / mile", "=IFERROR(FuelPrice/AvgMPG,0)", '"$"#,##0.00')]
    for i, (lab, fml, fmt) in enumerate(summ):
        r = sr + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Est. fuel cost":
            c.fill = fill(MINT_BG)
    cell_name(wb, "FuelEstimate", "Fuel", f"$C${sr+3}")


# ===========================================================================
# 7 — Vehicle Command Center
# ===========================================================================
def build_vehicle(wb):
    ws = wb.create_sheet("Vehicle"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 16, 16, 26, 2])
    luxe_header(ws, "G", "🔧  VEHICLE COMMAND CENTER",
                "Trip-ready & road-safe — every check, with automatic maintenance reminders.")
    # vehicle info
    merge_set(ws, "B5:C5", "VEHICLE INFO", "section")
    info = [("Vehicle", "=VehicleName"), ("VIN", "1FT****TRANSIT"), ("Plate", "NV · VAN-LYF"),
            ("Insurance", "Progressive — active"), ("Registration", "Exp. next year"), ("Roadside", "AAA Premier")]
    for i, (lab, val) in enumerate(info):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        ws.cell(row=r, column=3, value=val).style = "field_value"
    # checklist
    merge_set(ws, "D5:F5", "PRE-TRIP CHECKLIST", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Check", "Status", "Notes"], start_col=4)
    checks = [
        ("Oil change", "OK", "Done last week"), ("Tire pressure & tread", "OK", "All 40 psi"),
        ("Brake inspection", "OK", "Pads 60%"), ("Battery", "OK", "Tested good"),
        ("Wiper blades", "Due Soon", "Replace before trip"), ("Coolant / fluids", "OK", "Topped off"),
        ("Spare tire & jack", "OK", "In rear"), ("Lights & signals", "OK", "All working"),
        ("Air filter", "Due Soon", "Order online"), ("Emergency kit", "OK", "Restocked"),
    ]
    cs = 7
    for i, (chk, status, note) in enumerate(checks):
        r = cs + i
        ws.cell(row=r, column=4, value=chk).style = "td_left"
        ws.cell(row=r, column=5, value=status).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(4, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ce = cs + len(checks) - 1
    add_dv(ws, f"E{cs}:E{ce}", "VehStatusList")
    nrange(wb, "VehStatus", "Vehicle", "E", cs, ce)
    for st, cc in {"OK": MINT_BG, "Due Soon": WARN_BG, "Overdue": RED_BG}.items():
        ws.conditional_formatting.add(f"E{cs}:E{ce}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    # readiness
    merge_set(ws, "B14:C14", "READINESS", "section")
    ws.cell(row=15, column=2, value="Vehicle Readiness").style = "field_label"
    c = ws.cell(row=15, column=3, value='=IFERROR(COUNTIF(VehStatus,"OK")/COUNTA(VehStatus),0)'); c.style = "field_value"; c.number_format = "0%"; c.fill = fill(MINT_BG)
    ws.cell(row=16, column=2, value="Items to Address").style = "field_label"
    ws.cell(row=16, column=3, value='=COUNTIF(VehStatus,"Due Soon")+COUNTIF(VehStatus,"Overdue")').style = "field_value"


# ===========================================================================
# 8 — Accommodation Manager
# ===========================================================================
def build_stays(wb):
    rows = [
        ("SpringHill Suites Springdale", "Hotel", "SH-88421", 175, dplus(30), dplus(32), "Pool · breakfast", "Yes"),
        ("Bryce Canyon Lodge", "Hotel", "BCL-2207", 210, dplus(32), dplus(33), "In-park · dining", "Yes"),
        ("Capitol Reef Cabin", "Cabin", "CRC-114", 160, dplus(33), dplus(34), "Kitchenette · quiet", "Yes"),
        ("Moab Downtown Inn", "Hotel", "MDI-5590", 145, dplus(34), dplus(35), "Central · parking", "No"),
        ("Grand Canyon Yavapai", "Hotel", "GCY-7781", 235, dplus(39), dplus(40), "Rim shuttle", "Yes"),
        ("Monument Valley Cabin", "Vacation Rental", "MVC-330", 190, dplus(37), dplus(38), "Sunrise views", "Yes"),
    ]
    ws, start, end = build_log(
        wb, "Stays", "🏨", "ACCOMMODATION MANAGER",
        "Where you'll sleep — hotels, cabins & rentals with confirmations and dates.",
        ["Name", "Type", "Confirmation", "Cost", "Check-In", "Check-Out", "Amenities", "Booked?"],
        rows, [26, 16, 15, 11, 13, 13, 22, 11],
        text_left={1, 3, 7}, dates={5, 6}, money={4}, reserved=20,
        validations=[("B", "StayTypeList"), ("H", "YesNoList")])
    nrange(wb, "StayType", "Stays", "B", start, end)
    nrange(wb, "StayBooked", "Stays", "H", start, end)
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"H{start}:H{end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 9 — Campground Planner
# ===========================================================================
def build_camping(wb):
    rows = [
        ("Watchman Campground (Zion)", "B-24", "Yes", "Electric · restrooms", "30A hookup", "Yes", "Reserve 6mo out"),
        ("North Campground (Bryce)", "12", "Yes", "Vault toilets · water", "None", "Yes", "First-come backup"),
        ("Devils Garden (Arches)", "45", "Yes", "Grills · water", "None", "Yes", "Book at recreation.gov"),
        ("Mather Campground (GC)", "88", "Yes", "Showers nearby", "None", "Yes", "Shuttle to rim"),
        ("Goblin Valley SP", "9", "No", "Showers · dark sky", "30A hookup", "Yes", "Weeknight has space"),
        ("Kodachrome Basin", "7", "No", "Full hookups", "50A hookup", "Yes", "Rest-day option"),
    ]
    ws, start, end = build_log(
        wb, "Camping", "🏕", "CAMPGROUND PLANNER",
        "Nights under the stars — sites, hookups, rules & pet-friendly notes.",
        ["Campground", "Site #", "Reserved?", "Amenities", "Hookups", "Pet OK?", "Notes"],
        rows, [28, 10, 13, 22, 15, 11, 24],
        text_left={1, 4, 7}, reserved=20,
        validations=[("C", "YesNoList"), ("F", "YesNoList")])
    nrange(wb, "CampReserved", "Camping", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"C{start}:C{end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 10 — Attraction Planner
# ===========================================================================
def build_attractions(wb):
    rows = [
        ("Angels Landing", "Zion NP", 0, "Sunrise best", "Permit req.", "No", "High", "Chains section — not for kids"),
        ("The Narrows", "Zion NP", 0, "Midday", "None", "Yes", "High", "Rent water shoes"),
        ("Bryce Amphitheater", "Bryce", 0, "Sunrise", "None", "Yes", "High", "Navajo/Queens loop"),
        ("Scenic Byway 12", "Utah", 0, "All day", "None", "Yes", "High", "Just drive it slow"),
        ("Delicate Arch", "Arches", 0, "Sunset", "Timed entry", "Yes", "High", "3mi round trip"),
        ("Mesa Verde Cliff Tour", "Mesa Verde", 20, "10a / 2p", "Ranger tour", "Yes", "Medium", "Book ahead"),
        ("Monument Valley Loop", "Navajo", 8, "Golden hour", "Per vehicle", "Yes", "High", "17mi dirt loop"),
        ("Grand Canyon Rim Trail", "Grand Canyon", 0, "Sunset", "None", "Yes", "High", "Hopi Point shuttle"),
        ("Antelope Canyon (opt.)", "Page, AZ", 65, "Guided", "Tour only", "Yes", "Medium", "Detour if time"),
        ("Hoover Dam", "NV/AZ", 15, "9a-5p", "Tour opt.", "Yes", "Low", "Last-day stop"),
    ]
    ws, start, end = build_log(
        wb, "Attractions", "🏞", "ATTRACTION PLANNER",
        "The bucket list — cost, hours, tickets, priority & who it's for.",
        ["Attraction", "City / Park", "Cost", "Best Time", "Tickets", "Family?", "Priority", "Notes"],
        rows, [24, 16, 10, 13, 14, 10, 11, 26],
        text_left={1, 8}, money={3}, reserved=30,
        validations=[("F", "YesNoList"), ("G", "PriorityList")])
    nrange(wb, "AttractionName", "Attractions", "A", start, end)
    nrange(wb, "AttractionCost", "Attractions", "C", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="equal", formula=['"High"'], fill=fill(MINT_BG)))


# ===========================================================================
# 11 — Restaurant & Food Planner
# ===========================================================================
def build_food(wb):
    rows = [
        ("Oscar's Cafe", "Restaurant", "Springdale", 60, "No", "Post-hike dinner", "—"),
        ("Bryce general store", "Grocery", "Bryce", 45, "No", "Trail snacks & water", "Restock"),
        ("Hell's Backbone Grill", "Restaurant", "Boulder, UT", 90, "Yes 7pm", "Farm-to-table splurge", "Veg options"),
        ("Moab City Market", "Grocery", "Moab", 70, "No", "3-day resupply", "Cooler ice"),
        ("Milt's Stop & Eat", "Restaurant", "Moab", 40, "No", "Classic burgers", "—"),
        ("Van cook — pasta night", "Picnic", "Campground", 15, "No", "Camp stove dinner", "Quick & cheap"),
        ("Cameron Trading Post", "Restaurant", "Cameron, AZ", 50, "No", "Navajo tacos", "Gas stop combo"),
        ("El Tovar Dining", "Restaurant", "Grand Canyon", 110, "Yes 6:30pm", "Rim-view dinner", "Reserve early"),
    ]
    ws, start, end = build_log(
        wb, "Food", "🍽", "RESTAURANT & FOOD PLANNER",
        "Fed on the road — restaurants, grocery stops, budget & dietary needs.",
        ["Place", "Type", "City", "Est. Cost", "Reservation", "Notes", "Dietary"],
        rows, [24, 13, 16, 12, 15, 24, 14],
        text_left={1, 6, 7}, money={4}, reserved=30)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="beginsWith", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 12 — Packing Command Center
# ===========================================================================
def build_packing(wb):
    rows = [
        ("Base layers & tees", "Clothing", "Both", 14, "Yes"), ("Hiking pants/shorts", "Clothing", "Both", 8, "Yes"),
        ("Rain shells", "Clothing", "Both", 2, "Yes"), ("Warm layers (parks are cold)", "Clothing", "Both", 4, "No"),
        ("Hiking boots", "Hiking Gear", "Both", 2, "Yes"), ("Trail runners", "Hiking Gear", "Both", 2, "Yes"),
        ("Daypacks", "Hiking Gear", "Both", 2, "Yes"), ("Trekking poles", "Hiking Gear", "Jake", 2, "No"),
        ("Tent (backup)", "Camping Gear", "Jake", 1, "Yes"), ("Sleeping bags", "Camping Gear", "Both", 2, "Yes"),
        ("Camp stove + fuel", "Cooking Equipment", "Nora", 1, "Yes"), ("Cookset & utensils", "Cooking Equipment", "Nora", 1, "Yes"),
        ("Cooler + ice packs", "Cooking Equipment", "Nora", 1, "Yes"), ("Headlamps", "Camping Gear", "Both", 2, "Yes"),
        ("Camp chairs", "Camping Gear", "Jake", 2, "No"), ("Power bank + cords", "Electronics", "Both", 1, "Yes"),
        ("Dash cam / GPS", "Electronics", "Jake", 1, "Yes"), ("Camera + lenses", "Electronics", "Nora", 1, "No"),
        ("Jumper cables", "Vehicle Supplies", "Jake", 1, "Yes"), ("Tire repair kit", "Vehicle Supplies", "Jake", 1, "Yes"),
        ("Extra motor oil & coolant", "Vehicle Supplies", "Jake", 1, "No"), ("First-aid kit", "Emergency Kit", "Nora", 1, "Yes"),
        ("Emergency water (gal)", "Emergency Kit", "Both", 5, "Yes"), ("Reflective triangles", "Emergency Kit", "Jake", 1, "No"),
        ("Paper maps (no signal!)", "Documents", "Both", 3, "Yes"), ("Park passes / America the Beautiful", "Documents", "Nora", 1, "Yes"),
        ("Insurance & registration", "Documents", "Jake", 1, "Yes"), ("Sunscreen & bug spray", "Emergency Kit", "Both", 3, "Yes"),
        ("Water filter", "Hiking Gear", "Jake", 1, "No"), ("Trash bags", "Camping Gear", "Both", 10, "Yes"),
    ]
    ws, start, end = build_log(
        wb, "Packing", "🎒", "PACKING COMMAND CENTER",
        "Everything for the road — assign items, track quantity & check them off.",
        ["Item", "Category", "For", "Qty", "Packed?"],
        rows, [30, 18, 12, 8, 12],
        text_left={1}, ints={4},
        validations=[("B", "PackCatList"), ("E", "YesNoList")], reserved=60)
    nrange(wb, "PackStatus", "Packing", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 13 — Emergency & Safety Center
# ===========================================================================
def build_emergency(wb):
    ws = wb.create_sheet("Emergency"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 26, 4, 24, 28, 2])
    luxe_header(ws, "G", "🚨  EMERGENCY & SAFETY CENTER",
                "Peace of mind on remote roads — contacts, coverage & safety, one tap away.")
    merge_set(ws, "B5:C5", "CONTACTS & COVERAGE", "section_gold"); ws.row_dimensions[5].height = 22
    contacts = [("Emergency", "911"), ("Roadside (AAA)", "800-555-0199"),
                ("Insurance 24/7", "800-555-0100"), ("Home Contact (Mom)", "555-0110"),
                ("Vehicle Mfr. Roadside", "800-555-0177"), ("Poison Control", "800-222-1222"),
                ("Trip Buddy (Sam)", "555-0142"), ("Bank (lost card)", "800-555-0155")]
    for i, (name, num) in enumerate(contacts):
        r = 6 + i
        ws.cell(row=r, column=2, value=name).style = "field_label"
        ws.cell(row=r, column=3, value=num).style = "field_value"
        ws.row_dimensions[r].height = 24
    merge_set(ws, "E5:F5", "SAFETY & MEDICAL", "section_gold"); ws.row_dimensions[5].height = 22
    safety = [("Blood types", "On phone note"), ("Allergies", "Nora — penicillin"),
              ("Medications", "In first-aid kit"), ("Nearest hospitals", "Save offline maps"),
              ("Emergency supplies", "Water, blanket, flares, kit"), ("Weather alerts", "NOAA app + radio"),
              ("Cell dead zones", "Download offline maps"), ("Check-in plan", "Text home each night")]
    for i, (name, info) in enumerate(safety):
        r = 6 + i
        ws.cell(row=r, column=5, value=name).style = "field_label"
        ws.cell(row=r, column=6, value=info).style = "td_left"
        ws.row_dimensions[r].height = 24


# ===========================================================================
# 14 — Road Trip Journal
# ===========================================================================
def build_journal(wb):
    rows = [
        (dplus(30), "First glimpse of Zion's walls", "The canyon narrows", "Van-cooked tacos", "Jake's parallel-park fail", "Bighorn sheep on the cliff", "Leave earlier for shuttles"),
        (dplus(31), "Made it up Angels Landing", "Chains & the drop-offs", "Zion Lodge bison burger", "Nora's victory dance", "California condor overhead", "Start hikes at dawn"),
        (dplus(32), "Bryce hoodoos at sunrise", "Amphitheater glowing orange", "Trail mix summit snack", "Got 'lost' on Queens loop", "Pronghorn near the road", "Layers — it was 34°F"),
        (dplus(33), "Byway 12 is unreal", "Cliff-edge switchbacks", "Hell's Backbone farm dinner", "Sang the whole soundtrack", "Deer crossing at dusk", "Fuel up when you can"),
    ]
    ws, start, end = build_log(
        wb, "Journal", "📔", "ROAD TRIP JOURNAL",
        "One line a day keeps the adventure forever — write it before the campfire dies.",
        ["Date", "Best Memory", "Scenic Highlight", "Favorite Meal", "Funny Moment", "Wildlife", "Lesson Learned"],
        rows, [13, 24, 22, 22, 24, 20, 22],
        text_left={2, 3, 4, 5, 6, 7}, dates={1}, reserved=20)


# ===========================================================================
# 15 — Photo & Memory Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  PHOTO & MEMORY GALLERY",
                "Keep the views — drop in photos and caption the stops you'll never forget.")
    merge_set(ws, "B5:D5", "HOW TO ADD YOUR PHOTOS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Excel: Insert ▸ Pictures ▸ Place in Cell (or drag a photo) into any framed box below. "
                      "Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"paste-link-here\"). "
                      "Caption each memory underneath — scenic views, campsites, wildlife, landmarks, the rig & sunsets.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Scenic View", "Campsite", "Wildlife", "Landmark", "The Rig", "Sunset"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add photo)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Date · Location · caption…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 16 — National Park Checklist
# ===========================================================================
def build_parks(wb):
    rows = [
        ("Zion National Park", dplus(30), 35, "Angels Landing, Narrows", "Yes", "No", "Shuttle mandatory"),
        ("Bryce Canyon NP", dplus(32), 35, "Navajo/Queens Loop", "Yes", "No", "Sunrise Point first"),
        ("Capitol Reef NP", dplus(33), 20, "Hickman Bridge", "Yes", "No", "Free scenic drive"),
        ("Arches NP", dplus(35), 30, "Delicate Arch, Windows", "Yes", "No", "Timed entry ticket"),
        ("Canyonlands NP", dplus(35), 30, "Mesa Arch, Grand View", "Yes", "No", "Island in the Sky"),
        ("Mesa Verde NP", dplus(36), 30, "Cliff Palace tour", "Yes", "No", "Ranger tour ticket"),
        ("Grand Canyon NP", dplus(39), 35, "Rim Trail, Bright Angel", "Yes", "No", "South Rim"),
    ]
    ws, start, end = build_log(
        wb, "Parks", "🏔", "NATIONAL PARK CHECKLIST",
        "Collect the parks — fees, must-do trails & visitor centers, checked off as you go.",
        ["Park", "Visit Date", "Entrance Fee", "Must-Do Trails", "Visitor Center", "Completed?", "Notes"],
        rows, [24, 13, 13, 24, 14, 12, 22],
        text_left={1, 4, 7}, dates={2}, money={3}, reserved=16,
        validations=[("E", "YesNoList"), ("F", "YesNoList")])
    nrange(wb, "ParkName", "Parks", "A", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.cell(row=end + 2, column=1, value="Tip: an America the Beautiful pass ($80) covers all 7 parks — pays for itself.").font = Font(italic=True, color=ACCENT)


# ===========================================================================
# 17 — Travel Rewards Tracker
# ===========================================================================
def build_rewards(wb):
    rows = [
        ("Marriott Bonvoy", "Hotel", 52000, dplus(500), "2 free park-town nights", "Pooled account"),
        ("GetUpside (fuel)", "Fuel", 34, dplus(200), "Cash back per gallon", "Scan every receipt"),
        ("Chase Sapphire", "Credit Card", 71000, "None", "Transfer to travel", "3x on travel & gas"),
        ("Costco Gas Rewards", "Fuel", 62, "None", "Cheapest fill-ups", "Members only"),
        ("America the Beautiful", "Parks", 1, dplus(365), "All-park entry pass", "$80 — covers this trip"),
        ("Wyndham Rewards", "Hotel", 18000, dplus(400), "1 free night", "Backup for overflow"),
    ]
    ws, start, end = build_log(
        wb, "Rewards", "🎁", "TRAVEL REWARDS TRACKER",
        "Travel for less — points, fuel perks & park passes with expirations & goals.",
        ["Program", "Type", "Balance", "Expires", "Redemption Goal", "Notes"],
        rows, [22, 14, 14, 14, 26, 24],
        text_left={5, 6}, ints={3}, reserved=20)
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=72000, color=GOLD_LT, showValue=True))


# ===========================================================================
# 18 — Road Trip Analytics
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 3, 22, 12, 12, 2])
    luxe_header(ws, "G", "📊  ROAD TRIP ANALYTICS",
                "The whole journey by the numbers — plus a live Trip Readiness Score.")
    # readiness dims
    merge_set(ws, "B5:C5", "TRIP READINESS", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score"], start_col=2)
    dims = [
        ("Fund vs budget", "=IFERROR(MIN(SavedTotal/TripBudget,1),0)"),
        ("Packing complete", '=IFERROR(COUNTIF(PackStatus,"Yes")/COUNTA(PackStatus),0)'),
        ("Vehicle ready", '=IFERROR(COUNTIF(VehStatus,"OK")/COUNTA(VehStatus),0)'),
        ("Bookings confirmed", '=IFERROR((COUNTIF(StayBooked,"Yes")+COUNTIF(CampReserved,"Yes"))/(COUNTA(StayBooked)+COUNTA(CampReserved)),0)'),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "ReadinessRange", "Analytics", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    # trip snapshot + trip totals (define distance/hours)
    merge_set(ws, "E5:F5", "TRIP SNAPSHOT", "section")
    snap = [("Total distance (mi)", "=SUM(DailyMiles)", "#,##0", "TotalDistance"),
            ("Driving hours", "=SUM(DriveTime)", "0.0", "DriveHours"),
            ("Total budget", "=TripBudget", '"$"#,##0', None), ("Spent so far", "=SpentTotal", '"$"#,##0', None),
            ("Cost / traveler", "=IFERROR(TripBudget/TravelerCount,0)", '"$"#,##0', None),
            ("Attractions planned", "=COUNTA(AttractionName)", "#,##0", None),
            ("Parks on the list", "=COUNTA(ParkName)", "#,##0", None),
            ("Readiness score", "=IFERROR(AVERAGE(ReadinessRange),0)", "0%", None)]
    for i, (lab, fml, fmt, nm) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if nm:
            cell_name(wb, nm, "Analytics", f"$F${r}")
        if lab == "Readiness score":
            c.fill = fill(MINT_BG)
    # daily mileage mini-table for chart
    merge_set(ws, "B14:C14", "DAILY MILEAGE", "section"); ws.row_dimensions[14].height = 20
    ws.cell(row=15, column=2, value="Day").style = "th"; ws.cell(row=15, column=3, value="Miles").style = "th"
    daymiles = [("Day 1", 165), ("Day 4", 120), ("Day 5", 150), ("Day 7", 150), ("Day 9", 180), ("Day 12", 280)]
    for i, (d, m) in enumerate(daymiles):
        r = 16 + i
        ws.cell(row=r, column=2, value=d).style = "td_left"
        c = ws.cell(row=r, column=3, value=m); c.style = "td"; c.number_format = "#,##0"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "DayLabel", "Analytics", "$B$16:$B$21")
    cell_name(wb, "DayMiles", "Analytics", "$C$16:$C$21")


# ===========================================================================
# 1 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🚗  ROAD TRIP COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Route, budget, fuel, vehicle & memories — your whole road trip, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("DAYS TO DEPARTURE", "=MAX(TripStart-TODAY(),0)", "num"),
        ("TOTAL DISTANCE", "=TotalDistance", "num"),
        ("DRIVING HOURS", "=DriveHours", "dec"),
        ("TOTAL BUDGET", "=TripBudget", "money"),
        ("BUDGET REMAINING", "=TripBudget-SpentTotal", "money"),
        ("FUEL ESTIMATE", "=FuelEstimate", "money"),
    ]
    row2 = [
        ("STOPS PLANNED", "=COUNTA(AttractionName)", "num"),
        ("CAMPGROUNDS BOOKED", '=COUNTIF(CampReserved,"Yes")', "num"),
        ("HOTELS BOOKED", '=COUNTIFS(StayType,"Hotel",StayBooked,"Yes")', "num"),
        ("PACKING PROGRESS", '=IFERROR(COUNTIF(PackStatus,"Yes")/COUNTA(PackStatus),0)', "pct"),
        ("VEHICLE READINESS", '=IFERROR(COUNTIF(VehStatus,"OK")/COUNTA(VehStatus),0)', "pct"),
        ("TRIP READINESS", "=IFERROR(AVERAGE(ReadinessRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "BUDGET & MILEAGE", "section_gold")
    # budget donut
    d1 = DoughnutChart(); d1.title = "Budget by Category"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Budget"], min_col=2, min_row=4, max_row=16), titles_from_data=True)
    d1.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=16)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")
    # daily mileage bar
    cb = BarChart(); cb.type = "col"; cb.title = "Daily Mileage"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=15, max_row=21), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=16, max_row=21)); cb.legend = None
    ws.add_chart(cb, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "READINESS & SPENDING", "section_gold")
    # readiness bar
    rb = BarChart(); rb.type = "bar"; rb.title = "Trip Readiness"; rb.height = 8.2; rb.width = 11.5
    rb.add_data(Reference(wb["Analytics"], min_col=3, min_row=6, max_row=10), titles_from_data=True)
    rb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=7, max_row=10)); rb.legend = None
    ws.add_chart(rb, "B30")
    # planned vs actual
    bc = BarChart(); bc.type = "col"; bc.title = "Planned vs Actual"; bc.height = 8.2; bc.width = 11.5
    bc.add_data(Reference(wb["Budget"], min_col=2, min_row=4, max_col=3, max_row=16), titles_from_data=True)
    bc.set_categories(Reference(wb["Budget"], min_col=1, min_row=5, max_row=16))
    ws.add_chart(bc, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Road Trip Command Center™ — from the first pin to the last sunset, all in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_route(wb)
    build_itinerary(wb); build_budget(wb); build_fuel(wb); build_vehicle(wb)
    build_stays(wb); build_camping(wb); build_attractions(wb); build_food(wb)
    build_packing(wb); build_emergency(wb); build_journal(wb); build_gallery(wb)
    build_parks(wb); build_rewards(wb); build_analytics(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Trip Profile", "Route", "Itinerary", "Budget", "Fuel", "Vehicle",
             "Stays", "Camping", "Attractions", "Food", "Packing", "Emergency", "Journal", "Gallery",
             "Parks", "Rewards", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Road_Trip_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
