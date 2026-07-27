"""Build Daycare & Childcare Provider Command Center™ — The Provider's Operating System.

14 tabs · a premium home-daycare & childcare operating system in Google Sheets & Excel.
Dashboard, a rate & enrollment engine (weekly rate → net per child → how many children it
takes just to cover the house), children & families, tuition & payments with who's behind,
attendance, costs, the CACFP food program, ratios & schedule, compliance files, supplies,
a tax set-aside and a monthly summary — one dashboard. You're not babysitting. You're
running a business.

Run: python3 build_xlsx.py   ->  ../Daycare_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
SCHEDULE = ["Full time", "Part time", "Before/after school", "Drop-in", "Summer only"]
PAYSTATUS = ["Paid", "Due", "Late", "Partial", "Subsidy pending"]
AGEGROUP = ["Infant", "Toddler", "Preschool", "School age"]
EXPCAT = ["Payroll", "Insurance", "Home & utilities", "Food", "Supplies", "Training", "Vehicle", "Other"]

# --- Rate & enrollment engine ---
LICENSED_CAPACITY = 12
ENROLLED = 11
WEEKLY_RATE = 245.00
WEEKS_PER_MONTH = 4.33
FOOD_PER_CHILD = 135.00
SUPPLIES_PER_CHILD = 22.00
OPEN_HOURS_WEEK = 50
CAREGIVERS = 2
LATE_FEES = 180
WAITLIST = 3
TAX_SET_ASIDE = 6400
TAX_RESERVE_GOAL = 16000
TAX_RATE = 0.25

# --- Goals ---
OCCUPANCY_GOAL = 0.90
MARGIN_GOAL = 0.25
COVER_GOAL = 1.6
RATIO_MAX = 6
HOURLY_GOAL = 18

# Fixed costs — what the program costs before a single child walks in: (line, category, monthly)
FIXED_LINES = [
    ("Assistant wages", "Payroll", 2400), ("Payroll taxes on assistant", "Payroll", 220),
    ("Insurance & licensing", "Insurance", 185), ("Home business use & utilities share", "Home & utilities", 640),
    ("Curriculum, toys & billing software", "Supplies", 95), ("Cleaning, safety & first aid", "Supplies", 140),
    ("Vehicle & field trips", "Vehicle", 85), ("Training, CPR & renewals", "Training", 60),
]

# Children & families: (child, age group, schedule, weekly rate, family, start)
CHILDREN = [
    ("Mia R.", "Preschool", "Full time", 245, "Reyes", "01/08"),
    ("Noah R.", "Toddler", "Full time", 245, "Reyes", "01/08"),
    ("Ava T.", "Preschool", "Full time", 245, "Tran", "02/12"),
    ("Liam B.", "Toddler", "Full time", 245, "Brooks", "03/04"),
    ("Zoe K.", "Infant", "Full time", 285, "Kaur", "03/18"),
    ("Eli M.", "Preschool", "Full time", 245, "Moreno", "04/01"),
    ("Ruby C.", "Toddler", "Full time", 245, "Chen", "04/22"),
    ("Isaac D.", "School age", "Before/after school", 145, "Davis", "05/06"),
    ("Layla P.", "Preschool", "Full time", 245, "Patel", "05/20"),
    ("Owen H.", "Toddler", "Full time", 245, "Hayes", "06/10"),
    ("Nora W.", "Preschool", "Part time", 165, "Walsh", "06/24"),
]

# Tuition & payments: (child, due, paid, status)
PAYMENTS = [
    ("Mia R.", 1060.85, 1060.85, "Paid"), ("Noah R.", 1060.85, 1060.85, "Paid"),
    ("Ava T.", 1060.85, 1060.85, "Paid"), ("Liam B.", 1060.85, 1060.85, "Paid"),
    ("Zoe K.", 1234.05, 1234.05, "Paid"), ("Eli M.", 1060.85, 1060.85, "Paid"),
    ("Ruby C.", 1060.85, 700.00, "Partial"), ("Isaac D.", 627.85, 627.85, "Paid"),
    ("Layla P.", 1060.85, 1060.85, "Paid"), ("Owen H.", 1060.85, 0.00, "Late"),
    ("Nora W.", 714.45, 714.45, "Subsidy pending"),
]

# Attendance: (child, days present, days absent, late pickups)
ATTENDANCE = [
    ("Mia R.", 21, 0, 0), ("Noah R.", 20, 1, 1), ("Ava T.", 21, 0, 0), ("Liam B.", 19, 2, 0),
    ("Zoe K.", 21, 0, 2), ("Eli M.", 20, 1, 0), ("Ruby C.", 21, 0, 1), ("Isaac D.", 21, 0, 0),
    ("Layla P.", 18, 3, 0), ("Owen H.", 21, 0, 3), ("Nora W.", 12, 0, 0),
]

# CACFP food program: (meal, meals claimed, Tier I rate)
MEALS = [("Breakfast", 210, 1.66), ("Lunch", 231, 3.13), ("Afternoon snack", 231, 0.97)]

# Ratios & schedule: (age group, children, state ratio (children per adult), max group size)
RATIOS = [
    ("Infant", 1, 4, 8), ("Toddler", 4, 6, 12), ("Preschool", 5, 10, 20), ("School age", 1, 15, 30),
]

# Compliance files: (child, immunizations, emergency form, signed contract, allergy plan)
COMPLIANCE = [
    ("Mia R.", "Yes", "Yes", "Yes", "Yes"), ("Noah R.", "Yes", "Yes", "Yes", "Yes"),
    ("Ava T.", "Yes", "Yes", "Yes", "No"), ("Liam B.", "Yes", "Yes", "Yes", "Yes"),
    ("Zoe K.", "Yes", "Yes", "Yes", "Yes"), ("Eli M.", "No", "Yes", "Yes", "Yes"),
    ("Ruby C.", "Yes", "Yes", "No", "Yes"), ("Isaac D.", "Yes", "Yes", "Yes", "Yes"),
    ("Layla P.", "Yes", "No", "Yes", "Yes"), ("Owen H.", "Yes", "Yes", "Yes", "No"),
    ("Nora W.", "Yes", "Yes", "Yes", "Yes"),
]

# Supplies: (item, on hand, reorder at, unit cost)
SUPPLIES = [
    ("Diapers & wipes (case)", 3, 2, 42.00), ("Art & craft supplies", 5, 3, 28.00),
    ("Disinfectant & cleaning", 4, 3, 19.00), ("Paper goods & cups", 6, 4, 24.00),
    ("First aid restock", 2, 2, 35.00), ("Nap mats & bedding", 12, 10, 18.00),
    ("Curriculum printables & books", 8, 5, 15.00),
]

# Tax set-aside: (quarter, income, set aside, due)
TAXES = [
    ("Q1", 21400, 1700, 3900), ("Q2", 22800, 2100, 4000),
    ("Q3", 23600, 2600, 4050), ("Q4", 24200, 0, 4050),
]

# Monthly summary: (month, revenue)
MONTHS = [("Feb", 10940), ("Mar", 11480), ("Apr", 12060), ("May", 12510), ("Jun", 12880), ("Jul", 13145)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "money3": '"$"#,##0.000', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 20, 3] + [20] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your rates & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR RATES & GOALS", "section")
    controls = [
        ("Program", "Little Acorns Home Childcare", None, "Program"),
        ("Provider", "Dana", None, "Provider"),
        ("Licensed capacity", LICENSED_CAPACITY, "0", "Capacity"),
        ("Full-time weekly rate", WEEKLY_RATE, '"$"#,##0.00', "WeeklyRate"),
        ("Weeks per month", WEEKS_PER_MONTH, "0.00", "WeeksMonth"),
        ("Hours open per week", OPEN_HOURS_WEEK, "0", "OpenHours"),
        ("Caregivers on staff (incl. you)", CAREGIVERS, "0", "Caregivers"),
        ("Tax set-aside rate", TAX_RATE, "0%", "TaxRate"),
        ("Occupancy goal", OCCUPANCY_GOAL, "0%", "OccupancyGoal"),
        ("Margin goal", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Break-even cover goal (×)", COVER_GOAL, "0.0", "CoverGoal"),
        ("Your hourly goal", HOURLY_GOAL, '"$"#,##0', "HourlyGoal"),
        ("Max children per caregiver", RATIO_MAX, "0", "RatioMax"),
        ("Tax reserve goal (year)", TAX_RESERVE_GOAL, '"$"#,##0', "TaxReserveGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Schedule", SCHEDULE, "ScheduleList"), ("F", "Payment status", PAYSTATUS, "PayStatusList"),
             ("G", "Age group", AGEGROUP, "AgeGroupList"), ("H", "Expense category", EXPCAT, "ExpCatList"),
             ("I", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  \U0001f33f  DAYCARE & CHILDCARE PROVIDER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  You're not babysitting. You're running a business.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE PROGRAM, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Home childcare providers are some of the hardest-working business owners there are, and almost "
                      "none of them know their real numbers. What does one child actually net you after food and "
                      "supplies? How many children does it take just to cover the house, the insurance and your "
                      "assistant? What are you earning per hour you're actually open — and what's left after tax? "
                      "This workbook answers all of it, then runs enrollment, tuition and who's behind, attendance, "
                      "the food program, ratios, compliance files and a tax set-aside — all in ONE premium Google "
                      "Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — capacity, your weekly rate and your goals.",
             "2.  Costs & Expenses: list what the program costs every month.",
             "3.  Rate & Enrollment: read your net per child and break-even.",
             "4.  Add your children, their schedules and their rates.",
             "5.  Track tuition, attendance, the food program and files.",
             "6.  Check the Dashboard: your pay, per hour, and a Care Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional program (Little Acorns Home Childcare, provider Dana) is included so you "
               "can see how it all connects — just type over it with your own children and rates. Two things this "
               "workbook insists on: your PAY is what's left after every cost, not the tuition that came in; and a "
               "quarter of it is not yours — it belongs to the tax you haven't paid yet. Ratios, group sizes, "
               "licensing rules and CACFP reimbursement tiers vary by state, so enter your own in Settings. Twelve "
               "matching printable pages (enrollment form, daily sheet, attendance, meal count & more) are included. "
               "This is a business & organizing tool, not financial, tax, legal or licensing advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "The tuition that came in is not your pay. What's left after everything is your pay.", "section_gold")


# ===========================================================================
def build_costs(wb):
    ws = wb.create_sheet("Costs & Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 42, 20, 16, 16, 2])
    luxe_header(ws, "E", "\U0001f4b8  COSTS & EXPENSES",
                "What the program costs every month before a single child walks in.")
    table_headers(ws, 4, ["Line", "Category", "Monthly", "Yearly"], start_col=2)
    fs_ = L0
    for i, (lab, cat, amt) in enumerate(FIXED_LINES):
        r = fs_ + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        c = ws.cell(row=r, column=4, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        cy = ws.cell(row=r, column=5, value=f"=D{r}*12"); cy.style = "td"; cy.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 6):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    fe = fs_ + len(FIXED_LINES) - 1
    nrange(wb, "FixedLines", "Costs & Expenses", "D", fs_, fe)
    add_dv(ws, f"C{fs_}:C{fe}", "ExpCatList")
    tot = fe + 1
    ws.cell(row=tot, column=2, value="= FIXED COSTS / MONTH").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=4, value="=SUM(FixedLines)"); ct.style = "td"
    ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "FixedTotal", "Costs & Expenses", f"$D${tot}")
    cty = ws.cell(row=tot, column=5, value="=FixedTotal*12"); cty.style = "td"
    cty.font = Font(bold=True, size=12, color=PRIMARY); cty.fill = fill(MINT_BG); cty.number_format = '"$"#,##0'

    r = tot + 2
    ws.cell(row=r, column=2, value="COSTS THAT FOLLOW EACH CHILD").style = "section_gold"
    ws.cell(row=r + 1, column=2, value="Food per child / month").style = "field_label"
    cf = ws.cell(row=r + 1, column=4, value=FOOD_PER_CHILD); cf.style = "input"; cf.number_format = '"$"#,##0.00'
    cell_name(wb, "FoodPerChild", "Costs & Expenses", f"$D${r+1}")
    ws.cell(row=r + 2, column=2, value="Supplies & activities per child / month").style = "field_label"
    cs = ws.cell(row=r + 2, column=4, value=SUPPLIES_PER_CHILD); cs.style = "input"; cs.number_format = '"$"#,##0.00'
    cell_name(wb, "SuppliesPerChild", "Costs & Expenses", f"$D${r+2}")
    ws.cell(row=r + 3, column=2, value="= COST PER CHILD / MONTH").style = "th"
    cc2 = ws.cell(row=r + 3, column=4, value="=FoodPerChild+SuppliesPerChild"); cc2.style = "td"
    cc2.font = Font(bold=True, size=13, color=PRIMARY); cc2.fill = fill(SURFACE); cc2.number_format = '"$"#,##0.00'
    cell_name(wb, "CostPerChild", "Costs & Expenses", f"$D${r+3}")
    ws.cell(row=r + 4, column=2, value="× children enrolled = VARIABLE COSTS").style = "th"
    cv = ws.cell(row=r + 4, column=4, value="=CostPerChild*Enrolled"); cv.style = "td"
    cv.font = Font(bold=True, size=12, color=PRIMARY); cv.fill = fill(MINT_BG); cv.number_format = '"$"#,##0.00'
    cell_name(wb, "VariableCosts", "Costs & Expenses", f"$D${r+4}")
    ws.cell(row=r + 6, column=2, value="= TOTAL COSTS THIS MONTH").style = "th"
    ctc = ws.cell(row=r + 6, column=4, value="=FixedTotal+VariableCosts"); ctc.style = "td"
    ctc.font = Font(bold=True, size=14, color=PRIMARY); ctc.fill = fill(SURFACE); ctc.number_format = '"$"#,##0'
    cell_name(wb, "TotalCosts", "Costs & Expenses", f"$D${r+6}")
    ws.cell(row=r + 8, column=2, value="Every line above runs whether you're at eleven children or six.").style = "section_gold"


def build_rate(wb):
    ws = wb.create_sheet("Rate & Enrollment"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 46, 18, 2])
    luxe_header(ws, "C", "\U0001f9ee  RATE & ENROLLMENT — THE ENGINE",
                "What one child actually nets you — and how many it takes just to cover the house.")
    ws.cell(row=5, column=2, value="ONE CHILD").style = "section_gold"
    ws.cell(row=6, column=2, value="Full-time weekly rate").style = "field_label"
    cw = ws.cell(row=6, column=3, value="=WeeklyRate"); cw.style = "field_value"; cw.number_format = '"$"#,##0.00'
    ws.cell(row=7, column=2, value="× weeks per month").style = "field_label"
    cwm = ws.cell(row=7, column=3, value="=WeeksMonth"); cwm.style = "field_value"; cwm.number_format = "0.00"
    ws.cell(row=8, column=2, value="= TUITION PER CHILD / MONTH").style = "th"
    cg = ws.cell(row=8, column=3, value="=WeeklyRate*WeeksMonth"); cg.style = "td"
    cg.font = Font(bold=True, size=13, color=PRIMARY); cg.fill = fill(SURFACE); cg.number_format = '"$"#,##0.00'
    cell_name(wb, "TuitionPerChild", "Rate & Enrollment", "$C$8")
    ws.cell(row=9, column=2, value="− Food, supplies & activities for that child").style = "field_label"
    cc = ws.cell(row=9, column=3, value="=CostPerChild"); cc.style = "field_value"
    cc.number_format = '"$"#,##0.00'; cc.fill = fill(WARN_BG)
    ws.cell(row=10, column=2, value="= NET PER CHILD / MONTH").style = "th"
    cn = ws.cell(row=10, column=3, value="=TuitionPerChild-CostPerChild"); cn.style = "td"
    cn.font = Font(bold=True, size=16, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0.00'
    cell_name(wb, "NetPerChild", "Rate & Enrollment", "$C$10")

    ws.cell(row=12, column=2, value="HOW MANY CHILDREN JUST TO COVER THE HOUSE").style = "section_gold"
    ws.cell(row=13, column=2, value="Fixed costs / month").style = "field_label"
    cfx = ws.cell(row=13, column=3, value="=FixedTotal"); cfx.style = "field_value"; cfx.number_format = '"$"#,##0'
    ws.cell(row=14, column=2, value="÷ net per child").style = "field_label"
    cnn = ws.cell(row=14, column=3, value="=NetPerChild"); cnn.style = "field_value"; cnn.number_format = '"$"#,##0.00'
    ws.cell(row=15, column=2, value="= BREAK-EVEN CHILDREN").style = "th"
    cb = ws.cell(row=15, column=3, value="=IFERROR(ROUNDUP(FixedTotal/NetPerChild,0),0)"); cb.style = "td"
    cb.font = Font(bold=True, size=16, color=PRIMARY); cb.fill = fill(SURFACE); cb.number_format = "0"
    cell_name(wb, "BreakEven", "Rate & Enrollment", "$C$15")
    ws.cell(row=16, column=2, value="You have this many enrolled").style = "field_label"
    ce = ws.cell(row=16, column=3, value=ENROLLED); ce.style = "input"; ce.number_format = "0"
    cell_name(wb, "Enrolled", "Rate & Enrollment", "$C$16")
    ws.cell(row=17, column=2, value="= YOU COVER THE HOUSE THIS MANY TIMES OVER").style = "th"
    ccv = ws.cell(row=17, column=3, value="=IFERROR(Enrolled/BreakEven,0)"); ccv.style = "td"
    ccv.font = Font(bold=True, size=13, color=PRIMARY); ccv.fill = fill(MINT_BG); ccv.number_format = '0.00"×"'
    cell_name(wb, "CoverRatio", "Rate & Enrollment", "$C$17")
    ws.cell(row=18, column=2, value="= OCCUPANCY (enrolled ÷ licensed capacity)").style = "th"
    co = ws.cell(row=18, column=3, value="=IFERROR(Enrolled/Capacity,0)"); co.style = "td"
    co.font = Font(bold=True, size=13, color=PRIMARY); co.fill = fill(MINT_BG); co.number_format = "0.0%"
    cell_name(wb, "Occupancy", "Rate & Enrollment", "$C$18")
    ws.cell(row=19, column=2, value="Children on the waitlist").style = "field_label"
    cwl = ws.cell(row=19, column=3, value=WAITLIST); cwl.style = "input"; cwl.number_format = "0"
    cell_name(wb, "Waitlist", "Rate & Enrollment", "$C$19")

    ws.cell(row=21, column=2, value="WHAT ONE EMPTY SPOT COSTS YOU").style = "section_gold"
    ws.cell(row=22, column=2, value="Open spots right now").style = "field_label"
    cop = ws.cell(row=22, column=3, value="=Capacity-Enrolled"); cop.style = "field_value"; cop.number_format = "0"
    ws.cell(row=23, column=2, value="× net per child = lost every month").style = "field_label"
    cl = ws.cell(row=23, column=3, value="=(Capacity-Enrolled)*NetPerChild"); cl.style = "field_value"
    cl.number_format = '"$"#,##0'; cl.fill = fill(RED_BG)
    ws.cell(row=24, column=2, value="…which is this much a year").style = "field_label"
    cly = ws.cell(row=24, column=3, value="=(Capacity-Enrolled)*NetPerChild*12"); cly.style = "field_value"
    cly.number_format = '"$"#,##0'; cly.fill = fill(RED_BG)
    ws.cell(row=26, column=2, value="An empty spot is not a light day. It is the most expensive thing in the program.").style = "section_gold"


def build_children(wb):
    ws = wb.create_sheet("Children & Families"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 18, 15, 21, 15, 16, 12, 16, 2])
    luxe_header(ws, "H", "\U0001f9f8  CHILDREN & FAMILIES",
                "Everyone in the program — their schedule, their rate and what they're worth a month.")
    table_headers(ws, 4, ["Child", "Age group", "Schedule", "Weekly rate", "Family", "Start", "Monthly"], start_col=2)
    start = L0
    for i, (nm, ag, sch, rate, fam, st) in enumerate(CHILDREN):
        r = start + i
        ws.cell(row=r, column=2, value=nm).style = "td_left"
        ws.cell(row=r, column=3, value=ag).style = "td"
        ws.cell(row=r, column=4, value=sch).style = "td"
        cr = ws.cell(row=r, column=5, value=rate); cr.style = "input"; cr.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=6, value=fam).style = "td"
        ws.cell(row=r, column=7, value=st).style = "td"
        cm = ws.cell(row=r, column=8, value=f"=E{r}*WeeksMonth"); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CHILDREN) - 1
    nrange(wb, "ChildName", "Children & Families", "B", start, end)
    nrange(wb, "ChildRate", "Children & Families", "E", start, end)
    nrange(wb, "ChildMonthly", "Children & Families", "H", start, end)
    add_dv(ws, f"C{start}:C{end}", "AgeGroupList"); add_dv(ws, f"D{start}:D{end}", "ScheduleList")
    tot = end + 1
    ws.cell(row=tot, column=2, value="ENROLLED").style = "th"
    for c in (3, 4, 6, 7):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cn = ws.cell(row=tot, column=5, value="=COUNTA(ChildName)"); cn.style = "td"
    cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = "0"
    cm = ws.cell(row=tot, column=8, value="=SUM(ChildMonthly)"); cm.style = "td"
    cm.font = Font(bold=True, size=12, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0'
    cell_name(wb, "TuitionBilled", "Children & Families", f"$H${tot}")
    ws.cell(row=tot + 2, column=2, value="Rates vary by age and schedule — infants cost more to staff, so they cost more to enrol.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_payments(wb):
    ws = wb.create_sheet("Tuition & Payments"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 17, 17, 20, 17, 2])
    luxe_header(ws, "F", "\U0001f4b3  TUITION & PAYMENTS",
                "Who has paid, who is behind, and exactly how much is sitting out there.")
    table_headers(ws, 4, ["Child", "Due", "Paid", "Status", "Balance"], start_col=2)
    start = L0
    for i, (nm, due, paid, status) in enumerate(PAYMENTS):
        r = start + i
        ws.cell(row=r, column=2, value=nm).style = "td_left"
        cd = ws.cell(row=r, column=3, value=due); cd.style = "input"; cd.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=4, value=paid); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=5, value=status).style = "td"
        cb = ws.cell(row=r, column=6, value=f"=C{r}-D{r}"); cb.style = "td"; cb.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PAYMENTS) - 1
    nrange(wb, "PayDue", "Tuition & Payments", "C", start, end)
    nrange(wb, "PayPaid", "Tuition & Payments", "D", start, end)
    nrange(wb, "PayStatusCol", "Tuition & Payments", "E", start, end)
    nrange(wb, "PayBalance", "Tuition & Payments", "F", start, end)
    add_dv(ws, f"E{start}:E{end}", "PayStatusList")
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="greaterThan", formula=["0"],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL FAMILIES").style = "th"
    cd = ws.cell(row=tot, column=3, value="=SUM(PayDue)"); cd.style = "td"
    cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = '"$"#,##0'
    cell_name(wb, "TuitionDue", "Tuition & Payments", f"$C${tot}")
    cp = ws.cell(row=tot, column=4, value="=SUM(PayPaid)"); cp.style = "td"
    cp.font = Font(bold=True, size=12, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = '"$"#,##0'
    cell_name(wb, "TuitionCollected", "Tuition & Payments", f"$D${tot}")
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cb = ws.cell(row=tot, column=6, value="=SUM(PayBalance)"); cb.style = "td"
    cb.font = Font(bold=True, size=12, color=DANGER); cb.fill = fill(RED_BG); cb.number_format = '"$"#,##0'
    cell_name(wb, "Outstanding", "Tuition & Payments", f"$F${tot}")
    ws.cell(row=tot + 2, column=2, value="Collection rate").style = "field_label"
    ccr = ws.cell(row=tot + 2, column=4, value="=IFERROR(TuitionCollected/TuitionDue,0)"); ccr.style = "field_value"
    ccr.number_format = "0.0%"; ccr.fill = fill(MINT_BG)
    ws.cell(row=tot + 3, column=2, value="Families behind").style = "field_label"
    cfb = ws.cell(row=tot + 3, column=4, value='=COUNTIF(PayStatusCol,"Late")+COUNTIF(PayStatusCol,"Partial")')
    cfb.style = "field_value"; cfb.number_format = "0"
    ws.cell(row=tot + 5, column=2, value="Late fees are in your contract for a reason. Charging them is not unkind — it is the business.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_attendance(wb):
    ws = wb.create_sheet("Attendance"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 16, 16, 17, 16, 2])
    luxe_header(ws, "F", "\U0001f4c5  ATTENDANCE",
                "Days present, days absent and late pickups — the record that protects you.")
    table_headers(ws, 4, ["Child", "Days present", "Days absent", "Late pickups", "Attendance %"], start_col=2)
    start = L0
    for i, (nm, pres, absent, late) in enumerate(ATTENDANCE):
        r = start + i
        ws.cell(row=r, column=2, value=nm).style = "td_left"
        cp = ws.cell(row=r, column=3, value=pres); cp.style = "input"; cp.number_format = "0"
        ca = ws.cell(row=r, column=4, value=absent); ca.style = "input"; ca.number_format = "0"
        cl = ws.cell(row=r, column=5, value=late); cl.style = "input"; cl.number_format = "0"
        cpc = ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/(C{r}+D{r}),0)"); cpc.style = "td"; cpc.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(ATTENDANCE) - 1
    nrange(wb, "AttPresent", "Attendance", "C", start, end)
    nrange(wb, "AttAbsent", "Attendance", "D", start, end)
    nrange(wb, "AttLate", "Attendance", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="greaterThanOrEqual", formula=["2"],
                                                                fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL CHILDREN").style = "th"
    cp = ws.cell(row=tot, column=3, value="=SUM(AttPresent)"); cp.style = "td"
    cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(SURFACE); cp.number_format = "#,##0"
    cell_name(wb, "ChildDays", "Attendance", f"$C${tot}")
    ca = ws.cell(row=tot, column=4, value="=SUM(AttAbsent)"); ca.style = "td"
    ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = "#,##0"
    cl = ws.cell(row=tot, column=5, value="=SUM(AttLate)"); cl.style = "td"
    cl.font = Font(bold=True, size=12, color=ACCENT); cl.fill = fill(WARN_BG); cl.number_format = "#,##0"
    cell_name(wb, "LatePickups", "Attendance", f"$E${tot}")
    ws.cell(row=tot + 2, column=2, value="Late pickup fees earned (at $1/min, 15 min average)").style = "field_label"
    clf = ws.cell(row=tot + 2, column=5, value="=LatePickups*15"); clf.style = "field_value"
    clf.number_format = '"$"#,##0'; clf.fill = fill(MINT_BG)
    ws.cell(row=tot + 4, column=2, value="Child-days feed your food program claim. Every missed day is a missed reimbursement.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_food(wb):
    ws = wb.create_sheet("Food Program"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 18, 20, 2])
    luxe_header(ws, "E", "\U0001f34e  FOOD PROGRAM (CACFP)",
                "Claim every meal — the food program pays back most of what you spend feeding them.")
    table_headers(ws, 4, ["Meal", "Meals claimed", "Rate", "Reimbursement"], start_col=2)
    start = L0
    for i, (meal, n, rate) in enumerate(MEALS):
        r = start + i
        ws.cell(row=r, column=2, value=meal).style = "td_left"
        cn = ws.cell(row=r, column=3, value=n); cn.style = "input"; cn.number_format = "#,##0"
        cr = ws.cell(row=r, column=4, value=rate); cr.style = "input"; cr.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MEALS) - 1
    nrange(wb, "MealCount", "Food Program", "C", start, end)
    nrange(wb, "MealReimb", "Food Program", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="= REIMBURSEMENT THIS MONTH").style = "th"
    cn = ws.cell(row=tot, column=3, value="=SUM(MealCount)"); cn.style = "td"
    cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=5, value="=SUM(MealReimb)"); cm.style = "td"
    cm.font = Font(bold=True, size=14, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0.00'
    cell_name(wb, "CACFP", "Food Program", f"$E${tot}")

    r = tot + 2
    ws.cell(row=r, column=2, value="DOES IT COVER YOUR FOOD?").style = "section_gold"
    ws.cell(row=r + 1, column=2, value="You spent on food this month").style = "field_label"
    cs = ws.cell(row=r + 1, column=5, value="=FoodPerChild*Enrolled"); cs.style = "field_value"; cs.number_format = '"$"#,##0.00'
    cell_name(wb, "FoodSpend", "Food Program", f"$E${r+1}")
    ws.cell(row=r + 2, column=2, value="The program paid you back").style = "field_label"
    cp = ws.cell(row=r + 2, column=5, value="=CACFP"); cp.style = "field_value"; cp.number_format = '"$"#,##0.00'
    ws.cell(row=r + 3, column=2, value="= IT COVERS THIS MUCH OF YOUR FOOD BILL").style = "th"
    cc = ws.cell(row=r + 3, column=5, value="=IFERROR(CACFP/FoodSpend,0)"); cc.style = "td"
    cc.font = Font(bold=True, size=14, color=PRIMARY); cc.fill = fill(MINT_BG); cc.number_format = "0%"
    cell_name(wb, "FoodCoverage", "Food Program", f"$E${r+3}")
    ws.cell(row=r + 4, column=2, value="Out of pocket on food").style = "field_label"
    co = ws.cell(row=r + 4, column=5, value="=MAX(FoodSpend-CACFP,0)"); co.style = "field_value"
    co.number_format = '"$"#,##0.00'; co.fill = fill(WARN_BG)
    ws.cell(row=r + 6, column=2, value="Every unclaimed snack is money you already spent and simply didn't ask for back.").style = "section_gold"
    ws.cell(row=r + 7, column=2, value="Reimbursement tiers & rates vary by state and by year — enter your own above.").style = "field_label"


def build_ratios(wb):
    ws = wb.create_sheet("Ratios & Schedule"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 15, 18, 18, 20, 2])
    luxe_header(ws, "F", "⚖  RATIOS & SCHEDULE",
                "Children per caregiver, by age group — the rule that closes programs when it slips.")
    table_headers(ws, 4, ["Age group", "Children", "State ratio", "Max group", "Caregivers needed"], start_col=2)
    start = L0
    for i, (ag, n, ratio, mx) in enumerate(RATIOS):
        r = start + i
        ws.cell(row=r, column=2, value=ag).style = "td_left"
        cn = ws.cell(row=r, column=3, value=n); cn.style = "input"; cn.number_format = "0"
        cr = ws.cell(row=r, column=4, value=ratio); cr.style = "input"; cr.number_format = "0"
        cm = ws.cell(row=r, column=5, value=mx); cm.style = "input"; cm.number_format = "0"
        cc = ws.cell(row=r, column=6, value=f"=IFERROR(ROUNDUP(C{r}/D{r},0),0)"); cc.style = "td"; cc.number_format = "0"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(RATIOS) - 1
    nrange(wb, "RatioChildren", "Ratios & Schedule", "C", start, end)
    nrange(wb, "RatioNeeded", "Ratios & Schedule", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL GROUPS").style = "th"
    cn = ws.cell(row=tot, column=3, value="=SUM(RatioChildren)"); cn.style = "td"
    cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = "0"
    for c in (4, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cc = ws.cell(row=tot, column=6, value="=SUM(RatioNeeded)"); cc.style = "td"
    cc.font = Font(bold=True, size=12, color=PRIMARY); cc.fill = fill(MINT_BG); cc.number_format = "0"
    cell_name(wb, "CaregiversNeeded", "Ratios & Schedule", f"$F${tot}")
    r = tot + 2
    ws.cell(row=r, column=2, value="Caregivers you have (including you)").style = "field_label"
    ch = ws.cell(row=r, column=6, value="=Caregivers"); ch.style = "field_value"; ch.number_format = "0"
    ws.cell(row=r + 1, column=2, value="= CHILDREN PER CAREGIVER").style = "th"
    cpc = ws.cell(row=r + 1, column=6, value="=IFERROR(Enrolled/Caregivers,0)"); cpc.style = "td"
    cpc.font = Font(bold=True, size=14, color=PRIMARY); cpc.fill = fill(MINT_BG); cpc.number_format = "0.0"
    cell_name(wb, "ChildrenPerCaregiver", "Ratios & Schedule", f"$F${r+1}")
    ws.cell(row=r + 2, column=2, value="= ARE YOU COVERED?").style = "th"
    cv = ws.cell(row=r + 2, column=6, value='=IF(Caregivers>=CaregiversNeeded,"COVERED","SHORT STAFFED")'); cv.style = "td"
    cv.font = Font(bold=True, size=12, color=PRIMARY); cv.fill = fill(MINT_BG)
    ws.conditional_formatting.add(f"F{r+2}", CellIsRule(operator="equal", formula=['"SHORT STAFFED"'],
                                                       fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    ws.cell(row=r + 4, column=2, value="Ratios and group sizes are set by your state. Enter yours above — never guess this one.").style = "section_gold"


def build_compliance(wb):
    ws = wb.create_sheet("Compliance & Files"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 18, 18, 18, 16, 16, 2])
    luxe_header(ws, "G", "\U0001f4cb  COMPLIANCE & FILES",
                "The paperwork an inspector asks for — complete, or not, at a glance.")
    table_headers(ws, 4, ["Child", "Immunizations", "Emergency form", "Signed contract", "Allergy plan", "File complete"], start_col=2)
    start = L0
    for i, row in enumerate(COMPLIANCE):
        r = start + i
        ws.cell(row=r, column=2, value=row[0]).style = "td_left"
        for ci, v in enumerate(row[1:], 3):
            ws.cell(row=r, column=ci, value=v).style = "input"
        cf = ws.cell(row=r, column=7, value=f'=IF(COUNTIF(C{r}:F{r},"Yes")=4,"COMPLETE","MISSING")'); cf.style = "td"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(COMPLIANCE) - 1
    nrange(wb, "FileStatus", "Compliance & Files", "G", start, end)
    for col in ("C", "D", "E", "F"):
        add_dv(ws, f"{col}{start}:{col}{end}", "YesNoList")
    ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=['"MISSING"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 2
    ws.cell(row=tot, column=2, value="Files complete").style = "field_label"
    c1 = ws.cell(row=tot, column=7, value='=COUNTIF(FileStatus,"COMPLETE")'); c1.style = "field_value"
    c1.number_format = "0"; c1.fill = fill(MINT_BG)
    cell_name(wb, "FilesComplete", "Compliance & Files", f"$G${tot}")
    ws.cell(row=tot + 1, column=2, value="Files missing something").style = "field_label"
    c2 = ws.cell(row=tot + 1, column=7, value='=COUNTIF(FileStatus,"MISSING")'); c2.style = "field_value"
    c2.number_format = "0"; c2.fill = fill(RED_BG)
    ws.cell(row=tot + 3, column=2, value="An inspector will not wait while you find it. Clear the red rows this week.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_supplies(wb):
    ws, start, end = build_log(
        wb, "Supplies", "\U0001f4e6", "SUPPLIES & INVENTORY", "What's on the shelf — so you never run out mid-week.",
        ["Item", "On hand", "Reorder at", "Unit cost", "Value", "Reorder?"],
        [(i, oh, ra, uc) for (i, oh, ra, uc) in SUPPLIES],
        [2, 32, 14, 14, 14, 14, 15, 2], text_left={2}, ints={3, 4}, money2={5, 6}, start_col=2)
    for r in range(start, start + len(SUPPLIES)):
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7, value=f'=IF(C{r}<=D{r},"REORDER","OK")')
    nrange(wb, "SupValue", "Supplies", "F", start, end)
    nrange(wb, "SupFlag", "Supplies", "G", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=['"REORDER"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tr = end + 2
    ws.cell(row=tr, column=2, value="Value on the shelf").style = "field_label"
    c1 = ws.cell(row=tr, column=6, value="=SUM(SupValue)"); c1.style = "field_value"
    c1.number_format = '"$"#,##0'; c1.fill = fill(MINT_BG)
    ws.cell(row=tr + 1, column=2, value="Items to reorder").style = "field_label"
    c2 = ws.cell(row=tr + 1, column=6, value='=COUNTIF(SupFlag,"REORDER")'); c2.style = "field_value"; c2.number_format = "0"


def build_tax(wb):
    ws = wb.create_sheet("Tax Set-Aside"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 18, 18, 18, 18, 2])
    luxe_header(ws, "F", "\U0001f3e6  TAX SET-ASIDE",
                "A quarter of what you earned is not yours. Move it before you spend it.")
    table_headers(ws, 4, ["Quarter", "Income", "Set aside", "Estimated due", "Short by"], start_col=2)
    start = L0
    for i, (q, inc, saved, due) in enumerate(TAXES):
        r = start + i
        ws.cell(row=r, column=2, value=q).style = "td_left"
        ci = ws.cell(row=r, column=3, value=inc); ci.style = "input"; ci.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=5, value=due); cd.style = "input"; cd.number_format = '"$"#,##0'
        csh = ws.cell(row=r, column=6, value=f"=MAX(E{r}-D{r},0)"); csh.style = "td"; csh.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(TAXES) - 1
    nrange(wb, "TaxIncome", "Tax Set-Aside", "C", start, end)
    nrange(wb, "TaxSaved", "Tax Set-Aside", "D", start, end)
    nrange(wb, "TaxDue", "Tax Set-Aside", "E", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="greaterThan", formula=["0"],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="THE YEAR").style = "th"
    ci = ws.cell(row=tot, column=3, value="=SUM(TaxIncome)"); ci.style = "td"
    ci.font = Font(bold=True, color=PRIMARY); ci.fill = fill(SURFACE); ci.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=4, value="=SUM(TaxSaved)"); cs.style = "td"
    cs.font = Font(bold=True, size=12, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cell_name(wb, "TaxReserve", "Tax Set-Aside", f"$D${tot}")
    cd = ws.cell(row=tot, column=5, value="=SUM(TaxDue)"); cd.style = "td"
    cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = '"$"#,##0'
    csh = ws.cell(row=tot, column=6, value=f"=MAX(E{tot}-D{tot},0)"); csh.style = "td"
    csh.font = Font(bold=True, size=12, color=DANGER); csh.fill = fill(RED_BG); csh.number_format = '"$"#,##0'
    r = tot + 2
    ws.cell(row=r, column=2, value="You should set aside this month").style = "field_label"
    cm = ws.cell(row=r, column=4, value="=YourPay*TaxRate"); cm.style = "field_value"
    cm.number_format = '"$"#,##0'; cm.fill = fill(WARN_BG)
    ws.cell(row=r + 1, column=2, value="Reserve funded against goal").style = "field_label"
    cf = ws.cell(row=r + 1, column=4, value="=IFERROR(TaxReserve/TaxReserveGoal,0)"); cf.style = "field_value"
    cf.number_format = "0%"; cf.fill = fill(WARN_BG)
    ws.cell(row=r + 3, column=2, value="Self-employment tax is the bill that ends home daycares. Move the money the day it lands.").style = "section_gold"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 18, 2])
    luxe_header(ws, "C", "\U0001f4c8  MONTHLY SUMMARY",
                "Revenue in, costs out, and what is actually yours at the end of it.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    rows = [
        ("Tuition billed", "=TuitionBilled", '"$"#,##0', None),
        ("+ Late & activity fees", LATE_FEES, '"$"#,##0', "LateFees"),
        ("+ Food program (CACFP)", "=CACFP", '"$"#,##0', None),
        ("= TOTAL REVENUE", "=TuitionBilled+LateFees+CACFP", '"$"#,##0', "Revenue"),
        ("− Total costs (fixed + per child)", "=TotalCosts", '"$"#,##0', None),
    ]
    for i, (lab, val, fmt, nm) in enumerate(rows):
        r = 6 + i
        big = lab.startswith("=") or lab.startswith("= ")
        ws.cell(row=r, column=2, value=lab).style = "th" if big else "field_label"
        c = ws.cell(row=r, column=3, value=val)
        if big:
            c.style = "td"; c.font = Font(bold=True, size=13, color=PRIMARY); c.fill = fill(SURFACE)
        elif nm:
            c.style = "input"
        else:
            c.style = "field_value"
        c.number_format = fmt
        if nm:
            cell_name(wb, nm, "Monthly Summary", f"$C${r}")
    ws.cell(row=12, column=2, value="= YOUR PAY (BEFORE TAX)").style = "th"
    cp = ws.cell(row=12, column=3, value="=Revenue-TotalCosts"); cp.style = "td"
    cp.font = Font(bold=True, size=17, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = '"$"#,##0'
    cell_name(wb, "YourPay", "Monthly Summary", "$C$12")
    ws.cell(row=13, column=2, value="= MARGIN").style = "th"
    cm = ws.cell(row=13, column=3, value="=IFERROR(YourPay/Revenue,0)"); cm.style = "td"
    cm.font = Font(bold=True, size=13, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = "0.0%"
    cell_name(wb, "Margin", "Monthly Summary", "$C$13")
    ws.cell(row=14, column=2, value="= YOUR PAY PER HOUR OPEN").style = "th"
    ch = ws.cell(row=14, column=3, value="=IFERROR(YourPay/(OpenHours*WeeksMonth),0)"); ch.style = "td"
    ch.font = Font(bold=True, size=15, color=PRIMARY); ch.fill = fill(MINT_BG); ch.number_format = '"$"#,##0.00'
    cell_name(wb, "PayPerHour", "Monthly Summary", "$C$14")
    ws.cell(row=15, column=2, value="− tax set-aside = WHAT YOU REALLY KEEP PER HOUR").style = "th"
    cha = ws.cell(row=15, column=3, value="=PayPerHour*(1-TaxRate)"); cha.style = "td"
    cha.font = Font(bold=True, size=13, color=PRIMARY); cha.fill = fill(WARN_BG); cha.number_format = '"$"#,##0.00'
    cell_name(wb, "PayPerHourNet", "Monthly Summary", "$C$15")
    ws.cell(row=16, column=2, value="= RUN-RATE YEAR").style = "th"
    cy = ws.cell(row=16, column=3, value="=YourPay*12"); cy.style = "td"
    cy.font = Font(bold=True, size=13, color=PRIMARY); cy.fill = fill(SURFACE); cy.number_format = '"$"#,##0'
    cell_name(wb, "RunRate", "Monthly Summary", "$C$16")
    ws.cell(row=18, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 19, ["Month", "Revenue"], start_col=2)
    ts = 20
    for i, (m, rev) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        crv = ws.cell(row=r, column=3, value=rev); crv.style = "input"; crv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "RevTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Revenue by Month", ts, te, 3, 2), "E5")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  \U0001f33f  DAYCARE & CHILDCARE PROVIDER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  What you actually earn, per child and per hour — plus a Care Score, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("WEEKLY RATE", "=WeeklyRate", "money2"),
        ("TUITION / CHILD", "=TuitionPerChild", "money"),
        ("COST / CHILD", "=CostPerChild", "money2"),
        ("NET PER CHILD", "=NetPerChild", "money"),
        ("ENROLLED", "=Enrolled", "num"),
        ("OCCUPANCY", "=Occupancy", "pct1"),
    ]
    row2 = [
        ("MONTHLY REVENUE", "=Revenue", "money"),
        ("MONTHLY COSTS", "=TotalCosts", "money"),
        ("YOUR PAY", "=YourPay", "money"),
        ("BREAK-EVEN CHILDREN", "=BreakEven", "num"),
        ("YOUR $ / HOUR", "=PayPerHour", "money2"),
        ("CARE SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "PROGRAM HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Spots filled", "=IFERROR(MIN(Occupancy/OccupancyGoal,1),0)"),
        ("Margin healthy", "=IFERROR(MIN(Margin/MarginGoal,1),0)"),
        ("House more than covered", "=IFERROR(MIN(CoverRatio/CoverGoal,1),0)"),
        ("Ratios comfortable", "=IF(ChildrenPerCaregiver<=RatioMax,1,IFERROR(RatioMax/ChildrenPerCaregiver,0))"),
        ("Your hourly on target", "=IFERROR(MIN(PayPerHour/HourlyGoal,1),0)"),
        ("Tax reserve funded", "=IFERROR(MIN(TaxReserve/TaxReserveGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=20, max_row=19 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=20, max_row=19 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Daycare & Childcare Provider Command Center™ — you're not babysitting.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_costs(wb); build_rate(wb)
    build_children(wb); build_payments(wb); build_attendance(wb); build_food(wb)
    build_ratios(wb); build_compliance(wb); build_supplies(wb); build_tax(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Rate & Enrollment", "Children & Families", "Tuition & Payments",
             "Attendance", "Costs & Expenses", "Food Program", "Ratios & Schedule", "Compliance & Files",
             "Supplies", "Tax Set-Aside", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Daycare_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
