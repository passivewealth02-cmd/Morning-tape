"""Build Instacart Shopper Command Center™ — The Ultimate Grocery-Delivery Earnings & Budget System.

18 sheets + Welcome · a premium grocery-shopper operating system in Excel & Sheets.
Batch log, earnings, mileage, fuel, vehicle & maintenance, business expenses, a
household budget, a tax center (mileage deduction), savings & goals, peak boosts,
best stores, ratings and an Analytics shopper-health score — one dashboard.
Built for Instacart; works for Shipt, Amazon Flex or any grocery-delivery gig.

Run: python3 build_xlsx.py   ->  ../Instacart_Shopper_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

PLATFORMS = ["Instacart", "Shipt", "Amazon Flex", "Multi", "Other"]
EXP_CATS = ["Fuel", "Car Payment", "Insurance", "Maintenance & Repairs", "Phone & Data",
            "Bags & Coolers", "Tolls", "Car Wash & Cleaning", "Miscellaneous"]
BUDGET_CATS = ["Housing", "Utilities", "Groceries", "Health", "Debt", "Savings",
               "Personal", "Insurance", "Other"]
MAINT_TYPES = ["Oil Change", "Tires", "Brakes", "Inspection", "Rotation", "Battery", "Repair", "Other"]
GOAL_CATS = ["Earnings", "Hourly", "Batches", "Savings", "Miles", "Tax", "Debt"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
        "msg": NamedStyle(name="msg", font=f(10, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1), border=BOX),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 15 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%",
                        "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your goals & vehicle details once — every dashboard follows.")
    merge_set(ws, "B5:C5", "SHOPPER INPUTS", "section")
    controls = [
        ("Shopper Name", "Riley Chen", None, "DriverName"),
        ("Vehicle", "2020 Subaru Outback", None, "VehicleName"),
        ("Home City", "Portland, OR", None, "HomeCity"),
        ("Monthly Net Goal", 2000, '"$"#,##0', "NetGoal"),
        ("Target Net $ / Hour", 17, '"$"#,##0.00', "HourlyTarget"),
        ("Monthly Batches Goal", 90, "#,##0", "BatchGoal"),
        ("Monthly Shifts Goal", 16, "0", "ShiftGoal"),
        ("Emergency Fund Goal", 4000, '"$"#,##0', "SavingsGoal"),
        ("Tax Reserve Goal (mo)", 900, '"$"#,##0', "TaxReserveGoal"),
        ("IRS Mileage Rate", 0.70, '"$"#,##0.00', "MileageRate"),
        ("Tax Set-Aside %", 0.25, "0%", "TaxRate"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "App", PLATFORMS, "PlatformList"), ("F", "Expense Category", EXP_CATS, "ExpCatList"),
             ("G", "Budget Category", BUDGET_CATS, "BudgetCatList"), ("H", "Maintenance", MAINT_TYPES, "MaintList"),
             ("I", "Goal Category", GOAL_CATS, "GoalCatList"), ("J", "Priority", PRIORITIES, "PriorityList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    ws.cell(row=17, column=5, value="Yes / No").style = "th"
    for ri, v in enumerate(YESNO):
        ws.cell(row=18 + ri, column=5, value=v).style = "td_left"
    wb.defined_names["YesNoList"] = DefinedName("YesNoList", attr_text="Settings!$E$18:$E$19")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🥕  INSTACART SHOPPER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate grocery-delivery earnings & budget system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "RUN YOUR WHOLE SHOPPING BUSINESS FROM ONE FILE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Every batch, mile and dollar in ONE premium Excel & Google Sheets system. Log your batches and "
                      "the Command Center instantly shows your real take-home — net earnings, true $/hour, $/mile and "
                      "$/batch after gas and expenses, your mileage tax deduction, your budget and your savings. See "
                      "which stores, apps and hours actually pay, and stop guessing whether a batch was worth it. Built "
                      "for Instacart — works for Shipt, Amazon Flex or any grocery-delivery gig.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your vehicle, goals & the IRS mileage rate.",
             "2.  Log each shift in the Batch Log — hours, batches, items, pay, tips, miles & gas.",
             "3.  Track Mileage, Fuel & Vehicle Maintenance as you go.",
             "4.  Set your Business Expenses & the Monthly Budget — net income flows in live.",
             "5.  Use the Tax Center to bank your mileage deduction & set-aside.",
             "6.  Watch the Dashboard track net $/hour, savings & a Shopper Health Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (Riley, a full-time Portland shopper doing ~$2,470/mo across 16 shifts, 98 "
               "batches, 2,440 items and ~860 miles) is included so you can see how everything connects — just type "
               "over it with your own numbers. Net earnings, true $/hour, $/mile and $/batch, the mileage tax "
               "deduction, budget leftover and the Shopper Health Score all update automatically. Every sheet is "
               "print-friendly and works in Excel and Google Sheets, on desktop and phone — perfect for logging in the "
               "checkout line.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, every mile counted — let's grow your take-home.", "section_gold")


# ===========================================================================
# Shopper Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Shopper Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "🪪  SHOPPER PROFILE", "Your shopping business, defined — the details every number flows from.")
    blocks = [
        ("THE SHOPPER", [("Shopper Name", "=DriverName"), ("Home City", "=HomeCity"),
                         ("Apps", "Instacart + Shipt"), ("Shopping Since", "2020"),
                         ("Status", "Full-time"), ("Weekly Target", "5 shifts")]),
        ("THE VEHICLE", [("Vehicle", "=VehicleName"), ("MPG (avg)", "29 city/hwy"),
                         ("Odometer", "74,300 mi"), ("Cargo", "Wagon — big trunk"),
                         ("Insurance", "Delivery rider"), ("Est. Value", "$18,900")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    merge_set(ws, "B15:F15", "RATINGS & STANDING", "section_gold"); ws.row_dimensions[15].height = 22
    stand = [("Instacart Rating", "4.94 ★"), ("Shipt Rating", "4.91 ★"), ("On-Time", "97%"),
             ("Cancellation", "1%"), ("Lifetime Batches", "3,860"), ("Tier", "Cart Star")]
    for i, (p, h) in enumerate(stand):
        r = 16 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# Batch Log — core earnings engine
# ===========================================================================
SHIFTS = [
    (23, "Instacart", 6.0, 7, 168, 64, 92, 15, 58, 13),
    (22, "Instacart", 5.5, 6, 150, 56, 80, 12, 52, 12),
    (20, "Shipt", 5.0, 5, 128, 48, 66, 8, 46, 11),
    (19, "Multi", 6.5, 8, 190, 74, 105, 20, 64, 15),
    (17, "Instacart", 4.5, 4, 102, 40, 52, 6, 40, 9),
    (16, "Instacart", 6.5, 8, 196, 76, 110, 22, 66, 15),
    (15, "Instacart", 5.5, 6, 152, 58, 84, 14, 54, 12),
    (13, "Multi", 6.0, 7, 172, 66, 96, 18, 60, 14),
    (12, "Instacart", 5.0, 5, 130, 50, 68, 10, 48, 11),
    (10, "Shipt", 6.0, 7, 166, 64, 90, 16, 58, 13),
    (9, "Instacart", 5.0, 5, 128, 48, 64, 8, 46, 11),
    (8, "Multi", 6.0, 6, 158, 62, 88, 15, 56, 13),
    (6, "Instacart", 4.5, 4, 100, 38, 50, 6, 40, 9),
    (5, "Instacart", 6.5, 8, 192, 74, 108, 21, 64, 15),
    (3, "Shipt", 5.5, 6, 148, 56, 78, 12, 52, 12),
    (2, "Multi", 6.0, 6, 160, 62, 86, 16, 56, 13),
]


def build_batchlog(wb):
    ws = wb.create_sheet("Batch Log"); ws.sheet_view.showGridLines = False
    headers = ["Date", "App", "Hours", "Batches", "Items", "Batch Pay", "Tips", "Boost",
               "Earnings", "Miles", "Gas $"]
    set_widths(ws, [13, 13, 9, 9, 9, 11, 10, 10, 12, 9, 9])
    luxe_header(ws, "K", "🧺  BATCH LOG",
                "Log every shift — earnings, items, miles & gas roll straight into your dashboard.")
    table_headers(ws, 4, headers)
    start = L0
    reserved = 60
    end = start + reserved - 1
    for i, (off, plat, hrs, batches, items, base, tips, boost, miles, gas) in enumerate(SHIFTS):
        r = start + i
        ws.cell(row=r, column=1, value=dminus(off))
        ws.cell(row=r, column=2, value=plat)
        ws.cell(row=r, column=3, value=hrs)
        ws.cell(row=r, column=4, value=batches)
        ws.cell(row=r, column=5, value=items)
        ws.cell(row=r, column=6, value=base)
        ws.cell(row=r, column=7, value=tips)
        ws.cell(row=r, column=8, value=boost)
        ws.cell(row=r, column=9, value=f"=F{r}+G{r}+H{r}")
        ws.cell(row=r, column=10, value=miles)
        ws.cell(row=r, column=11, value=gas)
    style_rows(ws, start, end, len(headers), text_left=set(), dates={1},
               money={6, 7, 8, 9, 11}, ints={4, 5, 10}, dec={3})
    add_dv(ws, f"B{start}:B{end}", "PlatformList")
    ws.freeze_panes = "A5"
    trow = end + 1
    ws.cell(row=trow, column=1, value="TOTAL").style = "th"
    ws.cell(row=trow, column=2).style = "th"
    fmts = {3: "0.0", 4: "#,##0", 5: "#,##0", 6: '"$"#,##0', 7: '"$"#,##0', 8: '"$"#,##0',
            9: '"$"#,##0', 10: "#,##0", 11: '"$"#,##0'}
    for col in range(3, 12):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmts[col]
    nrange(wb, "ShiftDate", "Batch Log", "A", start, end)
    nrange(wb, "ShiftPlatform", "Batch Log", "B", start, end)
    nrange(wb, "ShiftHours", "Batch Log", "C", start, end)
    nrange(wb, "ShiftBatches", "Batch Log", "D", start, end)
    nrange(wb, "ShiftItems", "Batch Log", "E", start, end)
    nrange(wb, "ShiftBase", "Batch Log", "F", start, end)
    nrange(wb, "ShiftTips", "Batch Log", "G", start, end)
    nrange(wb, "ShiftBoost", "Batch Log", "H", start, end)
    nrange(wb, "ShiftEarn", "Batch Log", "I", start, end)
    nrange(wb, "ShiftMiles", "Batch Log", "J", start, end)
    nrange(wb, "ShiftFuel", "Batch Log", "K", start, end)
    cell_name(wb, "TotalHours", "Batch Log", f"$C${trow}")
    cell_name(wb, "TotalBatches", "Batch Log", f"$D${trow}")
    cell_name(wb, "TotalItems", "Batch Log", f"$E${trow}")
    cell_name(wb, "TotalBase", "Batch Log", f"$F${trow}")
    cell_name(wb, "TotalTips", "Batch Log", f"$G${trow}")
    cell_name(wb, "TotalBoost", "Batch Log", f"$H${trow}")
    cell_name(wb, "GrossEarn", "Batch Log", f"$I${trow}")
    cell_name(wb, "TotalMiles", "Batch Log", f"$J${trow}")
    cell_name(wb, "FuelTotal", "Batch Log", f"$K${trow}")
    ws.conditional_formatting.add(f"I{start}:I{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=210, color=PRIMARY, showValue=True))


# ===========================================================================
# Earnings Breakdown
# ===========================================================================
def build_earnings(wb):
    ws = wb.create_sheet("Earnings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 14, 12, 3, 18, 14, 12, 2])
    luxe_header(ws, "H", "💵  EARNINGS BREAKDOWN",
                "Where the money comes from — earnings mix and app split, live from your log.")
    merge_set(ws, "B5:D5", "EARNINGS MIX (THIS MONTH)", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Source", "Amount", "Share"], start_col=2)
    mix = [("Batch Pay", "=TotalBase"), ("Tips", "=TotalTips"), ("Peak Boost", "=TotalBoost")]
    ms = 7
    for i, (lab, fml) in enumerate(mix):
        r = ms + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cv = ws.cell(row=r, column=3, value=fml); cv.style = "td"; cv.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/GrossEarn,0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    me = ms + len(mix) - 1
    mt = me + 1
    ws.cell(row=mt, column=2, value="TOTAL EARNINGS").style = "th"
    ct = ws.cell(row=mt, column=3, value="=GrossEarn"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    ws.cell(row=mt, column=4).style = "td"; ws.cell(row=mt, column=4).fill = fill(SURFACE)
    nrange(wb, "MixSource", "Earnings", "B", ms, me)
    nrange(wb, "MixVal", "Earnings", "C", ms, me)
    ws.conditional_formatting.add(f"C{ms}:C{me}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1350, color=GOLD_LT, showValue=True))
    merge_set(ws, "F5:H5", "BY APP", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["App", "Earnings", "Share"], start_col=6)
    plats = ["Instacart", "Shipt", "Multi"]
    ps = 7
    for i, p in enumerate(plats):
        r = ps + i
        ws.cell(row=r, column=6, value=p).style = "td_left"
        cv = ws.cell(row=r, column=7, value=f'=SUMIF(ShiftPlatform,"{p}",ShiftEarn)'); cv.style = "td"; cv.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/GrossEarn,0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(6, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    pe = ps + len(plats) - 1
    nrange(wb, "PlatName", "Earnings", "F", ps, pe)
    nrange(wb, "PlatEarn", "Earnings", "G", ps, pe)
    merge_set(ws, "B13:D13", "PER-BATCH & ITEM", "section_gold"); ws.row_dimensions[13].height = 22
    avgs = [("Avg / batch", "=IFERROR(GrossEarn/TotalBatches,0)", '"$"#,##0.00'),
            ("$ / item", "=IFERROR(GrossEarn/TotalItems,0)", '"$"#,##0.00'),
            ("Gross $ / hour", "=IFERROR(GrossEarn/TotalHours,0)", '"$"#,##0.00'),
            ("Tips % of earnings", "=IFERROR(TotalTips/GrossEarn,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(avgs):
        r = 14 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        ws.cell(row=r, column=4).style = "td"; ws.cell(row=r, column=4).fill = fill(WHITE)
    merge_set(ws, "F13:H13", "TAKE-HOME (AFTER EXPENSES)", "section_gold"); ws.row_dimensions[13].height = 22
    take = [("Net earnings", "=NetEarn", '"$"#,##0'), ("Net $ / hour", "=PerHour", '"$"#,##0.00'),
            ("Net $ / mile", "=PerMile", '"$"#,##0.00'), ("Expense ratio", "=IFERROR(ExpenseTotal/GrossEarn,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(take):
        r = 14 + i
        ws.cell(row=r, column=6, value=lab).style = "field_label"
        c = ws.cell(row=r, column=7, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net earnings", "Net $ / hour"):
            c.fill = fill(MINT_BG)
        ws.cell(row=r, column=8).style = "td"; ws.cell(row=r, column=8).fill = fill(WHITE)


# ===========================================================================
# Mileage Tracker
# ===========================================================================
def build_mileage(wb):
    rows = [
        (dminus(23), "Costco + NW deliveries", "Business", 58, "Big weekend batches"),
        (dminus(22), "Whole Foods run", "Business", 52, "Dinner prep orders"),
        (dminus(20), "New Seasons + Pearl", "Business", 46, "Shipt orders"),
        (dminus(19), "Multi-store day", "Business", 64, "Instacart + Shipt"),
        (dminus(16), "Costco doubles", "Business", 66, "Two big batches"),
        (dminus(13), "Fred Meyer + SE", "Business", 60, "Steady"),
        (dminus(10), "Whole Foods + Hills", "Business", 58, "High-tip zone"),
        (dminus(8), "Multi-store dinner", "Business", 56, "Stacked batches"),
        (dminus(5), "Weekend rush", "Business", 64, "Best day this week"),
        (dminus(2), "Costco + downtown", "Business", 56, "End of month"),
        (dminus(11), "Personal errands", "Personal", 16, "Not deductible"),
    ]
    ws, start, end = build_log(
        wb, "Mileage", "🛣", "MILEAGE TRACKER",
        "Log business miles for the IRS deduction — the biggest tax break shoppers miss.",
        ["Date", "Route / Stores", "Type", "Miles", "Notes"],
        rows, [13, 30, 14, 12, 26],
        text_left={2, 5}, dates={1}, ints={4}, reserved=40)
    nrange(wb, "MileType", "Mileage", "C", start, end)
    nrange(wb, "MileMiles", "Mileage", "D", start, end)
    for t, cc in {"Business": MINT_BG, "Personal": WARN_BG}.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{t}"'], fill=fill(cc)))


# ===========================================================================
# Fuel Log
# ===========================================================================
def build_fuel(wb):
    rows = [
        (dminus(23), "Costco", 10.2, 3.35, "Cheapest in town"),
        (dminus(20), "Fred Meyer", 9.6, 3.49, "Fuel points"),
        (dminus(17), "Chevron", 9.1, 3.59, "Near Whole Foods"),
        (dminus(14), "Costco", 10.4, 3.35, "Membership pays off"),
        (dminus(11), "76", 8.8, 3.65, "Quick splash"),
        (dminus(8), "Fred Meyer", 9.7, 3.45, "Weekend fill"),
        (dminus(5), "Costco", 10.5, 3.35, "Before big shift"),
        (dminus(2), "Chevron", 9.2, 3.59, "End of month"),
    ]
    sample = [(d, st, g, ppg, f"=ROUND(C{i+L0}*D{i+L0},2)", note)
              for i, (d, st, g, ppg, note) in enumerate(rows)]
    ws, start, end = build_log(
        wb, "Fuel Log", "⛽", "FUEL LOG",
        "Track every fill-up — gallons, price & cost so your real gas spend is never a guess.",
        ["Date", "Station", "Gallons", "$ / Gal", "Cost", "Notes"],
        sample, [13, 16, 12, 12, 12, 24],
        text_left={2, 6}, dates={1}, dec={3}, money2={4, 5}, reserved=40)
    trow = end + 1
    ws.cell(row=trow, column=2, value="TOTAL").style = "th"
    for col, fmt in ((3, "0.0"), (5, '"$"#,##0.00')):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt
    for col in (1, 4, 6):
        ws.cell(row=trow, column=col).style = "td"; ws.cell(row=trow, column=col).fill = fill(SURFACE)


# ===========================================================================
# Vehicle & Maintenance
# ===========================================================================
def build_vehicle(wb):
    rows = [
        (dminus(18), "Oil Change", 68, 73400, "Synthetic — every 5k mi", dplus(60)),
        (dminus(40), "Rotation", 25, 72100, "With oil change", dplus(38)),
        (dminus(120), "Tires", 680, 68200, "4 new — all-weather", dplus(600)),
        (dminus(70), "Brakes", 300, 70500, "Front pads + rotors", dplus(700)),
        (dminus(9), "Inspection", 40, 74200, "State + emissions", dplus(355)),
        (dminus(200), "Battery", 150, 64800, "3-yr warranty", dplus(900)),
        (dminus(30), "Repair", 120, 72900, "Cabin filter + wipers", "—"),
    ]
    ws, start, end = build_log(
        wb, "Vehicle", "🔧", "VEHICLE & MAINTENANCE",
        "Keep your money-maker running — service history, cost & what's due next.",
        ["Date", "Service", "Cost", "Odometer", "Notes", "Next Due"],
        rows, [13, 16, 12, 13, 26, 13],
        text_left={2, 5}, dates={1, 6}, money={3}, ints={4}, reserved=30,
        validations=[("B", "MaintList")])
    trow = end + 1
    ws.cell(row=trow, column=2, value="TOTAL SPENT").style = "th"
    c = ws.cell(row=trow, column=3, value=f"=SUM(C{start}:C{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    for col in (1, 4, 5, 6):
        ws.cell(row=trow, column=col).style = "td"; ws.cell(row=trow, column=col).fill = fill(SURFACE)


# ===========================================================================
# Business Expenses  — defines ExpenseTotal
# ===========================================================================
def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 16, 3, 30, 2])
    luxe_header(ws, "F", "🧾  BUSINESS EXPENSES",
                "Every cost of shopping — because deductions are money back in your pocket.")
    table_headers(ws, 4, ["Category", "This Month", "Annual (est.)"])
    exp = {"Fuel": "=FuelTotal", "Car Payment": 300, "Insurance": 155, "Maintenance & Repairs": 95,
           "Phone & Data": 45, "Bags & Coolers": 35, "Tolls": 15, "Car Wash & Cleaning": 25, "Miscellaneous": 0}
    start = L0; eend = start + len(EXP_CATS) - 1
    for i, cat in enumerate(EXP_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        val = exp[cat]
        cm = ws.cell(row=r, column=2, value=val)
        cm.style = "field_value" if cat == "Fuel" else "input"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    etot = eend + 1
    ws.cell(row=etot, column=1, value="TOTAL EXPENSES").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=etot, column=col, value=f"=SUM({L}{start}:{L}{eend})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    nrange(wb, "ExpCat", "Expenses", "A", start, eend)
    nrange(wb, "ExpMonthly", "Expenses", "B", start, eend)
    cell_name(wb, "ExpenseTotal", "Expenses", f"$B${etot}")
    ws.conditional_formatting.add(f"B{start}:B{eend}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=300, color=ACCENT, showValue=True))
    merge_set(ws, "E4:F4", "COST PER MILE", "section_gold")
    rows2 = [("Total business miles", "=TotalMiles", "#,##0"),
             ("Expenses this month", "=ExpenseTotal", '"$"#,##0'),
             ("Cost per mile", "=IFERROR(ExpenseTotal/TotalMiles,0)", '"$"#,##0.00'),
             ("Gas per mile", "=IFERROR(FuelTotal/TotalMiles,0)", '"$"#,##0.00'),
             ("Expense ratio", "=IFERROR(ExpenseTotal/GrossEarn,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 5 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Cost per mile":
            c.fill = fill(WARN_BG)


# ===========================================================================
# Monthly Budget
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "📊  MONTHLY BUDGET",
                "Your take-home, put to work — shopping net income flows in; plan every dollar.")
    merge_set(ws, "B5:E5", "INCOME", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Source", "Planned", "Actual", ""], start_col=2)
    inc = [("Shopping (net)", "=NetGoal", "=NetEarn"), ("Side income", 350, 300), ("Other", 0, 0)]
    is_ = 7
    for i, (lab, pl, ac) in enumerate(inc):
        r = is_ + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cp = ws.cell(row=r, column=3, value=pl); cp.style = "td" if lab == "Shopping (net)" else "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=ac); ca.style = "td" if lab == "Shopping (net)" else "input"; ca.number_format = '"$"#,##0'
        ws.cell(row=r, column=5).style = "td"; ws.cell(row=r, column=5).fill = fill(WHITE)
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ie = is_ + len(inc) - 1; itot = ie + 1
    ws.cell(row=itot, column=2, value="TOTAL INCOME").style = "th"
    for col in (3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=itot, column=col, value=f"=SUM({L}{is_}:{L}{ie})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=itot, column=5).style = "td"; ws.cell(row=itot, column=5).fill = fill(SURFACE)
    cell_name(wb, "BudgetIncome", "Budget", f"$D${itot}")
    merge_set(ws, "B12:E12", "HOUSEHOLD EXPENSES", "section_gold"); ws.row_dimensions[12].height = 22
    table_headers(ws, 13, ["Category", "Planned", "Actual", "Δ"], start_col=2)
    hh = [("Housing", 1050, 1050), ("Utilities", 170, 185), ("Groceries", 400, 420),
          ("Health", 210, 210), ("Debt", 250, 250), ("Savings", 300, 300),
          ("Personal", 220, 195), ("Insurance", 130, 130), ("Other", 110, 90)]
    hs = 14
    for i, (cat, pl, ac) in enumerate(hh):
        r = hs + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=3, value=pl); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=ac); ca.style = "input"; ca.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cd.style = "td"; cd.number_format = '"$"#,##0;[Red]-"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    he = hs + len(hh) - 1; htot = he + 1
    ws.cell(row=htot, column=2, value="TOTAL EXPENSES").style = "th"
    for col in (3, 4):
        L = get_column_letter(col)
        c = ws.cell(row=htot, column=col, value=f"=SUM({L}{hs}:{L}{he})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.cell(row=htot, column=5).style = "td"; ws.cell(row=htot, column=5).fill = fill(SURFACE)
    nrange(wb, "BudgetCat", "Budget", "B", hs, he)
    nrange(wb, "BudgetActual", "Budget", "D", hs, he)
    cell_name(wb, "BudgetExpense", "Budget", f"$D${htot}")
    lo = htot + 2
    merge_set(ws, f"B{lo}:C{lo}", "LEFTOVER (INCOME − EXPENSES)", "section_gold")
    c = ws.cell(row=lo, column=4, value="=BudgetIncome-BudgetExpense"); c.style = "field_value"; c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    ws.conditional_formatting.add(f"D{lo}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))


# ===========================================================================
# Tax Center
# ===========================================================================
def build_tax(wb):
    ws = wb.create_sheet("Tax Center"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 4, 26, 16, 2])
    luxe_header(ws, "G", "🧮  TAX CENTER",
                "Keep more of what you earn — the mileage deduction, set-aside & quarterly estimates.")
    merge_set(ws, "B5:C5", "MILEAGE VS ACTUAL METHOD", "section_gold"); ws.row_dimensions[5].height = 22
    left = [("Business miles", "=TotalMiles", "#,##0"),
            ("IRS mileage rate", "=MileageRate", '"$"#,##0.00'),
            ("Mileage deduction", "=TotalMiles*MileageRate", '"$"#,##0'),
            ("Actual costs (expenses)", "=ExpenseTotal", '"$"#,##0'),
            ("Better method", '=IF(TotalMiles*MileageRate>=ExpenseTotal,"Mileage","Actual")', "text"),
            ("Deduction (best)", "=MAX(TotalMiles*MileageRate,ExpenseTotal)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(left):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"
        if fmt != "text":
            c.number_format = fmt
        if lab in ("Mileage deduction", "Deduction (best)"):
            c.fill = fill(MINT_BG)
    cell_name(wb, "TaxDeduction", "Tax Center", "$C$8")
    merge_set(ws, "E5:F5", "SET-ASIDE & ESTIMATE", "section_gold"); ws.row_dimensions[5].height = 22
    right = [("Gross earnings", "=GrossEarn", '"$"#,##0'),
             ("− Mileage deduction", "=-TotalMiles*MileageRate", '"$"#,##0'),
             ("Taxable (est.)", "=MAX(GrossEarn-TotalMiles*MileageRate,0)", '"$"#,##0'),
             ("Set-aside rate", "=TaxRate", "0%"),
             ("Set aside this month", "=MAX(GrossEarn-TotalMiles*MileageRate,0)*TaxRate", '"$"#,##0'),
             ("Quarterly estimate", "=MAX(GrossEarn-TotalMiles*MileageRate,0)*TaxRate*3", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(right):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Set aside this month", "Quarterly estimate"):
            c.fill = fill(WARN_BG)
    merge_set(ws, "B13:F13", "DEDUCTION CHECKLIST — DON'T LEAVE MONEY ON THE TABLE", "section_gold"); ws.row_dimensions[13].height = 22
    checks = [("Business mileage (biggest one)", "Yes"), ("Phone & data (business %)", "Yes"),
              ("Insulated bags & coolers", "Yes"), ("Tolls & parking", "Yes"),
              ("Phone mount & chargers", "Yes"), ("Hand sanitizer & wipes", "Yes"),
              ("Car wash & cleaning", "Yes"), ("Platform & bank fees", "Yes")]
    for i, (item, yn) in enumerate(checks):
        r = 14 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=item).style = "td_left"
        cc = ws.cell(row=r, column=col + 1, value=yn); cc.style = "td"
        add_dv(ws, f"{get_column_letter(col+1)}{r}", "YesNoList")
        ws.conditional_formatting.add(f"{get_column_letter(col+1)}{r}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# Savings & Goals  — defines SavingsSaved, TaxSetAside
# ===========================================================================
def build_savings(wb):
    ws = wb.create_sheet("Savings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 12, 2])
    luxe_header(ws, "E", "🐷  SAVINGS & GOALS",
                "Pay yourself first — build an emergency fund, tax reserve & work toward big goals.")
    merge_set(ws, "B5:E5", "SAVINGS BUCKETS", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Bucket", "Goal", "Saved", "Progress"], start_col=2)
    buckets = [("Emergency fund", "=SavingsGoal", 2800), ("Tax reserve", "=TaxReserveGoal", 650),
               ("New(er) car fund", 7000, 2200), ("Vacation", 2000, 700)]
    bs = 7
    for i, (lab, goal, saved) in enumerate(buckets):
        r = bs + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        cg = ws.cell(row=r, column=3, value=goal); cg.style = "field_value" if isinstance(goal, str) else "input"; cg.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=4, value=saved); cs.style = "input"; cs.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(MIN(D{r}/C{r},1),0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    be = bs + len(buckets) - 1
    cell_name(wb, "SavingsSaved", "Savings", f"$D${bs}")
    cell_name(wb, "TaxSetAside", "Savings", f"$D${bs+1}")
    ws.conditional_formatting.add(f"E{bs}:E{be}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    merge_set(ws, "B13:E13", "SHOPPER GOALS & OKRs", "section_gold"); ws.row_dimensions[13].height = 22
    table_headers(ws, 14, ["Goal", "Target", "Current", "Progress"], start_col=2)
    goals = [("$2,000 net / month", "$2,000", "$1,604", 0.80),
             ("$17 net / hour", "$17.00", "$17.82", 1.00),
             ("90 batches / month", "90", "98", 1.00),
             ("$4k emergency fund", "$4,000", "$2,800", 0.70),
             ("Pay off $3k card", "$3,000", "$1,950", 0.65),
             ("Lift tips to 55%", "55%", "53%", 0.96)]
    gs = 15
    for i, (g, tgt, cur, prog) in enumerate(goals):
        r = gs + i
        ws.cell(row=r, column=2, value=g).style = "td_left"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ge = gs + len(goals) - 1
    ws.conditional_formatting.add(f"E{gs}:E{ge}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GOLD_LT, showValue=True))


# ===========================================================================
# Boosts & Bonuses
# ===========================================================================
def build_boosts(wb):
    rows = [
        ("Instacart", "Peak Boost", "+$4 / batch dinner", "Active", 52, dplus(1), "Dinner rush only"),
        ("Instacart", "Cart Star Bonus", "Keep 4.8★ + 90% on-time", "On Track", 40, dplus(7), "Extra batch access"),
        ("Shipt", "Promo Pay", "+$2 / order lunch", "Active", 24, dplus(1), "Lunch shifts"),
        ("Instacart", "Batch Incentive", "20 batches (Fri-Sun)", "In Progress", 60, dplus(2), "14 / 20 done"),
        ("Shipt", "Weekly Bonus", "$700 / 45 orders", "In Progress", 75, dplus(3), "On pace"),
        ("Instacart", "Referral", "Refer a shopper", "Available", 100, dplus(20), "Send your link"),
        ("Instacart", "Heavy Pay", "Bulk / heavy orders", "Active", 35, dplus(1), "Costco water runs"),
    ]
    ws, start, end = build_log(
        wb, "Boosts", "🎯", "BOOSTS & BONUSES",
        "Never miss a promo — track every peak boost, incentive & bonus and the extra it pays.",
        ["App", "Type", "Requirement", "Status", "Reward", "By", "Notes"],
        rows, [13, 16, 22, 14, 11, 13, 22],
        text_left={3, 7}, money={5}, dates={6}, reserved=30,
        validations=[("A", "PlatformList")])
    nrange(wb, "BoostReward", "Boosts", "E", start, end)
    for st, cc in {"Active": MINT_BG, "In Progress": WARN_BG, "On Track": WARN_BG, "Available": SOFT_BG}.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# Best Stores
# ===========================================================================
def build_stores(wb):
    ws = wb.create_sheet("Stores"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 16, 14, 14, 26, 2])
    luxe_header(ws, "G", "📍  BEST STORES & HOURS",
                "Shop where it pays — the stores, days & hours that earn the most per hour.")
    merge_set(ws, "B5:F5", "BEST STORES", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Store", "Best Time", "Avg $/batch", "Demand", "Notes"], start_col=2)
    zones = [
        ("Costco", "Weekend AM", 34, "High", "Big orders + heavy pay"),
        ("Whole Foods", "Dinner prep 4-7pm", 30, "High", "High tips, affluent zones"),
        ("New Seasons", "Weekday PM", 26, "Medium", "Loyal customers, good tips"),
        ("Fred Meyer", "Weekend", 24, "Medium", "Everything under one roof"),
        ("Safeway", "Weekday AM", 22, "Low", "Steady, smaller orders"),
        ("Sprouts", "Lunch", 21, "Low", "Quick, produce-heavy"),
    ]
    zs = 7
    for i, (z, t, bp, dem, note) in enumerate(zones):
        r = zs + i
        ws.cell(row=r, column=2, value=z).style = "td_left"
        ws.cell(row=r, column=3, value=t).style = "td"
        ch = ws.cell(row=r, column=4, value=bp); ch.style = "td"; ch.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=dem).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ze = zs + len(zones) - 1
    nrange(wb, "ZoneName", "Stores", "B", zs, ze)
    nrange(wb, "ZoneRate", "Stores", "D", zs, ze)
    ws.conditional_formatting.add(f"D{zs}:D{ze}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=34, color=PRIMARY, showValue=True))
    for st, cc in {"High": MINT_BG, "Medium": WARN_BG, "Low": WHITE}.items():
        ws.conditional_formatting.add(f"E{zs}:E{ze}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    merge_set(ws, "B15:F15", "BEST HOURS & DAYS", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Window", "Avg $/hr", "Why", "Rating"], start_col=2)
    days = [("Sat AM", 30, "Big weekly shops", "★★★★★"), ("Sun AM", 29, "Weekend restocks", "★★★★★"),
            ("Weekday 4-7pm", 27, "Dinner prep rush", "★★★★☆"), ("Fri PM", 26, "Weekend prep", "★★★★☆"),
            ("Weekday AM", 22, "Slower, smaller", "★★★☆☆")]
    ds = 17
    for i, (d, hr, why, rate) in enumerate(days):
        r = ds + i
        ws.cell(row=r, column=2, value=d).style = "td_left"
        cr = ws.cell(row=r, column=3, value=hr); cr.style = "td"; cr.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=why).style = "td"
        ws.cell(row=r, column=5, value=rate).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)


# ===========================================================================
# Ratings & Feedback
# ===========================================================================
def build_ratings(wb):
    ws = wb.create_sheet("Ratings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 12, 12, 14, 30, 2])
    luxe_header(ws, "G", "⭐  RATINGS & FEEDBACK",
                "Protect your rating — it drives your access to the best batches & bonuses.")
    merge_set(ws, "B5:F5", "PLATFORM STANDING", "section_gold"); ws.row_dimensions[5].height = 22
    table_headers(ws, 6, ["Metric", "Score", "Count", "Trend", "Customers Say"], start_col=2)
    rows = [
        ("Instacart Rating", 4.94, 3120, "▲ +0.01", "Great replacements, careful bagging"),
        ("Shipt Rating", 4.91, 740, "► flat", "Fresh produce, on time"),
        ("On-Time / Early", 0.97, 0, "▲ +2%", "Knows the fast lanes"),
        ("Replacements OK'd", 0.95, 0, "► flat", "Texts before subbing"),
    ]
    start = 7
    for i, (plat, rt, cnt, trend, note) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=plat).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rt); cr.style = "td"
        cr.number_format = "0.00" if rt > 3 else "0%"
        cc = ws.cell(row=r, column=4, value=cnt if cnt else "—"); cc.style = "td"
        if cnt:
            cc.number_format = "#,##0"
        ws.cell(row=r, column=5, value=trend).style = "td"
        ws.cell(row=r, column=6, value=note).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1
    ws.conditional_formatting.add(f"C{start}:C{start+1}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=5, color=GOLD_LT, showValue=True))
    merge_set(ws, "B12:F12", "5-STAR CHECKLIST", "section_gold"); ws.row_dimensions[12].height = 22
    tips = ["Text about out-of-stocks & suggest good replacements", "Pick the freshest produce & check dates",
            "Follow delivery notes exactly (leave at door / hand it)", "Keep cold items cold — use insulated bags",
            "Send a quick 'shopping now' / 'on my way' update", "Double-check the order before you leave the store"]
    for i, t in enumerate(tips):
        r = 13 + (i // 2)
        col = 2 if i % 2 == 0 else 4
        merge_set(ws, f"{get_column_letter(col)}{r}:{get_column_letter(col+1)}{r}", "✓  " + t, "td_left")


# ===========================================================================
# Analytics — Shopper Health Score + computed cells
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 3, 20, 12, 12, 2])
    luxe_header(ws, "G", "📈  ANALYTICS COMMAND CENTER",
                "Your shopping business by the numbers — take-home, health dimensions & the Shopper Health Score.")
    merge_set(ws, "B5:C5", "TAKE-HOME SNAPSHOT", "section")
    snap = [("Gross earnings", "=GrossEarn", '"$"#,##0', None),
            ("Business expenses", "=ExpenseTotal", '"$"#,##0', None),
            ("Net earnings", "=GrossEarn-ExpenseTotal", '"$"#,##0', "NetEarn"),
            ("Online hours", "=TotalHours", "0.0", None),
            ("Net $ / hour", "=IFERROR((GrossEarn-ExpenseTotal)/TotalHours,0)", '"$"#,##0.00', "PerHour"),
            ("Net $ / mile", "=IFERROR((GrossEarn-ExpenseTotal)/TotalMiles,0)", '"$"#,##0.00', "PerMile"),
            ("Batches", "=TotalBatches", "#,##0", None),
            ("Mileage deduction", "=TotalMiles*MileageRate", '"$"#,##0', None)]
    for i, (lab, fml, fmt, nm) in enumerate(snap):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if nm:
            cell_name(wb, nm, "Analytics", f"$C${r}")
        if lab in ("Net earnings", "Net $ / hour"):
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    merge_set(ws, "E5:G5", "SHOPPER HEALTH", "section_gold")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=5)
    metrics = [
        ("Net earnings vs goal", "=IFERROR(MIN(NetEarn/NetGoal,1),0)"),
        ("Net $/hour vs target", "=IFERROR(MIN(PerHour/HourlyTarget,1),0)"),
        ("Batches vs goal", "=IFERROR(MIN(TotalBatches/BatchGoal,1),0)"),
        ("Emergency fund", "=IFERROR(MIN(SavingsSaved/SavingsGoal,1),0)"),
        ("Shopping consistency", "=IFERROR(MIN(COUNT(ShiftMiles)/ShiftGoal,1),0)"),
        ("Tax reserve", "=IFERROR(MIN(TaxSetAside/TaxReserveGoal,1),0)"),
    ]
    hs = 7
    for i, (dim, fml) in enumerate(metrics):
        r = hs + i
        ws.cell(row=r, column=5, value=dim).style = "td_left"
        c = ws.cell(row=r, column=6, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=7, value=f'=IF(F{r}>=0.75,"Strong",IF(F{r}>=0.5,"Growing","Focus"))').style = "td"
        if i % 2:
            for c2 in range(5, 8):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(metrics) - 1
    cell_name(wb, "HealthRange", "Analytics", f"$F${hs}:$F${he}")
    ws.conditional_formatting.add(f"F{hs}:F{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "B15:G15", "BEST SHIFTS (BY EARNINGS)", "section_gold"); ws.row_dimensions[15].height = 22
    table_headers(ws, 16, ["Shift", "Earnings", "Hours", "$/hr", "Batches"], start_col=2)
    best = [("Sat big batches", 208, 6.5, 8), ("Fri Instacart", 203, 6.5, 8),
            ("Sat multi-app", 199, 6.5, 8), ("Weeknight multi", 180, 6.0, 7),
            ("Instacart AM", 171, 6.0, 7), ("Shipt evening", 170, 6.0, 7)]
    vs = 17
    for i, (title, earn, hrs, ba) in enumerate(best):
        r = vs + i
        ws.cell(row=r, column=2, value=title).style = "td_left"
        ce = ws.cell(row=r, column=3, value=earn); ce.style = "td"; ce.number_format = '"$"#,##0'
        ch = ws.cell(row=r, column=4, value=hrs); ch.style = "td"; ch.number_format = "0.0"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cp.style = "td"; cp.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=6, value=ba); cm.style = "td"; cm.number_format = "#,##0"
        if i % 2:
            for c2 in range(2, 7):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    ve = vs + len(best) - 1
    nrange(wb, "BestShift", "Analytics", "B", vs, ve)
    nrange(wb, "BestEarn", "Analytics", "C", vs, ve)
    ws.conditional_formatting.add(f"C{vs}:C{ve}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=208, color=PRIMARY, showValue=True))
    merge_set(ws, "B25:C25", "NET EARNINGS — 6 MONTHS", "section")
    ws.cell(row=26, column=2, value="Month").style = "th"; ws.cell(row=26, column=3, value="Net ($)").style = "th"
    months = _recent_months(6); vals = [1360, 1440, 1520, 1490, 1560, 1604]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 27 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "TrendMonth", "Analytics", "$B$27:$B$32")
    cell_name(wb, "TrendVal", "Analytics", "$C$27:$C$32")


# ===========================================================================
# Weekly Planner
# ===========================================================================
def build_planner(wb):
    ws = wb.create_sheet("Planner"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 16, 14, 14, 26, 2])
    luxe_header(ws, "G", "🗓  WEEKLY PLANNER",
                "Plan the week that hits your goal — target hours, stores & the earnings you need.")
    table_headers(ws, 4, ["Day", "Plan", "Target Hrs", "$ Goal", "Store / Focus"])
    days = [
        ("Monday", "Off", 0, 0, "Rest / car maintenance"),
        ("Tuesday", "Afternoon", 5, 100, "New Seasons + Pearl"),
        ("Wednesday", "Dinner prep", 6, 120, "Whole Foods"),
        ("Thursday", "Afternoon", 5, 105, "Fred Meyer"),
        ("Friday", "Evening", 6, 130, "Costco + downtown"),
        ("Saturday", "AM rush", 6.5, 160, "Costco doubles"),
        ("Sunday", "AM rush", 6, 150, "Weekend restocks"),
    ]
    start = L0
    for i, (d, plan, hrs, goal, zone) in enumerate(days):
        r = start + i
        ws.cell(row=r, column=1, value=d).style = "td_left"
        ws.cell(row=r, column=2, value=plan).style = "input"
        ch = ws.cell(row=r, column=3, value=hrs); ch.style = "input"; ch.number_format = "0.0"
        cg = ws.cell(row=r, column=4, value=goal); cg.style = "input"; cg.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=zone).style = "td_left"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(days) - 1
    trow = end + 1
    ws.cell(row=trow, column=1, value="WEEK TOTAL").style = "th"
    ws.cell(row=trow, column=2).style = "th"
    for col, fmt in ((3, "0.0"), (4, '"$"#,##0')):
        L = get_column_letter(col)
        c = ws.cell(row=trow, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt
    ws.cell(row=trow, column=5).style = "td"; ws.cell(row=trow, column=5).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


# ===========================================================================
# Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "📸  RECEIPTS & VEHICLE GALLERY",
                "Your paper trail — snap receipts, odometer & maintenance photos for tax time.")
    merge_set(ws, "B5:D5", "HOW TO ADD PHOTOS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Excel: Insert ▸ Pictures ▸ Place in Cell (or drag an image) into any framed box below. "
                      "Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"paste-link-here\"). "
                      "Caption each one — date, amount & category — so tax season is a five-minute job.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["Gas Receipt", "Maintenance", "Odometer", "Bags / Coolers", "Repair Invoice", "Car Wash"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add photo)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Date · amount · category…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🥕  INSTACART SHOPPER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Earnings, items, miles, taxes & savings — your whole shopping business, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("NET EARNINGS", "=NetEarn", "money"),
        ("GROSS EARNINGS", "=GrossEarn", "money"),
        ("TIPS", "=TotalTips", "money"),
        ("NET $ / HOUR", "=PerHour", "money2"),
        ("NET $ / MILE", "=PerMile", "money2"),
        ("MILES DRIVEN", "=TotalMiles", "num"),
    ]
    row2 = [
        ("ONLINE HOURS", "=TotalHours", "dec"),
        ("BATCHES", "=TotalBatches", "num"),
        ("EXPENSES", "=ExpenseTotal", "money"),
        ("TAX DEDUCTION", "=TotalMiles*MileageRate", "money"),
        ("SAVINGS", "=IFERROR(MIN(SavingsSaved/SavingsGoal,1),0)", "pct"),
        ("SHOPPER HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "EARNINGS & TAKE-HOME", "section_gold")
    ln = LineChart(); ln.title = "Net Earnings ($) — 6 mo"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=26, max_row=32), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=27, max_row=32)); ln.legend = None
    ws.add_chart(ln, "B12")
    d1 = DoughnutChart(); d1.title = "Earnings Mix"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Earnings"], min_col=3, min_row=6, max_row=9), titles_from_data=True)
    d1.set_categories(Reference(wb["Earnings"], min_col=2, min_row=7, max_row=9)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "BEST SHIFTS & EXPENSES", "section_gold")
    cb = BarChart(); cb.type = "bar"; cb.title = "Best Shifts by Earnings"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    cb.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); cb.legend = None
    ws.add_chart(cb, "B30")
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Expenses"], min_col=2, min_row=4, max_row=13), titles_from_data=True)
    eb.set_categories(Reference(wb["Expenses"], min_col=1, min_row=5, max_row=13)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Instacart Shopper Command Center™ — every batch, mile & dollar in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_batchlog(wb)
    build_earnings(wb); build_mileage(wb); build_fuel(wb); build_vehicle(wb)
    build_expenses(wb); build_budget(wb); build_tax(wb); build_savings(wb)
    build_boosts(wb); build_stores(wb); build_ratings(wb); build_analytics(wb)
    build_planner(wb); build_gallery(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Shopper Profile", "Batch Log", "Earnings", "Mileage", "Fuel Log",
             "Vehicle", "Expenses", "Budget", "Tax Center", "Savings", "Boosts", "Stores", "Ratings",
             "Analytics", "Planner", "Gallery", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Instacart_Shopper_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
