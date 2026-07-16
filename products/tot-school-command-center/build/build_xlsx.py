"""Build Tot-School Command Center™ — The Complete At-Home Tot-School System.

16 planning tabs (+ Settings) · a premium, play-based tot-school operating system
in Google Sheets & Excel for toddlers (roughly 18 months to 3 years). Tot-school
dashboard, a profile for every tot, a toddler milestones tracker, first words &
concepts, gentle weekly themes, a daily rhythm, a tot-tray & sensory activity
planner, a board-book log, a messy-play idea bank, outings & nature, attendance,
a portfolio of firsts, a supplies budget, a ready-for-preschool checklist and
goals — one dashboard. No pressure, all play.

Run: python3 build_xlsx.py   ->  ../Tot_School_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

DOMAINS = ["Gross Motor", "Fine Motor", "Language", "Social-Emotional", "Self-Help", "Cognitive"]
MILE_STATUS = ["Met", "Emerging", "Not Yet"]
TRAY_STATUS = ["Done", "Planned", "Skipped"]
THEME_STATUS = ["Done", "In Progress", "Upcoming"]
YESNO = ["Yes", "No"]
LOVED = ["Loved it", "Liked it", "Meh"]
SUPPLY_STATUS = ["Have", "Ordered", "Need"]
CATEGORIES = ["Colors", "Shapes", "Body Parts", "Animal Sounds", "Counting", "Social Words"]

# ---- the sample tot-school — the data that drives every KPI ----
TOTS = [
    ("Theo Bennett", "Toddler (2)", 2, "On the move · into everything · big feelings"),
    ("Nora Bennett", "Baby (14 mo)", 1, "New walker · loves peekaboo & board books"),
]

# Toddler milestones: (domain, milestone, tot, status)
MILESTONES = [
    ("Gross Motor", "Walks well on their own", "Theo", "Met"),
    ("Gross Motor", "Runs without falling", "Theo", "Met"),
    ("Gross Motor", "Kicks a ball", "Theo", "Met"),
    ("Gross Motor", "Climbs stairs holding the rail", "Theo", "Emerging"),
    ("Gross Motor", "Jumps with both feet", "Theo", "Emerging"),
    ("Gross Motor", "Stands & cruises furniture", "Nora", "Met"),
    ("Gross Motor", "Takes a few steps alone", "Nora", "Met"),
    ("Fine Motor", "Stacks 4+ blocks", "Theo", "Met"),
    ("Fine Motor", "Scribbles with a crayon", "Theo", "Met"),
    ("Fine Motor", "Turns board-book pages", "Theo", "Met"),
    ("Fine Motor", "Uses a spoon", "Theo", "Met"),
    ("Fine Motor", "Neat pincer grasp", "Nora", "Met"),
    ("Fine Motor", "Puts objects in a container", "Nora", "Met"),
    ("Fine Motor", "Stacks 2 blocks", "Nora", "Emerging"),
    ("Language", "Says 20+ words", "Theo", "Met"),
    ("Language", "Two-word phrases", "Theo", "Emerging"),
    ("Language", "Points to name body parts", "Theo", "Met"),
    ("Language", "Follows a simple direction", "Theo", "Met"),
    ("Language", "Waves bye-bye", "Nora", "Met"),
    ("Language", "Says mama / dada with meaning", "Nora", "Met"),
    ("Language", "Points to ask for things", "Nora", "Emerging"),
    ("Social-Emotional", "Plays alongside other children", "Theo", "Met"),
    ("Social-Emotional", "Shows affection", "Theo", "Met"),
    ("Social-Emotional", "Imitates grown-ups", "Theo", "Met"),
    ("Social-Emotional", "Beginning pretend play", "Theo", "Emerging"),
    ("Social-Emotional", "Enjoys peekaboo & games", "Nora", "Met"),
    ("Self-Help", "Drinks from an open cup", "Theo", "Met"),
    ("Self-Help", "Helps with dressing", "Theo", "Emerging"),
    ("Self-Help", "Feeds self finger foods", "Theo", "Met"),
    ("Self-Help", "Shows interest in the potty", "Theo", "Emerging"),
    ("Self-Help", "Finger-feeds independently", "Nora", "Met"),
    ("Cognitive", "Sorts a few shapes", "Theo", "Emerging"),
    ("Cognitive", "Explores cause & effect", "Theo", "Met"),
    ("Cognitive", "Points to pictures in books", "Theo", "Met"),
    ("Cognitive", "Finds a hidden object", "Nora", "Met"),
]

# First words & concepts: (concept, category, got it? Yes/No)
CONCEPTS = [
    ("Red", "Colors", "Yes"), ("Blue", "Colors", "Yes"), ("Yellow", "Colors", "Yes"), ("Green", "Colors", "No"),
    ("Circle", "Shapes", "Yes"), ("Square", "Shapes", "No"),
    ("Nose", "Body Parts", "Yes"), ("Eyes", "Body Parts", "Yes"), ("Mouth", "Body Parts", "Yes"),
    ("Ears", "Body Parts", "Yes"), ("Hands", "Body Parts", "Yes"), ("Feet", "Body Parts", "Yes"),
    ("Tummy", "Body Parts", "No"), ("Hair", "Body Parts", "Yes"),
    ("Dog — woof", "Animal Sounds", "Yes"), ("Cat — meow", "Animal Sounds", "Yes"),
    ("Cow — moo", "Animal Sounds", "Yes"), ("Duck — quack", "Animal Sounds", "Yes"),
    ("Sheep — baa", "Animal Sounds", "No"), ("Pig — oink", "Animal Sounds", "No"),
    ("Counts to 3", "Counting", "Yes"), ("Counts to 5", "Counting", "No"),
    ("Waves hi & bye", "Social Words", "Yes"), ("Signs / says 'more'", "Social Words", "Yes"),
]

# Gentle weekly themes: (theme, focus, status)
THEMES = [
    ("All About Me", "Faces, names & family", "Done"),
    ("Animals & Sounds", "Farm & pet sounds", "Done"),
    ("Colors", "Red, blue, yellow play", "Done"),
    ("Big & Little", "Sorting & size words", "Done"),
    ("Water Play", "Pouring & splashing", "Done"),
    ("Things That Go", "Cars, trucks & ramps", "In Progress"),
    ("On the Farm", "Animals & textures", "Upcoming"),
    ("Weather & Sky", "Sun, rain & clouds", "Upcoming"),
    ("Bugs & Crawlies", "Little creatures", "Upcoming"),
    ("Food & Cooking", "Play kitchen fun", "Upcoming"),
    ("Music & Sounds", "Shakers & drums", "Upcoming"),
    ("Nature Walk", "Leaves, rocks & sticks", "Upcoming"),
]

# Tot trays & activities: (tray / activity, area, status)
TRAYS = [
    ("Pom-pom drop into a bottle", "Fine Motor", "Done"),
    ("Water pouring station", "Practical Life", "Done"),
    ("Animal-sound matching", "Language", "Done"),
    ("Color-sorting cups", "Cognitive", "Done"),
    ("Rainbow rice sensory bin", "Sensory", "Done"),
    ("Sticker peel & place", "Fine Motor", "Done"),
    ("Play-dough squish & poke", "Fine Motor", "Done"),
    ("Ball ramp & roll", "Gross Motor", "Done"),
    ("Peekaboo scarves", "Social-Emotional", "Done"),
    ("Stacking rings", "Fine Motor", "Done"),
    ("Shape sorter", "Cognitive", "Planned"),
    ("Posting coins in a slot", "Fine Motor", "Planned"),
    ("Bubble-wrap stomp", "Gross Motor", "Planned"),
    ("Nature basket explore", "Sensory", "Planned"),
    ("Finger painting", "Sensory", "Planned"),
    ("Kinetic-sand tray", "Sensory", "Planned"),
]

# Board-book log: (title, author, times read, loved)
BOARDBOOKS = [
    ("Goodnight Moon", "Margaret Wise Brown", 22, "Loved it"),
    ("Brown Bear, Brown Bear", "Bill Martin Jr.", 18, "Loved it"),
    ("Dear Zoo", "Rod Campbell", 15, "Loved it"),
    ("The Very Hungry Caterpillar", "Eric Carle", 12, "Loved it"),
    ("Peek-a Who?", "Nina Laden", 20, "Loved it"),
    ("Moo, Baa, La La La!", "Sandra Boynton", 16, "Loved it"),
    ("Pat the Bunny", "Dorothy Kunhardt", 9, "Liked it"),
    ("Where's Spot?", "Eric Hill", 11, "Loved it"),
    ("Global Babies", "The Global Fund", 7, "Liked it"),
    ("Little Blue Truck", "Alice Schertle", 13, "Loved it"),
]

# Messy & sensory idea bank (no KPI): (activity, area, materials, builds)
SENSORY = [
    ("Cloud-dough dig", "Sensory", "Flour, oil, scoops, cups", "Scooping · texture"),
    ("Frozen-fruit teether tray", "Taste-safe", "Frozen fruit, mesh feeder", "Soothing · self-feed"),
    ("Taped-road cars", "Fine Motor", "Painter's tape, toy cars", "Tracking · play"),
    ("Water & sponge transfer", "Practical Life", "Two tubs, sponges", "Squeeze · pour"),
    ("Edible finger paint", "Taste-safe Art", "Yogurt + food color", "Color · mark-making"),
    ("Muffin-tin sort", "Cognitive", "Muffin tin, pom-poms", "Sorting · counting"),
    ("Contact-paper collage", "Art", "Contact paper, tissue", "Sticky exploration"),
    ("Sensory bag squish", "Sensory", "Zip bag, hair gel, beads", "Mess-free play"),
    ("Basket of textures", "Sensory", "Fabric scraps, brushes", "Touch words"),
    ("Pom-pom & tongs", "Fine Motor", "Pom-poms, kitchen tongs", "Grip · release"),
    ("Bubbles outside", "Gross Motor", "Bubble wand", "Chase · reach"),
    ("Ice & warm water", "Science", "Ice cubes, warm water", "Melt · cause-effect"),
]

# Outings & nature: (date offset, place, theme tie-in, memory)
OUTINGS = [
    (-34, "Petting zoo", "Animals", "Fed the goats · so excited"),
    (-24, "Library baby time", "Books & songs", "Weekly · loves the bubbles"),
    (-16, "Duck pond walk", "Water play", "Watched & quacked back"),
    (-8, "Playground", "Gross motor", "First time down the slide"),
    (6, "Farm visit", "On the farm", "Booked for theme week"),
    (14, "Nature trail stroll", "Nature walk", "Collected leaves & rocks"),
]

WEEKS = [(f"Week {i}", 3, "") for i in range(1, 25)]
DAYS_GOAL = 120
BOARD_BOOKS = 24
SUPPLIES_BUDGET = 300
WORDS = 45

# Ready-for-preschool checklist: (item, done)
READINESS = [
    ("Separates from caregiver briefly", True), ("Follows a simple daily routine", True),
    ("Uses some words to ask for things", True), ("Points to communicate", True),
    ("Feeds self finger foods", True), ("Drinks from an open cup", True),
    ("Shows interest in the potty", False), ("Plays near other children", True),
    ("Sits for a short story", True), ("Stacks a few blocks", True),
    ("Scribbles with a crayon", True), ("Waves & greets people", True),
    ("Names a few body parts", True), ("Follows a one-step direction", True),
    ("Sleeps on a toddler nap schedule", True), ("Beginning to say 'no' & choose", False),
]

# Goals: (goal, tot, progress)
GOALS = [
    ("Say 50 words", "Theo", 0.80),
    ("Start two-word phrases", "Theo", 0.55),
    ("Show potty interest", "Theo", 0.60),
    ("Fall asleep in own bed", "Theo", 0.70),
    ("Walk confidently", "Nora", 0.85),
    ("Wave & clap on cue", "Nora", 0.75),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your family details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "TOT-SCHOOL INPUTS", "section")
    controls = [
        ("Family Name", "The Bennett Family", None, "FamilyName"),
        ("School Year", "2026–2027", None, "SchoolYear"),
        ("Tots", 2, "0", "Tots"),
        ("Days Goal (year)", DAYS_GOAL, "0", "DaysGoal"),
        ("Words Spoken", WORDS, "0", "Words"),
        ("Board Books (rotation)", BOARD_BOOKS, "#,##0", "BoardBooks"),
        ("Supplies Budget", SUPPLIES_BUDGET, '"$"#,##0', "SuppliesBudget"),
        ("Preschool Start", "Fall 2028", None, "PreschoolDate"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Domain", DOMAINS, "DomainList"), ("F", "Milestone", MILE_STATUS, "MileStatusList"),
             ("G", "Tray", TRAY_STATUS, "TrayStatusList"), ("H", "Theme", THEME_STATUS, "ThemeStatusList"),
             ("I", "Category", CATEGORIES, "CategoryList"), ("J", "Loved?", LOVED, "LovedList"),
             ("K", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🧸  TOT-SCHOOL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  No pressure, all play — gentle rhythms, tot trays & a keepsake of these tiny years.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE TOT-SCHOOL, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Toddler milestones, first words & concepts, gentle weekly themes, a tot-tray & sensory activity "
                      "planner, board-book read-alouds, messy play, outings, a portfolio of firsts and a "
                      "ready-for-preschool checklist — all in ONE premium Google Sheets & Excel system. A profile for "
                      "every tot and a live dashboard that shows how they're growing. Play-based and pressure-free, it's "
                      "made for toddlers (roughly 18 months to 3 years) — follow their lead and enjoy the little years.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your family & tots.",
             "2.  Fill a Tot Profile for each little one — age, personality, what they love.",
             "3.  Pick this week's gentle Theme, then set out a few Tot Trays from the play bank.",
             "4.  As they grow, mark Milestones and check off First Words & Concepts.",
             "5.  Log board-book reads, outings & save a few portfolio 'firsts' as you go.",
             "6.  Watch the Dashboard track milestones, words, themes & a Growing Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for the Bennett family (Theo, 2, and Nora, 14 months) is included so you can see how it all "
               "connects — just type over it with your own. Milestones, words, themes, trays and the ready-for-preschool "
               "checklist roll up into a live Growing Score. Twelve matching printable pages (weekly theme plan, milestone "
               "checklist, tot-tray planner, sensory bank, portfolio of firsts & more) are included to print and keep. "
               "Every toddler grows on their own timeline — this is a gentle guide & keepsake, never a test or a diagnosis.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Slow down and play — the little years are the big ones. You've got this.", "section_gold")


# ===========================================================================
# 3 — Tot Profiles
# ===========================================================================
def build_profiles(wb):
    ws = wb.create_sheet("Tot Profiles"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 22, 4, 18, 18, 2])
    luxe_header(ws, "G", "🧸  TOT PROFILES",
                "A profile for every little one — age, personality, comforts & what they love.")
    row = 5
    for name, stage, age, note in TOTS:
        merge_set(ws, f"B{row}:F{row}", name, "section_gold"); ws.row_dimensions[row].height = 22
        row += 1
        fields = [("Stage", stage), ("Age", age), ("Personality", note), ("Loves", "—"),
                  ("Comfort item", "—"), ("Nap schedule", "—"), ("Words so far", "See First Words tab"),
                  ("Allergies / notes", "—"), ("Working on", "See Milestones tab"), ("Big goal", "See Goals tab")]
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 22; i += 2; row += 1
        row += 1
    ws.freeze_panes = "A5"


# ===========================================================================
# 4 — Milestones  (defines MileDomain / MileName / MileStatus)
# ===========================================================================
def build_milestones(wb):
    ws, start, end = build_log(
        wb, "Milestones", "🌱", "TODDLER MILESTONES",
        "The whole-child picture — gross & fine motor, language, social-emotional, self-help & cognitive.",
        ["Domain", "Milestone", "Tot", "Status"],
        MILESTONES, [18, 34, 14, 14], text_left={2}, reserved=45,
        validations=[("A", "DomainList"), ("D", "MileStatusList")])
    nrange(wb, "MileDomain", "Milestones", "A", start, end)
    nrange(wb, "MileName", "Milestones", "B", start, end)
    nrange(wb, "MileStatus", "Milestones", "D", start, end)
    cmap = {"Met": MINT_BG, "Emerging": WARN_BG, "Not Yet": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    sr = end + 3
    ws.cell(row=sr - 1, column=2, value="MET BY DOMAIN").style = "section_gold"
    for i, dom in enumerate(DOMAINS):
        r = sr + i
        ws.cell(row=r, column=2, value=dom).style = "td_left"
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(MileDomain,B{r},MileStatus,"Met")')
        c.style = "td"; c.number_format = "#,##0"
    cell_name(wb, "DomSumLabels", "Milestones", f"$B${sr}:$B${sr + len(DOMAINS) - 1}")
    cell_name(wb, "DomSumVals", "Milestones", f"$C${sr}:$C${sr + len(DOMAINS) - 1}")


# ===========================================================================
# 5 — First Words & Concepts  (defines ConceptDone / ConceptName)
# ===========================================================================
def build_concepts(wb):
    ws, start, end = build_log(
        wb, "First Words & Concepts", "💬", "FIRST WORDS & CONCEPTS",
        "The early basics — colors, shapes, body parts, animal sounds & first counting. Set the word count in Settings.",
        ["Concept", "Category", "Got It?"],
        CONCEPTS, [26, 18, 14], text_left={1}, reserved=30,
        validations=[("B", "CategoryList"), ("C", "YesNoList")])
    nrange(wb, "ConceptName", "First Words & Concepts", "A", start, end)
    nrange(wb, "ConceptCat", "First Words & Concepts", "B", start, end)
    nrange(wb, "ConceptDone", "First Words & Concepts", "C", start, end)
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))
    ws.column_dimensions["E"].width = 24; ws.column_dimensions["F"].width = 12
    ws.cell(row=4, column=5, value="WORD COUNT").style = "th"
    rows_w = [("Words spoken so far", "=Words"), ("Concepts got (of 24)", '=COUNTIF(ConceptDone,"Yes")'),
              ("New words this month", 8), ("Points to name", 6)]
    for i, (lab, val) in enumerate(rows_w):
        r = 5 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=val); c.style = "input" if isinstance(val, int) else "field_value"; c.number_format = "#,##0"
        if not isinstance(val, int):
            c.fill = fill(MINT_BG)


# ===========================================================================
# 6 — Weekly Themes  (defines ThemeStatus / ThemeName)
# ===========================================================================
def build_themes(wb):
    ws, start, end = build_log(
        wb, "Weekly Themes", "🗓", "GENTLE WEEKLY THEMES",
        "A soft theme a week — a few books, songs and trays. No worksheets, just playful exposure.",
        ["Theme", "Focus", "Status"],
        THEMES, [24, 28, 14], text_left={1, 2}, reserved=30,
        validations=[("C", "ThemeStatusList")])
    nrange(wb, "ThemeName", "Weekly Themes", "A", start, end)
    nrange(wb, "ThemeStatus", "Weekly Themes", "C", start, end)
    cmap = {"Done": MINT_BG, "In Progress": WARN_BG, "Upcoming": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 7 — Daily Rhythm
# ===========================================================================
def build_rhythm(wb):
    ws = wb.create_sheet("Daily Rhythm"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 26, 30, 2])
    luxe_header(ws, "D", "🌤  DAILY RHYTHM",
                "A loose, repeatable flow around meals & naps — toddlers thrive on rhythm, not a clock.")
    table_headers(ws, 4, ["When", "Block", "What we do"], start_col=2)
    rows = [
        ("Morning", "Wake & breakfast", "Cuddles, milk, free play as the day starts"),
        ("~9:00", "Tot time", "A short theme song, a book & one tot tray"),
        ("~9:45", "Snack", "Snack + a board-book read-aloud"),
        ("~10:15", "Outside / movement", "Backyard, stroller walk or the park"),
        ("~11:00", "Sensory or messy play", "The day's sensory bin or messy activity"),
        ("Noon", "Lunch", "Self-feeding practice & lots of mess"),
        ("~1:00", "Nap", "Quiet room, comfort item, rest"),
        ("Afternoon", "Open play", "Blocks, stacking, pretend, board books"),
        ("Late day", "Tidy & wind-down", "Clean-up song, bath, one more story"),
    ]
    start = L0
    for i, (when, block, what) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=when).style = "td"
        ws.cell(row=r, column=3, value=block).style = "td_left"
        ws.cell(row=r, column=4, value=what).style = "td_left"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"
    merge_set(ws, f"B{start+10}:D{start+10}",
              "Follow your tot: skip, swap or shorten any block. On hard days, connection beats the plan every time.",
              "section_gold")


# ===========================================================================
# 8 — Tot Trays & Activities  (defines TrayName / TrayStatus)
# ===========================================================================
def build_trays(wb):
    ws, start, end = build_log(
        wb, "Tot Trays", "🧩", "TOT TRAYS & ACTIVITIES",
        "Simple invitations to play — one tray at a time, tagged to a skill. Set it out, watch, check it off.",
        ["Tray / Activity", "Skill Area", "Status"],
        TRAYS, [30, 18, 14], text_left={1}, reserved=40,
        validations=[("B", "DomainList"), ("C", "TrayStatusList")])
    nrange(wb, "TrayName", "Tot Trays", "A", start, end)
    nrange(wb, "TrayStatus", "Tot Trays", "C", start, end)
    cmap = {"Done": MINT_BG, "Planned": WARN_BG, "Skipped": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 9 — Board-Book Log
# ===========================================================================
def build_books(wb):
    ws, start, end = build_log(
        wb, "Board-Book Log", "📖", "BOARD-BOOK LOG",
        "The heart of tot-school — every well-loved board book & how many times. Set the rotation size in Settings.",
        ["Title", "Author", "Times Read", "Loved?"],
        BOARDBOOKS, [30, 22, 14, 14], text_left={1, 2}, ints={3}, reserved=40,
        validations=[("D", "LovedList")])
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"Loved it"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=22, color=PRIMARY, showValue=True))


# ===========================================================================
# 10 — Messy & Sensory bank
# ===========================================================================
def build_sensory(wb):
    build_log(wb, "Sensory Play", "🖐", "MESSY & SENSORY PLAY BANK",
              "A grab-and-go idea bank for busy little hands — most from what you already have. Taste-safe noted.",
              ["Activity", "Area", "Materials", "Builds"],
              SENSORY, [26, 16, 30, 20], text_left={1, 3, 4}, reserved=30)


# ===========================================================================
# 11 — Outings & Nature  (defines TripName)
# ===========================================================================
def build_outings(wb):
    rows = [(dplus(off) if off >= 0 else dminus(-off), place, theme, memory)
            for (off, place, theme, memory) in OUTINGS]
    ws, start, end = build_log(
        wb, "Outings", "🌳", "OUTINGS & NATURE",
        "Little adventures — every outing, what it tied to & a memory to keep.",
        ["Date", "Place", "Theme Tie-in", "Memory"],
        rows, [14, 26, 18, 28], text_left={2, 4}, dates={1}, reserved=30)
    nrange(wb, "TripName", "Outings", "B", start, end)


# ===========================================================================
# 12 — Attendance & Days  (defines DaysDone)
# ===========================================================================
def build_attendance(wb):
    ws, start, end = build_log(
        wb, "Attendance", "📅", "TOT-SCHOOL DAYS",
        "Tot-school days by week — a light record of your gentle rhythm. Set your days goal in Settings.",
        ["Week", "Days", "Notes"],
        WEEKS, [16, 12, 40], text_left={3}, ints={2}, reserved=40)
    nrange(wb, "AttDays", "Attendance", "B", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL").style = "th"
    cd = ws.cell(row=tot, column=2, value="=SUM(AttDays)"); cd.style = "td"; cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = "#,##0"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cell_name(wb, "DaysDone", "Attendance", f"$B${tot}")


# ===========================================================================
# 13 — Portfolio of Firsts
# ===========================================================================
def build_portfolio(wb):
    ws = wb.create_sheet("Portfolio"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 26, 26, 2])
    luxe_header(ws, "E", "🖼  PORTFOLIO OF FIRSTS",
                "Snap the firsts & the sweet ordinary days — these tiny years go by so fast.")
    merge_set(ws, "B5:D5", "HOW TO ADD PHOTOS", "section_gold"); ws.row_dimensions[5].height = 22
    ws.merge_cells("B6:D7")
    ws["B6"].value = ("Google Sheets: click a box, Insert ▸ Image ▸ Image in cell, or paste =IMAGE(\"link\"). "
                      "Excel: Insert ▸ Pictures ▸ Place in Cell. Photograph a first step, first scribble, a messy-play "
                      "grin or a proud stack of blocks and caption it — tot, what it is & the date. A few a month is plenty.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(WARN_BG); ws["B6"].border = BOX
    for rr in (6, 7):
        ws.row_dimensions[rr].height = 26
        for cc in range(2, 5):
            ws.cell(row=rr, column=cc).fill = fill(WARN_BG)
    captions = ["First Steps", "First Scribble", "Messy-Play Grin", "Block Tower", "Favorite Book", "A Milestone Moment"]
    idx = 0
    for band in range(2):
        img_row = 9 + band * 6
        cap_row = img_row + 4
        ws.row_dimensions[img_row].height = 120
        for col in (2, 3, 4):
            ws.merge_cells(start_row=img_row, start_column=col, end_row=img_row + 3, end_column=col)
            ic = ws.cell(row=img_row, column=col, value=f"🖼\n{captions[idx]}\n(add photo)")
            ic.style = "imgbox"
            cap = ws.cell(row=cap_row, column=col, value="Tot · what it is · date…")
            cap.style = "td_left"; cap.fill = fill(SOFT_BG)
            ws.row_dimensions[cap_row].height = 30
            idx += 1


# ===========================================================================
# 14 — Supplies & Budget  (defines SuppliesSpent)
# ===========================================================================
def build_supplies(wb):
    ws = wb.create_sheet("Supplies"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 14, 14, 3, 2])
    luxe_header(ws, "E", "🧺  SUPPLIES & BUDGET",
                "Sensory fillers, board books & a few good toys — planned vs actual, so tot-school stays low-cost.")
    table_headers(ws, 4, ["Category", "Planned", "Actual", "Remaining"], start_col=2)
    rows = [
        ("Sensory bin fillers", 60, 48), ("Board books", 60, 44), ("Art & craft basics", 50, 39),
        ("Open-ended toys", 70, 58), ("Printables & laminating", 30, 22),
        ("Snacks for play (taste-safe)", 20, 14), ("Outings", 40, 26), ("Misc", 20, 8),
    ]
    start = L0
    for i, (cat, plan, actual) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=3, value=plan); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=4, value=actual); ca.style = "input"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(rows) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    for col in (3, 4, 5):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "SuppliesPlanTotal", "Supplies", f"$C${tot}")
    cell_name(wb, "SuppliesSpent", "Supplies", f"$D${tot}")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=70, color=PRIMARY, showValue=True))
    merge_set(ws, "B15:E15", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Total budget", "=SuppliesPlanTotal", '"$"#,##0'), ("Spent so far", "=SuppliesSpent", '"$"#,##0'),
             ("Remaining", "=SuppliesPlanTotal-SuppliesSpent", '"$"#,##0'),
             ("Per tot (avg)", "=IFERROR(SuppliesSpent/Tots,0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 16 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Remaining", "Per tot (avg)"):
            c.fill = fill(MINT_BG)


# ===========================================================================
# 15 — Ready for Preschool  (defines RdyName / RdyDone)
# ===========================================================================
def build_readiness(wb):
    rows = [(item, "Yes" if done else "No", "—") for (item, done) in READINESS]
    ws, start, end = build_log(
        wb, "Ready for Preschool", "🎒", "READY FOR PRESCHOOL",
        "The gentle 'getting there' list — the little independences preschools love. A guide, never a test.",
        ["Readiness Skill", "Ready?", "Notes"],
        rows, [40, 12, 30], text_left={1, 3}, reserved=20,
        validations=[("B", "YesNoList")])
    nrange(wb, "RdyName", "Ready for Preschool", "A", start, end)
    nrange(wb, "RdyDone", "Ready for Preschool", "B", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 16 — Goals & Little Wins  (defines GoalProgress)
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 14, 2])
    luxe_header(ws, "D", "🎯  GOALS & LITTLE WINS",
                "The wins that matter at this age — one gentle goal per tot, plus family goals.")
    table_headers(ws, 4, ["Goal", "Tot", "Progress"], start_col=2)
    start = L0
    for i, (goal, who, prog) in enumerate(GOALS):
        r = start + i
        ws.cell(row=r, column=2, value=goal).style = "td_left"
        ws.cell(row=r, column=3, value=who).style = "td"
        cp = ws.cell(row=r, column=4, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(GOALS) - 1
    nrange(wb, "GoalProgress", "Goals", "D", start, end)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 2 — Tot-School Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🧸  TOT-SCHOOL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Milestones, first words, gentle themes & sweet firsts — your whole tot year, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("TOTS", "=Tots", "num"),
        ("TOT-SCHOOL DAYS", "=DaysDone", "num"),
        ("WORDS SPOKEN", "=Words", "num"),
        ("MILESTONES MET", '=COUNTIF(MileStatus,"Met")', "num"),
        ("THEMES DONE", '=COUNTIF(ThemeStatus,"Done")', "num"),
        ("TRAYS DONE", '=IFERROR(COUNTIF(TrayStatus,"Done")/COUNTA(TrayName),0)', "pct"),
    ]
    row2 = [
        ("BOARD BOOKS", "=BoardBooks", "num"),
        ("OUTINGS", "=COUNTA(TripName)", "num"),
        ("SUPPLIES SPENT", "=SuppliesSpent", "money"),
        ("PRESCHOOL-READY", '=IFERROR(COUNTIF(RdyDone,"Yes")/COUNTA(RdyName),0)', "pct"),
        ("GOALS", "=IFERROR(AVERAGE(GoalProgress),0)", "pct"),
        ("GROWING SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "GROWING WELL?", "section_gold")
    merge_set(ws, "H11:M11", "MILESTONES MET BY DOMAIN", "section_gold")
    table_headers(ws, 12, ["Area", "Progress", "Status"], start_col=2)
    dims = [
        ("Milestones met", '=IFERROR(COUNTIF(MileStatus,"Met")/COUNTA(MileName),0)'),
        ("Words & concepts", '=IFERROR(COUNTIF(ConceptDone,"Yes")/COUNTA(ConceptName),0)'),
        ("Themes done", '=IFERROR(COUNTIF(ThemeStatus,"Done")/COUNTA(ThemeName),0)'),
        ("Trays & play done", '=IFERROR(COUNTIF(TrayStatus,"Done")/COUNTA(TrayName),0)'),
        ("Ready for preschool", '=IFERROR(COUNTIF(RdyDone,"Yes")/COUNTA(RdyName),0)'),
        ("Goals & little wins", "=IFERROR(AVERAGE(GoalProgress),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Thriving",IF(C{r}>=0.6,"Growing","Emerging"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    sk = wb["Milestones"]
    dn = wb.defined_names["DomSumVals"].value
    import re
    m = re.search(r"\$C\$(\d+):\$C\$(\d+)", dn)
    r0, r1 = int(m.group(1)), int(m.group(2))
    bar = BarChart(); bar.type = "bar"; bar.title = "Milestones Met by Domain"; bar.height = 7.4; bar.width = 8.6
    bar.add_data(Reference(sk, min_col=3, min_row=r0, max_row=r1), titles_from_data=False)
    bar.set_categories(Reference(sk, min_col=2, min_row=r0, max_row=r1))
    bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
    bar.legend = None
    ws.add_chart(bar, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Tot-School Command Center™ — no pressure, all play. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profiles(wb); build_milestones(wb)
    build_concepts(wb); build_themes(wb); build_rhythm(wb); build_trays(wb)
    build_books(wb); build_sensory(wb); build_outings(wb); build_attendance(wb)
    build_portfolio(wb); build_supplies(wb); build_readiness(wb); build_goals(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Tot Profiles", "Milestones", "First Words & Concepts",
             "Weekly Themes", "Daily Rhythm", "Tot Trays", "Board-Book Log", "Sensory Play",
             "Outings", "Attendance", "Portfolio", "Supplies", "Ready for Preschool", "Goals", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Tot_School_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
