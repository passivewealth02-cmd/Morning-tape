"""Build Bakery Command Center™ — The Complete Bakery Business System.

14 tabs · a premium bakery operating system in Google Sheets & Excel. Dashboard,
a recipe cost-per-batch engine, a product list with retail & wholesale margins, a
pre-orders board, wholesale accounts, a production plan, inventory & par, a waste
log, a sales log, ordering, cash & deposits and market days — one dashboard.
Cost every batch, price retail & wholesale, and bake more profit.

Run: python3 build_xlsx.py   ->  ../Bakery_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["g", "kg", "oz", "lb", "each", "batch", "dozen"]
CATS = ["Bread", "Pastry", "Cookie", "Muffin", "Cake", "Savory", "Seasonal"]
ORDER_STATUS = ["Confirmed", "Deposit", "Pending", "Done"]

TARGET_FC = 0.30
PREORDER_GOAL = 7

# Recipe engine — flagship Sourdough batch: (component, qty, unit, cost/unit)
SOURDOUGH = [
    ("Bread flour", 2000, "g", 0.0022),
    ("Whole wheat flour", 400, "g", 0.003),
    ("Water", 1400, "g", 0.0003),
    ("Sea salt", 40, "g", 0.005),
    ("Levain / starter", 400, "g", 0.0025),
    ("Oven / energy", 1, "batch", 2.38),
    ("Bags & labels", 8, "each", 0.40),
]
SOURDOUGH_YIELD = 8

# Products: (item, batch cost, yield, retail, wholesale, units/wk). Sourdough batch = SourdoughBatch.
PRODUCTS = [
    ("Sourdough Loaf", None, 8, 8.00, 5.00, 120),
    ("Butter Croissant", 21.60, 24, 4.25, 2.65, 300),
    ("Baguette", 7.20, 10, 4.50, 2.80, 90),
    ("Cinnamon Roll", 24.00, 12, 5.50, 3.50, 150),
    ("Choc Chip Cookie", 19.20, 48, 3.00, 1.85, 400),
    ("Blueberry Muffin", 21.60, 18, 4.00, 2.50, 160),
    ("Focaccia", 11.20, 8, 9.00, 5.75, 60),
    ("Everything Bagel", 9.60, 12, 3.25, 2.00, 200),
]

# Pre-orders: (customer, item, days out, price, deposit, status)
PREORDERS = [
    ("Sarah M.", "Birthday Cake (2-tier)", 3, 85, 40, "Confirmed"),
    ("James T.", "Croissant platter x3", 5, 45, 0, "Confirmed"),
    ("Office Corp", "Muffin boxes x10", 2, 60, 30, "Confirmed"),
    ("Wedding — Lee", "Dessert table", 12, 320, 150, "Deposit"),
    ("Mia R.", "Custom cookies x4dz", 6, 96, 48, "Confirmed"),
    ("Dan K.", "Focaccia catering", 4, 72, 0, "Pending"),
]

# Wholesale accounts: (account, item, weekly qty, unit price)
WHOLESALE = [
    ("Corner Cafe", "Croissants", 80, 2.65),
    ("Green Grocer", "Sourdough", 40, 5.00),
    ("Bistro 21", "Baguettes", 60, 2.80),
    ("Hotel Vera", "Assorted", 120, 3.50),
    ("Deli Fresh", "Bagels", 150, 2.00),
    ("Market Co-op", "Muffins", 90, 2.50),
    ("Campus Cafe", "Cookies", 200, 1.85),
    ("Wine Bar", "Focaccia", 50, 5.75),
]

# Production plan: (day, morning bake, afternoon bake)
PRODUCTION = [
    ("Monday", "Sourdough, Baguette", "Cookies, Muffins"),
    ("Tuesday", "Croissants, Focaccia", "Cinnamon Rolls"),
    ("Wednesday", "Sourdough, Bagels", "Cookies, Muffins"),
    ("Thursday", "Croissants, Baguette", "Pre-orders"),
    ("Friday", "Sourdough, Focaccia", "Cookies, Muffins"),
    ("Saturday", "Everything (market)", "Restock"),
    ("Sunday", "Rest / prep levain", "—"),
]

# Ingredient inventory & par: (item, par, on hand, unit)
INVENTORY = [
    ("Bread flour", 100, 48, "lb"), ("Whole wheat flour", 40, 18, "lb"), ("Butter", 60, 22, "lb"),
    ("Sugar", 50, 30, "lb"), ("Eggs", 30, 12, "dozen"), ("Chocolate chips", 25, 9, "lb"),
    ("Yeast", 10, 6, "lb"), ("Bags & boxes", 800, 320, "each"),
]

# Waste / day-old log: (item, reason, cost)
WASTE = [
    ("Day-old bread", "Donated", 120.00),
    ("Broken pastries", "Handling", 85.00),
    ("Failed bake", "Oven issue", 52.00),
    ("Trim & scraps", "Normal loss", 35.00),
]

# Sales log (7 days): (day, retail sales, units)
SALES = [
    ("Monday", 820, 190), ("Tuesday", 910, 210), ("Wednesday", 980, 226),
    ("Thursday", 1040, 240), ("Friday", 1180, 272), ("Saturday", 1465, 342), ("Sunday", 100, 0),
]

# Ordering / suppliers: (item, supplier, par order, cost)
ORDERING = [
    ("Bread flour (50 lb)", "Mill Direct", 4, 96.00),
    ("Butter (case)", "Dairy Co.", 2, 168.00),
    ("Eggs (15 dz)", "Farm Fresh", 2, 90.00),
    ("Chocolate chips (25 lb)", "SweetSupply", 1, 78.00),
    ("Bags & boxes", "PackCo", 1000, 140.00),
]

# Cash & deposits (7 days): (day, cash, card, deposits)
CASHDEP = [
    ("Monday", 210, 610, 0), ("Tuesday", 240, 670, 40), ("Wednesday", 250, 730, 30),
    ("Thursday", 260, 780, 0), ("Friday", 300, 880, 150), ("Saturday", 420, 1045, 0),
    ("Sunday", 40, 60, 48),
]

# Market days: (market, days out, sales, fee)
MARKETS = [
    ("Saturday Farmers Market", 2, 1465, 50),
    ("Sunday Artisan Fair", 9, 980, 40),
    ("Holiday Market", 20, 2200, 120),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dplus(n):
    return dt.date.today() + dt.timedelta(days=abs(n))


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your bakery, targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR BAKERY", "section")
    controls = [
        ("Bakery name", "Rise & Crumb Bakery", None, "Bakery"),
        ("Owner / baker", "Nora", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Margin goal ($/unit)", 3, '"$"#,##0', "MarginGoal"),
        ("Pre-order goal / wk", PREORDER_GOAL, "0", "PreOrderGoal"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Category", CATS, "CatList"),
             ("G", "Order Status", ORDER_STATUS, "StatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🥖  BAKERY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every batch, price retail & wholesale, and bake more profit.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE BAKERY, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Cost each recipe by the batch, then let the sheet divide by yield for a true cost per loaf, "
                      "cookie or croissant. See each product's food-cost % and both retail and wholesale margin, keep "
                      "a pre-orders board for custom cakes, manage wholesale accounts, and run a production plan, "
                      "inventory, waste, sales, ordering, cash and market days — all in ONE premium Google Sheets & "
                      "Excel system built for bakers.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your bakery name & target food-cost %.",
             "2.  Cost a recipe by the batch; the sheet divides by yield for cost per unit.",
             "3.  Fill the Product List — retail & wholesale price, margin & food-cost %.",
             "4.  Keep the Pre-Orders board & Wholesale Accounts current.",
             "5.  Run the Production Plan, Inventory, Waste, Sales, Ordering & Markets.",
             "6.  Check the Dashboard: revenue, margins, pre-orders & a Bakery Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional bakery (Rise & Crumb) is included so you can see how it all connects — just "
               "type over it with your own recipes and prices. Costing by the batch and dividing by yield is the "
               "secret to real bakery pricing, and it rolls up into a live Bakery Score. Twelve matching printable "
               "pages (recipe cost card, pre-order form, production plan, waste log & more) are included for the "
               "kitchen. This is a business tool, not financial or accounting advice — confirm figures with your books.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Cost by the batch, divide by yield — that's how bakeries actually price.", "section_gold")


# ===========================================================================
def build_recipe(wb):
    ws = wb.create_sheet("Recipe Costing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 10, 10, 14, 14, 2])
    luxe_header(ws, "F", "🧾  RECIPE COSTING",
                "Cost a recipe by the batch, then divide by yield — the true cost per loaf, cookie or roll.")
    merge_set(ws, "B5:C5", "RECIPE", "section_gold")
    ws.cell(row=5, column=4, value="Sourdough Loaf").font = Font(bold=True, color=PRIMARY)
    ws.cell(row=5, column=5, value="Yield →").font = Font(italic=True, color=ACCENT)
    ws.cell(row=5, column=6, value=SOURDOUGH_YIELD).font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Component", "Qty", "Unit", "Cost/Unit", "Ext. Cost"], start_col=2)
    start = 7
    for i, (comp, qty, unit, cu) in enumerate(SOURDOUGH):
        r = start + i
        ws.cell(row=r, column=2, value=comp).style = "td_left"
        cq = ws.cell(row=r, column=3, value=qty); cq.style = "input"; cq.number_format = "0.###"
        ws.cell(row=r, column=4, value=unit).style = "td"
        cc = ws.cell(row=r, column=5, value=cu); cc.style = "input"; cc.number_format = '"$"#,##0.0000'
        ce = ws.cell(row=r, column=6, value=f"=C{r}*E{r}"); ce.style = "td"; ce.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SOURDOUGH) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="BATCH COST").style = "th"
    for c in range(3, 6):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0.00'
    cell_name(wb, "SourdoughBatch", "Recipe Costing", f"$F${tot}")
    cs = tot + 1
    merge_set(ws, f"B{cs}:E{cs}", "COST PER UNIT  (batch ÷ yield)", "section")
    cps = ws.cell(row=cs, column=6, value=f"=IFERROR(F{tot}/F5,0)"); cps.style = "field_value"; cps.font = Font(bold=True, size=13, color=PRIMARY); cps.number_format = '"$"#,##0.00'; cps.fill = fill(MINT_BG)
    ws.cell(row=cs + 2, column=2, value="Copy this build for every recipe — bread, pastry, cookies, cakes.").style = "section"


def build_products(wb):
    ws = wb.create_sheet("Product List"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 12, 10, 12, 12, 12, 12, 12, 10, 12, 2])
    luxe_header(ws, "I", "🍞  PRODUCT LIST",
                "Every product's unit cost, retail & wholesale price, margin & food-cost %.")
    table_headers(ws, 4, ["Product", "Batch Cost", "Yield", "Unit Cost", "Retail", "Wholesale", "Retail Margin", "Food %"], start_col=2)
    start = L0
    for i, (item, bc, y, retail, whole, units) in enumerate(PRODUCTS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cb = ws.cell(row=r, column=3, value="=SourdoughBatch" if bc is None else bc); cb.style = "input"; cb.number_format = '"$"#,##0.00'
        cy = ws.cell(row=r, column=4, value=y); cy.style = "input"; cy.number_format = "#,##0"
        cu = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cu.style = "td"; cu.number_format = '"$"#,##0.00'
        cr = ws.cell(row=r, column=6, value=retail); cr.style = "input"; cr.number_format = '"$"#,##0.00'
        cw = ws.cell(row=r, column=7, value=whole); cw.style = "input"; cw.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=8, value=f"=F{r}-E{r}"); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        cf = ws.cell(row=r, column=9, value=f"=IFERROR(E{r}/F{r},0)"); cf.style = "td"; cf.number_format = "0%"
        if i % 2:
            for c in range(2, 10):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRODUCTS) - 1
    # units/wk (col J) and weekly revenue (col K) helpers
    ws.cell(row=4, column=10, value="Units/wk").style = "th"
    ws.cell(row=4, column=11, value="Wk Rev").style = "th"
    for i, (item, bc, y, retail, whole, units) in enumerate(PRODUCTS):
        r = start + i
        cu = ws.cell(row=r, column=10, value=units); cu.style = "td"; cu.number_format = "#,##0"
        cr = ws.cell(row=r, column=11, value=f"=F{r}*J{r}"); cr.style = "td"; cr.number_format = '"$"#,##0'
    nrange(wb, "ProdItem", "Product List", "B", start, end)
    nrange(wb, "ProdUnitCost", "Product List", "E", start, end)
    nrange(wb, "ProdRetail", "Product List", "F", start, end)
    nrange(wb, "ProdMargin", "Product List", "H", start, end)
    nrange(wb, "ProdWkUnits", "Product List", "J", start, end)
    nrange(wb, "ProdRev", "Product List", "K", start, end)
    br = end + 2
    ws.cell(row=br, column=2, value="Retail revenue / wk").style = "field_label"
    c = ws.cell(row=br, column=6, value="=SUMPRODUCT(ProdRetail,ProdWkUnits)"); c.style = "field_value"; c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    cell_name(wb, "RetailRev", "Product List", f"$F${br}")
    ws.cell(row=br + 1, column=2, value="Overall food-cost %").style = "field_label"
    c2 = ws.cell(row=br + 1, column=6, value="=IFERROR(SUMPRODUCT(ProdUnitCost,ProdWkUnits)/SUMPRODUCT(ProdRetail,ProdWkUnits),0)"); c2.style = "field_value"; c2.number_format = "0%"; c2.fill = fill(MINT_BG)
    cell_name(wb, "FoodCostPct", "Product List", f"$F${br+1}")
    ws.conditional_formatting.add(f"I{start}:I{end}",
        ColorScaleRule(start_type="num", start_value=0.12, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.30, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.45, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


def build_preorders(wb):
    ws = wb.create_sheet("Pre-Orders"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 26, 12, 12, 12, 14, 2])
    luxe_header(ws, "G", "🎂  PRE-ORDERS",
                "The custom & special-order board — cakes, platters & catering, with deposits.")
    table_headers(ws, 4, ["Customer", "Order", "Due", "Price", "Deposit", "Status"], start_col=2)
    start = L0
    for i, (cust, item, days, price, dep, status) in enumerate(PREORDERS):
        r = start + i
        ws.cell(row=r, column=2, value=cust).style = "td_left"
        ws.cell(row=r, column=3, value=item).style = "td_left"
        cd = ws.cell(row=r, column=4, value=dplus(days)); cd.style = "input"; cd.number_format = "mm/dd"
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cdp = ws.cell(row=r, column=6, value=dep); cdp.style = "input"; cdp.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=7, value=status); cs.style = "td"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PREORDERS) - 1
    nrange(wb, "PreOrderName", "Pre-Orders", "B", start, end)
    nrange(wb, "PreOrderPrice", "Pre-Orders", "E", start, end)
    nrange(wb, "PreOrderStatus", "Pre-Orders", "G", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="PRE-ORDER VALUE").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(PreOrderPrice)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    for c2 in range(6, 8):
        ws.cell(row=tot, column=c2).style = "td"; ws.cell(row=tot, column=c2).fill = fill(SURFACE)
    cell_name(wb, "PreOrderValue", "Pre-Orders", f"$E${tot}")
    cmap = {"Confirmed": MINT_BG, "Deposit": WARN_BG, "Pending": SURFACE, "Done": MINT_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    ws.freeze_panes = "A5"


def build_wholesale(wb):
    ws = wb.create_sheet("Wholesale"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 20, 12, 14, 14, 2])
    luxe_header(ws, "F", "🏪  WHOLESALE ACCOUNTS",
                "Standing wholesale orders — steady revenue that fills the ovens between rushes.")
    table_headers(ws, 4, ["Account", "Item", "Weekly Qty", "Unit Price", "Weekly Rev"], start_col=2)
    start = L0
    for i, (acct, item, qty, price) in enumerate(WHOLESALE):
        r = start + i
        ws.cell(row=r, column=2, value=acct).style = "td_left"
        ws.cell(row=r, column=3, value=item).style = "td_left"
        cq = ws.cell(row=r, column=4, value=qty); cq.style = "input"; cq.number_format = "#,##0"
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cr = ws.cell(row=r, column=6, value=f"=D{r}*E{r}"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(WHOLESALE) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WHOLESALE REV / WK").style = "th"
    for c in range(3, 6):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "WholeRev", "Wholesale", f"$F${tot}")
    ws.conditional_formatting.add(f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


def build_production(wb):
    ws = wb.create_sheet("Production Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 30, 30, 2])
    luxe_header(ws, "D", "🗓  PRODUCTION PLAN",
                "The weekly bake schedule — morning & afternoon, so nothing's missed.")
    table_headers(ws, 4, ["Day", "Morning Bake", "Afternoon Bake"], start_col=2)
    start = L0
    for i, (day, am, pm) in enumerate(PRODUCTION):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"; ws.cell(row=r, column=2).font = Font(bold=True, color=PRIMARY)
        ws.cell(row=r, column=3, value=am).style = "td_left"
        ws.cell(row=r, column=4, value=pm).style = "td_left"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


def build_inventory(wb):
    ws = wb.create_sheet("Inventory & Par"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 12, 10, 14, 2])
    luxe_header(ws, "F", "📦  INVENTORY & PAR",
                "Par vs on hand — order flour & butter before the Saturday bake, not during it.")
    table_headers(ws, 4, ["Ingredient", "Par", "On Hand", "Unit", "To Order"], start_col=2)
    start = L0
    for i, (item, par, oh, unit) in enumerate(INVENTORY):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=par); cp.style = "input"; cp.number_format = "#,##0"
        co = ws.cell(row=r, column=4, value=oh); co.style = "input"; co.number_format = "#,##0"
        ws.cell(row=r, column=5, value=unit).style = "td"
        cb = ws.cell(row=r, column=6, value=f"=MAX(C{r}-D{r},0)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.number_format = "#,##0"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVENTORY) - 1
    for c in range(2, 7):
        add_dv(ws, f"E{start}:E{end}", "UnitList")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=[f"C{start}*0.5"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE & DAY-OLD LOG",
        "Day-old, broken & failed bakes — track it, donate it, and shrink it over time.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="WEEKLY WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_sales(wb):
    ws = wb.create_sheet("Sales Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 2])
    luxe_header(ws, "E", "📈  SALES LOG",
                "Retail sales & units by day — the pulse of the counter.")
    table_headers(ws, 4, ["Day", "Retail Sales", "Units", "Avg Sale"], start_col=2)
    start = L0
    for i, (day, sales, units) in enumerate(SALES):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        cu = ws.cell(row=r, column=4, value=units); cu.style = "input"; cu.number_format = "#,##0"
        ca = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); ca.style = "td"; ca.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SALES) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    cs = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cu = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "#,##0"
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    ws.add_chart(_barchart(ws, "Sales by Day", start, end, 3, 2), "G4")
    ws.freeze_panes = "A5"


def build_ordering(wb):
    ws, start, end = build_log(
        wb, "Ordering", "🚚", "ORDERING & SUPPLIERS",
        "Your standing order — par quantities & cost, by supplier.",
        ["Item", "Supplier", "Par Order", "Cost"],
        ORDERING, [2, 26, 20, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)
    nrange(wb, "OrderCost", "Ordering", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ORDER TOTAL").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(OrderCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_cash(wb):
    ws = wb.create_sheet("Cash & Deposits"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 14, 2])
    luxe_header(ws, "F", "💵  CASH & DEPOSITS",
                "Reconcile the till daily — cash, card & pre-order deposits.")
    table_headers(ws, 4, ["Day", "Cash", "Card", "Deposits", "Total"], start_col=2)
    start = L0
    for i, (day, cash, card, dep) in enumerate(CASHDEP):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cash); cc.style = "input"; cc.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=4, value=card); cd.style = "input"; cd.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=dep); cp.style = "input"; cp.number_format = '"$"#,##0'
        cto = ws.cell(row=r, column=6, value=f"=C{r}+D{r}+E{r}"); cto.style = "td"; cto.font = Font(bold=True, color=PRIMARY); cto.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CASHDEP) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    for col in (3, 4, 5, 6):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


def build_markets(wb):
    ws = wb.create_sheet("Market Days"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 14, 12, 14, 2])
    luxe_header(ws, "F", "⛺  MARKET DAYS",
                "Farmers markets & fairs — sales, booth fee & the net you actually take home.")
    table_headers(ws, 4, ["Market", "Date", "Sales", "Booth Fee", "Net"], start_col=2)
    start = L0
    for i, (name, days, sales, fee) in enumerate(MARKETS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cd = ws.cell(row=r, column=3, value=dplus(days)); cd.style = "input"; cd.number_format = "mm/dd"
        cs = ws.cell(row=r, column=4, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        cf = ws.cell(row=r, column=5, value=fee); cf.style = "input"; cf.number_format = '"$"#,##0'
        cn = ws.cell(row=r, column=6, value=f"=D{r}-E{r}"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🥖  BAKERY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Revenue, retail & wholesale margins, pre-orders & a Bakery Score — your whole bakery, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("PRODUCTS", "=COUNTA(ProdItem)", "num"),
        ("AVG UNIT COST", "=AVERAGE(ProdUnitCost)", "money2"),
        ("AVG RETAIL", "=AVERAGE(ProdRetail)", "money2"),
        ("AVG MARGIN", "=AVERAGE(ProdMargin)", "money2"),
        ("FOOD COST", "=FoodCostPct", "pct"),
        ("TOP SELLER", "=INDEX(ProdItem,MATCH(MAX(ProdRev),ProdRev,0))", "text"),
    ]
    row2 = [
        ("WEEKLY REVENUE", "=RetailRev+WholeRev", "money"),
        ("WEEKLY UNITS", "=SUM(ProdWkUnits)", "num"),
        ("PRE-ORDERS", "=COUNTA(PreOrderName)", "num"),
        ("WHOLESALE REV", "=WholeRev", "money"),
        ("WASTE %", "=IFERROR(WasteTotal/RetailRev,0)", "pct"),
        ("BAKERY SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "BAKERY HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SALES BY DAY", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Food cost on target", "=IFERROR(MIN(TargetFC/FoodCostPct,1),0)"),
        ("Margin per unit", "=IFERROR(MIN(AVERAGE(ProdMargin)/MarginGoal,1),0)"),
        ("Products fully costed", "=IFERROR(COUNTIF(ProdUnitCost,\">0\")/COUNTA(ProdItem),0)"),
        ("Pre-orders vs goal", "=IFERROR(MIN(COUNTA(PreOrderName)/PreOrderGoal,1),0)"),
        ("Low waste", "=IFERROR(1-MIN((WasteTotal/RetailRev)/0.08,1),0)"),
        ("Gross margin", "=IFERROR(MIN((1-FoodCostPct)/0.7,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    sl = wb["Sales Log"]
    ch = BarChart(); ch.type = "col"; ch.title = "Sales by Day"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(sl, min_col=3, min_row=5, max_row=4 + len(SALES)), titles_from_data=False)
    ch.set_categories(Reference(sl, min_col=2, min_row=5, max_row=4 + len(SALES))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Bakery Command Center™ — cost every batch, price retail & wholesale, bake more profit.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_recipe(wb); build_products(wb)
    build_preorders(wb); build_wholesale(wb); build_production(wb); build_inventory(wb)
    build_waste(wb); build_sales(wb); build_ordering(wb); build_cash(wb)
    build_markets(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Recipe Costing", "Product List", "Pre-Orders", "Wholesale",
             "Production Plan", "Inventory & Par", "Waste Log", "Sales Log", "Ordering",
             "Cash & Deposits", "Market Days", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Bakery_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
