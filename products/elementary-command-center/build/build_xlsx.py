"""Build Elementary Homeschool Command Center™ — The Complete K–5 Learning System.

16 tabs (+ Settings) · a premium elementary (grades K–5) homeschool operating
system in Google Sheets & Excel. Dashboard, a profile per student, weekly lesson
plan, subject-mastery tracker, reading log & levels, sight words & spelling, math
facts fluency, assignments & grades, report card, attendance, curriculum, field
trips, habits & character, awards and a book list — one dashboard.

Run: python3 build_xlsx.py   ->  ../Elementary_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

SUBJECTS = ["Math", "Reading", "Writing", "Spelling", "Science", "Social Studies", "Art", "PE"]
MASTERY = ["Mastered", "Practicing", "Not Yet"]
GRADED = ["Graded", "Turned In", "Not Yet"]
STATUS = ["Done", "In Progress", "Not Started"]
ACQUIRED = ["Have", "Ordered", "Need"]
YESNO = ["Yes", "No"]
GRADES = ["K", "1", "2", "3", "4", "5"]

# ---- sample family — the data that drives every KPI ----
STUDENTS = [
    ("Lucy Bennett", "3", 8, "Level P reader", "Loves reading & science · strong writer"),
    ("Jack Bennett", "1", 6, "Level E reader", "Hands-on · learning to read · loves math"),
]
DAYS_REQUIRED = 180
BOOKS_READ = 52
READING_GOAL = 60
SIGHT_WORDS_KNOWN = 210
SIGHT_WORD_GOAL = 220
SUBJECT_COUNT = 8

# Subject mastery: (subject, skill, student, status)
MASTERY_ROWS = [
    ("Math", "Add & subtract within 1,000", "Lucy", "Mastered"),
    ("Math", "Multiplication facts to 12", "Lucy", "Practicing"),
    ("Math", "Fractions on a number line", "Lucy", "Practicing"),
    ("Math", "Tell time to the minute", "Lucy", "Mastered"),
    ("Math", "Count & add within 20", "Jack", "Mastered"),
    ("Math", "Place value to 100", "Jack", "Practicing"),
    ("Reading", "Read grade-level text fluently", "Lucy", "Mastered"),
    ("Reading", "Find the main idea", "Lucy", "Mastered"),
    ("Reading", "Make inferences", "Lucy", "Practicing"),
    ("Reading", "Blend CVC words", "Jack", "Mastered"),
    ("Reading", "Read short vowels & digraphs", "Jack", "Practicing"),
    ("Writing", "Write a paragraph with details", "Lucy", "Mastered"),
    ("Writing", "Use capitals & end marks", "Lucy", "Mastered"),
    ("Writing", "Write a complete sentence", "Jack", "Practicing"),
    ("Spelling", "Grade-3 spelling list", "Lucy", "Practicing"),
    ("Spelling", "Short-vowel word families", "Jack", "Mastered"),
    ("Science", "Life cycles & habitats", "Lucy", "Mastered"),
    ("Science", "States of matter", "Lucy", "Practicing"),
    ("Science", "Weather & seasons", "Jack", "Mastered"),
    ("Social Studies", "Community & maps", "Lucy", "Mastered"),
    ("Social Studies", "Past & present", "Jack", "Practicing"),
    ("Art", "Color mixing & shapes", "Lucy", "Mastered"),
    ("Art", "Cutting & gluing control", "Jack", "Mastered"),
    ("PE", "Skip, gallop & jump rope", "Lucy", "Mastered"),
    ("PE", "Throw & catch a ball", "Jack", "Practicing"),
    ("Math", "Measure length & mass", "Lucy", "Not Yet"),
    ("Writing", "Spell high-frequency words", "Jack", "Not Yet"),
    ("Reading", "Retell a story in order", "Jack", "Mastered"),
]

# Math facts fluency: (operation, student, percent fluent)
MATHFACTS = [
    ("Addition (0–12)", "Lucy", 1.00),
    ("Subtraction (0–12)", "Lucy", 0.95),
    ("Multiplication (0–12)", "Lucy", 0.60),
    ("Division (0–12)", "Lucy", 0.30),
    ("Addition (0–10)", "Jack", 0.70),
    ("Subtraction (0–10)", "Jack", 0.45),
]

# Sight words & spelling progress: (list, student, known, total)
SIGHTWORDS = [
    ("Dolch Pre-Primer (40)", "Jack", 38, 40),
    ("Dolch Primer (52)", "Jack", 40, 52),
    ("Fry First 100", "Lucy", 100, 100),
    ("Fry Second 100", "Lucy", 92, 100),
    ("Grade-3 spelling units", "Lucy", 18, 24),
    ("Short-vowel families", "Jack", 12, 15),
]

# Reading log: (student, title, level, finished?)
READING = [
    ("Lucy", "Charlotte's Web", "R", "Yes"), ("Lucy", "The BFG", "R", "Yes"),
    ("Lucy", "Ramona the Pest", "O", "Yes"), ("Lucy", "The Magic Tree House #1", "M", "Yes"),
    ("Jack", "Bob Books Set 2", "C", "Yes"), ("Jack", "Frog and Toad", "K", "In Progress"),
    ("Jack", "Elephant & Piggie", "G", "Yes"), ("Lucy", "Because of Winn-Dixie", "R", "In Progress"),
]

# Assignments & grades: (date offset, student, subject, assignment, grade, status)
ASSIGNMENTS = [
    (12, "Lucy", "Math", "Ch. 6 test", "95%", "Graded"), (11, "Lucy", "Writing", "Animal report", "A", "Graded"),
    (10, "Lucy", "Reading", "Comprehension quiz", "92%", "Graded"), (9, "Jack", "Math", "Add within 20 quiz", "S+", "Graded"),
    (8, "Lucy", "Spelling", "Unit 12 test", "88%", "Graded"), (7, "Jack", "Reading", "Sight-word check", "S", "Graded"),
    (6, "Lucy", "Science", "Life-cycle project", "A", "Graded"), (5, "Jack", "Writing", "Sentence practice", "S", "Graded"),
    (4, "Lucy", "Math", "Multiplication drill", "80%", "Graded"), (3, "Jack", "Math", "Place value sort", "S+", "Graded"),
    (2, "Lucy", "Social Studies", "Map quiz", "A-", "Graded"), (1, "Lucy", "Spelling", "Unit 13 test", "", "Turned In"),
    (0, "Jack", "Reading", "Fluency check", "", "Not Yet"), (0, "Lucy", "Science", "Matter worksheet", "", "Not Yet"),
]

# Habits & character: (habit, done today, streak)
HABITS = [
    ("Made the bed", "Yes", 12), ("Morning chores", "Yes", 8), ("Read 20 minutes", "Yes", 24),
    ("Kind words / helped a sibling", "Yes", 6), ("Screen-time limit kept", "No", 0),
    ("Tidied school space", "Yes", 9), ("Practiced piano", "Yes", 4),
]

# Field trips & enrichment: (date offset, place, subject tie-in, who)
FIELDTRIPS = [
    (-30, "Science museum", "Science", "Both"), (-22, "Public library program", "Reading", "Both"),
    (-14, "Local farm", "Science", "Both"), (-6, "Art class at the co-op", "Art", "Lucy"),
    (8, "Children's theater", "Reading", "Both"), (16, "Nature-center hike", "Science", "Both"),
]

# Curriculum: (subject, resource, student, cost, status)
CURRICULUM = [
    ("Math", "Singapore 3A/3B", "Lucy", 90, "Have"), ("Math", "Math-U-See Alpha", "Jack", 85, "Have"),
    ("Reading", "The Good & the Beautiful", "Both", 0, "Have"), ("Writing", "IEW-B", "Lucy", 60, "Have"),
    ("Spelling", "All About Spelling 2", "Lucy", 40, "Have"), ("Science", "Nancy Larson Science", "Both", 120, "Have"),
    ("Social Studies", "Story of the World 1", "Both", 40, "Have"), ("Art", "Art supply box", "Both", 45, "Have"),
]

# Awards & milestones: (award, student, date offset)
AWARDS = [
    ("Read 50 books!", "Lucy", -10), ("Mastered addition facts", "Jack", -18),
    ("Finished Singapore 3A", "Lucy", -25), ("First chapter book read alone", "Jack", -40),
    ("Perfect spelling test", "Lucy", -5),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 14 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "date": "mmm d", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your family details & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "SCHOOL INPUTS", "section")
    controls = [
        ("Family Name", "The Bennett Family", None, "FamilyName"),
        ("School Year", "2026–2027", None, "SchoolYear"),
        ("Students", 2, "0", "Students"),
        ("Subjects", SUBJECT_COUNT, "0", "Subjects"),
        ("Required Days", DAYS_REQUIRED, "0", "DaysRequired"),
        ("Reading Goal (books)", READING_GOAL, "0", "ReadingGoal"),
        ("Books Read", BOOKS_READ, "0", "BooksRead"),
        ("Sight Words Known", SIGHT_WORDS_KNOWN, "0", "SightWordsKnown"),
        ("Sight-Word Goal", SIGHT_WORD_GOAL, "0", "SightWordGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Subject", SUBJECTS, "SubjectList"), ("F", "Mastery", MASTERY, "MasteryList"),
             ("G", "Graded", GRADED, "GradedList"), ("H", "Status", STATUS, "StatusList"),
             ("I", "Curriculum", ACQUIRED, "AcquiredList"), ("J", "Grade", GRADES, "GradeList"),
             ("K", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# 1 — Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ✏  ELEMENTARY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Grades K–5, one calm system — plan the week, track mastery, celebrate every win.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE ELEMENTARY YEAR, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A weekly lesson plan, a subject-mastery tracker, reading log & levels, sight words & spelling, "
                      "math-facts fluency, assignments & grades, a print-ready report card, attendance, curriculum, "
                      "field trips, habits & character, awards and a book list — all in ONE premium Google Sheets & "
                      "Excel system. A profile for every child and a live dashboard that shows exactly where each one is "
                      "thriving and what to work on next. Built for the elementary years (grades K–5), any method.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your family, students & goals.",
             "2.  Fill a Student Profile for each child — grade, reading level, strengths.",
             "3.  Map the Weekly Lesson Plan and load Curriculum & Resources.",
             "4.  Track Subject Mastery, Math Facts & Sight Words as skills click.",
             "5.  Log reading, grade assignments and update the Report Card.",
             "6.  Watch the Dashboard track mastery, facts, reading & an On-Track Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for the Bennett family (Lucy, grade 3, and Jack, grade 1) is included so you can see how it "
               "all connects — just type over it with your own. Subject mastery, math facts, sight words, reading and "
               "graded work roll up into a live On-Track Score. Twelve matching printable pages (weekly plan, mastery "
               "checklist, math-facts chart, reading log, report card & more) are included to print and keep. Every "
               "child grows at their own pace — use this as an encouraging guide, and confirm your state's rules.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "See what's mastered and what's next — teach with confidence.", "section_gold")


# ===========================================================================
# 3 — Student Profiles
# ===========================================================================
def build_profiles(wb):
    ws = wb.create_sheet("Student Profiles"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 22, 4, 18, 18, 2])
    luxe_header(ws, "G", "🧑‍🎓  STUDENT PROFILES",
                "A profile for every child — grade, reading level, strengths & what's next.")
    row = 5
    for name, grade, age, level, note in STUDENTS:
        merge_set(ws, f"B{row}:F{row}", name, "section_gold"); ws.row_dimensions[row].height = 22
        row += 1
        fields = [("Grade", grade), ("Age", age), ("Reading level", level), ("Strengths", note),
                  ("Working on", "See Mastery tab"), ("Math facts", "See Math Facts tab"),
                  ("Curriculum", "See Curriculum tab"), ("Favorite subject", "—"),
                  ("Interests", "—"), ("Big goal", "See Awards tab")]
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 22; i += 2; row += 1
        row += 1
    ws.freeze_panes = "A5"


# ===========================================================================
# 4 — Weekly Lesson Plan
# ===========================================================================
def build_weekly(wb):
    ws = wb.create_sheet("Weekly Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 18, 18, 18, 18, 18, 2])
    luxe_header(ws, "G", "🗓  WEEKLY LESSON PLAN",
                "Map the week, subject by subject — then check it off as you teach.")
    table_headers(ws, 4, ["Subject", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"], start_col=2)
    rows = ["Math", "Reading", "Writing", "Spelling", "Science", "Social Studies", "Art", "PE"]
    start = L0
    for i, subj in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=subj).style = "td_left"; ws.cell(row=r, column=2).fill = fill(SURFACE)
        for c in range(3, 8):
            cell = ws.cell(row=r, column=c, value=""); cell.style = "td"
            cell.fill = fill(MUTED_ROW if i % 2 else WHITE)
    ws.freeze_panes = "A5"
    merge_set(ws, f"B{start+9}:G{start+9}",
              "Tip: block by subject or by child. Fill lessons here, then track what's mastered on the Mastery tab.",
              "section_gold")


# ===========================================================================
# 5 — Subject Mastery  (defines MasterySubject / MasteryName / MasteryStatus)
# ===========================================================================
def build_mastery(wb):
    ws, start, end = build_log(
        wb, "Subject Mastery", "🎯", "SUBJECT MASTERY",
        "The skills that matter this year — per subject & student, marked Mastered, Practicing or Not Yet.",
        ["Subject", "Skill", "Student", "Status"],
        MASTERY_ROWS, [18, 36, 14, 14], text_left={2}, reserved=40,
        validations=[("A", "SubjectList"), ("D", "MasteryList")])
    nrange(wb, "MasterySubject", "Subject Mastery", "A", start, end)
    nrange(wb, "MasteryName", "Subject Mastery", "B", start, end)
    nrange(wb, "MasteryStatus", "Subject Mastery", "D", start, end)
    cmap = {"Mastered": MINT_BG, "Practicing": WARN_BG, "Not Yet": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    sr = end + 3
    ws.cell(row=sr - 1, column=2, value="MASTERED BY SUBJECT").style = "section_gold"
    for i, subj in enumerate(SUBJECTS):
        r = sr + i
        ws.cell(row=r, column=2, value=subj).style = "td_left"
        c = ws.cell(row=r, column=3, value=f'=COUNTIFS(MasterySubject,B{r},MasteryStatus,"Mastered")')
        c.style = "td"; c.number_format = "#,##0"
    cell_name(wb, "SubjSumLabels", "Subject Mastery", f"$B${sr}:$B${sr + len(SUBJECTS) - 1}")
    cell_name(wb, "SubjSumVals", "Subject Mastery", f"$C${sr}:$C${sr + len(SUBJECTS) - 1}")


# ===========================================================================
# 6 — Math Facts Fluency  (defines MathPct)
# ===========================================================================
def build_mathfacts(wb):
    ws = wb.create_sheet("Math Facts"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 16, 2])
    luxe_header(ws, "D", "➗  MATH FACTS FLUENCY",
                "Addition, subtraction, multiplication & division — % fluent, per child. Watch it climb.")
    table_headers(ws, 4, ["Operation", "Student", "% Fluent"], start_col=2)
    start = L0
    for i, (op, stu, pct) in enumerate(MATHFACTS):
        r = start + i
        ws.cell(row=r, column=2, value=op).style = "td_left"
        ws.cell(row=r, column=3, value=stu).style = "td"
        cp = ws.cell(row=r, column=4, value=pct); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MATHFACTS) - 1
    nrange(wb, "MathPct", "Math Facts", "D", start, end)
    cell_name(wb, "MathAvg", "Math Facts", f"$D${end + 1}")
    ws.cell(row=end + 1, column=2, value="AVERAGE FLUENCY").style = "th"
    ws.cell(row=end + 1, column=3).style = "td"; ws.cell(row=end + 1, column=3).fill = fill(SURFACE)
    c = ws.cell(row=end + 1, column=4, value="=IFERROR(AVERAGE(MathPct),0)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0%"
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 7 — Sight Words & Spelling  (defines SWKnown / SWTotal)
# ===========================================================================
def build_sightwords(wb):
    ws = wb.create_sheet("Sight Words"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 12, 12, 12, 2])
    luxe_header(ws, "F", "🔤  SIGHT WORDS & SPELLING",
                "Word lists & spelling units — known vs total, with progress bars for every list.")
    table_headers(ws, 4, ["List / Unit", "Student", "Known", "Total", "% "], start_col=2)
    start = L0
    for i, (lst, stu, known, total) in enumerate(SIGHTWORDS):
        r = start + i
        ws.cell(row=r, column=2, value=lst).style = "td_left"
        ws.cell(row=r, column=3, value=stu).style = "td"
        ck = ws.cell(row=r, column=4, value=known); ck.style = "input"; ck.number_format = "#,##0"
        ct = ws.cell(row=r, column=5, value=total); ct.style = "input"; ct.number_format = "#,##0"
        cp = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/E{r},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SIGHTWORDS) - 1
    nrange(wb, "SWKnown", "Sight Words", "D", start, end)
    nrange(wb, "SWTotal", "Sight Words", "E", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
# 8 — Reading Log & Levels
# ===========================================================================
def build_reading(wb):
    ws, start, end = build_log(
        wb, "Reading Log", "📕", "READING LOG & LEVELS",
        "Every book, per child, with its level — reading is the heart of it. Set the year's goal in Settings.",
        ["Student", "Title", "Level", "Finished?"],
        READING, [16, 32, 12, 14], text_left={2}, reserved=60,
        validations=[("D", "StatusList")])
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 9 — Assignments & Grades  (defines AssignName / AssignGraded)
# ===========================================================================
def build_assignments(wb):
    rows = [(dminus(off), stu, subj, name, grade, gstat) for (off, stu, subj, name, grade, gstat) in ASSIGNMENTS]
    ws, start, end = build_log(
        wb, "Assignments", "📝", "ASSIGNMENTS & GRADES",
        "Every graded assignment, per student — the raw material for the report card.",
        ["Date", "Student", "Subject", "Assignment", "Grade", "Status"],
        rows, [13, 14, 14, 24, 10, 14], text_left={4}, dates={1}, reserved=60,
        validations=[("C", "SubjectList"), ("F", "GradedList")])
    nrange(wb, "AssignName", "Assignments", "D", start, end)
    nrange(wb, "AssignGraded", "Assignments", "F", start, end)
    cmap = {"Graded": MINT_BG, "Turned In": WARN_BG, "Not Yet": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 10 — Report Card
# ===========================================================================
def build_report(wb):
    ws = wb.create_sheet("Report Card"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 12, 12, 12, 12, 16, 2])
    luxe_header(ws, "G", "🏅  REPORT CARD",
                "A clean quarter-by-quarter record per student — print-ready for your files.")
    table_headers(ws, 4, ["Student — Subject", "Q1", "Q2", "Q3", "Q4", "Final"], start_col=2)
    rows = [
        ("Lucy — Math", "A", "", "", "", ""), ("Lucy — Reading", "A", "", "", "", ""),
        ("Lucy — Writing", "A-", "", "", "", ""), ("Lucy — Spelling", "B+", "", "", "", ""),
        ("Lucy — Science", "A", "", "", "", ""), ("Jack — Math", "S+", "", "", "", ""),
        ("Jack — Reading", "S", "", "", "", ""), ("Jack — Writing", "S", "", "", "", ""),
    ]
    start = L0
    for i, row in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=2, value=row[0]).style = "td_left"
        for j in range(5):
            ws.cell(row=r, column=3 + j, value=row[1 + j]).style = "input"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"
    merge_set(ws, "B15:F15", "GRADING KEY", "section_gold"); ws.row_dimensions[15].height = 22
    keys = [("A / 90–100", "Excellent"), ("B / 80–89", "Good"), ("S+ / S", "Satisfactory (younger)"),
            ("N", "Needs practice"), ("I", "In progress")]
    for i, (k, v) in enumerate(keys):
        r = 16 + i
        ws.cell(row=r, column=2, value=k).style = "field_label"
        ws.cell(row=r, column=3, value=v).style = "field_value"


# ===========================================================================
# 11 — Attendance & Days  (defines DaysDone)
# ===========================================================================
def build_attendance(wb):
    rows = [(f"Week {i}", 5, "") for i in range(1, 25)]
    ws, start, end = build_log(
        wb, "Attendance", "📅", "ATTENDANCE & DAYS",
        "Days by week — the record your state may require. Set required days in Settings.",
        ["Week", "Days", "Notes"],
        rows, [16, 12, 40], text_left={3}, ints={2}, reserved=40)
    nrange(wb, "AttDays", "Attendance", "B", start, end)
    tot = end + 1
    ws.cell(row=tot, column=1, value="TOTAL").style = "th"
    cd = ws.cell(row=tot, column=2, value="=SUM(AttDays)"); cd.style = "td"; cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(SURFACE); cd.number_format = "#,##0"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cell_name(wb, "DaysDone", "Attendance", f"$B${tot}")


# ===========================================================================
# 12 — Curriculum & Resources
# ===========================================================================
def build_curriculum(wb):
    ws, start, end = build_log(
        wb, "Curriculum", "📖", "CURRICULUM & RESOURCES",
        "What you're using per subject & child, what it cost & what's still needed.",
        ["Subject", "Resource", "Student", "Cost", "Status"],
        CURRICULUM, [18, 26, 14, 12, 12], text_left={2}, money={4}, reserved=30,
        validations=[("A", "SubjectList"), ("E", "AcquiredList")])
    cmap = {"Have": MINT_BG, "Ordered": WARN_BG, "Need": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 13 — Field Trips & Enrichment  (defines TripName)
# ===========================================================================
def build_trips(wb):
    rows = [(dplus(off) if off >= 0 else dminus(-off), place, subj, who) for (off, place, subj, who) in FIELDTRIPS]
    ws, start, end = build_log(
        wb, "Field Trips", "🚌", "FIELD TRIPS & ENRICHMENT",
        "Learning beyond the table — every trip, class & what it covered.",
        ["Date", "Place / Activity", "Subject Tie-in", "Who"],
        rows, [14, 28, 18, 14], text_left={2}, dates={1}, reserved=30)
    nrange(wb, "TripName", "Field Trips", "B", start, end)


# ===========================================================================
# 14 — Habits & Character  (defines HabitDone)
# ===========================================================================
def build_habits(wb):
    ws, start, end = build_log(
        wb, "Habits", "🌟", "HABITS & CHARACTER",
        "The little routines that build great learners — check them off & watch the streaks grow.",
        ["Habit", "Done Today?", "Streak (days)"],
        HABITS, [34, 14, 14], ints={3}, reserved=20,
        validations=[("B", "YesNoList")])
    nrange(wb, "HabitDone", "Habits", "B", start, end)
    ws.conditional_formatting.add(f"B{start}:B{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"C{start}:C{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=30, color=PRIMARY, showValue=True))


# ===========================================================================
# 15 — Awards & Milestones
# ===========================================================================
def build_awards(wb):
    rows = [(award, stu, dminus(-off) if off < 0 else dplus(off)) for (award, stu, off) in AWARDS]
    build_log(wb, "Awards", "🏆", "AWARDS & MILESTONES",
              "Every win worth celebrating — the milestones that keep little learners motivated.",
              ["Award / Milestone", "Student", "Date"],
              rows, [40, 16, 16], text_left={1}, dates={3}, reserved=24)


# ===========================================================================
# 16 — Book List
# ===========================================================================
def build_booklist(wb):
    rows = [
        ("The Chronicles of Narnia", "C.S. Lewis", "Read-aloud", "Have"),
        ("Magic Tree House series", "Osborne", "Lucy", "Have"),
        ("Frog and Toad", "Lobel", "Jack", "Have"),
        ("Charlotte's Web", "E.B. White", "Read-aloud", "Have"),
        ("Elephant & Piggie set", "Willems", "Jack", "Have"),
        ("The Boxcar Children", "Warner", "Lucy", "Need"),
        ("Poetry for Young People", "various", "Both", "Need"),
        ("D'Aulaires' Greek Myths", "D'Aulaire", "Read-aloud", "Have"),
    ]
    build_log(wb, "Book List", "📚", "BOOK LIST & WISH LIST",
              "The year's reading & the wish list — what to borrow, buy or save for later.",
              ["Title", "Author", "For", "Status"],
              rows, [30, 18, 16, 12], text_left={1}, reserved=40,
              validations=[("D", "AcquiredList")])


# ===========================================================================
# 2 — Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ✏  ELEMENTARY COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Mastery, math facts, reading & report cards — your whole K–5 year, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("STUDENTS", "=Students", "num"),
        ("SCHOOL DAYS", "=DaysDone", "num"),
        ("SUBJECTS", "=Subjects", "num"),
        ("SKILLS MASTERED", '=COUNTIF(MasteryStatus,"Mastered")', "num"),
        ("MATH FACTS", "=MathAvg", "pct"),
        ("SIGHT WORDS", "=SightWordsKnown", "num"),
    ]
    row2 = [
        ("BOOKS READ", "=BooksRead", "num"),
        ("GRADED", '=IFERROR(COUNTIF(AssignGraded,"Graded")/COUNTA(AssignName),0)', "pct"),
        ("FIELD TRIPS", "=COUNTA(TripName)", "num"),
        ("HABITS", '=IFERROR(COUNTIF(HabitDone,"Yes")/COUNTA(HabitDone),0)', "pct"),
        ("MASTERY", '=IFERROR(COUNTIF(MasteryStatus,"Mastered")/COUNTA(MasteryName),0)', "pct"),
        ("ON-TRACK", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "ARE WE ON TRACK?", "section_gold")
    merge_set(ws, "H11:M11", "SKILLS MASTERED BY SUBJECT", "section_gold")
    table_headers(ws, 12, ["Area", "Progress", "Status"], start_col=2)
    dims = [
        ("Skills mastered", '=IFERROR(COUNTIF(MasteryStatus,"Mastered")/COUNTA(MasteryName),0)'),
        ("Math facts fluency", "=MathAvg"),
        ("Sight words", "=IFERROR(SightWordsKnown/SightWordGoal,0)"),
        ("Assignments graded", '=IFERROR(COUNTIF(AssignGraded,"Graded")/COUNTA(AssignName),0)'),
        ("Reading goal", "=IFERROR(BooksRead/ReadingGoal,0)"),
        ("Habits & character", '=IFERROR(COUNTIF(HabitDone,"Yes")/COUNTA(HabitDone),0)'),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"On track",IF(C{r}>=0.6,"Going well","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    sk = wb["Subject Mastery"]
    dn = wb.defined_names["SubjSumVals"].value
    import re
    m = re.search(r"\$C\$(\d+):\$C\$(\d+)", dn)
    r0, r1 = int(m.group(1)), int(m.group(2))
    bar = BarChart(); bar.type = "bar"; bar.title = "Skills Mastered by Subject"; bar.height = 7.4; bar.width = 8.6
    bar.add_data(Reference(sk, min_col=3, min_row=r0, max_row=r1), titles_from_data=False)
    bar.set_categories(Reference(sk, min_col=2, min_row=r0, max_row=r1))
    bar.dataLabels = DataLabelList(); bar.dataLabels.showVal = True
    bar.legend = None
    ws.add_chart(bar, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Elementary Command Center™ — see what's mastered, teach what's next. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profiles(wb); build_weekly(wb)
    build_mastery(wb); build_mathfacts(wb); build_sightwords(wb); build_reading(wb)
    build_assignments(wb); build_report(wb); build_attendance(wb); build_curriculum(wb)
    build_trips(wb); build_habits(wb); build_awards(wb); build_booklist(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Student Profiles", "Weekly Plan", "Subject Mastery",
             "Math Facts", "Sight Words", "Reading Log", "Assignments", "Report Card",
             "Attendance", "Curriculum", "Field Trips", "Habits", "Awards", "Book List", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Elementary_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
