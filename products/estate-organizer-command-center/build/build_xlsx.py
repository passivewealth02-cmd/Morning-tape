"""Build Estate & Emergency Organizer Command Center™ — Everything They'd Need to Find.

14 tabs · a premium estate & emergency organizing system in Google Sheets & Excel.
Dashboard, an estate snapshot (assets − debts → net estate, and the share of it heading
for probate), assets & accounts, debts & bills, beneficiaries, legal documents, insurance,
digital life, key contacts, medical & care wishes, final wishes and household
instructions — one dashboard. Not a binder. A system, so nobody has to guess.

Run: python3 build_xlsx.py   ->  ../Estate_Organizer_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
TRANSFER = ["Joint with right of survivorship", "Payable on death (POD)", "Transfer on death (TOD)",
            "Beneficiary named", "Held in trust", "Sole title", "NO beneficiary named"]
ASSETCAT = ["Cash", "Retirement", "Investment", "Property", "Vehicle", "Insurance", "Personal"]
DOCSTATUS = ["Signed", "Drafted", "Not started", "Needs review"]
DIGITALACTION = ["Close", "Memorialize", "Transfer", "Download & keep", "Cancel billing"]

# --- Estate snapshot engine ---
MONTHLY_NEED = 6200
PROBATE_PCT = 0.05
LIFE_INSURANCE = 500000
CONTACT_GOAL = 10
RUNWAY_GOAL = 6
COVER_GOAL = 1.5
DOC_GOAL = 10

# Assets & accounts: (asset, category, value, how it transfers, probate?, where to find it)
ASSETS = [
    ("Checking \u2014 joint", "Cash", 8400, "Joint with right of survivorship", "No", "First Cedar Bank \u2014 app & statements in file drawer"),
    ("Savings \u2014 joint", "Cash", 22000, "Joint with right of survivorship", "No", "First Cedar Bank \u2014 same login"),
    ("Emergency fund \u2014 individual", "Cash", 15000, "Payable on death (POD)", "No", "Harbor Credit Union \u2014 POD form on file"),
    ("401(k)", "Retirement", 312000, "Beneficiary named", "No", "Fidelity \u2014 employer plan, see HR contact"),
    ("Roth IRA", "Retirement", 96000, "Beneficiary named", "No", "Vanguard \u2014 online only"),
    ("Brokerage \u2014 individual", "Investment", 84000, "NO beneficiary named", "Yes", "Schwab \u2014 NEEDS A TOD FORM"),
    ("Home", "Property", 410000, "Sole title", "Yes", "Deed in fire safe \u2014 sole name, needs review"),
    ("Vehicle \u2014 sedan", "Vehicle", 18000, "Sole title", "Yes", "Title in fire safe"),
    ("Vehicle \u2014 truck", "Vehicle", 9500, "Sole title", "Yes", "Title in fire safe"),
    ("Life insurance \u2014 term", "Insurance", 500000, "Beneficiary named", "No", "Northfield Mutual \u2014 policy in fire safe"),
    ("HSA", "Retirement", 11000, "Beneficiary named", "No", "HealthEquity \u2014 online"),
    ("Personal property & collectibles", "Personal", 14000, "Sole title", "Yes", "Inventory list & photos in this file"),
]

# Debts & bills: (debt, holder, balance, monthly, who to call)
DEBTS = [
    ("Mortgage", "First Cedar Bank", 218000, 1640, "800-555-0142 \u2014 loan #4471"),
    ("Auto loan", "Harbor Credit Union", 12400, 385, "800-555-0198 \u2014 loan #2210"),
    ("Credit cards", "Two cards \u2014 see list", 4600, 210, "Numbers on the back of each card"),
    ("Student loan", "Federal servicer", 9200, 165, "Discharged on death \u2014 send certificate"),
]

# Beneficiaries: (account, primary, contingent, last reviewed, needs action)
BENEFICIARIES = [
    ("401(k)", "Spouse \u2014 100%", "Children equally", "03/2026", "No"),
    ("Roth IRA", "Spouse \u2014 100%", "Children equally", "03/2026", "No"),
    ("HSA", "Spouse \u2014 100%", "Estate", "03/2026", "No"),
    ("Life insurance \u2014 term", "Spouse \u2014 100%", "Children equally", "03/2026", "No"),
    ("Emergency fund (POD)", "Spouse", "\u2014", "01/2026", "No"),
    ("Checking \u2014 joint", "Joint owner", "\u2014", "01/2026", "No"),
    ("Savings \u2014 joint", "Joint owner", "\u2014", "01/2026", "No"),
]

# Legal documents: (document, status, where kept, prepared by / date)
DOCUMENTS = [
    ("Will", "Signed", "Fire safe \u2014 original; copy with attorney", "Reyes Law, 06/2023"),
    ("Living trust", "Not started", "\u2014", "\u2014"),
    ("Financial power of attorney", "Not started", "\u2014", "\u2014"),
    ("Healthcare power of attorney", "Signed", "Fire safe + copy with doctor", "Reyes Law, 06/2023"),
    ("Living will / advance directive", "Not started", "\u2014", "\u2014"),
    ("HIPAA release", "Not started", "\u2014", "\u2014"),
    ("Beneficiary designations reviewed", "Signed", "Confirmations in this file", "Self, 03/2026"),
    ("Letter of instruction", "Not started", "\u2014", "\u2014"),
    ("Digital asset authorization", "Not started", "\u2014", "\u2014"),
    ("Funeral & burial wishes", "Signed", "Final Wishes tab + copy with spouse", "Self, 03/2026"),
]

# Insurance policies: (policy, company, type, coverage, policy #, contact)
INSURANCE = [
    ("Term life \u2014 20 yr", "Northfield Mutual", "Life", 500000, "NM-4471902", "800-555-0110"),
    ("Homeowners", "Cedar Ridge Insurance", "Property", 410000, "CR-88120", "800-555-0166"),
    ("Auto \u2014 both vehicles", "Cedar Ridge Insurance", "Auto", 300000, "CR-88121", "800-555-0166"),
    ("Umbrella liability", "Cedar Ridge Insurance", "Liability", 1000000, "CR-88122", "800-555-0166"),
    ("Health", "Meridian Health", "Health", 0, "MH-33417", "800-555-0177"),
    ("Long-term disability", "Employer plan", "Disability", 0, "Through HR", "See HR contact"),
]

# Digital life: (service, what it is, where the login lives, what to do)
DIGITAL = [
    ("Password manager", "The key to everything else", "Master password in sealed envelope, fire safe", "Transfer"),
    ("Primary email", "Recovery for every other account", "Password manager", "Download & keep"),
    ("Phone & carrier", "Two-factor codes live here", "PIN in sealed envelope", "Transfer"),
    ("Banking apps", "Checking, savings, credit union", "Password manager", "Transfer"),
    ("Investment logins", "Fidelity, Vanguard, Schwab", "Password manager", "Transfer"),
    ("Cloud photo storage", "Family photos \u2014 please keep these", "Password manager", "Download & keep"),
    ("Social accounts", "Facebook, Instagram", "Password manager", "Memorialize"),
    ("Subscriptions", "Streaming, software, delivery", "Password manager", "Cancel billing"),
    ("Domain & website", "Renews yearly, do not let it lapse", "Password manager", "Transfer"),
]

# Key contacts: (name, role, phone, email)
CONTACTS = [
    ("Reyes Law \u2014 Ana Reyes", "Attorney", "555-0121", "ana@reyeslaw.example"),
    ("Bell & Co \u2014 Tom Bell", "Accountant / CPA", "555-0132", "tom@bellco.example"),
    ("Harper Wealth \u2014 Jo Harper", "Financial advisor", "555-0143", "jo@harperwealth.example"),
    ("Dr. Nadia Osei", "Primary physician", "555-0154", "\u2014"),
    ("Northfield Mutual", "Life insurance", "800-555-0110", "\u2014"),
    ("Cedar Ridge Insurance", "Home & auto", "800-555-0166", "\u2014"),
    ("First Cedar Bank", "Mortgage & banking", "800-555-0142", "\u2014"),
    ("Employer HR \u2014 Dana Cole", "Benefits & 401(k)", "555-0165", "hr@employer.example"),
    ("Sam \u2014 brother", "Executor named in will", "555-0176", "sam@family.example"),
    ("Ellie \u2014 daughter", "Family", "555-0187", "ellie@family.example"),
    ("Marcus \u2014 son", "Family", "555-0198", "marcus@family.example"),
    ("Willow Veterinary", "Pets \u2014 Biscuit & Fig", "555-0209", "\u2014"),
]

# Medical & care: (item, detail)
MEDICAL = [
    ("Blood type", "O positive"),
    ("Allergies", "Penicillin \u2014 severe; shellfish \u2014 mild"),
    ("Ongoing conditions", "Type 2 diabetes, managed; high blood pressure"),
    ("Current medications", "Metformin 500mg 2x daily; lisinopril 10mg daily"),
    ("Pharmacy", "Cedar Pharmacy, 555-0210"),
    ("Primary physician", "Dr. Nadia Osei, 555-0154"),
    ("Specialists", "Dr. Amrit Kaur \u2014 endocrinology, 555-0211"),
    ("Health insurance", "Meridian Health, member MH-33417"),
    ("Resuscitation wishes", "See signed healthcare power of attorney \u2014 fire safe"),
    ("Organ donation", "Yes \u2014 registered on driver's licence"),
    ("Preferred hospital", "Cedar Regional"),
    ("Who decides if I can't", "Spouse, then Sam (brother) \u2014 named in healthcare POA"),
]

# Final wishes: (item, wish)
WISHES = [
    ("Burial or cremation", "Cremation"),
    ("Service", "Small graveside gathering, no formal funeral"),
    ("Where", "Willow Grove Cemetery \u2014 family plot, deed in fire safe"),
    ("Readings or music", "Nothing sad. Play the Nina Simone record."),
    ("Who should speak", "Whoever wants to. Keep it short and funny."),
    ("Flowers or donations", "Donations to the county library, please"),
    ("Obituary", "Draft is in this file \u2014 please use it, I already wrote it"),
    ("Who to notify first", "See Key Contacts \u2014 start with Sam"),
    ("Pre-paid arrangements", "None \u2014 estimate $4,200, use the emergency fund"),
    ("What I'd like said", "That I was easy to reach and hard to rattle."),
]

# Household instructions: (item, instruction)
HOUSEHOLD = [
    ("Water shut-off", "Basement, north wall, behind the shelf"),
    ("Electrical panel", "Garage, left of the door \u2014 labelled"),
    ("Furnace & filter", "Basement \u2014 filter is 16x25x1, change every 3 months"),
    ("Spare keys", "Fire safe, and Sam has a set"),
    ("Fire safe combination", "Sealed envelope with the attorney"),
    ("Pets", "Biscuit (dog) & Fig (cat) \u2014 to Ellie. Vet: Willow Veterinary"),
    ("Lawn & snow", "Cedar Grounds, 555-0220, billed quarterly"),
    ("Utilities", "See Debts & Bills \u2014 all on autopay from joint checking"),
    ("Mail", "Stop or forward at usps.com \u2014 needs a death certificate"),
    ("Safe deposit box", "First Cedar Bank, box 214 \u2014 key in fire safe"),
]

# What a survivor can actually reach in the first weeks: (source, amount)
REACHABLE = [
    ("Checking \u2014 joint", 8400), ("Savings \u2014 joint", 22000), ("Emergency fund \u2014 POD", 15000),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "money3": '"$"#,##0.000', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 20, 3] + [26] * 5)
    luxe_header(ws, "I", "⚙  SETTINGS", "Set your household numbers & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR HOUSEHOLD & GOALS", "section")
    controls = [
        ("Prepared by", "Margot", None, "Keeper"),
        ("Household", "Lantern & Oak", None, "Household"),
        ("Monthly household running cost", MONTHLY_NEED, '"$"#,##0', "MonthlyNeed"),
        ("Probate cost estimate (% of estate)", PROBATE_PCT, "0%", "ProbatePct"),
        ("Survivor runway goal (months)", RUNWAY_GOAL, "0", "RunwayGoal"),
        ("Insurance-covers-debt goal (×)", COVER_GOAL, "0.0", "CoverGoal"),
        ("Key contacts goal", CONTACT_GOAL, "0", "ContactGoal"),
        ("Core documents to complete", DOC_GOAL, "0", "DocGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "How it transfers", TRANSFER, "TransferList"), ("F", "Asset category", ASSETCAT, "AssetCatList"),
             ("G", "Document status", DOCSTATUS, "DocStatusList"), ("H", "Digital action", DIGITALACTION, "DigitalActionList"),
             ("I", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:I5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  \U0001f3ee  ESTATE & EMERGENCY ORGANIZER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Everything they'd need to find — so nobody has to guess.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "WHY THIS EXISTS", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("When something happens, the people who love you are handed a filing cabinet and a lot of "
                      "guessing. Which bank. Which policy. Where the deed is. Whether there was a will. This is the "
                      "answer to all of it, in one place. It also does something no binder does: it adds up what you "
                      "own, subtracts what you owe, and shows you how much of your estate would go through probate "
                      "because it has no beneficiary named on it — plus how much cash a survivor could actually reach "
                      "in the first few weeks, before anything is released. Fill it in once, keep it current, and "
                      "tell one person where it is.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Assets & Accounts — list what you own and how each one transfers.",
             "2.  Debts & Bills — what you owe and who to call.",
             "3.  Read the Estate Snapshot: net estate, and what's exposed to probate.",
             "4.  Beneficiaries & Legal Documents — fix the gaps it shows you.",
             "5.  Contacts, medical, digital life, final wishes, the house.",
             "6.  Print the pack, put it somewhere safe, and TELL SOMEONE."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  PLEASE READ", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+4}")
    c = ws[f"B{dr+1}"]
    c.value = ("This is an ORGANIZING tool, not legal, tax or financial advice, and it is not a will. Nothing you "
               "write here transfers anything to anyone — only properly executed documents and beneficiary "
               "designations do that. Probate rules, costs and timelines vary enormously by state and country; the "
               "estimate here is a rough percentage you set yourself. Please have real documents prepared by a "
               "qualified attorney. NEVER TYPE ACTUAL PASSWORDS INTO THIS FILE — the Digital Life tab records where "
               "each login lives, not what it is. Sample data for a fictional household is included so you can see "
               "how it connects; type over it with your own.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 5):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+6}:B{dr+6}", "A finished file nobody knows about helps nobody. Tell one person where it is.", "section_gold")


# ===========================================================================
def build_assets(wb):
    ws = wb.create_sheet("Assets & Accounts"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 16, 30, 12, 46, 2])
    luxe_header(ws, "G", "\U0001f4b0  ASSETS & ACCOUNTS",
                "Everything you own, what it's worth, how it transfers — and where to find it.")
    table_headers(ws, 4, ["Asset", "Category", "Value", "How it transfers", "Probate?", "Where to find it"], start_col=2)
    start = L0
    for i, (nm, cat, val, xfer, prob, where) in enumerate(ASSETS):
        r = start + i
        ws.cell(row=r, column=2, value=nm).style = "td_left"
        ws.cell(row=r, column=3, value=cat).style = "td"
        cv = ws.cell(row=r, column=4, value=val); cv.style = "input"; cv.number_format = '"$"#,##0'
        ws.cell(row=r, column=5, value=xfer).style = "td_left"
        ws.cell(row=r, column=6, value=prob).style = "input"
        ws.cell(row=r, column=7, value=where).style = "td_left"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(ASSETS) - 1
    nrange(wb, "AssetName", "Assets & Accounts", "B", start, end)
    nrange(wb, "AssetValue", "Assets & Accounts", "D", start, end)
    nrange(wb, "AssetProbate", "Assets & Accounts", "F", start, end)
    add_dv(ws, f"C{start}:C{end}", "AssetCatList"); add_dv(ws, f"E{start}:E{end}", "TransferList")
    add_dv(ws, f"F{start}:F{end}", "YesNoList")
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=['"Yes"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL ASSETS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=4, value="=SUM(AssetValue)"); ct.style = "td"
    ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "TotalAssets", "Assets & Accounts", f"$D${tot}")
    ws.cell(row=tot, column=5, value="of which exposed to probate").style = "field_label"
    cp = ws.cell(row=tot, column=6, value='=SUMIF(AssetProbate,"Yes",AssetValue)'); cp.style = "td"
    cp.font = Font(bold=True, size=12, color=DANGER); cp.fill = fill(RED_BG); cp.number_format = '"$"#,##0'
    cell_name(wb, "ProbateExposed", "Assets & Accounts", f"$F${tot}")
    ws.cell(row=tot + 1, column=2, value="Accounts documented").style = "field_label"
    ca = ws.cell(row=tot + 1, column=4, value="=COUNTA(AssetName)"); ca.style = "field_value"; ca.number_format = "0"
    cell_name(wb, "AssetCount", "Assets & Accounts", f"$D${tot+1}")
    ws.cell(row=tot + 3, column=2, value="Anything marked Yes has no named beneficiary and no joint owner. That is the list to fix.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_debts(wb):
    ws = wb.create_sheet("Debts & Bills"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 28, 16, 16, 40, 2])
    luxe_header(ws, "F", "\U0001f4c4  DEBTS & BILLS",
                "What is owed, to whom, and the number someone would need to call.")
    table_headers(ws, 4, ["Debt", "Held by", "Balance", "Monthly", "Who to call"], start_col=2)
    start = L0
    for i, (nm, holder, bal, mo, call) in enumerate(DEBTS):
        r = start + i
        ws.cell(row=r, column=2, value=nm).style = "td_left"
        ws.cell(row=r, column=3, value=holder).style = "td_left"
        cb = ws.cell(row=r, column=4, value=bal); cb.style = "input"; cb.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=5, value=mo); cm.style = "input"; cm.number_format = '"$"#,##0'
        ws.cell(row=r, column=6, value=call).style = "td_left"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DEBTS) - 1
    nrange(wb, "DebtBalance", "Debts & Bills", "D", start, end)
    nrange(wb, "DebtMonthly", "Debts & Bills", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL DEBTS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    ct = ws.cell(row=tot, column=4, value="=SUM(DebtBalance)"); ct.style = "td"
    ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "TotalDebts", "Debts & Bills", f"$D${tot}")
    cm = ws.cell(row=tot, column=5, value="=SUM(DebtMonthly)"); cm.style = "td"
    cm.font = Font(bold=True, size=12, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0'
    cell_name(wb, "DebtMonthlyTotal", "Debts & Bills", f"$E${tot}")
    ws.cell(row=tot + 2, column=2, value="Some debts die with you and some do not. Write down which is which while you can.").style = "section_gold"


def build_snapshot(wb):
    ws = wb.create_sheet("Estate Snapshot"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 48, 20, 2])
    luxe_header(ws, "C", "\U0001f9ee  ESTATE SNAPSHOT — THE ENGINE",
                "What you own, what you owe — and how much of it would have to go through probate.")
    ws.cell(row=5, column=2, value="THE ESTATE").style = "section_gold"
    ws.cell(row=6, column=2, value="Total assets").style = "field_label"
    ca = ws.cell(row=6, column=3, value="=TotalAssets"); ca.style = "field_value"; ca.number_format = '"$"#,##0'
    ws.cell(row=7, column=2, value="− Total debts").style = "field_label"
    cd = ws.cell(row=7, column=3, value="=TotalDebts"); cd.style = "field_value"
    cd.number_format = '"$"#,##0'; cd.fill = fill(WARN_BG)
    ws.cell(row=8, column=2, value="= NET ESTATE").style = "th"
    cn = ws.cell(row=8, column=3, value="=TotalAssets-TotalDebts"); cn.style = "td"
    cn.font = Font(bold=True, size=16, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetEstate", "Estate Snapshot", "$C$8")

    ws.cell(row=10, column=2, value="⚠ WHAT WOULD GO THROUGH PROBATE").style = "section_gold"
    ws.cell(row=11, column=2, value="Assets with no beneficiary, no joint owner, no trust").style = "field_label"
    cp = ws.cell(row=11, column=3, value="=ProbateExposed"); cp.style = "field_value"
    cp.number_format = '"$"#,##0'; cp.fill = fill(RED_BG)
    ws.cell(row=12, column=2, value="= SHARE OF YOUR ESTATE EXPOSED").style = "th"
    cs = ws.cell(row=12, column=3, value="=IFERROR(ProbateExposed/TotalAssets,0)"); cs.style = "td"
    cs.font = Font(bold=True, size=16, color=PRIMARY); cs.fill = fill(RED_BG); cs.number_format = "0.0%"
    cell_name(wb, "ProbateShare", "Estate Snapshot", "$C$12")
    ws.cell(row=13, column=2, value="× your probate cost estimate").style = "field_label"
    cpp = ws.cell(row=13, column=3, value="=ProbatePct"); cpp.style = "field_value"; cpp.number_format = "0%"
    ws.cell(row=14, column=2, value="= ROUGH COST OF PROBATE").style = "th"
    cc = ws.cell(row=14, column=3, value="=ProbateExposed*ProbatePct"); cc.style = "td"
    cc.font = Font(bold=True, size=14, color=PRIMARY); cc.fill = fill(WARN_BG); cc.number_format = '"$"#,##0'
    cell_name(wb, "ProbateCost", "Estate Snapshot", "$C$14")
    ws.cell(row=15, column=2, value="…and it typically takes months, not weeks, before anything is released.").style = "field_label"

    ws.cell(row=17, column=2, value="WHAT A SURVIVOR COULD ACTUALLY REACH").style = "section_gold"
    rs = 18
    for i, (nm, amt) in enumerate(REACHABLE):
        r = rs + i
        ws.cell(row=r, column=2, value=nm).style = "field_label"
        c = ws.cell(row=r, column=3, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
    re_ = rs + len(REACHABLE) - 1
    nrange(wb, "ReachableCol", "Estate Snapshot", "C", rs, re_)
    ws.cell(row=re_ + 1, column=2, value="= CASH REACHABLE IN DAYS, NOT MONTHS").style = "th"
    cr = ws.cell(row=re_ + 1, column=3, value="=SUM(ReachableCol)"); cr.style = "td"
    cr.font = Font(bold=True, size=14, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = '"$"#,##0'
    cell_name(wb, "CashReachable", "Estate Snapshot", f"$C${re_+1}")
    ws.cell(row=re_ + 2, column=2, value="÷ monthly household running cost").style = "field_label"
    cm = ws.cell(row=re_ + 2, column=3, value="=MonthlyNeed"); cm.style = "field_value"; cm.number_format = '"$"#,##0'
    ws.cell(row=re_ + 3, column=2, value="= MONTHS THEY COULD KEEP GOING").style = "th"
    cw = ws.cell(row=re_ + 3, column=3, value="=IFERROR(CashReachable/MonthlyNeed,0)"); cw.style = "td"
    cw.font = Font(bold=True, size=16, color=PRIMARY); cw.fill = fill(MINT_BG); cw.number_format = '0.0" months"'
    cell_name(wb, "Runway", "Estate Snapshot", f"$C${re_+3}")

    b = re_ + 5
    ws.cell(row=b, column=2, value="DOES THE INSURANCE COVER WHAT'S OWED?").style = "section_gold"
    ws.cell(row=b + 1, column=2, value="Life insurance in force").style = "field_label"
    ci = ws.cell(row=b + 1, column=3, value=LIFE_INSURANCE); ci.style = "input"; ci.number_format = '"$"#,##0'
    cell_name(wb, "LifeInsurance", "Estate Snapshot", f"$C${b+1}")
    ws.cell(row=b + 2, column=2, value="÷ total debts").style = "field_label"
    cdd = ws.cell(row=b + 2, column=3, value="=TotalDebts"); cdd.style = "field_value"; cdd.number_format = '"$"#,##0'
    ws.cell(row=b + 3, column=2, value="= IT COVERS THE DEBTS THIS MANY TIMES").style = "th"
    ccv = ws.cell(row=b + 3, column=3, value="=IFERROR(LifeInsurance/TotalDebts,0)"); ccv.style = "td"
    ccv.font = Font(bold=True, size=14, color=PRIMARY); ccv.fill = fill(MINT_BG); ccv.number_format = '0.00"×"'
    cell_name(wb, "DebtCover", "Estate Snapshot", f"$C${b+3}")
    ws.cell(row=b + 5, column=2, value="A beneficiary form takes ten minutes and skips probate entirely. That is the whole lesson.").style = "section_gold"


def build_beneficiaries(wb):
    ws = wb.create_sheet("Beneficiaries"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 26, 26, 17, 16, 2])
    luxe_header(ws, "F", "\U0001f465  BENEFICIARIES",
                "The forms that override your will — check them, because they win.")
    table_headers(ws, 4, ["Account", "Primary", "Contingent", "Last reviewed", "Needs action"], start_col=2)
    start = L0
    for i, (acct, pri, con, rev, act) in enumerate(BENEFICIARIES):
        r = start + i
        ws.cell(row=r, column=2, value=acct).style = "td_left"
        ws.cell(row=r, column=3, value=pri).style = "input"
        ws.cell(row=r, column=4, value=con).style = "input"
        ws.cell(row=r, column=5, value=rev).style = "input"
        ws.cell(row=r, column=6, value=act).style = "input"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BENEFICIARIES) - 1
    nrange(wb, "BeneAccount", "Beneficiaries", "B", start, end)
    nrange(wb, "BenePrimary", "Beneficiaries", "C", start, end)
    nrange(wb, "BeneAction", "Beneficiaries", "F", start, end)
    add_dv(ws, f"F{start}:F{end}", "YesNoList")
    ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=['"Yes"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 2
    ws.cell(row=tot, column=2, value="Accounts that can carry a beneficiary").style = "field_label"
    c1 = ws.cell(row=tot, column=5, value="=COUNTA(BeneAccount)"); c1.style = "field_value"; c1.number_format = "0"
    cell_name(wb, "BeneEligible", "Beneficiaries", f"$E${tot}")
    ws.cell(row=tot + 1, column=2, value="…with someone actually named").style = "field_label"
    c2 = ws.cell(row=tot + 1, column=5, value="=COUNTA(BenePrimary)"); c2.style = "field_value"
    c2.number_format = "0"; c2.fill = fill(MINT_BG)
    cell_name(wb, "BeneNamed", "Beneficiaries", f"$E${tot+1}")
    ws.cell(row=tot + 3, column=2, value="A beneficiary designation beats your will. If they disagree, the form wins.").style = "section_gold"


def build_documents(wb):
    ws = wb.create_sheet("Legal Documents"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 18, 44, 28, 2])
    luxe_header(ws, "E", "\U0001f4dc  LEGAL DOCUMENTS",
                "The ten documents that decide everything — and honestly, which ones you have.")
    table_headers(ws, 4, ["Document", "Status", "Where it is kept", "Prepared by / date"], start_col=2)
    start = L0
    for i, (doc, status, where, by) in enumerate(DOCUMENTS):
        r = start + i
        ws.cell(row=r, column=2, value=doc).style = "td_left"
        ws.cell(row=r, column=3, value=status).style = "input"
        ws.cell(row=r, column=4, value=where).style = "td_left"
        ws.cell(row=r, column=5, value=by).style = "td_left"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DOCUMENTS) - 1
    nrange(wb, "DocName", "Legal Documents", "B", start, end)
    nrange(wb, "DocStatusCol", "Legal Documents", "C", start, end)
    add_dv(ws, f"C{start}:C{end}", "DocStatusList")
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"Signed"'],
                                                                fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    ws.conditional_formatting.add(f"C{start}:C{end}", CellIsRule(operator="equal", formula=['"Not started"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 2
    ws.cell(row=tot, column=2, value="SIGNED").style = "th"
    c1 = ws.cell(row=tot, column=3, value='=COUNTIF(DocStatusCol,"Signed")'); c1.style = "td"
    c1.font = Font(bold=True, size=14, color=PRIMARY); c1.fill = fill(MINT_BG); c1.number_format = "0"
    cell_name(wb, "DocsSigned", "Legal Documents", f"$C${tot}")
    ws.cell(row=tot, column=4, value="…out of the ten that matter").style = "field_label"
    ws.cell(row=tot + 1, column=2, value="Not started").style = "field_label"
    c2 = ws.cell(row=tot + 1, column=3, value='=COUNTIF(DocStatusCol,"Not started")'); c2.style = "field_value"
    c2.number_format = "0"; c2.fill = fill(RED_BG)
    ws.cell(row=tot + 2, column=2, value="= HOW READY THE PAPERWORK IS").style = "th"
    c3 = ws.cell(row=tot + 2, column=3, value="=IFERROR(DocsSigned/DocGoal,0)"); c3.style = "td"
    c3.font = Font(bold=True, size=14, color=PRIMARY); c3.fill = fill(WARN_BG); c3.number_format = "0%"
    cell_name(wb, "DocReadiness", "Legal Documents", f"$C${tot+2}")
    ws.cell(row=tot + 4, column=2, value="Most people have a will and nothing else. The other nine are the ones that get used first.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_insurance(wb):
    ws, start, end = build_log(
        wb, "Insurance", "\U0001f6e1", "INSURANCE POLICIES", "Every policy, the number on it, and who to call to claim it.",
        ["Policy", "Company", "Type", "Coverage", "Policy #", "Contact"],
        [(p, c, t, v, n, ct) for (p, c, t, v, n, ct) in INSURANCE],
        [2, 26, 26, 14, 16, 18, 18, 2], text_left={2, 3}, money={5}, start_col=2)
    nrange(wb, "InsCoverage", "Insurance", "E", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="Total coverage listed").style = "field_label"
    c1 = ws.cell(row=tr, column=5, value="=SUM(InsCoverage)"); c1.style = "field_value"
    c1.number_format = '"$"#,##0'; c1.fill = fill(MINT_BG)
    ws.cell(row=tr + 2, column=2, value="Nobody claims a policy they don't know exists. This page is why it gets claimed.").style = "section_gold"


def build_digital(wb):
    ws = wb.create_sheet("Digital Life"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 38, 44, 20, 2])
    luxe_header(ws, "E", "\U0001f4bb  DIGITAL LIFE",
                "Where each login LIVES — never the password itself — and what to do with the account.")
    ws.cell(row=4, column=2, value="⚠  Do not type real passwords here. Record where they live: a password manager, a sealed envelope, a safe.").style = "section_gold"
    table_headers(ws, 5, ["Service", "What it is", "Where the login lives", "What to do"], start_col=2)
    start = 6
    for i, (svc, what, where, action) in enumerate(DIGITAL):
        r = start + i
        ws.cell(row=r, column=2, value=svc).style = "td_left"
        ws.cell(row=r, column=3, value=what).style = "td_left"
        ws.cell(row=r, column=4, value=where).style = "input"
        ws.cell(row=r, column=5, value=action).style = "input"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DIGITAL) - 1
    nrange(wb, "DigitalService", "Digital Life", "B", start, end)
    add_dv(ws, f"E{start}:E{end}", "DigitalActionList")
    tot = end + 2
    ws.cell(row=tot, column=2, value="Accounts documented").style = "field_label"
    c1 = ws.cell(row=tot, column=4, value="=COUNTA(DigitalService)"); c1.style = "field_value"; c1.number_format = "0"
    cell_name(wb, "DigitalCount", "Digital Life", f"$D${tot}")
    ws.cell(row=tot + 2, column=2, value="The password manager and the phone are the two keys. Everything else unlocks from those.").style = "section_gold"
    ws.freeze_panes = "A6"


def build_contacts(wb):
    ws, start, end = build_log(
        wb, "Key Contacts", "\U0001f4de", "KEY CONTACTS", "The people someone would need to reach, in the order they'd need them.",
        ["Name", "Role", "Phone", "Email"], CONTACTS,
        [2, 34, 26, 18, 32, 2], text_left={2, 3, 5}, start_col=2)
    nrange(wb, "ContactName", "Key Contacts", "B", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="Contacts listed").style = "field_label"
    c1 = ws.cell(row=tr, column=4, value="=COUNTA(ContactName)"); c1.style = "field_value"
    c1.number_format = "0"; c1.fill = fill(MINT_BG)
    cell_name(wb, "ContactCount", "Key Contacts", f"$D${tr}")
    ws.cell(row=tr + 2, column=2, value="Start with the executor and the attorney. Everything else follows from those two calls.").style = "section_gold"


def _two_col(wb, sheet, icon, title, subtitle, headers, data, widths, note):
    ws = wb.create_sheet(sheet); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    luxe_header(ws, get_column_letter(len(widths)), f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=2)
    start = L0
    for i, (a, b) in enumerate(data):
        r = start + i
        ws.cell(row=r, column=2, value=a).style = "td_left"
        ws.cell(row=r, column=3, value=b).style = "input"
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(data) - 1
    ws.cell(row=end + 2, column=2, value=note).style = "section_gold"
    ws.freeze_panes = "A5"
    return ws, start, end


def build_medical(wb):
    _two_col(wb, "Medical & Care", "⚕", "MEDICAL & CARE",
             "What a paramedic, a nurse or a family member would need to know in five minutes.",
             ["Item", "Detail"], MEDICAL, [2, 34, 72, 2],
             "Print this one and put a copy on the fridge. That is genuinely where they look.")


def build_wishes(wb):
    _two_col(wb, "Final Wishes", "\U0001f56f", "FINAL WISHES",
             "Written down, in your words, so nobody has to argue about what you would have wanted.",
             ["Item", "What I want"], WISHES, [2, 34, 72, 2],
             "Every line here is a decision someone else won't have to make on the worst week of their life.")


def build_household(wb):
    _two_col(wb, "Household", "\U0001f3e0", "HOUSEHOLD INSTRUCTIONS",
             "Where the shut-off is, who mows the lawn, and what happens to the dog.",
             ["Item", "Instruction"], HOUSEHOLD, [2, 30, 76, 2],
             "The small practical things are what actually overwhelm people. Write them down.")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  \U0001f3ee  ESTATE & EMERGENCY ORGANIZER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Net estate, what's exposed to probate & a Readiness Score — at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("TOTAL ASSETS", "=TotalAssets", "money"),
        ("TOTAL DEBTS", "=TotalDebts", "money"),
        ("NET ESTATE", "=NetEstate", "money"),
        ("EXPOSED TO PROBATE", "=ProbateExposed", "money"),
        ("PROBATE SHARE", "=ProbateShare", "pct1"),
        ("ROUGH PROBATE COST", "=ProbateCost", "money"),
    ]
    row2 = [
        ("CASH THEY CAN REACH", "=CashReachable", "money"),
        ("MONTHS OF RUNWAY", "=Runway", "dec"),
        ("LIFE INSURANCE", "=LifeInsurance", "money"),
        ("DOCUMENTS SIGNED", "=DocsSigned", "num"),
        ("ACCOUNTS DOCUMENTED", "=AssetCount+DigitalCount", "num"),
        ("READINESS SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "READINESS", "section_gold")
    merge_set(ws, "H11:M11", "WHERE THE ESTATE SITS", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Survivors could keep going", "=IFERROR(MIN(Runway/RunwayGoal,1),0)"),
        ("Beneficiaries named", "=IFERROR(MIN(BeneNamed/BeneEligible,1),0)"),
        ("Insurance covers the debts", "=IFERROR(MIN(DebtCover/CoverGoal,1),0)"),
        ("Key contacts listed", "=IFERROR(MIN(ContactCount/ContactGoal,1),0)"),
        ("Digital life documented", "=IFERROR(MIN(DigitalCount/9,1),0)"),
        ("Core documents signed", "=IFERROR(MIN(DocsSigned/DocGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    aw = wb["Assets & Accounts"]
    ch = BarChart(); ch.type = "bar"; ch.title = "Assets by Value"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(aw, min_col=4, min_row=L0, max_row=L0 + len(ASSETS) - 1), titles_from_data=False)
    ch.set_categories(Reference(aw, min_col=2, min_row=L0, max_row=L0 + len(ASSETS) - 1))
    ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Estate & Emergency Organizer Command Center™ — an organizing tool, not legal advice.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_assets(wb); build_debts(wb)
    build_snapshot(wb); build_beneficiaries(wb); build_documents(wb); build_insurance(wb)
    build_digital(wb); build_contacts(wb); build_medical(wb); build_wishes(wb)
    build_household(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Estate Snapshot", "Assets & Accounts", "Debts & Bills",
             "Beneficiaries", "Legal Documents", "Insurance", "Digital Life", "Key Contacts",
             "Medical & Care", "Final Wishes", "Household", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Estate_Organizer_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
