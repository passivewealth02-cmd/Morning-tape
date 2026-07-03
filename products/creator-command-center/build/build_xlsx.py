"""Build Creator Command Center™ — The Ultimate Content Creator Business
Operating System.

24 sheets + Welcome · a premium creator business OS in Excel & Sheets.
Content, calendar, pipeline, ideas, performance, revenue, sponsorships,
affiliates, products, expenses, goals, assets, email, launches, SEO,
repurposing, analytics & more — one elegant dashboard.

Run: python3 build_xlsx.py   ->  ../Creator_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

PLATFORMS = ["YouTube", "TikTok", "Instagram", "IG Reels", "IG Stories", "Facebook", "X",
             "Threads", "Pinterest", "LinkedIn", "Newsletter", "Podcast", "Blog"]
CONTENT_TYPES = ["Long-form Video", "Short", "Reel", "Carousel", "Story", "Post", "Newsletter", "Podcast", "Blog", "Live"]
CONTENT_STATUS = ["Idea", "In Progress", "Review", "Scheduled", "Published", "Repurposed"]
PIPE_STAGES = ["Brainstorm", "Research", "Outline", "Script", "Filming", "Editing", "Graphics", "Review", "Scheduled", "Published"]
REV_CATS = ["Sponsorships", "Affiliate Marketing", "Ad Revenue", "Digital Products", "Memberships",
            "Courses", "Consulting", "Merchandise", "Donations", "Freelance"]
EXP_CATS = ["Software", "Camera Gear", "Lighting", "Microphones", "Editing Tools", "Marketing",
            "Contractors", "Travel", "Education", "Office", "Miscellaneous"]
SPON_STAGES = ["Lead", "Contacted", "Negotiating", "Signed", "Delivered", "Paid", "Follow-Up"]
CAMPAIGN_TYPES = ["Sponsored", "Affiliate", "Product Launch", "Organic", "Series", "Collab"]
GOAL_CATS = ["Business", "Content", "Audience", "Revenue", "Platform", "Personal"]
PRIORITIES = ["High", "Medium", "Low"]
YESNO = ["Yes", "No"]
INVOICE_STATUS = ["Not Sent", "Sent", "Paid", "Overdue"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vc.font = Font(size=18, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "ppct": "+0%;-0%",
                        "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None,
              validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _recent_months(n):
    today = dt.date.today().replace(day=1)
    y, m = today.year, today.month
    seq = []
    for _ in range(n):
        seq.append(dt.date(y, m, 1)); m -= 1
        if m == 0:
            m = 12; y -= 1
    return [d.strftime("%b") for d in reversed(seq)]


def totals_row(ws, row, cols, start, end, fmt='"$"#,##0'):
    ws.cell(row=row, column=1, value="TOTAL").style = "th"
    for col in cols:
        L = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = fmt


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [16] * 8)
    luxe_header(ws, "L", "⚙  SETTINGS", "Set your creator business details once — every dashboard follows.")
    merge_set(ws, "B5:C5", "BUSINESS INPUTS", "section")
    controls = [
        ("Creator Name", "Ava Sky", None, "CreatorName"),
        ("Brand Name", "Nova Creative", None, "BrandName"),
        ("Primary Platform", "YouTube", None, "PrimaryPlatform"),
        ("Currency", "USD", None, "HomeCurr"),
        ("Monthly Revenue Goal", 10000, '"$"#,##0', "RevenueGoal"),
        ("Monthly Content Goal", 16, "0", "ContentGoal"),
        ("Active Deal Target", 6, "0", "DealTarget"),
        ("Profit Margin Target", 0.60, "0%", "MarginTarget"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Platform", PLATFORMS, "PlatformList"), ("F", "Content Type", CONTENT_TYPES, "ContentTypeList"),
             ("G", "Content Status", CONTENT_STATUS, "ContentStatusList"), ("H", "Revenue Category", REV_CATS, "RevCatList"),
             ("I", "Expense Category", EXP_CATS, "ExpCatList"), ("J", "Campaign Type", CAMPAIGN_TYPES, "CampaignList"),
             ("K", "Goal Category", GOAL_CATS, "GoalCatList"), ("L", "Sponsor Stage", SPON_STAGES, "SponStageList")]
    merge_set(ws, "E5:L5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    small = [("E", 21, "Priority", PRIORITIES, "PriorityList"),
             ("F", 21, "Invoice Status", INVOICE_STATUS, "InvoiceStatusList"),
             ("G", 21, "Yes / No", YESNO, "YesNoList")]
    for col, top, h, data, nm in small:
        ci = column_index_from_string(col)
        ws.cell(row=top, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=top + 1 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}${top+1}:${col}${top+len(data)}")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🎥  CREATOR COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate content creator business operating system.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "TURN YOUR CONTENT INTO A BUSINESS", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Most creators juggle a dozen disconnected tools for planning, finances, sponsorships and "
                      "analytics. Creator Command Center™ combines business management, content production, "
                      "publishing, revenue, sponsorships, audience insights and asset organization into ONE premium "
                      "Excel & Google Sheets operating system — so you spend less time organizing and more time "
                      "creating. This isn't a content calendar. It's your creator business OS.")
    ws["B6"].style = "body"
    for r in (6, 7, 8, 9):
        ws.row_dimensions[r].height = 22
    merge_set(ws, "B11:B11", "START HERE", "section")
    steps = ["1.  Open Settings and add your name, brand, platform & monthly goals.",
             "2.  Fill in the Brand Command Center (mission, voice, colors, handles).",
             "3.  Plan posts in the Content Calendar & move them through the Pipeline.",
             "4.  Log revenue, sponsorships, affiliates & products — profit updates live.",
             "5.  Track performance & audience growth; set Goals & OKRs.",
             "6.  Watch the Executive Dashboard track revenue, output & business health."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Realistic sample data (a multi-platform creator earning $8,160/mo) is included so you can see how "
               "everything connects — just type over it with your own. Revenue, net profit, publishing consistency, "
               "sponsorship pipeline, audience growth and the Business Health Score all update automatically. "
               "Note: the Password & Account Index references your password manager — never store real passwords "
               "here. Every sheet is print-friendly and works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized system, more time to create — let's build your empire.", "section_gold")


# ===========================================================================
# 2 — Brand Command Center
# ===========================================================================
def build_brand(wb):
    ws = wb.create_sheet("Brand"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 30, 6, 24, 24, 2])
    luxe_header(ws, "G", "🎨  BRAND COMMAND CENTER", "Your brand, defined — the identity every piece of content flows from.")
    blocks = [
        ("BRAND IDENTITY", [("Brand Name", "=BrandName"), ("Creator", "=CreatorName"),
                            ("Website", "novacreative.co"), ("Niche", "Creative business & tech"),
                            ("Founded", "2021"), ("Primary Platform", "=PrimaryPlatform")]),
        ("VOICE & AUDIENCE", [("Voice & Tone", "Warm, sharp, encouraging"), ("Target Audience", "Aspiring creators 24–38"),
                              ("Core Values", "Clarity · Craft · Generosity"), ("Mission", "Help creators build real businesses"),
                              ("Elevator Pitch", "Systems that turn content into income"), ("Content Pillars", "Systems · Money · Growth")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            ws.cell(row=row, column=3, value=fields[i][1]).style = "field_value"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                ws.cell(row=row, column=6, value=fields[i + 1][1]).style = "field_value"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    # brand colors
    merge_set(ws, "B15:F15", "BRAND COLORS & TYPE", "section_gold"); ws.row_dimensions[15].height = 22
    colors = [("Primary", PRIMARY), ("Accent", ACCENT), ("Surface", SURFACE), ("Highlight", HIGHLIGHT)]
    for i, (nm, hexv) in enumerate(colors):
        c = 2 + i
        ws.cell(row=16, column=c, value=nm).style = "field_label"
        sw = ws.cell(row=17, column=c, value=f"#{hexv}"); sw.fill = fill(hexv)
        sw.font = Font(bold=True, color="FFFFFF" if nm in ("Primary", "Accent") else PRIMARY)
        sw.alignment = Alignment(horizontal="center", vertical="center"); sw.border = BOX
        ws.row_dimensions[17].height = 26
    ws.cell(row=18, column=2, value="Headings").style = "field_label"
    ws.cell(row=18, column=3, value="Serif display").style = "field_value"
    ws.cell(row=18, column=5, value="Body").style = "field_label"
    ws.cell(row=18, column=6, value="Clean sans-serif").style = "field_value"
    # handles
    merge_set(ws, "B20:F20", "SOCIAL HANDLES", "section_gold"); ws.row_dimensions[20].height = 22
    handles = [("YouTube", "@novacreative"), ("TikTok", "@novacreative"), ("Instagram", "@nova.creative"),
               ("X", "@novacreative"), ("Newsletter", "The Nova Note"), ("LinkedIn", "Ava Sky")]
    for i, (p, h) in enumerate(handles):
        r = 21 + (i // 2)
        col = 2 if i % 2 == 0 else 5
        ws.cell(row=r, column=col, value=p).style = "field_label"
        ws.cell(row=r, column=col + 1, value=h).style = "field_value"


# ===========================================================================
# 3 — Content Master Calendar
# ===========================================================================
def build_calendar(wb):
    # (dayoffset(+future/-past), platform, topic, campaign, cta, priority, status)
    rows = [
        (-20, "YouTube", "5 Systems That 10x Your Output", "Series", "Subscribe", "High", "Published"),
        (-18, "TikTok", "The 2-folder content system", "Organic", "Follow", "Medium", "Published"),
        (-15, "Newsletter", "How I plan a month in 90 min", "Organic", "Read more", "High", "Published"),
        (-12, "Instagram", "Behind the scenes: my studio", "Organic", "Save", "Low", "Published"),
        (-10, "YouTube", "$0 to $8k/mo as a creator", "Sponsored", "Sponsor link", "High", "Published"),
        (-8, "Podcast", "Interview: scaling to full-time", "Series", "Listen", "Medium", "Published"),
        (-6, "IG Reels", "3 tools I can't live without", "Affiliate", "Link in bio", "Medium", "Published"),
        (-4, "Blog", "The creator's tax checklist", "Organic", "Download", "Low", "Published"),
        (-2, "TikTok", "Reply: how much I actually make", "Organic", "Follow", "Medium", "Published"),
        (-1, "Instagram", "Carousel: 7 hooks that work", "Organic", "Save", "Medium", "Published"),
        (1, "YouTube", "My full content workflow (2024)", "Product Launch", "Get the template", "High", "Scheduled"),
        (2, "Newsletter", "Launch week: behind the build", "Product Launch", "Buy now", "High", "Scheduled"),
        (3, "TikTok", "POV: launch day", "Product Launch", "Link in bio", "High", "Scheduled"),
        (4, "IG Reels", "Template walkthrough", "Product Launch", "Shop", "High", "Scheduled"),
        (6, "YouTube", "Sponsor integration: Riverside", "Sponsored", "Try free", "High", "Scheduled"),
        (8, "Instagram", "Q&A: your top questions", "Organic", "Comment", "Low", "Scheduled"),
        (10, "Podcast", "Solo: pricing your products", "Series", "Listen", "Medium", "Scheduled"),
        (12, "Blog", "SEO deep-dive for creators", "Organic", "Read", "Medium", "Scheduled"),
        (5, "YouTube", "Repurpose: shorts from ep. 12", "Series", "Subscribe", "Medium", "In Progress"),
        (7, "TikTok", "Trend audio idea", "Organic", "Follow", "Low", "In Progress"),
        (0, "Instagram", "Reel: editing hack", "Organic", "Save", "Medium", "Review"),
        (14, "Newsletter", "Monthly recap + wins", "Organic", "Read", "Low", "Idea"),
        (16, "YouTube", "Gear upgrade tour", "Affiliate", "Links below", "Low", "Idea"),
        (18, "LinkedIn", "Lessons from 3 years creating", "Organic", "Connect", "Low", "Idea"),
    ]
    sample = []
    for doff, plat, topic, camp, cta, prio, status in rows:
        d = dplus(doff) if doff >= 0 else dminus(-doff)
        sample.append((d, plat, topic, camp, cta, prio, status))
    ws, start, end = build_log(
        wb, "Calendar", "🗓", "CONTENT MASTER CALENDAR",
        "Plan every platform in one view — publishing status calculates itself.",
        ["Publish Date", "Platform", "Topic", "Campaign", "CTA", "Priority", "Status"],
        sample, [14, 14, 34, 15, 15, 11, 14],
        text_left={3, 4, 5}, dates={1},
        validations=[("B", "PlatformList"), ("D", "CampaignList"), ("F", "PriorityList"), ("G", "ContentStatusList")], reserved=60)
    nrange(wb, "CalDate", "Calendar", "A", start, end)
    nrange(wb, "CalPlatform", "Calendar", "B", start, end)
    nrange(wb, "CalStatus", "Calendar", "G", start, end)
    cmap = {"Published": MINT_BG, "Scheduled": WARN_BG, "In Progress": SOFT_BG, "Review": SURFACE, "Idea": WHITE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 4 — Content Pipeline
# ===========================================================================
def build_pipeline(wb):
    rows = [
        ("Full content workflow (2024)", "YouTube", "Ava", 1, "Editing", 0.80, "Final cut review"),
        ("Launch week newsletter", "Newsletter", "Ava", 2, "Review", 0.90, "Proofing"),
        ("Template walkthrough reel", "IG Reels", "Editor — Jon", 4, "Graphics", 0.65, "Captions next"),
        ("Riverside sponsor video", "YouTube", "Ava", 6, "Script", 0.40, "Brief approved"),
        ("Pricing your products (pod)", "Podcast", "Ava", 10, "Filming", 0.55, "Recorded, editing"),
        ("SEO deep-dive blog", "Blog", "Writer — Mia", 12, "Outline", 0.30, "Keyword map done"),
        ("Shorts from ep. 12", "YouTube", "Editor — Jon", 5, "Editing", 0.70, "3 of 5 cut"),
        ("Monthly recap", "Newsletter", "Ava", 14, "Brainstorm", 0.10, "Collecting wins"),
    ]
    sample = [(t, p, o, dplus(doff), st, prog, note) for (t, p, o, doff, st, prog, note) in rows]
    ws, start, end = build_log(
        wb, "Pipeline", "🎬", "CONTENT PIPELINE",
        "Every piece, every stage — from brainstorm to published, with live progress.",
        ["Title", "Platform", "Owner", "Due Date", "Stage", "Progress", "Notes"],
        sample, [30, 14, 16, 13, 14, 12, 24],
        text_left={1, 3, 7}, dates={4}, pcts={6},
        validations=[("B", "PlatformList")], reserved=40)
    nrange(wb, "PipeTitle", "Pipeline", "A", start, end)
    nrange(wb, "PipeProgress", "Pipeline", "F", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 5 — Content Idea Vault
# ===========================================================================
def build_ideas(wb):
    # (idea, category, platform, keyword, hook, impact, effort, status)
    rows = [
        ("Creator tax masterclass", "Education", "YouTube", "creator taxes", "The tax bill that shocked me", 9, 6, "Planned"),
        ("Notion vs Sheets for creators", "Tools", "YouTube", "notion vs sheets", "I switched back. Here's why", 8, 4, "Idea"),
        ("30-day posting challenge", "Growth", "TikTok", "posting consistency", "Day 1 of posting daily", 7, 3, "Idea"),
        ("How I price digital products", "Money", "Newsletter", "digital product pricing", "$9 vs $49: what sold more", 9, 2, "Planned"),
        ("Studio tour + gear", "Behind-scenes", "YouTube", "creator studio setup", "My $2k home studio", 6, 5, "Idea"),
        ("Faceless content playbook", "Growth", "Blog", "faceless content", "No face, 100k followers", 8, 4, "Idea"),
        ("Sponsorship email templates", "Money", "Newsletter", "sponsorship outreach", "The email that landed $3k", 9, 2, "Planned"),
        ("Batch a month in a weekend", "Systems", "YouTube", "batch content", "How I film 12 videos in 2 days", 8, 5, "Idea"),
        ("Repurposing 1 → 10", "Systems", "Carousel", "repurpose content", "1 video, 10 posts", 7, 3, "Idea"),
        ("AI tools I actually use", "Tools", "TikTok", "ai for creators", "3 AI tools that save hours", 6, 2, "Idea"),
    ]
    ws = wb.create_sheet("Ideas"); ws.sheet_view.showGridLines = False
    set_widths(ws, [28, 15, 14, 18, 28, 9, 9, 11, 12])
    luxe_header(ws, "I", "💡  CONTENT IDEA VAULT",
                "Never run dry — capture ideas and score impact vs effort so the best rise up.")
    table_headers(ws, 4, ["Idea", "Category", "Platform", "Keyword", "Hook", "Impact", "Effort", "Score", "Status"])
    start = L0
    for i, (idea, cat, plat, kw, hook, imp, eff, status) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([idea, cat, plat, kw, hook, imp, eff], 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=8, value=f"=IFERROR(F{r}/G{r},0)")
        ws.cell(row=r, column=9, value=status)
    end = start + 40 - 1
    style_rows(ws, start, end, 9, text_left={1, 4, 5}, ints={6, 7}, dec={8})
    add_dv(ws, f"C{start}:C{end}", "PlatformList")
    ws.freeze_panes = "A5"
    ws.conditional_formatting.add(f"H{start}:H{end}",
        ColorScaleRule(start_type="num", start_value=0.5, start_color="FF" + WARN_BG,
                       end_type="num", end_value=4, end_color="FF" + HIGHLIGHT))


# ===========================================================================
# 6 — Content Performance Tracker
# ===========================================================================
def build_performance(wb):
    # (dago, title, platform, views, engagement, followers_gained)
    rows = [
        (2, "Reply: how much I actually make", "TikTok", 148000, 0.112, 1820),
        (1, "Carousel: 7 hooks that work", "Instagram", 42000, 0.086, 640),
        (4, "The creator's tax checklist", "Blog", 9800, 0.045, 90),
        (6, "3 tools I can't live without", "IG Reels", 96000, 0.094, 1120),
        (8, "Interview: scaling to full-time", "Podcast", 12400, 0.061, 210),
        (10, "$0 to $8k/mo as a creator", "YouTube", 84000, 0.078, 2100),
        (12, "Behind the scenes: my studio", "Instagram", 31000, 0.072, 380),
        (15, "How I plan a month in 90 min", "Newsletter", 6200, 0.412, 140),
        (18, "The 2-folder content system", "TikTok", 112000, 0.101, 1450),
        (20, "5 Systems That 10x Your Output", "YouTube", 67000, 0.069, 1580),
        (24, "Editing hack reel", "IG Reels", 58000, 0.088, 720),
        (27, "Weekly Q&A thread", "X", 21000, 0.034, 160),
    ]
    ws = wb.create_sheet("Performance"); ws.sheet_view.showGridLines = False
    set_widths(ws, [32, 14, 13, 13, 14, 15, 13])
    luxe_header(ws, "G", "📊  CONTENT PERFORMANCE TRACKER",
                "What's working — views, engagement & followers gained, with top performers ranked.")
    table_headers(ws, 4, ["Content Title", "Platform", "Publish Date", "Views", "Engagement", "Followers +", "Rank"])
    start = L0
    for i, (dago, title, plat, views, eng, fol) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=title)
        ws.cell(row=r, column=2, value=plat)
        ws.cell(row=r, column=3, value=dminus(dago))
        ws.cell(row=r, column=4, value=views)
        ws.cell(row=r, column=5, value=eng)
        ws.cell(row=r, column=6, value=fol)
        ws.cell(row=r, column=7, value=f"=IF(D{r}=\"\",\"\",RANK(D{r},$D${start}:$D${start+len(rows)-1}))")
    end = start + 40 - 1
    style_rows(ws, start, end, 7, text_left={1}, dates={3}, ints={4, 6, 7}, pcts={5})
    add_dv(ws, f"B{start}:B{end}", "PlatformList")
    ws.freeze_panes = "A5"
    nrange(wb, "PerfTitle", "Performance", "A", start, end)
    nrange(wb, "PerfPlatform", "Performance", "B", start, end)
    nrange(wb, "PerfDate", "Performance", "C", start, end)
    nrange(wb, "PerfViews", "Performance", "D", start, end)
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=150000, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="lessThanOrEqual", formula=["3"], fill=fill(MINT_BG), font=Font(bold=True, color=PRIMARY)))
    # platform counts (feeds dashboard bar)
    ws.cell(row=4, column=9, value="Platform").style = "th"
    ws.cell(row=4, column=10, value="Posts").style = "th"
    plats = ["YouTube", "TikTok", "Instagram", "IG Reels", "Podcast", "Newsletter", "Blog"]
    for i, p in enumerate(plats):
        r = 5 + i
        ws.cell(row=r, column=9, value=p).style = "td_left"
        ws.cell(row=r, column=10, value=f'=COUNTIF(PerfPlatform,I{r})').style = "td"
    set_widths(ws, [32, 14, 13, 13, 14, 15, 13, 3, 14, 10])
    cell_name(wb, "PlatLabel", "Performance", "$I$5:$I$11")
    cell_name(wb, "PlatCount", "Performance", "$J$5:$J$11")


# ===========================================================================
# 7 — Revenue Command Center
# ===========================================================================
def build_revenue(wb):
    ws = wb.create_sheet("Revenue"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 16, 12, 3, 22, 14])
    luxe_header(ws, "G", "💰  REVENUE COMMAND CENTER",
                "Every income stream in one place — monthly, annual run-rate & profit, live.")
    table_headers(ws, 4, ["Income Source", "This Month", "Annual (est.)", "% of Rev"])
    monthly = {"Sponsorships": 3200, "Affiliate Marketing": 1150, "Ad Revenue": 840, "Digital Products": 1680,
               "Memberships": 420, "Courses": 0, "Consulting": 600, "Merchandise": 180, "Donations": 90, "Freelance": 0}
    start = L0; end = start + len(REV_CATS) - 1
    for i, cat in enumerate(REV_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cm = ws.cell(row=r, column=2, value=monthly[cat]); cm.style = "input"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/$B${end+1},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL REVENUE").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=0.45, color=PRIMARY, showValue=True))
    nrange(wb, "RevSource", "Revenue", "A", start, end)
    nrange(wb, "RevMonthly", "Revenue", "B", start, end)
    cell_name(wb, "RevenueTotal", "Revenue", f"$B${total}")
    cell_name(wb, "RevSponsor", "Revenue", "$B$5")
    cell_name(wb, "RevAffiliate", "Revenue", "$B$6")
    cell_name(wb, "RevDigital", "Revenue", "$B$8")
    # bottom line box
    merge_set(ws, "F4:G4", "THE BOTTOM LINE", "section_gold")
    rows2 = [("Revenue (month)", "=RevenueTotal", '"$"#,##0'), ("Annual run-rate", "=RevenueTotal*12", '"$"#,##0'),
             ("Expenses (month)", "=ExpenseTotal", '"$"#,##0'), ("Net profit", "=RevenueTotal-ExpenseTotal", '"$"#,##0'),
             ("Profit margin", "=IFERROR((RevenueTotal-ExpenseTotal)/RevenueTotal,0)", "0%")]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 5 + i
        ws.cell(row=r, column=6, value=lab).style = "field_label"
        c = ws.cell(row=r, column=7, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab in ("Net profit", "Profit margin"):
            ws.cell(row=r, column=7).fill = fill(MINT_BG)
    cell_name(wb, "NetProfit", "Revenue", "$G$8")


# ===========================================================================
# 11 — Expense Tracker
# ===========================================================================
def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 14, 16, 12, 3, 22, 14])
    luxe_header(ws, "G", "🧾  EXPENSE TRACKER",
                "Know what the business costs — every category, monthly & annual, with share bars.")
    table_headers(ws, 4, ["Expense Category", "This Month", "Annual (est.)", "% of Spend"])
    monthly = {"Software": 240, "Camera Gear": 480, "Lighting": 0, "Microphones": 0, "Editing Tools": 60,
               "Marketing": 350, "Contractors": 900, "Travel": 120, "Education": 99, "Office": 80, "Miscellaneous": 61}
    start = L0; end = start + len(EXP_CATS) - 1
    for i, cat in enumerate(EXP_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cm = ws.cell(row=r, column=2, value=monthly[cat]); cm.style = "input"; cm.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f"=B{r}*12"); ca.style = "td"; ca.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(B{r}/$B${end+1},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL EXPENSES").style = "th"
    for col in (2, 3):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.conditional_formatting.add(f"D{start}:D{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=0.4, color=ACCENT, showValue=True))
    nrange(wb, "ExpCat", "Expenses", "A", start, end)
    nrange(wb, "ExpMonthly", "Expenses", "B", start, end)
    cell_name(wb, "ExpenseTotal", "Expenses", f"$B${total}")


# ===========================================================================
# 8 — Sponsorship CRM
# ===========================================================================
def build_sponsorship(wb):
    rows = [
        ("Riverside", "Dana K.", "dana@riverside.fm", "Q3 integration", "1 video + 2 stories", 3200, "Signed", "Sent", 12),
        ("Skillshare", "Marco P.", "marco@skillshare.com", "Course promo", "1 dedicated video", 2600, "Negotiating", "Not Sent", 30),
        ("Notion", "Lena R.", "partners@notion.so", "Template collab", "1 video + newsletter", 2200, "Delivered", "Sent", -5),
        ("Squarespace", "Ivy T.", "creators@squarespace.com", "Website build", "2 reels", 1400, "Paid", "Paid", -25),
        ("Epidemic Sound", "Sam W.", "sam@epidemicsound.com", "Music partner", "Ongoing mention", 900, "Signed", "Sent", 45),
        ("BetterHelp", "Chris N.", "brand@betterhelp.com", "Wellness spot", "1 mid-roll", 1800, "Contacted", "Not Sent", 60),
        ("Adobe", "Priya S.", "creators@adobe.com", "Express feature", "1 tutorial", 2400, "Lead", "Not Sent", 75),
        ("HelloFresh", "Owen B.", "influencer@hellofresh.com", "Meal kit", "1 reel + story", 1100, "Paid", "Paid", -40),
    ]
    sample = [(b, c, e, camp, deliv, rate, stage, inv, dplus(doff)) for (b, c, e, camp, deliv, rate, stage, inv, doff) in rows]
    ws, start, end = build_log(
        wb, "Sponsorships", "🤝", "SPONSORSHIP CRM",
        "Turn brand deals into a pipeline — track every stage from lead to paid.",
        ["Brand", "Contact", "Email", "Campaign", "Deliverables", "Rate", "Stage", "Invoice", "Due / Paid"],
        sample, [16, 14, 24, 18, 22, 11, 14, 12, 13],
        text_left={2, 3, 4, 5}, money={6}, dates={9},
        validations=[("G", "SponStageList"), ("H", "InvoiceStatusList")], reserved=30)
    nrange(wb, "SponBrand", "Sponsorships", "A", start, end)
    nrange(wb, "SponRate", "Sponsorships", "F", start, end)
    nrange(wb, "SponStage", "Sponsorships", "G", start, end)
    for st, cc in {"Paid": MINT_BG, "Signed": WARN_BG, "Delivered": SOFT_BG, "Lead": WHITE}.items():
        ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    ws.conditional_formatting.add(f"H{start}:H{end}", CellIsRule(operator="equal", formula=['"Overdue"'], fill=fill(RED_BG)))


# ===========================================================================
# 9 — Affiliate Tracker
# ===========================================================================
def build_affiliate(wb):
    rows = [
        ("Amazon Associates", "Gear picks", 0.04, 3200, 96, 410, -10),
        ("Notion", "Template affiliate", 0.50, 890, 42, 380, 15),
        ("Riverside", "Recording tool", 0.30, 540, 18, 210, 5),
        ("Epidemic Sound", "Music referrals", 0.25, 720, 34, 118, -8),
        ("ConvertKit", "Email tool", 0.30, 260, 9, 32, 20),
    ]
    ws = wb.create_sheet("Affiliate"); ws.sheet_view.showGridLines = False
    set_widths(ws, [22, 20, 14, 12, 11, 13, 13])
    luxe_header(ws, "G", "🔗  AFFILIATE TRACKER",
                "Passive income, tracked — clicks, conversions & commission by program.")
    table_headers(ws, 4, ["Program", "Product", "Commission", "Clicks", "Sales", "Revenue", "Payout"])
    start = L0
    for i, (prog, prod, comm, clicks, sales, rev, doff) in enumerate(rows):
        r = start + i
        for ci, v in enumerate([prog, prod, comm, clicks, sales, rev], 1):
            ws.cell(row=r, column=ci, value=v)
        ws.cell(row=r, column=7, value=dplus(doff))
    end = start + 30 - 1
    style_rows(ws, start, end, 7, text_left={1, 2}, pcts={3}, ints={4, 5}, money={6}, dates={7})
    totals_row(ws, end + 1, [6], start, end)
    c = ws.cell(row=end + 1, column=4, value=f"=SUM(D{start}:D{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "#,##0"
    c = ws.cell(row=end + 1, column=5, value=f"=SUM(E{start}:E{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "#,##0"
    ws.freeze_panes = "A5"
    nrange(wb, "AffRevenue", "Affiliate", "F", start, end)


# ===========================================================================
# 10 — Digital Product Dashboard
# ===========================================================================
def build_products(wb):
    rows = [
        ("Content OS (Notion)", -120, 49, 186, 12, 4.9),
        ("Creator Tax Kit", -80, 29, 142, 6, 4.8),
        ("Thumbnail Pack", -200, 19, 240, 15, 4.7),
        ("Pricing Playbook", -40, 39, 68, 2, 4.9),
        ("Sponsorship Templates", -60, 24, 96, 5, 4.8),
    ]
    ws = wb.create_sheet("Products"); ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 14, 10, 12, 13, 11, 13, 12])
    luxe_header(ws, "H", "📦  DIGITAL PRODUCT DASHBOARD",
                "Products that sell while you sleep — units, revenue, refunds & profit per product.")
    table_headers(ws, 4, ["Product", "Launched", "Price", "Units Sold", "Revenue", "Refunds", "Net Profit", "Rating"])
    start = L0
    for i, (nm, dago, price, units, refunds, rating) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=nm)
        ws.cell(row=r, column=2, value=dminus(-dago) if dago < 0 else dplus(dago))
        ws.cell(row=r, column=3, value=price)
        ws.cell(row=r, column=4, value=units)
        ws.cell(row=r, column=5, value=f"=C{r}*D{r}")
        ws.cell(row=r, column=6, value=refunds)
        ws.cell(row=r, column=7, value=f"=E{r}-C{r}*F{r}")
        ws.cell(row=r, column=8, value=rating)
    end = start + 30 - 1
    style_rows(ws, start, end, 8, text_left={1}, dates={2}, money={3, 5, 7}, ints={4, 6}, dec={8})
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in (4, 5, 7):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE)
        c.number_format = "#,##0" if col == 4 else '"$"#,##0'
    ws.freeze_panes = "A5"
    nrange(wb, "ProdName", "Products", "A", start, end)
    nrange(wb, "ProdProfit", "Products", "G", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=5000, color=PRIMARY, showValue=True))


# ===========================================================================
# 12 — Brand Deal Calendar
# ===========================================================================
def build_dealcal(wb):
    rows = [
        ("Riverside", "Deliver final video", -2, "Sent for approval", "Approved"),
        ("Riverside", "Post stories (x2)", 3, "Scheduled", "On track"),
        ("Notion", "Newsletter mention", -5, "Delivered", "Complete"),
        ("Skillshare", "Approve script", 5, "Awaiting brand", "Pending"),
        ("Skillshare", "Film dedicated video", 12, "Not started", "Upcoming"),
        ("Epidemic Sound", "Monthly mention", 8, "Recurring", "On track"),
        ("Squarespace", "Invoice payment", -25, "Paid", "Complete"),
    ]
    sample = [(b, d, dplus(doff) if doff >= 0 else dminus(-doff), note, status) for (b, d, doff, note, status) in rows]
    ws, start, end = build_log(
        wb, "Deal Calendar", "📅", "BRAND DEAL CALENDAR",
        "Never miss a deliverable — due dates, approvals & payment milestones in one line.",
        ["Brand", "Deliverable / Milestone", "Due Date", "Notes", "Status"],
        sample, [16, 30, 13, 24, 16],
        text_left={2, 4}, dates={3}, reserved=30)
    for st, cc in {"Complete": MINT_BG, "On track": SOFT_BG, "Pending": WARN_BG, "Upcoming": WHITE}.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


# ===========================================================================
# 13 — Goals & OKRs
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 14, 16, 12, 12, 2])
    luxe_header(ws, "F", "🎯  GOALS & OKRs",
                "Run your business on objectives — measurable key results with live progress.")
    table_headers(ws, 4, ["Objective / Key Result", "Category", "Target", "Current", "Progress"])
    goals = [
        ("Hit $10k/mo revenue", "Revenue", "$10,000", "$8,160", 0.82),
        ("Publish 16 pieces / month", "Content", "16", "12", 0.75),
        ("Grow to 100k followers", "Audience", "100,000", "72,400", 0.72),
        ("Launch Content OS course", "Business", "Ship it", "80% built", 0.80),
        ("Land 6 active brand deals", "Platform", "6", "4", 0.67),
        ("Read 1 business book / mo", "Personal", "12", "7", 0.58),
    ]
    start = L0
    for i, (g, cat, tgt, cur, prog) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=g).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td"
        ws.cell(row=r, column=3, value=tgt).style = "td"
        ws.cell(row=r, column=4, value=cur).style = "td"
        cp = ws.cell(row=r, column=5, value=prog); cp.style = "input"; cp.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalCatList")
    nrange(wb, "GoalCategory", "Goals", "B", start, end)
    nrange(wb, "GoalProgress", "Goals", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))
    merge_set(ws, "B13:E13", "THIS QUARTER'S FOCUS", "section_gold"); ws.row_dimensions[13].height = 22
    focus = ["Ship the Content OS course (biggest revenue lever)",
             "Systematize sponsorship outreach → 6 active deals",
             "Double down on YouTube (best growth + ad revenue)",
             "Build the email list to 10k (owned audience)"]
    for i, fx in enumerate(focus):
        r = 14 + i
        cb = ws.cell(row=r, column=2, value="◆"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=11, color=ACCENT); cb.border = BOX
        merge_set(ws, f"C{r}:E{r}", fx, "td_left")


# ===========================================================================
# 14 — Asset Library Index
# ===========================================================================
def build_assets(wb):
    rows = [
        ("Logo — primary (SVG)", "Logo", "Brand/Logos", "Yes", "Master vector"),
        ("Logo — white / mono", "Logo", "Brand/Logos", "Yes", "For dark bg"),
        ("Channel banner set", "Graphic", "Brand/Banners", "Yes", "All platforms"),
        ("B-roll — studio", "Video", "Drive/B-roll/Studio", "Yes", "4K, 40 clips"),
        ("B-roll — city", "Video", "Drive/B-roll/City", "Yes", "Establishing shots"),
        ("Music license — Epidemic", "Music", "Licenses/Epidemic", "Yes", "Active subscription"),
        ("Thumbnail template (PSD)", "Template", "Templates/Thumbnails", "Yes", "3 layouts"),
        ("Media kit (PDF)", "Document", "Brand/Media Kit", "Yes", "Updated Q3"),
        ("Sponsor contract template", "Contract", "Legal/Contracts", "Yes", "Lawyer-reviewed"),
        ("Intro / outro stingers", "Video", "Drive/Motion", "Yes", "5s each"),
    ]
    ws, start, end = build_log(
        wb, "Assets", "🗂", "ASSET LIBRARY INDEX",
        "Find any file in seconds — a searchable index of every brand asset & where it lives.",
        ["Asset", "Type", "Storage Location", "Backed Up", "Notes"],
        rows, [26, 14, 24, 12, 24],
        text_left={1, 3, 5}, validations=[("D", "YesNoList")], reserved=40)
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(RED_BG)))


# ===========================================================================
# 15 — Password & Account Index
# ===========================================================================
def build_passwords(wb):
    rows = [
        ("YouTube", "@novacreative", "ava@novacreative.co", "Yes", "Yes"),
        ("TikTok", "@novacreative", "ava@novacreative.co", "Yes", "Yes"),
        ("Instagram", "@nova.creative", "ava@novacreative.co", "Yes", "Yes"),
        ("X", "@novacreative", "ava@novacreative.co", "Yes", "No"),
        ("ConvertKit", "ava@novacreative.co", "backup@novacreative.co", "Yes", "Yes"),
        ("Stripe", "ava@novacreative.co", "backup@novacreative.co", "Yes", "Yes"),
        ("Gumroad", "ava@novacreative.co", "backup@novacreative.co", "No", "No"),
        ("Notion", "ava@novacreative.co", "backup@novacreative.co", "Yes", "Yes"),
    ]
    ws = wb.create_sheet("Accounts"); ws.sheet_view.showGridLines = False
    set_widths(ws, [18, 24, 26, 12, 15])
    luxe_header(ws, "E", "🔐  PASSWORD & ACCOUNT INDEX",
                "Every account, secured — never store real passwords here; use your password manager.")
    merge_set(ws, "A5:E5", "  ⚠  SECURITY: store passwords in a manager (1Password, Bitwarden). This is an index only.", "th")
    ws.row_dimensions[5].height = 24
    for c in range(1, 6):
        ws.cell(row=5, column=c).fill = fill(WARN_BG); ws.cell(row=5, column=c).font = Font(bold=True, color=ACCENT)
    table_headers(ws, 6, ["Platform", "Username", "Recovery Email", "MFA On?", "Backup Codes"])
    start = 7
    for i, row in enumerate(rows):
        r = start + i
        for ci, v in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=v)
    end = start + 30 - 1
    style_rows(ws, start, end, 5, text_left={1, 2, 3})
    for col_letter in ("D", "E"):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", "YesNoList")
    ws.freeze_panes = "A7"
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(RED_BG)))
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# 16 — Email List Dashboard
# ===========================================================================
def build_email(wb):
    ws = wb.create_sheet("Email List"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 4, 22, 16, 16, 2])
    luxe_header(ws, "H", "📧  EMAIL LIST DASHBOARD",
                "Your owned audience — subscribers, growth & campaign performance.")
    merge_set(ws, "B5:C5", "LIST HEALTH", "section")
    stats = [("Subscribers", 8420, "#,##0"), ("New this month", 640, "+#,##0"),
             ("Growth rate", 0.082, "0.0%"), ("Avg open rate", 0.446, "0.0%"),
             ("Avg click rate", 0.089, "0.0%"), ("Unsub rate", 0.006, "0.0%")]
    for i, (lab, val, fmt) in enumerate(stats):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "field_value"; c.number_format = fmt
        if lab in ("Growth rate", "Avg open rate"):
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    # recent campaigns
    merge_set(ws, "E5:G5", "RECENT CAMPAIGNS", "section_gold")
    table_headers(ws, 6, ["Campaign", "Opens", "Clicks"], start_col=5)
    camps = [("How I plan a month", "46%", "9.2%"), ("Launch week teaser", "51%", "12.4%"),
             ("Pricing your products", "43%", "7.8%"), ("Monthly recap + wins", "44%", "6.1%")]
    for i, (nm, op, cl) in enumerate(camps):
        r = 7 + i
        ws.cell(row=r, column=5, value=nm).style = "td_left"
        ws.cell(row=r, column=6, value=op).style = "td"
        ws.cell(row=r, column=7, value=cl).style = "td"
        if i % 2:
            for cc in (5, 6, 7):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    # lead magnets
    merge_set(ws, "B14:G14", "LEAD MAGNETS", "section_gold"); ws.row_dimensions[14].height = 22
    lm = [("Free Content Calendar", "1,240 signups"), ("Sponsorship Email Swipe", "890 signups"),
          ("Creator Tax Checklist", "610 signups")]
    for i, (nm, s) in enumerate(lm):
        r = 15 + i
        ws.cell(row=r, column=2, value=nm).style = "field_label"
        merge_set(ws, f"C{r}:D{r}", s, "field_value")


# ===========================================================================
# 17 — Collaboration Manager
# ===========================================================================
def build_collab(wb):
    rows = [
        ("Leo Marín (@leomakes)", "Podcast swap episode", "Podcast", 10, "0.50", "Confirmed"),
        ("Studio Kwan", "Thumbnail redesign", "YouTube", -3, "Flat $400", "Delivered"),
        ("Mia (writer)", "Blog SEO series", "Blog", 12, "Flat $250/post", "In Progress"),
        ("Jon (editor)", "Shorts editing", "YouTube", 5, "Flat $60/short", "In Progress"),
        ("Priya D. (@priyacodes)", "Collab reel", "IG Reels", 8, "0.50 split", "Scheduled"),
    ]
    sample = [(who, proj, plat, dplus(doff) if doff >= 0 else dminus(-doff), split, status) for (who, proj, plat, doff, split, status) in rows]
    ws, start, end = build_log(
        wb, "Collabs", "👥", "COLLABORATION MANAGER",
        "Partner up without the chaos — projects, deadlines, splits & status.",
        ["Collaborator", "Project", "Platform", "Deadline", "Revenue Split", "Status"],
        sample, [24, 24, 14, 13, 16, 15],
        text_left={1, 2, 5}, dates={4},
        validations=[("C", "PlatformList")], reserved=30)


# ===========================================================================
# 18 — Launch Planner
# ===========================================================================
def build_launch(wb):
    ws = wb.create_sheet("Launch Planner"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 30, 14, 2])
    luxe_header(ws, "F", "🚀  LAUNCH PLANNER",
                "Ship products with a plan — timeline, promo & KPIs for Content OS launch.")
    merge_set(ws, "B5:E5", "LAUNCH: CONTENT OS — 2-WEEK RUNWAY", "section")
    table_headers(ws, 6, ["Phase", "Date", "Action", "Status"], start_col=2)
    phases = [
        ("Pre-launch", -6, "Waitlist opens + teaser reel", "Done"),
        ("Pre-launch", -3, "Behind-the-build newsletter", "Done"),
        ("Launch day", 1, "YouTube video + email #1", "Scheduled"),
        ("Launch week", 2, "Launch newsletter + TikTok", "Scheduled"),
        ("Launch week", 4, "Template walkthrough reel", "Scheduled"),
        ("Launch week", 6, "Sponsor cross-promo", "Scheduled"),
        ("Close", 8, "Cart-close email + last call", "Planned"),
        ("Post-launch", 12, "Recap + testimonials", "Planned"),
    ]
    start = 7
    for i, (ph, doff, action, status) in enumerate(phases):
        r = start + i
        ws.cell(row=r, column=2, value=ph).style = "td_left"
        cd = ws.cell(row=r, column=3, value=dplus(doff) if doff >= 0 else dminus(-doff)); cd.style = "td"; cd.number_format = "mm/dd"
        ws.cell(row=r, column=4, value=action).style = "td_left"
        ws.cell(row=r, column=5, value=status).style = "td"
        if i % 2:
            for cc in range(2, 6):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    for st, cc in {"Done": MINT_BG, "Scheduled": WARN_BG, "Planned": WHITE}.items():
        ws.conditional_formatting.add(f"E{start}:E{start+len(phases)-1}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    # KPIs / budget
    merge_set(ws, "B16:E16", "LAUNCH TARGETS", "section_gold"); ws.row_dimensions[16].height = 22
    targets = [("Revenue goal", "$6,000"), ("Units goal", "150"), ("Promo budget", "$400"),
               ("Waitlist size", "1,240"), ("Email sequence", "5 emails"), ("Affiliate partners", "3")]
    for i, (lab, val) in enumerate(targets):
        r = 17 + (i // 2)
        col = 2 if i % 2 == 0 else 4
        ws.cell(row=r, column=col, value=lab).style = "field_label"
        ws.cell(row=r, column=col + 1, value=val).style = "field_value"


# ===========================================================================
# 19 — Audience Insights
# ===========================================================================
def build_audience(wb):
    ws = wb.create_sheet("Audience"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 30, 4, 24, 24, 2])
    luxe_header(ws, "G", "👤  AUDIENCE INSIGHTS",
                "Know who you serve — personas, pain points & the content they're asking for.")
    merge_set(ws, "B5:C5", "PRIMARY PERSONA", "section")
    persona = [("Name", "Side-Hustle Sam"), ("Age", "24–38"), ("Stage", "0–10k followers, part-time"),
               ("Goal", "Turn content into income"), ("Blocker", "Overwhelmed, disorganized"),
               ("Where", "YouTube + TikTok"), ("Spends on", "Tools & courses that save time")]
    for i, (lab, val) in enumerate(persona):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        ws.cell(row=r, column=3, value=val).style = "field_value"
    merge_set(ws, "E5:F5", "TOP PAIN POINTS", "section_gold")
    pains = ["\"I don't know what to post\"", "\"I'm not making money yet\"",
             "\"I can't stay consistent\"", "\"Editing takes forever\"", "\"How do I get sponsors?\""]
    for i, p in enumerate(pains):
        r = 6 + i
        ws.cell(row=r, column=5, value=f"•  {p}").style = "td_left"
        merge_set(ws, f"E{r}:F{r}", f"•  {p}", "td_left")
    merge_set(ws, "B14:C14", "MOST-REQUESTED TOPICS", "section_gold"); ws.row_dimensions[14].height = 22
    reqs = [("Sponsorship outreach", 42), ("Editing workflows", 38), ("Pricing products", 31),
            ("Content batching", 28), ("Faceless content", 22)]
    ws.cell(row=15, column=2, value="Topic").style = "th"; ws.cell(row=15, column=3, value="Requests").style = "th"
    for i, (t, n) in enumerate(reqs):
        r = 16 + i
        ws.cell(row=r, column=2, value=t).style = "td_left"
        ws.cell(row=r, column=3, value=n).style = "td"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    merge_set(ws, "E14:F14", "COMMUNITY FEEDBACK", "section_gold"); ws.row_dimensions[14].height = 22
    fb = ["\"Your systems changed my channel\"", "\"Finally hit $1k this month!\"",
          "\"The tax kit paid for itself\"", "\"More faceless content please\""]
    for i, f in enumerate(fb):
        r = 15 + i
        merge_set(ws, f"E{r}:F{r}", f"“{f.strip(chr(34))}”", "td_left")


# ===========================================================================
# 20 — SEO & Keyword Planner
# ===========================================================================
def build_seo(wb):
    rows = [
        ("how to start a youtube channel", "Informational", "YouTube", "Beginner guide video", "High", "Published"),
        ("best camera for youtube 2024", "Commercial", "Blog", "Gear roundup", "High", "In Progress"),
        ("content calendar template", "Transactional", "Blog", "Free template + upsell", "High", "Published"),
        ("how much do creators make", "Informational", "YouTube", "Income breakdown", "Medium", "Idea"),
        ("digital product ideas", "Informational", "Newsletter", "Idea list", "Medium", "Idea"),
        ("notion for content creators", "Commercial", "YouTube", "Notion setup tutorial", "Medium", "Planned"),
        ("faceless youtube automation", "Informational", "Blog", "Playbook post", "High", "Idea"),
        ("sponsorship rate calculator", "Transactional", "Blog", "Calculator tool", "Medium", "Planned"),
    ]
    ws, start, end = build_log(
        wb, "SEO", "🔍", "SEO & KEYWORD PLANNER",
        "Get found — map keywords to content by intent, priority & status.",
        ["Keyword", "Search Intent", "Platform", "Target Content", "Priority", "Status"],
        rows, [30, 16, 14, 24, 12, 14],
        text_left={1, 4}, validations=[("C", "PlatformList"), ("E", "PriorityList"), ("F", "ContentStatusList")], reserved=40)
    ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=['"High"'], fill=fill(WARN_BG), font=Font(bold=True, color=ACCENT)))


# ===========================================================================
# 21 — Content Repurposing Matrix
# ===========================================================================
def build_repurpose(wb):
    ws = wb.create_sheet("Repurposing"); ws.sheet_view.showGridLines = False
    formats = ["Shorts", "Reel", "TikTok", "Carousel", "Blog", "Newsletter", "Pins", "Pod Clip", "Quote Gfx"]
    set_widths(ws, [30] + [11] * len(formats) + [12])
    luxe_header(ws, get_column_letter(1 + len(formats) + 1), "♻  CONTENT REPURPOSING MATRIX",
                "One piece → ten assets — turn every long-form into a week of content.")
    table_headers(ws, 4, ["Source Content"] + formats + ["Done"])
    # (title, [done flags per format])
    sources = [
        ("$0 to $8k/mo as a creator", [1, 1, 1, 1, 1, 1, 0, 1, 1]),
        ("5 Systems That 10x Output", [1, 1, 1, 1, 0, 1, 1, 1, 0]),
        ("The 2-folder content system", [1, 1, 1, 0, 0, 0, 0, 0, 1]),
        ("Interview: scaling full-time", [1, 0, 1, 1, 1, 1, 0, 1, 0]),
        ("Full content workflow (2024)", [0, 0, 0, 0, 0, 0, 0, 0, 0]),
    ]
    start = L0
    for i, (title, flags) in enumerate(sources):
        r = start + i
        ws.cell(row=r, column=1, value=title).style = "td_left"
        for j, f in enumerate(flags):
            c = ws.cell(row=r, column=2 + j, value="✓" if f else "○")
            c.style = "td"; c.font = Font(bold=True, color=PRIMARY if f else BORDER)
            c.fill = fill(MINT_BG if f else WHITE); c.border = BOX
        done = sum(flags)
        cd = ws.cell(row=r, column=2 + len(formats), value=f"={done}/{len(formats)}")
        cd.style = "td"; cd.number_format = "0%"; cd.font = Font(bold=True, color=PRIMARY)
    end = start + len(sources) - 1
    ws.freeze_panes = "B5"
    ws.conditional_formatting.add(f"{get_column_letter(2+len(formats))}{start}:{get_column_letter(2+len(formats))}{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=PRIMARY, showValue=True))


# ===========================================================================
# 22 — Photo & Brand Gallery
# ===========================================================================
def build_gallery(wb):
    ws = wb.create_sheet("Gallery"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "📸  PHOTO & BRAND GALLERY",
                "Your visual brand in one board — logos, mood, thumbnails, mockups & media kit.")
    sections = ["Brand Logos", "Mood Board", "Thumbnail Inspo", "IG Feed Layout", "Product Mockups", "Media Kit"]
    top0 = 5; card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th"); ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+5}", "📷\nPaste image here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 6):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 6, column=col, value="Notes").style = "field_label"
        merge_set(ws, f"{M}{row+6}:{R}{row+6}", "", "field_value")
        ws.cell(row=row + 7, column=col, value="Status").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")


# ===========================================================================
# 23 — Analytics Command Center
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 18, 3, 16, 12, 12, 2])
    luxe_header(ws, "H", "📈  ANALYTICS COMMAND CENTER",
                "Your business by the numbers — every area scored into one Business Health Score.")
    merge_set(ws, "B5:D5", "BUSINESS HEALTH DIMENSIONS", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Revenue vs goal", "=IFERROR(MIN(RevenueTotal/RevenueGoal,1),0)"),
        ("Profit margin", "=IFERROR(MIN(((RevenueTotal-ExpenseTotal)/RevenueTotal)/MarginTarget,1),0)"),
        ("Publishing consistency", '=IFERROR(MIN(COUNTIFS(PerfDate,">="&TODAY()-30)/ContentGoal,1),0)'),
        ("Content completion", "=IFERROR(AVERAGE(PipeProgress),0)"),
        ("Sponsorship pipeline", '=IFERROR(MIN((COUNTIF(SponStage,"Signed")+COUNTIF(SponStage,"Delivered")+COUNTIF(SponStage,"Negotiating"))/DealTarget,1),0)'),
        ("Goal progress", "=IFERROR(AVERAGE(GoalProgress),0)"),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.75,"Strong",IF(C{r}>=0.5,"Growing","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(f"C{start}:C{end}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "F5:H5", "BUSINESS HEALTH SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{end}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY); cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "Revenue · margin · output · completion · pipeline · goals.", "subtitle")
    ws["F10"].fill = fill(IVORY)
    cell_name(wb, "HealthRange", "Analytics", f"$C${start}:$C${end}")
    # audience trend table
    merge_set(ws, "B15:D15", "AUDIENCE — LAST 6 MONTHS (K)", "section")
    ws.cell(row=16, column=2, value="Month").style = "th"; ws.cell(row=16, column=3, value="Followers (K)").style = "th"
    months = _recent_months(6); vals = [55.0, 58.4, 62.1, 65.8, 68.2, 72.4]
    for i, (m, v) in enumerate(zip(months, vals)):
        r = 17 + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        c = ws.cell(row=r, column=3, value=v); c.style = "td"; c.number_format = "0.0"
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    cell_name(wb, "AudMonth", "Analytics", "$B$17:$B$22")
    cell_name(wb, "AudVal", "Analytics", "$C$17:$C$22")
    cell_name(wb, "AudNow", "Analytics", "$C$22")
    cell_name(wb, "AudPrev", "Analytics", "$C$21")
    bar = BarChart(); bar.type = "bar"; bar.title = "Health by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None; ws.add_chart(bar, "F15")


# ===========================================================================
# 1 — Executive Business Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🎥  CREATOR COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Content, revenue, sponsorships & audience — your whole creator business, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("PUBLISHED (30D)", '=COUNTIFS(PerfDate,">="&TODAY()-30)', "num"),
        ("SCHEDULED", '=COUNTIF(CalStatus,"Scheduled")', "num"),
        ("REVENUE / MO", "=RevenueTotal", "money"),
        ("SPONSORSHIPS", "=RevSponsor", "money"),
        ("AFFILIATE", "=RevAffiliate", "money"),
        ("DIGITAL PRODUCTS", "=RevDigital", "money"),
    ]
    row2 = [
        ("EXPENSES / MO", "=ExpenseTotal", "money"),
        ("NET PROFIT", "=RevenueTotal-ExpenseTotal", "money"),
        ("BRAND DEALS ACTIVE", '=COUNTIF(SponStage,"Signed")+COUNTIF(SponStage,"Delivered")+COUNTIF(SponStage,"Negotiating")', "num"),
        ("AUDIENCE GROWTH", "=IFERROR(AudNow/AudPrev-1,0)", "ppct"),
        ("COMPLETION RATE", '=IFERROR(MIN(COUNTIFS(PerfDate,">="&TODAY()-30)/ContentGoal,1),0)', "pct"),
        ("BUSINESS HEALTH", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "REVENUE & AUDIENCE", "section_gold")
    # revenue by source donut
    d1 = DoughnutChart(); d1.title = "Revenue by Source"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Revenue"], min_col=2, min_row=4, max_row=14), titles_from_data=True)
    d1.set_categories(Reference(wb["Revenue"], min_col=1, min_row=5, max_row=14)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")
    # audience growth line
    ln = LineChart(); ln.title = "Audience Growth (K)"; ln.height = 8.2; ln.width = 11.5
    ln.add_data(Reference(wb["Analytics"], min_col=3, min_row=16, max_row=22), titles_from_data=True)
    ln.set_categories(Reference(wb["Analytics"], min_col=2, min_row=17, max_row=22)); ln.legend = None
    ws.add_chart(ln, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "CONTENT & EXPENSES", "section_gold")
    # content by platform bar
    cb = BarChart(); cb.type = "col"; cb.title = "Content by Platform"; cb.height = 8.2; cb.width = 11.5
    cb.add_data(Reference(wb["Performance"], min_col=10, min_row=4, max_row=11), titles_from_data=True)
    cb.set_categories(Reference(wb["Performance"], min_col=9, min_row=5, max_row=11)); cb.legend = None
    ws.add_chart(cb, "B30")
    # expense breakdown donut
    eb = DoughnutChart(); eb.title = "Expense Breakdown"; eb.height = 8.2; eb.width = 11.5
    eb.add_data(Reference(wb["Expenses"], min_col=2, min_row=4, max_row=15), titles_from_data=True)
    eb.set_categories(Reference(wb["Expenses"], min_col=1, min_row=5, max_row=15)); eb.dataLabels = no_labels()
    ws.add_chart(eb, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Creator Command Center™ — plan, publish, monetize & scale. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_brand(wb); build_calendar(wb)
    build_pipeline(wb); build_ideas(wb); build_performance(wb); build_revenue(wb)
    build_sponsorship(wb); build_affiliate(wb); build_products(wb); build_expenses(wb)
    build_dealcal(wb); build_goals(wb); build_assets(wb); build_passwords(wb)
    build_email(wb); build_collab(wb); build_launch(wb); build_audience(wb)
    build_seo(wb); build_repurpose(wb); build_gallery(wb); build_analytics(wb)
    build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Brand", "Calendar", "Pipeline", "Ideas", "Performance",
             "Revenue", "Sponsorships", "Affiliate", "Products", "Expenses", "Deal Calendar",
             "Goals", "Assets", "Accounts", "Email List", "Collabs", "Launch Planner",
             "Audience", "SEO", "Repurposing", "Gallery", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Creator_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
