"""Build Meal Prep Business Command Center™ — The Meal-Prep Operating System.

14 tabs · a premium meal-prep / subscription-meal operating system in Google Sheets
& Excel. Dashboard, a per-meal cost engine, meal-plan pricing, a menu, subscribers
& MRR, a production plan, ingredient costs, packaging & delivery, weekly orders,
waste, income & expenses and a monthly summary — one dashboard. Cost every meal and
grow your recurring revenue.

Run: python3 build_xlsx.py   ->  ../Meal_Prep_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["each", "lb", "oz", "case", "tray", "gallon", "bag", "meal"]
PLAN_LIST = ["5 meals / week", "10 meals / week", "15 meals / week", "21 meals / week"]
STATUS = ["Active", "Paused", "New", "Cancelled"]

TARGET_FC = 0.35
MARGIN_GOAL = 0.50
SUB_GOAL = 45
PROFIT_GOAL = 6000
WASTE_LIMIT = 0.04

# Meal cost engine — flagship Grilled Chicken Bowl: (component, cost)
MEAL = [
    ("Protein (chicken 6 oz)", 1.80), ("Carb (rice / potato)", 0.45), ("Roasted veg", 0.60),
    ("Sauce & seasoning", 0.35), ("Container + label", 0.55), ("Kitchen labor / meal", 1.25),
]

# Meal plans: (plan, meals/week, price/week)
PLANS = [
    ("5 meals / week", 5, 55), ("10 meals / week", 10, 105),
    ("15 meals / week", 15, 150), ("21 meals / week", 21, 200),
]

# Menu / individual meals: (meal, protein, cost, price)
MENU = [
    ("Grilled Chicken Bowl", "Chicken", 5.00, 11), ("Beef & Broccoli", "Beef", 5.80, 12),
    ("Salmon & Quinoa", "Salmon", 6.50, 13), ("Turkey Meatballs", "Turkey", 4.90, 11),
    ("Veggie Buddha Bowl", "Tofu", 4.20, 10), ("Shrimp Stir-Fry", "Shrimp", 6.20, 12),
    ("Chicken Fajita", "Chicken", 5.10, 11), ("Pesto Pasta", "—", 4.00, 10),
]

# Subscribers: (plan, count)
SUBSCRIBERS = [
    ("5 meals / week", 20), ("10 meals / week", 15), ("15 meals / week", 8), ("21 meals / week", 3),
]

# Production plan: (meal, qty this week, prep minutes)
PRODUCTION = [
    ("Grilled Chicken Bowl", 120, 90), ("Beef & Broccoli", 80, 75), ("Salmon & Quinoa", 60, 60),
    ("Turkey Meatballs", 70, 70), ("Veggie Buddha Bowl", 50, 45), ("Chicken Fajita", 53, 50),
]

# Ingredient cost library: (ingredient, pack price, pack size, unit)
INGREDIENTS = [
    ("Chicken breast", 2.80, "1 lb", "lb"), ("Ground beef", 4.50, "1 lb", "lb"), ("Salmon", 9.00, "1 lb", "lb"),
    ("Rice", 0.90, "1 lb", "lb"), ("Mixed veg", 1.60, "1 lb", "lb"), ("Sauces", 3.20, "16 oz", "oz"),
    ("Containers", 0.42, "each", "each"), ("Labels", 0.06, "each", "each"),
]

# Packaging & delivery: (item, cost, notes)
PACKAGING = [
    ("Meal container", 0.42, "3-compartment"), ("Label + sleeve", 0.13, "Printed weekly"),
    ("Insulated bag", 1.20, "Reusable"), ("Ice pack", 0.35, "Per delivery"),
    ("Delivery — Zone A", 4.00, "Local"), ("Delivery — Zone B", 6.00, "Suburbs"),
]

# Weekly orders: (client, plan, meals, delivery day)
ORDERS = [
    ("Alex P.", "10 meals / week", 10, "Sunday"), ("Bri T.", "5 meals / week", 5, "Sunday"),
    ("Chris L.", "21 meals / week", 21, "Monday"), ("Dana W.", "15 meals / week", 15, "Sunday"),
    ("Evan R.", "10 meals / week", 10, "Monday"), ("Faye M.", "5 meals / week", 5, "Sunday"),
]

# Waste log: (item, reason, cost) — monthly total ~$430
WASTE = [
    ("Over-prep", "Batch surplus", 160.00), ("Cancelled orders", "Late cancel", 120.00),
    ("Spoiled produce", "Cold-chain", 90.00), ("Remakes", "Order error", 60.00),
]

# Income & expenses (monthly): (item, amount)
LEDGER = [
    ("Ingredients", 5542, "E"), ("Packaging", 953, "E"), ("Kitchen labor", 2165, "E"),
    ("Delivery & fuel", 1200, "E"), ("Overhead & marketing", 1300, "E"),
]

# Monthly summary: (month, MRR, expenses)
MONTHS = [("Jul", 13200, 9400), ("Aug", 14800, 10100), ("Sep", 17900, 11160),
          ("Oct", 20400, 12200), ("Nov", 23600, 13800), ("Dec", 26000, 14900)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Business name", "Fresh Fuel Co.", None, "Business"),
        ("Owner", "Kai", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Subscriber goal", SUB_GOAL, "#,##0", "SubGoal"),
        ("Monthly profit goal", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Waste limit %", WASTE_LIMIT, "0%", "WasteLimit"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Plan", PLAN_LIST, "PlanList"),
             ("G", "Status", STATUS, "StatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🍱  MEAL PREP BUSINESS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every meal, and grow your recurring revenue.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE MEAL-PREP BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Meal prep lives or dies on two numbers: the cost of a single meal and how many subscribers you keep. "
                      "This makes both visible: a per-meal cost engine (protein, carb, veg, sauce, packaging & labor), "
                      "meal-plan pricing at 5/10/15/21 meals a week, and a subscriber board that turns counts into weekly "
                      "revenue and monthly recurring revenue. Plan production, track ingredient & packaging costs, manage "
                      "weekly orders & delivery, and keep income & expenses — all in ONE premium Google Sheets & Excel "
                      "system built for meal-prep businesses.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your target food-cost % & subscriber goal.",
             "2.  Cost a meal in Meal Cost — protein, sides, packaging & labor.",
             "3.  Price your Meal Plans; the margin per plan calculates live.",
             "4.  Add Subscribers by plan — weekly revenue & MRR roll up.",
             "5.  Plan production, track packaging, delivery & orders.",
             "6.  Check the Dashboard: MRR, profit & a Prep Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional meal-prep business (Fresh Fuel Co.) is included so you can see how it all "
               "connects — just type over it with your own meals and plans. Cost per meal and monthly recurring revenue "
               "are the two numbers that decide whether a meal-prep business scales, and they roll into a live Prep Score. "
               "Twelve matching printable pages (meal cost card, prep list, delivery run sheet, subscriber list & more) are "
               "included. This is a business tool, not financial, legal or food-safety advice — follow your local food-"
               "handling rules and confirm figures with your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Cost per meal × subscribers = a business that scales — measure both.", "section_gold")


# ===========================================================================
def build_mealcost(wb):
    ws = wb.create_sheet("Meal Cost"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 2])
    luxe_header(ws, "C", "🍲  MEAL COST",
                "Cost a meal to the component — protein, carb, veg, sauce, packaging & labor. The engine behind pricing.")
    ws.cell(row=5, column=2, value="MEAL").style = "section_gold"
    ws.cell(row=5, column=3, value="Grilled Chicken Bowl").font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Component", "Cost"], start_col=2)
    start = 7
    for i, (comp, cost) in enumerate(MEAL):
        r = start + i
        ws.cell(row=r, column=2, value=comp).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MEAL) - 1; tot = end + 1
    ing_end = start + 3  # first 4 rows are food ingredients
    ws.cell(row=tot, column=2, value="COST PER MEAL").style = "th"
    cm = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0.00'
    cell_name(wb, "MealCost", "Meal Cost", f"$C${tot}")
    ws.cell(row=tot + 2, column=2, value="Ingredient cost (food only)").style = "field_label"
    ci = ws.cell(row=tot + 2, column=3, value=f"=SUM(C{start}:C{ing_end})"); ci.style = "field_value"; ci.number_format = '"$"#,##0.00'; ci.fill = fill(MINT_BG)
    cell_name(wb, "IngredientCost", "Meal Cost", f"$C${tot+2}")
    ws.cell(row=tot + 4, column=2, value="Copy this build for every meal — swap the protein, carb & sauce.").style = "section"


def build_plans(wb):
    ws = wb.create_sheet("Meal Plans"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 12, 14, 14, 14, 12, 2])
    luxe_header(ws, "G", "📦  MEAL PLANS",
                "Price each weekly plan — meals × cost = plan cost, and the margin calculates live.")
    table_headers(ws, 4, ["Plan", "Meals", "Price/wk", "Cost/wk", "Margin", "Margin %"], start_col=2)
    start = L0
    for i, (plan, meals, price) in enumerate(PLANS):
        r = start + i
        ws.cell(row=r, column=2, value=plan).style = "td_left"
        cme = ws.cell(row=r, column=3, value=meals); cme.style = "input"; cme.number_format = "#,##0"
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cc = ws.cell(row=r, column=5, value=f"=C{r}*MealCost"); cc.style = "td"; cc.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=6, value=f"=D{r}-E{r}"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.number_format = '"$"#,##0'
        cmp = ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/D{r},0)"); cmp.style = "td"; cmp.number_format = "0%"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PLANS) - 1
    nrange(wb, "PlanName", "Meal Plans", "B", start, end)
    nrange(wb, "PlanMeals", "Meal Plans", "C", start, end)
    nrange(wb, "PlanPrice", "Meal Plans", "D", start, end)
    nrange(wb, "PlanCost", "Meal Plans", "E", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        ColorScaleRule(start_type="num", start_value=0.35, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=0.50, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.65, end_color="FF" + HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_menu(wb):
    ws = wb.create_sheet("Menu"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 12, 12, 12, 2])
    luxe_header(ws, "F", "🍽  MENU",
                "Every meal on rotation — cost, price, margin & food-cost %. Keep the low-cost stars in.")
    table_headers(ws, 4, ["Meal", "Protein", "Cost", "Price", "Food %"], start_col=2)
    start = L0
    for i, (meal, protein, cost, price) in enumerate(MENU):
        r = start + i
        ws.cell(row=r, column=2, value=meal).style = "td_left"
        ws.cell(row=r, column=3, value=protein).style = "td"
        cc = ws.cell(row=r, column=4, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cf = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/E{r},0)"); cf.style = "td"; cf.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MENU) - 1
    nrange(wb, "MenuItem", "Menu", "B", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        ColorScaleRule(start_type="num", start_value=0.30, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.45, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.60, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


def build_subscribers(wb):
    ws = wb.create_sheet("Subscribers"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 14, 14, 14, 16, 2])
    luxe_header(ws, "F", "👥  SUBSCRIBERS",
                "How many on each plan — the counts turn into weekly revenue and monthly recurring revenue.")
    table_headers(ws, 4, ["Plan", "Subscribers", "Price/wk", "Meals/wk", "Weekly Rev"], start_col=2)
    start = L0
    for i, (plan, count) in enumerate(SUBSCRIBERS):
        r = start + i
        ws.cell(row=r, column=2, value=plan).style = "td_left"
        cn = ws.cell(row=r, column=3, value=count); cn.style = "input"; cn.number_format = "#,##0"
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(INDEX(PlanPrice,MATCH(B{r},PlanName,0)),0)"); cp.style = "td"; cp.number_format = '"$"#,##0'
        cme = ws.cell(row=r, column=5, value=f"=C{r}*IFERROR(INDEX(PlanMeals,MATCH(B{r},PlanName,0)),0)"); cme.style = "td"; cme.number_format = "#,##0"
        crv = ws.cell(row=r, column=6, value=f"=C{r}*D{r}"); crv.style = "td"; crv.font = Font(bold=True, color=PRIMARY); crv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"B{r}", "PlanList")
    end = start + len(SUBSCRIBERS) - 1
    nrange(wb, "SubPlan", "Subscribers", "B", start, end)
    nrange(wb, "SubCount", "Subscribers", "C", start, end)
    nrange(wb, "SubMeals", "Subscribers", "E", start, end)
    nrange(wb, "SubRev", "Subscribers", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTALS").style = "th"
    cn = ws.cell(row=tot, column=3, value="=SUM(SubCount)"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cme = ws.cell(row=tot, column=5, value="=SUM(SubMeals)"); cme.style = "td"; cme.font = Font(bold=True, color=PRIMARY); cme.fill = fill(SURFACE); cme.number_format = "#,##0"
    crv = ws.cell(row=tot, column=6, value="=SUM(SubRev)"); crv.style = "td"; crv.font = Font(bold=True, color=PRIMARY); crv.fill = fill(MINT_BG); crv.number_format = '"$"#,##0'
    cell_name(wb, "TotalSubs", "Subscribers", f"$C${tot}")
    cell_name(wb, "MealsWeek", "Subscribers", f"$E${tot}")
    cell_name(wb, "WeeklyRev", "Subscribers", f"$F${tot}")
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Monthly recurring revenue (×4)").style = "field_label"
    c = ws.cell(row=sr, column=6, value="=WeeklyRev*4"); c.style = "field_value"; c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    cell_name(wb, "MRR", "Subscribers", f"$F${sr}")
    ws.cell(row=sr + 1, column=2, value="Avg price per meal").style = "field_label"
    c2 = ws.cell(row=sr + 1, column=6, value="=IFERROR(WeeklyRev/MealsWeek,0)"); c2.style = "field_value"; c2.number_format = '"$"#,##0.00'; c2.fill = fill(MINT_BG)
    cell_name(wb, "AvgPriceMeal", "Subscribers", f"$F${sr+1}")
    ws.cell(row=sr + 2, column=2, value="Food cost % / Margin %").style = "field_label"
    c3 = ws.cell(row=sr + 2, column=6, value="=IFERROR(IngredientCost/AvgPriceMeal,0)"); c3.style = "field_value"; c3.number_format = "0%"; c3.fill = fill(MINT_BG)
    cell_name(wb, "FoodCostPct", "Subscribers", f"$F${sr+2}")
    ws.cell(row=sr + 3, column=2, value="Margin % per meal").style = "field_label"
    c4 = ws.cell(row=sr + 3, column=6, value="=IFERROR((AvgPriceMeal-MealCost)/AvgPriceMeal,0)"); c4.style = "field_value"; c4.number_format = "0%"; c4.fill = fill(MINT_BG)
    cell_name(wb, "MarginPct", "Subscribers", f"$F${sr+3}")
    ws.freeze_panes = "A5"


def build_production(wb):
    ws = wb.create_sheet("Production Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 14, 14, 2])
    luxe_header(ws, "E", "👨‍🍳  PRODUCTION PLAN",
                "The weekly batch cook — how many of each meal to make, and the prep time it takes.")
    table_headers(ws, 4, ["Meal", "Qty", "Prep (min)", "Prep (hrs)"], start_col=2)
    start = L0
    for i, (meal, qty, prep) in enumerate(PRODUCTION):
        r = start + i
        ws.cell(row=r, column=2, value=meal).style = "td_left"
        cq = ws.cell(row=r, column=3, value=qty); cq.style = "input"; cq.number_format = "#,##0"
        cp = ws.cell(row=r, column=4, value=prep); cp.style = "input"; cp.number_format = "#,##0"
        ch = ws.cell(row=r, column=5, value=f"=D{r}/60"); ch.style = "td"; ch.number_format = "0.0"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRODUCTION) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    cq = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cq.style = "td"; cq.font = Font(bold=True, color=PRIMARY); cq.fill = fill(SURFACE); cq.number_format = "#,##0"
    cp = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(SURFACE); cp.number_format = "#,##0"
    ch = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); ch.style = "td"; ch.font = Font(bold=True, color=PRIMARY); ch.fill = fill(SURFACE); ch.number_format = "0.0"
    ws.freeze_panes = "A5"


def build_ingredients(wb):
    ws, start, end = build_log(
        wb, "Ingredient Costs", "🥕", "INGREDIENT COSTS",
        "Your buying list — what each ingredient costs, so meal costing stays accurate.",
        ["Ingredient", "Pack Price", "Pack Size", "Unit"],
        INGREDIENTS, [2, 22, 14, 14, 12, 2], text_left={2, 4}, money2={3}, reserved=24, start_col=2,
        validations=[("E", "UnitList")])


def build_packaging(wb):
    ws, start, end = build_log(
        wb, "Packaging & Delivery", "🚚", "PACKAGING & DELIVERY",
        "Containers, labels & delivery — the real cost of getting a meal to the door.",
        ["Item", "Cost", "Notes"],
        PACKAGING, [2, 24, 14, 24, 2], text_left={2, 4}, money2={3}, reserved=24, start_col=2)


def build_orders(wb):
    ws = wb.create_sheet("Weekly Orders"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 12, 16, 2])
    luxe_header(ws, "E", "📋  WEEKLY ORDERS",
                "This week's deliveries — who gets what, how many meals & which day.")
    table_headers(ws, 4, ["Client", "Plan", "Meals", "Delivery Day"], start_col=2)
    start = L0
    for i, (client, plan, meals, day) in enumerate(ORDERS):
        r = start + i
        ws.cell(row=r, column=2, value=client).style = "td_left"
        ws.cell(row=r, column=3, value=plan).style = "td_left"
        cm = ws.cell(row=r, column=4, value=meals); cm.style = "input"; cm.number_format = "#,##0"
        ws.cell(row=r, column=5, value=day).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "PlanList")
    end = start + len(ORDERS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="THIS WEEK").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = "#,##0"
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Over-prep, cancels & spoilage — the leaks that eat your meal margin.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_ledger(wb):
    ws = wb.create_sheet("Income & Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  INCOME & EXPENSES",
                "Monthly books — recurring revenue in, costs out, and the profit that's really yours.")
    ws.cell(row=5, column=2, value="INCOME").style = "section_gold"
    ws.cell(row=6, column=2, value="Monthly recurring revenue").style = "td_left"
    ci = ws.cell(row=6, column=3, value="=MRR"); ci.style = "td"; ci.font = Font(bold=True, color=PRIMARY); ci.number_format = '"$"#,##0'; ci.fill = fill(MINT_BG)
    cell_name(wb, "TotalIncome", "Income & Expenses", "$C$6")
    ws.cell(row=8, column=2, value="EXPENSES").style = "section_gold"
    table_headers(ws, 9, ["Expense", "Amount"], start_col=2)
    start = 10
    for i, (item, amt, _t) in enumerate(LEDGER):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LEDGER) - 1
    nrange(wb, "ExpAmt", "Income & Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Income & Expenses", f"$C${tot}")
    nr = tot + 2
    ws.cell(row=nr, column=2, value="= MONTHLY PROFIT").style = "th"
    cn = ws.cell(row=nr, column=3, value="=TotalIncome-ExpTotal"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetProfit", "Income & Expenses", f"$C${nr}")


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 14, 2])
    luxe_header(ws, "E", "📈  MONTHLY SUMMARY",
                "MRR, expenses & profit by month — watch your recurring revenue compound.")
    table_headers(ws, 4, ["Month", "MRR", "Expenses", "Profit"], start_col=2)
    start = L0
    for i, (m, mrr, exp) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ci = ws.cell(row=r, column=3, value=mrr); ci.style = "input"; ci.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=4, value=exp); ce.style = "input"; ce.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthMRR", "Monthly Summary", "C", start, end)
    ws.add_chart(_barchart(ws, "MRR by Month", start, end, 3, 2), "G4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🍱  MEAL PREP BUSINESS COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Cost per meal, subscribers, MRR & a Prep Score — your whole meal-prep business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MEAL COST", "=MealCost", "money2"),
        ("AVG PRICE/MEAL", "=AvgPriceMeal", "money2"),
        ("FOOD COST", "=FoodCostPct", "pct"),
        ("MARGIN/MEAL", "=MarginPct", "pct"),
        ("SUBSCRIBERS", "=TotalSubs", "num"),
        ("TOP PLAN", "=INDEX(SubPlan,MATCH(MAX(SubRev),SubRev,0))", "text"),
    ]
    row2 = [
        ("WEEKLY REVENUE", "=WeeklyRev", "money"),
        ("MRR", "=MRR", "money"),
        ("MEALS / WEEK", "=MealsWeek", "num"),
        ("MONTHLY PROFIT", "=NetProfit", "money"),
        ("WASTE %", "=IFERROR(WasteTotal/MRR,0)", "pct"),
        ("PREP SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "PREP HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "MRR BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Food cost on target", "=IFERROR(MIN(TargetFC/FoodCostPct,1),0)"),
        ("Margin healthy", "=IFERROR(MIN(MarginPct/MarginGoal,1),0)"),
        ("Plans costed", "=IFERROR(COUNTIF(PlanCost,\">0\")/COUNTA(PlanName),0)"),
        ("Subscribers vs goal", "=IFERROR(MIN(TotalSubs/SubGoal,1),0)"),
        ("Profitable", "=IFERROR(MIN(NetProfit/ProfitGoal,1),0)"),
        ("Waste low", "=IFERROR(1-MIN((WasteTotal/MRR)/WasteLimit,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "MRR by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Meal Prep Business Command Center™ — cost every meal, and grow your recurring revenue.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_mealcost(wb); build_plans(wb)
    build_menu(wb); build_subscribers(wb); build_production(wb); build_ingredients(wb)
    build_packaging(wb); build_orders(wb); build_waste(wb); build_ledger(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Meal Cost", "Meal Plans", "Menu", "Subscribers",
             "Production Plan", "Ingredient Costs", "Packaging & Delivery", "Weekly Orders", "Waste Log",
             "Income & Expenses", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Meal_Prep_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
