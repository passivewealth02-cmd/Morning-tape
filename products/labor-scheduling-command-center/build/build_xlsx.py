"""Build Restaurant Labor & Scheduling Command Center™ — The Labor-Cost System.

14 tabs · a premium labor & scheduling operating system in Google Sheets & Excel.
Dashboard, a labor-cost engine (scheduled hours × wage ÷ sales), a weekly schedule
grid, an employee roster, a sales forecast with labor targets, sales-per-labor-hour,
overtime, roles & rates, availability, prime cost, tips and labor-by-day — one
dashboard. Schedule to your sales, and never over-spend on labor again.

Run: python3 build_xlsx.py   ->  ../Labor_Scheduling_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
ROLES = ["Manager", "Line Cook", "Prep Cook", "Server", "Bartender", "Host", "Dishwasher"]
STATUS = ["Full-time", "Part-time", "On-call", "New"]
DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

TARGET_LABOR = 0.30
SPLH_GOAL = 50
OT_LIMIT = 10
PRIME_GOAL = 0.60
WAGE_GOAL = 16
FOOD_PCT = 0.30

# Weekly schedule: (name, role, wage, [Mon..Sun hours])
SCHEDULE = [
    ("Alex", "Line Cook", 19, [8, 8, 8, 8, 8, 0, 0]),
    ("Bailey", "Line Cook", 18, [0, 8, 8, 8, 8, 8, 0]),
    ("Casey", "Server", 12, [6, 6, 0, 6, 8, 8, 6]),
    ("Drew", "Server", 12, [0, 0, 6, 6, 8, 8, 6]),
    ("Emerson", "Dishwasher", 15, [6, 6, 6, 0, 8, 8, 0]),
    ("Finley", "Host", 14, [5, 5, 5, 5, 6, 6, 0]),
    ("Gray", "Bartender", 13, [0, 0, 0, 6, 8, 8, 6]),
    ("Harper", "Prep Cook", 17, [8, 8, 8, 8, 0, 0, 0]),
]

# Employee roster: (name, role, wage, status)
EMPLOYEES = [
    ("Alex", "Line Cook", 19, "Full-time"), ("Bailey", "Line Cook", 18, "Full-time"),
    ("Casey", "Server", 12, "Full-time"), ("Drew", "Server", 12, "Part-time"),
    ("Emerson", "Dishwasher", 15, "Part-time"), ("Finley", "Host", 14, "Part-time"),
    ("Gray", "Bartender", 13, "Part-time"), ("Harper", "Prep Cook", 17, "Part-time"),
]

# Sales forecast: (day, forecast sales)
FORECAST = [("Mon", 1500), ("Tue", 1700), ("Wed", 1800), ("Thu", 2100), ("Fri", 2900), ("Sat", 3200), ("Sun", 1800)]

# Overtime watch: (employee, OT hours, wage)
OVERTIME = [("Alex", 3, 19), ("Casey", 2, 12), ("Emerson", 1, 15)]

# Roles & rates: (role, wage, notes)
RATES = [
    ("Manager", 28, "Salary equivalent"), ("Line Cook", 19, "Station lead"), ("Prep Cook", 17, "AM prep"),
    ("Server", 12, "+ tips"), ("Bartender", 13, "+ tips"), ("Host", 14, "Front door"), ("Dishwasher", 15, "Back of house"),
]

# Availability / time-off: (employee, days off, notes)
AVAILABILITY = [
    ("Alex", "Sat, Sun", "Weekends off"), ("Bailey", "Mon, Sun", "School Mon AM"),
    ("Casey", "Wed", "Class"), ("Drew", "Mon, Tue", "Second job"),
    ("Emerson", "Thu, Sun", "—"), ("Finley", "Sun", "—"),
    ("Gray", "Mon, Tue, Wed", "Part-time"), ("Harper", "Fri, Sat, Sun", "AM only"),
]

# Tips: (day, tips, tip-out %)
TIPS = [("Mon", 320, 0.15), ("Tue", 360, 0.15), ("Wed", 380, 0.15), ("Thu", 520, 0.15),
        ("Fri", 780, 0.15), ("Sat", 860, 0.15), ("Sun", 440, 0.15)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Restaurant name", "Maple & Ash", None, "Restaurant"),
        ("Manager", "Jordan", None, "Manager"),
        ("Target labor %", TARGET_LABOR, "0%", "TargetLaborPct"),
        ("Sales-per-labor-hr goal", SPLH_GOAL, '"$"#,##0', "SPLHGoal"),
        ("Overtime limit (hrs)", OT_LIMIT, "#,##0", "OTLimit"),
        ("Prime-cost goal %", PRIME_GOAL, "0%", "PrimeGoal"),
        ("Avg-wage budget", WAGE_GOAL, '"$"#,##0', "WageGoal"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Role", ROLES, "RoleList"), ("F", "Status", STATUS, "StatusList"),
             ("G", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🕐  RESTAURANT LABOR & SCHEDULING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Schedule to your sales — and never over-spend on labor again.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE LABOR PICTURE, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Labor is the cost you control every single week — and it's the second half of prime cost. This system "
                      "makes it visible: build a weekly schedule and every employee's hours and cost calculate live, then "
                      "the labor-cost engine divides labor by sales for a true labor %. Forecast sales to set a labor "
                      "target, track sales per labor hour, watch overtime, and roll food cost + labor into a prime-cost "
                      "number — all in ONE premium Google Sheets & Excel system built for restaurants.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your target labor % & sales-per-hour goal.",
             "2.  Add your team in Employees; set roles & wages.",
             "3.  Build the Weekly Schedule — hours & cost calculate as you type.",
             "4.  Forecast sales; the labor target shows what you can spend.",
             "5.  Watch Sales-per-Labor-Hour, Overtime & Prime Cost.",
             "6.  Check the Dashboard: labor %, SPLH & a Labor Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional restaurant (Maple & Ash) is included so you can see how it all connects — just "
               "type over it with your own team and hours. Labor % and prime cost are the two numbers that decide whether "
               "a shift makes money, and they roll into a live Labor Score. Twelve matching printable pages (blank weekly "
               "schedule, roster, time-off request, tip sheet & more) are included. This is a business tool, not financial, "
               "legal or HR advice — confirm wages, overtime rules and labor law with your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Every hour you schedule is a dollar you spend — schedule to the forecast.", "section_gold")


# ===========================================================================
def build_schedule(wb):
    ws = wb.create_sheet("Weekly Schedule"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 18, 16, 8] + [7] * 7 + [10, 12, 2])
    luxe_header(ws, "M", "🗓  WEEKLY SCHEDULE",
                "Type each shift's hours — weekly hours & labor cost calculate live for every employee.")
    table_headers(ws, 4, ["Employee", "Role", "Wage"] + DAYS + ["Hours", "Cost"], start_col=2)
    start = L0
    for i, (name, role, wage, hours) in enumerate(SCHEDULE):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=role).style = "td_left"
        cw = ws.cell(row=r, column=4, value=wage); cw.style = "input"; cw.number_format = '"$"#,##0'
        for di, h in enumerate(hours):
            ch = ws.cell(row=r, column=5 + di, value=h if h else None); ch.style = "input"; ch.number_format = "0"
        chh = ws.cell(row=r, column=12, value=f"=SUM(E{r}:K{r})"); chh.style = "td"; chh.font = Font(bold=True, color=PRIMARY); chh.number_format = "#,##0"
        cc = ws.cell(row=r, column=13, value=f"=D{r}*L{r}"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 14):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SCHEDULE) - 1
    nrange(wb, "SchedName", "Weekly Schedule", "B", start, end)
    nrange(wb, "SchedWage", "Weekly Schedule", "D", start, end)
    nrange(wb, "SchedHours", "Weekly Schedule", "L", start, end)
    nrange(wb, "SchedCost", "Weekly Schedule", "M", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    for c in range(3, 12):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    chh = ws.cell(row=tot, column=12, value="=SUM(SchedHours)"); chh.style = "td"; chh.font = Font(bold=True, color=PRIMARY); chh.fill = fill(SURFACE); chh.number_format = "#,##0"
    cc = ws.cell(row=tot, column=13, value="=SUM(SchedCost)"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(MINT_BG); cc.number_format = '"$"#,##0'
    cell_name(wb, "TotalHours", "Weekly Schedule", f"$L${tot}")
    cell_name(wb, "LaborCost", "Weekly Schedule", f"$M${tot}")
    ws.freeze_panes = "E5"


def build_laborcalc(wb):
    ws = wb.create_sheet("Labor Cost Calc"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "🧮  LABOR COST CALC",
                "Scheduled labor ÷ sales = your true labor %. Sales ÷ hours = sales per labor hour.")
    table_headers(ws, 5, ["Line", "Amount"], start_col=2)
    rows = [
        ("Scheduled labor cost (from schedule)", "=LaborCost", "money"),
        ("÷ Weekly sales (from forecast)", "=ForecastSales", "money"),
        ("= LABOR %", "=IFERROR(LaborCost/ForecastSales,0)", "pctbold"),
        ("Total scheduled hours", "=TotalHours", "num"),
        ("Sales per labor hour", "=IFERROR(ForecastSales/TotalHours,0)", "money2b"),
        ("Average wage", "=IFERROR(LaborCost/TotalHours,0)", "money2b"),
    ]
    r0 = 6
    for i, (lab, val, kind) in enumerate(rows):
        r = r0 + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=val)
        if kind == "money":
            c.style = "td"; c.number_format = '"$"#,##0'
        elif kind == "num":
            c.style = "td"; c.number_format = "#,##0"
        elif kind == "pctbold":
            c.style = "td"; c.font = Font(bold=True, size=13, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = "0.0%"
        elif kind == "money2b":
            c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0.00'
        if i % 2 and kind not in ("pctbold",):
            for cc in range(2, 4):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    cell_name(wb, "LaborPct", "Labor Cost Calc", "$C$8")
    cell_name(wb, "SPLH", "Labor Cost Calc", "$C$10")
    cell_name(wb, "AvgWage", "Labor Cost Calc", "$C$11")
    ws.cell(row=r0 + len(rows) + 1, column=2, value="Change a single shift and your labor % updates instantly.").style = "section"


def build_employees(wb):
    ws, start, end = build_log(
        wb, "Employees", "👥", "EMPLOYEES",
        "Your team — role, wage & status. Hours come from the Weekly Schedule.",
        ["Name", "Role", "Wage", "Status"],
        [(n, r, w, s) for (n, r, w, s) in EMPLOYEES], [2, 18, 16, 12, 14, 2],
        text_left={2, 3}, money={4}, reserved=24, start_col=2,
        validations=[("C", "RoleList"), ("E", "StatusList")])
    nrange(wb, "EmpName", "Employees", "B", start, end)


def build_forecast(wb):
    ws = wb.create_sheet("Sales Forecast"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 14, 2])
    luxe_header(ws, "E", "📈  SALES FORECAST",
                "Forecast sales by day → a labor target (at your target %) and the hours it buys.")
    table_headers(ws, 4, ["Day", "Forecast Sales", "Labor Target", "Target Hrs"], start_col=2)
    start = L0
    for i, (day, sales) in enumerate(FORECAST):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        clt = ws.cell(row=r, column=4, value=f"=C{r}*TargetLaborPct"); clt.style = "td"; clt.number_format = '"$"#,##0'
        cth = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/WageGoal,0)"); cth.style = "td"; cth.number_format = "#,##0"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FORECAST) - 1
    nrange(wb, "ForecastDay", "Sales Forecast", "B", start, end)
    nrange(wb, "ForecastAmt", "Sales Forecast", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    cs = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    clt = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); clt.style = "td"; clt.font = Font(bold=True, color=PRIMARY); clt.fill = fill(SURFACE); clt.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "ForecastSales", "Sales Forecast", f"$C${tot}")
    cell_name(wb, "LaborTarget", "Sales Forecast", f"$D${tot}")
    ws.add_chart(_barchart(ws, "Forecast by Day", start, end, 3, 2), "G4")
    ws.freeze_panes = "A5"


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


def build_splh(wb):
    ws = wb.create_sheet("Sales per Labor Hr"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 12, 16, 2])
    luxe_header(ws, "E", "⚡  SALES PER LABOR HOUR",
                "The productivity number — each day's sales ÷ labor hours. Higher is a leaner shift.")
    table_headers(ws, 4, ["Day", "Sales", "Labor Hrs", "SPLH"], start_col=2)
    start = L0
    # labor hours per day = SUMPRODUCT of schedule day column (nonzero) -> count of hours
    sched_end = L0 + len(SCHEDULE) - 1
    for i, (day, sales) in enumerate(FORECAST):
        r = start + i
        daycol = get_column_letter(5 + i)
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        chh = ws.cell(row=r, column=4, value=f"=SUM('Weekly Schedule'!{daycol}{L0}:{daycol}{sched_end})"); chh.style = "td"; chh.number_format = "#,##0"
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FORECAST) - 1
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=30, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=50, mid_color="FFFFF3CD",
                       end_type="num", end_value=70, end_color="FF" + HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_overtime(wb):
    ws = wb.create_sheet("Overtime"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 14, 12, 16, 2])
    luxe_header(ws, "E", "⏰  OVERTIME",
                "Hours over 40 cost time-and-a-half — watch this before it eats your labor budget.")
    table_headers(ws, 4, ["Employee", "OT Hours", "Wage", "OT Cost (1.5×)"], start_col=2)
    start = L0
    for i, (name, ot, wage) in enumerate(OVERTIME):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        co = ws.cell(row=r, column=3, value=ot); co.style = "input"; co.number_format = "#,##0"
        cw = ws.cell(row=r, column=4, value=wage); cw.style = "input"; cw.number_format = '"$"#,##0'
        cc = ws.cell(row=r, column=5, value=f"=C{r}*D{r}*1.5"); cc.style = "td"; cc.font = Font(bold=True, color=DANGER); cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(OVERTIME) - 1
    nrange(wb, "OTHrs", "Overtime", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="OT TOTALS").style = "th"
    co = ws.cell(row=tot, column=3, value="=SUM(OTHrs)"); co.style = "td"; co.font = Font(bold=True, color=PRIMARY); co.fill = fill(SURFACE); co.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cc = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cc.style = "td"; cc.font = Font(bold=True, color=DANGER); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0.00'
    cell_name(wb, "OTHours", "Overtime", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_rates(wb):
    ws = wb.create_sheet("Roles & Rates"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 14, 24, 2])
    luxe_header(ws, "D", "💲  ROLES & RATES",
                "Your pay scale — one wage per role, so every new hire starts on the right number.")
    table_headers(ws, 4, ["Role", "Wage / hr", "Notes"], start_col=2)
    start = L0
    for i, (role, wage, notes) in enumerate(RATES):
        r = start + i
        ws.cell(row=r, column=2, value=role).style = "td_left"
        cw = ws.cell(row=r, column=3, value=wage); cw.style = "input"; cw.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=notes).style = "td_left"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


def build_availability(wb):
    ws, start, end = build_log(
        wb, "Availability", "📅", "AVAILABILITY & TIME-OFF",
        "Who can't work when — check it before you publish the schedule.",
        ["Employee", "Days Off", "Notes"],
        AVAILABILITY, [2, 20, 20, 26, 2], text_left={2, 3, 4}, reserved=24, start_col=2)


def build_prime(wb):
    ws = wb.create_sheet("Prime Cost"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "🎯  PRIME COST",
                "Food cost % + labor % = prime cost — the single number that decides if a restaurant profits.")
    table_headers(ws, 5, ["Line", "Amount"], start_col=2)
    ws.cell(row=6, column=2, value="Food cost % (enter yours)").style = "td_left"
    cf = ws.cell(row=6, column=3, value=FOOD_PCT); cf.style = "input"; cf.number_format = "0.0%"
    cell_name(wb, "FoodPct", "Prime Cost", "$C$6")
    ws.cell(row=7, column=2, value="+ Labor % (from Labor Cost Calc)").style = "td_left"
    cl = ws.cell(row=7, column=3, value="=LaborPct"); cl.style = "td"; cl.number_format = "0.0%"; cl.fill = fill(MUTED_ROW)
    ws.cell(row=7, column=2).fill = fill(MUTED_ROW)
    ws.cell(row=8, column=2, value="= PRIME COST").style = "th"
    cp = ws.cell(row=8, column=3, value="=FoodPct+LaborPct"); cp.style = "td"; cp.font = Font(bold=True, size=13, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = "0.0%"
    cell_name(wb, "PrimeCost", "Prime Cost", "$C$8")
    ws.cell(row=10, column=2, value="Target prime cost").style = "field_label"
    cg = ws.cell(row=10, column=3, value="=PrimeGoal"); cg.style = "field_value"; cg.number_format = "0%"
    ws.cell(row=12, column=2, value="Keep prime cost under 60-65% and there's room for rent, utilities & profit.").style = "section"


def build_tips(wb):
    ws = wb.create_sheet("Tips"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 2])
    luxe_header(ws, "E", "💵  TIPS & TIP-OUT",
                "Daily tips and the back-of-house tip-out — so the pool splits fair, every night.")
    table_headers(ws, 4, ["Day", "Tips", "Tip-out %", "Net Tips"], start_col=2)
    start = L0
    for i, (day, tips, pct) in enumerate(TIPS):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        ct = ws.cell(row=r, column=3, value=tips); ct.style = "input"; ct.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=pct); cp.style = "input"; cp.number_format = "0%"
        cn = ws.cell(row=r, column=5, value=f"=C{r}*(1-D{r})"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(TIPS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    ct = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cn = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


def build_laborday(wb):
    ws = wb.create_sheet("Labor by Day"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 14, 2])
    luxe_header(ws, "D", "📊  LABOR BY DAY",
                "Each day's labor dollars & labor % against the forecast — spot the heavy days.")
    table_headers(ws, 4, ["Day", "Labor $", "Labor %"], start_col=2)
    start = L0
    sched_end = L0 + len(SCHEDULE) - 1
    for i, (day, sales) in enumerate(FORECAST):
        r = start + i
        daycol = get_column_letter(5 + i)
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cl = ws.cell(row=r, column=3, value=f"=SUMPRODUCT('Weekly Schedule'!$D${L0}:$D${sched_end},'Weekly Schedule'!{daycol}{L0}:{daycol}{sched_end})")
        cl.style = "td"; cl.font = Font(bold=True, color=PRIMARY); cl.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/INDEX(ForecastAmt,{i+1}),0)"); cp.style = "td"; cp.number_format = "0.0%"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FORECAST) - 1
    nrange(wb, "LaborDayAmt", "Labor by Day", "C", start, end)
    ws.add_chart(_barchart(ws, "Labor by Day", start, end, 3, 2), "F4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🕐  LABOR & SCHEDULING COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Labor %, sales per labor hour, prime cost & a Score — your whole labor picture, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("LABOR COST", "=LaborCost", "money"),
        ("LABOR %", "=LaborPct", "pct"),
        ("WEEKLY SALES", "=ForecastSales", "money"),
        ("TOTAL HOURS", "=TotalHours", "num"),
        ("SALES / LABOR HR", "=SPLH", "money2"),
        ("EMPLOYEES", "=COUNTA(SchedName)", "num"),
    ]
    row2 = [
        ("LABOR TARGET", "=LaborTarget", "money"),
        ("VS TARGET", "=LaborTarget-LaborCost", "money"),
        ("AVG WAGE", "=AvgWage", "money2"),
        ("PRIME COST", "=PrimeCost", "pct"),
        ("OT HOURS", "=OTHours", "num"),
        ("LABOR SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "LABOR HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "LABOR BY DAY", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Labor on target", "=IFERROR(MIN(TargetLaborPct/LaborPct,1),0)"),
        ("Sales per labor hr healthy", "=IFERROR(MIN(SPLH/SPLHGoal,1),0)"),
        ("Overtime low", "=IFERROR(1-MIN(OTHours/OTLimit,1),0)"),
        ("Prime cost in check", "=IFERROR(MIN(PrimeGoal/PrimeCost,1),0)"),
        ("Schedule covered", "=IFERROR(COUNTIF(SchedHours,\">0\")/COUNTA(SchedName),0)"),
        ("Wage in budget", "=IFERROR(MIN(WageGoal/AvgWage,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    lbd = wb["Labor by Day"]
    ch = BarChart(); ch.type = "col"; ch.title = "Labor by Day"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(lbd, min_col=3, min_row=5, max_row=4 + len(FORECAST)), titles_from_data=False)
    ch.set_categories(Reference(lbd, min_col=2, min_row=5, max_row=4 + len(FORECAST))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Labor & Scheduling Command Center™ — schedule to your sales, protect your prime cost.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_laborcalc(wb); build_schedule(wb)
    build_employees(wb); build_forecast(wb); build_splh(wb); build_overtime(wb)
    build_rates(wb); build_availability(wb); build_prime(wb); build_tips(wb)
    build_laborday(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Labor Cost Calc", "Weekly Schedule", "Employees", "Sales Forecast",
             "Sales per Labor Hr", "Overtime", "Roles & Rates", "Availability", "Prime Cost", "Tips",
             "Labor by Day", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Labor_Scheduling_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
