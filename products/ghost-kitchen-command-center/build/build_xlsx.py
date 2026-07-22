"""Build Ghost Kitchen Command Center™ — The Delivery-Only Business System.

14 tabs · a premium delivery-only / virtual-restaurant operating system in Google
Sheets & Excel. Dashboard, an item-margin engine that shows true net after the
app's commission, a menu with net margin per item, per-platform P&L, virtual
brands, packaging, order volume, inventory, waste, ordering, payouts and promos —
one dashboard. Beat the apps: know your real margin after commission.

Run: python3 build_xlsx.py   ->  ../Ghost_Kitchen_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["each", "lb", "oz", "case", "box", "jug", "bag", "tray"]
PLATFORM_LIST = ["DoorDash", "Uber Eats", "Grubhub", "Direct Web", "Other"]
STATUS = ["Live", "Paused", "Testing", "New"]

TARGET_FC = 0.30
MARGIN_GOAL = 0.40
COMM_LIMIT = 0.40
DIRECT_GOAL = 0.15
COMMISSION = 0.25  # default blended commission for the item-margin engine

# Item-margin engine — flagship Signature Burrito: price, food cost, packaging
SIG_ITEM = ("Signature Burrito", 13.95, 3.80, 0.65)

# Menu & margins: (item, app price, food cost, packaging). Net uses Commission.
MENU = [
    ("Signature Burrito", 13.95, 3.80, 0.65),
    ("Loaded Fries", 8.95, 2.10, 0.45),
    ("Crispy Chicken Sandwich", 11.95, 3.20, 0.55),
    ("Birria Tacos (3)", 14.95, 4.40, 0.70),
    ("Vegan Power Bowl", 12.95, 3.30, 0.60),
    ("Wings (10 pc)", 13.95, 4.10, 0.65),
    ("Loaded Nachos", 10.95, 2.80, 0.55),
    ("Milkshake", 6.95, 1.60, 0.40),
]

# Platform P&L: (platform, orders, avg order value, commission %)
PLATFORMS = [
    ("DoorDash", 320, 22.00, 0.27),
    ("Uber Eats", 240, 21.00, 0.30),
    ("Grubhub", 120, 20.00, 0.25),
    ("Direct Web", 120, 24.00, 0.08),
]

# Virtual brands run from one kitchen: (brand, cuisine, orders, revenue)
BRANDS = [
    ("Burrito Barrio", "Mexican", 300, 6600),
    ("Cluckin' Wings", "Wings", 220, 4900),
    ("Green Fork", "Vegan Bowls", 160, 3300),
    ("Shake Yard", "Shakes & Sweets", 120, 2560),
]

# Packaging: (item, cost/order)
PACKAGING = [
    ("Entrée box + bag", 0.65), ("Fries sleeve", 0.45), ("Sandwich clamshell", 0.55),
    ("Taco tray", 0.70), ("Bowl + lid", 0.60), ("Wing box", 0.65),
    ("Nacho tray", 0.55), ("Shake cup + lid", 0.40),
]

# Order volume: (day, orders, avg order value)
ORDERVOL = [
    ("Monday", 90, 20.00), ("Tuesday", 95, 20.00), ("Wednesday", 110, 21.00),
    ("Thursday", 120, 22.00), ("Friday", 160, 23.00), ("Saturday", 150, 23.00), ("Sunday", 75, 21.00),
]

# Inventory & par: (item, par, on hand, unit)
INVENTORY = [
    ("Tortillas", 40, 18, "case"), ("Chicken", 60, 25, "lb"), ("Ground beef", 50, 20, "lb"),
    ("Fryer oil", 20, 8, "jug"), ("Cheese", 30, 12, "case"), ("Produce", 25, 10, "box"),
    ("Packaging", 40, 15, "case"), ("Beverages", 30, 12, "case"),
]

# Waste log: (item, reason, cost) — total $270
WASTE = [
    ("Prep over-run", "Batch surplus", 95.00),
    ("Cancelled orders", "App glitch", 70.00),
    ("Spoilage", "Cold-chain break", 60.00),
    ("Remakes", "Order error", 45.00),
]

# Ordering / suppliers: (item, supplier, par order, cost)
ORDERING = [
    ("Proteins (case)", "US Foods", 1, 420.00),
    ("Produce (box)", "Green Farms", 2, 180.00),
    ("Packaging (case)", "PackCo", 2, 240.00),
    ("Dry goods", "Restaurant Depot", 1, 160.00),
    ("Beverages", "BevCo", 1, 150.00),
]

# Payouts: (platform, gross, fees, status)
PAYOUTS = [
    ("DoorDash", 7040, 1900, "Received"),
    ("Uber Eats", 5040, 1512, "Received"),
    ("Grubhub", 2400, 600, "Pending"),
    ("Direct Web", 2880, 230, "Received"),
]

# Promotions & ad spend: (promo, platform, spend, orders driven)
PROMOS = [
    ("New-customer 20% off", "DoorDash", 180, 42),
    ("Free delivery", "Uber Eats", 150, 38),
    ("BOGO wings", "Grubhub", 90, 20),
    ("Loyalty 10%", "Direct Web", 60, 25),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your kitchen, targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR KITCHEN", "section")
    controls = [
        ("Kitchen name", "Midnight Kitchen Collective", None, "Kitchen"),
        ("Owner", "Devin", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Net-margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Commission limit %", COMM_LIMIT, "0%", "CommLimit"),
        ("Direct-order goal %", DIRECT_GOAL, "0%", "DirectGoal"),
        ("Default commission %", COMMISSION, "0%", "Commission"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Platform", PLATFORM_LIST, "PlatformList"),
             ("G", "Status", STATUS, "StatusList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  👻  GHOST KITCHEN COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Beat the apps — know your real margin after commission.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE DELIVERY BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Delivery apps take 15–30% of every order — so the menu price is not your margin. This system fixes "
                      "that: an item-margin engine that starts from the app price, subtracts the app's commission, your "
                      "food cost and your packaging, and shows the true net you keep. Run multiple virtual brands from "
                      "one kitchen, see a full P&L for each platform, and track packaging, order volume, inventory, waste, "
                      "ordering, payouts and promos — all in ONE premium Google Sheets & Excel system built for "
                      "delivery-only kitchens.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your kitchen name & default commission %.",
             "2.  Cost an item in Item Margin — see the true net after the app's cut.",
             "3.  Build your Menu — net margin $ and % show on every item.",
             "4.  Enter each platform's orders & commission in Platform P&L.",
             "5.  Add your Virtual Brands, packaging, order volume & promos.",
             "6.  Check the Dashboard: revenue, net payout & a Kitchen Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional ghost kitchen (Midnight Kitchen Collective) is included so you can see how it "
               "all connects — just type over it with your own items and platforms. Commission and food cost are the two "
               "numbers that decide whether a delivery-only kitchen makes money, and they roll into a live Kitchen Score. "
               "Twelve matching printable pages (item-margin card, platform P&L, packaging sheet, prep list & more) are "
               "included. This is a business tool, not financial or legal advice — confirm figures with your own books.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "The menu price is not your margin — the app takes its cut first. Know the real number.", "section_gold")


# ===========================================================================
def build_itemmargin(wb):
    ws = wb.create_sheet("Item Margin"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 16, 2])
    luxe_header(ws, "C", "🧮  ITEM MARGIN",
                "Start from the app price, subtract the commission, food & packaging — see the true net you keep.")
    name, price, food, pack = SIG_ITEM
    ws.cell(row=5, column=2, value="ITEM").style = "section_gold"
    ws.cell(row=5, column=3, value=name).font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Line", "Amount"], start_col=2)
    ws.cell(row=7, column=2, value="App menu price").style = "td_left"
    cp = ws.cell(row=7, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
    ws.cell(row=8, column=2, value="App commission (=Commission)").style = "td_left"
    cc = ws.cell(row=8, column=3, value="=-C7*Commission"); cc.style = "td"; cc.number_format = '"$"#,##0.00'; cc.fill = fill(MUTED_ROW)
    ws.cell(row=8, column=2).fill = fill(MUTED_ROW)
    ws.cell(row=9, column=2, value="Food cost").style = "td_left"
    cf = ws.cell(row=9, column=3, value=-food); cf.style = "input"; cf.number_format = '"$"#,##0.00'
    ws.cell(row=10, column=2, value="Packaging").style = "td_left"
    cpk = ws.cell(row=10, column=3, value=-pack); cpk.style = "input"; cpk.number_format = '"$"#,##0.00'; cpk.fill = fill(MUTED_ROW)
    ws.cell(row=10, column=2).fill = fill(MUTED_ROW)
    ws.cell(row=11, column=2, value="NET MARGIN YOU KEEP").style = "th"
    cn = ws.cell(row=11, column=3, value="=C7+C8+C9+C10"); cn.style = "td"
    cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0.00'
    cell_name(wb, "SigItemNet", "Item Margin", "$C$11")
    ws.cell(row=13, column=2, value="Net margin % of app price").style = "field_label"
    cpct = ws.cell(row=13, column=3, value="=IFERROR(C11/C7,0)"); cpct.style = "field_value"; cpct.number_format = "0%"; cpct.fill = fill(MINT_BG)
    ws.cell(row=15, column=2, value="The app takes its cut FIRST — this is the number that actually pays your rent.").style = "section"


def build_menu(wb):
    ws = wb.create_sheet("Menu & Margins"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 12, 12, 13, 12, 12, 2])
    luxe_header(ws, "H", "🍔  MENU & MARGINS",
                "App price, food, packaging & the app's commission — the true net margin on every item.")
    table_headers(ws, 4, ["Item", "App Price", "Food", "Packaging", "Net Margin", "Net %", "Food %"], start_col=2)
    start = L0
    for i, (item, price, food, pack) in enumerate(MENU):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cf = ws.cell(row=r, column=4, value=food); cf.style = "input"; cf.number_format = '"$"#,##0.00'
        ck = ws.cell(row=r, column=5, value=pack); ck.style = "input"; ck.number_format = '"$"#,##0.00'
        cn = ws.cell(row=r, column=6, value=f"=C{r}*(1-Commission)-D{r}-E{r}"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = '"$"#,##0.00'
        cnp = ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/C{r},0)"); cnp.style = "td"; cnp.number_format = "0%"
        cfp = ws.cell(row=r, column=8, value=f"=IFERROR(D{r}/C{r},0)"); cfp.style = "td"; cfp.number_format = "0%"
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MENU) - 1
    nrange(wb, "MenuItem", "Menu & Margins", "B", start, end)
    nrange(wb, "MenuPrice", "Menu & Margins", "C", start, end)
    nrange(wb, "MenuFood", "Menu & Margins", "D", start, end)
    nrange(wb, "MenuNet", "Menu & Margins", "F", start, end)
    nrange(wb, "MenuNetPct", "Menu & Margins", "G", start, end)
    nrange(wb, "MenuFoodPct", "Menu & Margins", "H", start, end)
    br = end + 2
    ws.cell(row=br, column=2, value="Avg net margin %").style = "field_label"
    c = ws.cell(row=br, column=7, value="=IFERROR(AVERAGE(MenuNetPct),0)"); c.style = "field_value"; c.number_format = "0%"; c.fill = fill(MINT_BG)
    cell_name(wb, "AvgNetPct", "Menu & Margins", f"$G${br}")
    ws.cell(row=br + 1, column=2, value="Avg food cost %").style = "field_label"
    c2 = ws.cell(row=br + 1, column=7, value="=IFERROR(AVERAGE(MenuFoodPct),0)"); c2.style = "field_value"; c2.number_format = "0%"; c2.fill = fill(MINT_BG)
    cell_name(wb, "FoodCostPct", "Menu & Margins", f"$G${br+1}")
    ws.conditional_formatting.add(f"G{start}:G{end}",
        ColorScaleRule(start_type="num", start_value=0.25, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=0.40, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.55, end_color="FF" + HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_platforms(wb):
    ws = wb.create_sheet("Platform P&L"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 12, 14, 14, 14, 14, 2])
    luxe_header(ws, "G", "📱  PLATFORM P&L",
                "Every platform's orders, commission & net payout — see what each app really pays you.")
    table_headers(ws, 4, ["Platform", "Orders", "Avg Order", "Gross", "Commission", "Net Payout"], start_col=2)
    start = L0
    for i, (plat, orders, aov, comm) in enumerate(PLATFORMS):
        r = start + i
        ws.cell(row=r, column=2, value=plat).style = "td_left"
        co = ws.cell(row=r, column=3, value=orders); co.style = "input"; co.number_format = "#,##0"
        ca = ws.cell(row=r, column=4, value=aov); ca.style = "input"; ca.number_format = '"$"#,##0.00'
        cg = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); cg.style = "td"; cg.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=6, value=comm); cm.style = "input"; cm.number_format = "0%"
        cn = ws.cell(row=r, column=7, value=f"=E{r}*(1-F{r})"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PLATFORMS) - 1
    nrange(wb, "PlatName", "Platform P&L", "B", start, end)
    nrange(wb, "PlatOrders", "Platform P&L", "C", start, end)
    nrange(wb, "PlatGross", "Platform P&L", "E", start, end)
    nrange(wb, "PlatNet", "Platform P&L", "G", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    co = ws.cell(row=tot, column=3, value="=SUM(PlatOrders)"); co.style = "td"; co.font = Font(bold=True, color=PRIMARY); co.fill = fill(SURFACE); co.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cg = ws.cell(row=tot, column=5, value="=SUM(PlatGross)"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(SURFACE); cg.number_format = '"$"#,##0'
    cm = ws.cell(row=tot, column=6, value=f"=IFERROR((E{tot}-G{tot})/E{tot},0)"); cm.style = "td"; cm.font = Font(bold=True, color=DANGER); cm.fill = fill(SURFACE); cm.number_format = "0.0%"
    cn = ws.cell(row=tot, column=7, value="=SUM(PlatNet)"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = '"$"#,##0'
    cell_name(wb, "TotalOrders", "Platform P&L", f"$C${tot}")
    cell_name(wb, "TotalGross", "Platform P&L", f"$E${tot}")
    cell_name(wb, "BlendedComm", "Platform P&L", f"$F${tot}")
    cell_name(wb, "NetPayout", "Platform P&L", f"$G${tot}")
    direct_row = start + 3
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Direct-order share").style = "field_label"
    c = ws.cell(row=sr, column=3, value=f"=IFERROR(C{direct_row}/TotalOrders,0)"); c.style = "field_value"; c.number_format = "0%"; c.fill = fill(MINT_BG)
    cell_name(wb, "DirectShare", "Platform P&L", f"$C${sr}")
    ws.cell(row=sr + 1, column=2, value="Avg order value").style = "field_label"
    c2 = ws.cell(row=sr + 1, column=3, value="=IFERROR(TotalGross/TotalOrders,0)"); c2.style = "field_value"; c2.number_format = '"$"#,##0.00'; c2.fill = fill(MINT_BG)
    cell_name(wb, "AvgOrder", "Platform P&L", f"$C${sr+1}")
    ws.conditional_formatting.add(f"F{start}:F{end}",
        CellIsRule(operator="greaterThan", formula=["0.28"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


def build_brands(wb):
    ws = wb.create_sheet("Virtual Brands"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 20, 12, 14, 14, 2])
    luxe_header(ws, "F", "🏷  VIRTUAL BRANDS",
                "Run many brands from one kitchen — orders, revenue & share, all from the same line.")
    table_headers(ws, 4, ["Brand", "Cuisine", "Orders", "Revenue", "Rev Share"], start_col=2)
    start = L0
    for i, (brand, cuisine, orders, rev) in enumerate(BRANDS):
        r = start + i
        ws.cell(row=r, column=2, value=brand).style = "td_left"
        ws.cell(row=r, column=3, value=cuisine).style = "td_left"
        co = ws.cell(row=r, column=4, value=orders); co.style = "input"; co.number_format = "#,##0"
        cr = ws.cell(row=r, column=5, value=rev); cr.style = "input"; cr.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/SUM(BrandRev),0)"); cs.style = "td"; cs.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BRANDS) - 1
    nrange(wb, "BrandName", "Virtual Brands", "B", start, end)
    nrange(wb, "BrandOrders", "Virtual Brands", "D", start, end)
    nrange(wb, "BrandRev", "Virtual Brands", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL BRANDS").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    co = ws.cell(row=tot, column=4, value="=SUM(BrandOrders)"); co.style = "td"; co.font = Font(bold=True, color=PRIMARY); co.fill = fill(SURFACE); co.number_format = "#,##0"
    cr = ws.cell(row=tot, column=5, value="=SUM(BrandRev)"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = '"$"#,##0'
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


def build_packaging(wb):
    ws = wb.create_sheet("Packaging"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 16, 2])
    luxe_header(ws, "C", "📦  PACKAGING",
                "Every order ships in a box — packaging is a real cost per order, so track it here.")
    table_headers(ws, 4, ["Package", "Cost / Order"], start_col=2)
    start = L0
    for i, (item, cost) in enumerate(PACKAGING):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PACKAGING) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="AVG PACKAGING / ORDER").style = "th"
    c = ws.cell(row=tot, column=3, value=f"=AVERAGE(C{start}:C{end})"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    ws.freeze_panes = "A5"


def _ordervol_chart(ws, start, end):
    ch = BarChart(); ch.type = "col"; ch.title = "Orders by Day"; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=3, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


def build_ordervol(wb):
    ws = wb.create_sheet("Order Volume"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 2])
    luxe_header(ws, "E", "📊  ORDER VOLUME",
                "Orders & average order value by day — see when the delivery rush really hits.")
    table_headers(ws, 4, ["Day", "Orders", "Avg Order", "Gross"], start_col=2)
    start = L0
    for i, (day, orders, aov) in enumerate(ORDERVOL):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        co = ws.cell(row=r, column=3, value=orders); co.style = "input"; co.number_format = "#,##0"
        ca = ws.cell(row=r, column=4, value=aov); ca.style = "input"; ca.number_format = '"$"#,##0.00'
        cg = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(ORDERVOL) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    co = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); co.style = "td"; co.font = Font(bold=True, color=PRIMARY); co.fill = fill(SURFACE); co.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cg = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(SURFACE); cg.number_format = '"$"#,##0'
    ws.add_chart(_ordervol_chart(ws, start, end), "G4")
    ws.freeze_panes = "A5"


def build_inventory(wb):
    ws = wb.create_sheet("Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 12, 12, 14, 2])
    luxe_header(ws, "F", "📦  INVENTORY & PAR",
                "Par vs on hand — never 86 an item mid-rush when the orders are stacking up.")
    table_headers(ws, 4, ["Item", "Par", "On Hand", "Unit", "To Order"], start_col=2)
    start = L0
    for i, (item, par, oh, unit) in enumerate(INVENTORY):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=par); cp.style = "input"; cp.number_format = "#,##0"
        co = ws.cell(row=r, column=4, value=oh); co.style = "input"; co.number_format = "#,##0"
        ws.cell(row=r, column=5, value=unit).style = "td"
        cb = ws.cell(row=r, column=6, value=f"=MAX(C{r}-D{r},0)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.number_format = "#,##0"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVENTORY) - 1
    add_dv(ws, f"E{start}:E{end}", "UnitList")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=[f"C{start}*0.5"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Cancelled orders, remakes & spoilage — the delivery-only leaks that eat your net.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_ordering(wb):
    ws, start, end = build_log(
        wb, "Ordering", "🚚", "ORDERING & SUPPLIERS",
        "Your standing order — par quantities & cost, by supplier.",
        ["Item", "Supplier", "Par Order", "Cost"],
        ORDERING, [2, 26, 20, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)
    nrange(wb, "OrderCost", "Ordering", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ORDER TOTAL").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=5, value="=SUM(OrderCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_payouts(wb):
    ws = wb.create_sheet("Payouts"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 14, 14, 14, 14, 2])
    luxe_header(ws, "F", "💵  PAYOUTS",
                "What each platform actually deposits — gross, fees & the net that lands in your bank.")
    table_headers(ws, 4, ["Platform", "Gross", "Fees", "Net Payout", "Status"], start_col=2)
    start = L0
    for i, (plat, gross, fees, status) in enumerate(PAYOUTS):
        r = start + i
        ws.cell(row=r, column=2, value=plat).style = "td_left"
        cg = ws.cell(row=r, column=3, value=gross); cg.style = "input"; cg.number_format = '"$"#,##0'
        cf = ws.cell(row=r, column=4, value=fees); cf.style = "input"; cf.number_format = '"$"#,##0'
        cn = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=6, value=status); cs.style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"F{r}", "StatusList")
    end = start + len(PAYOUTS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL NET PAYOUT").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cn = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = '"$"#,##0'
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


def build_promos(wb):
    ws = wb.create_sheet("Promotions"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 12, 14, 14, 2])
    luxe_header(ws, "F", "🎯  PROMOTIONS",
                "Promo & ad spend by platform — cost per order driven, so you only pay for what works.")
    table_headers(ws, 4, ["Promotion", "Platform", "Spend", "Orders", "Cost / Order"], start_col=2)
    start = L0
    for i, (promo, plat, spend, orders) in enumerate(PROMOS):
        r = start + i
        ws.cell(row=r, column=2, value=promo).style = "td_left"
        ws.cell(row=r, column=3, value=plat).style = "td"
        cs = ws.cell(row=r, column=4, value=spend); cs.style = "input"; cs.number_format = '"$"#,##0'
        co = ws.cell(row=r, column=5, value=orders); co.style = "input"; co.number_format = "#,##0"
        cc = ws.cell(row=r, column=6, value=f"=IFERROR(D{r}/E{r},0)"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "PlatformList")
    end = start + len(PROMOS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL PROMO SPEND").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cs = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    co = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); co.style = "td"; co.font = Font(bold=True, color=PRIMARY); co.fill = fill(SURFACE); co.number_format = "#,##0"
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  👻  GHOST KITCHEN COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Revenue, commission, net payout & a Kitchen Score — your whole delivery business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MENU ITEMS", "=COUNTA(MenuItem)", "num"),
        ("AVG APP PRICE", "=AVERAGE(MenuPrice)", "money2"),
        ("FOOD COST", "=FoodCostPct", "pct"),
        ("AVG NET MARGIN", "=AvgNetPct", "pct"),
        ("BLENDED COMM", "=BlendedComm", "pct"),
        ("TOP ITEM", "=INDEX(MenuItem,MATCH(MAX(MenuNet),MenuNet,0))", "text"),
    ]
    row2 = [
        ("WEEKLY ORDERS", "=TotalOrders", "num"),
        ("WEEKLY REVENUE", "=TotalGross", "money"),
        ("NET PAYOUT", "=NetPayout", "money"),
        ("AVG ORDER", "=AvgOrder", "money2"),
        ("VIRTUAL BRANDS", "=COUNTA(BrandName)", "num"),
        ("KITCHEN SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "KITCHEN HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "NET PAYOUT BY PLATFORM", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Food cost on target", "=IFERROR(MIN(TargetFC/FoodCostPct,1),0)"),
        ("Net margin healthy", "=IFERROR(MIN(AvgNetPct/MarginGoal,1),0)"),
        ("Menu fully costed", "=IFERROR(COUNTIF(MenuFood,\">0\")/COUNTA(MenuItem),0)"),
        ("Commission in control", "=IFERROR(1-MIN(BlendedComm/CommLimit,1),0)"),
        ("Direct-order mix", "=IFERROR(MIN(DirectShare/DirectGoal,1),0)"),
        ("Gross margin", "=IFERROR(MIN((1-FoodCostPct)/0.7,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    pl = wb["Platform P&L"]
    ch = BarChart(); ch.type = "col"; ch.title = "Net Payout by Platform"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(pl, min_col=7, min_row=5, max_row=4 + len(PLATFORMS)), titles_from_data=False)
    ch.set_categories(Reference(pl, min_col=2, min_row=5, max_row=4 + len(PLATFORMS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Ghost Kitchen Command Center™ — beat the apps, know your real margin after commission.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_itemmargin(wb); build_menu(wb)
    build_platforms(wb); build_brands(wb); build_packaging(wb); build_ordervol(wb)
    build_inventory(wb); build_waste(wb); build_ordering(wb); build_payouts(wb)
    build_promos(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Item Margin", "Menu & Margins", "Platform P&L", "Virtual Brands",
             "Packaging", "Order Volume", "Inventory", "Waste Log", "Ordering", "Payouts",
             "Promotions", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Ghost_Kitchen_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
