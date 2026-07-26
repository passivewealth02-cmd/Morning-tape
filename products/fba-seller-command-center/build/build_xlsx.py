"""Build Amazon FBA & Online Seller Profit Command Center™ — The Seller's Operating System.

14 tabs · a premium Amazon FBA / online-seller operating system in Google Sheets & Excel.
Dashboard, a true-profit engine (sale price − referral − FBA − storage − COGS − inbound →
net per unit, margin and ROI), a product catalog, a fee breakdown, inventory & reorder,
sales, PPC & ACoS, returns, suppliers, reviews, expenses and a monthly summary — one
dashboard. Know your real profit per unit, after every fee Amazon takes.

Run: python3 build_xlsx.py   ->  ../FBA_Seller_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
CHANNEL = ["Amazon FBA", "Amazon FBM", "Shopify", "Walmart", "eBay", "Etsy"]
STOCKSTATUS = ["In stock", "Low", "Reorder now", "Out of stock"]
RETURNREASON = ["Damaged", "Wrong item", "Didn't fit", "Changed mind", "Defective", "Other"]

# Profit engine — flagship SKU
SALE_PRICE = 29.99
REFERRAL_RATE = 0.15
FBA_FEE = 5.40
STORAGE_FEE = 0.35
UNIT_COGS = 7.50
INBOUND_SHIP = 1.20

# Goals
MARGIN_GOAL = 0.30
ROI_GOAL = 1.00
ACOS_GOAL = 0.25
COVER_GOAL = 60
REVIEW_GOAL = 10
SKU_GOAL = 8

UNITS_MONTH = 420
AD_SPEND = 1200
AD_SALES = 6000
ON_HAND = 840
DAILY_VELOCITY = 14
AVG_REVIEWS = 4

# Product catalog: (SKU, product, price, COGS, units/mo)
PRODUCTS = [
    ("NP-CAN-08", "Soy Candle 8oz", 29.99, 7.50, 420), ("NP-DIF-01", "Reed Diffuser", 34.99, 9.00, 180),
    ("NP-MLT-06", "Wax Melts 6pk", 16.99, 4.20, 260), ("NP-SPR-04", "Room Spray 4oz", 14.99, 3.60, 210),
    ("NP-SET-03", "Gift Set Trio", 69.99, 21.00, 95), ("NP-REF-02", "Refill Pouch", 12.99, 3.10, 150),
    ("NP-WIC-20", "Wick Trimmer", 11.99, 2.40, 130), ("NP-TRV-02", "Travel Tin 2oz", 9.99, 2.20, 175),
]

# Fee breakdown reference: (fee, how it works)
FEES = [
    ("Referral fee", "% of sale price — most categories 15%"),
    ("FBA fulfilment fee", "Flat, by size tier & weight"),
    ("Monthly storage", "Per cubic foot, higher Oct–Dec"),
    ("Long-term storage", "Charged after 181 days — watch aged stock"),
    ("Inbound shipping", "Your cost to send units to Amazon"),
    ("Returns processing", "Charged on some categories"),
    ("Advertising (PPC)", "Not a fee — but it comes off the same profit"),
]

# Inventory: (SKU, on hand, daily velocity, reorder at)
INVENTORY = [
    ("NP-CAN-08", 840, 14, 420), ("NP-DIF-01", 300, 6, 180), ("NP-MLT-06", 380, 9, 270),
    ("NP-SPR-04", 200, 7, 210), ("NP-SET-03", 160, 3, 90), ("NP-REF-02", 240, 5, 150),
    ("NP-WIC-20", 190, 4, 120), ("NP-TRV-02", 310, 6, 180),
]

# Sales log: (month, units, revenue)
SALES = [
    ("Feb", 300, 8997), ("Mar", 330, 9897), ("Apr", 360, 10796),
    ("May", 380, 11396), ("Jun", 400, 11996), ("Jul", 420, 12596),
]

# PPC: (campaign, spend, sales, ACoS)
PPC = [
    ("Auto — Candles", 380, 2100, 0.181), ("Exact — soy candle", 300, 1650, 0.182),
    ("Broad — home fragrance", 260, 1100, 0.236), ("Product targeting", 160, 700, 0.229),
    ("Brand defence", 100, 450, 0.222),
]

# Returns: (SKU, units returned, reason)
RETURNS = [
    ("NP-CAN-08", 6, "Damaged"), ("NP-DIF-01", 3, "Changed mind"), ("NP-MLT-06", 2, "Didn't fit"),
    ("NP-SET-03", 2, "Damaged"), ("NP-SPR-04", 1, "Defective"),
]

# Suppliers / POs: (supplier, SKU, units, unit cost, status)
SUPPLIERS = [
    ("Cedar Supply Co.", "NP-CAN-08", 1000, 7.50, "Shipped"), ("Glasswork Ltd", "NP-DIF-01", 400, 9.00, "In production"),
    ("Cedar Supply Co.", "NP-MLT-06", 600, 4.20, "Delivered"), ("Packwell", "NP-SPR-04", 500, 3.60, "Ordered"),
]

# Reviews: (SKU, reviews, avg rating)
REVIEWS = [
    ("NP-CAN-08", 9, 4.6), ("NP-DIF-01", 5, 4.4), ("NP-MLT-06", 4, 4.7), ("NP-SPR-04", 3, 4.3),
    ("NP-SET-03", 2, 4.8), ("NP-REF-02", 3, 4.5), ("NP-WIC-20", 4, 4.2), ("NP-TRV-02", 2, 4.6),
]

# Expenses (monthly): (item, amount)
EXPENSES = [
    ("Amazon PPC", 1200), ("Software & tools", 180), ("Photography", 150),
    ("Samples & testing", 120), ("Accounting", 90), ("Misc", 80),
]

# Monthly summary: (month, net profit)
MONTHS = [("Feb", 3312), ("Mar", 3643), ("Apr", 3974), ("May", 4195), ("Jun", 4416), ("Jul", 4637)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 20, 3] + [20] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your fee rates & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR RATES & GOALS", "section")
    controls = [
        ("Brand name", "Northport Goods", None, "BrandName"),
        ("Owner", "Sam", None, "Owner"),
        ("Referral fee rate", REFERRAL_RATE, "0%", "ReferralRate"),
        ("Net-margin goal", MARGIN_GOAL, "0%", "MarginGoal"),
        ("ROI goal", ROI_GOAL, "0%", "ROIGoal"),
        ("ACoS goal (max)", ACOS_GOAL, "0%", "ACoSGoal"),
        ("Days-of-cover goal", COVER_GOAL, "0", "CoverGoal"),
        ("Reviews-per-SKU goal", REVIEW_GOAL, "0", "ReviewGoal"),
        ("SKU count goal", SKU_GOAL, "0", "SKUGoal"),
        ("Ad spend (month)", AD_SPEND, '"$"#,##0', "AdSpend"),
        ("Ad-attributed sales", AD_SALES, '"$"#,##0', "AdSales"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Channel", CHANNEL, "ChannelList"), ("F", "Stock status", STOCKSTATUS, "StockStatusList"),
             ("G", "Return reason", RETURNREASON, "ReturnReasonList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  📦  AMAZON FBA & ONLINE SELLER PROFIT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know your real profit per unit — after every fee Amazon takes.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE SELLER BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A $29.99 sale is not $29.99 of revenue. By the time the referral fee, the FBA fulfilment fee, "
                      "storage, your cost of goods and inbound shipping come out, what's left can be a fraction of what "
                      "you thought — and PPC comes off that too. This makes it exact: a true-profit engine strips out "
                      "every fee to show your net per unit, your margin and your ROI. Then run your catalog, inventory "
                      "and reorder points, PPC and ACoS, returns, suppliers and reviews — all in ONE premium Google "
                      "Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your referral rate and goals.",
             "2.  Run the Profit Calculator on your best-selling SKU.",
             "3.  Read your net per unit, margin and ROI. That's the truth.",
             "4.  Load your Product Catalog and Inventory reorder points.",
             "5.  Track PPC & ACoS, returns, suppliers and reviews.",
             "6.  Check the Dashboard: profit, ACoS & a Seller Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional brand (Northport Goods, owner Sam) is included so you can see how it all "
               "connects — just type over it with your own SKUs and fees. Your net profit per unit and your ROI are the "
               "two numbers that decide whether a product is worth selling at all, and they roll into a live Seller "
               "Score. Amazon's fee schedule changes — check your own Seller Central fee preview and update Settings. "
               "Twelve matching printable pages (profit worksheet, SKU sheet, reorder planner, PPC log & more) are "
               "included. This is a business & organizing tool, not financial, tax or accounting advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Revenue is vanity. Net per unit after fees is the only number that pays you.", "section_gold")


# ===========================================================================
def build_profit(wb):
    ws = wb.create_sheet("Profit Calculator"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 18, 2])
    luxe_header(ws, "C", "🧮  PROFIT CALCULATOR",
                "Sale price minus every fee, your cost and inbound shipping — the real money you keep per unit.")
    ws.cell(row=5, column=2, value="THE SALE").style = "section_gold"
    ws.cell(row=6, column=2, value="Sale price").style = "field_label"
    cp = ws.cell(row=6, column=3, value=SALE_PRICE); cp.style = "input"; cp.number_format = '"$"#,##0.00'
    cell_name(wb, "SalePrice", "Profit Calculator", "$C$6")
    ws.cell(row=8, column=2, value="WHAT AMAZON TAKES").style = "section_gold"
    ws.cell(row=9, column=2, value="− Referral fee (% of price)").style = "field_label"
    cr = ws.cell(row=9, column=3, value="=ROUND(SalePrice*ReferralRate,2)"); cr.style = "field_value"; cr.number_format = '"$"#,##0.00'
    cell_name(wb, "ReferralFee", "Profit Calculator", "$C$9")
    ws.cell(row=10, column=2, value="− FBA fulfilment fee").style = "field_label"
    cf = ws.cell(row=10, column=3, value=FBA_FEE); cf.style = "input"; cf.number_format = '"$"#,##0.00'
    cell_name(wb, "FBAFee", "Profit Calculator", "$C$10")
    ws.cell(row=11, column=2, value="− Monthly storage (per unit)").style = "field_label"
    cs = ws.cell(row=11, column=3, value=STORAGE_FEE); cs.style = "input"; cs.number_format = '"$"#,##0.00'
    cell_name(wb, "StorageFee", "Profit Calculator", "$C$11")
    ws.cell(row=12, column=2, value="= TOTAL AMAZON FEES").style = "th"
    ctf = ws.cell(row=12, column=3, value="=ReferralFee+FBAFee+StorageFee"); ctf.style = "td"; ctf.font = Font(bold=True, size=12, color=DANGER); ctf.fill = fill(SURFACE); ctf.number_format = '"$"#,##0.00'
    cell_name(wb, "TotalFees", "Profit Calculator", "$C$12")
    ws.cell(row=14, column=2, value="WHAT THE PRODUCT COSTS YOU").style = "section_gold"
    ws.cell(row=15, column=2, value="− Cost of goods (per unit)").style = "field_label"
    cc = ws.cell(row=15, column=3, value=UNIT_COGS); cc.style = "input"; cc.number_format = '"$"#,##0.00'
    cell_name(wb, "UnitCOGS", "Profit Calculator", "$C$15")
    ws.cell(row=16, column=2, value="− Inbound shipping (per unit)").style = "field_label"
    ci = ws.cell(row=16, column=3, value=INBOUND_SHIP); ci.style = "input"; ci.number_format = '"$"#,##0.00'
    cell_name(wb, "InboundShip", "Profit Calculator", "$C$16")
    ws.cell(row=17, column=2, value="= YOUR LANDED COST").style = "th"
    cl = ws.cell(row=17, column=3, value="=UnitCOGS+InboundShip"); cl.style = "td"; cl.font = Font(bold=True, color=PRIMARY); cl.fill = fill(SURFACE); cl.number_format = '"$"#,##0.00'
    cell_name(wb, "LandedCost", "Profit Calculator", "$C$17")
    ws.cell(row=19, column=2, value="= NET PROFIT PER UNIT").style = "th"
    cn = ws.cell(row=19, column=3, value="=SalePrice-TotalFees-LandedCost"); cn.style = "td"; cn.font = Font(bold=True, size=14, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0.00'
    cell_name(wb, "NetPerUnit", "Profit Calculator", "$C$19")
    ws.cell(row=20, column=2, value="= NET MARGIN").style = "th"
    cm = ws.cell(row=20, column=3, value="=IFERROR(NetPerUnit/SalePrice,0)"); cm.style = "td"; cm.font = Font(bold=True, size=13, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = "0.0%"
    cell_name(wb, "NetMargin", "Profit Calculator", "$C$20")
    ws.cell(row=21, column=2, value="= ROI (return on your cash)").style = "th"
    cro = ws.cell(row=21, column=3, value="=IFERROR(NetPerUnit/LandedCost,0)"); cro.style = "td"; cro.font = Font(bold=True, size=13, color=PRIMARY); cro.fill = fill(MINT_BG); cro.number_format = "0.0%"
    cell_name(wb, "ROI", "Profit Calculator", "$C$21")
    ws.cell(row=23, column=2, value="Units sold per month").style = "field_label"
    cu = ws.cell(row=23, column=3, value=UNITS_MONTH); cu.style = "input"; cu.number_format = "#,##0"
    cell_name(wb, "UnitsMonth", "Profit Calculator", "$C$23")
    ws.cell(row=24, column=2, value="= MONTHLY PROFIT (this SKU)").style = "th"
    cmp = ws.cell(row=24, column=3, value="=NetPerUnit*UnitsMonth"); cmp.style = "td"; cmp.font = Font(bold=True, size=13, color=PRIMARY); cmp.fill = fill(MINT_BG); cmp.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyProfit", "Profit Calculator", "$C$24")
    ws.cell(row=25, column=2, value="= MONTHLY REVENUE (this SKU)").style = "field_label"
    cmr = ws.cell(row=25, column=3, value="=SalePrice*UnitsMonth"); cmr.style = "field_value"; cmr.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyRevenue", "Profit Calculator", "$C$25")
    ws.cell(row=27, column=2, value="A 30%+ margin and 100%+ ROI is a product worth scaling. Below that, fix it or drop it.").style = "section_gold"


def build_catalog(wb):
    ws = wb.create_sheet("Product Catalog"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 24, 12, 12, 12, 14, 14, 2])
    luxe_header(ws, "H", "🏷  PRODUCT CATALOG",
                "Every SKU with its price, cost and volume — and the estimated net each one throws off.")
    table_headers(ws, 4, ["SKU", "Product", "Price", "COGS", "Units/mo", "Est. Net/Unit", "Est. Monthly"], start_col=2)
    start = L0
    for i, (sku, prod, price, cogs, units) in enumerate(PRODUCTS):
        r = start + i
        ws.cell(row=r, column=2, value=sku).style = "td_left"
        ws.cell(row=r, column=3, value=prod).style = "td_left"
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cc = ws.cell(row=r, column=5, value=cogs); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cu = ws.cell(row=r, column=6, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cn = ws.cell(row=r, column=7, value=f"=D{r}-ROUND(D{r}*ReferralRate,2)-FBAFee-StorageFee-E{r}-InboundShip")
        cn.style = "td"; cn.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=8, value=f"=G{r}*F{r}"); cm.style = "td"; cm.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRODUCTS) - 1
    nrange(wb, "SKUName", "Product Catalog", "B", start, end)
    nrange(wb, "SKUNet", "Product Catalog", "G", start, end)
    nrange(wb, "SKUMonthly", "Product Catalog", "H", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL SKUS").style = "th"
    for c in (3, 4, 5, 6, 7):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cm = ws.cell(row=tot, column=8, value="=SUM(SKUMonthly)"); cm.style = "td"; cm.font = Font(bold=True, size=12, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = '"$"#,##0'
    cell_name(wb, "AllSKUProfit", "Product Catalog", f"$H${tot}")
    ws.cell(row=tot + 2, column=2, value="SKUs listed").style = "field_label"
    cs = ws.cell(row=tot + 2, column=4, value="=COUNTA(SKUName)"); cs.style = "field_value"; cs.number_format = "#,##0"
    cell_name(wb, "SKUCount", "Product Catalog", f"$D${tot+2}")
    ws.freeze_panes = "A5"


def build_fees(wb):
    ws = wb.create_sheet("Fee Breakdown"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 46, 2])
    luxe_header(ws, "C", "💸  FEE BREAKDOWN",
                "Every fee Amazon charges, in plain English — so nothing on your statement is a mystery.")
    table_headers(ws, 4, ["Fee", "How it works"], start_col=2)
    start = L0
    for i, (fee, how) in enumerate(FEES):
        r = start + i
        ws.cell(row=r, column=2, value=fee).style = "td_left"
        ws.cell(row=r, column=3, value=how).style = "td_left"
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(FEES) - 1
    ws.cell(row=end + 2, column=2, value="ON THIS SKU").style = "section_gold"
    rows = [("Referral fee", "=ReferralFee"), ("FBA fulfilment", "=FBAFee"),
            ("Storage", "=StorageFee"), ("Total fees", "=TotalFees"),
            ("Fees as % of price", "=IFERROR(TotalFees/SalePrice,0)")]
    for i, (lab, f) in enumerate(rows):
        r = end + 3 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=f); c.style = "field_value"
        c.number_format = "0.0%" if "%" in lab else '"$"#,##0.00'
        if lab == "Total fees":
            c.fill = fill(WARN_BG)
    ws.freeze_panes = "A5"


def build_inventory(wb):
    ws = wb.create_sheet("Inventory"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 16, 14, 16, 16, 2])
    luxe_header(ws, "G", "📦  INVENTORY & REORDER",
                "Units on hand versus how fast they sell — the days-of-cover number that stops a stockout.")
    table_headers(ws, 4, ["SKU", "On Hand", "Units / Day", "Reorder At", "Days of Cover", "Status"], start_col=2)
    start = L0
    for i, (sku, onhand, vel, reorder) in enumerate(INVENTORY):
        r = start + i
        ws.cell(row=r, column=2, value=sku).style = "td_left"
        co = ws.cell(row=r, column=3, value=onhand); co.style = "input"; co.number_format = "#,##0"
        cv = ws.cell(row=r, column=4, value=vel); cv.style = "input"; cv.number_format = "#,##0"
        cr = ws.cell(row=r, column=5, value=reorder); cr.style = "input"; cr.number_format = "#,##0"
        cd = ws.cell(row=r, column=6, value=f"=IFERROR(C{r}/D{r},0)"); cd.style = "td"; cd.number_format = "#,##0"
        cst = ws.cell(row=r, column=7, value=f'=IF(C{r}<=0,"Out of stock",IF(C{r}<=E{r},"Reorder now",IF(F{r}<30,"Low","In stock")))')
        cst.style = "td"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVENTORY) - 1
    nrange(wb, "DaysCover", "Inventory", "F", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        ColorScaleRule(start_type="num", start_value=15, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=45, mid_color="FFFFF3CD",
                       end_type="num", end_value=90, end_color="FF" + HIGHLIGHT))
    ws.cell(row=end + 2, column=2, value="Flagship days of cover").style = "field_label"
    cf = ws.cell(row=end + 2, column=6, value=f"=F{start}"); cf.style = "field_value"; cf.number_format = "#,##0"; cf.fill = fill(MINT_BG)
    cell_name(wb, "FlagshipCover", "Inventory", f"$F${end+2}")
    ws.cell(row=end + 3, column=2, value="SKUs needing a reorder").style = "field_label"
    cn = ws.cell(row=end + 3, column=6, value=f'=COUNTIF(G{start}:G{end},"Reorder now")+COUNTIF(G{start}:G{end},"Out of stock")')
    cn.style = "field_value"; cn.number_format = "#,##0"
    ws.freeze_panes = "A5"


def build_sales(wb):
    ws = wb.create_sheet("Sales"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 16, 18, 18, 2])
    luxe_header(ws, "E", "🛒  SALES",
                "Units and revenue month by month — the growth curve your inventory has to keep up with.")
    table_headers(ws, 4, ["Month", "Units", "Revenue", "Avg Order"], start_col=2)
    start = L0
    for i, (m, units, rev) in enumerate(SALES):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cu = ws.cell(row=r, column=3, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cr = ws.cell(row=r, column=4, value=rev); cr.style = "input"; cr.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); ca.style = "td"; ca.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SALES) - 1
    nrange(wb, "SaleUnits", "Sales", "C", start, end)
    nrange(wb, "SaleRev", "Sales", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    cu = ws.cell(row=tot, column=3, value="=SUM(SaleUnits)"); cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "#,##0"
    cr = ws.cell(row=tot, column=4, value="=SUM(SaleRev)"); cr.style = "td"; cr.font = Font(bold=True, size=12, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "RevenueYTD", "Sales", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_ppc(wb):
    ws = wb.create_sheet("PPC & ACoS"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 16, 14, 2])
    luxe_header(ws, "E", "📣  PPC & ACOS",
                "Every campaign's spend against the sales it produced — ACoS is the number that decides if ads pay.")
    table_headers(ws, 4, ["Campaign", "Ad Spend", "Ad Sales", "ACoS"], start_col=2)
    start = L0
    for i, (camp, spend, sales, acos) in enumerate(PPC):
        r = start + i
        ws.cell(row=r, column=2, value=camp).style = "td_left"
        cs = ws.cell(row=r, column=3, value=spend); cs.style = "input"; cs.number_format = '"$"#,##0'
        csa = ws.cell(row=r, column=4, value=sales); csa.style = "input"; csa.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); ca.style = "td"; ca.number_format = "0.0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PPC) - 1
    nrange(wb, "AdCampSpend", "PPC & ACoS", "C", start, end)
    nrange(wb, "AdCampSales", "PPC & ACoS", "D", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        CellIsRule(operator="greaterThan", formula=["ACoSGoal"], fill=fill(RED_BG)))
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL CAMPAIGNS").style = "th"
    cs = ws.cell(row=tot, column=3, value="=SUM(AdCampSpend)"); cs.style = "td"; cs.font = Font(bold=True, color=DANGER); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    csa = ws.cell(row=tot, column=4, value="=SUM(AdCampSales)"); csa.style = "td"; csa.font = Font(bold=True, color=PRIMARY); csa.fill = fill(SURFACE); csa.number_format = '"$"#,##0'
    ca = ws.cell(row=tot, column=5, value=f"=IFERROR(C{tot}/D{tot},0)"); ca.style = "td"; ca.font = Font(bold=True, size=12, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = "0.0%"
    cell_name(wb, "ACoS", "PPC & ACoS", f"$E${tot}")
    ws.cell(row=tot + 2, column=2, value="TACoS (ad spend ÷ total revenue)").style = "field_label"
    ct = ws.cell(row=tot + 2, column=5, value="=IFERROR(AdSpend/MonthlyRevenue,0)"); ct.style = "field_value"; ct.number_format = "0.0%"; ct.fill = fill(MINT_BG)
    cell_name(wb, "TACoS", "PPC & ACoS", f"$E${tot+2}")
    ws.cell(row=tot + 4, column=2, value="ACoS above your margin means you're paying to lose money on that campaign.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_returns(wb):
    ws, start, end = build_log(
        wb, "Returns", "↩", "RETURNS & REFUNDS",
        "What comes back and why — a return rate creeping up is a listing or a quality problem.",
        ["SKU", "Units Returned", "Reason"],
        RETURNS, [2, 18, 18, 20, 2], text_left={2}, ints={3}, reserved=24, start_col=2,
        validations=[("D", "ReturnReasonList")])
    nrange(wb, "RetUnits", "Returns", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL RETURNED").style = "th"
    cr = ws.cell(row=tot, column=3, value="=SUM(RetUnits)"); cr.style = "td"; cr.font = Font(bold=True, color=DANGER); cr.fill = fill(SURFACE); cr.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    ws.cell(row=tot + 1, column=2, value="Return rate (of units sold)").style = "field_label"
    crr = ws.cell(row=tot + 1, column=3, value="=IFERROR(SUM(RetUnits)/UnitsMonth,0)"); crr.style = "field_value"; crr.number_format = "0.0%"; crr.fill = fill(WARN_BG)
    cell_name(wb, "ReturnRate", "Returns", f"$C${tot+1}")


def build_suppliers(wb):
    ws, start, end = build_log(
        wb, "Suppliers & POs", "🚚", "SUPPLIERS & POS",
        "Who makes what, at what cost, and where each purchase order is right now.",
        ["Supplier", "SKU", "Units", "Unit Cost", "Status"],
        SUPPLIERS, [2, 22, 16, 12, 14, 16, 2], text_left={2, 6}, ints={4}, money2={5}, reserved=24, start_col=2)
    nrange(wb, "POUnits", "Suppliers & POs", "D", start, end)
    nrange(wb, "POCost", "Suppliers & POs", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ON ORDER").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cu = ws.cell(row=tot, column=4, value="=SUM(POUnits)"); cu.style = "td"; cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "#,##0"
    cv = ws.cell(row=tot, column=5, value="=SUMPRODUCT(POUnits,POCost)"); cv.style = "td"; cv.font = Font(bold=True, color=PRIMARY); cv.fill = fill(MINT_BG); cv.number_format = '"$"#,##0'
    ws.cell(row=tot, column=6).style = "td"; ws.cell(row=tot, column=6).fill = fill(SURFACE)
    cell_name(wb, "OnOrderValue", "Suppliers & POs", f"$E${tot}")


def build_reviews(wb):
    ws = wb.create_sheet("Reviews"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 18, 16, 16, 2])
    luxe_header(ws, "D", "⭐  REVIEWS",
                "Review count and rating per SKU — review velocity is what unlocks conversion on Amazon.")
    table_headers(ws, 4, ["SKU", "Reviews", "Avg Rating"], start_col=2)
    start = L0
    for i, (sku, revs, rating) in enumerate(REVIEWS):
        r = start + i
        ws.cell(row=r, column=2, value=sku).style = "td_left"
        cr = ws.cell(row=r, column=3, value=revs); cr.style = "input"; cr.number_format = "#,##0"
        cg = ws.cell(row=r, column=4, value=rating); cg.style = "input"; cg.number_format = "0.0"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(REVIEWS) - 1
    nrange(wb, "RevCount", "Reviews", "C", start, end)
    nrange(wb, "RevRating", "Reviews", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVERAGE").style = "th"
    cr = ws.cell(row=tot, column=3, value="=IFERROR(AVERAGE(RevCount),0)"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = "0.0"
    cg = ws.cell(row=tot, column=4, value="=IFERROR(AVERAGE(RevRating),0)"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(MINT_BG); cg.number_format = "0.0"
    cell_name(wb, "AvgReviews", "Reviews", f"$C${tot}")
    cell_name(wb, "AvgRating", "Reviews", f"$D${tot}")
    ws.cell(row=tot + 2, column=2, value="Use Amazon's Request a Review button on every order — it's free and compliant.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_expenses(wb):
    ws = wb.create_sheet("Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "💰  EXPENSES",
                "The costs that sit outside the per-unit maths — ads, software, photography, accounting.")
    table_headers(ws, 4, ["Expense", "Monthly"], start_col=2)
    start = L0
    for i, (item, amt) in enumerate(EXPENSES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EXPENSES) - 1
    nrange(wb, "ExpAmt", "Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Expenses", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "Net profit month by month — the only line that matters once the fees are out.")
    ws.cell(row=5, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 6, ["Month", "Net Profit"], start_col=2)
    ts = 7
    for i, (m, np_) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cn = ws.cell(row=r, column=3, value=np_); cn.style = "input"; cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "ProfitTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Net Profit by Month", ts, te, 3, 2), "E5")


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  📦  AMAZON FBA & ONLINE SELLER PROFIT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Net per unit, margin, ROI, ACoS & a Seller Score — your whole brand, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("SALE PRICE", "=SalePrice", "money2"),
        ("AMAZON FEES", "=TotalFees", "money2"),
        ("LANDED COST", "=LandedCost", "money2"),
        ("NET / UNIT", "=NetPerUnit", "money2"),
        ("NET MARGIN", "=NetMargin", "pct1"),
        ("ROI", "=ROI", "pct1"),
    ]
    row2 = [
        ("UNITS / MONTH", "=UnitsMonth", "num"),
        ("MONTHLY REVENUE", "=MonthlyRevenue", "money"),
        ("MONTHLY PROFIT", "=MonthlyProfit", "money"),
        ("ACOS", "=ACoS", "pct1"),
        ("DAYS OF COVER", "=FlagshipCover", "num"),
        ("SELLER SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "SELLER HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "NET PROFIT BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Margin healthy", "=IFERROR(MIN(NetMargin/MarginGoal,1),0)"),
        ("ROI healthy", "=IFERROR(MIN(ROI/ROIGoal,1),0)"),
        ("ACoS in check", "=IFERROR(MIN(ACoSGoal/ACoS,1),0)"),
        ("Inventory covered", "=IFERROR(MIN(FlagshipCover/CoverGoal,1),0)"),
        ("Catalog built out", "=IFERROR(MIN(SKUCount/SKUGoal,1),0)"),
        ("Reviews per SKU", "=IFERROR(MIN(AvgReviews/ReviewGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Net Profit by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=7, max_row=6 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=7, max_row=6 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Amazon FBA & Online Seller Profit Command Center™ — revenue is vanity, net per unit is truth.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_profit(wb); build_catalog(wb)
    build_fees(wb); build_inventory(wb); build_sales(wb); build_ppc(wb)
    build_returns(wb); build_suppliers(wb); build_reviews(wb); build_expenses(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Profit Calculator", "Product Catalog", "Fee Breakdown", "Inventory",
             "Sales", "PPC & ACoS", "Returns", "Suppliers & POs", "Reviews", "Expenses",
             "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "FBA_Seller_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
