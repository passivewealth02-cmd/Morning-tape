"""Build Real Estate Agent Command Center™ — The Realtor Operating System.

14 tabs · a premium real-estate-agent operating system in Google Sheets & Excel.
Dashboard, a commission calculator (sale price → GCI → split → net per deal), a deal
pipeline, closings, buyers & sellers, listings, lead sources, a database/sphere
tracker, a GCI goal tracker, business expenses, mileage and a monthly summary — one
dashboard. Know your GCI, your net, and your real business.

Run: python3 build_xlsx.py   ->  ../Realtor_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
SIDE = ["Listing", "Buyer", "Both", "Referral"]
STAGE = ["Lead", "Showing", "Offer", "Under contract", "Closed", "Lost"]
SOURCE = ["Referral", "Sphere", "Zillow", "Open house", "Social", "Sign call", "Past client"]

MARGIN_GOAL = 0.70
COMM_GOAL = 8000
PACE_GOAL = 6
PIPE_GOAL = 6
PROFIT_GOAL = 40000
TOUCH_GOAL = 30

# Commission calculator — flagship deal
SALE_PRICE = 450000
COMM_RATE = 0.03
AGENT_SPLIT = 0.70
# Per-deal transaction costs: (item, amount)
DEAL_COSTS = [("Brokerage transaction fee", 395), ("Marketing & staging", 600),
              ("Transaction coordinator", 400), ("E&O / misc", 55)]

# Closings YTD: (address, sale price, side, close date)
CLOSINGS = [
    ("102 Elm St", 400000, "Listing", "Jan 18"), ("55 Oak Ave", 450000, "Buyer", "Feb 09"),
    ("8 Birch Ln", 500000, "Listing", "Mar 02"), ("240 Cedar Ct", 350000, "Buyer", "Mar 21"),
    ("77 Maple Dr", 550000, "Listing", "Apr 12"), ("16 Pine Rd", 425000, "Buyer", "May 04"),
    ("301 Ash Blvd", 475000, "Both", "May 27"), ("9 Willow Way", 450000, "Listing", "Jun 15"),
]

# Pipeline: (address / client, stage, price, est agent GCI)
PIPELINE = [
    ("Ford — buyer", "Under contract", 470000, 9870), ("Nguyen — listing", "Offer", 520000, 10920),
    ("Ramos — buyer", "Showing", 380000, 7980), ("Bexar Condo", "Under contract", 300000, 6300),
    ("Whit — listing", "Lead", 610000, 12810), ("Ivy Duplex", "Showing", 540000, 11340),
    ("Osei — buyer", "Offer", 415000, 8715), ("Vance — listing", "Lead", 495000, 10395),
]

# Buyers & sellers: (client, type, stage, source)
CLIENTS = [
    ("Ford Family", "Buyer", "Under contract", "Referral"), ("Nguyen", "Seller", "Offer", "Sphere"),
    ("Ramos", "Buyer", "Showing", "Zillow"), ("Whitfield", "Seller", "Lead", "Past client"),
    ("Osei", "Buyer", "Offer", "Open house"),
]

# Listings: (address, list price, DOM, status)
LISTINGS = [
    ("77 Maple Dr", 550000, 12, "Active"), ("Nguyen — 520 Vine", 520000, 8, "Pending"),
    ("Whit — 610 Hill", 610000, 21, "Active"), ("Vance — 495 Park", 495000, 4, "Coming soon"),
]

# Lead sources — ROI: (source, leads, closings, spend)
LEADSOURCES = [
    ("Referral", 14, 4, 0), ("Sphere / past client", 10, 2, 300), ("Zillow", 22, 1, 1200),
    ("Open house", 8, 1, 150), ("Social media", 30, 0, 400),
]

# Database / sphere: (group, contacts, touches this month)
DATABASE = [
    ("A — raving fans", 40, 8), ("B — past clients", 120, 3), ("C — sphere", 260, 1), ("D — new leads", 55, 0),
]
DB_TOUCHES = 12

# Business expenses (YTD): (item, amount)
EXPENSES = [
    ("Marketing", 6000), ("CRM & tech", 1800), ("Dues & MLS", 2400), ("Car & fuel", 4200),
    ("TC & E&O", 3000), ("Staging & photos", 5200), ("Coaching", 3000),
]

# Mileage & auto: (date, purpose, miles, cost)
MILEAGE = [
    ("Jun 3", "Showings — 4 homes", 42, 28.14), ("Jun 9", "Listing appt", 18, 12.06),
    ("Jun 15", "Closing — Willow", 26, 17.42), ("Jun 22", "Open house", 14, 9.38),
]

# GCI goal
GCI_GOAL = 100000

# Monthly summary: (month, agent GCI, closings)
MONTHS = [("Jan", 8400, 1), ("Feb", 10500, 1), ("Mar", 14700, 2),
          ("Apr", 11550, 1), ("May", 16800, 2), ("Jun", 13650, 1)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General", "vol": '"$"#,##0,,"M"'}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Agent name", "Taylor Brooks", None, "Business"),
        ("Brokerage", "Summit Realty", None, "Owner"),
        ("Net-margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("GCI-per-deal goal ($)", COMM_GOAL, '"$"#,##0', "CommGoal"),
        ("Closings pace goal (YTD)", PACE_GOAL, "#,##0", "PaceGoal"),
        ("Pipeline deals goal", PIPE_GOAL, "#,##0", "PipeGoal"),
        ("Net income goal (YTD)", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Database touches goal", TOUCH_GOAL, "#,##0", "TouchGoal"),
        ("Annual GCI goal", GCI_GOAL, '"$"#,##0', "GCIGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Side", SIDE, "SideList"), ("F", "Stage", STAGE, "StageList"),
             ("G", "Source", SOURCE, "SourceList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🏡  REAL ESTATE AGENT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know your GCI, your net, and your real business.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE REAL-ESTATE BUSINESS, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("An agent's business lives or dies on two numbers: the commission each deal really nets after the "
                      "split and costs, and whether the pipeline will hit the goal. This makes both visible: a commission "
                      "calculator that turns a sale price into GCI, agent commission and your true net per deal, and a "
                      "pipeline that shows what's coming. Track closings, buyers & sellers, listings, lead-source ROI, "
                      "your database, expenses and mileage, and a GCI goal — all in ONE premium Google Sheets & Excel "
                      "system built for real-estate agents.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your split, GCI goal & database touches goal.",
             "2.  Use the Commission Calculator to net any deal before you take it.",
             "3.  Log your Closings; GCI, volume and net income roll up.",
             "4.  Fill the Pipeline and track buyers, sellers and listings.",
             "5.  Track lead-source ROI, your database, expenses and mileage.",
             "6.  Check the Dashboard: GCI, net & an Agent Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional agent (Taylor Brooks) is included so you can see how it all connects — just "
               "type over it with your own deals and numbers. GCI and net income are the two numbers that decide whether "
               "a real-estate business is really working, and they roll into a live Agent Score. Twelve matching "
               "printable pages (commission worksheet, pipeline, closing checklist, database plan & more) are included. "
               "This is a business tool, not financial, legal or tax advice — confirm your split, fees and figures with "
               "your broker and your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Net every deal, fill the pipeline, and touch the database — that's how agents win.", "section_gold")


# ===========================================================================
def build_commission(wb):
    ws = wb.create_sheet("Commission Calculator"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 16, 2])
    luxe_header(ws, "C", "🧮  COMMISSION CALCULATOR",
                "Sale price → GCI → your split → net per deal — know what a deal really pays before you take it.")
    ws.cell(row=5, column=2, value="Sale price").style = "field_label"
    cp = ws.cell(row=5, column=3, value=SALE_PRICE); cp.style = "input"; cp.number_format = '"$"#,##0'
    cell_name(wb, "SalePrice", "Commission Calculator", "$C$5")
    ws.cell(row=6, column=2, value="Commission rate (your side)").style = "field_label"
    cr = ws.cell(row=6, column=3, value=COMM_RATE); cr.style = "input"; cr.number_format = "0.00%"
    cell_name(wb, "CommRate", "Commission Calculator", "$C$6")
    ws.cell(row=7, column=2, value="= GROSS COMMISSION (GCI)").style = "th"
    cg = ws.cell(row=7, column=3, value="=SalePrice*CommRate"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(MINT_BG); cg.number_format = '"$"#,##0'
    cell_name(wb, "GCIDeal", "Commission Calculator", "$C$7")
    ws.cell(row=8, column=2, value="Agent split % (you keep)").style = "field_label"
    cs = ws.cell(row=8, column=3, value=AGENT_SPLIT); cs.style = "input"; cs.number_format = "0%"
    cell_name(wb, "AgentSplit", "Commission Calculator", "$C$8")
    ws.cell(row=9, column=2, value="= AGENT COMMISSION").style = "th"
    ca = ws.cell(row=9, column=3, value="=GCIDeal*AgentSplit"); ca.style = "td"; ca.font = Font(bold=True, size=12, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = '"$"#,##0'
    cell_name(wb, "AgentComm", "Commission Calculator", "$C$9")
    ws.cell(row=11, column=2, value="TRANSACTION COSTS").style = "section_gold"
    table_headers(ws, 12, ["Cost", "Amount"], start_col=2)
    start = 13
    for i, (item, amt) in enumerate(DEAL_COSTS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=amt); cc.style = "input"; cc.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DEAL_COSTS) - 1
    nrange(wb, "CostAmt", "Commission Calculator", "C", start, end)
    dtot = end + 1
    ws.cell(row=dtot, column=2, value="= DEAL COSTS").style = "th"
    cdc = ws.cell(row=dtot, column=3, value="=SUM(CostAmt)"); cdc.style = "td"; cdc.font = Font(bold=True, color=DANGER); cdc.fill = fill(SURFACE); cdc.number_format = '"$"#,##0'
    cell_name(wb, "DealCosts", "Commission Calculator", f"$C${dtot}")
    nr = dtot + 1
    ws.cell(row=nr, column=2, value="= NET PER DEAL").style = "th"
    cn = ws.cell(row=nr, column=3, value="=AgentComm-DealCosts"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetDeal", "Commission Calculator", f"$C${nr}")
    ws.cell(row=nr + 1, column=2, value="Net margin (net ÷ agent commission)").style = "field_label"
    cm = ws.cell(row=nr + 1, column=3, value="=IFERROR(NetDeal/AgentComm,0)"); cm.style = "field_value"; cm.number_format = "0%"; cm.fill = fill(MINT_BG)
    cell_name(wb, "NetMargin", "Commission Calculator", f"$C${nr+1}")


def build_closings(wb):
    ws = wb.create_sheet("Closings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 16, 14, 14, 16, 2])
    luxe_header(ws, "F", "✅  CLOSINGS",
                "Every closed deal this year — sale price and agent commission roll up to your GCI and volume.")
    table_headers(ws, 4, ["Address", "Sale Price", "Side", "Closed", "Agent GCI"], start_col=2)
    start = L0
    for i, (addr, price, side, dt) in enumerate(CLOSINGS):
        r = start + i
        ws.cell(row=r, column=2, value=addr).style = "td_left"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=side).style = "td"
        ws.cell(row=r, column=5, value=dt).style = "td"
        cg = ws.cell(row=r, column=6, value=f"=C{r}*CommRate*AgentSplit"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"D{r}", "SideList")
    end = start + len(CLOSINGS) - 1
    nrange(wb, "ClosePrice", "Closings", "C", start, end)
    nrange(wb, "CloseGCI", "Closings", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="YTD TOTALS").style = "th"
    cv = ws.cell(row=tot, column=3, value="=SUM(ClosePrice)"); cv.style = "td"; cv.font = Font(bold=True, color=PRIMARY); cv.fill = fill(SURFACE); cv.number_format = '"$"#,##0'
    for c in (4, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cg = ws.cell(row=tot, column=6, value="=SUM(CloseGCI)"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(MINT_BG); cg.number_format = '"$"#,##0'
    cell_name(wb, "VolumeYTD", "Closings", f"$C${tot}")
    cell_name(wb, "GCI_YTD", "Closings", f"$F${tot}")
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Closings YTD").style = "field_label"
    cc = ws.cell(row=sr, column=6, value="=COUNTA(ClosePrice)"); cc.style = "field_value"; cc.number_format = "#,##0"; cc.fill = fill(MINT_BG)
    cell_name(wb, "ClosingsYTD", "Closings", f"$F${sr}")
    ws.cell(row=sr + 1, column=2, value="Average sale price").style = "field_label"
    cap = ws.cell(row=sr + 1, column=6, value="=IFERROR(VolumeYTD/ClosingsYTD,0)"); cap.style = "field_value"; cap.number_format = '"$"#,##0'; cap.fill = fill(MINT_BG)
    cell_name(wb, "AvgPrice", "Closings", f"$F${sr+1}")
    ws.freeze_panes = "A5"


def build_pipeline(wb):
    ws = wb.create_sheet("Pipeline"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 22, 18, 16, 14, 2])
    luxe_header(ws, "E", "🔭  PIPELINE",
                "What's coming — stage, price and estimated agent commission, so you can see the goal from here.")
    table_headers(ws, 4, ["Deal", "Stage", "Price", "Est. GCI"], start_col=2)
    start = L0
    for i, (deal, stage, price, gci) in enumerate(PIPELINE):
        r = start + i
        ws.cell(row=r, column=2, value=deal).style = "td_left"
        ws.cell(row=r, column=3, value=stage).style = "td"
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cg = ws.cell(row=r, column=5, value=gci); cg.style = "input"; cg.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"C{r}", "StageList")
    end = start + len(PIPELINE) - 1
    nrange(wb, "PipeName", "Pipeline", "B", start, end)
    nrange(wb, "PipeGCI", "Pipeline", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="PIPELINE").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cg = ws.cell(row=tot, column=5, value="=SUM(PipeGCI)"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(MINT_BG); cg.number_format = '"$"#,##0'
    cell_name(wb, "PipelineGCI", "Pipeline", f"$E${tot}")
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Deals in pipeline").style = "field_label"
    cd = ws.cell(row=sr, column=5, value="=COUNTA(PipeName)"); cd.style = "field_value"; cd.number_format = "#,##0"; cd.fill = fill(MINT_BG)
    cell_name(wb, "PipelineDeals", "Pipeline", f"$E${sr}")
    ws.freeze_panes = "A5"


def build_clients(wb):
    ws, start, end = build_log(
        wb, "Buyers & Sellers", "👥", "BUYERS & SELLERS",
        "Your active clients — type, stage and where they came from.",
        ["Client", "Type", "Stage", "Source"],
        CLIENTS, [2, 22, 14, 18, 16, 2], text_left={2}, reserved=24, start_col=2,
        validations=[("D", "StageList"), ("E", "SourceList")])


def build_listings(wb):
    ws = wb.create_sheet("Listings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 12, 16, 2])
    luxe_header(ws, "E", "🏠  LISTINGS",
                "Your active listings — price, days on market and status, so nothing goes stale.")
    table_headers(ws, 4, ["Address", "List Price", "DOM", "Status"], start_col=2)
    start = L0
    for i, (addr, price, dom, status) in enumerate(LISTINGS):
        r = start + i
        ws.cell(row=r, column=2, value=addr).style = "td_left"
        cp = ws.cell(row=r, column=3, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=4, value=dom); cd.style = "input"; cd.number_format = "#,##0"
        ws.cell(row=r, column=5, value=status).style = "td"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LISTINGS) - 1
    nrange(wb, "ListDOM", "Listings", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVG DAYS ON MARKET").style = "th"
    for c in (3,):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cd = ws.cell(row=tot, column=4, value="=IFERROR(AVERAGE(ListDOM),0)"); cd.style = "td"; cd.font = Font(bold=True, color=PRIMARY); cd.fill = fill(MINT_BG); cd.number_format = "0"
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)
    cell_name(wb, "AvgDOM", "Listings", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_leadsources(wb):
    ws = wb.create_sheet("Lead Sources"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 12, 14, 14, 16, 2])
    luxe_header(ws, "F", "📣  LEAD SOURCES",
                "Where deals really come from — leads, closings and spend, so you invest where it pays.")
    table_headers(ws, 4, ["Source", "Leads", "Closings", "Spend", "Cost/Deal"], start_col=2)
    start = L0
    for i, (src, leads, closings, spend) in enumerate(LEADSOURCES):
        r = start + i
        ws.cell(row=r, column=2, value=src).style = "td_left"
        cl = ws.cell(row=r, column=3, value=leads); cl.style = "input"; cl.number_format = "#,##0"
        cc = ws.cell(row=r, column=4, value=closings); cc.style = "input"; cc.number_format = "#,##0"
        cs = ws.cell(row=r, column=5, value=spend); cs.style = "input"; cs.number_format = '"$"#,##0'
        ccd = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/D{r},0)"); ccd.style = "td"; ccd.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    ws.freeze_panes = "A5"


def build_database(wb):
    ws = wb.create_sheet("Database"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 18, 2])
    luxe_header(ws, "D", "📇  DATABASE",
                "Your sphere by group — contacts and touches this month. The database is the business.")
    table_headers(ws, 4, ["Group", "Contacts", "Touches (mo)"], start_col=2)
    start = L0
    for i, (grp, contacts, touches) in enumerate(DATABASE):
        r = start + i
        ws.cell(row=r, column=2, value=grp).style = "td_left"
        cc = ws.cell(row=r, column=3, value=contacts); cc.style = "input"; cc.number_format = "#,##0"
        ct = ws.cell(row=r, column=4, value=touches); ct.style = "input"; ct.number_format = "#,##0"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DATABASE) - 1
    nrange(wb, "DBContacts", "Database", "C", start, end)
    nrange(wb, "DBTouchCol", "Database", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTALS").style = "th"
    cc = ws.cell(row=tot, column=3, value="=SUM(DBContacts)"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = "#,##0"
    ct = ws.cell(row=tot, column=4, value="=SUM(DBTouchCol)"); ct.style = "td"; ct.font = Font(bold=True, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = "#,##0"
    cell_name(wb, "DBTouches", "Database", f"$D${tot}")
    ws.freeze_panes = "A5"


def build_expenses(wb):
    ws = wb.create_sheet("Business Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  BUSINESS EXPENSES",
                "Your business costs (YTD) — netted against your GCI to show what you really made.")
    table_headers(ws, 4, ["Expense", "Amount"], start_col=2)
    start = L0
    for i, (item, amt) in enumerate(EXPENSES):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EXPENSES) - 1
    nrange(wb, "ExpAmt", "Business Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES (YTD)").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "BizExp", "Business Expenses", f"$C${tot}")
    nr = tot + 2
    ws.cell(row=nr, column=2, value="= NET INCOME (GCI − expenses)").style = "th"
    cn = ws.cell(row=nr, column=3, value="=GCI_YTD-BizExp"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetIncome", "Business Expenses", f"$C${nr}")
    ws.freeze_panes = "A5"


def build_mileage(wb):
    ws, start, end = build_log(
        wb, "Mileage & Auto", "🚗", "MILEAGE & AUTO",
        "Every business drive — showings, listings and closings, deductible at tax time.",
        ["Date", "Purpose", "Miles", "Cost"],
        MILEAGE, [2, 14, 26, 12, 14, 2], text_left={2, 3}, ints={4}, money2={5}, reserved=24, start_col=2)


def build_goals(wb):
    ws = wb.create_sheet("Goals & GCI"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "🎯  GOALS & GCI",
                "Your GCI goal vs where you are — and the pace it takes to hit it.")
    rows = [
        ("Annual GCI goal", "=GCIGoal", '"$"#,##0', WARN_BG),
        ("GCI earned (YTD)", "=GCI_YTD", '"$"#,##0', MINT_BG),
        ("= REMAINING TO GOAL", "=MAX(GCIGoal-GCI_YTD,0)", '"$"#,##0', SURFACE),
        ("Goal progress", "=IFERROR(GCI_YTD/GCIGoal,0)", "0%", MINT_BG),
        ("Pipeline GCI (coming)", "=PipelineGCI", '"$"#,##0', MINT_BG),
        ("Avg GCI per closing", "=IFERROR(GCI_YTD/ClosingsYTD,0)", '"$"#,##0', MINT_BG),
        ("Closings still needed", "=IFERROR(ROUNDUP(MAX(GCIGoal-GCI_YTD,0)/(GCI_YTD/ClosingsYTD),0),0)", "#,##0", WARN_BG),
    ]
    for i, (lab, fml, fmt, bg) in enumerate(rows):
        r = 5 + i
        style = "th" if lab.startswith("=") else "field_label"
        ws.cell(row=r, column=2, value=lab).style = style
        c = ws.cell(row=r, column=3, value=fml); c.style = "td" if lab.startswith("=") else "field_value"
        c.number_format = fmt; c.fill = fill(bg)
        if lab.startswith("="):
            c.font = Font(bold=True, size=12, color=PRIMARY)
    ws.cell(row=13, column=2, value="Fill the pipeline, and the goal takes care of itself.").style = "section"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 14, 2])
    luxe_header(ws, "D", "📈  MONTHLY SUMMARY",
                "GCI & closings by month — watch the year build toward your goal.")
    table_headers(ws, 4, ["Month", "Agent GCI", "Closings"], start_col=2)
    start = L0
    for i, (m, gci, cl) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        cg = ws.cell(row=r, column=3, value=gci); cg.style = "input"; cg.number_format = '"$"#,##0'
        cc = ws.cell(row=r, column=4, value=cl); cc.style = "input"; cc.number_format = "#,##0"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthGCI", "Monthly Summary", "C", start, end)
    ws.add_chart(_barchart(ws, "GCI by Month", start, end, 3, 2), "F4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🏡  REAL ESTATE AGENT COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  GCI, net per deal, pipeline & an Agent Score — your whole real-estate business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("AVG SALE PRICE", "=AvgPrice", "money"),
        ("COMMISSION", "=CommRate", "pct"),
        ("GCI / DEAL", "=GCIDeal", "money"),
        ("NET / DEAL", "=NetDeal", "money"),
        ("CLOSINGS YTD", "=ClosingsYTD", "num"),
        ("VOLUME YTD", "=VolumeYTD", "vol"),
    ]
    row2 = [
        ("GCI YTD", "=GCI_YTD", "money"),
        ("NET INCOME", "=NetIncome", "money"),
        ("PIPELINE", "=PipelineDeals", "num"),
        ("PIPELINE GCI", "=PipelineGCI", "money"),
        ("GOAL PROGRESS", "=IFERROR(GCI_YTD/GCIGoal,0)", "pct"),
        ("AGENT SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "AGENT HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "GCI BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Net margin healthy", "=IFERROR(MIN(NetMargin/MarginGoal,1),0)"),
        ("GCI per deal on target", "=IFERROR(MIN(AgentComm/CommGoal,1),0)"),
        ("Closings on pace", "=IFERROR(MIN(ClosingsYTD/PaceGoal,1),0)"),
        ("Pipeline full", "=IFERROR(MIN(PipelineDeals/PipeGoal,1),0)"),
        ("Profitable", "=IFERROR(MIN(NetIncome/ProfitGoal,1),0)"),
        ("Database nurtured", "=IFERROR(MIN(DBTouches/TouchGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "GCI by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Real Estate Agent Command Center™ — know your GCI, your net, and your real business.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_commission(wb); build_closings(wb)
    build_pipeline(wb); build_clients(wb); build_listings(wb); build_leadsources(wb)
    build_database(wb); build_goals(wb); build_expenses(wb); build_mileage(wb); build_summary(wb)
    build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Commission Calculator", "Pipeline", "Closings", "Buyers & Sellers",
             "Listings", "Lead Sources", "Database", "Goals & GCI", "Business Expenses", "Mileage & Auto",
             "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Realtor_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
