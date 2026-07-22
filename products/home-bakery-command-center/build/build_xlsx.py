"""Build Home Bakery & Cottage Food Command Center™ — The Home-Baker's System.

14 tabs · a premium home-bakery / cottage-food operating system in Google Sheets &
Excel. Dashboard, a "price it right" engine that pays you for your time, recipe
costing, a product list, custom orders, ingredient costs, labeling & allergens,
markets, income & expenses, waste, customers and a monthly summary — one dashboard.
Stop underpricing: cost every recipe and pay yourself for your time.

Run: python3 build_xlsx.py   ->  ../Home_Bakery_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["each", "dozen", "loaf", "cake", "batch", "cup", "lb", "oz"]
STATUS = ["Confirmed", "Deposit", "Inquiry", "Delivered"]
ALLERGENS = ["Wheat/Gluten", "Eggs", "Dairy", "Nuts", "Soy", "None"]

TARGET_FC = 0.25
MARGIN_GOAL = 0.70
WAGE_GOAL = 20
PROFIT_GOAL = 1500
WASTE_LIMIT = 0.05

# Price-It-Right engine — flagship Custom Cookies (dozen)
PIR = {"ingredients": 6.00, "packaging": 1.50, "minutes": 45, "rate": 20, "overhead": 1.00, "price": 33}

# Recipe costing — Chocolate Chip Cookies batch (2 dozen): (ingredient, cost)
RECIPE = [
    ("Flour (3 cups)", 0.90), ("Butter (1 cup)", 2.20), ("Sugar (1.5 cups)", 0.85),
    ("Eggs (2)", 0.60), ("Chocolate chips (2 cups)", 3.80), ("Vanilla + baking soda", 0.65),
]
RECIPE_YIELD = 24

# Product list: (item, ingredient cost, price, monthly units)
PRODUCTS = [
    ("Custom Cookies (dz)", 6.00, 33, 20),
    ("Cupcakes (dz)", 5.50, 30, 15),
    ("Layer Cake 8\"", 9.00, 55, 8),
    ("Cinnamon Rolls (6)", 3.20, 18, 12),
    ("Banana Bread", 2.40, 12, 25),
    ("Brownies (dz)", 4.10, 22, 14),
    ("Macarons (dz)", 5.80, 30, 10),
    ("Pie", 4.50, 26, 9),
]

# Custom orders: (customer, item, days out, price, deposit, status)
ORDERS = [
    ("Sarah M.", "Birthday cake", 5, 55, 25, "Confirmed"),
    ("Mike R.", "3 dz cookies", 3, 99, 0, "Confirmed"),
    ("Ana L.", "Wedding cupcakes", 12, 180, 90, "Deposit"),
    ("Tom B.", "Pie x2", 2, 52, 0, "Confirmed"),
    ("Lily K.", "Macarons", 8, 30, 0, "Inquiry"),
]

# Ingredient cost library: (ingredient, pack price, pack size, unit)
INGREDIENTS = [
    ("Flour", 4.50, "5 lb bag", "lb"), ("Butter", 4.80, "1 lb", "lb"), ("Sugar", 3.20, "4 lb bag", "lb"),
    ("Eggs", 3.60, "dozen", "dozen"), ("Chocolate chips", 3.80, "12 oz", "oz"), ("Vanilla", 8.00, "4 oz", "oz"),
    ("Cream cheese", 3.40, "8 oz", "oz"), ("Powdered sugar", 2.90, "2 lb", "lb"),
]

# Labeling & allergens: (product, allergens, net weight)
LABELS = [
    ("Custom Cookies", "Wheat/Gluten, Eggs, Dairy", "12 oz"),
    ("Cupcakes", "Wheat/Gluten, Eggs, Dairy", "16 oz"),
    ("Layer Cake", "Wheat/Gluten, Eggs, Dairy", "32 oz"),
    ("Banana Bread", "Wheat/Gluten, Eggs, Nuts", "18 oz"),
    ("Macarons", "Eggs, Nuts", "8 oz"),
]

# Markets & events: (market, days out, fee, sales)
MARKETS = [
    ("Downtown Farmers Market", 4, 35, 320),
    ("Craft Fair", 11, 60, 480),
    ("Holiday Market", 25, 75, 640),
    ("Neighborhood Pop-up", 6, 0, 210),
]

# Income & expenses: (item, amount, type)  type: I=income, E=expense
LEDGER = [
    ("Ingredients", 720, "E"), ("Packaging & labels", 180, "E"), ("Market fees", 120, "E"),
    ("Supplies & equipment", 90, "E"), ("Marketing", 60, "E"), ("Mileage / other", 80, "E"),
]

# Waste log: (item, reason, cost) — total ~$90
WASTE = [
    ("Failed batch", "Over-baked", 32.00),
    ("Leftover / unsold", "Market end", 30.00),
    ("Spoiled ingredients", "Expired", 18.00),
    ("Broken / dropped", "Handling", 10.00),
]

# Customers: (name, favorite, contact, repeat)
CUSTOMERS = [
    ("Sarah M.", "Custom cakes", "sarah@mail.com", "Repeat"),
    ("Mike R.", "Cookies", "mike@mail.com", "Repeat"),
    ("Ana L.", "Cupcakes", "ana@mail.com", "New"),
    ("Tom B.", "Pies", "tom@mail.com", "Repeat"),
    ("Lily K.", "Macarons", "lily@mail.com", "New"),
]

# Monthly summary: (month, income, expenses)
MONTHS = [("Jul", 2450, 1120), ("Aug", 2680, 1180), ("Sep", 2908, 1250),
          ("Oct", 3200, 1340), ("Nov", 3850, 1520), ("Dec", 4600, 1780)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Bakery name", "Sugar & Thyme", None, "Bakery"),
        ("Owner", "Mel", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Your hourly-wage goal", WAGE_GOAL, '"$"#,##0', "WageGoal"),
        ("Monthly profit goal", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Waste limit %", WASTE_LIMIT, "0%", "WasteLimit"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Status", STATUS, "StatusList"),
             ("G", "Allergen", ALLERGENS, "AllergenList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🧁  HOME BAKERY & COTTAGE FOOD COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every recipe and pay yourself for your time.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE HOME BAKERY, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Home bakers undercharge more than anyone — because they forget to pay themselves for their time. "
                      "This fixes that: a 'price it right' engine that adds your ingredients, packaging, YOUR labor and "
                      "overhead, then shows your profit and your true hourly wage. Cost your recipes, price your products, "
                      "manage custom orders, track ingredient costs, build cottage-food labels with allergens, log markets, "
                      "and keep income & expenses for taxes — all in ONE premium Google Sheets & Excel system built for "
                      "home & cottage bakers.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your hourly-wage goal & margin goal.",
             "2.  Use Price It Right — it pays you for your time on every item.",
             "3.  Cost recipes & set prices in the Product List.",
             "4.  Book Custom Orders; build labels with allergens.",
             "5.  Log Markets, Income & Expenses and Waste.",
             "6.  Check the Dashboard: income, profit & a Bakery Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional home bakery (Sugar & Thyme) is included so you can see how it all connects — "
               "just type over it with your own recipes and prices. Your effective hourly wage and monthly profit are the "
               "two numbers that decide whether your bakery is a business or an expensive hobby, and they roll into a live "
               "Bakery Score. Twelve matching printable pages (order form, cottage-food label, recipe cost card, market "
               "checklist & more) are included. This is a business tool, not financial, legal or cottage-food-law advice — "
               "check your state's cottage food rules and confirm figures with your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "If you're not paying yourself, it's a hobby — this makes it a business.", "section_gold")


# ===========================================================================
def build_priceit(wb):
    ws = wb.create_sheet("Price It Right"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 16, 2])
    luxe_header(ws, "C", "💷  PRICE IT RIGHT",
                "Ingredients + packaging + YOUR time + overhead = true cost. See your profit and your real hourly wage.")
    table_headers(ws, 5, ["Line", "Amount"], start_col=2)
    ws.cell(row=6, column=2, value="Ingredient cost").style = "td_left"
    ci = ws.cell(row=6, column=3, value=PIR["ingredients"]); ci.style = "input"; ci.number_format = '"$"#,##0.00'
    ws.cell(row=7, column=2, value="Packaging").style = "td_left"
    cpk = ws.cell(row=7, column=3, value=PIR["packaging"]); cpk.style = "input"; cpk.number_format = '"$"#,##0.00'
    ws.cell(row=8, column=2, value="Time to make (minutes)").style = "td_left"
    ct = ws.cell(row=8, column=3, value=PIR["minutes"]); ct.style = "input"; ct.number_format = "#,##0"
    ws.cell(row=9, column=2, value="Your hourly rate").style = "td_left"
    cr = ws.cell(row=9, column=3, value=PIR["rate"]); cr.style = "input"; cr.number_format = '"$"#,##0'
    ws.cell(row=10, column=2, value="Overhead / item").style = "td_left"
    co = ws.cell(row=10, column=3, value=PIR["overhead"]); co.style = "input"; co.number_format = '"$"#,##0.00'
    for rr in (7, 9):
        for cc in (2, 3):
            ws.cell(row=rr, column=cc).fill = fill(MUTED_ROW)
    ws.cell(row=11, column=2, value="= Your labor (time × rate)").style = "td_left"
    cl = ws.cell(row=11, column=3, value="=C8/60*C9"); cl.style = "td"; cl.number_format = '"$"#,##0.00'
    ws.cell(row=12, column=2, value="= TRUE COST (with your time)").style = "th"
    ctc = ws.cell(row=12, column=3, value="=C6+C7+C11+C10"); ctc.style = "td"; ctc.font = Font(bold=True, color=PRIMARY); ctc.fill = fill(SURFACE); ctc.number_format = '"$"#,##0.00'
    cell_name(wb, "TrueCost", "Price It Right", "$C$12")
    ws.cell(row=14, column=2, value="Your price").style = "td_left"
    cp = ws.cell(row=14, column=3, value=PIR["price"]); cp.style = "input"; cp.number_format = '"$"#,##0'
    ws.cell(row=15, column=2, value="Profit above your wage").style = "field_label"
    cpr = ws.cell(row=15, column=3, value="=C14-C12"); cpr.style = "field_value"; cpr.number_format = '"$"#,##0.00'; cpr.fill = fill(MINT_BG)
    ws.cell(row=16, column=2, value="= YOUR EFFECTIVE HOURLY WAGE").style = "th"
    ce = ws.cell(row=16, column=3, value="=IFERROR((C14-C6-C7-C10)/(C8/60),0)"); ce.style = "td"; ce.font = Font(bold=True, size=13, color=PRIMARY); ce.fill = fill(MINT_BG); ce.number_format = '"$"#,##0.00'
    cell_name(wb, "EffHourly", "Price It Right", "$C$16")
    ws.cell(row=18, column=2, value="If your effective wage is below your goal, raise the price — your time is worth it.").style = "section"


def build_recipe(wb):
    ws = wb.create_sheet("Recipe Costing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "📖  RECIPE COSTING",
                "Cost a batch, divide by yield — the true ingredient cost of each cookie, loaf or slice.")
    ws.cell(row=5, column=2, value="RECIPE").style = "section_gold"
    ws.cell(row=5, column=3, value="Choc-Chip Cookies").font = Font(bold=True, color=PRIMARY)
    table_headers(ws, 6, ["Ingredient", "Cost"], start_col=2)
    start = 7
    for i, (ing, cost) in enumerate(RECIPE):
        r = start + i
        ws.cell(row=r, column=2, value=ing).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(RECIPE) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="BATCH COST").style = "th"
    cb = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.fill = fill(SURFACE); cb.number_format = '"$"#,##0.00'
    cell_name(wb, "BatchCost", "Recipe Costing", f"$C${tot}")
    ws.cell(row=tot + 2, column=2, value=f"Yield ({RECIPE_YIELD} cookies)").style = "field_label"
    cy = ws.cell(row=tot + 2, column=3, value=RECIPE_YIELD); cy.style = "input"; cy.number_format = "#,##0"
    ws.cell(row=tot + 3, column=2, value="Cost per cookie").style = "field_label"
    cpc = ws.cell(row=tot + 3, column=3, value=f"=IFERROR(BatchCost/C{tot+2},0)"); cpc.style = "field_value"; cpc.number_format = '"$"#,##0.00'; cpc.fill = fill(MINT_BG)


def build_products(wb):
    ws = wb.create_sheet("Product List"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 12, 12, 12, 14, 2])
    luxe_header(ws, "G", "🍰  PRODUCT LIST",
                "Ingredient cost, price, margin & monthly revenue on every product — price for profit.")
    table_headers(ws, 4, ["Product", "Ing. Cost", "Price", "Margin", "Units/mo", "Revenue"], start_col=2)
    start = L0
    for i, (item, cost, price, units) in enumerate(PRODUCTS):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=5, value=f"=IFERROR((D{r}-C{r})/D{r},0)"); cm.style = "td"; cm.number_format = "0%"
        cu = ws.cell(row=r, column=6, value=units); cu.style = "input"; cu.number_format = "#,##0"
        crv = ws.cell(row=r, column=7, value=f"=D{r}*F{r}"); crv.style = "td"; crv.font = Font(bold=True, color=PRIMARY); crv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRODUCTS) - 1
    nrange(wb, "ProdItem", "Product List", "B", start, end)
    nrange(wb, "ProdCost", "Product List", "C", start, end)
    nrange(wb, "ProdPrice", "Product List", "D", start, end)
    nrange(wb, "ProdMargin", "Product List", "E", start, end)
    nrange(wb, "ProdUnits", "Product List", "F", start, end)
    nrange(wb, "ProdRev", "Product List", "G", start, end)
    br = end + 2
    ws.cell(row=br, column=2, value="Monthly product income").style = "field_label"
    c = ws.cell(row=br, column=7, value="=SUM(ProdRev)"); c.style = "field_value"; c.number_format = '"$"#,##0'; c.fill = fill(MINT_BG)
    cell_name(wb, "MonthlyIncome", "Product List", f"$G${br}")
    ws.cell(row=br + 1, column=2, value="Overall food cost %").style = "field_label"
    c2 = ws.cell(row=br + 1, column=7, value="=IFERROR(SUMPRODUCT(ProdCost,ProdUnits)/SUMPRODUCT(ProdPrice,ProdUnits),0)"); c2.style = "field_value"; c2.number_format = "0%"; c2.fill = fill(MINT_BG)
    cell_name(wb, "FoodCostPct", "Product List", f"$G${br+1}")
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=0.5, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=0.7, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.9, end_color="FF" + HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_orders(wb):
    ws = wb.create_sheet("Custom Orders"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 18, 22, 12, 12, 12, 14, 2])
    luxe_header(ws, "G", "📝  CUSTOM ORDERS",
                "Every custom order — due date, price, deposit & status, so nothing gets missed.")
    table_headers(ws, 4, ["Customer", "Order", "Due", "Price", "Deposit", "Status"], start_col=2)
    start = L0
    import datetime as dt
    for i, (cust, item, days, price, dep, status) in enumerate(ORDERS):
        r = start + i
        ws.cell(row=r, column=2, value=cust).style = "td_left"
        ws.cell(row=r, column=3, value=item).style = "td_left"
        cd = ws.cell(row=r, column=4, value=dt.date.today() + dt.timedelta(days=days)); cd.style = "input"; cd.number_format = "mm/dd"
        cp = ws.cell(row=r, column=5, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cdp = ws.cell(row=r, column=6, value=dep); cdp.style = "input"; cdp.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=7, value=status); cs.style = "td"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"G{r}", "StatusList")
    end = start + len(ORDERS) - 1
    nrange(wb, "OrderCust", "Custom Orders", "B", start, end)
    nrange(wb, "OrderPrice", "Custom Orders", "E", start, end)
    nrange(wb, "OrderStatus", "Custom Orders", "G", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="OPEN ORDER VALUE").style = "th"
    for c in range(3, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cp = ws.cell(row=tot, column=5, value="=SUMIF(OrderStatus,\"<>Delivered\",OrderPrice)"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(MINT_BG); cp.number_format = '"$"#,##0'
    for c in range(6, 8):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cell_name(wb, "OrderValue", "Custom Orders", f"$E${tot}")
    cmap = {"Confirmed": MINT_BG, "Deposit": WARN_BG, "Inquiry": RED_BG, "Delivered": SURFACE}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    ws.freeze_panes = "A5"


def build_ingredients(wb):
    ws, start, end = build_log(
        wb, "Ingredient Costs", "🥣", "INGREDIENT COSTS",
        "Your pantry price list — what each ingredient costs, so recipe costing is accurate.",
        ["Ingredient", "Pack Price", "Pack Size", "Unit"],
        INGREDIENTS, [2, 20, 14, 16, 12, 2], text_left={2, 4}, money2={3}, reserved=24, start_col=2,
        validations=[("E", "UnitList")])


def build_labels(wb):
    ws, start, end = build_log(
        wb, "Labeling & Allergens", "🏷", "LABELING & ALLERGENS",
        "Cottage-food labels — product, allergens & net weight. Check your state's rules for what's required.",
        ["Product", "Allergens", "Net Weight"],
        LABELS, [2, 22, 30, 14, 2], text_left={2, 3, 4}, reserved=24, start_col=2)


def build_markets(wb):
    ws = wb.create_sheet("Markets & Events"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 12, 14, 14, 2])
    luxe_header(ws, "F", "⛺  MARKETS & EVENTS",
                "Farmers markets & pop-ups — the fee, the sales & the profit after the booth cost.")
    table_headers(ws, 4, ["Market / Event", "Date", "Fee", "Sales", "Net"], start_col=2)
    start = L0
    import datetime as dt
    for i, (mkt, days, fee, sales) in enumerate(MARKETS):
        r = start + i
        ws.cell(row=r, column=2, value=mkt).style = "td_left"
        cd = ws.cell(row=r, column=3, value=dt.date.today() + dt.timedelta(days=days)); cd.style = "input"; cd.number_format = "mm/dd"
        cf = ws.cell(row=r, column=4, value=fee); cf.style = "input"; cf.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=5, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        cn = ws.cell(row=r, column=6, value=f"=E{r}-D{r}"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MARKETS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTALS").style = "th"
    for c in (3,):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cf = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cf.style = "td"; cf.font = Font(bold=True, color=PRIMARY); cf.fill = fill(SURFACE); cf.number_format = '"$"#,##0'
    cs = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cn = ws.cell(row=tot, column=6, value=f"=SUM(F{start}:F{end})"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


def build_ledger(wb):
    ws = wb.create_sheet("Income & Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  INCOME & EXPENSES",
                "Your books for tax time — income in, expenses out, and the profit that's really yours.")
    ws.cell(row=5, column=2, value="INCOME").style = "section_gold"
    ws.cell(row=6, column=2, value="Product & order income").style = "td_left"
    ci = ws.cell(row=6, column=3, value="=MonthlyIncome"); ci.style = "td"; ci.font = Font(bold=True, color=PRIMARY); ci.number_format = '"$"#,##0'; ci.fill = fill(MINT_BG)
    cell_name(wb, "TotalIncome", "Income & Expenses", "$C$6")
    ws.cell(row=8, column=2, value="EXPENSES").style = "section_gold"
    table_headers(ws, 9, ["Expense", "Amount"], start_col=2)
    start = 10
    for i, (item, amt, _t) in enumerate(LEDGER):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LEDGER) - 1
    nrange(wb, "ExpAmt", "Income & Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Income & Expenses", f"$C${tot}")
    nr = tot + 2
    ws.cell(row=nr, column=2, value="= NET PROFIT").style = "th"
    cn = ws.cell(row=nr, column=3, value="=TotalIncome-ExpTotal"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetProfit", "Income & Expenses", f"$C${nr}")


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Failed batches & unsold stock — the leaks that quietly eat your profit.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_customers(wb):
    ws, start, end = build_log(
        wb, "Customers", "🤝", "CUSTOMERS",
        "Your regulars — what they order & how to reach them. Repeat customers are your best marketing.",
        ["Customer", "Favorite", "Contact", "Status"],
        CUSTOMERS, [2, 20, 18, 24, 14, 2], text_left={2, 3, 4}, reserved=24, start_col=2,
        validations=[("E", "StatusList")])
    cmap = {"Repeat": MINT_BG, "New": WARN_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"E{start}:E{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 14, 2])
    luxe_header(ws, "E", "📊  MONTHLY SUMMARY",
                "Income, expenses & profit by month — watch the busy season and your profit grow.")
    table_headers(ws, 4, ["Month", "Income", "Expenses", "Profit"], start_col=2)
    start = L0
    for i, (m, inc, exp) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ci = ws.cell(row=r, column=3, value=inc); ci.style = "input"; ci.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=4, value=exp); ce.style = "input"; ce.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthProfit", "Monthly Summary", "E", start, end)
    ws.add_chart(_barchart(ws, "Profit by Month", start, end, 5, 2), "G4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🧁  HOME BAKERY & COTTAGE FOOD COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Income, profit, your hourly wage & a Bakery Score — your whole home bakery, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("PRODUCTS", "=COUNTA(ProdItem)", "num"),
        ("AVG PRICE", "=AVERAGE(ProdPrice)", "money2"),
        ("FOOD COST", "=FoodCostPct", "pct"),
        ("AVG MARGIN", "=AVERAGE(ProdMargin)", "pct"),
        ("MONTHLY INCOME", "=MonthlyIncome", "money"),
        ("TOP SELLER", "=INDEX(ProdItem,MATCH(MAX(ProdRev),ProdRev,0))", "text"),
    ]
    row2 = [
        ("MONTHLY PROFIT", "=NetProfit", "money"),
        ("YOUR HOURLY", "=EffHourly", "money2"),
        ("OPEN ORDERS", "=COUNTIF(OrderStatus,\"<>Delivered\")", "num"),
        ("ORDER VALUE", "=OrderValue", "money"),
        ("MONTHLY UNITS", "=SUM(ProdUnits)", "num"),
        ("BAKERY SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "BAKERY HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "PROFIT BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Food cost on target", "=IFERROR(MIN(TargetFC/FoodCostPct,1),0)"),
        ("Margins healthy", "=IFERROR(MIN(AVERAGE(ProdMargin)/MarginGoal,1),0)"),
        ("Products priced", "=IFERROR(COUNTIF(ProdPrice,\">0\")/COUNTA(ProdItem),0)"),
        ("Paying yourself", "=IFERROR(MIN(EffHourly/WageGoal,1),0)"),
        ("Profitable", "=IFERROR(MIN(NetProfit/ProfitGoal,1),0)"),
        ("Waste low", "=IFERROR(1-MIN((WasteTotal/MonthlyIncome)/WasteLimit,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Profit by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=5, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Home Bakery & Cottage Food Command Center™ — cost every recipe, and pay yourself.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_priceit(wb); build_recipe(wb)
    build_products(wb); build_orders(wb); build_ingredients(wb); build_labels(wb)
    build_markets(wb); build_ledger(wb); build_waste(wb); build_customers(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Price It Right", "Recipe Costing", "Product List", "Custom Orders",
             "Ingredient Costs", "Labeling & Allergens", "Markets & Events", "Income & Expenses", "Waste Log",
             "Customers", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Home_Bakery_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
