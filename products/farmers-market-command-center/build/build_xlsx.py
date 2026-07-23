"""Build Farmers Market Vendor Command Center™ — The Market-Stall Operating System.

14 tabs · a premium farmers-market / market-stall operating system in Google Sheets &
Excel. Dashboard, a per-market booth P&L engine (did this market day actually pay?),
a product price & margin list, ingredient costs, a bake/prep plan, a markets log,
booth & stall costs, packaging, customers, waste, income & expenses and a monthly
summary — one dashboard. Cost every product, and know if the market day paid.

Run: python3 build_xlsx.py   ->  ../Farmers_Market_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["each", "lb", "oz", "loaf", "jar", "bag", "dozen", "bunch"]
CATEGORY = ["Bread", "Pastry", "Preserves", "Pantry", "Produce", "Other"]
CTYPE = ["Regular", "CSA / subscriber", "Wholesale", "New"]

TARGET_COGS = 0.30
MARGIN_GOAL = 0.45
SALES_GOAL = 3000
PROFIT_GOAL = 1500
WASTE_LIMIT = 0.04

# Product price list: (product, cost, price)
PRODUCTS = [
    ("Sourdough Loaf", 1.85, 8), ("Cinnamon Rolls (6)", 2.40, 12), ("Baguette", 1.10, 5),
    ("Focaccia", 1.60, 7), ("Cookies (dz)", 2.20, 10), ("Jam (jar)", 2.80, 9),
    ("Granola (bag)", 3.10, 11), ("Honey (jar)", 4.50, 14),
]

# Market Day — units sold at this Saturday's market (same order as PRODUCTS)
DAY_SOLD = [22, 10, 18, 12, 9, 8, 6, 5]
CUSTOMERS = 64
MARKET_HOURS = 8
# Booth costs for the day: (item, cost)
BOOTH = [("Stall / space fee", 45), ("Fuel & travel", 18), ("Hired help", 60),
         ("Bags & supplies", 12), ("Samples", 8)]

# Ingredient cost library: (ingredient, pack price, pack size, unit)
INGREDIENTS = [
    ("Flour", 0.85, "1 lb", "lb"), ("Butter", 4.20, "1 lb", "lb"), ("Sugar", 0.70, "1 lb", "lb"),
    ("Eggs", 3.60, "dozen", "each"), ("Yeast", 0.15, "packet", "each"), ("Oats", 1.10, "1 lb", "lb"),
    ("Fruit (jam)", 2.40, "1 lb", "lb"), ("Jars & lids", 0.55, "each", "each"),
]

# Bake / prep plan for the market: (product, make qty, prep minutes)
BAKEPLAN = [
    ("Sourdough Loaf", 26, 40), ("Cinnamon Rolls (6)", 12, 35), ("Baguette", 22, 30),
    ("Focaccia", 14, 25), ("Cookies (dz)", 12, 30), ("Jam (jar)", 10, 20),
]

# Markets log (this month): (date, market, sales)
MARKETS = [
    ("Sat · wk 1", "Downtown Market", 768), ("Sat · wk 2", "Downtown Market", 690),
    ("Sat · wk 3", "Riverside Market", 815), ("Sat · wk 4", "Downtown Market", 727),
]

# Booth & stall costs by market: (market, stall fee, notes)
STALLS = [
    ("Downtown Market", 45.00, "Weekly · reserved"), ("Riverside Market", 60.00, "Premium spot"),
    ("Winter Indoor Market", 35.00, "Nov–Mar"), ("Sunday Craft Fair", 50.00, "Occasional"),
]

# Packaging & supplies: (item, cost, notes)
PACKAGING = [
    ("Kraft bags", 0.10, "Per loaf"), ("Bakery boxes", 0.45, "Rolls & pastries"),
    ("Jam labels", 0.06, "Printed"), ("Twine & tags", 0.04, "Finishing"),
    ("Tent & table", 0.00, "Owned"), ("Card reader fees", 0.03, "Per $1 sold"),
]

# Customers / regulars: (name, type, notes)
CUSTOMERS_LIST = [
    ("The Halls", "CSA / subscriber", "Weekly bread share"), ("Cafe on 5th", "Wholesale", "12 loaves/wk"),
    ("Marta R.", "Regular", "Always cinnamon rolls"), ("Devi & Sam", "CSA / subscriber", "Bread + jam"),
    ("Inn on Main", "Wholesale", "Focaccia Fridays"),
]

# Waste log: (item, reason, cost) — monthly total $72
WASTE = [
    ("Unsold bread", "End of market", 32.00), ("Over-bake", "Slow day", 24.00),
    ("Spoiled fruit", "Jam batch", 16.00),
]

# Income & expenses (monthly): (item, amount)
LEDGER = [
    ("Ingredients", 735, "E"), ("Booth & stall fees", 180, "E"), ("Fuel & travel", 75, "E"),
    ("Hired help", 240, "E"), ("Packaging & supplies", 95, "E"), ("Market & insurance fees", 80, "E"),
]

# Monthly summary: (month, sales, expenses)
MONTHS = [("Jul", 3000, 1405), ("Aug", 3200, 1470), ("Sep", 3450, 1560),
          ("Oct", 3100, 1440), ("Nov", 2600, 1290), ("Dec", 3600, 1650)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Business name", "Harvest Lane Market Co.", None, "Business"),
        ("Owner", "Sage", None, "Owner"),
        ("Target COGS %", TARGET_COGS, "0%", "TargetCOGS"),
        ("Net-margin goal %", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Monthly sales goal", SALES_GOAL, '"$"#,##0', "SalesGoal"),
        ("Monthly profit goal", PROFIT_GOAL, '"$"#,##0', "ProfitGoal"),
        ("Waste limit %", WASTE_LIMIT, "0%", "WasteLimit"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Category", CATEGORY, "CategoryList"),
             ("G", "Cust. type", CTYPE, "CTypeList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🧺  FARMERS MARKET VENDOR COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Cost every product, and know if the market day paid.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE MARKET STALL, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A market stall lives or dies on two numbers: the margin on each product, and whether a whole "
                      "market day actually paid after the stall fee, fuel and hired help. This makes both visible: a "
                      "product price & margin list, and a per-market booth P&L that takes what you sold, subtracts your "
                      "cost of goods and your booth costs, and tells you the booth net and what you made per hour. Plan "
                      "your bake, track ingredient and packaging costs, keep your regulars and CSA subscribers, and run "
                      "income & expenses — all in ONE premium Google Sheets & Excel system built for market vendors.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your target COGS % & monthly sales goal.",
             "2.  Price your Products — cost, price and the margin calculates live.",
             "3.  After a market, enter what you sold in Market Day.",
             "4.  Add your booth costs; the booth net & net-per-hour appear.",
             "5.  Plan your bake, track ingredients, packaging & customers.",
             "6.  Check the Dashboard: sales, profit & a Vendor Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional market stall (Harvest Lane Market Co.) is included so you can see how it "
               "all connects — just type over it with your own products and sales. Product margin and the booth net "
               "per market day are the two numbers that decide whether a stall pays, and they roll into a live Vendor "
               "Score. Twelve matching printable pages (booth P&L, product price list, bake plan, market checklist & "
               "more) are included. This is a business tool, not financial, legal or food-safety advice — follow your "
               "local cottage-food and market rules and confirm figures with your own advisors.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Margin per product × a booth day that pays = a stall that's worth it.", "section_gold")


# ===========================================================================
def build_products(wb):
    ws = wb.create_sheet("Products"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 12, 14, 12, 2])
    luxe_header(ws, "F", "🥖  PRODUCTS",
                "Your stall's price list — cost, price and the margin on every product you sell.")
    table_headers(ws, 4, ["Product", "Cost", "Price", "Margin", "Margin %"], start_col=2)
    start = L0
    for i, (prod, cost, price) in enumerate(PRODUCTS):
        r = start + i
        ws.cell(row=r, column=2, value=prod).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=5, value=f"=D{r}-C{r}"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.number_format = '"$"#,##0.00'
        cmp = ws.cell(row=r, column=6, value=f"=IFERROR((D{r}-C{r})/D{r},0)"); cmp.style = "td"; cmp.number_format = "0%"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PRODUCTS) - 1
    nrange(wb, "ProductName", "Products", "B", start, end)
    nrange(wb, "ProductCost", "Products", "C", start, end)
    nrange(wb, "ProductPrice", "Products", "D", start, end)
    ws.conditional_formatting.add(f"F{start}:F{end}",
        ColorScaleRule(start_type="num", start_value=0.40, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=0.60, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.80, end_color="FF" + HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_marketday(wb):
    ws = wb.create_sheet("Market Day"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 10, 12, 14, 14, 2])
    luxe_header(ws, "F", "🧺  MARKET DAY",
                "Enter what you sold — the booth P&L tells you if this market day actually paid.")
    ws.cell(row=5, column=2, value="MARKET").style = "section_gold"
    ws.cell(row=5, column=3, value="Saturday Market").font = Font(bold=True, color=PRIMARY)
    ws.cell(row=5, column=5, value="Customers").style = "field_label"
    cc = ws.cell(row=5, column=6, value=CUSTOMERS); cc.style = "input"; cc.number_format = "#,##0"
    cell_name(wb, "Customers", "Market Day", "$F$5")
    table_headers(ws, 7, ["Product", "Sold", "Price", "Revenue", "Line Cost"], start_col=2)
    start = 8
    for i, (prod, cost, price) in enumerate(PRODUCTS):
        r = start + i
        ws.cell(row=r, column=2, value=prod).style = "td_left"
        cs = ws.cell(row=r, column=3, value=DAY_SOLD[i]); cs.style = "input"; cs.number_format = "#,##0"
        cp = ws.cell(row=r, column=4, value=f"=IFERROR(INDEX(ProductPrice,MATCH(B{r},ProductName,0)),0)"); cp.style = "td"; cp.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=5, value=f"=C{r}*D{r}"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.number_format = '"$"#,##0'
        clc = ws.cell(row=r, column=6, value=f"=C{r}*IFERROR(INDEX(ProductCost,MATCH(B{r},ProductName,0)),0)"); clc.style = "td"; clc.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"B{r}", "ProductName")
    end = start + len(PRODUCTS) - 1
    nrange(wb, "DayProduct", "Market Day", "B", start, end)
    nrange(wb, "DaySoldRng", "Market Day", "C", start, end)
    nrange(wb, "DayRev", "Market Day", "E", start, end)
    nrange(wb, "DayLineCost", "Market Day", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="MARKET TOTALS").style = "th"
    cst = ws.cell(row=tot, column=3, value="=SUM(DaySoldRng)"); cst.style = "td"; cst.font = Font(bold=True, color=PRIMARY); cst.fill = fill(SURFACE); cst.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    crt = ws.cell(row=tot, column=5, value="=SUM(DayRev)"); crt.style = "td"; crt.font = Font(bold=True, color=PRIMARY); crt.fill = fill(MINT_BG); crt.number_format = '"$"#,##0'
    cct = ws.cell(row=tot, column=6, value="=SUM(DayLineCost)"); cct.style = "td"; cct.font = Font(bold=True, color=PRIMARY); cct.fill = fill(SURFACE); cct.number_format = '"$"#,##0'
    cell_name(wb, "DayUnits", "Market Day", f"$C${tot}")
    cell_name(wb, "DaySales", "Market Day", f"$E${tot}")
    cell_name(wb, "DayCOGS", "Market Day", f"$F${tot}")
    # booth costs block
    bh = tot + 2
    ws.cell(row=bh, column=2, value="BOOTH COSTS (THIS DAY)").style = "section_gold"
    bs = bh + 1
    for i, (item, cost) in enumerate(BOOTH):
        r = bs + i
        ws.cell(row=r, column=2, value=item).style = "field_label"
        cbx = ws.cell(row=r, column=6, value=cost); cbx.style = "input"; cbx.number_format = '"$"#,##0'
    be = bs + len(BOOTH) - 1
    nrange(wb, "BoothItems", "Market Day", "F", bs, be)
    btot = be + 1
    ws.cell(row=btot, column=2, value="TOTAL BOOTH COSTS").style = "th"
    cbt = ws.cell(row=btot, column=6, value="=SUM(BoothItems)"); cbt.style = "td"; cbt.font = Font(bold=True, color=DANGER); cbt.fill = fill(SURFACE); cbt.number_format = '"$"#,##0'
    cell_name(wb, "BoothCosts", "Market Day", f"$F${btot}")
    hr = btot + 1
    ws.cell(row=hr, column=2, value="Hours (setup + market + pack-down)").style = "field_label"
    chh = ws.cell(row=hr, column=6, value=MARKET_HOURS); chh.style = "input"; chh.number_format = "0.0"
    cell_name(wb, "MarketHours", "Market Day", f"$F${hr}")
    # results
    nr = hr + 2
    ws.cell(row=nr, column=2, value="= BOOTH NET (sales − COGS − booth costs)").style = "th"
    cbn = ws.cell(row=nr, column=6, value="=DaySales-DayCOGS-BoothCosts"); cbn.style = "td"; cbn.font = Font(bold=True, size=13, color=PRIMARY); cbn.fill = fill(MINT_BG); cbn.number_format = '"$"#,##0'
    cell_name(wb, "BoothNet", "Market Day", f"$F${nr}")
    ws.cell(row=nr + 1, column=2, value="= NET PER HOUR").style = "th"
    cnh = ws.cell(row=nr + 1, column=6, value="=IFERROR(BoothNet/MarketHours,0)"); cnh.style = "td"; cnh.font = Font(bold=True, size=12, color=PRIMARY); cnh.fill = fill(MINT_BG); cnh.number_format = '"$"#,##0.00'
    cell_name(wb, "NetHour", "Market Day", f"$F${nr+1}")
    fr = nr + 3
    ws.cell(row=fr, column=2, value="COGS % (of sales)").style = "field_label"
    ccp = ws.cell(row=fr, column=6, value="=IFERROR(DayCOGS/DaySales,0)"); ccp.style = "field_value"; ccp.number_format = "0%"; ccp.fill = fill(MINT_BG)
    cell_name(wb, "COGSPct", "Market Day", f"$F${fr}")
    ws.cell(row=fr + 1, column=2, value="Net margin % (booth net ÷ sales)").style = "field_label"
    cnm = ws.cell(row=fr + 1, column=6, value="=IFERROR(BoothNet/DaySales,0)"); cnm.style = "field_value"; cnm.number_format = "0%"; cnm.fill = fill(MINT_BG)
    cell_name(wb, "NetMarginPct", "Market Day", f"$F${fr+1}")
    ws.cell(row=fr + 2, column=2, value="Average basket (sales ÷ customers)").style = "field_label"
    cab = ws.cell(row=fr + 2, column=6, value="=IFERROR(DaySales/Customers,0)"); cab.style = "field_value"; cab.number_format = '"$"#,##0.00'; cab.fill = fill(MINT_BG)
    cell_name(wb, "AvgBasket", "Market Day", f"$F${fr+2}")
    ws.cell(row=fr + 3, column=2, value="Top seller (by revenue)").style = "field_label"
    cts = ws.cell(row=fr + 3, column=6, value="=IFERROR(INDEX(DayProduct,MATCH(MAX(DayRev),DayRev,0)),\"\")"); cts.style = "field_value"; cts.fill = fill(MINT_BG)
    cell_name(wb, "TopSeller", "Market Day", f"$F${fr+3}")


def build_ingredients(wb):
    ws, start, end = build_log(
        wb, "Ingredient Costs", "🌾", "INGREDIENT COSTS",
        "Your buying list — what each ingredient costs, so product costing stays accurate.",
        ["Ingredient", "Pack Price", "Pack Size", "Unit"],
        INGREDIENTS, [2, 22, 14, 14, 12, 2], text_left={2, 4}, money2={3}, reserved=24, start_col=2,
        validations=[("E", "UnitList")])


def build_bakeplan(wb):
    ws = wb.create_sheet("Bake Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 14, 14, 2])
    luxe_header(ws, "E", "👩‍🍳  BAKE PLAN",
                "What to make for the market — quantities and the prep time it takes.")
    table_headers(ws, 4, ["Product", "Make Qty", "Prep (min)", "Prep (hrs)"], start_col=2)
    start = L0
    for i, (prod, qty, prep) in enumerate(BAKEPLAN):
        r = start + i
        ws.cell(row=r, column=2, value=prod).style = "td_left"
        cq = ws.cell(row=r, column=3, value=qty); cq.style = "input"; cq.number_format = "#,##0"
        cp = ws.cell(row=r, column=4, value=prep); cp.style = "input"; cp.number_format = "#,##0"
        ch = ws.cell(row=r, column=5, value=f"=D{r}/60"); ch.style = "td"; ch.number_format = "0.0"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BAKEPLAN) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="MARKET TOTAL").style = "th"
    cq = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cq.style = "td"; cq.font = Font(bold=True, color=PRIMARY); cq.fill = fill(SURFACE); cq.number_format = "#,##0"
    cp = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.fill = fill(SURFACE); cp.number_format = "#,##0"
    ch = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); ch.style = "td"; ch.font = Font(bold=True, color=PRIMARY); ch.fill = fill(SURFACE); ch.number_format = "0.0"
    ws.freeze_panes = "A5"


def build_markets(wb):
    ws = wb.create_sheet("Markets"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 24, 16, 2])
    luxe_header(ws, "D", "📅  MARKETS",
                "Every market day this month — where you sold and the day's sales, so monthly sales roll up.")
    table_headers(ws, 4, ["Date", "Market", "Day Sales"], start_col=2)
    start = L0
    for i, (date, market, sales) in enumerate(MARKETS):
        r = start + i
        ws.cell(row=r, column=2, value=date).style = "td_left"
        ws.cell(row=r, column=3, value=market).style = "td_left"
        cs = ws.cell(row=r, column=4, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MARKETS) - 1
    nrange(wb, "MarketSales", "Markets", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="THIS MONTH").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    cs = ws.cell(row=tot, column=4, value="=SUM(MarketSales)"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(MINT_BG); cs.number_format = '"$"#,##0'
    cell_name(wb, "MonthlySales", "Markets", f"$D${tot}")
    sr = tot + 2
    ws.cell(row=sr, column=2, value="Markets this month").style = "field_label"
    c = ws.cell(row=sr, column=4, value="=COUNTA(MarketSales)"); c.style = "field_value"; c.number_format = "#,##0"; c.fill = fill(MINT_BG)
    cell_name(wb, "MarketCount", "Markets", f"$D${sr}")
    ws.freeze_panes = "A5"


def build_stalls(wb):
    ws, start, end = build_log(
        wb, "Booth Costs", "⛺", "BOOTH COSTS",
        "Your stall fees by market — what a space costs before you sell a thing.",
        ["Market", "Stall Fee", "Notes"],
        STALLS, [2, 26, 14, 24, 2], text_left={2, 4}, money2={3}, reserved=24, start_col=2)


def build_packaging(wb):
    ws, start, end = build_log(
        wb, "Packaging", "🛍", "PACKAGING & SUPPLIES",
        "Bags, boxes & labels — the small costs that quietly eat your margin.",
        ["Item", "Cost", "Notes"],
        PACKAGING, [2, 24, 14, 24, 2], text_left={2, 4}, money2={3}, reserved=24, start_col=2)


def build_customers(wb):
    ws, start, end = build_log(
        wb, "Customers", "🤝", "CUSTOMERS",
        "Your regulars, CSA subscribers & wholesale accounts — the base that comes back.",
        ["Name", "Type", "Notes"],
        CUSTOMERS_LIST, [2, 24, 20, 26, 2], text_left={2, 4}, reserved=24, start_col=2,
        validations=[("C", "CTypeList")])


def build_waste(wb):
    ws, start, end = build_log(
        wb, "Waste Log", "🗑", "WASTE LOG",
        "Unsold stock, over-bake & spoilage — the leaks that eat your booth net.",
        ["Item", "Reason", "Cost"],
        WASTE, [2, 26, 24, 14, 2], text_left={2, 3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "WasteCost", "Waste Log", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL WASTE").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(WasteCost)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    cell_name(wb, "WasteTotal", "Waste Log", f"$D${tot}")


def build_ledger(wb):
    ws = wb.create_sheet("Income & Expenses"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "💰  INCOME & EXPENSES",
                "Monthly books — market sales in, costs out, and the profit that's really yours.")
    ws.cell(row=5, column=2, value="INCOME").style = "section_gold"
    ws.cell(row=6, column=2, value="Monthly market sales").style = "td_left"
    ci = ws.cell(row=6, column=3, value="=MonthlySales"); ci.style = "td"; ci.font = Font(bold=True, color=PRIMARY); ci.number_format = '"$"#,##0'; ci.fill = fill(MINT_BG)
    cell_name(wb, "TotalIncome", "Income & Expenses", "$C$6")
    ws.cell(row=8, column=2, value="EXPENSES").style = "section_gold"
    table_headers(ws, 9, ["Expense", "Amount"], start_col=2)
    start = 10
    for i, (item, amt, _t) in enumerate(LEDGER):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LEDGER) - 1
    nrange(wb, "ExpAmt", "Income & Expenses", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL EXPENSES").style = "th"
    ce = ws.cell(row=tot, column=3, value="=SUM(ExpAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Income & Expenses", f"$C${tot}")
    nr = tot + 2
    ws.cell(row=nr, column=2, value="= MONTHLY PROFIT").style = "th"
    cn = ws.cell(row=nr, column=3, value="=TotalIncome-ExpTotal"); cn.style = "td"; cn.font = Font(bold=True, size=13, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NetProfit", "Income & Expenses", f"$C${nr}")


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 14, 2])
    luxe_header(ws, "E", "📈  MONTHLY SUMMARY",
                "Sales, expenses & profit by month — watch the season swing.")
    table_headers(ws, 4, ["Month", "Sales", "Expenses", "Profit"], start_col=2)
    start = L0
    for i, (m, sales, exp) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ci = ws.cell(row=r, column=3, value=sales); ci.style = "input"; ci.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=4, value=exp); ce.style = "input"; ce.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthSales", "Monthly Summary", "C", start, end)
    ws.add_chart(_barchart(ws, "Sales by Month", start, end, 3, 2), "G4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🧺  FARMERS MARKET VENDOR COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Product margin, booth net per market, monthly sales & a Vendor Score — at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("UNITS SOLD", "=DayUnits", "num"),
        ("MARKET SALES", "=DaySales", "money"),
        ("COGS %", "=COGSPct", "pct"),
        ("BOOTH NET", "=BoothNet", "money"),
        ("NET / HOUR", "=NetHour", "money2"),
        ("TOP SELLER", "=TopSeller", "text"),
    ]
    row2 = [
        ("AVG BASKET", "=AvgBasket", "money2"),
        ("MONTHLY SALES", "=MonthlySales", "money"),
        ("MARKETS / MO", "=MarketCount", "num"),
        ("MONTHLY PROFIT", "=NetProfit", "money"),
        ("WASTE %", "=IFERROR(WasteTotal/MonthlySales,0)", "pct"),
        ("VENDOR SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "VENDOR HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "SALES BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("COGS on target", "=IFERROR(MIN(TargetCOGS/COGSPct,1),0)"),
        ("Margin healthy", "=IFERROR(MIN(NetMarginPct/MarginGoal,1),0)"),
        ("Products priced", "=IFERROR(COUNTIF(ProductPrice,\">0\")/COUNTA(ProductName),0)"),
        ("Sales vs goal", "=IFERROR(MIN(MonthlySales/SalesGoal,1),0)"),
        ("Profitable", "=IFERROR(MIN(NetProfit/ProfitGoal,1),0)"),
        ("Waste low", "=IFERROR(1-MIN((WasteTotal/MonthlySales)/WasteLimit,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Sales by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Farmers Market Vendor Command Center™ — cost every product, and know if the market day paid.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_products(wb); build_marketday(wb)
    build_ingredients(wb); build_bakeplan(wb); build_markets(wb); build_stalls(wb)
    build_packaging(wb); build_customers(wb); build_waste(wb); build_ledger(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Market Day", "Products", "Ingredient Costs", "Bake Plan",
             "Markets", "Booth Costs", "Packaging", "Customers", "Waste Log",
             "Income & Expenses", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Farmers_Market_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
