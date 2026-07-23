"""Build Rental Property & Landlord Command Center™ — The Buy-&-Hold Operating System.

14 tabs · a premium rental-property / landlord operating system in Google Sheets &
Excel. Dashboard, a deal-analyzer engine (NOI, cap rate, cash-on-cash, DSCR, cash
flow), a rent roll, tenants, a rent ledger, expenses, maintenance & CapEx, mortgage,
reserves, mileage, renewals and a monthly summary — one dashboard. Know the real
return on every door.

Run: python3 build_xlsx.py   ->  ../Rental_Property_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITSTATUS = ["Occupied", "Vacant", "Notice given", "Turnover", "Rehab"]
PAYSTATUS = ["Paid", "Partial", "Late", "Unpaid"]
EXPCAT = ["Tax", "Insurance", "Management", "Repairs", "Utilities", "CapEx", "Other"]

CF_GOAL = 300
CAP_GOAL = 0.055
COC_GOAL = 0.07
DSCR_GOAL = 1.25
EXP_TARGET = 0.50
RESERVE_GOAL_MONTHS = 6

# Deal Analyzer — flagship unit income
MONTHLY_RENT = 1850
OTHER_INCOME = 75
# Operating expenses (monthly): (item, amount)
OPEX = [
    ("Property tax", 210), ("Insurance", 90), ("Property management (8%)", 154), ("Repairs & maintenance", 100),
    ("Vacancy reserve (5%)", 96), ("CapEx reserve", 85), ("Utilities (owner-paid)", 55), ("Misc / HOA", 40),
]
MORTGAGE_PI = 735
PURCHASE_PRICE = 219000
CASH_INVESTED = 54000
RESERVE_BALANCE = 3756

# Rent roll — portfolio units: (unit, rent, status, tenant)
UNITS = [
    ("123 Maple St · A", 1850, "Occupied", "Ramirez"), ("123 Maple St · B", 1650, "Occupied", "Kelly"),
    ("88 Oak Ave", 2100, "Occupied", "Osei"), ("12 Birch Ct", 1400, "Turnover", "—"),
]

# Tenants: (tenant, unit, lease ends, deposit)
TENANTS = [
    ("Ramirez", "Maple A", "2027-03-31", 1850), ("Kelly", "Maple B", "2026-11-30", 1650),
    ("Osei", "88 Oak", "2027-06-30", 2100), ("—", "Birch Ct", "vacant", 0),
]

# Rent ledger — payments this month: (date, unit, amount, status)
LEDGER_PAY = [
    ("Jul 1", "Maple A", 1850, "Paid"), ("Jul 1", "Maple B", 1650, "Paid"),
    ("Jul 1", "88 Oak", 2100, "Paid"), ("Jul 3", "Birch Ct", 0, "Vacant"),
]

# Expenses log — actuals: (date, category, payee, amount)
EXP_LOG = [
    ("Jul 2", "Repairs", "Ace Plumbing", 140.00), ("Jul 8", "Utilities", "City Water", 62.00),
    ("Jul 12", "Management", "Doorstep PM", 154.00), ("Jul 18", "Insurance", "StateFarm", 90.00),
    ("Jul 25", "CapEx", "Home Depot", 220.00),
]

# Maintenance & CapEx log: (date, unit, item, cost)
MAINT = [
    ("Jul 2", "Maple A", "Water heater flush", 120.00), ("Jul 10", "88 Oak", "Gutter repair", 180.00),
    ("Jul 19", "Maple B", "Garbage disposal", 95.00), ("Jul 26", "Birch Ct", "Turn paint & carpet", 640.00),
]

# Mortgage & loan: (field, value)
LOANS = [
    ("123 Maple St", 175200, 6.75, 735), ("88 Oak Ave", 168000, 6.50, 690),
]

# Reserves & escrow: (fund, balance, goal)
RESERVES = [
    ("Operating reserve", 3756, 9390), ("CapEx sinking fund", 2400, 6000),
    ("Tax & insurance escrow", 1800, 3600), ("Vacancy reserve", 1150, 2400),
]

# Mileage & travel: (date, property, miles, purpose)
MILEAGE = [
    ("Jul 2", "Maple St", 14, "Repair oversight"), ("Jul 10", "88 Oak", 22, "Gutter check"),
    ("Jul 19", "Maple St", 14, "Disposal fix"), ("Jul 26", "Birch Ct", 30, "Turnover walk"),
]

# Renewals & docs: (item, unit, due, done)
RENEWALS = [
    ("Lease renewal", "Maple B", "2026-11-30", "No"), ("Insurance policy", "Portfolio", "2026-09-15", "No"),
    ("Rental license", "88 Oak", "2026-10-01", "No"), ("Smoke/CO inspection", "Maple St", "2026-08-20", "Yes"),
]

# Monthly summary: (month, income, expenses)
MONTHS = [("Jul", 7000, 4200), ("Aug", 7000, 3900), ("Sep", 7100, 4050),
          ("Oct", 5600, 4600), ("Nov", 7000, 3850), ("Dec", 7000, 4100)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.00", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TARGETS", "section")
    controls = [
        ("Portfolio name", "Maple Lane Rentals", None, "Business"),
        ("Owner", "Jordan", None, "Owner"),
        ("Cash-flow goal ($/mo)", CF_GOAL, '"$"#,##0', "CFGoal"),
        ("Cap-rate goal %", CAP_GOAL, "0.0%", "CapGoal"),
        ("Cash-on-cash goal %", COC_GOAL, "0.0%", "CoCGoal"),
        ("DSCR goal", DSCR_GOAL, "0.00", "DSCRGoal"),
        ("Expense-ratio target %", EXP_TARGET, "0%", "ExpTarget"),
        ("Reserve goal (months)", RESERVE_GOAL_MONTHS, "#,##0", "ReserveGoalMonths"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit status", UNITSTATUS, "UnitStatusList"), ("F", "Pay status", PAYSTATUS, "PayStatusList"),
             ("G", "Exp category", EXPCAT, "ExpCatList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🏘  RENTAL PROPERTY & LANDLORD COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know the real return on every door.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE RENTAL PORTFOLIO, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A rental lives or dies on two numbers: the cash flow it throws off each month, and the return on "
                      "the cash you put in. This makes both visible: a deal analyzer that takes your rent, subtracts "
                      "every operating cost and the mortgage, and shows your NOI, cash flow, cap rate, cash-on-cash and "
                      "DSCR. Run your rent roll and tenants, log rent payments and expenses, track maintenance and CapEx, "
                      "your mortgage and reserves, mileage and renewals — all in ONE premium Google Sheets & Excel system "
                      "built for buy-and-hold landlords.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your cash-flow, cap-rate & DSCR goals.",
             "2.  Run the Deal Analyzer — rent, expenses, mortgage → the returns.",
             "3.  Build your Rent Roll & add tenants and lease dates.",
             "4.  Log rent payments and actual expenses as they happen.",
             "5.  Track maintenance, CapEx, reserves and renewals.",
             "6.  Check the Dashboard: cash flow, cap rate & a Landlord Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional portfolio (Maple Lane Rentals) is included so you can see how it all "
               "connects — just type over it with your own doors and numbers. Monthly cash flow and cash-on-cash return "
               "are the two numbers that decide whether a rental is worth owning, and they roll into a live Landlord "
               "Score. Twelve matching printable pages (deal analyzer, rent roll, tenant sheet, reserve tracker & more) "
               "are included. This is a business tool, not financial, legal, tax or investment advice — confirm every "
               "figure with your own advisors before you buy or refinance.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Analyze before you buy, and measure every month — that's how doors pay.", "section_gold")


# ===========================================================================
def build_dealanalyzer(wb):
    ws = wb.create_sheet("Deal Analyzer"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 36, 16, 2])
    luxe_header(ws, "C", "🧮  DEAL ANALYZER",
                "Rent in, every cost and the mortgage out — your NOI, cash flow, cap rate, cash-on-cash and DSCR.")
    ws.cell(row=5, column=2, value="INCOME (per month)").style = "section_gold"
    ws.cell(row=6, column=2, value="Monthly rent").style = "field_label"
    cr = ws.cell(row=6, column=3, value=MONTHLY_RENT); cr.style = "input"; cr.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyRent", "Deal Analyzer", "$C$6")
    ws.cell(row=7, column=2, value="Other income (pet, parking, laundry)").style = "field_label"
    co = ws.cell(row=7, column=3, value=OTHER_INCOME); co.style = "input"; co.number_format = '"$"#,##0'
    ws.cell(row=8, column=2, value="= GROSS MONTHLY INCOME").style = "th"
    cg = ws.cell(row=8, column=3, value="=C6+C7"); cg.style = "td"; cg.font = Font(bold=True, color=PRIMARY); cg.fill = fill(MINT_BG); cg.number_format = '"$"#,##0'
    cell_name(wb, "GrossIncome", "Deal Analyzer", "$C$8")
    ws.cell(row=10, column=2, value="OPERATING EXPENSES (per month)").style = "section_gold"
    table_headers(ws, 11, ["Expense", "Amount"], start_col=2)
    start = 12
    for i, (item, amt) in enumerate(OPEX):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        ca = ws.cell(row=r, column=3, value=amt); ca.style = "input"; ca.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(OPEX) - 1
    nrange(wb, "OpExAmt", "Deal Analyzer", "C", start, end)
    otot = end + 1
    ws.cell(row=otot, column=2, value="= TOTAL OPERATING EXPENSES").style = "th"
    ce = ws.cell(row=otot, column=3, value="=SUM(OpExAmt)"); ce.style = "td"; ce.font = Font(bold=True, color=DANGER); ce.fill = fill(SURFACE); ce.number_format = '"$"#,##0'
    cell_name(wb, "OpExTotal", "Deal Analyzer", f"$C${otot}")
    noi = otot + 1
    ws.cell(row=noi, column=2, value="= NET OPERATING INCOME (NOI)").style = "th"
    cn = ws.cell(row=noi, column=3, value="=GrossIncome-OpExTotal"); cn.style = "td"; cn.font = Font(bold=True, size=12, color=PRIMARY); cn.fill = fill(MINT_BG); cn.number_format = '"$"#,##0'
    cell_name(wb, "NOI", "Deal Analyzer", f"$C${noi}")
    dr = noi + 2
    ws.cell(row=dr, column=2, value="DEBT & CASH FLOW").style = "section_gold"
    ws.cell(row=dr + 1, column=2, value="Mortgage payment (P&I)").style = "field_label"
    cm = ws.cell(row=dr + 1, column=3, value=MORTGAGE_PI); cm.style = "input"; cm.number_format = '"$"#,##0'
    cell_name(wb, "Mortgage", "Deal Analyzer", f"$C${dr+1}")
    ws.cell(row=dr + 2, column=2, value="= MONTHLY CASH FLOW").style = "th"
    cf = ws.cell(row=dr + 2, column=3, value="=NOI-Mortgage"); cf.style = "td"; cf.font = Font(bold=True, size=13, color=PRIMARY); cf.fill = fill(MINT_BG); cf.number_format = '"$"#,##0'
    cell_name(wb, "CashFlow", "Deal Analyzer", f"$C${dr+2}")
    # deal metrics
    mr = dr + 4
    ws.cell(row=mr, column=2, value="THE RETURNS").style = "section_gold"
    ws.cell(row=mr + 1, column=2, value="Purchase price").style = "field_label"
    cp = ws.cell(row=mr + 1, column=3, value=PURCHASE_PRICE); cp.style = "input"; cp.number_format = '"$"#,##0'
    cell_name(wb, "Price", "Deal Analyzer", f"$C${mr+1}")
    ws.cell(row=mr + 2, column=2, value="Cash invested (down + closing + rehab)").style = "field_label"
    ci = ws.cell(row=mr + 2, column=3, value=CASH_INVESTED); ci.style = "input"; ci.number_format = '"$"#,##0'
    cell_name(wb, "CashInvested", "Deal Analyzer", f"$C${mr+2}")
    ws.cell(row=mr + 3, column=2, value="Operating reserve balance").style = "field_label"
    cres = ws.cell(row=mr + 3, column=3, value=RESERVE_BALANCE); cres.style = "input"; cres.number_format = '"$"#,##0'
    cell_name(wb, "Reserve", "Deal Analyzer", f"$C${mr+3}")
    metrics = [
        ("Cap rate (NOI × 12 ÷ price)", "=IFERROR(NOI*12/Price,0)", "0.00%", "CapRate"),
        ("Cash-on-cash (cash flow × 12 ÷ invested)", "=IFERROR(CashFlow*12/CashInvested,0)", "0.0%", "CoC"),
        ("DSCR (NOI ÷ mortgage)", "=IFERROR(NOI/Mortgage,0)", "0.00", "DSCR"),
        ("1% rule (rent ÷ price)", "=IFERROR(MonthlyRent/Price,0)", "0.00%", "OnePct"),
        ("Annualized cash flow (× 12)", "=CashFlow*12", '"$"#,##0', "AnnualCF"),
        ("Reserves funded (of goal months)", "=IFERROR(Reserve/(ReserveGoalMonths*(OpExTotal+Mortgage)),0)", "0%", "ReservePct"),
    ]
    for i, (lab, fml, fmt, nm) in enumerate(metrics):
        r = mr + 5 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt; c.fill = fill(MINT_BG)
        cell_name(wb, nm, "Deal Analyzer", f"$C${r}")


def build_rentroll(wb):
    ws = wb.create_sheet("Rent Roll"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 14, 16, 16, 2])
    luxe_header(ws, "E", "🏘  RENT ROLL",
                "Every door — rent, status and tenant, so your portfolio's monthly rent rolls up.")
    table_headers(ws, 4, ["Unit", "Rent", "Status", "Tenant"], start_col=2)
    start = L0
    for i, (unit, rent, status, tenant) in enumerate(UNITS):
        r = start + i
        ws.cell(row=r, column=2, value=unit).style = "td_left"
        cr = ws.cell(row=r, column=3, value=rent); cr.style = "input"; cr.number_format = '"$"#,##0'
        ws.cell(row=r, column=4, value=status).style = "td"
        ws.cell(row=r, column=5, value=tenant).style = "td_left"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
        add_dv(ws, f"D{r}", "UnitStatusList")
    end = start + len(UNITS) - 1
    nrange(wb, "UnitRent", "Rent Roll", "C", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="PORTFOLIO RENT").style = "th"
    cr = ws.cell(row=tot, column=3, value="=SUM(UnitRent)"); cr.style = "td"; cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0'
    for c in (4, 5):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cell_name(wb, "PortfolioRent", "Rent Roll", f"$C${tot}")
    ws.freeze_panes = "A5"


def build_tenants(wb):
    ws, start, end = build_log(
        wb, "Tenants", "👤", "TENANTS",
        "Who's in each unit, when the lease ends, and the deposit you hold.",
        ["Tenant", "Unit", "Lease ends", "Deposit"],
        TENANTS, [2, 20, 16, 16, 14, 2], text_left={2, 3, 4}, money={5}, reserved=24, start_col=2)


def build_ledger(wb):
    ws, start, end = build_log(
        wb, "Rent Ledger", "📒", "RENT LEDGER",
        "Rent received this period — who paid, how much, and what's still outstanding.",
        ["Date", "Unit", "Amount", "Status"],
        LEDGER_PAY, [2, 14, 18, 14, 14, 2], text_left={2, 3}, money={4}, reserved=30, start_col=2,
        validations=[("E", "PayStatusList")])
    nrange(wb, "RentPaid", "Rent Ledger", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="COLLECTED").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(RentPaid)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(MINT_BG); c.number_format = '"$"#,##0'
    ws.cell(row=tot, column=5).style = "td"; ws.cell(row=tot, column=5).fill = fill(SURFACE)


def build_explog(wb):
    ws, start, end = build_log(
        wb, "Expenses Log", "🧾", "EXPENSES LOG",
        "Actual spending as it happens — by category, so tax time is already done.",
        ["Date", "Category", "Payee", "Amount"],
        EXP_LOG, [2, 14, 16, 20, 14, 2], text_left={2, 4}, money2={5}, reserved=30, start_col=2,
        validations=[("C", "ExpCatList")])
    nrange(wb, "ExpLogAmt", "Expenses Log", "F", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL SPENT").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=6, value="=SUM(ExpLogAmt)"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


def build_maint(wb):
    ws, start, end = build_log(
        wb, "Maintenance & CapEx", "🔧", "MAINTENANCE & CAPEX",
        "Repairs and capital improvements by unit — protect the asset and the deduction.",
        ["Date", "Unit", "Item", "Cost"],
        MAINT, [2, 14, 16, 28, 14, 2], text_left={2, 3, 4}, money2={5}, reserved=28, start_col=2)


def build_loans(wb):
    ws = wb.create_sheet("Mortgage & Loan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 16, 12, 14, 2])
    luxe_header(ws, "E", "🏦  MORTGAGE & LOAN",
                "Every loan on the portfolio — balance, rate and payment, so debt is never a surprise.")
    table_headers(ws, 4, ["Property", "Balance", "Rate %", "Payment"], start_col=2)
    start = L0
    for i, (prop, bal, rate, pay) in enumerate(LOANS):
        r = start + i
        ws.cell(row=r, column=2, value=prop).style = "td_left"
        cb = ws.cell(row=r, column=3, value=bal); cb.style = "input"; cb.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=rate / 100); cr.style = "input"; cr.number_format = "0.00%"
        cp = ws.cell(row=r, column=5, value=pay); cp.style = "input"; cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(LOANS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL DEBT").style = "th"
    cb = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cb.style = "td"; cb.font = Font(bold=True, color=DANGER); cb.fill = fill(SURFACE); cb.number_format = '"$"#,##0'
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cp = ws.cell(row=tot, column=5, value=f"=SUM(E{start}:E{end})"); cp.style = "td"; cp.font = Font(bold=True, color=DANGER); cp.fill = fill(SURFACE); cp.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


def build_reserves(wb):
    ws = wb.create_sheet("Reserves & Escrow"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 16, 14, 2])
    luxe_header(ws, "E", "🛟  RESERVES & ESCROW",
                "Your safety funds — balance vs goal, so a big repair never becomes a crisis.")
    table_headers(ws, 4, ["Fund", "Balance", "Goal", "Funded %"], start_col=2)
    start = L0
    for i, (fund, bal, goal) in enumerate(RESERVES):
        r = start + i
        ws.cell(row=r, column=2, value=fund).style = "td_left"
        cb = ws.cell(row=r, column=3, value=bal); cb.style = "input"; cb.number_format = '"$"#,##0'
        cg = ws.cell(row=r, column=4, value=goal); cg.style = "input"; cg.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(RESERVES) - 1
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=0.3, start_color="FF" + RED_BG,
                       mid_type="num", mid_value=0.6, mid_color="FFFFF3CD",
                       end_type="num", end_value=1.0, end_color="FF" + HIGHLIGHT))
    ws.freeze_panes = "A5"


def build_mileage(wb):
    ws, start, end = build_log(
        wb, "Mileage & Travel", "🚗", "MILEAGE & TRAVEL",
        "Every drive to a property — miles and purpose, deductible at tax time.",
        ["Date", "Property", "Miles", "Purpose"],
        MILEAGE, [2, 14, 18, 12, 24, 2], text_left={2, 3, 5}, ints={4}, reserved=24, start_col=2)


def build_renewals(wb):
    ws, start, end = build_log(
        wb, "Renewals & Docs", "📁", "RENEWALS & DOCS",
        "Leases, licenses & inspections with a due date — never miss a renewal again.",
        ["Item", "Unit", "Due", "Done?"],
        RENEWALS, [2, 24, 16, 16, 12, 2], text_left={2, 3}, dates=set(), reserved=24, start_col=2,
        validations=[("E", "YesNoList")])


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 16, 16, 14, 2])
    luxe_header(ws, "E", "📈  MONTHLY SUMMARY",
                "Income, expenses & cash flow by month — watch the portfolio compound.")
    table_headers(ws, 4, ["Month", "Income", "Expenses", "Cash Flow"], start_col=2)
    start = L0
    for i, (m, inc, exp) in enumerate(MONTHS):
        r = start + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        ci = ws.cell(row=r, column=3, value=inc); ci.style = "input"; ci.number_format = '"$"#,##0'
        ce = ws.cell(row=r, column=4, value=exp); ce.style = "input"; ce.number_format = '"$"#,##0'
        cp = ws.cell(row=r, column=5, value=f"=C{r}-D{r}"); cp.style = "td"; cp.font = Font(bold=True, color=PRIMARY); cp.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MONTHS) - 1
    nrange(wb, "MonthCF", "Monthly Summary", "E", start, end)
    ws.add_chart(_barchart(ws, "Cash Flow by Month", start, end, 5, 2), "G4")
    ws.freeze_panes = "A5"


# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🏘  RENTAL PROPERTY & LANDLORD COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Cash flow, cap rate, cash-on-cash, DSCR & a Landlord Score — the real return on every door.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("MONTHLY RENT", "=MonthlyRent", "money"),
        ("GROSS INCOME", "=GrossIncome", "money"),
        ("OPERATING EXP", "=OpExTotal", "money"),
        ("NOI (MONTHLY)", "=NOI", "money"),
        ("MORTGAGE P&I", "=Mortgage", "money"),
        ("CASH FLOW", "=CashFlow", "money"),
    ]
    row2 = [
        ("CAP RATE", "=CapRate", "pct"),
        ("CASH-ON-CASH", "=CoC", "pct"),
        ("DSCR", "=DSCR", "dec"),
        ("1% RULE", "=OnePct", "pct"),
        ("ANNUAL CASH FLOW", "=AnnualCF", "money"),
        ("LANDLORD SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "LANDLORD HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "CASH FLOW BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Cash flow positive", "=IFERROR(MIN(CashFlow/CFGoal,1),0)"),
        ("Cap rate on target", "=IFERROR(MIN(CapRate/CapGoal,1),0)"),
        ("Cash-on-cash healthy", "=IFERROR(MIN(CoC/CoCGoal,1),0)"),
        ("DSCR safe", "=IFERROR(MIN(DSCR/DSCRGoal,1),0)"),
        ("Expenses in check", "=IFERROR(MIN(ExpTarget/(OpExTotal/GrossIncome),1),0)"),
        ("Reserves funded", "=IFERROR(MIN(ReservePct,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Cash Flow by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=5, min_row=5, max_row=4 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=5, max_row=4 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Rental Property & Landlord Command Center™ — know the real return on every door.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_dealanalyzer(wb); build_rentroll(wb)
    build_tenants(wb); build_ledger(wb); build_explog(wb); build_maint(wb)
    build_loans(wb); build_reserves(wb); build_mileage(wb); build_renewals(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Deal Analyzer", "Rent Roll", "Tenants", "Rent Ledger",
             "Expenses Log", "Maintenance & CapEx", "Mortgage & Loan", "Reserves & Escrow",
             "Mileage & Travel", "Renewals & Docs", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Rental_Property_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
