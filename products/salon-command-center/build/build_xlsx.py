"""Build Salon, Barber & Booth Renter Command Center™ — The Chair's Operating System.

14 tabs · a premium salon/barber/booth-renter operating system in Google Sheets & Excel.
Dashboard, a true-ticket engine (price − backbar − card fee − the rent your chair costs
per hour → what you actually keep), a chair & rent break-even, a service menu costed by
the hour, a client book, appointments, retail & backbar, income & tips, expenses,
rebooking & retention, product inventory and a monthly summary — one dashboard. Price the
chair, not just the cut.

Run: python3 build_xlsx.py   ->  ../Salon_Command_Center.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
APPTSTATUS = ["Booked", "Completed", "No-show", "Late cancel", "Rescheduled"]
SVCTYPE = ["Cut", "Color", "Style", "Treatment", "Barber", "Add-on"]
EXPCAT = ["Chair", "Product", "Insurance", "Software", "Education", "Marketing", "Other"]

# --- The true-ticket engine (flagship service) ---
SERVICE_PRICE = 65.00
BACKBAR_COST = 6.50
CARD_RATE = 0.029
CARD_FIXED = 0.30
SERVICE_HOURS = 1.25
OPEN_HOURS = 160
BOOKED_HOURS = 112

# --- This month ---
CLIENTS_MONTH = 92
REBOOKED = 72
NEW_CLIENTS = 4
NO_SHOWS = 3
TIPS_MONTH = 1150

# --- Goals ---
MARGIN_GOAL = 0.60
REBOOK_GOAL = 0.70
ATTACH_GOAL = 0.10
COVER_GOAL = 3.0
NOSHOW_GOAL = 0.05
NEW_CLIENT_GOAL = 10
UTIL_GOAL = 0.80

# Fixed monthly cost of having a chair at all: (line, monthly)
FIXED_LINES = [
    ("Booth rent / chair rental", 900), ("Supplies, tools & color waste", 120),
    ("Liability insurance", 45), ("Booking & payment software", 30),
    ("Education & classes", 55),
]

# Service menu: (service, type, price, minutes, backbar cost)
SERVICES = [
    ("Women's cut & style", "Cut", 65, 75, 6.50), ("Men's cut", "Barber", 35, 30, 2.50),
    ("Root touch-up", "Color", 95, 105, 14.00), ("Full highlight", "Color", 175, 180, 28.00),
    ("Balayage", "Color", 210, 210, 32.00), ("Gloss / toner", "Color", 45, 45, 9.00),
    ("Blowout", "Style", 40, 45, 4.00), ("Deep conditioning add-on", "Add-on", 25, 20, 5.00),
    ("Beard trim", "Barber", 20, 20, 1.50), ("Kids cut", "Barber", 28, 30, 2.00),
]

# Retail: (product, your cost, retail price, units sold)
RETAIL = [
    ("Hydrating shampoo", 11.00, 22.00, 8), ("Hydrating conditioner", 11.00, 22.00, 6),
    ("Leave-in treatment", 14.00, 28.00, 5), ("Texture spray", 13.00, 26.00, 4),
    ("Argan hair oil", 15.00, 30.00, 3), ("Styling cream — travel", 7.50, 15.00, 2),
]

# Client book: (client, last visit, visits YTD, spend YTD, rebooked?)
CLIENTS = [
    ("Ava R.", "07/18", 9, 720, "Yes"), ("Marcus T.", "07/20", 12, 480, "Yes"),
    ("Priya S.", "07/12", 6, 1140, "Yes"), ("Jordan K.", "07/22", 8, 560, "Yes"),
    ("Nina W.", "06/28", 4, 810, "No"), ("Devon L.", "07/19", 11, 415, "Yes"),
    ("Sofia M.", "07/15", 5, 985, "Yes"), ("Eli B.", "07/21", 10, 350, "Yes"),
    ("Hana C.", "05/30", 3, 630, "No"), ("Ruth A.", "07/17", 7, 1290, "Yes"),
]

# Appointments: (date, client, service, price, status)
APPTS = [
    ("07/22", "Jordan K.", "Women's cut & style", 65, "Completed"),
    ("07/22", "Eli B.", "Men's cut", 35, "Completed"),
    ("07/21", "Ruth A.", "Balayage", 210, "Completed"),
    ("07/21", "Devon L.", "Beard trim", 20, "Completed"),
    ("07/20", "Marcus T.", "Men's cut", 35, "Completed"),
    ("07/20", "Nina W.", "Root touch-up", 95, "No-show"),
    ("07/19", "Sofia M.", "Full highlight", 175, "Completed"),
    ("07/18", "Ava R.", "Women's cut & style", 65, "Completed"),
    ("07/18", "Hana C.", "Blowout", 40, "Late cancel"),
    ("07/17", "Priya S.", "Gloss / toner", 45, "Completed"),
    ("07/24", "Ava R.", "Deep conditioning add-on", 25, "Booked"),
    ("07/25", "Jordan K.", "Kids cut", 28, "Booked"),
]

# Expenses: (item, category, monthly)
EXPENSES = [
    ("Booth rent", "Chair", 900), ("Color & backbar restock", "Product", 285),
    ("Tools, shears & clipper service", "Product", 45), ("Liability insurance", "Insurance", 45),
    ("Booking software", "Software", 30), ("Card reader fees", "Software", 0),
    ("Education & classes", "Education", 55), ("Instagram ads & prints", "Marketing", 40),
    ("Laundry & towels", "Other", 35),
]

# Retention trend: (month, clients, rebooked, new, no-shows)
RETENTION = [
    ("Feb", 74, 52, 6, 5), ("Mar", 79, 58, 5, 4), ("Apr", 83, 63, 7, 4),
    ("May", 86, 66, 5, 3), ("Jun", 89, 69, 6, 3), ("Jul", 92, 72, 4, 3),
]

# Product inventory: (item, on hand, reorder at, unit cost)
INVENTORY = [
    ("Hydrating shampoo", 6, 4, 11.00), ("Hydrating conditioner", 5, 4, 11.00),
    ("Leave-in treatment", 3, 4, 14.00), ("Texture spray", 7, 3, 13.00),
    ("Argan hair oil", 2, 3, 15.00), ("Styling cream — travel", 9, 4, 7.50),
    ("Backbar color — level 6", 4, 3, 9.50), ("Developer 20 vol", 8, 4, 6.00),
]

# Monthly summary: (month, revenue)
MONTHS = [("Feb", 5420), ("Mar", 5780), ("Apr", 6090), ("May", 6280), ("Jun", 6470), ("Jul", 6652)]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "pct1": "0.0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.type = "col"; ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 20, 3] + [20] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your rates & goals once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR RATES & GOALS", "section")
    controls = [
        ("Business", "Gilded Chair Studio", None, "Studio"),
        ("Stylist / owner", "Remi", None, "Owner"),
        ("Card processing rate", CARD_RATE, "0.00%", "CardRate"),
        ("Card fee per transaction", CARD_FIXED, '"$"#,##0.00', "CardFixed"),
        ("True-margin goal", MARGIN_GOAL, "0%", "MarginGoal"),
        ("Rebooking goal", REBOOK_GOAL, "0%", "RebookGoal"),
        ("Retail attach goal", ATTACH_GOAL, "0%", "AttachGoal"),
        ("Break-even cover goal (×)", COVER_GOAL, "0.0", "CoverGoal"),
        ("No-show goal (max)", NOSHOW_GOAL, "0%", "NoShowGoal"),
        ("New clients / month goal", NEW_CLIENT_GOAL, "0", "NewClientGoal"),
        ("Chair utilization goal", UTIL_GOAL, "0%", "UtilGoal"),
        ("Tips this month", TIPS_MONTH, '"$"#,##0', "Tips"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Appointment status", APPTSTATUS, "ApptStatusList"), ("F", "Service type", SVCTYPE, "SvcTypeList"),
             ("G", "Expense category", EXPCAT, "ExpCatList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ✂  SALON, BARBER & BOOTH RENTER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Price the chair, not just the cut.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE CHAIR, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("A $65 service is not $65. Backbar product comes out, the card processor takes a cut, and — the "
                      "part almost nobody prices in — your booth rent is running whether that chair is full or empty, "
                      "so every service owes rent for the time it takes. This workbook does that math: the true-ticket "
                      "engine strips out product, fees and the rent your chair costs per hour to show what you ACTUALLY "
                      "keep. Then run your service menu by the hour, your client book, retail, tips, expenses, "
                      "rebooking and inventory — all in ONE premium Google Sheets & Excel system.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — set your card rate and your goals.",
             "2.  Chair & Rent: list what the chair costs you every month.",
             "3.  Service Pricing: enter a price, product cost and how long it takes.",
             "4.  Read what you actually keep — and your break-even client count.",
             "5.  Cost your whole menu by the hour. Some services aren't worth the chair.",
             "6.  Check the Dashboard: profit, rebooking & a Chair Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional studio (Gilded Chair Studio, stylist Remi) is included so you can see how "
               "it all connects — just type over it with your own prices and clients. Works whether you rent a booth, "
               "rent a suite, own the salon or cut at home: the chair costs what it costs, and this splits it across "
               "the hours you're open. Twelve matching printable pages (service pricing sheet, client card, day sheet, "
               "rebooking tracker & more) are included. This is a business & organizing tool, not financial, tax or "
               "accounting advice — and card processing rates vary by processor.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "An empty chair still charges rent. Price every service like it knows that.", "section_gold")


# ===========================================================================
def build_chair(wb):
    ws = wb.create_sheet("Chair & Rent"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 38, 18, 2])
    luxe_header(ws, "C", "🪑  CHAIR & RENT",
                "What the chair costs you every month — and how many clients it takes just to cover it.")
    ws.cell(row=5, column=2, value="FIXED COST OF HAVING A CHAIR").style = "section_gold"
    table_headers(ws, 6, ["Line", "Monthly"], start_col=2)
    fs = 7
    for i, (lab, amt) in enumerate(FIXED_LINES):
        r = fs + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            for cc in range(2, 4):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    fe = fs + len(FIXED_LINES) - 1
    nrange(wb, "FixedLines", "Chair & Rent", "C", fs, fe)
    tot = fe + 1
    ws.cell(row=tot, column=2, value="= FIXED COSTS / MONTH").style = "th"
    ct = ws.cell(row=tot, column=3, value="=SUM(FixedLines)"); ct.style = "td"
    ct.font = Font(bold=True, size=13, color=PRIMARY); ct.fill = fill(SURFACE); ct.number_format = '"$"#,##0'
    cell_name(wb, "FixedCosts", "Chair & Rent", f"$C${tot}")

    r = tot + 2
    ws.cell(row=r, column=2, value="THE CHAIR CHARGES BY THE HOUR").style = "section_gold"
    ws.cell(row=r + 1, column=2, value="Hours the chair is open / month").style = "field_label"
    co = ws.cell(row=r + 1, column=3, value=OPEN_HOURS); co.style = "input"; co.number_format = "#,##0"
    cell_name(wb, "OpenHours", "Chair & Rent", f"$C${r+1}")
    ws.cell(row=r + 2, column=2, value="Hours actually booked / month").style = "field_label"
    cbh = ws.cell(row=r + 2, column=3, value=BOOKED_HOURS); cbh.style = "input"; cbh.number_format = "#,##0"
    cell_name(wb, "BookedHours", "Chair & Rent", f"$C${r+2}")
    ws.cell(row=r + 3, column=2, value="= CHAIR UTILIZATION").style = "th"
    cu = ws.cell(row=r + 3, column=3, value="=IFERROR(BookedHours/OpenHours,0)"); cu.style = "td"
    cu.font = Font(bold=True, size=12, color=PRIMARY); cu.fill = fill(MINT_BG); cu.number_format = "0%"
    cell_name(wb, "Utilization", "Chair & Rent", f"$C${r+3}")
    ws.cell(row=r + 4, column=2, value="= RENT PER CHAIR-HOUR").style = "th"
    cr = ws.cell(row=r + 4, column=3, value="=IFERROR(FixedCosts/OpenHours,0)"); cr.style = "td"
    cr.font = Font(bold=True, size=13, color=PRIMARY); cr.fill = fill(MINT_BG); cr.number_format = '"$"#,##0.00'
    cell_name(wb, "RentPerHour", "Chair & Rent", f"$C${r+4}")

    b = r + 6
    ws.cell(row=b, column=2, value="BREAK-EVEN — JUST TO COVER THE CHAIR").style = "section_gold"
    ws.cell(row=b + 1, column=2, value="Clients / month to break even").style = "th"
    cbe = ws.cell(row=b + 1, column=3, value="=IFERROR(ROUNDUP(FixedCosts/ServiceNet,0),0)"); cbe.style = "td"
    cbe.font = Font(bold=True, size=15, color=PRIMARY); cbe.fill = fill(SURFACE); cbe.number_format = "#,##0"
    cell_name(wb, "BreakEven", "Chair & Rent", f"$C${b+1}")
    ws.cell(row=b + 2, column=2, value="…which is this many per week").style = "field_label"
    cbw = ws.cell(row=b + 2, column=3, value="=IFERROR(ROUNDUP(BreakEven/4.33,0),0)"); cbw.style = "field_value"; cbw.number_format = "#,##0"
    ws.cell(row=b + 3, column=2, value="You served this many").style = "field_label"
    ccm = ws.cell(row=b + 3, column=3, value="=ClientsMonth"); ccm.style = "field_value"; ccm.number_format = "#,##0"
    ws.cell(row=b + 4, column=2, value="= YOU COVERED THE CHAIR THIS MANY TIMES OVER").style = "th"
    ccv = ws.cell(row=b + 4, column=3, value="=IFERROR(ClientsMonth/BreakEven,0)"); ccv.style = "td"
    ccv.font = Font(bold=True, size=13, color=PRIMARY); ccv.fill = fill(MINT_BG); ccv.number_format = '0.0"×"'
    cell_name(wb, "CoverRatio", "Chair & Rent", f"$C${b+4}")
    ws.cell(row=b + 6, column=2, value="Every unbooked hour still owes rent. That is the whole game.").style = "section_gold"


def build_pricing(wb):
    ws = wb.create_sheet("Service Pricing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 40, 18, 2])
    luxe_header(ws, "C", "💇  SERVICE PRICING — THE TRUE-TICKET ENGINE",
                "Price − product − card fee − the rent your chair costs for that hour = what you actually keep.")
    ws.cell(row=5, column=2, value="THE TICKET").style = "section_gold"
    ws.cell(row=6, column=2, value="Service price (what the client pays)").style = "field_label"
    cp = ws.cell(row=6, column=3, value=SERVICE_PRICE); cp.style = "input"; cp.number_format = '"$"#,##0.00'
    cell_name(wb, "ServicePrice", "Service Pricing", "$C$6")
    ws.cell(row=7, column=2, value="− Backbar / product used").style = "field_label"
    cb = ws.cell(row=7, column=3, value=BACKBAR_COST); cb.style = "input"; cb.number_format = '"$"#,##0.00'
    cell_name(wb, "BackbarCost", "Service Pricing", "$C$7")
    ws.cell(row=8, column=2, value="− Card processing fee").style = "field_label"
    cc = ws.cell(row=8, column=3, value="=ServicePrice*CardRate+CardFixed"); cc.style = "field_value"; cc.number_format = '"$"#,##0.00'
    cell_name(wb, "CardFee", "Service Pricing", "$C$8")
    ws.cell(row=9, column=2, value="= SERVICE NET (before rent)").style = "th"
    cn = ws.cell(row=9, column=3, value="=ServicePrice-BackbarCost-CardFee"); cn.style = "td"
    cn.font = Font(bold=True, size=12, color=PRIMARY); cn.fill = fill(SURFACE); cn.number_format = '"$"#,##0.00'
    cell_name(wb, "ServiceNet", "Service Pricing", "$C$9")

    ws.cell(row=11, column=2, value="NOW THE PART EVERYONE SKIPS — THE CHAIR").style = "section_gold"
    ws.cell(row=12, column=2, value="Rent per chair-hour (from Chair & Rent)").style = "field_label"
    cr = ws.cell(row=12, column=3, value="=RentPerHour"); cr.style = "field_value"; cr.number_format = '"$"#,##0.00'
    ws.cell(row=13, column=2, value="× Hours this service takes").style = "field_label"
    ch = ws.cell(row=13, column=3, value=SERVICE_HOURS); ch.style = "input"; ch.number_format = "0.00"
    cell_name(wb, "ServiceHours", "Service Pricing", "$C$13")
    ws.cell(row=14, column=2, value="= Rent this service owes").style = "th"
    cl = ws.cell(row=14, column=3, value="=RentPerHour*ServiceHours"); cl.style = "td"
    cl.font = Font(bold=True, size=12, color=PRIMARY); cl.fill = fill(WARN_BG); cl.number_format = '"$"#,##0.00'
    cell_name(wb, "RentLoad", "Service Pricing", "$C$14")

    ws.cell(row=16, column=2, value="= YOU ACTUALLY KEEP").style = "th"
    ct = ws.cell(row=16, column=3, value="=ServiceNet-RentLoad"); ct.style = "td"
    ct.font = Font(bold=True, size=16, color=PRIMARY); ct.fill = fill(MINT_BG); ct.number_format = '"$"#,##0.00'
    cell_name(wb, "TrueNet", "Service Pricing", "$C$16")
    ws.cell(row=17, column=2, value="= TRUE MARGIN").style = "th"
    cm = ws.cell(row=17, column=3, value="=IFERROR(TrueNet/ServicePrice,0)"); cm.style = "td"
    cm.font = Font(bold=True, size=13, color=PRIMARY); cm.fill = fill(MINT_BG); cm.number_format = "0.0%"
    cell_name(wb, "TrueMargin", "Service Pricing", "$C$17")
    ws.cell(row=18, column=2, value="= YOUR REAL RATE PER HOUR").style = "th"
    crh = ws.cell(row=18, column=3, value="=IFERROR(TrueNet/ServiceHours,0)"); crh.style = "td"
    crh.font = Font(bold=True, size=13, color=PRIMARY); crh.fill = fill(MINT_BG); crh.number_format = '"$"#,##0.00'
    cell_name(wb, "TrueHourly", "Service Pricing", "$C$18")

    ws.cell(row=20, column=2, value="⚠ THE TICKET SAYS ONE THING. THE BANK SAYS ANOTHER.").style = "section_gold"
    ws.cell(row=21, column=2, value="The ticket says").style = "field_label"
    c1 = ws.cell(row=21, column=3, value="=ServicePrice"); c1.style = "field_value"; c1.number_format = '"$"#,##0.00'
    ws.cell(row=22, column=2, value="Product, card fee & the chair took").style = "field_label"
    c2 = ws.cell(row=22, column=3, value="=ServicePrice-TrueNet"); c2.style = "field_value"
    c2.number_format = '"$"#,##0.00'; c2.fill = fill(RED_BG)
    ws.cell(row=23, column=2, value="…which is this share of every ticket").style = "field_label"
    c3 = ws.cell(row=23, column=3, value="=IFERROR((ServicePrice-TrueNet)/ServicePrice,0)"); c3.style = "field_value"
    c3.number_format = "0.0%"; c3.fill = fill(RED_BG)
    ws.cell(row=24, column=2, value="Raise this service by $5 and you'd keep").style = "field_label"
    c4 = ws.cell(row=24, column=3, value="=TrueNet+5-(5*CardRate)"); c4.style = "field_value"
    c4.number_format = '"$"#,##0.00'; c4.fill = fill(MINT_BG)
    ws.cell(row=26, column=2, value="A $5 raise is almost pure profit. Discounting $5 is almost pure loss.").style = "section_gold"


def build_menu(wb):
    ws = wb.create_sheet("Services Menu"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 13, 13, 11, 14, 14, 14, 2])
    luxe_header(ws, "H", "📋  SERVICES MENU — COSTED BY THE HOUR",
                "Every service on your menu, net of product, fees and chair rent — and what it really pays per hour.")
    table_headers(ws, 4, ["Service", "Type", "Price", "Minutes", "Backbar", "You keep", "Per hour"], start_col=2)
    start = L0
    for i, (svc, typ, price, mins, back) in enumerate(SERVICES):
        r = start + i
        ws.cell(row=r, column=2, value=svc).style = "td_left"
        ws.cell(row=r, column=3, value=typ).style = "td"
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cm = ws.cell(row=r, column=5, value=mins); cm.style = "input"; cm.number_format = "#,##0"
        cbk = ws.cell(row=r, column=6, value=back); cbk.style = "input"; cbk.number_format = '"$"#,##0.00'
        ck = ws.cell(row=r, column=7, value=f"=D{r}-F{r}-(D{r}*CardRate+CardFixed)-RentPerHour*(E{r}/60)")
        ck.style = "td"; ck.number_format = '"$"#,##0.00'
        chr_ = ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/(E{r}/60),0)"); chr_.style = "td"; chr_.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 9):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(SERVICES) - 1
    nrange(wb, "MenuPrice", "Services Menu", "D", start, end)
    nrange(wb, "MenuKeep", "Services Menu", "G", start, end)
    nrange(wb, "MenuHourly", "Services Menu", "H", start, end)
    add_dv(ws, f"C{start}:C{end}", "SvcTypeList")
    ws.conditional_formatting.add(f"H{start}:H{end}", DataBarRule(start_type="min", end_type="max", color=HIGHLIGHT))
    ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="lessThan", formula=["0"],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tot = end + 2
    ws.cell(row=tot, column=2, value="Best per-hour service on the menu").style = "field_label"
    cbst = ws.cell(row=tot, column=8, value="=IFERROR(MAX(MenuHourly),0)"); cbst.style = "field_value"
    cbst.number_format = '"$"#,##0.00'; cbst.fill = fill(MINT_BG)
    ws.cell(row=tot + 1, column=2, value="Worst per-hour service on the menu").style = "field_label"
    cwst = ws.cell(row=tot + 1, column=8, value="=IFERROR(MIN(MenuHourly),0)"); cwst.style = "field_value"
    cwst.number_format = '"$"#,##0.00'; cwst.fill = fill(WARN_BG)
    ws.cell(row=tot + 3, column=2, value="A cheap service that eats an hour costs you more than an empty chair.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_clients(wb):
    ws, start, end = build_log(
        wb, "Client Book", "📖", "CLIENT BOOK", "Who they are, when they were last in, what they're worth — and whether they rebooked.",
        ["Client", "Last visit", "Visits YTD", "Spend YTD", "Rebooked?", "Avg ticket"],
        [(n, lv, v, s, rb) for (n, lv, v, s, rb) in CLIENTS],
        [2, 26, 14, 13, 14, 13, 14, 2],
        text_left={2}, ints={4}, money={5}, money2={7}, validations=[("F", "YesNoList")], start_col=2)
    for r in range(start, start + len(CLIENTS)):
        c = ws.cell(row=r, column=7, value=f"=IFERROR(E{r}/D{r},0)"); c.number_format = '"$"#,##0.00'
    nrange(wb, "ClientName", "Client Book", "B", start, end)
    nrange(wb, "ClientSpend", "Client Book", "E", start, end)
    nrange(wb, "ClientRebook", "Client Book", "F", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="Clients on the books").style = "field_label"
    c1 = ws.cell(row=tr, column=5, value="=COUNTA(ClientName)"); c1.style = "field_value"; c1.number_format = "#,##0"
    ws.cell(row=tr + 1, column=2, value="Total client value YTD").style = "field_label"
    c2 = ws.cell(row=tr + 1, column=5, value="=SUM(ClientSpend)"); c2.style = "field_value"
    c2.number_format = '"$"#,##0'; c2.fill = fill(MINT_BG)
    ws.cell(row=tr + 2, column=2, value="Average client value YTD").style = "field_label"
    c3 = ws.cell(row=tr + 2, column=5, value="=IFERROR(SUM(ClientSpend)/COUNTA(ClientName),0)"); c3.style = "field_value"
    c3.number_format = '"$"#,##0'


def build_appts(wb):
    ws, start, end = build_log(
        wb, "Appointments", "📅", "APPOINTMENTS", "Every booking — kept, cancelled or no-showed. The gaps are the real cost.",
        ["Date", "Client", "Service", "Price", "Status"], APPTS,
        [2, 12, 24, 28, 13, 16, 2], text_left={3, 4}, money2={5}, validations=[("F", "ApptStatusList")], start_col=2)
    nrange(wb, "ApptPrice", "Appointments", "E", start, end)
    nrange(wb, "ApptStatus", "Appointments", "F", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="Completed").style = "field_label"
    c1 = ws.cell(row=tr, column=5, value='=COUNTIF(ApptStatus,"Completed")'); c1.style = "field_value"; c1.number_format = "#,##0"
    ws.cell(row=tr + 1, column=2, value="No-shows & late cancels").style = "field_label"
    c2 = ws.cell(row=tr + 1, column=5, value='=COUNTIF(ApptStatus,"No-show")+COUNTIF(ApptStatus,"Late cancel")')
    c2.style = "field_value"; c2.number_format = "#,##0"; c2.fill = fill(RED_BG)
    ws.cell(row=tr + 2, column=2, value="Money that walked out the door").style = "field_label"
    c3 = ws.cell(row=tr + 2, column=5, value='=SUMIF(ApptStatus,"No-show",ApptPrice)+SUMIF(ApptStatus,"Late cancel",ApptPrice)')
    c3.style = "field_value"; c3.number_format = '"$"#,##0.00'; c3.fill = fill(RED_BG)
    ws.cell(row=tr + 4, column=2, value="A no-show is not a free hour. That hour already paid rent.").style = "section_gold"


def build_retail(wb):
    ws = wb.create_sheet("Retail & Backbar"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 14, 14, 12, 14, 14, 2])
    luxe_header(ws, "G", "🧴  RETAIL & BACKBAR",
                "Retail is the highest-margin thing in the room and it takes no chair time at all.")
    table_headers(ws, 4, ["Product", "Your cost", "Retail price", "Units sold", "Revenue", "Profit"], start_col=2)
    start = L0
    for i, (p, cost, price, units) in enumerate(RETAIL):
        r = start + i
        ws.cell(row=r, column=2, value=p).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cu = ws.cell(row=r, column=5, value=units); cu.style = "input"; cu.number_format = "#,##0"
        cr = ws.cell(row=r, column=6, value=f"=D{r}*E{r}"); cr.style = "td"; cr.number_format = '"$"#,##0.00'
        cpf = ws.cell(row=r, column=7, value=f"=(D{r}-C{r})*E{r}"); cpf.style = "td"; cpf.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(RETAIL) - 1
    nrange(wb, "RetailCost", "Retail & Backbar", "C", start, end)
    nrange(wb, "RetailPrice", "Retail & Backbar", "D", start, end)
    nrange(wb, "RetailUnitsCol", "Retail & Backbar", "E", start, end)
    nrange(wb, "RetailRev", "Retail & Backbar", "F", start, end)
    nrange(wb, "RetailProfitCol", "Retail & Backbar", "G", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="ALL RETAIL").style = "th"
    for c in (3, 4):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cu = ws.cell(row=tot, column=5, value="=SUM(RetailUnitsCol)"); cu.style = "td"
    cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "#,##0"
    cell_name(wb, "RetailUnits", "Retail & Backbar", f"$E${tot}")
    cr = ws.cell(row=tot, column=6, value="=SUM(RetailRev)"); cr.style = "td"
    cr.font = Font(bold=True, color=PRIMARY); cr.fill = fill(SURFACE); cr.number_format = '"$"#,##0'
    cell_name(wb, "RetailRevenue", "Retail & Backbar", f"$F${tot}")
    cpf = ws.cell(row=tot, column=7, value="=SUM(RetailProfitCol)"); cpf.style = "td"
    cpf.font = Font(bold=True, size=12, color=PRIMARY); cpf.fill = fill(MINT_BG); cpf.number_format = '"$"#,##0'
    cell_name(wb, "RetailProfit", "Retail & Backbar", f"$G${tot}")
    ws.cell(row=tot + 1, column=2, value="Retail cost of goods").style = "field_label"
    cg = ws.cell(row=tot + 1, column=6, value="=SUMPRODUCT(RetailCost,RetailUnitsCol)"); cg.style = "field_value"; cg.number_format = '"$"#,##0'
    cell_name(wb, "RetailCOGS", "Retail & Backbar", f"$F${tot+1}")
    ws.cell(row=tot + 3, column=2, value="RETAIL ATTACH RATE").style = "section_gold"
    ws.cell(row=tot + 4, column=2, value="Retail revenue ÷ service revenue").style = "th"
    ca = ws.cell(row=tot + 4, column=6, value="=IFERROR(RetailRevenue/ServiceRevenue,0)"); ca.style = "td"
    ca.font = Font(bold=True, size=13, color=PRIMARY); ca.fill = fill(MINT_BG); ca.number_format = "0.0%"
    cell_name(wb, "AttachRate", "Retail & Backbar", f"$F${tot+4}")
    ws.cell(row=tot + 5, column=2, value="Goal").style = "field_label"
    cgo = ws.cell(row=tot + 5, column=6, value="=AttachGoal"); cgo.style = "field_value"; cgo.number_format = "0%"
    ws.cell(row=tot + 7, column=2, value="Every bottle you sell is profit your chair didn't have to sit for.").style = "section_gold"
    ws.freeze_panes = "A5"


def build_income(wb):
    ws = wb.create_sheet("Income & Tips"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 40, 18, 2])
    luxe_header(ws, "C", "💰  INCOME & TIPS",
                "Services plus retail, minus what it cost to earn them — and what actually lands in your account.")
    ws.cell(row=5, column=2, value="THE MONTH").style = "section_gold"
    ws.cell(row=6, column=2, value="Clients served").style = "field_label"
    cc = ws.cell(row=6, column=3, value="=ClientsMonth"); cc.style = "field_value"; cc.number_format = "#,##0"
    ws.cell(row=7, column=2, value="Service revenue (clients × avg ticket)").style = "field_label"
    cs = ws.cell(row=7, column=3, value="=ClientsMonth*ServicePrice"); cs.style = "field_value"; cs.number_format = '"$"#,##0'
    cell_name(wb, "ServiceRevenue", "Income & Tips", "$C$7")
    ws.cell(row=8, column=2, value="+ Retail revenue").style = "field_label"
    cr = ws.cell(row=8, column=3, value="=RetailRevenue"); cr.style = "field_value"; cr.number_format = '"$"#,##0'
    ws.cell(row=9, column=2, value="= MONTHLY REVENUE").style = "th"
    cm = ws.cell(row=9, column=3, value="=ServiceRevenue+RetailRevenue"); cm.style = "td"
    cm.font = Font(bold=True, size=14, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyRevenue", "Income & Tips", "$C$9")

    ws.cell(row=11, column=2, value="WHAT IT COST TO EARN IT").style = "section_gold"
    ws.cell(row=12, column=2, value="− Backbar product (clients × cost)").style = "field_label"
    cb = ws.cell(row=12, column=3, value="=ClientsMonth*BackbarCost"); cb.style = "field_value"; cb.number_format = '"$"#,##0.00'
    cell_name(wb, "BackbarTotal", "Income & Tips", "$C$12")
    ws.cell(row=13, column=2, value="− Card processing (rate + per-swipe)").style = "field_label"
    cf = ws.cell(row=13, column=3, value="=MonthlyRevenue*CardRate+CardFixed*(ClientsMonth+RetailUnits)")
    cf.style = "field_value"; cf.number_format = '"$"#,##0.00'
    cell_name(wb, "CardTotal", "Income & Tips", "$C$13")
    ws.cell(row=14, column=2, value="− Retail cost of goods").style = "field_label"
    cg = ws.cell(row=14, column=3, value="=RetailCOGS"); cg.style = "field_value"; cg.number_format = '"$"#,##0.00'
    ws.cell(row=15, column=2, value="= VARIABLE COSTS").style = "th"
    cv = ws.cell(row=15, column=3, value="=BackbarTotal+CardTotal+RetailCOGS"); cv.style = "td"
    cv.font = Font(bold=True, size=12, color=PRIMARY); cv.fill = fill(WARN_BG); cv.number_format = '"$"#,##0.00'
    cell_name(wb, "VariableCosts", "Income & Tips", "$C$15")
    ws.cell(row=16, column=2, value="− Fixed costs (the chair)").style = "field_label"
    cx = ws.cell(row=16, column=3, value="=FixedCosts"); cx.style = "field_value"; cx.number_format = '"$"#,##0'

    ws.cell(row=18, column=2, value="= MONTHLY PROFIT").style = "th"
    cpf = ws.cell(row=18, column=3, value="=MonthlyRevenue-VariableCosts-FixedCosts"); cpf.style = "td"
    cpf.font = Font(bold=True, size=16, color=PRIMARY); cpf.fill = fill(MINT_BG); cpf.number_format = '"$"#,##0'
    cell_name(wb, "MonthlyProfit", "Income & Tips", "$C$18")
    ws.cell(row=19, column=2, value="= NET MARGIN").style = "th"
    cnm = ws.cell(row=19, column=3, value="=IFERROR(MonthlyProfit/MonthlyRevenue,0)"); cnm.style = "td"
    cnm.font = Font(bold=True, size=13, color=PRIMARY); cnm.fill = fill(MINT_BG); cnm.number_format = "0.0%"
    cell_name(wb, "NetMargin", "Income & Tips", "$C$19")
    ws.cell(row=21, column=2, value="TIPS (yours — track them, they're income)").style = "section_gold"
    ws.cell(row=22, column=2, value="Tips this month").style = "field_label"
    ct = ws.cell(row=22, column=3, value="=Tips"); ct.style = "field_value"; ct.number_format = '"$"#,##0'
    ws.cell(row=23, column=2, value="= TOTAL TAKE-HOME").style = "th"
    cth = ws.cell(row=23, column=3, value="=MonthlyProfit+Tips"); cth.style = "td"
    cth.font = Font(bold=True, size=14, color=PRIMARY); cth.fill = fill(MINT_BG); cth.number_format = '"$"#,##0'
    cell_name(wb, "TakeHome", "Income & Tips", "$C$23")
    ws.cell(row=24, column=2, value="Tips are taxable income. Set some aside — see the printable tax page.").style = "section_gold"


def build_expenses(wb):
    ws, start, end = build_log(
        wb, "Expenses", "🧾", "EXPENSES", "Everything the chair costs you — sorted, so nothing quietly grows.",
        ["Item", "Category", "Monthly", "Yearly"], [(i, c, m) for (i, c, m) in EXPENSES],
        [2, 34, 18, 14, 14, 2], text_left={2}, money={4, 5}, validations=[("C", "ExpCatList")], start_col=2)
    for r in range(start, start + len(EXPENSES)):
        ws.cell(row=r, column=5, value=f"=D{r}*12").number_format = '"$"#,##0'
    nrange(wb, "ExpMonthly", "Expenses", "D", start, end)
    tr = end + 2
    ws.cell(row=tr, column=2, value="TOTAL MONTHLY").style = "th"
    c1 = ws.cell(row=tr, column=4, value="=SUM(ExpMonthly)"); c1.style = "td"
    c1.font = Font(bold=True, size=12, color=PRIMARY); c1.fill = fill(SURFACE); c1.number_format = '"$"#,##0'
    cell_name(wb, "ExpTotal", "Expenses", f"$D${tr}")
    c2 = ws.cell(row=tr, column=5, value="=ExpTotal*12"); c2.style = "td"
    c2.font = Font(bold=True, size=12, color=PRIMARY); c2.fill = fill(MINT_BG); c2.number_format = '"$"#,##0'
    ws.cell(row=tr + 2, column=2, value="Everything on this list runs whether the chair is full or empty.").style = "section_gold"


def build_retention(wb):
    ws = wb.create_sheet("Rebooking & Retention"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 3, 12, 12, 12, 12, 12, 2])
    luxe_header(ws, "I", "🔁  REBOOKING & RETENTION",
                "A client who rebooks before they leave is worth more than a client you have to chase.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    rows = [
        ("Clients served", CLIENTS_MONTH, "#,##0", "ClientsMonth", "input"),
        ("Rebooked before they left", REBOOKED, "#,##0", "Rebooked", "input"),
        ("= REBOOKING RATE", "=IFERROR(Rebooked/ClientsMonth,0)", "0%", "RebookRate", "big"),
        ("New clients this month", NEW_CLIENTS, "#,##0", "NewClients", "input"),
        ("No-shows & late cancels", NO_SHOWS, "#,##0", "NoShows", "input"),
        ("= NO-SHOW RATE", "=IFERROR(NoShows/ClientsMonth,0)", "0.0%", "NoShowRate", "big"),
    ]
    for i, (lab, val, fmt, nm, kind) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "th" if kind == "big" else "field_label"
        c = ws.cell(row=r, column=3, value=val)
        if kind == "big":
            c.style = "td"; c.font = Font(bold=True, size=13, color=PRIMARY); c.fill = fill(MINT_BG)
        else:
            c.style = "input"
        c.number_format = fmt
        cell_name(wb, nm, "Rebooking & Retention", f"$C${r}")

    merge_set(ws, "E5:I5", "THE TREND", "section_gold")
    table_headers(ws, 6, ["Month", "Clients", "Rebooked", "New", "No-shows"], start_col=5)
    ts = 7
    for i, (m, cl, rb, nw, ns) in enumerate(RETENTION):
        r = ts + i
        ws.cell(row=r, column=5, value=m).style = "td_left"
        for ci, v in enumerate((cl, rb, nw, ns), 6):
            c = ws.cell(row=r, column=ci, value=v); c.style = "input"; c.number_format = "#,##0"
        if i % 2:
            for c in range(5, 10):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(RETENTION) - 1
    nrange(wb, "RetClients", "Rebooking & Retention", "F", ts, te)
    ws.add_chart(_barchart(ws, "Clients by Month", ts, te, 6, 5), "E15")
    ws.cell(row=14, column=2, value="Rebooking is the cheapest marketing there is. Ask before they stand up.").style = "section_gold"


def build_inventory(wb):
    ws, start, end = build_log(
        wb, "Inventory", "📦", "PRODUCT INVENTORY", "Retail and backbar stock — so you never have to say \"I'm out of that.\"",
        ["Item", "On hand", "Reorder at", "Unit cost", "Value", "Reorder?"],
        [(i, oh, ra, uc) for (i, oh, ra, uc) in INVENTORY],
        [2, 30, 13, 13, 13, 14, 14, 2], text_left={2}, ints={3, 4}, money2={5, 6}, start_col=2)
    for r in range(start, start + len(INVENTORY)):
        ws.cell(row=r, column=6, value=f"=C{r}*E{r}").number_format = '"$"#,##0.00'
        ws.cell(row=r, column=7, value=f'=IF(C{r}<=D{r},"REORDER","OK")')
    nrange(wb, "InvValue", "Inventory", "F", start, end)
    nrange(wb, "InvFlag", "Inventory", "G", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}", CellIsRule(operator="equal", formula=['"REORDER"'],
                                                                fill=fill(RED_BG), font=Font(bold=True, color=DANGER)))
    tr = end + 2
    ws.cell(row=tr, column=2, value="Stock on the shelf").style = "field_label"
    c1 = ws.cell(row=tr, column=6, value="=SUM(InvValue)"); c1.style = "field_value"
    c1.number_format = '"$"#,##0'; c1.fill = fill(MINT_BG)
    ws.cell(row=tr + 1, column=2, value="Items to reorder").style = "field_label"
    c2 = ws.cell(row=tr + 1, column=6, value='=COUNTIF(InvFlag,"REORDER")'); c2.style = "field_value"; c2.number_format = "#,##0"


def build_summary(wb):
    ws = wb.create_sheet("Monthly Summary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 32, 18, 2])
    luxe_header(ws, "C", "📈  MONTHLY SUMMARY",
                "Revenue month by month — and whether the chair is actually growing.")
    ws.cell(row=5, column=2, value="THIS MONTH").style = "section_gold"
    ws.cell(row=6, column=2, value="Revenue").style = "field_label"
    c1 = ws.cell(row=6, column=3, value="=MonthlyRevenue"); c1.style = "field_value"; c1.number_format = '"$"#,##0'
    ws.cell(row=7, column=2, value="Profit").style = "field_label"
    c2 = ws.cell(row=7, column=3, value="=MonthlyProfit"); c2.style = "field_value"; c2.number_format = '"$"#,##0'
    ws.cell(row=8, column=2, value="+ Tips").style = "field_label"
    c3 = ws.cell(row=8, column=3, value="=Tips"); c3.style = "field_value"; c3.number_format = '"$"#,##0'
    ws.cell(row=9, column=2, value="= TAKE-HOME").style = "th"
    c4 = ws.cell(row=9, column=3, value="=TakeHome"); c4.style = "td"
    c4.font = Font(bold=True, size=14, color=PRIMARY); c4.fill = fill(MINT_BG); c4.number_format = '"$"#,##0'
    ws.cell(row=10, column=2, value="= RUN-RATE YEAR").style = "th"
    c5 = ws.cell(row=10, column=3, value="=TakeHome*12"); c5.style = "td"
    c5.font = Font(bold=True, size=13, color=PRIMARY); c5.fill = fill(SURFACE); c5.number_format = '"$"#,##0'
    cell_name(wb, "RunRate", "Monthly Summary", "$C$10")
    ws.cell(row=12, column=2, value="THE TREND").style = "section_gold"
    table_headers(ws, 13, ["Month", "Revenue"], start_col=2)
    ts = 14
    for i, (m, rev) in enumerate(MONTHS):
        r = ts + i
        ws.cell(row=r, column=2, value=m).style = "td_left"
        crv = ws.cell(row=r, column=3, value=rev); crv.style = "input"; crv.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 4):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    te = ts + len(MONTHS) - 1
    nrange(wb, "RevTrend", "Monthly Summary", "C", ts, te)
    ws.add_chart(_barchart(ws, "Revenue by Month", ts, te, 3, 2), "E5")


def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ✂  SALON, BARBER & BOOTH RENTER COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  What you actually keep, what the chair costs & a Chair Score — at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("TICKET PRICE", "=ServicePrice", "money2"),
        ("BACKBAR / PRODUCT", "=BackbarCost", "money2"),
        ("CARD FEE", "=CardFee", "money2"),
        ("RENT / CHAIR-HOUR", "=RentPerHour", "money2"),
        ("YOU ACTUALLY KEEP", "=TrueNet", "money2"),
        ("TRUE MARGIN", "=TrueMargin", "pct1"),
    ]
    row2 = [
        ("CLIENTS / MONTH", "=ClientsMonth", "num"),
        ("MONTHLY REVENUE", "=MonthlyRevenue", "money"),
        ("MONTHLY PROFIT", "=MonthlyProfit", "money"),
        ("BREAK-EVEN CLIENTS", "=BreakEven", "num"),
        ("CHAIR UTILIZATION", "=Utilization", "pct"),
        ("CHAIR SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "CHAIR HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "REVENUE BY MONTH", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("True margin healthy", "=IFERROR(MIN(TrueMargin/MarginGoal,1),0)"),
        ("Clients rebooking", "=IFERROR(MIN(RebookRate/RebookGoal,1),0)"),
        ("Retail attaching", "=IFERROR(MIN(AttachRate/AttachGoal,1),0)"),
        ("Chair more than covered", "=IFERROR(MIN(CoverRatio/CoverGoal,1),0)"),
        ("No-shows under control", "=IF(NoShowRate<=NoShowGoal,1,IFERROR(NoShowGoal/NoShowRate,0))"),
        ("New clients coming in", "=IFERROR(MIN(NewClients/NewClientGoal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Watch"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ms = wb["Monthly Summary"]
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue by Month"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ms, min_col=3, min_row=14, max_row=13 + len(MONTHS)), titles_from_data=False)
    ch.set_categories(Reference(ms, min_col=2, min_row=14, max_row=13 + len(MONTHS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Salon, Barber & Booth Renter Command Center™ — price the chair, not just the cut.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_chair(wb); build_pricing(wb)
    build_menu(wb); build_clients(wb); build_appts(wb); build_retail(wb)
    build_income(wb); build_expenses(wb); build_retention(wb); build_inventory(wb)
    build_summary(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Service Pricing", "Chair & Rent", "Services Menu", "Client Book",
             "Appointments", "Retail & Backbar", "Income & Tips", "Expenses", "Rebooking & Retention",
             "Inventory", "Monthly Summary", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Salon_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
