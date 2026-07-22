"""Build Food Truck Command Center™ — The Complete Mobile-Food Business System.

14 tabs · a premium food-truck operating system in Google Sheets & Excel.
Dashboard, a menu & cost sheet, an event P&L engine, a break-even calculator, a
daily sales log, commissary & overhead, inventory & par, a fuel & mileage log, a
permits tracker, a supplies list, a bookings calendar and cash & tips — one
dashboard. Know your profit per event, cover your overhead & book the right gigs.

Run: python3 build_xlsx.py   ->  ../Food_Truck_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
UNITS = ["each", "lb", "oz", "case", "can", "bag", "gal"]
EVTYPES = ["Lunch", "Brewery", "Market", "Office", "Festival", "Catering", "Private", "Other"]
PERMIT_STATUS = ["OK", "Renew", "Expired"]

TARGET_FC = 0.30
EVENT_GOAL = 10

# Menu & cost: (item, cost, price)
MENU = [
    ("Pulled Pork Sandwich", 2.80, 12.00),
    ("Brisket Plate", 5.50, 18.00),
    ("Loaded Fries", 2.10, 9.00),
    ("Mac & Cheese", 1.40, 6.00),
    ("Smoked Wings", 3.20, 11.00),
    ("Cornbread", 0.70, 4.00),
    ("Craft Soda", 0.60, 3.00),
    ("Combo Plate", 6.20, 20.00),
]

# Events: (event, type, sales, food cost, fuel, event fee, staff)
EVENTS = [
    ("Downtown Lunch", "Lunch", 1450, 435, 40, 50, 240),
    ("Brewery Friday", "Brewery", 2100, 630, 55, 0, 320),
    ("Farmers Market", "Market", 980, 294, 30, 75, 160),
    ("Office Park", "Office", 1250, 375, 35, 40, 200),
    ("Music Festival", "Festival", 3400, 1020, 70, 350, 480),
    ("Corporate Catering", "Catering", 2800, 700, 45, 0, 360),
    ("Saturday Night Lot", "Private", 1900, 570, 50, 60, 280),
    ("Sunday Brunch", "Lunch", 1350, 405, 35, 40, 220),
]

# Overhead (monthly fixed): (item, amount)
OVERHEAD = [
    ("Commissary rent", 900),
    ("Insurance", 350),
    ("Truck loan", 680),
    ("Permits & licenses", 120),
    ("Phone / POS", 90),
]

# Daily sales: (day, sales, cogs)
DAILY = [
    ("Monday", 1250, 375), ("Tuesday", 0, 0), ("Wednesday", 1450, 435),
    ("Thursday", 1350, 405), ("Friday", 2100, 630), ("Saturday", 3400, 1020),
    ("Sunday", 1350, 405),
]

# Inventory & par: (item, par, on hand, unit)
INVENTORY = [
    ("Pork butt", 40, 22, "lb"), ("Brisket", 30, 18, "lb"), ("Burger buns", 200, 120, "each"),
    ("Fries", 50, 30, "lb"), ("Cheese", 20, 12, "lb"), ("Craft soda", 240, 140, "can"),
    ("Sauce base", 12, 5, "gal"), ("Paper goods", 500, 260, "each"),
]

# Fuel & mileage: (date offset, miles, gallons, cost)
FUEL = [
    (-2, 78, 12, 46.00), (-5, 64, 10, 38.50), (-8, 92, 14, 53.20),
    (-12, 70, 11, 42.00), (-16, 85, 13, 49.40),
]

# Permits: (permit, expires in days, status)
PERMITS = [
    ("Business License", 210, "OK"), ("Health Permit", 95, "OK"),
    ("Fire Certification", 140, "OK"), ("Commissary Agreement", 300, "OK"),
    ("Special Event Permit", 18, "Renew"),
]

# Supplies / shopping: (item, qty, est. cost)
SUPPLIES = [
    ("Pork butt (case)", 2, 96.00), ("Buns (dozen)", 10, 66.00), ("Fries (case)", 1, 32.00),
    ("Gloves & liners", 1, 24.00), ("Propane refill", 2, 44.00), ("Napkins & trays", 1, 38.00),
]

# Bookings: (event, days out, location, deposit, status)
BOOKINGS = [
    ("Brewery Anniversary", 6, "Hops & Barrel", 200, "Confirmed"),
    ("Corporate Lunch", 11, "TechPark Plaza", 300, "Confirmed"),
    ("Food Truck Rally", 18, "Riverside Lot", 150, "Pending"),
    ("Wedding Catering", 26, "Oak Hollow", 500, "Confirmed"),
    ("Farmers Market", 4, "Town Square", 0, "Confirmed"),
]

# Cash & tips: (day, cash, card, tips)
CASHTIPS = [
    ("Wednesday", 420, 1030, 145), ("Thursday", 380, 970, 130),
    ("Friday", 610, 1490, 220), ("Saturday", 980, 2420, 360),
    ("Sunday", 410, 940, 150),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 12 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "money2": '"$"#,##0.00', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=abs(n))


def dplus(n):
    return dt.date.today() + dt.timedelta(days=abs(n))


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [15] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your truck, targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOUR TRUCK", "section")
    controls = [
        ("Truck name", "Rolling Smoke BBQ", None, "Truck"),
        ("Owner", "Dana", None, "Owner"),
        ("Target food-cost %", TARGET_FC, "0%", "TargetFC"),
        ("Events goal / month", EVENT_GOAL, "0", "EventGoal"),
        ("Profit-margin goal", 0.35, "0%", "MarginGoal"),
        ("Currency", "USD ($)", None, "Currency"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Unit", UNITS, "UnitList"), ("F", "Event Type", EVTYPES, "EvTypeList"),
             ("G", "Permit Status", PERMIT_STATUS, "PermitList"), ("H", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🚚  FOOD TRUCK COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Know your profit per event, cover your overhead & book the right gigs.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE TRUCK, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Cost your menu, log every event's sales and costs, and see the net profit of each gig at a "
                      "glance. A break-even calculator tells you the sales you need to cover your commissary rent, "
                      "insurance and loan; a bookings calendar keeps the right events on the schedule; and inventory, "
                      "fuel, permits, supplies and cash-&-tips tools keep the wheels turning — all in ONE premium "
                      "Google Sheets & Excel system built for mobile food.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your truck name, target food-cost % & events goal.",
             "2.  Cost your Menu — each item's cost, price & margin.",
             "3.  Log each gig on Events — sales & costs give you net profit per event.",
             "4.  Enter your fixed costs on Commissary & Overhead; read your Break-Even.",
             "5.  Keep Inventory, Fuel, Permits, Bookings & Cash-&-Tips current.",
             "6.  Check the Dashboard: sales, profit per event, break-even & a Truck Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional truck (Rolling Smoke BBQ) is included so you can see how it all connects — "
               "just type over it with your own gigs and numbers. Your events, menu, overhead and permits roll up "
               "into a live Truck Score. Twelve matching printable pages (event P&L, prep & par list, cash-&-tips "
               "sheet, permit tracker & more) are included for the truck. This is a business tool, not financial, tax "
               "or legal advice — confirm figures and permit rules with the proper authorities.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Every gig is a mini-business — know the number before you park.", "section_gold")


# ===========================================================================
# Menu & Cost
# ===========================================================================
def build_menu(wb):
    ws = wb.create_sheet("Menu & Cost"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 14, 14, 12, 14, 2])
    luxe_header(ws, "F", "🍖  MENU & COST",
                "Every item's cost, price, food-cost % & margin — price with a healthy margin.")
    table_headers(ws, 4, ["Item", "Plate Cost", "Price", "Food %", "Margin $"], start_col=2)
    start = L0
    for i, (item, cost, price) in enumerate(MENU):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"; cc.number_format = '"$"#,##0.00'
        cp = ws.cell(row=r, column=4, value=price); cp.style = "input"; cp.number_format = '"$"#,##0.00'
        cf = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/D{r},0)"); cf.style = "td"; cf.number_format = "0%"
        cm = ws.cell(row=r, column=6, value=f"=D{r}-C{r}"); cm.style = "td"; cm.number_format = '"$"#,##0.00'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MENU) - 1
    nrange(wb, "MenuItem", "Menu & Cost", "B", start, end)
    nrange(wb, "MenuFC", "Menu & Cost", "E", start, end)
    ws.conditional_formatting.add(f"E{start}:E{end}",
        ColorScaleRule(start_type="num", start_value=0.15, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=0.30, mid_color="FFFFF3CD",
                       end_type="num", end_value=0.45, end_color="FF" + RED_BG))
    ws.freeze_panes = "A5"


# ===========================================================================
# Events — the P&L engine
# ===========================================================================
def build_events(wb):
    ws = wb.create_sheet("Events"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 12, 12, 12, 10, 10, 12, 14, 12, 2])
    luxe_header(ws, "J", "📍  EVENT P&L",
                "Log every gig — sales minus food, fuel, fees & staff = the net profit of each event.")
    table_headers(ws, 4, ["Event", "Type", "Sales", "Food", "Fuel", "Fee", "Staff", "Net Profit", "Food %"], start_col=2)
    start = L0
    for i, (name, typ, sales, food, fuel, fee, staff) in enumerate(EVENTS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ws.cell(row=r, column=3, value=typ).style = "td"
        for ci, val in zip((4, 5, 6, 7, 8), (sales, food, fuel, fee, staff)):
            cc = ws.cell(row=r, column=ci, value=val); cc.style = "input"; cc.number_format = '"$"#,##0'
        cn = ws.cell(row=r, column=9, value=f"=D{r}-E{r}-F{r}-G{r}-H{r}"); cn.style = "td"; cn.font = Font(bold=True, color=PRIMARY); cn.number_format = '"$"#,##0'
        cf = ws.cell(row=r, column=10, value=f"=IFERROR(E{r}/D{r},0)"); cf.style = "td"; cf.number_format = "0%"
        if i % 2:
            for c in range(2, 11):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(EVENTS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL / MONTH").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    for col in (4, 5, 6, 7, 8, 9):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cf = ws.cell(row=tot, column=10, value=f"=IFERROR(E{tot}/D{tot},0)"); cf.style = "td"; cf.font = Font(bold=True, color=PRIMARY); cf.fill = fill(SURFACE); cf.number_format = "0%"
    nrange(wb, "EvName", "Events", "B", start, end)
    nrange(wb, "EvSales", "Events", "D", start, end)
    nrange(wb, "EvFood", "Events", "E", start, end)
    nrange(wb, "EvNet", "Events", "I", start, end)
    cell_name(wb, "TotalSales", "Events", f"$D${tot}")
    cell_name(wb, "TotalFood", "Events", f"$E${tot}")
    cell_name(wb, "TotalProfit", "Events", f"$I${tot}")
    ws.conditional_formatting.add(f"I{start}:I{end}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
# Break-Even
# ===========================================================================
def build_breakeven(wb):
    ws = wb.create_sheet("Break-Even"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 34, 18, 4, 34, 18, 2])
    luxe_header(ws, "F", "⚖  BREAK-EVEN",
                "The sales you need to cover your fixed costs — know it before you park.")
    merge_set(ws, "B5:C5", "THE NUMBERS", "section_gold")
    rows = [
        ("Monthly fixed overhead", "=OverheadTotal", '"$"#,##0'),
        ("Avg net profit / event", "=IFERROR(AVERAGE(EvNet),0)", '"$"#,##0'),
        ("Break-even events / month", "=IFERROR(OverheadTotal/AVERAGE(EvNet),0)", "0.0"),
        ("Avg sales / event", "=IFERROR(AVERAGE(EvSales),0)", '"$"#,##0'),
        ("Break-even sales / month", "=IFERROR(OverheadTotal/(1-TotalFood/TotalSales-0.32),0)", '"$"#,##0'),
    ]
    for i, (lab, fml, fmt) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "input" if i in (0,) else "field_value"; c.number_format = fmt
        if i in (2, 4):
            c.fill = fill(MINT_BG); c.font = Font(bold=True, color=PRIMARY)
    cell_name(wb, "BreakEvenEvents", "Break-Even", "$C$8")
    merge_set(ws, "E5:F5", "EVERY GIG AFTER BREAK-EVEN", "section_gold")
    ws.cell(row=6, column=5, value="Once you clear break-even, most of each event's").style = "td_left"
    ws.cell(row=7, column=5, value="net profit drops straight to your pocket. Aim for").style = "td_left"
    ws.cell(row=8, column=5, value="events with high sales and low fees & fuel.").style = "td_left"
    ws.cell(row=10, column=2, value="Break-even sales assumes ~32% labor & other variable costs on top of food.").style = "section"


# ===========================================================================
# Daily Sales
# ===========================================================================
def build_daily(wb):
    ws = wb.create_sheet("Daily Sales"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 2])
    luxe_header(ws, "E", "🗓  DAILY SALES",
                "Sales & cost of goods each service day — the pulse of the week.")
    table_headers(ws, 4, ["Day", "Sales", "COGS", "Food %"], start_col=2)
    start = L0
    for i, (day, sales, cogs) in enumerate(DAILY):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cs = ws.cell(row=r, column=3, value=sales); cs.style = "input"; cs.number_format = '"$"#,##0'
        cc = ws.cell(row=r, column=4, value=cogs); cc.style = "input"; cc.number_format = '"$"#,##0'
        cf = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); cf.style = "td"; cf.number_format = "0%"
        if i % 2:
            for c in range(2, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(DAILY) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK TOTAL").style = "th"
    cs = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = '"$"#,##0'
    cc = ws.cell(row=tot, column=4, value=f"=SUM(D{start}:D{end})"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0'
    cf = ws.cell(row=tot, column=5, value=f"=IFERROR(D{tot}/C{tot},0)"); cf.style = "td"; cf.font = Font(bold=True, color=PRIMARY); cf.fill = fill(SURFACE); cf.number_format = "0%"
    ws.add_chart(_barchart(ws, "Sales by Day", start, end, 3, 2), "G4")
    ws.freeze_panes = "A5"


def _barchart(ws, title, start, end, val_col, cat_col):
    ch = BarChart(); ch.title = title; ch.height = 7.4; ch.width = 12
    ch.add_data(Reference(ws, min_col=val_col, min_row=start, max_row=end), titles_from_data=False)
    ch.set_categories(Reference(ws, min_col=cat_col, min_row=start, max_row=end)); ch.dataLabels = no_labels(); ch.legend = None
    return ch


# ===========================================================================
# Commissary & Overhead
# ===========================================================================
def build_overhead(wb):
    ws = wb.create_sheet("Commissary & Overhead"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 2])
    luxe_header(ws, "C", "🏭  COMMISSARY & OVERHEAD",
                "Your monthly fixed costs — the number your gigs have to beat every month.")
    table_headers(ws, 4, ["Fixed Cost", "Monthly"], start_col=2)
    start = L0
    for i, (item, amt) in enumerate(OVERHEAD):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        c = ws.cell(row=r, column=3, value=amt); c.style = "input"; c.number_format = '"$"#,##0'
        if i % 2:
            ws.cell(row=r, column=2).fill = fill(MUTED_ROW); ws.cell(row=r, column=3).fill = fill(MUTED_ROW)
    end = start + len(OVERHEAD) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL OVERHEAD / MO").style = "th"
    c = ws.cell(row=tot, column=3, value=f"=SUM(C{start}:C{end})"); c.style = "td"; c.font = Font(bold=True, color=DANGER); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cell_name(wb, "OverheadTotal", "Commissary & Overhead", f"$C${tot}")
    ws.freeze_panes = "A5"


# ===========================================================================
# Inventory & Par
# ===========================================================================
def build_inventory(wb):
    ws = wb.create_sheet("Inventory & Par"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 12, 12, 10, 14, 2])
    luxe_header(ws, "F", "📦  INVENTORY & PAR",
                "Par levels vs on hand — see what to restock before the next service.")
    table_headers(ws, 4, ["Item", "Par", "On Hand", "Unit", "To Buy"], start_col=2)
    start = L0
    for i, (item, par, oh, unit) in enumerate(INVENTORY):
        r = start + i
        ws.cell(row=r, column=2, value=item).style = "td_left"
        cp = ws.cell(row=r, column=3, value=par); cp.style = "input"; cp.number_format = "#,##0"
        co = ws.cell(row=r, column=4, value=oh); co.style = "input"; co.number_format = "#,##0"
        ws.cell(row=r, column=5, value=unit).style = "td"
        cb = ws.cell(row=r, column=6, value=f"=MAX(C{r}-D{r},0)"); cb.style = "td"; cb.font = Font(bold=True, color=PRIMARY); cb.number_format = "#,##0"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(INVENTORY) - 1
    for c in range(2, 7):
        add_dv(ws, f"E{start}:E{end}", "UnitList")
    ws.conditional_formatting.add(f"D{start}:D{end}",
        CellIsRule(operator="lessThan", formula=[f"C{start}*0.5"], fill=fill(RED_BG)))
    ws.freeze_panes = "A5"


# ===========================================================================
# Fuel & Mileage
# ===========================================================================
def build_fuel(wb):
    rows = [(dminus(off), miles, gal, cost) for (off, miles, gal, cost) in FUEL]
    rows.sort(key=lambda r: r[0], reverse=True)
    ws, start, end = build_log(
        wb, "Fuel & Mileage", "⛽", "FUEL & MILEAGE",
        "Track every fill-up — miles, gallons & cost. Fuel is a real line on every gig.",
        ["Date", "Miles", "Gallons", "Cost"],
        rows, [2, 14, 12, 12, 12, 2], dates={2}, ints={3}, dec={4}, money2={5}, reserved=30, start_col=2)
    nrange(wb, "FuelMiles", "Fuel & Mileage", "C", start, end)
    nrange(wb, "FuelCost", "Fuel & Mileage", "E", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    cm = ws.cell(row=tot, column=3, value="=SUM(FuelMiles)"); cm.style = "td"; cm.font = Font(bold=True, color=PRIMARY); cm.fill = fill(SURFACE); cm.number_format = "#,##0"
    ws.cell(row=tot, column=4).style = "td"; ws.cell(row=tot, column=4).fill = fill(SURFACE)
    cc = ws.cell(row=tot, column=5, value="=SUM(FuelCost)"); cc.style = "td"; cc.font = Font(bold=True, color=PRIMARY); cc.fill = fill(SURFACE); cc.number_format = '"$"#,##0.00'
    cell_name(wb, "MilesTotal", "Fuel & Mileage", f"$C${tot}")


# ===========================================================================
# Permits & Licenses
# ===========================================================================
def build_permits(wb):
    ws = wb.create_sheet("Permits"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 14, 2])
    luxe_header(ws, "D", "📋  PERMITS & LICENSES",
                "Every permit & its expiry — never get shut down at a gig for a lapsed license.")
    table_headers(ws, 4, ["Permit / License", "Expires", "Status"], start_col=2)
    start = L0
    for i, (name, days, status) in enumerate(PERMITS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        ce = ws.cell(row=r, column=3, value=dplus(days)); ce.style = "input"; ce.number_format = "mm/dd/yyyy"
        cs = ws.cell(row=r, column=4, value=status); cs.style = "input"
        add_dv(ws, f"D{r}", "PermitList")
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(PERMITS) - 1
    nrange(wb, "PermitStatus", "Permits", "D", start, end)
    cmap = {"OK": MINT_BG, "Renew": WARN_BG, "Expired": RED_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    ws.cell(row=end + 2, column=2, value="Permits current").style = "field_label"
    c = ws.cell(row=end + 2, column=3, value='=COUNTIF(PermitStatus,"OK")&" of "&COUNTA(PermitStatus)'); c.style = "field_value"; c.fill = fill(MINT_BG)
    ws.freeze_panes = "A5"


# ===========================================================================
# Supplies / Shopping
# ===========================================================================
def build_supplies(wb):
    ws, start, end = build_log(
        wb, "Supplies", "🛒", "SUPPLIES & SHOPPING",
        "The next restock run — quantities & estimated cost so nothing runs out mid-service.",
        ["Item", "Qty", "Est. Cost"],
        SUPPLIES, [2, 30, 10, 14, 2], text_left={2}, ints={3}, money2={4}, reserved=24, start_col=2)
    nrange(wb, "SupplyCost", "Supplies", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="SHOPPING TOTAL").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value="=SUM(SupplyCost)"); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'


# ===========================================================================
# Bookings
# ===========================================================================
def build_bookings(wb):
    ws = wb.create_sheet("Bookings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 14, 22, 12, 14, 2])
    luxe_header(ws, "F", "📅  BOOKINGS",
                "Upcoming gigs on the calendar — deposits, locations & confirmation status.")
    table_headers(ws, 4, ["Event", "Date", "Location", "Deposit", "Status"], start_col=2)
    start = L0
    for i, (name, days, loc, dep, status) in enumerate(BOOKINGS):
        r = start + i
        ws.cell(row=r, column=2, value=name).style = "td_left"
        cd = ws.cell(row=r, column=3, value=dplus(days)); cd.style = "input"; cd.number_format = "mm/dd/yyyy"
        ws.cell(row=r, column=4, value=loc).style = "td_left"
        cdp = ws.cell(row=r, column=5, value=dep); cdp.style = "input"; cdp.number_format = '"$"#,##0'
        cs = ws.cell(row=r, column=6, value=status); cs.style = "td"
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(BOOKINGS) - 1
    nrange(wb, "BookStatus", "Bookings", "F", start, end)
    cmap = {"Confirmed": MINT_BG, "Pending": WARN_BG}
    for st, cc in cmap.items():
        ws.conditional_formatting.add(f"F{start}:F{end}", CellIsRule(operator="equal", formula=[f'"{st}"'], fill=fill(cc)))
    ws.cell(row=end + 2, column=2, value="Confirmed gigs").style = "field_label"
    c = ws.cell(row=end + 2, column=3, value='=COUNTIF(BookStatus,"Confirmed")'); c.style = "field_value"; c.number_format = "0"; c.fill = fill(MINT_BG)
    ws.freeze_panes = "A5"


# ===========================================================================
# Cash & Tips
# ===========================================================================
def build_cash(wb):
    ws = wb.create_sheet("Cash & Tips"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 16, 14, 14, 14, 14, 2])
    luxe_header(ws, "F", "💵  CASH & TIPS",
                "Reconcile each day — cash, card & tips, so the drawer always balances.")
    table_headers(ws, 4, ["Day", "Cash", "Card", "Tips", "Total"], start_col=2)
    start = L0
    for i, (day, cash, card, tips) in enumerate(CASHTIPS):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"
        cc = ws.cell(row=r, column=3, value=cash); cc.style = "input"; cc.number_format = '"$"#,##0'
        cd = ws.cell(row=r, column=4, value=card); cd.style = "input"; cd.number_format = '"$"#,##0'
        ct = ws.cell(row=r, column=5, value=tips); ct.style = "input"; ct.number_format = '"$"#,##0'
        cto = ws.cell(row=r, column=6, value=f"=C{r}+D{r}+E{r}"); cto.style = "td"; cto.font = Font(bold=True, color=PRIMARY); cto.number_format = '"$"#,##0'
        if i % 2:
            for c in range(2, 7):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(CASHTIPS) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="TOTAL").style = "th"
    for col in (3, 4, 5, 6):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=SUM({L}{start}:{L}{end})")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    ws.freeze_panes = "A5"


# ===========================================================================
# Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🚚  FOOD TRUCK COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Sales, profit per event, break-even & a Truck Score — your whole mobile business, at a glance.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("EVENTS", "=COUNTA(EvName)", "num"),
        ("TOTAL SALES", "=TotalSales", "money"),
        ("TOTAL PROFIT", "=TotalProfit", "money"),
        ("AVG PROFIT/EVENT", "=IFERROR(AVERAGE(EvNet),0)", "money"),
        ("FOOD COST", "=IFERROR(TotalFood/TotalSales,0)", "pct"),
        ("AVG SALES/EVENT", "=IFERROR(AVERAGE(EvSales),0)", "money"),
    ]
    row2 = [
        ("BREAK-EVEN", "=BreakEvenEvents", "dec"),
        ("OVERHEAD/MO", "=OverheadTotal", "money"),
        ("TOP EVENT", "=INDEX(EvName,MATCH(MAX(EvNet),EvNet,0))", "text"),
        ("BEST SALES", "=INDEX(EvName,MATCH(MAX(EvSales),EvSales,0))", "text"),
        ("PERMITS OK", '=COUNTIF(PermitStatus,"OK")&"/"&COUNTA(PermitStatus)', "text"),
        ("TRUCK SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "TRUCK HEALTH", "section_gold")
    merge_set(ws, "H11:M11", "NET PROFIT BY EVENT", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Profit margin", "=IFERROR(MIN((TotalProfit/TotalSales)/MarginGoal,1),0)"),
        ("Food cost on target", "=IFERROR(MIN(TargetFC/(TotalFood/TotalSales),1),0)"),
        ("Events booked", "=IFERROR(MIN(COUNTA(EvName)/EventGoal,1),0)"),
        ("Permits current", '=IFERROR(COUNTIF(PermitStatus,"OK")/COUNTA(PermitStatus),0)'),
        ("Profitable events", "=IFERROR(COUNTIF(EvNet,\">0\")/COUNTA(EvName),0)"),
        ("Overhead covered", "=IFERROR(MIN(TotalProfit/OverheadTotal,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"OK","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    ev = wb["Events"]
    ch = BarChart(); ch.type = "bar"; ch.title = "Net Profit by Event"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(ev, min_col=9, min_row=5, max_row=4 + len(EVENTS)), titles_from_data=False)
    ch.set_categories(Reference(ev, min_col=2, min_row=5, max_row=4 + len(EVENTS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Food Truck Command Center™ — know your profit per event & cover your overhead.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_menu(wb); build_events(wb)
    build_breakeven(wb); build_daily(wb); build_overhead(wb); build_inventory(wb)
    build_fuel(wb); build_permits(wb); build_supplies(wb); build_bookings(wb)
    build_cash(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Menu & Cost", "Events", "Break-Even", "Daily Sales",
             "Commissary & Overhead", "Inventory & Par", "Fuel & Mileage", "Permits",
             "Supplies", "Bookings", "Cash & Tips", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Food_Truck_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
