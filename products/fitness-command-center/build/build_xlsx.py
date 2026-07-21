"""Build Fitness & Meal-Prep Command Center™ — The Complete Health System.

12 tabs (+ Settings) · a premium fitness & nutrition operating system in Google
Sheets & Excel. Dashboard, goals & stats, a weekly meal plan, a recipe bank, an
auto grocery list, a macro tracker, a workout plan, a workout log, body metrics,
a habit tracker — one dashboard. Plan the food, log the lifts, track the body &
build the habit.

Run: python3 build_xlsx.py   ->  ../Fitness_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

YESNO = ["Yes", "No"]
MEAL_SLOTS = ["Breakfast", "Lunch", "Dinner", "Snack"]
CATS = ["Breakfast", "Lunch", "Dinner", "Snack"]
AISLES = ["Produce", "Meat", "Dairy", "Pantry", "Frozen", "Bakery", "Other"]
FOCUS = ["Push", "Pull", "Legs", "Upper", "Lower", "Full Body", "Cardio", "Core", "Rest"]

START_WT = 185; CURRENT_WT = 176; GOAL_WT = 165
CAL_TARGET = 2100; PROTEIN_TARGET = 150; WATER_TARGET = 8; STEP_GOAL = 10000
WEEKLY_WORKOUT_GOAL = 5

# Body metrics: weekly weigh-ins
WEIGHTS = [("Wk 1", 185.0), ("Wk 2", 183.4), ("Wk 3", 182.1), ("Wk 4", 180.6),
           ("Wk 5", 179.5), ("Wk 6", 178.2), ("Wk 7", 177.3), ("Wk 8", 176.6), ("Wk 9", 176.0)]

# Macro log: (day offset, calories, protein, carbs, fat)
MACROS = [
    (-13, 2080, 148, 210, 62), (-12, 2010, 152, 190, 60), (-11, 2150, 140, 225, 66),
    (-10, 1980, 155, 185, 58), (-9, 2120, 146, 215, 64), (-8, 2050, 150, 200, 61),
    (-7, 2200, 138, 230, 70), (-6, 1990, 158, 180, 57), (-5, 2060, 149, 205, 62),
    (-4, 2040, 151, 198, 60), (-3, 2100, 144, 214, 65), (-2, 1970, 156, 182, 56),
    (-1, 2090, 147, 208, 63), (0, 2030, 153, 195, 59),
]

# Recipes: (recipe, servings, cal/serving, protein/serving, category)
RECIPES = [
    ("Overnight Oats & Berries", 1, 380, 24, "Breakfast"),
    ("Greek Yogurt Protein Bowl", 1, 320, 28, "Breakfast"),
    ("Veggie Egg Scramble", 1, 340, 26, "Breakfast"),
    ("Chicken & Rice Bowl", 1, 520, 45, "Lunch"),
    ("Turkey Hummus Wrap", 1, 430, 35, "Lunch"),
    ("Tuna Quinoa Salad", 1, 410, 38, "Lunch"),
    ("Salmon & Roasted Veg", 1, 540, 42, "Dinner"),
    ("Lean Beef Stir-Fry", 1, 600, 44, "Dinner"),
    ("Chicken Fajita Bowl", 1, 560, 46, "Dinner"),
    ("Protein Smoothie", 1, 280, 30, "Snack"),
    ("Cottage Cheese & Fruit", 1, 220, 24, "Snack"),
    ("Apple & Peanut Butter", 1, 260, 8, "Snack"),
]

# Weekly meal plan: (day, breakfast, lunch, dinner, snack, day calories)
MEALPLAN = [
    ("Monday", "Overnight Oats & Berries", "Chicken & Rice Bowl", "Salmon & Roasted Veg", "Protein Smoothie", 1720),
    ("Tuesday", "Greek Yogurt Protein Bowl", "Turkey Hummus Wrap", "Lean Beef Stir-Fry", "Cottage Cheese & Fruit", 1570),
    ("Wednesday", "Veggie Egg Scramble", "Tuna Quinoa Salad", "Chicken Fajita Bowl", "Apple & Peanut Butter", 1570),
    ("Thursday", "Overnight Oats & Berries", "Chicken & Rice Bowl", "Salmon & Roasted Veg", "Protein Smoothie", 1720),
    ("Friday", "Greek Yogurt Protein Bowl", "Turkey Hummus Wrap", "Lean Beef Stir-Fry", "Cottage Cheese & Fruit", 1570),
    ("Saturday", "Veggie Egg Scramble", "Tuna Quinoa Salad", "Chicken Fajita Bowl", "Protein Smoothie", 1720),
    ("Sunday", "Overnight Oats & Berries", "Turkey Hummus Wrap", "Chicken Fajita Bowl", "Cottage Cheese & Fruit", 1580),
]

# Grocery list: (item, aisle, have?)
GROCERY = [
    ("Chicken breast", "Meat", "No"), ("Salmon fillets", "Meat", "No"),
    ("Lean ground beef", "Meat", "No"), ("Greek yogurt", "Dairy", "Yes"),
    ("Eggs", "Dairy", "No"), ("Cottage cheese", "Dairy", "No"),
    ("Rolled oats", "Pantry", "Yes"), ("Brown rice", "Pantry", "No"),
    ("Quinoa", "Pantry", "No"), ("Protein powder", "Pantry", "Yes"),
    ("Olive oil", "Pantry", "Yes"), ("Broccoli", "Produce", "No"),
    ("Spinach", "Produce", "No"), ("Mixed berries", "Produce", "No"),
    ("Sweet potato", "Produce", "No"), ("Apples", "Produce", "No"),
]

# Workout plan: (day, focus, done?)
WORKOUT_PLAN = [
    ("Monday", "Upper Push", "Yes"), ("Tuesday", "Lower", "Yes"),
    ("Wednesday", "Rest", "No"), ("Thursday", "Upper Pull", "Yes"),
    ("Friday", "Lower + Core", "Yes"), ("Saturday", "Cardio", "No"),
    ("Sunday", "Rest", "No"),
]

# Workout log: (day offset, exercise, sets, reps, weight)
WORKOUT_LOG = [
    (-5, "Bench Press", 4, 8, 135), (-5, "Incline DB Press", 3, 10, 50), (-5, "Overhead Press", 3, 8, 75),
    (-4, "Back Squat", 4, 6, 185), (-4, "Romanian Deadlift", 3, 8, 155), (-4, "Leg Press", 3, 12, 270),
    (-2, "Pull-Up", 4, 8, 25), (-2, "Barbell Row", 4, 10, 115), (-2, "Face Pull", 3, 15, 40),
    (-1, "Deadlift", 3, 5, 225), (-1, "Walking Lunge", 3, 12, 40), (-1, "Plank (sec)", 3, 45, 0),
]

# Habit log: (day offset, water cups, sleep hrs, steps)
HABITS = [
    (-13, 7, 7.5, 9400), (-12, 8, 7.0, 10200), (-11, 6, 6.5, 8100), (-10, 7, 7.5, 9800),
    (-9, 8, 8.0, 11000), (-8, 7, 7.0, 9200), (-7, 6, 6.5, 7800), (-6, 8, 7.5, 10500),
    (-5, 7, 7.0, 9600), (-4, 7, 7.5, 9100), (-3, 8, 8.0, 10800), (-2, 6, 6.5, 8300),
    (-1, 7, 7.0, 9400), (0, 8, 7.5, 9700),
]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, dec=None, money2=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set(); dec = dec or set(); money2 = money2 or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in money2:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "#,##0"
            elif c in dec:
                cell.number_format = "0.0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vsize = 13 if kind == "text" else 18
    vc.font = Font(size=vsize, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "#,##0", "money": '"$"#,##0', "pct": "0%", "dec": "0.0", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=abs(n))


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, dec=None, money2=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", start_col=1):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers) + start_col - 1)
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers, start_col=start_col)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, start_col):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set(), dec=dec or set(), money2=money2 or set(), start_col=start_col)
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def cell_name(wb, name, sheet, ref):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!{ref}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 20, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your stats, targets & lists once — every tab follows.")
    merge_set(ws, "B5:C5", "YOU & YOUR TARGETS", "section")
    controls = [
        ("Name", "Jordan", None, "Name"),
        ("Start Weight (lb)", START_WT, "0.0", "StartWt"),
        ("Current Weight (lb)", CURRENT_WT, "0.0", "CurrentWt"),
        ("Goal Weight (lb)", GOAL_WT, "0.0", "GoalWt"),
        ("Daily Calorie Target", CAL_TARGET, "#,##0", "CalTarget"),
        ("Daily Protein Target (g)", PROTEIN_TARGET, "#,##0", "ProteinTarget"),
        ("Water Target (cups)", WATER_TARGET, "0", "WaterTarget"),
        ("Daily Step Goal", STEP_GOAL, "#,##0", "StepGoal"),
        ("Weekly Workout Goal", WEEKLY_WORKOUT_GOAL, "0", "WeeklyWorkoutGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    banks = [("E", "Meal Slot", MEAL_SLOTS, "SlotList"), ("F", "Category", CATS, "CatList"),
             ("G", "Aisle", AISLES, "AisleList"), ("H", "Focus", FOCUS, "FocusList"), ("I", "Yes / No", YESNO, "YesNoList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")


# ===========================================================================
# Start Here
# ===========================================================================
def build_start(wb):
    ws = wb.create_sheet("Start Here"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 84, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  💪  FITNESS & MEAL-PREP COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  Plan the food, log the lifts, track the body & build the habit — all in one place.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "YOUR WHOLE HEALTH ROUTINE, IN ONE PLACE", "section_gold")
    ws.merge_cells("B6:B9")
    ws["B6"].value = ("Set your stats & targets, plan a week of meals from a reusable recipe bank, auto-build a grocery "
                      "list, log your macros, follow a weekly workout split, record every lift, and track your weight, "
                      "water, sleep & steps — all in ONE premium Google Sheets & Excel system. Everything rolls up into "
                      "a live Fitness Score so you can see, at a glance, whether you're on track for your goal. Simple "
                      "enough for a beginner, detailed enough for a nerd.")
    ws["B6"].style = "body"; ws["B6"].fill = fill(MINT_BG); ws["B6"].border = BOX
    for rr in range(6, 10):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(MINT_BG)
    merge_set(ws, "B11:B11", "START HERE — 6 QUICK STEPS", "section")
    steps = ["1.  Open Settings — add your weight, calorie & protein targets, and weekly goals.",
             "2.  Fill the Recipe Bank, then drag recipes into the Weekly Meal Plan.",
             "3.  Build your Grocery List from the plan — check off what you already have.",
             "4.  Log your macros daily; the tracker averages calories & protein vs target.",
             "5.  Follow the Workout Plan & log each lift; update Body Metrics weekly.",
             "6.  Check the Dashboard: weight to goal, macros, workouts & a Fitness Score."]
    for i, s in enumerate(steps):
        r = 12 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 20
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data for a fictional person is included so you can see how it all connects — just type over it "
               "with your own. Your weight, macros, workouts & habits roll up into a live Fitness Score. Twelve "
               "matching printable pages (meal planner, grocery list, workout log, macro tracker, progress chart & "
               "more) are included to print and keep. This is a general wellness tool, not medical, nutrition or "
               "training advice — talk to a doctor or qualified professional before starting any new diet or program.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "Consistency beats intensity — show up, log it, and let the score climb.", "section_gold")


# ===========================================================================
# Goals & Stats
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals & Stats"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 4, 30, 18, 2])
    luxe_header(ws, "F", "🎯  GOALS & STATS",
                "Your targets in one place — the numbers every other tab measures against.")
    merge_set(ws, "B5:C5", "BODY", "section_gold")
    merge_set(ws, "E5:F5", "NUTRITION & ACTIVITY", "section_gold")
    left = [("Start weight (lb)", "=StartWt", "0.0"), ("Current weight (lb)", "=CurrentWt", "0.0"),
            ("Goal weight (lb)", "=GoalWt", "0.0"), ("Lost so far (lb)", "=StartWt-CurrentWt", "0.0"),
            ("To go (lb)", "=CurrentWt-GoalWt", "0.0"), ("% to goal", "=IFERROR((StartWt-CurrentWt)/(StartWt-GoalWt),0)", "0%")]
    for i, (lab, val, fmt) in enumerate(left):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "td_left"
        c = ws.cell(row=r, column=3, value=val); c.style = "td"; c.number_format = fmt
        if i == 5:
            c.fill = fill(MINT_BG); c.font = Font(bold=True, color=PRIMARY)
    right = [("Calorie target", "=CalTarget", "#,##0"), ("Avg calories (14d)", "=IFERROR(AVERAGE(MacroCal),0)", "#,##0"),
             ("Protein target (g)", "=ProteinTarget", "#,##0"), ("Avg protein (g)", "=IFERROR(AVERAGE(MacroProtein),0)", "#,##0"),
             ("Weekly workout goal", "=WeeklyWorkoutGoal", "0"), ("Daily step goal", "=StepGoal", "#,##0")]
    for i, (lab, val, fmt) in enumerate(right):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "td_left"
        c = ws.cell(row=r, column=6, value=val); c.style = "td"; c.number_format = fmt
    ws.cell(row=13, column=2, value="Set your targets in Settings — everything here follows automatically.").style = "section"


# ===========================================================================
# Recipe Bank  (defines RecipeName / RecipeCal / RecipeProtein)
# ===========================================================================
def build_recipes(wb):
    ws, start, end = build_log(
        wb, "Recipe Bank", "🍳", "RECIPE BANK",
        "Your go-to meals with calories & protein — build the plan by pulling from here.",
        ["Recipe", "Servings", "Cal / serving", "Protein (g)", "Category"],
        RECIPES, [2, 32, 12, 14, 14, 16], text_left={2}, ints={4, 5}, reserved=30,
        validations=[("F", "CatList")], start_col=2)
    nrange(wb, "RecipeName", "Recipe Bank", "B", start, end)
    nrange(wb, "RecipeCal", "Recipe Bank", "D", start, end)
    nrange(wb, "RecipeProtein", "Recipe Bank", "E", start, end)


# ===========================================================================
# Weekly Meal Plan
# ===========================================================================
def build_mealplan(wb):
    ws = wb.create_sheet("Meal Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 12, 26, 26, 26, 22, 14, 2])
    luxe_header(ws, "G", "🍽  WEEKLY MEAL PLAN",
                "A full week of meals — pull from your recipe bank and keep each day near target.")
    table_headers(ws, 4, ["Day", "Breakfast", "Lunch", "Dinner", "Snack", "Calories"], start_col=2)
    start = L0
    for i, (day, b, l, dn, sn, cal) in enumerate(MEALPLAN):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"; ws.cell(row=r, column=2).font = Font(bold=True, color=PRIMARY)
        for ci, val in zip((3, 4, 5, 6), (b, l, dn, sn)):
            ws.cell(row=r, column=ci, value=val).style = "td_left"
        cc = ws.cell(row=r, column=7, value=cal); cc.style = "td"; cc.number_format = "#,##0"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(MEALPLAN) - 1; tot = end + 1
    ws.cell(row=tot, column=2, value="WEEK AVG").style = "th"
    for c in range(3, 7):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    ca = ws.cell(row=tot, column=7, value=f"=ROUND(AVERAGE(G{start}:G{end}),0)"); ca.style = "td"; ca.font = Font(bold=True, color=PRIMARY); ca.fill = fill(SURFACE); ca.number_format = "#,##0"
    nrange(wb, "PlanCal", "Meal Plan", "G", start, end)
    ws.conditional_formatting.add(f"G{start}:G{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=2200, color=PRIMARY, showValue=True))
    ws.freeze_panes = "A5"


# ===========================================================================
# Grocery List
# ===========================================================================
def build_grocery(wb):
    ws, start, end = build_log(
        wb, "Grocery List", "🛒", "GROCERY LIST",
        "Everything the plan needs, by aisle — check off what you already have at home.",
        ["Item", "Aisle", "Have?"],
        GROCERY, [2, 30, 16, 12], text_left={2}, reserved=30,
        validations=[("C", "AisleList"), ("D", "YesNoList")], start_col=2)
    nrange(wb, "GroceryHave", "Grocery List", "D", start, end)
    nrange(wb, "GroceryItem", "Grocery List", "B", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="STILL TO BUY").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value='=COUNTIF(GroceryHave,"No")'); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0"
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))


# ===========================================================================
# Macro Tracker  (defines MacroCal / MacroProtein)
# ===========================================================================
def build_macros(wb):
    rows = [(dminus(off), cal, prot, carb, fat) for (off, cal, prot, carb, fat) in MACROS]
    rows.sort(key=lambda r: r[0], reverse=True)
    ws, start, end = build_log(
        wb, "Macro Tracker", "🥗", "MACRO TRACKER",
        "Log what you eat — the tracker averages calories & protein against your targets.",
        ["Date", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)"],
        rows, [2, 14, 14, 14, 14, 12], dates={2}, ints={3, 4, 5, 6}, reserved=45, start_col=2)
    nrange(wb, "MacroCal", "Macro Tracker", "C", start, end)
    nrange(wb, "MacroProtein", "Macro Tracker", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVERAGE").style = "th"
    for col, nm in ((3, None), (4, None), (5, None), (6, None)):
        L = get_column_letter(col)
        c = ws.cell(row=tot, column=col, value=f"=ROUND(AVERAGE({L}{start}:{L}{end}),0)")
        c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "#,##0"
    # color scale on calories vs target
    ws.conditional_formatting.add(f"C{start}:C{end}",
        ColorScaleRule(start_type="num", start_value=1600, start_color="FF" + HIGHLIGHT,
                       mid_type="num", mid_value=2100, mid_color="FFFFF3CD",
                       end_type="num", end_value=2600, end_color="FF" + RED_BG))


# ===========================================================================
# Workout Plan  (defines WorkoutDone)
# ===========================================================================
def build_workoutplan(wb):
    ws = wb.create_sheet("Workout Plan"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 22, 12, 2])
    luxe_header(ws, "D", "🗓  WORKOUT PLAN",
                "Your weekly split — check off each session as you finish it.")
    table_headers(ws, 4, ["Day", "Focus", "Done?"], start_col=2)
    start = L0
    for i, (day, focus, done) in enumerate(WORKOUT_PLAN):
        r = start + i
        ws.cell(row=r, column=2, value=day).style = "td_left"; ws.cell(row=r, column=2).font = Font(bold=True, color=PRIMARY)
        ws.cell(row=r, column=3, value=focus).style = "td_left"
        cd = ws.cell(row=r, column=4, value=done); cd.style = "input"
        add_dv(ws, f"D{r}", "YesNoList")
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(WORKOUT_PLAN) - 1
    nrange(wb, "WorkoutDone", "Workout Plan", "D", start, end)
    tot = end + 1
    ws.cell(row=tot, column=2, value="DONE THIS WEEK").style = "th"
    ws.cell(row=tot, column=3).style = "td"; ws.cell(row=tot, column=3).fill = fill(SURFACE)
    c = ws.cell(row=tot, column=4, value='=COUNTIF(WorkoutDone,"Yes")'); c.style = "td"; c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0"
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.freeze_panes = "A5"


# ===========================================================================
# Workout Log
# ===========================================================================
def build_workoutlog(wb):
    rows = [(dminus(off), ex, s, rp, w, f"=D{0}") for (off, ex, s, rp, w) in WORKOUT_LOG]
    # volume computed per-row below
    data = [(dminus(off), ex, s, rp, w) for (off, ex, s, rp, w) in WORKOUT_LOG]
    data.sort(key=lambda r: r[0], reverse=True)
    ws = wb.create_sheet("Workout Log"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 26, 10, 10, 12, 14, 2])
    luxe_header(ws, "G", "🏋  WORKOUT LOG",
                "Every lift, logged — sets × reps × weight = volume, the number that tracks progress.")
    table_headers(ws, 4, ["Date", "Exercise", "Sets", "Reps", "Weight", "Volume"], start_col=2)
    start = L0
    for i, (d, ex, s, rp, w) in enumerate(data):
        r = start + i
        cd = ws.cell(row=r, column=2, value=d); cd.style = "td"; cd.number_format = "mm/dd/yyyy"
        ws.cell(row=r, column=3, value=ex).style = "td_left"
        ws.cell(row=r, column=4, value=s).style = "input"
        ws.cell(row=r, column=5, value=rp).style = "input"
        cw = ws.cell(row=r, column=6, value=w); cw.style = "input"; cw.number_format = "#,##0"
        cv = ws.cell(row=r, column=7, value=f"=D{r}*E{r}*F{r}"); cv.style = "td"; cv.number_format = "#,##0"
        if i % 2:
            for c in range(2, 8):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(data) - 1
    # pad reserved rows
    for r in range(end + 1, end + 21):
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c); cell.style = "td"; cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
        ws.cell(row=r, column=7, value=f"=IF(AND(D{r}<>\"\",E{r}<>\"\",F{r}<>\"\"),D{r}*E{r}*F{r},\"\")").number_format = "#,##0"
    tot = end + 21
    ws.cell(row=tot, column=2, value="TOTAL VOLUME").style = "th"
    for c in range(3, 7):
        ws.cell(row=tot, column=c).style = "td"; ws.cell(row=tot, column=c).fill = fill(SURFACE)
    cv = ws.cell(row=tot, column=7, value=f"=SUM(G{start}:G{tot-1})"); cv.style = "td"; cv.font = Font(bold=True, color=PRIMARY); cv.fill = fill(SURFACE); cv.number_format = "#,##0"
    ws.freeze_panes = "A5"


# ===========================================================================
# Body Metrics  (defines WeightVal)
# ===========================================================================
def build_body(wb):
    ws = wb.create_sheet("Body Metrics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 14, 14, 14, 2])
    luxe_header(ws, "D", "📉  BODY METRICS",
                "Weigh in weekly — the trend line, not any single day, tells the real story.")
    table_headers(ws, 4, ["Week", "Weight (lb)", "Change"], start_col=2)
    start = L0
    for i, (wk, val) in enumerate(WEIGHTS):
        r = start + i
        ws.cell(row=r, column=2, value=wk).style = "td_left"
        cb = ws.cell(row=r, column=3, value=val); cb.style = "input"; cb.number_format = "0.0"
        cc = ws.cell(row=r, column=4, value=(f"=C{r}-C{r-1}" if i else 0)); cc.style = "td"; cc.number_format = "+0.0;[Red]-0.0"
        if i % 2:
            for c in range(2, 5):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    end = start + len(WEIGHTS) - 1
    nrange(wb, "WeightVal", "Body Metrics", "C", start, end)
    nrange(wb, "WeightWk", "Body Metrics", "B", start, end)
    ch = LineChart(); ch.title = "Weight Trend"; ch.height = 8; ch.width = 15
    ch.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end)); ch.dataLabels = no_labels()
    ws.add_chart(ch, "F4")
    ws.cell(row=end + 2, column=2, value="Total change").style = "field_label"
    c = ws.cell(row=end + 2, column=3, value=f"=C{end}-C{start}"); c.style = "field_value"; c.number_format = "+0.0;[Red]-0.0"; c.fill = fill(MINT_BG)
    ws.freeze_panes = "A5"


# ===========================================================================
# Habit Tracker  (defines HabitWater / HabitSteps / HabitSleep)
# ===========================================================================
def build_habits(wb):
    rows = [(dminus(off), water, sleep, steps) for (off, water, sleep, steps) in HABITS]
    rows.sort(key=lambda r: r[0], reverse=True)
    ws, start, end = build_log(
        wb, "Habit Tracker", "✅", "HABIT TRACKER",
        "Water, sleep & steps, every day — the small habits that make the big results stick.",
        ["Date", "Water (cups)", "Sleep (hrs)", "Steps"],
        rows, [2, 14, 16, 14, 14], dates={2}, ints={5}, dec={4}, reserved=45, start_col=2)
    nrange(wb, "HabitWater", "Habit Tracker", "C", start, end)
    nrange(wb, "HabitSleep", "Habit Tracker", "D", start, end)
    nrange(wb, "HabitSteps", "Habit Tracker", "E", start, end)
    # fix sleep format to one decimal
    for r in range(start, end + 1):
        ws.cell(row=r, column=4).number_format = "0.0"
    tot = end + 1
    ws.cell(row=tot, column=2, value="AVERAGE").style = "th"
    cw = ws.cell(row=tot, column=3, value=f"=ROUND(AVERAGE(C{start}:C{end}),1)"); cw.style = "td"; cw.font = Font(bold=True, color=PRIMARY); cw.fill = fill(SURFACE); cw.number_format = "0.0"
    cs = ws.cell(row=tot, column=4, value=f"=ROUND(AVERAGE(D{start}:D{end}),1)"); cs.style = "td"; cs.font = Font(bold=True, color=PRIMARY); cs.fill = fill(SURFACE); cs.number_format = "0.0"
    cst = ws.cell(row=tot, column=5, value=f"=ROUND(AVERAGE(E{start}:E{end}),0)"); cst.style = "td"; cst.font = Font(bold=True, color=PRIMARY); cst.fill = fill(SURFACE); cst.number_format = "#,##0"


# ===========================================================================
# Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  💪  FITNESS & MEAL-PREP COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Weight, macros, workouts & habits — your whole health routine, automatically tracked.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("CURRENT WEIGHT", "=CurrentWt", "dec"),
        ("GOAL WEIGHT", "=GoalWt", "dec"),
        ("LBS TO GO", "=CurrentWt-GoalWt", "dec"),
        ("LOST SO FAR", "=StartWt-CurrentWt", "dec"),
        ("CALORIE TARGET", "=CalTarget", "num"),
        ("AVG CALORIES", "=IFERROR(ROUND(AVERAGE(MacroCal),0),0)", "num"),
    ]
    row2 = [
        ("PROTEIN TARGET", "=ProteinTarget", "num"),
        ("AVG PROTEIN", "=IFERROR(ROUND(AVERAGE(MacroProtein),0),0)", "num"),
        ("WORKOUTS / WK", '=COUNTIF(WorkoutDone,"Yes")&" / "&WeeklyWorkoutGoal', "text"),
        ("STEPS AVG", "=IFERROR(ROUND(AVERAGE(HabitSteps),0),0)", "num"),
        ("WATER AVG", "=IFERROR(ROUND(AVERAGE(HabitWater),1),0)", "dec"),
        ("FITNESS SCORE", "=IFERROR(AVERAGE(HealthRange),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:G11", "FITNESS SCORE", "section_gold")
    merge_set(ws, "H11:M11", "WEIGHT TREND", "section_gold")
    table_headers(ws, 12, ["Check", "Score", "Status"], start_col=2)
    dims = [
        ("Weight to goal", "=IFERROR((StartWt-CurrentWt)/(StartWt-GoalWt),0)"),
        ("Protein hit", "=IFERROR(MIN(AVERAGE(MacroProtein)/ProteinTarget,1),0)"),
        ("Calories on target", "=IFERROR(1-ABS(AVERAGE(MacroCal)-CalTarget)/CalTarget,0)"),
        ("Workouts done", '=IFERROR(MIN(COUNTIF(WorkoutDone,"Yes")/WeeklyWorkoutGoal,1),0)'),
        ("Steps", "=IFERROR(MIN(AVERAGE(HabitSteps)/StepGoal,1),0)"),
        ("Water", "=IFERROR(MIN(AVERAGE(HabitWater)/WaterTarget,1),0)"),
    ]
    hs = 13
    for i, (dim, fml) in enumerate(dims):
        r = hs + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.85,"Strong",IF(C{r}>=0.6,"On track","Focus"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    he = hs + len(dims) - 1
    cell_name(wb, "HealthRange", "Dashboard", f"$C${hs}:$C${he}")
    ws.conditional_formatting.add(f"C{hs}:C{he}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    body = wb["Body Metrics"]
    ch = LineChart(); ch.title = "Weight Trend"; ch.height = 7.4; ch.width = 8.6
    ch.add_data(Reference(body, min_col=3, min_row=4, max_row=4 + len(WEIGHTS)), titles_from_data=True)
    ch.set_categories(Reference(body, min_col=2, min_row=5, max_row=4 + len(WEIGHTS))); ch.dataLabels = no_labels(); ch.legend = None
    ws.add_chart(ch, "H13")
    ws.row_dimensions[21].height = 26
    merge_set(ws, "B21:M21", "Fitness & Meal-Prep Command Center™ — plan the food, log the lifts, track the body.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_start(wb); build_goals(wb); build_recipes(wb)
    build_mealplan(wb); build_grocery(wb); build_macros(wb); build_workoutplan(wb)
    build_workoutlog(wb); build_body(wb); build_habits(wb); build_dashboard(wb)

    order = ["Start Here", "Dashboard", "Goals & Stats", "Meal Plan", "Recipe Bank", "Grocery List",
             "Macro Tracker", "Workout Plan", "Workout Log", "Body Metrics", "Habit Tracker", "Settings"]
    wb._sheets = [wb[n] for n in order]

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Fitness_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(wb._sheets)} sheets)")


if __name__ == "__main__":
    main()
