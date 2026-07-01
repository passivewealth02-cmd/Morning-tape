"""Build Etsy Seller Profit Dashboard™ — The Ultimate Etsy Finance System.

14 sheets + Welcome · a mini Etsy CFO system in Excel & Google Sheets.
Turns messy shop data into clean profit intelligence — automatically.

Run: python3 build_xlsx.py   ->  ../Etsy_Seller_Profit_Dashboard.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Brand tokens
# ---------------------------------------------------------------------------
PRIMARY = "1B4F48"
ACCENT = "937356"
GOLD_LT = "C9A86A"
SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"
MINT_BG = "E3F8EF"
WHITE = "FFFFFF"
TEXT = "333333"
DANGER = "C94C4C"
RED_BG = "FBE6E6"
WARN_BG = "FBF0E2"
MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"
SOFT_BG = "FAF7F1"
IVORY = "FBF8F2"

CATEGORIES = ["Printable", "Digital", "Handmade", "POD", "Template", "Other"]
EXPENSE_TYPES = ["Software / Tools", "Mockups", "Marketing", "Education", "Freelancers", "Website / Domain", "Other"]
FEE_TYPES = ["Listing", "Transaction", "Processing", "Offsite Ads", "Etsy Ads", "Subscription"]
AD_TYPES = ["Etsy Ads", "Offsite Ads", "Social", "Influencer"]
GOAL_TYPES = ["Revenue", "Profit", "Listings", "ROAS", "Conversion", "Passive"]
YESNO = ["Yes", "No"]
STATUSES = ["Active", "Paused", "Sold Out", "Draft"]

# Product catalog: (name, category, price, cost)
PRODUCTS = [
    ("Budget Planner", "Printable", 12, 0),
    ("Wedding Suite", "Printable", 18, 0),
    ("Resume Bundle", "Template", 14, 0),
    ("IG Templates", "Digital", 9, 0),
    ("Mega Bundle", "Printable", 29, 0),
    ("Sticker Pack", "Digital", 6, 0),
]
# order distribution (product index, count)
ORDER_DIST = [(0, 12), (1, 7), (2, 7), (3, 6), (4, 7), (5, 4)]


def build_orders_data():
    """Deterministic order list -> (order_id, date, prod_idx, qty, offsite, refunded).

    Products are interleaved (round-robin) so the order log reads like a real,
    varied sales history rather than being grouped by product.
    """
    remaining = {pi: cnt for pi, cnt in ORDER_DIST}
    seq = []
    while any(remaining.values()):
        for pi, _ in ORDER_DIST:
            if remaining[pi] > 0:
                seq.append(pi)
                remaining[pi] -= 1
    orders = []
    n = len(seq)
    for i, pi in enumerate(seq):
        qty = 2 if i % 7 == 0 else 1
        offsite = (i % 4 == 0)
        refunded = i in (5, 31)
        date = dt.date.today() - dt.timedelta(days=(28 - (i * 28 // n)))
        orders.append((f"#{2401 + i}", date, pi, qty, offsite, refunded))
    return orders


ORDERS = build_orders_data()

# fee rates (Etsy, 2026 typical US)
FEE_LISTING = 0.20
FEE_TXN = 0.065
FEE_PROC = 0.03
FEE_PROC_FIXED = 0.25
FEE_OFFSITE = 0.15
AD_SPEND_TOTAL = 55
EXPENSES_TOTAL = 32
SUBSCRIPTION = 0

LOG_ROWS = 40
L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
# Styles & shared helpers
# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)

    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"),
                            fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True),
                               fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY),
                              alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT),
                                   alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"),
                         fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                         border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT),
                              alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True),
                              border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY),
                            fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY),
                                  alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX),
        "prompt": NamedStyle(name="prompt", font=f(11, False, TEXT, italic=True),
                             alignment=Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True),
                             border=BOX, fill=PatternFill("solid", fgColor=IVORY)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT),
                           alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng)
    cell = ws[rng.split(":")[0]]
    cell.value = value
    cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None,
               dates=None, pcts=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set()
    dates = dates or set(); pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0.00'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}")
    ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label
    lc.font = Font(size=9, bold=True, color=ACCENT)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula
    vc.font = Font(size=13 if kind == "text" else 19, bold=True, color=PRIMARY)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "money2": '"$"#,##0.00',
                        "pct": "0%", "text": "General"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc)
            c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN,
                              top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18
    ws.row_dimensions[row + 1].height = 44


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None,
              validations=None, reserved=LOG_ROWS, freeze="A5", money_fmt='"$"#,##0.00'):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers),
               text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set())
    if money:
        for r in range(start, end + 1):
            for c in money:
                ws.cell(row=r, column=c).number_format = money_fmt
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def no_labels():
    dl = DataLabelList()
    dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 3] + [16] * 6)
    luxe_header(ws, "J", "⚙  SETTINGS", "Set your fee rates & goals once — every calculation follows. 2026 Etsy defaults included.")

    merge_set(ws, "B5:C5", "FEE RATES & GOALS", "section")
    controls = [
        ("Shop Name", "Willow & Oak Co.", None, "ShopName"),
        ("Listing Fee ($)", FEE_LISTING, '"$"#,##0.00', "FeeListing"),
        ("Transaction Fee %", FEE_TXN, "0.0%", "FeeTxn"),
        ("Processing Fee %", FEE_PROC, "0.0%", "FeeProc"),
        ("Processing Fixed ($)", FEE_PROC_FIXED, '"$"#,##0.00', "FeeProcFixed"),
        ("Offsite Ads Fee %", FEE_OFFSITE, "0%", "FeeOffsite"),
        ("Etsy Subscription ($/mo)", SUBSCRIPTION, '"$"#,##0', "FeeSubscription"),
        ("Tax Set-Aside %", 0.25, "0%", "TaxRate"),
        ("Monthly Revenue Goal", 1000, '"$"#,##0', "RevGoal"),
        ("Monthly Profit Goal", 700, '"$"#,##0', "ProfitGoal"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")

    banks = [("E", "Product Category", CATEGORIES, "CategoryList"),
             ("F", "Expense Type", EXPENSE_TYPES, "ExpenseTypeList"),
             ("G", "Fee Type", FEE_TYPES, "FeeTypeList"),
             ("H", "Ad Type", AD_TYPES, "AdTypeList"),
             ("I", "Goal Type", GOAL_TYPES, "GoalTypeList"),
             ("J", "Status / YesNo", STATUSES + ["—"] + YESNO, "StatusList")]
    merge_set(ws, "E5:J5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    wb.defined_names["YesNoList"] = DefinedName("YesNoList", attr_text="Settings!$J$11:$J$12")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 74, 3])
    ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  🧾  ETSY SELLER PROFIT DASHBOARD™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate Etsy finance & profit tracking system — 2026 edition.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    merge_set(ws, "B5:B5", "STOP GUESSING — SEE REAL PROFIT", "section_gold")
    ws.merge_cells("B6:B8")
    ws["B6"].value = (
        "Every Etsy seller eventually asks the same question: “How much am I ACTUALLY making "
        "after fees?” Etsy fees, transaction fees, ads, refunds and taxes make that almost "
        "impossible to see. This dashboard is your mini Etsy CFO — enter your orders and it "
        "shows exactly what you make per product, per order, and per month. Real profit, not "
        "just revenue.")
    ws["B6"].style = "body"
    for r in (6, 7, 8):
        ws.row_dimensions[r].height = 22

    merge_set(ws, "B10:B10", "HOW TO USE IT", "section")
    steps = [
        "1.  Open Settings — the 2026 Etsy fee rates are pre-filled (edit if yours differ).",
        "2.  Add your products in the Product Profit Calculator & Product Library.",
        "3.  Log sales in the Order Tracker — fees & profit calculate automatically.",
        "4.  Track Etsy Ads and business Expenses on their tabs.",
        "5.  Watch the Seller Dashboard reveal your true profit, margin & best sellers.",
    ]
    for i, s in enumerate(steps):
        r = 11 + i
        ws.merge_cells(f"B{r}:B{r}")
        ws[f"B{r}"].value = s
        ws[f"B{r}"].style = "body"
        ws.row_dimensions[r].height = 22

    dr = 18
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th")
    ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = (
        "Sample data is included so you can see how everything connects — just type over it "
        "with your own numbers. Fee rates are estimates; confirm your exact rates in your Etsy "
        "Payment account. This tool is for tracking only and is not tax or accounting advice.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT)
    c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22
        ws.cell(row=rr, column=2).fill = fill(WARN_BG)

    merge_set(ws, f"B{dr+5}:B{dr+5}",
              "Know your numbers, scale what works — let's find your real profit.", "section_gold")


# ===========================================================================
# 3 — Product Profit Calculator  (built before Orders so COGS lookup exists)
# ===========================================================================
def build_product_calc(wb):
    ws = wb.create_sheet("Product Calc")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 12, 11, 12, 12, 12, 12, 12, 13, 11])
    luxe_header(ws, "J", "🧮  PRODUCT PROFIT CALCULATOR",
                "Per-product profit after every Etsy fee — see your true margin on each listing.")
    table_headers(ws, 4, ["Product", "Price", "COGS", "Listing Fee", "Txn Fee",
                          "Proc Fee", "Ads / Sale", "Total Cost", "Net / Sale", "Margin"])
    start = L0
    end = start + len(PRODUCTS) - 1
    for i, (name, cat, price, cost) in enumerate(PRODUCTS):
        r = start + i
        ws.cell(row=r, column=1, value=name).style = "td_left"
        cp = ws.cell(row=r, column=2, value=price); cp.style = "input"
        cc = ws.cell(row=r, column=3, value=cost); cc.style = "input"
        ws.cell(row=r, column=4, value="=FeeListing")
        ws.cell(row=r, column=5, value=f"=FeeTxn*B{r}")
        ws.cell(row=r, column=6, value=f"=FeeProc*B{r}+FeeProcFixed")
        ws.cell(row=r, column=7, value=f"=IFERROR(SUMIF(AdProduct,A{r},AdSpend)/MAX(SUMIF(AdProduct,A{r},AdOrders),1),0)")
        ws.cell(row=r, column=8, value=f"=C{r}+D{r}+E{r}+F{r}+G{r}")
        ws.cell(row=r, column=9, value=f"=B{r}-H{r}")
        cm = ws.cell(row=r, column=10, value=f"=IFERROR(I{r}/B{r},0)"); cm.number_format = "0%"
        for c in range(2, 10):
            ws.cell(row=r, column=c).number_format = '"$"#,##0.00'
        for c in range(1, 11):
            cell = ws.cell(row=r, column=c)
            if c != 1:
                cell.style = "td"
            cell.fill = fill(MUTED_ROW if i % 2 else WHITE)
        if i % 2 == 0:
            ws.cell(row=r, column=2).fill = fill(SURFACE); ws.cell(row=r, column=3).fill = fill(SURFACE)
    ws.cell(row=start, column=2).number_format = '"$"#,##0.00'
    ws.conditional_formatting.add(
        f"J{start}:J{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))
    nrange(wb, "CalcName", "Product Calc", "A", start, end)
    nrange(wb, "CalcCOGS", "Product Calc", "C", start, end)
    nrange(wb, "CalcNet", "Product Calc", "I", start, end)
    ws.freeze_panes = "A5"


# ===========================================================================
# 2 — Order Tracker  (the engine)
# ===========================================================================
def build_orders(wb):
    ws = wb.create_sheet("Orders")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [11, 12, 18, 6, 11, 11, 11, 12, 12, 11, 12, 12])
    luxe_header(ws, "L", "🧾  ORDER TRACKER",
                "Log every sale — Etsy fees, net revenue & final profit calculate automatically.")
    table_headers(ws, 4, ["Order ID", "Date", "Product", "Qty", "Unit Price", "Sale Total",
                          "Shipping", "Offsite Ad?", "Etsy Fees", "COGS", "Refunded?", "Net Profit"])
    start = L0
    end = start + max(len(ORDERS), LOG_ROWS) - 1
    price_by_idx = {i: p[2] for i, p in enumerate(PRODUCTS)}
    for i, (oid, date, pi, qty, offsite, refunded) in enumerate(ORDERS):
        r = start + i
        ws.cell(row=r, column=1, value=oid)
        ws.cell(row=r, column=2, value=date)
        ws.cell(row=r, column=3, value=PRODUCTS[pi][0])
        ws.cell(row=r, column=4, value=qty)
        ws.cell(row=r, column=5, value=price_by_idx[pi])
        ws.cell(row=r, column=6, value=f"=D{r}*E{r}")
        ws.cell(row=r, column=7, value=0)
        ws.cell(row=r, column=8, value="Yes" if offsite else "No")
        ws.cell(row=r, column=9,
                value=f'=IF(F{r}="","",FeeListing*D{r}+FeeTxn*(F{r}+G{r})+(FeeProc*(F{r}+G{r})+FeeProcFixed)+IF(H{r}="Yes",FeeOffsite*(F{r}+G{r}),0))')
        ws.cell(row=r, column=10, value=f"=IFERROR(SUMIF(CalcName,C{r},CalcCOGS)*D{r},0)")
        ws.cell(row=r, column=11, value="Yes" if refunded else "No")
        ws.cell(row=r, column=12, value=f'=IF(F{r}="","",IF(K{r}="Yes",0,F{r}+G{r}-I{r}-J{r}))')
    style_rows(ws, start, end, 12, text_left={1, 3}, dates={2}, ints={4},
               money={5, 6, 7, 9, 10, 12})
    add_dv(ws, f"C{start}:C{end}", "CalcName")
    add_dv(ws, f"H{start}:H{end}", "YesNoList")
    add_dv(ws, f"K{start}:K{end}", "YesNoList")
    ws.conditional_formatting.add(
        f"A{start}:L{end}",
        FormulaRule(formula=[f'$K{start}="Yes"'], fill=fill(RED_BG)))
    ws.conditional_formatting.add(
        f"L{start}:L{end}",
        CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))
    nrange(wb, "OrdID", "Orders", "A", start, end)
    nrange(wb, "OrdDate", "Orders", "B", start, end)
    nrange(wb, "OrdProduct", "Orders", "C", start, end)
    nrange(wb, "OrdQty", "Orders", "D", start, end)
    nrange(wb, "OrdSale", "Orders", "F", start, end)
    nrange(wb, "OrdShip", "Orders", "G", start, end)
    nrange(wb, "OrdOffsite", "Orders", "H", start, end)
    nrange(wb, "OrdFees", "Orders", "I", start, end)
    nrange(wb, "OrdRefunded", "Orders", "K", start, end)
    nrange(wb, "OrdProfit", "Orders", "L", start, end)
    ws.freeze_panes = "A5"
    return end


# ===========================================================================
# 5 — Ads Performance Tracker  (built before Product Calc ads lookup? define names)
# ===========================================================================
def build_ads(wb):
    sample = [
        ("Budget Planner - Search", "Budget Planner", 12, 4200, 168, 6, 72, "Active"),
        ("Wedding Suite - Search", "Wedding Suite", 15, 3100, 108, 5, 90, "Active"),
        ("Mega Bundle - Promoted", "Mega Bundle", 18, 5400, 210, 7, 203, "Active"),
        ("Resume Bundle - Search", "Resume Bundle", 6, 1500, 45, 2, 28, "Paused"),
        ("Sticker Pack - Promoted", "Sticker Pack", 4, 2200, 55, 1, 6, "Paused"),
    ]
    ws, start, end = build_log(
        wb, "Ads", "📣", "ADS PERFORMANCE TRACKER",
        "Track Etsy Ads like a pro — ROAS and profit-after-ads calculate themselves.",
        ["Campaign", "Product", "Spend", "Impressions", "Clicks", "Orders", "Revenue", "Status"],
        sample, [24, 16, 10, 13, 10, 9, 11, 11],
        text_left={1, 2, 8}, ints={4, 5, 6}, money={3, 7}, reserved=25, money_fmt='"$"#,##0')
    # CTR / ROAS / profit helper columns
    ws.cell(row=4, column=9, value="CTR").style = "th"
    ws.cell(row=4, column=10, value="ROAS").style = "th"
    ws.column_dimensions["I"].width = 9
    ws.column_dimensions["J"].width = 9
    for r in range(start, end + 1):
        c = ws.cell(row=r, column=9, value=f'=IF(D{r}="","",E{r}/D{r})'); c.style = "td"; c.number_format = "0.0%"
        c.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
        c = ws.cell(row=r, column=10, value=f'=IF(C{r}="","",IFERROR(G{r}/C{r},0))'); c.style = "td"; c.number_format = "0.0\"x\""
        c.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
    nrange(wb, "AdProduct", "Ads", "B", start, end)
    nrange(wb, "AdSpend", "Ads", "C", start, end)
    nrange(wb, "AdOrders", "Ads", "F", start, end)
    nrange(wb, "AdRevenue", "Ads", "G", start, end)
    ws.conditional_formatting.add(
        f"J{start}:J{end}",
        CellIsRule(operator="greaterThanOrEqual", formula=["3"], fill=fill(MINT_BG)))
    # totals
    ws.cell(row=4, column=12, value="TOTALS").style = "section_gold"
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 12
    tot = [("Ad spend", "=SUM(AdSpend)", '"$"#,##0'),
           ("Ad revenue", "=SUM(AdRevenue)", '"$"#,##0'),
           ("Blended ROAS", "=IFERROR(SUM(AdRevenue)/SUM(AdSpend),0)", '0.0"x"'),
           ("Orders from ads", "=SUM(AdOrders)", "0")]
    for i, (lab, fml, fmt) in enumerate(tot):
        r = 5 + i
        ws.cell(row=r, column=12, value=lab).style = "field_label"
        c = ws.cell(row=r, column=13, value=fml); c.style = "field_value"; c.number_format = fmt


# ===========================================================================
# 4 — Etsy Fees Breakdown
# ===========================================================================
def build_fees(wb):
    ws = wb.create_sheet("Fees")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 16, 14, 4, 24, 16])
    luxe_header(ws, "F", "💳  ETSY FEES BREAKDOWN",
                "Every fee, fully transparent — see exactly what Etsy takes and what % of revenue.")
    merge_set(ws, "A4:C4", "FEES THIS MONTH", "section_gold")
    ws.cell(row=5, column=1, value="Fee Type").style = "th"
    ws.cell(row=5, column=2, value="Amount").style = "th"
    ws.cell(row=5, column=3, value="% of Rev").style = "th"
    fees = [
        ("Listing Fees", "=FeeListing*SUM(OrdQty)"),
        ("Transaction Fees", "=FeeTxn*(SUM(OrdSale)+SUM(OrdShip))"),
        ("Processing Fees", "=FeeProc*(SUM(OrdSale)+SUM(OrdShip))+FeeProcFixed*COUNTA(OrdID)"),
        ("Offsite Ads Fees", '=FeeOffsite*SUMIFS(OrdSale,OrdOffsite,"Yes")'),
        ("Etsy Ads Spend", "=SUM(AdSpend)"),
        ("Subscription", "=FeeSubscription"),
    ]
    fstart = 6
    for i, (lab, fml) in enumerate(fees):
        r = fstart + i
        ws.cell(row=r, column=1, value=lab).style = "td_left"
        c = ws.cell(row=r, column=2, value=fml); c.style = "td"; c.number_format = '"$"#,##0.00'
        c = ws.cell(row=r, column=3, value=f"=IFERROR(B{r}/(SUM(OrdSale)+SUM(OrdShip)),0)"); c.style = "td"; c.number_format = "0.0%"
        if i % 2:
            for cc in range(1, 4):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    fend = fstart + len(fees) - 1
    tr = fend + 1
    ws.cell(row=tr, column=1, value="TOTAL FEES").style = "th"
    c = ws.cell(row=tr, column=2, value=f"=SUM(B{fstart}:B{fend})"); c.style = "td"
    c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0.00'
    c = ws.cell(row=tr, column=3, value=f"=IFERROR(B{tr}/(SUM(OrdSale)+SUM(OrdShip)),0)"); c.style = "td"
    c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = "0.0%"
    nrange(wb, "FeeLabels", "Fees", "A", fstart, fend)
    nrange(wb, "FeeVals", "Fees", "B", fstart, fend)
    wb.defined_names["FeesTotal"] = DefinedName("FeesTotal", attr_text=f"Fees!$B${tr}")

    donut = DoughnutChart(); donut.title = "Fee Breakdown by Type"; donut.height = 8; donut.width = 12
    donut.add_data(Reference(ws, min_col=2, min_row=fstart, max_row=fend), titles_from_data=False)
    donut.set_categories(Reference(ws, min_col=1, min_row=fstart, max_row=fend))
    donut.dataLabels = no_labels()
    ws.add_chart(donut, "E5")


# ===========================================================================
# 6 — Expense Tracker
# ===========================================================================
def build_expenses(wb):
    sample = [
        ("Canva Pro", "Software / Tools", dminus(12), 13, "Monthly", "Design"),
        ("Creative Market fonts", "Software / Tools", dminus(20), 18, "One-time", ""),
        ("Placeit mockups", "Mockups", dminus(8), 15, "Monthly", ""),
        ("Etsy SEO course", "Education", dminus(40), 47, "One-time", "eRank guide"),
        ("eRank Pro", "Marketing", dminus(5), 6, "Monthly", "Keyword research"),
        ("Domain (shop link)", "Website / Domain", dminus(60), 12, "Annual", ""),
    ]
    ws, start, end = build_log(
        wb, "Expenses", "🧰", "EXPENSE TRACKER",
        "Non-product business costs — because tools & courses eat into profit too.",
        ["Expense", "Type", "Date", "Amount", "Frequency", "Notes"],
        sample, [24, 18, 13, 12, 13, 20],
        text_left={1, 6}, dates={3}, money={4}, reserved=40, money_fmt='"$"#,##0.00')
    nrange(wb, "ExpAmount", "Expenses", "D", start, end)
    ws.cell(row=4, column=8, value="TOTALS").style = "section_gold"
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 12
    ws.cell(row=5, column=8, value="Total expenses").style = "field_label"
    c = ws.cell(row=5, column=9, value="=SUM(ExpAmount)"); c.style = "field_value"; c.number_format = '"$"#,##0.00'
    wb.defined_names["ExpensesTotal"] = DefinedName("ExpensesTotal", attr_text="Expenses!$I$5")


# ===========================================================================
# 7 — Product Library
# ===========================================================================
def build_library(wb):
    ws = wb.create_sheet("Library")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 14, 11, 14, 12, 13, 13, 13])
    luxe_header(ws, "H", "📚  PRODUCT LIBRARY",
                "Every listing ranked — sales, revenue & profit pulled straight from your orders.")
    table_headers(ws, 4, ["Product", "Category", "Price", "Status", "Units Sold", "Revenue", "Profit", "Margin"])
    start = L0
    end = start + len(PRODUCTS) - 1
    for i, (name, cat, price, cost) in enumerate(PRODUCTS):
        r = start + i
        ws.cell(row=r, column=1, value=name).style = "td_left"
        ws.cell(row=r, column=2, value=cat).style = "td_left"
        c = ws.cell(row=r, column=3, value=price); c.style = "input"; c.number_format = '"$"#,##0.00'
        ws.cell(row=r, column=4, value="Active").style = "td"
        ws.cell(row=r, column=5, value=f"=SUMIF(OrdProduct,A{r},OrdQty)").style = "td"
        c = ws.cell(row=r, column=6, value=f"=SUMIF(OrdProduct,A{r},OrdSale)"); c.style = "td"; c.number_format = '"$"#,##0.00'
        c = ws.cell(row=r, column=7, value=f"=SUMIF(OrdProduct,A{r},OrdProfit)"); c.style = "td"; c.number_format = '"$"#,##0.00'
        c = ws.cell(row=r, column=8, value=f"=IFERROR(G{r}/F{r},0)"); c.style = "td"; c.number_format = "0%"
        if i % 2:
            for cc in range(1, 9):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    add_dv(ws, f"B{start}:B{end}", "CategoryList")
    add_dv(ws, f"D{start}:D{end}", "StatusList")
    ws.conditional_formatting.add(
        f"G{start}:G{end}",
        DataBarRule(start_type="num", start_value=0, end_type="max", color=PRIMARY, showValue=True))
    nrange(wb, "LibName", "Library", "A", start, end)
    nrange(wb, "LibCat", "Library", "B", start, end)
    nrange(wb, "LibUnits", "Library", "E", start, end)
    nrange(wb, "LibRevenue", "Library", "F", start, end)
    nrange(wb, "LibProfit", "Library", "G", start, end)
    nrange(wb, "LibMargin", "Library", "H", start, end)
    bar = BarChart(); bar.type = "bar"; bar.title = "Profit by Product"; bar.height = 8; bar.width = 13
    bar.add_data(Reference(ws, min_col=7, min_row=4, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    bar.legend = None
    ws.add_chart(bar, "A13")
    ws.freeze_panes = "A5"


# ===========================================================================
# 8 — Monthly Financial Snapshot
# ===========================================================================
def build_monthly(wb):
    ws = wb.create_sheet("Monthly")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [12, 13, 12, 12, 12, 13, 12, 3, 12, 13])
    luxe_header(ws, "J", "📆  MONTHLY FINANCIAL SNAPSHOT",
                "Your shop's story month by month — revenue, fees, profit & growth.")
    table_headers(ws, 4, ["Month", "Revenue", "Etsy Fees", "Ad Spend", "Expenses", "Net Profit", "Margin", "Growth"])
    hist = [
        ("Feb", 210, 42, 20, 30), ("Mar", 340, 64, 25, 30), ("Apr", 480, 88, 35, 45),
        ("May", 610, 110, 40, 30), ("Jun", 720, 128, 45, 35), ("Jul", 740, 132, 55, 32),
    ]
    start = L0
    for i, (mon, rev, fees, ads, exp) in enumerate(hist):
        r = start + i
        ws.cell(row=r, column=1, value=mon).style = "td_left"
        for c, v in ((2, rev), (3, fees), (4, ads), (5, exp)):
            cell = ws.cell(row=r, column=c, value=v); cell.style = "input" if c in (2,) else "td"
            cell.number_format = '"$"#,##0'
        ws.cell(row=r, column=6, value=f"=B{r}-C{r}-D{r}-E{r}").number_format = '"$"#,##0'
        ws.cell(row=r, column=6).style = "td"; ws.cell(row=r, column=6).number_format = '"$"#,##0'
        cm = ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/B{r},0)"); cm.style = "td"; cm.number_format = "0%"
        if i == 0:
            cg = ws.cell(row=r, column=8, value="—")
        else:
            cg = ws.cell(row=r, column=8, value=f"=IFERROR(B{r}/B{r-1}-1,0)"); cg.number_format = "0%"
        cg.style = "td"
        if i % 2:
            for cc in range(1, 9):
                ws.cell(row=r, column=cc).fill = fill(MUTED_ROW)
    end = start + len(hist) - 1
    nrange(wb, "MonLabel", "Monthly", "A", start, end)
    nrange(wb, "MonRevenue", "Monthly", "B", start, end)
    nrange(wb, "MonProfit", "Monthly", "F", start, end)
    line = LineChart(); line.title = "Revenue vs Profit"; line.height = 8.5; line.width = 14
    line.add_data(Reference(ws, min_col=2, min_row=4, max_row=end), titles_from_data=True)
    line.add_data(Reference(ws, min_col=6, min_row=4, max_row=end), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    ws.add_chart(line, "A13")
    ws.freeze_panes = "A5"


# ===========================================================================
# 9 — Customer & Order Insights
# ===========================================================================
def build_insights(wb):
    ws = wb.create_sheet("Insights")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 6, 24, 16, 2])
    luxe_header(ws, "G", "🌍  CUSTOMER & ORDER INSIGHTS",
                "Who's buying and how — refund rate, order value & repeat-buyer signals.")
    merge_set(ws, "B5:C5", "ORDER METRICS", "section_gold")
    metrics = [
        ("Total orders", "=COUNTA(OrdID)", "0"),
        ("Refund rate", '=IFERROR(COUNTIF(OrdRefunded,"Yes")/COUNTA(OrdID),0)', "0.0%"),
        ("Avg order value", "=IFERROR((SUM(OrdSale)+SUM(OrdShip))/COUNTA(OrdID),0)", '"$"#,##0.00'),
        ("Avg profit / order", "=IFERROR(SUM(OrdProfit)/COUNTA(OrdID),0)", '"$"#,##0.00'),
        ("Units per order", "=IFERROR(SUM(OrdQty)/COUNTA(OrdID),0)", "0.0"),
        ("Offsite-ad orders", '=COUNTIF(OrdOffsite,"Yes")', "0"),
    ]
    for i, (lab, fml, fmt) in enumerate(metrics):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt

    merge_set(ws, "E5:F5", "TOP COUNTRIES (sample)", "section_gold")
    ws.cell(row=6, column=5, value="Country").style = "th"
    ws.cell(row=6, column=6, value="Orders").style = "th"
    for i, (co, n) in enumerate([("United States", 28), ("Canada", 6), ("United Kingdom", 5),
                                 ("Australia", 3), ("Germany", 1)]):
        r = 7 + i
        ws.cell(row=r, column=5, value=co).style = "td_left"
        ws.cell(row=r, column=6, value=n).style = "td"

    merge_set(ws, "B14:F14", "TIP", "section")
    ws.merge_cells("B15:F16")
    ws["B15"].value = ("Digital sellers rarely get repeat data from Etsy — watch your refund rate and AOV "
                       "instead. A rising AOV (via bundles) is the fastest way to grow profit without more traffic.")
    ws["B15"].style = "body"; ws["B15"].fill = fill(IVORY)


# ===========================================================================
# 10 — Goal Tracker
# ===========================================================================
def build_goals(wb):
    ws = wb.create_sheet("Goals")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [24, 14, 14, 14, 14, 4, 22])
    luxe_header(ws, "G", "🎯  GOAL TRACKER",
                "Where your shop is headed — progress bars pull live from your real numbers.")
    table_headers(ws, 4, ["Goal", "Type", "Target", "Current", "Progress"])
    goals = [
        ("Monthly revenue", "Revenue", "=RevGoal", "=SUM(OrdSale)+SUM(OrdShip)", "money"),
        ("Monthly profit", "Profit", "=ProfitGoal", "=SUM(OrdProfit)-SUM(AdSpend)-ExpensesTotal", "money"),
        ("Active listings", "Listings", 25, "=COUNTA(LibName)", "num"),
        ("Blended ROAS", "ROAS", 3, "=IFERROR(SUM(AdRevenue)/SUM(AdSpend),0)", "x"),
        ("Profit margin", "Conversion", 0.6, "=IFERROR((SUM(OrdProfit)-SUM(AdSpend)-ExpensesTotal)/(SUM(OrdSale)+SUM(OrdShip)),0)", "pct"),
        ("Passive income / mo", "Passive", 500, "=SUM(OrdProfit)-SUM(AdSpend)-ExpensesTotal", "money"),
    ]
    start = L0
    for i, (name, typ, target, current, kind) in enumerate(goals):
        r = start + i
        ws.cell(row=r, column=1, value=name).style = "td_left"
        ws.cell(row=r, column=2, value=typ).style = "td"
        ct = ws.cell(row=r, column=3, value=target); ct.style = "input"
        cc = ws.cell(row=r, column=4, value=current); cc.style = "td"
        fmt = {"money": '"$"#,##0', "num": "0", "x": '0.0"x"', "pct": "0%"}[kind]
        ct.number_format = fmt; cc.number_format = fmt
        cp = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/C{r},0)"); cp.style = "td"; cp.number_format = "0%"
        if i % 2:
            for cc2 in range(1, 6):
                ws.cell(row=r, column=cc2).fill = fill(MUTED_ROW)
    end = start + len(goals) - 1
    add_dv(ws, f"B{start}:B{end}", "GoalTypeList")
    ws.conditional_formatting.add(
        f"E{start}:E{end}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=HIGHLIGHT, showValue=True))


# ===========================================================================
# 11 — Cash Flow Dashboard
# ===========================================================================
def build_cashflow(wb):
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 6, 28, 18, 2])
    luxe_header(ws, "G", "💵  CASH FLOW DASHBOARD",
                "Money in, money out, and what you actually keep — plus a tax reserve.")
    merge_set(ws, "B5:C5", "THIS MONTH", "section_gold")
    rows = [
        ("Money In (sales + shipping)", "=SUM(OrdSale)+SUM(OrdShip)"),
        ("Etsy Fees Out", "=-FeesTotal"),
        ("Ad Spend Out", "=-SUM(AdSpend)"),
        ("Expenses Out", "=-ExpensesTotal"),
        ("Refunds Out", '=-SUMIFS(OrdSale,OrdRefunded,"Yes")'),
        ("Net Cash Flow", "=(SUM(OrdSale)+SUM(OrdShip))-FeesTotal-SUM(AdSpend)-ExpensesTotal"),
    ]
    for i, (lab, fml) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = '"$"#,##0.00;[Red]-"$"#,##0.00'
    ws.cell(row=11, column=2).font = Font(bold=True, size=12, color=PRIMARY)

    merge_set(ws, "E5:F5", "RETENTION & TAX", "section_gold")
    rows2 = [
        ("Profit retention rate", "=IFERROR(((SUM(OrdSale)+SUM(OrdShip))-FeesTotal-SUM(AdSpend)-ExpensesTotal)/(SUM(OrdSale)+SUM(OrdShip)),0)", "0%"),
        ("Est. tax to set aside", "=MAX(((SUM(OrdSale)+SUM(OrdShip))-FeesTotal-SUM(AdSpend)-ExpensesTotal)*TaxRate,0)", '"$"#,##0.00'),
        ("Keep after tax", "=MAX(((SUM(OrdSale)+SUM(OrdShip))-FeesTotal-SUM(AdSpend)-ExpensesTotal)*(1-TaxRate),0)", '"$"#,##0.00'),
    ]
    for i, (lab, fml, fmt) in enumerate(rows2):
        r = 6 + i
        ws.cell(row=r, column=5, value=lab).style = "field_label"
        c = ws.cell(row=r, column=6, value=fml); c.style = "field_value"; c.number_format = fmt


# ===========================================================================
# 12 — Tax Preparation Sheet
# ===========================================================================
def build_tax(wb):
    ws = wb.create_sheet("Tax Prep")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 18, 2])
    luxe_header(ws, "D", "🧾  TAX PREPARATION",
                "A simple, tax-ready summary. Not tax advice — take this to your accountant.")
    merge_set(ws, "B5:C5", "TAX-READY SUMMARY", "section_gold")
    rows = [
        ("Gross revenue", "=SUM(OrdSale)+SUM(OrdShip)", '"$"#,##0.00'),
        ("Etsy fees (deductible)", "=FeesTotal", '"$"#,##0.00'),
        ("Ad spend (deductible)", "=SUM(AdSpend)", '"$"#,##0.00'),
        ("Business expenses (deductible)", "=ExpensesTotal", '"$"#,##0.00'),
        ("Net taxable profit", "=(SUM(OrdSale)+SUM(OrdShip))-FeesTotal-SUM(AdSpend)-ExpensesTotal", '"$"#,##0.00'),
        ("Tax set-aside %", "=TaxRate", "0%"),
        ("Estimated tax reserve", "=MAX(((SUM(OrdSale)+SUM(OrdShip))-FeesTotal-SUM(AdSpend)-ExpensesTotal)*TaxRate,0)", '"$"#,##0.00'),
    ]
    for i, (lab, fml, fmt) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if "Net taxable" in lab or "reserve" in lab:
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    merge_set(ws, "B14:C14",
              "Keep receipts for every deductible expense. Set the reserve aside each month so tax season is stress-free.",
              "subtitle")
    ws["B14"].fill = fill(IVORY)
    for c in ("B14", "C14"):
        ws[c].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)


# ===========================================================================
# 13 — Strategy Analyzer
# ===========================================================================
def build_strategy(wb):
    ws = wb.create_sheet("Strategy")
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 30, 2])
    luxe_header(ws, "D", "🧠  STRATEGY ANALYZER",
                "Formula-driven insights — what to scale, what to fix, what to cut.")
    merge_set(ws, "B5:C5", "AUTOMATED INSIGHTS", "section_gold")
    insights = [
        ("🏆 Best-selling product", "=INDEX(LibName,MATCH(MAX(LibRevenue),LibRevenue,0))"),
        ("💰 Highest-profit product", "=INDEX(LibName,MATCH(MAX(LibProfit),LibProfit,0))"),
        ("📈 Best profit margin", "=INDEX(LibName,MATCH(MAX(LibMargin),LibMargin,0))"),
        ("⚠ Lowest-profit product", "=INDEX(LibName,MATCH(MIN(LibProfit),LibProfit,0))"),
        ("🔻 Fewest units sold", "=INDEX(LibName,MATCH(MIN(LibUnits),LibUnits,0))"),
        ("🚀 Best category (by revenue)", "=INDEX(LibCat,MATCH(MAX(LibRevenue),LibRevenue,0))"),
    ]
    for i, (lab, fml) in enumerate(insights):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"
        ws.row_dimensions[r].height = 26

    merge_set(ws, "B13:C13", "PLAYBOOK", "section")
    plays = [
        ("SCALE", "Put ad budget behind your best-seller & highest-margin items."),
        ("IMPROVE", "Re-photograph or re-title low-unit listings before cutting them."),
        ("BUNDLE", "Combine two winners into a higher-priced bundle to lift AOV."),
        ("CUT", "Retire listings that lose money after ads & fees."),
        ("RESERVE", "Set aside your tax % every month — see the Tax Prep tab."),
    ]
    for i, (tag, txt) in enumerate(plays):
        r = 14 + i
        ws.cell(row=r, column=2, value=tag).style = "field_label"
        c = ws.cell(row=r, column=3, value=txt); c.style = "td_left"
        ws.row_dimensions[r].height = 24


# ===========================================================================
# 1 — Seller Executive Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0)
    ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2])
    ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  🧾  ETSY SELLER PROFIT DASHBOARD™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2",
              "  Revenue is vanity. Profit is sanity. See exactly what you make after every Etsy fee, ad & refund.",
              "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)

    rev = "(SUM(OrdSale)+SUM(OrdShip))"
    net = "(SUM(OrdProfit)-SUM(AdSpend)-ExpensesTotal)"
    row1 = [
        ("TOTAL REVENUE", f"={rev}", "money"),
        ("NET PROFIT", f"={net}", "money"),
        ("ETSY FEES PAID", "=FeesTotal", "money"),
        ("AD SPEND", "=SUM(AdSpend)", "money"),
        ("REFUNDS", '=SUMIFS(OrdSale,OrdRefunded,"Yes")', "money"),
    ]
    row2 = [
        ("PROFIT MARGIN", f"=IFERROR({net}/{rev},0)", "pct"),
        ("AVG ORDER VALUE", f"=IFERROR({rev}/COUNTA(OrdID),0)", "money2"),
        ("ORDERS", "=COUNTA(OrdID)", "num"),
        ("BEST SELLER", "=INDEX(LibName,MATCH(MAX(LibRevenue),LibRevenue,0))", "text"),
        ("NEEDS WORK", "=INDEX(LibName,MATCH(MIN(LibProfit),LibProfit,0))", "text"),
    ]
    cols5 = [2, 4, 6, 8, 10]
    for (lab, fml, kind), col in zip(row1, cols5):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols5):
        kpi_card(ws, 8, col, 2, lab, fml, kind)

    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "PROFIT & FEES", "section_gold")

    # Revenue vs Profit trend (from Monthly)
    line = LineChart(); line.title = "Revenue vs Profit Trend"; line.height = 8.2; line.width = 11.5
    line.add_data(Reference(wb["Monthly"], min_col=2, min_row=4, max_row=10), titles_from_data=True)
    line.add_data(Reference(wb["Monthly"], min_col=6, min_row=4, max_row=10), titles_from_data=True)
    line.set_categories(Reference(wb["Monthly"], min_col=1, min_row=5, max_row=10))
    ws.add_chart(line, "B12")

    # Etsy Fees breakdown donut
    d1 = DoughnutChart(); d1.title = "Etsy Fees Breakdown"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Fees"], min_col=2, min_row=6, max_row=11), titles_from_data=False)
    d1.set_categories(Reference(wb["Fees"], min_col=1, min_row=6, max_row=11))
    d1.dataLabels = no_labels()
    ws.add_chart(d1, "H12")

    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "PRODUCTS & ADS", "section_gold")

    # Product profit bar
    p1 = BarChart(); p1.type = "bar"; p1.title = "Profit by Product"; p1.height = 8.2; p1.width = 11.5
    p1.add_data(Reference(wb["Library"], min_col=7, min_row=4, max_row=4 + len(PRODUCTS) - 1), titles_from_data=True)
    p1.set_categories(Reference(wb["Library"], min_col=1, min_row=5, max_row=4 + len(PRODUCTS) - 1))
    p1.legend = None
    ws.add_chart(p1, "B30")

    # Ads revenue vs spend
    a1 = BarChart(); a1.type = "col"; a1.title = "Ad Revenue vs Spend"; a1.height = 8.2; a1.width = 11.5
    a1.add_data(Reference(wb["Ads"], min_col=3, min_row=4, max_row=9), titles_from_data=True)
    a1.add_data(Reference(wb["Ads"], min_col=7, min_row=4, max_row=9), titles_from_data=True)
    a1.set_categories(Reference(wb["Ads"], min_col=2, min_row=5, max_row=9))
    ws.add_chart(a1, "H30")

    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47",
              "Etsy Seller Profit Dashboard™ — your mini Etsy CFO. Enter orders; everything else is automatic.",
              "subtitle")


# ===========================================================================
# Summary (for marketing) — replicate order formulas in Python
# ===========================================================================
def compute_summary():
    revenue = fees = order_profit = refunds = units = 0
    prod_rev = {p[0]: 0 for p in PRODUCTS}
    prod_profit = {p[0]: 0 for p in PRODUCTS}
    prod_units = {p[0]: 0 for p in PRODUCTS}
    for oid, date, pi, qty, offsite, refunded in ORDERS:
        name, cat, price, cost = PRODUCTS[pi]
        sale = qty * price
        listing = FEE_LISTING * qty
        txn = FEE_TXN * sale
        proc = FEE_PROC * sale + FEE_PROC_FIXED
        off = FEE_OFFSITE * sale if offsite else 0
        f = listing + txn + proc + off
        cogs = cost * qty
        profit = 0 if refunded else sale - f - cogs
        revenue += sale
        fees += f
        order_profit += profit
        units += qty
        if refunded:
            refunds += sale
        prod_rev[name] += sale
        prod_profit[name] += profit
        prod_units[name] += qty
    ad_rev = 72 + 90 + 203 + 28 + 6
    net = order_profit - AD_SPEND_TOTAL - EXPENSES_TOTAL
    best = max(prod_rev, key=prod_rev.get)
    worst = min(prod_profit, key=prod_profit.get)
    return {
        "orders": len(ORDERS), "revenue": revenue, "fees": fees, "ad_spend": AD_SPEND_TOTAL,
        "expenses": EXPENSES_TOTAL, "order_profit": order_profit, "net": net,
        "margin": net / revenue if revenue else 0, "aov": revenue / len(ORDERS),
        "refunds": refunds, "units": units, "best": best, "worst": worst,
        "ad_revenue": ad_rev, "roas": ad_rev / AD_SPEND_TOTAL,
        "prod_profit": prod_profit, "prod_rev": prod_rev, "prod_units": prod_units,
    }


# ===========================================================================
# Build
# ===========================================================================
def main():
    wb = Workbook()
    wb.remove(wb.active)
    register_styles(wb)

    build_settings(wb)
    build_welcome(wb)
    build_ads(wb)              # defines AdProduct/AdSpend used by Product Calc
    build_product_calc(wb)
    build_orders(wb)
    build_fees(wb)
    build_expenses(wb)
    build_library(wb)
    build_monthly(wb)
    build_insights(wb)
    build_goals(wb)
    build_cashflow(wb)
    build_tax(wb)
    build_strategy(wb)
    build_dashboard(wb)        # index 0

    order = ["Welcome", "Dashboard", "Orders", "Product Calc", "Fees", "Ads",
             "Expenses", "Library", "Monthly", "Insights", "Goals", "Cash Flow",
             "Tax Prep", "Strategy", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                       "Etsy_Seller_Profit_Dashboard.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")
    s = compute_summary()
    print("SUMMARY:", {k: (round(v, 2) if isinstance(v, float) else v)
                       for k, v in s.items() if k not in ("prod_profit", "prod_rev", "prod_units")})
    print("  prod_profit:", {k: round(v, 2) for k, v in s["prod_profit"].items()})
    print("  prod_rev:", s["prod_rev"], "units:", s["prod_units"])
    print("  FIRST 12 ORDERS (id, product, qty, sale, fees, refund, profit):")
    for oid, date, pi, qty, offsite, refunded in ORDERS[:12]:
        name, cat, price, cost = PRODUCTS[pi]
        sale = qty * price
        f = FEE_LISTING * qty + FEE_TXN * sale + (FEE_PROC * sale + FEE_PROC_FIXED) + (FEE_OFFSITE * sale if offsite else 0)
        profit = 0 if refunded else sale - f - cost * qty
        print(f"    {oid} {name:15s} q{qty} ${sale:<4} fees=${f:.2f} refund={'Yes' if refunded else 'No':3s} profit=${profit:.2f}")


if __name__ == "__main__":
    main()
