"""Build Travel Command Center™ — The Ultimate Travel Budget & Trip Planning System.

21 sheets + Welcome · a premium travel operating system in Excel & Sheets.
Budget, itinerary, reservations, packing, documents, currency & memories.

Run: python3 build_xlsx.py   ->  ../Travel_Command_Center.xlsx
"""
from __future__ import annotations

import datetime as dt
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

PRIMARY = "1B4F48"; ACCENT = "937356"; GOLD_LT = "C9A86A"; SURFACE = "E5D3BA"
HIGHLIGHT = "75E6C1"; MINT_BG = "E3F8EF"; WHITE = "FFFFFF"; TEXT = "333333"
DANGER = "C94C4C"; RED_BG = "FBE6E6"; WARN_BG = "FBF0E2"; MUTED_ROW = "F4ECDE"
BORDER = "D6D2C8"; SOFT_BG = "FAF7F1"; IVORY = "FBF8F2"

EXPENSE_CATS = ["Flights", "Hotels", "Vacation Rentals", "Transportation", "Fuel",
                "Food", "Drinks", "Activities", "Shopping", "Entertainment",
                "Travel Insurance", "Visas", "Souvenirs", "Phone / Data", "Tips", "Miscellaneous"]
TRAVEL_STYLES = ["Budget", "Mid-range", "Luxury", "Backpacking", "Business", "Family", "Solo", "Couple"]
ACTIVITY_CATS = ["Sightseeing", "Tour", "Museum", "Adventure", "Food", "Show", "Beach", "Nightlife"]
PACK_CATS = ["Clothing", "Shoes", "Toiletries", "Electronics", "Documents", "Medication", "Camera Gear", "Beach Gear"]
TRANSPORT_TYPES = ["Flight", "Train", "Bus", "Ferry", "Rental Car", "Ride Share", "Metro", "Taxi"]
ACCOM_TYPES = ["Hotel", "Vacation Rental", "Hostel", "Resort", "B&B", "Cruise"]
CURRENCIES = ["USD", "EUR", "GBP", "JPY", "CAD", "AUD", "MXN", "THB"]
BOOK_STATUS = ["Booked", "Planned", "Confirmed", "Waitlist", "Cancelled"]
YESNO = ["Yes", "No"]

LOG_ROWS = 40; L0 = 5
THIN = Side(style="thin", color=BORDER)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
GOLD = Side(style="medium", color=GOLD_LT)


# ===========================================================================
def register_styles(wb):
    def f(size, bold=False, color=TEXT, italic=False):
        return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)
    styles = {
        "title": NamedStyle(name="title", font=f(24, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                            alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "subtitle": NamedStyle(name="subtitle", font=f(11, False, "E5D3BA", italic=True), fill=PatternFill("solid", fgColor=PRIMARY),
                               alignment=Alignment(horizontal="left", vertical="center", indent=2)),
        "section": NamedStyle(name="section", font=f(12, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center")),
        "section_gold": NamedStyle(name="section_gold", font=f(12, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center")),
        "th": NamedStyle(name="th", font=f(11, True, "FFFFFF"), fill=PatternFill("solid", fgColor=PRIMARY),
                         alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td": NamedStyle(name="td", font=f(11, False, TEXT), alignment=Alignment(horizontal="center", vertical="center", wrap_text=True), border=BOX),
        "td_left": NamedStyle(name="td_left", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True), border=BOX),
        "input": NamedStyle(name="input", font=f(11, True, PRIMARY), fill=PatternFill("solid", fgColor=SURFACE),
                            alignment=Alignment(horizontal="center", vertical="center"), border=BOX),
        "field_label": NamedStyle(name="field_label", font=f(10, True, ACCENT), alignment=Alignment(horizontal="left", vertical="center", indent=1),
                                  border=BOX, fill=PatternFill("solid", fgColor=SOFT_BG)),
        "field_value": NamedStyle(name="field_value", font=f(11, True, PRIMARY), alignment=Alignment(horizontal="left", vertical="center", indent=1), border=BOX),
        "imgbox": NamedStyle(name="imgbox", font=f(11, True, ACCENT, italic=True), fill=PatternFill("solid", fgColor=SOFT_BG),
                             alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                             border=Border(left=GOLD, right=GOLD, top=GOLD, bottom=GOLD)),
        "body": NamedStyle(name="body", font=f(11, False, TEXT), alignment=Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)),
    }
    for s in styles.values():
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


def fill(c):
    return PatternFill("solid", fgColor=c)


def merge_set(ws, rng, value, style):
    ws.merge_cells(rng); cell = ws[rng.split(":")[0]]; cell.value = value; cell.style = style
    return cell


def luxe_header(ws, last_col, title, subtitle):
    ws.row_dimensions[1].height = 46; ws.row_dimensions[2].height = 22; ws.row_dimensions[3].height = 6
    merge_set(ws, f"A1:{last_col}1", "  " + title, "title")
    merge_set(ws, f"A2:{last_col}2", "  " + subtitle, "subtitle")
    for c in range(1, column_index_from_string(last_col) + 1):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_dv(ws, rng, list_name):
    dv = DataValidation(type="list", formula1=f"={list_name}", allow_blank=True); ws.add_data_validation(dv); dv.add(rng)


def table_headers(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        ws.cell(row=row, column=start_col + i, value=h).style = "th"
    ws.row_dimensions[row].height = 30


def style_rows(ws, start, end, ncols, text_left=None, money=None, ints=None, dates=None, pcts=None, start_col=1):
    text_left = text_left or set(); money = money or set(); ints = ints or set(); dates = dates or set(); pcts = pcts or set()
    for r in range(start, end + 1):
        for c in range(start_col, start_col + ncols):
            cell = ws.cell(row=r, column=c)
            cell.style = "td_left" if c in text_left else "td"
            cell.fill = fill(MUTED_ROW if (r - start) % 2 else WHITE)
            if c in money:
                cell.number_format = '"$"#,##0'
            elif c in pcts:
                cell.number_format = "0%"
            elif c in ints:
                cell.number_format = "0"
            elif c in dates:
                cell.number_format = "mm/dd/yyyy"


def kpi_card(ws, row, col, span, label, formula, kind="num"):
    L, R = get_column_letter(col), get_column_letter(col + span - 1)
    ws.merge_cells(f"{L}{row}:{R}{row}"); ws.merge_cells(f"{L}{row+1}:{R}{row+1}")
    lc = ws[f"{L}{row}"]; lc.value = label; lc.font = Font(size=9, bold=True, color=ACCENT); lc.alignment = Alignment(horizontal="center", vertical="center")
    vc = ws[f"{L}{row+1}"]; vc.value = formula; vc.font = Font(size=18, bold=True, color=PRIMARY); vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.number_format = {"num": "General", "money": '"$"#,##0', "pct": "0%", "days": "0", "dec": "0.0"}[kind]
    for rr in (row, row + 1):
        for cc in range(col, col + span):
            c = ws.cell(row=rr, column=cc); c.fill = fill(WHITE)
            c.border = Border(left=THIN, right=THIN, top=GOLD if rr == row else THIN, bottom=THIN)
    ws.row_dimensions[row].height = 18; ws.row_dimensions[row + 1].height = 40


def dminus(n):
    return dt.date.today() - dt.timedelta(days=n)


def dplus(n):
    return dt.date.today() + dt.timedelta(days=n)


def build_log(wb, name, icon, title, subtitle, headers, sample, widths,
              text_left=None, dates=None, money=None, ints=None, pcts=None, validations=None, reserved=LOG_ROWS, freeze="A5"):
    ws = wb.create_sheet(name); ws.sheet_view.showGridLines = False
    set_widths(ws, widths)
    last = get_column_letter(len(headers))
    luxe_header(ws, last, f"{icon}  {title}", subtitle)
    table_headers(ws, 4, headers)
    start, end = L0, L0 + reserved - 1
    for i, row in enumerate(sample):
        r = start + i
        for ci, val in enumerate(row, 1):
            ws.cell(row=r, column=ci, value=val)
    style_rows(ws, start, end, len(headers), text_left=text_left or set(), dates=dates or set(),
               money=money or set(), ints=ints or set(), pcts=pcts or set())
    for col_letter, lst in (validations or []):
        add_dv(ws, f"{col_letter}{start}:{col_letter}{end}", lst)
    ws.freeze_panes = freeze
    return ws, start, end


def nrange(wb, name, sheet, col, start, end):
    wb.defined_names[name] = DefinedName(name, attr_text=f"'{sheet}'!${col}${start}:${col}${end}")


def no_labels():
    dl = DataLabelList(); dl.showSerName = dl.showCatName = dl.showLegendKey = False
    dl.showBubbleSize = dl.showVal = dl.showPercent = False
    return dl


# ===========================================================================
# Settings
# ===========================================================================
def build_settings(wb):
    ws = wb.create_sheet("Settings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 18, 3] + [16] * 7)
    luxe_header(ws, "K", "⚙  SETTINGS", "Set your trip details once — every dashboard follows. Edit the lists to fit any journey.")
    merge_set(ws, "B5:C5", "TRIP INPUTS", "section")
    controls = [
        ("Trip Name", "Europe Grand Tour", None, "TripName"),
        ("Departure Date", dplus(45), "mm/dd/yyyy", "DepartDate"),
        ("Return Date", dplus(59), "mm/dd/yyyy", "ReturnDate"),
        ("Travelers", 2, "0", "Travelers"),
        ("Home Currency", "USD", None, "HomeCurr"),
        ("Destination Currency", "EUR", None, "DestCurr"),
        ("Exchange Rate (1 home =)", 0.92, "0.000", "ExchRate"),
        ("Savings Saved So Far", 6000, '"$"#,##0', "SavingsSaved"),
    ]
    for i, (lab, val, fmt, nm) in enumerate(controls):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=val); c.style = "input"
        if fmt:
            c.number_format = fmt
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!$C${r}")
    ws.cell(row=15, column=2, value="Trip Days").style = "field_label"
    c = ws.cell(row=15, column=3, value="=ReturnDate-DepartDate+1"); c.style = "field_value"; c.number_format = "0"
    wb.defined_names["TripDays"] = DefinedName("TripDays", attr_text="Settings!$C$15")
    banks = [("E", "Expense Category", EXPENSE_CATS, "ExpenseCatList"), ("F", "Currency", CURRENCIES, "CurrencyList"),
             ("G", "Transport Type", TRANSPORT_TYPES, "TransportList"), ("H", "Accommodation", ACCOM_TYPES, "AccomList"),
             ("I", "Travel Style", TRAVEL_STYLES, "StyleList"), ("J", "Activity Category", ACTIVITY_CATS, "ActivityCatList"),
             ("K", "Packing Category", PACK_CATS, "PackCatList")]
    merge_set(ws, "E5:K5", "DROPDOWN LISTS", "section_gold")
    for col, h, data, nm in banks:
        ci = column_index_from_string(col)
        ws.cell(row=6, column=ci, value=h).style = "th"
        for ri, v in enumerate(data):
            ws.cell(row=7 + ri, column=ci, value=v).style = "td_left"
        wb.defined_names[nm] = DefinedName(nm, attr_text=f"Settings!${col}$7:${col}${6 + len(data)}")
    # status / yesno tucked in a lower area
    ws.cell(row=24, column=5, value="Booking Status").style = "th"
    for i, v in enumerate(BOOK_STATUS):
        ws.cell(row=25 + i, column=5, value=v).style = "td_left"
    wb.defined_names["StatusList"] = DefinedName("StatusList", attr_text="Settings!$E$25:$E$29")
    ws.cell(row=24, column=6, value="Yes / No").style = "th"
    for i, v in enumerate(YESNO):
        ws.cell(row=25 + i, column=6, value=v).style = "td_left"
    wb.defined_names["YesNoList"] = DefinedName("YesNoList", attr_text="Settings!$F$25:$F$26")


# ===========================================================================
# Welcome
# ===========================================================================
def build_welcome(wb):
    ws = wb.create_sheet("Welcome"); ws.sheet_view.showGridLines = False
    set_widths(ws, [3, 76, 3]); ws.row_dimensions[1].height = 58
    merge_set(ws, "A1:C1", "  ✈  TRAVEL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:C2", "  The ultimate travel budget & trip planning system — your whole journey in one place.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 4):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    merge_set(ws, "B5:B5", "PACK YOUR BAGS — WE'VE GOT THE PLANNING", "section_gold")
    ws.merge_cells("B6:B8")
    ws["B6"].value = ("From a weekend getaway to a multi-country adventure, Travel Command Center™ brings "
                      "budgeting, itinerary, flights, hotels, activities, packing, documents, currency and "
                      "memories into one beautifully organized dashboard — so you can plan with confidence, "
                      "spend with clarity, and actually relax when you get there.")
    ws["B6"].style = "body"
    for r in (6, 7, 8):
        ws.row_dimensions[r].height = 22
    merge_set(ws, "B10:B10", "START HERE", "section")
    steps = ["1.  Open Settings and add your trip name, dates, travelers & currency.",
             "2.  Set your budget in the Budget Command Center (16 categories).",
             "3.  Add flights, hotels & activities as you book them.",
             "4.  Log expenses — your budget actuals update automatically.",
             "5.  Watch the Executive Travel Dashboard track your countdown & readiness."]
    for i, s in enumerate(steps):
        r = 11 + i
        ws.merge_cells(f"B{r}:B{r}"); ws[f"B{r}"].value = s; ws[f"B{r}"].style = "body"; ws.row_dimensions[r].height = 22
    dr = 18
    merge_set(ws, f"B{dr}:B{dr}", "  GOOD TO KNOW", "th"); ws.row_dimensions[dr].height = 26
    ws.merge_cells(f"B{dr+1}:B{dr+3}")
    c = ws[f"B{dr+1}"]
    c.value = ("Sample data (a 14-day Europe trip) is included to show how everything connects — just type "
               "over it with your own. Countdowns, budgets, packing % and the readiness score all update "
               "automatically. Every sheet is print-friendly and works in Excel and Google Sheets, on desktop and mobile.")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    c.font = Font(size=11, color=TEXT); c.fill = fill(WARN_BG); c.border = BOX
    for rr in range(dr + 1, dr + 4):
        ws.row_dimensions[rr].height = 22; ws.cell(row=rr, column=2).fill = fill(WARN_BG)
    merge_set(ws, f"B{dr+5}:B{dr+5}", "One organized workbook, zero planning stress — bon voyage!", "section_gold")


# ===========================================================================
# 2 — Trip Profile
# ===========================================================================
def build_profile(wb):
    ws = wb.create_sheet("Trip Profile"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 24, 26, 6, 22, 20, 2])
    luxe_header(ws, "G", "🌍  TRIP PROFILE", "Your trip at a glance — the essentials, all in one place.")
    blocks = [
        ("THE TRIP", [("Trip Name", "=TripName"), ("Departure", "=DepartDate"), ("Return", "=ReturnDate"),
                      ("Trip Days", "=TripDays"), ("Travelers", "=Travelers"), ("Travel Style", "Mid-range")]),
        ("LOGISTICS", [("Primary Airline", "SkyEurope"), ("Home Airport", "JFK"),
                       ("Passport Expiry", dplus(900)), ("Travel Insurance", "SafeTrip #TR-2210"),
                       ("Emergency Contact", "Mom — (555) 210-4521"), ("Embassy / Consulate", "See Documents")]),
    ]
    row = 5
    for title, fields in blocks:
        merge_set(ws, f"B{row}:F{row}", title, "section_gold"); ws.row_dimensions[row].height = 22; row += 1
        i = 0
        while i < len(fields):
            ws.cell(row=row, column=2, value=fields[i][0]).style = "field_label"
            cv = ws.cell(row=row, column=3, value=fields[i][1]); cv.style = "field_value"
            if fields[i][0] in ("Departure", "Return", "Passport Expiry"):
                cv.number_format = "mm/dd/yyyy"
            if i + 1 < len(fields):
                ws.cell(row=row, column=5, value=fields[i + 1][0]).style = "field_label"
                cv2 = ws.cell(row=row, column=6, value=fields[i + 1][1]); cv2.style = "field_value"
                if fields[i + 1][0] in ("Departure", "Return", "Passport Expiry"):
                    cv2.number_format = "mm/dd/yyyy"
            ws.row_dimensions[row].height = 24; i += 2; row += 1
        row += 1
    # countries list (for KPI)
    merge_set(ws, "B14:C14", "COUNTRIES ON THIS TRIP", "section_gold")
    for i, co in enumerate(["France", "Italy", "Spain"]):
        r = 15 + i
        ws.cell(row=r, column=2, value=co).style = "td_left"
        ws.cell(row=r, column=3, value=["Paris", "Rome", "Barcelona"][i]).style = "td_left"
    nrange(wb, "TripCountries", "Trip Profile", "B", 15, 20)


# ===========================================================================
# 3 — Master Itinerary
# ===========================================================================
def build_itinerary(wb):
    ws = wb.create_sheet("Itinerary"); ws.sheet_view.showGridLines = False
    set_widths(ws, [7, 13, 14, 18, 16, 24, 22])
    luxe_header(ws, "G", "🗺  MASTER ITINERARY", "Day-by-day plan with automatic day numbering — your whole trip mapped out.")
    table_headers(ws, 4, ["Day", "Date", "City", "Hotel", "Transport", "Activities", "Notes"])
    plan = [
        ("Paris", "Le Marais Hotel", "Arrive / flight", "Check in, Seine walk", ""),
        ("Paris", "Le Marais Hotel", "Metro", "Louvre + Tuileries", "Book timed entry"),
        ("Paris", "Le Marais Hotel", "Walk", "Eiffel Tower, Montmartre", ""),
        ("Paris", "Le Marais Hotel", "Train to Rome", "Travel day", "Early train"),
        ("Rome", "Trastevere Suites", "—", "Colosseum + Forum", "Guided tour booked"),
        ("Rome", "Trastevere Suites", "Walk", "Vatican Museums", "Book skip-the-line"),
        ("Rome", "Trastevere Suites", "Walk", "Trevi, Pantheon, gelato", ""),
        ("Rome", "Trastevere Suites", "Flight to BCN", "Travel day", ""),
        ("Barcelona", "Eixample Loft", "—", "Sagrada Família", "Tickets booked"),
        ("Barcelona", "Eixample Loft", "Metro", "Park Güell + Gothic Qtr", ""),
        ("Barcelona", "Eixample Loft", "Walk", "Beach + tapas crawl", ""),
        ("Barcelona", "Eixample Loft", "—", "Day trip: Montserrat", "Train"),
        ("Barcelona", "Eixample Loft", "—", "Free day / shopping", ""),
        ("Barcelona", "Eixample Loft", "Flight home", "Depart", "Airport 3h early"),
    ]
    start = L0
    for i, (city, hotel, trans, act, note) in enumerate(plan):
        r = start + i
        dn = ws.cell(row=r, column=1, value=f"=IF(B{r}=\"\",\"\",B{r}-DepartDate+1)"); dn.number_format = "0"
        ws.cell(row=r, column=2, value=dplus(45 + i))
        ws.cell(row=r, column=3, value=city); ws.cell(row=r, column=4, value=hotel)
        ws.cell(row=r, column=5, value=trans); ws.cell(row=r, column=6, value=act); ws.cell(row=r, column=7, value=note)
    end = start + 20 - 1
    style_rows(ws, start, end, 7, text_left={3, 4, 5, 6, 7}, dates={2})
    for r in range(start, end + 1):
        ws.cell(row=r, column=1).number_format = "0"
    nrange(wb, "ItinDate", "Itinerary", "B", start, end)
    nrange(wb, "ItinCity", "Itinerary", "C", start, end)
    ws.freeze_panes = "A5"


# ===========================================================================
# 5 — Expense Tracker (feeds budget)
# ===========================================================================
def build_expenses(wb):
    sample = [
        (dminus(30), "Flights", "Round-trip flights (x2)", "USD", 1200, "Card", "Me"),
        (dminus(30), "Flights", "Intra-Europe segments", "USD", 1000, "Card", "Me"),
        (dminus(25), "Hotels", "Paris + Rome deposits", "USD", 1400, "Card", "Me"),
        (dminus(22), "Vacation Rentals", "Barcelona loft (prepaid)", "USD", 300, "Card", "Partner"),
        (dminus(20), "Travel Insurance", "SafeTrip policy (x2)", "USD", 180, "Card", "Me"),
        (dminus(15), "Activities", "Vatican skip-the-line (x2)", "USD", 220, "Card", "Me"),
        (dminus(12), "Activities", "Sagrada Família tickets", "USD", 200, "Card", "Partner"),
        (dminus(8), "Phone / Data", "eSIM data plans (x2)", "USD", 60, "Card", "Me"),
        (dminus(5), "Transportation", "Rome↔airport transfers", "USD", 120, "Card", "Me"),
        (dminus(2), "Miscellaneous", "Packing cubes & adapters", "USD", 40, "Card", "Me"),
    ]
    ws, start, end = build_log(
        wb, "Expenses", "🧾", "EXPENSE TRACKER",
        "Log every purchase — your budget actuals update automatically by category.",
        ["Date", "Category", "Description", "Currency", "Amount", "Method", "Traveler"],
        sample, [13, 16, 26, 10, 12, 12, 12],
        text_left={3}, dates={1}, money={5}, validations=[("B", "ExpenseCatList"), ("D", "CurrencyList")], reserved=60)
    nrange(wb, "ExpDate", "Expenses", "A", start, end)
    nrange(wb, "ExpCat", "Expenses", "B", start, end)
    nrange(wb, "ExpAmount", "Expenses", "E", start, end)


# ===========================================================================
# 4 — Travel Budget Command Center
# ===========================================================================
def build_budget(wb):
    ws = wb.create_sheet("Budget"); ws.sheet_view.showGridLines = False
    set_widths(ws, [20, 14, 14, 14, 12, 3, 20, 14])
    luxe_header(ws, "H", "💰  TRAVEL BUDGET COMMAND CENTER",
                "16 categories, plan vs actual — actuals pull from your Expense Tracker automatically.")
    table_headers(ws, 4, ["Category", "Budget", "Actual", "Remaining", "% Used"])
    planned = {"Flights": 2200, "Hotels": 2000, "Vacation Rentals": 300, "Transportation": 400,
               "Fuel": 0, "Food": 900, "Drinks": 250, "Activities": 700, "Shopping": 400,
               "Entertainment": 200, "Travel Insurance": 180, "Visas": 0, "Souvenirs": 200,
               "Phone / Data": 80, "Tips": 150, "Miscellaneous": 240}
    start = L0; end = start + len(EXPENSE_CATS) - 1
    for i, cat in enumerate(EXPENSE_CATS):
        r = start + i
        ws.cell(row=r, column=1, value=cat).style = "td_left"
        cp = ws.cell(row=r, column=2, value=planned[cat]); cp.style = "input"; cp.number_format = '"$"#,##0'
        ca = ws.cell(row=r, column=3, value=f'=IFERROR(SUMIF(ExpCat,A{r},ExpAmount),0)'); ca.style = "td"; ca.number_format = '"$"#,##0'
        cr = ws.cell(row=r, column=4, value=f"=B{r}-C{r}"); cr.style = "td"; cr.number_format = '"$"#,##0;[Red]-"$"#,##0'
        cu = ws.cell(row=r, column=5, value=f"=IFERROR(C{r}/B{r},0)"); cu.style = "td"; cu.number_format = "0%"
        if i % 2:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = fill(MUTED_ROW)
    total = end + 1
    ws.cell(row=total, column=1, value="TOTAL").style = "th"
    for col in range(2, 5):
        L = get_column_letter(col)
        c = ws.cell(row=total, column=col, value=f"=SUM({L}{start}:{L}{end})"); c.style = "td"
        c.font = Font(bold=True, color=PRIMARY); c.fill = fill(SURFACE); c.number_format = '"$"#,##0'
    cu = ws.cell(row=total, column=5, value=f"=IFERROR(C{total}/B{total},0)"); cu.style = "td"
    cu.font = Font(bold=True, color=PRIMARY); cu.fill = fill(SURFACE); cu.number_format = "0%"
    ws.conditional_formatting.add(f"E{start}:E{end}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.2, color=PRIMARY, showValue=True))
    ws.conditional_formatting.add(f"D{start}:D{end}", CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG)))
    nrange(wb, "BudgetCat", "Budget", "A", start, end)
    nrange(wb, "BudgetActual", "Budget", "C", start, end)
    wb.defined_names["BudgetTotalPlanned"] = DefinedName("BudgetTotalPlanned", attr_text=f"Budget!$B${total}")
    wb.defined_names["BudgetTotalActual"] = DefinedName("BudgetTotalActual", attr_text=f"Budget!$C${total}")
    ws.cell(row=4, column=7, value="TRIP MATH").style = "section_gold"
    kpis = [("Total budget", "=BudgetTotalPlanned", '"$"#,##0'), ("Spent so far", "=BudgetTotalActual", '"$"#,##0'),
            ("Remaining", "=BudgetTotalPlanned-BudgetTotalActual", '"$"#,##0;[Red]-"$"#,##0'),
            ("Budget / day", "=IFERROR(BudgetTotalPlanned/TripDays,0)", '"$"#,##0'),
            ("Cost / traveler", "=IFERROR(BudgetTotalPlanned/Travelers,0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(kpis):
        r = 5 + i
        ws.cell(row=r, column=7, value=lab).style = "field_label"
        c = ws.cell(row=r, column=8, value=fml); c.style = "field_value"; c.number_format = fmt
    donut = DoughnutChart(); donut.title = "Spending by Category"; donut.height = 8.5; donut.width = 12
    donut.add_data(Reference(ws, min_col=3, min_row=4, max_row=end), titles_from_data=True)
    donut.set_categories(Reference(ws, min_col=1, min_row=start, max_row=end))
    donut.dataLabels = no_labels(); ws.add_chart(donut, "G11")
    ws.freeze_panes = "A5"


# ===========================================================================
# 6 — Flight Manager
# ===========================================================================
def build_flights(wb):
    ws, s, e = build_log(
        wb, "Flights", "🛫", "FLIGHT MANAGER",
        "Every flight in one place — confirmations, seats, times & baggage.",
        ["Airline", "Flight #", "From → To", "Depart", "Arrive", "Confirmation", "Seat", "Baggage", "Cost"],
        [
            ("SkyEurope", "SE 108", "JFK → CDG", dplus(45), dplus(46), "SKY-88231", "22A", "1 checked", 620),
            ("SkyEurope", "SE 412", "CDG → FCO", dplus(49), dplus(49), "SKY-88232", "14C", "1 checked", 180),
            ("IberFly", "IB 621", "FCO → BCN", dplus(53), dplus(53), "IBR-55120", "9F", "1 checked", 160),
            ("SkyEurope", "SE 109", "BCN → JFK", dplus(59), dplus(59), "SKY-88240", "18B", "1 checked", 640),
        ],
        [14, 10, 16, 13, 13, 14, 8, 12, 10],
        text_left={3}, dates={4, 5}, money={9}, reserved=20)
    nrange(wb, "FlightAirline", "Flights", "A", s, e)


# ===========================================================================
# 7 — Accommodation Manager
# ===========================================================================
def build_hotels(wb):
    ws, s, e = build_log(
        wb, "Hotels", "🏨", "ACCOMMODATION MANAGER",
        "Hotels & rentals — confirmations, dates, cost & amenities in one view.",
        ["Property", "Type", "City", "Check-In", "Check-Out", "Confirmation", "Cost", "Amenities"],
        [
            ("Le Marais Hotel", "Hotel", "Paris", dplus(45), dplus(49), "LMH-3021", 720, "Breakfast, WiFi"),
            ("Trastevere Suites", "Vacation Rental", "Rome", dplus(49), dplus(53), "TRS-8890", 560, "Kitchen, AC"),
            ("Eixample Loft", "Vacation Rental", "Barcelona", dplus(53), dplus(59), "EXL-1180", 300, "Balcony, WiFi"),
        ],
        [20, 16, 14, 13, 13, 14, 10, 22],
        text_left={1, 8}, dates={4, 5}, money={7}, validations=[("B", "AccomList")], reserved=20)
    nrange(wb, "HotelName", "Hotels", "A", s, e)


# ===========================================================================
# 9 — Activity Planner
# ===========================================================================
def build_activities(wb):
    ws, s, e = build_log(
        wb, "Activities", "🎟", "ACTIVITY PLANNER",
        "Attractions & tours — tickets, cost & booking status, so nothing sells out on you.",
        ["Attraction", "City", "Date", "Category", "Tickets", "Cost", "Confirmation", "Status"],
        [
            ("Louvre Museum", "Paris", dplus(46), "Museum", "2", 40, "LV-2211", "Booked"),
            ("Eiffel Tower summit", "Paris", dplus(47), "Sightseeing", "2", 78, "ET-9930", "Booked"),
            ("Colosseum guided tour", "Rome", dplus(50), "Tour", "2", 120, "CL-4420", "Booked"),
            ("Vatican Museums", "Rome", dplus(51), "Museum", "2", 100, "VT-7781", "Booked"),
            ("Sagrada Família", "Barcelona", dplus(54), "Sightseeing", "2", 90, "SF-1102", "Booked"),
            ("Montserrat day trip", "Barcelona", dplus(57), "Adventure", "2", 130, "—", "Planned"),
        ],
        [22, 14, 13, 14, 9, 10, 14, 12],
        text_left={1}, dates={3}, money={6}, ints={5}, validations=[("D", "ActivityCatList"), ("H", "StatusList")], reserved=30)
    nrange(wb, "ActStatus", "Activities", "H", s, e)
    ws.conditional_formatting.add(f"H{s}:H{e}", CellIsRule(operator="equal", formula=['"Booked"'], fill=fill(MINT_BG)))


# ===========================================================================
# Simple logs (transport, restaurants, road trip, group, journal)
# ===========================================================================
def build_simple(wb):
    build_log(
        wb, "Transport", "🚆", "TRANSPORTATION PLANNER",
        "Trains, transfers, rentals & rides — every leg between the big flights.",
        ["Type", "From → To", "Date", "Provider", "Confirmation", "Cost", "Notes"],
        [
            ("Train", "Paris → Rome", dplus(48), "EuroRail", "ER-2201", 180, "High-speed"),
            ("Airport Transfer", "Rome hotel ↔ FCO", dplus(53), "Local", "—", 120, "Both ways"),
            ("Metro Pass", "Paris (5-day)", dplus(45), "RATP", "—", 60, "x2"),
            ("Metro Pass", "Barcelona (4-day)", dplus(53), "TMB", "—", 45, "x2"),
            ("Day-trip Train", "BCN → Montserrat", dplus(57), "FGC", "—", 40, "Round trip x2"),
        ],
        [16, 20, 13, 14, 14, 10, 20],
        text_left={2, 7}, dates={3}, money={6}, validations=[("A", "TransportList")], reserved=30)

    build_log(
        wb, "Restaurants", "🍽", "RESTAURANT & FOOD PLANNER",
        "Where to eat — reservations, must-try dishes & estimated cost.",
        ["Restaurant", "City", "Cuisine", "Reservation", "Est. Cost", "Rating", "Must-Try", "Notes"],
        [
            ("Le Comptoir", "Paris", "French", dplus(46), 90, "★★★★★", "Duck confit", "Book ahead"),
            ("Da Enzo", "Rome", "Italian", dplus(50), 70, "★★★★★", "Cacio e pepe", "Cash only"),
            ("Cervecería Catalana", "Barcelona", "Tapas", "Walk-in", 60, "★★★★☆", "Patatas bravas", "Go early"),
            ("Bar del Pla", "Barcelona", "Tapas", dplus(55), 75, "★★★★★", "Ham croquettes", ""),
            ("Gelateria del Teatro", "Rome", "Gelato", "Walk-in", 15, "★★★★★", "Pistachio", ""),
        ],
        [20, 14, 14, 14, 11, 12, 18, 18],
        text_left={1, 7, 8}, dates={4}, money={5}, reserved=40)

    build_log(
        wb, "Road Trip", "🚗", "ROAD TRIP PLANNER",
        "For the driving legs — routes, fuel & hotel stops, distance & drive time.",
        ["Leg", "From → To", "Distance", "Drive Time", "Fuel Stops", "Hotel Stop", "Est. Fuel", "Notes"],
        [
            ("Optional", "Rome → Florence", "170 mi", "3h", "1", "—", 40, "If time allows"),
            ("Optional", "BCN → Costa Brava", "60 mi", "1.5h", "0", "—", 20, "Beach day"),
        ],
        [14, 20, 12, 12, 11, 14, 11, 18],
        text_left={2, 8}, money={7}, reserved=20)

    build_log(
        wb, "Group Travel", "👥", "GROUP TRAVEL MANAGER",
        "Traveling together — who paid what, shared expenses & who owes whom.",
        ["Traveler", "Paid For", "Amount", "Split", "Their Share", "Owes / Owed", "Notes"],
        [
            ("Me", "Flights (x2)", 1200, "50/50", 600, "+$600", "Partner owes"),
            ("Partner", "Barcelona loft", 300, "50/50", 150, "-$150", "I owe"),
            ("Me", "Activities (x2)", 420, "50/50", 210, "+$210", ""),
            ("Partner", "Insurance", 180, "50/50", 90, "-$90", ""),
        ],
        [14, 20, 12, 12, 13, 13, 18],
        text_left={2, 6, 7}, money={3, 5}, reserved=20)

    build_log(
        wb, "Journal", "📓", "TRAVEL JOURNAL",
        "The trip you'll want to remember — a few lines a day is all it takes.",
        ["Date", "Location", "Highlight", "Favorite Moment", "Weather", "Spent", "Reflections"],
        [
            (dplus(45), "Paris", "Landed & Seine sunset", "First croissant!", "Sunny 68°", 45, ""),
            (dplus(46), "Paris", "Louvre + climbed Eiffel", "Top at golden hour", "Clear 70°", 120, ""),
        ],
        [13, 14, 22, 22, 12, 10, 24],
        text_left={3, 4, 7}, dates={1}, money={6}, reserved=40)


# ===========================================================================
# 11 — Packing Command Center
# ===========================================================================
def build_packing(wb):
    ws = wb.create_sheet("Packing"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 8, 22, 3, 20, 8, 22, 3, 14, 10])
    luxe_header(ws, "K", "🧳  PACKING COMMAND CENTER", "Category checklists with automatic completion — pack smart, forget nothing.")
    cats = {
        "CLOTHING": ["Tops x7", "Bottoms x4", "Dresses x2", "Jacket", "Sleepwear", "Underwear x8", "Socks x8", "Swimwear"],
        "TOILETRIES": ["Toothbrush + paste", "Deodorant", "Skincare", "Sunscreen", "Meds", "Razor", "Hairbrush"],
        "ELECTRONICS": ["Phone + charger", "EU adapters x2", "Power bank", "Headphones", "Camera + charger", "eSIM active"],
        "DOCUMENTS": ["Passports x2", "Boarding passes", "Hotel confirmations", "Insurance card", "Copies (digital)"],
    }
    blocks = [("CLOTHING", 2), ("TOILETRIES", 6), ("ELECTRONICS", 2), ("DOCUMENTS", 6)]
    positions = {"CLOTHING": (5, 2), "TOILETRIES": (5, 6), "ELECTRONICS": (18, 2), "DOCUMENTS": (18, 6)}
    pack_cells = []
    done_pattern = 0
    for cat, (top, cdone) in positions.items():
        ctext = cdone + 1
        merge_set(ws, f"{get_column_letter(cdone)}{top}:{get_column_letter(ctext)}{top}", cat, "section_gold")
        ws.row_dimensions[top].height = 22
        for i, it in enumerate(cats[cat]):
            r = top + 1 + i
            # ~30% done
            val = "Yes" if (len(pack_cells) % 10 < 3) else "No"
            cb = ws.cell(row=r, column=cdone, value=val); cb.style = "input"; add_dv(ws, f"{get_column_letter(cdone)}{r}", "YesNoList")
            ct = ws.cell(row=r, column=ctext, value=it); ct.style = "td_left"
            ws.merge_cells(f"{get_column_letter(ctext)}{r}:{get_column_letter(ctext)}{r}")
            pack_cells.append((cdone, r))
            if i % 2:
                ct.fill = fill(MUTED_ROW)
    # helper consolidation column
    hcol = 10
    for idx, (cd, r) in enumerate(pack_cells):
        ws.cell(row=5 + idx, column=hcol, value=f'={get_column_letter(cd)}{r}')
    nrange(wb, "PackDone", "Packing", get_column_letter(hcol), 5, 5 + len(pack_cells) - 1)
    # completion summary
    ws.cell(row=4, column=11, value="DONE").style = "section_gold"
    c = ws.cell(row=5, column=11, value='=IFERROR(COUNTIF(PackDone,"Yes")/MAX(COUNTA(PackDone),1),0)')
    c.number_format = "0%"; c.font = Font(size=20, bold=True, color=PRIMARY); c.alignment = Alignment(horizontal="center")


# ===========================================================================
# 12 — Document Vault Index
# ===========================================================================
def build_documents(wb):
    ws, s, e = build_log(
        wb, "Documents", "🗂", "DOCUMENT VAULT INDEX",
        "Every travel document tracked — expiry, where it lives & whether it's ready.",
        ["Document", "For Whom", "Expiry", "Storage", "Digital Copy?", "Ready?", "Notes"],
        [
            ("Passport", "Me", dplus(900), "Wallet + safe", "Yes", "Yes", ""),
            ("Passport", "Partner", dplus(700), "Wallet + safe", "Yes", "Yes", ""),
            ("Travel Insurance", "Both", dplus(60), "Email + app", "Yes", "Yes", "SafeTrip"),
            ("Boarding Passes", "Both", "", "Airline app", "Yes", "Yes", "Check-in 24h prior"),
            ("Hotel Confirmations", "Both", "", "Email folder", "Yes", "Yes", ""),
            ("Activity Tickets", "Both", "", "Email + wallet", "Yes", "Yes", ""),
            ("Driver's License", "Me", dplus(500), "Wallet", "Yes", "Yes", "For any driving"),
            ("EU Entry / ETIAS", "Both", "", "—", "No", "No", "Check if required"),
        ],
        [22, 14, 13, 16, 13, 11, 20],
        text_left={1, 7}, dates={3}, validations=[("E", "YesNoList"), ("F", "YesNoList")], reserved=30)
    nrange(wb, "DocName", "Documents", "A", s, e)
    nrange(wb, "DocReady", "Documents", "F", s, e)
    ws.conditional_formatting.add(f"F{s}:F{e}", CellIsRule(operator="equal", formula=['"Yes"'], fill=fill(MINT_BG)))
    ws.conditional_formatting.add(f"F{s}:F{e}", CellIsRule(operator="equal", formula=['"No"'], fill=fill(WARN_BG)))


# ===========================================================================
# 13 — Currency & Exchange
# ===========================================================================
def build_currency(wb):
    ws = wb.create_sheet("Currency"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 26, 16, 4, 24, 16, 2])
    luxe_header(ws, "G", "💱  CURRENCY & EXCHANGE TRACKER", "Home vs destination currency — quick conversions & spend summaries.")
    merge_set(ws, "B5:C5", "EXCHANGE", "section_gold")
    rows = [("Home currency", "=HomeCurr", None), ("Destination currency", "=DestCurr", None),
            ("Exchange rate (1 home =)", "=ExchRate", "0.000"),
            ("Total spent (home)", "=BudgetTotalActual", '"$"#,##0'),
            ("Total spent (dest)", "=BudgetTotalActual*ExchRate", "#,##0")]
    for i, (lab, fml, fmt) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"
        if fmt:
            c.number_format = fmt
    merge_set(ws, "E5:F5", "QUICK CONVERTER", "section_gold")
    ws.cell(row=6, column=5, value="Amount (home)").style = "field_label"
    amt = ws.cell(row=6, column=6, value=100); amt.style = "input"; amt.number_format = '"$"#,##0'
    ws.cell(row=7, column=5, value="= in destination").style = "field_label"
    c = ws.cell(row=7, column=6, value="=F6*ExchRate"); c.style = "field_value"; c.number_format = "#,##0.00"
    ws.cell(row=9, column=5, value="Amount (dest)").style = "field_label"
    amt2 = ws.cell(row=9, column=6, value=50); amt2.style = "input"; amt2.number_format = "#,##0"
    ws.cell(row=10, column=5, value="= in home").style = "field_label"
    c = ws.cell(row=10, column=6, value="=IFERROR(F9/ExchRate,0)"); c.style = "field_value"; c.number_format = '"$"#,##0.00'


# ===========================================================================
# 14 — Savings Planner
# ===========================================================================
def build_savings(wb):
    ws = wb.create_sheet("Savings"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 28, 18, 4, 22, 16, 2])
    luxe_header(ws, "G", "🐷  SAVINGS PLANNER", "Fund the trip stress-free — track progress toward your target budget.")
    merge_set(ws, "B5:C5", "SAVINGS GOAL", "section_gold")
    rows = [("Target (trip budget)", "=BudgetTotalPlanned", '"$"#,##0'),
            ("Saved so far", "=SavingsSaved", '"$"#,##0'),
            ("Still to save", "=MAX(BudgetTotalPlanned-SavingsSaved,0)", '"$"#,##0'),
            ("Progress", "=IFERROR(SavingsSaved/BudgetTotalPlanned,0)", "0%"),
            ("Months to departure", "=MAX(ROUND((DepartDate-TODAY())/30,1),0)", "0.0"),
            ("Monthly needed", "=IFERROR(MAX(BudgetTotalPlanned-SavingsSaved,0)/MAX((DepartDate-TODAY())/30,1),0)", '"$"#,##0')]
    for i, (lab, fml, fmt) in enumerate(rows):
        r = 6 + i
        ws.cell(row=r, column=2, value=lab).style = "field_label"
        c = ws.cell(row=r, column=3, value=fml); c.style = "field_value"; c.number_format = fmt
        if lab == "Progress":
            ws.cell(row=r, column=3).fill = fill(MINT_BG)
    wb.defined_names["SavProgress"] = DefinedName("SavProgress", attr_text="Savings!$C$9")
    # monthly contribution plan
    merge_set(ws, "E5:F5", "CONTRIBUTION PLAN", "section_gold")
    ws.cell(row=6, column=5, value="Month").style = "th"; ws.cell(row=6, column=6, value="Added").style = "th"
    for i, (m, v) in enumerate([("Jan", 1500), ("Feb", 1500), ("Mar", 1500), ("Apr", 1500)]):
        r = 7 + i
        ws.cell(row=r, column=5, value=m).style = "td_left"
        c = ws.cell(row=r, column=6, value=v); c.style = "td"; c.number_format = '"$"#,##0'


# ===========================================================================
# 17 — Checklist Command Center
# ===========================================================================
def build_checklists(wb):
    ws = wb.create_sheet("Checklists"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 4, 28, 4, 4, 28, 2])
    luxe_header(ws, "G", "✅  CHECKLIST COMMAND CENTER", "The little things that make travel smooth — ticked off at every stage.")
    lists = [
        ("BEFORE DEPARTURE", 5, 2, ["Check passport expiry", "Notify bank of travel", "Buy travel insurance",
                                    "Download offline maps", "Set up eSIM / roaming", "Arrange pet / plant care",
                                    "Hold mail / packages", "Charge all devices"]),
        ("AIRPORT & FLIGHT", 5, 5, ["Check in online (24h)", "Weigh bags", "Passport + boarding pass out",
                                    "Snacks + water bottle", "Neck pillow & meds", "Screenshot confirmations"]),
        ("HOTEL CHECK-IN", 15, 2, ["Confirm reservation", "Photo of room condition", "Find nearest pharmacy",
                                   "Save hotel address (offline)", "Note checkout time"]),
        ("RETURN HOME", 15, 5, ["Charge phone for travel", "Check under beds & drawers", "Keep receipts for taxes",
                                "Back up photos", "Leave a review", "Unpack + laundry"]),
    ]
    for title, top, cbox, items in lists:
        ctext = cbox + 1
        merge_set(ws, f"{get_column_letter(cbox)}{top}:{get_column_letter(ctext)}{top}", title, "section_gold")
        ws.row_dimensions[top].height = 22
        for i, it in enumerate(items):
            r = top + 1 + i
            cb = ws.cell(row=r, column=cbox, value="☐"); cb.alignment = Alignment(horizontal="center"); cb.font = Font(size=12, color=ACCENT); cb.border = BOX
            ct = ws.cell(row=r, column=ctext, value=it); ct.style = "td_left"
            if i % 2:
                cb.fill = fill(MUTED_ROW); ct.fill = fill(MUTED_ROW)


# ===========================================================================
# 18 — Photo & Memory Gallery
# ===========================================================================
def build_memories(wb):
    ws = wb.create_sheet("Memories"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 20, 20, 14, 3, 20, 20, 14, 2])
    luxe_header(ws, "I", "📸  PHOTO & MEMORY GALLERY", "The trip you'll relive for years — paste photos with the story behind each.")
    sections = ["Destination Highlight", "Landmark", "Best Meal", "Adventure",
                "Us / Family", "Souvenir", "Favorite Moment", "Sunset / View"]
    top0 = 5; card_h = 9
    for idx, name in enumerate(sections):
        col = 2 if idx % 2 == 0 else 6
        row = top0 + (idx // 2) * card_h
        L = get_column_letter(col); M = get_column_letter(col + 1); R = get_column_letter(col + 2)
        merge_set(ws, f"{L}{row}:{R}{row}", f"  {name}", "th"); ws.row_dimensions[row].height = 22
        merge_set(ws, f"{L}{row+1}:{R}{row+5}", "📷\nPaste photo here\n(Insert ▸ Picture)", "imgbox")
        for rr in range(row + 1, row + 6):
            ws.row_dimensions[rr].height = 24
        ws.cell(row=row + 6, column=col, value="Location / Date").style = "field_label"
        merge_set(ws, f"{M}{row+6}:{R}{row+6}", "", "field_value")
        ws.cell(row=row + 7, column=col, value="Caption").style = "field_label"
        merge_set(ws, f"{M}{row+7}:{R}{row+7}", "", "field_value")


# ===========================================================================
# 20 — Analytics Dashboard
# ===========================================================================
def build_analytics(wb):
    ws = wb.create_sheet("Analytics"); ws.sheet_view.showGridLines = False
    set_widths(ws, [2, 30, 16, 18, 3, 16, 12, 12, 12, 12, 12, 2])
    luxe_header(ws, "L", "📈  ANALYTICS DASHBOARD", "Your trip by the numbers — budget health, readiness & a Trip Readiness Score.")
    merge_set(ws, "B5:D5", "TRIP READINESS SCORES", "section")
    table_headers(ws, 6, ["Dimension", "Score", "Status"], start_col=2)
    metrics = [
        ("Documents ready", '=IFERROR(COUNTIF(DocReady,"Yes")/MAX(COUNTA(DocName),1),0)'),
        ("Packing complete", '=IFERROR(COUNTIF(PackDone,"Yes")/MAX(COUNTA(PackDone),1),0)'),
        ("Savings funded", "=SavProgress"),
        ("Activities booked", '=IFERROR(COUNTIF(ActStatus,"Booked")/MAX(COUNTA(ActStatus),1),0)'),
        ("Itinerary planned", '=IFERROR(COUNTA(ItinCity)/MAX(TripDays,1),0)'),
        ("On budget", '=IFERROR(1-BudgetTotalActual/BudgetTotalPlanned,0)'),
    ]
    start = 7
    for i, (dim, fml) in enumerate(metrics):
        r = start + i
        ws.cell(row=r, column=2, value=dim).style = "td_left"
        c = ws.cell(row=r, column=3, value=fml); c.style = "td"; c.number_format = "0%"
        ws.cell(row=r, column=4, value=f'=IF(C{r}>=0.75,"Ready",IF(C{r}>=0.4,"Getting There","To Do"))').style = "td"
        if i % 2:
            for c2 in range(2, 5):
                ws.cell(row=r, column=c2).fill = fill(MUTED_ROW)
    end = start + len(metrics) - 1
    ws.conditional_formatting.add(f"C{start}:C{end}", ColorScaleRule(
        start_type="num", start_value=0, start_color="FF" + WARN_BG, mid_type="num", mid_value=0.5,
        mid_color="FFFFF3CD", end_type="num", end_value=1, end_color="FF" + HIGHLIGHT))
    merge_set(ws, "F5:H5", "TRIP READINESS SCORE", "section_gold")
    ws.merge_cells("F6:H9")
    cell = ws["F6"]; cell.value = f"=IFERROR(AVERAGE(C{start}:C{start+4}),0)"
    cell.font = Font(size=46, bold=True, color=PRIMARY); cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = "0%"; cell.fill = fill(IVORY)
    for rr in range(6, 10):
        for cc in (6, 7, 8):
            ws.cell(row=rr, column=cc).fill = fill(IVORY)
            ws.cell(row=rr, column=cc).border = Border(top=GOLD if rr == 6 else THIN, bottom=THIN, left=THIN, right=THIN)
    merge_set(ws, "F10:H10", "A blend of documents, packing, savings, bookings & itinerary.", "subtitle")
    ws["F10"].fill = fill(IVORY)
    bar = BarChart(); bar.type = "bar"; bar.title = "Readiness by Area"; bar.height = 9; bar.width = 13
    bar.add_data(Reference(ws, min_col=3, min_row=6, max_row=end), titles_from_data=True)
    bar.set_categories(Reference(ws, min_col=2, min_row=start, max_row=end))
    bar.legend = None; ws.add_chart(bar, "B17")


# ===========================================================================
# 1 — Executive Travel Dashboard
# ===========================================================================
def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard", 0); ws.sheet_view.showGridLines = False
    set_widths(ws, [2] + [12] * 12 + [2]); ws.row_dimensions[1].height = 56
    merge_set(ws, "A1:N1", "  ✈  TRAVEL COMMAND CENTER™", "title")
    ws.row_dimensions[2].height = 24
    merge_set(ws, "A2:N2", "  Budget, itinerary, bookings, packing & readiness — your whole trip, automatically organized.", "subtitle")
    ws.row_dimensions[3].height = 6
    for c in range(1, 15):
        ws.cell(row=3, column=c).fill = fill(GOLD_LT)
    row1 = [
        ("DAYS TO DEPARTURE", "=MAX(DepartDate-TODAY(),0)", "days"),
        ("TOTAL BUDGET", "=BudgetTotalPlanned", "money"),
        ("BUDGET LEFT", "=BudgetTotalPlanned-BudgetTotalActual", "money"),
        ("SPENT SO FAR", "=BudgetTotalActual", "money"),
        ("FLIGHTS BOOKED", "=COUNTA(FlightAirline)", "num"),
        ("HOTELS CONFIRMED", "=COUNTA(HotelName)", "num"),
    ]
    row2 = [
        ("ACTIVITIES BOOKED", '=COUNTIF(ActStatus,"Booked")', "num"),
        ("COUNTRIES", "=COUNTA(TripCountries)", "num"),
        ("DOCS READY", '=IFERROR(COUNTIF(DocReady,"Yes")/MAX(COUNTA(DocName),1),0)', "pct"),
        ("PACKING", '=IFERROR(COUNTIF(PackDone,"Yes")/MAX(COUNTA(PackDone),1),0)', "pct"),
        ("SAVINGS GOAL", "=SavProgress", "pct"),
        ("READINESS", "=IFERROR(AVERAGE(Analytics!C7:C11),0)", "pct"),
    ]
    cols6 = [2, 4, 6, 8, 10, 12]
    for (lab, fml, kind), col in zip(row1, cols6):
        kpi_card(ws, 5, col, 2, lab, fml, kind)
    for (lab, fml, kind), col in zip(row2, cols6):
        kpi_card(ws, 8, col, 2, lab, fml, kind)
    ws.row_dimensions[11].height = 26
    merge_set(ws, "B11:M11", "BUDGET & SPENDING", "section_gold")
    bend = L0 + len(EXPENSE_CATS) - 1
    d1 = DoughnutChart(); d1.title = "Spending by Category"; d1.height = 8.2; d1.width = 11.5
    d1.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=bend), titles_from_data=True)
    d1.set_categories(Reference(wb["Budget"], min_col=1, min_row=L0, max_row=bend)); d1.dataLabels = no_labels()
    ws.add_chart(d1, "B12")
    # budget vs actual bar (planned/actual per category top rows)
    b1 = BarChart(); b1.type = "col"; b1.title = "Budget vs Actual"; b1.height = 8.2; b1.width = 11.5
    b1.add_data(Reference(wb["Budget"], min_col=2, min_row=4, max_row=L0 + 7), titles_from_data=True)
    b1.add_data(Reference(wb["Budget"], min_col=3, min_row=4, max_row=L0 + 7), titles_from_data=True)
    b1.set_categories(Reference(wb["Budget"], min_col=1, min_row=L0, max_row=L0 + 7))
    ws.add_chart(b1, "H12")
    ws.row_dimensions[29].height = 26
    merge_set(ws, "B29:M29", "READINESS", "section_gold")
    r1 = BarChart(); r1.type = "bar"; r1.title = "Trip Readiness by Area"; r1.height = 8.2; r1.width = 11.5
    r1.add_data(Reference(wb["Analytics"], min_col=3, min_row=6, max_row=12), titles_from_data=True)
    r1.set_categories(Reference(wb["Analytics"], min_col=2, min_row=7, max_row=12)); r1.legend = None
    ws.add_chart(r1, "B30")
    # savings donut-ish -> use bar of saved vs remaining
    ws2 = wb["Savings"]
    sv = DoughnutChart(); sv.title = "Savings Progress"; sv.height = 8.2; sv.width = 11.5
    ws2["E12"] = "Saved"; ws2["F12"] = "=SavingsSaved"; ws2["E13"] = "To go"; ws2["F13"] = "=MAX(BudgetTotalPlanned-SavingsSaved,0)"
    ws2["F12"].number_format = '"$"#,##0'; ws2["F13"].number_format = '"$"#,##0'
    sv.add_data(Reference(ws2, min_col=6, min_row=12, max_row=13), titles_from_data=False)
    sv.set_categories(Reference(ws2, min_col=5, min_row=12, max_row=13)); sv.dataLabels = no_labels()
    ws.add_chart(sv, "H30")
    ws.row_dimensions[47].height = 26
    merge_set(ws, "B47:M47", "Travel Command Center™ — your whole trip, organized in one place. Edit anything in Settings.", "subtitle")


# ===========================================================================
def main():
    wb = Workbook(); wb.remove(wb.active); register_styles(wb)
    build_settings(wb); build_welcome(wb); build_profile(wb); build_itinerary(wb)
    build_expenses(wb); build_budget(wb); build_flights(wb); build_hotels(wb)
    build_activities(wb); build_simple(wb)   # transport, restaurants, road trip, group, journal
    build_packing(wb); build_documents(wb); build_currency(wb); build_savings(wb)
    build_checklists(wb); build_memories(wb); build_analytics(wb); build_dashboard(wb)

    order = ["Welcome", "Dashboard", "Trip Profile", "Itinerary", "Budget", "Expenses",
             "Flights", "Hotels", "Transport", "Activities", "Restaurants", "Packing",
             "Documents", "Currency", "Savings", "Road Trip", "Group Travel",
             "Checklists", "Memories", "Journal", "Analytics", "Settings"]
    wb._sheets = [wb[n] for n in order]
    palette = [PRIMARY, ACCENT, HIGHLIGHT, SURFACE]
    for i, n in enumerate(order):
        wb[n].sheet_properties.tabColor = palette[i % len(palette)]
    wb["Welcome"].sheet_properties.tabColor = PRIMARY
    wb["Dashboard"].sheet_properties.tabColor = PRIMARY
    wb["Settings"].sheet_properties.tabColor = SURFACE
    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Travel_Command_Center.xlsx")
    wb.save(out)
    print(f"Wrote {out}  ({len(order)} sheets)")


if __name__ == "__main__":
    main()
